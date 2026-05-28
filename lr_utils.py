# -*- coding: utf-8 -*-
"""
lr_utils.py  ─  汎用ユーティリティ / Excel書式ヘルパー
"""


# ============================================================
# 決まり手カラムマップ（コース別マスタのカラム名エイリアス対応）
# ============================================================
KIMARI_COL_MAP = {
    "逃げ%":       ["逃げ%",       "逃げ率",   "nige_pct"],
    "差し%":       ["差し%",       "差し率",   "sashi_pct"],
    "まくり%":     ["まくり%",     "捲り%",    "makuri_pct"],
    "まくり差し%": ["まくり差し%", "捲差%",    "makurisa_pct"],
    "抜き%":       ["抜き%",       "抜き率",   "nuki_pct"],
    "差され%":     ["差され%",     "差され率"],
    "捲られ%":     ["捲られ%",     "まくられ%"],
    "捲り差され%": ["捲り差され%", "まくり差され%"],
}


def safe_float(value, default=None):
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _get_cm_val(cm: dict, col: str):
    if not cm:
        return None
    if col in cm:
        return cm[col]
    aliases = KIMARI_COL_MAP.get(col, [])
    for alias in aliases:
        if alias in cm:
            return cm[alias]
    return None


def sep(char="=", width=44):
    print(char * width)


# ============================================================
# openpyxl スタイルヘルパー
# ============================================================
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def make_fill(rgb: str):
        return PatternFill(fill_type="solid", fgColor=rgb)

    def make_font(bold=False, color="FF000000", size=9, name="メイリオ"):
        return Font(bold=bold, color=color, size=size, name=name)

    def center_align(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def left_align(wrap=False):
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    def thin_border():
        side = Side(style="thin")
        return Border(left=side, right=side, top=side, bottom=side)

    def write_cell(ws, row: int, col: int, value=None,
                   fill=None, font=None, alignment=None, border=None):
        cell = ws.cell(row=row, column=col)
        if value is not None:
            cell.value = value
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if border is not None:
            cell.border = border
        return cell

except ImportError:
    def make_fill(rgb): return None
    def make_font(bold=False, color="FF000000", size=9, name="メイリオ"): return None
    def center_align(wrap=False): return None
    def left_align(wrap=False): return None
    def thin_border(): return None
    def write_cell(ws, row, col, value=None, fill=None, font=None, alignment=None, border=None): return None
