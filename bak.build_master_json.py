"""
build_master_json.py  (v3.3 — winner_course_order に3着率を追加)
ボートリサーチ_マスタ.xlsx から master_data.json を生成する。

使い方:
    python build_master_json.py <xlsx_path> <output_json_path>

出力するデータ:
    course_master      : 選手×コース別（全国）実績 ＋ 決まり手%
    venue_course_master: 選手×会場×コース別実績
    venue_stats        : 会場×コース×R番号別1着率・荒れスコア ＋ inn_2place
    player_index       : 選手指数（フォーム・ST順位・FLY情報）
    meta               : ビルドパラメータ・荒れスコア統計
    venue_kimari       : 会場別 決まり手比率（逃げ/差し/まくり等）

【v3.3 変更点】
  1. winner_course_order に rate3（3着率）を追加
     （勝者コース別着順分析シート: 選手×自コース×勝者コース→2着率・3着率・3着以内率）
     ※ rate3i（3着以内率）は後方互換のため引き続き格納。rate3 は純粋な3着率。

【v3.2 変更点】
  1. tenkai_remaining を復活
     （展開別残存_全国シート: 決まり手×1着コース→各コースの2着率）
  2. winner_course_order を復活
     （勝者コース別着順分析シート: 選手×自コース×勝者コース→2着率）
  3. 両データを展開シナリオの2着確率計算に活用できるよう整備
"""

import sys
import json
import time
import statistics
from pathlib import Path
from openpyxl import load_workbook

# ── コース別 reliable 閾値（全コース一律20走）────────────────────────
COURSE_MIN_RUNS = {
    "1": 20,
    "2": 20,
    "3": 20,
    "4": 20,
    "5": 20,
    "6": 20,
}
VENUE_COURSE_MIN_RUNS = 20   # 会場別コースマスタの最小出走数

# ── ベイズ平均のパラメータ ──────────────────────────────────────────
# prior_weight: overall_win に付与する「仮想サンプル数」
# 大きいほど recent10 の影響を抑える（20 = 20走分の重みで全体実績を優先）
BAYESIAN_PRIOR_WEIGHT = 20
RECENT10_N = 10  # recent10_win のサンプル数

# ── ST順位の全選手平均（コース別）──────────────────────────────────
# 実データから算出済み。補正感度もコース1を最大として外側に向かって減衰
ST_AVG = {"1": 3.14, "2": 3.57, "3": 3.28, "4": 3.41, "5": 3.52, "6": 4.04}
ST_SENSITIVITY = {"1": 0.15, "2": 0.10, "3": 0.08, "4": 0.06, "5": 0.04, "6": 0.03}


# ──────────────────────────────────────────────────────────────────
# ユーティリティ
# ──────────────────────────────────────────────────────────────────

def safe_float(v, default=None):
    try:
        f = float(v)
        return f if f == f else default  # NaN チェック
    except (TypeError, ValueError):
        return default


def safe_int(v, default=None):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def is_reliable_course(runs: int, course: str, player_type: str = "", is_joshi: bool = False) -> bool:
    """コース別の閾値で信頼性を判定。
    グレードメイン選手は一般戦出走数が構造的に少ないため閾値を緩和する。
    　通常（一般メイン/混合/不明）: 20走
    　混合                         : 20走（一般戦と同じ扱い）
    　グレードメイン               :  1走（実績が1走でもあれば信頼できる）
    　　※ 6コース等は構造的に出走自体が少ない。5走閾値だと峰竜太クラスでも
    　　　 データなし扱いになるため1走に緩和。runs=0（完全に実績なし）は除外。
    女子戦マスタ（is_joshi=True）:
    　女子選手は年間レース数が男子より少なく20走閾値ではほぼデータ不足になるため
    　閾値を8走に緩和する。8走未満は ts_win_rate（ベイズ補正済み）へのフォールバック
    　で対応するため精度への影響は最小限に抑えられる。
    """
    if runs == 0:
        return False  # 実績ゼロは選手タイプ問わず信頼不可
    if player_type == "グレードメイン":
        threshold = 1
    elif is_joshi:
        threshold = 8  # 女子戦: 年間出走数の構造的少なさを考慮して緩和
    else:
        threshold = COURSE_MIN_RUNS.get(course, 20)
    return runs >= threshold


def calc_trust(runs: int, win_rate, venue_avg_win_rate) -> float:
    """
    venue_course_master の信頼度をスクリプト内で計算（改修: Excel依存を排除）

    Excel の trust 列が全件0.0だったため reliable が機能していなかった。
    runs とコース平均との乖離から 0〜1 の信頼度を算出する。

    Args:
        runs             : 出走数
        win_rate         : 選手の会場別コース勝率
        venue_avg_win_rate: 同会場・同コースの全選手平均勝率（ベースレート）
    Returns:
        float: 0.0〜1.0 の信頼度
    """
    if runs < VENUE_COURSE_MIN_RUNS:
        return 0.0
    # 出走数スコア（20走で満点、10〜20走は線形補完）
    run_score = min(runs / 20.0, 1.0)
    # 会場平均との乖離（極端に乖離していれば信頼度を下げる）
    if win_rate is not None and venue_avg_win_rate is not None:
        dev = abs(win_rate - venue_avg_win_rate)
        dev_score = max(0.0, 1.0 - dev * 4.0)
    else:
        dev_score = 0.5
    return round(run_score * 0.7 + dev_score * 0.3, 4)


def bayesian_win_rate(recent10: float, overall: float) -> float:
    """
    ベイズ平均による recent10_win の補正（直近10走の統計ノイズを抑制）

    recent10_win=1.000（直近10走全勝）でも overall で引き戻す。
    prior_weight=20 は「全体実績を20走分の重みで信頼する」設定。

    例: 坪井康晴 overall=0.374, recent10=0.800
        → (0.374*20 + 0.800*10) / (20+10) = 0.516（過大評価を抑制）
    """
    if recent10 is None:
        return overall
    if overall is None:
        return recent10
    return round(
        (overall * BAYESIAN_PRIOR_WEIGHT + recent10 * RECENT10_N)
        / (BAYESIAN_PRIOR_WEIGHT + RECENT10_N),
        4,
    )


def calc_composite_win_rate(
    overall: float,
    bayesian: float,
    arek_score: float,
    runs: int,
) -> float:
    """
    荒れスコアに応じて overall と bayesian の重みを動的合成（改修: 新規追加）

    荒れやすい会場（arek_score 高）ほど直近フォームが重要。
    サンプルが少ない場合は recent5 の重みを抑制する。

    arek_score の実データ範囲: 39.2（大村）〜 60.1（戸田）
    """
    if overall is None:
        return bayesian
    if bayesian is None:
        return overall

    # 荒れスコアを 0〜1 に正規化（39〜60の実範囲を使用）
    arek_norm = max(0.0, min((arek_score - 39.0) / (60.0 - 39.0), 1.0))
    # 最大で 0.4 の重みを recent 側に付与
    recent_weight = arek_norm * 0.4
    # サンプルが少なければ抑制（30走で満点）
    recent_weight *= min(runs / 30.0, 1.0)

    return round(overall * (1.0 - recent_weight) + bayesian * recent_weight, 4)


def st_adjusted_win_rate(base_win_rate: float, st_rank: float, course: str) -> float:
    """
    ST順位による1着率補正（改修: 新規追加）

    全選手平均ST順位との差分をベースに補正係数を計算。
    コース1はSTの影響が大きく、外コースは小さい。
    """
    if base_win_rate is None or st_rank is None:
        return base_win_rate
    avg = ST_AVG.get(course, 3.5)
    sensitivity = ST_SENSITIVITY.get(course, 0.05)
    diff = avg - st_rank  # 正 = 平均より早い
    adjusted = base_win_rate * (1.0 + diff * sensitivity)
    return round(max(0.0, min(adjusted, 1.0)), 4)


# ──────────────────────────────────────────────────────────────────
# 各シートのビルド関数
# ──────────────────────────────────────────────────────────────────

def build_course_master(wb, player_index: dict = None, sheet_name: str = "📊コース別マスタ", is_joshi: bool = False) -> dict:
    """
    📊コース別マスタ → course_master[選手名][コース文字列] = {...}

    格納するフィールド:
        runs          : 出走数
        win_rate      : 1着率（生値）
        ts_win_rate   : 時系列補正1着率
        st_rank       : コース別ST順位（小さいほど早い）
        reliable      : コース別閾値による信頼フラグ（改修）
                        グレードメイン選手は閾値5走に緩和（is_reliable_course参照）
        kimari        : 決まり手%（1着時の決まり手分布）
        被kimari      : 被決まり手%（コース1のみ: 差され/捲られ/捲り差され）
    """
    if sheet_name not in wb.sheetnames:
        print(f"  [SKIP] シート '{sheet_name}' が存在しないためスキップ")
        return {}
    ws = wb[sheet_name]
    _pi = player_index or {}
    master = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        name = row[0]
        if not name or not isinstance(name, str):
            continue

        course = safe_int(row[1])
        if course is None or course not in range(1, 7):
            continue

        runs     = safe_int(row[2], 0)
        win_rate = safe_float(row[4], None)
        st_rank  = safe_float(row[27], None)
        ts_win   = safe_float(row[38], None)

        if name not in master:
            master[name] = {}

        c = str(course)
        kimari = {}
        for kname, kidx in COURSE_KIMARI_COLS.items():
            v = safe_float(row[kidx])
            if v is not None:
                kimari[kname] = v

        entry = {
            "runs":        runs,
            "win_rate":    win_rate,
            "ts_win_rate": ts_win,
            "st_rank":     st_rank,
            "reliable":    is_reliable_course(runs, c, _pi.get(name, {}).get("player_type", ""), is_joshi=is_joshi),
            "kimari":      kimari,  # 決まり手%（1着時の決まり手分布）
        }

        # 被決まり手はコース1のみ（差される側は常に1コース）
        if course == 1:
            hi_kimari = {}
            for kname, kidx in COURSE_被KIMARI_COLS.items():
                v = safe_float(row[kidx])
                if v is not None:
                    hi_kimari[kname] = v
            if hi_kimari:
                entry["被kimari"] = hi_kimari

        master[name][c] = entry

    print(f"  {sheet_name}: {len(master)} 選手")

    # ── デバッグログ: 選手タイプ別 reliable 適用状況 ──
    # base_rate の信頼性を後から追跡できるよう集計して出力する
    type_stats = {}
    for name, courses in master.items():
        ptype = _pi.get(name, {}).get("player_type", "不明")
        for c, entry in courses.items():
            key = (ptype, c)
            if key not in type_stats:
                type_stats[key] = {"total": 0, "reliable": 0}
            type_stats[key]["total"] += 1
            if entry["reliable"]:
                type_stats[key]["reliable"] += 1
    grade_main_reliable = sum(
        v["reliable"] for k, v in type_stats.items() if k[0] == "グレードメイン"
    )
    grade_main_total = sum(
        v["total"] for k, v in type_stats.items() if k[0] == "グレードメイン"
    )
    if grade_main_total > 0:
        print(f"  [reliable比率] グレードメイン: {grade_main_reliable}/{grade_main_total} "
              f"({grade_main_reliable/grade_main_total*100:.1f}%) — 閾値=1走で緩和済み")
    return master


def build_venue_course_master(wb, venue_stats: dict) -> dict:
    """
    会場別コースマスタ → venue_course_master[選手名][会場名][コース文字列] = {...}

    格納するフィールド:
        runs              : 出走数
        win_rate          : 1着率（生値）
        ts_win_rate       : 時系列補正1着率
        trust             : 信頼度（スクリプト内で再計算、改修）
        reliable          : runs >= 閾値 かつ trust > 0（改修）
        composite_win_rate: 荒れスコア加味の合成1着率（新規追加）
    """
    ws = wb["会場別コースマスタ"]

    # 会場×コース別の全選手平均勝率を事前計算（trust 算出に使用）
    # まず全行を読み込んでから平均を出す
    raw_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name  = row[0]
        venue = row[1]
        if not name or not venue or not isinstance(name, str):
            continue
        course = safe_int(row[2])
        if course is None or course not in range(1, 7):
            continue
        runs     = safe_int(row[3], 0)
        win_rate = safe_float(row[5], None)
        ts_win   = safe_float(row[16], None)
        raw_rows.append((name, venue, str(course), runs, win_rate, ts_win))

    # 会場×コース別 平均勝率（ベースレート）の算出
    venue_course_avg = {}  # (venue, course) -> list of win_rates
    for _, venue, c, runs, win_rate, _ in raw_rows:
        if win_rate is not None and runs >= VENUE_COURSE_MIN_RUNS:
            key = (venue, c)
            venue_course_avg.setdefault(key, []).append(win_rate)
    venue_course_avg = {
        k: statistics.mean(v) for k, v in venue_course_avg.items() if v
    }

    # 本組み立て
    master = {}
    for name, venue, c, runs, win_rate, ts_win in raw_rows:
        avg_wr = venue_course_avg.get((venue, c))
        trust  = calc_trust(runs, win_rate, avg_wr)  # 改修: スクリプト内で計算

        arek_score = venue_stats.get(venue, {}).get("arek_score", 47.4)

        if name not in master:
            master[name] = {}
        if venue not in master[name]:
            master[name][venue] = {}

        master[name][venue][c] = {
            "runs":               runs,
            "win_rate":           win_rate,
            "ts_win_rate":        ts_win,
            "trust":              trust,
            "reliable":           runs >= VENUE_COURSE_MIN_RUNS and trust > 0,
            # composite_win_rate は player_index の bayesian を使うので後で補完
        }

    player_count = len(master)
    print(f"  venue_course_master: {player_count} 選手")
    return master


def build_venue_stats(wb) -> dict:
    """
    会場統計 → venue_stats[会場名] = {
        arek_score       : 荒れスコア（全体）
        arek_by_race     : {R番号: score}          R別荒れスコア
        course_rates     : {コース: 1着率}          会場×コース全体
        race_course_rates: {R番号: {コース: 1着率}} R×コース別1着率（改修: 積極活用推奨）
        inn_rate         : インコース率
    }
    """
    ws = wb["会場統計"]

    headers = []
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        headers = list(row)
        break

    # R別荒れ列: "1R\n荒れ" 〜 "12R\n荒れ"
    arek_cols = {}
    for rno in range(1, 13):
        for i, h in enumerate(headers):
            if h and f"{rno}R\n荒れ" == str(h):
                arek_cols[rno] = i
                break

    # R×コース別1着率列: "1C\n1R" 〜 "6C\n12R"
    rc_cols = {}
    for c in range(1, 7):
        for r in range(1, 13):
            label = f"{c}C\n{r}R"
            for i, h in enumerate(headers):
                if h and str(h) == label:
                    rc_cols[(c, r)] = i
                    break

    # コース別1着率列: "1C\n1着率" など
    course_rate_cols = {}
    for c in range(1, 7):
        label = f"{c}C\n1着率"
        for i, h in enumerate(headers):
            if h and str(h) == label:
                course_rate_cols[c] = i
                break

    # arek_score のデフォルト値を先行計算するため2パス
    arek_all = []
    raw_venue_rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        venue = row[0]
        if not venue or not isinstance(venue, str):
            continue
        a = safe_float(row[8])
        if a is not None:
            arek_all.append(a)
        raw_venue_rows.append((venue, row))

    # 改修: デフォルト荒れスコアを全会場の実平均から算出
    arek_default = round(statistics.mean(arek_all), 1) if arek_all else 47.4

    stats = {}
    for venue, row in raw_venue_rows:
        arek_score = safe_float(row[8], arek_default)

        arek_by_race = {}
        for rno, ci in arek_cols.items():
            v = safe_float(row[ci])
            if v is not None:
                arek_by_race[rno] = v

        course_rates = {}
        for c, ci in course_rate_cols.items():
            v = safe_float(row[ci])
            if v is not None:
                course_rates[str(c)] = v

        race_course_rates = {}
        for (c, r), ci in rc_cols.items():
            v = safe_float(row[ci])
            if v is not None:
                rkey = str(r)
                if rkey not in race_course_rates:
                    race_course_rates[rkey] = {}
                race_course_rates[rkey][str(c)] = v

        stats[venue] = {
            "arek_score":        arek_score,
            "arek_by_race":      arek_by_race,
            "course_rates":      course_rates,
            "race_course_rates": race_course_rates,
            "inn_rate":          safe_float(row[2], 0.5),
        }

    # ── イン逃げ分析シートから inn_2place を注入 ─────────────────────────
    # シート構成: 行1=タイトル 行2=ヘッダ 行3〜=データ
    # 列構成: [0]会場名 [1]イン逃げ回数 [2〜7]1〜6枠2着率 [8〜13]1〜6枠3着以内率
    try:
        ws_inn = wb["イン逃げ分析"]
        for row in ws_inn.iter_rows(min_row=3, values_only=True):
            venue = row[0]
            if not venue or not isinstance(venue, str):
                continue
            if venue not in stats:
                continue
            inn_2place = {}
            for c in range(1, 7):
                v = safe_float(row[1 + c])  # index 2〜7
                if v is not None:
                    inn_2place[str(c)] = v
            if inn_2place:
                stats[venue]["inn_2place"] = inn_2place
        print(f"  inn_2place: {sum(1 for v in stats.values() if 'inn_2place' in v)} 会場に注入")
    except Exception as e:
        print(f"  ⚠ inn_2place 注入スキップ: {e}")

    print(f"  venue_stats: {len(stats)} 会場  arek_default={arek_default}")
    return stats, arek_default


def build_player_index(wb, venue_stats: dict) -> tuple:
    """
    選手指数マスタ → (player_index, player_id_map)

    player_index[選手名] = {
        reg_no            : 登録番号
        st_rank           : {コース: ST順位}
        form_index        : フォーム指数（生値）
        form_index_norm   : フォーム指数（パーセンタイル正規化 0〜1）（新規追加）
        recent10_win      : 直近10走1着率（生値）
        recent5_place3    : 直近5走3連対率
        overall_win       : 1着率（一般戦）
        bayesian_win      : ベイズ補正後1着率（新規追加）
        fly_days          : FLY経過日数
        fly_impact        : FLY影響度
        fly_after_runs    : FLY後走数
    }
    """
    ws = wb["選手指数マスタ"]
    index  = {}
    id_map = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        reg_no = str(row[0]).strip() if row[0] else None
        name   = row[1]
        if not name or not isinstance(name, str):
            continue

        st_rank = {}
        for c in range(1, 7):
            v = safe_float(row[20 + (c - 1)])
            if v is not None:
                st_rank[str(c)] = v

        fly_days     = safe_float(row[37])
        overall_win  = safe_float(row[5])
        recent10_win = safe_float(row[31])

        index[name] = {
            "reg_no":          reg_no,
            "st_rank":         st_rank,
            "form_index":      safe_float(row[26]),
            "form_index_norm":  None,  # 後で全選手分を正規化
            "recent10_win":    recent10_win,
            "recent5_place3":  safe_float(row[30]),
            "overall_win":     overall_win,
            "bayesian_win":    bayesian_win_rate(recent10_win, overall_win),
            "fly_days":        fly_days,
            "fly_impact":      safe_float(row[41]),
            "fly_after_runs":  safe_int(row[42]),
            # 【追加】選手タイプ: update_master.py が付与する分類
            # "グレードメイン" / "混合" / "一般メイン" / "不明"
            # reliable判定・dq='insufficient'の除外に使用
            "player_type":    str(row[44]).strip() if row[44] else "不明",
            # 【追加】直近1年・一般戦3連対率（row[45]: 0始まりindex）
            "annual_place3":  safe_float(row[45]),
        }

        if reg_no:
            id_map[reg_no] = name

    # ── form_index のパーセンタイル正規化（改修）──────────────────
    fi_values = sorted(
        [v["form_index"] for v in index.values() if v["form_index"] is not None]
    )
    fi_n = len(fi_values)
    if fi_n > 0:
        for name, v in index.items():
            fi = v["form_index"]
            if fi is None:
                v["form_index_norm"] = None
                continue
            # パーセンタイルランク: fi より小さい値の割合
            rank = sum(1 for x in fi_values if x < fi)
            v["form_index_norm"] = round(rank / fi_n, 4)

    print(f"  player_index: {len(index)} 選手 / player_id_map: {len(id_map)} 件")
    return index, id_map


def enrich_venue_course_master(vcm: dict, player_index: dict, venue_stats: dict) -> dict:
    """
    venue_course_master に composite_win_rate を付与（改修: 新規追加）

    player_index の bayesian_win と会場の arek_score を組み合わせて
    会場特性を加味した合成1着率を算出する。
    """
    enriched = 0
    for name, venues in vcm.items():
        pi = player_index.get(name, {})
        bayesian = pi.get("bayesian_win")
        for venue, courses in venues.items():
            arek_score = venue_stats.get(venue, {}).get("arek_score", 47.4)
            for c, v in courses.items():
                runs    = v["runs"]
                overall = v["win_rate"]
                # 会場別 overall を優先、なければ全国 bayesian で代用
                base    = overall if (overall is not None and v["reliable"]) else bayesian
                cwr = calc_composite_win_rate(base, bayesian, arek_score, runs)
                v["composite_win_rate"] = cwr
                if cwr is not None:
                    enriched += 1
    print(f"  composite_win_rate 付与: {enriched} 件")
    return vcm


# ──────────────────────────────────────────────────────────────────
# 決まり手・展開推定 関連
# ──────────────────────────────────────────────────────────────────

# 会場統計シートの決まり手列インデックス
VENUE_KIMARI_COLS = {
    "逃げ":       3,
    "差し":       4,
    "まくり":     5,
    "まくり差し": 6,
    "抜き":       7,
    # 恵まれは転覆等による繰り上がりのため予測不可 → 除外
}

# コース別マスタシートの決まり手%列インデックス
COURSE_KIMARI_COLS = {
    "逃げ":       13,
    "差し":       14,
    "まくり":     15,
    "まくり差し": 16,
    "抜き":       17,
    # 恵まれ（18列）は転覆等による繰り上がりのため予測不可 → 除外
}

# 1コース選手の被決まり手%（コース1のみ集計）
# 2〜6コースの被決まり手は予測に意味がないため除外
COURSE_被KIMARI_COLS = {
    "差され":       35,
    "捲られ":       36,
    "捲り差され":   37,
}

# 物理的に絶対ありえない組み合わせ（ハード除外）
# 恵まれは転覆等による繰り上がりのため予測不可 → 使用しない
KIMARI_HARD_EXCLUDE = {
    "逃げ":       {"2","3","4","5","6"},  # 逃げで勝てるのは1コースのみ
    "差し":       {"1"},                  # 1コースは差される側
    "まくり":     {"1"},                  # 1コースはまくる相手がいない
    "まくり差し": {"1","2"},              # 助走距離が必要なため1・2コースは不可
    "抜き":       set(),                  # 全コース理論上あり得る
}

# グレーゾーン：個人kimari%が閾値以上なら有効とみなす
KIMARI_SOFT_THRESHOLD = {
    "まくり": {"2": 0.05},   # 2コースのまくりは実績5%以上で有効
    "抜き":   {"1": 0.03},   # 1コースの抜きは実績3%以上で有効
}




def build_venue_kimari(wb) -> dict:
    """
    会場統計 → venue_kimari[会場名] = {
        "逃げ": 0.48, "差し": 0.13, "まくり": 0.15,
        "まくり差し": 0.17, "抜き": 0.05
    }
    会場全体の決まり手比率。1着推定の事前確率として使用。
    """
    ws = wb["会場統計"]
    result = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        venue = row[0]
        if not venue or not isinstance(venue, str):
            continue
        rates = {}
        for kimari, col in VENUE_KIMARI_COLS.items():
            v = safe_float(row[col])
            if v is not None:
                rates[kimari] = v
        if rates:
            result[venue] = rates
    print(f"  venue_kimari: {len(result)} 会場")
    return result


# ──────────────────────────────────────────────────────────────────
# 展開別残存マスタ（決まり手 × 1着コース → 各コースの2着率・3着以内率）
# ──────────────────────────────────────────────────────────────────

def build_tenkai_remaining(wb) -> dict:
    """
    「展開別残存_会場別」シートから読み込む。
    tenkai_remaining[決まり手][1着コース][進入コース] = {
        "rate2": 2着率,
        "rate3": 3着以内率,
        "count": レース数,
        "trust": 信頼度
    }

    会場別データを全会場合算（レース数加重平均）して全国値として使用。
    展開シナリオの2着確率計算に使用する。

    列構造（1行目ヘッダー、2行目からデータ）:
        A: (空)  B: 会場名  C: 決まり手  D: 1着コース
        E: レース数  F: 信頼度  G: 進入コース  H: 2着率  I: 3着率  J: 3着以内率
    """
    SHEET = "展開別残存_会場別"
    if SHEET not in wb.sheetnames:
        print(f"  ⚠ シート'{SHEET}'が見つかりません → tenkai_remaining はスキップ")
        return {}

    ws = wb[SHEET]

    # 会場別データを収集してレース数加重平均で合算
    # accum[決まり手][1着コース][進入コース] = {"sum_rate2*count": ..., "sum_count": ..., "sum_trust*count": ...}
    accum = {}
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # A列は空、B列から始まる
        venue  = row[1]   # B: 会場名
        kimari = row[2]   # C: 決まり手
        ichi_c = row[3]   # D: 1着コース
        count  = safe_int(row[4])         # E: レース数
        trust  = safe_float(row[5])       # F: 信頼度
        zan_c  = row[6]                   # G: 進入コース
        rate2  = safe_float(row[7])       # H: 2着率（小数 or %）
        rate3  = safe_float(row[8])       # I: 3着率
        rate3i = safe_float(row[9])       # J: 3着以内率

        if not kimari or ichi_c is None or zan_c is None:
            skipped += 1
            continue
        if count is None or count <= 0:
            skipped += 1
            continue

        # % 表記（例: 8.5）を小数（0.085）に変換
        if rate2 is not None and rate2 > 1.0:
            rate2 = rate2 / 100.0
        if rate3  is not None and rate3  > 1.0:
            rate3  = rate3  / 100.0
        if rate3i is not None and rate3i > 1.0:
            rate3i = rate3i / 100.0

        kimari_s = str(kimari).strip()
        ichi_s   = str(int(ichi_c)) if isinstance(ichi_c, (int, float)) else str(ichi_c).strip()
        zan_s    = str(int(zan_c))  if isinstance(zan_c,  (int, float)) else str(zan_c).strip()

        key = (kimari_s, ichi_s, zan_s)
        if key not in accum:
            accum[key] = {"sum_r2": 0.0, "sum_r3": 0.0, "sum_trust": 0.0, "total_count": 0}

        r2 = rate2 if rate2 is not None else 0.0
        r3 = rate3 if rate3 is not None else 0.0  # I列: 単独3着率
        tr = trust if trust is not None else 0.0

        accum[key]["sum_r2"]      += r2 * count
        accum[key]["sum_r3"]      += r3 * count
        accum[key]["sum_trust"]   += tr * count
        accum[key]["total_count"] += count

    # 加重平均してresultに格納
    result = {}
    for (kimari_s, ichi_s, zan_s), a in accum.items():
        n = a["total_count"]
        if n <= 0:
            continue
        result.setdefault(kimari_s, {}).setdefault(ichi_s, {})[zan_s] = {
            "rate2": round(a["sum_r2"]    / n, 4),
            "rate3": round(a["sum_r3"] / n, 4),  # 単独3着率（I列）
            "count": n,
            "trust": round(a["sum_trust"] / n, 4),
        }

    total = sum(
        len(zans)
        for kimaris in result.values()
        for zans in kimaris.values()
    )
    print(f"  tenkai_remaining: {len(result)}決まり手 / {total}エントリ (スキップ: {skipped}行)")
    return result


# ──────────────────────────────────────────────────────────────────
# 勝者コース別着順分析（選手 × 自コース × 勝者コース → 2着率・3着以内率）
# ──────────────────────────────────────────────────────────────────

def build_winner_course_order(wb) -> dict:
    """
    「勝者コース別着順分析」シートから読み込む。
    winner_course_order[選手名][自コース][勝者コース] = {
        "rate2": 2着率,
        "rate3": 3着率,
        "rate3i": 3着以内率,
        "avg_rank": 平均着順,
        "count": レース数,
        "trust": 信頼度
    }

    「この選手がXコースに乗り、Yコースが1着の展開」での
    実績2着率・3着率を展開シナリオの2着・3着確率補正に使用する。
    個人実績があれば tenkai_remaining（全国平均）より優先する。

    【trust の算出方針】
    Excel列の trust は生成ロジックが不透明なためスクリプト内で上書き計算する。
    count（レース数）をベースに以下の区分で信頼度を設定：

        count <  10 : trust = 0.0          （サンプル不足、個人補正を使わない）
        count   10  : trust = 0.30          （下限: 全国実績70%:個人30%）
        count   30  : trust = 0.60          （中程度: 全国実績40%:個人60%）
        count >= 60 : trust = 1.0           （十分なサンプル、個人実績を全面採用）
        10〜60 の間 : 対数補間で滑らかに増加

    JS側のブレンド式:  final = personRate * trust + baseRate * (1 - trust)
    JS側の採用閾値:    trust > 0.3（10レース未満を確実に除外）
    """
    # trust 算出定数 ── 勝者コース別に設定
    #
    # 1号艇勝利時: データが豊富（平均22.4走）→ 従来に近い基準を維持
    #   MIN=8  : 8走未満は除外
    #   FULL=40 : 40走でフル反映（従来60から引き下げ）
    #
    # 2〜6号艇勝利時: データが少ない（平均5.7走）→ 基準を大幅に引き下げ
    #   MIN=3  : 3走以上で採用開始
    #   FULL=15 : 15走でフル反映
    #
    WCO_TRUST_1_MIN_COUNT   = 8    # 1号艇勝利: これ未満は trust=0.0
    WCO_TRUST_1_FULL_COUNT  = 40   # 1号艇勝利: これ以上は trust=1.0
    WCO_TRUST_26_MIN_COUNT  = 3    # 2〜6号艇勝利: これ未満は trust=0.0
    WCO_TRUST_26_FULL_COUNT = 15   # 2〜6号艇勝利: これ以上は trust=1.0
    WCO_TRUST_AT_MIN        = 0.30 # MIN到達時の trust 下限値（共通）
    WCO_TRUST_AT_FULL       = 1.00 # FULL到達時の trust 上限値（共通）

    import math

    def calc_wco_trust(count: int, is_winner_1: bool) -> float:
        """count と勝者コース区分から trust を対数補間で算出する。"""
        min_c  = WCO_TRUST_1_MIN_COUNT  if is_winner_1 else WCO_TRUST_26_MIN_COUNT
        full_c = WCO_TRUST_1_FULL_COUNT if is_winner_1 else WCO_TRUST_26_FULL_COUNT
        if count is None or count < min_c:
            return 0.0
        if count >= full_c:
            return WCO_TRUST_AT_FULL
        # log(count/MIN) / log(FULL/MIN) で 0〜1 に正規化し、trust 範囲にスケール
        t = math.log(count / min_c) / math.log(full_c / min_c)
        return round(WCO_TRUST_AT_MIN + t * (WCO_TRUST_AT_FULL - WCO_TRUST_AT_MIN), 4)

    SHEET = "勝者コース別着順分析"
    if SHEET not in wb.sheetnames:
        print(f"  ⚠ シート'{SHEET}'が見つかりません → winner_course_order はスキップ")
        return {}

    ws = wb[SHEET]
    result = {}
    skipped = 0

    # ヘッダー行を確認（行3がヘッダー）
    for row in ws.iter_rows(min_row=4, values_only=True):
        name     = row[0]  # 選手名
        jiko_c   = row[1]  # 自コース（進入コース）
        win_c    = row[2]  # 勝者コース
        count    = safe_int(row[3])     # レース数
        avg_rank = safe_float(row[4])   # 平均着順
        rate2    = safe_float(row[5])   # 2着率
        rate3    = safe_float(row[6])   # 3着率（新規追加）
        rate3i   = safe_float(row[7])   # 3着以内率
        # row[8] の Excel trust は使用せず count から再計算する
        # 勝者コースが1かどうかで信頼度基準を切り替える（データ量が大きく異なるため）
        is_winner_1 = (win_c == 1 or win_c == '1')
        trust    = calc_wco_trust(count, is_winner_1)

        if not name or not isinstance(name, str) or jiko_c is None or win_c is None:
            skipped += 1
            continue

        # % 表記（例: 28.6）を小数（0.286）に変換
        if rate2  is not None and rate2  > 1.0: rate2  = rate2  / 100.0
        if rate3  is not None and rate3  > 1.0: rate3  = rate3  / 100.0
        if rate3i is not None and rate3i > 1.0: rate3i = rate3i / 100.0

        jiko_s = str(int(jiko_c)) if isinstance(jiko_c, (int, float)) else str(jiko_c)
        win_s  = str(int(win_c))  if isinstance(win_c,  (int, float)) else str(win_c)

        result.setdefault(name, {}).setdefault(jiko_s, {})[win_s] = {
            "rate2":    rate2,
            "rate3":    rate3,
            "rate3i":   rate3i,
            "avg_rank": avg_rank,
            "count":    count,
            "trust":    trust,
        }

    total = sum(
        len(wins)
        for jikos in result.values()
        for wins in jikos.values()
    )
    print(f"  winner_course_order: {len(result)}選手 / {total}エントリ (スキップ: {skipped}行)")
    return result







# ──────────────────────────────────────────────────────────────────
# エントリポイント
# ──────────────────────────────────────────────────────────────────

def main():
    # 引数が指定されていない場合は、デフォルトパスで実行できるようにする
    # （bat/ダブルクリック起動でも動くようにするため）
    if len(sys.argv) < 3:
        default_xlsx = Path(r"C:\Users\user\Desktop\データ収集\ボートリサーチ_マスタ.xlsx")
        default_out  = Path(__file__).resolve().parent / "master_data.json"

        print("引数が指定されていないためデフォルトパスで実行します。")
        print(f"  xlsx : {default_xlsx}")
        print(f"  out  : {default_out}")
        print("使い方: python build_master_json.py <xlsx_path> <output_json_path>")
        xlsx_path = default_xlsx
        out_path  = default_out
    else:
        xlsx_path = Path(sys.argv[1])
        out_path  = Path(sys.argv[2])

    if not xlsx_path.exists():
        print(f"エラー: {xlsx_path} が見つかりません")
        sys.exit(1)

    print(f"読み込み中: {xlsx_path}")
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)

    print("ビルド開始...")

    # venue_stats を最初にビルド（他の関数が arek_score を参照するため）
    venue_stats, arek_default = build_venue_stats(wb)

    # 【修正】player_index を先にビルドして player_type を course_master の
    # reliable判定に渡す（グレードメイン選手の閾値緩和のため）
    player_index, player_id_map = build_player_index(wb, venue_stats)
    course_master               = build_course_master(wb, player_index)
    venue_course_master         = build_venue_course_master(wb, venue_stats)

    # 女子戦コースマスタ（シートが存在する場合のみ）
    # is_joshi=True により reliable 閾値を8走に緩和（女子は年間出走数が構造的に少ないため）
    course_master_joshi = build_course_master(
        wb, player_index, sheet_name="📊コース別マスタ_女子", is_joshi=True
    )

    # composite_win_rate を後付けで付与
    venue_course_master = enrich_venue_course_master(
        venue_course_master, player_index, venue_stats
    )

    # ── 展開推定マスタ ────────────────────────────────────────────
    venue_kimari        = build_venue_kimari(wb)

    # ── 展開別残存・勝者コース別着順分析 ──────────────────────────
    tenkai_remaining    = build_tenkai_remaining(wb)
    winner_course_order = build_winner_course_order(wb)

    # ── メタ情報 ─────────────────────────────────────────────────
    arek_scores = [v["arek_score"] for v in venue_stats.values()]
    meta = {
        "version": "3.3",
        "built_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "parameters": {
            "venue_course_min_runs":    VENUE_COURSE_MIN_RUNS,
            "course_min_runs":          COURSE_MIN_RUNS,
            "bayesian_prior_weight":    BAYESIAN_PRIOR_WEIGHT,
            "recent10_n":               RECENT10_N,
            "st_avg_by_course":         ST_AVG,
            "st_sensitivity_by_course": ST_SENSITIVITY,
        },
        "arek_score_stats": {
            "default": arek_default,
            "mean":    round(statistics.mean(arek_scores), 2),
            "median":  round(statistics.median(arek_scores), 2),
            "min":     min(arek_scores),
            "max":     max(arek_scores),
        },
    }

    data = {
        "meta":                  meta,
        # 後方互換のためトップレベルにも残す
        "built_at":              meta["built_at"],
        "venue_course_min_runs": VENUE_COURSE_MIN_RUNS,
        "course_master_min_runs": COURSE_MIN_RUNS["1"],
        "course_master":         course_master,
        "course_master_joshi":   course_master_joshi,  # 女子戦コースマスタ（空の場合は{}）
        "venue_course_master":   venue_course_master,
        "venue_stats":           venue_stats,
        "player_index":          player_index,
        "player_id_map":         player_id_map,
        "venue_kimari":           venue_kimari,
        "tenkai_remaining":       tenkai_remaining,
        "winner_course_order":    winner_course_order,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = out_path.stat().st_size / 1024
    print(f"完了: {out_path} ({size_kb:.0f} KB)")

    # player_id_map.json を同フォルダに別途出力
    id_map_path = out_path.parent / "player_id_map.json"
    with open(id_map_path, "w", encoding="utf-8") as f:
        json.dump(player_id_map, f, ensure_ascii=False, separators=(",", ":"))
    print(f"player_id_map.json: {id_map_path} ({id_map_path.stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    main()
