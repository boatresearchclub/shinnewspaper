# -*- coding: utf-8 -*-
"""
lr_probs.py  ─  3連単確率計算 / EV計算
分割元: load_race.py
※ _calc_3rentan_probs（旧版）は削除済み。_calc_3rentan_probs_v2 を使用すること。
"""
import re, sys, itertools
import pandas as pd
from openpyxl import load_workbook
from lr_utils import safe_float, _get_cm_val
from lr_suggest import _calc_attack_effectiveness

# ──────────────────────────────────────────────────────────────────────
# 【削除済み】_calc_3rentan_probs (旧v1) → _calc_3rentan_probs_v2 に移行済み
# ──────────────────────────────────────────────────────────────────────

def _calc_sc_weight(results, cm_map, win3_map, rel_map, jizen_eval=None):
    """
    SCシナリオ（飛び役自滅→漁夫の利）の重みと、
    SC発動時の受益者スコアを計算して返す。

    【競艇物理法則に基づく設計】

    ■ 飛び役の決まり手タイプを「最大決まり手（1位）」で分類
        まくり系   : 最大決まり手が「まくり」または「まくり差し」
        差し系     : 最大決まり手が「差し」
        逃げ系(1号): SCシナリオ対象外（逃げが自滅しても外がそのまま1着になる）

    ■ まくり系飛び役が自滅した場合（外に膨らむ）
        物理法則: まくり艇が旋回で外に膨らむ
                  → コース(飛び役)より内側の艇のコースが開く（内側開放）
                  → コース(飛び役)より外側の艇が続いて捲る（外側継続）
        受益者:
          内側開放候補: コース < 飛び役コース の艇（1号艇除く）
                        ただし1号艇はS1逃げで既に前提計算済みなので2号艇以降
          外側継続候補: コース > 飛び役コース の艇
        重み付け:
          内側開放ボーナス: ×1.5（コースが開く = 進入路が確保される）
          外側継続ボーナス: ×0.8（勢いはあるが距離損）

    ■ 差し系飛び役が自滅した場合（蓋をされる）
        物理法則: 差し艇が1号艇に蓋をされて失速
                  → 後方から来た外側艇（まくり差し）が浮上する
        受益者: コース > 飛び役コース の艇（まくり差しで来た艇）
        重み付け:
          後方浮上ボーナス: ×1.2

    ■ 漁夫スコア計算
        漁夫スコア = 位置ボーナス
                   × (1 - 攻撃性正規化) × 0.5 + 0.5  ← 仕掛けに行かない艇を優遇
                   × win3_rate                          ← 地力（荒れても残る力）

        攻撃性 = (まくり% + まくり差し% + 差し%) / 100
                 高い → 自分も仕掛けに行く → 自滅リスクあり → 漁夫スコア低め
                 低い → 控えて後ろから残る → 漁夫スコア高め

    ■ SCシナリオの1着
        飛び役が自滅 → 1着を取り直すのは:
          パターンA: 1号艇（逃げ取り戻し） ← まくり系飛び役が自滅した場合に多い
          パターンB: 第2飛び役（sub_fly）   ← 差し系飛び役が自滅した場合
          → 両方をSCシナリオ内で重み付き按分

    ■ SCシナリオ重み（S4の代替）
        SC_base = S4_old の重みをベースに、以下で調整:
          飛び役の攻撃性が高い（自滅リスク大）→ SC重みを増やす
          飛び役の実績（win3_rate）が低い      → SC重みを増やす
        SC重み = SC_base × (1 + 飛び役攻撃性 × 0.5) × (1 + (1 - 飛び役win3) × 0.3)

    Returns
    -------
    dict:
        sc_weight       : SCシナリオの全体重み（p_s4の代替として使用）
        sc_1st_weights  : {waku: float} SC時の1着按分重み
        sc_beneficiary  : {waku: float} SC時の2・3着漁夫スコア
        sc_fly_type     : "まくり系" / "差し系" / "不明"
        sc_fly_waku     : 主要飛び役の艇番
    """
    wakus = [r["waku"] for r in results]
    if len(wakus) < 4:
        return {
            "sc_weight": 0.02,
            "sc_1st_weights": {"1": 1.0},
            "sc_beneficiary": {w: 0.1 for w in wakus},
            "sc_fly_type": "不明",
            "sc_fly_waku": None,
        }

    COURSE_NATIONAL_WIN = {"1": 0.555, "2": 0.137, "3": 0.134,
                           "4": 0.111, "5": 0.066, "6": 0.021}

    def safe_pct(cm, key):
        v = cm.get(key)
        try:
            return max(float(v), 0.0) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    # ── 飛び役の特定と決まり手タイプ分類 ──────────────────────────────────────
    # 1号艇以外で最も rel_win1 が高い艇を主要飛び役とする
    fly_candidates = sorted(
        [(r["waku"], rel_map.get(r["waku"], 0)) for r in results if r["waku"] != "1"],
        key=lambda x: x[1], reverse=True
    )
    main_fly_waku = fly_candidates[0][0] if fly_candidates else None
    sub_fly_waku  = fly_candidates[1][0] if len(fly_candidates) >= 2 else None

    if main_fly_waku is None:
        return {
            "sc_weight": 0.02,
            "sc_1st_weights": {"1": 1.0},
            "sc_beneficiary": {w: 0.1 for w in wakus},
            "sc_fly_type": "不明",
            "sc_fly_waku": None,
        }

    # 決まり手タイプ判定（最大決まり手1位で分類）
    fly_cm = cm_map.get(main_fly_waku, {})
    mak_pct     = safe_pct(fly_cm, "まくり%") + safe_pct(fly_cm, "まくり差し%")
    sashi_pct   = safe_pct(fly_cm, "差し%")
    nige_pct    = safe_pct(fly_cm, "逃げ%")

    scores = {"まくり系": mak_pct, "差し系": sashi_pct, "逃げ系": nige_pct}
    sc_fly_type = max(scores, key=scores.get)
    if scores[sc_fly_type] == 0:
        sc_fly_type = "まくり系"  # データなしはまくり系で代替

    # 攻撃性スコア（高いほど仕掛けに行く→自滅リスクあり）
    fly_attack = min((mak_pct + sashi_pct) / 100.0, 1.0)
    fly_win3   = win3_map.get(main_fly_waku, 0.5)

    # ── SCシナリオ重み ──────────────────────────────────────────────────────
    # 旧S4相当のベース（5-6号艇の全国平均微小値）を飛び役の特性で増幅
    sc_base = (
        COURSE_NATIONAL_WIN.get("5", 0.066) * 0.5 +
        COURSE_NATIONAL_WIN.get("6", 0.021) * 0.5
    )
    sc_weight = sc_base * (1.0 + fly_attack * 0.5) * (1.0 + (1.0 - fly_win3) * 0.3)

    # ── SC発動時の1着按分 ──────────────────────────────────────────────────
    # まくり系自滅 → 1号艇が取り戻す確率が高い（コースが開く）
    # 差し系自滅   → 第2飛び役が続く確率が高い
    if sc_fly_type == "まくり系":
        w1_1st = 0.65   # 1号艇が逃げ取り戻し
        w2_1st = 0.35   # 第2飛び役
    elif sc_fly_type == "差し系":
        w1_1st = 0.30   # 差し自滅では1号艇も蓋の影響を受ける
        w2_1st = 0.70   # 第2飛び役が浮上
    else:
        w1_1st = 0.50
        w2_1st = 0.50

    sc_1st_weights = {}
    sc_1st_weights["1"] = w1_1st
    if sub_fly_waku:
        sc_1st_weights[sub_fly_waku] = w2_1st
    else:
        sc_1st_weights["1"] = 1.0  # 第2飛び役なければ1号艇に集約

    # ── 漁夫スコア計算 ──────────────────────────────────────────────────────
    try:
        main_fly_course = int(main_fly_waku)
    except (ValueError, TypeError):
        main_fly_course = 3

    sc_beneficiary = {}
    for r in results:
        w = r["waku"]
        if w == main_fly_waku:
            # 自滅した飛び役自身は漁夫スコア0
            sc_beneficiary[w] = 0.0
            continue

        try:
            w_course = int(w)
        except (ValueError, TypeError):
            w_course = 3

        # 位置ボーナス（コース物理法則）
        if sc_fly_type == "まくり系":
            if w == "1":
                # 1号艇はSC時の1着候補なので2・3着漁夫スコアは低め
                pos_bonus = 0.6
            elif w_course < main_fly_course:
                # 内側開放候補（まくり艇が膨らんでコースが開く）
                pos_bonus = 1.5
            else:
                # 外側継続候補（まくり艇に続いて捲る）
                pos_bonus = 0.8
        elif sc_fly_type == "差し系":
            if w_course > main_fly_course:
                # 後方浮上候補（まくり差しで来た艇）
                pos_bonus = 1.2
            elif w == "1":
                # 蓋をした1号艇は先に行っているが差し系なので若干恩恵あり
                pos_bonus = 0.9
            else:
                pos_bonus = 0.7
        else:
            pos_bonus = 1.0

        # 攻撃性の逆数（仕掛けに行かない艇 = 漁夫スコア高め）
        w_cm      = cm_map.get(w, {})
        w_mak     = safe_pct(w_cm, "まくり%") + safe_pct(w_cm, "まくり差し%")
        w_sashi   = safe_pct(w_cm, "差し%")
        w_attack  = min((w_mak + w_sashi) / 100.0, 1.0)
        passivity = (1.0 - w_attack) * 0.5 + 0.5  # 0.5〜1.0 の範囲

        # 地力（荒れた展開でも着に残る力）
        ground    = win3_map.get(w, 0.5)

        sc_beneficiary[w] = pos_bonus * passivity * ground

    return {
        "sc_weight":      sc_weight,
        "sc_1st_weights": sc_1st_weights,
        "sc_beneficiary": sc_beneficiary,
        "sc_fly_type":    sc_fly_type,
        "sc_fly_waku":    main_fly_waku,
    }


def _calc_3rentan_probs_v2(results, venue_course_1c_rate=None, jizen_eval=None, race_judgment=None,
                            tenkai_national=None, tenkai_venue=None, venue_stats=None, st_kimete_master=None):
    """
    【改善(2)】展開シナリオ×条件付き確率モデル

    旧方式の問題：各艇が独立に競争すると仮定 → 展開の依存関係を無視
    　例）1号艇が逃げた時とまくりが決まった時では2着・3着の確率分布が全く異なる

    新方式：4つの展開シナリオをまず確率決定し、
            シナリオ内で条件付き確率（P(2着|シナリオ)、P(3着|シナリオ,2着)）を計算する。

    ─────────────────────────────────────────────────────────────
    シナリオ定義：
      S1: イン逃げ   … 1号艇が1着（コース1の逃げ決まり手%で推定）
      S2: 差し       … 内側艇(2-3)が差して1着
      S3: まくり系   … 外側艇(3-5)がまくり/まくり差しで1着
      S4: 大荒れ     … 5-6号艇1着（確率は小さいが高配当に寄与）

    各シナリオの重み：
      S1 ∝ 1号艇の逃げ%（コース別マスタ）× 会場イン逃げ率補正
      S2 ∝ 2-3号艇の差し% × rel_win1
      S3 ∝ 3-5号艇のまくり% × rel_win1
      S4 ∝ 5-6号艇の rel_win1（極小）

    条件付き2着確率（シナリオごとに異なる）：
      S1（イン逃げ時）: circle_pct（イン逃げ時2着率）を使用
      S2/S3/S4 時:      展開別残存マスタの実2着率を優先参照（マスタなしは内側残存補正テーブルで代替）
                        会場別マスタ（信頼度>=0.3）→ 全国マスタ → ハードコードテーブル の優先順位

    条件付き3着確率（シナリオごとに異なる）：
      S2/S3/S4 時:      展開別残存マスタの実3着以内率を優先参照（同上優先順位）

    引数追加（v6）：
      tenkai_national: dict  {(決まり手, 1着コース): row_dict}  全国展開別残存マスタ
      tenkai_venue:    dict  {(会場名, 決まり手, 1着コース): row_dict}  会場別展開別残存マスタ
      venue_stats:     dict  会場統計（差し率/まくり率/まくり差し率をS2〜S4重みへ反映）

    ─────────────────────────────────────────────────────────────
    戻り値: prob降順ソート済み 全120通りの combinationリスト
    """
    wakus = [r["waku"] for r in results]
    if len(wakus) < 3:
        return []

    # ── ルックアップテーブル構築 ──
    # 【改善】rel_win1=0 の艇にコース別フロア確率を適用（0%買い目根絶）
    # Laplace smoothingフロア: calc_race_indices で既に適用済みだが
    # 外部calc(_ext_calc_3rentan)経由や直接呼び出し時にも保護する。
    _RELWIN_FLOOR = {"1": 0.5, "2": 0.15, "3": 0.15, "4": 0.10, "5": 0.06, "6": 0.02}
    rel_map   = {r["waku"]: max(r.get("rel_win1") or 0,
                                _RELWIN_FLOOR.get(r["waku"], 0.02)) for r in results}
    # 【修正(1)】_circ_raw（正規化前絶対スコア）を参照して二重正規化を解消
    # circle_pct は表示用相対%（合計100%）のため確率計算では使用しない。
    circ_map  = {r["waku"]: max(
        r.get("_circ_raw") if r.get("_circ_raw") is not None
        else (r.get("circle_pct") or 0) / 100.0, 0
    ) for r in results}
    idx3_map  = {r["waku"]: max(r.get("idx3")  or r.get("rel_win1") or 0, 0) for r in results}
    win3_map  = {r["waku"]: r.get("win3_rate") or 0.5 for r in results}
    cm_map    = {r["waku"]: r.get("raw_cm", {}) for r in results}
    avg_st_map = {r["waku"]: r.get("avg_st") for r in results}

    def safe_pct(cm, key):
        v = safe_float(_get_cm_val(cm, key))
        return max(v or 0.0, 0.0)

    # ──────────────────────────────────────────────────────────────────────────
    # 【修正(5)】メンバー相互作用を反映したシナリオ重み計算
    # ──────────────────────────────────────────────────────────────────────────
    # 旧問題: 各艇の決まり手%を独立に合算 → 「誰が誰に対してどう動くか」が無視される
    #   例(1)) 2号艇の差し%が高くても、3〜4号艇のまくりが速ければS2は潰される
    #   例(2)) 1号艇のSTが遅くても S1重みは下がらなかった
    #   例(3)) 今節成績・モーター調子が確率計算に全く反映されていなかった
    #
    # 新方式: 以下3つの相互作用補正を追加する。
    #
    #   補正A ── STの相対関係によるS1重み調整
    #     1号艇と2号艇の平均STを比較し、2号艇が速い（先マイ状態）ならS1を減衰させる。
    #     ST差 >= +0.03秒（2号艇が速い）: S1 × 0.75
    #     ST差 <= -0.03秒（1号艇が速い）: S1 × 1.15（上限補正）
    #
    #   補正B ── 内側艇の差し力によるS3重み抑制（まくり潰し確率）
    #     2号艇の差し%が高いほど、外側艇のまくりは潰されやすい。
    #     まくり潰し係数 = 1.0 - 0.5 × (2号艇差し% / 全国平均差し率の2倍)
    #     ただし係数の下限は0.5（完全には潰れない）
    #
    #   補正C ── 今節成績・モーター調子によるシナリオ重み調整
    #     今節成績: "1-1-2"のような文字列を解析し、直近の1着数が多いほど
    #               その艇が関与するシナリオ重みを最大+20%上昇させる
    #     モーター2連率: 全艇平均と比較し、突出した艇はシナリオ関与重みを補正する
    #     ※ 補正Cは各シナリオの基礎重みに乗算するスケーラーとして適用
    # ──────────────────────────────────────────────────────────────────────────
    COURSE_NATIONAL_WIN = {"1": 0.555, "2": 0.137, "3": 0.134,
                           "4": 0.111, "5": 0.066, "6": 0.021}
    # 会場別1コース1着率があれば使用（荒れやすい会場の過信補正）
    # 例: 戸田=0.430、平和島=0.446 → 全国平均0.555より大幅に低い
    _c1_win_base = float(venue_course_1c_rate) if venue_course_1c_rate is not None else COURSE_NATIONAL_WIN["1"]

    # ── 補正C用: 今節成績スコアとモーター調子スコアを艇番ごとに算出 ──────────
    def _kosetsu_score(kosetsu_str):
        """
        今節成績文字列（例: "1-2-3" "1-1-F" "2-1-1"）から1着率相当のスコアを返す。
        形式は「着順-着順-着順...」を想定。1着=1.0、2着=0.5、3着=0.25、それ以外=0。
        データなし or 解析不能の場合は None を返す。
        """
        if not kosetsu_str or kosetsu_str in ("", "None", "nan", "-"):
            return None
        scores = []
        for token in re.split(r"[-・/]", kosetsu_str):
            token = token.strip()
            if token == "1":
                scores.append(1.0)
            elif token == "2":
                scores.append(0.5)
            elif token == "3":
                scores.append(0.25)
            elif token.isdigit():
                scores.append(0.0)
            # F/L/S/K等の失格・欠場は無視（スコアに含めない）
        return sum(scores) / len(scores) if scores else None

    def _kosetsu_course_match_score(kosetsu_str, today_course_str):
        """
        【(2)改善】今節成績のコース別一致度スコア。

        今節の各走が「今日と同じコース」で走ったものかどうかを判定し、
        一致走の着順を2倍重みで評価する。
        CSVに今節コース列（例: "1-2-3" 形式）がある場合のみ有効。
        列がない・データが空の場合は通常の _kosetsu_score と同じ結果を返す。

        例: 今日が3コース、今節成績 "1-3-2"、今節コース "3-4-3"
          → 1走目3C一致:1着(2倍) / 2走目4C不一致:3着(1倍) / 3走目3C一致:2着(2倍)
          → 加重平均でスコアUP

        設計思想:
          3コースで今節2連続まくり差しを決めている選手と
          3コースで2連続6着の選手を今節成績文字列だけでは区別できない問題を解消。
          同コース実績を重視することで「今日のコースへの適性」を正確に評価する。
        """
        if not kosetsu_str or kosetsu_str in ("", "None", "nan", "-"):
            return None

        order_tokens = [t.strip() for t in re.split(r"[-・/]", kosetsu_str)]

        # コース情報なし → 通常スコアにフォールバック
        if not today_course_str or today_course_str in ("", "None", "nan"):
            return _kosetsu_score(kosetsu_str)

        try:
            today_course = int(str(today_course_str).strip())
        except (ValueError, TypeError):
            return _kosetsu_score(kosetsu_str)

        # 今節コース列（CSVに "今節コース" 列があれば使用）
        # 形式例: "3-4-3" = 1走目3C, 2走目4C, 3走目3C
        # 列がなければ全走を一致なしとして通常重みで計算
        kosetsu_course_key = None  # build時に渡される予定（現状はNone）

        scores = []
        weights = []
        rank_score_map = {"1": 1.0, "2": 0.5, "3": 0.25}

        for i, token in enumerate(order_tokens):
            token = token.strip()
            rank_score = rank_score_map.get(token, 0.0 if token.isdigit() else None)
            if rank_score is None:
                continue  # F/L/S/K等はスキップ

            # 今節コース列があれば一致判定（現状は全走1倍で計算、列実装後に2倍に拡張）
            weight = 1.0
            scores.append(rank_score)
            weights.append(weight)

        if not scores:
            return None

        total_weight = sum(weights)
        return sum(s * w for s, w in zip(scores, weights)) / total_weight if total_weight > 0 else None

    kosetsu_score_map = {}
    for r in results:
        today_course = str(r.get("course", r.get("waku", "")))
        sc = _kosetsu_course_match_score(str(r.get("kosetsu", "")), today_course)
        kosetsu_score_map[r["waku"]] = sc  # None の場合は補正しない

    motor2_map = {}
    for r in results:
        v = safe_float(r.get("motor2"))
        motor2_map[r["waku"]] = v  # None の場合は補正しない

    # モーター2連率の全艇平均（Noneを除く）
    valid_motor2 = [v for v in motor2_map.values() if v is not None]
    motor2_mean = sum(valid_motor2) / len(valid_motor2) if valid_motor2 else None

    # ── 【v6.2】全選手指数マップを事前構築 ──────────────────────────────────
    # 選手指数マスタ（pm）から各艇の指数を取得してマップ化
    # ここで構築した指数が _member_scenario_scale に全て反映される
    _form_map   = {}   # フォーム指数（0〜16.69、中央値~=2.1）
    _recent3_map= {}   # 直近3走1着率（0〜1.0）
    _recent5_map= {}   # 直近5走1着率（0〜1.0）
    _st_std_map = {}   # ST標準偏差（小さいほど安定）
    _st_stab_map= {}   # ST安定スコア（大きいほど安定、7〜85）
    _jizai_map  = {}   # 自在性加重1着率（外枠実力）
    _ippan_map  = {}   # 一般戦1着率（一般戦専門の実力）
    _recent10_map={}   # 直近10走平均着順（小さいほど好調）

    for r in results:
        pm_r = r.get("raw_pm") or r.get("pm") or {}
        # フォーム指数
        v = safe_float(pm_r.get("フォーム\n指数") or pm_r.get("フォーム指数"))
        _form_map[r["waku"]] = v

        # 直近3走1着率
        v = safe_float(pm_r.get("直近3走\n1着率") or pm_r.get("直近3走1着率"))
        _recent3_map[r["waku"]] = v

        # 直近5走1着率
        v = safe_float(pm_r.get("直近5走\n1着率") or pm_r.get("直近5走1着率"))
        _recent5_map[r["waku"]] = v

        # ST標準偏差
        v = safe_float(pm_r.get("ST\n標準偏差") or pm_r.get("ST標準偏差"))
        _st_std_map[r["waku"]] = v

        # ST安定スコア
        v = safe_float(pm_r.get("ST安定\nスコア") or pm_r.get("ST安定スコア"))
        _st_stab_map[r["waku"]] = v

        # 自在性加重1着率
        v = safe_float(pm_r.get("自在性\n加重1着率") or pm_r.get("自在性加重1着率"))
        _jizai_map[r["waku"]] = v

        # 一般戦1着率
        v = safe_float(pm_r.get("1着率\n(一般戦)") or pm_r.get("1着率(一般戦)"))
        _ippan_map[r["waku"]] = v

        # 直近10走平均着順
        v = safe_float(pm_r.get("直近10走\n平均着順") or pm_r.get("直近10走平均着順"))
        _recent10_map[r["waku"]] = v

        # ── 【v6.2追加】コース別マスタの最速ST・最遅ST からSTレンジを算出 ──
        # STレンジ（最遅-最速）が小さいほど発艇が安定している
        # raw_cm（コース別マスタ）から取得
        cm_r = r.get("raw_cm") or {}
        _st_max = safe_float(cm_r.get("最遅ST") or cm_r.get("最遅\nST"))
        _st_min = safe_float(cm_r.get("最速ST") or cm_r.get("最速\nST"))
        if _st_max is not None and _st_min is not None:
            r["_st_range"] = round(_st_max - _st_min, 4)
        else:
            r["_st_range"] = None

    # STレンジマップ（0〜0.8程度、小さいほど安定）
    _st_range_map = {r["waku"]: r.get("_st_range") for r in results}
    # STレンジの全艇有効値で平均を計算（比較用）
    _valid_st_ranges = [v for v in _st_range_map.values() if v is not None]
    _st_range_mean = sum(_valid_st_ranges) / len(_valid_st_ranges) if _valid_st_ranges else 0.30

    # ── 【v6.4新設】★STフラグマップ ──────────────────────────────────────────
    # update_master.py: base["★ST"] = base["出走数"] < THRESH_ST (=10)
    # Excelには "★"（文字列）または "" で書き込まれる。
    # ★STが立っている艇はST標準偏差・STレンジの計算サンプルが10未満であり
    # 平均値の信頼性が低いため、ST系の補正をスキップする。
    _star_st_map = {}
    for r in results:
        cm_r = r.get("raw_cm") or {}
        val = cm_r.get("★ST")
        # Trueの場合: Excelから読んだ"★"文字列 or update_master側のbool True
        _star_st_map[r["waku"]] = bool(val and str(val).strip() in ("★", "True", "1"))

    # 各艇のpm参照（_member_scenario_scaleのフォールバック用にresultsからも取得試行）
    for r in results:
        if _form_map.get(r["waku"]) is None:
            # raw_pmにない場合、resultsのpm_*フィールドからも取得を試みる
            v = safe_float(r.get("form_index") or r.get("フォーム指数"))
            if v is not None:
                _form_map[r["waku"]] = v
        if _recent3_map.get(r["waku"]) is None:
            v = safe_float(r.get("recent3_win") or r.get("直近3走1着率"))
            if v is not None:
                _recent3_map[r["waku"]] = v
        if _recent5_map.get(r["waku"]) is None:
            v = safe_float(r.get("recent5_win") or r.get("直近5走1着率"))
            if v is not None:
                _recent5_map[r["waku"]] = v

    def _member_scenario_scale(waku, base_kosetsu_weight=0.18, base_motor_weight=0.08):
        """
        【v6.2 全指数統合版】
        その艇が関与するシナリオへの乗算スケーラーを返す。
        スケーラー = 1.0 ± 補正値（範囲: 0.75 〜 1.25）

        使用指数と重み配分（合計補正上限±0.25）:
          今節成績        (0.12): 今節の着順文字列から算出
          モーター2連率   (0.08): 機力の相対評価
          フォーム指数    (0.05): 直近調子の総合指標（中央値2.1基準）
          直近3走1着率    (0.04): 超短期フォーム（3走で評価が安定しやすい）
          直近5走1着率    (0.03): 短期フォーム（3走と5走の差で上昇/下降を見る）
          ST標準偏差      (0.025): STばらつき小→スタートが安定→シナリオ実現度高
          STレンジ        (0.015): 最速〜最遅のレンジ小→コース内での安定度
          ST安定スコア    (0.02): ST品質の総合評価（平均60基準）
          自在性加重1着率 (0.03): 外枠からの攻め実力（S2〜S4シナリオに特に重要）
          一般戦1着率     (0.02): 格付け補正（特別戦選手の一般戦での実力）
          直近10走平均着順(0.02): 中期トレンド（3.5以下=好調、4.5超=不調）
        """
        scale = 1.0

        # ── (1) 今節成績（既存） ────────────────────────────────────────────
        ks = kosetsu_score_map.get(waku)
        if ks is not None:
            scale += (ks - 0.5) * 2 * base_kosetsu_weight

        # ── (2) モーター2連率（既存） ──────────────────────────────────────
        m2 = motor2_map.get(waku)
        if m2 is not None and motor2_mean is not None and motor2_mean > 0:
            ratio = (m2 - motor2_mean) / motor2_mean
            scale += ratio * base_motor_weight

        # ── (3) フォーム指数 ────────────────────────────────────────────────
        form = _form_map.get(waku)
        if form is not None:
            form_norm = max(-1.0, min(1.0, (form / 3.0) - 1.0))
            scale += form_norm * 0.05

        # ── (4) 直近3走1着率 ───────────────────────────────────────────────
        r3 = _recent3_map.get(waku)
        if r3 is not None:
            scale += max(-1.0, min(1.0, (r3 - 0.17) / 0.17)) * 0.04

        # ── (5) 直近5走1着率 ───────────────────────────────────────────────
        r5 = _recent5_map.get(waku)
        if r5 is not None:
            scale += max(-1.0, min(1.0, (r5 - 0.17) / 0.17)) * 0.03

        # ── (6) ST標準偏差（小さいほど安定）──────────────────────────────
        # ★STフラグ = サンプル10未満 → ST値が不安定なためST系補正をスキップ
        _st_unreliable = _star_st_map.get(waku, False)
        st_std = _st_std_map.get(waku)
        if st_std is not None and not _st_unreliable:
            scale += max(-1.0, min(1.0, (0.071 - st_std) / 0.071)) * 0.025

        # ── (7) STレンジ（最速〜最遅、小さいほど安定）【v6.2追加】────────
        st_range = _st_range_map.get(waku)
        if st_range is not None and _st_range_mean > 0 and not _st_unreliable:
            # レンジが平均より小さいほど正（安定）、大きいほど負（不安定）
            scale += max(-1.0, min(1.0, (_st_range_mean - st_range) / _st_range_mean)) * 0.015

        # ── (8) ST安定スコア ───────────────────────────────────────────────
        st_stab = _st_stab_map.get(waku)
        if st_stab is not None and not _st_unreliable:
            scale += max(-1.0, min(1.0, (st_stab - 60.0) / 30.0)) * 0.02

        # ── (9) 自在性加重1着率 ────────────────────────────────────────────
        jizai = _jizai_map.get(waku)
        if jizai is not None:
            scale += max(-1.0, min(1.0, (jizai - 0.06) / 0.06)) * 0.03

        # ── ⑩ 一般戦1着率 ────────────────────────────────────────────────
        ippan = _ippan_map.get(waku)
        if ippan is not None:
            scale += max(-1.0, min(1.0, (ippan - 0.17) / 0.17)) * 0.02

        # ── ⑪ 直近10走平均着順 ───────────────────────────────────────────
        r10 = _recent10_map.get(waku)
        if r10 is not None:
            scale += max(-1.0, min(1.0, (3.5 - r10) / 1.5)) * 0.02

        return max(0.75, min(1.25, scale))

    # ── 補正A: STの相対関係によるS1重み調整係数 ──────────────────────────────
    st1 = avg_st_map.get("1")
    st2 = avg_st_map.get("2")
    st_adj_s1 = 1.0  # デフォルト: 補正なし
    if st1 is not None and st2 is not None:
        st_diff = st1 - st2  # 正 → 1号艇が遅い、負 → 1号艇が速い
        if st_diff >= 0.03:
            # 2号艇の方が速い（先マイになりやすい）→ S1を減衰
            st_adj_s1 = max(0.75, 1.0 - st_diff * 5.0)
        elif st_diff <= -0.03:
            # 1号艇の方が速い（楽逃げ）→ S1を増強（上限1.20）
            st_adj_s1 = min(1.20, 1.0 + abs(st_diff) * 3.0)

    # ── 【修正(5)】S1基礎重み: rel_win1が既に会場特性を織り込み済みのため二重補正を除去 ──
    # 【旧問題】s1_base = _c1_win_base × (0.5 + w1_nige) としていたが、
    #   rel_win1 の計算時点で既に venue_rate（会場別1コース1着率）が w_venue 分混入している。
    #   そこへさらに _c1_win_base を掛けると「会場特性が二重に効く」構造になっていた。
    #   特に荒れやすい会場（戸田・平和島等）では rel_win1 が既に低めになっているのに
    #   _c1_win_base でさらに下げてしまい、1号艇を過剰に不利評価していた。
    #
    # 【修正(3)】S1重みの二重補正を解消: nige_adj を乗算からブレンド補正に変更。
    #   旧方式: s1 = rel_win1 × (0.5 + nige%) × ST補正 - 逃げ%が rel_win1 に既に含まれている
    #   新方式: s1 = nige_blend × ST補正 × member補正
    #     nige_blend = 個人逃げ% × 0.6 + 会場1コース1着率 × 0.4
    #     → 個人実績と会場特性を直接ブレンドし、rel_win1 による二重乗算を排除。
    #     → 逃げ%データなし時は会場1着率100%にフォールバック。
    w1_nige      = safe_pct(cm_map.get("1", {}), "逃げ%")
    if w1_nige > 0:
        nige_blend = w1_nige * 0.6 + _c1_win_base * 0.4
    else:
        nige_blend = _c1_win_base  # データなし → 会場特性のみ
    s1_weight = nige_blend * st_adj_s1 * _member_scenario_scale("1")

    # ── 【v6.2新設】S1重みをRNo別1C1着率で補正 ──────────────────────────────
    # 同じ会場でもレース番号で1コース1着率が大きく異なる
    # （例: 大村1R=0.494 vs 大村12R=0.747）
    # race_judgmentに格納済みの venue_1c_race_rate を使用
    _venue_1c_race = safe_float((race_judgment or {}).get("venue_1c_race_rate")) if race_judgment else None
    if _venue_1c_race is not None and _c1_win_base and _c1_win_base > 0:
        # RNo別1着率 / 会場平均1着率 で補正係数を算出（0.75〜1.25 にクリップ）
        _race_s1_adj = max(0.75, min(1.25, _venue_1c_race / _c1_win_base))
        s1_weight *= _race_s1_adj

    # ── 【修正(2)】事前評価（jizen_eval）による追加補正 ────────────────────────
    # evaluate_jizen.evaluate_all() の結果（◎○△）をシナリオ重みに反映する。
    # (1)イン逃げ評価: S1の補正（◎→+15%、○→+7%、空白→-10%）
    # (2)相性評価:     S2/S3 の按分を相性◎の艇を優遇
    # 補正幅は ±15%以内に抑えて過剰な振れを防ぐ。
    _jizen_symbol_s1 = ""
    _jizen_aisho = {}  # {waku_str: symbol}
    if jizen_eval is not None:
        try:
            _jizen_symbol_s1 = jizen_eval.get("in_nige", [""])[0]  # 1号艇のイン逃げ評価
            for idx, sym in enumerate(jizen_eval.get("aisho", [])):
                _jizen_aisho[str(idx + 1)] = sym
        except Exception:
            pass

    _s1_jizen_adj = {"◎": 1.15, "◎?": 1.10, "○": 1.07, "△": 1.0, "": 0.90}.get(
        _jizen_symbol_s1, 1.0
    )
    s1_weight *= _s1_jizen_adj

    def _jizen_aisho_scale(waku):
        """相性評価記号をシナリオ重み乗数に変換（S2/S3の1着按分で使用）"""
        sym = _jizen_aisho.get(str(waku), "")
        return {"◎": 1.15, "○": 1.07, "△": 1.0, "": 0.95}.get(sym, 1.0)

    # ── 補正B: 2号艇の差し力によるS3抑制係数（まくり潰し） ─────────────────────
    sashi_pct_2   = safe_pct(cm_map.get("2", {}), "差し%")
    nat_sashi_avg = COURSE_NATIONAL_WIN["2"] * 0.5   # 全国平均の差し率代替値（約0.07）
    if sashi_pct_2 > 0:
        makuri_suppress = max(0.50, 1.0 - 0.5 * (sashi_pct_2 / (nat_sashi_avg * 2 + 1e-6)))
    else:
        makuri_suppress = 1.0  # 差し%データなし → 補正しない

    # ── S2基礎重み: 差し（全艇対象・コース別距離補正付き）─────────────────────
    # 旧: 2・3号艇限定 → 4〜6コースからの差しを完全無視していた問題を解消
    # 新: 全艇の差し%を対象に、コースが外になるほど距離補正で減衰させる
    #   1号艇: 差し%は逃げ側なので除外（差しシナリオで1着は物理的にほぼない）
    #   2号艇: 補正1.0（内側差しは最も決まりやすい）
    #   3号艇: 補正1.0
    #   4号艇: 補正0.7（外側差しは距離が伸びて決まりにくい）
    #   5号艇: 補正0.5
    #   6号艇: 補正0.3
    # ── S2/S3/S4重み: _calc_attack_effectivenessの決まり手別スコアを直接使用 ──
    # 旧: 枠番固定スケール（2枠=差し, 3〜6枠=まくり等）で機械的に按分
    # 新: 各艇の「最有力決まり手スコア」を決まり手別に集計
    #     → 「誰がどの決まり手で主役になるか」が確率モデルに直接反映される
    _r1_w = next((r for r in results if r["waku"] == "1"), None)
    _w1cm_s = _r1_w.get("raw_cm", {}) if _r1_w else {}

    s2_weight = 0.0  # 差し展開の合計重み
    s3_weight = 0.0  # まくり展開の合計重み
    s4_weight = 0.0  # まくり差し展開の合計重み

    for r in results:
        w = r["waku"]
        if w == "1":
            continue
        _eff = _calc_attack_effectiveness(r, _w1cm_s, venue_stats or {}, results, st_kimete_master=st_kimete_master)
        _bd  = _eff.get("breakdown", {})
        # 各決まり手の個別スコアを取得（採用された最大だけでなく全て使う）
        _sc_sashi = _bd.get("差しスコア",      0.0) or 0.0
        _sc_maku  = _bd.get("まくりスコア",    0.0) or 0.0
        _sc_maksa = _bd.get("まくり差しスコア",0.0) or 0.0
        _scale    = _member_scenario_scale(w) * _jizen_aisho_scale(w)
        s2_weight += _sc_sashi * _scale
        s3_weight += _sc_maku  * _scale
        s4_weight += _sc_maksa * _scale

    s3_weight *= makuri_suppress  # 補正B: 2号艇差し力によるまくり抑制
    s4_weight *= makuri_suppress  # まくり差しも同様

    # ── 【v6新設】[黄] 会場統計の決まり手率をS2〜S4重みに反映 ─────────────────────
    # 「この会場はまくりが多い」という傾向をS3/S4重みに乗算補正する。
    # venue_stats が渡されていない場合はスキップ（補正なし）。
    # 全国平均: 差し≒0.155 / まくり≒0.150 / まくり差し≒0.095（boatrace.jp 全国統計）
    _NATIONAL_AVG_SASHI    = 0.155
    _NATIONAL_AVG_MAKURI   = 0.150
    _NATIONAL_AVG_MAKUSA   = 0.095
    if venue_stats:
        _kimari = venue_stats.get("kimari_avg") or {}
        # 差し補正 → S2重み
        _vs_sashi = safe_float(_kimari.get("差し"))
        if _vs_sashi is not None and _NATIONAL_AVG_SASHI > 0:
            _s2_venue_boost = max(0.70, min(1.30, _vs_sashi / _NATIONAL_AVG_SASHI))
            s2_weight *= _s2_venue_boost
        # まくり補正 → S3重み
        _vs_makuri = safe_float(_kimari.get("まくり"))
        if _vs_makuri is not None and _NATIONAL_AVG_MAKURI > 0:
            _s3_venue_boost = max(0.70, min(1.30, _vs_makuri / _NATIONAL_AVG_MAKURI))
            s3_weight *= _s3_venue_boost
        # まくり差し補正 → S4重み
        _vs_makusa = safe_float(_kimari.get("まくり差し"))
        if _vs_makusa is not None and _NATIONAL_AVG_MAKUSA > 0:
            _s4_venue_boost = max(0.70, min(1.30, _vs_makusa / _NATIONAL_AVG_MAKUSA))
            s4_weight *= _s4_venue_boost

    # ── 【v6.2新設】会場コース別1着率（2C〜6C）でS2〜S4シナリオ重みを補強 ──────
    # 決まり手率補正だけでは「どのコースから飛んでくるか」が反映できない。
    # 会場統計の実コース別1着率を使って「この会場では外枠が有力か」を補正する。
    # 全国平均: 2C=0.137 / 3C=0.134 / 4C=0.111 / 5C=0.066 / 6C=0.021
    _NATIONAL_C_WIN = {"2": 0.137, "3": 0.134, "4": 0.111, "5": 0.066, "6": 0.021}
    _venue_cw = (race_judgment or {}).get("venue_course_win_rates") or {} if race_judgment else {}
    if _venue_cw:
        # S2（差し）: 主に2〜3コースが担う → 2C・3Cの会場実績で補正
        _s2_c_adj = 1.0
        for _c in ["2", "3"]:
            _vr = safe_float(_venue_cw.get(_c))
            _nr = _NATIONAL_C_WIN.get(_c, 0.13)
            if _vr is not None and _nr > 0:
                _s2_c_adj *= max(0.80, min(1.20, _vr / _nr)) ** 0.5  # 平方根で緩和
        s2_weight *= _s2_c_adj

        # S3（まくり）: 主に3〜5コースが担う → 3C・4C・5Cの会場実績で補正
        _s3_c_adj = 1.0
        for _c in ["3", "4", "5"]:
            _vr = safe_float(_venue_cw.get(_c))
            _nr = _NATIONAL_C_WIN.get(_c, 0.10)
            if _vr is not None and _nr > 0:
                _s3_c_adj *= max(0.80, min(1.20, _vr / _nr)) ** 0.5
        s3_weight *= _s3_c_adj

        # S4（まくり差し）: 主に3〜6コースが担う → 4C・5C・6Cの会場実績で補正
        _s4_c_adj = 1.0
        for _c in ["4", "5", "6"]:
            _vr = safe_float(_venue_cw.get(_c))
            _nr = _NATIONAL_C_WIN.get(_c, 0.066)
            if _vr is not None and _nr > 0:
                _s4_c_adj *= max(0.80, min(1.20, _vr / _nr)) ** 0.5
        s4_weight *= _s4_c_adj

    # ── (2) main_type による S2〜S4 重みの直結補正（STEP1→確率計算の断絶を解消）──
    # _judge_main_player が算出した main_waku（主役候補）と main_type（差し/まくり/まくり差し）を
    # 対応するシナリオの重みに直接乗算して STEP1 の定性判断を確率に反映する。
    #
    # 【(4)改善】主軸強度（mc_strength）と主役スコア（main_score）の両方が高い場合に
    # 上限を1.3倍→2.0倍に引き上げ。
    # 条件: mc_strength >= 60 かつ _main_score >= 0.55
    #   → 「展開が読めているレース」で確信を持って点数を絞る根拠になる。
    # 条件未達の場合は従来通り最大+30%。
    # mc_strength は _build_conflict_map が算出する主軸対立の強度（0〜100程度）。
    _mp_data    = (race_judgment or {}).get("main_player", {}) or {}
    _main_waku  = _mp_data.get("main_waku")
    _main_type  = _mp_data.get("main_type", "")
    _main_score = float(_mp_data.get("main_score", 0) or 0)
    _mc_strength = float((race_judgment or {}).get("conflict_map", {}).get("main_conflict", {}).get("strength", 0) or 0) \
        if race_judgment else 0.0

    # 条件付き上限引き上げ
    if _mc_strength >= 60 and _main_score >= 0.55:
        # 展開が明確に読めているレース → 最大+50%
        _main_boost = 1.0 + min(_main_score * 0.9, 0.50)
    else:
        # 通常ケース → 最大+30%（従来通り）
        _main_boost = 1.0 + min(_main_score * 0.6, 0.30)

    if _main_waku and _main_waku != "1":
        if _main_type == "差し":
            s2_weight *= _main_boost
        elif _main_type == "まくり":
            s3_weight *= _main_boost
        elif _main_type == "まくり差し":
            s4_weight *= _main_boost

    # ── SCシナリオ重み（S4「大荒れ」を「潰れ展開・漁夫の利」SCに差し替え） ────────
    # 旧S4は「何が起きるかわからない雑な確率」だった。
    # SCは「飛び役が自滅したとき誰が漁夫の利を取るか」という具体的な展開を表す。
    # 確率の合計は変わらず、展開の解像度だけ上がる。
    _sc_info = _calc_sc_weight(
        results, cm_map, win3_map, rel_map, jizen_eval=jizen_eval
    )
    sc_weight         = _sc_info["sc_weight"]
    _sc_1st_weights   = _sc_info["sc_1st_weights"]    # {waku: float} SC時1着按分
    _sc_beneficiary   = _sc_info["sc_beneficiary"]    # {waku: float} 漁夫スコア
    _sc_fly_type      = _sc_info["sc_fly_type"]       # "まくり系"/"差し系"/"不明"

    # ── 【(8)改善】複数シナリオ同時成立の相関補正 ─────────────────────────────
    # 現在S1〜S4+SCは排他的シナリオとして扱っているが、
    # 実際には「3号艇まくりが決まりながら2号艇が差して2着に入る」複合展開が多い。
    #
    # 対処: シナリオ間相関として「S3成立時にS2的2着が起きる確率」を
    # S3重みとS2重みの幾何平均で表現し、S3_S2_CORR重みとして別途管理する。
    # これはシナリオを増やすのではなく「S3の2着分布を差し方向に引っ張る」補正。
    #
    # 相関係数: s2_weight と s3_weight の幾何平均 × 相関強度
    # 条件: 2号艇の差し%が全艇平均より高い場合のみ発動（強い差し屋がいる場合）
    _sashi_pct_2 = safe_pct(cm_map.get("2", {}), "差し%")
    _sashi_avg   = sum(safe_pct(cm_map.get(w, {}), "差し%") or 0 for w in ["2","3","4","5","6"]) / 5
    _s3s2_corr_weight = 0.0
    if _sashi_pct_2 > _sashi_avg * 1.2 and s3_weight > 0 and s2_weight > 0:
        # 2号艇が平均より20%以上差しが強い場合: S3展開でも2号艇2着の可能性を追加
        import math as _math
        _s3s2_corr_weight = _math.sqrt(s3_weight * s2_weight) * 0.25
        # 相関重みをS3から借りる（S3重みを薄める）ことで総重みは変わらない
        s3_weight = max(0.0, s3_weight - _s3s2_corr_weight * 0.5)
        s2_weight = max(0.0, s2_weight - _s3s2_corr_weight * 0.5)

    total_s = s1_weight + s2_weight + s3_weight + s4_weight + sc_weight + _s3s2_corr_weight
    if total_s <= 0:
        total_s = 1.0
    p_s1 = s1_weight / total_s
    p_s2 = s2_weight / total_s
    p_s3 = s3_weight / total_s
    p_s4 = s4_weight / total_s
    p_sc = sc_weight / total_s
    p_s3s2 = _s3s2_corr_weight / total_s  # まくり+差し複合シナリオ確率

    # ── 条件付き1着確率（シナリオ × 1着艇）──
    # S2/S3は決まり手%が主軸
    # SC は _calc_sc_weight が算出した1着按分重みを使用
    def prob_first_given_scenario(scenario, first_w):
        """P(1着=first_w | scenario)"""
        if scenario == "S1":
            # イン逃げシナリオ → 1号艇固定
            return 1.0 if first_w == "1" else 0.0
        elif scenario == "S2":
            # 差しシナリオ → 全艇の差し%で按分（コース別距離補正付き）
            _s2_scale = {"2": 1.0, "3": 1.0, "4": 0.7, "5": 0.5, "6": 0.3}
            candidates = {}
            for w, sc in _s2_scale.items():
                pct = safe_pct(cm_map.get(w, {}), "差し%")
                candidates[w] = (pct if pct > 0 else COURSE_NATIONAL_WIN.get(w, 0.1) * 0.3) * sc
            total = sum(candidates.values()) or 1
            return candidates.get(first_w, 0) / total
        elif scenario == "S3":
            # まくりシナリオ → まくり%のみで按分（まくり差し除外）
            outer_scale = {"3": 1.0, "4": 1.0, "5": 0.8, "6": 0.5}
            candidates = {}
            for w, sc in outer_scale.items():
                pct = safe_pct(cm_map.get(w, {}), "まくり%")   # まくり%のみ
                candidates[w] = (pct if pct > 0 else COURSE_NATIONAL_WIN.get(w, 0.04) * 0.4) * sc
            total = sum(candidates.values()) or 1
            return candidates.get(first_w, 0) / total
        elif scenario == "S4":
            # まくり差しシナリオ → まくり差し%のみで按分
            outer_scale = {"3": 1.0, "4": 1.0, "5": 0.8, "6": 0.5}
            candidates = {}
            for w, sc in outer_scale.items():
                pct = safe_pct(cm_map.get(w, {}), "まくり差し%")   # まくり差し%のみ
                candidates[w] = (pct if pct > 0 else COURSE_NATIONAL_WIN.get(w, 0.04) * 0.3) * sc
            total = sum(candidates.values()) or 1
            return candidates.get(first_w, 0) / total
        else:  # SC（潰れ展開）
            # 飛び役自滅後の1着は _calc_sc_weight が算出した按分重みで決まる
            total = sum(_sc_1st_weights.values()) or 1
            return _sc_1st_weights.get(first_w, 0) / total

    # ── 条件付き2着確率 P(2着=B | 1着=A, scenario) ──
    # ── 【v6新設】展開別残存マスタから残存補正係数を取得するヘルパー ──────────────
    # 戻り値: {コース文字列: {"2着率": float, "3着以内率": float}} または None（データなし）
    #
    # シナリオ→決まり手 対応テーブル（全シナリオ対応・v6.1拡張）
    # S4 で first_w=2 のデータが存在しない（物理的にまくり差し2着はほぼない）場合は
    # まくり/2 のデータをフォールバックとして使用。
    _SCENARIO_TO_KIMETE = {
        "S1": "逃げ",        # イン逃げ: 逃げ/1着コースのデータを使用
        "S2": "差し",
        "S3": "まくり",
        "S4": "まくり差し",
        "SC": "恵まれ",      # 潰れ展開: 恵まれ/抜きのデータを使用（漁夫の利）
    }
    # S4でデータなし時のフォールバック決まり手（まくり差しデータがないコースはまくりで代替）
    _S4_FALLBACK_KIMETE = "まくり"

    def _get_tenkai_rates(scenario, first_w):
        """
        展開別残存マスタから、1着コース=first_w のシナリオにおける
        各進入コースの 2着率・3着以内率を返す。

        戻り値: {進入コース文字列: {"2着率": float, "3着以内率": float}} または None

        キー構造（v6）:
          会場別: (会場名, 決まり手, 1着コース, 進入コース) → row
          全国版: (決まり手, 1着コース, 進入コース)         → row

        シナリオ対応（v6.1拡張）:
          S1 → 決まり手=逃げ, 1着コース=1 固定（イン逃げ）
          S2 → 決まり手=差し
          S3 → 決まり手=まくり
          S4 → 決まり手=まくり差し（データなし時はまくりで代替）
          SC → 決まり手=恵まれ + 抜き をブレンド（潰れ展開の漁夫の利）
        """
        _venue = (race_judgment or {}).get("venue") if race_judgment else None

        # S1: 1着コースは常に1固定（逃げは1コースから）
        if scenario == "S1":
            kimete = "逃げ"
            actual_course = "1"
        else:
            kimete = _SCENARIO_TO_KIMETE.get(scenario)
            if not kimete:
                return None
            r_data = next((r for r in results if r["waku"] == first_w), {})
            actual_course = str(int(float(r_data.get("course") or r_data.get("進入コース") or first_w)))

        def _fetch(kimete_k, course_k, c_str):
            """(決まり手, 1着コース, 進入コース) の行データを会場別→全国の優先順で取得

            【修正(2)】信頼度閾値を0.3→0.15に緩和し、低信頼度の会場別データも
            全国マスタとブレンドして活用する（旧実装: 0.3未満は完全に捨てていた）

            ブレンド方式:
              信頼度 >= 0.50 → 会場別100%（十分な実績あり）
              信頼度 >= 0.15 → 会場別 × 信頼度/0.50 + 全国 × (1 - 信頼度/0.50)
                               例: 信頼度0.30 → 会場60% / 全国40%
                               例: 信頼度0.15 → 会場30% / 全国70%
              信頼度 < 0.15  → 全国マスタのみ（会場データが少なすぎて誤学習リスク大）
            """
            if c_str == course_k:
                return None  # 1着艇自身は除外

            row_v    = None
            trust_v  = 0.0
            if tenkai_venue and _venue:
                key_v = (str(_venue), kimete_k, course_k, c_str)
                _rv = tenkai_venue.get(key_v)
                if _rv:
                    try:
                        trust_v = float(_rv.get("信頼度") or 0)
                        if trust_v >= 0.15:
                            row_v = _rv
                    except (ValueError, TypeError):
                        pass

            row_n = None
            if tenkai_national:
                key_n = (kimete_k, course_k, c_str)
                row_n = tenkai_national.get(key_n)

            # 会場別データなし → 全国マスタのみ返す（従来どおり）
            if row_v is None:
                return row_n

            # 会場別データあり・信頼度 >= 0.50 → 会場別データをそのまま返す
            if trust_v >= 0.50:
                return row_v

            # 会場別データあり・0.15 <= 信頼度 < 0.50 → 全国マスタとブレンド
            if row_n is None:
                # 全国マスタがなければ会場別をそのまま返す
                return row_v
            try:
                w_v = trust_v / 0.50          # 0.30 → 0.60 / 0.15 → 0.30
                w_n = 1.0 - w_v
                r2_blended = (
                    float(row_v.get("2着率")     or 0) * w_v +
                    float(row_n.get("2着率")     or 0) * w_n
                )
                r3_blended = (
                    float(row_v.get("3着以内率") or 0) * w_v +
                    float(row_n.get("3着以内率") or 0) * w_n
                )
                # 元の行を壊さないよう新dictで返す
                return {"2着率": r2_blended, "3着以内率": r3_blended, "_blended": True}
            except (ValueError, TypeError):
                return row_v

        # SC（潰れ展開）: 自滅タイプに応じて参照マスタを切り替える
        # 【(6)改善】
        #   まくり系自滅 → 外に膨らんで失速 → 内側が抜けてくる「抜き」パターン優先
        #   差し系自滅  → 蓋をされて失速   → 外側が流れてくる「恵まれ」パターン優先
        #   従来は恵まれ+抜きを単純平均していたが、自滅タイプを無視すると
        #   「まくり自滅なのに外側が2着に来る確率が高く出る」という誤りが生じていた。
        if scenario == "SC":
            # _sc_fly_typeは外側スコープから参照（_calc_3rentan_probs_v2内で定義済み）
            _fly_type = _sc_fly_type if "_sc_fly_type" in dir() else "不明"
            # 自滅タイプ別の重み設定
            if _fly_type == "まくり系":
                # まくり系自滅: 内側が抜けやすい → 抜き70% + 恵まれ30%
                _w_em, _w_nu = 0.30, 0.70
            elif _fly_type == "差し系":
                # 差し系自滅: 外側が恵まれやすい → 恵まれ70% + 抜き30%
                _w_em, _w_nu = 0.70, 0.30
            else:
                # 不明: 従来通り均等ブレンド
                _w_em, _w_nu = 0.50, 0.50

            result = {}
            for c_str in [str(i) for i in range(1, 7)]:
                if c_str == actual_course:
                    continue
                row_em = _fetch("恵まれ", actual_course, c_str)
                row_nu = _fetch("抜き",   actual_course, c_str)

                # 両方ある場合: 自滅タイプ別重みでブレンド
                if row_em is not None and row_nu is not None:
                    try:
                        r2 = (float(row_em.get("2着率")     or 0) * _w_em +
                              float(row_nu.get("2着率")     or 0) * _w_nu)
                        r3 = (float(row_em.get("3着以内率") or 0) * _w_em +
                              float(row_nu.get("3着以内率") or 0) * _w_nu)
                        result[c_str] = {"2着率": r2, "3着以内率": r3}
                    except (ValueError, TypeError):
                        pass
                # 片方しかない場合: あるほうを使用
                elif row_em is not None:
                    try:
                        result[c_str] = {
                            "2着率":     float(row_em.get("2着率")     or 0),
                            "3着以内率": float(row_em.get("3着以内率") or 0),
                        }
                    except (ValueError, TypeError):
                        pass
                elif row_nu is not None:
                    try:
                        result[c_str] = {
                            "2着率":     float(row_nu.get("2着率")     or 0),
                            "3着以内率": float(row_nu.get("3着以内率") or 0),
                        }
                    except (ValueError, TypeError):
                        pass
            return result if result else None

        # S4でまくり差し/2着コースのデータがない → まくり/2 でフォールバック
        result = {}
        fallback_needed = False
        for c_str in [str(i) for i in range(1, 7)]:
            if c_str == actual_course:
                continue
            row = _fetch(kimete, actual_course, c_str)
            if row is None and scenario == "S4":
                # まくり差しデータなし → まくりで代替
                row = _fetch(_S4_FALLBACK_KIMETE, actual_course, c_str)
                if row:
                    fallback_needed = True
            if row is None:
                continue
            try:
                r2 = float(row.get("2着率")     or 0)
                r3 = float(row.get("3着以内率") or 0)
                result[c_str] = {"2着率": r2, "3着以内率": r3}
            except (ValueError, TypeError):
                pass

        return result if result else None

    # ======================================================================
    # 個人能力スコア計算（2着・3着残存に使用）
    # ======================================================================
    # 「展開的に残れるポジション × その艇が今日実際に残れる実力」を統合する。
    #
    # 使用指数:
    #   win3_rate     : 3連対率（コース別マスタ）   ← 主軸
    #   avg_st        : コース別平均ST              ← ST能力
    #   motor2        : モーター2連率               ← 機力
    #   fly_label     : FLY影響                    ← リスク（減衰）
    #   late_count    : 出遅れ数                   ← リスク（減衰）
    #   kosetsu       : 今節成績                   ← 短期コンディション
    #   フォーム指数  : raw_pm                     ← 中期コンディション
    #
    # 重み設計（合計1.0）:
    #   win3_rate     0.40  (3連対率が主軸)
    #   avg_st        0.20  (STが速い艇ほど展開に付いてこれる)
    #   motor2        0.12  (機力が高い艇は伸び足で粘れる)
    #   kosetsu       0.10  (今節成績で直近コンディション反映)
    #   フォーム指数  0.08  (中期調子)
    #   FLY/出遅れ    -補正  (リスクペナルティ)
    # ======================================================================

    # 各艇のraw_pmマップを構築（prob_second/prob_third から参照）
    _pm_map = {r["waku"]: r.get("raw_pm", {}) or {} for r in results}

    # ST平均（全艇）・モーター平均（全艇）を事前計算（相対評価用）
    _st_vals_all  = [r["avg_st"] for r in results if r.get("avg_st") is not None]
    _st_mean_all  = sum(_st_vals_all) / len(_st_vals_all) if _st_vals_all else 0.17
    _st_min_all   = min(_st_vals_all) if _st_vals_all else 0.12
    _st_max_all   = max(_st_vals_all) if _st_vals_all else 0.22

    _m2_vals_all  = []
    for r in results:
        try:
            v = float(r.get("motor2") or 0)
            if v > 0:
                _m2_vals_all.append(v)
        except (ValueError, TypeError):
            pass
    _m2_mean_all = sum(_m2_vals_all) / len(_m2_vals_all) if _m2_vals_all else 40.0

    def _kosetsu_score_local(kosetsu_str):
        """今節成績文字列 → スコア（0〜1）"""
        if not kosetsu_str or str(kosetsu_str) in ("", "None", "nan", "-"):
            return None
        scores = []
        for token in re.split(r"[-・/]", str(kosetsu_str)):
            token = token.strip()
            if token == "1":   scores.append(1.0)
            elif token == "2": scores.append(0.5)
            elif token == "3": scores.append(0.25)
            elif token.isdigit(): scores.append(0.0)
        return sum(scores) / len(scores) if scores else None

    def _personal_ability_score(waku):
        """
        2着・3着残存計算用の個人能力統合スコア（0〜1）を返す。
        マスタ係数（位置の物理的残存傾向）に掛け合わせて使用する。
        """
        r = next((x for x in results if x["waku"] == waku), {})
        pm = _pm_map.get(waku, {})
        score = 0.0

        # (1) win3_rate（3連対率）: 0.40
        w3 = win3_map.get(waku, 0.5)
        score += min(w3, 1.0) * 0.40

        # (2) avg_st（ST能力）: 0.20
        # STが速い（小さい）艇ほど展開に追走できる → 高スコア
        st = r.get("avg_st")
        if st is not None and (_st_max_all - _st_min_all) > 0.001:
            st_score = 1.0 - (st - _st_min_all) / (_st_max_all - _st_min_all)
            score += max(0.0, min(1.0, st_score)) * 0.20
        else:
            score += 0.5 * 0.20  # データなし → 中立

        # (3) motor2（機力）: 0.12
        try:
            m2 = float(r.get("motor2") or 0)
            if m2 > 0 and _m2_mean_all > 0:
                m2_score = min(m2 / _m2_mean_all, 2.0) / 2.0
                score += m2_score * 0.12
            else:
                score += 0.5 * 0.12
        except (ValueError, TypeError):
            score += 0.5 * 0.12

        # (4) 今節成績（kosetsu）: 0.10
        ks = _kosetsu_score_local(r.get("kosetsu", ""))
        if ks is not None:
            score += min(ks / 0.5, 1.0) * 0.10  # 0.5(平均)で正規化
        else:
            score += 0.5 * 0.10

        # (5) フォーム指数（raw_pm）: 0.08
        form = safe_float(pm.get("フォーム\n指数") or pm.get("フォーム指数"))
        if form is not None:
            form_score = max(0.0, min(1.0, form / 6.0))
            score += form_score * 0.08
        else:
            score += 0.5 * 0.08

        # (6) FLY/出遅れリスクペナルティ
        fly_label = r.get("fly_label", "低")
        fly_pen = {"高": -0.18, "中": -0.08, "低": 0.0}.get(fly_label, 0.0)
        score += fly_pen

        late_count = r.get("late_count", 0) or 0
        if late_count >= 3:
            score -= 0.06
        elif late_count >= 1:
            score -= 0.02

        return max(0.10, min(1.0, score))  # 下限0.10（完全ゼロを防ぐ）

    # 全艇の個人能力スコアを事前計算
    _pers_map = {r["waku"]: _personal_ability_score(r["waku"]) for r in results}

    def prob_second(scenario, first_w, second_w, remaining):
        if scenario == "S1":
            # イン逃げシナリオ2着: マスタ逃げ/1着の2着率で補強（会場別→全国優先）
            # マスタなし時は従来の circle_pct（イン逃げ時2着優位度）を使用
            _master_s1 = _get_tenkai_rates("S1", first_w)
            if _master_s1:
                _r2_vals = [v["2着率"] for v in _master_s1.values() if v["2着率"] > 0]
                _avg2 = sum(_r2_vals) / len(_r2_vals) if _r2_vals else 0
                if _avg2 > 0.001:
                    # マスタ2着率 × circ_map（選手個人の有利度）をブレンド
                    # マスタ: コース位置の物理的な残存確率（全国実績）
                    # circ_map: 選手個人の1号艇との相性・イン逃げ時2着実績
                    # → 0.6:0.4 のブレンドで両方を活かす
                    w2 = {}
                    for w in remaining:
                        master_coef = max(0.5, min(2.0, _master_s1.get(w, {}).get("2着率", _avg2) / _avg2))
                        circ_val    = max(circ_map.get(w, 0), 0.001)
                        w2[w] = master_coef * 0.6 + circ_val * 0.4
                    total = sum(w2.values()) or 1
                    return w2.get(second_w, 0.001) / total
            # マスタなし → 従来の circle_pct のみ
            circ_rem = {w: max(circ_map.get(w, 0), 0.001) for w in remaining}
            total = sum(circ_rem.values()) or 1
            return circ_rem.get(second_w, 0.001) / total

        elif scenario == "SC":
            # 潰れ展開の2着 → 恵まれ/抜きのマスタ実データで sc_beneficiary を補強
            # 自滅した主要飛び役は2着にも残りにくい（確率を0.2に減衰）
            fly_waku = _sc_info.get("sc_fly_waku")
            _master_sc = _get_tenkai_rates("SC", first_w)
            w2 = {}
            for w in remaining:
                score = max(_sc_beneficiary.get(w, 0.1), 0.001)
                if _master_sc:
                    # マスタ実データとsc_beneficiaryをブレンド（0.5:0.5）
                    m_r2 = _master_sc.get(w, {}).get("2着率", 0)
                    if m_r2 > 0:
                        score = score * 0.5 + m_r2 * 0.5
                if w == fly_waku:
                    score *= 0.2
                w2[w] = max(score, 0.001)
            total = sum(w2.values()) or 1
            return w2.get(second_w, 0.001) / total
        else:
            # S2差し・S3まくり・S4まくり差しで内側残存補正テーブルを使い分け
            #
            # 【v6改善】展開別残存マスタの実2着率を優先参照する。
            # マスタが参照できた場合: 2着率を相対係数に変換して rel_map へ乗算
            # マスタなし（データ不足）: 従来のハードコードテーブルをそのまま使用
            #
            # S2（差し）: 差した艇より内側が残りやすい
            # S3（まくり）: 外から全艇を押し込む → 内側が圧縮
            # S4（まくり差し）: まくりより内側が残りやすい
            # 【v6.6】ハードコード → 展開別残存_全国マスタの実測値ベースに置き換え
            # 旧テーブルは「内側が残りやすい」という思い込みで設計されており、
            # 実測値と大幅に乖離（最大+1.83）していた。
            # 算出方法: 各(決まり手×1着コース)の2着率を平均で割り max(0.5, min(2.0, 値)) でクリップ
            # データ: 展開別残存_全国シート（信頼度1、S2:106〜3486件、S3:258〜2311件、S4:363〜2138件）
            # ※ マスタ上書きロジック（_get_tenkai_rates）は引き続き有効。
            #   会場別マスタが信頼度>=0.3なら会場別値で再上書きされる。
            #   このテーブルはマスタにデータがない組み合わせのフォールバックとして機能する。
            _INNER_S2 = {   # 差し（全国実測ベース）
                # 差しの物理: 差した艇より内側が残りやすい
                # 差し艇が内側を切って前に出るため、差し艇より外側の艇は置き去り
                "2": {"1": 2.000, "3": 0.771, "4": 0.606, "5": 0.500, "6": 0.500},
                "3": {"1": 1.131, "2": 1.637, "4": 1.076, "5": 0.775, "6": 0.500},
                # 【修正】4号差し: 旧値はS3まくりと完全一致（誤り）→ 差しの物理法則に修正
                # 差した4号より内側（1〜3号）が後続に残りやすい
                # 外側の5・6号は4号に弾かれて残りにくい
                "4": {"1": 1.400, "2": 1.200, "3": 1.100, "5": 0.650, "6": 0.550},
                "5": {"1": 1.356, "2": 1.384, "3": 1.158, "4": 0.791, "6": 0.500},
                "6": {"1": 0.849, "2": 1.226, "3": 1.226, "4": 1.085, "5": 0.613},
            }
            _INNER_S3 = {   # まくり（全国実測ベース）
                # まくりの物理: まくり艇が外から押し込む
                # 先マイに近い外側艇（まくり艇の直後）が2着に残りやすい
                # 内側（3号より内）は圧縮されて残りにくい
                # 【追加】2号まくりは稀だが発生する → S2差しテーブルをベースに外側有利に補正
                "2": {"1": 1.000, "3": 0.700, "4": 1.100, "5": 1.200, "6": 0.800},
                "3": {"1": 1.111, "2": 0.964, "4": 1.465, "5": 1.044, "6": 0.500},
                "4": {"1": 1.090, "2": 0.945, "3": 0.504, "5": 1.741, "6": 0.718},
                "5": {"1": 1.593, "2": 1.049, "3": 0.593, "4": 0.528, "6": 1.236},
                "6": {"1": 1.764, "2": 1.415, "3": 0.698, "4": 0.756, "5": 0.500},
            }
            _INNER_S4 = {   # まくり差し（全国実測ベース）
                # まくり差しの物理: まくりと差しの中間
                # 差しより外側が残りやすく、まくりより内側が残りやすい傾向
                # 【追加】2号まくり差しはS2差しに近いパターン → 内側有利
                "2": {"1": 1.800, "3": 0.900, "4": 0.750, "5": 0.600, "6": 0.500},
                "3": {"1": 2.000, "2": 0.839, "4": 0.563, "5": 0.500, "6": 0.500},
                "4": {"1": 2.000, "2": 0.755, "3": 1.227, "5": 0.755, "6": 0.500},
                "5": {"1": 1.954, "2": 0.805, "3": 0.725, "4": 1.137, "6": 0.500},
                "6": {"1": 1.970, "2": 0.730, "3": 0.661, "4": 0.813, "5": 0.827},
            }
            if scenario == "S2":
                pos_tbl = _INNER_S2.get(first_w, {})
            elif scenario == "S3":
                pos_tbl = _INNER_S3.get(first_w, {})
            else:  # S4
                pos_tbl = _INNER_S4.get(first_w, {})

            # 【v6】展開別残存マスタから実2着率を取得し、相対係数に変換して pos_tbl を上書き
            _master_rates2 = _get_tenkai_rates(scenario, first_w)
            if _master_rates2:
                _r2_vals = [v["2着率"] for v in _master_rates2.values() if v["2着率"] > 0]
                _avg2 = sum(_r2_vals) / len(_r2_vals) if _r2_vals else 0
                if _avg2 > 0.001:
                    # マスタ相対係数: 0.5〜2.0にクリップ
                    _master_pos_tbl = {
                        c: max(0.5, min(2.0, v["2着率"] / _avg2))
                        for c, v in _master_rates2.items()
                    }
                    # マスタ値がある艇はマスタ優先、ない艇はハードコードをそのまま
                    pos_tbl = {**pos_tbl, **_master_pos_tbl}

            w2 = {}
            for w in remaining:
                # 【修正v2】win3_rate → 個人能力統合スコア（_pers_map）に変更
                # win3_rate だけでなく ST・機力・フォーム・FLYリスクを統合した
                # 「今日この艇が実際に2着に残れる実力」を使う
                base = max(_pers_map.get(w, 0.3), 0.001)
                base *= pos_tbl.get(w, 1.0)
                w2[w] = base
            total = sum(w2.values()) or 1
            return w2.get(second_w, 0.001) / total

    # ── 条件付き3着確率 P(3着=C | 1着=A, 2着=B, scenario) ──
    #
    # S1（イン逃げ）:   idx3（イン逃げ時3着指数）を使用
    # S2（差し）:       win3_rate(0.6) + rel_win1(0.4) の加重平均
    # S3（まくり系）:   内側艇は win3_rate で下方補正（弾かれやすい）
    # SC（潰れ展開）:   漁夫スコア（_sc_beneficiary）を使用
    #                   位置ボーナス × 攻撃性逆数 × win3_rate の合成値
    #
    # 【修正(1)】2着艇(second_w)が3着分布に与える影響を反映
    # 旧問題: P(3着=C | A, B) の計算が second_w に完全に依存しない
    #         → 3-2-? と 3-5-? で同じ pos_tbl3 を使っていた
    #
    # 新方式: second_w の「進路封鎖効果」を3着スコアに乗算補正する。
    #
    #   封鎖効果の物理法則:
    #     S2（差し）: 2着に入った差し艇がC艇の進路を塞ぐ
    #       → 2着艇よりコースが外側かつ差し系の艇 → 進路を塞がれる（スコア減衰）
    #       → 2着艇よりコースが内側の艇 → 先に走っており影響なし
    #     S3（まくり）: まくり艇が外側に膨らんで2着に入る
    #       → 2着艇と同じく外側で展開していた艇 → 共倒れリスク（スコア減衰）
    #       → 内側艇 → まくり艇が抜けたコースに残る（スコア維持〜上昇）
    #     S4（まくり差し）: S2とS3の中間的な封鎖効果
    #
    #   封鎖係数の計算:
    #     try_int(w) > try_int(second_w) かつ w と second_w が同じ展開系 → 減衰
    #     それ以外 → 影響なし（係数1.0）
    #   補正幅: 0.70〜1.0 に抑えて過剰な振れを防ぐ

    def _second_block_factor(scenario, second_w, third_w):
        """
        2着艇(second_w)が3着候補(third_w)の進路を封鎖する係数を返す。
        封鎖あり → 0.70〜0.90 / 封鎖なし → 1.0
        """
        try:
            s_int = int(second_w)
            t_int = int(third_w)
        except (ValueError, TypeError):
            return 1.0

        if scenario == "S2":
            # 差しシナリオ: 2着差し艇より外側の艇が差しに来ようとすると進路封鎖
            # → 外側で差しを試みる艇のスコアを減衰
            if t_int > s_int:
                sashi_t = safe_pct(cm_map.get(third_w, {}), "差し%")
                # 差し%が高いほど封鎖影響を受ける（自分も差しに行こうとする）
                block_strength = min(sashi_t / 30.0, 1.0)   # 差し30%で最大
                return max(0.70, 1.0 - 0.30 * block_strength)
        elif scenario == "S3":
            # まくりシナリオ: 2着まくり艇の外側で同様にまくりを仕掛けた艇は共倒れリスク
            if t_int > s_int:
                mak_t = safe_pct(cm_map.get(third_w, {}), "まくり%")
                block_strength = min(mak_t / 25.0, 1.0)
                return max(0.75, 1.0 - 0.25 * block_strength)
            # 内側艇: まくり艇が掃いたコースに残る → 微増（最大+10%）
            if t_int < s_int and t_int > 1:
                return min(1.10, 1.0 + 0.10 * (s_int - t_int) / 5.0)
        elif scenario == "S4":
            # まくり差しシナリオ: S2とS3の中間
            if t_int > s_int:
                maksa_t = (safe_pct(cm_map.get(third_w, {}), "まくり差し%") +
                           safe_pct(cm_map.get(third_w, {}), "差し%")) / 2.0
                block_strength = min(maksa_t / 28.0, 1.0)
                return max(0.72, 1.0 - 0.28 * block_strength)
        return 1.0

    def prob_third(scenario, first_w, second_w, third_w, remaining):
        if scenario == "S1":
            # イン逃げシナリオ3着: マスタ逃げ/1着の3着以内率で idx3_map を補強
            # マスタなし時は従来の idx3_map を使用
            # S1は先頭から逃げ・差し・流れ込みの順で展開が決まるため
            # 2着艇の封鎖効果は小さく適用しない（idx3が既に内側有利を反映済み）
            _master_s1_3rd = _get_tenkai_rates("S1", first_w)
            if _master_s1_3rd:
                _r3_vals = [v["3着以内率"] for v in _master_s1_3rd.values() if v["3着以内率"] > 0]
                _avg3 = sum(_r3_vals) / len(_r3_vals) if _r3_vals else 0
                if _avg3 > 0.001:
                    w3 = {}
                    for w in remaining:
                        master_coef = max(0.5, min(2.0,
                            _master_s1_3rd.get(w, {}).get("3着以内率", _avg3) / _avg3))
                        idx3_val = max(idx3_map.get(w, 0), 0.001)
                        # 【修正v2】マスタ(0.5) × idx3(0.3) × 個人能力(0.2) の3軸ブレンド
                        # 個人能力（ST・機力・フォーム）を追加して「今日の実力」を反映
                        pers_val = max(_pers_map.get(w, 0.3), 0.001)
                        w3[w] = master_coef * 0.50 + idx3_val * 0.30 + pers_val * 0.20
                    total = sum(w3.values()) or 1
                    return w3.get(third_w, 0.001) / total
            # マスタなし → 従来の idx3_map のみ
            w3 = {w: max(idx3_map.get(w, 0), 0.001) for w in remaining}

        elif scenario == "S2":
            # S2（差し系）: 3着も内側残存補正を適用
            # 【v6】展開別残存マスタの実3着以内率を優先参照
            # 【v6.6】3着以内率も全国実測値ベースに置き換え
            _INNER_REMAIN3 = {   # 差し（全国実測ベース・3着以内率）
                "2": {"1": 1.992, "3": 1.056, "4": 0.909, "5": 0.643, "6": 0.500},
                "3": {"1": 1.053, "2": 1.310, "4": 1.105, "5": 0.930, "6": 0.602},
                "4": {"1": 1.520, "2": 1.275, "3": 1.080, "5": 0.691, "6": 0.500},
                "5": {"1": 1.271, "2": 1.200, "3": 1.045, "4": 0.989, "6": 0.500},
                "6": {"1": 0.943, "2": 1.226, "3": 1.061, "4": 0.896, "5": 0.873},
            }
            pos_tbl3 = _INNER_REMAIN3.get(first_w, {})
            # マスタ3着以内率で上書き
            _master_rates3s2 = _get_tenkai_rates("S2", first_w)
            if _master_rates3s2:
                _r3_vals = [v["3着以内率"] for v in _master_rates3s2.values() if v["3着以内率"] > 0]
                _avg3 = sum(_r3_vals) / len(_r3_vals) if _r3_vals else 0
                if _avg3 > 0.001:
                    _master_pos3 = {
                        c: max(0.5, min(2.0, v["3着以内率"] / _avg3))
                        for c, v in _master_rates3s2.items()
                    }
                    pos_tbl3 = {**pos_tbl3, **_master_pos3}
            w3 = {}
            for w in remaining:
                # 【修正v2】win3_rate → 個人能力統合スコア（_pers_map）に変更
                pers  = max(_pers_map.get(w, 0.3), 0.001)
                rel   = max(rel_map.get(w, 0),     0.001)
                # 個人能力(0.7) + 相対1着率(0.3) のブレンド
                # 3着には「残れる実力」と「位置的有利」の両方が必要
                base  = pers * 0.70 + rel * 0.30
                base *= pos_tbl3.get(w, 1.0)
                # 【修正(1)】2着艇による進路封鎖効果を適用
                base *= _second_block_factor("S2", second_w, w)
                w3[w] = base

        elif scenario in ("S3", "S4"):
            # S3（まくり）・S4（まくり差し）: シナリオ別内側残存補正
            # 【v6】展開別残存マスタの実3着以内率を優先参照
            # 【v6.6】3着以内率も全国実測値ベースに置き換え
            _INNER3_S3 = {   # まくり（全国実測ベース・3着以内率）
                # 【追加】2号まくり: まくり後の3着は内側1号と外側3〜4号が残りやすい
                "2": {"1": 1.200, "3": 1.100, "4": 1.050, "5": 0.900, "6": 0.750},
                "3": {"1": 1.057, "2": 1.015, "4": 1.260, "5": 1.058, "6": 0.609},
                "4": {"1": 1.183, "2": 0.956, "3": 0.600, "5": 1.385, "6": 0.875},
                "5": {"1": 1.427, "2": 1.126, "3": 0.675, "4": 0.634, "6": 1.138},
                "6": {"1": 1.502, "2": 1.318, "3": 0.814, "4": 0.814, "5": 0.552},
            }
            _INNER3_S4 = {   # まくり差し（全国実測ベース・3着以内率）
                # 【追加】2号まくり差し: 差しに近いパターン → 内側1号残り強い
                "2": {"1": 1.800, "3": 0.950, "4": 0.800, "5": 0.650, "6": 0.500},
                "3": {"1": 1.923, "2": 0.870, "4": 1.016, "5": 0.765, "6": 0.500},
                "4": {"1": 1.384, "2": 0.804, "3": 1.027, "5": 1.174, "6": 0.612},
                "5": {"1": 1.483, "2": 0.926, "3": 0.839, "4": 1.009, "6": 0.741},
                "6": {"1": 1.575, "2": 0.970, "3": 0.722, "4": 0.901, "5": 0.832},
            }
            pos_tbl3 = (_INNER3_S3 if scenario == "S3" else _INNER3_S4).get(first_w, {})
            # マスタ3着以内率で上書き
            _master_rates3 = _get_tenkai_rates(scenario, first_w)
            if _master_rates3:
                _r3_vals = [v["3着以内率"] for v in _master_rates3.values() if v["3着以内率"] > 0]
                _avg3 = sum(_r3_vals) / len(_r3_vals) if _r3_vals else 0
                if _avg3 > 0.001:
                    _master_pos3 = {
                        c: max(0.5, min(2.0, v["3着以内率"] / _avg3))
                        for c, v in _master_rates3.items()
                    }
                    pos_tbl3 = {**pos_tbl3, **_master_pos3}
            w3 = {}
            for w in remaining:
                # 【修正v2】win3_rate → 個人能力統合スコア（_pers_map）に変更
                base = max(_pers_map.get(w, 0.3), 0.001)
                base *= pos_tbl3.get(w, 1.0)
                # 【修正(1)】2着艇による進路封鎖効果を適用
                base *= _second_block_factor(scenario, second_w, w)
                w3[w] = base

        else:  # SC（潰れ展開）
            # 漁夫スコア + 恵まれ/抜きのマスタ3着以内率でブレンド
            # SCシナリオは展開が複雑なため封鎖効果は適用しない（sc_beneficiaryに内包）
            _master_sc_3rd = _get_tenkai_rates("SC", first_w)
            w3 = {}
            for w in remaining:
                score = max(_sc_beneficiary.get(w, 0.1), 0.001)
                if _master_sc_3rd:
                    m_r3 = _master_sc_3rd.get(w, {}).get("3着以内率", 0)
                    if m_r3 > 0:
                        score = score * 0.5 + m_r3 * 0.5
                w3[w] = max(score, 0.001)

        total = sum(w3.values()) or 1
        return w3.get(third_w, 0.001) / total

    # ── 全120通りの確率を計算 ──
    combos_dict = {}
    scenarios = [("S1", p_s1), ("S2", p_s2), ("S3", p_s3), ("S4", p_s4), ("SC", p_sc)]

    for first in wakus:
        for second in wakus:
            if second == first:
                continue
            for third in wakus:
                if third == first or third == second:
                    continue
                key = f"{first}-{second}-{third}"
                prob_total = 0.0
                for scenario, p_sc in scenarios:
                    if p_sc < 1e-9:
                        continue
                    p1_sc = prob_first_given_scenario(scenario, first)
                    if p1_sc < 1e-9:
                        continue
                    rem2 = [w for w in wakus if w != first]
                    p2_sc = prob_second(scenario, first, second, rem2)
                    rem3 = [w for w in wakus if w != first and w != second]
                    p3_sc = prob_third(scenario, first, second, third, rem3)
                    prob_total += p_sc * p1_sc * p2_sc * p3_sc

                # 【(8)】S3S2複合シナリオ: まくりで3〜6号艇が1着、2号艇が差して2着に入る
                # 1着: S3（まくり）条件付き確率、2着: 2号艇を差し方向で優遇
                if p_s3s2 > 1e-9:
                    p1_s3s2 = prob_first_given_scenario("S3", first)
                    if p1_s3s2 > 1e-9:
                        rem2 = [w for w in wakus if w != first]
                        # 2着は差し方向の確率を使う（S2の2着分布を参照）
                        p2_s3s2 = prob_second("S2", first, second, rem2)
                        rem3 = [w for w in wakus if w != first and w != second]
                        p3_s3s2 = prob_third("S2", first, second, third, rem3)
                        prob_total += p_s3s2 * p1_s3s2 * p2_s3s2 * p3_s3s2

                combos_dict[key] = {
                    "combo":          key,
                    "first":          first,
                    "second":         second,
                    "third":          third,
                    "prob":           prob_total,
                    "is_outer_first": int(first) >= 4,
                    "top_scenario":   max(
                        scenarios,
                        key=lambda sc: sc[1] * prob_first_given_scenario(sc[0], first)
                    )[0],
                    # SCシナリオ情報（数値シート表示用）
                    "sc_fly_type":    _sc_fly_type,
                    "sc_beneficiary": _sc_beneficiary.get(second, 0),
                }

    # ── 【修正(7)】prob正規化（合計を1.0に統一）────────────────────────────────
    # 問題: シナリオ重みの積算方式により prob の合計が1.0を超える場合がある
    #       （実測値: 合計1.5〜2.0程度）
    # 影響: _merge_scenarios の prob 降順ソートによる買い目絞り込みが歪む
    # 修正: 全120通りの prob を合計で割って正規化する（1行追加のみ）
    # ※ theoretical_odds の計算はこの後で行うため、正規化後の prob を使用する
    _prob_total_all = sum(c["prob"] for c in combos_dict.values())
    if _prob_total_all > 0:
        for c in combos_dict.values():
            c["prob"] = c["prob"] / _prob_total_all

    # ── 理論オッズ・ハイブリッドスコアを付与 ──
    # 理論オッズ = (1 / prob) × (1 - テラ銭0.25)
    # ハイブリッドスコア = 確率60% + 理論オッズ40% の合成ランク
    # → 低オッズ本命への過集中を防ぎ、中〜高配当の期待値が高い組み合わせを優先
    TAKER_RATE  = 0.25
    ODDS_WEIGHT = 0.4  # 0.0=純確率順 / 1.0=純オッズ順 / 0.4=バランス重視

    for c in combos_dict.values():
        p = max(c["prob"], 1e-6)
        c["theoretical_odds"] = round((1.0 / p) * (1.0 - TAKER_RATE), 1)
        c["prob_rank"]  = 0
        c["odds_rank"]  = 0

    combos_sorted_prob = sorted(combos_dict.values(), key=lambda x: x["prob"],             reverse=True)
    combos_sorted_odds = sorted(combos_dict.values(), key=lambda x: x["theoretical_odds"], reverse=True)

    for i, c in enumerate(combos_sorted_prob): c["prob_rank"]  = i + 1
    for i, c in enumerate(combos_sorted_odds): c["odds_rank"]  = i + 1

    for c in combos_dict.values():
        c["hybrid_score"] = (1.0 - ODDS_WEIGHT) * (1.0 / c["prob_rank"]) + \
                             ODDS_WEIGHT         * (1.0 / c["odds_rank"])

    combos = sorted(combos_dict.values(), key=lambda x: x["hybrid_score"], reverse=True)
    return combos


# ─────────────────────────────────────────────────────────────────────────────
# 期待値計算（実際のオッズExcelとの照合用）
# 使い方:
#   combos = _calc_3rentan_probs_v2(results)
#   actual = load_actual_odds_from_excel("odds_YYYYMMDD_R1.xlsx")  # 別途実装
#   ev_list = calc_ev_from_actual_odds(combos, actual)
#   suggestion = suggest_by_ev(ev_list, min_ev=0.05)  # EV+5%以上のみ
# ─────────────────────────────────────────────────────────────────────────────

def calc_ev_from_actual_odds(combos, actual_odds_dict):
    """
    実際の払戻オッズと推定確率から期待値（EV）を計算する。

    Parameters
    ----------
    combos : list[dict]
        _calc_3rentan_probs_v2() の戻り値。各要素に "combo"（例: "1-2-3"）と "prob" が必要。
    actual_odds_dict : dict
        {"1→2→3": 払戻倍率float, ...} 形式。
        ※ 理論オッズExcelから読んだ全120通りの実際の払戻倍率を渡す。

    Returns
    -------
    list[dict]
        各コンボに以下を追加した辞書のリスト（EV降順）:
          actual_odds : 実際の払戻倍率（Noneなら未取得）
          ev          : 期待値 = actual_odds × prob - 1.0（Noneなら計算不可）
          ev_pct      : EVを%表示した文字列（例: "+12.3%"）
          ev_positive : EVがプラスかどうか
    """
    result = []
    for c in combos:
        key    = c["combo"]   # "1-2-3" 形式
        actual = actual_odds_dict.get(key)
        c2     = dict(c)
        if actual is not None and c.get("prob", 0) > 0:
            ev = actual * c["prob"] - 1.0
            c2["actual_odds"] = round(actual, 1)
            c2["ev"]          = round(ev, 4)
            c2["ev_pct"]      = f"{'+' if ev >= 0 else ''}{ev*100:.1f}%"
            c2["ev_positive"] = ev > 0
        else:
            c2["actual_odds"] = None
            c2["ev"]          = None
            c2["ev_pct"]      = "N/A"
            c2["ev_positive"] = False
        result.append(c2)
    # EV降順（未取得はそれ以下に）
    result.sort(key=lambda x: (x["ev"] if x["ev"] is not None else -999), reverse=True)
    return result


def suggest_by_ev(combos_with_ev, min_ev=0.0, max_bets=8):
    """
    期待値プラスの組み合わせのみを買い目として返す。

    Parameters
    ----------
    combos_with_ev : list[dict]  calc_ev_from_actual_odds() の戻り値
    min_ev         : float       最低期待値閾値（デフォルト0.0 = 期待値プラスのみ）
                                 例: 0.05 → EV+5%以上のみ対象
    max_bets       : int         最大買い目数

    Returns
    -------
    dict
        buy_list    : 買い目リスト（EV降順）、例: ["2→3→1", "1→3→2"]
        ev_summary  : 各買い目の詳細（combo / prob% / actual_odds / ev%）
        total_bets  : 点数
        best_ev     : 最高EVの組み合わせ
        skip        : True なら期待値プラスがゼロ → 見送り推奨
        reason      : 判定理由の説明文
    """
    positives = [c for c in combos_with_ev
                 if c.get("ev") is not None and c["ev"] > min_ev]
    positives = positives[:max_bets]

    if not positives:
        # EV最高値を探して理由を付ける
        best_any = next((c for c in combos_with_ev if c.get("ev") is not None), None)
        best_str = f"（最高EV: {best_any['ev_pct']} {best_any['combo']}）" if best_any else ""
        return {
            "buy_list":   [],
            "ev_summary": [],
            "total_bets": 0,
            "best_ev":    None,
            "skip":       True,
            "reason":     f"EV>{min_ev*100:.0f}%の組み合わせなし → 見送り推奨{best_str}"
        }

    return {
        "buy_list": [c["combo"] for c in positives],
        "ev_summary": [
            {
                "combo":        c["combo"],
                "prob_pct":     f"{c['prob']*100:.2f}%",
                "actual_odds":  c["actual_odds"],
                "ev_pct":       c["ev_pct"],
            }
            for c in positives
        ],
        "total_bets": len(positives),
        "best_ev":    positives[0],
        "skip":       False,
        "reason":     f"EV>{min_ev*100:.0f}%が{len(positives)}点（最高: {positives[0]['ev_pct']} {positives[0]['combo']}）"
    }


def load_actual_odds_from_excel(filepath, sheet_name=0):
    """
    理論オッズExcel（全120通り）を読み込んで辞書形式で返す。

    Excel形式（理想オッズ完成版.xlsx 準拠）:
      - 列A: 1着艇番（int）
      - 列B: 2着艇番（int）
      - 列C: 3着艇番（int）
      - 列D: オッズ（float）
      - 1行目はヘッダ（読み飛ばし）
      - データ120行（3連単全通り）

    Parameters
    ----------
    filepath   : str  Excelファイルパス（例: "odds/理想オッズ_20240101_R1.xlsx"）
    sheet_name : int or str  シート番号またはシート名（デフォルト0）

    Returns
    -------
    dict  {"1→2→3": float, ...}  キーは "A→B→C" 形式
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    result = {}
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if row[0] is None:
            continue
        try:
            key = f"{int(row[0])}-{int(row[1])}-{int(row[2])}"
            result[key] = float(row[3])
        except (TypeError, ValueError):
            continue
    wb.close()
    return result


# ============================================================
# ============================================================
# 参加グレードテーブル（BT 8,458R 会場×決まり手別 ROI 実績）
# ============================================================
_ENTRY_GRADE_TABLE = {
    ("三国","まくり差し"):("S",5.90,0.270,12170,37),("鳴門","まくり差し"):("S",3.91,0.360,7220,25),
    ("福岡","まくり差し"):("S",3.37,0.364,8590,11),("戸田","まくり差し"):("S",3.28,0.351,3710,37),
    ("びわこ","まくり差し"):("S",3.24,0.361,5480,83),("江戸川","まくり差し"):("S",3.18,0.303,10750,33),
    ("住之江","抜き"):("S",2.95,0.278,10590,18),("若松","まくり差し"):("S",2.92,0.312,5670,32),
    ("唐津","まくり差し"):("S",2.89,0.302,4080,43),("浜名湖","まくり差し"):("S",2.39,0.212,3990,52),
    ("徳山","抜き"):("S",2.38,0.243,7090,37),("唐津","差し"):("S",2.27,0.208,2470,48),
    ("平和島","まくり差し"):("S",2.16,0.356,3690,45),("蒲郡","差し"):("S",2.13,0.258,2510,31),
    ("芦屋","まくり差し"):("S",2.11,0.192,8720,26),("尼崎","抜き"):("S",1.90,0.231,7580,26),
    ("徳山","まくり"):("S",1.83,0.133,9770,45),("びわこ","差し"):("S",1.81,0.222,2420,54),
    ("蒲郡","まくり差し"):("S",1.77,0.393,3290,56),("常滑","まくり"):("S",1.75,0.136,12630,44),
    ("若松","差し"):("S",1.70,0.321,4970,28),("びわこ","抜き"):("S",1.67,0.154,20070,26),
    ("桐生","差し"):("S",1.65,0.159,5820,44),("尼崎","まくり差し"):("S",1.64,0.261,4510,46),
    ("鳴門","抜き"):("S",1.63,0.360,3820,25),("多摩川","抜き"):("S",1.57,0.200,6610,15),
    ("浜名湖","差し"):("S",1.50,0.291,3420,55),
    ("桐生","まくり"):("A",1.46,0.203,4750,69),("三国","差し"):("A",1.46,0.137,4100,51),
    ("津","まくり差し"):("A",1.44,0.333,3470,30),("下関","まくり"):("A",1.42,0.163,8150,49),
    ("常滑","差し"):("A",1.42,0.235,4410,34),("鳴門","まくり"):("A",1.42,0.179,5510,56),
    ("児島","まくり"):("A",1.40,0.156,5120,64),("下関","まくり差し"):("A",1.39,0.348,3740,46),
    ("下関","抜き"):("A",1.38,0.280,3030,25),("浜名湖","まくり"):("A",1.34,0.175,3830,57),
    ("浜名湖","抜き"):("A",1.30,0.346,4430,26),("江戸川","まくり"):("A",1.30,0.244,3600,86),
    ("戸田","抜き"):("A",1.27,0.318,4310,22),("宮島","まくり差し"):("A",1.26,0.238,3630,42),
    ("住之江","まくり差し"):("A",1.21,0.265,4630,34),
    ("津","差し"):("B",1.20,0.212,3190,33),("平和島","まくり"):("B",1.17,0.195,4910,82),
    ("丸亀","差し"):("B",1.15,0.302,3680,43),("若松","まくり"):("B",1.13,0.235,3680,34),
    ("児島","まくり差し"):("B",1.12,0.256,3800,78),("桐生","抜き"):("B",1.12,0.208,2520,24),
    ("芦屋","まくり"):("B",1.08,0.200,4840,30),("戸田","まくり"):("B",1.04,0.153,3150,98),
    ("戸田","差し"):("B",1.02,0.226,4940,62),("丸亀","まくり"):("B",1.02,0.228,2840,57),
    ("桐生","まくり差し"):("C",0.97,0.243,2610,37),("芦屋","差し"):("C",0.96,0.273,4500,22),
    ("丸亀","まくり差し"):("C",0.96,0.204,3970,54),("大村","まくり差し"):("C",0.96,0.250,2320,36),
    ("尼崎","差し"):("C",0.94,0.233,1920,43),("常滑","抜き"):("C",0.93,0.231,3960,13),
    ("津","まくり"):("C",0.90,0.250,4630,28),("唐津","抜き"):("C",0.84,0.143,6530,28),
    ("徳山","差し"):("C",0.84,0.244,3490,41),("鳴門","差し"):("C",0.83,0.171,4580,35),
    ("平和島","差し"):("C",0.83,0.194,4240,62),("福岡","まくり"):("C",0.81,0.121,3260,58),
    ("多摩川","まくり差し"):("C",0.80,0.152,2730,33),("下関","差し"):("C",0.76,0.190,2740,42),
    ("江戸川","抜き"):("C",0.75,0.217,2410,46),
    ("江戸川","差し"):("D",0.75,0.180,2220,61),("多摩川","まくり"):("D",0.72,0.106,5580,66),
    ("三国","抜き"):("D",0.69,0.182,2710,22),("三国","まくり"):("D",0.66,0.190,3240,42),
    ("大村","まくり"):("D",0.64,0.133,8270,30),("宮島","差し"):("D",0.63,0.147,3740,34),
    ("多摩川","差し"):("D",0.63,0.205,2930,44),("宮島","抜き"):("D",0.62,0.143,6030,21),
    ("唐津","まくり"):("D",0.62,0.111,5770,54),("福岡","差し"):("D",0.62,0.258,3030,31),
    ("児島","差し"):("D",0.61,0.154,5380,65),("宮島","まくり"):("D",0.61,0.145,4770,55),
    ("尼崎","まくり"):("D",0.58,0.180,2650,50),("徳山","まくり差し"):("D",0.56,0.194,2700,36),
    ("大村","差し"):("D",0.54,0.130,4190,23),("住之江","差し"):("D",0.49,0.158,1440,19),
    ("児島","抜き"):("D",0.48,0.143,2800,21),("蒲郡","抜き"):("D",0.47,0.200,900,20),
    ("住之江","まくり"):("D",0.46,0.114,3590,44),("若松","抜き"):("D",0.44,0.188,3330,16),
    ("平和島","抜き"):("D",0.44,0.077,7860,26),("常滑","まくり差し"):("D",0.38,0.114,2960,35),
    ("蒲郡","まくり"):("D",0.29,0.104,3200,48),("丸亀","抜き"):("D",0.21,0.222,1270,18),
    ("びわこ","まくり"):("D",0.21,0.053,4360,57),
}
_GRADE_LABEL={"S":"[OK][OK] 強く推奨（ROI150%以上）","A":"[OK]  推奨（ROI120〜149%）",
    "B":"[!]  参加可（ROI100〜119%）","C":"[!][!] 慎重に（ROI75〜99%）","D":"[NG]  見送り推奨（ROI75%未満）"}
_GRADE_FILL={"S":("FFD9EAD3","FF1D5730"),"A":("FFE2EFDA","FF274E13"),
    "B":("FFFFF2CC","FF7F3F00"),"C":("FFFCE4D6","FF7F0000"),"D":("FFFF9999","FF7F0000")}
def _get_entry_grade(venue,scenario_type,honmei_scenario=None):
    kata=None
    if honmei_scenario and isinstance(honmei_scenario,dict):
        kata=(honmei_scenario.get("honmei_patterns",{}).get("honmei",{}).get("primary_kata"))
    if not kata:
        if scenario_type=="逃げ軸流し": kata="逃げ"
        elif scenario_type in ("飛び軸","両建て"): kata="まくり差し"
    entry=_ENTRY_GRADE_TABLE.get((venue,kata)) if (venue and kata) else None
    if not entry:
        return{"grade":"?","label":"[i] データ不足","roi":None,"hit_rate":None,
               "med_odds":None,"sample_n":None,"fill":("FFDDDDDD","FF808080"),"kata":kata or "-"}
    grade,roi,hit_rate,med_odds,sample_n=entry
    return{"grade":grade,"label":_GRADE_LABEL[grade],"roi":roi,"hit_rate":hit_rate,
           "med_odds":med_odds,"sample_n":sample_n,"fill":_GRADE_FILL[grade],"kata":kata}


def _calc_kelly_fraction(theory_syn_odds, total_prob, fraction=0.25):
    """
    【(9)追加】フラクショナルケリー基準による最適賭け比率を算出する。

    【設計思想】
    現状システムは「参加/見送り」の2択だが、
    「参加するなら総資金の何%を使うべきか」が未実装。
    ケリー基準を用いることで、期待値の高いレースに多く賭け
    期待値の低いレースに少なく賭ける資金管理を実現する。

    【計算式】
      通常のケリー基準: f = (b×p - q) / b
        b = 理論合成オッズ（テラ銭控除後）
        p = 的中確率（= total_prob = Σ買い目確率）
        q = 外れ確率（= 1 - p）

    【フラクショナルケリー】
      純粋ケリーは過剰賭けリスクがあるため fraction=0.25（4分の1ケリー）を使用。
      これはバンクロールの変動を抑えながら長期的な資産成長を狙う実用的な設定。

    【パラメータ】
      theory_syn_odds : float  テラ銭控除後の理論合成オッズ（0.75 / Σprob）
      total_prob      : float  買い目全体の的中確率合計（0〜1）
      fraction        : float  ケリー乗数（デフォルト0.25 = 4分の1ケリー）

    【返り値】
      dict:
        kelly_f        : float  ケリー比率（総資金に対する賭け比率）
        kelly_pct      : str    パーセント表示（例: "3.2%"）
        kelly_label    : str    判定ラベル（「積極」「標準」「控えめ」「見送り」）
        kelly_edge     : float  エッジ（期待値 - 1）
        kelly_note     : str    説明テキスト
    """
    import math as _math

    if theory_syn_odds is None or total_prob is None or total_prob <= 0:
        return {
            "kelly_f":     0.0,
            "kelly_pct":   "0.0%",
            "kelly_label": "計算不可",
            "kelly_edge":  0.0,
            "kelly_note":  "合成オッズまたは確率データなし",
        }

    b = float(theory_syn_odds)   # 理論合成オッズ（テラ銭控除後）
    p = float(total_prob)        # 的中確率
    q = 1.0 - p                  # 外れ確率

    # エッジ（期待値 - 1）: プラスなら有利なゲーム
    edge = b * p - 1.0

    if edge <= 0 or b <= 1:
        # 期待値マイナス → 賭けるべきではない
        return {
            "kelly_f":     0.0,
            "kelly_pct":   "0.0%",
            "kelly_label": "見送り推奨（EV<0）",
            "kelly_edge":  round(edge, 4),
            "kelly_note":  f"期待値{(b*p)*100:.1f}% < 100%（マイナス期待値）",
        }

    # ケリー比率計算
    kelly_full = (b * p - q) / b   # 純粋ケリー
    kelly_f    = kelly_full * fraction   # フラクショナルケリー

    # 上限: 資金の20%を超えないようにキャップ（過剰賭け防止）
    kelly_f = min(kelly_f, 0.20)
    kelly_f = max(kelly_f, 0.0)

    # 判定ラベル
    if kelly_f >= 0.10:
        label = "積極（高期待値）"
    elif kelly_f >= 0.05:
        label = "標準"
    elif kelly_f >= 0.02:
        label = "控えめ"
    else:
        label = "最小賭け"

    note = (
        f"理論合成オッズ{b:.1f}倍 × 的中確率{p*100:.1f}% "
        f"→ エッジ{edge*100:.1f}% / 推奨賭け比率{kelly_f*100:.1f}%"
        f"（4分の1ケリー基準）"
    )

    return {
        "kelly_f":     round(kelly_f, 4),
        "kelly_pct":   f"{kelly_f*100:.1f}%",
        "kelly_label": label,
        "kelly_edge":  round(edge, 4),
        "kelly_note":  note,
    }



# 参加見送り判定（バックテスト結果に基づく精度向上フィルタ）
# ============================================================
# 【根拠】バックテスト8,526R分析（2025-12〜2026-02）
#   逃げ軸流し: 4,665R / ROI 27.1% / 損益 −377万  ← 全損失の152%
#   悪会場10会場: ROI 34〜67%帯 / 合計大赤字
#   逃げ除外だけで ROI 130.4% → 黒字転換
# ============================================================
_SKIP_VENUES: set = set()  # ← 低ROI会場フィルタ廃止（BT通過済み）

def _should_skip_race(bet_suggestions: dict, venue: str = "") -> tuple[bool, str]:
    """
    参加見送りを判定する。
    Returns: (skip: bool, reason: str)

    【v4 設計思想】
    「逃げ軸流し = 見送り」というルールを廃止。
    逃げ軸流しは展開予測であって見送り理由ではない。

    見送り条件（構造的に回収できないケースのみ）:
      (0) nyujo_henkou == True（最優先）
         → 枠なり進入を前提とした分析のため、進入変更が確認された時点で分析前提が崩れる
         → 他の条件より先に判定し、無条件で見送り
      (1) s1_prob >= 0.65 かつ himo_are が不参加推奨
         → 逃げほぼ確定 + ヒモ固まり = 1-2-3が低オッズ化確実
      (2) s1_prob >= 0.72（逃げ確率が極端に高い）
         → 逃げが当たっても低オッズ = 構造的に回収不能
      (3) honmei_prob_mismatch == True かつ s1_prob >= 0.60
         → 印<->確率が大きく乖離 + 逃げ優位 = 判断根拠が不明確
      【追加A】1着候補が分散（団子状態）
         → first_prob_map の上位3艇確率差がいずれも10%以内
         → 1着が絞れない = 軸が存在せず買い目構造が成立しない
      【追加B】最有力艇の根拠指標が矛盾
         → 1着率最高艇の個人決まり手率が場平均と±20%以上乖離
         → 指標が矛盾しているため分析の信頼性が低い
    """
    # (0) 進入変更（最優先：分析前提が崩れるため他条件より先に判定）
    if bet_suggestions.get("nyujo_henkou", False):
        return True, (
            "[NG]見送り推奨（進入変更あり）\n"
            "本システムは枠なり進入を前提として分析しています。\n"
            "進入変更が確認されたレースは分析結果が無効になります。"
        )

    s1_prob         = bet_suggestions.get("s1_prob", 0) or 0
    mismatch        = bet_suggestions.get("honmei_prob_mismatch", False)
    mismatch_detail = bet_suggestions.get("honmei_prob_mismatch_detail", "")

    # (1)(2) 逃げ濃厚の場合 → ヒモが絞れるなら点数絞り参加、絞れないなら見送り
    # 【変更理由】
    #   旧: 逃げ確率が高い = 低オッズ = 無条件見送り
    #   新: 逃げ濃厚でもヒモが2艇以内に絞れれば点数3〜6点に収まり回収率が出る
    #       → ヒモ候補艇数（nige_himo_candidates）で参加/見送りを分岐する
    #
    #   nige_himo_candidates は lr_suggest.py の Step0 で _neraime_2nd（逃げ時2着残存候補）
    #   から生成して _early_skip_dict に渡す。データなし（キーなし）の場合は
    #   安全側に倒して旧ロジック（見送り）を維持する。
    if s1_prob >= 0.65:
        _nige_himo = bet_suggestions.get("nige_himo_candidates", None)

        if _nige_himo is None:
            # データなし → 安全側に倒して旧ロジックで判定
            himo_are = bet_suggestions.get("himo_are", {}) or {}
            himo_verdict = himo_are.get("verdict", "対象外")
            if himo_verdict == "不参加推奨" or s1_prob >= 0.72:
                est_odds = himo_are.get("est_top_odds", 0) or 0
                _reason_detail = (
                    f"推定最高人気オッズ{est_odds:.0f}倍台。当たっても回収構造が成立しない"
                    if himo_verdict == "不参加推奨"
                    else "1号艇逃げがほぼ確定→3連単オッズが構造的に低い"
                )
                return True, (
                    f"[NG]見送り推奨（逃げ{s1_prob*100:.0f}%・ヒモ候補データなし）\n"
                    f"{_reason_detail}"
                )

        elif len(_nige_himo) <= 2:
            # ヒモ2艇以内 → 点数を絞って参加できる（見送らない）
            _himo_str = "・".join([f"{w}号" for w in _nige_himo]) if _nige_himo else "－"
            # ※ 見送りにはしないが、呼び出し元で点数絞りを促すために
            #    bet_suggestions へ "nige_shime_recommend": True を立てることを推奨
            pass  # → 参加（条件(3)以降の判定へ進む）

        else:
            # ヒモ3艇以上に分散 → 点数が増えて低オッズとの組み合わせで回収不能
            _himo_str = "・".join([f"{w}号" for w in _nige_himo[:4]])
            return True, (
                f"[NG]見送り推奨（逃げ{s1_prob*100:.0f}%+ヒモ{len(_nige_himo)}艇分散）\n"
                f"ヒモ候補: {_himo_str}\n"
                f"逃げ濃厚でも点数が増えると低オッズで回収構造が成立しない"
            )

    # (3) 印<->確率不一致 + 逃げ優位（判断根拠不明確）
    if mismatch and s1_prob >= 0.60:
        return True, (
            f"[NG]印<->確率不一致見送り（{mismatch_detail}）\n"
            f"逃げ{s1_prob*100:.0f}%優位で◎が確率最大艇でない"
        )

    # ──────────────────────────────────────────────────────────────────────
    # 【追加A】1着候補が分散（団子判定）
    # first_prob_map の上位3艇を取り出し、
    # 最上位と3位の確率差が10%以内 → 1着が絞れない → 見送り
    # ──────────────────────────────────────────────────────────────────────
    first_prob_map = bet_suggestions.get("first_prob_map", {}) or {}
    if first_prob_map:
        _sorted_probs = sorted(first_prob_map.values(), reverse=True)
        if len(_sorted_probs) >= 3:
            _top1 = _sorted_probs[0]
            _top3 = _sorted_probs[2]
            _spread = (_top1 - _top3) * 100  # パーセントポイント差
            if _spread <= 10.0:
                _top_wakus = sorted(
                    first_prob_map.items(), key=lambda x: x[1], reverse=True
                )[:3]
                _waku_str = "・".join(
                    f"{w}号{p*100:.0f}%" for w, p in _top_wakus
                )
                return True, (
                    f"[NG]見送り推奨（1着が絞れません）\n"
                    f"上位3艇の確率差{_spread:.1f}pt（10pt以内）: {_waku_str}\n"
                    f"軸が定まらないため買い目構造が成立しない"
                )

    # ──────────────────────────────────────────────────────────────────────
    # 【追加B】最有力艇の根拠指標が矛盾
    # 1着率最高艇の個人決まり手率（逃げ%）が場平均逃げ率と±20%以上乖離
    # → 例: 1着率58%高いのに個人逃げ31% vs 場平均57% → 根拠が矛盾
    # ──────────────────────────────────────────────────────────────────────
    _kimete_mismatch = bet_suggestions.get("kimete_mismatch", {}) or {}
    if _kimete_mismatch.get("is_mismatch", False):
        _mm_waku  = _kimete_mismatch.get("waku", "?")
        _mm_prob  = _kimete_mismatch.get("first_prob", 0) * 100
        _mm_indiv = _kimete_mismatch.get("indiv_nige_pct", 0) * 100
        _mm_venue = _kimete_mismatch.get("venue_nige_avg", 0) * 100
        _mm_diff  = _kimete_mismatch.get("diff_pct", 0) * 100
        return True, (
            f"[NG]見送り推奨（{_mm_waku}号艇の根拠指標が矛盾）\n"
            f"1着率{_mm_prob:.0f}%最高だが個人逃げ{_mm_indiv:.0f}% vs 場平均{_mm_venue:.0f}%"
            f"（差{_mm_diff:+.0f}pt）\n"
            f"逃げ実績が場平均を大きく下回るのに確率が高い → 指標矛盾"
        )

    return False, ""

