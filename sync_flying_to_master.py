# -*- coding: utf-8 -*-
"""
sync_flying_to_master.py
========================
scrape_flying.py が出力した flying_YYYYMMDD.xlsx を読み込み、
ボートリサーチ_マスタ.xlsx の「選手指数マスタ」シートにある
「FLY数」「FLY経過日数」列を自動更新する。

このスクリプトを実行してからload_race.py を起動することで、
フライングデータが新聞に正しく反映される。

【使い方】
  python scripts/sync_flying_to_master.py               # 当日のflyingファイルを自動検出
  python scripts/sync_flying_to_master.py --date 20260403  # 日付指定 (YYYYMMDD)
  python scripts/sync_flying_to_master.py --flying C:\\path\\to\\flying_20260403.xlsx

【処理内容】
  1. flying_YYYYMMDD.xlsx から選手名・FLY累計数・FLY経過日数を読み取る
  2. ボートリサーチ_マスタ.xlsx「選手指数マスタ」の各選手行を照合
  3. FLY数・FLY経過日数・FLY影響度を更新（FLYなし選手は変更しない）

【FLY経過日数の計算】
  scrape_flying.py は「F1(F後3走)」のような表記を含むため、
  経過日数は「当日を起点に0日」として更新する。
  （出場停止明けの場合は停止期間から逆算するが、
    ここでは0日とすることで最も厳しい判定になる）

【注意】
  - マスタExcelはこのスクリプト実行中は閉じておいてください。
  - バックアップを自動作成します（.bak.xlsx）。
"""

import sys
import pathlib
import argparse
import shutil
import glob
import re
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl がインストールされていません。")
    print("   pip install openpyxl を実行してください。")
    sys.exit(1)

# ============================================================
# パス設定
# ============================================================
BASE_DIR    = pathlib.Path(r"C:\Users\user\Desktop\データ収集")
SCRIPTS_DIR = BASE_DIR / "scripts"
MASTER_PATH = BASE_DIR / "ボートリサーチ_マスタ.xlsx"

# flying ファイルは データ収集/ 直下に出力される（scrape_flying.py のデフォルト）
# ※ scrape_flying.py を scripts/ から実行している場合は SCRIPTS_DIR を参照
FLYING_SEARCH_DIRS = [SCRIPTS_DIR, BASE_DIR]


# ============================================================
# FLY影響度の判定（fill_newspaper.py / load_race.py と同一ロジック）
# ============================================================
def calc_fly_impact_label(fly_count: int, fly_days: int | None) -> str | None:
    """
    FLY数と経過日数から影響度ラベルを返す。
    戻り値: "高" | "中" | "低" | None(FLYなし)
    """
    if fly_count == 0:
        return None
    if fly_days is not None:
        if fly_days >= 180:
            return None   # 影響消滅
        elif fly_days < 90:
            return "高"
        else:
            return "高" if fly_count >= 2 else "中"
    else:
        return "高" if fly_count >= 2 else "中"


def calc_fly_impact_level(fly_count: int, fly_days: int | None) -> str:
    """
    マスタの「FLY影響度」列（大/中/小/なし）に書き込む値を返す。
    """
    label = calc_fly_impact_label(fly_count, fly_days)
    if label is None:
        return "なし"
    if label == "高":
        return "大" if fly_count >= 2 else "中"
    return "小"


# ============================================================
# flying_YYYYMMDD.xlsx を読み込む
# ============================================================
def load_flying_excel(path: pathlib.Path) -> dict:
    """
    flying_YYYYMMDD.xlsx の「フライング一覧」シートを読み込み、
    {正規化選手名: {"fly_total": N, "fly_days": 0}} を返す。

    scrape_flying.py の出力列:
      会場 / レース / 枠 / 選手名 / 登録/級別 / フライング / 合計F数
    """
    print(f"  [読込]  {path.name}")
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        print(f"  [NG]   Excelを開けませんでした: {e}")
        return {}

    # シート名候補: 「フライング一覧」または最初のシート
    sheet_name = "フライング一覧"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        print(f"  [!]   「フライング一覧」シートが見つからないため {sheet_name} を使用")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"  [!]   データが空です")
        return {}

    # ヘッダ行を特定
    header = None
    header_idx = 0
    for i, row in enumerate(rows):
        if row and "選手名" in [str(c) for c in row if c]:
            header = [str(c).strip() if c else "" for c in row]
            header_idx = i
            break

    if header is None:
        print(f"  [!]   ヘッダ行（選手名列）が見つかりません")
        return {}

    # 列インデックス
    try:
        col_name  = header.index("選手名")
        col_total = header.index("合計F数")
    except ValueError as e:
        print(f"  [!]   必要な列が見つかりません: {e}")
        print(f"         ヘッダ: {header}")
        return {}

    # 選手ごとの最大FLY数を集計（同一選手が複数行あれば合算）
    # ★ scrape_flying.pyの「合計F数」はそのレースでの累積F数なので MAX を使う
    player_fly: dict[str, int] = {}

    for row in rows[header_idx + 1:]:
        if not row or not row[col_name]:
            continue
        name = str(row[col_name]).strip()
        if not name:
            continue
        norm = name.replace("　", "").replace(" ", "")
        try:
            fly_total = int(float(row[col_total])) if row[col_total] is not None else 1
        except (TypeError, ValueError):
            fly_total = 1
        # 同一選手は最大値を採用
        player_fly[norm] = max(player_fly.get(norm, 0), fly_total)

    # fly_days は「本日取得 = 出場停止明け直近」として 0 日扱い
    result = {
        norm: {"fly_total": total, "fly_days": 0}
        for norm, total in player_fly.items()
    }

    print(f"  [OK]   {len(result)}名のフライングデータを読み込みました")
    for norm, data in result.items():
        print(f"          {norm}: F{data['fly_total']}回 / 経過{data['fly_days']}日")

    return result


# ============================================================
# マスタ Excel を更新する
# ============================================================
def update_master(master_path: pathlib.Path, fly_data: dict) -> int:
    """
    ボートリサーチ_マスタ.xlsx の「選手指数マスタ」シートを更新する。
    戻り値: 更新した行数
    """
    print()
    print(f"  [マスタ] {master_path.name} を更新します")

    # バックアップ作成
    bak_path = master_path.with_suffix(".bak.xlsx")
    try:
        shutil.copy2(str(master_path), str(bak_path))
        print(f"  [BAK]  バックアップ作成: {bak_path.name}")
    except Exception as e:
        print(f"  [!]   バックアップ失敗（続行します）: {e}")

    # マスタ読み込み
    try:
        wb = openpyxl.load_workbook(str(master_path))
    except Exception as e:
        print(f"  [NG]   マスタを開けませんでした: {e}")
        return 0

    sheet_name = "選手指数マスタ"
    if sheet_name not in wb.sheetnames:
        print(f"  [NG]   「{sheet_name}」シートが見つかりません")
        return 0

    ws = wb[sheet_name]
    rows_list = list(ws.iter_rows())

    # ヘッダ行を検索（B列=「選手名」の行）
    header_row_idx = None
    for i, row in enumerate(rows_list):
        if row[1].value == "選手名":
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f"  [NG]   ヘッダ行（選手名）が見つかりません")
        return 0

    header_cells = rows_list[header_row_idx]
    headers = [str(c.value).strip().replace("\n", "") if c.value else "" for c in header_cells]

    # 必要な列のインデックスを取得（なければ作成）
    def find_or_create_col(col_name: str) -> int:
        """列名を検索し、なければ末尾に追加してインデックスを返す（1始まり）"""
        for j, h in enumerate(headers):
            if h == col_name:
                return j + 1  # openpyxl は1始まり
        # 末尾に新規追加
        new_col = len(headers) + 1
        ws.cell(row=header_row_idx + 1, column=new_col, value=col_name)
        headers.append(col_name)
        print(f"  [NEW]  列を追加しました: {col_name} (列{new_col})")
        return new_col

    col_fly_count  = find_or_create_col("FLY数")
    col_fly_days   = find_or_create_col("FLY経過日数")
    col_fly_impact = find_or_create_col("FLY影響度")

    # 名前列のインデックス（B列=2）
    col_name_idx = 2  # B列固定

    # 各選手行を更新
    updated = 0
    not_found = []

    for i in range(header_row_idx + 1, len(rows_list)):
        row = rows_list[i]
        name_cell = row[col_name_idx - 1]  # 0始まり
        if not name_cell.value:
            continue
        norm = str(name_cell.value).replace("　", "").replace(" ", "").strip()

        # 完全一致 → 部分一致の順で検索
        data = fly_data.get(norm)
        if data is None:
            for key, val in fly_data.items():
                if norm and (norm in key or key in norm):
                    data = val
                    break

        if data is None:
            continue  # この選手のFLYデータなし → 更新しない

        fly_total = data["fly_total"]
        fly_days  = data["fly_days"]
        impact    = calc_fly_impact_level(fly_total, fly_days)

        excel_row = i + 1  # openpyxl は1始まり
        ws.cell(row=excel_row, column=col_fly_count,  value=fly_total)
        ws.cell(row=excel_row, column=col_fly_days,   value=fly_days)
        ws.cell(row=excel_row, column=col_fly_impact, value=impact)

        print(f"  [更新]  {str(name_cell.value).strip()}: "
              f"F{fly_total}回 / {fly_days}日 / 影響度={impact}")
        updated += 1

    # 一致しなかった選手をリスト
    master_names = set()
    for i in range(header_row_idx + 1, len(rows_list)):
        v = rows_list[i][col_name_idx - 1].value
        if v:
            master_names.add(str(v).replace("　", "").replace(" ", "").strip())

    for norm in fly_data:
        if norm not in master_names:
            # 部分一致も確認
            matched = any(norm in m or m in norm for m in master_names if norm)
            if not matched:
                not_found.append(norm)

    if not_found:
        print()
        print(f"  [!]   マスタに存在しない選手（新規登録が必要かもしれません）:")
        for n in not_found:
            print(f"          {n}: F{fly_data[n]['fly_total']}回")

    # 保存
    try:
        wb.save(str(master_path))
        print()
        print(f"  [OK]   マスタを保存しました ({updated}名を更新)")
    except PermissionError:
        print(f"  [NG]   マスタExcelが開いています。閉じてから再実行してください。")
        return 0
    except Exception as e:
        print(f"  [NG]   保存エラー: {e}")
        return 0

    return updated


# ============================================================
# flying ファイルの自動検出
# ============================================================
def find_flying_file(date_str: str | None) -> pathlib.Path | None:
    """
    指定日付（YYYYMMDD）または当日の flying_*.xlsx を検索する。
    複数見つかった場合は最新のものを返す。
    """
    if date_str:
        patterns = [f"flying_{date_str}.xlsx"]
    else:
        today = datetime.now().strftime("%Y%m%d")
        patterns = [f"flying_{today}.xlsx", "flying_*.xlsx"]

    for search_dir in FLYING_SEARCH_DIRS:
        for pattern in patterns:
            matches = sorted(glob.glob(str(search_dir / pattern)), reverse=True)
            if matches:
                return pathlib.Path(matches[0])

    return None


# ============================================================
# メイン
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="フライングデータをマスタに反映")
    parser.add_argument(
        "--date", type=str, default=None,
        help="flying ファイルの日付 (例: 20260403)。省略時は当日"
    )
    parser.add_argument(
        "--flying", type=str, default=None,
        help="flying_*.xlsx のパスを直接指定"
    )
    parser.add_argument(
        "--master", type=str, default=None,
        help="マスタExcelのパスを直接指定（省略時は既定パス）"
    )
    args = parser.parse_args()

    print("=" * 55)
    print("  フライングデータ → マスタ同期ツール")
    print("=" * 55)
    print()

    # flying ファイルを特定
    if args.flying:
        flying_path = pathlib.Path(args.flying)
        if not flying_path.exists():
            print(f"❌ ファイルが見つかりません: {flying_path}")
            sys.exit(1)
    else:
        flying_path = find_flying_file(args.date)
        if flying_path is None:
            ds = args.date or datetime.now().strftime("%Y%m%d")
            print(f"❌ flying_{ds}.xlsx が見つかりません。")
            print(f"   先に scrape_flying.py を実行してください。")
            print(f"   検索フォルダ: {[str(d) for d in FLYING_SEARCH_DIRS]}")
            sys.exit(1)

    print(f"  対象フライングファイル: {flying_path}")

    # マスタパスを決定
    master_path = pathlib.Path(args.master) if args.master else MASTER_PATH
    if not master_path.exists():
        print(f"❌ マスタExcelが見つかりません: {master_path}")
        sys.exit(1)

    print(f"  マスタ             : {master_path}")
    print()

    # flying データ読み込み
    fly_data = load_flying_excel(flying_path)
    if not fly_data:
        print("⚠️  フライングデータが空です。処理を終了します。")
        sys.exit(0)

    # マスタ更新
    updated = update_master(master_path, fly_data)

    print()
    print("=" * 55)
    if updated > 0:
        print(f"✅ 完了！ {updated}名のFLYデータをマスタに反映しました")
        print(f"   次に load_race.py を実行してください")
    else:
        print("⚠️  更新対象の選手が見つかりませんでした")
        print("   選手名の表記がマスタと一致しているか確認してください")
    print("=" * 55)


if __name__ == "__main__":
    main()
