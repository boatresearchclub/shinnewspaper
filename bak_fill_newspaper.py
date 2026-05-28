# -*- coding: utf-8 -*-
"""
fill_newspaper.py
=================
load_race.py 実行後に ボートリサーチ新聞_軽量版.xlsx の「{会場}_数値」シートを読み取り、
新聞出力テンプレート（新聞出力.xlsx）に転記する。
マスタデータ（選手指数・会場統計・イン逃げ分析）は ボートリサーチ_マスタ.xlsx から参照する。

xlwings を使用するため、図形・外部リンク・書式を一切壊さない。
※ Windows 上で Excel がインストールされている環境が必要。

【前提】
  先に load_race.py を実行して「{会場}_数値」シートが作成されていること。

【配置場所】
  scripts/fill_newspaper.py  ← load_race.py と同じフォルダ

【テンプレートの置き場所】
  プロジェクトルート/新聞出力.xlsx

【使い方】
  # 引数なし → 対話式で入力
  python scripts/fill_newspaper.py

  # 会場のみ指定（全レース）
  python scripts/fill_newspaper.py --venue 大村

  # レース番号も指定
  python scripts/fill_newspaper.py --venue 大村 --race 5

【出力先】
  プロジェクトルート/新聞出力_filled.xlsx
  レースが複数の場合はシートをコピーして 大村_1R, 大村_2R ... と追加する。
  ※テンプレートの書式・フォント・色・結合セル・図形は一切変更しない。値のみ上書き。
"""

import os
import sys
import re
import argparse
import pathlib
import shutil

try:
    import xlwings as xw
except ImportError:
    print("❌ xlwings がインストールされていません。")
    print("   pip install xlwings を実行してください。")
    sys.exit(1)

import openpyxl  # 数値シートの読み取りのみに使用
import io
import zipfile
import tempfile
from PIL import Image, ImageDraw, ImageFont

# ============================================================
# パス設定
# ============================================================
BASE_DIR      = pathlib.Path(__file__).parent.parent
TEMPLATE_PATH = BASE_DIR / "新聞出力.xlsx"
OUTPUT_PATH   = BASE_DIR / "新聞出力_filled.xlsx"
MASTER_PATH   = BASE_DIR / "ボートリサーチ_マスタ.xlsx"
NUM_PATH      = BASE_DIR / "ボートリサーチ新聞_軽量版.xlsx"


# ============================================================
# 選手指数マスタからコース別ST順位を取得
# ============================================================
def load_st_rank_master(wb):
    """
    選手指数マスタシートから全選手のコース別ST順位を読み込む。

    戻り値:
        {
            "選手名": {1: 2.72, 2: 3.60, 3: 4.08, 4: 3.72, 5: 4.11, 6: 4.30},
            ...
        }
        ※ キーはコース番号(1〜6)、値はST順位(小さいほど早い)
    """
    sheet_name = "選手指数マスタ"
    if sheet_name not in wb.sheetnames:
        print(f"⚠️  シートが見つかりません: {sheet_name} → ST順位なしで処理")
        return {}

    ws = wb[sheet_name]
    # 2行目がヘッダー行（A=登録番号, B=選手名, U〜Z=ST順位コース1〜6）
    master = {}
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 2).value  # B列: 選手名
        if not name:
            continue
        # スペース（全角・半角）をすべて除去した正規化名をキーにする
        norm_name = str(name).replace("\u3000", "").replace(" ", "").strip()
        st_ranks = {}
        for course, col in enumerate(range(21, 27), start=1):  # U=21〜Z=26
            v = ws.cell(r, col).value
            if v is not None:
                try:
                    st_ranks[course] = float(v)
                except (TypeError, ValueError):
                    pass
        if st_ranks:
            master[norm_name] = st_ranks
    return master


# ============================================================
# 会場統計（決まり手場平均・場平均1着率）を読み込む
# ============================================================
def load_venue_stats(wb, venue):
    """
    会場統計シートから指定会場の決まり手場平均・1コース1着率を返す。
    戻り値:
        {
            "in_rate":    0.612,          # イン逃げ率（小数）
            "kimari_avg": {"差し": 0.12, "まくり": 0.15, "まくり差し": 0.13},
            "win1_rate":  0.612,          # 場平均1コース1着率
        }
    """
    if "会場統計" not in wb.sheetnames:
        return {}
    ws = wb["会場統計"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == "会場名":
            header_idx = i
            break
    if header_idx is None:
        return {}
    headers = rows[header_idx]
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        d = dict(zip(headers, row))
        if str(d.get("会場名", "")).strip() != venue:
            continue
        def sf(v):
            try:
                return float(str(v).replace("%", "").strip())
            except Exception:
                return None
        in_rate = sf(d.get("イン逃げ率"))
        # まくり差し率はヘッダに改行が入る場合あり
        makuri_sashi = sf(d.get("まくり差し率") or d.get("まくり\n差し率"))
        # コース別1着率（1〜6コース）— マスタに "{c}コース1着率" で格納されている
        course_win_rates = {}
        for c in range(1, 7):
            raw = (d.get(f"{c}コース1着率") or d.get(f"{c}C\n1着率") or d.get(f"{c}C_1着率"))
            v = sf(raw)
            if v is not None:
                course_win_rates[c] = v
        win1_rate = course_win_rates.get(1)  # 後方互換: 1コース1着率
        return {
            "in_rate":           in_rate,
            "kimari_avg": {
                "差し":          sf(d.get("差し率")),
                "まくり":        sf(d.get("まくり率")),
                "まくり差し":    makuri_sashi,
            },
            "win1_rate":         win1_rate,
            "course_win_rates":  course_win_rates,  # {1: 0.558, 2: 0.12, ..., 6: 0.05}
        }
    return {}


# ============================================================
# FLY影響度をマスタから直接取得する
# ============================================================
def calc_fly_impact(wb_master, name, course=None, avg_st_alltime=None):
    """
    【刷新 v2】ST標準偏差による代替推定を廃止。
    update_master.py が集計した「FLY影響度」列（大/中/小/なし）を
    選手指数マスタから直接読んで返す。

    FLYなし選手は load_fly_label() 側で空白にするため、
    この関数は FLYあり選手の手入力ケースでのみ呼ばれる。

    戻り値:
        "大" | "中" | "小" | "なし" | None（データなし → 呼び出し元で "中" フォールバック）

    引数 course / avg_st_alltime は後方互換のため残存（未使用）。
    """
    if wb_master is None:
        return None
    sheet_name = "選手指数マスタ"
    if sheet_name not in wb_master.sheetnames:
        return None

    ws = wb_master[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(rows):
        if row and row[1] == "選手名":
            header_idx = i
            break
    if header_idx is None:
        return None

    headers = rows[header_idx]
    col_idx = {str(h).strip().replace("\n", ""): j
               for j, h in enumerate(headers) if h is not None}

    # 「FLY影響度」列を探す
    impact_col = col_idx.get("FLY影響度")
    if impact_col is None:
        return None

    norm_name = str(name).replace("　", "").replace(" ", "").strip()
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        row_name = str(row[1]).replace("　", "").replace(" ", "").strip() if row[1] else ""
        if row_name != norm_name:
            continue
        v = row[impact_col]
        if v is None or str(v).strip() in ("", "nan", "None"):
            return None
        return str(v).strip()   # "大" | "中" | "小" | "なし"

    return None


# ============================================================
# イン逃げ分析シートから枠別2着率（1-○）を読み込む
# ============================================================
def load_ininage_stats(wb, venue):
    """
    イン逃げ分析シートから指定会場の枠別2着率を返す。
    戻り値: {"1": 32.7, "2": 32.7, "3": 33.1, "4": 19.6, "5": 11.1, "6": 3.5}
    """
    sheet_name = "イン逃げ分析"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == "会場名":
            header_idx = i
            break
    if header_idx is None:
        return {}
    headers = rows[header_idx]
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        d = dict(zip(headers, row))
        if str(d.get("会場名", "")).strip() != venue:
            continue
        result = {}
        for waku_no in range(1, 7):
            # load_race.py と同じキー名: "{N}枠\n2着率" (実際の改行文字)
            col_key = f"{waku_no}枠" + "\n" + "2着率"  # 実際の改行文字
            val = d.get(col_key)
            if val is not None:
                try:
                    result[str(waku_no)] = float(val)
                except Exception:
                    pass
        return result
    return {}



# ============================================================
# 選手指数マスタから FLY数・FLY経過日数を読み込み、F/ST影響ラベルを返す
# ============================================================
def load_fly_label(wb, players):
    """
    選手指数マスタから各艇の FLY数・FLY経過日数 を取得し、
    FLY影響ラベルを返す。

    【修正②】FLY経過日数を load_race.py と同じロジックで判定。
    旧実装: FLY数≥2→高、≥1→中 のみ（経過日数を全く無視）
    新実装: FLY経過日数を加味し、以下のルールを適用する。
      FLY数=0                        → 低（FLYなし）
      FLY数≥1 かつ 経過日数 ≥ 180日  → 低（影響消滅）
      FLY数≥1 かつ 経過日数 <  90日  → 高（出場停止明け直後）
      FLY数≥1 かつ 90≤経過日数<180日 → 高(FLY≥2) / 中(FLY=1)
      FLY経過日数が取得できない場合   → 旧来判定にフォールバック

    戻り値: {waku(1〜6): {"count": int, "days": int or None, "label": "高"|"中"|"低"|None}}
    """
    sheet_name = "選手指数マスタ"
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # ヘッダ行を検出（B列="選手名"）
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[1] == "選手名":
            header_idx = i
            break
    if header_idx is None:
        return {}

    headers = rows[header_idx]

    # 列インデックスを特定（FLY数・FLY経過日数・FLY影響度）
    fly_col    = None
    days_col   = None
    impact_col = None
    for j, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip().replace("\n", "")
        if h_str == "FLY数":
            fly_col = j
        elif h_str in ("FLY経過日数", "FLY経過日数"):
            days_col = j
        elif h_str == "FLY影響度":
            impact_col = j

    if fly_col is None:
        return {}

    # 選手名 → (FLY数, FLY経過日数, FLY影響度) の辞書を構築
    fly_map = {}  # {正規化選手名: (fly_count, fly_days, impact)}
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        name = str(row[1]).strip() if row[1] else ""
        if not name:
            continue
        norm = name.replace("　", "").replace(" ", "").strip()
        try:
            fly = int(float(row[fly_col])) if row[fly_col] is not None else 0
        except (TypeError, ValueError):
            fly = 0
        days = None
        if days_col is not None:
            try:
                days_raw = row[days_col]
                if days_raw is not None and str(days_raw).strip() not in ("", "None", "nan"):
                    days = int(float(days_raw))
            except (TypeError, ValueError):
                days = None
        # FLY影響度（大/中/小/なし）：FLYなし選手はNaN→None
        impact = None
        if impact_col is not None:
            v = row[impact_col]
            if v is not None and str(v).strip() not in ("", "nan", "None"):
                impact = str(v).strip()
        fly_map[norm] = (fly, days, impact)

    def _fly_label_from(fly_count, fly_days):
        """load_race.py と同一の判定ロジック"""
        if fly_count == 0:
            return None   # FLYなし → 空白
        if fly_days is not None:
            if fly_days >= 180:
                return None   # 影響消滅 → 空白
            elif fly_days < 90:
                return "高"
            else:
                return "高" if fly_count >= 2 else "中"
        else:
            return "高" if fly_count >= 2 else "中"

    # 各艇の選手名で照合 → {waku: {"count": N, "days": D, "label": ..., "impact": ...}}
    result = {}
    for p in players:
        waku = p["waku"]
        norm_name = str(p.get("name", "")).replace("　", "").replace(" ", "").strip()
        entry = fly_map.get(norm_name)
        if entry is None:
            # 部分一致フォールバック
            for k, v in fly_map.items():
                if norm_name and (norm_name in k or k in norm_name):
                    entry = v
                    break
        if entry is None:
            result[waku] = {"count": None, "days": None, "label": None, "impact": None}
        else:
            fly_count, fly_days, impact = entry
            label = _fly_label_from(fly_count, fly_days)
            result[waku] = {
                "count":  fly_count,
                "days":   fly_days,
                "label":  label,
                "impact": impact,   # "大"|"中"|"小"|"なし"|None
            }
    return result



def get_player_st_ranks(players, st_master):
    """
    各艇の選手名と想定コースから、コース別ST順位を返す。

    戻り値:
        {waku(1〜6): st_rank(float) or None}
    """
    result = {}
    for p in players:
        waku = p["waku"]
        name = str(p.get("name", "")).strip()
        # スペース（全角・半角）除去した正規化名で照合
        norm_name = name.replace("\u3000", "").replace(" ", "")
        course = p.get("course")  # 想定コース

        st_ranks = st_master.get(norm_name)
        if st_ranks and course is not None:
            result[waku] = st_ranks.get(int(course))
        else:
            # 正規化名でも一致しない場合、部分一致で探す
            matched = None
            for master_name, ranks in st_master.items():
                if norm_name and (norm_name in master_name or master_name in norm_name):
                    matched = ranks
                    break
            if matched and course is not None:
                result[waku] = matched.get(int(course))
            else:
                result[waku] = None
    return result


# ============================================================
# スリット図生成
# ============================================================
# 艇カラー定義（艇番 → 塗り色, 文字色, 枠色）
BOAT_COLORS = {
    1: {"fill": (255, 255, 255, 255), "text": (0, 0, 0, 255),   "border": (100, 100, 100, 255)},
    2: {"fill": (26,  26,  26,  255), "text": (255, 255, 255, 255), "border": (0, 0, 0, 255)},
    3: {"fill": (204,  0,   0,  255), "text": (255, 255, 255, 255), "border": (153, 0, 0, 255)},
    4: {"fill": (34,  85,  204, 255), "text": (255, 255, 255, 255), "border": (17, 68, 170, 255)},
    5: {"fill": (221, 204,  0,  255), "text": (0, 0, 0, 255),   "border": (180, 160, 0, 255)},
    6: {"fill": (17,  85,  17,  255), "text": (255, 255, 255, 255), "border": (0, 50, 0, 255)},
}

# 画像全体のサイズ（元画像に合わせる）
SLIT_W = 208
SLIT_H = 500
BOAT_H = 48        # 艇の高さ(px)
BOAT_W = 118       # 艇の標準幅(px)（テンプレート実測値）
BOAT_SPACING = 83  # 艇間のY間隔(px)
BOAT_START_Y = 17  # 最初の艇のY開始位置
# X方向: 早い(順位小)=右、遅い(順位大)=左
# 基準X（ST順位=3.5のときの艇左端）= 画像中央
BASE_X = (SLIT_W - BOAT_W) // 2   # = 45px
BASE_RANK = 3.5
# 0.5順位差 = 艇幅の1/4(約30px) → 1順位差 = 59px
# SLIT_W=208での有効rank範囲: 2.77〜4.23
PX_PER_RANK = BOAT_W / 4 / 0.5    # = 59px/順位
X_MIN = 2                           # 艇左端の最小値
X_MAX = SLIT_W - BOAT_W - 2        # 艇左端の最大値 = 88px


def draw_boat_shape(draw, x, y, w, h, colors, number):
    """
    テンプレートに合わせた七角形ボート（後端六角形カット＋右矢印先端）を描画。
    x, y: 左上座標  w, h: 幅・高さ

    テンプレート実測(h=48, w=118)から導いた形状:
      後端テーパー: 上下それぞれ h*0.271 行で w*0.178 px 削る
      先端(右):     y中心で最大幅(w-1), 上下端でw*0.712幅
    """
    cy = h // 2                       # 中心y相対値 = 24
    tp = round(h * 0.271)             # 後端テーパー行数 = 13
    lx = round(w * 0.178)             # 後端水平テーパー量 = 21
    rx_tip = w - 1                    # 先端の最大右端(相対) = 117
    rx_base = round(w * 0.712)        # 先端の上下端右位置(相対) = 84

    points = [
        (x + lx,      y),             # 後端 上隅
        (x,           y + tp),        # 後端 中央上
        (x,           y + h - tp),    # 後端 中央下
        (x + lx,      y + h),         # 後端 下隅
        (x + rx_base, y + h),         # 先端側 下
        (x + rx_tip,  y + cy),        # 先端 頂点
        (x + rx_base, y),             # 先端側 上
    ]
    draw.polygon(points, fill=colors["fill"], outline=colors["border"])

    # 艇番号テキスト
    font_size = int(h * 0.55)
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
    except Exception:
        font = ImageFont.load_default()

    text = str(number)
    bbox = draw.textbbox((0, 0), text, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    tx = x + lx + (rx_base - lx) // 2 - tw // 2
    ty = y + h // 2 - th // 2 - bbox[1]
    draw.text((tx, ty), text, fill=colors["text"], font=font)


def generate_slit_image(st_ranks_by_waku):
    """
    各艇のST順位に基づいてスリット図PNGを生成する。

    st_ranks_by_waku: {waku(1〜6): st_rank(float) or None}
    戻り値: PNGバイト列
    """
    img = Image.new("RGBA", (SLIT_W, SLIT_H), (255, 255, 255, 255))
    draw = ImageDraw.Draw(img)

    # ST順位が取得できない艇は平均順位(3.5)とする
    ranks = {}
    for waku in range(1, 7):
        r = st_ranks_by_waku.get(waku)
        ranks[waku] = r if r is not None else BASE_RANK

    for waku in range(1, 7):
        boat_y = BOAT_START_Y + (waku - 1) * BOAT_SPACING
        rank = ranks[waku]
        shift = (BASE_RANK - rank) * PX_PER_RANK
        boat_x = max(X_MIN, min(int(BASE_X + shift), X_MAX))

        colors = BOAT_COLORS[waku]
        draw_boat_shape(draw, boat_x, boat_y, BOAT_W, BOAT_H, colors, waku)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def replace_slit_images_in_xlsx(xlsx_path, png_bytes_by_sheet_idx):
    """
    xlsxファイル内の各シートのスリット図を、シートごとの個別PNG画像で差し替える。

    png_bytes_by_sheet_idx: {sheet_idx(1始まり): png_bytes}
        sheet_idx=1 → xl/media/image1.png（既存）を上書き
        sheet_idx=2以降 → xl/media/imageN_slit.png を新規追加し、
                          対応する drawing の relsを書き換える

    各シートは sheetN.xml → drawingN.xml → 画像 という構造を持つ前提。
    """
    import re as _re

    tmp_path = str(xlsx_path) + ".tmp_slit"

    with zipfile.ZipFile(str(xlsx_path), 'r') as zin:
        # drawing1.xml.rels の rId を取得（スリット画像のrId）
        drawing1_rels_xml = zin.read("xl/drawings/_rels/drawing1.xml.rels").decode("utf-8")
        # image1.png を参照している rId を特定
        m = _re.search(r'Id="([^"]*)"[^>]*Target="[^"]*image1\.png"', drawing1_rels_xml)
        slit_rid = m.group(1) if m else "rId2"  # デフォルトrId2

        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                fname = item.filename

                # sheet_idx=1 のスリット画像（image1.png）は上書き
                if fname == "xl/media/image1.png":
                    data = png_bytes_by_sheet_idx.get(1, zin.read(fname))
                    zout.writestr(item, data)

                # drawingN.xml.rels（N>=2）: スリット画像の参照先を imageN_slit.png に書き換え
                elif _re.match(r"xl/drawings/_rels/drawing(\d+)\.xml\.rels", fname):
                    n = int(_re.match(r"xl/drawings/_rels/drawing(\d+)\.xml\.rels", fname).group(1))
                    if n >= 2 and n in png_bytes_by_sheet_idx:
                        new_img = f"../media/image{n}_slit.png"
                        xml = zin.read(fname).decode("utf-8")
                        xml = _re.sub(
                            r'(Id="{rid}"[^>]*Target=")[^"]*image1\.png"'.format(rid=slit_rid),
                            r'\g<1>' + new_img + '"',
                            xml
                        )
                        zout.writestr(item, xml.encode("utf-8"))
                    else:
                        zout.writestr(item, zin.read(fname))

                else:
                    zout.writestr(item, zin.read(fname))

            # imageN_slit.png を新規追加（N>=2）
            for sheet_idx, png_bytes in png_bytes_by_sheet_idx.items():
                if sheet_idx >= 2:
                    img_path = f"xl/media/image{sheet_idx}_slit.png"
                    zout.writestr(img_path, png_bytes)

    import os
    os.replace(tmp_path, str(xlsx_path))


# ============================================================
# 数値シートに FLY入力行がなければ追記する
# ============================================================
def ensure_fly_input_rows(wb, venue):
    """
    {venue}_数値 シートの「選手情報_想定コース」ブロック直後に
    FLY入力行（艇番1〜6）が存在しなければ追記する。
    ユーザーが各レース列に手動でフライング数を入力する。
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    sheet_name = f"{venue}_数値"
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]

    # iter_rowsで一括読み込み（ws.cell個別アクセスより高速）
    all_rows_fly = list(ws.iter_rows(values_only=True))

    # 既にFLY入力行があればスキップ
    if any(row[0] == "FLY入力" for row in all_rows_fly if row and row[0] is not None):
        return

    # 「想定コース」ブロックの艇番6行を探す
    insert_after_row = None
    in_course_block  = False
    for r_idx, row_vals in enumerate(all_rows_fly):
        r = r_idx + 1
        a = str(row_vals[0] if len(row_vals) > 0 and row_vals[0] is not None else "").strip()
        b = str(row_vals[1] if len(row_vals) > 1 and row_vals[1] is not None else "").strip()
        c = str(row_vals[2] if len(row_vals) > 2 and row_vals[2] is not None else "").strip()
        if b == "想定コース":
            in_course_block = True
        if in_course_block and c == "6":
            insert_after_row = r
            break

    if insert_after_row is None:
        print(f"  ⚠️  FLY入力行の挿入位置が見つかりませんでした")
        return

    insert_row = insert_after_row + 1
    ws.insert_rows(insert_row, amount=6)

    # レース列数を確認（2行目 = index1 のcol4以降）
    header_vals = all_rows_fly[1] if len(all_rows_fly) >= 2 else []
    n_race_cols = 0
    for col_idx in range(3, len(header_vals)):  # col4以降(0-indexed: 3〜)
        if header_vals[col_idx] is not None:
            n_race_cols += 1
        else:
            break

    fill_red    = PatternFill("solid", fgColor="FFCC0000")
    fill_yellow = PatternFill("solid", fgColor="FFFFFF00")
    fill_boat   = {1:"FFFFFFFF",2:"FF1A1A1A",3:"FFCC0000",4:"FF2255CC",5:"FFDDCC00",6:"FF115511"}
    fill_font   = {1:"FF000000",2:"FFFFFFFF",3:"FFFFFFFF",4:"FFFFFFFF",5:"FF000000",6:"FFFFFFFF"}
    thin = Side(style="thin", color="FFCCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    fn_w = Font(name="Meiryo UI", size=9, bold=True,  color="FFFFFFFF")
    fn_b = Font(name="Meiryo UI", size=9, bold=False, color="FF000000")
    al_c = Alignment(horizontal="center", vertical="center")

    for i, waku_no in enumerate(range(1, 7)):
        r = insert_row + i
        c = ws.cell(r, 1)
        c.value, c.fill, c.font, c.alignment, c.border = "FLY入力", fill_red, fn_w, al_c, bdr
        c = ws.cell(r, 2)
        c.value, c.fill, c.font, c.alignment, c.border = "FLY（1=あり）", fill_yellow, fn_b, al_c, bdr
        c = ws.cell(r, 3)
        c.value = waku_no
        c.fill  = PatternFill("solid", fgColor=fill_boat[waku_no])
        c.font  = Font(name="Meiryo UI", size=9, bold=True, color=fill_font[waku_no])
        c.alignment, c.border = al_c, bdr
        for col in range(4, 4 + n_race_cols):
            c = ws.cell(r, col)
            c.value, c.fill, c.border = None, fill_yellow, bdr
        ws.row_dimensions[r].height = 15.0

    print(f"  ✅ {sheet_name} に FLY入力行を追加しました（Row{insert_row}〜{insert_row+5}）")


# ============================================================
# 数値シート読み取り
# ============================================================
def read_numeric_sheet(wb, venue):
    """
    {venue}_数値 シートからレースデータを辞書形式で読み取る。

    戻り値:
        {
            "1": {  # レース番号（文字列）
                "deadline": "15:21",
                "rank": "S",
                "players": [
                    {
                        "waku": 1,
                        "name": "山田太郎",
                        "kumi": "A1",
                        "motor2": 37.80,
                        "honmei": "◎",
                        "rel_win1": 60.7,    # オリジナル1着率(%)
                        "abs_win3": 89.3,    # コース別3連対率(%)
                        "circle_pct": 46.9,  # 2着優位度(%)
                        "idx3": 100,         # 3着指数
                        "kimari": {"逃げ": "61%", "差し": "4%", ...},
                    }, ...
                ]
            }, ...
        }
    """
    sheet_name = f"{venue}_数値"
    if sheet_name not in wb.sheetnames:
        print(f"❌ シートが見つかりません: {sheet_name}")
        print(f"   先に load_race.py を実行してください。")
        return None

    ws = wb[sheet_name]

    # ── 列マッピング（2行目からレース番号を取得）─────────────────────
    # ── iter_rows で全セルを一括読み込み（ws.cell個別アクセスより大幅に高速）──
    all_rows = list(ws.iter_rows(values_only=True))

    # 列マッピング（2行目 = index1 からレース番号を取得）
    race_col_map = {}
    if len(all_rows) >= 2:
        for c_idx, val in enumerate(all_rows[1]):
            if val and isinstance(val, str) and "R" in val:
                m = re.match(r"(\d+)R\s*[\n\r]*(.*)", val.strip())
                if m:
                    race_col_map[m.group(1)] = {"col": c_idx + 1, "deadline": m.group(2).strip()}

    if not race_col_map:
        print(f"❌ {sheet_name} にレース列が見つかりません")
        return None

    # セルキャッシュ: (row_1indexed, col_1indexed) → value
    cell_cache = {}
    for r_idx, row_vals in enumerate(all_rows):
        r = r_idx + 1
        for c_idx, v in enumerate(row_vals):
            if v is not None:
                cell_cache[(r, c_idx + 1)] = v

    # 行マッピング・ランク行を iter_rows の結果から一括構築
    row_map       = {}
    current_label = None
    rank_row      = None
    step2_row     = None   # Step2: 展開シナリオ行（ランク・点数を取得）
    _SKIP_PREFIXES = ("▼", "━", "★", "※", "⚠", "【", "候補", "展示前",
                      "全速型", "標準型", "万舟型", "狙い撃ち型", "3択判定")
    for r_idx, row_vals in enumerate(all_rows):
        r    = r_idx + 1
        cat  = row_vals[0] if len(row_vals) > 0 else None
        item = row_vals[1] if len(row_vals) > 1 else None
        waku = row_vals[2] if len(row_vals) > 2 else None

        if cat and str(cat) not in ("分類",) and not str(cat).startswith(_SKIP_PREFIXES):
            current_label = f"{cat}_{item}" if item else str(cat)

        if current_label and waku is not None:
            waku_str = str(waku).strip()
            if waku_str.isdigit():
                row_map.setdefault(current_label, {})[waku_str] = r

        if rank_row is None and cat is not None:
            c1 = str(cat)
            c2 = str(item or "")
            if (c1 == "判定" and c2.startswith("ランク")) or c1 == "①判定スコア" or "堅実度" in c1:
                rank_row = r

        # Step2行（A列またはB列に「Step2」を含む）
        if step2_row is None and cat is not None:
            if "Step2" in str(cat) or (item and "Step2" in str(item)):
                step2_row = r
    # ── 各レースのデータを組み立て ────────────────────────────────────
    def pct_to_float(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace("%", "").strip()
        try:
            return float(s)
        except ValueError:
            return None

    def get_val(label, waku, col):
        r = row_map.get(label, {}).get(str(waku))
        if r is None:
            return None
        return cell_cache.get((r, col))  # cell_cacheで高速ルックアップ

    races = {}
    for race_no, rc in race_col_map.items():
        col      = rc["col"]
        deadline = rc["deadline"]

        # ランク・スコア：rank_row（①判定スコア行）を優先、次にstep2_row
        # ①判定スコア行のD列: 「【ランクS】75点\n戦略: ...\n信頼度: ...%」形式
        rank  = "-"
        score = None
        _score_rows = []
        if rank_row:
            _score_rows.append(rank_row)
        if step2_row:
            _score_rows.append(step2_row)
        for _sr in _score_rows:
            raw = str(cell_cache.get((_sr, col), "") or "")
            if not raw:
                raw = str(ws.cell(_sr, col).value or "")
            m = re.search(r"【ランク(\w+)】", raw)
            if not m:
                m = re.search(r"ランク[：:]\s*([A-Za-z]+)", raw)
            if not m:
                m = re.search(r"堅実\s*([A-Za-zA-Z]+)", raw)
            if m:
                rank = m.group(1)
            # 点数: 「52点(+15) / 堅実B」形式 or 「75点」形式
            ms = re.search(r"(?<![%\d])(\d{2,3})点", raw)
            if ms:
                try:
                    score = int(ms.group(1))
                except ValueError:
                    pass
            if rank != "-" or score is not None:
                break

        # 選手データ
        players = []
        for waku in range(1, 7):
            players.append({
                "waku":       waku,
                "name":       get_val("選手情報_選手名",    waku, col) or "",
                "kumi":       get_val("選手情報_組",        waku, col) or "",
                "motor2":     get_val("選手情報_モータ2連", waku, col),
                "honmei":     get_val("選手情報_攻め記号",  waku, col) or "",
                "course":     get_val("選手情報_想定コース", waku, col),
                "rel_win1":   pct_to_float(get_val("数値指標_オリジナル1着率(%)",              waku, col)),
                "abs_win3":   pct_to_float(get_val("数値指標_コース別3連対率(%)",              waku, col)),
                "circle_pct": pct_to_float(get_val("数値指標_2着優位度(%)[相対・イン逃げ時]", waku, col)),
                "idx3":       get_val("数値指標_3着指数", waku, col),
                "jizen": {
                    "in_nige":  get_val("事前評価_①逃げ",   waku, col),
                    "aisho":    get_val("事前評価_②相性",   waku, col),
                    "kiryoku":  get_val("事前評価_③機力",   waku, col),
                    "tenkai":   get_val("事前評価_④展開",   waku, col),
                    "jizaisei": get_val("事前評価_⑤S安定",  waku, col),
                },
                "kimari": {
                    key.replace("%", ""): str(v) if "%" in str(v) else f"{v}"
                    for key in ["逃げ%", "差し%", "まくり%", "まくり差し%"]
                    if (v := get_val(f"決まり手_{key}", waku, col)) is not None
                    and str(v) not in ("-", "", "None")
                },
                # FLY手動入力（FLY入力_FLY（1=あり）に 1 が入っていたら対象）
                "fly_input": get_val("FLY入力_FLY（1=あり）", waku, col),
                # コース別全期間平均ST（マスタ由来、数値シート経由で渡されていない場合はNone）
                "avg_st":    get_val("選手情報_コース別平均ST", waku, col),
            })


        # ── 買い目リスト・理想合成オッズ・判定スコアをcell_cacheから取得 ──
        # cell_cacheはiter_rows一括読み込み済みなのでws.cell不要（高速）
        bet_list_raw    = None
        theory_syn_odds = None
        ev_warning      = False
        strategy        = None
        trust_pct       = None
        kosatsu_raw     = None   # ⑧考察の結論（生テキスト）
        tenkai_raw      = None   # ④展開予測（生テキスト）
        grade_raw       = None   # 参加グレード（bet_list_raw から抽出）
        buy_honsen_raw  = None   # 主役展開（本線・押さえ買い目テキスト）

        # col2キャッシュ（item列）を事前に辞書化してrow→v2のルックアップを高速化
        max_r = len(all_rows)

        # ── 主役展開行を「Step1」含む行の次行として取得 ─────────────────
        # A列に「Step1」を含む行を探し、その次の行のD列（col）の値を buy_honsen_raw に格納
        # また同じブロック内でD列テキストが【展開考察】を含む行を kosatsu_raw に格納
        for _r in range(1, max_r + 1):
            _v1 = str(cell_cache.get((_r, 1), "") or "")
            _v2 = str(cell_cache.get((_r, 2), "") or "")
            _vc = str(cell_cache.get((_r, col), "") or "")
            # Step1行: A列またはB列に "Step1" を含む
            if "Step1" in _v1 or "Step1" in _v2:
                # この行自体のD列が展開考察テキスト（【展開考察】含む）
                if "展開考察" in _vc or "逃げるなら" in _vc or "まくりなら" in _vc:
                    if not kosatsu_raw:
                        kosatsu_raw = _vc
                # 次行以降を走査して主役展開テキストを取得
                for _nr in range(_r + 1, min(_r + 10, max_r + 1)):
                    _nv2 = str(cell_cache.get((_nr, 2), "") or "")
                    _nvc = str(cell_cache.get((_nr, col), "") or "")
                    if "主役展開" in _nv2:
                        buy_honsen_raw = _nvc
                        break
                    # Step2が来たら終了
                    _nv1 = str(cell_cache.get((_nr, 1), "") or "")
                    if "Step2" in _nv1 or "Step2" in _nv2:
                        break
                break

        # Step1行のD列から展開考察が取れなかった場合、D列に直接テキストがある行を探す
        if not kosatsu_raw:
            for _r in range(100, max_r + 1):
                _vc = str(cell_cache.get((_r, col), "") or "")
                if ("展開考察" in _vc or "逃げるなら" in _vc) and len(_vc) > 20:
                    kosatsu_raw = _vc
                    break

        for r in range(1, max_r + 1):
            v2 = str(cell_cache.get((r, 2), "") or "")
            v1 = str(cell_cache.get((r, 1), "") or "")

            if ("combo順" in v2 or "買い目リスト" in v2) and bet_list_raw is None:
                bet_list_raw = cell_cache.get((r, col))

            if "理想合成オッズ" in v2:
                raw_odds = cell_cache.get((r, col))
                if raw_odds:
                    s = str(raw_odds).replace("理想合成オッズ", "").replace("倍", "").strip()
                    for token in s.split():
                        try:
                            theory_syn_odds = float(token)
                            break
                        except ValueError:
                            continue
                    raw_tso_str = str(raw_odds or "")
                    if "期待値基準を下回っています" in raw_tso_str or "⚠" in raw_tso_str:
                        ev_warning = True

            if v1 == "①判定スコア" and strategy is None:
                raw_kosatsu_v = str(cell_cache.get((r, col), "") or "")
                ms = re.search(r"戦略[:：]\s*(.+?)(?:\n|$)", raw_kosatsu_v)
                if ms:
                    strategy = ms.group(1).strip()
                mt = re.search(r"信頼度[:：]\s*(\d+)", raw_kosatsu_v)
                if mt:
                    try:
                        trust_pct = int(mt.group(1))
                    except ValueError:
                        pass

            # ⑧考察の結論・④展開予測を追加取得
            # ※数値シートでは「⑧考察の結論」は存在せず、A列「④展開予測」/B列「展開◎○艇の役割」が相当する。
            #   A列ラベル（v1）でも B列ラベル（v2）でも両方チェックしてフォールバック取得する。
            if ("⑧考察の結論" in v2 or "⑧考察の結論" in v1) and kosatsu_raw is None:
                kosatsu_raw = str(cell_cache.get((r, col), "") or "")
            if ("④展開予測" in v2 or "④展開予測" in v1) and tenkai_raw is None:
                tenkai_raw = str(cell_cache.get((r, col), "") or "")
            # 「展開◎○艇の役割」（④展開予測のB列ラベル）で取得できていない場合のフォールバック
            if ("展開◎○艇の役割" in v2 or "展開◎○艇" in v2) and kosatsu_raw is None:
                kosatsu_raw = str(cell_cache.get((r, col), "") or "")

        # 参加グレードを bet_list_raw から抽出
        if bet_list_raw:
            gm = re.search(r"参加グレード\n(.+?)(?:\n|$)", str(bet_list_raw))
            if gm:
                grade_raw = gm.group(1).strip()

        races[race_no] = {
            "deadline":        deadline,
            "rank":            rank,
            "score":           score,
            "strategy":        strategy,
            "trust_pct":       trust_pct,
            "players":         players,
            "bet_list_raw":    bet_list_raw,
            "theory_syn_odds": theory_syn_odds,
            "ev_warning":      ev_warning,
            "kosatsu_raw":     kosatsu_raw,
            "tenkai_raw":      tenkai_raw,
            "grade_raw":       grade_raw,
            "buy_honsen_raw":  buy_honsen_raw,
        }

    return races


# ============================================================
# 1レース分をテンプレートシートに書き込む（xlwings）
# ============================================================
def fill_one_race(ws, race_no, venue, race_date, race_data, st_ranks_by_waku=None,
                  venue_stats=None, ininage_stats=None, fly_labels=None, wb_master=None,
                  missing_wakus=None):
    """xlwings シートに1レース分の値を書き込む。
    st_ranks_by_waku: {waku(1〜6): st_rank(float) or None}
    venue_stats:      load_venue_stats() の戻り値
    ininage_stats:    load_ininage_stats() の戻り値
    fly_labels:       未使用（後方互換のため残存）
    wb_master:        openpyxl workbook（選手指数マスタ参照用）
    """
    deadline = race_data["deadline"]
    rank     = race_data["rank"]
    score    = race_data.get("score")
    players  = race_data["players"]

    if st_ranks_by_waku is None:
        st_ranks_by_waku = {}

    def wv(row, col, value):
        ws.cells(row, col).value = value

    import openpyxl.styles as oxs

    def red_fill(row, col):
        """セルを赤塗りつぶしにする（xlwingsのapi経由）"""
        try:
            ws.cells(row, col).api.Interior.Color = 0x0000FF  # BGRなので赤=0x0000FF
        except Exception:
            pass

    def clear_fill(row, col):
        """セルの塗りつぶしを白（元の色）に戻す"""
        try:
            ws.cells(row, col).api.Interior.Color = 0xFFFFFF  # 白
        except Exception:
            pass

    # Row 1: 締切時刻・発行日
    wv(1, 1, deadline)
    wv(1, 9, race_date)

    strategy  = race_data.get("strategy")
    trust_pct = race_data.get("trust_pct")

    # Row 2: 会場名・ランク（G2: ランク記号）
    wv(2, 1, venue)
    wv(2, 7, rank)

    # Row 3: レース番号・決まり手場平均 ＋ 堅実度スコア（G3: スコア＋戦略＋信頼度）
    wv(3, 1, f"{race_no}R")
    if score is not None:
        wv(3, 7, f"{score}点")
    if venue_stats:
        kimari = venue_stats.get("kimari_avg", {})
        in_rate = venue_stats.get("in_rate")
        # C3: 逃げ（イン逃げ率）— セル書式が 0% なので小数で渡す（例: 0.612）
        wv(3, 3, in_rate)
        # D3: 差し, E3: まくり, F3: まくり差し — 同様に小数で渡す
        for col, key in [(4, "差し"), (5, "まくり"), (6, "まくり差し")]:
            v = kimari.get(key)
            wv(3, col, v)

    # ── K4: 1号艇の平均ST順位（小数1桁、ヘッダーテキストごと更新）────
    st_rank_1 = st_ranks_by_waku.get(1)
    if st_rank_1 is not None:
        wv(4, 11, f"1号艇平均ST\n{st_rank_1:.1f}位")

    def fmt_pct(v):
        """float を小数点なし%文字列に変換。例: 60.7 → '61%'"""
        if v is None:
            return "-"
        return f"{round(v)}%"

    def fmt_motor(v):
        """モータ2連を小数点1桁%文字列に変換。例: 37.80 → '37.8%'"""
        if v is None:
            return None
        try:
            f = float(v)
            return f"{f:.1f}%"
        except (TypeError, ValueError):
            return str(v)

    def fmt_num(v):
        """数値を小数点なし文字列に変換。例: 37.80 → '38'"""
        if v is None:
            return None
        try:
            return str(round(float(v)))
        except (TypeError, ValueError):
            return str(v)

    # Row 5〜10: 選手データ
    _fly = fly_labels or {}   # {waku: {"count": N, "label": "高"|"中"|"低"}}
    _missing = set(missing_wakus or [])  # データ不足艇番セット
    prev_st = None  # 直前の艇のST順位
    for i, p in enumerate(players[:6]):
        row = 5 + i
        waku = p["waku"]
        wv(row,  1, waku)
        # データ不足選手は名前の前に ⚠（全角スペース区切り）
        name_raw = p["name"]
        name_display = f"⚠　{name_raw}" if waku in _missing else name_raw
        wv(row,  2, name_display)
        wv(row,  5, p["kumi"])
        wv(row,  6, fmt_motor(p["motor2"]))
        # G列（列7）: フライング数、H列（列8）: F/ST影響ラベル
        # 数値シートの手入力（fly_input: 1=あり）を優先して使用
        fly_input_val = p.get("fly_input")
        has_fly_input = fly_input_val is not None and str(fly_input_val).strip() not in ("", "0", "None")

        if has_fly_input:
            # 手入力あり → fly_input の値をG列に出力
            # H列: マスタのFLY影響度を直接取得（FLYなしなら空白）
            try:
                fly_count_from_input = int(float(str(fly_input_val).strip()))
            except (TypeError, ValueError):
                fly_count_from_input = 1
            wv(row, 7, fly_count_from_input if fly_count_from_input > 0 else None)
            if wb_master is not None:
                fly_label_val = calc_fly_impact(wb_master, p.get("name", ""))
                # calc_fly_impact が None を返した場合のフォールバック
                if fly_label_val is None:
                    fly_label_val = "中"
            else:
                fly_label_val = "中"
        else:
            # 手入力なし → 選手指数マスタ由来のデータを使用
            fly_data      = _fly.get(waku) or {}
            fly_count_val = fly_data.get("count")
            fly_label_val = fly_data.get("label")    # FLYなし=None → 空白
            # FLYあり かつ impactがある場合はimpactを優先表示
            impact_val = fly_data.get("impact")
            if fly_label_val is not None and impact_val is not None:
                fly_label_val = impact_val   # "大"|"中"|"小"|"なし"
            # G列: FLY数（0 や None は空白）
            _fc = None
            if fly_count_val is not None:
                try:
                    _fc = int(fly_count_val) if int(fly_count_val) > 0 else None
                except (TypeError, ValueError):
                    _fc = None
            wv(row, 7, _fc)

        # H列: FLY影響度（FLYなし=空白、FLYあり="大"|"中"|"小"|"なし"）
        wv(row, 8, fly_label_val)
        # L列: 評価（1号艇は逃◎/逃○/逃△/逃× のため評価欄には出さない）
        honmei_val = p["honmei"] if waku != 1 else None
        wv(row, 12, honmei_val or None)
        wv(row, 13, fmt_pct(p["rel_win1"]))
        wv(row, 15, fmt_pct(p["abs_win3"]))

        # I列（列9）の赤塗り条件: 直前の艇より0.5以上ST順位が低い（早い）
        st_w = st_ranks_by_waku.get(waku)
        if row == 5:  # 行5（艇番1）は常に白
            clear_fill(row, 9)
        else:  # 行6〜10 = 艇番2〜6
            if prev_st is not None and st_w is not None and (prev_st - st_w) >= 0.5:
                red_fill(row, 9)
            else:
                clear_fill(row, 9)

        prev_st = st_w

    # Row 14〜19: 事前評価（B〜F列）+ 3着指数（L列）
    # テンプレート列配置: B=逃げ, C=相性, D=機力, E=自在性, F=展開, L=3着指数
    JIZEN_COL_MAP = [
        ("in_nige",  2),   # B列: ①逃げ
        ("aisho",    3),   # C列: ②相性
        ("kiryoku",  4),   # D列: ③機力
        ("tenkai",   5),   # E列: ④展開
        ("jizaisei", 6),   # F列: ⑤S安定
    ]
    for i, p in enumerate(players[:6]):
        row = 14 + i
        wv(row, 1, p["waku"])
        jizen = p.get("jizen") or {}
        for key, col in JIZEN_COL_MAP:
            val = jizen.get(key)
            wv(row, col, val if val not in (None, "", "None", "-", "－") else None)  # "－"（全角）も空扱いに修正
        idx = p["idx3"]
        wv(row, 12, round(float(idx)) if idx is not None else None)

    # Row 22〜26: イン逃げ時2着率（艇番2〜6）
    frame_labels = {2: "2枠(黒)", 3: "3枠(赤)", 4: "4枠(青)", 5: "5枠(黄)", 6: "6枠(緑)"}
    for i, p in enumerate(players[1:6]):
        row = 22 + i
        wv(row, 17, frame_labels.get(p["waku"], f"{p['waku']}枠"))
        pct = p["circle_pct"]
        wv(row, 18, f"{round(pct)}%" if pct is not None else None)

    # Row 23〜28: コース別場平均1着率（F列=6）・1-○ 枠別2着率（G列=7）
    # テンプレート: F22='1着率', G22='1-○' → Row23〜28が枠番1〜6の値行
    # F列には各枠の想定コース別1着率を入れる（1枠→1C1着率, 2枠→2C1着率, ...）
    course_win = (venue_stats or {}).get("course_win_rates", {})
    # フォールバック: course_win_rates がなければ win1_rate を全行に使う
    win1_fallback = (venue_stats or {}).get("win1_rate")
    ininage = ininage_stats or {}
    for i, p in enumerate(players[:6]):
        row = 23 + i
        waku = p["waku"]
        # F列(=6): コース別場平均1着率（枠番=コース番号と対応）
        # Row23(i==0)のみテンプレートがGeneral書式のため '%' 文字列で渡す
        # Row24以降は 0% 書式なので小数で渡す
        win_val = course_win.get(waku, win1_fallback)
        if win_val is not None:
            if i == 0:
                wv(row, 6, f"{round(win_val * 100)}%")  # Row23: General書式 → '56%' 文字列
            else:
                wv(row, 6, win_val)                      # Row24-28: 0%書式 → 小数で渡す
        # G列(=7): 1-○（イン逃げ時の枠別2着率）— 元データが既に小数形式なのでそのまま渡す
        # ininage_stats は小数（例: 0.3029 = 30.29%）で格納されている
        val_2nd = ininage.get(str(waku))
        if val_2nd is not None:
            wv(row, 7, val_2nd)

    # Row 23〜28: 各艇決まり手データ
    def round_pct_str(v):
        """'61.5%' → '62%'、すでに整数ならそのまま"""
        if v is None:
            return None
        s = str(v).replace("%", "").strip()
        try:
            return f"{round(float(s))}%"
        except ValueError:
            return str(v)

    for i, p in enumerate(players[:6]):
        row = 23 + i
        km  = p["kimari"]
        wv(row, 1, str(p["waku"]))
        wv(row, 2, round_pct_str(km.get("逃げ")))
        wv(row, 3, round_pct_str(km.get("差し")))
        wv(row, 4, round_pct_str(km.get("まくり")))
        wv(row, 5, round_pct_str(km.get("まくり差し")))

    # ══════════════════════════════════════════════════════════════
    # 考察・参考買い目・理想合成オッズ
    #
    # I22:P24  考察文（3行）
    # I25:K29  参考買い目・左（参加グレード＋combo前半）
    # L25:P28  参考買い目・右（combo後半）
    # L29:P29  理想合成オッズ
    # ※テンプレートの結合セルをそのまま使用（_remerge不要）
    # ══════════════════════════════════════════════════════════════

    kosatsu_raw    = race_data.get("kosatsu_raw")    or ""
    tenkai_raw     = race_data.get("tenkai_raw")     or ""
    grade_raw      = race_data.get("grade_raw")      or ""
    buy_honsen_raw = race_data.get("buy_honsen_raw") or ""
    rank       = race_data.get("rank", "-")
    score      = race_data.get("score")
    strategy   = race_data.get("strategy") or ""
    trust_pct  = race_data.get("trust_pct")

    # ── ①考察文 → I22（数値シート144行 Step1 のD列テキスト）────────────
    # 「参加見送り」「見送り推奨」「※」行を除去して表示。
    def _strip_unwanted_lines(text):
        skip_kw = ("参加見送り", "見送り推奨", "[!]見送り", "！見送り")
        lines = []
        for l in str(text).splitlines():
            s = l.strip()
            if s.startswith("※"):
                continue
            if any(kw in s for kw in skip_kw):
                continue
            lines.append(l)
        return "\n".join(lines).strip()

    if kosatsu_raw:
        kosatsu_text = _strip_unwanted_lines(kosatsu_raw)[:400]
    elif tenkai_raw:
        kosatsu_text = _strip_unwanted_lines(tenkai_raw)[:400]
    else:
        kosatsu_text = "考察データなし"
    wv(22, 9, kosatsu_text)   # I22（結合セル I22:P24）

    # ── ②参考買い目 → 数値シート145行「主役展開」のD列テキスト ──────────
    # 「参加見送り」「見送り推奨」「※」行を除去。
    import re as _re2, math as _math
    buy_left = buy_right = None
    skip_kw2 = ("参加見送り", "見送り推奨", "[!]見送り", "！見送り")

    if buy_honsen_raw:
        buy_lines = [
            l.strip() for l in buy_honsen_raw.splitlines()
            if l.strip()
            and not l.strip().startswith("※")
            and not any(kw in l for kw in skip_kw2)
        ]
        if buy_lines:
            half = _math.ceil(len(buy_lines) / 2)
            buy_left  = "\n".join(buy_lines[:half])
            buy_right = "\n".join(buy_lines[half:]) if len(buy_lines) > half else None

    # フォールバック: 主役展開データがなければ bet_list_raw の combo 行を使用
    if buy_left is None:
        bet_list_raw = race_data.get("bet_list_raw")
        if bet_list_raw:
            combo_lines = [
                l.strip() for l in str(bet_list_raw).splitlines()
                if _re2.search(r"\d[\d\-]+.*[（(][\d.]+%[）)]", l.strip())
                and not l.strip().startswith("※")
                and not any(kw in l for kw in skip_kw2)
            ]
            if combo_lines:
                half = _math.ceil(len(combo_lines) / 2)
                buy_left  = "\n".join(combo_lines[:half])
                buy_right = "\n".join(combo_lines[half:]) if len(combo_lines) > half else None

    wv(25, 9,  buy_left)   # I25（結合セル I25:K29）
    wv(25, 12, buy_right)  # L25（結合セル L25:P28）
    wv(25, 12, buy_right)  # L25（結合セル L25:P28）

    # ── ③理想合成オッズ → L29 明示的にクリア（テンプレートの残存値を消す）
    wv(29, 12, None)   # テンプレートに残っている理想合成オッズの値を空白で上書き

# ============================================================
# メイン
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="新聞出力テンプレートへの書き込み（数値シートから転記）")
    parser.add_argument("--venue", type=str, default=None, help="会場名 (例: 大村)")
    parser.add_argument("--race",  type=int, default=None, help="レース番号 (省略時: 全レース自動処理)")
    parser.add_argument("--all",   action="store_true",    help="数値Excelにある全会場を一括処理")
    args = parser.parse_args()

    print("=" * 55)
    print("  新聞出力テンプレート 書き込みツール")
    print("=" * 55)

    # ── ファイル確認 ──────────────────────────────────────────────────
    for path, label in [
        (TEMPLATE_PATH, "テンプレート"),
        (MASTER_PATH,   "マスタExcel"),
        (NUM_PATH,      "数値Excel（軽量版）"),
    ]:
        if not path.exists():
            print(f"❌ {label}が見つかりません: {path}")
            if label == "数値Excel（軽量版）":
                print("   先に load_race.py を実行してください。")
            return

    # ── マスタExcelを開く（選手指数・会場統計・イン逃げ分析）─────────
    print(f"  📊 マスタデータを読み込み中...")
    try:
        wb_master = openpyxl.load_workbook(str(MASTER_PATH), data_only=True)
    except Exception as e:
        print(f"❌ マスタExcelを開けませんでした: {e}")
        return

    # ── 選手指数マスタからST順位を読み込み ───────────────────────────
    print(f"  🏁 選手指数マスタからST順位を読み込み中...")
    st_master = load_st_rank_master(wb_master)

    # ── 数値Excel（軽量版）を開く（{会場}_数値シート）────────────────
    print(f"  📊 数値シートを読み込み中...")
    try:
        wb_num = openpyxl.load_workbook(str(NUM_PATH), data_only=True)
    except Exception as e:
        print(f"❌ 数値Excelを開けませんでした: {e}")
        return

    # ── 会場リスト確定（--all / --venue / 対話式）───────────────────
    all_candidates = sorted(
        s.replace("_数値", "") for s in wb_num.sheetnames if s.endswith("_数値")
    )
    if not all_candidates:
        print("❌ 数値シートが見つかりません。先に load_race.py を実行してください。")
        return

    if args.all:
        venues_to_process = all_candidates
        print(f"  🏁 全会場一括処理モード: {', '.join(venues_to_process)}")
    elif args.venue:
        venues_to_process = [args.venue]
        print(f"  🏁 会場: {args.venue}")
    else:
        print()
        print("  📂 会場を選択してください:")
        for i, v in enumerate(all_candidates, start=1):
            print(f"    {i}. {v}")
        print()
        try:
            sel = input(f"  番号を入力 (1〜{len(all_candidates)}): ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n❌ キャンセルしました")
            return
        if not sel.isdigit() or not (1 <= int(sel) <= len(all_candidates)):
            print(f"❌ 無効な番号です: {sel}")
            return
        venues_to_process = [all_candidates[int(sel) - 1]]
        print(f"  🏁 会場: {venues_to_process[0]}")

    # ── 残存xlwings/Excelプロセスを解放してからコピー ────────────────
    import time, subprocess, platform
    if OUTPUT_PATH.exists():
        try:
            import xlwings as _xw_check
            for _app in _xw_check.apps:
                for _bk in list(_app.books):
                    try:
                        if pathlib.Path(_bk.fullname).resolve() == OUTPUT_PATH.resolve():
                            _bk.close()
                    except Exception:
                        pass
        except Exception:
            pass
        if platform.system() == "Windows":
            try:
                subprocess.run(
                    ["powershell", "-Command",
                     f"Get-Process EXCEL -ErrorAction SilentlyContinue | "
                     f"Where-Object {{$_.MainWindowTitle -like '*新聞出力_filled*'}} | Stop-Process -Force"],
                    capture_output=True, timeout=5
                )
            except Exception:
                pass
        time.sleep(1)

    for attempt in range(3):
        try:
            shutil.copy2(str(TEMPLATE_PATH), str(OUTPUT_PATH))
            break
        except PermissionError:
            if attempt < 2:
                print(f"  ⏳ 出力ファイルがロック中... 3秒後にリトライ ({attempt+1}/3)")
                time.sleep(3)
            else:
                print(f"❌ 出力ファイルを書き込めません。手動でExcelを閉じてから再実行してください: {OUTPUT_PATH.name}")
                return

    # ── 全会場分のデータを事前収集 ──────────────────────────────────
    # venue → (venue_stats, ininage_stats, races, fly_labels_all, race_date, race_nos)
    venue_data = {}
    for venue in venues_to_process:
        print()
        print(f"  {'─'*45}")
        print(f"  🏟️  [{venue}] データ収集中...")

        # 会場統計・イン逃げ
        venue_stats   = load_venue_stats(wb_master, venue)
        ininage_stats = load_ininage_stats(wb_master, venue)
        if venue_stats:
            win1 = venue_stats.get("win1_rate")
            in_r = venue_stats.get("in_rate")
            print(f"  ✅ 場平均1着率: {win1*100:.1f}% / イン逃げ率: {in_r*100:.1f}%" if win1 and in_r else "  ⚠️  会場統計データなし")
        else:
            print(f"  ⚠️  会場統計シートが見つからないか、{venue}のデータがありません")

        # FLY入力行確認・追加
        sheet_name_check = f"{venue}_数値"
        fly_already_exists = False
        if sheet_name_check in wb_num.sheetnames:
            ws_check = wb_num[sheet_name_check]
            for r in range(1, ws_check.max_row + 1):
                if ws_check.cell(r, 1).value == "FLY入力":
                    fly_already_exists = True
                    break
        if not fly_already_exists:
            try:
                wb_num_rw = openpyxl.load_workbook(str(NUM_PATH))
                ensure_fly_input_rows(wb_num_rw, venue)
                wb_num_rw.save(str(NUM_PATH))
                wb_num_rw.close()
                wb_num.close()
                wb_num = openpyxl.load_workbook(str(NUM_PATH), data_only=True)
            except Exception as e:
                print(f"  ⚠️  FLY入力行の追加に失敗しました（続行します）: {e}")

        races = read_numeric_sheet(wb_num, venue)
        if not races:
            print(f"  ⚠️  {venue}: 数値シートが読み取れないためスキップ")
            continue

        # FLY数収集
        fly_labels_all = {}
        for rno, rd in races.items():
            fly_labels_all[rno] = load_fly_label(wb_master, rd["players"])
        fly_msgs = [
            f"{rno}R-{waku}号艇: FLY{v.get('count',0)}回/{v.get('days','?')}日/{v['label']}"
            for rno in fly_labels_all
            for waku, v in fly_labels_all[rno].items()
            if v and v.get("label") in ("高", "中")
        ]
        print(f"  ✅ F/ST影響: FLYあり {len(fly_msgs)}艇検出")

        # 日付・レース番号
        ws_num_tmp = wb_num[f"{venue}_数値"]
        header     = str(ws_num_tmp.cell(1, 1).value or "")
        m_date     = re.search(r"(\d{4}-\d{2}-\d{2})", header)
        race_date  = m_date.group(1) if m_date else ""

        all_race_nos = sorted(races.keys(), key=lambda x: int(x) if x.isdigit() else 99)
        if args.race:
            race_nos = [r for r in all_race_nos if r == str(args.race)]
            if not race_nos:
                print(f"  ⚠️  {venue}: {args.race}Rのデータが見つかりません（スキップ）")
                continue
        else:
            race_nos = all_race_nos
        print(f"  📋 [{venue}] 対象レース: {race_nos}")

        venue_data[venue] = dict(
            venue_stats=venue_stats, ininage_stats=ininage_stats,
            races=races, fly_labels_all=fly_labels_all,
            race_date=race_date, race_nos=race_nos
        )

    if not venue_data:
        print("❌ 処理できる会場がありませんでした。")
        return

    print()
    print(f"  📂 Excelを起動中...")

    # ── xlwings でExcelを起動（全会場まとめて1回だけ）──────────────
    app = xw.App(visible=False)
    app.display_alerts = False
    app.screen_updating = False

    wb_out = None
    png_bytes_by_sheet_idx = {}   # 全会場通しのシートインデックス
    sheet_counter = 0             # 書き込み済みシート数
    try:
        wb_out = app.books.open(str(OUTPUT_PATH.resolve()))
        template_sheet = wb_out.sheets[0]

        for venue, vd in venue_data.items():
            print()
            print(f"  🏟️  [{venue}] 新聞書き込み中...")
            races          = vd["races"]
            fly_labels_all = vd["fly_labels_all"]
            race_date      = vd["race_date"]
            race_nos       = vd["race_nos"]
            venue_stats    = vd["venue_stats"]
            ininage_stats  = vd["ininage_stats"]

            for race_no in race_nos:
                sheet_counter += 1
                sheet_name = f"{venue}_{race_no}R"
                if sheet_counter == 1:
                    ws = template_sheet
                    ws.name = sheet_name
                else:
                    template_sheet.api.Copy(After=wb_out.sheets[-1].api)
                    ws = wb_out.sheets[-1]
                    ws.name = sheet_name

                st_ranks = get_player_st_ranks(races[race_no]["players"], st_master)
                missing  = [w for w, v in st_ranks.items() if v is None]
                if missing:
                    print(f"  ⚠️  {race_no}R 艇番 {missing} のST順位が取得できませんでした（平均値で代替）")
                fill_one_race(ws, race_no, venue, race_date, races[race_no], st_ranks,
                             venue_stats=venue_stats, ininage_stats=ininage_stats,
                             fly_labels=fly_labels_all.get(race_no, {}),
                             wb_master=wb_master,
                             missing_wakus=missing)
                print(f"  ✅ {venue} {race_no}R 書き込み完了")

                # スリット図PNG生成（close前）
                png_bytes_by_sheet_idx[sheet_counter] = generate_slit_image(st_ranks)

        print()
        print(f"  💾 保存中...")
        wb_out.save()
        print(f"  ✅ 完了！ → {OUTPUT_PATH.name}")

    except Exception as e:
        import traceback
        print()
        print("=" * 55)
        print("❌ エラーが発生しました:")
        traceback.print_exc()
        print("=" * 55)

    finally:
        if wb_out is not None:
            try:
                wb_out.close()
            except Exception:
                pass
        try:
            app.quit()
        except Exception:
            pass

    # ── スリット図差し替え（close/quit後にzipを操作）─────────────────
    try:
        replace_slit_images_in_xlsx(OUTPUT_PATH, png_bytes_by_sheet_idx)
        print(f"  🖼️  スリット図をST順位で更新完了（{len(png_bytes_by_sheet_idx)}レース分）")
    except Exception as e:
        print(f"  ⚠️  スリット図の更新に失敗しました: {e}")

    print("=" * 55)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        traceback.print_exc()