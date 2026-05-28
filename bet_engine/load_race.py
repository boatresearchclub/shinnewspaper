# -*- coding: utf-8 -*-
"""
ボートリサーチ新聞 - 全レース一括新聞出力スクリプト
=====================================================
【使い方】
  python scripts/load_race.py --venue 大村
  python scripts/load_race.py --venue 大村 --race 5
  python scripts/load_race.py --venue 大村 --date 2026-02-15
  python scripts/load_race.py --venue 大村 --date 2026-02-15 --race 5
  python scripts/load_race.py   # csv_output/ の最新ファイルを自動検出

【出力先】
  ボートリサーチ新聞_軽量版.xlsx の「出力_新聞」シート
  ※「出力_新聞サンプル」のレイアウト（行高・列幅・結合・色）を完全再現
  ※データのみ上書き。行高・列幅は一切変更しない
"""
_SCRIPT_VERSION = "2026-03-15_❶❷❸❹_v6.9_orkaeshi-fix"  # ← このバージョン文字列が起動時に表示されれば正しいファイルを実行中
# v6.9変更点:
#   修正①: prob_third に _second_block_factor を追加。
#           S2/S3/S4 の3着確率計算で second_w（2着艇）の進路封鎖効果を反映。
#           旧: P(3着=C|A,B) が second_w に依存しない → 3-2-? と 3-5-? が同じ3着分布
#           新: 2着艇より外側かつ同系展開の艇を 0.70〜0.90 に減衰（封鎖係数乗算）
#   修正②: _fetch の信頼度閾値を 0.3 → 0.15 に緩和し、
#           低信頼度（0.15〜0.50）の会場別データを全国マスタとブレンドして活用。
#           旧: trust<0.3 は会場データを完全に捨てていた
#           新: trust=0.30 → 会場60%/全国40% / trust=0.15 → 会場30%/全国70%

import os
import sys
import glob
import argparse
import re
import pathlib
# import copy  # 未使用
import shutil
import platform
import subprocess
import statistics
# import tempfile  # 未使用（BytesIO移行後に不要）
import json
import itertools
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    # GradientFill,  # 未使用
)
# from openpyxl.styles.differential import DifferentialStyle  # 未使用
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    print("⚠️  Pillowが見つかりません。ST舟図は出力されません。")

try:
    from calc_3rentan_probs_v2 import calc_3rentan_probs_v2 as _ext_calc_3rentan
    _EXT_CALC_AVAILABLE = True
except ImportError:
    _EXT_CALC_AVAILABLE = False

# 倶楽部流 事前評価モジュール（scripts/ フォルダに配置）
sys.path.insert(0, str(pathlib.Path(__file__).parent))
try:
    from evaluate_jizen import evaluate_all, calculate_diversity_rate
    JIZEN_AVAILABLE = True
except ImportError:
    JIZEN_AVAILABLE = False
    print("⚠️  evaluate_jizen.py が見つかりません。事前評価はスキップされます。")

try:
    from honmei_scenario import integrate_with_suggest_3rentan
    HONMEI_SCENARIO_AVAILABLE = True
except ImportError:
    HONMEI_SCENARIO_AVAILABLE = False
    print("⚠️  honmei_scenario.py が見つかりません。既存ロジックで動作します。")

# ============================================================
# パス設定
# ============================================================
BASE_DIR    = pathlib.Path(__file__).parent.parent
CSV_DIR     = pathlib.Path(__file__).parent / "csv_output"
EXCEL_FILE  = BASE_DIR / "ボートリサーチ新聞_軽量版.xlsx"
MASTER_FILE = BASE_DIR / "ボートリサーチ_マスタ.xlsx"
# 会場別コースマスタのCSVキャッシュ（update_master.py が生成）
VENUE_COURSE_CSV = BASE_DIR / "data" / "venue_course_master.csv"
# 展開別残存マスタのCSVキャッシュ（update_master.py が生成）
TENKAI_VENUE_CSV    = BASE_DIR / "data" / "tenkai_survival_venue.csv"
TENKAI_NATIONAL_CSV = BASE_DIR / "data" / "tenkai_survival_national.csv"

# ── グレード別CSVキャッシュ（update_master.py --grade g1 / --grade sg で生成）──
# ファイルが存在しない場合は一般戦マスタにフォールバックする
VENUE_COURSE_CSV_G1 = BASE_DIR / "data" / "venue_course_master_g1.csv"
VENUE_COURSE_CSV_SG = BASE_DIR / "data" / "venue_course_master_sg.csv"
TENKAI_VENUE_CSV_G1    = BASE_DIR / "data" / "tenkai_survival_venue_g1.csv"
TENKAI_NATIONAL_CSV_G1 = BASE_DIR / "data" / "tenkai_survival_national_g1.csv"
TENKAI_VENUE_CSV_SG    = BASE_DIR / "data" / "tenkai_survival_venue_sg.csv"
TENKAI_NATIONAL_CSV_SG = BASE_DIR / "data" / "tenkai_survival_national_sg.csv"

# グレード → CSVパスのマッピング（load_masters / main で参照）
_GRADE_CSV_MAP = {
    "一般": {
        "venue_course":      VENUE_COURSE_CSV,
        "tenkai_venue":      TENKAI_VENUE_CSV,
        "tenkai_national":   TENKAI_NATIONAL_CSV,
        "sheet_master":      "📊コース別マスタ",
        "sheet_ininage":     "イン逃げ分析",
    },
    "G1": {
        "venue_course":      VENUE_COURSE_CSV_G1,
        "tenkai_venue":      TENKAI_VENUE_CSV_G1,
        "tenkai_national":   TENKAI_NATIONAL_CSV_G1,
        "sheet_master":      "📊コース別マスタ_G1",   # update_master.py が生成するシート名
        "sheet_ininage":     "イン逃げ分析_G1",
    },
    "SG": {
        "venue_course":      VENUE_COURSE_CSV_SG,
        "tenkai_venue":      TENKAI_VENUE_CSV_SG,
        "tenkai_national":   TENKAI_NATIONAL_CSV_SG,
        "sheet_master":      "📊コース別マスタ_SG",
        "sheet_ininage":     "イン逃げ分析_SG",
    },
}
# G2・G3 は独立マスタ
_GRADE_CSV_MAP["G2"] = {
    "venue_course":      BASE_DIR / "data" / "venue_course_master_g2.csv",
    "tenkai_venue":      BASE_DIR / "data" / "tenkai_survival_venue_g2.csv",
    "tenkai_national":   BASE_DIR / "data" / "tenkai_survival_national_g2.csv",
    "sheet_master":      "📊コース別マスタ_G2",
    "sheet_ininage":     "イン逃げ分析_G2",
}
_GRADE_CSV_MAP["G3"] = _GRADE_CSV_MAP["G2"]  # G3はG2マスタを流用
_GRADE_CSV_MAP["SG"] = _GRADE_CSV_MAP["G1"]  # SG はG1マスタを共用
# 【⑤追加】会場別コース距離補正値CSVキャッシュ（update_master.py が生成）
VENUE_COURSE_ADJ_CSV = BASE_DIR / "data" / "venue_course_adj.csv"

# 【⑤追加】会場別コース距離補正値をCSVから読み込む
# キー: 会場名 / 値: {コース番号str: 補正秒数float}
# CSVが存在しない場合は空dictを返し、_predict_first_turnが固定値にフォールバック
def _load_venue_course_adj() -> dict:
    if not VENUE_COURSE_ADJ_CSV.exists():
        return {}
    try:
        df = pd.read_csv(str(VENUE_COURSE_ADJ_CSV), encoding="utf-8-sig", dtype=str)
        result = {}
        for row in df.to_dict(orient="records"):
            venue = str(row.get("会場名", "") or "").strip()
            trust = safe_float(row.get("信頼度"), 0.0) or 0.0
            if not venue or trust < 0.3:
                continue  # 信頼度0.3未満は固定値を使用
            adj_map = {}
            for c in range(2, 7):
                v = safe_float(row.get(f"{c}C補正"))
                if v is not None:
                    adj_map[str(c)] = v
            if adj_map:
                result[venue] = adj_map
        print(f"  📐 会場別コース距離補正値読込: {len(result)}会場")
        return result
    except Exception as e:
        print(f"  ⚠️  会場別コース距離補正値読込失敗（固定値を使用）: {e}")
        return {}

# 起動時に一度だけ読み込む
_VENUE_COURSE_ADJ: dict = _load_venue_course_adj()
# ─── オッズ設定 ───────────────────────────────────────────────────────────
# 【方法A】boatrace.jp のURLを指定（推奨）
#   レース直前にURLを貼るだけで実際のオッズを自動取得してEV計算
#   例: ACTUAL_ODDS_URL = "https://www.boatrace.jp/owpc/pc/race/odds3t?rno=3&jcd=01&hd=20240601"
#   ※ レース番号(rno)と場コード(jcd)と日付(hd)を書き換えて使用
#   ※ Noneにすると方法Bにフォールバック
ACTUAL_ODDS_URL = None  # ← ここにURLを貼る（レースごとに書き換え）

# 【修正⑤】会場名 → boatrace.jp 場コード（jcd）マッピング
# 公式場コード: 01=桐生 02=戸田 03=江戸川 04=平和島 05=多摩川 06=浜名湖
#              07=蒲郡 08=常滑 09=津  10=三国  11=びわこ 12=住之江
#              13=尼崎 14=鳴門 15=丸亀 16=児島 17=宮島  18=徳山
#              19=下関 20=若松 21=芦屋 22=福岡 23=唐津  24=大村
VENUE_JCD_MAP = {
    "桐生":   "01", "戸田":   "02", "江戸川": "03", "平和島": "04",
    "多摩川": "05", "浜名湖": "06", "蒲郡":   "07", "常滑":   "08",
    "津":     "09", "三国":   "10", "びわこ": "11", "住之江": "12",
    "尼崎":   "13", "鳴門":   "14", "丸亀":   "15", "児島":   "16",
    "宮島":   "17", "徳山":   "18", "下関":   "19", "若松":   "20",
    "芦屋":   "21", "福岡":   "22", "唐津":   "23", "大村":   "24",
}

def build_odds_url(venue: str, race_no, race_date: str) -> str | None:
    """
    【修正⑤】会場名・レース番号・日付から boatrace.jp の3連単オッズURLを自動生成。

    Parameters
    ----------
    venue     : str  会場名（例: "大村"）
    race_no   : int or str  レース番号（例: 3）
    race_date : str  日付文字列（例: "2026-03-08" または "20260308"）

    Returns
    -------
    str  URL文字列 / None（会場コード不明の場合）

    使用例
    ------
    url = build_odds_url("大村", 5, "2026-03-08")
    # → "https://www.boatrace.jp/owpc/pc/race/odds3t?rno=5&jcd=24&hd=20260308"
    """
    jcd = VENUE_JCD_MAP.get(str(venue).strip())
    if not jcd:
        return None
    rno = int(race_no) if str(race_no).isdigit() else race_no
    # 日付フォーマット統一（"2026-03-08" → "20260308"）
    hd = str(race_date).replace("-", "").replace("/", "")
    if len(hd) != 8 or not hd.isdigit():
        return None
    return f"https://www.boatrace.jp/owpc/pc/race/odds3t?rno={rno}&jcd={jcd}&hd={hd}"

# 【方法B】理論オッズExcel（全会場共通・固定ファイル）
#   URLが指定されていない場合のフォールバック
#   odds/ フォルダに「理想オッズ完成版.xlsx」を置くだけでOK
ODDS_DIR      = BASE_DIR / "odds"
ODDS_FILEPATH = ODDS_DIR / "理想オッズ完成版.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

SHEET_MASTER    = "📊コース別マスタ"
SHEET_PLAYER    = "選手指数マスタ"
SHEET_ININAGE   = "イン逃げ分析"
SHEET_OUTPUT    = "出力_新聞"
SHEET_SAMPLE    = "出力_新聞サンプル"

ROWS_PER_RACE = 30  # サンプル29行ブロック + 1空行

# 艇色（競艇公式カラー）
BOAT_COLORS = {
    1: {"bg": "FFFFFFFF", "fg": "FF000000"},  # 白
    2: {"bg": "FF111111", "fg": "FFFFFFFF"},  # 黒
    3: {"bg": "FFFF0000", "fg": "FFFFFFFF"},  # 赤
    4: {"bg": "FF00B0F0", "fg": "FFFFFFFF"},  # 青
    5: {"bg": "FFFFFF00", "fg": "FF000000"},  # 黄
    6: {"bg": "FF00B050", "fg": "FFFFFFFF"},  # 緑
}
BOAT_BG_LIGHT = {
    1: "FFFFFFFF",
    2: "FF555555",
    3: "FFFFCCCC",
    4: "FFCCE8FF",
    5: "FFFFFACC",
    6: "FFCCFFEE",
}

# ============================================================
# ユーティリティ
# ============================================================
def sep(char="=", n=55):
    print(char * n)

def safe_float(val, default=None):
    """文字列・数値を安全にfloatへ変換する共通ユーティリティ。"""
    try:
        v = str(val).replace("%","").strip()
        return float(v) if v not in ("", "None", "nan", "-", "★") else default
    except Exception:
        return default

# ============================================================
# カラム名マッピング定数（load_race ↔ update_master 間で統一）
# update_master.py は「決まり手_逃げ%」形式で書き込む。
# Excelマスタシートのヘッダは「逃げ%」「差し%」…と短縮されているため
# ここで両方を吸収するエイリアスを定義する。
# ============================================================
KIMARI_COL_MAP = {
    "逃げ%":        ["逃げ%",        "決まり手_逃げ%"],
    "差し%":        ["差し%",        "決まり手_差し%"],
    "まくり%":      ["まくり%",      "決まり手_まくり%"],
    "まくり差し%":  ["まくり差し%",  "決まり手_まくり差し%"],
    "抜き%":        ["抜き%",        "決まり手_抜き%"],
    "恵まれ%":      ["恵まれ%",      "決まり手_恵まれ%"],
}

def _get_cm_val(cm: dict, canonical_key: str):
    """カラム名マッピングを使ってコース別マスタから値を取得する。"""
    for alias in KIMARI_COL_MAP.get(canonical_key, [canonical_key]):
        v = cm.get(alias)
        if v is not None:
            return v
    return None

def make_fill(hex_color):
    if not hex_color or hex_color in ("00000000", "FFFFFFFF", ""):
        return PatternFill(fill_type=None)
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, size=9, color="FF000000", name="Noto Sans CJK SC"):
    return Font(name=name, size=size, bold=bold, color=color)

def center_align(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="FFCCCCDD")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill():
    return make_fill("FF404060")

def header_font():
    return make_font(bold=True, color="FFFFFFFF")

def subheader_fill():
    return make_fill("FFCCCCDD")

def subheader_font():
    return make_font(bold=True, color="FF000000")

def write_cell(ws, row, col, value, fill=None, font=None, alignment=None, border=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    if fill:      c.fill = fill
    if font:      c.font = font
    if alignment: c.alignment = alignment
    if border:    c.border = border


# ============================================================
# モーターCSV読み込み（scrape_motor.py の出力）
# ============================================================
def load_motor_csv(venue, race_date_str, race_df=None):
    """
    出走表CSV（load_csvで読み込み済みのDataFrame）から
    モーター番号・2連対率を抽出して返す。

    出走表CSVの列名対応:
        レース列  : 「レース」または「R」
        艇番列    : 「艇番」または「枠」
        モーター番号 : 「M番」
        2連対率   : 「M2率」

    戻り値は build_jizen_members が期待する形式:
        columns = [race_no, boat_no, motor_no, motor_2rate]
    """
    if race_df is None or len(race_df) == 0:
        return None

    # 列名を特定
    race_col = next((c for c in race_df.columns if c in ("レース", "R", "レース番号")), None)
    boat_col = next((c for c in race_df.columns if c in ("艇番", "枠")), None)
    mno_col  = next((c for c in race_df.columns if c in ("M番", "モーター番号", "motor_no")), None)
    m2r_col  = next((c for c in race_df.columns if c in ("M2率", "モーター2連率", "M2連対率", "motor_2rate")), None)

    if mno_col is None or m2r_col is None:
        print(f"  ⚠️  出走表CSVにモーター列（M番/M2率）が見つかりません。機力評価はスキップします。")
        return None

    try:
        rows = []
        for _, row in race_df.iterrows():
            race_no  = pd.to_numeric(row.get(race_col), errors="coerce") if race_col else None
            boat_no  = pd.to_numeric(row.get(boat_col), errors="coerce") if boat_col else None
            motor_no = pd.to_numeric(row.get(mno_col),  errors="coerce")
            m2rate   = pd.to_numeric(row.get(m2r_col),  errors="coerce")
            rows.append({
                "race_no":     race_no,
                "boat_no":     boat_no,
                "motor_no":    motor_no,
                "motor_2rate": m2rate,
            })
        df = pd.DataFrame(rows)
        print(f"  ⚙️  モーターデータを出走表CSVから読み込みました（{len(df)}艇分）")
        return df
    except Exception as e:
        print(f"  ⚠️  モーターデータ抽出エラー: {e}")
        return None

# ============================================================
# マスタ読み込み
# ============================================================

# 【軽微②改善】選手名正規化とマスタルックアップを1か所に集約
# 旧方式: 登録時にエイリアスを追加 → 4文字名が5文字名に誤マッチする可能性
# 新方式: ルックアップ関数で「完全一致 → 先頭N文字一致」の順で検索
def _lookup_name_course(master_dict, name, course_str):
    """
    course_masterから(name, course)でルックアップする。
    完全一致 → 4文字前方一致 → 5文字前方一致 の順で検索。
    """
    key_exact = (name, course_str)
    if key_exact in master_dict:
        return master_dict[key_exact]
    # 4文字前方一致（5文字名→4文字で照合）
    if len(name) >= 5:
        key4 = (name[:4], course_str)
        if key4 in master_dict:
            return master_dict[key4]
    # 4文字キーで登録されているマスタと5文字クエリの照合
    if len(name) == 4:
        # マスタ側が5文字で登録されていてこちらが4文字の場合は対応不要（登録時に済み）
        pass
    return None

def _lookup_player(master_dict, name):
    """player_masterから選手名でルックアップ。完全一致 → 4文字前方一致 の順。"""
    if name in master_dict:
        return master_dict[name]
    if len(name) >= 5:
        alias = name[:4]
        if alias in master_dict:
            return master_dict[alias]
    return None

def load_masters(wb, race_grade: str = "一般"):
    """コース別マスタ・選手指数マスタ・イン逃げ分析を読み込む

    Parameters
    ----------
    wb          : openpyxl Workbook（ボートリサーチ_マスタ.xlsx）
    race_grade  : "一般" / "G1" / "G2" / "G3" / "SG"
                  グレードに応じて参照するCSV・シートを切り替える。
                  対応CSVが存在しない場合は自動的に一般戦マスタにフォールバック。
    """
    # グレード → CSV/シートマップを解決（未知グレードは一般扱い）
    _gmap = _GRADE_CSV_MAP.get(race_grade, _GRADE_CSV_MAP["一般"])
    _gmap_ippan = _GRADE_CSV_MAP["一般"]  # フォールバック用

    def _resolve_csv(key):
        """グレード別パスが存在すればそれを、なければ一般戦パスを返す"""
        p = _gmap[key]
        if isinstance(p, pathlib.Path) and not p.exists():
            p_fallback = _gmap_ippan[key]
            print(f"  ⚠️  {race_grade}用マスタなし → 一般戦マスタを使用: {p.name}")
            return p_fallback
        return p

    _vc_csv_path       = _resolve_csv("venue_course")
    _tenkai_venue_path = _resolve_csv("tenkai_venue")
    _tenkai_nat_path   = _resolve_csv("tenkai_national")

    # コース別マスタシート名（グレード別シートがなければ一般戦シートにフォールバック）
    _sheet_master  = _gmap["sheet_master"]
    _sheet_ininage = _gmap["sheet_ininage"]
    if _sheet_master not in wb.sheetnames:
        print(f"  ⚠️  シート「{_sheet_master}」なし → 一般戦シートを使用")
        _sheet_master = SHEET_MASTER
    if _sheet_ininage not in wb.sheetnames:
        _sheet_ininage = SHEET_ININAGE

    # コース別マスタ
    ws_m = wb[_sheet_master]
    master_rows = list(ws_m.iter_rows(values_only=True))
    # ヘッダ行を探す
    header_row = None
    for i, row in enumerate(master_rows):
        if row and row[0] == "選手名":
            header_row = i
            break
    if header_row is None:
        return {}, {}, {}
    headers_m = master_rows[header_row]
    course_master = {}  # {(選手名, コース): {指数dict}}
    for row in master_rows[header_row+1:]:
        if not row or row[0] is None:
            continue
        d = dict(zip(headers_m, row))
        name_full = str(d.get("選手名","")).strip()
        course_str = str(d.get("コース","")).strip()
        key = (name_full, course_str)
        course_master[key] = d
        # 【軽微②改善】5文字名の4文字エイリアスは、4文字キーが未登録の場合のみ追加
        # ※ただし4文字名の選手と衝突しないよう、ルックアップ側で優先度制御する
        if len(name_full) == 5:
            alias_key = (name_full[:4], course_str)
            if alias_key not in course_master:
                course_master[alias_key] = d

    # 選手指数マスタ
    ws_p = wb[SHEET_PLAYER]
    player_rows = list(ws_p.iter_rows(values_only=True))
    header_row_p = None
    for i, row in enumerate(player_rows):
        # A列が「登録番号」、B列が「選手名」の行をヘッダとして検出
        if row and row[1] == "選手名":
            header_row_p = i
            break
    headers_p = player_rows[header_row_p] if header_row_p is not None else []
    player_master = {}  # {選手名: {指数dict}}
    if header_row_p is not None:
        for row in player_rows[header_row_p+1:]:
            if not row or row[0] is None:
                continue
            d = dict(zip(headers_p, row))
            name = str(d.get("選手名","")).strip()
            player_master[name] = d
            # 改善D: 5文字名の先頭4文字エイリアスを登録時点で追加（O(N)フォールバックループを除去）
            if len(name) == 5:
                alias = name[:4]
                if alias not in player_master:
                    player_master[alias] = d

    # イン逃げ分析（1行=1会場、枠別2着率が横展開）
    ws_i = wb[_sheet_ininage]
    ininage_rows = list(ws_i.iter_rows(values_only=True))
    header_row_i = None
    for i, row in enumerate(ininage_rows):
        if row and row[0] == "会場名":
            header_row_i = i
            break
    headers_i = ininage_rows[header_row_i] if header_row_i is not None else []
    # ininage_master: {会場名: {"2nd": {"2": 率,...}, "3rd": {"2": 率,...}}} に変換
    # "2nd" = 枠別イン逃げ時2着率（circle_pct算出用）
    # "3rd" = 枠別イン逃げ時3着以内率（idx3独立算出用）
    ininage_master = {}
    if header_row_i is not None:
        for row in ininage_rows[header_row_i+1:]:
            if not row or row[0] is None:
                continue
            d = dict(zip(headers_i, row))
            v = str(d.get("会場名", "")).strip()
            frame_map_2nd = {}
            frame_map_3rd = {}
            for waku_no in range(1, 7):
                val_2nd = d.get(f"{waku_no}枠\n2着率")
                if val_2nd is not None:
                    frame_map_2nd[str(waku_no)] = float(val_2nd)
                val_3rd = d.get(f"{waku_no}枠\n3着以内率")
                if val_3rd is not None:
                    frame_map_3rd[str(waku_no)] = float(val_3rd)
            # 後方互換: ininage_master[v] をそのままdictとして参照していた箇所のため
            # 旧形式（2着率のみの flat dict）も維持しつつ "3rd" キーを追加する
            ininage_master[v] = dict(frame_map_2nd)      # 旧形式互換（2着率フラット）
            ininage_master[v]["_3rd"] = frame_map_3rd    # 3着以内率（枠番→率）

    # 会場統計（イン逃げ率・決まり手場平均）
    # ──────────────────────────────────────────────────────────────────
    # 【キー名正規化】Excelヘッダは改行あり（例: "1C\n1着率"、"1C\n1R"）だが
    # load_race.py 内では "1コース1着率"、"1C_1R1着率" を参照していたため
    # 全て不一致だった。ここで読み込み時に両方のキー形式を登録する。
    # ──────────────────────────────────────────────────────────────────
    venue_stats_master = {}
    if "会場統計" in wb.sheetnames:
        ws_vs = wb["会場統計"]
        vs_rows = list(ws_vs.iter_rows(values_only=True))
        header_vs = None
        for i, row in enumerate(vs_rows):
            if row and row[0] == "会場名":
                header_vs = i
                break
        if header_vs is not None:
            headers_vs = vs_rows[header_vs]
            for row in vs_rows[header_vs+1:]:
                if not row or row[0] is None:
                    continue
                d = dict(zip(headers_vs, row))
                v = str(d.get("会場名", "")).strip()

                # ── キー名エイリアスを追加登録 ──
                normalized = dict(d)
                for h, val in d.items():
                    if h is None:
                        continue
                    h_str = str(h)
                    # "XC\n1着率" → "Xコース1着率" と "XC_1着率"
                    # re はトップレベルで import 済みのため ループ内 import 不要
                    m = re.match(r"^(\d)C\n1着率$", h_str)
                    if m:
                        c = m.group(1)
                        normalized[f"{c}コース1着率"]  = val
                        normalized[f"{c}C_1着率"]      = val
                    # "XC\nYR" → "XC_YR1着率"（R×C別1着率）
                    m2 = re.match(r"^(\d)C\n(\d+)R$", h_str)
                    if m2:
                        c, r_no = m2.group(1), m2.group(2)
                        normalized[f"{c}C_{r_no}R1着率"] = val
                    # "XR\n荒れ" → "XR荒れスコア"
                    m3 = re.match(r"^(\d+)R\n荒れ$", h_str)
                    if m3:
                        r_no = m3.group(1)
                        normalized[f"{r_no}R荒れスコア"] = val
                    # "まくり\n差し率" → "まくり差し率"
                    if h_str == "まくり\n差し率":
                        normalized["まくり差し率"] = val

                venue_stats_master[v] = normalized

    # ── 会場別コースマスタ（選手×会場×コース の3次元実績）──
    # CSVキャッシュが存在すればそちらを優先（openpyxl iter_rowsより5〜10倍速）
    # CSVがなければ従来通りExcelシートから読み込む（後方互換）
    venue_course_master = {}
    if _vc_csv_path.exists():
        try:
            df_vc = pd.read_csv(str(_vc_csv_path), encoding="utf-8-sig", dtype=str)
            # 列名を load_race.py 側が期待するキー名に正規化
            if "進入コース" in df_vc.columns and "コース" not in df_vc.columns:
                df_vc = df_vc.rename(columns={"進入コース": "コース"})
            # iterrows()は91,491行で遅いため to_dict(orient="records") で一括変換
            for row in df_vc.to_dict(orient="records"):
                name   = str(row.get("選手名", "") or "").strip()
                kaijo  = str(row.get("会場名", "") or "").strip()
                course = str(row.get("コース",  "") or "").strip()
                if not name or not kaijo or not course:
                    continue
                key = (name, kaijo, course)
                venue_course_master[key] = row
                # 5文字名の4文字エイリアス
                if len(name) == 5:
                    alias_key = (name[:4], kaijo, course)
                    if alias_key not in venue_course_master:
                        venue_course_master[alias_key] = row
            print(f"  📍 会場別コースマスタ読込（CSV）: {len(venue_course_master):,}件")
        except Exception as e:
            print(f"  ⚠️  会場別コースマスタCSV読み込み失敗、Excelにフォールバックします: {e}")
            venue_course_master = {}

    # CSVが存在しないか読み込み失敗 → Excelシートから読む（従来通り）
    if not venue_course_master:
        if "会場別コースマスタ" in wb.sheetnames:
            ws_vc = wb["会場別コースマスタ"]
            vc_rows = list(ws_vc.iter_rows(values_only=True))
            header_vc = None
            for i, row in enumerate(vc_rows):
                if row and row[0] == "選手名":
                    header_vc = i
                    break
            if header_vc is not None:
                headers_vc = vc_rows[header_vc]
                for row in vc_rows[header_vc+1:]:
                    if not row or row[0] is None:
                        continue
                    d = dict(zip(headers_vc, row))
                    name   = str(d.get("選手名", "")).strip()
                    kaijo  = str(d.get("会場名", "")).strip()
                    course = str(d.get("コース", "")).strip()
                    key = (name, kaijo, course)
                    venue_course_master[key] = d
                    if len(name) == 5:
                        alias_key = (name[:4], kaijo, course)
                        if alias_key not in venue_course_master:
                            venue_course_master[alias_key] = d
            print(f"  📍 会場別コースマスタ読込（Excel）: {len(venue_course_master):,}件")
        else:
            print("  ⚠️  会場別コースマスタ未作成。update_master.py を実行してください。")


    # ── 展開別残存マスタ読み込み ─────────────────────────────────────────────
    # 会場別: キー=(会場名, 決まり手, 1着コース, 進入コース) → row_dict
    # 全国版: キー=(決まり手, 1着コース, 進入コース)         → row_dict
    # ※ 旧設計では (決まり手, 1着コース) をキーとしていたため、同一シナリオの
    #   進入コース別行（5行）が最後の1行しか残らなかった。キーに進入コースを
    #   追加して全行を保持するよう修正（v6）。
    tenkai_venue_master    = {}
    tenkai_national_master = {}
    for csv_path, master_dict, key_cols in [
        (_tenkai_venue_path,  tenkai_venue_master,    ["会場名", "決まり手", "1着コース", "進入コース"]),
        (_tenkai_nat_path,    tenkai_national_master, ["決まり手", "1着コース", "進入コース"]),
    ]:
        if csv_path.exists():
            try:
                _df = pd.read_csv(str(csv_path), encoding="utf-8-sig", dtype=str)
                # iterrows()は遅いため to_dict(orient="records") で一括変換
                for row in _df.to_dict(orient="records"):
                    key = tuple(str(row.get(c, "") or "").strip() for c in key_cols)
                    master_dict[key] = row
                print(f"  📊 展開別残存マスタ読込: {csv_path.name} {len(master_dict):,}件")
            except Exception as e:
                print(f"  ⚠️  展開別残存マスタ読み込み失敗 ({csv_path.name}): {e}")

    # ── CSVが存在しない場合はExcelシートから直接読み込む（v6フォールバック）──
    # update_master.py 未実行の環境でもマスタが使えるよう、Excelシートを優先参照する。
    # CSVが既に読み込まれている場合はスキップ。
    _TENKAI_SHEET_MAP = [
        ("展開別残存_全国",  tenkai_national_master, ["決まり手", "1着コース", "進入コース"]),
        ("展開別残存_会場別", tenkai_venue_master,    ["会場名", "決まり手", "1着コース", "進入コース"]),
    ]
    for sheet_name, master_dict, key_cols in _TENKAI_SHEET_MAP:
        if master_dict:
            continue  # CSV読み込み済みはスキップ
        if sheet_name not in wb.sheetnames:
            continue
        try:
            ws_tk = wb[sheet_name]
            tk_rows = list(ws_tk.iter_rows(values_only=True))
            # ヘッダ行を探す（最初の非Noneセルが「決まり手」か「会場名」の行）
            hdr_idx = None
            for i, row in enumerate(tk_rows):
                cells = [c for c in row if c is not None]
                if cells and str(cells[0]).strip() in ("決まり手", "会場名"):
                    hdr_idx = i
                    break
            if hdr_idx is None:
                continue
            headers_tk = [str(h).strip() if h is not None else "" for h in tk_rows[hdr_idx]]
            for row in tk_rows[hdr_idx + 1:]:
                if not row or all(c is None for c in row):
                    continue
                d = {headers_tk[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
                # 必須キー列が欠けている行はスキップ
                if any(not d.get(c) for c in key_cols):
                    continue
                key = tuple(d.get(c, "") for c in key_cols)
                master_dict[key] = d
            print(f"  📊 展開別残存マスタ(Excel): {sheet_name} {len(master_dict):,}件")
        except Exception as e:
            print(f"  ⚠️  展開別残存マスタ(Excel)読み込み失敗 ({sheet_name}): {e}")

    # ── グレード別マスタのブレンド処理 ────────────────────────────────────────
    # race_grade が G1/SG または G2/G3 の場合、一般戦マスタとグレードマスタを
    # G1出走数に応じた比率でブレンドして course_master / venue_course_master を補正する。
    #
    # ブレンド比率（G1/G2出走数ベース）:
    #   50走以上  → 一般0.6 : グレード0.4
    #   20〜49走  → 一般0.7 : グレード0.3
    #   20走未満  → 一般0.9 : グレード0.1
    #   データなし → 一般1.0 : グレード0.0（一般戦マスタのみ）
    if race_grade != "一般":
        # ══════════════════════════════════════════════════════════════
        # 指数ごとのブレンド方針
        # ══════════════════════════════════════════════════════════════
        #
        # ❌ ブレンドしない（グレードマスタが閾値以上あればグレード優先、なければ一般戦をそのまま使う）
        #   1C1着率・3連対率・イン逃げ率・決まり手%
        #   → 相手レベルで意味が全く変わる。一般戦の「逃げ率」とG1の「逃げ率」は別物。
        #
        # ✅ ブレンドしてよい（体の癖・直近調子など相手によらない指数）
        #   STタイミング・フォーム指数・時系列補正値（直近調子）
        #   → 選手固有の傾向なのでグレードに関係なく参考になる。
        #
        # 閾値: グレード出走数 >= 30走 → グレードマスタをそのまま使用
        #        グレード出走数 <  30走 → 一般戦マスタをそのまま使用（補正なし）
        #                                  ＋「グレード初出場注意」フラグ
        # ══════════════════════════════════════════════════════════════

        # ブレンドしてよいキー（STのみ。フォーム指数は player_master 側なので除外）
        _BLEND_KEYS_COURSE = [
            "コース別平均ST",
            "時系列補正1着率", "時系列補正3連対率",  # 直近調子は参考にする
        ]
        _BLEND_KEYS_VC = [
            "平均ST",
            "時系列補正1着率", "時系列補正3連対率",
        ]

        # グレードマスタを優先使用する閾値（これ以上の出走数があればグレードデータを信頼）
        _GRADE_TRUST_THRESHOLD = 30

        def _blend_ratio(n_grade):
            """
            G1出走数 → (一般戦比率, グレード比率)
            ブレンド対象は ST・時系列補正のみ。
            閾値未満は一般戦100%（グレードデータ不足）。
            閾値以上はST等をわずかにグレード側に寄せる。
            """
            try:
                n = int(float(n_grade)) if n_grade not in (None, "", "nan") else 0
            except (ValueError, TypeError):
                n = 0
            if n >= _GRADE_TRUST_THRESHOLD:
                return 0.4, 0.6   # ST等：一般4割・グレード6割
            else:
                return 1.0, 0.0   # データ不足：一般戦100%

        def _grade_trusted(n_grade):
            """グレードマスタを信頼できるか（閾値以上の出走数があるか）"""
            try:
                n = int(float(n_grade)) if n_grade not in (None, "", "nan") else 0
            except (ValueError, TypeError):
                n = 0
            return n >= _GRADE_TRUST_THRESHOLD

        def _safe_float(v):
            try:
                return float(v) if v not in (None, "", "nan") else None
            except (ValueError, TypeError):
                return None

        def _blend_val(v_ippan, v_grade, w_ippan, w_grade):
            """2値をブレンド。どちらかがNoneなら存在する方を使う"""
            vi = _safe_float(v_ippan)
            vg = _safe_float(v_grade)
            if vi is not None and vg is not None:
                return round(vi * w_ippan + vg * w_grade, 4)
            return vi if vi is not None else vg

        # ── 一般戦コース別マスタを追加読み込み（ブレンド用） ──────────────
        _ippan_course_master = {}
        try:
            _ws_ippan = wb[SHEET_MASTER]
            _rows_i   = list(_ws_ippan.iter_rows(values_only=True))
            _hdr_i    = next((i for i, r in enumerate(_rows_i) if r and r[0] == "選手名"), None)
            if _hdr_i is not None:
                _hdrs_i = _rows_i[_hdr_i]
                for _row in _rows_i[_hdr_i + 1:]:
                    if not _row or _row[0] is None:
                        continue
                    _d = dict(zip(_hdrs_i, _row))
                    _nm = str(_d.get("選手名", "")).strip()
                    _cs = str(_d.get("コース",  "")).strip()
                    _ippan_course_master[(_nm, _cs)] = _d
                    if len(_nm) == 5:
                        _ak = (_nm[:4], _cs)
                        if _ak not in _ippan_course_master:
                            _ippan_course_master[_ak] = _d
            print(f"  📊 一般戦コース別マスタ（ブレンド用）: {len(_ippan_course_master):,}件")
        except Exception as _e:
            print(f"  ⚠️  一般戦マスタ読み込み失敗（ブレンドなしで続行）: {_e}")

        # ── 一般戦 venue_course_master をCSVから追加読み込み ──────────────
        _ippan_vc_master = {}
        try:
            if VENUE_COURSE_CSV.exists():
                _df_vc_i = pd.read_csv(str(VENUE_COURSE_CSV), encoding="utf-8-sig", dtype=str)
                if "進入コース" in _df_vc_i.columns and "コース" not in _df_vc_i.columns:
                    _df_vc_i = _df_vc_i.rename(columns={"進入コース": "コース"})
                for _row in _df_vc_i.to_dict(orient="records"):
                    _nm  = str(_row.get("選手名", "") or "").strip()
                    _kj  = str(_row.get("会場名", "") or "").strip()
                    _cs  = str(_row.get("コース",  "") or "").strip()
                    if not _nm or not _kj or not _cs:
                        continue
                    _ippan_vc_master[(_nm, _kj, _cs)] = _row
                    if len(_nm) == 5:
                        _ak = (_nm[:4], _kj, _cs)
                        if _ak not in _ippan_vc_master:
                            _ippan_vc_master[_ak] = _row
            print(f"  📍 一般戦会場別コースマスタ（ブレンド用）: {len(_ippan_vc_master):,}件")
        except Exception as _e:
            print(f"  ⚠️  一般戦会場別コースマスタ読み込み失敗: {_e}")

        # ── course_master を閾値切り替え方式で更新 ─────────────────────────
        # 閾値以上（グレード信頼）:
        #   ❌ 混ぜない指数（1着率・3連対率・決まり手%・イン逃げ率）→ グレードマスタをそのまま使用
        #   ✅ 混ぜてよい指数（ST・時系列補正）→ 一般戦とグレードをブレンド
        # 閾値未満（データ不足）:
        #   全指数 → 一般戦マスタで上書き ＋「グレード初出場注意」フラグを付与
        _blended_count = 0
        _ippan_fallback_count = 0
        for _key, _gd in course_master.items():
            _id = _ippan_course_master.get(_key)
            _n_grade = _safe_float(_gd.get("出走数"))
            _trusted = _grade_trusted(_n_grade)

            if not _trusted:
                # データ不足 → 一般戦マスタで完全上書き（フラグ付き）
                if _id is not None:
                    _fallback = dict(_id)
                    _fallback["_grade_data_shortage"] = True   # 予想時に注意フラグとして参照可
                    course_master[_key] = _fallback
                    _ippan_fallback_count += 1
                # 一般戦データもなければグレードマスタのまま（初出場選手等）
                continue

            # 閾値以上: STのみブレンド、それ以外はグレードマスタ優先
            if _id is None:
                continue  # 一般戦データなし → グレードマスタのまま
            _wi, _wg = _blend_ratio(_n_grade)
            _blended = dict(_gd)  # グレードマスタをベースにコピー
            for _k in _BLEND_KEYS_COURSE:
                _bv = _blend_val(_id.get(_k), _gd.get(_k), _wi, _wg)
                if _bv is not None:
                    _blended[_k] = _bv
            course_master[_key] = _blended
            _blended_count += 1

        print(f"  🔀 course_master: グレード優先={_blended_count:,}件 / 一般戦フォールバック={_ippan_fallback_count:,}件")

        # ── venue_course_master を閾値切り替え方式で更新 ──────────────────
        _blended_vc = 0
        _ippan_fb_vc = 0
        for _key, _gd in venue_course_master.items():
            _id = _ippan_vc_master.get(_key)
            _n_grade = _safe_float(_gd.get("出走数"))
            _trusted = _grade_trusted(_n_grade)

            if not _trusted:
                if _id is not None:
                    _fallback = dict(_id)
                    _fallback["_grade_data_shortage"] = True
                    venue_course_master[_key] = _fallback
                    _ippan_fb_vc += 1
                continue

            if _id is None:
                continue
            _wi, _wg = _blend_ratio(_n_grade)
            _blended = dict(_gd)
            for _k in _BLEND_KEYS_VC:
                _bv = _blend_val(_id.get(_k), _gd.get(_k), _wi, _wg)
                if _bv is not None:
                    _blended[_k] = _bv
            venue_course_master[_key] = _blended
            _blended_vc += 1

        print(f"  🔀 venue_course_master: グレード優先={_blended_vc:,}件 / 一般戦フォールバック={_ippan_fb_vc:,}件")

        # ── グレードマスタにしか存在しない選手はそのまま残す（フォールバック済み）──
        # ── 一般戦にしかいない選手も course_master に追加（出場しない可能性高いが念のため）──
        for _key, _id in _ippan_course_master.items():
            if _key not in course_master:
                course_master[_key] = _id
        for _key, _id in _ippan_vc_master.items():
            if _key not in venue_course_master:
                venue_course_master[_key] = _id

    return course_master, player_master, ininage_master, venue_stats_master, venue_course_master, tenkai_venue_master, tenkai_national_master

# ============================================================
# CSVからレースデータ読み込み
# ============================================================
def load_csv(venue, race_no=None, date_str=None):
    """csv_output/会場名/ サブフォルダ（優先）または scripts/csv_output/ から会場CSVを読み込む"""
    # 会場サブフォルダを優先検索、なければルートにフォールバック
    search_dirs = [CSV_DIR / venue, CSV_DIR]

    def find_files(pattern_name):
        for d in search_dirs:
            files = sorted(glob.glob(str(d / pattern_name)))
            if files:
                return files
        return []

    if date_str:
        files = find_files(f"{venue}_{date_str}.csv")
        if not files:
            print(f"  ❌ 指定日付のCSVが見つかりません: {venue}_{date_str}.csv")
            return None, None
        csv_path = files[0]
        print(f"  📂 日付指定: {csv_path}")
    else:
        files = find_files(f"{venue}_*.csv")
        if not files:
            # 最終手段: csv_output/ 全体から最新を検索
            all_files = sorted(glob.glob(str(CSV_DIR / "**" / "*.csv"), recursive=True))
            if not all_files:
                return None, None
            files = [all_files[-1]]
            venue_detected = os.path.basename(files[0]).split("_")[0]
            print(f"  📂 自動検出: {os.path.basename(files[0])} (会場: {venue_detected})")
        csv_path = files[-1]  # 最新

    date_match = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(csv_path))
    race_date = date_match.group(1) if date_match else datetime.today().strftime("%Y/%m/%d")

    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
    df.columns = df.columns.str.strip()
    for col in df.columns:
        df[col] = df[col].astype(str).str.lstrip("'").str.strip()

    # CSVの「日付」列がある場合はそちらを優先（ファイル名より確実）
    if "日付" in df.columns:
        date_vals = df["日付"].dropna()
        date_vals = date_vals[date_vals.str.strip() != ""]
        if len(date_vals) > 0:
            raw_date = date_vals.iloc[0].strip()
            date_from_csv = re.sub(r"(\d{4})[/\-](\d{2})[/\-](\d{2}).*", r"\1-\2-\3", raw_date)
            if re.match(r"\d{4}-\d{2}-\d{2}", date_from_csv):
                race_date = date_from_csv

    if race_no:
        if "レース番号" in df.columns:
            df = df[df["レース番号"].astype(str) == str(race_no)]
        elif "R" in df.columns:
            df = df[df["R"].astype(str) == str(race_no)]

    return df, race_date

# ============================================================
# 指数計算
# ============================================================
def calc_race_indices(venue, race_no, players, course_master, player_master, ininage_master, venue_stats_master, venue_course_master=None, tenkai_venue_master=None, tenkai_national_master=None):
    """
    1レース分の指数を計算して返す。

    【進入コース前提】
    本システムは「枠なり進入」を前提として分析する。
    CSVに「想定コース」列がある場合はその値を使用するが、
    ない場合・空値の場合は 枠番 = コース として扱う（枠なり進入とみなす）。

    展示後に進入変更が確認された場合は race_judgment["nyujo_henkou"] = True を
    セットすることで _should_skip_race が無条件に見送り推奨を返す。

    改善①  会場別コースマスタ (venue_course_master) を優先参照。
            選手の当該会場での実績があればそちらを使い、なければ全国マスタにフォールバック。
    改善③  ハイブリッド係数を動的計算。
            会場の実績量（会場統計レース数）と選手の当該会場実績量（信頼度）に応じて
            「選手実績 : 会場特性」の比率を動的に調整する。
    """
    results = []
    if venue_course_master is None:
        venue_course_master = {}
    if tenkai_venue_master is None:
        tenkai_venue_master = {}
    if tenkai_national_master is None:
        tenkai_national_master = {}

    # 【問題C修正】_blend_tsをループ外に定義（毎ループの再生成を防ぐ）
    def _blend_ts(ts_val, flat_val, eff_n, full_n=20.0):
        """時系列補正値とフラット平均を有効走数に応じてブレンドして返す。
        eff_n ≥ full_n → 時系列補正100%採用
        eff_n = 0      → フラット平均100%採用"""
        if ts_val is None and flat_val is None:
            return None
        if ts_val is None:
            return flat_val
        if flat_val is None:
            return ts_val
        t = min(eff_n / full_n, 1.0)
        return ts_val * t + flat_val * (1.0 - t)

    for p in players:
        name_raw = str(p.get("選手名","")).strip()
        # 末尾の年齢数字を除去（例: "熊本英一47" → "熊本英一"、"芦澤　望48" → "芦澤　望"）
        name_raw = re.sub(r'\s*\d+\s*$', '', name_raw).strip()
        name = name_raw.replace("　", "").replace(" ", "")  # スペース除去でマスタキーと統一
        # [修正2] lzh_to_csv.py の出力列名は「艇番」。「枠」にも後方互換対応
        waku = str(p.get("枠", p.get("艇番", ""))).strip()
        # 【枠なり前提】CSVに「想定コース」列があればその値を使用。
        # ない場合・空値（nan/None/空文字）の場合は 枠番 = コース として扱う（枠なり進入を想定）。
        # 展示後に進入変更が判明した場合は race_judgment["nyujo_henkou"] = True をセットすること。
        _course_raw = str(p.get("想定コース", "")).strip()
        course = _course_raw if _course_raw not in ("", "nan", "None") else waku

        # ── 改善①: 会場別コースマスタを優先参照 ────────────────────────────
        # 【軽微②改善】ルックアップ関数経由で完全一致→4文字前方一致の順で検索
        # 優先度: 会場別コースマスタ（選手×会場×コース） > 全国コース別マスタ
        vc_key = (name, venue, course)
        vcm = venue_course_master.get(vc_key)
        if vcm is None and len(name) >= 5:
            vcm = venue_course_master.get((name[:4], venue, course))

        cm = _lookup_name_course(course_master, name, course)
        cm = cm or {}
        pm = _lookup_player(player_master, name)
        pm = pm or {}

        # ── 【修正①②④】会場別コースマスタと全国マスタの統合 ────────────────────
        # 【修正①】vcm存在かつvc_trust=0.0 と vcm=None を明示的に区別する。
        # 【修正②】win1_rate は純粋な「選手個人の1着率推定値」のみ。venue_rateは後段で混ぜる。
        # 【修正④】時系列有効走数（eff_n）による時系列補正の信頼制御
        #   eff_n ≥ 20走 → 時系列補正100%採用 / eff_n=0 → フラット平均100%
        #   ※ _blend_ts はループ外で定義済み

        # 全国マスタの有効走数を取得
        cm_eff_n = safe_float(cm.get("時系列有効走数"), 0.0) or 0.0
        _ts  = safe_float(cm.get("時系列補正1着率"))
        _fl  = safe_float(cm.get("1着率"))
        global_win1 = _blend_ts(_ts, _fl, cm_eff_n)
        _ts  = safe_float(cm.get("時系列補正3連対率"))
        _fl  = safe_float(cm.get("3連対率"))
        global_win3 = _blend_ts(_ts, _fl, cm_eff_n)

        if vcm is not None:
            vc_trust_raw = safe_float(vcm.get("信頼度"), 0.0) or 0.0  # 0.0〜1.0
            if vc_trust_raw > 0.0:
                # 会場別実績あり＆信頼度あり → vc_trustに応じて会場別と全国をブレンド
                vc_trust = vc_trust_raw
                # 会場別コースマスタの有効走数で時系列補正を制御
                vc_eff_n = safe_float(vcm.get("時系列有効走数"), 0.0) or 0.0
                _ts  = safe_float(vcm.get("時系列補正1着率"))
                _fl  = safe_float(vcm.get("1着率"))
                vc_win1 = _blend_ts(_ts, _fl, vc_eff_n)
                _ts  = safe_float(vcm.get("時系列補正3連対率"))
                _fl  = safe_float(vcm.get("3連対率"))
                vc_win3 = _blend_ts(_ts, _fl, vc_eff_n)
                if vc_win1 is not None and global_win1 is not None:
                    win1_rate = vc_win1 * vc_trust + global_win1 * (1.0 - vc_trust)
                elif vc_win1 is not None:
                    win1_rate = vc_win1
                else:
                    win1_rate = global_win1
                if vc_win3 is not None and global_win3 is not None:
                    win3_rate = vc_win3 * vc_trust + global_win3 * (1.0 - vc_trust)
                elif vc_win3 is not None:
                    win3_rate = vc_win3
                else:
                    win3_rate = global_win3
            else:
                # vcm存在するが信頼度=0 → 実質データ不足。全国マスタのみ使用
                vc_trust  = 0.0
                win1_rate = global_win1
                win3_rate = global_win3
        else:
            # vcm=None → 会場別実績なし。全国マスタのみ使用
            vc_trust  = 0.0
            win1_rate = global_win1
            win3_rate = global_win3

        gen_win3 = safe_float(pm.get("3連対率\n(一般戦)"))
        avg_st   = safe_float(cm.get("コース別\n平均ST"))
        st_rank  = safe_float(pm.get("ST順位\n(1コース)"))

        # ── 進入コース別ST順位（相性評価のフォールバック用） ──────────────────
        # avg_st（コース別平均ST秒）が取れない選手向けの代替データ。
        # 選手指数マスタには「ST順位\n(Nコース)」が全コース分格納されている。
        # ST順位 = 同レース内での速さ順位（1=最速〜6=最遅）
        # build_jizen_members 側で avg_st が None のとき推定ST秒に変換して使用。
        course_st_rank = safe_float(
            pm.get(f"ST順位\n({course}コース)") or pm.get(f"ST順位({course}コース)")
        )

        # 決まり手（会場別があればそちらを優先）
        kimari = _build_kimari(vcm if vcm else cm)

        # ── データ不足チェック ──────────────────────────────────────────────
        cm_missing = not cm
        cm_count   = safe_float(cm.get("出走数"), 0)
        cm_scarce  = (not cm_missing) and (cm_count is not None) and (cm_count < 5)
        pm_missing = not pm
        pm_count   = safe_float(pm.get("総出走数"), 0)
        pm_scarce  = (not pm_missing) and (pm_count is not None) and (pm_count < 10)
        data_missing = cm_missing or cm_scarce or pm_missing or pm_scarce
        missing_reasons = []
        if cm_missing:
            missing_reasons.append(f"コース{course}実績なし")
        elif cm_scarce:
            missing_reasons.append(f"コース{course}実績{int(cm_count)}走")
        if pm_missing:
            missing_reasons.append("選手マスタ未登録")
        elif pm_scarce:
            missing_reasons.append(f"総実績{int(pm_count)}走")
        missing_reason_str = " / ".join(missing_reasons) if missing_reasons else ""

        # FLY数・出遅れ数・FLY経過日数（選手指数マスタ pm から取得）
        # 【修正④】FLY経過日数を使って影響度を精密判定
        # 旧: FLY数≥2→高、≥1→中 のみ → FLY明け直後(60日)と1年後(365日)が同判定
        # 新: FLY経過日数を加味:
        #   経過日数 < 90  日: 出場停止明け直後 → 判定を1段階引き上げ（中→高等）
        #   経過日数 < 180 日: 影響残存期間    → 素直に使用
        #   経過日数 ≥ 180 日: 影響ほぼ消滅    → FLY数に依らず「低」
        _fly_count    = int(safe_float(pm.get("FLY数"),    0) or 0)
        _late_count   = int(safe_float(pm.get("出遅れ数"), 0) or 0)
        _fly_days_raw = pm.get("FLY経過日数")
        _fly_days     = safe_float(_fly_days_raw) if _fly_days_raw not in (None, "", "nan") else None

        if _fly_count == 0:
            _fly_label = "低"
        elif _fly_days is not None:
            if _fly_days >= 180:
                # FLY後180日超 → 影響ほぼ消滅
                _fly_label = "低"
            elif _fly_days < 90:
                # 出場停止明け直後 → 1段階引き上げ
                _fly_label = "高"  # (FLY1回でも高)
            else:
                # 90〜180日: 通常判定
                _fly_label = "高" if _fly_count >= 2 else "中"
        else:
            # 経過日数不明（update_master.py 更新前の旧データ）→ 旧来判定にフォールバック
            _fly_label = "高" if _fly_count >= 2 else "中"

        results.append({
            "waku":       waku,
            "name":       name_raw,  # 表示用（スペースあり）
            "name_norm":  name,      # マスタ検索用（スペースなし）
            "kumi":       str(p.get("級別", p.get("組",""))).strip(),
            "motor2":     str(p.get("モーター2連率", p.get("M2率", p.get("モータ2連","")))).strip(),
            "course":     course,
            "win1_rate":      win1_rate,
            "win3_rate":      win3_rate,
            "avg_st":         avg_st,
            "st_rank":        st_rank,
            "course_st_rank": course_st_rank,   # 進入コース別ST順位（フォールバック用）
            "kimari":     kimari,
            "kosetsu":    str(p.get("今節成績","")).strip(),
            "tenji_time": safe_float(p.get("展示タイム", p.get("展示", p.get("展示ST","")))),
            "raw_cm":     cm,
            "raw_pm":     pm,
            "raw_vcm":    vcm or {},     # 会場別コースマスタ（デバッグ用）
            "vc_trust":   vc_trust,      # 会場別実績の信頼度（0〜1）
            "data_missing":   data_missing,
            "missing_reason": missing_reason_str,
            "fly_count":  _fly_count,   # FLY数（0以上の整数）
            "fly_label":  _fly_label,   # F/ST影響ラベル（高/中/低）
            "late_count": _late_count,  # 出遅れ数
        })

    # ── 【修正②⑦】動的ハイブリッド係数: 「選手実績 vs 会場特性」のブレンド ──────────
    # win1_rate は上段で「会場別マスタ×全国マスタ」をブレンド済みの純粋な選手個人実績値。
    # 「その選手実績値をどれだけ信頼するか」を cm_count と vc_trust の両方で決定する。
    #
    # 【修正⑦】vc_trust を後段にも引き継いで w_player 上限を拡張する。
    #   修正②で「上段ブレンドに vc_trust を消費」したが、それは
    #   「会場別 vs 全国の個人実績ブレンド」への使用であり、
    #   「個人実績全体 vs 会場特性」のブレンド比率を決める後段とは別の軸。
    #   → 二重混合ではなく「直交した2つの調整」なので後段でも参照してよい。
    #
    # w_player の計算式（修正⑦版）:
    #   cm_trust = min(cm_count / 30, 1.0) × 0.60   全国出走数による基礎信頼度（最大0.60）
    #   vc_bonus = vc_trust × 0.30                   会場別実績が十分なら最大+0.30
    #   w_player = min(cm_trust + vc_bonus, 0.90)    上限0.90（会場特性を最低10%保証）
    #
    # 具体例:
    #   会場別実績なし(vc=0)  全国30走 → 0.60+0.00 = 0.60
    #   会場別実績あり(vc=1)  全国30走 → 0.60+0.30 = 0.90（上限）
    #   会場別実績あり(vc=0.5) 全国10走 → 0.20+0.15 = 0.35
    #   全実績なし(vc=0, cm=0)         → 0.00+0.00 = 0.00 → 会場特性100%
    #
    # ※ キー名正規化は load_masters() 内で実施済み（"1コース1着率"等のエイリアスを追加）

    COURSE_AVG_WIN = {
        "1": 0.555, "2": 0.137, "3": 0.134,
        "4": 0.111, "5": 0.066, "6": 0.021,
    }

    vs = venue_stats_master.get(venue, {})

    # ── 【修正⑤】会場のコース別1着率: R番号別と全体平均を加重平均でブレンド ────────
    # 【旧方式の問題】R番号別1着率が存在すれば無条件に優先していた。
    #   R別はサンプルが少ない（例: 同会場の同R番号は年間50〜200レース程度）ため
    #   たまたまの偏りをシグナルと誤認する「過学習的なノイズ混入」が起きていた。
    #
    # 【修正⑥】R別1着率とコース全体平均のブレンド比率を動的化。
    # 旧方式: W_RC = 0.30 固定 → サンプルが少ない会場でもRC別を30%参照してノイズが混入。
    # 新方式: 会場の総レース数（= 統計の信頼性）に応じてW_RCを動的決定。
    #   計算式: W_RC = clip(レース数 / 3000, 0.05, 0.30)
    #     レース数3000以上（約12年分）→ W_RC = 0.30（上限：RC別を最大30%参照）
    #     レース数1500程度（約6年）  → W_RC = 0.15
    #     レース数300以下（約1年）   → W_RC = 0.05（ほぼ全体平均のみ）
    #   根拠: R別は年間約120〜150レースのデータ。
    #   統計的に安定するには最低5〜8年(600〜1200件)が目安。
    #   上限を0.30に抑えてRC別に過剰依存しないよう設計。
    _venue_race_count = safe_float(vs.get("レース数"), 0) or 0
    W_RC = float(min(max(_venue_race_count / 3000.0, 0.05), 0.30))

    venue_course_rate = {}
    for c in range(1, 7):
        rc_key     = f"{c}C_{race_no}R1着率"   # キー名正規化済み
        course_key = f"{c}コース1着率"          # キー名正規化済み
        nat_avg    = COURSE_AVG_WIN[str(c)]
        rc_val     = safe_float(vs.get(rc_key))
        course_val = safe_float(vs.get(course_key))
        if rc_val is not None and course_val is not None:
            # 両方存在 → 加重平均（R別30%、全体70%）
            blended = rc_val * W_RC + course_val * (1.0 - W_RC)
        elif course_val is not None:
            blended = course_val
        elif rc_val is not None:
            blended = rc_val
        else:
            blended = nat_avg
        venue_course_rate[str(c)] = blended

    for r in results:
        course     = str(r["course"])
        venue_rate = venue_course_rate.get(course, COURSE_AVG_WIN.get(course, 0.10))

        # 【修正⑦】全国出走数 + 会場別実績信頼度の両方でw_playerを決定（上限0.90）
        cm_count_w = safe_float(r.get("raw_cm", {}).get("出走数"), 0) or 0
        vc_trust   = r.get("vc_trust", 0.0)
        cm_trust   = min(cm_count_w / 30.0, 1.0) * 0.60   # 全国出走数による基礎信頼度（最大0.60）
        vc_bonus   = vc_trust * 0.30                        # 会場別実績ボーナス（最大+0.30）
        w_player   = min(cm_trust + vc_bonus, 0.90)        # 上限0.90（会場特性を最低10%保証）
        w_venue    = 1.0 - w_player                         # 0.10〜1.0

        # ── 【新修正: 全国平均比スケール換算（vc_trust対応版）】────────────────────────
        # 問題: win1_rate のスケールが vcm の有無によって変わる。
        #   - vcm=None:  win1_rate は全国コース別実績（全国スケール）
        #   - vcm有り:   win1_rate = vc_win1 * vc_trust + global_win1 * (1-vc_trust)
        #               vc_win1 は「当該会場での絶対的1着率（会場スケール）」
        #               → win1_rate は会場スケールと全国スケールの混合
        #
        # 解決: win1_rate のスケールに合わせた基準値 base でratio計算する。
        #   base = venue_rate * vc_trust + nat_avg_c * (1 - vc_trust)
        #     vc_trust=0.0 → base = nat_avg_c（全国比換算のみ）
        #     vc_trust=1.0 → base = venue_rate → ratio=win1/venue_rate
        #                    win1_scaled = venue_rate * ratio = win1（スケール変換なし）
        #
        # 効果:
        #   vcm無し・荒れ会場: 全国比換算で過大評価を解消（戸田1号艇 54.8%→50.9%）
        #   vcm有り・荒れ会場: 会場スケールの実績をそのまま反映（過小評価なし）
        #   全国平均会場: 変化なし
        #   ratio は [0.25, 4.0] にクリップして外れ値を防ぐ。
        _RATIO_MIN, _RATIO_MAX = 0.25, 4.0
        nat_avg_c = COURSE_AVG_WIN.get(course, 0.10)  # コースの全国平均1着率
        # win1_rate のスケールに合わせた基準値（vc_trust で全国↔会場を補間）
        _ratio_base = venue_rate * vc_trust + nat_avg_c * (1.0 - vc_trust)
        _ratio_base = max(_ratio_base, 1e-6)  # ゼロ除算防止

        if r["win1_rate"] is not None and r["win1_rate"] > 0:
            ratio = max(_RATIO_MIN, min(r["win1_rate"] / _ratio_base, _RATIO_MAX))
            win1_scaled = venue_rate * ratio  # 会場スケールでの個人実績推定値
            r["_raw_win"] = win1_scaled * w_player + venue_rate * w_venue
        elif r["win1_rate"] == 0.0:
            # 【問題B修正】出走したが1着なし → 会場特性をw_venue分だけ使用。
            r["_raw_win"] = venue_rate * w_venue
        else:
            # データなし → 会場特性100%
            r["_raw_win"] = venue_rate

    # ── 【修正④】Laplace smoothingフロア: 均等フロアで正規化歪みを防ぐ ────────────
    # 【旧方式の問題】フロア値を「全国平均×10%」にしていたため
    #   6コースのフロア ≈ 0.0021（0.2%）と極端に小さく、
    #   この値が正規化の分母に混入することで1〜5コースの確率が不当に圧縮されていた。
    #
    # 【新方式】全コース共通で「全国平均の最小値（6コース=2.1%）の50%」= 0.0105 を下限とする。
    #   これにより「確率0の艇を救済する」目的は維持しつつ、
    #   極小フロアによる正規化への影響を実質ゼロに近づける。
    COURSE_WIN_FLOOR_UNIFORM = 0.021 * 0.50   # ≈ 0.0105（全コース共通）
    for r in results:
        r["_raw_win"] = max(r["_raw_win"], COURSE_WIN_FLOOR_UNIFORM)

    # ── 【修正③⑥・問題A修正】キャリブレーション補正（_raw_winスケール対応版）──────────
    # 【問題A】修正③で「_raw_winに補正を適用」に変えたが、
    #   breakpointsの閾値が旧方式（rel_win1 = 0〜1スケール）のままだった。
    #   _raw_win の実際の範囲は 1号艇:0.02〜0.60、外コース:0.01〜0.10 であり、
    #   旧閾値 p=0.35 未満は補正なし → 2〜6号艇が全て補正なしになっていた。
    #
    # 【修正】breakpointsを _raw_win の実際のスケールに合わせて再設計する。
    #   目標: バックテストCal誤差0.0698を反映しつつ全艇に段階的に補正を掛ける。
    #   設計方針:
    #     ・p < 0.10 → 補正なし（6コース等の極小確率帯は信頼できる）
    #     ・p = 0.15 → scale = 0.990（2%の緩い補正）
    #     ・p = 0.25 → scale = 0.975
    #     ・p = 0.35 → scale = 0.960
    #     ・p = 0.45 → scale = 0.945
    #     ・p = 0.55 → scale = 0.930（1号艇強いケースの核心補正 ≒ Cal誤差0.0698相当）
    #     ・p ≥ 0.65 → scale = 0.920（上限なし: _raw_winは最大0.60程度のため実質到達しない）
    #
    # 【修正⑥連携】荒れ会場では緩和スケール(relax)を適用:
    #   relax = clip(venue_c1_rate / 0.555, 0.70, 1.0)
    #   戸田(c1≈0.430) → relax=0.775 → 補正を22.5%緩和

    _venue_c1_rate = venue_course_rate.get("1", 0.555)
    _calib_relax   = float(min(max(_venue_c1_rate / 0.555, 0.70), 1.0))

    def _calibrate(p_raw, relax=1.0):
        """_raw_winスケール(0〜1)にキャリブレーション補正を適用して返す。
        relax: 1.0=通常補正、<1.0=補正を緩和（荒れ会場向け）"""
        if p_raw is None:
            return None
        p = float(p_raw)
        # _raw_winスケール向けbreakpoints（旧の0.35〜1.01から0.10〜1.01に再設計）
        breakpoints = [
            (0.10, 1.000),
            (0.15, 0.990),
            (0.25, 0.975),
            (0.35, 0.960),
            (0.45, 0.945),
            (0.55, 0.930),
            (0.65, 0.920),
            (1.01, 0.920),
        ]
        if p < breakpoints[0][0]:
            return p  # 極小確率帯は補正なし
        for i in range(len(breakpoints) - 1):
            x0, s0 = breakpoints[i]
            x1, s1 = breakpoints[i + 1]
            if x0 <= p < x1:
                t = (p - x0) / (x1 - x0)
                scale_full = s0 + t * (s1 - s0)
                scale = 1.0 - (1.0 - scale_full) * relax
                return p * scale
        scale_full = breakpoints[-1][1]
        scale = 1.0 - (1.0 - scale_full) * relax
        return p * scale

    # _raw_win は 0〜1 程度のスケール。_calibrate に relax（荒れ会場緩和係数）を渡す。
    for r in results:
        r["_raw_win"] = _calibrate(r["_raw_win"], relax=_calib_relax) or r["_raw_win"]

    total_raw = sum(r["_raw_win"] for r in results)
    for r in results:
        if total_raw > 0:
            r["rel_win1"] = r["_raw_win"] / total_raw * 100
        else:
            r["rel_win1"] = None

    # rel_win1_cal は rel_win1 と同値（補正済み値の別名として下流コードと互換性を保つ）
    for r in results:
        r["rel_win1_cal"] = r["rel_win1"]
    
    # 3連対率：絶対評価（コース別実績をそのまま%表示）
    # 相対化するとメンバーレベルが見えなくなるため絶対値を使用
    for r in results:
        if r["win3_rate"] is not None:
            r["abs_win3"] = round(r["win3_rate"] * 100, 1)  # 例: 0.625 → 62.5%
        else:
            r["abs_win3"] = None
    
    # ══════════════════════════════════════════════════════════════════════════
    # 本命記号（多面評価スコアによる総合印）◎→○→▲→△
    # ──────────────────────────────────────────────────────────────────────────
    # 【旧方式の問題】rel_win1（オリジナル1着率）のみで機械的に順位付け
    #   → FLYリスクが高い艇でも◎がつく
    #   → 飛びシナリオ高確率なのに1号艇◎という矛盾が起きる
    #
    # 【新方式: 総合印スコア】
    #   基礎点  = rel_win1 × 0.60
    #   FLY高  : -8pt  / FLY中: -4pt
    #   イン逃げ◎(1号艇): +8pt / ○: +4pt / 空白: -5pt  ← jizen確定後に適用
    #   飛び相性◎(tobi_prob≥55の2〜6号艇): +10pt / ○: +5pt ← jizen確定後に適用
    #
    # 【印の意味】◎>○>▲>△
    #   ◎: 総合スコア1位（本命）  ○: 2位（対抗）
    #   ▲: 3位（単穴）           △: 4位（穴）
    #
    # ※ここではFLYペナルティのみ適用した仮印を付ける。
    #   jizen_eval確定後に main() で _apply_jizen_honmei() を呼んで最終確定する。
    # ══════════════════════════════════════════════════════════════════════════

    # _calc_honmei_score / _apply_jizen_honmei はトップレベルで定義（下記参照）

    # 仮印（jizen未確定。相互作用モデルで計算し、jizen確定後に _apply_jizen_honmei で上書き）
    # venue_stats をここで構築して相互作用モデルに渡す
    _venue_stats_pre = _calc_venue_stats(venue_stats_master, venue)
    _honmei_scores_pre = [
        (i, _calc_honmei_score(r, 0, jizen_ev=None, results_ctx=results,
                               venue_stats=_venue_stats_pre,
                               race_judgment=None))   # 仮印計算時点ではrace_judgment未確定
        for i, r in enumerate(results)
        if r["rel_win1"] is not None
    ]
    _honmei_scores_pre.sort(key=lambda x: x[1], reverse=True)

    honmei_map = {0: "◎", 1: "○", 2: "▲", 3: "△"}  # ◎>○>▲>△
    for rank, (idx, _) in enumerate(_honmei_scores_pre[:4]):
        results[idx]["honmei"] = honmei_map[rank]
    for r in results:
        if "honmei" not in r:
            r["honmei"] = " "

    # _apply_jizen_honmei はトップレベルで定義（下記参照）

    # ── 新機能④：展示タイム偏差値（レース内相対評価） ──────────────────────
    # 【軽微①注意】展示タイムは当日展示航走後にしか取得できない。
    # 前日予想CSV（締め切り前スクレイプ）では値が存在しないため tenji_hensa は None になる。
    # 表示側では None の場合「-（前日）」と表示し、当日版で上書きされることを明示する。
    tenji_vals = [r["tenji_time"] for r in results if r.get("tenji_time") is not None]
    _has_tenji_data = len(tenji_vals) >= 2  # 展示タイムデータが揃っているか
    if _has_tenji_data:
        t_mean = statistics.mean(tenji_vals)
        t_stdev = statistics.stdev(tenji_vals) if len(tenji_vals) > 1 else 1
        for r in results:
            t = r.get("tenji_time")
            if t is not None and t_stdev > 0:
                # 競艇の展示タイムは速い（小さい）ほど良いので逆転
                r["tenji_hensa"] = round(50 - (t - t_mean) / t_stdev * 10, 1)
            else:
                r["tenji_hensa"] = None
    else:
        # 前日時点では展示タイム未取得 → 明示的にNoneを設定
        for r in results:
            r["tenji_hensa"] = None
        if not tenji_vals:
            pass  # 前日出力では正常（展示前）

    # 想定スリット（平均STでソート）
    sortable = [(r["waku"], r["avg_st"]) for r in results if r["avg_st"] is not None]
    sortable.sort(key=lambda x: x[1])
    slit = "-".join([s[0] for s in sortable])
    if not slit:
        slit = "-".join([r["waku"] for r in results])
    
    # ══════════════════════════════════════════════════════════════════════════
    # circle_pct（2着優位度）・idx3（3着指数）
    # ──────────────────────────────────────────────────────────────────────────
    # 【設計思想】
    #   「この会場でこの出走メンバーが戦ったとき何が起きるか」を計算する。
    #   枠番の固定統計ではなく、各艇の決まり手・ST・1号艇との関係性から
    #   相互作用モデルでスコアを算出する。
    #
    # ──────────────────────────────────────────────────────────────────────────
    # ■ circle_pct（2着優位度）：「イン逃げ時に2着を取れる攻め手の強さ」
    #
    #   コース別に異なる攻め手を評価し、1号艇との具体的な関係性で補正する。
    #
    #   2枠: 差し特化コース
    #     差し% × ST優位(vs 1号艇) × 1号艇の被差し脆弱性
    #
    #   3枠: まくり差し主体
    #     (まくり差し% × 1.2 + まくり%) × ST優位
    #     × 2枠差し力ペナルティ（2枠が強いと進路を塞がれる）
    #
    #   4枠: まくり主体
    #     (まくり% + まくり差し%) × ST優位
    #     × 3枠壁ペナルティ（3枠のまくり力が強いと被る）
    #
    #   5・6枠: 展開待ち・まくり
    #     (まくり% + まくり差し%) × ST優位 × 外枠減衰
    #
    #   ※ 全艇の絶対スコアを算出し、合計100%に正規化して circle_pct とする。
    #   ※ 選手実績データが不足する場合は全国コース別平均でフォールバック。
    #
    # ──────────────────────────────────────────────────────────────────────────
    # ■ idx3（3着指数）：「イン逃げ時に3着に残る固有の力」
    #
    #   純3着率（3着以内率 - 2着率）を主軸に据える。
    #   「2着を取りこぼしても3着に粘り込む能力」を選手個人の実績から評価する。
    #
    #   ベーススコア = 純3着率(選手実績) × trust + 会場枠別純3着率 × (1-trust)
    #   → 選手実績が豊富なほど個人差が出る。データ不足なら会場値に寄せる。
    #   → 最大値=100にスケーリング（ただし枠番固定にならないよう分散を確保）
    # ══════════════════════════════════════════════════════════════════════════

    venue_frame   = ininage_master.get(venue, {})            # {枠番str: 会場2着率float}
    venue_3rd_map = ininage_master.get(venue, {}).get("_3rd", {})  # {枠番str: 会場3着以内率float}
    MIN_ININAGE_COUNT = 5

    # ── 全国コース別平均（フォールバック用） ──────────────────────────────────
    NATIONAL_2ND  = {"2": 0.35, "3": 0.33, "4": 0.18, "5": 0.09, "6": 0.05}
    NATIONAL_3RD_PURE = {"2": 0.08, "3": 0.14, "4": 0.16, "5": 0.13, "6": 0.09}
    # 純3着率の全国平均（3着以内 - 2着）:
    #   2枠は差しで2着か沈むか → 純3着が少ない
    #   3〜5枠は流れ込みが多い → 純3着が比較的多い

    # ── 1号艇の情報を事前取得 ────────────────────────────────────────────────
    res1   = next((r for r in results if r["waku"] == "1"), None)
    cm1    = res1.get("raw_cm", {}) if res1 else {}
    st1    = res1.get("avg_st") if res1 else None

    # 1号艇の被差し脆弱性（0〜1、高いほど差されやすい）
    sasar_vuln = safe_float(cm1.get("差され%"), 0) or 0.0    # 被差し%
    makur_vuln = safe_float(cm1.get("捲られ%"), 0) or 0.0   # 被まくり%
    nige_pct1  = safe_float(_get_cm_val(cm1, "逃げ%"), 0.6) or 0.6
    # 被差し脆弱性スコア: 被差し%を主軸にし、逃げ%が低いほど補強
    vuln_sashi  = max(0.0, min(1.0, sasar_vuln * 0.7 + (1.0 - nige_pct1) * 0.3))
    vuln_makuri = max(0.0, min(1.0, makur_vuln * 0.7 + (1.0 - nige_pct1) * 0.3))

    # ── 全艇のST平均（ST優位スコア計算の基準） ───────────────────────────────
    st_vals = [r["avg_st"] for r in results if r.get("avg_st") is not None]
    st_mean = sum(st_vals) / len(st_vals) if st_vals else 0.15

    def _st_advantage(avg_st, reference=None):
        """
        avg_st が reference（デフォルト: メンバー平均ST）より速い（小さい）ほど高い。
        差±0.05秒を基準に [-1, +1] → スコア [0.5〜1.5] に変換。
        """
        ref = reference if reference is not None else st_mean
        if avg_st is None:
            return 1.0   # データなし → 中立
        diff = ref - avg_st   # 正 = 自分が速い
        return max(0.5, min(1.5, 1.0 + diff / 0.05))

    def _st_advantage_vs1(avg_st):
        """1号艇とのST比較。1号艇より速いほど2着争いで有利。"""
        return _st_advantage(avg_st, reference=st1)

    # ══════════════════════════════════════════════════════════════════════════
    # circle_pct（イン逃げ時2着優位度）
    # ──────────────────────────────────────────────────────────────────────────
    # 【修正版設計思想】
    #   主軸: コース別マスタの「イン逃げ時2着率」（直接実績）
    #   補正: ST優位（速い艇は差しやすい）× 1号艇脆弱性（差されやすいほど2枠有利）
    #         決まり手%は補正因子として残す（主軸ではなく傾向補正）
    #
    #   ブレンド方式:
    #     player_2nd   : コース別マスタの「2着率」（イン逃げ時2着率）
    #     venue_2nd    : イン逃げ分析シートの「枠番別2着率」（会場ベースライン）
    #     national_2nd : 全国平均（フォールバック）
    #
    #     trust = min(ininage_count / 30, 1.0)
    #     base  = player_2nd * (0.5 + 0.4 * trust) + venue_2nd * (0.5 - 0.4 * trust)
    #           → 実績30走以上: player 90% / venue 10%
    #           → 実績0走(trust=0): player 50% / venue 50%（データ不足でも完全棄却しない）
    #
    #   ST補正（対1号艇）:
    #     速い艇（ST<1号艇）→ 差しやすい → 2着スコアUP（最大×1.3）
    #     遅い艇             → ×0.85程度
    #
    #   脆弱性補正（2・3枠のみ）:
    #     1号艇が差されやすい → 2枠の差しスコアをさらに補強
    #     1号艇が捲られやすい → 3枠のまくり差しスコアを補強
    #
    #   コース変数は「想定コース（waku フォールバック）」を使用。
    #   会場ベースラインは「枠番」ベースのイン逃げ分析シートを参照（シート定義に準拠）。
    # ══════════════════════════════════════════════════════════════════════════

    raw_scores = {}
    for r in results:
        w = r["waku"]
        if w == "1":
            raw_scores[w] = None
            continue

        cm    = r.get("raw_cm", {})
        avg_st = r.get("avg_st")
        try:
            wno = int(w)
        except (ValueError, TypeError):
            wno = 3

        # ── イン逃げ時2着率（直接実績）────────────────────────────────────
        ininage_count = safe_float(cm.get("イン逃げ\n出走数"), 0) or 0
        player_2nd    = safe_float(cm.get("2着率"))   # コース別マスタのイン逃げ時2着率
        has_data      = ininage_count >= MIN_ININAGE_COUNT

        # 会場ベースライン（枠番別）
        venue_2nd    = venue_frame.get(w)            # イン逃げ分析シート（枠番ベース）
        national_2nd = NATIONAL_2ND.get(w, 0.10)    # 全国平均フォールバック

        # trust加重ブレンド
        trust = min(ininage_count / 30.0, 1.0) if has_data else 0.0
        w_player = 0.5 + 0.4 * trust   # 0.50（trust=0）〜 0.90（trust=1）
        w_venue  = 1.0 - w_player       # 0.50 〜 0.10

        baseline = venue_2nd if venue_2nd is not None else national_2nd
        if player_2nd is not None and has_data:
            base_2nd = player_2nd * w_player + baseline * w_venue
        else:
            base_2nd = baseline   # データ不足 → 会場/全国ベースライン

        # ── ST優位補正（対1号艇）──────────────────────────────────────────
        # 速い艇は1号艇より先にスリットを切れる → 2着争いで有利（最大×1.3）
        st_adv = _st_advantage_vs1(avg_st)
        # ST補正を穏やかに適用（主軸が2着率なので過剰にかけない）
        st_factor = 0.85 + (st_adv - 0.5) * 0.3   # 0.85〜1.30 の範囲

        # ── コース別・脆弱性補正（細かい傾向調整）──────────────────────────
        if wno == 2:
            # 2枠差し: 1号艇が差されやすいほど有利
            vuln_factor = 1.0 + vuln_sashi * 0.3    # 最大+30%
        elif wno == 3:
            # 3枠まくり差し: 1号艇が捲られやすいほど有利
            vuln_factor = 1.0 + vuln_makuri * 0.25  # 最大+25%
        elif wno == 4:
            # 4枠まくり: 3枠選手のまくり力が強いと外枠に押し出されペナルティ
            r3      = next((x for x in results if x["waku"] == "3"), None)
            cm3     = r3.get("raw_cm", {}) if r3 else {}
            mk3_pct = (safe_float(_get_cm_val(cm3, "まくり%"), 0) or 0.0) + \
                      (safe_float(_get_cm_val(cm3, "まくり差し%"), 0) or 0.0)
            st3     = r3.get("avg_st") if r3 else None
            wall3   = mk3_pct * _st_advantage_vs1(st3) * 0.30
            vuln_factor = max(0.7, 1.0 - wall3)
        elif wno == 5:
            vuln_factor = 0.80   # 外枠減衰
        else:  # 6枠
            vuln_factor = 0.60   # さらに減衰

        score = base_2nd * st_factor * vuln_factor
        raw_scores[w] = max(score, 0.001)   # ゼロ除算防止

    # レース内正規化（合計100%）
    valid_scores = {w: s for w, s in raw_scores.items() if w != "1" and s is not None}
    total_score  = sum(valid_scores.values()) or 1.0

    for r in results:
        w = r["waku"]
        s = raw_scores.get(w)
        if w != "1" and s is not None:
            r["circle_pct"] = round(s / total_score * 100, 1)
            r["_circ_raw"]  = s          # 確率計算用（正規化前絶対スコア）
        else:
            r["circle_pct"] = None
            r["_circ_raw"]  = None

    # ══════════════════════════════════════════════════════════════════════════
    # idx3（イン逃げ時3着残存指数）
    # ──────────────────────────────────────────────────────────────────────────
    # 【修正版設計思想】
    #   主軸: 純3着率 = イン逃げ時3着以内率 − イン逃げ時2着率（直接実績）
    #
    #   修正点:
    #     ① win3_rate（全体3連対率）を除去 → イン逃げ局面と無関係のため
    #     ② 会場ベースラインも純3着率ベース（3着以内率 − 2着率）で統一
    #     ③ ST補正は「遅い艇ほど流れ込みで3着に残りやすい」傾向を穏やかに反映
    #
    #   trust加重ブレンド（circle_pctと同じ方式）:
    #     base_pure3 = player_pure3 * w_player + venue_pure3 * w_venue
    # ══════════════════════════════════════════════════════════════════════════

    raw_idx3_scores = {}
    for r in results:
        w = r["waku"]
        if w == "1":
            raw_idx3_scores[w] = None
            continue

        cm = r.get("raw_cm", {})
        ininage_count = safe_float(cm.get("イン逃げ\n出走数"), 0) or 0
        player_3rd    = safe_float(cm.get("3着以内率"))   # イン逃げ時3着以内率
        player_2nd    = safe_float(cm.get("2着率"))       # イン逃げ時2着率
        has_data      = ininage_count >= MIN_ININAGE_COUNT

        # 選手の純3着率（イン逃げ時3着以内率 − 2着率）
        if player_3rd is not None and player_2nd is not None and has_data:
            player_pure3 = max(player_3rd - player_2nd, 0.0)
        elif player_3rd is not None and has_data:
            # 2着率不明 → 3着以内率の35%を純3着と推定（全国平均比から導出）
            player_pure3 = player_3rd * 0.35
        else:
            player_pure3 = None

        # 会場ベースラインの純3着率（枠番別）
        venue_3rd_rate = venue_3rd_map.get(w)
        venue_2nd_rate = venue_frame.get(w)
        if venue_3rd_rate is not None and venue_2nd_rate is not None:
            venue_pure3 = max(venue_3rd_rate - venue_2nd_rate, 0.0)
        else:
            venue_pure3 = None

        national_pure3 = NATIONAL_3RD_PURE.get(w, 0.10)

        # trust加重ブレンド（circle_pctと同じ方式）
        trust    = min(ininage_count / 30.0, 1.0) if has_data else 0.0
        w_player = 0.5 + 0.4 * trust   # 0.50〜0.90
        w_venue  = 1.0 - w_player       # 0.50〜0.10

        venue_val = venue_pure3 if venue_pure3 is not None else national_pure3
        if player_pure3 is not None and has_data:
            score_3rd = player_pure3 * w_player + venue_val * w_venue
        else:
            score_3rd = venue_val

        # ST補正（遅め艇ほど流れ込みで3着に残りやすい）
        avg_st = r.get("avg_st")
        if avg_st is not None:
            # 平均より遅いほど微加算、速いほど微減（影響は小さく±15%以内）
            st_slow_bonus = max(0.85, min(1.15, 1.0 + (avg_st - st_mean) / 0.05 * 0.10))
        else:
            st_slow_bonus = 1.0
        score_3rd = max(score_3rd * st_slow_bonus, 0.0)

        raw_idx3_scores[w] = score_3rd

    # 最大値=100にスケーリング（分散を保つため下限は設けない）
    valid_idx3 = [s for w, s in raw_idx3_scores.items() if w != "1" and s is not None]
    max_idx3   = max(valid_idx3) if valid_idx3 else 1.0
    for r in results:
        w  = r["waku"]
        s3 = raw_idx3_scores.get(w)
        if w != "1" and s3 is not None and max_idx3 > 0:
            r["idx3"] = min(int(s3 / max_idx3 * 100), 100)
        else:
            r["idx3"] = 0

    # frame_2nd: write_race_flat の2着率テキストブロック描画用に渡す
    frame_2nd = {w: s for w, s in raw_scores.items() if s is not None}
    
    # 会場イン逃げ場平均・決まり手場平均
    venue_stats = _calc_venue_stats(venue_stats_master, venue)

    # ── ★新機能: 6人相性スコア・イン飛び条件定量化 ────────────────────────
    # ── 展開考察エンジン（STxコース連動 → 対立構造 → 展開quality）────────────
    # _judge_tobi_scenario の前段として実行し、結果を race_judgment に連携させる。
    # first_turn: 1M到達順序と展開パターン
    # conflict_map: 誰が誰を潰しに行くかの対立構造
    # scenario_quality: 展開がどれだけ絞れているかのqualityスコア
    first_turn      = _predict_first_turn(results, venue=venue)
    conflict_map    = _build_conflict_map(results, first_turn)
    # scenario_quality は _suggest_3rentan 後に s1_prob が確定してから補完する
    # ここでは先行優位度と対立構造のみで暫定計算
    scenario_quality = _calc_scenario_quality(first_turn, conflict_map, s1_prob_est=None)

    affinity = _calc_affinity_score(results, venue_stats_master, venue)
    tobi     = _judge_tobi_scenario(results, affinity, venue_stats)

    # ── 荒れ/堅いレース判定 ＋ 3連単買い目提案（数値シート出力用） ──────────
    race_judgment  = _judge_race_type(results, venue_stats, venue_frame, race_no,
                                      venue_stats_master=venue_stats_master, venue=venue,
                                      tobi_scenario=tobi)
    # ── ★新機能: イン逃げ/イン飛び/両建て 3択判定（暫定：s1_prob未確定）────────
    # この時点では s1_prob がまだ確定していないため、暫定値として計算する。
    # s1_prob 確定後に main() または calc_race_indices の末尾で再計算する。
    ryotate_judgment = _judge_ryotate(race_judgment, tobi, venue_stats, s1_prob=None)
    race_judgment["ryotate"] = ryotate_judgment

    # 会場別1コース1着率を race_judgment に追加（_suggest_3rentan → _calc_3rentan_probs_v2 で使用）
    # 荒れやすい会場（戸田=43%, 平和島=45%等）で過剰なS1重みを抑制するため
    _vs_raw = venue_stats_master.get(venue, {})
    _vc1r = safe_float(_vs_raw.get("1コース1着率") or _vs_raw.get("1C_1着率"))
    race_judgment["venue_c1_win_rate"] = _vc1r  # Noneなら全国平均(0.555)にフォールバック
    race_judgment["affinity"]          = affinity  # ③相性考察で参照
    race_judgment["venue"]             = venue     # 参加見送り判定用

    # ── 【v6.2】会場統計の全データを race_judgment に格納 ─────────────────────
    # RNo別1C1着率: レース番号によるS1重み補正に使用
    # コース別1着率（2C〜6C）: シナリオ重みの会場補正に使用
    # Rレース別荒れスコア: MAX_BETS・買い目点数の調整に使用
    try:
        _rno = int(str(race_no))
    except (ValueError, TypeError):
        _rno = None
    # 当該レース番号の1C1着率（例: 5Rなら "1C_5R1着率"）
    _venue_1c_race_rate = None
    if _rno:
        _venue_1c_race_rate = safe_float(_vs_raw.get(f"1C_{_rno}R1着率"))
    race_judgment["venue_1c_race_rate"] = _venue_1c_race_rate  # R番号補正後の1C1着率

    # コース別1着率（全コース）
    race_judgment["venue_course_win_rates"] = {
        str(c): safe_float(_vs_raw.get(f"{c}C_1着率") or _vs_raw.get(f"{c}コース1着率"))
        for c in range(1, 7)
    }

    # Rレース別荒れスコア
    _venue_are_score = None
    if _rno:
        _venue_are_score = safe_float(_vs_raw.get(f"{_rno}R荒れスコア"))
    race_judgment["venue_race_are_score"] = _venue_are_score  # 当該Rの荒れスコア

    # 【枠なり前提】進入変更フラグ（初期値 False）
    # 展示後にコース変更が確認された場合は True にセットすること。
    # True の場合 _should_skip_race が最優先で見送り推奨を返す。
    race_judgment.setdefault("nyujo_henkou", False)

    # ── ❶ 1号艇逃げ力判定（6人構成メンバーを相手に逃げ切れるか）────────────
    # race_judgment や ryotate を参照しない純粋な前向き計算。
    # この結果が後続の ryotate 再計算・印スコアの入力として使われる。
    w1_escape = _judge_w1_escape(results, venue_stats, race_judgment=None)
    race_judgment["w1_escape"] = w1_escape
    print(f"  ❶ 1号艇逃げ判定: {w1_escape['escape_pct']}【{w1_escape['escape_rank']}】"
          f" 最大脅威={w1_escape['top_threat_waku']}号艇({w1_escape['top_threat_type']})")

    # ── ❷ 主役候補判定（逃げない場合に誰が主役でどの展開か）────────────────
    # w1_escape の threat_list を引き継ぎ、上位2艇・展開タイプ・2/3着候補を確定する。
    # escape_rank が「低」または「中」のとき特に重要。
    main_player = _judge_main_player(results, venue_stats, race_judgment)
    race_judgment["main_player"] = main_player
    _mp_sub = (f"  対抗={main_player['sub_waku']}号艇({main_player['sub_type']})"
               if main_player["sub_waku"] else "")
    print(f"  ❷ 主役候補: {main_player['main_waku']}号艇【{main_player['main_type']}】"
          f" スコア{main_player['main_score']*100:.0f}%{_mp_sub}")

    # ── ❸ 主役が来れなかった時の逃げ残存確率 ─────────────────────────────────
    # main_player（主役候補）が自滅した場合に1号艇が2着以内に残れるかを算出。
    # conflict_map はこの時点でローカル変数として確定済み（race_judgmentへの格納は後工程）
    # のでローカル変数を直接渡す。
    escape_fallback = _judge_escape_fallback(results, venue_stats, race_judgment,
                                             conflict_map=conflict_map)
    race_judgment["escape_fallback"] = escape_fallback
    print(f"  ❸ 逃げ残存({main_player['main_waku']}号艇自滅時):"
          f" {escape_fallback['fallback_pct']}【{escape_fallback['fallback_rank']}】"
          f" 自滅タイプ={escape_fallback['fly_type']}")

    # ── ❹ 主役展開の穴をつく艇判定 ──────────────────────────────────────────
    # 主軸対立（main_waku vs 1号艇）の外側で美味しいポジションに入れる艇を特定。
    # collapse_beneficiary は conflict_map から取得するためローカル変数を直接渡す。
    dark_horse = _judge_dark_horse(results, venue_stats, race_judgment,
                                   conflict_map=conflict_map)
    race_judgment["dark_horse"] = dark_horse
    if dark_horse["is_valid"]:
        _dh_str = "  ".join(
            f"{w}号艇({tag}:{s*100:.0f}%)"
            for w, s, tag in dark_horse["dark_horse_candidates"]
        )
        print(f"  ❹ 穴候補: {_dh_str}")
    else:
        print(f"  ❹ 穴候補: 有効な穴なし")

    # ── ★ヒモ荒れ判定 ────────────────────────────────────────────────────────
    # 1号艇が強本命（rel_win1 >= 65%）のとき「2・3着ヒモが荒れるか」を判定し
    # 参加可否と買い目点数調整の根拠として race_judgment に格納する。
    race_judgment["himo_are"] = _judge_himo_are(results, race_judgment)

    bet_suggestions = _suggest_3rentan(results, race_judgment, tenkai_venue=tenkai_venue_master, tenkai_national=tenkai_national_master, venue=venue, venue_stats=venue_stats)

    # ── s1_prob 確定後の各種最終補完 ─────────────────────────────────────────
    s1_prob_final = bet_suggestions.get("s1_prob") or race_judgment.get("s1_prob")

    # ① scenario_quality を s1_prob ベースで最終補完
    scenario_quality = _calc_scenario_quality(
        first_turn, conflict_map, s1_prob_est=s1_prob_final
    )

    # ② ryotate（3択判定）を s1_prob 確定後に再計算 ← 【断絶修正】
    #    s1_prob を渡すことで確率モデルと定性スコアの整合性チェックを実行し、
    #    3択の verdict・表示%・consistency_warn をすべて確率ベースに統一する。
    if s1_prob_final is not None:
        ryotate_judgment = _judge_ryotate(
            race_judgment, tobi, venue_stats, s1_prob=s1_prob_final
        )
        race_judgment["ryotate"] = ryotate_judgment

    # ③ ◎艇番とs1_prob最大艇の乖離チェック ← 【内部矛盾検出】
    #    バックテスト除外フラグとして活用可能
    first_prob_map_final = bet_suggestions.get("first_prob_map", {})
    if first_prob_map_final:
        top_prob_waku = max(first_prob_map_final, key=first_prob_map_final.get)
        honmei_waku_check = next(
            (str(r["waku"]) for r in results if r.get("honmei") == "◎"), None
        )
        if honmei_waku_check and honmei_waku_check != top_prob_waku:
            # 印◎ != 確率最大艇 → 矛盾フラグ（展開シナリオが印を覆している）
            race_judgment["honmei_prob_mismatch"] = True
            race_judgment["honmei_prob_mismatch_detail"] = (
                f"◎={honmei_waku_check}号艇 vs 確率最大={top_prob_waku}号艇"
                f"({first_prob_map_final.get(top_prob_waku, 0)*100:.1f}%)"
            )
        else:
            race_judgment["honmei_prob_mismatch"] = False
            race_judgment["honmei_prob_mismatch_detail"] = ""

    # ── 展開考察エンジン結果を race_judgment に格納 ───────────────────────────
    race_judgment["first_turn"]      = first_turn
    race_judgment["conflict_map"]    = conflict_map
    race_judgment["scenario_quality"] = scenario_quality

    return results, slit, venue_stats, frame_2nd, race_judgment, bet_suggestions

# ============================================================
# ★ 新機能① 6人相性スコア計算
# ============================================================

# ════════════════════════════════════════════════════════════════════════
# 展開考察エンジン（3関数セット）
# _predict_first_turn → _build_conflict_map → _calc_scenario_quality
# ════════════════════════════════════════════════════════════════════════

def _predict_first_turn(results, venue=None):
    """
    STとコース位置から第1ターンマーク（1M）到達順序と展開パターンを推定する。

    【競艇の物理法則】
      第1ターンへの到達時間 ≈ コース距離 + スタートタイム差
      内側コース（1号艇）ほど距離が短い。
      ただし外側艇がST大幅有利なら距離ハンデを逆転できる。

      到達順序（早い順）= コース距離補正後のST実効値でソート。
      実効ST = avg_st + コース距離補正

    【⑤改善】venue引数を受け取り、venue_course_adj.csv が存在すれば
      固定値の代わりに会場別実測値補正を使用する。
      信頼度 < 0.3 の会場は固定値にフォールバック。

    【ST不明時のフォールバック】
      avg_stがNoneの艇はコース別全国平均STで代替する。

    Returns
    -------
    dict:
        entry_order      : [(waku, eff_st), ...] 1M到達順（早い順・推定）
        lead_waku        : 先行艇（第1ターン最速艇）
        chase_waku       : 追走艇（2番手）
        lead_margin      : 先行差（秒）
        pattern          : "A（逃げ有利）"/"B（差し有利）"/"C（まくり差し）"/"D（大外）"
        pattern_strength : "強"（先行差>0.05）/"中"（>0.02）/"弱"（接戦）
        narrative        : 展開の絵（自然言語）
        st_details       : {waku: {"avg_st", "eff_st", "course_adj"}} デバッグ用
    """
    # コース距離補正（秒）— デフォルト固定値
    COURSE_ADJ = {"1": 0.00, "2": 0.03, "3": 0.06,
                  "4": 0.10, "5": 0.15, "6": 0.21}
    # STデータなし時の全国平均
    ST_NATIONAL = {"1": 0.18, "2": 0.17, "3": 0.17,
                   "4": 0.18, "5": 0.19, "6": 0.20}

    # 【⑤】会場別補正値CSVが存在すれば上書き（信頼度0.3以上の会場のみ）
    if venue and VENUE_COURSE_ADJ_CSV.exists():
        try:
            _adj_df = pd.read_csv(str(VENUE_COURSE_ADJ_CSV), encoding="utf-8-sig")
            _row = _adj_df[_adj_df["会場名"] == venue]
            if not _row.empty and float(_row["信頼度"].iloc[0]) >= 0.3:
                for c in range(2, 7):
                    col = f"{c}C補正"
                    if col in _row.columns:
                        COURSE_ADJ[str(c)] = float(_row[col].iloc[0])
        except Exception:
            pass  # 読み込み失敗時は固定値で継続

    entries = []
    st_details = {}
    for r in results:
        w    = r["waku"]
        st   = r.get("avg_st")
        if st is None or st <= 0:
            st = ST_NATIONAL.get(w, 0.18)
            is_estimated = True
        else:
            is_estimated = False
        adj     = COURSE_ADJ.get(w, 0.10)
        eff_st  = round(st + adj, 4)
        entries.append((w, eff_st))
        st_details[w] = {
            "avg_st":      st,
            "course_adj":  adj,
            "eff_st":      eff_st,
            "estimated":   is_estimated,
        }

    entries.sort(key=lambda x: x[1])

    lead_waku  = entries[0][0]
    chase_waku = entries[1][0] if len(entries) >= 2 else None
    lead_margin = round(entries[1][1] - entries[0][1], 4) if len(entries) >= 2 else 0.0

    # 展開パターン
    try:
        lead_course = int(lead_waku)
    except (ValueError, TypeError):
        lead_course = 1

    if lead_course == 1:
        pattern = "A（逃げ有利）"
    elif lead_course == 2:
        pattern = "B（差し有利）"
    elif lead_course <= 4:
        pattern = "C（まくり差し）"
    else:
        pattern = "D（大外まくり）"

    if lead_margin > 0.05:
        pattern_strength = "強"
    elif lead_margin > 0.02:
        pattern_strength = "中"
    else:
        pattern_strength = "弱（接戦）"

    # 展開の絵（自然言語）
    entry_str = " → ".join([f"{w}号({st:.3f})" for w, st in entries])
    if pattern_strength == "強":
        strength_desc = f"{lead_waku}号艇が{lead_margin:.3f}秒差で先行確定的"
    elif pattern_strength == "中":
        strength_desc = f"{lead_waku}号艇が優位だが{chase_waku}号艇が追走できる差"
    else:
        strength_desc = f"{lead_waku}号艇と{chase_waku}号艇が接戦（展開は流動的）"

    narrative = (
        f"【1M到達順（推定）】{entry_str}\n"
        f"【先行】{lead_waku}号艇 → {pattern} / 強度:{pattern_strength}\n"
        f"{strength_desc}"
    )

    return {
        "entry_order":       entries,
        "lead_waku":         lead_waku,
        "chase_waku":        chase_waku,
        "lead_margin":       lead_margin,
        "pattern":           pattern,
        "pattern_strength":  pattern_strength,
        "narrative":         narrative,
        "st_details":        st_details,
    }


def _build_conflict_map(results, first_turn, cm_map_ext=None):
    """
    1M到達順序を受けて「誰が誰を潰しに行くか」の対立構造を計算する。

    【競艇の対立構造の本質】
      1M到達順序が決まれば、各艇の「攻撃対象」が物理的に決まる。

      攻撃対象の決まり方:
        先行艇（1番手）: 攻撃対象なし（前に誰もいない）
        2番手艇: 先行艇を差す or まくる
        3番手艇: 2番手艇を外から包む（まくり）or 1番手を差す（まくり差し）
        4番手以降: 外からまくる or 内側の争いを待って2着圏に滑り込む

      攻撃強度 = 決まり手適性 × ST差（接戦ほど攻撃が届きやすい）× 位置補正

    【対立の主軸と副軸】
      主軸対立: 1M到達2番手が先行艇を攻撃する構図
      副軸対立: 1M到達3〜4番手が2番手を潰す構図（漁夫の利の発生源）

    【出力する「展開の絵」】
      例:
        主軸: 3号艇(まくり差し系)が1号艇を包む構図
        副軸: 4号艇(まくり系)が3号艇を外から被せる可能性
        潰れ受益: 3号艇が自滅した場合、5号艇（攻撃性低・win3_rate高）が漁夫

    Returns
    -------
    dict:
        main_conflict    : {"attacker", "target", "method", "strength", "desc"}
        sub_conflict     : {"attacker", "target", "method", "strength", "desc"} or None
        collapse_beneficiary : [(waku, score), ...]  潰れ受益候補（スコア降順）
        narrative        : 対立構造の展開の絵（自然言語）
        conflict_entries : 全艇の対立エントリ（詳細デバッグ用）
    """
    entry_order = first_turn["entry_order"]  # [(waku, eff_st), ...]
    wakus_in_order = [w for w, _ in entry_order]

    # cm_mapを構築（resultsから）
    cm_map = {r["waku"]: r.get("raw_cm", {}) for r in results}
    win3_map = {r["waku"]: r.get("win3_rate") or 0.5 for r in results}

    def safe_pct(cm, key):
        v = cm.get(key)
        try:
            return max(float(v), 0.0) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    def _kime_type(waku):
        """決まり手タイプを返す（まくり系/差し系/逃げ系）"""
        cm = cm_map.get(waku, {})
        mak  = safe_pct(cm, "まくり%") + safe_pct(cm, "まくり差し%")
        sash = safe_pct(cm, "差し%")
        nige = safe_pct(cm, "逃げ%")
        best = max(mak, sash, nige)
        if best == 0:
            return "不明"
        if mak == best:
            return "まくり系"
        if sash == best:
            return "差し系"
        return "逃げ系"

    def _attack_strength(attacker_w, target_w, entry_idx_diff):
        """
        攻撃強度 = 決まり手適性スコア × ST接近度補正 × 位置距離補正
        entry_idx_diff: 1M到達順序上の距離（1=直後、2=2つ後ろ etc.）
        """
        cm = cm_map.get(attacker_w, {})
        mak  = safe_pct(cm, "まくり%") + safe_pct(cm, "まくり差し%")
        sash = safe_pct(cm, "差し%")
        base = max(mak, sash)  # 攻撃決まり手の主力
        if base == 0:
            base = 10.0  # データなしフォールバック

        # 位置距離減衰（直後が最も攻撃しやすい）
        dist_factor = {1: 1.0, 2: 0.7, 3: 0.5, 4: 0.3, 5: 0.2}
        dist_dec = dist_factor.get(entry_idx_diff, 0.15)

        # ST接近度（eff_st差が小さいほど追いつきやすい）
        lead_eff = dict(entry_order).get(target_w, 0.5)
        att_eff  = dict(entry_order).get(attacker_w, 0.5)
        st_gap   = abs(lead_eff - att_eff)
        st_factor = max(0.3, 1.0 - st_gap * 8.0)  # 0.1秒差で0.2低下

        return round(base * dist_dec * st_factor, 1)

    # ── 各艇の攻撃対象・手法・強度を計算 ──────────────────────────────────────
    conflict_entries = []
    for idx, (w, eff_st) in enumerate(entry_order):
        if idx == 0:
            # 先行艇は攻撃対象なし（前方クリア）
            conflict_entries.append({
                "waku": w, "role": "先行",
                "target": None, "method": "-", "strength": 0
            })
            continue

        # 攻撃対象 = 1M到達順で1つ前の艇（最も自然な攻撃対象）
        target_w = entry_order[idx - 1][0]
        kime     = _kime_type(w)
        strength = _attack_strength(w, target_w, 1)

        # 攻撃手法の決定
        try:
            w_course      = int(w)
            target_course = int(target_w)
        except (ValueError, TypeError):
            w_course = target_course = 3

        if kime == "差し系" and w_course < target_course + 2:
            method = "差し"
        elif kime == "まくり系":
            method = "まくり"
        elif kime in ("まくり系", "不明"):
            method = "まくり"
        else:
            method = "まくり差し"

        conflict_entries.append({
            "waku":     w,
            "role":     f"{idx+1}番手",
            "target":   target_w,
            "method":   method,
            "kime":     kime,
            "strength": strength,
        })

    # ── 主軸対立（進入2番手 vs 先行艇）──────────────────────────────────────────
    main_entry = conflict_entries[1] if len(conflict_entries) >= 2 else None
    if main_entry and main_entry["target"]:
        main_conflict = {
            "attacker": main_entry["waku"],
            "target":   main_entry["target"],
            "method":   main_entry["method"],
            "strength": main_entry["strength"],
            "desc": (
                f"{main_entry['waku']}号艇が{main_entry['target']}号艇に"
                f"「{main_entry['method']}」で挑む"
                f"（強度{main_entry['strength']:.0f}）"
            ),
        }
    else:
        main_conflict = None

    # ── 副軸対立（進入3番手 vs 2番手）──────────────────────────────────────────
    sub_entry = conflict_entries[2] if len(conflict_entries) >= 3 else None
    if sub_entry and sub_entry["target"]:
        sub_conflict = {
            "attacker": sub_entry["waku"],
            "target":   sub_entry["target"],
            "method":   sub_entry["method"],
            "strength": sub_entry["strength"],
            "desc": (
                f"{sub_entry['waku']}号艇が{sub_entry['target']}号艇を"
                f"「{sub_entry['method']}」で外から被せる"
                f"（強度{sub_entry['strength']:.0f}）"
            ),
        }
    else:
        sub_conflict = None

    # ── 潰れ受益者（主軸攻撃艇が自滅した場合の漁夫候補）────────────────────────
    # 主軸攻撃艇が自滅 → 受益者の条件:
    #   ① 攻撃性が低い（自分は仕掛けに行かない）
    #   ② win3_rateが高い（荒れても残る力がある）
    #   ③ 主軸対立の外側にいる（巻き込まれない位置）
    main_attacker = main_conflict["attacker"] if main_conflict else None
    beneficiary_scores = {}
    for r in results:
        w = r["waku"]
        if w == main_attacker or w == (main_conflict["target"] if main_conflict else None):
            continue  # 主軸の当事者は除外
        cm   = cm_map.get(w, {})
        mak  = safe_pct(cm, "まくり%") + safe_pct(cm, "まくり差し%")
        sash = safe_pct(cm, "差し%")
        attack_rate  = min((mak + sash) / 100.0, 1.0)
        passivity    = 1.0 - attack_rate * 0.6  # 低攻撃性 = 高漁夫スコア
        ground       = win3_map.get(w, 0.5)
        beneficiary_scores[w] = round(passivity * ground, 4)

    collapse_beneficiary = sorted(
        beneficiary_scores.items(), key=lambda x: x[1], reverse=True
    )

    # ── 展開の絵（自然言語ナラティブ）──────────────────────────────────────────
    parts = []
    if main_conflict:
        parts.append(f"主軸: {main_conflict['desc']}")
    if sub_conflict:
        parts.append(f"副軸: {sub_conflict['desc']}")
    if collapse_beneficiary:
        top_b = collapse_beneficiary[0]
        parts.append(
            f"潰れ受益候補: {top_b[0]}号艇"
            f"（攻撃性低・win3高・主軸争いの外側）"
        )
    narrative = "\n".join(parts) if parts else "展開の絵を生成できませんでした"

    return {
        "main_conflict":         main_conflict,
        "sub_conflict":          sub_conflict,
        "collapse_beneficiary":  collapse_beneficiary,
        "narrative":             narrative,
        "conflict_entries":      conflict_entries,
        "lead_waku":             wakus_in_order[0] if wakus_in_order else None,
    }


def _calc_scenario_quality(first_turn, conflict_map, s1_prob_est=None):
    """
    展開の「絞れ度」= 展開が特定シナリオに集中しているか（quality）を計算する。

    【質スコアの概念】
      競艇のレースには「読みやすいレース」と「読みにくいレース」がある。

      読みやすいレース（quality高）:
        → 先行艇が圧倒的有利、攻撃艇が1艇に絞れる
        → 買い目を少点数に絞れる根拠になる
        → 合成オッズが低くても期待値が確保できる

      読みにくいレース（quality低）:
        → 先行艇と追走艇が接戦、攻撃艇が複数
        → 何でも起きる → 買い目を絞ることに根拠がない
        → 高いqualityのレースを待つ方が合理的

    【quality構成要素】
      ① 先行優位度  (0〜40点): lead_margin が大きいほど高い
      ② 主軸集中度  (0〜30点): 主軸攻撃が強く副軸が弱いほど高い
      ③ 展開2択度   (0〜20点): 逃げか1つの飛びかに絞れているほど高い
      ④ 接戦ペナルティ (-10〜0点): 3艇以上が接戦のとき減点

    Returns
    -------
    dict:
        quality_score  : 0〜100
        quality_rank   : "S"(>=75)/"A"(>=55)/"B"(>=40)/"C"(>=25)/"D"(<25)
        quality_verdict: "展開が絞れている" / "要注意（展開流動的）" / "見送り推奨（混戦）"
        bet_size_guide : "点数を絞れる" / "標準点数" / "増やすか見送り"
        components     : 各要素のスコア（デバッグ用）
        narrative      : qualityの説明（自然言語）
    """
    score = 0.0
    components = {}

    # ① 先行優位度（0〜40点）
    lead_margin = first_turn.get("lead_margin", 0)
    if lead_margin > 0.08:
        lead_score = 40.0
    elif lead_margin > 0.05:
        lead_score = 30.0
    elif lead_margin > 0.02:
        lead_score = 20.0
    elif lead_margin > 0.00:
        lead_score = 10.0
    else:
        lead_score = 0.0
    score += lead_score
    components["先行優位度"] = round(lead_score, 1)

    # ② 主軸集中度（0〜30点）
    # 主軸強・副軸弱 = 集中度高
    main = conflict_map.get("main_conflict") or {}
    sub  = conflict_map.get("sub_conflict") or {}
    main_str = main.get("strength", 0) or 0
    sub_str  = sub.get("strength", 0) or 0
    if main_str > 0:
        concentration = main_str / max(main_str + sub_str, 1)
        axis_score = min(30.0, concentration * 30.0 * (main_str / 50.0))
    else:
        axis_score = 0.0
    score += axis_score
    components["主軸集中度"] = round(axis_score, 1)

    # ③ 展開2択度（0〜20点）
    # s1_prob_est が渡されている場合は逃げ確率を使用
    # 渡されていない場合はパターンで推定
    if s1_prob_est is not None:
        # 逃げ一択 or 飛び一択に近いほど高い
        two_choice = abs(s1_prob_est - 0.5) * 2  # 0〜1
        two_score  = two_choice * 20.0
    else:
        pattern = first_turn.get("pattern", "")
        if first_turn.get("pattern_strength") == "強":
            two_score = 20.0
        elif first_turn.get("pattern_strength") == "中":
            two_score = 12.0
        else:
            two_score = 5.0
    score += two_score
    components["展開2択度"] = round(two_score, 1)

    # ④ 接戦ペナルティ
    entry_order = first_turn.get("entry_order", [])
    if len(entry_order) >= 3:
        eff_sts = [st for _, st in entry_order[:4]]
        if len(eff_sts) >= 3:
            spread = eff_sts[2] - eff_sts[0]
            if spread < 0.03:  # 上位3艇が0.03秒以内 = 接戦
                penalty = -10.0
                score += penalty
                components["接戦ペナルティ"] = round(penalty, 1)

    score = max(0.0, min(100.0, score))

    # ── ランク判定（v2: 閾値を緩和してS/Aが出るように再設計）─────────────────
    # 旧版の問題: 最大90点満点のスコアに対し「S>=75」という閾値が高すぎた。
    # lead_margin > 0.08秒（40pt）+ 主軸集中度（最大30pt）+ 2択度（最大20pt）= 90pt が天井。
    # 現実の競艇では lead_margin > 0.08 は珍しいため、Sが出ない構造になっていた。
    #
    # v2 設計思想:
    #   先行差 0.05秒以上（30pt）+ 主軸集中度 >= 20pt + 2択度 >= 12pt → 合計62pt以上でA到達
    #   先行差 0.08秒以上（40pt）+ 同上合計 → 72pt以上でS到達
    #   よく出る lead_margin（0.03〜0.06秒）でもA/Bが取れるよう調整。
    #
    # 新閾値:
    #   S: >= 65  （絞れる展開。少点数推奨）
    #   A: >= 48  （ある程度集中。標準点数）
    #   B: >= 33  （やや流動的。点数増か厳選）
    #   C: >= 18  （読みにくい。参加基準UP）
    #   D: < 18   （混戦。見送り推奨）
    if score >= 65:
        quality_rank    = "S"
        quality_verdict = "展開が絞れている（買い目を絞る根拠あり）"
        bet_size_guide  = "点数を絞れる（少点数で合成オッズを高く保てる）"
    elif score >= 48:
        quality_rank    = "A"
        quality_verdict = "ある程度の展開集中（標準的な点数で対応可）"
        bet_size_guide  = "標準点数"
    elif score >= 33:
        quality_rank    = "B"
        quality_verdict = "展開やや流動的（点数増か厳選が必要）"
        bet_size_guide  = "点数増か主軸に絞る"
    elif score >= 18:
        quality_rank    = "C"
        quality_verdict = "展開が読みにくい（参加基準を上げる）"
        bet_size_guide  = "参加するなら大幅厳選"
    else:
        quality_rank    = "D"
        quality_verdict = "混戦（展開予測が困難。見送り推奨）"
        bet_size_guide  = "見送り推奨"

    # 自然言語
    narrative = (
        f"展開quality: {quality_rank}（{score:.0f}点）\n"
        f"→ {quality_verdict}\n"
        f"→ {bet_size_guide}"
    )

    return {
        "quality_score":   round(score, 1),
        "quality_rank":    quality_rank,
        "quality_verdict": quality_verdict,
        "bet_size_guide":  bet_size_guide,
        "components":      components,
        "narrative":       narrative,
    }


def _calc_affinity_score(results, venue_stats_master, venue):
    """
    出走6選手の相互作用を定量化し、各艇が「インを脅かす力」を数値化する。

    【設計思想】
    イン飛びは「インが弱い」だけでなく「誰かが積極的に攻める」ときに起きる。
    この関数は各艇の攻撃力と、1号艇の被攻撃脆弱性を組み合わせて評価する。

    出力 (dict):
      attack_score[waku]  : 各艇のイン攻撃ポテンシャル (0〜100)
      threat_total        : 全艇合計攻撃力 (イン脅威度)
      boat1_vulnerability : 1号艇の被攻撃脆弱性スコア (0〜100、高いほど飛びやすい)
      dominant_attacker   : 最も脅威の高い艇番 (str)
      affinity_summary    : 各艇の攻撃根拠テキスト dict

    【スコア計算要素】
    A. 差し力 (2コース艇)  : 差し% × ST速さ（1号艇との相対ST差）
    B. まくり力 (3〜5C)    : (まくり% + まくり差し%) × コース別1着率 × ST出足補正
    C. 1号艇の脆弱性       : 差され% + 捲られ% + ST遅さペナルティ + STばらつき
    D. A1選手補正          : 外枠にA1級選手がいる場合に攻撃力を+20%
    E. 今節成績補正        : 直近好成績の艇は攻撃スコアを最大+15%加算
    """
    NATIONAL_ATTACK_BASE = {
        "2": 0.137, "3": 0.134, "4": 0.111, "5": 0.066, "6": 0.021
    }

    res1 = next((r for r in results if r["waku"] == "1"), None)
    st1  = res1.get("avg_st") if res1 else None
    cm1  = res1.get("raw_cm", {}) if res1 else {}

    # 1号艇の脆弱性スコア（0〜100）
    # ※ "差し%"（_get_cm_val経由）は「自分が差して勝った割合」であり脆弱性評価には使わない
    nige_pct    = safe_float(_get_cm_val(cm1, "逃げ%"), 0) or 0    # 逃げ率（高いほど脆弱性低）

    # 被攻撃実績（update_masterが生成する被決まり手%）
    sasar_vuln = safe_float(cm1.get("差され%"), 0) or 0
    makur_vuln = safe_float(cm1.get("捲られ%"), 0) or 0
    maksa_vuln = safe_float(cm1.get("捲り差され%"), 0) or 0  # update_master出力キーに統一

    # ST脆弱性（1号艇が遅いほど脆弱）
    st_vuln = 0.0
    if st1 is not None:
        if st1 > 0.18:
            st_vuln = 30.0
        elif st1 > 0.15:
            st_vuln = 15.0
        elif st1 > 0.12:
            st_vuln = 5.0

    # STばらつき（不安定なほど脆弱）
    pm1 = res1.get("raw_pm", {}) if res1 else {}
    st_stable = safe_float(pm1.get("ST安定\nスコア") or pm1.get("ST安定スコア"))
    st_unstable_vuln = 0.0
    if st_stable is not None:
        if st_stable < 40:
            st_unstable_vuln = 20.0
        elif st_stable < 60:
            st_unstable_vuln = 10.0

    # ── FLY数・出遅れ数によるST脆弱性補正 ────────────────────────────────
    # 選手指数マスタに FLY数・出遅れ数 が集計されているが、従来は未使用だった。
    # FLY（フライング） → ペナルティ後の緊張・心理的影響で次走STが不安定になりやすい
    # 出遅れ（出遅れ数が多い）→ スロースタートの癖がある → イン逃げ失敗リスクUP
    #
    # FLY補正:
    #   FLY数が1走内に存在 → 直近にフライングあり → ST脆弱性+15
    #   FLY数が2走以上    → 繰り返しフライング → ST脆弱性+25（最大値）
    # 出遅れ補正:
    #   出遅れ数/総出走数 が 5%超 → 出遅れ癖あり → ST脆弱性+10
    #   出遅れ数/総出走数 が 10%超 → 出遅れ癖強い → ST脆弱性+20
    fly_count    = safe_float(pm1.get("FLY数"),    0) or 0
    late_count   = safe_float(pm1.get("出遅れ数"), 0) or 0
    total_runs_pm = safe_float(pm1.get("ST計測件数") or pm1.get("総出走数"), 0) or 1

    fly_vuln = 0.0
    if fly_count >= 2:
        fly_vuln = 25.0
    elif fly_count >= 1:
        fly_vuln = 15.0

    late_rate = late_count / max(total_runs_pm, 1)
    late_vuln = 0.0
    if late_rate >= 0.10:
        late_vuln = 20.0
    elif late_rate >= 0.05:
        late_vuln = 10.0

    boat1_vulnerability = min(100.0,
        (sasar_vuln * 100) * 0.30 +          # 差された実績（旧0.35から微調整）
        (makur_vuln * 100) * 0.22 +          # 捲られた実績
        (maksa_vuln * 100) * 0.12 +          # 捲り差された実績
        st_vuln * 0.12 +                     # ST遅さペナルティ
        st_unstable_vuln * 0.08 +            # STばらつきペナルティ
        fly_vuln * 0.10 +                    # 【新追加】FLY履歴ペナルティ
        late_vuln * 0.06 +                   # 【新追加】出遅れ癖ペナルティ
        max(0.0, (40.0 - nige_pct) * 0.5)   # 逃げ%が低いほど加算
    )

    # 各艇の攻撃スコアを計算
    attack_score  = {}
    affinity_summary = {}

    for r in results:
        w = r["waku"]
        if w == "1":
            attack_score[w] = 0.0
            affinity_summary[w] = "1号艇（逃げ側）"
            continue

        cm  = r.get("raw_cm", {})
        pm  = r.get("raw_pm", {})
        st  = r.get("avg_st")

        # 基礎攻撃力（決まり手%）
        sashi_pct = safe_float(_get_cm_val(cm, "差し%"), 0) or 0
        makuri_pct = (safe_float(_get_cm_val(cm, "まくり%"), 0) or 0) + \
                     (safe_float(_get_cm_val(cm, "まくり差し%"), 0) or 0)
        nat_base   = NATIONAL_ATTACK_BASE.get(w, 0.05) * 100

        if w == "2":
            # 2号艇：差し特化
            base_attack = sashi_pct if sashi_pct > 0 else nat_base * 0.3
            attack_type = f"差し{sashi_pct:.0f}%"
        else:
            # 3〜6号艇：まくり系
            base_attack = makuri_pct if makuri_pct > 0 else nat_base * 0.4
            attack_type = f"まくり系{makuri_pct:.0f}%"

        # ST相対補正（1号艇より速い艇は攻撃力UP）
        st_boost = 1.0
        if st is not None and st1 is not None:
            diff = st1 - st  # 正 = この艇が速い
            if diff > 0.03:
                st_boost = 1.20
            elif diff > 0.01:
                st_boost = 1.10
            elif diff < -0.03:
                st_boost = 0.85

        # 今節成績補正（直近の実走成績から個人の調子を反映）
        kosetsu = str(r.get("kosetsu", ""))
        kosetsu_boost = 1.0
        if kosetsu and kosetsu not in ("", "None", "nan", "-"):
            try:
                tokens = [t.strip() for t in re.split(r"[-・/]", kosetsu)]
                win_count = tokens.count("1")
                if win_count >= 2:
                    kosetsu_boost = 1.15
                elif win_count == 1:
                    kosetsu_boost = 1.07
            except Exception:
                pass

        # コース別実績補正（会場別コースマスタ優先）
        # ※ 級別（A1/A2等）は使用しない。
        #   個々の選手が「実際にそのコースで何をしてきたか」
        #   （決まり手%・ST・今節成績）のみで評価する。
        venue_win = safe_float(r.get("win1_rate")) or 0

        score = base_attack * st_boost * kosetsu_boost
        # 実績との整合（選手の実際のコース別1着率で加重）
        score = score * 0.7 + venue_win * 100 * 0.3

        attack_score[w] = round(score, 2)
        summary_parts = [attack_type]
        if st_boost > 1.05:
            summary_parts.append(f"ST優位(×{st_boost:.2f})")
        elif st_boost < 0.90:
            summary_parts.append(f"ST劣位(×{st_boost:.2f})")
        if kosetsu_boost > 1.0:
            summary_parts.append(f"今節好調(×{kosetsu_boost:.2f})")
        affinity_summary[w] = " / ".join(summary_parts)

    # 合計攻撃力（1号艇除く）
    threat_total = sum(v for k, v in attack_score.items() if k != "1")

    # 最大攻撃艇
    outer_scores = {k: v for k, v in attack_score.items() if k != "1"}
    dominant_attacker = max(outer_scores, key=lambda k: outer_scores[k]) if outer_scores else "-"

    return {
        "attack_score":       attack_score,
        "threat_total":       round(threat_total, 2),
        "boat1_vulnerability": round(boat1_vulnerability, 1),
        "dominant_attacker":  dominant_attacker,
        "affinity_summary":   affinity_summary,
    }


# ============================================================
# ★ 新機能② イン飛び条件総合判定
# ============================================================
def _judge_tobi_scenario(results, affinity, venue_stats):
    """
    イン飛び（1号艇が1着にならない）の総合確率と根拠を返す。

    【判定ロジック】
    以下5つの条件を重みづけして飛び確率(0〜100)を算出する。

    条件①  1号艇の脆弱性スコア（被差され・捲られ実績）
    条件②  攻撃艇の合計脅威スコア（全艇の攻撃力合算）
    条件③  会場のイン逃げ率（低い会場ほど飛びやすい）
    条件④  支配的攻撃艇の存在感（最強攻撃艇のスコアが突出しているか）
    条件⑤  スリット不利（1号艇STが相対的に遅い場合）

    出力:
      tobi_prob   : イン飛び推定確率 (0〜100)
      tobi_rank   : 飛び確率ランク  S(>70)/A(>55)/B(>40)/C(>25)/D(<=25)
      main_threat : 最も危険な飛ばし役の艇番
      reasons     : 判定根拠リスト
      tobi_type   : 予想される飛び方 ("差し" / "まくり" / "まくり差し" / "不明")
    """
    reasons = []
    score = 0.0  # 0〜100: 高いほどイン飛び確率大

    vuln  = affinity["boat1_vulnerability"]
    total = affinity["threat_total"]
    dom   = affinity["dominant_attacker"]
    atk   = affinity["attack_score"]

    # 条件①: 1号艇の脆弱性（最大35点）
    vul_contrib = min(35.0, vuln * 0.35)
    score += vul_contrib
    if vuln >= 40:
        reasons.append(f"1号艇脆弱性スコア{vuln:.0f}（被差し・捲られ実績大）")
    elif vuln >= 20:
        reasons.append(f"1号艇脆弱性スコア{vuln:.0f}（やや脆弱）")

    # 条件②: 攻撃艇の合計脅威（最大30点）
    # threat_total の自然な範囲は0〜150程度 → 100以上で満点
    threat_contrib = min(30.0, total / 100.0 * 30.0)
    score += threat_contrib
    if total >= 80:
        reasons.append(f"攻撃艇合計スコア{total:.0f}（攻撃力が強い）")
    elif total >= 50:
        reasons.append(f"攻撃艇合計スコア{total:.0f}（中程度の攻撃）")

    # 条件③: 会場イン逃げ率（最大20点）
    in_rate = venue_stats.get("in_rate")
    if in_rate is not None:
        # イン逃げ率が低い会場ほどスコアUP（逃げ率50%以下で最大）
        venue_contrib = max(0.0, (0.65 - float(in_rate)) / 0.65 * 20.0)
        score += venue_contrib
        if float(in_rate) < 0.45:
            reasons.append(f"会場イン逃げ率{float(in_rate)*100:.0f}%（飛びやすい会場）")

    # 条件④: 支配的攻撃艇の突出度（最大10点）
    if dom != "-":
        dom_score = atk.get(dom, 0)
        other_scores = [v for k, v in atk.items() if k != "1" and k != dom]
        if other_scores:
            avg_other = sum(other_scores) / len(other_scores)
            if avg_other > 0 and dom_score > avg_other * 1.5:
                score += 10.0
                reasons.append(f"{dom}号艇が突出した攻撃力({dom_score:.0f}pt)：明確な飛ばし役")
            elif dom_score > 0:
                score += 5.0
                reasons.append(f"{dom}号艇が攻撃力最大({dom_score:.0f}pt)")

    # 条件⑤: スリット不利（1号艇のSTが遅い）（最大5点）
    res1 = next((r for r in results if r["waku"] == "1"), None)
    if res1:
        st1 = res1.get("avg_st")
        all_sts = [r.get("avg_st") for r in results if r.get("avg_st") is not None and r["waku"] != "1"]
        if st1 is not None and all_sts:
            faster_count = sum(1 for s in all_sts if s < st1 - 0.02)
            if faster_count >= 3:
                score += 5.0
                reasons.append(f"1号艇STが相対的に遅い（{faster_count}艇が有意に速い）")

    score = min(100.0, max(0.0, score))

    # ランク判定
    if score >= 70:
        tobi_rank = "S"
    elif score >= 55:
        tobi_rank = "A"
    elif score >= 40:
        tobi_rank = "B"
    elif score >= 25:
        tobi_rank = "C"
    else:
        tobi_rank = "D"

    # 飛び方の予測（最強攻撃艇のコースと決まり手から推定）
    tobi_type = "不明"
    if dom != "-":
        dom_r = next((r for r in results if r["waku"] == dom), None)
        if dom_r:
            cm_dom = dom_r.get("raw_cm", {})
            sashi  = safe_float(_get_cm_val(cm_dom, "差し%"), 0) or 0
            makuri = safe_float(_get_cm_val(cm_dom, "まくり%"), 0) or 0
            maksa  = safe_float(_get_cm_val(cm_dom, "まくり差し%"), 0) or 0
            if dom == "2":
                tobi_type = "差し"
            elif makuri >= maksa and makuri >= sashi:
                tobi_type = "まくり"
            elif maksa >= makuri and maksa >= sashi:
                tobi_type = "まくり差し"
            elif sashi > 0:
                tobi_type = "差し"

    return {
        "tobi_prob":   round(score, 1),
        "tobi_rank":   tobi_rank,
        "main_threat": dom,
        "reasons":     reasons,
        "tobi_type":   tobi_type,
        "affinity":    affinity,
    }


# ============================================================
# ★ 買い方ヒント生成（一気通貫：根拠→展開→買い目）
# ============================================================
def _generate_tenkai_story(results, venue, venue_stats, race_judgment, bet_suggestions):
    """
    あなたの7ステップ思考フローをそのまま1つのストーリーとして出力する。

    ① 1号艇は逃げるか
    ② 逃げる → 2・3着は誰か
    ③ 逃げない → 誰が主役か・決まり手は何か
    ④ その決まり手なら2・3着は誰か
    ⑤ 主役が崩れたとき誰が浮上するか
    ⑥ 崩れ後に1号が逃げを拾う確率
    ⑦ 他の艇が展開を突いたとき2・3着は誰か
    → 最終買い目サマリー
    """
    rj  = race_judgment  or {}
    bet = bet_suggestions or {}

    # ── データ取り出し ──────────────────────────────────────────────────
    s1_prob      = bet.get("s1_prob")  or rj.get("s1_prob")  or 0.0
    fp_map       = bet.get("first_prob_map", {}) or {}
    w1_escape    = rj.get("w1_escape",    {}) or {}
    main_player  = rj.get("main_player",  {}) or {}
    escape_fb    = rj.get("escape_fallback", {}) or {}
    dark_horse   = rj.get("dark_horse",   {}) or {}
    conflict_map = rj.get("conflict_map", {}) or {}
    neraime_2nd  = bet.get("neraime_2nd", []) or []
    candidates   = bet.get("candidates",  []) or []
    tenkai_pat   = bet.get("tenkai_pattern", rj.get("tenkai_pattern", "?"))

    rd = [r for r in (results or []) if isinstance(r, dict)]
    name_map = {str(r.get("waku","")): r.get("name","") for r in rd}

    def wn(w):
        """艇番と名前を返す（名前なしなら艇番のみ）"""
        n = name_map.get(str(w), "")
        return f"{w}号{n}" if n else f"{w}号艇"

    def pct(v):
        """0〜1のfloatを%文字列に変換"""
        try:
            return f"{float(v)*100:.0f}%"
        except (TypeError, ValueError):
            return "-"

    lines = []

    # ══════════════════════════════════════════════════════════════════
    # ① 1号艇は逃げるか
    # ══════════════════════════════════════════════════════════════════
    esc_prob  = w1_escape.get("escape_prob", s1_prob)
    esc_rank  = w1_escape.get("escape_rank", "中")
    esc_pct   = w1_escape.get("escape_pct",  pct(s1_prob))
    thr_w     = w1_escape.get("top_threat_waku",  "-")
    thr_t     = w1_escape.get("top_threat_type",  "-")
    thr_s     = w1_escape.get("top_threat_score", 0)

    rank_emoji = {"高": "🟢", "中": "🟡", "低": "🔴"}.get(esc_rank, "⚪")
    lines.append(
        f"① 1号艇 逃げ力：{esc_pct}【{esc_rank}】{rank_emoji}\n"
        f"   最大脅威：{wn(thr_w)}（{thr_t}）{pct(thr_s)}"
    )

    # ══════════════════════════════════════════════════════════════════
    # ② 逃げた場合の2・3着（展開別残存マスタ活用）
    # ══════════════════════════════════════════════════════════════════
    if neraime_2nd:
        # 残存型狙い目から2着候補を取得（展開別残存マスタ参照済み）
        top3_2nd = neraime_2nd[:3]
        s2_str = "  ".join(
            f"{n['waku']}号{n['r2_rate']*100:.0f}%"
            for n in top3_2nd
        )
        # 3着以内率から2着率を引いて純3着率を計算
        s3_str = "  ".join(
            f"{n['waku']}号{max(n['r3i_rate']-n['r2_rate'],0)*100:.0f}%"
            for n in top3_2nd
        )
    else:
        # circle_pctから2着候補を生成
        circ_sorted = sorted(
            [(r["waku"], r.get("circle_pct") or 0)
             for r in rd if r["waku"] != "1"],
            key=lambda x: x[1], reverse=True
        )
        s2_str = "  ".join(f"{w}号{v:.0f}%" for w, v in circ_sorted[:3])
        s3_str = "（展示後確認）"

    # 逃げ時買い目上位3点
    nige_buys = sorted(
        [c["combo"] for c in candidates if c.get("combo","").split("-")[0] == "1"],
        key=lambda x: next((c["prob"] for c in candidates if c["combo"]==x), 0),
        reverse=True
    )[:3]
    nige_buy_str = " / ".join(nige_buys) if nige_buys else "─"

    lines.append(
        f"\n② 逃げた場合\n"
        f"   2着候補：{s2_str}\n"
        f"   純3着候補：{s3_str}\n"
        f"   └ 買い目：{nige_buy_str}"
    )

    # ══════════════════════════════════════════════════════════════════
    # ③ 主役は誰か・決まり手は何か
    # ══════════════════════════════════════════════════════════════════
    main_w     = main_player.get("main_waku",  "-")
    main_type  = main_player.get("main_type",  "-")
    main_score = main_player.get("main_score", 0)
    sub_w      = main_player.get("sub_waku")
    sub_type   = main_player.get("sub_type",   "-")
    sub_score  = main_player.get("sub_score",  0)

    main_prob  = fp_map.get(str(main_w), 0)

    sub_line = ""
    if sub_w:
        sub_line = f"\n   対抗主役：{wn(sub_w)}【{sub_type}】{pct(sub_score)}"

    lines.append(
        f"\n③ 主役（逃げない場合）\n"
        f"   {wn(main_w)}【{main_type}】攻撃力{pct(main_score)} → 1着確率{pct(main_prob)}"
        f"{sub_line}"
    )

    # ══════════════════════════════════════════════════════════════════
    # ④ 主役が来た場合の2・3着（展開別残存マスタ活用）
    # ══════════════════════════════════════════════════════════════════
    p2_cands = main_player.get("place2_candidates", []) or []
    p3_cands = main_player.get("place3_candidates", []) or []

    p2_str = "  ".join(f"{w}号({s:.0f}pt)" for w, s in p2_cands[:3]) if p2_cands else "─"
    p3_str = "  ".join(f"{w}号({s:.0f}pt)" for w, s in p3_cands[:3]) if p3_cands else "─"

    # 主役頭の買い目上位3点
    main_buys = sorted(
        [c["combo"] for c in candidates
         if c.get("combo","").split("-")[0] == str(main_w)
         and not c.get("is_fallback_bet") and not c.get("is_dh_bet")],
        key=lambda x: next((c["prob"] for c in candidates if c["combo"]==x), 0),
        reverse=True
    )[:3]
    main_buy_str = " / ".join(main_buys) if main_buys else "─"

    lines.append(
        f"\n④ {wn(main_w)}が来た場合\n"
        f"   2着残存：{p2_str}\n"
        f"   3着残存：{p3_str}\n"
        f"   └ 買い目：{main_buy_str}"
    )

    # ══════════════════════════════════════════════════════════════════
    # ⑤ 主役が崩れたとき誰が浮上するか
    # ══════════════════════════════════════════════════════════════════
    dh_ok    = dark_horse.get("is_valid", False)
    dh_top   = dark_horse.get("top_waku",  "-")
    dh_score = dark_horse.get("top_score", 0)
    dh_cands = dark_horse.get("dark_horse_candidates", []) or []
    cb       = conflict_map.get("collapse_beneficiary", []) or []

    if dh_ok and dh_cands:
        dh_str = "  ".join(
            f"{w}号【{tag}】{s*100:.0f}%"
            for w, s, tag in dh_cands[:2]
        )
    elif cb:
        dh_str = "  ".join(f"{w}号({s*100:.0f}%)" for w, s in cb[:2])
    else:
        dh_str = "─"

    fly_type = escape_fb.get("fly_type", "-")
    lines.append(
        f"\n⑤ {wn(main_w)}が崩れた場合\n"
        f"   浮上候補：{dh_str}\n"
        f"   崩れ方：{fly_type}"
    )

    # ══════════════════════════════════════════════════════════════════
    # ⑥ 崩れ後に1号が逃げを拾う確率
    # ══════════════════════════════════════════════════════════════════
    fb_prob = escape_fb.get("fallback_prob", 0)
    fb_rank = escape_fb.get("fallback_rank", "-")
    fb_pct  = escape_fb.get("fallback_pct",  pct(fb_prob))

    fb_emoji = {"高": "🟢", "中": "🟡", "低": "🔴"}.get(fb_rank, "⚪")
    lines.append(
        f"\n⑥ 崩れ後に1号が残す確率：{fb_pct}【{fb_rank}】{fb_emoji}"
    )

    # ══════════════════════════════════════════════════════════════════
    # ⑦ 他の艇が展開を突いたとき（SC漁夫・2着3着）
    # ══════════════════════════════════════════════════════════════════
    sc_buys = sorted(
        [c["combo"] for c in candidates if c.get("is_sc_bet")],
        key=lambda x: next((c["prob"] for c in candidates if c["combo"]==x), 0),
        reverse=True
    )[:2]
    fb_buys = sorted(
        [c["combo"] for c in candidates if c.get("is_fallback_bet")],
        key=lambda x: next((c["prob"] for c in candidates if c["combo"]==x), 0),
        reverse=True
    )[:2]

    other_buys = sc_buys + fb_buys
    other_buy_str = " / ".join(other_buys) if other_buys else "─"

    if dh_ok:
        tobi_line = f"   展開突き候補：{wn(dh_top)}（浮上スコア{dh_score*100:.0f}%）"
    elif cb:
        top_cb = cb[0]
        tobi_line = f"   展開突き候補：{wn(top_cb[0])}（漁夫{top_cb[1]*100:.0f}%）"
    else:
        tobi_line = "   展開突き候補：─"

    lines.append(
        f"\n⑦ 展開を他の艇が突く場合\n"
        f"{tobi_line}\n"
        f"   └ 潰れ・残存買い目：{other_buy_str}"
    )

    # ══════════════════════════════════════════════════════════════════
    # → 最終買い目サマリー
    # ══════════════════════════════════════════════════════════════════
    total_pts  = len(candidates)
    syn_odds   = bet.get("theory_syn_odds")
    skip       = bet.get("skip", False)
    skip_rsn   = bet.get("skip_reason", "")
    entry_grade = bet.get("entry_grade", "-")

    # 展開パターン絵文字
    tp_emoji = {"A": "🟢鉄板", "B": "🔴主役", "C": "🟡拮抗", "D": "🟣荒れ"}.get(tenkai_pat, "⚪")

    # 全買い目をセクション別に集計
    n_nige  = sum(1 for c in candidates if c.get("combo","").split("-")[0] == "1"
                  and not c.get("is_fallback_bet") and not c.get("is_dh_bet"))
    n_main  = sum(1 for c in candidates if c.get("combo","").split("-")[0] == str(main_w)
                  and not c.get("is_fallback_bet") and not c.get("is_dh_bet"))
    n_other = total_pts - n_nige - n_main

    skip_line = f"\n   ⛔ {skip_rsn}" if skip else ""
    syn_line  = f"理論合成{syn_odds}倍" if syn_odds else ""

    lines.append(
        f"\n{'━'*30}\n"
        f"展開：{tp_emoji}  参加：{entry_grade}  {syn_line}\n"
        f"逃げ軸{n_nige}点 / 主役軸{n_main}点 / その他{n_other}点 = 計{total_pts}点"
        f"{skip_line}"
    )

    return "\n".join(lines)


def _generate_buy_hint(results, venue, venue_stats, race_judgment, bet_suggestions):
    """
    レース考察を初心者向けの短い文章にまとめて返す。
    fill_newspaper.py の kosatsu_raw（300文字制限）に収まるよう設計。

    【設計方針】
    buy_listを統計的に分析（1着/2着の分布・印◎の位置）してから考察を書く。
    「buy_list先頭だけ見る」方式では拾えなかった両建て・印◎対抗を正確に反映。
    """
    rj  = race_judgment  or {}
    bet = bet_suggestions or {}

    s1_prob      = rj.get("s1_prob")  or bet.get("s1_prob")  or 0.0
    affinity     = rj.get("affinity",    {}) or {}
    himo_are     = rj.get("himo_are",    {}) or {}
    conflict_map = rj.get("conflict_map", {}) or {}

    aff_summary   = affinity.get("affinity_summary", {}) or {}
    attack_score  = affinity.get("attack_score", {}) or {}
    ha_verdict    = himo_are.get("verdict", "対象外")
    scenario_type = bet.get("scenario_type", "")
    vc1           = rj.get("venue_c1_win_rate")
    collapse_bene = conflict_map.get("collapse_beneficiary", []) or []

    rd = [r for r in (results or []) if isinstance(r, dict)]
    name_map   = {str(r.get("waku","")): r.get("name","")   for r in rd}
    honmei_map = {str(r.get("waku","")): r.get("honmei","") for r in rd}

    def wn(w):
        n = name_map.get(str(w), "")
        h = honmei_map.get(str(w), "")
        mark = f"【{h}】" if h and h.strip() else ""
        return f"{w}号{mark}{n}" if n else f"{w}号艇"

    r1    = next((r for r in rd if str(r.get("waku","")) == "1"), {})
    r1_st = r1.get("avg_st")
    r1_name = wn(1)

    # ── 印◎の艇番を取得 ─────────────────────────────────────────────
    honmei_w = next(
        (str(r.get("waku","")) for r in rd if str(r.get("honmei","")).strip() == "◎"),
        None
    )

    # ── buy_listを統計分析 ───────────────────────────────────────────
    buy_list = bet.get("buy_list") or []

    # 1着・2着・3着ごとに艇番の出現回数を集計
    from collections import Counter as _Counter
    first_counter  = _Counter()
    second_counter = _Counter()
    third_counter  = _Counter()
    for bl in buy_list:
        parts = str(bl).split("-")
        if len(parts) >= 1 and parts[0].isdigit():
            first_counter[parts[0]] += 1
        if len(parts) >= 2 and parts[1].isdigit():
            second_counter[parts[1]] += 1
        if len(parts) >= 3 and parts[2].isdigit():
            third_counter[parts[2]] += 1

    total = len(buy_list)

    # 1着軸（最多頭数の艇）
    main_axis = first_counter.most_common(1)[0][0] if first_counter else "1"

    # 1号以外の頭買い点数
    non1_first = sum(v for k, v in first_counter.items() if k != "1")

    # 両建て判定: 1号頭と非1号頭が両方3点以上、かつ1号頭が全体の2/3未満
    # 1号頭が過半数を大きく超える場合は逃げ軸として扱う
    w1_first = first_counter.get("1", 0)
    is_ryotate = (
        w1_first >= 3
        and non1_first >= 3
        and (total == 0 or w1_first / total < 0.60)  # 1号頭6割未満のときだけ両建て
    )

    # ST評価
    if r1_st is not None:
        st_w = "速め" if r1_st <= 0.150 else "遅め" if r1_st >= 0.185 else "標準"
    else:
        st_w = "標準"

    # 会場特性は考察文に出さない（システムが内部で判断済み・混乱を招くため削除）
    venue_w = ""

    # 決まり手テキスト（長い順にマッチ）
    _TT_ITEMS = [("まくり差し","まくり差し"),("まくり系","まくり・まくり差し"),("まくり","まくり"),("差し","差し")]
    def weapon_txt(w):
        s = aff_summary.get(str(w), "")
        return next((v for k, v in _TT_ITEMS if k in s), "攻め")

    # 2着・3着に多く登場する艇（1着軸以外）
    himo_candidates = [
        w for w, _ in (second_counter + third_counter).most_common()
        if w != main_axis
    ]

    # 漁夫候補（主役上位・1号・軸艇以外）
    outer = sorted([(w,v) for w,v in attack_score.items() if w!="1"], key=lambda x:x[1], reverse=True)
    main_wakus = {str(w) for w,_ in outer[:2]} | {"1", main_axis}
    gyofu = [(w,s) for w,s in collapse_bene if str(w) not in main_wakus]

    parts = []

    # ── ①レースの基本構図 ────────────────────────────────────────────
    if is_ryotate:
        # 両建て: 1号逃げと非1号頭が拮抗
        sub_axes = [k for k,v in first_counter.most_common() if k != "1"][:1]
        sub_name = wn(sub_axes[0]) if sub_axes else "主役候補"
        line = f"{r1_name}の逃げと{sub_name}の攻めが拮抗。"
        if venue_w:
            line += f"（{venue_w}）"
        parts.append(line)
    elif main_axis == "1":
        # 逃げ軸
        if s1_prob >= 0.65:
            line = f"{r1_name}はST{st_w}で逃げ切り濃厚。"
        elif s1_prob >= 0.50:
            line = f"{r1_name}はST{st_w}で逃げ優勢。"
        else:
            line = f"{r1_name}はST{st_w}で逃げに期待。"
        if venue_w:
            line += f"（{venue_w}）"
        # 印◎が1号以外で買い目にも◎頭がある → ①の行末に自然につなげる
        if honmei_w and honmei_w != "1" and first_counter.get(honmei_w, 0) >= 1:
            line += f" {wn(honmei_w)}も1着あり。"
        parts.append(line)
    else:
        # 非1号軸
        if s1_prob >= 0.50:
            parts.append(f"{wn(main_axis)}を1着本命に。{r1_name}は相手候補筆頭。")
        else:
            parts.append(f"{r1_name}の逃げより{wn(main_axis)}の攻めが上回る展開。")

    # ── ②警戒艇（attack_score上位・軸艇以外）────────────────────────
    warn_outer = [(w,v) for w,v in outer if str(w) != main_axis]
    if warn_outer:
        if main_axis == "1" or is_ryotate:
            top_w, _ = warn_outer[0]
            warn_line = f"警戒は{wn(top_w)}の{weapon_txt(top_w)}。"
            if len(warn_outer) >= 2:
                sec_w, _ = warn_outer[1]
                warn_line += f"次点は{wn(sec_w)}。"
        else:
            # 非1号軸: 軸艇の武器と対抗を1行に
            axis_line = f"{wn(main_axis)}の{weapon_txt(main_axis)}が軸。"
            sub = warn_outer[0][0]
            axis_line += f"対抗は{wn(sub)}。"
            warn_line = axis_line
        parts.append(warn_line)

    # ── ③2・3着候補（ヒモ）──────────────────────────────────────────
    # second+third_counterの上位から軸艇・既出艇を除いて2艇まで
    himo_show = [w for w in himo_candidates
                 if w not in {main_axis} and w != "1" or (w == "1" and main_axis != "1")][:2]
    # 1号が相手候補筆頭の場合は先頭に
    if main_axis != "1" and "1" in (second_counter + third_counter):
        himo_show = ["1"] + [w for w in himo_show if w != "1"]

    if himo_show:
        names = "・".join(wn(w) for w in himo_show[:2])
        parts.append(f"{names}も2・3着に入ってくる可能性がある。")

    # ── ④ヒモ広め/絞り ──────────────────────────────────────────────
    if ha_verdict == "参加推奨":
        parts.append("ヒモは広めに。")
    elif ha_verdict == "点数絞り":
        parts.append("ヒモは絞り推奨。")
    elif ha_verdict == "不参加推奨":
        parts.append("見送りも検討。")

    # ── ⑤まとめ ─────────────────────────────────────────────────────
    if is_ryotate:
        sub_axes2 = [k for k,v in first_counter.most_common() if k != "1"][:1]
        fly_w = sub_axes2[0] if sub_axes2 else main_axis
        parts.append(f"→{r1_name}軸と{wn(fly_w)}軸を両建て。展示後に絞る。")
    elif main_axis == "1":
        # 印◎が別艇で買い目に◎頭がある場合は「中心＋押さえ」表現に
        if honmei_w and honmei_w != "1" and first_counter.get(honmei_w, 0) >= 1:
            parts.append(f"→{r1_name}1着中心。{wn(honmei_w)}も1着で押さえ。")
        else:
            parts.append(f"→{r1_name}1着固定で流す。")
    else:
        parts.append(f"→{wn(main_axis)}1着固定で流す。")

    return "\n".join(parts)




# ============================================================
# ★ 新機能③ イン逃げ / イン飛び / 両建て 3択判定
# ============================================================
def _judge_ryotate(race_judgment, tobi_scenario, venue_stats, s1_prob=None):
    """
    「イン逃げ狙い」「イン飛び狙い」「両建て」の3択を判定する。

    【判定ロジック v2 — 確率モデルとの一貫性確保】
    s1_prob（確率モデル由来）が渡された場合はそれを主軸にし、
    escape_score（定性スコア）との乖離を検出・調整する。

    - s1_prob が渡される場合（bet_suggestions確定後の再呼び出し）:
        確率値を直接 verdict の判定基準として使用し、
        定性スコアとの整合チェックを行う。
    - s1_prob が渡されない場合（初回暫定計算）:
        従来通り escape_score / tobi_score の差分で判定。

    整合性チェック:
      s1_prob が高い（≥0.65）のに escape_score が低い（<50）場合 →
        確率モデル優先で逃げスコアを上方補正し警告フラグを立てる。
      tobi_score が高い（≥55）のに s1_prob も高い（≥0.60）場合 →
        両建て推奨とし矛盾フラグを立てる。

    出力:
      verdict          : "逃げ狙い" / "飛び狙い" / "両建て推奨"
      confidence       : 判断の確信度 (0〜100)
      escape_score     : 逃げスコア（補正後）
      tobi_score       : 飛びスコア
      escape_pct       : 逃げ確率%（s1_prob * 100、確率モデル由来）
      tobi_pct         : 飛び確率%（(1 - s1_prob) * 100）
      reason           : 判定根拠
      buy_style        : 買い方の具体的指示
      consistency_warn : 定性スコアと確率モデルの乖離警告（True/False）
      realtime_hook    : リアルタイム情報で確認すべき事項
    """
    escape_score = float(race_judgment.get("score", 50))
    tobi_score   = float(tobi_scenario.get("tobi_prob", 30))

    # ── s1_prob による整合性チェックと補正 ──────────────────────────────
    consistency_warn = False
    consistency_note = ""

    if s1_prob is not None:
        fly_prob = 1.0 - s1_prob
        # s1_probを0〜100スケールのスコアに変換して escape_score と比較
        s1_score_equiv = s1_prob * 100.0

        # ケース①: 確率モデルは逃げ優勢なのに定性スコアが低い
        if s1_prob >= 0.65 and escape_score < 50:
            consistency_warn = True
            consistency_note = (
                f"⚠️ 確率モデル逃げ{s1_prob*100:.0f}%だが定性スコア{escape_score:.0f}。"
                f"確率モデル優先で補正。"
            )
            # 確率モデルに寄せて escape_score を上方補正（加重平均）
            escape_score = round(escape_score * 0.35 + s1_score_equiv * 0.65, 1)

        # ケース②: 確率モデルは飛び優勢なのに定性スコアが高い
        elif s1_prob < 0.45 and escape_score >= 65:
            consistency_warn = True
            consistency_note = (
                f"⚠️ 確率モデル逃げ{s1_prob*100:.0f}%（飛び優勢）だが定性スコア{escape_score:.0f}（高）。"
                f"両建てに引き寄せ。"
            )
            # 確率モデルに寄せて escape_score を下方補正
            escape_score = round(escape_score * 0.35 + s1_score_equiv * 0.65, 1)

        # ケース③: 飛びスコアが高いのに s1_prob も高い（真の矛盾）
        elif tobi_score >= 55 and s1_prob >= 0.60:
            consistency_warn = True
            consistency_note = (
                f"⚠️ 飛びスコア{tobi_score:.0f}と逃げ確率{s1_prob*100:.0f}%が矛盾。"
                f"両建て推奨に強制。"
            )
            # この場合は両建てに強制するため escape_score を中間値に
            escape_score = round((escape_score + s1_score_equiv) / 2, 1)
            tobi_score   = round(tobi_score * 0.8, 1)   # 飛びスコアも緩める
    else:
        fly_prob = None

    diff = escape_score - tobi_score

    # ── 表示用確率（s1_prob 確定後のみ意味を持つ） ───────────────────────
    escape_pct = round(s1_prob * 100, 1) if s1_prob is not None else None
    tobi_pct   = round((1.0 - s1_prob) * 100, 1) if s1_prob is not None else None

    # リアルタイム情報差し込み口
    realtime_hook = {
        "展示タイム_確認事項":  "1号艇の展示タイムがレース内偏差値50未満なら飛びスコア+10",
        "直前オッズ_確認事項":  "1号艇単勝オッズが1.2倍未満なら過剰人気→飛び狙いの妙味UP",
        "進入_確認事項":        "枠なり進入確認後にコース変更があれば再判定を推奨",
        "hook_fn":              None,
    }

    if escape_score >= 65 and tobi_score < 40:
        verdict     = "逃げ狙い"
        confidence  = min(100, int(escape_score * 0.7 + (65 - tobi_score) * 0.3))
        buy_style   = (
            f"1号艇1着固定。2着は「circle_pct（イン逃げ時2着率）上位2〜3艇」に絞る。"
            f"3連単で5〜8点。"
        )
        reason = (
            f"逃げスコア{escape_score:.0f}（高）・飛びスコア{tobi_score:.0f}（低）。"
            f"イン逃げが成立しやすい局面。"
        )
        if escape_pct is not None:
            reason += f" 確率モデル逃げ{escape_pct:.0f}%。"

    elif tobi_score >= 55 and escape_score < 55:
        threat = tobi_scenario.get("main_threat", "-")
        ttype  = tobi_scenario.get("tobi_type", "不明")
        verdict     = "飛び狙い"
        confidence  = min(100, int(tobi_score * 0.7 + (65 - escape_score) * 0.3))
        buy_style   = (
            f"{threat}号艇1着候補（{ttype}）。2・3着は残りの内寄り艇を広めに。"
            f"3連単で6〜10点。1号艇2・3着付けは除外か最小限に。"
        )
        reason = (
            f"飛びスコア{tobi_score:.0f}（高）・逃げスコア{escape_score:.0f}（低）。"
            f"{threat}号艇が主な脅威（{ttype}）。"
        )
        if tobi_pct is not None:
            reason += f" 確率モデル飛び{tobi_pct:.0f}%。"

    else:
        verdict     = "両建て推奨"
        confidence  = max(10, 80 - int(abs(diff) * 1.5))
        threat = tobi_scenario.get("main_threat", "-")
        ttype  = tobi_scenario.get("tobi_type", "不明")
        buy_style   = (
            f"【逃げ軸】1号艇1着の買い目を確保（4〜5点）。"
            f"【飛び軸】{threat}号艇1着の買い目（3〜4点）。"
            f"合計7〜9点。両軸を保持し展示後に軸の比重を傾ける。"
        )
        reason = (
            f"逃げスコア{escape_score:.0f}・飛びスコア{tobi_score:.0f}（差{abs(diff):.0f}pt）。"
            f"拮抗しており単軸は危険。{threat}号艇({ttype})が飛ばし役候補。"
        )
        if escape_pct is not None:
            reason += f" 確率モデル逃げ{escape_pct:.0f}%/飛び{tobi_pct:.0f}%。"

    if consistency_warn and consistency_note:
        reason = consistency_note + " " + reason

    return {
        "verdict":          verdict,
        "confidence":       confidence,
        "escape_score":     round(escape_score, 1),
        "tobi_score":       round(tobi_score, 1),
        "escape_pct":       escape_pct,    # 確率モデル由来の逃げ%（None=初回暫定）
        "tobi_pct":         tobi_pct,      # 確率モデル由来の飛び%
        "reason":           reason,
        "buy_style":        buy_style,
        "consistency_warn": consistency_warn,
        "realtime_hook":    realtime_hook,
    }


def _judge_himo_are(results, race_judgment):
    """
    ヒモ荒れ判定：1号艇が有力本命のときに「2・3着ヒモが荒れるか」を評価する。

    【v2 変更点】
    旧版は rel_win1 >= 60% という高すぎる閾値で「対象外」が大半を占めていた。
    番組表確定時点では 45% 以上あれば参加/見送り判断の材料として十分有効。

    【判定フロー v2】
    Step1: rel_win1 >= 45% → 判定対象（旧60%→45%に緩和）
    Step2: 最有力3連単の推定確率と2着集中度でヒモ固まり度を評価
    Step3: 展示前推奨アクションを明示（展示確認トリガーを含む）

    【閾値基準 v2】
    max_combo_prob:
      >= 0.25 → 不参加推奨（推定最高人気オッズ ≈ 5倍以下）
      0.12〜0.25 → 点数絞り
      < 0.12   → 参加推奨（ヒモ分散・広め流し）
    """
    res1     = next((r for r in results if r["waku"] == "1"), None)
    rel_win1 = res1.get("rel_win1") if res1 else None

    NOT_TARGET = {
        "is_target": False, "verdict": "対象外", "max_combo_prob": None,
        "est_top_odds": None, "circle_concentration": None,
        "eligible_count": 0, "bet_adj": 0,
        "reason": "1号艇rel_win1 < 45%: 通常判定に委ねる",
        "tenji_trigger": "",
    }
    if rel_win1 is None or rel_win1 < 45.0:
        return NOT_TARGET

    # ── Step2: 組み合わせ確率を計算 ──────────────────────────────────────
    wakus_rest = [r["waku"] for r in results if r["waku"] != "1"]
    sum_rel    = sum(r.get("rel_win1") or 0 for r in results) or 100.0
    p1_win     = rel_win1 / sum_rel

    circ_raw  = {r["waku"]: max(r.get("circle_pct") or 0, 0.001)
                 for r in results if r["waku"] != "1"}
    total_circ = sum(circ_raw.values()) or 1.0
    idx3_raw  = {r["waku"]: max(float(r.get("idx3") or 0), 0.001)
                 for r in results if r["waku"] != "1"}

    combo_probs = []
    for second in wakus_rest:
        p2 = circ_raw[second] / total_circ
        remaining = [w for w in wakus_rest if w != second]
        total_i3  = sum(idx3_raw[w] for w in remaining) or 1.0
        for third in remaining:
            p3   = idx3_raw[third] / total_i3
            prob = p1_win * p2 * p3
            combo_probs.append((f"1-{second}-{third}", prob))

    combo_probs.sort(key=lambda x: x[1], reverse=True)
    max_combo_prob = combo_probs[0][1] if combo_probs else 0.0
    est_top_odds   = round(1.0 / (max_combo_prob * 0.75), 1) if max_combo_prob > 0 else 999.9

    # ── Step3: 2着集中度 ──────────────────────────────────────────────────
    circ_sorted   = sorted(circ_raw.items(), key=lambda x: x[1], reverse=True)
    top2_circ_sum = (
        (circ_sorted[0][1] + circ_sorted[1][1]) / total_circ * 100
        if len(circ_sorted) >= 2 else 100.0
    )

    # ── Step4: 有効組み合わせ数 ────────────────────────────────────────────
    rank     = (race_judgment or {}).get("rank", "B")
    TARGET_BETS = {"S": 6, "A": 8, "B": 9, "C": 6, "D": 6}
    target_n = TARGET_BETS.get(rank, 8)
    _vc1 = (race_judgment or {}).get("venue_c1_win_rate")
    if _vc1 is not None:
        target_n += 2 if _vc1 < 0.45 else (1 if _vc1 < 0.50 else 0)
    eligible_count = min(target_n, len(combo_probs))

    # ── Step5: 総合判定（v2: 閾値を緩和・展示確認トリガー追加）─────────────
    circ_adj = 0.04 if top2_circ_sum >= 70 else (-0.04 if top2_circ_sum < 55 else 0.0)
    prob_adj = max_combo_prob + circ_adj

    # 展示確認トリガー（展示前システムとして必須）
    # 1号艇のST実績をチェックしてトリガー文言を組み立てる
    st1 = res1.get("avg_st") if res1 else None
    st_trigger = ""
    if st1 is not None:
        if st1 > 0.18:
            st_trigger = f"展示で1号艇スタート遅め({st1:.3f}秒)確認→ヒモが荒れやすい"
        elif st1 < 0.14:
            st_trigger = f"展示で1号艇スタート超安定({st1:.3f}秒)→ヒモ固まり方向に注意"

    reasons = [
        f"rel_win1={rel_win1:.1f}%（本命度）",
        f"最有力組み合わせ確率: {max_combo_prob:.3f}（推定実オッズ≈{est_top_odds:.0f}倍）",
        f"2着集中度: 上位2艇 {top2_circ_sum:.1f}%（補正{circ_adj:+.2f}）",
        f"総合判定確率: {prob_adj:.3f}",
    ]

    if prob_adj >= 0.25:
        verdict = "不参加推奨"
        bet_adj = -99
        tenji_trigger = (
            f"展示確認: 1号艇の伸びが平凡以下なら見送り確定。"
            f"好伸びでも推定オッズ{est_top_odds:.0f}倍台のため回収期待値低。"
            + (f"\n{st_trigger}" if st_trigger else "")
        )
        reasons.append(f"→ ヒモ固まり（推定1番人気{est_top_odds:.0f}倍台）。回収率が構造的に低い")
    elif prob_adj >= 0.12:
        verdict = "点数絞り"
        bet_adj = 0
        tenji_trigger = (
            f"展示確認: 1号艇伸び良→そのまま採用。悪→飛び組に差し替え。"
            f"circle_pct上位2艇に絞り込むこと。"
            + (f"\n{st_trigger}" if st_trigger else "")
        )
        reasons.append("→ ヒモはやや固め。展示後に上位2艇以外を切ること")
    else:
        verdict = "参加推奨"
        bet_adj = +2
        tenji_trigger = (
            f"展示確認: ヒモ分散レース。1号艇伸び確認後、広めのヒモ流しを採用。"
            f"穴ヒモ（5・6号艇）を積極的に含める。"
            + (f"\n{st_trigger}" if st_trigger else "")
        )
        reasons.append("→ ヒモ分散。1号艇1着固定で広めのヒモ流し推奨（買い目+2点）")

    return {
        "is_target":            True,
        "verdict":              verdict,
        "max_combo_prob":       round(max_combo_prob, 4),
        "est_top_odds":         est_top_odds,
        "circle_concentration": round(top2_circ_sum, 1),
        "eligible_count":       eligible_count,
        "bet_adj":              bet_adj,
        "reason":               " / ".join(reasons),
        "tenji_trigger":        tenji_trigger,
    }


def _build_kimari(cm):
    parts = []
    mapping = [
        ("逃げ%", "逃"),
        ("差し%", "差"),
        ("まくり%", "ま"),
        ("まくり差し%", "差ま"),
        ("抜き%", "抜"),
    ]
    for col, label in mapping:
        v = safe_float(_get_cm_val(cm, col))
        if v and v > 0:
            parts.append(f"{label}{int(round(v))}%")
    return " ".join(parts) if parts else "-"

def _judge_race_type(results, venue_stats, venue_frame, race_no=None,
                     venue_stats_master=None, venue=None, tobi_scenario=None):
    """
    荒れ/堅いレース自動判定。乗法モデルで複合確率スコアを計算して S/A/B/C/D を返す。
    score 高（1.0に近い）→ 堅い、低（0.0に近い）→ 荒れ

    【旧方式の問題点と修正理由】
    旧方式: 各指標を単純に加算/減算 → score が 50±αで計算
      問題①: 加算モデルは独立事象の積を無視する
        例）「1号艇3連対率92%」かつ「会場イン逃げ率60%」の複合は
             P(堅い) = 0.92 × 0.60 ≒ 0.55 だが、加算では過大評価される
      問題②: score が 0〜100 に張り付きやすく、ランク境界が不安定
      問題③: 各指標の重みが明示されず、調整根拠が不明

    【新方式: 乗法モデル】
      base_prob = 各確率因子の積（0〜1）
      修正因子（乗算）= 連続値で補正（0.5〜1.5程度）
      最終スコア = base_prob × 修正因子群 を 0〜100 にスケール

      確率因子（直接確率として解釈できるもの）:
        P1: 1号艇の相対1着率 / 理論値(40%)   正規化乗数
        P2: 1号艇の3連対率                  直接使用
        P3: 会場イン逃げ率                  直接使用
        P4: 1号艇がTop-1の確率（1位集中度）  Top1比率

      修正因子（確率ではなく補正係数）:
        M1: フォーム指数補正     0.85〜1.15
        M2: ST安定スコア補正     0.90〜1.10
        M3: FLY/出遅れペナルティ 0.75〜1.00
        M4: 飛びシナリオ補正     0.60〜1.00
        M5: R番号補正            0.90〜1.05
        M6: データ不足補正       0.70〜1.00
    """
    reasons = []
    honmei_concentrated = False
    two_top_race = False

    res1 = next((r for r in results if r["waku"] == "1"), None)

    # ─── 確率因子 ───────────────────────────────────────────────────────────

    # P1: 1号艇の相対1着率（理論値40%との比で正規化、上限1.5・下限0.3）
    p1 = 1.0
    if res1 and res1.get("rel_win1") is not None:
        r1 = res1["rel_win1"]
        p1 = max(0.30, min(1.50, r1 / 40.0))
        if r1 >= 45:
            reasons.append(f"1号艇相対1着率{r1:.1f}%（突出 ×{p1:.2f}）")
        elif r1 >= 35:
            reasons.append(f"1号艇相対1着率{r1:.1f}%（高い ×{p1:.2f}）")
        elif r1 <= 20:
            reasons.append(f"1号艇相対1着率{r1:.1f}%（低い→荒れ要素 ×{p1:.2f}）")
        elif r1 <= 28:
            reasons.append(f"1号艇相対1着率{r1:.1f}%（やや低い ×{p1:.2f}）")

    # P2: 1号艇の3連対率（直接確率として使用、基準値0.75）
    # ※ 95%超は少数データ疑いのため0.90に上限クリップ
    p2 = 0.75  # データなし時のデフォルト
    if res1 and res1.get("win3_rate") is not None:
        w3 = float(res1["win3_rate"])
        if w3 >= 0.95:
            p2 = 0.90   # 高すぎ → 少数データ疑い → 上限クリップ
            reasons.append(f"1号艇3連対率{w3*100:.0f}%（高すぎ→少数データ疑い→0.90にクリップ）")
        elif 0.92 <= w3 < 0.95:
            p2 = w3
            reasons.append(f"1号艇3連対率{w3*100:.0f}%（最安定帯 92-95%）")
        elif w3 >= 0.90:
            p2 = w3
            reasons.append(f"1号艇3連対率{w3*100:.0f}%（安定型）")
        elif w3 <= 0.55:
            p2 = w3
            reasons.append(f"1号艇3連対率{w3*100:.0f}%（不安定）")
        else:
            p2 = w3

    # P3: 会場イン逃げ率（直接確率として使用、デフォルト0.55）
    p3 = 0.55
    in_rate = venue_stats.get("in_rate")
    if in_rate is not None:
        p3 = float(in_rate)
        if in_rate >= 0.60:
            reasons.append(f"会場イン逃げ率{in_rate*100:.1f}%（逃げ率高）")
        elif in_rate <= 0.42:
            reasons.append(f"会場イン逃げ率{in_rate*100:.1f}%（荒れやすい会場）")

    # P4: 1号艇のトップ集中度（sorted_rel[0] / sorted_rel[1] の比率）
    p4 = 1.0
    sorted_rel = sorted(
        [r["rel_win1"] for r in results if r.get("rel_win1") is not None],
        reverse=True
    )
    if len(sorted_rel) >= 2:
        gap = sorted_rel[0] - sorted_rel[1]
        h2_threat = sorted_rel[1]
        if h2_threat >= 22 and gap < 15:
            # 実質2強 → 集中度低下
            p4 = 0.80
            two_top_race = True
            reasons.append(f"2番手脅威{h2_threat:.1f}%・差{gap:.1f}pt（実質2強→2軸推奨 ×{p4:.2f}）")
        elif gap >= 20:
            # 本命突出
            p4 = 1.20
            honmei_concentrated = True
            reasons.append(f"1・2位差{gap:.1f}pt（本命突出 ×{p4:.2f}）")
        elif gap <= 8:
            # 実力拮抗
            p4 = 0.85
            reasons.append(f"1・2位差{gap:.1f}pt（実力拮抗 ×{p4:.2f}）")

    # base_prob: P1〜P4の積（0〜1.5程度、後でスケール化）
    base_prob = p1 * p2 * p3 * p4

    # ─── 修正因子 ───────────────────────────────────────────────────────────

    # M1: フォーム指数補正
    m1 = 1.0
    if res1:
        pm1 = res1.get("raw_pm", {})
        form_idx = safe_float(pm1.get("フォーム\n指数") or pm1.get("フォーム指数"))
        if form_idx is not None:
            if form_idx >= 8.0:
                m1 = 1.12; reasons.append(f"1号艇フォーム指数{form_idx:.1f}（🔥ホット ×{m1:.2f}）")
            elif form_idx < 4.0:
                m1 = 0.88; reasons.append(f"1号艇フォーム指数{form_idx:.1f}（❄コールド ×{m1:.2f}）")
            elif form_idx < 6.0:
                m1 = 0.94; reasons.append(f"1号艇フォーム指数{form_idx:.1f}（やや不調 ×{m1:.2f}）")

    # M2: ST安定スコア補正
    m2 = 1.0
    if res1:
        pm1 = res1.get("raw_pm", {})
        st_stable = safe_float(pm1.get("ST安定\nスコア") or pm1.get("ST安定スコア"))
        if st_stable is not None:
            if st_stable >= 80:
                m2 = 1.08; reasons.append(f"1号艇ST安定スコア{st_stable:.0f}（◎超安定 ×{m2:.2f}）")
            elif st_stable >= 60:
                m2 = 1.04; reasons.append(f"1号艇ST安定スコア{st_stable:.0f}（○安定 ×{m2:.2f}）")
            elif st_stable < 40:
                m2 = 0.92; reasons.append(f"1号艇ST安定スコア{st_stable:.0f}（×不安定→ST信頼性低 ×{m2:.2f}）")

    # M3: FLY数・出遅れ数によるペナルティ
    m3 = 1.0
    if res1:
        pm1 = res1.get("raw_pm", {})
        fly_count_j  = safe_float(pm1.get("FLY数"),    0) or 0
        late_count_j = safe_float(pm1.get("出遅れ数"), 0) or 0
        st_meas_j    = safe_float(pm1.get("ST計測件数") or pm1.get("総出走数"), 1) or 1
        late_rate_j  = late_count_j / max(st_meas_j, 1)

        if fly_count_j >= 2:
            m3 *= 0.82; reasons.append(f"1号艇FLY{int(fly_count_j)}回（直近複数FLY ×0.82）")
        elif fly_count_j >= 1:
            m3 *= 0.91; reasons.append(f"1号艇FLY{int(fly_count_j)}回（直近FLYあり ×0.91）")

        if late_rate_j >= 0.10:
            m3 *= 0.88; reasons.append(f"1号艇出遅れ率{late_rate_j*100:.0f}%（出遅れ癖あり ×0.88）")
        elif late_rate_j >= 0.05:
            m3 *= 0.94; reasons.append(f"1号艇出遅れ率{late_rate_j*100:.0f}%（出遅れやや多い ×0.94）")

    # M4: イン飛びシナリオ補正（飛び確率が高いほど逃げ確率を下げる）
    m4 = 1.0
    if tobi_scenario is not None:
        tobi_prob = tobi_scenario.get("tobi_prob", 0)
        if tobi_prob >= 70:
            m4 = 0.62; reasons.append(f"飛び確率{tobi_prob:.0f}%（S：強い飛び示唆 ×{m4:.2f}）")
        elif tobi_prob >= 55:
            m4 = 0.75; reasons.append(f"飛び確率{tobi_prob:.0f}%（A：飛び傾向あり ×{m4:.2f}）")
        elif tobi_prob >= 40:
            m4 = 0.88; reasons.append(f"飛び確率{tobi_prob:.0f}%（B：やや飛びの余地 ×{m4:.2f}）")

    # M5: R番号補正（後半Rほど荒れやすい）
    m5 = 1.0
    NATIONAL_R_AREYASUSA = {
        1: 55, 2: 52, 3: 50, 4: 48, 5: 46, 6: 44,
        7: 43, 8: 42, 9: 41, 10: 40, 11: 39, 12: 38,
    }
    try:
        rno = int(race_no) if hasattr(race_no, '__int__') else int(str(race_no))
        r_are_score = None
        if venue_stats_master and venue:
            vs_m = venue_stats_master.get(venue, {})
            r_are_score = safe_float(vs_m.get(f"{rno}R荒れスコア"))
        if r_are_score is None:
            r_are_score = NATIONAL_R_AREYASUSA.get(rno, 45)
        # 荒れスコア50を基準に乗算係数へ変換（50→1.0、55→1.03、38→0.94）
        m5 = max(0.90, min(1.06, 1.0 + (float(r_are_score) - 50) * 0.006))
        if abs(m5 - 1.0) >= 0.02:
            direction = "堅め" if m5 > 1.0 else "荒れやすい"
            reasons.append(f"{rno}R荒れスコア{r_are_score:.0f}（{direction} ×{m5:.3f}）")
    except (ValueError, TypeError) as e:
        print(f"  ⚠️  R番号補正の計算をスキップしました: {e}")

    # M6: データ不足補正
    missing_list = [r for r in results if r.get("data_missing")]
    missing = len(missing_list)
    m6 = max(0.70, 1.0 - missing * 0.08)  # 1人欠け→×0.92、3人→×0.76
    data_trust_score = round((6 - missing) / 6 * 100)
    if missing >= 3:
        reasons.append(f"データ不足{missing}人（信頼性低 ×{m6:.2f}）")
    elif missing >= 1:
        reasons.append(f"データ不足{missing}人（×{m6:.2f}）")
        missing_names = "・".join(
            f"{r['waku']}号艇{r['name']}({r.get('missing_reason','')})"
            for r in missing_list
        )
        reasons.append(f"データ信頼度{data_trust_score}%（不足:{missing_names}）")

    # ─── 最終スコア計算 ────────────────────────────────────────────────────
    # base_prob（0〜約1.5）× 修正因子群 を 0〜100 にスケール
    # 基準値（全てデフォルトの場合）: 1.0 × 0.75 × 0.55 × 1.0 ≒ 0.41 → score ≒ 50
    raw_score  = base_prob * m1 * m2 * m3 * m4 * m5 * m6
    # 正規化: raw_score=0.41 → 50点 となるよう線形スケール（slope = 50/0.41 ≒ 122）
    NORMALIZATION_CENTER = 0.41   # 全デフォルト値での期待出力
    score = int(round(min(100, max(0, raw_score / NORMALIZATION_CENTER * 50))))

    reasons.insert(0, f"複合確率スコア: {raw_score:.4f} → {score}点")

    if score >= 80:
        rank, strategy, skip = "S", "◎1着固定 2-3着流し（5〜8点）", False
    elif score >= 60:
        rank, strategy, skip = "A", "◎-○軸 3着3〜4艇流し（10〜15点）", False
    elif score >= 45:
        rank, strategy, skip = "B", "◎○△フォーメーション（15〜20点）", False
    elif score >= 35:
        # 【改善】旧閾値30→35に厳格化。過去バックテストでC帯の回収率が損益分岐を下回った。
        rank, strategy, skip = "C", "荒れ傾向。高配当狙いか見送り", False
    else:
        # score < 35 → 大荒れ確実。見送り推奨。
        rank, strategy, skip = "D", "大荒れ。見送り推奨", True

    return {
        "rank":                rank,
        "score":               score,
        "reason":              reasons,
        "skip":                skip,
        "strategy":            strategy,
        "honmei_concentrated": honmei_concentrated,
        "two_top_race":        two_top_race,
        "data_trust_score":    data_trust_score,
    }


def _calc_3rentan_probs(results):
    """
    全120通りの3連単推定確率を計算して返す。
    計算モデル：
      P(1着=A)         = rel_win1 を正規化
      P(2着=B | 1着=A) = A除外後の残艇で再正規化
                         1号艇1着時はイン逃げ時2着率(circle_pct)で補正
      P(3着=C | 1着=A,2着=B) = A・B除外後の残艇でidx3を正規化
    戻り値: prob降順ソート済みリスト
    """
    wakus = [r["waku"] for r in results]
    if len(wakus) < 3:
        return []

    # 【改善】rel_win1=0 の艇にフロア確率を適用（0%買い目根絶）
    _FLOOR = {"1": 0.5, "2": 0.15, "3": 0.15, "4": 0.10, "5": 0.06, "6": 0.02}
    rel_map  = {r["waku"]: max(r.get("rel_win1") or 0, _FLOOR.get(r["waku"], 0.02)) for r in results}
    circ_map = {r["waku"]: max(r.get("circle_pct") or 0, 0.01) for r in results}
    idx3_map = {r["waku"]: max(r.get("idx3") or r.get("rel_win1") or 0, _FLOOR.get(r["waku"], 0.02)) for r in results}

    total_rel = sum(rel_map.values()) or 1
    p1 = {w: rel_map[w] / total_rel for w in wakus}

    combos = []
    for first in wakus:
        for second in wakus:
            if second == first:
                continue
            for third in wakus:
                if third == first or third == second:
                    continue
                prob_1 = p1[first]

                remaining_2 = [w for w in wakus if w != first]
                if first == "1":
                    w2 = {w: max(circ_map[w], 0.001) for w in remaining_2}
                else:
                    w2 = {w: max(rel_map[w], 0.001) for w in remaining_2}
                prob_2 = w2[second] / sum(w2.values())

                remaining_3 = [w for w in wakus if w != first and w != second]
                w3 = {w: max(idx3_map[w], 0.001) for w in remaining_3}
                prob_3 = w3[third] / sum(w3.values())

                combos.append({
                    "combo":          f"{first}-{second}-{third}",
                    "first":          first,
                    "second":         second,
                    "third":          third,
                    "prob":           prob_1 * prob_2 * prob_3,
                    "is_outer_first": int(first) >= 4,
                })

    combos.sort(key=lambda x: x["prob"], reverse=True)
    return combos


# ════════════════════════════════════════════════════════════════════════
# SCシナリオ（潰れ展開・漁夫の利）計算関数
# ════════════════════════════════════════════════════════════════════════

def _calc_sc_weight(results, cm_map, win3_map, rel_map, jizen_eval=None):
    """
    SCシナリオ（飛び役自滅→漁夫の利）の重みと、
    SC発動時の受益者スコアを計算して返す。

    【競艇物理法則に基づく設計】

    ■ 飛び役の決まり手タイプを「最大決まり手（1位）」で分類
        まくり系   : 最大決まり手が「まくり」または「まくり差し」
        差し系     : 最大決まり手が「差し」
        逃げ系(1号): SCシナリオ対象外（逃げが自滅しても外がそのまま1着になる）

    ■ まくり系飛び役が自滅した場合（外に膨らむ）
        物理法則: まくり艇が旋回で外に膨らむ
                  → コース(飛び役)より内側の艇のコースが開く（内側開放）
                  → コース(飛び役)より外側の艇が続いて捲る（外側継続）
        受益者:
          内側開放候補: コース < 飛び役コース の艇（1号艇除く）
                        ただし1号艇はS1逃げで既に前提計算済みなので2号艇以降
          外側継続候補: コース > 飛び役コース の艇
        重み付け:
          内側開放ボーナス: ×1.5（コースが開く = 進入路が確保される）
          外側継続ボーナス: ×0.8（勢いはあるが距離損）

    ■ 差し系飛び役が自滅した場合（蓋をされる）
        物理法則: 差し艇が1号艇に蓋をされて失速
                  → 後方から来た外側艇（まくり差し）が浮上する
        受益者: コース > 飛び役コース の艇（まくり差しで来た艇）
        重み付け:
          後方浮上ボーナス: ×1.2

    ■ 漁夫スコア計算
        漁夫スコア = 位置ボーナス
                   × (1 - 攻撃性正規化) × 0.5 + 0.5  ← 仕掛けに行かない艇を優遇
                   × win3_rate                          ← 地力（荒れても残る力）

        攻撃性 = (まくり% + まくり差し% + 差し%) / 100
                 高い → 自分も仕掛けに行く → 自滅リスクあり → 漁夫スコア低め
                 低い → 控えて後ろから残る → 漁夫スコア高め

    ■ SCシナリオの1着
        飛び役が自滅 → 1着を取り直すのは:
          パターンA: 1号艇（逃げ取り戻し） ← まくり系飛び役が自滅した場合に多い
          パターンB: 第2飛び役（sub_fly）   ← 差し系飛び役が自滅した場合
          → 両方をSCシナリオ内で重み付き按分

    ■ SCシナリオ重み（S4の代替）
        SC_base = S4_old の重みをベースに、以下で調整:
          飛び役の攻撃性が高い（自滅リスク大）→ SC重みを増やす
          飛び役の実績（win3_rate）が低い      → SC重みを増やす
        SC重み = SC_base × (1 + 飛び役攻撃性 × 0.5) × (1 + (1 - 飛び役win3) × 0.3)

    Returns
    -------
    dict:
        sc_weight       : SCシナリオの全体重み（p_s4の代替として使用）
        sc_1st_weights  : {waku: float} SC時の1着按分重み
        sc_beneficiary  : {waku: float} SC時の2・3着漁夫スコア
        sc_fly_type     : "まくり系" / "差し系" / "不明"
        sc_fly_waku     : 主要飛び役の艇番
    """
    wakus = [r["waku"] for r in results]
    if len(wakus) < 4:
        return {
            "sc_weight": 0.02,
            "sc_1st_weights": {"1": 1.0},
            "sc_beneficiary": {w: 0.1 for w in wakus},
            "sc_fly_type": "不明",
            "sc_fly_waku": None,
        }

    COURSE_NATIONAL_WIN = {"1": 0.555, "2": 0.137, "3": 0.134,
                           "4": 0.111, "5": 0.066, "6": 0.021}

    def safe_pct(cm, key):
        v = cm.get(key)
        try:
            return max(float(v), 0.0) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    # ── 飛び役の特定と決まり手タイプ分類 ──────────────────────────────────────
    # 1号艇以外で最も rel_win1 が高い艇を主要飛び役とする
    fly_candidates = sorted(
        [(r["waku"], rel_map.get(r["waku"], 0)) for r in results if r["waku"] != "1"],
        key=lambda x: x[1], reverse=True
    )
    main_fly_waku = fly_candidates[0][0] if fly_candidates else None
    sub_fly_waku  = fly_candidates[1][0] if len(fly_candidates) >= 2 else None

    if main_fly_waku is None:
        return {
            "sc_weight": 0.02,
            "sc_1st_weights": {"1": 1.0},
            "sc_beneficiary": {w: 0.1 for w in wakus},
            "sc_fly_type": "不明",
            "sc_fly_waku": None,
        }

    # 決まり手タイプ判定（最大決まり手1位で分類）
    fly_cm = cm_map.get(main_fly_waku, {})
    mak_pct     = safe_pct(fly_cm, "まくり%") + safe_pct(fly_cm, "まくり差し%")
    sashi_pct   = safe_pct(fly_cm, "差し%")
    nige_pct    = safe_pct(fly_cm, "逃げ%")

    scores = {"まくり系": mak_pct, "差し系": sashi_pct, "逃げ系": nige_pct}
    sc_fly_type = max(scores, key=scores.get)
    if scores[sc_fly_type] == 0:
        sc_fly_type = "まくり系"  # データなしはまくり系で代替

    # 攻撃性スコア（高いほど仕掛けに行く→自滅リスクあり）
    fly_attack = min((mak_pct + sashi_pct) / 100.0, 1.0)
    fly_win3   = win3_map.get(main_fly_waku, 0.5)

    # ── SCシナリオ重み ──────────────────────────────────────────────────────
    # 旧S4相当のベース（5-6号艇の全国平均微小値）を飛び役の特性で増幅
    sc_base = (
        COURSE_NATIONAL_WIN.get("5", 0.066) * 0.5 +
        COURSE_NATIONAL_WIN.get("6", 0.021) * 0.5
    )
    sc_weight = sc_base * (1.0 + fly_attack * 0.5) * (1.0 + (1.0 - fly_win3) * 0.3)

    # ── SC発動時の1着按分 ──────────────────────────────────────────────────
    # まくり系自滅 → 1号艇が取り戻す確率が高い（コースが開く）
    # 差し系自滅   → 第2飛び役が続く確率が高い
    if sc_fly_type == "まくり系":
        w1_1st = 0.65   # 1号艇が逃げ取り戻し
        w2_1st = 0.35   # 第2飛び役
    elif sc_fly_type == "差し系":
        w1_1st = 0.30   # 差し自滅では1号艇も蓋の影響を受ける
        w2_1st = 0.70   # 第2飛び役が浮上
    else:
        w1_1st = 0.50
        w2_1st = 0.50

    sc_1st_weights = {}
    sc_1st_weights["1"] = w1_1st
    if sub_fly_waku:
        sc_1st_weights[sub_fly_waku] = w2_1st
    else:
        sc_1st_weights["1"] = 1.0  # 第2飛び役なければ1号艇に集約

    # ── 漁夫スコア計算 ──────────────────────────────────────────────────────
    try:
        main_fly_course = int(main_fly_waku)
    except (ValueError, TypeError):
        main_fly_course = 3

    sc_beneficiary = {}
    for r in results:
        w = r["waku"]
        if w == main_fly_waku:
            # 自滅した飛び役自身は漁夫スコア0
            sc_beneficiary[w] = 0.0
            continue

        try:
            w_course = int(w)
        except (ValueError, TypeError):
            w_course = 3

        # 位置ボーナス（コース物理法則）
        if sc_fly_type == "まくり系":
            if w == "1":
                # 1号艇はSC時の1着候補なので2・3着漁夫スコアは低め
                pos_bonus = 0.6
            elif w_course < main_fly_course:
                # 内側開放候補（まくり艇が膨らんでコースが開く）
                pos_bonus = 1.5
            else:
                # 外側継続候補（まくり艇に続いて捲る）
                pos_bonus = 0.8
        elif sc_fly_type == "差し系":
            if w_course > main_fly_course:
                # 後方浮上候補（まくり差しで来た艇）
                pos_bonus = 1.2
            elif w == "1":
                # 蓋をした1号艇は先に行っているが差し系なので若干恩恵あり
                pos_bonus = 0.9
            else:
                pos_bonus = 0.7
        else:
            pos_bonus = 1.0

        # 攻撃性の逆数（仕掛けに行かない艇 = 漁夫スコア高め）
        w_cm      = cm_map.get(w, {})
        w_mak     = safe_pct(w_cm, "まくり%") + safe_pct(w_cm, "まくり差し%")
        w_sashi   = safe_pct(w_cm, "差し%")
        w_attack  = min((w_mak + w_sashi) / 100.0, 1.0)
        passivity = (1.0 - w_attack) * 0.5 + 0.5  # 0.5〜1.0 の範囲

        # 地力（荒れた展開でも着に残る力）
        ground    = win3_map.get(w, 0.5)

        sc_beneficiary[w] = pos_bonus * passivity * ground

    return {
        "sc_weight":      sc_weight,
        "sc_1st_weights": sc_1st_weights,
        "sc_beneficiary": sc_beneficiary,
        "sc_fly_type":    sc_fly_type,
        "sc_fly_waku":    main_fly_waku,
    }


def _calc_3rentan_probs_v2(results, venue_course_1c_rate=None, jizen_eval=None, race_judgment=None,
                            tenkai_national=None, tenkai_venue=None, venue_stats=None):
    """
    【改善②】展開シナリオ×条件付き確率モデル

    旧方式の問題：各艇が独立に競争すると仮定 → 展開の依存関係を無視
    　例）1号艇が逃げた時とまくりが決まった時では2着・3着の確率分布が全く異なる

    新方式：4つの展開シナリオをまず確率決定し、
            シナリオ内で条件付き確率（P(2着|シナリオ)、P(3着|シナリオ,2着)）を計算する。

    ─────────────────────────────────────────────────────────────
    シナリオ定義：
      S1: イン逃げ   … 1号艇が1着（コース1の逃げ決まり手%で推定）
      S2: 差し       … 内側艇(2-3)が差して1着
      S3: まくり系   … 外側艇(3-5)がまくり/まくり差しで1着
      S4: 大荒れ     … 5-6号艇1着（確率は小さいが高配当に寄与）

    各シナリオの重み：
      S1 ∝ 1号艇の逃げ%（コース別マスタ）× 会場イン逃げ率補正
      S2 ∝ 2-3号艇の差し% × rel_win1
      S3 ∝ 3-5号艇のまくり% × rel_win1
      S4 ∝ 5-6号艇の rel_win1（極小）

    条件付き2着確率（シナリオごとに異なる）：
      S1（イン逃げ時）: circle_pct（イン逃げ時2着率）を使用
      S2/S3/S4 時:      展開別残存マスタの実2着率を優先参照（マスタなしは内側残存補正テーブルで代替）
                        会場別マスタ（信頼度≥0.3）→ 全国マスタ → ハードコードテーブル の優先順位

    条件付き3着確率（シナリオごとに異なる）：
      S2/S3/S4 時:      展開別残存マスタの実3着以内率を優先参照（同上優先順位）

    引数追加（v6）：
      tenkai_national: dict  {(決まり手, 1着コース): row_dict}  全国展開別残存マスタ
      tenkai_venue:    dict  {(会場名, 決まり手, 1着コース): row_dict}  会場別展開別残存マスタ
      venue_stats:     dict  会場統計（差し率/まくり率/まくり差し率をS2〜S4重みへ反映）

    ─────────────────────────────────────────────────────────────
    戻り値: prob降順ソート済み 全120通りの combinationリスト
    """
    wakus = [r["waku"] for r in results]
    if len(wakus) < 3:
        return []

    # ── ルックアップテーブル構築 ──
    # 【改善】rel_win1=0 の艇にコース別フロア確率を適用（0%買い目根絶）
    # Laplace smoothingフロア: calc_race_indices で既に適用済みだが
    # 外部calc(_ext_calc_3rentan)経由や直接呼び出し時にも保護する。
    _RELWIN_FLOOR = {"1": 0.5, "2": 0.15, "3": 0.15, "4": 0.10, "5": 0.06, "6": 0.02}
    rel_map   = {r["waku"]: max(r.get("rel_win1") or 0,
                                _RELWIN_FLOOR.get(r["waku"], 0.02)) for r in results}
    # 【修正①】_circ_raw（正規化前絶対スコア）を参照して二重正規化を解消
    # circle_pct は表示用相対%（合計100%）のため確率計算では使用しない。
    circ_map  = {r["waku"]: max(
        r.get("_circ_raw") if r.get("_circ_raw") is not None
        else (r.get("circle_pct") or 0) / 100.0, 0
    ) for r in results}
    idx3_map  = {r["waku"]: max(r.get("idx3")  or r.get("rel_win1") or 0, 0) for r in results}
    win3_map  = {r["waku"]: r.get("win3_rate") or 0.5 for r in results}
    cm_map    = {r["waku"]: r.get("raw_cm", {}) for r in results}
    avg_st_map = {r["waku"]: r.get("avg_st") for r in results}

    def safe_pct(cm, key):
        v = safe_float(_get_cm_val(cm, key))
        return max(v or 0.0, 0.0)

    # ──────────────────────────────────────────────────────────────────────────
    # 【修正⑤】メンバー相互作用を反映したシナリオ重み計算
    # ──────────────────────────────────────────────────────────────────────────
    # 旧問題: 各艇の決まり手%を独立に合算 → 「誰が誰に対してどう動くか」が無視される
    #   例①) 2号艇の差し%が高くても、3〜4号艇のまくりが速ければS2は潰される
    #   例②) 1号艇のSTが遅くても S1重みは下がらなかった
    #   例③) 今節成績・モーター調子が確率計算に全く反映されていなかった
    #
    # 新方式: 以下3つの相互作用補正を追加する。
    #
    #   補正A ── STの相対関係によるS1重み調整
    #     1号艇と2号艇の平均STを比較し、2号艇が速い（先マイ状態）ならS1を減衰させる。
    #     ST差 ≥ +0.03秒（2号艇が速い）: S1 × 0.75
    #     ST差 ≤ -0.03秒（1号艇が速い）: S1 × 1.15（上限補正）
    #
    #   補正B ── 内側艇の差し力によるS3重み抑制（まくり潰し確率）
    #     2号艇の差し%が高いほど、外側艇のまくりは潰されやすい。
    #     まくり潰し係数 = 1.0 - 0.5 × (2号艇差し% / 全国平均差し率の2倍)
    #     ただし係数の下限は0.5（完全には潰れない）
    #
    #   補正C ── 今節成績・モーター調子によるシナリオ重み調整
    #     今節成績: "1-1-2"のような文字列を解析し、直近の1着数が多いほど
    #               その艇が関与するシナリオ重みを最大+20%上昇させる
    #     モーター2連率: 全艇平均と比較し、突出した艇はシナリオ関与重みを補正する
    #     ※ 補正Cは各シナリオの基礎重みに乗算するスケーラーとして適用
    # ──────────────────────────────────────────────────────────────────────────
    COURSE_NATIONAL_WIN = {"1": 0.555, "2": 0.137, "3": 0.134,
                           "4": 0.111, "5": 0.066, "6": 0.021}
    # 会場別1コース1着率があれば使用（荒れやすい会場の過信補正）
    # 例: 戸田=0.430、平和島=0.446 → 全国平均0.555より大幅に低い
    _c1_win_base = float(venue_course_1c_rate) if venue_course_1c_rate is not None else COURSE_NATIONAL_WIN["1"]

    # ── 補正C用: 今節成績スコアとモーター調子スコアを艇番ごとに算出 ──────────
    def _kosetsu_score(kosetsu_str):
        """
        今節成績文字列（例: "1-2-3" "1-1-F" "2-1-1"）から1着率相当のスコアを返す。
        形式は「着順-着順-着順...」を想定。1着=1.0、2着=0.5、3着=0.25、それ以外=0。
        データなし or 解析不能の場合は None を返す。
        """
        if not kosetsu_str or kosetsu_str in ("", "None", "nan", "-"):
            return None
        scores = []
        for token in re.split(r"[-・/]", kosetsu_str):
            token = token.strip()
            if token == "1":
                scores.append(1.0)
            elif token == "2":
                scores.append(0.5)
            elif token == "3":
                scores.append(0.25)
            elif token.isdigit():
                scores.append(0.0)
            # F/L/S/K等の失格・欠場は無視（スコアに含めない）
        return sum(scores) / len(scores) if scores else None

    def _kosetsu_course_match_score(kosetsu_str, today_course_str):
        """
        【②改善】今節成績のコース別一致度スコア。

        今節の各走が「今日と同じコース」で走ったものかどうかを判定し、
        一致走の着順を2倍重みで評価する。
        CSVに今節コース列（例: "1-2-3" 形式）がある場合のみ有効。
        列がない・データが空の場合は通常の _kosetsu_score と同じ結果を返す。

        例: 今日が3コース、今節成績 "1-3-2"、今節コース "3-4-3"
          → 1走目3C一致:1着(2倍) / 2走目4C不一致:3着(1倍) / 3走目3C一致:2着(2倍)
          → 加重平均でスコアUP

        設計思想:
          3コースで今節2連続まくり差しを決めている選手と
          3コースで2連続6着の選手を今節成績文字列だけでは区別できない問題を解消。
          同コース実績を重視することで「今日のコースへの適性」を正確に評価する。
        """
        if not kosetsu_str or kosetsu_str in ("", "None", "nan", "-"):
            return None

        order_tokens = [t.strip() for t in re.split(r"[-・/]", kosetsu_str)]

        # コース情報なし → 通常スコアにフォールバック
        if not today_course_str or today_course_str in ("", "None", "nan"):
            return _kosetsu_score(kosetsu_str)

        try:
            today_course = int(str(today_course_str).strip())
        except (ValueError, TypeError):
            return _kosetsu_score(kosetsu_str)

        # 今節コース列（CSVに "今節コース" 列があれば使用）
        # 形式例: "3-4-3" = 1走目3C, 2走目4C, 3走目3C
        # 列がなければ全走を一致なしとして通常重みで計算
        kosetsu_course_key = None  # build時に渡される予定（現状はNone）

        scores = []
        weights = []
        rank_score_map = {"1": 1.0, "2": 0.5, "3": 0.25}

        for i, token in enumerate(order_tokens):
            token = token.strip()
            rank_score = rank_score_map.get(token, 0.0 if token.isdigit() else None)
            if rank_score is None:
                continue  # F/L/S/K等はスキップ

            # 今節コース列があれば一致判定（現状は全走1倍で計算、列実装後に2倍に拡張）
            weight = 1.0
            scores.append(rank_score)
            weights.append(weight)

        if not scores:
            return None

        total_weight = sum(weights)
        return sum(s * w for s, w in zip(scores, weights)) / total_weight if total_weight > 0 else None

    kosetsu_score_map = {}
    for r in results:
        today_course = str(r.get("course", r.get("waku", "")))
        sc = _kosetsu_course_match_score(str(r.get("kosetsu", "")), today_course)
        kosetsu_score_map[r["waku"]] = sc  # None の場合は補正しない

    motor2_map = {}
    for r in results:
        v = safe_float(r.get("motor2"))
        motor2_map[r["waku"]] = v  # None の場合は補正しない

    # モーター2連率の全艇平均（Noneを除く）
    valid_motor2 = [v for v in motor2_map.values() if v is not None]
    motor2_mean = sum(valid_motor2) / len(valid_motor2) if valid_motor2 else None

    # ── 【v6.2】全選手指数マップを事前構築 ──────────────────────────────────
    # 選手指数マスタ（pm）から各艇の指数を取得してマップ化
    # ここで構築した指数が _member_scenario_scale に全て反映される
    _form_map   = {}   # フォーム指数（0〜16.69、中央値≈2.1）
    _recent3_map= {}   # 直近3走1着率（0〜1.0）
    _recent5_map= {}   # 直近5走1着率（0〜1.0）
    _st_std_map = {}   # ST標準偏差（小さいほど安定）
    _st_stab_map= {}   # ST安定スコア（大きいほど安定、7〜85）
    _jizai_map  = {}   # 自在性加重1着率（外枠実力）
    _ippan_map  = {}   # 一般戦1着率（一般戦専門の実力）
    _recent10_map={}   # 直近10走平均着順（小さいほど好調）

    for r in results:
        pm_r = r.get("raw_pm") or r.get("pm") or {}
        # フォーム指数
        v = safe_float(pm_r.get("フォーム\n指数") or pm_r.get("フォーム指数"))
        _form_map[r["waku"]] = v

        # 直近3走1着率
        v = safe_float(pm_r.get("直近3走\n1着率") or pm_r.get("直近3走1着率"))
        _recent3_map[r["waku"]] = v

        # 直近5走1着率
        v = safe_float(pm_r.get("直近5走\n1着率") or pm_r.get("直近5走1着率"))
        _recent5_map[r["waku"]] = v

        # ST標準偏差
        v = safe_float(pm_r.get("ST\n標準偏差") or pm_r.get("ST標準偏差"))
        _st_std_map[r["waku"]] = v

        # ST安定スコア
        v = safe_float(pm_r.get("ST安定\nスコア") or pm_r.get("ST安定スコア"))
        _st_stab_map[r["waku"]] = v

        # 自在性加重1着率
        v = safe_float(pm_r.get("自在性\n加重1着率") or pm_r.get("自在性加重1着率"))
        _jizai_map[r["waku"]] = v

        # 一般戦1着率
        v = safe_float(pm_r.get("1着率\n(一般戦)") or pm_r.get("1着率(一般戦)"))
        _ippan_map[r["waku"]] = v

        # 直近10走平均着順
        v = safe_float(pm_r.get("直近10走\n平均着順") or pm_r.get("直近10走平均着順"))
        _recent10_map[r["waku"]] = v

        # ── 【v6.2追加】コース別マスタの最速ST・最遅ST からSTレンジを算出 ──
        # STレンジ（最遅-最速）が小さいほど発艇が安定している
        # raw_cm（コース別マスタ）から取得
        cm_r = r.get("raw_cm") or {}
        _st_max = safe_float(cm_r.get("最遅ST") or cm_r.get("最遅\nST"))
        _st_min = safe_float(cm_r.get("最速ST") or cm_r.get("最速\nST"))
        if _st_max is not None and _st_min is not None:
            r["_st_range"] = round(_st_max - _st_min, 4)
        else:
            r["_st_range"] = None

    # STレンジマップ（0〜0.8程度、小さいほど安定）
    _st_range_map = {r["waku"]: r.get("_st_range") for r in results}
    # STレンジの全艇有効値で平均を計算（比較用）
    _valid_st_ranges = [v for v in _st_range_map.values() if v is not None]
    _st_range_mean = sum(_valid_st_ranges) / len(_valid_st_ranges) if _valid_st_ranges else 0.30

    # ── 【v6.4新設】★STフラグマップ ──────────────────────────────────────────
    # update_master.py: base["★ST"] = base["出走数"] < THRESH_ST (=10)
    # Excelには "★"（文字列）または "" で書き込まれる。
    # ★STが立っている艇はST標準偏差・STレンジの計算サンプルが10未満であり
    # 平均値の信頼性が低いため、ST系の補正をスキップする。
    _star_st_map = {}
    for r in results:
        cm_r = r.get("raw_cm") or {}
        val = cm_r.get("★ST")
        # Trueの場合: Excelから読んだ"★"文字列 or update_master側のbool True
        _star_st_map[r["waku"]] = bool(val and str(val).strip() in ("★", "True", "1"))

    # 各艇のpm参照（_member_scenario_scaleのフォールバック用にresultsからも取得試行）
    for r in results:
        if _form_map.get(r["waku"]) is None:
            # raw_pmにない場合、resultsのpm_*フィールドからも取得を試みる
            v = safe_float(r.get("form_index") or r.get("フォーム指数"))
            if v is not None:
                _form_map[r["waku"]] = v
        if _recent3_map.get(r["waku"]) is None:
            v = safe_float(r.get("recent3_win") or r.get("直近3走1着率"))
            if v is not None:
                _recent3_map[r["waku"]] = v
        if _recent5_map.get(r["waku"]) is None:
            v = safe_float(r.get("recent5_win") or r.get("直近5走1着率"))
            if v is not None:
                _recent5_map[r["waku"]] = v

    def _member_scenario_scale(waku, base_kosetsu_weight=0.18, base_motor_weight=0.08):
        """
        【v6.2 全指数統合版】
        その艇が関与するシナリオへの乗算スケーラーを返す。
        スケーラー = 1.0 ± 補正値（範囲: 0.75 〜 1.25）

        使用指数と重み配分（合計補正上限±0.25）:
          今節成績        (0.12): 今節の着順文字列から算出
          モーター2連率   (0.08): 機力の相対評価
          フォーム指数    (0.05): 直近調子の総合指標（中央値2.1基準）
          直近3走1着率    (0.04): 超短期フォーム（3走で評価が安定しやすい）
          直近5走1着率    (0.03): 短期フォーム（3走と5走の差で上昇/下降を見る）
          ST標準偏差      (0.025): STばらつき小→スタートが安定→シナリオ実現度高
          STレンジ        (0.015): 最速〜最遅のレンジ小→コース内での安定度
          ST安定スコア    (0.02): ST品質の総合評価（平均60基準）
          自在性加重1着率 (0.03): 外枠からの攻め実力（S2〜S4シナリオに特に重要）
          一般戦1着率     (0.02): 格付け補正（特別戦選手の一般戦での実力）
          直近10走平均着順(0.02): 中期トレンド（3.5以下=好調、4.5超=不調）
        """
        scale = 1.0

        # ── ① 今節成績（既存） ────────────────────────────────────────────
        ks = kosetsu_score_map.get(waku)
        if ks is not None:
            scale += (ks - 0.5) * 2 * base_kosetsu_weight

        # ── ② モーター2連率（既存） ──────────────────────────────────────
        m2 = motor2_map.get(waku)
        if m2 is not None and motor2_mean is not None and motor2_mean > 0:
            ratio = (m2 - motor2_mean) / motor2_mean
            scale += ratio * base_motor_weight

        # ── ③ フォーム指数 ────────────────────────────────────────────────
        form = _form_map.get(waku)
        if form is not None:
            form_norm = max(-1.0, min(1.0, (form / 3.0) - 1.0))
            scale += form_norm * 0.05

        # ── ④ 直近3走1着率 ───────────────────────────────────────────────
        r3 = _recent3_map.get(waku)
        if r3 is not None:
            scale += max(-1.0, min(1.0, (r3 - 0.17) / 0.17)) * 0.04

        # ── ⑤ 直近5走1着率 ───────────────────────────────────────────────
        r5 = _recent5_map.get(waku)
        if r5 is not None:
            scale += max(-1.0, min(1.0, (r5 - 0.17) / 0.17)) * 0.03

        # ── ⑥ ST標準偏差（小さいほど安定）──────────────────────────────
        # ★STフラグ = サンプル10未満 → ST値が不安定なためST系補正をスキップ
        _st_unreliable = _star_st_map.get(waku, False)
        st_std = _st_std_map.get(waku)
        if st_std is not None and not _st_unreliable:
            scale += max(-1.0, min(1.0, (0.071 - st_std) / 0.071)) * 0.025

        # ── ⑦ STレンジ（最速〜最遅、小さいほど安定）【v6.2追加】────────
        st_range = _st_range_map.get(waku)
        if st_range is not None and _st_range_mean > 0 and not _st_unreliable:
            # レンジが平均より小さいほど正（安定）、大きいほど負（不安定）
            scale += max(-1.0, min(1.0, (_st_range_mean - st_range) / _st_range_mean)) * 0.015

        # ── ⑧ ST安定スコア ───────────────────────────────────────────────
        st_stab = _st_stab_map.get(waku)
        if st_stab is not None and not _st_unreliable:
            scale += max(-1.0, min(1.0, (st_stab - 60.0) / 30.0)) * 0.02

        # ── ⑨ 自在性加重1着率 ────────────────────────────────────────────
        jizai = _jizai_map.get(waku)
        if jizai is not None:
            scale += max(-1.0, min(1.0, (jizai - 0.06) / 0.06)) * 0.03

        # ── ⑩ 一般戦1着率 ────────────────────────────────────────────────
        ippan = _ippan_map.get(waku)
        if ippan is not None:
            scale += max(-1.0, min(1.0, (ippan - 0.17) / 0.17)) * 0.02

        # ── ⑪ 直近10走平均着順 ───────────────────────────────────────────
        r10 = _recent10_map.get(waku)
        if r10 is not None:
            scale += max(-1.0, min(1.0, (3.5 - r10) / 1.5)) * 0.02

        return max(0.75, min(1.25, scale))

    # ── 補正A: STの相対関係によるS1重み調整係数 ──────────────────────────────
    st1 = avg_st_map.get("1")
    st2 = avg_st_map.get("2")
    st_adj_s1 = 1.0  # デフォルト: 補正なし
    if st1 is not None and st2 is not None:
        st_diff = st1 - st2  # 正 → 1号艇が遅い、負 → 1号艇が速い
        if st_diff >= 0.03:
            # 2号艇の方が速い（先マイになりやすい）→ S1を減衰
            st_adj_s1 = max(0.75, 1.0 - st_diff * 5.0)
        elif st_diff <= -0.03:
            # 1号艇の方が速い（楽逃げ）→ S1を増強（上限1.20）
            st_adj_s1 = min(1.20, 1.0 + abs(st_diff) * 3.0)

    # ── 【修正⑤】S1基礎重み: rel_win1が既に会場特性を織り込み済みのため二重補正を除去 ──
    # 【旧問題】s1_base = _c1_win_base × (0.5 + w1_nige) としていたが、
    #   rel_win1 の計算時点で既に venue_rate（会場別1コース1着率）が w_venue 分混入している。
    #   そこへさらに _c1_win_base を掛けると「会場特性が二重に効く」構造になっていた。
    #   特に荒れやすい会場（戸田・平和島等）では rel_win1 が既に低めになっているのに
    #   _c1_win_base でさらに下げてしまい、1号艇を過剰に不利評価していた。
    #
    # 【修正③】S1重みの二重補正を解消: nige_adj を乗算からブレンド補正に変更。
    #   旧方式: s1 = rel_win1 × (0.5 + nige%) × ST補正 — 逃げ%が rel_win1 に既に含まれている
    #   新方式: s1 = nige_blend × ST補正 × member補正
    #     nige_blend = 個人逃げ% × 0.6 + 会場1コース1着率 × 0.4
    #     → 個人実績と会場特性を直接ブレンドし、rel_win1 による二重乗算を排除。
    #     → 逃げ%データなし時は会場1着率100%にフォールバック。
    w1_nige      = safe_pct(cm_map.get("1", {}), "逃げ%")
    if w1_nige > 0:
        nige_blend = w1_nige * 0.6 + _c1_win_base * 0.4
    else:
        nige_blend = _c1_win_base  # データなし → 会場特性のみ
    s1_weight = nige_blend * st_adj_s1 * _member_scenario_scale("1")

    # ── 【v6.2新設】S1重みをRNo別1C1着率で補正 ──────────────────────────────
    # 同じ会場でもレース番号で1コース1着率が大きく異なる
    # （例: 大村1R=0.494 vs 大村12R=0.747）
    # race_judgmentに格納済みの venue_1c_race_rate を使用
    _venue_1c_race = safe_float((race_judgment or {}).get("venue_1c_race_rate")) if race_judgment else None
    if _venue_1c_race is not None and _c1_win_base and _c1_win_base > 0:
        # RNo別1着率 / 会場平均1着率 で補正係数を算出（0.75〜1.25 にクリップ）
        _race_s1_adj = max(0.75, min(1.25, _venue_1c_race / _c1_win_base))
        s1_weight *= _race_s1_adj

    # ── 【修正②】事前評価（jizen_eval）による追加補正 ────────────────────────
    # evaluate_jizen.evaluate_all() の結果（◎○△）をシナリオ重みに反映する。
    # ①イン逃げ評価: S1の補正（◎→+15%、○→+7%、空白→-10%）
    # ②相性評価:     S2/S3 の按分を相性◎の艇を優遇
    # 補正幅は ±15%以内に抑えて過剰な振れを防ぐ。
    _jizen_symbol_s1 = ""
    _jizen_aisho = {}  # {waku_str: symbol}
    if jizen_eval is not None:
        try:
            _jizen_symbol_s1 = jizen_eval.get("in_nige", [""])[0]  # 1号艇のイン逃げ評価
            for idx, sym in enumerate(jizen_eval.get("aisho", [])):
                _jizen_aisho[str(idx + 1)] = sym
        except Exception:
            pass

    _s1_jizen_adj = {"◎": 1.15, "◎?": 1.10, "○": 1.07, "△": 1.0, "": 0.90}.get(
        _jizen_symbol_s1, 1.0
    )
    s1_weight *= _s1_jizen_adj

    def _jizen_aisho_scale(waku):
        """相性評価記号をシナリオ重み乗数に変換（S2/S3の1着按分で使用）"""
        sym = _jizen_aisho.get(str(waku), "")
        return {"◎": 1.15, "○": 1.07, "△": 1.0, "": 0.95}.get(sym, 1.0)

    # ── 補正B: 2号艇の差し力によるS3抑制係数（まくり潰し） ─────────────────────
    sashi_pct_2   = safe_pct(cm_map.get("2", {}), "差し%")
    nat_sashi_avg = COURSE_NATIONAL_WIN["2"] * 0.5   # 全国平均の差し率代替値（約0.07）
    if sashi_pct_2 > 0:
        makuri_suppress = max(0.50, 1.0 - 0.5 * (sashi_pct_2 / (nat_sashi_avg * 2 + 1e-6)))
    else:
        makuri_suppress = 1.0  # 差し%データなし → 補正しない

    # ── S2基礎重み: 差し（全艇対象・コース別距離補正付き）─────────────────────
    # 旧: 2・3号艇限定 → 4〜6コースからの差しを完全無視していた問題を解消
    # 新: 全艇の差し%を対象に、コースが外になるほど距離補正で減衰させる
    #   1号艇: 差し%は逃げ側なので除外（差しシナリオで1着は物理的にほぼない）
    #   2号艇: 補正1.0（内側差しは最も決まりやすい）
    #   3号艇: 補正1.0
    #   4号艇: 補正0.7（外側差しは距離が伸びて決まりにくい）
    #   5号艇: 補正0.5
    #   6号艇: 補正0.3
    _S2_OUTER_SCALE = {"2": 1.0, "3": 1.0, "4": 0.7, "5": 0.5, "6": 0.3}
    s2_weight = 0.0
    for w, outer_scale in _S2_OUTER_SCALE.items():
        pct_sashi = safe_pct(cm_map.get(w, {}), "差し%")
        nat_avg   = COURSE_NATIONAL_WIN.get(w, 0.1) * 0.3
        base      = pct_sashi if pct_sashi > 0 else nat_avg
        s2_weight += base * outer_scale * _member_scenario_scale(w) * _jizen_aisho_scale(w)

    # ── S3基礎重み: まくり（全艇対象・まくり%のみ）────────────────────────────
    # まくり差しとまくりを分離: まくりは外から全艇を押し込む展開
    # → 内側艇が圧縮される（S2より内側残存が小さい）
    _S3_OUTER_SCALE = {"3": 1.0, "4": 1.0, "5": 0.8, "6": 0.5}
    s3_weight = 0.0
    for w, outer_scale in _S3_OUTER_SCALE.items():
        pct_mak = safe_pct(cm_map.get(w, {}), "まくり%")   # まくり%のみ（まくり差し除外）
        nat_avg = COURSE_NATIONAL_WIN.get(w, 0.05) * 0.4
        base    = pct_mak if pct_mak > 0 else nat_avg
        s3_weight += base * outer_scale * _member_scenario_scale(w) * _jizen_aisho_scale(w)
    s3_weight *= makuri_suppress  # 補正B: 2号艇の差し力でまくりを抑制

    # ── S4基礎重み: まくり差し（全艇対象・まくり差し%のみ）新設 ─────────────
    # まくり差しは外から入って内側に切り込む展開
    # → まくりより内側が残りやすい（差しとまくりの中間）
    _S4_OUTER_SCALE = {"3": 1.0, "4": 1.0, "5": 0.8, "6": 0.5}
    s4_weight = 0.0
    for w, outer_scale in _S4_OUTER_SCALE.items():
        pct_maksa = safe_pct(cm_map.get(w, {}), "まくり差し%")   # まくり差し%のみ
        nat_avg   = COURSE_NATIONAL_WIN.get(w, 0.05) * 0.3
        base      = pct_maksa if pct_maksa > 0 else nat_avg
        s4_weight += base * outer_scale * _member_scenario_scale(w) * _jizen_aisho_scale(w)
    s4_weight *= makuri_suppress  # まくり差しもまくり潰しの影響を受ける

    # ── 【v6新設】🟡 会場統計の決まり手率をS2〜S4重みに反映 ─────────────────────
    # 「この会場はまくりが多い」という傾向をS3/S4重みに乗算補正する。
    # venue_stats が渡されていない場合はスキップ（補正なし）。
    # 全国平均: 差し≒0.155 / まくり≒0.150 / まくり差し≒0.095（boatrace.jp 全国統計）
    _NATIONAL_AVG_SASHI    = 0.155
    _NATIONAL_AVG_MAKURI   = 0.150
    _NATIONAL_AVG_MAKUSA   = 0.095
    if venue_stats:
        _kimari = venue_stats.get("kimari_avg") or {}
        # 差し補正 → S2重み
        _vs_sashi = safe_float(_kimari.get("差し"))
        if _vs_sashi is not None and _NATIONAL_AVG_SASHI > 0:
            _s2_venue_boost = max(0.70, min(1.30, _vs_sashi / _NATIONAL_AVG_SASHI))
            s2_weight *= _s2_venue_boost
        # まくり補正 → S3重み
        _vs_makuri = safe_float(_kimari.get("まくり"))
        if _vs_makuri is not None and _NATIONAL_AVG_MAKURI > 0:
            _s3_venue_boost = max(0.70, min(1.30, _vs_makuri / _NATIONAL_AVG_MAKURI))
            s3_weight *= _s3_venue_boost
        # まくり差し補正 → S4重み
        _vs_makusa = safe_float(_kimari.get("まくり差し"))
        if _vs_makusa is not None and _NATIONAL_AVG_MAKUSA > 0:
            _s4_venue_boost = max(0.70, min(1.30, _vs_makusa / _NATIONAL_AVG_MAKUSA))
            s4_weight *= _s4_venue_boost

    # ── 【v6.2新設】会場コース別1着率（2C〜6C）でS2〜S4シナリオ重みを補強 ──────
    # 決まり手率補正だけでは「どのコースから飛んでくるか」が反映できない。
    # 会場統計の実コース別1着率を使って「この会場では外枠が有力か」を補正する。
    # 全国平均: 2C=0.137 / 3C=0.134 / 4C=0.111 / 5C=0.066 / 6C=0.021
    _NATIONAL_C_WIN = {"2": 0.137, "3": 0.134, "4": 0.111, "5": 0.066, "6": 0.021}
    _venue_cw = (race_judgment or {}).get("venue_course_win_rates") or {} if race_judgment else {}
    if _venue_cw:
        # S2（差し）: 主に2〜3コースが担う → 2C・3Cの会場実績で補正
        _s2_c_adj = 1.0
        for _c in ["2", "3"]:
            _vr = safe_float(_venue_cw.get(_c))
            _nr = _NATIONAL_C_WIN.get(_c, 0.13)
            if _vr is not None and _nr > 0:
                _s2_c_adj *= max(0.80, min(1.20, _vr / _nr)) ** 0.5  # 平方根で緩和
        s2_weight *= _s2_c_adj

        # S3（まくり）: 主に3〜5コースが担う → 3C・4C・5Cの会場実績で補正
        _s3_c_adj = 1.0
        for _c in ["3", "4", "5"]:
            _vr = safe_float(_venue_cw.get(_c))
            _nr = _NATIONAL_C_WIN.get(_c, 0.10)
            if _vr is not None and _nr > 0:
                _s3_c_adj *= max(0.80, min(1.20, _vr / _nr)) ** 0.5
        s3_weight *= _s3_c_adj

        # S4（まくり差し）: 主に3〜6コースが担う → 4C・5C・6Cの会場実績で補正
        _s4_c_adj = 1.0
        for _c in ["4", "5", "6"]:
            _vr = safe_float(_venue_cw.get(_c))
            _nr = _NATIONAL_C_WIN.get(_c, 0.066)
            if _vr is not None and _nr > 0:
                _s4_c_adj *= max(0.80, min(1.20, _vr / _nr)) ** 0.5
        s4_weight *= _s4_c_adj

    # ── ❷ main_type による S2〜S4 重みの直結補正（STEP1→確率計算の断絶を解消）──
    # _judge_main_player が算出した main_waku（主役候補）と main_type（差し/まくり/まくり差し）を
    # 対応するシナリオの重みに直接乗算して STEP1 の定性判断を確率に反映する。
    #
    # 【④改善】主軸強度（mc_strength）と主役スコア（main_score）の両方が高い場合に
    # 上限を1.3倍→2.0倍に引き上げ。
    # 条件: mc_strength >= 60 かつ _main_score >= 0.55
    #   → 「展開が読めているレース」で確信を持って点数を絞る根拠になる。
    # 条件未達の場合は従来通り最大+30%。
    # mc_strength は _build_conflict_map が算出する主軸対立の強度（0〜100程度）。
    _mp_data    = (race_judgment or {}).get("main_player", {}) or {}
    _main_waku  = _mp_data.get("main_waku")
    _main_type  = _mp_data.get("main_type", "")
    _main_score = float(_mp_data.get("main_score", 0) or 0)
    _mc_strength = float((race_judgment or {}).get("conflict_map", {}).get("main_conflict", {}).get("strength", 0) or 0) \
        if race_judgment else 0.0

    # 条件付き上限引き上げ
    if _mc_strength >= 60 and _main_score >= 0.55:
        # 展開が明確に読めているレース → 最大+50%
        _main_boost = 1.0 + min(_main_score * 0.9, 0.50)
    else:
        # 通常ケース → 最大+30%（従来通り）
        _main_boost = 1.0 + min(_main_score * 0.6, 0.30)

    if _main_waku and _main_waku != "1":
        if _main_type == "差し":
            s2_weight *= _main_boost
        elif _main_type == "まくり":
            s3_weight *= _main_boost
        elif _main_type == "まくり差し":
            s4_weight *= _main_boost

    # ── SCシナリオ重み（S4「大荒れ」を「潰れ展開・漁夫の利」SCに差し替え） ────────
    # 旧S4は「何が起きるかわからない雑な確率」だった。
    # SCは「飛び役が自滅したとき誰が漁夫の利を取るか」という具体的な展開を表す。
    # 確率の合計は変わらず、展開の解像度だけ上がる。
    _sc_info = _calc_sc_weight(
        results, cm_map, win3_map, rel_map, jizen_eval=jizen_eval
    )
    sc_weight         = _sc_info["sc_weight"]
    _sc_1st_weights   = _sc_info["sc_1st_weights"]    # {waku: float} SC時1着按分
    _sc_beneficiary   = _sc_info["sc_beneficiary"]    # {waku: float} 漁夫スコア
    _sc_fly_type      = _sc_info["sc_fly_type"]       # "まくり系"/"差し系"/"不明"

    # ── 【⑧改善】複数シナリオ同時成立の相関補正 ─────────────────────────────
    # 現在S1〜S4+SCは排他的シナリオとして扱っているが、
    # 実際には「3号艇まくりが決まりながら2号艇が差して2着に入る」複合展開が多い。
    #
    # 対処: シナリオ間相関として「S3成立時にS2的2着が起きる確率」を
    # S3重みとS2重みの幾何平均で表現し、S3_S2_CORR重みとして別途管理する。
    # これはシナリオを増やすのではなく「S3の2着分布を差し方向に引っ張る」補正。
    #
    # 相関係数: s2_weight と s3_weight の幾何平均 × 相関強度
    # 条件: 2号艇の差し%が全艇平均より高い場合のみ発動（強い差し屋がいる場合）
    _sashi_pct_2 = safe_pct(cm_map.get("2", {}), "差し%")
    _sashi_avg   = sum(safe_pct(cm_map.get(w, {}), "差し%") or 0 for w in ["2","3","4","5","6"]) / 5
    _s3s2_corr_weight = 0.0
    if _sashi_pct_2 > _sashi_avg * 1.2 and s3_weight > 0 and s2_weight > 0:
        # 2号艇が平均より20%以上差しが強い場合: S3展開でも2号艇2着の可能性を追加
        import math as _math
        _s3s2_corr_weight = _math.sqrt(s3_weight * s2_weight) * 0.25
        # 相関重みをS3から借りる（S3重みを薄める）ことで総重みは変わらない
        s3_weight = max(0.0, s3_weight - _s3s2_corr_weight * 0.5)
        s2_weight = max(0.0, s2_weight - _s3s2_corr_weight * 0.5)

    total_s = s1_weight + s2_weight + s3_weight + s4_weight + sc_weight + _s3s2_corr_weight
    if total_s <= 0:
        total_s = 1.0
    p_s1 = s1_weight / total_s
    p_s2 = s2_weight / total_s
    p_s3 = s3_weight / total_s
    p_s4 = s4_weight / total_s
    p_sc = sc_weight / total_s
    p_s3s2 = _s3s2_corr_weight / total_s  # まくり+差し複合シナリオ確率

    # ── 条件付き1着確率（シナリオ × 1着艇）──
    # S2/S3は決まり手%が主軸
    # SC は _calc_sc_weight が算出した1着按分重みを使用
    def prob_first_given_scenario(scenario, first_w):
        """P(1着=first_w | scenario)"""
        if scenario == "S1":
            # イン逃げシナリオ → 1号艇固定
            return 1.0 if first_w == "1" else 0.0
        elif scenario == "S2":
            # 差しシナリオ → 全艇の差し%で按分（コース別距離補正付き）
            _s2_scale = {"2": 1.0, "3": 1.0, "4": 0.7, "5": 0.5, "6": 0.3}
            candidates = {}
            for w, sc in _s2_scale.items():
                pct = safe_pct(cm_map.get(w, {}), "差し%")
                candidates[w] = (pct if pct > 0 else COURSE_NATIONAL_WIN.get(w, 0.1) * 0.3) * sc
            total = sum(candidates.values()) or 1
            return candidates.get(first_w, 0) / total
        elif scenario == "S3":
            # まくりシナリオ → まくり%のみで按分（まくり差し除外）
            outer_scale = {"3": 1.0, "4": 1.0, "5": 0.8, "6": 0.5}
            candidates = {}
            for w, sc in outer_scale.items():
                pct = safe_pct(cm_map.get(w, {}), "まくり%")   # まくり%のみ
                candidates[w] = (pct if pct > 0 else COURSE_NATIONAL_WIN.get(w, 0.04) * 0.4) * sc
            total = sum(candidates.values()) or 1
            return candidates.get(first_w, 0) / total
        elif scenario == "S4":
            # まくり差しシナリオ → まくり差し%のみで按分
            outer_scale = {"3": 1.0, "4": 1.0, "5": 0.8, "6": 0.5}
            candidates = {}
            for w, sc in outer_scale.items():
                pct = safe_pct(cm_map.get(w, {}), "まくり差し%")   # まくり差し%のみ
                candidates[w] = (pct if pct > 0 else COURSE_NATIONAL_WIN.get(w, 0.04) * 0.3) * sc
            total = sum(candidates.values()) or 1
            return candidates.get(first_w, 0) / total
        else:  # SC（潰れ展開）
            # 飛び役自滅後の1着は _calc_sc_weight が算出した按分重みで決まる
            total = sum(_sc_1st_weights.values()) or 1
            return _sc_1st_weights.get(first_w, 0) / total

    # ── 条件付き2着確率 P(2着=B | 1着=A, scenario) ──
    # ── 【v6新設】展開別残存マスタから残存補正係数を取得するヘルパー ──────────────
    # 戻り値: {コース文字列: {"2着率": float, "3着以内率": float}} または None（データなし）
    #
    # シナリオ→決まり手 対応テーブル（全シナリオ対応・v6.1拡張）
    # S4 で first_w=2 のデータが存在しない（物理的にまくり差し2着はほぼない）場合は
    # まくり/2 のデータをフォールバックとして使用。
    _SCENARIO_TO_KIMETE = {
        "S1": "逃げ",        # イン逃げ: 逃げ/1着コースのデータを使用
        "S2": "差し",
        "S3": "まくり",
        "S4": "まくり差し",
        "SC": "恵まれ",      # 潰れ展開: 恵まれ/抜きのデータを使用（漁夫の利）
    }
    # S4でデータなし時のフォールバック決まり手（まくり差しデータがないコースはまくりで代替）
    _S4_FALLBACK_KIMETE = "まくり"

    def _get_tenkai_rates(scenario, first_w):
        """
        展開別残存マスタから、1着コース=first_w のシナリオにおける
        各進入コースの 2着率・3着以内率を返す。

        戻り値: {進入コース文字列: {"2着率": float, "3着以内率": float}} または None

        キー構造（v6）:
          会場別: (会場名, 決まり手, 1着コース, 進入コース) → row
          全国版: (決まり手, 1着コース, 進入コース)         → row

        シナリオ対応（v6.1拡張）:
          S1 → 決まり手=逃げ, 1着コース=1 固定（イン逃げ）
          S2 → 決まり手=差し
          S3 → 決まり手=まくり
          S4 → 決まり手=まくり差し（データなし時はまくりで代替）
          SC → 決まり手=恵まれ + 抜き をブレンド（潰れ展開の漁夫の利）
        """
        _venue = (race_judgment or {}).get("venue") if race_judgment else None

        # S1: 1着コースは常に1固定（逃げは1コースから）
        if scenario == "S1":
            kimete = "逃げ"
            actual_course = "1"
        else:
            kimete = _SCENARIO_TO_KIMETE.get(scenario)
            if not kimete:
                return None
            r_data = next((r for r in results if r["waku"] == first_w), {})
            actual_course = str(int(float(r_data.get("course") or r_data.get("進入コース") or first_w)))

        def _fetch(kimete_k, course_k, c_str):
            """(決まり手, 1着コース, 進入コース) の行データを会場別→全国の優先順で取得

            【修正②】信頼度閾値を0.3→0.15に緩和し、低信頼度の会場別データも
            全国マスタとブレンドして活用する（旧実装: 0.3未満は完全に捨てていた）

            ブレンド方式:
              信頼度 >= 0.50 → 会場別100%（十分な実績あり）
              信頼度 >= 0.15 → 会場別 × 信頼度/0.50 + 全国 × (1 - 信頼度/0.50)
                               例: 信頼度0.30 → 会場60% / 全国40%
                               例: 信頼度0.15 → 会場30% / 全国70%
              信頼度 < 0.15  → 全国マスタのみ（会場データが少なすぎて誤学習リスク大）
            """
            if c_str == course_k:
                return None  # 1着艇自身は除外

            row_v    = None
            trust_v  = 0.0
            if tenkai_venue and _venue:
                key_v = (str(_venue), kimete_k, course_k, c_str)
                _rv = tenkai_venue.get(key_v)
                if _rv:
                    try:
                        trust_v = float(_rv.get("信頼度") or 0)
                        if trust_v >= 0.15:
                            row_v = _rv
                    except (ValueError, TypeError):
                        pass

            row_n = None
            if tenkai_national:
                key_n = (kimete_k, course_k, c_str)
                row_n = tenkai_national.get(key_n)

            # 会場別データなし → 全国マスタのみ返す（従来どおり）
            if row_v is None:
                return row_n

            # 会場別データあり・信頼度 >= 0.50 → 会場別データをそのまま返す
            if trust_v >= 0.50:
                return row_v

            # 会場別データあり・0.15 <= 信頼度 < 0.50 → 全国マスタとブレンド
            if row_n is None:
                # 全国マスタがなければ会場別をそのまま返す
                return row_v
            try:
                w_v = trust_v / 0.50          # 0.30 → 0.60 / 0.15 → 0.30
                w_n = 1.0 - w_v
                r2_blended = (
                    float(row_v.get("2着率")     or 0) * w_v +
                    float(row_n.get("2着率")     or 0) * w_n
                )
                r3_blended = (
                    float(row_v.get("3着以内率") or 0) * w_v +
                    float(row_n.get("3着以内率") or 0) * w_n
                )
                # 元の行を壊さないよう新dictで返す
                return {"2着率": r2_blended, "3着以内率": r3_blended, "_blended": True}
            except (ValueError, TypeError):
                return row_v

        # SC（潰れ展開）: 自滅タイプに応じて参照マスタを切り替える
        # 【⑥改善】
        #   まくり系自滅 → 外に膨らんで失速 → 内側が抜けてくる「抜き」パターン優先
        #   差し系自滅  → 蓋をされて失速   → 外側が流れてくる「恵まれ」パターン優先
        #   従来は恵まれ+抜きを単純平均していたが、自滅タイプを無視すると
        #   「まくり自滅なのに外側が2着に来る確率が高く出る」という誤りが生じていた。
        if scenario == "SC":
            # _sc_fly_typeは外側スコープから参照（_calc_3rentan_probs_v2内で定義済み）
            _fly_type = _sc_fly_type if "_sc_fly_type" in dir() else "不明"
            # 自滅タイプ別の重み設定
            if _fly_type == "まくり系":
                # まくり系自滅: 内側が抜けやすい → 抜き70% + 恵まれ30%
                _w_em, _w_nu = 0.30, 0.70
            elif _fly_type == "差し系":
                # 差し系自滅: 外側が恵まれやすい → 恵まれ70% + 抜き30%
                _w_em, _w_nu = 0.70, 0.30
            else:
                # 不明: 従来通り均等ブレンド
                _w_em, _w_nu = 0.50, 0.50

            result = {}
            for c_str in [str(i) for i in range(1, 7)]:
                if c_str == actual_course:
                    continue
                row_em = _fetch("恵まれ", actual_course, c_str)
                row_nu = _fetch("抜き",   actual_course, c_str)

                # 両方ある場合: 自滅タイプ別重みでブレンド
                if row_em is not None and row_nu is not None:
                    try:
                        r2 = (float(row_em.get("2着率")     or 0) * _w_em +
                              float(row_nu.get("2着率")     or 0) * _w_nu)
                        r3 = (float(row_em.get("3着以内率") or 0) * _w_em +
                              float(row_nu.get("3着以内率") or 0) * _w_nu)
                        result[c_str] = {"2着率": r2, "3着以内率": r3}
                    except (ValueError, TypeError):
                        pass
                # 片方しかない場合: あるほうを使用
                elif row_em is not None:
                    try:
                        result[c_str] = {
                            "2着率":     float(row_em.get("2着率")     or 0),
                            "3着以内率": float(row_em.get("3着以内率") or 0),
                        }
                    except (ValueError, TypeError):
                        pass
                elif row_nu is not None:
                    try:
                        result[c_str] = {
                            "2着率":     float(row_nu.get("2着率")     or 0),
                            "3着以内率": float(row_nu.get("3着以内率") or 0),
                        }
                    except (ValueError, TypeError):
                        pass
            return result if result else None

        # S4でまくり差し/2着コースのデータがない → まくり/2 でフォールバック
        result = {}
        fallback_needed = False
        for c_str in [str(i) for i in range(1, 7)]:
            if c_str == actual_course:
                continue
            row = _fetch(kimete, actual_course, c_str)
            if row is None and scenario == "S4":
                # まくり差しデータなし → まくりで代替
                row = _fetch(_S4_FALLBACK_KIMETE, actual_course, c_str)
                if row:
                    fallback_needed = True
            if row is None:
                continue
            try:
                r2 = float(row.get("2着率")     or 0)
                r3 = float(row.get("3着以内率") or 0)
                result[c_str] = {"2着率": r2, "3着以内率": r3}
            except (ValueError, TypeError):
                pass

        return result if result else None

    # ══════════════════════════════════════════════════════════════════════
    # 個人能力スコア計算（2着・3着残存に使用）
    # ══════════════════════════════════════════════════════════════════════
    # 「展開的に残れるポジション × その艇が今日実際に残れる実力」を統合する。
    #
    # 使用指数:
    #   win3_rate     : 3連対率（コース別マスタ）   ← 主軸
    #   avg_st        : コース別平均ST              ← ST能力
    #   motor2        : モーター2連率               ← 機力
    #   fly_label     : FLY影響                    ← リスク（減衰）
    #   late_count    : 出遅れ数                   ← リスク（減衰）
    #   kosetsu       : 今節成績                   ← 短期コンディション
    #   フォーム指数  : raw_pm                     ← 中期コンディション
    #
    # 重み設計（合計1.0）:
    #   win3_rate     0.40  (3連対率が主軸)
    #   avg_st        0.20  (STが速い艇ほど展開に付いてこれる)
    #   motor2        0.12  (機力が高い艇は伸び足で粘れる)
    #   kosetsu       0.10  (今節成績で直近コンディション反映)
    #   フォーム指数  0.08  (中期調子)
    #   FLY/出遅れ    -補正  (リスクペナルティ)
    # ══════════════════════════════════════════════════════════════════════

    # 各艇のraw_pmマップを構築（prob_second/prob_third から参照）
    _pm_map = {r["waku"]: r.get("raw_pm", {}) or {} for r in results}

    # ST平均（全艇）・モーター平均（全艇）を事前計算（相対評価用）
    _st_vals_all  = [r["avg_st"] for r in results if r.get("avg_st") is not None]
    _st_mean_all  = sum(_st_vals_all) / len(_st_vals_all) if _st_vals_all else 0.17
    _st_min_all   = min(_st_vals_all) if _st_vals_all else 0.12
    _st_max_all   = max(_st_vals_all) if _st_vals_all else 0.22

    _m2_vals_all  = []
    for r in results:
        try:
            v = float(r.get("motor2") or 0)
            if v > 0:
                _m2_vals_all.append(v)
        except (ValueError, TypeError):
            pass
    _m2_mean_all = sum(_m2_vals_all) / len(_m2_vals_all) if _m2_vals_all else 40.0

    def _kosetsu_score_local(kosetsu_str):
        """今節成績文字列 → スコア（0〜1）"""
        if not kosetsu_str or str(kosetsu_str) in ("", "None", "nan", "-"):
            return None
        scores = []
        for token in re.split(r"[-・/]", str(kosetsu_str)):
            token = token.strip()
            if token == "1":   scores.append(1.0)
            elif token == "2": scores.append(0.5)
            elif token == "3": scores.append(0.25)
            elif token.isdigit(): scores.append(0.0)
        return sum(scores) / len(scores) if scores else None

    def _personal_ability_score(waku):
        """
        2着・3着残存計算用の個人能力統合スコア（0〜1）を返す。
        マスタ係数（位置の物理的残存傾向）に掛け合わせて使用する。
        """
        r = next((x for x in results if x["waku"] == waku), {})
        pm = _pm_map.get(waku, {})
        score = 0.0

        # ① win3_rate（3連対率）: 0.40
        w3 = win3_map.get(waku, 0.5)
        score += min(w3, 1.0) * 0.40

        # ② avg_st（ST能力）: 0.20
        # STが速い（小さい）艇ほど展開に追走できる → 高スコア
        st = r.get("avg_st")
        if st is not None and (_st_max_all - _st_min_all) > 0.001:
            st_score = 1.0 - (st - _st_min_all) / (_st_max_all - _st_min_all)
            score += max(0.0, min(1.0, st_score)) * 0.20
        else:
            score += 0.5 * 0.20  # データなし → 中立

        # ③ motor2（機力）: 0.12
        try:
            m2 = float(r.get("motor2") or 0)
            if m2 > 0 and _m2_mean_all > 0:
                m2_score = min(m2 / _m2_mean_all, 2.0) / 2.0
                score += m2_score * 0.12
            else:
                score += 0.5 * 0.12
        except (ValueError, TypeError):
            score += 0.5 * 0.12

        # ④ 今節成績（kosetsu）: 0.10
        ks = _kosetsu_score_local(r.get("kosetsu", ""))
        if ks is not None:
            score += min(ks / 0.5, 1.0) * 0.10  # 0.5(平均)で正規化
        else:
            score += 0.5 * 0.10

        # ⑤ フォーム指数（raw_pm）: 0.08
        form = safe_float(pm.get("フォーム\n指数") or pm.get("フォーム指数"))
        if form is not None:
            form_score = max(0.0, min(1.0, form / 6.0))
            score += form_score * 0.08
        else:
            score += 0.5 * 0.08

        # ⑥ FLY/出遅れリスクペナルティ
        fly_label = r.get("fly_label", "低")
        fly_pen = {"高": -0.18, "中": -0.08, "低": 0.0}.get(fly_label, 0.0)
        score += fly_pen

        late_count = r.get("late_count", 0) or 0
        if late_count >= 3:
            score -= 0.06
        elif late_count >= 1:
            score -= 0.02

        return max(0.10, min(1.0, score))  # 下限0.10（完全ゼロを防ぐ）

    # 全艇の個人能力スコアを事前計算
    _pers_map = {r["waku"]: _personal_ability_score(r["waku"]) for r in results}

    def prob_second(scenario, first_w, second_w, remaining):
        if scenario == "S1":
            # イン逃げシナリオ2着: マスタ逃げ/1着の2着率で補強（会場別→全国優先）
            # マスタなし時は従来の circle_pct（イン逃げ時2着優位度）を使用
            _master_s1 = _get_tenkai_rates("S1", first_w)
            if _master_s1:
                _r2_vals = [v["2着率"] for v in _master_s1.values() if v["2着率"] > 0]
                _avg2 = sum(_r2_vals) / len(_r2_vals) if _r2_vals else 0
                if _avg2 > 0.001:
                    # マスタ2着率 × circ_map（選手個人の有利度）をブレンド
                    # マスタ: コース位置の物理的な残存確率（全国実績）
                    # circ_map: 選手個人の1号艇との相性・イン逃げ時2着実績
                    # → 0.6:0.4 のブレンドで両方を活かす
                    w2 = {}
                    for w in remaining:
                        master_coef = max(0.5, min(2.0, _master_s1.get(w, {}).get("2着率", _avg2) / _avg2))
                        circ_val    = max(circ_map.get(w, 0), 0.001)
                        w2[w] = master_coef * 0.6 + circ_val * 0.4
                    total = sum(w2.values()) or 1
                    return w2.get(second_w, 0.001) / total
            # マスタなし → 従来の circle_pct のみ
            circ_rem = {w: max(circ_map.get(w, 0), 0.001) for w in remaining}
            total = sum(circ_rem.values()) or 1
            return circ_rem.get(second_w, 0.001) / total

        elif scenario == "SC":
            # 潰れ展開の2着 → 恵まれ/抜きのマスタ実データで sc_beneficiary を補強
            # 自滅した主要飛び役は2着にも残りにくい（確率を0.2に減衰）
            fly_waku = _sc_info.get("sc_fly_waku")
            _master_sc = _get_tenkai_rates("SC", first_w)
            w2 = {}
            for w in remaining:
                score = max(_sc_beneficiary.get(w, 0.1), 0.001)
                if _master_sc:
                    # マスタ実データとsc_beneficiaryをブレンド（0.5:0.5）
                    m_r2 = _master_sc.get(w, {}).get("2着率", 0)
                    if m_r2 > 0:
                        score = score * 0.5 + m_r2 * 0.5
                if w == fly_waku:
                    score *= 0.2
                w2[w] = max(score, 0.001)
            total = sum(w2.values()) or 1
            return w2.get(second_w, 0.001) / total
        else:
            # S2差し・S3まくり・S4まくり差しで内側残存補正テーブルを使い分け
            #
            # 【v6改善】展開別残存マスタの実2着率を優先参照する。
            # マスタが参照できた場合: 2着率を相対係数に変換して rel_map へ乗算
            # マスタなし（データ不足）: 従来のハードコードテーブルをそのまま使用
            #
            # S2（差し）: 差した艇より内側が残りやすい
            # S3（まくり）: 外から全艇を押し込む → 内側が圧縮
            # S4（まくり差し）: まくりより内側が残りやすい
            # 【v6.6】ハードコード → 展開別残存_全国マスタの実測値ベースに置き換え
            # 旧テーブルは「内側が残りやすい」という思い込みで設計されており、
            # 実測値と大幅に乖離（最大+1.83）していた。
            # 算出方法: 各(決まり手×1着コース)の2着率を平均で割り max(0.5, min(2.0, 値)) でクリップ
            # データ: 展開別残存_全国シート（信頼度1、S2:106〜3486件、S3:258〜2311件、S4:363〜2138件）
            # ※ マスタ上書きロジック（_get_tenkai_rates）は引き続き有効。
            #   会場別マスタが信頼度≥0.3なら会場別値で再上書きされる。
            #   このテーブルはマスタにデータがない組み合わせのフォールバックとして機能する。
            _INNER_S2 = {   # 差し（全国実測ベース）
                # 差しの物理: 差した艇より内側が残りやすい
                # 差し艇が内側を切って前に出るため、差し艇より外側の艇は置き去り
                "2": {"1": 2.000, "3": 0.771, "4": 0.606, "5": 0.500, "6": 0.500},
                "3": {"1": 1.131, "2": 1.637, "4": 1.076, "5": 0.775, "6": 0.500},
                # 【修正】4号差し: 旧値はS3まくりと完全一致（誤り）→ 差しの物理法則に修正
                # 差した4号より内側（1〜3号）が後続に残りやすい
                # 外側の5・6号は4号に弾かれて残りにくい
                "4": {"1": 1.400, "2": 1.200, "3": 1.100, "5": 0.650, "6": 0.550},
                "5": {"1": 1.356, "2": 1.384, "3": 1.158, "4": 0.791, "6": 0.500},
                "6": {"1": 0.849, "2": 1.226, "3": 1.226, "4": 1.085, "5": 0.613},
            }
            _INNER_S3 = {   # まくり（全国実測ベース）
                # まくりの物理: まくり艇が外から押し込む
                # 先マイに近い外側艇（まくり艇の直後）が2着に残りやすい
                # 内側（3号より内）は圧縮されて残りにくい
                # 【追加】2号まくりは稀だが発生する → S2差しテーブルをベースに外側有利に補正
                "2": {"1": 1.000, "3": 0.700, "4": 1.100, "5": 1.200, "6": 0.800},
                "3": {"1": 1.111, "2": 0.964, "4": 1.465, "5": 1.044, "6": 0.500},
                "4": {"1": 1.090, "2": 0.945, "3": 0.504, "5": 1.741, "6": 0.718},
                "5": {"1": 1.593, "2": 1.049, "3": 0.593, "4": 0.528, "6": 1.236},
                "6": {"1": 1.764, "2": 1.415, "3": 0.698, "4": 0.756, "5": 0.500},
            }
            _INNER_S4 = {   # まくり差し（全国実測ベース）
                # まくり差しの物理: まくりと差しの中間
                # 差しより外側が残りやすく、まくりより内側が残りやすい傾向
                # 【追加】2号まくり差しはS2差しに近いパターン → 内側有利
                "2": {"1": 1.800, "3": 0.900, "4": 0.750, "5": 0.600, "6": 0.500},
                "3": {"1": 2.000, "2": 0.839, "4": 0.563, "5": 0.500, "6": 0.500},
                "4": {"1": 2.000, "2": 0.755, "3": 1.227, "5": 0.755, "6": 0.500},
                "5": {"1": 1.954, "2": 0.805, "3": 0.725, "4": 1.137, "6": 0.500},
                "6": {"1": 1.970, "2": 0.730, "3": 0.661, "4": 0.813, "5": 0.827},
            }
            if scenario == "S2":
                pos_tbl = _INNER_S2.get(first_w, {})
            elif scenario == "S3":
                pos_tbl = _INNER_S3.get(first_w, {})
            else:  # S4
                pos_tbl = _INNER_S4.get(first_w, {})

            # 【v6】展開別残存マスタから実2着率を取得し、相対係数に変換して pos_tbl を上書き
            _master_rates2 = _get_tenkai_rates(scenario, first_w)
            if _master_rates2:
                _r2_vals = [v["2着率"] for v in _master_rates2.values() if v["2着率"] > 0]
                _avg2 = sum(_r2_vals) / len(_r2_vals) if _r2_vals else 0
                if _avg2 > 0.001:
                    # マスタ相対係数: 0.5〜2.0にクリップ
                    _master_pos_tbl = {
                        c: max(0.5, min(2.0, v["2着率"] / _avg2))
                        for c, v in _master_rates2.items()
                    }
                    # マスタ値がある艇はマスタ優先、ない艇はハードコードをそのまま
                    pos_tbl = {**pos_tbl, **_master_pos_tbl}

            w2 = {}
            for w in remaining:
                # 【修正v2】win3_rate → 個人能力統合スコア（_pers_map）に変更
                # win3_rate だけでなく ST・機力・フォーム・FLYリスクを統合した
                # 「今日この艇が実際に2着に残れる実力」を使う
                base = max(_pers_map.get(w, 0.3), 0.001)
                base *= pos_tbl.get(w, 1.0)
                w2[w] = base
            total = sum(w2.values()) or 1
            return w2.get(second_w, 0.001) / total

    # ── 条件付き3着確率 P(3着=C | 1着=A, 2着=B, scenario) ──
    #
    # S1（イン逃げ）:   idx3（イン逃げ時3着指数）を使用
    # S2（差し）:       win3_rate(0.6) + rel_win1(0.4) の加重平均
    # S3（まくり系）:   内側艇は win3_rate で下方補正（弾かれやすい）
    # SC（潰れ展開）:   漁夫スコア（_sc_beneficiary）を使用
    #                   位置ボーナス × 攻撃性逆数 × win3_rate の合成値
    #
    # 【修正①】2着艇(second_w)が3着分布に与える影響を反映
    # 旧問題: P(3着=C | A, B) の計算が second_w に完全に依存しない
    #         → 3-2-? と 3-5-? で同じ pos_tbl3 を使っていた
    #
    # 新方式: second_w の「進路封鎖効果」を3着スコアに乗算補正する。
    #
    #   封鎖効果の物理法則:
    #     S2（差し）: 2着に入った差し艇がC艇の進路を塞ぐ
    #       → 2着艇よりコースが外側かつ差し系の艇 → 進路を塞がれる（スコア減衰）
    #       → 2着艇よりコースが内側の艇 → 先に走っており影響なし
    #     S3（まくり）: まくり艇が外側に膨らんで2着に入る
    #       → 2着艇と同じく外側で展開していた艇 → 共倒れリスク（スコア減衰）
    #       → 内側艇 → まくり艇が抜けたコースに残る（スコア維持〜上昇）
    #     S4（まくり差し）: S2とS3の中間的な封鎖効果
    #
    #   封鎖係数の計算:
    #     try_int(w) > try_int(second_w) かつ w と second_w が同じ展開系 → 減衰
    #     それ以外 → 影響なし（係数1.0）
    #   補正幅: 0.70〜1.0 に抑えて過剰な振れを防ぐ

    def _second_block_factor(scenario, second_w, third_w):
        """
        2着艇(second_w)が3着候補(third_w)の進路を封鎖する係数を返す。
        封鎖あり → 0.70〜0.90 / 封鎖なし → 1.0
        """
        try:
            s_int = int(second_w)
            t_int = int(third_w)
        except (ValueError, TypeError):
            return 1.0

        if scenario == "S2":
            # 差しシナリオ: 2着差し艇より外側の艇が差しに来ようとすると進路封鎖
            # → 外側で差しを試みる艇のスコアを減衰
            if t_int > s_int:
                sashi_t = safe_pct(cm_map.get(third_w, {}), "差し%")
                # 差し%が高いほど封鎖影響を受ける（自分も差しに行こうとする）
                block_strength = min(sashi_t / 30.0, 1.0)   # 差し30%で最大
                return max(0.70, 1.0 - 0.30 * block_strength)
        elif scenario == "S3":
            # まくりシナリオ: 2着まくり艇の外側で同様にまくりを仕掛けた艇は共倒れリスク
            if t_int > s_int:
                mak_t = safe_pct(cm_map.get(third_w, {}), "まくり%")
                block_strength = min(mak_t / 25.0, 1.0)
                return max(0.75, 1.0 - 0.25 * block_strength)
            # 内側艇: まくり艇が掃いたコースに残る → 微増（最大+10%）
            if t_int < s_int and t_int > 1:
                return min(1.10, 1.0 + 0.10 * (s_int - t_int) / 5.0)
        elif scenario == "S4":
            # まくり差しシナリオ: S2とS3の中間
            if t_int > s_int:
                maksa_t = (safe_pct(cm_map.get(third_w, {}), "まくり差し%") +
                           safe_pct(cm_map.get(third_w, {}), "差し%")) / 2.0
                block_strength = min(maksa_t / 28.0, 1.0)
                return max(0.72, 1.0 - 0.28 * block_strength)
        return 1.0

    def prob_third(scenario, first_w, second_w, third_w, remaining):
        if scenario == "S1":
            # イン逃げシナリオ3着: マスタ逃げ/1着の3着以内率で idx3_map を補強
            # マスタなし時は従来の idx3_map を使用
            # S1は先頭から逃げ・差し・流れ込みの順で展開が決まるため
            # 2着艇の封鎖効果は小さく適用しない（idx3が既に内側有利を反映済み）
            _master_s1_3rd = _get_tenkai_rates("S1", first_w)
            if _master_s1_3rd:
                _r3_vals = [v["3着以内率"] for v in _master_s1_3rd.values() if v["3着以内率"] > 0]
                _avg3 = sum(_r3_vals) / len(_r3_vals) if _r3_vals else 0
                if _avg3 > 0.001:
                    w3 = {}
                    for w in remaining:
                        master_coef = max(0.5, min(2.0,
                            _master_s1_3rd.get(w, {}).get("3着以内率", _avg3) / _avg3))
                        idx3_val = max(idx3_map.get(w, 0), 0.001)
                        # 【修正v2】マスタ(0.5) × idx3(0.3) × 個人能力(0.2) の3軸ブレンド
                        # 個人能力（ST・機力・フォーム）を追加して「今日の実力」を反映
                        pers_val = max(_pers_map.get(w, 0.3), 0.001)
                        w3[w] = master_coef * 0.50 + idx3_val * 0.30 + pers_val * 0.20
                    total = sum(w3.values()) or 1
                    return w3.get(third_w, 0.001) / total
            # マスタなし → 従来の idx3_map のみ
            w3 = {w: max(idx3_map.get(w, 0), 0.001) for w in remaining}

        elif scenario == "S2":
            # S2（差し系）: 3着も内側残存補正を適用
            # 【v6】展開別残存マスタの実3着以内率を優先参照
            # 【v6.6】3着以内率も全国実測値ベースに置き換え
            _INNER_REMAIN3 = {   # 差し（全国実測ベース・3着以内率）
                "2": {"1": 1.992, "3": 1.056, "4": 0.909, "5": 0.643, "6": 0.500},
                "3": {"1": 1.053, "2": 1.310, "4": 1.105, "5": 0.930, "6": 0.602},
                "4": {"1": 1.520, "2": 1.275, "3": 1.080, "5": 0.691, "6": 0.500},
                "5": {"1": 1.271, "2": 1.200, "3": 1.045, "4": 0.989, "6": 0.500},
                "6": {"1": 0.943, "2": 1.226, "3": 1.061, "4": 0.896, "5": 0.873},
            }
            pos_tbl3 = _INNER_REMAIN3.get(first_w, {})
            # マスタ3着以内率で上書き
            _master_rates3s2 = _get_tenkai_rates("S2", first_w)
            if _master_rates3s2:
                _r3_vals = [v["3着以内率"] for v in _master_rates3s2.values() if v["3着以内率"] > 0]
                _avg3 = sum(_r3_vals) / len(_r3_vals) if _r3_vals else 0
                if _avg3 > 0.001:
                    _master_pos3 = {
                        c: max(0.5, min(2.0, v["3着以内率"] / _avg3))
                        for c, v in _master_rates3s2.items()
                    }
                    pos_tbl3 = {**pos_tbl3, **_master_pos3}
            w3 = {}
            for w in remaining:
                # 【修正v2】win3_rate → 個人能力統合スコア（_pers_map）に変更
                pers  = max(_pers_map.get(w, 0.3), 0.001)
                rel   = max(rel_map.get(w, 0),     0.001)
                # 個人能力(0.7) + 相対1着率(0.3) のブレンド
                # 3着には「残れる実力」と「位置的有利」の両方が必要
                base  = pers * 0.70 + rel * 0.30
                base *= pos_tbl3.get(w, 1.0)
                # 【修正①】2着艇による進路封鎖効果を適用
                base *= _second_block_factor("S2", second_w, w)
                w3[w] = base

        elif scenario in ("S3", "S4"):
            # S3（まくり）・S4（まくり差し）: シナリオ別内側残存補正
            # 【v6】展開別残存マスタの実3着以内率を優先参照
            # 【v6.6】3着以内率も全国実測値ベースに置き換え
            _INNER3_S3 = {   # まくり（全国実測ベース・3着以内率）
                # 【追加】2号まくり: まくり後の3着は内側1号と外側3〜4号が残りやすい
                "2": {"1": 1.200, "3": 1.100, "4": 1.050, "5": 0.900, "6": 0.750},
                "3": {"1": 1.057, "2": 1.015, "4": 1.260, "5": 1.058, "6": 0.609},
                "4": {"1": 1.183, "2": 0.956, "3": 0.600, "5": 1.385, "6": 0.875},
                "5": {"1": 1.427, "2": 1.126, "3": 0.675, "4": 0.634, "6": 1.138},
                "6": {"1": 1.502, "2": 1.318, "3": 0.814, "4": 0.814, "5": 0.552},
            }
            _INNER3_S4 = {   # まくり差し（全国実測ベース・3着以内率）
                # 【追加】2号まくり差し: 差しに近いパターン → 内側1号残り強い
                "2": {"1": 1.800, "3": 0.950, "4": 0.800, "5": 0.650, "6": 0.500},
                "3": {"1": 1.923, "2": 0.870, "4": 1.016, "5": 0.765, "6": 0.500},
                "4": {"1": 1.384, "2": 0.804, "3": 1.027, "5": 1.174, "6": 0.612},
                "5": {"1": 1.483, "2": 0.926, "3": 0.839, "4": 1.009, "6": 0.741},
                "6": {"1": 1.575, "2": 0.970, "3": 0.722, "4": 0.901, "5": 0.832},
            }
            pos_tbl3 = (_INNER3_S3 if scenario == "S3" else _INNER3_S4).get(first_w, {})
            # マスタ3着以内率で上書き
            _master_rates3 = _get_tenkai_rates(scenario, first_w)
            if _master_rates3:
                _r3_vals = [v["3着以内率"] for v in _master_rates3.values() if v["3着以内率"] > 0]
                _avg3 = sum(_r3_vals) / len(_r3_vals) if _r3_vals else 0
                if _avg3 > 0.001:
                    _master_pos3 = {
                        c: max(0.5, min(2.0, v["3着以内率"] / _avg3))
                        for c, v in _master_rates3.items()
                    }
                    pos_tbl3 = {**pos_tbl3, **_master_pos3}
            w3 = {}
            for w in remaining:
                # 【修正v2】win3_rate → 個人能力統合スコア（_pers_map）に変更
                base = max(_pers_map.get(w, 0.3), 0.001)
                base *= pos_tbl3.get(w, 1.0)
                # 【修正①】2着艇による進路封鎖効果を適用
                base *= _second_block_factor(scenario, second_w, w)
                w3[w] = base

        else:  # SC（潰れ展開）
            # 漁夫スコア + 恵まれ/抜きのマスタ3着以内率でブレンド
            # SCシナリオは展開が複雑なため封鎖効果は適用しない（sc_beneficiaryに内包）
            _master_sc_3rd = _get_tenkai_rates("SC", first_w)
            w3 = {}
            for w in remaining:
                score = max(_sc_beneficiary.get(w, 0.1), 0.001)
                if _master_sc_3rd:
                    m_r3 = _master_sc_3rd.get(w, {}).get("3着以内率", 0)
                    if m_r3 > 0:
                        score = score * 0.5 + m_r3 * 0.5
                w3[w] = max(score, 0.001)

        total = sum(w3.values()) or 1
        return w3.get(third_w, 0.001) / total

    # ── 全120通りの確率を計算 ──
    combos_dict = {}
    scenarios = [("S1", p_s1), ("S2", p_s2), ("S3", p_s3), ("S4", p_s4), ("SC", p_sc)]

    for first in wakus:
        for second in wakus:
            if second == first:
                continue
            for third in wakus:
                if third == first or third == second:
                    continue
                key = f"{first}-{second}-{third}"
                prob_total = 0.0
                for scenario, p_sc in scenarios:
                    if p_sc < 1e-9:
                        continue
                    p1_sc = prob_first_given_scenario(scenario, first)
                    if p1_sc < 1e-9:
                        continue
                    rem2 = [w for w in wakus if w != first]
                    p2_sc = prob_second(scenario, first, second, rem2)
                    rem3 = [w for w in wakus if w != first and w != second]
                    p3_sc = prob_third(scenario, first, second, third, rem3)
                    prob_total += p_sc * p1_sc * p2_sc * p3_sc

                # 【⑧】S3S2複合シナリオ: まくりで3〜6号艇が1着、2号艇が差して2着に入る
                # 1着: S3（まくり）条件付き確率、2着: 2号艇を差し方向で優遇
                if p_s3s2 > 1e-9:
                    p1_s3s2 = prob_first_given_scenario("S3", first)
                    if p1_s3s2 > 1e-9:
                        rem2 = [w for w in wakus if w != first]
                        # 2着は差し方向の確率を使う（S2の2着分布を参照）
                        p2_s3s2 = prob_second("S2", first, second, rem2)
                        rem3 = [w for w in wakus if w != first and w != second]
                        p3_s3s2 = prob_third("S2", first, second, third, rem3)
                        prob_total += p_s3s2 * p1_s3s2 * p2_s3s2 * p3_s3s2

                combos_dict[key] = {
                    "combo":          key,
                    "first":          first,
                    "second":         second,
                    "third":          third,
                    "prob":           prob_total,
                    "is_outer_first": int(first) >= 4,
                    "top_scenario":   max(
                        scenarios,
                        key=lambda sc: sc[1] * prob_first_given_scenario(sc[0], first)
                    )[0],
                    # SCシナリオ情報（数値シート表示用）
                    "sc_fly_type":    _sc_fly_type,
                    "sc_beneficiary": _sc_beneficiary.get(second, 0),
                }

    # ── 理論オッズ・ハイブリッドスコアを付与 ──
    # 理論オッズ = (1 / prob) × (1 - テラ銭0.25)
    # ハイブリッドスコア = 確率60% + 理論オッズ40% の合成ランク
    # → 低オッズ本命への過集中を防ぎ、中〜高配当の期待値が高い組み合わせを優先
    TAKER_RATE  = 0.25
    ODDS_WEIGHT = 0.4  # 0.0=純確率順 / 1.0=純オッズ順 / 0.4=バランス重視

    for c in combos_dict.values():
        p = max(c["prob"], 1e-6)
        c["theoretical_odds"] = round((1.0 / p) * (1.0 - TAKER_RATE), 1)
        c["prob_rank"]  = 0
        c["odds_rank"]  = 0

    combos_sorted_prob = sorted(combos_dict.values(), key=lambda x: x["prob"],             reverse=True)
    combos_sorted_odds = sorted(combos_dict.values(), key=lambda x: x["theoretical_odds"], reverse=True)

    for i, c in enumerate(combos_sorted_prob): c["prob_rank"]  = i + 1
    for i, c in enumerate(combos_sorted_odds): c["odds_rank"]  = i + 1

    for c in combos_dict.values():
        c["hybrid_score"] = (1.0 - ODDS_WEIGHT) * (1.0 / c["prob_rank"]) + \
                             ODDS_WEIGHT         * (1.0 / c["odds_rank"])

    combos = sorted(combos_dict.values(), key=lambda x: x["hybrid_score"], reverse=True)
    return combos


# ─────────────────────────────────────────────────────────────────────────────
# 期待値計算（実際のオッズExcelとの照合用）
# 使い方:
#   combos = _calc_3rentan_probs_v2(results)
#   actual = load_actual_odds_from_excel("odds_YYYYMMDD_R1.xlsx")  # 別途実装
#   ev_list = calc_ev_from_actual_odds(combos, actual)
#   suggestion = suggest_by_ev(ev_list, min_ev=0.05)  # EV+5%以上のみ
# ─────────────────────────────────────────────────────────────────────────────

def calc_ev_from_actual_odds(combos, actual_odds_dict):
    """
    実際の払戻オッズと推定確率から期待値（EV）を計算する。

    Parameters
    ----------
    combos : list[dict]
        _calc_3rentan_probs_v2() の戻り値。各要素に "combo"（例: "1-2-3"）と "prob" が必要。
    actual_odds_dict : dict
        {"1→2→3": 払戻倍率float, ...} 形式。
        ※ 理論オッズExcelから読んだ全120通りの実際の払戻倍率を渡す。

    Returns
    -------
    list[dict]
        各コンボに以下を追加した辞書のリスト（EV降順）:
          actual_odds : 実際の払戻倍率（Noneなら未取得）
          ev          : 期待値 = actual_odds × prob - 1.0（Noneなら計算不可）
          ev_pct      : EVを%表示した文字列（例: "+12.3%"）
          ev_positive : EVがプラスかどうか
    """
    result = []
    for c in combos:
        key    = c["combo"]   # "1-2-3" 形式
        actual = actual_odds_dict.get(key)
        c2     = dict(c)
        if actual is not None and c.get("prob", 0) > 0:
            ev = actual * c["prob"] - 1.0
            c2["actual_odds"] = round(actual, 1)
            c2["ev"]          = round(ev, 4)
            c2["ev_pct"]      = f"{'+' if ev >= 0 else ''}{ev*100:.1f}%"
            c2["ev_positive"] = ev > 0
        else:
            c2["actual_odds"] = None
            c2["ev"]          = None
            c2["ev_pct"]      = "N/A"
            c2["ev_positive"] = False
        result.append(c2)
    # EV降順（未取得はそれ以下に）
    result.sort(key=lambda x: (x["ev"] if x["ev"] is not None else -999), reverse=True)
    return result


def suggest_by_ev(combos_with_ev, min_ev=0.0, max_bets=8):
    """
    期待値プラスの組み合わせのみを買い目として返す。

    Parameters
    ----------
    combos_with_ev : list[dict]  calc_ev_from_actual_odds() の戻り値
    min_ev         : float       最低期待値閾値（デフォルト0.0 = 期待値プラスのみ）
                                 例: 0.05 → EV+5%以上のみ対象
    max_bets       : int         最大買い目数

    Returns
    -------
    dict
        buy_list    : 買い目リスト（EV降順）、例: ["2→3→1", "1→3→2"]
        ev_summary  : 各買い目の詳細（combo / prob% / actual_odds / ev%）
        total_bets  : 点数
        best_ev     : 最高EVの組み合わせ
        skip        : True なら期待値プラスがゼロ → 見送り推奨
        reason      : 判定理由の説明文
    """
    positives = [c for c in combos_with_ev
                 if c.get("ev") is not None and c["ev"] > min_ev]
    positives = positives[:max_bets]

    if not positives:
        # EV最高値を探して理由を付ける
        best_any = next((c for c in combos_with_ev if c.get("ev") is not None), None)
        best_str = f"（最高EV: {best_any['ev_pct']} {best_any['combo']}）" if best_any else ""
        return {
            "buy_list":   [],
            "ev_summary": [],
            "total_bets": 0,
            "best_ev":    None,
            "skip":       True,
            "reason":     f"EV>{min_ev*100:.0f}%の組み合わせなし → 見送り推奨{best_str}"
        }

    return {
        "buy_list": [c["combo"] for c in positives],
        "ev_summary": [
            {
                "combo":        c["combo"],
                "prob_pct":     f"{c['prob']*100:.2f}%",
                "actual_odds":  c["actual_odds"],
                "ev_pct":       c["ev_pct"],
            }
            for c in positives
        ],
        "total_bets": len(positives),
        "best_ev":    positives[0],
        "skip":       False,
        "reason":     f"EV>{min_ev*100:.0f}%が{len(positives)}点（最高: {positives[0]['ev_pct']} {positives[0]['combo']}）"
    }


def load_actual_odds_from_excel(filepath, sheet_name=0):
    """
    理論オッズExcel（全120通り）を読み込んで辞書形式で返す。

    Excel形式（理想オッズ完成版.xlsx 準拠）:
      - 列A: 1着艇番（int）
      - 列B: 2着艇番（int）
      - 列C: 3着艇番（int）
      - 列D: オッズ（float）
      - 1行目はヘッダ（読み飛ばし）
      - データ120行（3連単全通り）

    Parameters
    ----------
    filepath   : str  Excelファイルパス（例: "odds/理想オッズ_20240101_R1.xlsx"）
    sheet_name : int or str  シート番号またはシート名（デフォルト0）

    Returns
    -------
    dict  {"1→2→3": float, ...}  キーは "A→B→C" 形式
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    result = {}
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if row[0] is None:
            continue
        try:
            key = f"{int(row[0])}-{int(row[1])}-{int(row[2])}"
            result[key] = float(row[3])
        except (TypeError, ValueError):
            continue
    wb.close()
    return result


# ============================================================
# ============================================================
# 参加グレードテーブル（BT 8,458R 会場×決まり手別 ROI 実績）
# ============================================================
_ENTRY_GRADE_TABLE = {
    ("三国","まくり差し"):("S",5.90,0.270,12170,37),("鳴門","まくり差し"):("S",3.91,0.360,7220,25),
    ("福岡","まくり差し"):("S",3.37,0.364,8590,11),("戸田","まくり差し"):("S",3.28,0.351,3710,37),
    ("びわこ","まくり差し"):("S",3.24,0.361,5480,83),("江戸川","まくり差し"):("S",3.18,0.303,10750,33),
    ("住之江","抜き"):("S",2.95,0.278,10590,18),("若松","まくり差し"):("S",2.92,0.312,5670,32),
    ("唐津","まくり差し"):("S",2.89,0.302,4080,43),("浜名湖","まくり差し"):("S",2.39,0.212,3990,52),
    ("徳山","抜き"):("S",2.38,0.243,7090,37),("唐津","差し"):("S",2.27,0.208,2470,48),
    ("平和島","まくり差し"):("S",2.16,0.356,3690,45),("蒲郡","差し"):("S",2.13,0.258,2510,31),
    ("芦屋","まくり差し"):("S",2.11,0.192,8720,26),("尼崎","抜き"):("S",1.90,0.231,7580,26),
    ("徳山","まくり"):("S",1.83,0.133,9770,45),("びわこ","差し"):("S",1.81,0.222,2420,54),
    ("蒲郡","まくり差し"):("S",1.77,0.393,3290,56),("常滑","まくり"):("S",1.75,0.136,12630,44),
    ("若松","差し"):("S",1.70,0.321,4970,28),("びわこ","抜き"):("S",1.67,0.154,20070,26),
    ("桐生","差し"):("S",1.65,0.159,5820,44),("尼崎","まくり差し"):("S",1.64,0.261,4510,46),
    ("鳴門","抜き"):("S",1.63,0.360,3820,25),("多摩川","抜き"):("S",1.57,0.200,6610,15),
    ("浜名湖","差し"):("S",1.50,0.291,3420,55),
    ("桐生","まくり"):("A",1.46,0.203,4750,69),("三国","差し"):("A",1.46,0.137,4100,51),
    ("津","まくり差し"):("A",1.44,0.333,3470,30),("下関","まくり"):("A",1.42,0.163,8150,49),
    ("常滑","差し"):("A",1.42,0.235,4410,34),("鳴門","まくり"):("A",1.42,0.179,5510,56),
    ("児島","まくり"):("A",1.40,0.156,5120,64),("下関","まくり差し"):("A",1.39,0.348,3740,46),
    ("下関","抜き"):("A",1.38,0.280,3030,25),("浜名湖","まくり"):("A",1.34,0.175,3830,57),
    ("浜名湖","抜き"):("A",1.30,0.346,4430,26),("江戸川","まくり"):("A",1.30,0.244,3600,86),
    ("戸田","抜き"):("A",1.27,0.318,4310,22),("宮島","まくり差し"):("A",1.26,0.238,3630,42),
    ("住之江","まくり差し"):("A",1.21,0.265,4630,34),
    ("津","差し"):("B",1.20,0.212,3190,33),("平和島","まくり"):("B",1.17,0.195,4910,82),
    ("丸亀","差し"):("B",1.15,0.302,3680,43),("若松","まくり"):("B",1.13,0.235,3680,34),
    ("児島","まくり差し"):("B",1.12,0.256,3800,78),("桐生","抜き"):("B",1.12,0.208,2520,24),
    ("芦屋","まくり"):("B",1.08,0.200,4840,30),("戸田","まくり"):("B",1.04,0.153,3150,98),
    ("戸田","差し"):("B",1.02,0.226,4940,62),("丸亀","まくり"):("B",1.02,0.228,2840,57),
    ("桐生","まくり差し"):("C",0.97,0.243,2610,37),("芦屋","差し"):("C",0.96,0.273,4500,22),
    ("丸亀","まくり差し"):("C",0.96,0.204,3970,54),("大村","まくり差し"):("C",0.96,0.250,2320,36),
    ("尼崎","差し"):("C",0.94,0.233,1920,43),("常滑","抜き"):("C",0.93,0.231,3960,13),
    ("津","まくり"):("C",0.90,0.250,4630,28),("唐津","抜き"):("C",0.84,0.143,6530,28),
    ("徳山","差し"):("C",0.84,0.244,3490,41),("鳴門","差し"):("C",0.83,0.171,4580,35),
    ("平和島","差し"):("C",0.83,0.194,4240,62),("福岡","まくり"):("C",0.81,0.121,3260,58),
    ("多摩川","まくり差し"):("C",0.80,0.152,2730,33),("下関","差し"):("C",0.76,0.190,2740,42),
    ("江戸川","抜き"):("C",0.75,0.217,2410,46),
    ("江戸川","差し"):("D",0.75,0.180,2220,61),("多摩川","まくり"):("D",0.72,0.106,5580,66),
    ("三国","抜き"):("D",0.69,0.182,2710,22),("三国","まくり"):("D",0.66,0.190,3240,42),
    ("大村","まくり"):("D",0.64,0.133,8270,30),("宮島","差し"):("D",0.63,0.147,3740,34),
    ("多摩川","差し"):("D",0.63,0.205,2930,44),("宮島","抜き"):("D",0.62,0.143,6030,21),
    ("唐津","まくり"):("D",0.62,0.111,5770,54),("福岡","差し"):("D",0.62,0.258,3030,31),
    ("児島","差し"):("D",0.61,0.154,5380,65),("宮島","まくり"):("D",0.61,0.145,4770,55),
    ("尼崎","まくり"):("D",0.58,0.180,2650,50),("徳山","まくり差し"):("D",0.56,0.194,2700,36),
    ("大村","差し"):("D",0.54,0.130,4190,23),("住之江","差し"):("D",0.49,0.158,1440,19),
    ("児島","抜き"):("D",0.48,0.143,2800,21),("蒲郡","抜き"):("D",0.47,0.200,900,20),
    ("住之江","まくり"):("D",0.46,0.114,3590,44),("若松","抜き"):("D",0.44,0.188,3330,16),
    ("平和島","抜き"):("D",0.44,0.077,7860,26),("常滑","まくり差し"):("D",0.38,0.114,2960,35),
    ("蒲郡","まくり"):("D",0.29,0.104,3200,48),("丸亀","抜き"):("D",0.21,0.222,1270,18),
    ("びわこ","まくり"):("D",0.21,0.053,4360,57),
}
_GRADE_LABEL={"S":"✅✅ 強く推奨（ROI150%以上）","A":"✅  推奨（ROI120〜149%）",
    "B":"⚠️  参加可（ROI100〜119%）","C":"⚠️⚠️ 慎重に（ROI75〜99%）","D":"⛔  見送り推奨（ROI75%未満）"}
_GRADE_FILL={"S":("FFD9EAD3","FF1D5730"),"A":("FFE2EFDA","FF274E13"),
    "B":("FFFFF2CC","FF7F3F00"),"C":("FFFCE4D6","FF7F0000"),"D":("FFFF9999","FF7F0000")}
def _get_entry_grade(venue,scenario_type,honmei_scenario=None):
    kata=None
    if honmei_scenario and isinstance(honmei_scenario,dict):
        kata=(honmei_scenario.get("honmei_patterns",{}).get("honmei",{}).get("primary_kata"))
    if not kata:
        if scenario_type=="逃げ軸流し": kata="逃げ"
        elif scenario_type in ("飛び軸","両建て"): kata="まくり差し"
    entry=_ENTRY_GRADE_TABLE.get((venue,kata)) if (venue and kata) else None
    if not entry:
        return{"grade":"?","label":"ℹ️ データ不足","roi":None,"hit_rate":None,
               "med_odds":None,"sample_n":None,"fill":("FFDDDDDD","FF808080"),"kata":kata or "-"}
    grade,roi,hit_rate,med_odds,sample_n=entry
    return{"grade":grade,"label":_GRADE_LABEL[grade],"roi":roi,"hit_rate":hit_rate,
           "med_odds":med_odds,"sample_n":sample_n,"fill":_GRADE_FILL[grade],"kata":kata}


def _calc_kelly_fraction(theory_syn_odds, total_prob, fraction=0.25):
    """
    【⑨追加】フラクショナルケリー基準による最適賭け比率を算出する。

    【設計思想】
    現状システムは「参加/見送り」の2択だが、
    「参加するなら総資金の何%を使うべきか」が未実装。
    ケリー基準を用いることで、期待値の高いレースに多く賭け
    期待値の低いレースに少なく賭ける資金管理を実現する。

    【計算式】
      通常のケリー基準: f = (b×p - q) / b
        b = 理論合成オッズ（テラ銭控除後）
        p = 的中確率（= total_prob = Σ買い目確率）
        q = 外れ確率（= 1 - p）

    【フラクショナルケリー】
      純粋ケリーは過剰賭けリスクがあるため fraction=0.25（4分の1ケリー）を使用。
      これはバンクロールの変動を抑えながら長期的な資産成長を狙う実用的な設定。

    【パラメータ】
      theory_syn_odds : float  テラ銭控除後の理論合成オッズ（0.75 / Σprob）
      total_prob      : float  買い目全体の的中確率合計（0〜1）
      fraction        : float  ケリー乗数（デフォルト0.25 = 4分の1ケリー）

    【返り値】
      dict:
        kelly_f        : float  ケリー比率（総資金に対する賭け比率）
        kelly_pct      : str    パーセント表示（例: "3.2%"）
        kelly_label    : str    判定ラベル（「積極」「標準」「控えめ」「見送り」）
        kelly_edge     : float  エッジ（期待値 - 1）
        kelly_note     : str    説明テキスト
    """
    import math as _math

    if theory_syn_odds is None or total_prob is None or total_prob <= 0:
        return {
            "kelly_f":     0.0,
            "kelly_pct":   "0.0%",
            "kelly_label": "計算不可",
            "kelly_edge":  0.0,
            "kelly_note":  "合成オッズまたは確率データなし",
        }

    b = float(theory_syn_odds)   # 理論合成オッズ（テラ銭控除後）
    p = float(total_prob)        # 的中確率
    q = 1.0 - p                  # 外れ確率

    # エッジ（期待値 - 1）: プラスなら有利なゲーム
    edge = b * p - 1.0

    if edge <= 0 or b <= 1:
        # 期待値マイナス → 賭けるべきではない
        return {
            "kelly_f":     0.0,
            "kelly_pct":   "0.0%",
            "kelly_label": "見送り推奨（EV<0）",
            "kelly_edge":  round(edge, 4),
            "kelly_note":  f"期待値{(b*p)*100:.1f}% < 100%（マイナス期待値）",
        }

    # ケリー比率計算
    kelly_full = (b * p - q) / b   # 純粋ケリー
    kelly_f    = kelly_full * fraction   # フラクショナルケリー

    # 上限: 資金の20%を超えないようにキャップ（過剰賭け防止）
    kelly_f = min(kelly_f, 0.20)
    kelly_f = max(kelly_f, 0.0)

    # 判定ラベル
    if kelly_f >= 0.10:
        label = "積極（高期待値）"
    elif kelly_f >= 0.05:
        label = "標準"
    elif kelly_f >= 0.02:
        label = "控えめ"
    else:
        label = "最小賭け"

    note = (
        f"理論合成オッズ{b:.1f}倍 × 的中確率{p*100:.1f}% "
        f"→ エッジ{edge*100:.1f}% / 推奨賭け比率{kelly_f*100:.1f}%"
        f"（4分の1ケリー基準）"
    )

    return {
        "kelly_f":     round(kelly_f, 4),
        "kelly_pct":   f"{kelly_f*100:.1f}%",
        "kelly_label": label,
        "kelly_edge":  round(edge, 4),
        "kelly_note":  note,
    }



# 参加見送り判定（バックテスト結果に基づく精度向上フィルタ）
# ============================================================
# 【根拠】バックテスト8,526R分析（2025-12〜2026-02）
#   逃げ軸流し: 4,665R / ROI 27.1% / 損益 −377万  ← 全損失の152%
#   悪会場10会場: ROI 34〜67%帯 / 合計大赤字
#   逃げ除外だけで ROI 130.4% → 黒字転換
# ============================================================
_SKIP_VENUES: set = set()  # ← 低ROI会場フィルタ廃止（BT通過済み）

def _should_skip_race(bet_suggestions: dict, venue: str = "") -> tuple[bool, str]:
    """
    参加見送りを判定する。
    Returns: (skip: bool, reason: str)

    【v3 設計思想】
    「逃げ軸流し = 見送り」というルールを廃止。
    逃げ軸流しは展開予測であって見送り理由ではない。

    見送り条件（構造的に回収できないケースのみ）:
      ⓪ nyujo_henkou == True（最優先）
         → 枠なり進入を前提とした分析のため、進入変更が確認された時点で分析前提が崩れる
         → 他の条件より先に判定し、無条件で見送り
      ① s1_prob >= 0.65 かつ himo_are が不参加推奨
         → 逃げほぼ確定 + ヒモ固まり = 1-2-3が低オッズ化確実
      ② s1_prob >= 0.72（逃げ確率が極端に高い）
         → 逃げが当たっても低オッズ = 構造的に回収不能
      ③ honmei_prob_mismatch == True かつ s1_prob >= 0.60
         → 印↔確率が大きく乖離 + 逃げ優位 = 判断根拠が不明確
    """
    # ⓪ 進入変更（最優先：分析前提が崩れるため他条件より先に判定）
    if bet_suggestions.get("nyujo_henkou", False):
        return True, (
            "⛔見送り推奨（進入変更あり）\n"
            "本システムは枠なり進入を前提として分析しています。\n"
            "進入変更が確認されたレースは分析結果が無効になります。"
        )

    s1_prob         = bet_suggestions.get("s1_prob", 0) or 0
    mismatch        = bet_suggestions.get("honmei_prob_mismatch", False)
    mismatch_detail = bet_suggestions.get("honmei_prob_mismatch_detail", "")

    # ① 逃げ確率高 + ヒモ固まり（最も確実な見送り根拠）
    himo_are = bet_suggestions.get("himo_are", {}) or {}
    himo_verdict = himo_are.get("verdict", "対象外")
    if s1_prob >= 0.65 and himo_verdict == "不参加推奨":
        est_odds = himo_are.get("est_top_odds", 0) or 0
        return True, (
            f"⛔見送り推奨（逃げ{s1_prob*100:.0f}%+ヒモ固まり）\n"
            f"推定最高人気オッズ≈{est_odds:.0f}倍台。当たっても回収構造が成立しない"
        )

    # ② 逃げ確率が極端に高い（ヒモ判定関係なく低オッズ確実）
    if s1_prob >= 0.72:
        return True, (
            f"⛔見送り推奨（逃げ確率{s1_prob*100:.0f}%超）\n"
            f"1号艇逃げがほぼ確定→3連単オッズが構造的に低い"
        )

    # ③ 印↔確率不一致 + 逃げ優位（判断根拠不明確）
    if mismatch and s1_prob >= 0.60:
        return True, (
            f"⛔印↔確率不一致見送り（{mismatch_detail}）\n"
            f"逃げ{s1_prob*100:.0f}%優位で◎が確率最大艇でない"
        )

    return False, ""

def _suggest_3rentan(results, race_judgment, jizen_eval=None, honmei_map=None, tenkai_venue=None, tenkai_national=None, venue=None, venue_stats=None):
    """
    3連単買い目提案 v5 ── 印ベース軸決定＋内側残存補正ヒモ選択

    【設計思想】
      実オッズ・参加可否は人間が最終判断する前提。
      このロジックは「どの展開が来そうか」「その展開で何を買うべきか」を
      展示前の素材として提供することに専念する。

    【シナリオ判定と買い目形式】
      S1確率 >= 60%           → 逃げ軸流し（1着=1号艇固定）
      S1確率 40〜60%（拮抗）  → 両建てフォーメーション（逃げ軸＋飛び軸）
      S1確率 < 40%（飛び有力）→ 飛び軸フォーメーション

    【買い目の構成原則】
      ・全120通りの推定確率をシナリオ別に分類
      ・各シナリオの確率上位から累積80%を目安に買い目を選出
      ・折り返し買い（1-A-B と A-1-B）を自動的に対で追加
      ・カットオフなし・点数上限なし（確率で自動決定）

    【出力】
      scenario_type    : "逃げ軸流し" / "両建て" / "飛び軸"
      buy_list         : 推奨買い目（combo文字列リスト）
      point_count      : 点数
      theory_syn_odds  : 理論合成オッズ（= 0.75 / Σprob）
      required_syn_odds: 必要合成オッズ（= 点数 × 1.10）
      margin_ratio     : theory_syn_odds / required_syn_odds
      margin_verdict   : "余裕あり" / "要確認" / "見送り有力"
      candidates       : 買い目ごとの確率・シナリオ種別リスト
    """
    ryotate         = race_judgment.get("ryotate", {})
    ryotate_verdict = ryotate.get("verdict", "逃げ狙い")

    if not results:
        return {
            "axis1": "-", "axis2": "-", "buy_list": [], "point_count": 0,
            "comment": "データ不足", "combos": [], "candidates": [],
            "scenario_type": "-", "scenario_verdict": "-",
            "theory_syn_odds": None, "required_syn_odds": None,
            "margin_ratio": None, "margin_verdict": "-",
            "escape_score": 50, "tobi_score": 30, "fly_axes": [],
            "candidates_s1": [], "candidates_s2": [],
            "axis_candidates": [], "himo_candidates": [],
            "jizen_formation": {}, "ryotate_verdict": ryotate_verdict,
            "ryotate_detail": ryotate,
        }

    combos = _calc_3rentan_probs_v2(
        results,
        venue_course_1c_rate=race_judgment.get("venue_c1_win_rate"),
        jizen_eval=jizen_eval,
        race_judgment=race_judgment,
        tenkai_national=tenkai_national,
        tenkai_venue=tenkai_venue,
        venue_stats=venue_stats,
    )
    if not combos:
        return {
            "axis1": "-", "axis2": "-", "buy_list": [], "point_count": 0,
            "comment": "確率計算不能", "combos": [], "candidates": [],
            "scenario_type": "-", "scenario_verdict": "-",
            "theory_syn_odds": None, "required_syn_odds": None,
            "margin_ratio": None, "margin_verdict": "-",
            "escape_score": 50, "tobi_score": 30, "fly_axes": [],
            "candidates_s1": [], "candidates_s2": [],
            "axis_candidates": [], "himo_candidates": [],
            "jizen_formation": {}, "ryotate_verdict": ryotate_verdict,
            "ryotate_detail": ryotate, "first_prob_map": {},
        }

    # ══════════════════════════════════════════════════════════════════════
    # Step A: 1着別確率集計
    # ══════════════════════════════════════════════════════════════════════
    first_prob_map = {}
    for c in combos:
        w = c["first"]
        first_prob_map[w] = first_prob_map.get(w, 0) + c["prob"]

    s1_prob  = first_prob_map.get("1", 0.0)
    fly_prob = sum(p for w, p in first_prob_map.items() if w != "1")

    # ══════════════════════════════════════════════════════════════════════
    # Step A-2: 個人攻撃有効性による first_prob_map 補正
    # ══════════════════════════════════════════════════════════════════════
    # 【設計思想】
    # 現状の first_prob_map は「統計的な1着確率（コース別マスタ×会場特性）」で
    # 計算されており、「このメンバー6人の個人間の勝負」が反映されていない。
    #
    # _calc_attack_effectiveness は既に
    #   攻撃力 × 1号艇脆弱性 × STアドバンテージ × 会場適性
    # を個人レベルで計算しているため、これを first_prob_map に直接フィードバックする。
    #
    # 補正方式:
    #   2〜6号艇: first_prob × (1 + atk_eff × ATKCORR_WEIGHT)
    #   1号艇:    first_prob × (1 - max_threat × NIGE_SUPPRESS_WEIGHT)
    #   → 全艇を再正規化して確率の総和を1.0に保つ
    #
    # 係数設計:
    #   ATKCORR_WEIGHT = 0.40: 攻撃有効性スコアが1.0のとき最大+40%の確率ブースト
    #   NIGE_SUPPRESS_WEIGHT = 0.30: 最大脅威スコア1.0のとき1号艇を最大-30%抑制
    #   → 統計確率を完全に置き換えるのではなく「メンバー補正」として重ねる設計
    # ══════════════════════════════════════════════════════════════════════
    ATKCORR_WEIGHT      = 0.22   # 修正: 0.40→0.22（統計確率を守る）
    NIGE_SUPPRESS_WEIGHT = 0.20  # 修正: 0.30→0.20（1号艇確率の過剰抑制を防ぐ）

    _r1_for_atk = next((r for r in results if r["waku"] == "1"), None)
    _w1_cm_atk  = _r1_for_atk.get("raw_cm", {}) if _r1_for_atk else {}

    # 各攻撃艇の有効性スコアを計算
    _atk_eff_map = {}   # {waku: total_score}
    _atk_type_map = {}  # {waku: attack_type}
    _atk_breakdown = {} # {waku: breakdown dict} ← 狙い目コメント生成用
    for _r in results:
        if _r["waku"] == "1":
            continue
        _eff = _calc_attack_effectiveness(_r, _w1_cm_atk, venue_stats or {}, results)
        _atk_eff_map[_r["waku"]]   = _eff["total_score"]
        _atk_type_map[_r["waku"]]  = _eff["attack_type"]
        _atk_breakdown[_r["waku"]] = _eff

    # 最大脅威スコア（1号艇の逃げ抑制用）
    _max_threat_score = max(_atk_eff_map.values()) if _atk_eff_map else 0.0

    # first_prob_map を個人攻撃有効性で補正
    _corrected_prob = {}
    for _w, _p in first_prob_map.items():
        if _w == "1":
            # 1号艇: 最大脅威が高いほど逃げ確率を抑制
            _suppress = 1.0 - _max_threat_score * NIGE_SUPPRESS_WEIGHT
            _corrected_prob[_w] = max(_p * _suppress, 0.001)
        else:
            # 2〜6号艇: 個人攻撃有効性でブースト
            _boost = 1.0 + _atk_eff_map.get(_w, 0.0) * ATKCORR_WEIGHT
            _corrected_prob[_w] = _p * _boost

    # 再正規化（総和を1.0に保つ）
    _total_corrected = sum(_corrected_prob.values()) or 1.0
    first_prob_map = {
        _w: round(_p / _total_corrected, 5)
        for _w, _p in _corrected_prob.items()
    }

    # ── 狙い目（neraime）生成 ────────────────────────────────────────────
    # 攻め型: 攻撃有効性が閾値を超えた艇（1着狙い）
    # 残存型: 逃げ本命時に展開別残存マスタから2着残存確率が高い艇（2着狙い）
    #
    # 閾値: 0.15（旧0.25から引き下げ。各因子の積で0.25超えは稀すぎた）
    # 信頼度: 高(≥0.35) / 中(≥0.22) / 低(≥0.15) の3段階で表示
    NERAIME_THRESHOLD = 0.15
    _neraime_cands = sorted(
        [(w, s) for w, s in _atk_eff_map.items() if s >= NERAIME_THRESHOLD],
        key=lambda x: x[1], reverse=True
    )

    def _build_neraime_reason(waku, breakdown):
        """攻撃有効性の内訳から狙い目の根拠文を生成"""
        atk_type  = _atk_type_map.get(waku, "攻撃")
        atk_score = _atk_eff_map.get(waku, 0.0)
        bd        = breakdown or {}
        parts     = [f"{waku}号艇【{atk_type}】 攻撃有効性{atk_score*100:.0f}%"]
        if atk_type == "差し":
            v = bd.get("breakdown", {}).get("差し%", 0)
            if v:
                parts.append(f"差し実績{v*100:.0f}%")
        elif atk_type in ("まくり", "まくり差し"):
            mk  = bd.get("breakdown", {}).get("まくり%", 0)
            mks = bd.get("breakdown", {}).get("まくり差し%", 0)
            if mk or mks:
                parts.append(f"まくり系{(mk+mks)*100:.0f}%")
        vuln = bd.get("w1_vulnerability", 0)
        if vuln >= 0.3:
            parts.append(f"1号艇脆弱性高({vuln*100:.0f}%)")
        st_adv = bd.get("st_advantage", 1.0)
        if st_adv >= 1.3:
            parts.append(f"ST優位({st_adv:.2f}倍)")
        elif st_adv <= 0.7:
            parts.append(f"ST不利({st_adv:.2f}倍)")
        return " / ".join(parts)

    _neraime_list = []
    for _nw, _ns in _neraime_cands:
        _nbd = _atk_breakdown.get(_nw, {})
        _level = "高" if _ns >= 0.35 else "中" if _ns >= 0.22 else "低"
        _neraime_list.append({
            "waku":        _nw,
            "score":       round(_ns, 4),
            "level":       _level,
            "attack_type": _atk_type_map.get(_nw, "-"),
            "reason":      _build_neraime_reason(_nw, _nbd),
            "prob_after":  first_prob_map.get(_nw, 0.0),
            "type":        "攻め",   # 攻め型狙い目
        })

    # ── 残存型狙い目（逃げ本命時：展開別残存マスタの2着率直接参照）──────────
    # 逃げ本命（s1_prob ≥ 0.60）のとき「1号艇逃げ確定なら誰が2着に残るか」を
    # 展開別残存マスタの S1（逃げ/1着コース=1）から直接取得する。
    # circle_pct（イン逃げ時2着優位度）と 0.7:0.3 でブレンド。
    _neraime_2nd = []
    if s1_prob >= 0.60:
        _tenkai_s1_2nd = {}
        for _c in range(2, 7):
            _c_str = str(_c)
            _row_s1 = None
            _venue_rj = race_judgment.get("venue") if race_judgment else None
            if tenkai_venue and _venue_rj:
                _key_v = (str(_venue_rj), "逃げ", "1", _c_str)
                _rv = tenkai_venue.get(_key_v)
                if _rv:
                    try:
                        if float(_rv.get("信頼度") or 0) >= 0.15:
                            _row_s1 = _rv
                    except (ValueError, TypeError):
                        pass
            if _row_s1 is None and tenkai_national:
                _row_s1 = tenkai_national.get(("逃げ", "1", _c_str))
            if _row_s1:
                try:
                    _r2_v  = float(_row_s1.get("2着率")     or 0)
                    _r3i_v = float(_row_s1.get("3着以内率") or 0)
                    _tenkai_s1_2nd[_c_str] = {"2着率": _r2_v, "3着以内率": _r3i_v}
                except (ValueError, TypeError):
                    pass

        if _tenkai_s1_2nd:
            _circ_total = max(sum(circ_map.values()), 1)
            for _c_str, _rates in sorted(
                _tenkai_s1_2nd.items(), key=lambda x: x[1]["2着率"], reverse=True
            ):
                _r2_v   = _rates["2着率"]
                _r3i_v  = _rates["3着以内率"]
                _circ_v = circ_map.get(_c_str, 0)
                _circ_n = _circ_v / _circ_total   # 正規化
                # マスタ2着率(0.70) + 正規化circle_pct(0.30) でブレンド
                _blend2 = _r2_v * 0.70 + _circ_n * 0.30
                if _r2_v >= 0.15:
                    _neraime_2nd.append({
                        "waku":     _c_str,
                        "r2_rate":  round(_r2_v, 4),
                        "r3i_rate": round(_r3i_v, 4),
                        "blend":    round(_blend2, 4),
                        "reason":   (
                            f"{_c_str}号艇 逃げ時2着残存{_r2_v*100:.0f}%"
                            f"（展開別残存マスタ）/ 2着優位度{_circ_v:.0f}%"
                        ),
                        "type":     "残存",   # 残存型狙い目
                    })

    # 最有力狙い目（攻め型のスコア最大の1艇）
    neraime_top = _neraime_list[0] if _neraime_list else None

    # ── s1_prob / fly_prob を補正後の first_prob_map で再計算 ───────────
    s1_prob  = first_prob_map.get("1", 0.0)
    fly_prob = sum(p for w, p in first_prob_map.items() if w != "1")

    # 飛び候補艇を確率降順でリストアップ
    fly_candidates_sorted = sorted(
        [(w, p) for w, p in first_prob_map.items() if w != "1"],
        key=lambda x: x[1], reverse=True
    )
    main_fly_waku = fly_candidates_sorted[0][0] if fly_candidates_sorted else None
    sub_fly_waku  = fly_candidates_sorted[1][0] if len(fly_candidates_sorted) >= 2 else None

    # ══════════════════════════════════════════════════════════════════════
    # Step B: シナリオ判定（確率優先型 v3 ── 印は確率を補強する証拠として使用）
    # ══════════════════════════════════════════════════════════════════════
    # 【設計思想】
    # 旧版は「印◎の艇番でシナリオを決定」していたため、
    # 確率と印が逆向きになるとシナリオも買い目も確率と矛盾していた。
    # v3 では確率を起点にシナリオを決定し、印はその補強証拠として扱う。
    #
    # 判定フロー:
    #   Step B-1: s1_prob（確率）でシナリオの「下地」を決める
    #   Step B-2: 印◎の位置で補強・修正する（確率と同方向なら強化、逆方向なら両建て）
    #   Step B-3: ryotate（定性スコア）で最終調整
    # ══════════════════════════════════════════════════════════════════════
    first_turn      = race_judgment.get("first_turn", {}) or {}
    conflict_map    = race_judgment.get("conflict_map", {}) or {}
    sq              = race_judgment.get("scenario_quality", {}) or {}
    quality_rank    = sq.get("quality_rank", "B")
    lead_waku       = first_turn.get("lead_waku", "1")
    main_conflict   = conflict_map.get("main_conflict") or {}
    sub_conflict    = conflict_map.get("sub_conflict") or {}
    collapse_bene   = conflict_map.get("collapse_beneficiary", [])
    lead_is_1       = (lead_waku == "1")
    p_strength      = first_turn.get("pattern_strength", "中")
    mc_strength     = main_conflict.get("strength", 0) or 0

    # ── 印から軸・ヒモ情報を取得 ─────────────────────────────────────────
    # 【v7.0修正】1号艇の honmei は「逃◎→◎」に変換済みだが、
    # 2〜6号艇にも攻め◎が存在するため inv 辞書で後勝ちになり
    # 1号艇の◎が攻め◎で上書きされる問題を修正。
    #
    # 1号艇の軸判断: s1_prob の水準で直接決める（honmei_map を介さない）
    # 2〜6号艇の軸判断: 1号艇を除外した honmei_map の◎○▲△を参照
    if honmei_map:
        # 2〜6号艇のみで inv を構築（1号艇の"◎"を除外して上書き衝突を防ぐ）
        inv = {v: k for k, v in honmei_map.items() if v.strip() and k != "1"}
        atk_honmei_waku  = inv.get("◎")   # 攻め◎（2〜6号艇）
        taiko_waku       = inv.get("○")
        tanhana_waku     = inv.get("▲")
        ana_waku         = inv.get("△")
    else:
        sorted_fp = sorted(
            [(w, p) for w, p in first_prob_map.items() if w != "1"],
            key=lambda x: x[1], reverse=True
        )
        atk_honmei_waku  = sorted_fp[0][0] if len(sorted_fp) > 0 else None
        taiko_waku       = sorted_fp[1][0] if len(sorted_fp) > 1 else None
        tanhana_waku     = sorted_fp[2][0] if len(sorted_fp) > 2 else None
        ana_waku         = sorted_fp[3][0] if len(sorted_fp) > 3 else None

    # ── 1号艇の軸判断（v8.0: s1_prob バイパスで逃げ軸転落を防止）──────────
    # 【v7.0の副作用と修正】
    #   v7.0で1号艇の honmei は「逃◎」専用となり攻め◎は付かなくなった。
    #   そのため honmei_map 逆引きの atk_honmei_waku は常に2〜6号艇を指す。
    #   旧ロジックでは s1_prob>=0.60 で honmei_waku="1" としていたが、
    #   直後の Step B-2 で「honmei_waku!=1 → 両建て転落」に上書きされていた。
    #   （_prob_gap_pct の大小に関わらず全分岐が両建てを返す構造だった）
    #
    #   v8.0修正: s1_prob >= 0.60 の場合は honmei_waku="1" かつ
    #   scenario_type="逃げ軸流し" を直接確定し Step B-2 をバイパスする。
    #   飛び軸・両建て域（s1_prob < 0.60）は旧ロジックと同一。
    _s1_bypass = (s1_prob >= 0.60)  # True のとき Step B-2 をスキップ

    if s1_prob >= 0.60:
        honmei_waku   = "1"
        scenario_type = "逃げ軸流し"   # Step B-2 をバイパスして直接確定
    elif s1_prob >= 0.42:
        # 拮抗域: 攻め◎がいれば飛び軸候補、いなければ1号艇
        honmei_waku   = atk_honmei_waku if atk_honmei_waku else "1"
        scenario_type = None            # Step B-2 で確定
    else:
        # 飛び有力: 攻め◎を主軸に
        honmei_waku   = atk_honmei_waku if atk_honmei_waku else "1"
        scenario_type = None            # Step B-2 で確定

    # ── Step B-1: 確率でシナリオ下地を決定 ──────────────────────────────
    # s1_prob が高い → 1号艇が1着になる確率が高い → 逃げ軸が基本
    # s1_prob が低い → 他艇が1着になる確率が高い → 飛び軸が基本
    if s1_prob >= 0.60:
        scenario_base = "逃げ軸流し"
    elif s1_prob >= 0.42:
        scenario_base = "両建て"
    else:
        scenario_base = "飛び軸"

    # ── Step B-2: 印◎で補強・修正（s1_prob>=0.60 はバイパス済み）──────────
    # 確率と印が同方向 → シナリオ確定
    # 確率と印が逆方向 → 矛盾の「強度」で判断
    #
    # 【v2 改善: 矛盾解消ロジック】
    # 旧方式: 逆方向なら無条件に両建て
    # 新方式: 乖離幅に応じて判断
    #   ・乖離が大きい（確率差 >= 20pt）→ 確率優先でシナリオ決定、印は◎位置を修正
    #   ・乖離が中程度（10〜20pt）      → 両建て（どちらが正しいか不確実）
    #   ・乖離が小さい（< 10pt）         → 印優先（印の方が人間の定性判断が入っている）
    #
    # 「確率vs印の乖離幅」: honmei_waku の first_prob を確率上位艇と比較
    _honmei_first_prob = first_prob_map.get(honmei_waku, 0) if honmei_waku else 0
    _top_prob_waku     = max(first_prob_map, key=first_prob_map.get) if first_prob_map else "1"
    _top_first_prob    = first_prob_map.get(_top_prob_waku, 0)
    _prob_gap_pct      = (_top_first_prob - _honmei_first_prob) * 100  # 正 = 確率最大艇 > ◎

    if not _s1_bypass:
        # s1_prob < 0.60 のみ Step B-2 を実行（逃げ軸は既に確定済みのためスキップ）
        if scenario_base == "逃げ軸流し":
            if honmei_waku == "1":
                # 確率高×印◎=1号艇 → 最強の逃げシナリオ
                scenario_type = "逃げ軸流し"
            elif honmei_waku is not None:
                # 確率高なのに印◎≠1号艇
                if _prob_gap_pct >= 20:
                    # 確率と印の乖離が大きい → 確率優先（印が誤っている可能性）
                    scenario_type = "逃げ軸流し"
                    # ただし飛び軸候補として印◎艇も両建てに含める
                    main_fly_waku = honmei_waku
                    scenario_type = "両建て"
                elif _prob_gap_pct >= 10:
                    scenario_type = "両建て"
                    main_fly_waku = honmei_waku
                else:
                    # 乖離小 → 印優先（定性が確率を微修正している）
                    scenario_type = "両建て"
                    main_fly_waku = honmei_waku
            else:
                scenario_type = "逃げ軸流し"

        elif scenario_base == "飛び軸":
            if honmei_waku != "1":
                # 確率低×印◎≠1号艇 → 飛び軸確定
                scenario_type = "飛び軸"
                main_fly_waku = honmei_waku
            elif honmei_waku == "1":
                # 確率低なのに印◎=1号艇
                if _prob_gap_pct >= 20:
                    # 確率と印の乖離大 → 確率優先して飛び軸
                    scenario_type = "飛び軸"
                else:
                    # 乖離小 → 印を尊重して両建て
                    scenario_type = "両建て"
            else:
                scenario_type = "飛び軸"

        else:  # 両建て
            scenario_type = "両建て"
            if honmei_waku and honmei_waku != "1":
                main_fly_waku = honmei_waku

    # ── Step B-3: ryotate（定性スコア）で最終調整 ───────────────────────
    _ryotate_verdict   = ryotate.get("verdict", "逃げ狙い")
    _consistency_warn  = ryotate.get("consistency_warn", False)

    # ryotate が明確に飛び狙いと言っているのにシナリオが逃げ軸なら両建てに
    if _ryotate_verdict == "飛び狙い" and scenario_type == "逃げ軸流し":
        scenario_type = "両建て"
    # quality D は混戦 → 逃げ軸流しは危険
    if quality_rank == "D" and scenario_type == "逃げ軸流し":
        scenario_type = "両建て"

    # ── 飛び軸主軸の確定 ─────────────────────────────────────────────────
    if scenario_type in ("飛び軸", "両建て"):
        if main_fly_waku is None or main_fly_waku == "1":
            # 飛び軸主軸が未確定 or 1号艇になっている → 確率2位の艇に
            main_fly_waku = fly_candidates_sorted[0][0] if fly_candidates_sorted else "2"
        if sub_fly_waku is None or sub_fly_waku == main_fly_waku:
            sub_fly_waku = fly_candidates_sorted[1][0] if len(fly_candidates_sorted) >= 2 else taiko_waku

    # ══════════════════════════════════════════════════════════════════════
    # Step C: 買い目構成 ── 考察フル連動モデル v8.0
    # ══════════════════════════════════════════════════════════════════════
    #
    # 考察行 → 買い目への因果
    # ❶逃げ力(escape_rank)   → tenkai_pattern の基盤
    # ❷主役(main_score/type) → 飛び頭の軸・スロット比率
    # ❸残存(fallback_rank)   → 逃げ残存フォロー点数
    # ❹穴(dark_horse)        → 穴ヒモ挿入
    # ⑦展開quality(rank)     → 累積閾値・総点数
    # tenkai_pattern          → 1号頭/飛び頭 スロット比率
    #
    # ─── tenkai_pattern 別スロット比率 ──────────────────────────────────
    # A（鉄板逃げ）: 1号頭100%（逃げのみ）
    # B（主役展開）: 1号頭30% / 飛び頭70%  ← 飛び主体・逃げ残存フォロー付き
    # C（拮抗）    : 1号頭50% / 飛び頭50%  ← 均等両建て
    # D（荒れ）    : 1号頭30% / 飛び頭70%  ← 広め・穴ヒモ付き

    # ── quality 別 累積閾値 ──────────────────────────────────────────────
    THRESHOLD_BY_QUALITY = {"S": 0.70, "A": 0.75, "B": 0.80, "C": 0.85, "D": 0.85}
    CUMULATIVE_THRESHOLD = THRESHOLD_BY_QUALITY.get(quality_rank, 0.78)
    MIN_BETS = 3

    _race_are = safe_float((race_judgment or {}).get("venue_race_are_score")) if race_judgment else None
    if _race_are is not None:
        if _race_are >= 65:
            CUMULATIVE_THRESHOLD = min(0.92, CUMULATIVE_THRESHOLD + 0.05)
        elif _race_are >= 55:
            CUMULATIVE_THRESHOLD = min(0.90, CUMULATIVE_THRESHOLD + 0.02)
        elif _race_are <= 30:
            CUMULATIVE_THRESHOLD = max(0.65, CUMULATIVE_THRESHOLD - 0.05)
        elif _race_are <= 40:
            CUMULATIVE_THRESHOLD = max(0.68, CUMULATIVE_THRESHOLD - 0.02)

    # ── tenkai_pattern 確定 ───────────────────────────────────────────────
    _mp      = race_judgment.get("main_player", {}) or {}
    _dh      = race_judgment.get("dark_horse", {}) or {}
    _ef      = race_judgment.get("escape_fallback", {}) or {}
    _er      = (race_judgment.get("w1_escape", {}) or {}).get("escape_rank", "中")
    _ms      = float(_mp.get("main_score", 0) or 0)
    _fb_rank = _ef.get("fallback_rank", "中")
    _dh_ok   = _dh.get("is_valid", False)

    # v8.0: s1_prob を第1軸として tenkai_pattern を決定する
    # 【旧版の問題】escape_rank×main_score の2変数マトリクスのみで判定していたため、
    #   _er="高" でも _ms>=0.40 であれば常にBとなり、Aがほぼ発動しなかった。
    #   競艇では「逃げ力が高くても攻め手がいる」のは普通の状況なので
    #   _ms>=0.40 はほぼ全レースで成立し、結果1号頭スロット30%に固定されていた。
    # 【新版の設計】
    #   確率モデルが逃げ鉄板（s1_prob>=0.65）と判断しているなら、
    #   主役スコアに関わらず tenkai_pattern A を発動させる。
    #   s1_prob と escape_rank の両方が逃げを示す場合のみA確定とすることで
    #   確率と展開パターンの一貫性を保つ。
    if s1_prob >= 0.65 and _er == "高":
        # 確率モデル・定性スコア双方が逃げ鉄板 → A確定
        tenkai_pattern = "A"
    elif s1_prob >= 0.60 and _er == "高":
        # 逃げ有力だが主役候補が強い場合はB、そうでなければA
        tenkai_pattern = "A" if _ms < 0.55 else "B"
    elif s1_prob >= 0.60 and _er == "中":
        # 確率高・逃げ力中程度 → 拮抗C（逃げ軸寄り）
        tenkai_pattern = "C"
    elif s1_prob >= 0.50 and _er == "高" and _ms >= 0.50:
        # 逃げ力高・主役明確 → 主役展開B
        tenkai_pattern = "B"
    elif s1_prob >= 0.50 and _er == "高":
        # 逃げ力高・主役弱 → 拮抗C（逃げ軸寄り）
        tenkai_pattern = "C"
    elif s1_prob >= 0.42 and _ms >= 0.45:
        # 拮抗域・主役候補明確 → 主役展開B
        tenkai_pattern = "B"
    elif s1_prob >= 0.42:
        # 拮抗域・主役弱 → 拮抗C
        tenkai_pattern = "C"
    elif _er == "低" and _dh_ok:
        # 逃げ弱・穴候補あり → 荒れD
        tenkai_pattern = "D"
    elif _er == "低" and _ms >= 0.45:
        # 逃げ弱・主役強 → 主役展開B
        tenkai_pattern = "B"
    elif _er == "低":
        # 逃げ弱・主役弱 → 荒れD
        tenkai_pattern = "D"
    else:
        tenkai_pattern = "C"

    _TENKAI_POLICY = {
        "A": "1着1号艇固定・ヒモ絞り（逃げ圧倒・主役展開リスク低）",
        "B": "主役1着軸・逃げ残存フォロー（escape_fallback補強）",
        "C": "1号艇＋主役の2頭軸・ヒモ広め",
        "D": "穴候補込みの広め買い（dark_horse補強）",
    }
    _tenkai_policy_text = (
        "広め買い・参加慎重（逃げ弱・主役弱・穴候補なし＝全面荒れリスク）"
        if tenkai_pattern == "D" and not _dh_ok
        else _TENKAI_POLICY[tenkai_pattern]
    )
    race_judgment["tenkai_pattern"]        = tenkai_pattern
    race_judgment["tenkai_pattern_policy"] = _tenkai_policy_text

    # ── スロット配分 ─────────────────────────────────────────────────────
    MAX_BETS = 20 if tenkai_pattern == "D" else 18
    if _race_are is not None:
        if _race_are >= 65:
            MAX_BETS = min(MAX_BETS + 2, 22)
        elif _race_are >= 55:
            MAX_BETS = min(MAX_BETS + 1, 21)
        elif _race_are <= 30:
            MAX_BETS = max(MAX_BETS - 2, 10)
        elif _race_are <= 40:
            MAX_BETS = max(MAX_BETS - 1, 12)

    add_sc_bets   = (mc_strength >= 40 and len(collapse_bene) >= 1)
    _dh_cands_all = _dh.get("dark_horse_candidates", []) or []
    _dh_slots = (4 if len(_dh_cands_all) >= 2 else 2) if (tenkai_pattern in ("B", "D") and _dh_ok) else 0
    SC_SLOTS  = 2 if add_sc_bets and collapse_bene else 0
    base_max  = MAX_BETS - SC_SLOTS - _dh_slots

    # tenkai_pattern 別の 1号頭スロット比率
    _slot_ratio = {"A": 1.0, "B": 0.30, "C": 0.50, "D": 0.30}
    _w1_ratio   = _slot_ratio.get(tenkai_pattern, 0.50)
    # v8.0: s1_prob 連動の1号頭最低保証点数
    # 旧版 MIN_BETS=3 固定では打ち切り時に1号頭買い目が削られていた
    # s1_prob が高いほど1号頭スロットを積み増すことで逃げ軸を確保する
    _s1_min_bets = (
        5 if s1_prob >= 0.60 else
        4 if s1_prob >= 0.50 else
        MIN_BETS
    )
    _w1_slots   = base_max if tenkai_pattern == "A" else max(_s1_min_bets, round(base_max * _w1_ratio))
    _fly_slots  = 0 if tenkai_pattern == "A" else base_max - _w1_slots

    # ── 飛び軸スロットを◎頭（main）と○頭（taiko）に分割 ────────────────
    # 飛び軸（B/C/D）で taiko_waku（印○）が存在する場合、
    # ◎の確率に応じて◎頭と○頭にスロットを配分する。
    # ◎単独（taiko_waku なし）の場合は全スロットを◎頭に集中。
    #
    # 配分ロジック:
    #   確率比率 = ◎確率 / (◎確率 + ○確率) を基本に
    #   min 6:4（◎有利）〜 max 8:2 でクランプ
    _main_fly_prob  = first_prob_map.get(main_fly_waku, 0) if main_fly_waku else 0
    _taiko_prob     = first_prob_map.get(taiko_waku, 0)    if taiko_waku    else 0
    _has_taiko      = (
        taiko_waku is not None
        and taiko_waku != main_fly_waku
        and taiko_waku != "1"
        and _taiko_prob > 0
        and tenkai_pattern in ("B", "C", "D")
        and _fly_slots >= 4          # 最低4スロットないと○頭に割く余裕がない
    )
    if _has_taiko and (_main_fly_prob + _taiko_prob) > 0:
        _raw_ratio   = _main_fly_prob / (_main_fly_prob + _taiko_prob)
        _main_ratio  = max(0.60, min(0.80, _raw_ratio))   # 6:4〜8:2 でクランプ
    else:
        _main_ratio  = 1.0

    _main_fly_slots  = _fly_slots if not _has_taiko else max(MIN_BETS, round(_fly_slots * _main_ratio))
    _taiko_fly_slots = 0          if not _has_taiko else max(2, _fly_slots - _main_fly_slots)

    combo_lookup = {c["combo"]: c for c in combos}

    # 買い目根拠を生成する補助関数
    def _build_reason(e, scenario_ctx):
        """
        各買い目コンボの「買う理由」を自然言語で生成する。
        考察（展開エンジン）→ 買い目への因果を明示する。
        """
        first  = e["first"]
        second = e["second"]
        third  = e["third"]
        is_rev = e.get("_orkaeshi", False)

        # 1着根拠
        if first == "1":
            if lead_is_1 and p_strength == "強":
                r1 = f"1号先行確定的({p_strength})"
            elif lead_is_1:
                r1 = f"1号先行優位"
            else:
                r1 = f"1号逃げ残り(先行は{lead_waku}号)"
        elif first == main_fly_waku:
            mc_method = main_conflict.get("method", "攻撃")
            r1 = f"{first}号{mc_method}（主軸攻撃・強度{mc_strength:.0f}）"
        elif first == sub_fly_waku:
            sc_method = sub_conflict.get("method", "攻撃") if sub_conflict else "攻撃"
            r1 = f"{first}号{sc_method}（副軸）"
        else:
            r1 = f"{first}号（その他）"

        # 2着根拠
        top_bene = [w for w, _ in collapse_bene[:2]]
        if is_rev:
            r2 = f"{second}号折返（{first}号自滅時の逃げ取り戻し想定）"
        elif second in top_bene:
            r2 = f"{second}号漁夫（{main_conflict.get('attacker','?')}号自滅時の受益）"
        else:
            r2 = f"{second}号残存"

        return f"{r1} / {r2}"

    # ── 折り返し要否の判定関数 ──────────────────────────────────────────────────

    def _needs_orkaeshi_12(base_combo, rev_key):
        """
        1着折り返し（A-1-B vs 1-A-B）が必要かを判定する。

        不要と判断する条件:
          ① s1_prob >= 0.75: 逃げ確率が圧倒的 → 飛び役1着はほぼない
          ② 折り返しコンボの確率が本体の1/4未満: 展開として非現実的
          ③ 折り返し1着艇の1着確率が全艇平均の0.5倍未満: その艇が1着になる素地がない

        いずれか1つでも該当すれば不要と判断。
        """
        if rev_key not in combo_lookup:
            return False
        base = combo_lookup.get(base_combo)
        rev  = combo_lookup[rev_key]
        if not base:
            return True
        # ① 逃げ圧倒的
        if s1_prob >= 0.75:
            return False
        # ② 確率比
        if base["prob"] > 0 and rev["prob"] / base["prob"] < 0.25:
            return False
        # ③ 折り返し1着艇の1着確率（記号○以上は閾値を0.5→0.3に緩和）
        # 「攻め力を認められた艇なら確率が低めでも折り返しを追加する価値がある」
        rev_first_waku = rev["first"]
        rev_first_prob = first_prob_map.get(rev_first_waku, 0)
        avg_first_prob = sum(first_prob_map.values()) / max(len(first_prob_map), 1)
        _honmei_of_rev = next(
            (r.get("honmei", "") for r in results if r["waku"] == rev_first_waku), ""
        )
        _thresh = 0.3 if _honmei_of_rev in ("◎", "○") else 0.5
        if rev_first_prob < avg_first_prob * _thresh:
            return False
        return True

    def _needs_orkaeshi_23(base_combo, rev_key):
        """
        2着3着折り返し（1-A-B vs 1-B-A）が必要かを判定する。

        不要と判断する条件:
          ① 本体と折り返しの確率比が3倍以上: 順序がほぼ固定的
          ② 2着と3着の circle_pct（イン逃げ時2着優位度）差が2倍以上:
             2着候補がほぼ固定されている
          ③ 折り返しコンボの確率が全買い目平均の0.4倍未満: 薄すぎる

        いずれか1つでも該当すれば不要と判断。
        """
        if rev_key not in combo_lookup:
            return False
        base = combo_lookup.get(base_combo)
        rev  = combo_lookup[rev_key]
        if not base:
            return True
        # ① 確率比（3倍以上なら逆順はほぼ来ない）
        if base["prob"] > 0 and base["prob"] / max(rev["prob"], 1e-9) >= 3.0:
            return False
        # ② circle_pct差（2着固定度）
        circ_a = next((r.get("circle_pct") or 0 for r in results if r["waku"] == base["second"]), 0)
        circ_b = next((r.get("circle_pct") or 0 for r in results if r["waku"] == base["third"]),  0)
        if circ_a > 0 and circ_b > 0 and circ_a / circ_b >= 2.0:
            return False
        if circ_b > 0 and circ_a > 0 and circ_b / circ_a >= 2.0:
            return False
        # ③ 折り返し確率が薄すぎる
        avg_prob = sum(c["prob"] for c in combos) / max(len(combos), 1)
        if rev["prob"] < avg_prob * 0.4:
            return False
        return True

    # ══════════════════════════════════════════════════════════════════════
    # 共通データマップ（_build_buys内で参照）
    # ══════════════════════════════════════════════════════════════════════
    _r_map     = {r["waku"]: r for r in results}
    _cm_map    = {r["waku"]: r.get("raw_cm", {}) for r in results}
    _circ_map  = {r["waku"]: (r.get("circle_pct") or 0) for r in results}
    _win3_map  = {r["waku"]: (r.get("win3_rate") or 0.0) for r in results}
    _st_map    = {r["waku"]: r.get("avg_st") for r in results}
    _motor_map = {}
    for r in results:
        try:
            v = float(r.get("motor2") or 0)
            _motor_map[r["waku"]] = v if v > 0 else None
        except (ValueError, TypeError):
            _motor_map[r["waku"]] = None

    # ── 【v6.2】_suggest_3rentan用 選手指数マップ（raw_pmから全指数を取得）──
    # _calc_3rentan_probs_v2内の同名変数と独立して構築（スコープが異なるため）
    _sg_form_map    = {}  # フォーム指数
    _sg_recent3_map = {}  # 直近3走1着率
    _sg_recent5_map = {}  # 直近5走1着率
    _sg_st_std_map  = {}  # ST標準偏差
    _sg_st_stab_map = {}  # ST安定スコア
    _sg_jizai_map   = {}  # 自在性加重1着率
    _sg_ippan_map   = {}  # 一般戦1着率
    _sg_recent10_map= {}  # 直近10走平均着順
    for r in results:
        w   = r["waku"]
        pm_r = r.get("raw_pm") or {}
        _sg_form_map[w]     = safe_float(pm_r.get("フォーム\n指数")    or pm_r.get("フォーム指数"))
        _sg_recent3_map[w]  = safe_float(pm_r.get("直近3走\n1着率")    or pm_r.get("直近3走1着率"))
        _sg_recent5_map[w]  = safe_float(pm_r.get("直近5走\n1着率")    or pm_r.get("直近5走1着率"))
        _sg_st_std_map[w]   = safe_float(pm_r.get("ST\n標準偏差")      or pm_r.get("ST標準偏差"))
        _sg_st_stab_map[w]  = safe_float(pm_r.get("ST安定\nスコア")    or pm_r.get("ST安定スコア"))
        _sg_jizai_map[w]    = safe_float(pm_r.get("自在性\n加重1着率") or pm_r.get("自在性加重1着率"))
        _sg_ippan_map[w]    = safe_float(pm_r.get("1着率\n(一般戦)")   or pm_r.get("1着率(一般戦)"))
        _sg_recent10_map[w] = safe_float(pm_r.get("直近10走\n平均着順")or pm_r.get("直近10走平均着順"))

    # ── 【v6.3】STレンジマップ（_member_scenario_scaleのr["_st_range"]を再利用）──
    # コース別マスタの最速ST・最遅STから算出済みの値をそのまま使う
    _sg_st_range_map  = {r["waku"]: r.get("_st_range") for r in results}
    _sg_valid_ranges  = [v for v in _sg_st_range_map.values() if v is not None]
    _sg_st_range_mean = sum(_sg_valid_ranges) / len(_sg_valid_ranges) if _sg_valid_ranges else 0.30

    # ── 【v6.4新設】★STフラグマップ（_suggest_3rentan スコープ）────────────
    # _member_scenario_scale側の_star_st_mapと同じ判定ロジック（スコープが独立しているため再構築）
    _sg_star_st_map = {}
    for r in results:
        cm_r = r.get("raw_cm") or {}
        val  = cm_r.get("★ST")
        _sg_star_st_map[r["waku"]] = bool(val and str(val).strip() in ("★", "True", "1"))

    # jizen評価マップ（jizen_eval から艇別に取り出す）
    _jizen_aisho  = {}
    _jizen_tenkai = {}
    _jizen_jizai  = {}
    if jizen_eval is not None:
        for idx in range(6):
            w = str(idx + 1)
            _jizen_aisho[w]  = (jizen_eval.get("aisho")    or [""] * 6)[idx]
            _jizen_tenkai[w] = (jizen_eval.get("tenkai")   or [""] * 6)[idx]
            _jizen_jizai[w]  = (jizen_eval.get("jizaisei") or [""] * 6)[idx]

    # 1着軸ごとの2着残存補正テーブル（動的版 v2）
    # 【設計根拠 v2】
    # 旧版は静的テーブル（物理的な内側残存のみ）。
    # 実際には「2着に来る艇は、残存するだけでなく自分も仕掛けて生き残った艇」。
    # → POSITION_REMAIN[first][second] = 内側残存補正 × max(active_score, 0.5)
    #   active_score: その艇が仕掛け試行して生き残れる確率（calc_attack_probabilityで計算）
    #
    # 1号艇1着のときは circle_pct（イン逃げ時2着優位度）を使うためニュートラル維持。
    # ── 展開別残存マスタから動的にPOSITION_REMAIN_BASEを構築 ──────────────────
    # Phase 1: 純粋に「場の論理」（コース位置）だけで補正係数を実データから算出
    # 【構築方法】
    #   1. 各軸艇の「主な決まり手」を決まり手%から推定（最頻決まり手）
    #   2. (会場, 決まり手, 1着コース) でtenkai_venue_masterを検索
    #   3. 会場データが薄い(信頼度<0.3)場合はtenkai_national_masterでフォールバック
    #   4. 各残存コースの2着率を全体平均で割って相対補正係数に変換（1.0=平均）
    #   5. データなし → フォールバック静的テーブルを使用
    _POSITION_REMAIN_FALLBACK = {
        "1": {"2": 1.0, "3": 1.0, "4": 1.0, "5": 1.0, "6": 1.0},
        "2": {"1": 1.35, "3": 0.75, "4": 0.90, "5": 0.95, "6": 0.95},
        "3": {"1": 1.25, "2": 1.20, "4": 0.85, "5": 0.90, "6": 0.90},
        "4": {"1": 1.20, "2": 1.15, "3": 1.10, "5": 0.80, "6": 0.85},
        "5": {"1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "6": 0.75},
        "6": {"1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "5": 0.95},
    }

    # 各軸艇の主な決まり手を推定（コース別マスタの決まり手%から）
    KIMETE_PRIORITY = {
        "1": ["逃げ"],
        "2": ["差し", "まくり差し"],
        "3": ["まくり", "まくり差し", "差し"],
        "4": ["まくり", "まくり差し", "差し"],
        "5": ["まくり", "まくり差し"],
        "6": ["まくり", "まくり差し"],
    }

    def _get_main_kimete(waku):
        """軸艇の主な決まり手をコース別マスタから推定する"""
        r = _r_map.get(waku, {})
        candidates = KIMETE_PRIORITY.get(waku, ["まくり"])
        best_k, best_pct = candidates[0], 0.0
        for k in candidates:
            pct_key = f"決まり手_{k}%"
            try:
                v = float(r.get(pct_key) or r.get("raw_cm", {}).get(pct_key, 0) or 0)
                if v > best_pct:
                    best_pct, best_k = v, k
            except (ValueError, TypeError):
                pass
        return best_k

    def _build_pos_remain_from_master(first_w):
        """
        展開別残存マスタから first_w 軸の残存補正係数を構築する。
        戻り値: {second_w: 補正係数} （1.0=全国平均、>1.0=残りやすい、<1.0=残りにくい）

        【修正】決まり手を1つに固定せず、全決まり手の2着率を実データ比率でブレンド。
        例: 3号艇が差し6%・まくり6%の会場では「差し/3」と「まくり/3」のマスタを
            それぞれ参照し、比率に応じてブレンドした係数を返す。
        これにより決まり手が拮抗する場合のマスタ精度が向上する。
        """
        if first_w == "1":
            return _POSITION_REMAIN_FALLBACK["1"]  # 逃げはcircle_pctで判断

        # 実際の進入コースを results から取得
        r_data = _r_map.get(first_w, {})
        actual_course = str(int(float(r_data.get("course") or r_data.get("進入コース") or first_w)))

        # 各決まり手の実データ比率を取得（cm_map から）
        cm_fw = _cm_map.get(first_w, {})
        def _safe_pct_local(key):
            v = cm_fw.get(key)
            try:
                return max(float(v), 0.0) if v is not None else 0.0
            except (ValueError, TypeError):
                return 0.0

        kimete_weights = {}
        if actual_course == "1":
            kimete_weights["逃げ"] = 1.0
        else:
            sashi   = _safe_pct_local("差し%")
            makuri  = _safe_pct_local("まくり%")
            maksa   = _safe_pct_local("まくり差し%")
            total_k = sashi + makuri + maksa
            if total_k > 0:
                kimete_weights["差し"]      = sashi  / total_k
                kimete_weights["まくり"]    = makuri / total_k
                kimete_weights["まくり差し"] = maksa  / total_k
            else:
                # データなし → コース別デフォルト
                _DEFAULT_KIMETE = {
                    "2": {"差し": 0.70, "まくり": 0.10, "まくり差し": 0.20},
                    "3": {"差し": 0.20, "まくり": 0.35, "まくり差し": 0.45},
                    "4": {"差し": 0.10, "まくり": 0.60, "まくり差し": 0.30},
                    "5": {"差し": 0.10, "まくり": 0.65, "まくり差し": 0.25},
                    "6": {"差し": 0.10, "まくり": 0.60, "まくり差し": 0.30},
                }
                kimete_weights = _DEFAULT_KIMETE.get(actual_course, {"まくり": 1.0})

        def _fetch_row(kimete_k, c_str):
            """(決まり手, 進入コース) の行を会場別→全国の優先順で取得"""
            if c_str == actual_course:
                return None
            if tenkai_venue and venue:
                key_v = (str(venue), kimete_k, actual_course, c_str)
                row_v = tenkai_venue.get(key_v)
                if row_v:
                    try:
                        if float(row_v.get("信頼度") or 0) >= 0.3:
                            return row_v
                    except (ValueError, TypeError):
                        pass
            if tenkai_national:
                key_n = (kimete_k, actual_course, c_str)
                return tenkai_national.get(key_n)
            return None

        # 全決まり手をブレンドして2着率を集計
        blended_rates = {}
        for c in range(1, 7):
            c_str = str(c)
            if c_str == actual_course:
                continue
            blended_r2 = 0.0
            total_weight = 0.0
            for kimete_k, w_k in kimete_weights.items():
                if w_k <= 0:
                    continue
                row = _fetch_row(kimete_k, c_str)
                if row is None:
                    continue
                try:
                    r2 = float(row.get("2着率") or 0)
                    blended_r2    += r2 * w_k
                    total_weight  += w_k
                except (ValueError, TypeError):
                    pass
            if total_weight > 0:
                blended_rates[c_str] = blended_r2 / total_weight

        if not blended_rates:
            return _POSITION_REMAIN_FALLBACK.get(first_w, {})

        avg_rate = sum(blended_rates.values()) / max(len(blended_rates), 1)
        if avg_rate < 0.001:
            return _POSITION_REMAIN_FALLBACK.get(first_w, {})

        result = {}
        for wk, rate in blended_rates.items():
            coef = rate / avg_rate
            result[wk] = max(0.5, min(2.0, coef))
        return result

    # 全軸艇の残存補正テーブルを事前構築
    POSITION_REMAIN_BASE = {}
    for _fw in [str(i) for i in range(1, 7)]:
        POSITION_REMAIN_BASE[_fw] = _build_pos_remain_from_master(_fw)

    # 各艇の仕掛け積極性スコアを事前計算（動的補正用）
    # active_score = P_attempt × P_survive（仕掛けて生き残れる確率）
    _r1_for_ap = next((r for r in results if r["waku"] == "1"), None)
    _active_scores = {}
    if HONMEI_SCENARIO_AVAILABLE and _r1_for_ap:
        try:
            from honmei_scenario import calc_attack_probability as _cap
            for _r in results:
                if _r["waku"] == "1":
                    _active_scores["1"] = 0.5   # 1号艇自身は逃げなので中立
                else:
                    _ap = _cap(_r, _r1_for_ap, results)
                    _active_scores[_r["waku"]] = _ap["attempt_prob"] * _ap["survive_prob"]
        except Exception:
            pass

    def _get_pos_remain(first_w, second_w):
        """POSITION_REMAIN_BASE × active_score補正で動的残存補正係数を返す"""
        base = POSITION_REMAIN_BASE.get(first_w, {}).get(second_w, 1.0)
        if first_w == "1":
            return base   # 逃げ時は circle_pct で別途判断
        # active_score: 0.5未満でも0.5を下限（完全に来ない艇は除外されるが補正は緩く）
        active = _active_scores.get(second_w, 0.5)
        active_adj = max(0.5, min(1.2, active * 1.5 + 0.25))
        return base * active_adj

    POSITION_REMAIN = POSITION_REMAIN_BASE   # 後方互換 (circle_pct処理で直接参照する箇所用)

    def _calc_himo_score(first_w, second_w, third_w, combo_prob):
        """
        1着軸が決まったときの 2着・3着組み合わせの総合スコアを算出。

        【スコア構成】
          ① combo確率ベース     (35%): シナリオ計算済みの3連単確率
          ② 位置残存補正        (25%): 内側残存の物理補正（対話ログ合意版）
              差し1着   → 1号艇内残存×1.35、3号艇競合×0.75
              まくり1着 → 内側艇↑（×1.05〜1.25）、外側艇↓（×0.75〜0.95）
              1号艇1着  → circle_pct（イン逃げ時2着優位度）で補正
          ③ 個人能力            (25%): コース別3連対率 × ST能力 × 機力 × 自在性
          ④ jizen展開・相性     (15%): 2着艇の相性評価 × 3着艇の展開評価
        """
        # ── ① combo確率（S1シナリオとは別物。3連単の確率値） ─────────────
        combo_prob_val = combo_prob

        # ── ② 位置残存補正 ───────────────────────────────────────────────
        # 1号艇1着のときは circle_pct（イン逃げ時2着優位度）を使う
        if first_w == "1":
            circ2 = _circ_map.get(second_w, 50) / 100.0
            circ3 = _circ_map.get(third_w, 50) / 100.0
            pos2  = 0.5 + circ2 * 0.5   # 0.5〜1.0にスケール
            pos3  = 0.5 + circ3 * 0.5
        else:
            # 動的残存補正: 物理的内側残存 × 仕掛け積極性
            pos2 = _get_pos_remain(first_w, second_w)
            pos3 = _get_pos_remain(first_w, third_w)
        pos_score = (pos2 + pos3) / 2.0

        # ── ③ 個人能力スコア ─────────────────────────────────────────────
        def _personal(w):
            """
            【v6.3強化版】2着・3着候補の個人能力を多面的に評価。

            構成要素（重み合計 = 1.0）:
              コース別3連対率  (0.30): 場の論理に即した実績
              STスコア         (0.18): このレースでの発艇優位
              機力スコア       (0.12): モーター2連率の相対評価
              フォーム指数     (0.10): 直近調子の総合指標
              直近3走1着率     (0.08): 超短期フォーム
              直近5走1着率     (0.05): 短期フォーム
              自在性加重1着率  (0.07): 外枠攻め実力（S2〜S4で重要）
              ST標準偏差(逆)   (0.025): STばらつき小=安定 【v6.3: 0.04→0.025に調整】
              STレンジ(逆)     (0.015): 最速〜最遅レンジ小=コース内安定 【v6.3新追加】
              一般戦1着率      (0.03): 格付け補正
              直近10走平均着順 (0.03): 中期トレンド
            """
            # コース別3連対率（0〜1）
            w3 = _win3_map.get(w, 0.0)

            # STスコア: 速い艇ほど高い（艇間相対、0〜1）
            st_self = _st_map.get(w)
            all_sts = [v for v in _st_map.values() if v is not None]
            if st_self is not None and len(all_sts) >= 2:
                st_min, st_max = min(all_sts), max(all_sts)
                st_score = 1.0 - (st_self - st_min) / max(st_max - st_min, 0.001)
            else:
                st_score = 0.5

            # 機力スコア（艇間相対、0〜1）
            valid_motors = [v for v in _motor_map.values() if v is not None]
            mv = _motor_map.get(w)
            if mv is not None and len(valid_motors) >= 2:
                m_min, m_max = min(valid_motors), max(valid_motors)
                motor_score = (mv - m_min) / max(m_max - m_min, 0.001)
            else:
                motor_score = 0.5

            # フォーム指数（中央値3.0基準、0〜1にスケール）
            form = _sg_form_map.get(w)
            form_score = 0.5
            if form is not None:
                form_score = max(0.0, min(1.0, form / 6.0))  # 0〜6+で0〜1+

            # 直近3走1着率（全国平均0.17基準、0〜1）
            r3 = _sg_recent3_map.get(w)
            r3_score = 0.5
            if r3 is not None:
                r3_score = max(0.0, min(1.0, r3 / 0.34))   # 0.34(2倍平均)で1.0

            # 直近5走1着率
            r5 = _sg_recent5_map.get(w)
            r5_score = 0.5
            if r5 is not None:
                r5_score = max(0.0, min(1.0, r5 / 0.34))

            # 自在性加重1着率（外枠攻め実力、全国平均0.06基準）
            jizai = _sg_jizai_map.get(w)
            jizai_score = 0.5
            if jizai is not None:
                jizai_score = max(0.0, min(1.0, jizai / 0.12))  # 0.12(2倍平均)で1.0

            # ST標準偏差逆スコア（小さいほど高スコア）
            # ★STフラグ = サンプル10未満 → ST値が不安定なためスキップ（0.5=中立を維持）
            _sg_st_unreliable = _sg_star_st_map.get(w, False)
            st_std = _sg_st_std_map.get(w)
            st_std_score = 0.5
            if st_std is not None and not _sg_st_unreliable:
                # 0.044〜0.143の範囲: 0.044→1.0、0.143→0.0
                st_std_score = max(0.0, min(1.0, 1.0 - (st_std - 0.044) / 0.099))

            # ── 【v6.3新追加 / v6.4★STガード追加】STレンジ逆スコア ─────────
            # _sg_st_range_map は _suggest_3rentan スコープで構築済み
            st_range = _sg_st_range_map.get(w)
            st_range_score = 0.5
            if st_range is not None and _sg_st_range_mean > 0 and not _sg_st_unreliable:
                # レンジが平均より小さいほど1.0に近づく（最大1.0、最小0.0）
                st_range_score = max(0.0, min(1.0,
                    0.5 + (_sg_st_range_mean - st_range) / (2.0 * _sg_st_range_mean)
                ))

            # 一般戦1着率（全国平均0.17基準）
            ippan = _sg_ippan_map.get(w)
            ippan_score = 0.5
            if ippan is not None:
                ippan_score = max(0.0, min(1.0, ippan / 0.34))

            # 直近10走平均着順（3.5基準、2.0→1.0、5.0→0.0）
            r10 = _sg_recent10_map.get(w)
            r10_score = 0.5
            if r10 is not None:
                r10_score = max(0.0, min(1.0, (5.0 - r10) / 3.0))

            return (w3              * 0.30
                  + st_score        * 0.18
                  + motor_score     * 0.12
                  + form_score      * 0.10
                  + r3_score        * 0.08
                  + jizai_score     * 0.07
                  + r5_score        * 0.05
                  + st_std_score    * 0.025   # v6.3: 0.04→0.025（STレンジと合計0.04を分担）
                  + st_range_score  * 0.015   # v6.3新追加
                  + ippan_score     * 0.03
                  + r10_score       * 0.03)

        personal2 = _personal(second_w)
        personal3 = _personal(third_w)
        personal_score = (personal2 + personal3) / 2.0

        # ── ④ jizen展開・相性スコア ──────────────────────────────────────
        sym4 = {"◎": 1.0, "○": 0.75, "△": 0.40, "": 0.25}

        # 2着艇の相性（1号艇に対する攻め適性）
        aisho2  = sym4.get(_jizen_aisho.get(second_w, ""), 0.25)
        # 3着艇の展開（外枠での展開形成力）
        tenkai3 = sym4.get(_jizen_tenkai.get(third_w, ""), 0.25)
        jizen_score = (aisho2 + tenkai3) / 2.0

        # ── ⑤ 課題1修正: ❷主役候補のplace2/3_candidatesボーナス ────────────
        # _judge_main_playerが「このメンバー構成でこの展開なら2・3着に来やすい」
        # と判断した艇番リストに一致するコンボを加点する。
        # 加点幅: 2着一致=+0.08 / 3着一致=+0.05（合成比率の範囲内に収まるよう設計）
        _mp_data      = race_judgment.get("main_player", {}) or {}
        _p2_wakus     = {w for w, _ in (_mp_data.get("place2_candidates") or [])}
        _p3_wakus     = {w for w, _ in (_mp_data.get("place3_candidates") or [])}
        _p2_bonus     = 0.08 if second_w in _p2_wakus else 0.0
        _p3_bonus     = 0.05 if third_w  in _p3_wakus else 0.0
        main_cand_bonus = _p2_bonus + _p3_bonus   # 最大 0.13

        # ── 合成（⑤ボーナスを加算、上限1.0にクランプ）──────────────────────
        score = min(1.0,
               combo_prob_val * 0.35
               + pos_score        * 0.25
               + personal_score   * 0.25
               + jizen_score      * 0.15
               + main_cand_bonus)

        return score

    def _build_buys(base_first_waku, orkaeshi_first=None, max_bets=None):
        source = [c for c in combos if c["first"] == base_first_waku]
        if not source:
            return [], set()

        scored = []
        for c in source:
            hs = _calc_himo_score(c["first"], c["second"], c["third"], c["prob"])
            scored.append((c, hs))
        scored.sort(key=lambda x: x[1], reverse=True)
        total_hs = sum(hs for _, hs in scored) or 1.0
        scored_with_share = [(c, hs, hs / total_hs) for c, hs in scored]

        _limit = max_bets if max_bets is not None else base_max
        selected  = []
        combo_set = set()
        cum_share = 0.0

        for c, hs, share in scored_with_share:
            key = c["combo"]
            if key not in combo_set:
                selected.append(c)
                combo_set.add(key)
                cum_share += share
            if orkaeshi_first is not None:
                rev_key = f"{c['second']}-{orkaeshi_first}-{c['third']}"
                if rev_key not in combo_set and _needs_orkaeshi_12(c["combo"], rev_key):
                    rev = dict(combo_lookup[rev_key])
                    rev["_orkaeshi"] = True
                    combo_set.add(rev_key)
                    selected.append(rev)
                    rev_hs = _calc_himo_score(rev["first"], rev["second"], rev["third"], rev["prob"])
                    cum_share += rev_hs / total_hs
            if len(selected) >= MIN_BETS and cum_share >= CUMULATIVE_THRESHOLD:
                break
            if len(selected) >= _limit:
                break
        return selected, combo_set

    def _trim_to_max(entries, max_n):
        return sorted(entries, key=lambda x: x["prob"], reverse=True)[:max_n]

    buy_entries = []
    seen_combos = set()

    # ════════════════════════════════════════════════════════════════════
    # パターン別 買い目生成（考察→買い目 フル連動）
    # ════════════════════════════════════════════════════════════════════
    if tenkai_pattern == "A":
        # A: 鉄板逃げ ── 1号頭のみ
        entries, _ = _build_buys("1", orkaeshi_first="1", max_bets=base_max)
        buy_entries = entries
        # 2着3着折り返し（逃げ軸専用）
        orkaeshi_23_set = set()
        for e in list(buy_entries):
            if e["first"] != "1":
                continue
            rev_key  = f"1-{e['third']}-{e['second']}"
            pair_key = tuple(sorted([e["combo"], rev_key]))
            if pair_key in orkaeshi_23_set:
                continue
            orkaeshi_23_set.add(pair_key)
            if not any(x["combo"] == rev_key for x in buy_entries) and _needs_orkaeshi_23(e["combo"], rev_key):
                rev = dict(combo_lookup[rev_key])
                rev["_orkaeshi_23"] = True
                buy_entries.append(rev)

    elif tenkai_pattern == "C":
        # C: 拮抗 ── 1号頭50% + 飛び頭50%（◎＋○に分割）
        s1_entries,  s1_seen  = _build_buys("1", orkaeshi_first="1", max_bets=_w1_slots)
        fly_entries, fly_seen = (
            _build_buys(main_fly_waku, orkaeshi_first=main_fly_waku, max_bets=_main_fly_slots)
            if main_fly_waku else ([], set())
        )
        taiko_entries, taiko_seen = (
            _build_buys(taiko_waku, orkaeshi_first=None, max_bets=_taiko_fly_slots)
            if _has_taiko else ([], set())
        )
        all_raw = (
            s1_entries
            + [e for e in fly_entries    if e["combo"] not in s1_seen]
            + [e for e in taiko_entries  if e["combo"] not in s1_seen and e["combo"] not in fly_seen]
        )
        all_raw.sort(key=lambda x: x["prob"], reverse=True)
        for e in all_raw:
            if e["combo"] not in seen_combos:
                buy_entries.append(e)
                seen_combos.add(e["combo"])

    else:
        # B/D: 主役/荒れ ── 飛び頭主体（◎＋○） + 1号頭最低保証
        fly_entries, fly_seen = (
            _build_buys(main_fly_waku, orkaeshi_first=main_fly_waku, max_bets=_main_fly_slots)
            if main_fly_waku else ([], set())
        )
        taiko_entries, taiko_seen = (
            _build_buys(taiko_waku, orkaeshi_first=None, max_bets=_taiko_fly_slots)
            if _has_taiko else ([], set())
        )
        s1_entries, s1_seen = _build_buys("1", orkaeshi_first="1", max_bets=_w1_slots)
        all_raw = (
            fly_entries
            + [e for e in taiko_entries if e["combo"] not in fly_seen]
            + [e for e in s1_entries    if e["combo"] not in fly_seen
                                          and e["combo"] not in taiko_seen]
        )
        all_raw.sort(key=lambda x: x["prob"], reverse=True)
        for e in all_raw:
            if e["combo"] not in seen_combos:
                buy_entries.append(e)
                seen_combos.add(e["combo"])

    # ── base_max 打ち切り（1号頭 _w1_slots 点を保証してから打ち切る）────────
    if len(buy_entries) > base_max:
        w1_in  = [e for e in buy_entries if e["first"] == "1"]
        non_w1 = [e for e in buy_entries if e["first"] != "1"]
        _keep_w1 = min(len(w1_in), _w1_slots if tenkai_pattern != "A" else base_max)
        kept_w1  = sorted(w1_in,  key=lambda x: x["prob"], reverse=True)[:_keep_w1]
        kept_fly = sorted(non_w1, key=lambda x: x["prob"], reverse=True)[:base_max - _keep_w1]
        buy_entries = kept_w1 + kept_fly
        seen_combos = {e["combo"] for e in buy_entries}

    # ── ❺ SC1着按分を escape_fallback で補正 ────────────────────────────────────
    _fb_prob = float(_ef.get("fallback_prob", 0.5) or 0.5)
    _fb_type = _ef.get("fly_type", "不明")
    _rel_map_for_sc = {w: p for w, p in first_prob_map.items()}
    _sc_info2 = _calc_sc_weight(
        results, _cm_map, _win3_map, _rel_map_for_sc, jizen_eval=jizen_eval
    )
    _sc_1st_weights = dict(_sc_info2["sc_1st_weights"])
    if _fb_type == "まくり系":
        _corrected_w1 = max(0.35, min(0.95, 0.35 + _fb_prob * 0.60))
    elif _fb_type == "差し系":
        _corrected_w1 = max(0.10, min(0.50, 0.10 + _fb_prob * 0.40))
    else:
        _corrected_w1 = _sc_1st_weights.get("1", 0.50)
    _old_w1 = _sc_1st_weights.get("1", 0.50)
    _delta  = _corrected_w1 - _old_w1
    _sc_1st_weights["1"] = _corrected_w1
    if sub_fly_waku and sub_fly_waku in _sc_1st_weights:
        _sc_1st_weights[sub_fly_waku] = max(0.05, _sc_1st_weights[sub_fly_waku] - _delta)

    # ── SCシナリオ（潰れ展開）漁夫候補を補完 ─────────────────────────────
    if add_sc_bets and collapse_bene:
        top_bene_w = collapse_bene[0][0]
        sc_additions = sorted(
            [c for c in combos
             if c["third"] == top_bene_w and c["combo"] not in seen_combos],
            key=lambda x: x["prob"], reverse=True
        )[:SC_SLOTS]
        for c in sc_additions:
            c = dict(c); c["_sc_bet"] = True
            buy_entries.append(c); seen_combos.add(c["combo"])

    # ── ❸ 逃げ残存フォロー ────────────────────────────────────────────────
    _fallback_limit = 0
    if tenkai_pattern == "B" and _fb_rank == "高":
        _fallback_limit = 2
    elif tenkai_pattern == "C" and _fb_rank == "高":
        _fallback_limit = 1
    if _fallback_limit > 0:
        _main_w = _mp.get("main_waku")
        if _main_w and _main_w != "1":
            for c in sorted(
                [c for c in combos
                 if c["first"] == _main_w and c["second"] == "1"
                 and c["combo"] not in seen_combos],
                key=lambda x: x["prob"], reverse=True
            )[:_fallback_limit]:
                c = dict(c); c["_fallback_bet"] = True
                buy_entries.append(c); seen_combos.add(c["combo"])

    # ── ❹ ダークホース3着挿入 ─────────────────────────────────────────────
    _venue_c1    = float(race_judgment.get("venue_c1_win_rate") or 0.555)
    _venue_ratio = max(0.70, min(1.30, _venue_c1 / 0.555))
    _dh_thresh   = round({"D": 0.15, "B": 0.20}.get(tenkai_pattern, 0.20) * _venue_ratio, 3)
    if _dh_ok and tenkai_pattern in ("B", "D"):
        _dh_top_w = _dh.get("top_waku")
        _dh_score = float(_dh.get("top_score", 0) or 0)
        if _dh_top_w and _dh_score >= _dh_thresh:
            for c in sorted(
                [c for c in combos
                 if c["third"] == _dh_top_w and c["combo"] not in seen_combos],
                key=lambda x: x["prob"], reverse=True
            )[:2]:
                c = dict(c); c["_dh_bet"] = True
                buy_entries.append(c); seen_combos.add(c["combo"])
            for _rank_i, (_dh_w, _dh_s, _dh_tag) in enumerate(_dh_cands_all[1:3], start=2):
                if _dh_s < _dh_thresh * (1.0 + _rank_i * 0.25):
                    continue
                for c in sorted(
                    [c for c in combos
                     if c["third"] == _dh_w and c["combo"] not in seen_combos],
                    key=lambda x: x["prob"], reverse=True
                )[:1]:
                    c = dict(c); c["_dh_bet"] = True
                    buy_entries.append(c); seen_combos.add(c["combo"])

    # ── 最終上限チェック ──────────────────────────────────────────────────
    if len(buy_entries) > MAX_BETS:
        buy_entries = _trim_to_max(buy_entries, MAX_BETS)

    # ── 飛び狙い判定時の1号頭フィルタ ────────────────────────────────────
    # v8.0修正: s1_prob >= 0.55 の場合は確率モデルが逃げを支持しているため
    # 定性スコア由来の「飛び狙い」判定で1号頭を消してはいけない
    # （escape_score は race_judgment["score"] 由来でs1_probと別軸のため矛盾が生じやすい）
    if _ryotate_verdict == "飛び狙い" and tenkai_pattern in ("B", "D") and s1_prob < 0.55:
        non_w1 = [e for e in buy_entries if e["first"] != "1"]
        if len(non_w1) >= MIN_BETS:
            buy_entries = non_w1

    # 艇番若い順ソート
    buy_entries.sort(key=lambda e: (int(e["first"]), int(e["second"]), int(e["third"])))
    buy_list    = [e["combo"] for e in buy_entries]
    point_count = len(buy_list)

    # ══════════════════════════════════════════════════════════════════════
    # Step D: 理論合成オッズ・余裕度・期待値警告
    # ══════════════════════════════════════════════════════════════════════
    # 【設計思想】
    # 点数で削らない。合成オッズが基準を下回った場合は「警告」として表示し、
    # 判断はユーザーに委ねる。
    #   → 回収重視ユーザー: 警告を見て見送り
    #   → 的中重視ユーザー: 警告を無視して参考買い目を使う
    #
    # 期待値基準:
    #   quality S/A → 3.0倍未満で警告
    #   quality B/C/D → 2.5倍未満で警告
    # ══════════════════════════════════════════════════════════════════════
    total_prob        = sum(e["prob"] for e in buy_entries)
    theory_syn_odds   = round(0.75 / total_prob, 1) if total_prob > 0 else None
    required_syn_odds = round(point_count * 1.10, 1) if point_count > 0 else None

    # 期待値警告フラグ
    # 胴元控除25%を考慮したブレークイーブンは1.33倍。
    # 実用的な下限を2.0倍とし、それを下回る場合のみ警告する。
    # （旧設定のS/A=3.0倍、B以下=2.5倍は厳しすぎて正常なレースにも警告が出ていた）
    EV_THRESHOLD = 2.0
    ev_warning = (theory_syn_odds is not None and theory_syn_odds < EV_THRESHOLD)
    ev_warning_msg = (
        f"⚠ 理想合成オッズ{theory_syn_odds}倍（期待値基準{EV_THRESHOLD}倍を下回っています）\n"
        f"  → 回収重視なら見送り推奨 / 的中重視なら参考買い目を使用可"
    ) if ev_warning else ""

    if theory_syn_odds and required_syn_odds and required_syn_odds > 0:
        margin_ratio = round(theory_syn_odds / required_syn_odds, 2)
        if margin_ratio >= 2.0:
            margin_verdict = "余裕あり（実オッズが理論値の半分でも成立）"
        elif margin_ratio >= 1.2:
            margin_verdict = "要確認（展示後に実オッズで判断）"
        else:
            margin_verdict = "見送り有力（理論値に余裕なし）"
    else:
        margin_ratio   = None
        margin_verdict = "-"

    # ══════════════════════════════════════════════════════════════════════
    # Step E: コメント・候補リスト生成（考察の結論を明示）
    # ══════════════════════════════════════════════════════════════════════
    fly_str    = f"{main_fly_waku}号艇" if main_fly_waku else "-"
    q_guide    = sq.get("bet_size_guide", "-")
    mc_desc    = main_conflict.get("desc", "-") if main_conflict else "-"

    # 考察→買い目の因果を1行で要約
    reasoning_line = (
        f"1M先行{lead_waku}号/{p_strength}・{mc_desc}"
        f"・quality{quality_rank}→{q_guide}"
    )

    # 整合性警告（ryotate の consistency_warn / honmei_prob_mismatch）を付加
    _warn_parts = []
    if ryotate.get("consistency_warn"):
        _warn_parts.append("⚠️スコア↔確率乖離補正済")
    _mismatch = race_judgment.get("honmei_prob_mismatch", False)
    if _mismatch:
        _mismatch_detail = race_judgment.get("honmei_prob_mismatch_detail", "")
        _warn_parts.append(f"⚠️印↔確率不一致({_mismatch_detail})")
    warn_suffix = "  " + " / ".join(_warn_parts) if _warn_parts else ""

    # 狙い目サマリー（comment用）
    if neraime_top:
        _neraime_summary = (
            f"🎯狙い目: {neraime_top['waku']}号艇【{neraime_top['attack_type']}】"
            f" 有効性{neraime_top['score']*100:.0f}%"
            f" → 補正後1着確率{neraime_top['prob_after']*100:.1f}%"
        )
    else:
        _neraime_summary = "🎯狙い目: 明確な攻め手なし（逃げ本命レース）"

    comment = (
        f"【{scenario_type}】【展開:{tenkai_pattern}】"
        f"逃げ{s1_prob*100:.0f}%/飛び{fly_prob*100:.0f}%（主:{fly_str}）"
        f" ／ {point_count}点 / 理論合成{theory_syn_odds}倍"
        f" / 余裕度{margin_ratio}倍 → {margin_verdict}\n考察根拠: {reasoning_line}"
        f"\n展開方針: {_TENKAI_POLICY.get(tenkai_pattern, '-')}"
        f"\n{_neraime_summary}"
        f"{warn_suffix}"
    )

    def _scenario_label(e):
        f = e["first"]
        if e.get("_sc_bet"):
            return f"潰れ受益({f}号頭)"
        elif e.get("_fallback_bet"):
            return f"❸逃げ残存({f}号頭-1号2着)"
        elif e.get("_dh_bet"):
            return f"❹穴ヒモ({f}号頭)"
        elif e.get("_orkaeshi_23"):
            return f"2着3着折返({f}号頭)"
        elif f == "1":
            return "逃げ（1号艇頭）"
        elif f == main_fly_waku:
            return f"◎飛び（{f}号艇頭）"
        elif f == taiko_waku and _has_taiko:
            return f"○飛び（{f}号艇頭）"
        elif e.get("_orkaeshi"):
            return f"1着折返({f}号頭)"
        else:
            return f"その他（{f}号艇頭）"

    # ヒモスコアを買い目エントリに付与（candidatesで表示用）
    _himo_score_cache = {}
    for e in buy_entries:
        k = e["combo"]
        if k not in _himo_score_cache:
            _himo_score_cache[k] = _calc_himo_score(
                e["first"], e["second"], e["third"], e["prob"]
            )

    candidates = [
        {
            "combo":          e["combo"],
            "prob":           round(e["prob"], 5),
            "prob_pct":       round(e["prob"] * 100, 2),
            "himo_score":     round(_himo_score_cache.get(e["combo"], 0), 4),
            "scenario":       _scenario_label(e),
            "is_orkaeshi":    e.get("_orkaeshi", False),
            "is_orkaeshi_23": e.get("_orkaeshi_23", False),
            "is_sc_bet":      e.get("_sc_bet", False),
            "is_fallback_bet": e.get("_fallback_bet", False),
            "is_dh_bet":      e.get("_dh_bet", False),
            "reason":         _build_reason(e, scenario_type),
        }
        for e in buy_entries
    ]

    # ── 下流コードとの互換性維持 ──────────────────────────────────────────
    fly_axes = [w for w, _ in fly_candidates_sorted[:2]]

    from collections import Counter
    himo_counter = Counter()
    for c in combos[:30]:
        himo_counter[c["second"]] += 2
        himo_counter[c["third"]]  += 1
    axis_candidates = list(first_prob_map.keys())[:3]
    himo_candidates = [w for w, _ in himo_counter.most_common()
                       if w not in axis_candidates][:4]

    # race_judgmentへ分析結果を追記
    race_judgment["scenario_type"]   = scenario_type
    race_judgment["s1_prob"]         = round(s1_prob, 4)
    race_judgment["fly_prob"]        = round(fly_prob, 4)
    race_judgment["main_fly_waku"]   = main_fly_waku
    race_judgment["theory_syn_odds"] = theory_syn_odds
    race_judgment["margin_ratio"]    = margin_ratio
    race_judgment["margin_verdict"]  = margin_verdict
    race_judgment["ev_axes"]         = (
        ["1"] if scenario_type == "逃げ軸流し"
        else [main_fly_waku] if scenario_type == "飛び軸" and main_fly_waku
        else (["1", main_fly_waku] if main_fly_waku else ["1"])
    )
    race_judgment["w1_vs_venue"]     = round(
        (first_prob_map.get("1", 0) - (race_judgment.get("venue_c1_win_rate") or 0.555)) * 100, 1
    )
    race_judgment["w1_ev_pos_count"] = 0   # 互換維持（EV概念廃止）
    race_judgment["ev_axis_summary"] = [
        f"{w}号艇(1着確率{p*100:.1f}%)" for w, p in fly_candidates_sorted[:3]
    ]
    # SCシナリオ情報（数値シート表示用）
    _sc_combos_info = combos[0].get("sc_fly_type", None) if combos else None
    race_judgment["sc_fly_type"]    = _sc_combos_info or "-"
    race_judgment["sc_fly_waku"]    = main_fly_waku  # 飛び役兼潰れ役
    # 漁夫候補: 漁夫スコアが高い上位3艇（buy_listに含まれるもの優先）
    _sc_b = combos[0].get("sc_beneficiary") if combos else None
    if _sc_b is None:
        race_judgment["sc_gyofu_top3"] = []
    else:
        # combos全体からsc_beneficiaryを一度だけ取り出す（全艇共通値）
        # _calc_sc_weightの結果はcombo単位ではなく全艇共通なので最初のcomboから取得
        _all_b = {c["second"]: c.get("sc_beneficiary", 0) for c in combos[:30]}
        race_judgment["sc_gyofu_top3"] = sorted(
            _all_b.keys(), key=lambda w: _all_b.get(w, 0), reverse=True
        )[:3]

    h1 = "1" if scenario_type != "飛び軸" else (main_fly_waku or "1")
    h2 = main_fly_waku or "-"

    # ── honmei_scenario v2 統合 ──────────────────────────────────────────
    _base_result = {
        "axis1":             h1,
        "axis2":             h2,
        "buy_list":          buy_list,
        "point_count":       point_count,
        "comment":           comment,
        "combos":            combos,
        "candidates":        candidates,
        "scenario_type":     scenario_type,
        "scenario_verdict":  scenario_type,
        "s1_prob":           round(s1_prob, 4),
        "fly_prob":          round(fly_prob, 4),
        "theory_syn_odds":   theory_syn_odds,
        "required_syn_odds": required_syn_odds,
        "margin_ratio":      margin_ratio,
        "margin_verdict":    margin_verdict,
        "escape_score":      ryotate.get("escape_score", 50),
        "tobi_score":        ryotate.get("tobi_score", 30),
        "fly_axes":          fly_axes,
        "candidates_s1":     [],
        "candidates_s2":     [],
        "axis_candidates":   axis_candidates,
        "himo_candidates":   himo_candidates,
        "jizen_formation":   {},
        "ryotate_verdict":   ryotate_verdict,
        "ryotate_detail":    ryotate,
        "first_prob_map":    {w: round(p, 4) for w, p in first_prob_map.items()},
        # ── 狙い目（個人攻撃有効性ベース）──────────────────────────────
        "neraime":           _neraime_list,       # 攻め型狙い目候補リスト
        "neraime_2nd":       _neraime_2nd,        # 残存型2着狙い目（逃げ本命時）
        "neraime_top":       neraime_top,         # 最有力攻め型狙い目（1艇）
        "atk_eff_map":       {w: round(s, 4) for w, s in _atk_eff_map.items()},
        # 整合性フラグ（バックテスト除外・警告表示用）
        "consistency_warn":          ryotate.get("consistency_warn", False),
        "honmei_prob_mismatch":      race_judgment.get("honmei_prob_mismatch", False),
        "honmei_prob_mismatch_detail": race_judgment.get("honmei_prob_mismatch_detail", ""),
        # 期待値警告（合成オッズが基準を下回っている場合）
        "ev_warning":     ev_warning,
        "ev_warning_msg": ev_warning_msg,
        # ❻ 展開4分類（A:鉄板逃げ / B:主役展開 / C:拮抗 / D:荒れ）
        "tenkai_pattern":        tenkai_pattern,
        "tenkai_pattern_policy": _tenkai_policy_text,  # D判定の_dh_ok別出し分けを反映
    }

    if HONMEI_SCENARIO_AVAILABLE and honmei_map:
        # 【v8.0】逃げ鉄板判定: s1_prob高 かつ tenkai_pattern=A の場合は
        # integrate_with_suggest_3rentan を呼ばず _base_result をそのまま使う。
        # honmei_scenario は印◎艇（2〜6号）を軸にするため、
        # 逃げ鉄板レースで呼ぶと1号頭の買い目が4号頭等に差し替えられる。
        _hs_s1p = _base_result.get("s1_prob", 0) or 0
        _hs_tp  = _base_result.get("tenkai_pattern", "")
        _skip_integrate = (_hs_s1p >= 0.60 and _hs_tp == "A")

        if _skip_integrate:
            # 逃げ鉄板: honmei_scenario を呼ばず確率モデルの結果を使用
            _result = _base_result
        else:
            # honmei_scenario.py は results[waku]["honmei"] を直接参照する。
            # v7.0で1号艇の honmei が「逃◎」になったため、
            # resultsのコピーで1号艇 honmei を変換してから渡す。
            _nige_conv = {"逃◎": "◎", "逃○": "○", "逃△": "△", "逃×": " "}
            _results_for_hs = []
            for _r in results:
                if str(_r.get("waku", "")) == "1":
                    _rc = dict(_r)
                    _rc["honmei"] = _nige_conv.get(_rc.get("honmei", " "), " ")
                    _results_for_hs.append(_rc)
                else:
                    _results_for_hs.append(_r)

            _result = integrate_with_suggest_3rentan(
                original_result = _base_result,
                results         = _results_for_hs,
                honmei_map      = honmei_map,
                combos          = combos,
                race_judgment   = race_judgment,
                jizen_eval      = jizen_eval,
                )
        # integrate_with_suggest_3rentan が theory_syn_odds を上書きするため
        # ev_warning を最終的な theory_syn_odds で再計算する
        _tso = _result.get("theory_syn_odds")
        _result["ev_warning"] = (_tso is not None and _tso < EV_THRESHOLD)
        _result["ev_warning_msg"] = (
            f"⚠ 理想合成オッズ{_tso}倍（期待値基準{EV_THRESHOLD}倍を下回っています）\n"
            f"  → 回収重視なら見送り推奨 / 的中重視なら参考買い目を使用可"
        ) if _result["ev_warning"] else ""
    else:
        _result = _base_result

    # ── 参加グレードを付与 ──
    _venue   = (race_judgment or {}).get("venue", "")
    _result["entry_grade"] = _get_entry_grade(
        venue           = _venue,
        scenario_type   = _result.get("scenario_type", ""),
        honmei_scenario = _result.get("honmei_scenario"),
    )

    # ── 【⑨追加】ケリー基準による最適賭け比率を付与 ──
    # buy_listの要素はdictが前提だが、文字列等が混入するケースに備えて型チェック
    _buy_list_safe = [c for c in (_result.get("buy_list") or []) if isinstance(c, dict)]
    _kelly = _calc_kelly_fraction(
        theory_syn_odds = _result.get("theory_syn_odds"),
        total_prob      = sum(c.get("prob", 0) for c in _buy_list_safe),
    )
    _result["kelly"] = _kelly

    # ── 参加見送り判定をフラグとして付与 ──
    # venue は race_judgment 経由で受け取る（_suggest_3rentan は venue を直接知らない）
    # himo_are を race_judgment から _result に渡す（_should_skip_race が参照するため）
    if "himo_are" not in _result:
        _result["himo_are"] = (race_judgment or {}).get("himo_are", {}) or {}
    # nyujo_henkou（進入変更フラグ）を race_judgment から転記
    # _should_skip_race は _result から読むため、race_judgment にセットされた値を同期する
    _result["nyujo_henkou"] = (race_judgment or {}).get("nyujo_henkou", False)
    # main_player（主役候補判定）を race_judgment から転記して bet_suggestions でも参照できるようにする
    _result["main_player"] = (race_judgment or {}).get("main_player", {})
    # escape_fallback（❸ 主役自滅時の逃げ残存確率）を転記
    _result["escape_fallback"] = (race_judgment or {}).get("escape_fallback", {})
    # dark_horse（❹ 穴候補）を転記
    _result["dark_horse"] = (race_judgment or {}).get("dark_horse", {})

    # ── scenario_type の実態補正（v8.0）──────────────────────────────────
    # integrate_with_suggest_3rentan は buy_list/candidates を印◎ベースで差し替えるが
    # scenario_type を更新しない。その結果「逃げ軸流し」と書かれているのに
    # 買い目が全て非1号頭という矛盾が生じる。
    # ここで candidates の実際の1着分布から scenario_type を補正する。
    # ── is_orkaeshi フラグを base_result から復元 ──────────────────────────
    # integrate が candidates を差し替えるとき is_orkaeshi / is_orkaeshi_23 フラグが
    # 失われる。base_result の candidates から combo をキーにしてフラグを引き継ぐ。
    _base_cands_map = {
        c["combo"]: c
        for c in (_base_result.get("candidates") or [])
    }
    _cands_final = _result.get("candidates", [])
    for _c in _cands_final:
        _ck = _c.get("combo", "")
        if _ck in _base_cands_map:
            _bc = _base_cands_map[_ck]
            if "is_orkaeshi"    not in _c: _c["is_orkaeshi"]    = _bc.get("is_orkaeshi", False)
            if "is_orkaeshi_23" not in _c: _c["is_orkaeshi_23"] = _bc.get("is_orkaeshi_23", False)
    # ───────────────────────────────────────────────────────────────────────
    if _cands_final:
        from collections import Counter as _SC
        _first_dist = _SC(
            c["combo"].split("-")[0]
            for c in _cands_final
            if c.get("combo") and not c.get("is_fallback_bet") and not c.get("is_dh_bet")
        )
        _w1_cnt  = _first_dist.get("1", 0)
        _fly_cnt = sum(v for k, v in _first_dist.items() if k != "1")
        _top_fly = _first_dist.most_common(1)[0][0] if _fly_cnt > 0 else None

        _declared = _result.get("scenario_type", "")

        # 「逃げ軸流し」なのに1号頭が1点もない → 実態は飛び軸か両建て
        if _declared == "逃げ軸流し" and _w1_cnt == 0 and _fly_cnt > 0:
            _result["scenario_type"]    = "飛び軸"
            _result["scenario_verdict"] = "飛び軸"
            _result["scenario_type_note"] = (
                f"⚠ 確率モデル逃げ{s1_prob*100:.0f}%→逃げ軸流しだが"
                f"印◎{_top_fly}号により買い目は{_top_fly}号頭軸に変更"
            )

        # 「逃げ軸流し」だが1号頭と非1号頭が混在 → 両建て
        elif _declared == "逃げ軸流し" and _w1_cnt > 0 and _fly_cnt > 0:
            _result["scenario_type"]    = "両建て"
            _result["scenario_verdict"] = "両建て"
            _result["scenario_type_note"] = (
                f"1号頭{_w1_cnt}点 / {_top_fly}号頭{_fly_cnt}点の両建て"
                f"（確率逃げ{s1_prob*100:.0f}%・印◎{_top_fly}号）"
            )

        # 「飛び軸」なのに1号頭が過半数 → 逃げ軸流しに補正
        elif _declared == "飛び軸" and _w1_cnt > _fly_cnt:
            _result["scenario_type"]    = "逃げ軸流し"
            _result["scenario_verdict"] = "逃げ軸流し"
            _result["scenario_type_note"] = (
                f"1号頭{_w1_cnt}点が多数 → 逃げ軸流しに補正"
            )

        else:
            _result["scenario_type_note"] = ""
    # ─────────────────────────────────────────────────────────────────────
    _skip, _skip_reason = _should_skip_race(_result, _venue)
    _result["skip"]        = _skip
    _result["skip_reason"] = _skip_reason

    return _result


# ============================================================
# 本命記号スコア計算（トップレベル関数）
# ※ 元は calc_race_indices 内のローカル関数だったが
#   main() から直接呼ぶためトップレベルに移動。
# ============================================================

# ============================================================
# 6人相互作用モデル：攻撃有効性スコア計算
# ============================================================

def _calc_attack_effectiveness(attacker, w1_cm, venue_stats, results):
    """
    「この攻撃艇が、今日の1号艇を実際に崩せるか」を定量化する。

    【設計思想】
    単独の決まり手%や平均STではなく、
    「攻撃艇の攻撃力 × 1号艇の当該攻撃への脆弱性 × STアドバンテージ × 会場適性」
    の積として攻撃有効性を算出する。

    これにより「1号艇の逃げ率が高くても、2号艇のSTが圧倒的で差し実績が高ければ
    2号艇◎」という競艇の実態に即した印付けが可能になる。

    Parameters
    ----------
    attacker    : dict  攻撃艇のresultsエントリ
    w1_cm       : dict  1号艇のraw_cm（被決まり手データ含む）
    venue_stats : dict  会場統計（決まり手別比率）
    results     : list  全艇のresultsリスト（ST平均計算用）

    Returns
    -------
    dict:
        total_score     : 総合攻撃有効性スコア（0〜1）
        attack_type     : 主要攻撃手段（"差し"/"まくり"/"まくり差し"/"逃げ"）
        attack_power    : 攻撃力（0〜1）
        w1_vulnerability: 1号艇の当該攻撃への脆弱性（0〜1）
        st_advantage    : STアドバンテージ（0.5〜1.5）
        venue_affinity  : 会場の当該決まり手適性（0〜1）
        breakdown       : 各因子の詳細
    """
    waku      = str(attacker.get("waku", ""))
    cm        = attacker.get("raw_cm", {})
    avg_st    = attacker.get("avg_st")

    # 1号艇の情報
    st1       = next((r.get("avg_st") for r in results if r["waku"] == "1"), None)
    nige_pct1 = safe_float(w1_cm.get("逃げ%"), 0.6) or 0.6
    sasar_v   = safe_float(w1_cm.get("差され%"),     0.0) or 0.0
    makur_v   = safe_float(w1_cm.get("捲られ%"),     0.0) or 0.0
    maksa_v   = safe_float(w1_cm.get("捲り差され%"), 0.0) or 0.0

    # 全艇ST平均（基準）
    st_vals = [r["avg_st"] for r in results if r.get("avg_st") is not None]
    st_mean = sum(st_vals) / len(st_vals) if st_vals else 0.15

    # 会場の決まり手別成功率（攻撃が会場に合うか）
    kimari_avg = venue_stats.get("kimari_avg", {}) or {}
    v_sashi   = safe_float(kimari_avg.get("差し"),      0.15) or 0.15
    v_makuri  = safe_float(kimari_avg.get("まくり"),    0.20) or 0.20
    v_maksa   = safe_float(kimari_avg.get("まくり差し"),0.10) or 0.10
    # 全国平均でスケール正規化（差し:15%、まくり:20%、まくり差し:10%）
    v_sashi_n  = min(v_sashi  / 0.15, 2.0)
    v_makuri_n = min(v_makuri / 0.20, 2.0)
    v_maksa_n  = min(v_maksa  / 0.10, 2.0)

    # ── STアドバンテージ（1号艇比較）─────────────────────────────────────
    if avg_st is not None and st1 is not None:
        st_diff   = st1 - avg_st          # 正 = 自分が速い
        st_adv_v1 = max(0.3, min(2.0, 1.0 + st_diff / 0.04))
    else:
        st_adv_v1 = 1.0

    # ── コース別の主攻撃手段と有効性を計算 ──────────────────────────────
    if waku == "1":
        # 1号艇：逃げ力 × 被攻撃への耐性（別関数で計算）
        nige_pct = safe_float(cm.get("逃げ%"), 0.6) or 0.6
        return {
            "total_score":      nige_pct,
            "attack_type":      "逃げ",
            "attack_power":     nige_pct,
            "w1_vulnerability": 0.0,
            "st_advantage":     1.0,
            "venue_affinity":   1.0,
            "breakdown":        {"逃げ%": nige_pct},
        }

    elif waku == "2":
        # 2枠：差し特化
        # 差し有効性 = 差し% × 被差し脆弱性 × STアドバンテージ × 会場差し率
        sashi_pct = safe_float(cm.get("差し%"), 0.0) or 0.0
        # 被差し脆弱性：差され%だけでなく逃げ%の低さも加味
        vuln_s = min(1.0, sasar_v * 0.65 + (1.0 - nige_pct1) * 0.35)
        score  = sashi_pct * vuln_s * st_adv_v1 * v_sashi_n
        return {
            "total_score":      min(score, 1.0),
            "attack_type":      "差し",
            "attack_power":     sashi_pct,
            "w1_vulnerability": vuln_s,
            "st_advantage":     st_adv_v1,
            "venue_affinity":   v_sashi_n,
            "breakdown":        {
                "差し%": sashi_pct, "被差し脆弱性": round(vuln_s,3),
                "ST優位": round(st_adv_v1,3), "会場差し率": round(v_sashi_n,3),
            },
        }

    elif waku == "3":
        # 3枠：まくり差し主体（まくりも一部）
        # 3枠は2枠の差し力が高いと進路を塞がれるペナルティ
        maksa_pct = safe_float(cm.get("まくり差し%"), 0.0) or 0.0
        maku_pct  = safe_float(cm.get("まくり%"),    0.0) or 0.0
        attack_p  = maksa_pct * 1.2 + maku_pct * 0.8

        # 2枠の差し力が高いほど3枠は動きにくい（進路妨害）
        r2  = next((r for r in results if r["waku"] == "2"), None)
        if r2:
            w2_sashi = safe_float(r2.get("raw_cm", {}).get("差し%"), 0.0) or 0.0
            w2_st    = r2.get("avg_st")
            if w2_st is not None and st1 is not None:
                w2_adv = max(0.0, (st1 - w2_st) / 0.04)
            else:
                w2_adv = 0.0
            block_penalty = max(0.5, 1.0 - w2_sashi * w2_adv * 0.5)
        else:
            block_penalty = 1.0

        vuln_m = min(1.0, maksa_v * 0.6 + makur_v * 0.4)
        score  = attack_p * vuln_m * st_adv_v1 * v_maksa_n * block_penalty
        return {
            "total_score":      min(score, 1.0),
            "attack_type":      "まくり差し",
            "attack_power":     attack_p,
            "w1_vulnerability": vuln_m,
            "st_advantage":     st_adv_v1,
            "venue_affinity":   v_maksa_n,
            "breakdown":        {
                "まくり差し%": maksa_pct, "まくり%": maku_pct,
                "被脆弱性": round(vuln_m,3), "2枠妨害": round(block_penalty,3),
            },
        }

    else:
        # 4〜6枠：まくり主体（外枠ほど難易度高い）
        maku_pct  = safe_float(cm.get("まくり%"),    0.0) or 0.0
        maksa_pct = safe_float(cm.get("まくり差し%"),0.0) or 0.0
        attack_p  = maku_pct + maksa_pct * 0.8

        # 外枠補正（4→1.0、5→0.85、6→0.70）
        outer_mult = {4: 1.0, 5: 0.85, 6: 0.70}.get(int(waku), 0.70)

        # 内側艇による壁ペナルティ（自分より内側にまくり力の強い艇がいると被る）
        inner_block = 1.0
        for r in results:
            rw = int(r["waku"])
            if 2 <= rw < int(waku):
                r_maku = safe_float(r.get("raw_cm", {}).get("まくり%"), 0.0) or 0.0
                r_st   = r.get("avg_st")
                if r_st is not None and avg_st is not None:
                    r_adv = max(0.0, (avg_st - r_st) / 0.04)  # 内側艇が速いほど壁
                else:
                    r_adv = 0.0
                inner_block = max(0.4, inner_block - r_maku * r_adv * 0.15)

        vuln_m = min(1.0, makur_v * 0.7 + (1.0 - nige_pct1) * 0.3)
        score  = attack_p * vuln_m * st_adv_v1 * v_makuri_n * outer_mult * inner_block
        return {
            "total_score":      min(score, 1.0),
            "attack_type":      "まくり",
            "attack_power":     attack_p,
            "w1_vulnerability": vuln_m,
            "st_advantage":     st_adv_v1,
            "venue_affinity":   v_makuri_n,
            "breakdown":        {
                "まくり%": maku_pct, "被脆弱性": round(vuln_m,3),
                "外枠補正": outer_mult, "内壁": round(inner_block,3),
            },
        }


def _calc_w1_escape_score(r1, results, venue_stats, race_judgment=None):
    """
    1号艇が「今日このメンバーを相手に逃げ切れる確率」を定量化する。

    【v2: 展開エンジン連動】
    race_judgment の escape_score（逃げ定性スコア）と tobi_score（飛び定性スコア）を
    直接参照する。展開エンジンが「飛び優勢」と判定しているのに
    1号艇◎になるという矛盾を根本から解消する。

    計算式:
      逃げ有効スコア
        = 選手固有の逃げ力（nige_pct × ST安定性）   … 過去実績
        × 今日の展開的な逃げ支持度（ryotate補正）    … 展開エンジン判定
        × 攻撃艇脅威ペナルティ                       … メンバー相互作用

    ryotate補正の設計:
      escape_score > tobi_score + 15 → 逃げ明確優勢 → 補正なし（×1.0）
      escape_score ≒ tobi_score（差±15以内）→ 拮抗  → 軽い抑制（×0.75）
      tobi_score > escape_score + 15  → 飛び優勢    → 強い抑制（×0.45）
      tobi_score > escape_score + 30  → 飛び圧倒    → 最大抑制（×0.25）

    Returns
    -------
    float: 逃げ有効スコア（0〜1）
    """
    cm1      = r1.get("raw_cm", {})
    nige_pct = safe_float(cm1.get("逃げ%"), 0.6) or 0.6
    st1      = r1.get("avg_st")

    # ── 展開エンジン（ryotate）の逃げ/飛びスコアを参照 ──────────────────
    rj = race_judgment or {}
    ryotate = rj.get("ryotate", {}) or {}
    esc_s = float(ryotate.get("escape_score", 50) or 50)   # 逃げ定性スコア（0〜100）
    tob_s = float(ryotate.get("tobi_score",   30) or 30)   # 飛び定性スコア（0〜100）
    diff  = esc_s - tob_s   # 正=逃げ優勢、負=飛び優勢

    if diff >= 15:
        ryotate_mult = 1.00   # 逃げ明確優勢
    elif diff >= -15:
        ryotate_mult = 0.75   # 拮抗（両建て領域）
    elif diff >= -30:
        ryotate_mult = 0.45   # 飛び優勢（画像のケース: esc=23, tob=35 → diff=-12 → 0.75）
    else:
        ryotate_mult = 0.25   # 飛び圧倒

    # ── 攻撃艇脅威ペナルティ（メンバー相互作用）───────────────────────
    threats = []
    for r in results:
        if r["waku"] == "1":
            continue
        eff = _calc_attack_effectiveness(r, cm1, venue_stats, results)
        threats.append(eff["total_score"])

    if threats:
        max_threat = max(threats)
        avg_threat = sum(threats) / len(threats)
        threat_penalty = max(0.4, 1.0 - max_threat * 0.5 - avg_threat * 0.15)
    else:
        threat_penalty = 1.0

    # ── ST安定性 ────────────────────────────────────────────────────────
    fly_label  = r1.get("fly_label", "低")
    fly_mult   = {"低": 1.0, "中": 0.88, "高": 0.72}.get(fly_label, 1.0)
    st_vals    = [r["avg_st"] for r in results if r.get("avg_st") is not None]
    if st1 is not None and st_vals:
        st_mean    = sum(st_vals) / len(st_vals)
        st_penalty = max(0.7, min(1.15, 1.0 - (st1 - st_mean) / 0.05 * 0.15))
    else:
        st_penalty = 1.0

    escape_score = nige_pct * ryotate_mult * threat_penalty * fly_mult * st_penalty
    return min(escape_score, 1.0)


def _judge_w1_escape(results, venue_stats, race_judgment=None):
    """
    ❶ 1号艇逃げ力判定 ── 6人構成メンバーを相手に逃げ切れるか

    【設計思想】
    「1号艇単体の逃げ実績」ではなく、
    「このメンバー6人の構成の中で、1号艇が逃げ切れる確率」を評価する。

    循環依存を排除するため race_judgment（ryotate/escape_score）は参照しない。
    本関数の出力が後続の ryotate 判定・印スコアの入力になる一方向の流れ。

    計算ステップ:
      Step1: 1号艇の固有逃げ力
             = nige_pct（コース別逃げ決まり手%）× ST安定性 × FLYリスク
      Step2: 2〜6号艇それぞれの対1号艇攻撃力（_calc_attack_effectiveness）
             → 最大脅威艇と攻撃タイプを特定
      Step3: 会場イン逃げ率による補正
             = 逃げやすい会場ならボーナス、荒れ会場ならペナルティ
      Step4: 総合逃げ確率（0〜1）と根拠テキストを出力

    Returns
    -------
    dict:
        escape_prob      : 逃げ確率（0〜1）
        escape_pct       : 逃げ確率%表示（例: "62.3%"）
        escape_rank      : "高"(>=0.60) / "中"(>=0.40) / "低"(<0.40)
        nige_power       : 1号艇固有逃げ力（0〜1）
        top_threat_waku  : 最大脅威艇の艇番
        top_threat_type  : 最大脅威の攻撃タイプ（"差し"/"まくり"等）
        top_threat_score : 最大脅威スコア（0〜1）
        threat_list      : 全攻撃艇の脅威スコア一覧 [(waku, type, score), ...]
        venue_modifier   : 会場補正係数
        reason           : 根拠テキスト（コンソール・新聞出力用）
    """
    r1 = next((r for r in results if r["waku"] == "1"), None)
    if r1 is None:
        return {
            "escape_prob": 0.5, "escape_pct": "50.0%", "escape_rank": "中",
            "nige_power": 0.5, "top_threat_waku": None, "top_threat_type": "-",
            "top_threat_score": 0.0, "threat_list": [], "venue_modifier": 1.0,
            "reason": "1号艇データなし",
        }

    cm1      = r1.get("raw_cm", {})
    nige_pct = safe_float(cm1.get("逃げ%"), 0.6) or 0.6
    st1      = r1.get("avg_st")
    fly_label = r1.get("fly_label", "低")
    fly_days  = safe_float((r1.get("raw_pm") or {}).get("FLY経過日数"))

    # ── Step1: 1号艇固有逃げ力 ─────────────────────────────────────────
    # FLYリスク補正（平均STへの影響）
    fly_mult = {"低": 1.0, "中": 0.88, "高": 0.72}.get(fly_label, 1.0)

    # 【⑦改善】FLY明けST分散拡大補正
    # FLY明け選手はSTが遅くなるだけでなく、ばらつきが大きくなる（読めない）。
    # これを「逃げ確率の信頼区間幅拡大」として表現:
    #   fly_uncertainty = ST分散拡大率（0.0〜0.20）
    #   fly_uncert_penalty = uncertainty × 0.5 として逃げ確率から減算
    # 根拠: ばらつきが大きいほど「たまに速いが信頼できない」状態を確率に反映
    # fly_daysが短いほど分散が大きい（出場停止明け直後が最大）
    if fly_label == "高":
        if fly_days is not None and fly_days < 90:
            fly_uncertainty = 0.20   # 出場停止明け直後: 最大分散
        else:
            fly_uncertainty = 0.14   # FLY複数回・日数不明
    elif fly_label == "中":
        if fly_days is not None and fly_days < 90:
            fly_uncertainty = 0.10
        else:
            fly_uncertainty = 0.06
    else:
        fly_uncertainty = 0.0        # FLYなし: 分散拡大なし

    # ST安定性補正（全艇平均より遅いほど逃げ力低下）
    st_vals = [r["avg_st"] for r in results if r.get("avg_st") is not None]
    if st1 is not None and st_vals:
        st_mean    = sum(st_vals) / len(st_vals)
        st_penalty = max(0.70, min(1.15, 1.0 - (st1 - st_mean) / 0.05 * 0.15))
    else:
        st_penalty = 1.0

    nige_power = min(nige_pct * fly_mult * st_penalty, 1.0)

    # ── Step2: 2〜6号艇の対1号艇攻撃力 ────────────────────────────────
    threat_list = []
    for r in results:
        if r["waku"] == "1":
            continue
        eff = _calc_attack_effectiveness(r, cm1, venue_stats, results)
        threat_list.append((
            r["waku"],
            eff["attack_type"],
            round(eff["total_score"], 4),
        ))
    threat_list.sort(key=lambda x: x[2], reverse=True)

    if threat_list:
        top_threat_waku  = threat_list[0][0]
        top_threat_type  = threat_list[0][1]
        top_threat_score = threat_list[0][2]
        avg_threat       = sum(s for _, _, s in threat_list) / len(threat_list)
        # 脅威ペナルティ: 最大脅威×0.50 + 平均脅威×0.15（下限0.40）
        threat_penalty = max(0.40, 1.0 - top_threat_score * 0.50 - avg_threat * 0.15)
    else:
        top_threat_waku  = None
        top_threat_type  = "-"
        top_threat_score = 0.0
        threat_penalty   = 1.0

    # ── Step3: 会場イン逃げ率補正 ───────────────────────────────────────
    in_rate = safe_float(venue_stats.get("in_rate"))
    if in_rate is not None:
        # 全国平均0.555を基準に補正（上限1.15、下限0.75）
        venue_modifier = max(0.75, min(1.15, in_rate / 0.555))
    else:
        venue_modifier = 1.0

    # ── Step4: 総合逃げ確率 ────────────────────────────────────────────
    escape_prob_raw = min(nige_power * threat_penalty * venue_modifier, 1.0)

    # 【⑦】FLY明け分散拡大ペナルティを逃げ確率に適用
    # ばらつきが大きい = 「高い時もあるが低い時もある」 → 期待値を下方修正
    fly_uncert_penalty = fly_uncertainty * 0.5
    escape_prob = max(0.0, min(1.0, escape_prob_raw - fly_uncert_penalty))

    # ── 逃げランク ──────────────────────────────────────────────────────
    if escape_prob >= 0.60:
        escape_rank = "高"
    elif escape_prob >= 0.40:
        escape_rank = "中"
    else:
        escape_rank = "低"

    # ── 根拠テキスト生成 ────────────────────────────────────────────────
    reason_parts = []
    reason_parts.append(
        f"1号艇逃げ力: nige%={nige_pct*100:.0f}% "
        f"× FLY({fly_label})={fly_mult:.2f} "
        f"× ST={st_penalty:.2f} → 固有逃げ力{nige_power*100:.0f}%"
    )
    if fly_uncertainty > 0:
        reason_parts.append(
            f"FLY明けST分散拡大: 不確実度±{fly_uncertainty*100:.0f}% "
            f"→ 逃げ確率を{fly_uncert_penalty*100:.1f}%下方修正"
        )
    if threat_list:
        top_w, top_t, top_s = threat_list[0]
        reason_parts.append(
            f"最大脅威: {top_w}号艇（{top_t}） 脅威スコア{top_s*100:.0f}% "
            f"→ 脅威ペナルティ×{threat_penalty:.2f}"
        )
        if len(threat_list) >= 2:
            w2, t2, s2 = threat_list[1]
            reason_parts.append(f"次点脅威: {w2}号艇（{t2}） {s2*100:.0f}%")
    if in_rate is not None:
        reason_parts.append(
            f"会場イン逃げ率{in_rate*100:.1f}% → 会場補正×{venue_modifier:.2f}"
        )
    reason_parts.append(
        f"→ 総合逃げ確率 {escape_prob*100:.1f}%【{escape_rank}】"
        + (f"（FLY分散補正前:{escape_prob_raw*100:.1f}%）" if fly_uncertainty > 0 else "")
    )

    return {
        "escape_prob":      round(escape_prob, 4),
        "escape_pct":       f"{escape_prob*100:.1f}%",
        "escape_rank":      escape_rank,
        "nige_power":       round(nige_power, 4),
        "top_threat_waku":  top_threat_waku,
        "top_threat_type":  top_threat_type,
        "top_threat_score": top_threat_score,
        "threat_list":      threat_list,
        "venue_modifier":   round(venue_modifier, 4),
        "reason":           " / ".join(reason_parts),
    }


def _judge_main_player(results, venue_stats, race_judgment):
    """
    ❷ 主役候補判定 ── 逃げない場合に誰が主役で、どの展開になるか

    【設計思想】
    「1号艇が逃げられない」と判断された場合、2〜6号艇の中から
    「このメンバー構成で最も主導権を握れる艇」を最大2艇選び、
    展開タイプ（差し/まくり/まくり差し）と
    そのときの2・3着候補を合わせて出力する。

    入力として _judge_w1_escape の threat_list（脅威スコア順）を活用し、
    さらに first_prob_map（確率モデル）と合成して最終的な主役順位を決める。

    ステップ:
      Step1: threat_list（攻撃有効性スコア）から上位艇を抽出
      Step2: _calc_attack_effectiveness の attack_type で展開タイプを確定
      Step3: 主役展開時の2着候補 → circle_pct（イン逃げ時2着率）を
             「主役が1着のときの内側残存補正」に読み替えて選出
      Step4: 主役展開時の3着候補 → idx3（3着残存指数）上位から選出
      Step5: 根拠テキストを生成してコンソール出力

    Returns
    -------
    dict:
        main_waku        : 主役候補1位の艇番
        main_type        : 展開タイプ（"差し"/"まくり"/"まくり差し"）
        main_score       : 主役スコア（0〜1）
        sub_waku         : 主役候補2位の艇番（Noneの場合あり）
        sub_type         : 候補2位の展開タイプ
        sub_score        : 候補2位スコア（0〜1）
        place2_candidates: 主役展開時の2着候補 [(waku, score), ...] 上位3艇
        place3_candidates: 主役展開時の3着候補 [(waku, score), ...] 上位3艇
        reason           : 根拠テキスト（コンソール・新聞出力用）
    """
    # ── w1_escape の threat_list を取得 ────────────────────────────────
    w1_escape   = race_judgment.get("w1_escape", {})
    threat_list = w1_escape.get("threat_list", [])  # [(waku, type, score), ...]

    # threat_list が空なら直接計算
    if not threat_list:
        r1 = next((r for r in results if r["waku"] == "1"), None)
        w1_cm = r1.get("raw_cm", {}) if r1 else {}
        for r in results:
            if r["waku"] == "1":
                continue
            eff = _calc_attack_effectiveness(r, w1_cm, venue_stats, results)
            threat_list.append((r["waku"], eff["attack_type"], round(eff["total_score"], 4)))
        threat_list.sort(key=lambda x: x[2], reverse=True)

    if not threat_list:
        return {
            "main_waku": None, "main_type": "-", "main_score": 0.0,
            "sub_waku": None,  "sub_type":  "-", "sub_score":  0.0,
            "place2_candidates": [], "place3_candidates": [],
            "reason": "攻撃艇データなし",
        }

    # ── Step1: 主役候補1・2位を確定 ────────────────────────────────────
    main_waku  = threat_list[0][0]
    main_type  = threat_list[0][1]
    main_score = threat_list[0][2]
    sub_waku   = threat_list[1][0] if len(threat_list) >= 2 else None
    sub_type   = threat_list[1][1] if len(threat_list) >= 2 else "-"
    sub_score  = threat_list[1][2] if len(threat_list) >= 2 else 0.0

    # ── Step2: 主役展開時の2着候補 ──────────────────────────────────────
    # 「主役が1着に来るとき、2着に残りやすい艇」
    # = 主役艇・1号艇以外で circle_pct（内側残存補正）が高い艇
    # ただし主役が内枠（差し・まくり差し）なら1号艇の2着残りも候補に入れる
    place2_raw = []
    for r in results:
        w = r["waku"]
        if w == main_waku:
            continue
        # 主役が外から来るとき（まくり）は内側が2着に残りやすい
        # 主役が差しのとき（2枠）は1号艇が2着残りしやすい
        circ = r.get("circle_pct") or 0.0
        win3 = r.get("win3_rate") or 0.5
        # 1号艇は差し・まくり差し展開で2着残りしやすいのでボーナス
        if w == "1" and main_type in ("差し", "まくり差し"):
            score_2nd = win3 * 0.7 + circ * 0.01  # 1号艇は3連対率主体で評価
        else:
            score_2nd = circ * 0.6 + win3 * 100 * 0.4  # circle_pct(0-100)とwin3(0-1)をスケール統一
        place2_raw.append((w, round(score_2nd, 2)))

    place2_candidates = sorted(place2_raw, key=lambda x: x[1], reverse=True)[:3]

    # ── Step3: 主役展開時の3着候補 ──────────────────────────────────────
    # 主役・1着候補（主役＋2着最有力）以外で idx3 が高い艇
    top2_wakus = {main_waku, place2_candidates[0][0]} if place2_candidates else {main_waku}
    place3_raw = []
    for r in results:
        w = r["waku"]
        if w in top2_wakus:
            continue
        idx3 = r.get("idx3") or 0
        win3 = r.get("win3_rate") or 0.5
        score_3rd = idx3 * 0.65 + win3 * 100 * 0.35
        place3_raw.append((w, round(score_3rd, 2)))

    place3_candidates = sorted(place3_raw, key=lambda x: x[1], reverse=True)[:3]

    # ── Step4: 根拠テキスト生成 ─────────────────────────────────────────
    p2_str = "  ".join(
        f"{w}号({s:.0f})" for w, s in place2_candidates
    )
    p3_str = "  ".join(
        f"{w}号({s:.0f})" for w, s in place3_candidates
    )

    reason_parts = [
        f"主役候補①: {main_waku}号艇【{main_type}】 攻撃スコア{main_score*100:.0f}%",
    ]
    if sub_waku:
        reason_parts.append(
            f"主役候補②: {sub_waku}号艇【{sub_type}】 攻撃スコア{sub_score*100:.0f}%"
        )
    reason_parts.append(f"→ 2着候補: {p2_str}")
    reason_parts.append(f"→ 3着候補: {p3_str}")

    return {
        "main_waku":         main_waku,
        "main_type":         main_type,
        "main_score":        round(main_score, 4),
        "sub_waku":          sub_waku,
        "sub_type":          sub_type,
        "sub_score":         round(sub_score, 4),
        "place2_candidates": place2_candidates,
        "place3_candidates": place3_candidates,
        "reason":            " / ".join(reason_parts),
    }


def _judge_escape_fallback(results, venue_stats, race_judgment, conflict_map=None):
    """
    ❸ 主役が来れなかった時の逃げ残存確率

    【設計思想】
    主役候補（main_player.main_waku）が自滅・失速した場合に、
    1号艇が2着以内に「生き残れるか」を数値化する。

    現状の SC シナリオには「1号艇が2着に残れる確率」が明示的に出ないため、
    本関数でそれを算出し買い目の「逃げ残存フォロー」判断に使う。

    計算ステップ:
      Step1: 1号艇の逃げ力ベース（w1_escape.escape_prob）を取得
      Step2: 主役の自滅タイプ（sc_fly_type）に応じた1号艇残存補正
             まくり系自滅 → コースが開く → 1号艇残存ボーナス（×1.30）
             差し系自滅   → 蓋の影響あり → 1号艇残存ペナルティ（×0.75）
      Step3: 漁夫（collapse_beneficiary）の上位艇が外から来ると1号艇の
             2着スペースが圧迫される → 上位受益艇数に応じた追加ペナルティ
      Step4: 最終的な逃げ残存確率（0〜1）と根拠テキストを出力

    Returns
    -------
    dict:
        fallback_prob   : 主役自滅時の1号艇2着以内残存確率（0〜1）
        fallback_pct    : fallback_prob の%表示
        fallback_rank   : "高"(>=0.55) / "中"(>=0.35) / "低"(<0.35)
        fly_type        : 主役の自滅タイプ（"まくり系"/"差し系"/"不明"）
        pressure_wakus  : 1号艇を圧迫する上位受益艇リスト [(waku, score), ...]
        reason          : 根拠テキスト
    """
    w1_escape   = race_judgment.get("w1_escape", {}) or {}
    main_player = race_judgment.get("main_player", {}) or {}
    # conflict_map はキーワード引数優先、なければ race_judgment から取得
    conflict_map = conflict_map or race_judgment.get("conflict_map", {}) or {}

    escape_prob = float(w1_escape.get("escape_prob", 0.5) or 0.5)
    main_waku   = main_player.get("main_waku")
    main_type   = main_player.get("main_type", "-")

    # ── Step2: 主役の自滅タイプを展開タイプから推定 ──────────────────────────
    # main_type（差し/まくり/まくり差し）を _calc_sc_weight の sc_fly_type に対応させる
    if main_type in ("まくり", "まくり差し"):
        fly_type       = "まくり系"
        type_modifier  = 1.30   # コースが開く → 1号艇が取り戻しやすい
    elif main_type == "差し":
        fly_type       = "差し系"
        type_modifier  = 0.75   # 差し自滅では蓋の影響が1号艇に及ぶ
    else:
        fly_type       = "不明"
        type_modifier  = 1.00

    base_prob = min(escape_prob * type_modifier, 1.0)

    # ── Step3: 漁夫受益艇による圧迫ペナルティ ────────────────────────────────
    # collapse_beneficiary は主軸当事者を除いた外側の艇 → これらが来ると
    # 1号艇の2着スペースがさらに圧迫される
    collapse_bene = conflict_map.get("collapse_beneficiary", []) or []

    # 主役艇自身は受益者リストから除外（主役が自滅した場合の話なので）
    pressure_wakus = [
        (w, s) for w, s in collapse_bene
        if w != main_waku and w != "1"
    ][:3]

    # 上位受益艇のスコア合計に応じてペナルティ加算（最大 -0.15）
    top_pressure_sum = sum(s for _, s in pressure_wakus[:2])
    pressure_penalty = min(top_pressure_sum * 0.20, 0.15)

    fallback_prob = max(0.0, min(base_prob - pressure_penalty, 1.0))

    # ── ランク付け ────────────────────────────────────────────────────────────
    if fallback_prob >= 0.55:
        fallback_rank = "高"
    elif fallback_prob >= 0.35:
        fallback_rank = "中"
    else:
        fallback_rank = "低"

    # ── 根拠テキスト生成 ──────────────────────────────────────────────────────
    reason_parts = [
        f"逃げ力ベース={escape_prob*100:.0f}%"
        f" × {fly_type}補正×{type_modifier:.2f} → {base_prob*100:.0f}%",
    ]
    if pressure_wakus:
        pw_str = "  ".join(f"{w}号({s:.2f})" for w, s in pressure_wakus)
        reason_parts.append(
            f"圧迫受益艇: {pw_str} → 圧迫ペナルティ-{pressure_penalty*100:.0f}%"
        )
    reason_parts.append(
        f"→ 逃げ残存確率 {fallback_prob*100:.1f}%【{fallback_rank}】"
    )

    return {
        "fallback_prob":  round(fallback_prob, 4),
        "fallback_pct":   f"{fallback_prob*100:.1f}%",
        "fallback_rank":  fallback_rank,
        "fly_type":       fly_type,
        "pressure_wakus": pressure_wakus,
        "reason":         " / ".join(reason_parts),
    }


def _judge_dark_horse(results, venue_stats, race_judgment, conflict_map=None):
    """
    ❹ 主役展開の穴をつく艇（ダークホース）判定

    【設計思想】
    主役展開（main_player.main_waku が1着に来る）が成立するとき、
    主軸の対立構造（main_waku vs 1号艇）の「外側」で美味しいポジションに
    入れる艇を特定する。

    既に計算済みの collapse_beneficiary（_build_conflict_map の出力）を活用し、
    さらに以下3条件でフィルタリングして「本物の穴」を絞り込む：
      条件①: 主軸対立の外側にいる（main_waku でも "1" でもない）
      条件②: win3_rate が高い（地力がある = 荒れても残れる）
      条件③: 決まり手が抜き・逃げ系（受動型 = 自滅しない）

    計算ステップ:
      Step1: collapse_beneficiary から主軸当事者を除いた外側艇リストを取得
      Step2: 各艇に対して「抜き系受動スコア」を計算
             = win3_rate × 受動性 × 漁夫スコア（collapse_beneficiaryスコア）
      Step3: スコア上位3艇を dark_horse_candidates として返す
      Step4: 根拠テキスト生成

    Returns
    -------
    dict:
        dark_horse_candidates : [(waku, score, reason_tag), ...] 上位3艇
        top_waku              : 最有力ダークホース艇番（Noneの場合あり）
        top_score             : 最有力艇のスコア（0〜1）
        is_valid              : True = 有効なダークホース候補あり
        reason                : 根拠テキスト
    """
    main_player  = race_judgment.get("main_player", {}) or {}
    # conflict_map はキーワード引数優先、なければ race_judgment から取得
    conflict_map = conflict_map or race_judgment.get("conflict_map", {}) or {}

    main_waku     = main_player.get("main_waku")
    collapse_bene = conflict_map.get("collapse_beneficiary", []) or []

    if not collapse_bene:
        return {
            "dark_horse_candidates": [],
            "top_waku":   None,
            "top_score":  0.0,
            "is_valid":   False,
            "reason":     "潰れ受益候補データなし",
        }

    # results を waku をキーにした辞書に変換
    results_map = {r["waku"]: r for r in results}

    # ── Step1: 外側艇リスト（主軸当事者を除外） ───────────────────────────────
    outer_bene = [
        (w, s) for w, s in collapse_bene
        if w != main_waku and w != "1"
    ]

    if not outer_bene:
        return {
            "dark_horse_candidates": [],
            "top_waku":   None,
            "top_score":  0.0,
            "is_valid":   False,
            "reason":     "主軸外側の受益候補なし",
        }

    # ── Step2: 各艇の「穴スコア」計算 ─────────────────────────────────────────
    def _safe_pct(cm, key):
        v = cm.get(key) if cm else None
        try:
            return max(float(v), 0.0) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    dark_horse_raw = []
    for w, bene_score in outer_bene:
        r = results_map.get(w, {})
        cm = r.get("raw_cm", {}) or {}

        win3  = r.get("win3_rate") or 0.5
        # 受動性スコア: 抜き・逃げ系の決まり手比率が高いほど受動型
        nuki_pct  = _safe_pct(cm, "抜き%")
        nige_pct  = _safe_pct(cm, "逃げ%")
        mak_pct   = _safe_pct(cm, "まくり%") + _safe_pct(cm, "まくり差し%")
        sash_pct  = _safe_pct(cm, "差し%")

        passive_score = (nuki_pct + nige_pct * 0.5) / 100.0
        attack_score  = (mak_pct + sash_pct) / 100.0
        # 受動性 = 攻撃性の低さ（仕掛けに行かない → 自滅しない）
        passivity = max(0.0, 1.0 - attack_score * 0.7) + passive_score * 0.3

        # 穴スコア = win3 × 受動性 × 漁夫スコア（collapse_beneficiaryスコア）
        dark_score = round(win3 * passivity * (bene_score + 0.1), 4)

        # 根拠タグ（印コメント用）
        if nuki_pct >= 15.0:
            tag = "抜き系"
        elif win3 >= 0.65:
            tag = "高win3"
        elif passivity >= 0.7:
            tag = "受動型"
        else:
            tag = "漁夫"

        dark_horse_raw.append((w, dark_score, tag))

    dark_horse_raw.sort(key=lambda x: x[1], reverse=True)
    dark_horse_candidates = dark_horse_raw[:3]

    # ── Step3: 有効判定 ───────────────────────────────────────────────────────
    # is_valid 閾値を venue_c1_win_rate 連動に変更（④⑤対応）
    # race_judgment に格納済みの venue_c1_win_rate を参照する。
    # 荒れやすい会場（戸田=0.43等） → 閾値↓（穴が出やすい）
    # 堅い会場（住之江=0.65等）     → 閾値↑（本当に強い候補のみ）
    _vc1_dh          = float(race_judgment.get('venue_c1_win_rate') or 0.555)
    _vr_dh           = max(0.70, min(1.30, _vc1_dh / 0.555))
    _is_valid_thresh = round(0.15 * _vr_dh, 3)  # 基準閾値0.15を会場補正
    top_waku  = dark_horse_candidates[0][0] if dark_horse_candidates else None
    top_score = dark_horse_candidates[0][1] if dark_horse_candidates else 0.0
    is_valid  = top_score >= _is_valid_thresh

    # ── Step4: 根拠テキスト生成 ──────────────────────────────────────────────
    reason_parts = []
    if main_waku:
        reason_parts.append(f"主軸: {main_waku}号艇展開の外側を探索")
    for w, s, tag in dark_horse_candidates:
        reason_parts.append(f"{w}号艇【{tag}】 穴スコア{s*100:.0f}%")

    if not is_valid:
        reason_parts.append("→ 有効な穴候補なし（スコア不十分）")
    else:
        reason_parts.append(f"→ 最有力穴: {top_waku}号艇 スコア{top_score*100:.0f}%")

    return {
        "dark_horse_candidates": dark_horse_candidates,
        "top_waku":   top_waku,
        "top_score":  round(top_score, 4),
        "is_valid":   is_valid,
        "reason":     " / ".join(reason_parts),
    }


def _calc_honmei_score(r, tobi_prob_val, jizen_ev=None, first_prob_map=None, results_ctx=None,
                       venue_stats=None, race_judgment=None):
    """
    【6人相互作用モデル 印スコア v4】

    設計思想:
      「確率が高い艇を◎にする」でも「補正点が高い艇を◎にする」でもなく、
      「このレースのこのメンバー構成で最も主導権を握れる艇を◎にする」。

      Stage1: 相互作用スコア（0〜50pt）← 土台
        攻撃艇: _calc_attack_effectiveness による「この1号艇を崩せる力」
        1号艇:  _calc_w1_escape_score による「このメンバーに逃げられる力」
        → 確率モデル（first_prob_map）との加重平均でブレンド
          相互作用 × 0.55 + 確率シェア × 0.45

      Stage2: 状態補正（0〜30pt）← このレースの「今の状態」
        ② 機力（0〜10pt）: モーター2連率ランク
        ③ jizen（0〜14pt）: 事前評価（展示での状態確認）
        ④ ST安定（0〜6pt）: FLYリスク・出遅れ癖

      確率ガード:
        first_prob_map の確率最大艇と自艇の確率差が大きいとき、
        相互作用スコアが高くても◎になれる上限を設ける。
        ただし旧版より緩め（25pt差で発動）: 相互作用モデルが
        確率を覆す場合もあり得るため。

      合計最大: 80pt
    """
    results   = results_ctx or []
    waku_str  = str(r["waku"])
    waku_idx  = int(r["waku"]) - 1
    _venue_stats = venue_stats or {}

    EQUAL_SHARE = 1.0 / 6.0

    # ══════════════════════════════════════════════════════
    # Stage1: 相互作用スコア (0〜50pt)
    # ══════════════════════════════════════════════════════

    # 1号艇の raw_cm を取得（全艇の攻撃有効性計算に必要）
    res1 = next((x for x in results if x["waku"] == "1"), None)
    w1_cm = res1.get("raw_cm", {}) if res1 else {}

    # 相互作用スコア（0〜1）
    if waku_str == "1":
        if res1:
            interaction_raw = _calc_w1_escape_score(res1, results, _venue_stats,
                                                    race_judgment=race_judgment)
        else:
            interaction_raw = 0.5
    else:
        eff = _calc_attack_effectiveness(r, w1_cm, _venue_stats, results)
        interaction_raw = eff["total_score"]
        # 攻撃タイプを記録（展開考察のコメント用）
        r["_attack_type"] = eff["attack_type"]
        r["_attack_eff"]  = round(eff["total_score"], 4)

    # first_prob_map との加重ブレンド
    if first_prob_map:
        total_p = sum(first_prob_map.values()) or 1.0
        p_share = first_prob_map.get(waku_str, 0.0) / total_p
    else:
        total_r = sum((x.get("rel_win1") or 0) for x in results) or 1.0
        p_share = (r.get("rel_win1") or 0) / total_r

    # ── ryotate（展開エンジン）の逃げ/飛びスコアを取得 ──────────────────
    rj       = race_judgment or {}
    ryotate  = rj.get("ryotate", {}) or {}
    esc_s    = float(ryotate.get("escape_score", 50) or 50)
    tob_s    = float(ryotate.get("tobi_score",   30) or 30)
    ryo_diff = esc_s - tob_s   # 正=逃げ優勢、負=飛び優勢

    # ── Stage1スコアを ryotate で補正 ───────────────────────────────────
    # 【核心設計】
    # first_prob_map は「1号艇55%の統計的優位」を内包している。
    # 飛び優勢レースでも1号艇の確率シェアが高くなるバイアスがある。
    # このバイアスを ryotate_diff で直接打ち消す。
    #
    # 1号艇: 飛び優勢のとき → 逃げ力スコアを削る（interaction_rawが既に削られているが不十分）
    # 飛び艇（2〜6号艇）: 飛び優勢のとき → 攻撃有効性スコアにボーナスを加算
    if waku_str == "1":
        # 飛び優勢度に応じて1号艇の相互作用スコアをさらに抑制
        if ryo_diff >= 10:
            w1_adj = 1.0    # 逃げ優勢 → 抑制なし
        elif ryo_diff >= -10:
            w1_adj = 0.80   # 拮抗
        elif ryo_diff >= -25:
            w1_adj = 0.60   # 飛び優勢（画像のケース diff=-12 → 0.80）
        else:
            w1_adj = 0.40   # 飛び圧倒
        interaction_raw = interaction_raw * w1_adj

        # 確率ボーナスも飛び優勢時は削る（統計バイアスを打ち消す）
        p_share_adj = p_share * max(0.5, (ryo_diff + 50) / 100.0)
    else:
        # 飛び艇: 飛び優勢のとき攻撃有効性ボーナス
        if ryo_diff <= -10:
            # 飛び優勢レースで攻撃力がある艇を後押し
            tobi_bonus = min(abs(ryo_diff) / 100.0, 0.30)  # 最大+0.30
            interaction_raw = min(interaction_raw * (1.0 + tobi_bonus), 1.0)
        p_share_adj = p_share

    blended = interaction_raw * 0.55 + p_share_adj * 0.45
    excess  = p_share_adj - EQUAL_SHARE
    prob_bonus = max(0.0, excess) * 20.0
    pt_stage1  = min(blended * 50 + prob_bonus, 50.0)

    # ══════════════════════════════════════════════════════
    # Stage2: 状態補正 (0〜30pt)
    # ══════════════════════════════════════════════════════

    # ── ② 機力点 (0〜10pt) ──────────────────────────────────────────────
    motor_vals = []
    for x in results:
        try:
            mv = float(x.get("motor2") or 0)
            if mv > 0:
                motor_vals.append((x["waku"], mv))
        except (ValueError, TypeError):
            pass
    if motor_vals:
        sorted_motors = sorted(motor_vals, key=lambda x: x[1], reverse=True)
        rank_map = {w: i for i, (w, _) in enumerate(sorted_motors)}
        my_rank  = rank_map.get(waku_str, len(sorted_motors) - 1)
        pt_motor = (1.0 - my_rank / max(len(sorted_motors) - 1, 1)) * 10.0
    else:
        pt_motor = 5.0

    # ── ③ jizen総合点 (0〜14pt) ─────────────────────────────────────────
    pt_jizen = 0.0
    if jizen_ev is not None:
        sym4 = {"◎": 3, "◎?": 2, "○": 2, "△": 1, "": 0, "-": 0}
        sym3 = {"A": 3, "B": 2, "C": 1, "D": 0, "E": 0, "-": 1}
        if waku_str == "1":
            in_sym  = (jizen_ev.get("in_nige")   or [""]   )[0]
            pt_jizen += sym4.get(in_sym,  0) / 3.0 * 4.0
        if waku_str != "1":
            ai_sym  = (jizen_ev.get("aisho")      or [""] * 6)[waku_idx]
            pt_jizen += sym4.get(ai_sym,  0) / 3.0 * 4.0
        ki_sym      = (jizen_ev.get("kiryoku")    or ["C"] * 6)[waku_idx]
        pt_jizen += sym3.get(ki_sym,  1) / 3.0 * 2.0
        jz_sym      = (jizen_ev.get("jizaisei")   or [""] * 6)[waku_idx]
        pt_jizen += sym4.get(jz_sym,  0) / 3.0 * 4.0
        if waku_str not in ("1", "2"):
            tk_sym  = (jizen_ev.get("tenkai")     or [""] * 6)[waku_idx]
            pt_jizen += sym4.get(tk_sym, 0) / 3.0 * 4.0
    pt_jizen = min(pt_jizen, 14.0)

    # ── ④ ST安定点 (0〜6pt) ─────────────────────────────────────────────
    all_sts = [(x["waku"], x.get("avg_st")) for x in results if x.get("avg_st") is not None]
    if len(all_sts) >= 2:
        sorted_sts  = sorted(all_sts, key=lambda x: x[1])
        st_rank_map = {w: i for i, (w, _) in enumerate(sorted_sts)}
        my_st_rank  = st_rank_map.get(waku_str, len(all_sts) - 1)
        pt_st       = (1.0 - my_st_rank / max(len(all_sts) - 1, 1)) * 6.0
    else:
        pt_st = 3.0

    # FLYリスク控除（状態補正からペナルティ）
    fly_pen = {"高": -15, "中": -7, "低": 0}.get(r.get("fly_label", "低"), 0)

    raw_score = pt_stage1 + pt_motor + pt_jizen + pt_st + fly_pen

    # ══════════════════════════════════════════════════════
    # 確率ガード（旧版より緩和: 差25pt以上で発動）
    # 相互作用モデルが確率を覆す正当な根拠がある場合を許容する
    # ══════════════════════════════════════════════════════
    if first_prob_map:
        max_p      = max(first_prob_map.values()) / (sum(first_prob_map.values()) or 1.0)
        max_excess = max_p - EQUAL_SHARE
        if max_excess >= 0:
            max_pt_stage1 = min(EQUAL_SHARE * 50 + max_excess * 2 * 50 + max_excess * 20, 50.0)
        else:
            max_pt_stage1 = max(0.0, max_p * 50)

        # 自艇の確率スコア相当値
        own_excess     = p_share - EQUAL_SHARE
        own_pt_stage1  = min(p_share * 50 + max(0.0, own_excess) * 20, 50.0)
        prob_gap       = max_pt_stage1 - own_pt_stage1

        if prob_gap >= 25:
            # 確率差が大きい → 相互作用が高くても上限キャップ
            cap = max_pt_stage1 + 30.0   # Stage2上限30ptまでは届ける
            raw_score = min(raw_score, cap)

    return raw_score


def _apply_jizen_honmei(results_ref, tobi_prob_val, jizen_ev, first_prob_map=None,
                        venue_stats=None, race_judgment=None):
    """
    【v7.0 攻め評価モデル】

    1号艇: _calc_w1_escape_score の結果で逃げ評価専用記号を付与。
           逃◎(>=0.55) / 逃○(>=0.40) / 逃△(>=0.25) / 逃×(<0.25)
           攻め記号（◎○▲△）の対象外とする。

    2〜6号艇: _calc_honmei_score（攻撃有効性スコア）で◎○▲△を付与。
              「この1号艇に対して攻め切れる力のランク」が記号の意味。

    設計意図:
      旧版は全6艇を同じ軸（6人相互作用スコア）で順位付けしていたため、
      1号艇の◎（逃げ評価）と2〜6号艇の◎（攻め評価）が混在していた。
      分離することで「1号が逃げられるか」「誰が崩すか」を別軸で読める。
    """
    _venue_stats = venue_stats or {}

    # ── 1号艇: 逃げ評価専用記号を付与 ──────────────────────────────────
    res1 = next((r for r in results_ref if r["waku"] == "1"), None)
    if res1 is not None:
        escape_score = _calc_w1_escape_score(
            res1, results_ref, _venue_stats, race_judgment=race_judgment
        )
        if escape_score >= 0.55:
            nige_mark = "逃◎"
        elif escape_score >= 0.40:
            nige_mark = "逃○"
        elif escape_score >= 0.25:
            nige_mark = "逃△"
        else:
            nige_mark = "逃×"
        res1["honmei"]       = nige_mark
        res1["nige_mark"]    = nige_mark
        res1["escape_score"] = round(escape_score, 4)

    # ── 2〜6号艇: 攻め評価スコアで◎○▲△を付与 ──────────────────────
    attack_scores = []
    for i, r in enumerate(results_ref):
        if r["waku"] == "1":
            continue
        if r.get("rel_win1") is None:
            continue
        atk_score = _calc_honmei_score(
            r, tobi_prob_val,
            jizen_ev=jizen_ev,
            first_prob_map=first_prob_map,
            results_ctx=results_ref,
            venue_stats=_venue_stats,
            race_judgment=race_judgment,
        )
        attack_scores.append((i, atk_score))

    attack_scores.sort(key=lambda x: x[1], reverse=True)
    _hmap = {0: "◎", 1: "○", 2: "▲", 3: "△"}

    for r in results_ref:
        if r["waku"] != "1":
            r["honmei"] = " "

    for rank, (idx, _) in enumerate(attack_scores[:4]):
        results_ref[idx]["honmei"] = _hmap[rank]


def _calc_venue_stats(venue_stats_master, venue):
    """会場イン逃げ率・決まり手場平均・コース別1着率を返す（会場統計シートから取得）"""
    vs = venue_stats_master.get(venue, {})
    in_rate  = safe_float(vs.get("イン逃げ率"))
    kimari_avg = {
        "差し":      safe_float(vs.get("差し率")),
        "まくり":    safe_float(vs.get("まくり率")),
        "まくり差し": safe_float(vs.get("まくり差し率")),
    }
    areyasusa = round((1.0 - float(in_rate)) * 100, 1) if in_rate is not None else None
    # コース別1着率（2C〜6C）: S2〜S4シナリオ重み補正に使用
    course_win_rates = {
        str(c): safe_float(vs.get(f"{c}C_1着率") or vs.get(f"{c}コース1着率"))
        for c in range(1, 7)
    }
    return {
        "in_rate":          in_rate,
        "kimari_avg":       kimari_avg,
        "areyasusa_score":  areyasusa,
        "course_win_rates": course_win_rates,
    }


# ============================================================
# 倶楽部流 事前評価 - メンバーデータ組み立て
# ============================================================
def build_jizen_members(results, course_master, player_master, motor_df, race_no):
    """
    calc_race_indicesの戻り値(results)から evaluate_all 用データを組み立てる。

    Parameters
    ----------
    results       : list[dict]  calc_race_indices の戻り値（艇番順）
    course_master : dict        {(選手名, コース文字列): {指数dict}}
    player_master : dict        {選手名: {指数dict}}
    motor_df      : pd.DataFrame or None  scrape_motor.py の出力
    race_no       : int or str

    Returns
    -------
    list[dict]  evaluate_all() に渡せる6要素リスト（インデックス0=1号艇）
    """
    def _s(val, default=0.0):
        """文字列/数値を float に変換。変換不能な場合は default を返す。
        【修正⑧】default=None の場合は 0.0 も "データなし" と同一視せず None として扱う。
        つまり:
          _s("0.0", 0.0)  → 0.0  （正常変換）
          _s(None, 0.0)   → 0.0  （デフォルト）
          _s(None, None)  → None （「データなし」を明示したい場合）
          _s("0.0", None) → 0.0  （正常変換: 0.0 は有効なデータ）
        """
        try:
            v = str(val).replace("%", "").strip()
            if v in ("", "None", "nan", "-", "★"):
                return default
            return float(v)
        except Exception:
            return default

    def _s_nullable(val):
        """float変換を試み、変換不能(欠損)は None を返す。0.0は有効値として区別する。"""
        try:
            v = str(val).replace("%", "").strip()
            if v in ("", "None", "nan", "-", "★"):
                return None
            return float(v)
        except Exception:
            return None

    # モーターデータをインデックス化 {艇番int: 2連対率float or None}
    motor_index = {}
    if motor_df is not None:
        race_motor = motor_df[motor_df["race_no"] == int(race_no)]
        for _, row in race_motor.iterrows():
            bn = int(row["boat_no"]) if pd.notna(row["boat_no"]) else 0
            rate = row["motor_2rate"] if pd.notna(row["motor_2rate"]) else None
            motor_index[bn] = rate

    # 1枠の逃げ率（相性計算用）
    res0 = results[0] if results else {}
    cm0 = res0.get("raw_cm", {})
    nige_rate_1 = _s(cm0.get("逃げ%"), 1.0)

    members = []
    for i, res in enumerate(results[:6]):
        waku = res.get("waku", str(i + 1))
        boat_no = int(waku) if str(waku).isdigit() else (i + 1)
        name = res.get("name_norm", res.get("name", "").replace("　","").replace(" ",""))
        course = str(res.get("course", str(i + 1))).strip()
        cm = res.get("raw_cm", {})
        pm = res.get("raw_pm", {})

        # ── イン逃げ ──
        # 1コースのイン1着率（選手指数マスタ）
        rate_1st_c1 = _s(pm.get("イン\n1着率") or pm.get("イン1着率"))
        # 1コースのST順位（選手指数マスタ）
        st_rank_raw = pm.get("ST順位\n(1コース)") or pm.get("ST順位(1コース)")
        st_rank_c1 = _s(st_rank_raw) if st_rank_raw not in (None, "", "None") else None
        star_rate = bool(cm.get("★1着率"))

        # ── 相性（自コースの攻め決まり手割合） ──
        sashi_pct  = _s(cm.get("差し%"))
        makuri_pct = _s(cm.get("まくり%"))
        mz_pct     = _s(cm.get("まくり差し%"))
        attack_rate = sashi_pct + makuri_pct + mz_pct  # 合計割合（0〜1）

        # ── 相性用: 自コースの平均ST（秒）── evaluate_jizen.calc_aisho に必要
        avg_st_self = res.get("avg_st")  # calc_race_indices で算出済み

        # ── 【修正】avg_st が None のとき ST順位から推定値でフォールバック ──────
        # 問題: コース別マスタにデータが少ない選手は avg_st = None になり
        #       _st_advantage_score が中立値0.5を返す → ST比較が完全に無効化される。
        #       2〜6号艇で全員Noneだと相性評価が全員空白になる（鳴門2R等で確認）。
        #
        # 解決: 選手指数マスタの「ST順位(Nコース)」を使って推定ST秒を計算する。
        #   ST順位 = 同レース内での速さ順位（1=最速〜6=最遅）
        #   推定式: avg_st_est = 0.12 + (rank - 1) / 5 × 0.08
        #     ST順位1.0 → 0.120秒（最速クラス）
        #     ST順位3.5 → 0.160秒（平均的）
        #     ST順位6.0 → 0.200秒（最遅クラス）
        #   これはコース別マスタ実測値より精度は落ちるが、
        #   「中立0.5固定」よりは大幅に正確なST比較が可能になる。
        if avg_st_self is None:
            _csr = res.get("course_st_rank")   # 進入コース別ST順位
            if _csr is not None:
                avg_st_self = round(0.12 + (_csr - 1) / 5 * 0.08, 4)

        # ── 相性用: 1号艇の被決まり手%（members[0] = 1号艇にのみ格納）──
        # 2号艇以降の相性評価で「この1号艇は差されやすいか/捲られやすいか」を参照する
        # 【修正⑧】_s_nullable を使って「差された回数ゼロ(0.0)」と「データなし(None)」を区別
        lose_sashi_rate  = None
        lose_makuri_rate = None
        lose_rate_reliable = False   # 【v6.4新設】C1敗戦数が十分あるか
        if i == 0:
            # cm0（1号艇のコース別マスタ）から被決まり手%を取得
            lose_sashi_rate  = _s_nullable(cm.get("差され%"))
            lose_makuri_rate = _s_nullable(cm.get("捲られ%"))
            # キー名ゆれ対応（0.0 は有効値なので None の場合のみ代替キーを試みる）
            if lose_sashi_rate is None:
                lose_sashi_rate  = _s_nullable(cm.get("差し被%") or cm.get("被差し%"))
            if lose_makuri_rate is None:
                lose_makuri_rate = _s_nullable(cm.get("まくり被%") or cm.get("被まくり%"))

            # ── 【v6.4新設】C1敗戦数による信頼度判定 ────────────────────────
            # update_master.py はC1敗戦数=「1コース出走で負けた回数」を集計している。
            # 件数が少ない（目安: 10件未満）場合、差され%・捲られ%はノイズが大きいため
            # evaluate_jizen 側で信頼度フラグを参照して重みを下げることができる。
            # C1敗戦数はコース別マスタの「C1敗戦数」列に格納されている（コース1行のみ有効）。
            c1_lose_cnt = safe_float(cm.get("C1敗戦数"))
            LOSE_RATE_MIN_SAMPLES = 10  # 信頼度ありとみなす最低件数
            lose_rate_reliable = (
                c1_lose_cnt is not None and c1_lose_cnt >= LOSE_RATE_MIN_SAMPLES
            )

        # ── 機力 ──
        motor_2rate = motor_index.get(boat_no)

        # ── 自在性（2〜6コースの多様性割合） ──
        course_rows = []
        for c in range(2, 7):
            ck = (name, str(c))  # name is already normalized
            crow = course_master.get(ck) or {}
            crow_copy = dict(crow)
            crow_copy["course"] = c
            crow_copy["1着数"]    = _s(crow.get("1着数"))
            crow_copy["差し(件)"]  = _s(crow.get("差し(件)"))
            crow_copy["まくり(件)"] = _s(crow.get("まくり(件)"))
            crow_copy["まくり差し%"] = _s(crow.get("まくり差し%"))
            course_rows.append(crow_copy)
        diversity_rate = calculate_diversity_rate(course_rows)
        jizaisei_rate  = _s(pm.get("自在性\n1着率") or pm.get("自在性1着率"))
        star_kimete    = bool(cm.get("★決手"))

        # ── 安定性評価用キー（evaluate_jizen.calc_jizaisei に渡す） ──
        # ST安定スコア（0〜100）
        st_stable_score = _s(
            pm.get("ST安定\nスコア") or pm.get("ST安定スコア"), default=None
        )
        # FLY数・FLY経過日数・出遅れ数・ST計測件数
        _fly_count_raw  = pm.get("FLY数")
        _fly_days_raw   = pm.get("FLY経過\n日数") or pm.get("FLY経過日数")
        _late_count_raw = pm.get("出遅れ数")
        _st_count_raw   = pm.get("ST\n計測件数") or pm.get("ST計測件数")
        fly_count_stab  = int(safe_float(_fly_count_raw,  0) or 0)
        fly_days_stab   = safe_float(_fly_days_raw) if _fly_days_raw not in (None, "", "nan") else None
        late_count_stab = int(safe_float(_late_count_raw, 0) or 0)
        st_count_stab   = int(safe_float(_st_count_raw,   1) or 1)

        # ── 展開（自コースの3連対率・まくり系） ──
        rate_3ren          = _s(cm.get("3連対率"))
        makuri_rate_t      = _s(cm.get("まくり%"))
        mz_rate_t          = _s(cm.get("まくり差し%"))

        members.append({
            # イン逃げ
            "rate_1st_c1":  rate_1st_c1,
            "st_rank_c1":   st_rank_c1,
            "star_rate":    star_rate,
            # 相性（v3対応: 攻め武器・ST・被決まり手）
            "nigé_rate":         nige_rate_1,
            "attack_rate":       attack_rate,
            "sashi_rate":        sashi_pct,         # 差し%（相性用・展開用共通）
            "makuri_rate":       makuri_pct,         # まくり%（相性用・展開用共通）
            "makuri_zashi_rate": mz_pct,             # まくり差し%（相性用・展開用共通）
            "avg_st_self":       avg_st_self,         # 【修正②】自コース平均ST秒（相性用）
            "lose_sashi_rate":   lose_sashi_rate,     # 【修正②】1号艇のみ: 差され%（相性用）
            "lose_makuri_rate":  lose_makuri_rate,    # 【修正②】1号艇のみ: 捲られ%（相性用）
            "lose_rate_reliable": lose_rate_reliable, # 【v6.4新設】C1敗戦数≥10ならTrue
            # 機力
            "motor_2rate":  motor_2rate,
            # 自在性（後方互換のため残存・evaluate_jizen側では未使用）
            "diversity_rate": diversity_rate,
            "jizaisei_rate":  jizaisei_rate,
            "star_kimete":    star_kimete,
            # 安定性評価用（evaluate_jizen.calc_jizaisei v4）
            "st_stable_score": st_stable_score,
            "fly_count":       fly_count_stab,
            "fly_days":        fly_days_stab,
            "late_count":      late_count_stab,
            "st_count":        st_count_stab,
            # 展開（makuri_rate / makuri_zashi_rate は相性用と同値のため共用）
            "rate_3ren":  rate_3ren,
            # 【v7追加】コース番号（evaluate_jizen._calc_weapon_score で使用）
            "course_int": int(course) if str(course).isdigit() else (i + 1),
        })

    return members

# ============================================================
# サンプルシートからレイアウトをコピーして新シートを作成
# ============================================================
def clone_sample_layout(wb):
    """出力_新聞サンプルのレイアウトを 出力_新聞 に完全コピー（行高・列幅・結合・色）

    ⚠️  DEPRECATED: main() は write_race_flat() を直接呼び出すためこの関数は未使用。
    削除候補。外部スクリプトから参照されていないことを確認後に除去すること。
    """
    ws_src = wb[SHEET_SAMPLE]
    
    # 既存の出力_新聞を削除して再作成
    if SHEET_OUTPUT in wb.sheetnames:
        del wb[SHEET_OUTPUT]
    
    # サンプルシートをコピーして名前変更
    ws_new = wb.copy_worksheet(ws_src)
    ws_new.title = SHEET_OUTPUT
    
    # シートをサンプルの直後に移動
    sample_idx = wb.sheetnames.index(SHEET_SAMPLE)
    wb.move_sheet(SHEET_OUTPUT, offset=-(len(wb.sheetnames) - sample_idx - 2))
    
    return ws_new

# ============================================================
# 1レース分をサンプルレイアウトに書き込む
# ============================================================
def write_race_to_sample_layout(ws, row_offset, race_no, venue, race_date,
                                  results, slit, venue_stats, frame_2nd, total_races,
                                  _jizen_members=None):
    """
    出力_新聞サンプルの29行レイアウトに合わせてデータ書き込み
    row_offset: このレースの開始行（1始まり）

    ⚠️  DEPRECATED: main() は write_race_flat() を使用。この関数は呼び出されていない。
    削除候補。外部スクリプトから参照されていないことを確認後に除去すること。
    """
    r = row_offset  # 行番号ショートカット

    # === Row 1: ヘッダ行（会場・レース番号・新聞名・日付） ===
    ws.cell(r+0, 1).value = venue
    ws.cell(r+0, 3).value = f"{race_no}R"
    ws.cell(r+0, 9).value = race_date
    # 「ボートリサーチ新聞」は結合セル I2:P3 → row2のI列（相対row+1, col=9）

    # === Row 2: イン逃げ場平均・決まり手場平均・スリット ===
    in_rate = venue_stats.get("in_rate")
    in_str = f"イン逃げ場平均: {in_rate*100:.1f}%" if in_rate is not None else "イン逃げ場平均: -"
    ws.cell(r+2, 3).value = in_str  # C3（相対）→ col=3
    ws.cell(r+5, 9).value = f"想定スリット: {slit}"  # スリットはrow3（header行の右側）

    # === Row 4: ヘッダ行（カラム名） - サンプルから継承されるので更新不要 ===
    # 「1号艇平均ST N.N位」を更新
    st1_rank = None
    for res in results:
        if res["waku"] == "1":
            st1_rank = res.get("st_rank")
            break
    if st1_rank:
        ws.cell(r+3, 11).value = f"1号艇平均ST\n{st1_rank:.1f}位"
    else:
        ws.cell(r+3, 11).value = "1号艇平均ST\n-"

    # === Row 5-10: 6選手データ ===
    for i, res in enumerate(results[:6]):
        data_row = r + 4 + i  # row5〜10（0-indexed: +4,+5,+6,+7,+8,+9）
        waku_no = int(res["waku"]) if res["waku"].isdigit() else (i+1)
        
        name_display = f"{res['honmei']} {res['name']}" if res["honmei"].strip() else f"  {res['name']}"
        rel_w1 = f"{res['rel_win1']:.1f}%" if res["rel_win1"] is not None else "-"
        abs_w3 = f"{res['abs_win3']:.1f}%" if res.get("abs_win3") is not None else "-"
        idx3   = res.get("idx3", 0)
        avg_st = f"{res['avg_st']:.3f}" if res.get("avg_st") is not None else "-"

        ws.cell(data_row, 2).value = name_display         # B: 選手名
        ws.cell(data_row, 5).value = res["kumi"]          # E: 組
        ws.cell(data_row, 6).value = res["motor2"]        # F: モータ2連
        ws.cell(data_row, 9).value = res["course"]        # I: 想定スリット（コース）
        ws.cell(data_row, 12).value = res["honmei"] if res["honmei"].strip() else None  # L: 本命記号
        ws.cell(data_row, 13).value = rel_w1              # M: ◆相対1着率
        ws.cell(data_row, 15).value = abs_w3              # O: コース別3連対率（絶対評価）
        
        # [修正5] 上記の二重書き込みブロックを削除（col9/12/13の上書きバグ修正）
        # H列: F/St影響（モーター由来・未実装のため空欄）
        # I列: 想定スリット内コース番号（上記col9=res["course"]で書込済）
        # L列: 評価記号（上記col12=res["honmei"]で書込済）
        # M列: オリジナル1着率（上記col13=rel_w1で書込済）

    # === Row 12: 事前評価セクションヘッダ - 継承 ===

    # === Row 14-19: 事前評価（5項目）+ 3着指数 ===
    # 3着指数は既存通り
    for i, res in enumerate(results[:6]):
        data_row = r + 13 + i
        idx3 = res.get("idx3", 0)
        ws.cell(data_row, 16).value = idx3 if idx3 else None  # P: 3着指数

    # 倶楽部流 事前評価（evaluate_jizen.py が利用可能な場合のみ）
    if JIZEN_AVAILABLE and _jizen_members:
        jizen_result = evaluate_all(_jizen_members)
        # 列マッピング: B=2(イン逃げ) C=3(相性) D=4(機力) E=5(自在性) F=6(展開)
        col_map = {
            "in_nige":  2,
            "aisho":    3,
            "kiryoku":  4,
            "jizaisei": 5,
            "tenkai":   6,
        }
        for key, col in col_map.items():
            for i, symbol in enumerate(jizen_result[key]):
                data_row = r + 13 + i
                ws.cell(data_row, col).value = symbol if symbol else None

    # === Row 21-28: 決まり手・場平均ブロック ===
    # G14:O19 の2着率テキストブロックを更新
    # （結合セルなので左上セルに書く）
    circle_lines = ["2着優位度（イン逃げ決着時・相対%）\n"]
    color_labels = {2: "黒", 3: "赤", 4: "青", 5: "黄", 6: "緑"}
    circle_pct_map2 = {res["waku"]: res.get("circle_pct") for res in results}
    for waku_no in range(2, 7):
        w = str(waku_no)
        pct = circle_pct_map2.get(w)
        label = color_labels.get(waku_no, "")
        pct_str = f"{pct:.0f}%" if pct is not None else "-"
        circle_lines.append(f"{label}({waku_no}枠): {pct_str:>4}")
    ws.cell(r+13, 7).value = "\n".join(circle_lines)

    # === 区切り行（Row 29）: 注意書き継承 ===
    # サンプルから自動引き継ぎ

# ============================================================
# ST順位 舟図生成（Pillowのみ使用・matplotlibは不要）
# ============================================================
def _make_st_boat_chart(results, player_master, outpath):
    """
    各選手の出走コース平均STをもとに、横位置でST順位を表現した舟図を生成。
    速い（ST値が小さい）ほど右側に配置。
    上から1号艇〜6号艇の固定順。
    Pillowのみ使用（matplotlibは不要）。
    """
    if not PIL_AVAILABLE:
        print("  ⚠️  Pillowが未インストールのためST舟図をスキップします。")
        return

    # 画像サイズ（EMU 865179 x 2082127 → 96dpi換算で約91x220px、220dpiで約204x490px）
    W, H = 204, 490

    # フォント
    font_candidates = [
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    fnt_path = next((f for f in font_candidates if os.path.exists(f)), None)
    fnt_size = max(int(H / 6 * 0.38), 10)
    fnt = ImageFont.truetype(fnt_path, fnt_size) if fnt_path else ImageFont.load_default()

    BOAT_STYLES = [
        ('#FFFFFF', '#333333', '#000000'),  # 1白
        ('#1a1a1a', '#555555', '#ffffff'),  # 2黒
        ('#CC0000', '#990000', '#ffffff'),  # 3赤
        ('#2255CC', '#1144AA', '#ffffff'),  # 4青
        ('#DDCC00', '#BBAA00', '#000000'),  # 5黄
        ('#115511', '#003300', '#ffffff'),  # 6緑
    ]

    # 各選手の出走コース平均STを取得
    st_col_keys = [
        '平均ST\n(1コース)', '平均ST\n(2コース)', '平均ST\n(3コース)',
        '平均ST\n(4コース)', '平均ST\n(5コース)', '平均ST\n(6コース)',
    ]
    boats = []
    for res in results[:6]:
        waku = int(res['waku']) if str(res['waku']).isdigit() else (len(boats) + 1)
        name = res.get('name_norm', res.get('name', '').replace('　', '').replace(' ', ''))
        course = str(res.get('course', str(waku))).strip()
        pm = player_master.get(name, {})
        st = None
        try:
            cidx = int(course) - 1
            if 0 <= cidx < 6:
                raw = pm.get(st_col_keys[cidx])
                if raw is not None:
                    v = str(raw).replace('%', '').strip()
                    if v not in ('', 'None', 'nan', '-'):
                        st = float(v)
        except Exception as e:
            print(f"  ⚠️  ST値取得エラー（{name} / コース{course}）: {e}")
        boats.append({'boat': waku, 'st': st})

    # ST値の範囲
    st_vals = [b['st'] for b in boats if b['st'] is not None]
    st_min = min(st_vals) if st_vals else 0.15
    st_max = max(st_vals) if st_vals else 0.20
    st_range = max(st_max - st_min, 0.001)

    # レイアウト
    slot_h = H / 6
    bh = slot_h * 0.55          # 舟の高さ
    bw = W * 0.52               # 舟の長さ
    x_fast = W - bw / 2 - 4    # 速い=右端
    x_slow = bw / 2 + 4        # 遅い=左端

    img = PILImage.new('RGB', (W, H), 'white')
    draw = ImageDraw.Draw(img)

    def hex2rgb(h):
        h = h.lstrip('#')
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

    def draw_boat_pil(draw, cx, cy, bw, bh, fill, outline, num, tc, fnt):
        pts = [
            (cx + bw * 0.50, cy),
            (cx + bw * 0.22, cy - bh * 0.45),
            (cx - bw * 0.32, cy - bh * 0.45),
            (cx - bw * 0.50, cy - bh * 0.20),
            (cx - bw * 0.50, cy + bh * 0.20),
            (cx - bw * 0.32, cy + bh * 0.45),
            (cx + bw * 0.22, cy + bh * 0.45),
        ]
        draw.polygon(pts, fill=hex2rgb(fill), outline=hex2rgb(outline))
        draw.text((cx - bw * 0.06, cy), str(num),
                  fill=hex2rgb(tc), font=fnt, anchor='mm')

    for i, b in enumerate(boats):
        fill, outline, tc = BOAT_STYLES[b['boat'] - 1]
        cy = (i + 0.5) * slot_h
        if b['st'] is not None:
            norm = (b['st'] - st_min) / st_range
            cx = x_fast - norm * (x_fast - x_slow)
        else:
            cx = (x_fast + x_slow) / 2
        draw_boat_pil(draw, cx, cy, bw, bh, fill, outline, b['boat'], tc, fnt)

    # 改善C: outpathがNoneの場合はBytesIOに保存してそのまま返す（ディスクI/O削減）
    if outpath is None:
        from io import BytesIO
        buf = BytesIO()
        img.save(buf, format='PNG', dpi=(220, 220))
        buf.seek(0)
        return buf
    img.save(outpath, dpi=(220, 220))



def write_race_flat(ws, row_offset, race_no, venue, race_date,
                    results, slit, venue_stats, frame_2nd, _jizen_members=None,
                    player_master=None, _tmp_image_paths=None, deadline=None,
                    race_judgment=None):
    """
    出力_新聞サンプルのレイアウトに完全準拠して1レース分を書き込む。
    29行ブロック（Row1-29）＋空行1行 = 30行/レース
    """
    r = row_offset
    c = center_align()
    lft = left_align()
    hf = header_fill()
    hfont = header_font()
    shf = subheader_fill()
    shfont = subheader_font()
    bdr = thin_border()

    def wc(row, col, val, **kwargs):
        write_cell(ws, row, col, val, **kwargs)

    def merge(r1, c1, r2, c2):
        try:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        except Exception as e:
            print(f"  ⚠️  セル結合スキップ（行{r1}列{c1}〜行{r2}列{c2}）: {e}")

    # ── 結合セル定義（サンプル完全再現） ──
    merge(r+0, 1, r+1, 2)    # A1:B2  会場名
    merge(r+0, 3, r+1, 3)    # C1:C2  イン逃げ場平均
    merge(r+0, 4, r+0, 6)    # D1:F1  決まり手場平均
    merge(r+0, 7, r+0, 8)    # G1:H1  波乱度
    merge(r+0, 9, r+0,15)    # I1:O1  発行日
    merge(r+1, 7, r+1, 8)    # G2:H2
    merge(r+1, 9, r+2,16)    # I2:P3  ボートリサーチ新聞
    merge(r+2, 1, r+2, 2)    # A3:B3  レース番号
    merge(r+2, 7, r+2, 8)    # G3:H3  想定スリット
    merge(r+3, 1, r+3, 4)    # A4:D4  選手名ヘッダ
    merge(r+3, 7, r+3, 8)    # G4:H4  F/St
    merge(r+3, 9, r+3,10)    # I4:J4  想定スリット
    merge(r+3,13, r+3,14)    # M4:N4  オリジナル1着率
    merge(r+3,15, r+3,16)    # O4:P4  一般戦3連対
    for i in range(6):       # 選手行 B:D結合, I:J結合（K列はST図用に開放）, M:N結合, O:P結合
        rr = r+4+i
        merge(rr, 2, rr, 4)
        merge(rr, 9, rr,10)
        merge(rr,13, rr,14)
        merge(rr,15, rr,16)
    merge(r+11, 1, r+11, 6)  # A12:F12 事前評価タイトル
    merge(r+11, 7, r+12,13)  # G12:M13 2着率テキスト
    merge(r+11,14, r+12,14)  # N12:N13 3着指数
    merge(r+11,15, r+12,16)  # O12:P13 オリジナル3連対
    for i in range(6):       # 事前評価 O:P結合
        merge(r+13+i,15, r+13+i,16)
    merge(r+20, 1, r+20, 5)  # A21:E21 決まり手タイトル
    merge(r+20, 6, r+27,11)  # F21:K28 説明テキスト
    merge(r+20,12, r+20,16)  # L21:P21 場平均タイトル
    for i in range(7):       # 場平均 M:N結合, O:P結合
        merge(r+21+i,13, r+21+i,14)
        merge(r+21+i,15, r+21+i,16)
    merge(r+28, 1, r+28,16)  # A29:P29 注記

    # ── Row 1-2: ヘッダ ──
    wc(r+0, 1, venue,  font=Font(name="Noto Sans CJK SC", size=12, bold=True), alignment=c)
    wc(r+0, 3, 'イン\n逃げ\n場平均', fill=hf, font=Font(name="ＭＳ ゴシック", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+0, 4, '決まり手場平均', fill=hf, font=hfont, alignment=c)
    # 波乱度セルに会場の荒れやすさスコアを表示
    areyasusa = venue_stats.get('areyasusa_score')
    haranddo_str = f'波乱度\n{areyasusa:.0f}pt' if areyasusa is not None else '波乱度'
    wc(r+0, 7, haranddo_str, fill=hf, font=hfont, alignment=c)
    in_rate = venue_stats.get('in_rate')
    wc(r+0, 9, race_date, fill=hf, font=Font(name="Meiryo UI", size=11, bold=True, color="FFFFFFFF"),
       alignment=Alignment(horizontal="right", vertical="center", wrap_text=True))
    wc(r+0,16, '発行', fill=hf, font=Font(name="Noto Sans CJK SC", size=11, bold=True, color="FFFFFFFF"), alignment=c)
    # ── タイトル「ボートリサーチ新聞」（I2:P3 結合セルの左上 = col9, row+1） ──
    wc(r+1, 9, 'ボートリサーチ新聞',
       font=Font(name="Noto Sans CJK SC", size=24, bold=True),
       alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

    # 決まり手場平均（会場統計から）
    kimari = venue_stats.get('kimari_avg', {})
    for col, key in [(4,'差し'), (5,'まくり'), (6,'まくり差し')]:
        wc(r+1, col, key, fill=hf, font=hfont, alignment=c)
        v = kimari.get(key)
        vs_str = f'{v*100:.0f}%' if v else '-'
        wc(r+2, col, vs_str, alignment=c)

    # イン逃げ場平均（C列, Row1-2結合）
    in_str = f'{in_rate*100:.1f}%' if in_rate is not None else '-'
    wc(r+2, 3, in_str, alignment=c)

    # ── Row 3: レース番号・スリット ──
    race_no_val = f'{race_no}R\n{deadline}' if deadline else f'{race_no}R'
    wc(r+2, 1, race_no_val, font=Font(name="Meiryo UI", size=14, bold=True), alignment=c)
    wc(r+2, 7, f'想定スリット: {slit}', font=Font(name="Arial", size=13, bold=True), alignment=c)


    # ── Row 4: 列ヘッダ ──
    st1_rank = next((res.get('st_rank') for res in results if res['waku']=='1'), None)
    st1_str  = f'1号艇平均ST\n{st1_rank:.1f}位' if st1_rank else '1号艇平均ST\n-'
    for col, txt in [(1,'選手名'),(5,'組'),(9,'想定\nスリット'),(13,'オリジナル\n1着率'),(15,'コース別\n3連対')]:
        wc(r+3, col, txt, fill=hf, font=Font(name="Noto Sans CJK SC", size=9 if col!=1 else 9, bold=True, color="FFFFFFFF"), alignment=c)
    for col, txt in [(6,'モータ\n2連'),(7,'展示\n偏差値')]:
        wc(r+3, col, txt, fill=hf, font=Font(name="Meiryo UI", size=8, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+3, 11, st1_str, fill=hf, font=Font(name="Meiryo UI", size=8, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+3, 12, '評価', fill=hf, font=Font(name="Meiryo UI", size=10, bold=True, color="FFFFFFFF"), alignment=c)

    # ── Row 5-10: 選手データ ──
    # 艇番の濃い背景色（サンプル準拠）
    color_fills = {
        1: 'FFFFFFFF', 2: 'FF111111', 3: 'FFFF0000',
        4: 'FF00B0F0', 5: 'FFFFFF00', 6: 'FF00B050'
    }
    # 艇番列の文字色
    color_fonts = {
        1: 'FF000000', 2: 'FFFFFFFF', 3: 'FFFFFFFF',
        4: 'FFFFFFFF', 5: 'FF000000', 6: 'FFFFFFFF'
    }
    # 隣接列（B,E,F,G,H,L,M,O）の薄い背景色（サンプル準拠）
    color_fills_light = {
        1: 'FFFFFFFF', 2: 'FF000000', 3: 'FFFFCCCC',
        4: 'FFCCE8FF', 5: 'FFFFFACC', 6: 'FFCCFFEE'
    }
    # 黒艇（2号艇）の隣接列は黒背景だが、サンプルではtheme=0（黒）→実際は黒
    # ただしB列はtheme=0 tint=-0.15→ほぼ黒。文字は白。
    color_fonts_light = {
        1: 'FF000000', 2: 'FFFFFFFF', 3: 'FF000000',
        4: 'FF000000', 5: 'FF000000', 6: 'FF000000'
    }
    # 改善A: Font/Fillオブジェクトをループ外で一括生成してキャッシュ
    _bf_cache       = {w: make_fill(color_fills[w])       for w in range(1, 7)}
    _bf_light_cache = {w: make_fill(color_fills_light[w]) for w in range(1, 7)}
    _ff_waku_cache  = {w: Font(name="Meiryo UI",        size=11, bold=True,  color=color_fonts[w])       for w in range(1, 7)}
    _ff_name_cache  = {w: Font(name="Noto Sans CJK SC", size=11, bold=True,  color=color_fonts_light[w]) for w in range(1, 7)}
    _ff_data_cache  = {w: Font(name="Meiryo UI",        size=10, bold=False, color=color_fonts_light[w]) for w in range(1, 7)}
    _ff_rate_cache  = {w: Font(name="Meiryo UI",        size=11, bold=True,  color=color_fonts_light[w]) for w in range(1, 7)}
    _ff_honmei_cache= {w: Font(name="Noto Sans CJK SC", size=10, bold=False, color=color_fonts_light[w]) for w in range(1, 7)}
    _fill_white     = make_fill('FFFFFFFF')
    _fill_missing   = make_fill("FFFCE4D6")
    _font_missing_rate = Font(name="Meiryo UI", size=10, bold=False, color="FFBF8F00")
    _font_course    = Font(name="Meiryo UI", size=10, bold=False, color='FF000000')
    for i, res in enumerate(results[:6]):
        dr = r + 4 + i
        waku_no = int(res['waku']) if str(res['waku']).isdigit() else i+1
        bf      = _bf_cache[waku_no]
        bf_light = _bf_light_cache[waku_no]
        fc      = color_fonts[waku_no]
        fc_light = color_fonts_light[waku_no]
        ff_waku = _ff_waku_cache[waku_no]
        ff_name = _ff_name_cache[waku_no]
        ff_data = _ff_data_cache[waku_no]
        ff_rate = _ff_rate_cache[waku_no]

        # 選手名: 姓（2文字）＋全角スペース2つ＋名
        # （年齢除去は calc_race_indices で実施済み）
        name = res['name']
        # 全角スペース・半角スペースで分割
        parts = [p for p in re.split(r'[\s\u3000]+', name) if p]
        if len(parts) >= 2:
            name_fmt = parts[0] + '　　' + parts[1]
        else:
            name_fmt = name
        # データ不足の場合は選手名に★を付加
        if res.get('data_missing'):
            name_fmt = '★' + name_fmt

        rel_w1 = f"{res['rel_win1']:.1f}%" if res['rel_win1'] is not None else '-'
        abs_w3 = f"{res['abs_win3']:.1f}%" if res.get('abs_win3') is not None else '-'

        wc(dr, 1, waku_no,   fill=bf,       font=ff_waku, alignment=c, border=bdr)
        wc(dr, 2, name_fmt,  fill=bf_light,  font=ff_name, alignment=c, border=bdr)
        wc(dr, 5, res['kumi'],   fill=bf_light, font=ff_data, alignment=c, border=bdr)
        wc(dr, 6, res['motor2'], fill=bf_light, font=ff_data, alignment=c, border=bdr)
        # G/H列：展示タイム偏差値（F/St影響列に表示）
        # 【軽微①】前日出力では展示未実施のため None → 「-前日」と表示して当日版と区別
        tenji_h = res.get("tenji_hensa")
        tenji_str = f"{tenji_h:.1f}" if tenji_h is not None else "前日"
        wc(dr, 7, tenji_str,     fill=bf_light, font=ff_data, alignment=c, border=bdr)
        wc(dr, 8, None,          fill=bf_light, font=ff_data, alignment=c, border=bdr)
        # I列はサンプルでFFFFFFFF（白）
        wc(dr, 9, res['course'], fill=_fill_white, font=_font_course, alignment=c, border=bdr)
        wc(dr,12, res['honmei'] if res['honmei'].strip() else None,
                                 fill=bf_light, font=_ff_honmei_cache[waku_no], alignment=c, border=bdr)
        # データ不足の場合は薄オレンジ背景に変更
        if res.get('data_missing'):
            rel_w1_disp = "－" if res['rel_win1'] is None else rel_w1
            abs_w3_disp = "－" if res.get('abs_win3') is None else abs_w3
            wc(dr,13, rel_w1_disp, fill=_fill_missing, font=_font_missing_rate, alignment=c, border=bdr)
            wc(dr,15, abs_w3_disp, fill=_fill_missing, font=_font_missing_rate, alignment=c, border=bdr)
        else:
            wc(dr,13, rel_w1,        fill=bf_light, font=ff_rate, alignment=c, border=bdr)
            wc(dr,15, abs_w3,        fill=bf_light, font=ff_data, alignment=c, border=bdr)

    # ── Row 11: 区切り（空行） ──
    sep_fill = make_fill("FFCCCCDD")
    for col in range(1, 17):
        ws.cell(r+10, col).fill = sep_fill

    # ── Row 12-13: 事前評価ヘッダ ──
    wc(r+11, 1, 'ボートリサーチ流  事前評価', fill=hf, font=Font(name="Noto Sans CJK SC", size=11, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+11, 7, 'イン逃げ決着時の2着優位度（相対%）', fill=shf, font=Font(name="Noto Sans CJK SC", size=11, bold=True, color="FF000000"), alignment=c)
    wc(r+11,14, '3着\n指数', fill=hf, font=Font(name="Meiryo UI", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+11,15, 'コース別\n3連対', fill=hf, font=Font(name="Meiryo UI", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    for col, txt in [(2,'イン\n逃げ'),(3,'相性'),(4,'機力'),(5,'自在性'),(6,'展開')]:
        wc(r+12, col, txt, fill=shf, font=Font(name="Meiryo UI" if col in (2,6) else "Noto Sans CJK SC", size=9, bold=True, color="FF000000"), alignment=c)

    # ── Row 14-19: 事前評価データ ──
    # 2着率テキストブロック（G14結合）
    circle_labels = {2:'黒', 3:'赤', 4:'青', 5:'黄', 6:'緑'}
    # 【重大②明確化】circle_pctはレース内正規化済みの「相対シェア%」
    # 「絶対的な2着確率」ではなく「このレースで2着になる相対優位度」を示す
    circle_lines = ['2着優位度（イン逃げ決着時・レース内相対%）', '']
    # circle_pct は calc_race_indices でレース内正規化済み
    circle_pct_map = {res['waku']: res.get('circle_pct') for res in results}
    for wk in range(2, 7):
        pct = circle_pct_map.get(str(wk))
        pct_str = f'{pct:>4.0f}%' if pct is not None else '   -%'
        circle_lines.append(f'{circle_labels[wk]}({wk}枠): {pct_str}')
    wc(r+13, 7, '\n'.join(circle_lines), alignment=left_align(wrap=True))

    # 事前評価記号
    jizen_result = None
    if JIZEN_AVAILABLE and _jizen_members:
        try:
            jizen_result = evaluate_all(_jizen_members)
        except Exception as e:
            print(f"  ⚠️  事前評価（write_race_flat内）でエラーが発生しました: {e}")

    # 2着率テキストブロック: results の circle_pct（正規化済み）を使用
    # total_2nd は描画用の再計算（frame_2nd は raw スコアなので circle_pct が正値）
    total_2nd = sum(v for k, v in frame_2nd.items() if k != "1" and v is not None)

    # 事前評価行・決まり手行の薄い背景色（サンプル準拠）
    jizen_light_fills = {
        1: 'FFFFFFFF', 2: 'FF000000', 3: 'FFFFCCCC',
        4: 'FFCCE8FF', 5: 'FFFFFACC', 6: 'FFCCFFEE'
    }
    jizen_light_fonts = {
        1: 'FF000000', 2: 'FFFFFFFF', 3: 'FF000000',
        4: 'FF000000', 5: 'FF000000', 6: 'FF000000'
    }
    # 改善A: 事前評価・決まり手ループ用キャッシュ
    _bf_j_cache    = {w: make_fill(jizen_light_fills[w]) for w in range(1, 7)}
    _ff_j11_cache  = {w: Font(name="Meiryo UI", size=11, bold=True,  color=color_fonts[w])       for w in range(1, 7)}
    _ff_j11l_cache = {w: Font(name="Meiryo UI", size=11, bold=True,  color=jizen_light_fonts[w]) for w in range(1, 7)}
    _ff_j10l_cache = {w: Font(name="Meiryo UI", size=10, bold=False, color=jizen_light_fonts[w]) for w in range(1, 7)}
    _fill_m5       = make_fill('FFE5DFEC')  # 黄艇M/N列専用
    _font_red      = Font(name="Meiryo UI", size=10, bold=False, color="FFCC0000")
    _align_cwrap   = center_align(wrap=True)
    for i, res in enumerate(results[:6]):
        dr = r + 13 + i
        waku_no = int(res['waku']) if str(res['waku']).isdigit() else i+1
        bf   = _bf_cache[waku_no]
        bf_j = _bf_j_cache[waku_no]
        ff   = _ff_j11_cache[waku_no]
        ff_j = _ff_j11l_cache[waku_no]
        wc(dr, 1, waku_no, fill=bf, font=ff, alignment=c, border=bdr)

        # B列: ほぼ白
        wc(dr, 2, None, fill=_fill_white, alignment=c)
        # C/D/E列: 薄い艇番色
        for col in [3, 4, 5]:
            wc(dr, col, None, fill=bf_j, alignment=c)
        # F列: waku1-3は白、waku4-6は薄い色
        f_fill = bf_j if waku_no >= 4 else _fill_white
        wc(dr, 6, None, fill=f_fill, alignment=c)

        # 事前評価記号
        if jizen_result:
            col_map = {'in_nige':2,'aisho':3,'kiryoku':4,'jizaisei':5,'tenkai':6}
            for key, col in col_map.items():
                sym = jizen_result[key][i]
                cell = ws.cell(dr, col)
                cell.value = sym if sym else None

        # N列（3着指数）: 薄い色
        wc(dr, 14, None, fill=bf_j, alignment=c)
        idx3 = res.get('idx3', 0)
        ws.cell(dr, 14).value = idx3 if idx3 else None
        # O/P列（オリジナル3連対）: 薄い色
        wc(dr, 15, None, fill=bf_j, alignment=c)
        abs_w3_ev = f"{res['abs_win3']:.1f}%" if res.get('abs_win3') is not None else '-'
        ws.cell(dr, 15).value = abs_w3_ev

    # ── Row 20: 区切り（空行） ──
    for col in range(1, 17):
        ws.cell(r+19, col).fill = sep_fill

    # ── Row 21: 決まり手タイトル・説明・場平均タイトル ──
    wc(r+20, 1, '決まり手 直近1年\n1=被決まり手 2〜6=決まり手', fill=hf, font=Font(name="Noto Sans CJK SC", size=8, bold=True, color="FFFFFFFF"), alignment=c)
    _exp = (
        'ボートリサーチ流 事前評価（展示前の評価）\n\n'
        '◇イン逃げ … 出走メンバーで相対評価。\n  イン逃げ信頼度 (◎>○>△>×)\n\n'
        '◇相性 … 1枠に対する相性\n  (◎>○>△ ※空白は平凡)\n\n'
        '◇機力 … 出走メンバーで機力を相対評価\n  (A>B>C>D>E)\n\n'
        '◇自在性 … 自ら動きレース展開を作れるか\n  (◎>○>△ ※空白は平凡)\n\n'
        '◇展開 … 4~6枠で展開が向いた時の対応力\n  (◎>○>△ ※空白は平凡)'
    )
    wc(r+20, 6, _exp, alignment=left_align(wrap=True))
    wc(r+20,12, '場平均', fill=hf, font=hfont, alignment=c)

    # ── Row 22: 決まり手ヘッダ・場平均ヘッダ ──
    for col, txt in [(1,'枠'),(2,'逃げ%'),(3,'差し%\n差された%'),(4,'まくり%\n捲られた%'),(5,'まくり差し%\n捲り差された%')]:
        wc(r+21, col, txt, fill=hf, font=Font(name="Noto Sans CJK SC", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    for col, txt in [(12,'枠'),(13,'1-○'),(15,'1着率')]:
        wc(r+21, col, txt, fill=shf, font=Font(name="Noto Sans CJK SC", size=9, bold=True, color="FF000000"), alignment=c)

    # ── Row 23-28: 決まり手データ・場平均データ ──
    for i, res in enumerate(results[:6]):
        dr = r + 22 + i
        waku_no = int(res['waku']) if str(res['waku']).isdigit() else i+1
        bf       = _bf_cache[waku_no]
        bf_light = _bf_j_cache[waku_no]
        ff       = _ff_j11_cache[waku_no]
        ff_light = _ff_j10l_cache[waku_no]
        cm = res.get('raw_cm', {})

        wc(dr, 1, str(waku_no), fill=bf, font=ff, alignment=c, border=bdr)

        # 決まり手%（コース別マスタから）B〜E列に薄い背景色を設定
        def pct_str(key):
            v = safe_float(_get_cm_val(cm, key))
            return f'{v*100:.0f}%' if v else '-'

        # B列: 逃げ% → 1号艇のみ表示、2〜6はハイフン
        if waku_no == 1:
            wc(dr, 2, pct_str('逃げ%'), fill=_fill_white, font=ff_light, alignment=c)
        else:
            wc(dr, 2, '-', fill=_fill_white, font=ff_light, alignment=c)

        # C/D/E列:
        #   1号艇 → 被決まり手%（差された%・捲られた%・捲り差された%）赤字
        #   2〜6号艇 → 決まり手%（差し%・まくり%・まくり差し%）
        if waku_no == 1:
            def lose_pct_str(key):
                v = safe_float(cm.get(key))
                return f'{v*100:.0f}%' if v else '-'
            wc(dr, 3, lose_pct_str('差され%'),
               fill=bf_light, font=_font_red, alignment=_align_cwrap)
            wc(dr, 4, lose_pct_str('捲られ%'),
               fill=bf_light, font=_font_red, alignment=_align_cwrap)
            wc(dr, 5, lose_pct_str('捲り差され%'),  # update_master出力キーに統一
               fill=bf_light, font=_font_red, alignment=_align_cwrap)
        else:
            wc(dr, 3, pct_str('差し%'),       fill=bf_light, font=ff_light, alignment=c)
            wc(dr, 4, pct_str('まくり%'),     fill=bf_light, font=ff_light, alignment=c)
            wc(dr, 5, pct_str('まくり差し%'), fill=bf_light, font=ff_light, alignment=c)

        # 場平均（イン逃げ時2着率）L列は艇番色
        wc(dr,12, str(waku_no), fill=bf, font=ff, alignment=c, border=bdr)
        # circle_pct はレース内正規化済みの値を results から取得
        pct_2nd_norm = res.get('circle_pct', 0) or 0
        # M/N列: 薄い色（黄艇のみT7 = FFE5DFEC）
        m_fill = _fill_m5 if waku_no == 5 else bf_light
        wc(dr,13, f'{pct_2nd_norm:.0f}%' if total_2nd > 0 else '-', fill=m_fill, font=ff_light, alignment=c)
        # 1着率（相対1着率）O/P列
        rel_w1 = f"{res['rel_win1']:.1f}%" if res['rel_win1'] is not None else '-'
        wc(dr,15, rel_w1, fill=m_fill, font=ff_light, alignment=c)

    # ── Row 29: 注記 ＋ ヒモ荒れ判定 ──
    himo_are = (race_judgment or {}).get("himo_are", {}) if race_judgment else {}
    himo_verdict = himo_are.get("verdict", "対象外")
    if himo_verdict == "不参加推奨":
        mcp = himo_are.get("max_combo_prob", 0.0) or 0.0
        eto = himo_are.get("est_top_odds",   0.0) or 0.0
        cc  = himo_are.get("circle_concentration", 0.0) or 0.0
        himo_str = (
            f"【🚫 ヒモ固まり・見送り推奨】"
            f"最有力確率{mcp:.3f}（推定最高人気{eto:.0f}倍台）/ 2着集中度{cc:.0f}%\n"
            f"→ 1号艇1着固定でも3連単オッズが構造的に低い。参加しない方が期待値が高い。\n"
        )
        himo_fill = make_fill("FFFCE4D6")   # 薄オレンジ（警告色）
        himo_font = Font(name="Noto Sans CJK SC", size=9, bold=True, color="FFCC0000")
    elif himo_verdict == "点数絞り":
        mcp = himo_are.get("max_combo_prob", 0.0) or 0.0
        eto = himo_are.get("est_top_odds",   0.0) or 0.0
        cc  = himo_are.get("circle_concentration", 0.0) or 0.0
        himo_str = (
            f"【⚠ ヒモやや固め・点数絞り】"
            f"最有力確率{mcp:.3f}（推定最高人気{eto:.0f}倍台）/ 2着集中度{cc:.0f}%\n"
            f"→ 展示後にヒモを上位2艇に絞り込むこと。\n"
        )
        himo_fill = make_fill("FFFFFFF0")   # 薄黄（注意色）
        himo_font = Font(name="Noto Sans CJK SC", size=9, bold=False, color="FF996600")
    elif himo_verdict == "参加推奨":
        mcp = himo_are.get("max_combo_prob", 0.0) or 0.0
        eto = himo_are.get("est_top_odds",   0.0) or 0.0
        cc  = himo_are.get("circle_concentration", 0.0) or 0.0
        himo_str = (
            f"【✅ ヒモ分散・参加推奨】"
            f"最有力確率{mcp:.3f}（推定最高人気{eto:.0f}倍台）/ 2着集中度{cc:.0f}%\n"
            f"→ 1号艇1着固定でヒモ広めに流す。買い目を+2点追加推奨。\n"
        )
        himo_fill = make_fill("FFE2EFDA")   # 薄緑（推奨色）
        himo_font = Font(name="Noto Sans CJK SC", size=9, bold=False, color="FF375623")
    else:
        himo_str  = ""
        himo_fill = None
        himo_font = None

    note_base = (
        '※オリジナル1着率、2着率、3着指数は出走メンバーで相対評価しており、天候、潮、展示、場特性などを考慮していません。\n'
        '※枠内を想定します。進入変更があった場合使用できません。'
    )
    note_val  = (himo_str + note_base) if himo_str else note_base
    note_fill = himo_fill or make_fill("FFFFFFFF")
    note_font = himo_font or Font(name="Meiryo UI", size=8)
    wc(r+28, 1, note_val,
       fill=note_fill, font=note_font,
       alignment=left_align(wrap=True))

    # ── 行高（サンプル完全再現） ──
    row_heights = {
        0:12.75, 1:30.0, 2:24.75, 3:31.5,
        4:27.75, 5:27.75, 6:27.75, 7:27.75, 8:27.75, 9:27.75,
        10:4.5, 11:25.5, 12:21.75,
        13:25.5, 14:25.5, 15:25.5, 16:25.5, 17:25.5, 18:25.5,
        19:4.5, 20:25.5, 21:21.75,
        22:25.5, 23:25.5, 24:25.5, 25:25.5, 26:25.5, 27:25.5,
        28:39.0
    }
    for offset, height in row_heights.items():
        ws.row_dimensions[r + offset].height = height

    # ── ST順位 舟図をK5:K10に埋め込む ──
    # 改善C: tempfileをやめてBytesIOインメモリ処理に変更（ディスクI/O削減）
    if player_master is not None:
        try:
            buf = _make_st_boat_chart(results, player_master, outpath=None)
            if buf is not None:
                img = XLImage(buf)
                # K列(col=10, 0始まり), 選手データ開始行(r+4, 0始まり=r+3)
                marker = AnchorMarker(col=10, colOff=0, row=r+3, rowOff=0)
                size   = XDRPositiveSize2D(cx=865179, cy=2082127)
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
        except Exception as e:
            print(f"  ⚠️  ST舟図生成エラー ({race_no}R): {e}")


# ============================================================
# 数値データシート書き込み（出力_数値と同一テーブル形式）
# ============================================================
def write_numeric_sheet(wb, all_race_data, course_master=None, venue_course_master=None):
    """
    「会場名_数値」シートを「出力_数値」と完全同一のテーブル形式で書き出す。
    縦軸: 分類・項目・艇番、横軸: 1R〜12R（最大レース数分）
    """
    if not all_race_data:
        return

    venue_name = all_race_data[0]["venue"]
    sheet_name = f"{venue_name}_数値"

    # ── FLY入力値の準備（flying_*.xlsx を優先、なければ手入力値を退避して復元）──
    fly_input_backup = {}  # {(race_no_str, waku_int): value}

    # まず flying_*.xlsx から自動取得を試みる
    _fly_auto = {}   # {(race_no_int, waku_int): 1}
    _venue_name_for_fly = all_race_data[0]["venue"] if all_race_data else ""
    _flying_path = None
    for _d in [pathlib.Path(__file__).parent, BASE_DIR, pathlib.Path(".")]:
        _cands = sorted(_d.glob("flying_*.xlsx"), reverse=True)
        if _cands:
            _flying_path = _cands[0]
            break
    if _flying_path:
        try:
            _wb_fly = openpyxl.load_workbook(str(_flying_path), read_only=True, data_only=True)
            _ws_fly = _wb_fly["フライング一覧"]
            for _frow in _ws_fly.iter_rows(min_row=2, values_only=True):
                _fcols = list(_frow) + [None] * 6
                _fv, _fr, _fw, _, _, _fi = _fcols[:6]
                if _fv != _venue_name_for_fly or not _fr or not _fw or not _fi:
                    continue
                try:
                    _fly_auto[(int(_fr), int(str(_fw).strip()))] = 1
                except (TypeError, ValueError):
                    pass
            _wb_fly.close()
            if _fly_auto:
                print(f"  ✈️  FLY自動取得: {_venue_name_for_fly} {len(_fly_auto)}件 ({_flying_path.name})")
        except Exception as _e:
            print(f"  ⚠️  flying xlsx 読み込み失敗（手入力値を使用）: {_e}")

    if sheet_name in wb.sheetnames:
        _ws_old = wb[sheet_name]
        # ヘッダ行(2行目)からレース番号→列のマッピングを取得
        _old_race_col = {}
        for _c in range(1, _ws_old.max_column + 1):
            _v = _ws_old.cell(2, _c).value
            if _v and isinstance(_v, str) and "R" in _v:
                import re as _re
                _m = _re.match(r"(\d+)R", _v.strip())
                if _m:
                    _old_race_col[_m.group(1)] = _c
        # flying_*.xlsx がない場合のみ手入力値をバックアップ
        if not _fly_auto:
            for _r in range(1, _ws_old.max_row + 1):
                if _ws_old.cell(_r, 1).value == "FLY入力":
                    _waku_val = _ws_old.cell(_r, 3).value
                    try:
                        _waku_int = int(_waku_val)
                    except (TypeError, ValueError):
                        continue
                    for _rno, _col in _old_race_col.items():
                        _fv = _ws_old.cell(_r, _col).value
                        if _fv is not None and str(_fv).strip() not in ("", "0", "None"):
                            fly_input_backup[(_rno, _waku_int)] = _fv
            if fly_input_backup:
                print(f"  ✈️  FLY手入力値を退避: {len(fly_input_backup)}件")
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ── スタイル定数（出力_数値から実測） ──
    BOAT_FILL   = {1:"FFFFFFFF", 2:"FF1A1A1A", 3:"FFCC0000", 4:"FF2255CC", 5:"FFDDCC00", 6:"FF115511"}
    BOAT_FONT   = {1:"FF000000", 2:"FFFFFFFF", 3:"FFFFFFFF", 4:"FFFFFFFF", 5:"FF000000", 6:"FFFFFFFF"}
    FILL_HDR    = "FF1F4E79"   # ヘッダ行（濃紺）
    FILL_MISSING = "FFFCE4D6"  # データ不足セル（薄オレンジ）
    FILL_SEC_B  = "FF2E75B6"   # 数値指標・選手情報セクション（青）
    FILL_SEC_G  = "FF70AD47"   # 決まり手セクション（緑）
    FILL_ITEM_B = "FFDCE6F1"   # 数値指標・選手情報 項目セル（薄青）
    FILL_ITEM_G = "FFE2EFDA"   # 決まり手 項目セル（薄緑）
    FILL_ITEM_P = "FFD9E1F2"   # 選手情報 項目セル（薄紺）

    race_nos = [rd["race_no"] for rd in all_race_data]
    n_races  = len(race_nos)
    _first_results = all_race_data[0].get("results", [])
    _first_rel = _first_results[0].get("rel_win1") if _first_results else "なし"
    print(f"  📊 {sheet_name}: {n_races}レース, 1R results件数={len(_first_results)}, rel_win1={_first_rel}")

    def sf(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def fn(bold=True, size=9, color="FF000000", name="Meiryo UI"):
        return Font(name=name, size=size, bold=bold, color=color)

    def al(h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def wc(row, col, val, fill=None, font=None, align=None, border=None):
        c = ws.cell(row=row, column=col)
        c.value = val
        if fill:   c.fill   = fill
        if font:   c.font   = font
        if align:  c.alignment = align
        if border: c.border = border

    thin = Side(style="thin", color="FFCCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 列幅 ──
    ws.column_dimensions["A"].width = 14.0
    ws.column_dimensions["B"].width = 20.0
    ws.column_dimensions["C"].width = 7.0
    for i in range(n_races):
        ws.column_dimensions[get_column_letter(4 + i)].width = 14.0

    # ── Row 1: タイトル ──
    race_date = all_race_data[0].get("race_date", "")
    title = f"ボートリサーチ数値データ　【{venue_name}】　{race_date}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + n_races)
    wc(1, 1, title, font=fn(bold=True, size=12), align=al("left"))
    ws.row_dimensions[1].height = 22.0

    # ── Row 2: ヘッダ（分類・項目・艇番・1R〜nR） ──
    hf = sf(FILL_HDR)
    hfn = fn(bold=True, color="FFFFFFFF")
    for col, txt in [(1,"分類"),(2,"項目"),(3,"艇番")]:
        wc(2, col, txt, fill=hf, font=hfn, align=al(), border=bdr)
    for i, rno in enumerate(race_nos):
        deadline = all_race_data[i].get("deadline") if i < len(all_race_data) else None
        header_val = f"{rno}R\n{deadline}" if deadline else f"{rno}R"
        wc(2, 4+i, header_val, fill=hf, font=hfn, align=al(wrap=True), border=bdr)
    ws.row_dimensions[2].height = 28.0

    # ── セクション書き込みヘルパー ──
    def write_section_header(row, label, fill_color):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
        wc(row, 1, label, fill=sf(fill_color), font=fn(bold=True, color="FFFFFFFF"), align=al("left"))
        ws.row_dimensions[row].height = 13.0

    def write_item_block(row_start, cat_label, item_label, cat_fill, item_fill, data_by_waku, missing_by_waku=None):
        """
        data_by_waku:   {1: [r1val, r2val, ...], 2: [...], ...}
        missing_by_waku:{1: [bool, bool, ...], 2: [...], ...}  データ不足フラグ
        6艇分 × nレース のデータブロックを書く
        """
        for i in range(6):
            waku = i + 1
            row  = row_start + i
            # A列（分類）: 最初の艇番のみ
            if waku == 1:
                wc(row, 1, cat_label,  fill=sf(cat_fill),  font=fn(bold=True, color="FFFFFFFF"), align=al())
                wc(row, 2, item_label, fill=sf(item_fill), font=fn(bold=True, color="FF000000"), align=al("left"))
            else:
                wc(row, 1, None, fill=sf(cat_fill), align=al())
                wc(row, 2, None, fill=sf(item_fill), align=al())
            # C列（艇番）
            wc(row, 3, str(waku),
               fill=sf(BOAT_FILL[waku]),
               font=fn(bold=True, color=BOAT_FONT[waku]),
               align=al(), border=bdr)
            # D〜 データ列
            vals    = data_by_waku.get(waku,    [None]*n_races)
            missing = (missing_by_waku or {}).get(waku, [False]*n_races)
            for j, v in enumerate(vals):
                is_missing = missing[j] if j < len(missing) else False
                if is_missing and v is None:
                    # データ不足セル: 薄オレンジ背景 ＋ 「－」テキスト
                    wc(row, 4+j, "－",
                       fill=sf(FILL_MISSING),
                       font=fn(bold=False, color="FFBF8F00"),
                       align=al(), border=bdr)
                else:
                    wc(row, 4+j, v, font=fn(bold=False), align=al(), border=bdr)
            ws.row_dimensions[row].height = 15.0

    # ── データ取り出しヘルパー ──
    def get_vals(rd, key):
        """results から艇番→値のdictを返す"""
        d = {}
        for res in rd.get("results", []):
            try: waku = int(res["waku"])
            except (ValueError, TypeError): continue
            d[waku] = res.get(key)
        return d

    def get_cm_vals(rd, cm_key, multiplier=1.0):
        d = {}
        for res in rd.get("results", []):
            try: waku = int(res["waku"])
            except (ValueError, TypeError): continue
            v = safe_float(res.get("raw_cm", {}).get(cm_key))
            d[waku] = round(v * multiplier, 1) if v is not None else None
        return d

    # 6艇×nレース のデータを転置して {waku: [r1,r2,...]} に変換
    def build_waku_data(extract_fn):
        result = {w: [] for w in range(1, 7)}
        for rd in all_race_data:
            vals = extract_fn(rd)
            for w in range(1, 7):
                result[w].append(vals.get(w))
        return result

    # ── データ不足フラグを {waku: [bool×nレース]} に変換するヘルパー ──
    def build_missing_waku():
        """各艇番・各レースについて data_missing フラグを返す"""
        result = {w: [] for w in range(1, 7)}
        for rd in all_race_data:
            missing_map = {}
            for res in rd.get("results", []):
                try: waku = int(res["waku"])
                except (ValueError, TypeError): continue
                missing_map[waku] = bool(res.get("data_missing", False))
            for w in range(1, 7):
                result[w].append(missing_map.get(w, False))
        return result

    missing_waku = build_missing_waku()

    # ── ▼ 数値指標セクション ──
    row = 3
    write_section_header(row, "▼ 数値指標", FILL_SEC_B)
    row += 1

    # オリジナル1着率
    write_item_block(row, "数値指標", "オリジナル1着率(%)", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): f"{round(r['rel_win1'], 1)}%" if r.get("rel_win1") is not None else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }), missing_by_waku=missing_waku)
    row += 6

    # オリジナル3連対率
    write_item_block(row, "数値指標", "コース別3連対率(%)", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): f"{round(r['abs_win3'], 1)}%" if r.get("abs_win3") is not None else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }), missing_by_waku=missing_waku)
    row += 6

    # イン逃げ時2着率（1号艇は対象外のため表示なし）
    write_item_block(row, "数値指標", "2着優位度(%)[相対・イン逃げ時]", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): (
                None if r["waku"] == "1"
                else f"{round(r.get('circle_pct'), 1)}%" if r.get("circle_pct") is not None else "-"
            )
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 3着指数
    write_item_block(row, "数値指標", "3着指数", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("idx3") if r.get("idx3") else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── ▼ 決まり手セクション ──
    write_section_header(row, "▼ 決まり手（直近1年｜1号艇=被決まり手%、2〜6号艇=決まり手%）", FILL_SEC_G)
    row += 1

    # 被決まり手マップ: {(レースindex, waku=1): {差され%: v, 捲られ%: v, 捲り差され%: v}}
    # 差し%行→1号艇欄に差され%、まくり%行→捲られ%、まくり差し%行→捲り差され% を赤字で表示
    LOSE_KEY_MAP = {
        "差し%":      "差され%",
        "まくり%":    "捲られ%",
        "まくり差し%":"捲り差され%",
    }

    for cm_key, label in [
        ("逃げ%",      "逃げ%"),
        ("差し%",      "差し%"),
        ("まくり%",    "まくり%"),
        ("まくり差し%","まくり差し%"),
        ("抜き%",      "抜き%"),
    ]:
        lose_key = LOSE_KEY_MAP.get(cm_key)  # 対応する被決まり手キー（なければNone）

        # 各艇番×レースのデータを構築
        main_data = build_waku_data(lambda rd, k=cm_key: {
            int(r["waku"]): safe_float(r.get("raw_cm", {}).get(k), 0)
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        })
        lose_data = build_waku_data(lambda rd, lk=lose_key: {
            int(r["waku"]): safe_float(r.get("raw_cm", {}).get(lk)) if lk and int(r["waku"]) == 1 else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }) if lose_key else None

        for i in range(6):
            waku = i + 1
            dr   = row + i
            if waku == 1:
                wc(dr, 1, "決まり手", fill=sf(FILL_SEC_G),  font=fn(bold=True, color="FFFFFFFF"), align=al())
                wc(dr, 2, label,      fill=sf(FILL_ITEM_G), font=fn(bold=True, color="FF000000"), align=al("left"))
            else:
                wc(dr, 1, None, fill=sf(FILL_SEC_G),  align=al())
                wc(dr, 2, None, fill=sf(FILL_ITEM_G), align=al())
            wc(dr, 3, str(waku),
               fill=sf(BOAT_FILL[waku]),
               font=fn(bold=True, color=BOAT_FONT[waku]),
               align=al(), border=bdr)

            main_vals = main_data.get(waku, [None]*n_races)
            lose_vals = lose_data.get(waku, [None]*n_races) if lose_data else [None]*n_races

            for j, mv in enumerate(main_vals):
                lv = lose_vals[j] if j < len(lose_vals) else None
                main_str = f"{round((mv or 0) * 100, 1)}%" if mv not in (None,) else "-"

                if waku == 1 and lose_key:
                    # 1号艇：被決まり手%のみ表示（差された・捲られた・捲り差された）
                    lose_str = f"{round(lv * 100, 1)}%" if lv not in (None, 0.0) else "-"
                    wc(dr, 4+j, lose_str,
                       fill=sf(FILL_ITEM_G),
                       font=fn(bold=False, size=9, color="FFCC0000"),
                       align=al(), border=bdr)
                elif waku == 1:
                    # 1号艇の逃げ%行はそのまま表示
                    wc(dr, 4+j, main_str,
                       font=fn(bold=False), align=al(), border=bdr)
                elif cm_key == "逃げ%":
                    # 2〜6号艇の逃げ%行は表示不要（コース1専用）
                    wc(dr, 4+j, "-",
                       font=fn(bold=False, color="FF999999"), align=al(), border=bdr)
                else:
                    # 2〜6号艇：決まり手%をそのまま表示
                    wc(dr, 4+j, main_str,
                       font=fn(bold=False), align=al(), border=bdr)
            ws.row_dimensions[dr].height = 15.0
        row += 6

    # ── ▼ 事前評価セクション ──
    FILL_SEC_J  = "FF7030A0"   # 事前評価セクション（紫）
    FILL_ITEM_J = "FFE9D7F5"   # 事前評価 項目セル（薄紫）

    if JIZEN_AVAILABLE:
        write_section_header(row, "▼ 事前評価（ボートリサーチ流｜展示前）", FILL_SEC_J)
        row += 1

        JIZEN_ITEMS = [
            ("in_nige",  "①逃げ"),
            ("aisho",    "②相性"),
            ("kiryoku",  "③機力"),
            ("tenkai",   "④展開"),
            ("jizaisei", "⑤S安定"),
        ]

        JIZEN_FONT_COLOR = {
            "◎": "FF7F6000", "◎?": "FF7F6000",
            "○": "FF1F3864",
            "△": "FF843C0C",
            "A": "FF375623", "B": "FF375623",
            "C": "FF404040",
            "D": "FF843C0C", "E": "FF7F0000",
        }
        JIZEN_FILL_COLOR = {
            "◎": "FFFFF2CC", "◎?": "FFFFF2CC",
            "○": "FFDAE3F3",
            "△": "FFFCE4D6",
            "A": "FFE2EFDA", "B": "FFE2EFDA",
            "C": "FFF2F2F2",
            "D": "FFFCE4D6", "E": "FFFFCCCC",
        }

        # レースごとに evaluate_all を呼んで結果をキャッシュ
        jizen_cache = []
        for rd in all_race_data:
            jm = rd.get("jizen_members")
            if jm:
                try:
                    jizen_cache.append(evaluate_all(jm))
                except Exception:
                    jizen_cache.append(None)
            else:
                jizen_cache.append(None)

        for jizen_key, jizen_label in JIZEN_ITEMS:
            for i in range(6):
                waku = i + 1
                dr   = row + i
                if waku == 1:
                    wc(dr, 1, "事前評価",   fill=sf(FILL_SEC_J),  font=fn(bold=True, color="FFFFFFFF"), align=al())
                    wc(dr, 2, jizen_label,  fill=sf(FILL_ITEM_J), font=fn(bold=True, color="FF000000"), align=al("left"))
                else:
                    wc(dr, 1, None, fill=sf(FILL_SEC_J),  align=al())
                    wc(dr, 2, None, fill=sf(FILL_ITEM_J), align=al())
                wc(dr, 3, str(waku),
                   fill=sf(BOAT_FILL[waku]),
                   font=fn(bold=True, color=BOAT_FONT[waku]),
                   align=al(), border=bdr)
                for j, jr in enumerate(jizen_cache):
                    sym = (jr.get(jizen_key, [""] * 6)[i] or "") if jr else ""
                    cell_fill = sf(JIZEN_FILL_COLOR.get(sym, "FFFFFFFF")) if sym else None
                    cell_font = fn(bold=(sym in ("◎", "◎?", "A")),
                                   size=9,
                                   color=JIZEN_FONT_COLOR.get(sym, "FF808080"))
                    # sym が空のとき None を書く（旧: "－"）
                    # "－"（全角）は fill_newspaper.py のフィルタ（半角"-"のみ）を
                    # すり抜けて新聞に転記され、印が出なくなるバグを修正
                    wc(dr, 4 + j, sym if sym else None,
                       fill=cell_fill, font=cell_font,
                       align=al(), border=bdr)
                ws.row_dimensions[dr].height = 15.0
            row += 6

    # ── ▼ 選手情報セクション ──
    write_section_header(row, "▼ 選手情報", FILL_SEC_B)
    row += 1

    # 選手名（マスタのフルネームで補完）
    # 公式サイトの選手名が4文字に切り詰められている場合（例: 安河内鈴 → 安河内鈴之介）、
    # course_master / venue_course_master のキーから前方一致でフルネームを補完する。
    def _resolve_full_name(name_raw):
        """4文字名をマスタのフルネームに補完して返す。見つからなければ元の名前を返す。"""
        if not name_raw:
            return name_raw
        norm = name_raw.replace("\u3000", "").replace(" ", "").strip()
        # course_master: キーは (選手名, コース文字列)
        if course_master:
            for master_key in course_master:
                master_nm = master_key[0].replace("\u3000", "").replace(" ", "").strip()
                if master_nm.startswith(norm) and len(master_nm) > len(norm):
                    return master_key[0]
        # venue_course_master: キーは (選手名, 会場名, コース文字列) または (選手名, 会場名, int)
        if venue_course_master:
            for master_key in venue_course_master:
                master_nm = str(master_key[0]).replace("\u3000", "").replace(" ", "").strip()
                if master_nm.startswith(norm) and len(master_nm) > len(norm):
                    return master_key[0]
        return name_raw

    write_item_block(row, "選手情報", "選手名", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): _resolve_full_name(r.get("name", ""))
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 組
    write_item_block(row, "選手情報", "組", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("kumi", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # モータ2連
    write_item_block(row, "選手情報", "モータ2連", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("motor2", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 攻め記号（2〜6号艇）/ 逃げ評価（1号艇）
    # ※ item ラベルを "攻め記号" に統一: fill_newspaper.py が "選手情報_攻め記号" で参照するため
    write_item_block(row, "選手情報", "攻め記号", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("honmei", "").strip() or None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 想定コース
    write_item_block(row, "選手情報", "想定コース", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("course", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── FLY手動入力行（想定コース直後に挿入）────────────────────────────────
    # ユーザーが各レース列に手動でフライング数を入力するための行
    from openpyxl.styles import PatternFill as _PF, Font as _Fn, Alignment as _Al, Border as _Br, Side as _Sd
    _fill_red    = _PF("solid", fgColor="FFCC0000")
    _fill_yellow = _PF("solid", fgColor="FFFFFF00")
    _fill_boat   = {1:"FFFFFFFF",2:"FF1A1A1A",3:"FFCC0000",4:"FF2255CC",5:"FFDDCC00",6:"FF115511"}
    _fill_font   = {1:"FF000000",2:"FFFFFFFF",3:"FFFFFFFF",4:"FFFFFFFF",5:"FF000000",6:"FFFFFFFF"}
    _thin = _Sd(style="thin", color="FFCCCCCC")
    _bdr  = _Br(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _fn_w = _Fn(name="Meiryo UI", size=9, bold=True,  color="FFFFFFFF")
    _fn_b = _Fn(name="Meiryo UI", size=9, bold=False, color="FF000000")
    _al_c = _Al(horizontal="center", vertical="center")
    for _i, _waku_no in enumerate(range(1, 7)):
        _r = row + _i
        _c = ws.cell(_r, 1)
        _c.value, _c.fill, _c.font, _c.alignment, _c.border = "FLY入力", _fill_red, _fn_w, _al_c, _bdr
        _c = ws.cell(_r, 2)
        _c.value, _c.fill, _c.font, _c.alignment, _c.border = "FLY（1=あり）", _fill_yellow, _fn_b, _al_c, _bdr
        _c = ws.cell(_r, 3)
        _c.value = _waku_no
        _c.fill  = _PF("solid", fgColor=_fill_boat[_waku_no])
        _c.font  = _Fn(name="Meiryo UI", size=9, bold=True, color=_fill_font[_waku_no])
        _c.alignment, _c.border = _al_c, _bdr
        for _col in range(4, 4 + n_races):
            _cell = ws.cell(_r, _col)
            # flying_*.xlsx のデータを優先、なければバックアップ手入力値を復元
            _rno_str = race_nos[_col - 4] if (_col - 4) < len(race_nos) else None
            if _rno_str is not None and _fly_auto.get((int(_rno_str), _waku_no)):
                _cell.value = 1
            else:
                _restored = fly_input_backup.get((_rno_str, _waku_no)) if _rno_str else None
                _cell.value = _restored
            _cell.fill   = _fill_yellow
            _cell.border = _bdr
        ws.row_dimensions[_r].height = 15.0
    row += 6

    # FLY数
    write_item_block(row, "選手情報", "FLY数", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("fly_count", 0)
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # F/ST影響ラベル
    write_item_block(row, "選手情報", "F/ST影響", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("fly_label", "低")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── 注記 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "※オリジナル1着率・オリジナル3連対率・イン逃げ時2着率・3着指数は出走メンバーで正規化。決まり手%は過去コース別1着時の割合。",
       font=fn(bold=False, size=8), align=al("left"))
    ws.row_dimensions[row].height = 13.0
    row += 1

    # ── データ不足凡例 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "★凡例：薄オレンジ「－」はデータ不足（コース別実績 < 5走 or 総実績 < 10走）。新人・移籍直後の選手はデータ蓄積後に正確な指数が算出されます。",
       fill=sf(FILL_MISSING),
       font=fn(bold=False, size=8, color="FF7F3F00"),
       align=al("left"))
    ws.row_dimensions[row].height = 14.0

    # ── データ不足選手一覧（シート末尾サマリー） ──
    missing_summary = []
    for rd in all_race_data:
        for res in rd.get("results", []):
            if res.get("data_missing") and res.get("missing_reason"):
                missing_summary.append(
                    f"{rd['race_no']}R-{res['waku']}号艇 {res['name']}：{res['missing_reason']}"
                )
    if missing_summary:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
        wc(row, 1, "【データ不足選手一覧】",
           fill=sf("FFFCE4D6"),
           font=fn(bold=True, size=9, color="FF7F3F00"),
           align=al("left"))
        ws.row_dimensions[row].height = 14.0
        for summary_line in missing_summary:
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
            wc(row, 1, summary_line,
               fill=sf(FILL_MISSING),
               font=fn(bold=False, size=9, color="FF7F3F00"),
               align=al("left"))
            ws.row_dimensions[row].height = 13.0

    # ════════════════════════════════════════════════════════════════════
    # 📋 展示前参考買い目 ＋ 考察セクション
    # ════════════════════════════════════════════════════════════════════
    FILL_JF_HDR  = "FF1F4E79"   # 濃紺（ヘッダ）
    FILL_JF_SUB  = "FF2E75B6"   # 中青（行ラベル）
    FILL_JF_BODY = "FFDCE6F1"   # 薄青（データ）
    FILL_JF_BET  = "FFFFFF99"   # 薄黄（買い目リスト）
    FILL_JF_NOTE = "FFFFE699"   # 黄（絞込ガイド）

    FILL_KS_HDR  = "FF203864"   # 濃紺（考察ヘッダ）
    FILL_KS_SUB  = "FF305496"   # 中紺（考察行ラベル）
    FILL_KS_BODY = "FFD9E1F2"   # 薄紺（考察データ）
    FILL_KS_GOOD = "FFE2EFDA"   # 薄緑（良好）
    FILL_KS_WARN = "FFFFF2CC"   # 薄黄（注意）
    FILL_KS_BAD  = "FFFCE4D6"   # 薄橙（警告）

    row += 2

    # ── 考察セクションヘッダ ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "━━━  🔍 Step1: レース考察  ━━━  "
       "展開予想の根拠（判定スコア・相性・展開・ヒモ荒れ）→ 次の買い目の根拠になります",
       fill=sf(FILL_KS_HDR), font=fn(bold=True, size=9, color="FFFFFFFF"), align=al("left"))
    ws.row_dimensions[row].height = 16.0
    row += 1

    KOSATSU_ROWS = [
        # ── 🎯 最優先：展開ストーリー（7ステップ統合） ───────────────────────
        ("🎯展開ストーリー",  "①逃 ②逃時2・3着 ③主役 ④主役時2・3着 ⑤崩れ ⑥残存 ⑦突き → 買い目"),
        # ── 層1: 判定サマリー（詳細確認用） ────────────────────────────────
        ("🎯展開パターン",  "A鉄板/B主役/C拮抗/D荒れ ＋ 方針"),
        ("🎯狙い目",        "個人攻撃有効性ベース（攻撃力×脆弱性×ST×会場）"),
        ("④展開予測",      "❶❷❸❹サマリー・1M順・主軸・漁夫"),
        ("①判定スコア",    "ランク/スコア/戦略"),
        # ── 層2: 根拠詳細 ─────────────────────────────────────────────────
        ("②3択判定",       "逃げ/飛び/両建て 根拠"),
        ("③相性考察",      "攻撃艇の根拠（攻撃力順）"),
        ("⑤注意事項",      "FLY・データ不足・ST不安定"),
        ("⑥ヒモ荒れ",      "1号艇強本命時のヒモ分散判定"),
        ("⑦展開quality",  "展開の絞れ度・買い目点数ガイド"),
        # ── 層3: 結論 ─────────────────────────────────────────────────────
        ("⑧考察の結論",    "考察まとめ → Step2へ"),
    ]

    for sec_label, sec_item in KOSATSU_ROWS:
        wc(row, 1, sec_label,
           fill=sf(FILL_KS_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        wc(row, 2, sec_item,
           fill=sf(FILL_KS_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())

        max_lines = 2
        for i, rd in enumerate(all_race_data):
            rj  = rd.get("race_judgment", {}) or {}
            bet = rd.get("bet_suggestions", {}) or {}
            je  = rd.get("jizen_eval", {}) or {}
            jf  = bet.get("jizen_formation", {}) or {}
            ryotate = bet.get("ryotate_detail", rj.get("ryotate", {})) or {}
            txt = ""
            fc  = FILL_KS_BODY

            if sec_label == "🎯展開ストーリー":
                # ── 7ステップ統合ストーリー ───────────────────────────────────
                try:
                    txt = _generate_tenkai_story(
                        results         = rd.get("results", []),
                        venue           = rd.get("venue", ""),
                        venue_stats     = rd.get("venue_stats", {}),
                        race_judgment   = rj,
                        bet_suggestions = bet,
                    )
                    if not txt or not txt.strip():
                        txt = "（データ不足のためストーリーを生成できませんでした）"
                except Exception as _e:
                    txt = f"（展開ストーリー生成エラー: {_e}）"

                # 濃紺背景・白文字（⑧考察の結論と同じスタイルだが行高を大きく取る）
                fc = "FF1A2E4A"
                wc(row, 4+i, txt,
                   fill=sf(fc), font=fn(bold=False, size=8, color="FFFFFFFF"),
                   align=al(h="left", wrap=True), border=bdr)
                max_lines = max(max_lines, txt.count("\n") + 1)
                continue

            elif sec_label == "🎯展開パターン":
                # 層1先頭: 展開パターン（A/B/C/D）＋方針を最上部に表示
                tp   = bet.get("tenkai_pattern", rj.get("tenkai_pattern", "?"))
                tp_pol = bet.get("tenkai_pattern_policy", rj.get("tenkai_pattern_policy", "-"))
                rank = rj.get("rank", "-")
                _TP_FULL = {
                    "A": "🟢 A: 鉄板逃げ",
                    "B": "🔴 B: 主役展開",
                    "C": "🟡 C: 拮抗",
                    "D": "🟣 D: 荒れ",
                }
                tp_label_full = _TP_FULL.get(tp, f"⚪ ?: {tp}")
                txt = f"【{tp_label_full}】\n方針: {tp_pol}\nランク: {rank}"
                fc = {
                    "A": FILL_KS_GOOD,
                    "B": FILL_KS_BAD,
                    "C": FILL_KS_WARN,
                    "D": "FFEDE7F6",
                }.get(tp, FILL_KS_BODY)

            elif sec_label == "🎯狙い目":
                # ── 個人攻撃有効性ベースの狙い目 ──────────────────────────────
                neraime_list = bet.get("neraime", []) or []
                neraime_2nd  = bet.get("neraime_2nd", []) or []
                neraime_top  = bet.get("neraime_top") or {}
                atk_eff_map  = bet.get("atk_eff_map", {}) or {}
                s1p_nr = bet.get("s1_prob", 0) or 0
                fp_map = bet.get("first_prob_map", {}) or {}

                # ── 全艇スコア行（共通）──
                atk_lines = []
                for _w, _s in sorted(atk_eff_map.items(), key=lambda x: x[1], reverse=True):
                    _p = fp_map.get(_w, 0)
                    _top_n = neraime_top.get("waku") if neraime_top else None
                    _marker = "★" if _w == _top_n else "  "
                    _level_tag = ""
                    for _n in neraime_list:
                        if _n["waku"] == _w:
                            _level_tag = f"[{_n.get('level','?')}]"
                            break
                    atk_lines.append(
                        f"{_marker}{_w}号: {_level_tag}有効{_s*100:.0f}%"
                        f" → 確率{_p*100:.1f}%"
                    )

                if neraime_top:
                    # 攻め型狙い目あり
                    nw    = neraime_top.get("waku", "-")
                    ns    = neraime_top.get("score", 0)
                    nlv   = neraime_top.get("level", "?")
                    nat   = neraime_top.get("attack_type", "-")
                    nr    = neraime_top.get("reason", "-")
                    np    = neraime_top.get("prob_after", 0)

                    # 残存型も合わせて表示
                    _2nd_lines = []
                    if neraime_2nd:
                        _2nd_lines.append("─ 逃げ時2着残存 ─")
                        for _n2 in neraime_2nd[:3]:
                            _2nd_lines.append(
                                f"  {_n2['waku']}号: {_n2['r2_rate']*100:.0f}%"
                                f"（マスタ）/ {_n2['r3i_rate']*100:.0f}%（3着以内）"
                            )

                    txt = (
                        f"🎯 攻め型: {nw}号艇【{nat}】[{nlv}]\n"
                        f"   有効性{ns*100:.0f}% → 補正後1着確率{np*100:.1f}%\n"
                        f"   {nr}\n"
                        f"─────────────────\n"
                        f"全艇スコア:\n"
                        + "\n".join(atk_lines)
                        + ("\n" + "\n".join(_2nd_lines) if _2nd_lines else "")
                    )
                    fc = (FILL_KS_BAD   if ns >= 0.35 else
                          FILL_KS_WARN  if ns >= 0.22 else
                          FILL_KS_BODY)

                elif neraime_2nd:
                    # 攻め型なし＋逃げ本命 → 残存型のみ
                    _2nd_top = neraime_2nd[0]
                    _2nd_lines = ["─ 逃げ時2着残存（展開別残存マスタ） ─"]
                    for _n2 in neraime_2nd[:4]:
                        _2nd_lines.append(
                            f"  {_n2['waku']}号: 2着{_n2['r2_rate']*100:.0f}%"
                            f" / 3着以内{_n2['r3i_rate']*100:.0f}%"
                        )
                    txt = (
                        f"🎯 残存型: {_2nd_top['waku']}号艇が2着有力\n"
                        f"   逃げ時2着率{_2nd_top['r2_rate']*100:.0f}%"
                        f"（1号艇1着確率{fp_map.get('1',0)*100:.1f}%）\n"
                        f"─────────────────\n"
                        f"全艇スコア:\n"
                        + "\n".join(atk_lines) + "\n"
                        + "\n".join(_2nd_lines)
                    )
                    fc = FILL_KS_GOOD   # 緑：逃げ安定・残存狙い

                else:
                    # 攻め型も残存型も該当なし
                    txt = (
                        f"🎯 狙い目: 明確な攻め手なし\n"
                        f"   （攻撃有効性15%未満・逃げ2着残存も集計中）\n"
                        f"─────────────────\n"
                        f"全艇スコア:\n"
                        + "\n".join(atk_lines)
                    )
                    fc = FILL_KS_BODY

            elif sec_label == "①判定スコア":
                rank  = rj.get("rank", "-")
                score = rj.get("score", "-")
                strat = rj.get("strategy", "-")
                trust = rj.get("data_trust_score", "-")
                txt = f"【ランク{rank}】{score}点\n戦略: {strat}\n信頼度: {trust}%"
                fc = FILL_KS_GOOD if rank in ("S","A") else FILL_KS_WARN if rank == "B" else FILL_KS_BAD

            elif sec_label == "①判定根拠":
                reasons = rj.get("reason", []) or []
                filtered = [r for r in reasons if not r.startswith("複合確率スコア")][:3]
                txt = "\n".join(f"・{r}" for r in filtered) if filtered else "-"

            elif sec_label == "②3択判定":
                verdict = ryotate.get("verdict", "-")
                conf    = ryotate.get("confidence", "-")
                esc_pct = ryotate.get("escape_pct")
                tobi_pct = ryotate.get("tobi_pct")

                # 1行サマリー: 「なぜこの3択か」を1文で表現
                if verdict == "逃げ狙い":
                    if esc_pct is not None:
                        summary = f"→ 逃げ{esc_pct:.0f}%: 1号艇先行優勢・ヒモ流しで参加"
                    else:
                        summary = "→ 1号艇先行優勢・逃げ軸"
                elif verdict == "飛び狙い":
                    main_th = (rj.get("affinity") or {}).get("dominant_attacker", "?")
                    if tobi_pct is not None:
                        summary = f"→ 飛び{tobi_pct:.0f}%: {main_th}号艇が主な脅威・飛び軸"
                    else:
                        summary = f"→ {main_th}号艇先行・飛び軸"
                else:
                    if esc_pct is not None and tobi_pct is not None:
                        summary = f"→ 逃げ{esc_pct:.0f}%/飛び{tobi_pct:.0f}%拮抗・展示で軸決定"
                    else:
                        summary = "→ 拮抗展開・展示後に軸を傾ける"

                reason = ryotate.get("reason", "-")
                # reason は長いので先頭の警告部分のみ抽出（⚠以降の重要部分）
                reason_short = reason.split("。")[0] if reason != "-" else "-"

                txt = (
                    f"【{verdict}】確信度{conf}%\n"
                    f"{summary}\n"
                    f"根拠: {reason_short}"
                )
                fc = FILL_KS_GOOD if verdict == "逃げ狙い" else FILL_KS_BAD if verdict == "飛び狙い" else FILL_KS_WARN

            elif sec_label == "②買い方指示":
                txt = ryotate.get("buy_style", "-") or "-"

            elif sec_label == "③相性考察":
                affinity = rj.get("affinity", {}) or {}
                summary  = affinity.get("affinity_summary", {}) or {}
                attack   = affinity.get("attack_score", {}) or {}
                outer = sorted(
                    [(w, summary.get(w, "-"), attack.get(w, 0))
                     for w in ["2","3","4","5","6"] if w in summary],
                    key=lambda x: x[2], reverse=True
                )[:4]
                txt = "\n".join(f"{w}号: {s}（攻{a:.0f}pt）" for w, s, a in outer) or "データなし"

            elif sec_label == "④展開予測":
                ft   = rj.get("first_turn", {}) or {}
                cm   = rj.get("conflict_map", {}) or {}
                w1e  = rj.get("w1_escape", {}) or {}
                mp   = rj.get("main_player", {}) or {}
                ef   = rj.get("escape_fallback", {}) or {}
                dh   = rj.get("dark_horse", {}) or {}
                tp   = bet.get("tenkai_pattern", rj.get("tenkai_pattern", "?"))

                # 展開パターンラベル
                _TP_LABEL = {
                    "A": "A: 鉄板逃げ",
                    "B": "B: 主役展開",
                    "C": "C: 拮抗",
                    "D": "D: 荒れ",
                }
                tp_label = _TP_LABEL.get(tp, f"?: {tp}")

                # ❶
                er   = w1e.get("escape_rank", "-")
                epct = w1e.get("escape_pct", "-")
                thr  = w1e.get("top_threat_waku", "-")
                tht  = w1e.get("top_threat_type", "-")
                # ❷
                mw   = mp.get("main_waku", "-")
                mt   = mp.get("main_type", "-")
                ms   = mp.get("main_score", 0) or 0
                sw   = mp.get("sub_waku")
                # ❸
                fbp  = ef.get("fallback_pct", "-")
                fbr  = ef.get("fallback_rank", "-")
                fbt  = ef.get("fly_type", "-")
                # ❹
                dh_ok  = dh.get("is_valid", False)
                dh_top = dh.get("top_waku", "-")
                dh_sc  = dh.get("top_score", 0) or 0
                dh_cands = dh.get("dark_horse_candidates", [])
                dh_str = "  ".join(
                    f"{w}号【{tag}】{s*100:.0f}%"
                    for w, s, tag in dh_cands[:2]
                ) if dh_cands else "なし"

                # 1M到達・対立構造
                entry = ft.get("entry_order", [])
                entry_str = "→".join([f"{w}号" for w, _ in entry[:4]]) if entry else "-"
                mc = cm.get("main_conflict") or {}
                mc_desc = mc.get("desc", "-") if mc else "-"
                cb = cm.get("collapse_beneficiary", [])
                cb_str = "・".join([f"{w}号" for w, _ in cb[:2]]) if cb else "-"

                sub_str = f"/{sw}号" if sw else ""
                dh_line = f"✅{dh_top}号{dh_sc*100:.0f}%" if dh_ok else "─"
                txt = (
                    f"❶逃 {epct}【{er}】脅={thr}号({tht})\n"
                    f"❷主 {mw}号【{mt}】{ms*100:.0f}%{sub_str}\n"
                    f"❸残 {fbp}【{fbr}】  ❹穴 {dh_line}\n"
                    f"─────────────────\n"
                    f"1M: {entry_str}\n"
                    f"主軸: {mc_desc} / 漁夫: {cb_str}"
                )
                # パターン別カラー
                fc = {
                    "A": FILL_KS_GOOD,
                    "B": FILL_KS_BAD,
                    "C": FILL_KS_WARN,
                    "D": "FFEDE7F6",  # 薄紫（荒れ）
                }.get(tp, FILL_KS_BODY)

            elif sec_label == "⑤注意事項":
                reasons = rj.get("reason", []) or []
                warn_keys = ["FLY","出遅れ","データ不足","データ信頼","ST不安定","実績なし","暫定"]
                warns = [r for r in reasons if any(k in r for k in warn_keys)]

                # 展示確認トリガー（ヒモ荒れ判定から）
                ha = rj.get("himo_are", {}) or {}
                tenji_trigger = ha.get("tenji_trigger", "")

                # 注意事項がなくても展示トリガーは必ず表示
                lines = []
                if warns:
                    lines += [f"⚠ {w}" for w in warns]
                else:
                    lines.append("特記なし")

                if tenji_trigger:
                    lines.append("─ 展示確認ポイント ─")
                    lines += [f"📋 {t}" for t in tenji_trigger.split("\n") if t.strip()]

                txt = "\n".join(lines)
                fc = FILL_KS_WARN if warns else FILL_KS_BODY

            elif sec_label == "⑥ヒモ荒れ":
                ha = rj.get("himo_are", {}) or {}
                ha_verdict = ha.get("verdict", "対象外")
                tenji_t = ha.get("tenji_trigger", "")
                tenji_line = f"\n📋 {tenji_t}" if tenji_t else ""
                if ha_verdict == "対象外":
                    txt = "対象外（rel_win1 < 45%）\n通常判定に委ねる"
                    fc  = FILL_KS_BODY
                elif ha_verdict == "不参加推奨":
                    mcp = ha.get("max_combo_prob", 0.0) or 0.0
                    eto = ha.get("est_top_odds",   0.0) or 0.0
                    cc  = ha.get("circle_concentration", 0.0) or 0.0
                    txt = (
                        f"🚫 見送り推奨\n"
                        f"最有力確率{mcp:.3f}（推定{eto:.0f}倍台）\n"
                        f"2着集中度{cc:.0f}%"
                        f"{tenji_line}"
                    )
                    fc  = FILL_KS_BAD
                elif ha_verdict == "点数絞り":
                    mcp = ha.get("max_combo_prob", 0.0) or 0.0
                    eto = ha.get("est_top_odds",   0.0) or 0.0
                    cc  = ha.get("circle_concentration", 0.0) or 0.0
                    txt = (
                        f"⚠ 点数絞り\n"
                        f"最有力確率{mcp:.3f}（推定{eto:.0f}倍台）\n"
                        f"2着集中度{cc:.0f}%"
                        f"{tenji_line}"
                    )
                    fc  = FILL_KS_WARN
                else:  # 参加推奨
                    mcp = ha.get("max_combo_prob", 0.0) or 0.0
                    eto = ha.get("est_top_odds",   0.0) or 0.0
                    cc  = ha.get("circle_concentration", 0.0) or 0.0
                    txt = (
                        f"✅ 参加推奨（ヒモ分散）\n"
                        f"最有力確率{mcp:.3f}（推定{eto:.0f}倍台）\n"
                        f"2着集中度{cc:.0f}%"
                        f"{tenji_line}"
                    )
                    fc  = FILL_KS_GOOD

            # ⑦展開quality
            elif sec_label == "⑦展開quality":
                sq = rj.get("scenario_quality", {}) or {}
                q_score   = sq.get("quality_score", "-")
                q_rank    = sq.get("quality_rank", "-")
                q_verdict = sq.get("quality_verdict", "-")
                bet_guide = sq.get("bet_size_guide", "-")
                comps     = sq.get("components", {})
                comp_str  = "  ".join([f"{k}:{v:.0f}" for k, v in comps.items()])

                # シナリオタイプ別ROI目安（バックテスト実績ベース）
                stype_for_roi = bet.get("scenario_type", "-")
                ROI_GUIDE = {
                    "逃げ軸流し": "ROI目安: 逃げ軸=約27%(中央値¥1,080)  ※ヒモ参加推奨時のみ",
                    "飛び軸":     "ROI目安: 飛び軸=約38%(中央値¥2,200)  ※quality A以上推奨",
                    "両建て":     "ROI目安: 両建て=約32%(中央値¥1,400)  ※展示後に軸絞ること",
                }
                roi_line = ROI_GUIDE.get(stype_for_roi, "ROI目安: シナリオ未確定")

                txt = (
                    f"quality: {q_rank}（{q_score}点）\n"
                    f"→ {q_verdict}\n"
                    f"→ {bet_guide}\n"
                    f"構成: {comp_str}\n"
                    f"─\n"
                    f"{roi_line}"
                )
                if q_rank in ("S", "A"):
                    fc = FILL_KS_GOOD
                elif q_rank == "B":
                    fc = FILL_KS_BODY
                elif q_rank == "C":
                    fc = FILL_KS_WARN
                else:
                    fc = FILL_KS_BAD

            # ⑧考察の結論: 初心者向け短文（_generate_buy_hint → fill_newspaper.py kosatsu_raw に転記）
            if sec_label == "⑧考察の結論":
                try:
                    txt = _generate_buy_hint(
                        results         = rd.get("results", []),
                        venue           = rd.get("venue", ""),
                        venue_stats     = rd.get("venue_stats", {}),
                        race_judgment   = rj,
                        bet_suggestions = bet,
                    )
                    if not txt or not txt.strip():
                        txt = "（データ不足のため考察を生成できませんでした）"
                except Exception as _e:
                    txt = f"（考察生成エラー: {_e}）"
                fc = "FF1F4E79"   # 濃紺（固定）
                wc(row, 4+i, txt,
                   fill=sf(fc), font=fn(bold=False, size=8, color="FFFFFFFF"),
                   align=al(h="left", wrap=True), border=bdr)
                max_lines = max(max_lines, txt.count("\n") + 1)
                continue

            # ⑨買い方ヒント: KOSATSU_ROWSから削除済み（デッドコード）
            # 必要な場合は KOSATSU_ROWS に ("⑨買い方ヒント", ...) を追加すれば復活可能
            if sec_label == "⑨買い方ヒント":
                try:
                    txt = _generate_buy_hint(
                        results   = rd.get("results", []),
                        venue     = rd.get("venue", ""),
                        venue_stats = rd.get("venue_stats", {}),
                        race_judgment  = rj,
                        bet_suggestions = bet,
                    )
                    if not txt or not txt.strip():
                        txt = "（データ不足：s1_prob未計算の可能性）"
                except Exception as _e:
                    txt = f"（買い方ヒント生成エラー: {_e}）"
                fc = "FF1A3A5C"   # 深紺（専用色）
                wc(row, 4+i, txt,
                   fill=sf(fc), font=fn(bold=False, size=8, color="FFFFFFFF"),
                   align=al(h="left", wrap=True), border=bdr)
                max_lines = max(max_lines, txt.count("\n") + 1)
                continue

            wc(row, 4+i, txt,
               fill=sf(fc), font=fn(bold=False, size=8),
               align=al(h="left", wrap=True), border=bdr)
            max_lines = max(max_lines, txt.count("\n") + 1)

        # 変更6: 行高を情報密度に合わせて最適化
        # 層1（先頭3行）は低め固定、層2（根拠詳細）は可変、層3（⑧結論）はコンパクト
        _ROW_HEIGHT_MAP = {
            "🎯展開ストーリー": max(13.0 * max_lines, 160.0),  # 最優先・7ステップ分を確保
            "🎯展開パターン": max(14.0 * max_lines, 36.0),
            "🎯狙い目":       max(13.0 * max_lines, 80.0),
            "④展開予測":     max(14.0 * max_lines, 40.0),
            "①判定スコア":   max(14.0 * max_lines, 36.0),
            "②3択判定":      max(16.0 * max_lines, 36.0),
            "③相性考察":     max(14.0 * max_lines, 36.0),
            "⑤注意事項":     max(14.0 * max_lines, 28.0),
            "⑥ヒモ荒れ":     max(14.0 * max_lines, 28.0),
            "⑦展開quality":  max(14.0 * max_lines, 36.0),
            "⑧考察の結論":   max(14.0 * max_lines, 40.0),
        }
        ws.row_dimensions[row].height = _ROW_HEIGHT_MAP.get(
            sec_label, max(18.0 * max_lines, 36.0)
        )
        row += 1

    # ════════════════════════════════════════════════════════════════════
    # 📋 Step2: 展開シナリオ＋買い目候補
    # ════════════════════════════════════════════════════════════════════
    # 【設計方針】
    #   シナリオ自動切替（逃げ軸/両建て/飛び軸）の判定結果と
    #   各シナリオの買い目候補を確率順に提示する。
    #   実オッズ・参加可否は人間が展示後に最終判断する。
    # ════════════════════════════════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "━━━  📋 Step2: 展開シナリオ＋買い目候補  ━━━  "
       "確率から展開を自動判定。買い目候補と理想合成オッズを確認して展示後の最終判断へ",
       fill=sf(FILL_JF_HDR), font=fn(bold=True, size=9, color="FFFFFFFF"), align=al("left"))
    ws.row_dimensions[row].height = 16.0
    row += 1

    # ─ 説明行 ─
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "【逃げ軸流し】印◎が1号艇かつ逃げ確率高め　"
       "【両建て】印◎が1号艇で確率低め、または印◎外枠で印○が1号艇　"
       "【飛び軸】印◎◎が2〜6号艇　"
       "2〜3着ヒモは展開位置補正×個人能力×jizen評価で自動選択。折り返しは条件付き自動追加",
       fill=sf(FILL_JF_NOTE), font=fn(bold=False, size=8, color="FF7F3F00"), align=al("left"))
    ws.row_dimensions[row].height = 14.0
    row += 1

    # ─ シナリオ判定行 ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "展開シナリオ判定",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        bet  = rd.get("bet_suggestions", {}) or {}
        stype = bet.get("scenario_type", bet.get("scenario_verdict", "-"))
        s1p  = bet.get("s1_prob", 0) or 0
        flyp = bet.get("fly_prob", 0) or 0
        fly_axes = bet.get("fly_axes", [])
        fly_str  = "・".join(fly_axes[:2]) + "号" if fly_axes else "-"
        rj   = rd.get("race_judgment", {}) or {}
        sc_fly_type  = rj.get("sc_fly_type", "-")
        gyofu_top3   = rj.get("sc_gyofu_top3", [])
        gyofu_str    = "・".join(gyofu_top3) + "号" if gyofu_top3 else "-"

        # ❶〜❹サマリー
        w1e  = rj.get("w1_escape", {}) or {}
        mp2  = rj.get("main_player", {}) or {}
        ef2  = rj.get("escape_fallback", {}) or {}
        dh2  = rj.get("dark_horse", {}) or {}
        tp2  = bet.get("tenkai_pattern", rj.get("tenkai_pattern", "?"))
        tp2_pol = bet.get("tenkai_pattern_policy", rj.get("tenkai_pattern_policy", "-"))
        _TP2_EMOJI = {"A": "🟢", "B": "🔴", "C": "🟡", "D": "🟣"}
        tp2_icon = _TP2_EMOJI.get(tp2, "⚪")

        er2   = w1e.get("escape_rank", "-")
        epct2 = w1e.get("escape_pct", "-")
        mw2   = mp2.get("main_waku", "-")
        mt2   = mp2.get("main_type", "-")
        ms2   = mp2.get("main_score", 0) or 0
        fbp2  = ef2.get("fallback_pct", "-")
        fbr2  = ef2.get("fallback_rank", "-")
        dh_ok2  = dh2.get("is_valid", False)
        dh_top2 = dh2.get("top_waku", "-")
        dh_sc2  = dh2.get("top_score", 0) or 0

        hs2      = bet.get("honmei_scenario") or {}
        hs2_pats = hs2.get("honmei_patterns", {}) or {}
        hs2_hp   = hs2_pats.get("honmei") or {}
        hs2_sj   = hs2.get("scenario_judgment", {}) or {}
        hs2_narr = hs2_hp.get("win_narrative", "")
        hs2_conf = hs2.get("confidence")
        hs2_conf_str = f"{hs2_conf*100:.0f}%" if hs2_conf is not None else "-"
        hs2_reas = (hs2_sj.get("reasons") or [""])[1:3]
        hs2_reas_str = "\n".join(f"  {r}" for r in hs2_reas)

        # 実際の買い目1着分布を集計してシナリオとの整合性を明示
        _cands2 = bet.get("candidates", [])
        from collections import Counter as _C2
        _fd2 = _C2(
            c["combo"].split("-")[0] for c in _cands2
            if c.get("combo") and not c.get("is_fallback_bet") and not c.get("is_dh_bet")
        )
        _w1c2  = _fd2.get("1", 0)
        _flyc2 = sum(v for k, v in _fd2.items() if k != "1")
        _topf2 = _fd2.most_common(1)[0][0] if _flyc2 > 0 else "-"
        _stn2  = bet.get("scenario_type_note", "")
        # 買い目軸サマリー（実態）
        if _w1c2 > 0 and _flyc2 > 0:
            _axis_summary = f"軸: 1号頭{_w1c2}点 / {_topf2}号頭{_flyc2}点"
        elif _w1c2 > 0:
            _axis_summary = f"軸: 1号頭{_w1c2}点（逃げ一本）"
        elif _flyc2 > 0:
            _axis_summary = f"軸: {_topf2}号頭{_flyc2}点（飛び一本）"
        else:
            _axis_summary = "軸: -"

        txt = (
            f"{tp2_icon}【展開:{tp2}】{stype}\n"
            f"逃げ{s1p*100:.0f}% / 飛び{flyp*100:.0f}%（主:{fly_str}）\n"
            f"─────────────────\n"
            f"❶ 逃げ力: {epct2}【{er2}】\n"
            f"❷ 主役:   {mw2}号【{mt2}】{ms2*100:.0f}%\n"
            f"❸ 残存:   {fbp2}【{fbr2}】\n"
            f"❹ 穴:     {'✅' + str(dh_top2) + '号 ' + f'{dh_sc2*100:.0f}%' if dh_ok2 else '─'}\n"
            f"─────────────────\n"
            f"買い目: {_axis_summary}\n"
            + (f"⚠ {_stn2}\n" if _stn2 else "")
            + f"方針: {tp2_pol}\n"
            f"─ 潰れ展開(SC) ─\n"
            f"飛び役:{sc_fly_type} / 漁夫:{gyofu_str}\n"
            f"─ ◎勝ちパターン ─\n"
            f"{hs2_narr} 信頼度{hs2_conf_str}\n"
            f"{hs2_reas_str}"
        )
        # 変更4: tenkai_pattern（A/B/C/D）で色分け（Step1の展開パターン行と統一）
        fc = {
            "A": "FFE2EFDA",   # 緑（鉄板逃げ）
            "B": "FFFCE4D6",   # 橙（主役展開）
            "C": "FFFFF2CC",   # 黄（拮抗）
            "D": "FFEDE7F6",   # 薄紫（荒れ）
        }.get(tp2, "FFFFF2CC")
        wc(row, 4+i, txt,
           fill=sf(fc), font=fn(bold=True, size=9),
           align=al(h="left", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 56.0
    row += 1

    # ─ 変更5: 展示後チェックポイント行（買い目候補の直上） ─────────────────────
    FILL_CHECK = "FF4A235A"   # 深紫（展示後確認専用色）
    wc(row, 1, "展示後\n確認",
       fill=sf(FILL_CHECK), font=fn(bold=True, size=8, color="FFFFFFFF"), align=al(wrap=True))
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "展示後チェックポイント",
       fill=sf(FILL_CHECK), font=fn(bold=True, size=8, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        rj_c  = rd.get("race_judgment", {}) or {}
        bet_c = rd.get("bet_suggestions", {}) or {}
        tp_c  = bet_c.get("tenkai_pattern", rj_c.get("tenkai_pattern", "?"))
        mw_c  = (rj_c.get("main_player", {}) or {}).get("main_waku", "-")
        mt_c  = (rj_c.get("main_player", {}) or {}).get("main_type", "-")
        er_c  = (rj_c.get("w1_escape", {}) or {}).get("escape_rank", "-")
        ha_c  = (rj_c.get("himo_are", {}) or {}).get("tenji_trigger", "")
        _CHECK_MSG = {
            "A": f"1号艇ST確認（逃げ力【{er_c}】→ 出遅れなければ軸固定）",
            "B": f"{mw_c}号艇({mt_c})の展示タイム確認（主役スタート・伸び足が鍵）",
            "C": f"1号艇 vs {mw_c}号艇のST比較（拮抗→展示で軸を絞る）",
            "D": f"展示タイム全艇チェック（荒れ展開→突出艇を再確認）",
        }
        check_msg = _CHECK_MSG.get(tp_c, "展示タイム・ST確認後に軸決定")
        if ha_c:
            check_msg += f"\n📋 {ha_c.split(chr(10))[0]}"
        wc(row, 4+i, check_msg,
           fill=sf("FFFBEAFE"), font=fn(bold=False, size=8, color="FF4A235A"),
           align=al(h="left", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 36.0
    row += 1

    # ─ 買い目候補行（確率順・シナリオ別） ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "買い目候補（確率順）\nシナリオ種別 / 確率%",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al(wrap=True))
    max_lines_cand = 3
    for i, rd in enumerate(all_race_data):
        bet   = rd.get("bet_suggestions", {}) or {}
        rj_b  = rd.get("race_judgment",   {}) or {}
        cands = bet.get("candidates", [])
        # ── デバッグ: candidates の重複チェック ──
        _seen_debug = set()
        _dup_debug  = []
        for _c in cands:
            _ck = _c.get("combo", "")
            if _ck in _seen_debug:
                _dup_debug.append(_ck)
            _seen_debug.add(_ck)
        if _dup_debug:
            print(f"[DEBUG] candidates重複検出: {_dup_debug} (race={rd.get('race_no','-')})")
        # ────────────────────────────────────────
        if not cands:
            wc(row, 4+i, "候補なし",
               fill=sf("FFDDDDDD"), font=fn(bold=False, size=8, color="FF808080"),
               align=al(h="left", wrap=True), border=bdr)
            continue

        # ── 考察エンジン出力を収集（Step1→買い目の橋渡し用） ──────────────
        _tp_b    = bet.get("tenkai_pattern", rj_b.get("tenkai_pattern", "?"))
        _stype_b = bet.get("scenario_type", "-")
        _s1p_b   = bet.get("s1_prob", 0) or 0
        _flyp_b  = bet.get("fly_prob", 0) or 0
        _fly_ax  = bet.get("fly_axes", [])
        _mp_b    = rj_b.get("main_player", {}) or {}
        _mw_b    = _mp_b.get("main_waku", "-")
        _mt_b    = _mp_b.get("main_type", "-")
        _ms_b    = float(_mp_b.get("main_score", 0) or 0)
        _w1e_b   = rj_b.get("w1_escape", {}) or {}
        _er_b    = _w1e_b.get("escape_rank", "-")
        _ep_b    = _w1e_b.get("escape_pct", "-")
        _ef_b    = rj_b.get("escape_fallback", {}) or {}
        _fbr_b   = _ef_b.get("fallback_rank", "-")
        _fbp_b   = _ef_b.get("fallback_pct", "-")
        _dh_b    = rj_b.get("dark_horse", {}) or {}
        _dh_ok_b = _dh_b.get("is_valid", False)
        _dh_w_b  = _dh_b.get("top_waku", "-")

        # ── candidatesの実態から1着分布を集計（表示との乖離を防ぐ） ──────────
        # honmei_scenario等がbuy_listを差し替えた場合もここで実態を反映する
        _w1_count  = sum(1 for c in cands if c.get("combo","").split("-")[0] == "1"
                         and not c.get("is_fallback_bet") and not c.get("is_dh_bet"))
        _fly_count = sum(1 for c in cands if c.get("combo","").split("-")[0] != "1"
                         and not c.get("is_fallback_bet") and not c.get("is_dh_bet")
                         and not c.get("is_sc_bet"))
        # 実際の飛び軸（買い目に最も多く出てくる非1号頭の1着艇）
        from collections import Counter as _Counter
        _fly_first_dist = _Counter(
            c["combo"].split("-")[0] for c in cands
            if c.get("combo","").split("-")[0] != "1"
            and not c.get("is_fallback_bet") and not c.get("is_dh_bet")
        )
        _actual_fly_waku = _fly_first_dist.most_common(1)[0][0] if _fly_first_dist else _mw_b

        # 実態と考察の整合性チェック
        # scenario_typeが「逃げ軸流し」なのに1号頭買い目がゼロ → 矛盾を検出して表示
        _stype_display = _stype_b
        if _stype_b == "逃げ軸流し" and _w1_count == 0 and _fly_count > 0:
            _stype_display = f"⚠印◎軸（{_actual_fly_waku}号頭）※確率逃げ{_s1p_b*100:.0f}%"
        elif _stype_b == "逃げ軸流し" and _w1_count > 0:
            _stype_display = f"逃げ軸流し（1号頭{_w1_count}点+{_actual_fly_waku}号頭{_fly_count}点）"

        # 考察❶❷❸❹ → 実際の買い目構成 の因果を1行に圧縮
        # ポイント: "実際の軸" を使う（main_fly_wakuではなく_actual_fly_wakuを使う）
        if _tp_b == "A" and _w1_count > 0:
            _bridge = f"❶逃げ{_ep_b}【{_er_b}】× 逃げ{_s1p_b*100:.0f}% → 1号頭固定({_w1_count}点)"
        elif _w1_count == 0:
            # 1号頭が一切ない → 印◎軸が確率モデルを上回っている
            _honmei_w = next((r.get("waku") for r in rj_b.get("results",[]) if r.get("honmei") == "◎"), None) if rj_b.get("results") else None
            _h_label  = f"印◎{_honmei_w}号" if _honmei_w else f"印◎{_actual_fly_waku}号"
            _bridge = (
                f"{_h_label}頭軸（全{len(cands)}点）"
                f"  ※逃げ{_s1p_b*100:.0f}%だが印◎が{_actual_fly_waku}号 → 印優先"
            )
        elif _tp_b == "B":
            _fly_lbl = f"{_actual_fly_waku}号【{_mt_b}】{_ms_b*100:.0f}%"
            _bridge  = f"❷主役{_fly_lbl}→{_actual_fly_waku}号頭{_fly_count}点  ❸残存{_fbp_b}【{_fbr_b}】{_w1_count}点"
        elif _tp_b == "C":
            _bridge = (
                f"逃げ{_s1p_b*100:.0f}%/飛び{_flyp_b*100:.0f}%拮抗"
                f" → 1号頭{_w1_count}点 vs {_actual_fly_waku}号頭{_fly_count}点"
            )
        elif _tp_b == "D":
            _dh_lbl = f"❹{_dh_w_b}号穴" if _dh_ok_b else "穴候補不明"
            _bridge = f"❶逃げ力【{_er_b}】低・混戦 → 1号{_w1_count}点/{_actual_fly_waku}号{_fly_count}点 {_dh_lbl}"
        else:
            _bridge = f"1号頭{_w1_count}点 / {_actual_fly_waku}号頭{_fly_count}点"

        # セクションヘッダ用の考察橋渡しラベル（実際の軸を使う）
        _fly_w_lbl = f"{_actual_fly_waku}号{_mt_b}"
        _nige_why  = f"❶{_ep_b}【{_er_b}】逃げ{_s1p_b*100:.0f}%"
        _tobi_why  = f"❷{_fly_w_lbl} {_ms_b*100:.0f}%"
        _fb_why    = f"❸残存{_fbp_b}【{_fbr_b}】1号2着狙い"
        _dh_why    = f"❹{_dh_w_b}号穴" if _dh_ok_b else "❹穴"

        combo_map_local = {c["combo"]: c for c in cands}
        used_combos     = set()
        lines_nige      = []
        lines_tobi      = []
        lines_sc        = []
        lines_fallback  = []
        lines_dh        = []
        lines_other     = []

        def _append_to_section(line, sc, c):
            # フラグ優先（明示的に付与されたフラグで分類）
            if c.get("is_fallback_bet"):
                lines_fallback.append(line)
                return
            if c.get("is_dh_bet"):
                lines_dh.append(line)
                return
            if c.get("is_sc_bet") or "潰れ" in sc:
                lines_sc.append(line)
                return
            # first waku で正しく分類（scenarioテキストへの依存をやめる）
            # honmei_scenario版でscenario文字列が変わっても確実に機能する
            _first_w = c.get("combo", "-").split("-")[0] if c.get("combo") else sc
            if _first_w == "1":
                lines_nige.append(line)
            else:
                lines_tobi.append(line)

        # is_orkaeshi フラグ優先でペアを構築（フラグが消えた場合はフラグなし扱い）
        # integrate が candidates を差し替えても、直前の行6122でフラグを復元済み
        paired_cands = {}
        for c in cands:
            key = c["combo"]
            parts = key.split("-")
            if len(parts) != 3:
                continue
            f, s, t = parts
            if c.get("is_orkaeshi"):
                # 1着折り返し: A-B-C の本体は B-A-C
                base_key = f"{s}-{f}-{t}"
                if base_key in combo_map_local:
                    paired_cands[key]      = base_key
                    paired_cands[base_key] = key
            elif c.get("is_orkaeshi_23"):
                # 2着3着折り返し: A-B-C の本体は A-C-B
                base_key = f"{f}-{t}-{s}"
                if base_key in combo_map_local:
                    paired_cands[key]      = base_key
                    paired_cands[base_key] = key
        # ── デバッグ: paired_cands の内容を出力 ──
        if paired_cands:
            print(f"[DEBUG] paired_cands: {paired_cands} (race={rd.get('race_no','-')})")
        # ── デバッグ: cands の順番とフラグを出力 ──
        print(f"[DEBUG] cands order (race={rd.get('race_no','-')}): {[(c['combo'], c.get('is_orkaeshi'), c.get('is_orkaeshi_23')) for c in cands]}")
        # ── デバッグ: is_orkaeshiフラグが消えているcomboを検出 ──
        _orkaeshi_combos = {c["combo"] for c in cands if c.get("is_orkaeshi")}
        _orkaeshi_23_combos = {c["combo"] for c in cands if c.get("is_orkaeshi_23")}
        print(f"[DEBUG] is_orkaeshi={_orkaeshi_combos}, is_orkaeshi_23={_orkaeshi_23_combos} (race={rd.get('race_no','-')})")

        for c in cands:
            key = c["combo"]
            if key in used_combos:
                continue
            sc  = c.get("scenario", "")
            rsn = c.get("reason", "")
            # reason の1着根拠部分のみ抽出（「1号先行優位 / 2号残存」→「1号先行優位」）
            rsn_short = rsn.split(" / ")[0] if rsn else ""
            parts = key.split("-")

            if len(parts) != 3:
                used_combos.add(key)
                line = f"  {key}  {c['prob_pct']:.1f}%"
                if rsn_short:
                    line += f"\n    ← {rsn_short}"
                _append_to_section(line, sc, c)
                continue

            # 折り返しペアも単独行として個別表示（＝表記なし）
            used_combos.add(key)
            line = f"  {key}  {c['prob_pct']:.1f}%"
            if rsn_short:
                line += f"\n    ← {rsn_short}"
            _append_to_section(line, sc, c)

        # セクションヘッダに考察との橋渡しを付与
        sections = []
        if lines_nige:
            sections.append(f"── 🟢逃げ軸 {len(lines_nige)}点（{_nige_why}） ──")
            sections.extend(lines_nige)
        if lines_tobi:
            sections.append(f"── 🔴飛び軸 {len(lines_tobi)}点（{_tobi_why}） ──")
            sections.extend(lines_tobi)
        if lines_sc:
            sections.append(f"── 🎣潰れ受益 {len(lines_sc)}点 ──")
            sections.extend(lines_sc)
        if lines_fallback:
            sections.append(f"── ❸逃げ残存 {len(lines_fallback)}点（{_fb_why}） ──")
            sections.extend(lines_fallback)
        if lines_dh:
            sections.append(f"── ❹穴ヒモ {len(lines_dh)}点（{_dh_why}） ──")
            sections.extend(lines_dh)
        if lines_other:
            sections.append(f"── その他 {len(lines_other)}点 ──")
            sections.extend(lines_other)

        total  = len(cands)
        # 冒頭ブリッジ: 考察→シナリオ→買い目の一気通貫を1〜2行で表示
        header = (
            f"【計{total}点】{_stype_display}\n"
            f"▶ {_bridge}\n"
            f"＝折返 🎣潰れ ❸残存 ❹穴"
        )
        txt = header + "\n" + "\n".join(sections)
        wc(row, 4+i, txt,
           fill=sf(FILL_JF_BET), font=fn(bold=False, size=8, color="FF1F3864"),
           align=al(h="left", wrap=True), border=bdr)
        max_lines_cand = max(max_lines_cand, txt.count("\n") + 1)
    ws.row_dimensions[row].height = max(13.0 * max_lines_cand, 80.0)
    row += 1

    # ─ 買い目シンプル表示行（combo順・確率%） ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "買い目リスト\n（combo順・確率%）",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al(wrap=True))
    max_lines_simple = 3
    for i, rd in enumerate(all_race_data):
        bet   = rd.get("bet_suggestions", {}) or {}
        cands = bet.get("candidates", [])

        skip        = bet.get("skip", False)
        skip_reason = bet.get("skip_reason", "")

        if not cands:
            wc(row, 4+i, "候補なし",
               fill=sf("FFDDDDDD"), font=fn(bold=False, size=8, color="FF808080"),
               align=al(h="left", wrap=True), border=bdr)
            continue
        def _combo_sort_key(c):
            parts = str(c.get("combo", "")).split("-")
            try:
                return [int(p) for p in parts]
            except ValueError:
                return [99, 99, 99]
        # 艇番若い順ソート
        sorted_cands = sorted(cands, key=lambda c: [int(x) for x in c.get("combo", "9-9-9").split("-")])
        # ── 折り返しペアをフラグベースで事前に対応付け ──────────────────────
        # is_orkaeshi フラグを持つ側が見つかった時点で両方向に登録する。
        # 「base_key not in paired」を外すことで、本体側が先に処理された場合でも
        # 折返側が後から正しく上書き登録される。
        combo_to_cand = {c["combo"]: c for c in sorted_cands}
        # is_orkaeshi フラグ優先でペアを構築（フラグ復元済みのため確実）
        paired = {}
        for c in sorted_cands:
            key = c["combo"]
            parts = key.split("-")
            if len(parts) != 3:
                continue
            f, s, t = parts
            if c.get("is_orkaeshi"):
                base_key = f"{s}-{f}-{t}"
                if base_key in combo_to_cand:
                    paired[key]      = base_key
                    paired[base_key] = key
            elif c.get("is_orkaeshi_23"):
                base_key = f"{f}-{t}-{s}"
                if base_key in combo_to_cand:
                    paired[key]      = base_key
                    paired[base_key] = key

        # ── ペア対応付けを使ってlines生成 ────────────────────────────────────
        used_simple = set()
        lines = []
        for c in sorted_cands:
            key = c["combo"]
            if key in used_simple:
                continue
            parts = key.split("-")
            if len(parts) != 3:
                used_simple.add(key)
                lines.append(f"{key}（{c['prob_pct']:.1f}%）")
                continue
            f, s, t = parts

            # 折り返しペアも単独行として個別表示（＝表記なし）
            used_simple.add(key)
            lines.append(f"{key}（{c['prob_pct']:.1f}%）")
        total = len(sorted_cands)
        eg       = bet.get("entry_grade") or {}
        eg_grade = eg.get("grade", "?")
        eg_bg, eg_fc = eg.get("fill", (FILL_JF_BET, "FF1F3864"))

        if skip:
            skip_label = skip_reason.split("\n")[0] if skip_reason else "⛔見送り推奨"
            txt = (
                f"{skip_label}\n"
                f"─────────\n"
                + "\n".join(lines)
            )
            cell_fill  = "FFFCE4D6"
            cell_color = "FF7F0000"
        else:
            eg_roi = eg.get("roi")
            if eg_grade != "?":
                roi_str = f"ROI{eg_roi*100:.0f}%+" if eg_roi else ""
                badge   = f"参加グレード\n{eg_grade}（{roi_str}）\n─────────\n"
            else:
                badge   = ""
            txt        = badge + "\n".join(lines)
            cell_fill  = eg_bg
            cell_color = eg_fc
        wc(row, 4+i, txt,
           fill=sf(cell_fill), font=fn(bold=True, size=9, color=cell_color),
           align=al(h="left", wrap=True), border=bdr)
        max_lines_simple = max(max_lines_simple, txt.count("\n") + 1)
    ws.row_dimensions[row].height = max(13.0 * max_lines_simple, 60.0)
    row += 1

    # ─ 理想合成オッズ行 ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "理想合成オッズ",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        bet  = rd.get("bet_suggestions", {}) or {}
        tso  = bet.get("theory_syn_odds")
        mr   = bet.get("margin_ratio")
        skip = bet.get("skip", False)
        skip_reason = bet.get("skip_reason", "")

        if skip:
            # 見送りレースは合成オッズ行に何も表示しない
            # （買い目リスト行に見送り理由＋参考買い目が既に表示されているため）
            txt    = "-"
            bg     = "FFDDDDDD"
            fc_col = "FF808080"
        elif tso:
            ev_warn     = bet.get("ev_warning", False)
            ev_warn_msg = bet.get("ev_warning_msg", "")
            if ev_warn:
                txt = (
                    f"理想合成オッズ\n{tso}倍\n"
                    f"⚠ 期待値基準を下回っています\n"
                    f"回収重視→見送り推奨\n"
                    f"的中重視→参考買い目を使用可"
                )
                bg, fc_col = "FFFFF2CC", "FF7F3F00"
            else:
                # 【⑨】ケリー賭け比率を合成オッズ行に追記
                _kelly2 = bet.get("kelly") or {}
                _kelly_pct = _kelly2.get("kelly_pct", "")
                _kelly_label = _kelly2.get("kelly_label", "")
                _kelly_line = f"\n推奨賭け比率: {_kelly_pct}（{_kelly_label}）" if _kelly_pct else ""
                txt = f"理想合成オッズ\n{tso}倍{_kelly_line}"
                _eg2 = bet.get("entry_grade") or {}
                _g2  = _eg2.get("grade", "?")
                _b2, _f2 = _eg2.get("fill", ("", ""))
                if _g2 in ("S", "A") and _b2:
                    bg, fc_col = _b2, _f2
                elif _g2 == "B":
                    bg, fc_col = "FFFFF2CC", "FF7F3F00"
                elif _g2 in ("C", "D"):
                    bg, fc_col = "FFFCE4D6", "FF7F0000"
                elif mr and mr >= 2.0:
                    bg, fc_col = "FFE2EFDA", "FF1D5730"
                elif mr and mr >= 1.2:
                    bg, fc_col = "FFFFF2CC", "FF7F3F00"
                else:
                    bg, fc_col = "FFFCE4D6", "FF7F0000"
        else:
            txt    = "計算不能"
            bg     = "FFDDDDDD"
            fc_col = "FF808080"
        wc(row, 4+i, txt,
           fill=sf(bg), font=fn(bold=True, size=16 if not skip else 11, color=fc_col),
           align=al(h="center", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 42.0
    row += 1

    # ─ 展示後判断ガイド行 ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "展示後の判断フロー",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        bet   = rd.get("bet_suggestions", {}) or {}
        stype = bet.get("scenario_type", "-")
        if stype == "逃げ軸流し":
            guide = "① 展示で1号艇の伸び確認\n   良ければそのまま採用\n   悪ければ飛び組に組み替え"
        elif stype == "飛び軸":
            guide = "① 展示で1号艇の伸び確認\n   悪ければ飛び組採用\n   良ければ逃げ組に組み替え"
        else:
            guide = "① 展示で1号艇の伸び確認\n   良→逃げ組採用\n   悪→飛び組採用"
        txt = (
            f"{guide}\n"
            f"② 実オッズで合成オッズを計算\n"
            f"③ 必要合成オッズと比較 → Step3"
        )
        wc(row, 4+i, txt,
           fill=sf(FILL_JF_NOTE), font=fn(bold=False, size=9, color="FF7F3F00"),
           align=al(h="left", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 80.0
    row += 1

    # ════════════════════════════════════════════════════════════════════



    # FLY退避値の復元ログ
    if fly_input_backup:
        print(f"  ✈️  FLY手入力値を復元済み: {len(fly_input_backup)}件")

# ============================================================
# 予想ログ保存（Step0 / refine_tenji.py 連携用）
# ============================================================
def _build_prediction_entry(race_no, bet_suggestions, race_judgment):
    """
    1レース分の予想ログエントリ（dict）を生成して返す。
    ファイルへの書き込みは行わない。_flush_prediction_log() で一括書き込みする。
    """
    _combos_raw = bet_suggestions.get("combos", [])
    combos_full = [
        {
            "combo":            c["combo"],
            "prob":             round(float(c["prob"]), 6),
            "theoretical_odds": c.get("theoretical_odds"),
            "hybrid_score":     round(float(c.get("hybrid_score", 0)), 5),
            "top_scenario":     c.get("top_scenario", "-"),
        }
        for c in _combos_raw
        if float(c.get("prob", 0)) >= 0.0005
    ]
    return {
        "race_no":          int(race_no) if str(race_no).isdigit() else race_no,
        "buy_list":         bet_suggestions.get("buy_list", []),
        "point_count":      bet_suggestions.get("point_count", 0),
        "candidates":       bet_suggestions.get("candidates", []),
        "axis_candidates":  bet_suggestions.get("axis_candidates", []),
        "himo_candidates":  bet_suggestions.get("himo_candidates", []),
        "comment":          bet_suggestions.get("comment", ""),
        "rank":             (race_judgment or {}).get("rank",     bet_suggestions.get("rank",     "-")),
        "score":            (race_judgment or {}).get("score",    bet_suggestions.get("score",    0)),
        "strategy":         (race_judgment or {}).get("strategy", bet_suggestions.get("strategy", "")),
        "ryotate_verdict":  bet_suggestions.get("ryotate_verdict", "-"),
        "ryotate_reason":   bet_suggestions.get("ryotate_detail", {}).get("reason", ""),
        "himo_are_verdict": ((race_judgment or {}).get("himo_are") or {}).get("verdict", "対象外"),
        "himo_are_mcp":     ((race_judgment or {}).get("himo_are") or {}).get("max_combo_prob"),
        "himo_are_est_odds":((race_judgment or {}).get("himo_are") or {}).get("est_top_odds"),
        "himo_are_cc":      ((race_judgment or {}).get("himo_are") or {}).get("circle_concentration"),
        "combos_full":      combos_full,
        # 結果欄（レース後に手動記入）- 初期値はNone
        "result_1st": None,
        "result_2nd": None,
        "result_3rd": None,
        "hit":        None,
        "dividend":   None,
    }


def _save_prediction_log(venue, race_date, race_no, results, bet_suggestions, race_judgment=None):
    """
    予想ログを logs/YYYY-MM-DD_会場名.json に保存する。
    refine_tenji.py が candidates / buy_list を参照するために必須。
    check_ev.py が combos_full を使って当日オッズとEV計算を行う。
    レース後に result_1st / result_2nd / result_3rd / hit / dividend を手動記入すること。

    【高速化】エントリ生成のみ行い、ファイル書き込みは _flush_prediction_log() に委譲。
    ただし単体呼び出し時の後方互換のため、単独でも書き込む。
    """
    logs_dir = pathlib.Path(__file__).parent.parent / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    date_str = str(race_date).replace("/", "-")[:10]
    log_path = logs_dir / f"{date_str}_{venue}.json"

    log_data = {"venue": venue, "date": date_str, "races": []}
    if log_path.exists():
        try:
            with open(log_path, encoding="utf-8") as f:
                log_data = json.load(f)
        except Exception:
            pass

    new_entry = _build_prediction_entry(race_no, bet_suggestions, race_judgment)
    # 既存エントリの結果欄を引き継ぐ
    for ex in log_data.get("races", []):
        if str(ex.get("race_no")) == str(race_no):
            for k in ("result_1st", "result_2nd", "result_3rd", "hit", "dividend"):
                new_entry[k] = ex.get(k)
            log_data["races"].remove(ex)
            break
    log_data.setdefault("races", []).append(new_entry)

    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(log_data, f, ensure_ascii=False, indent=2)
    print(f"  📝 予想ログ保存: {log_path.name} ({race_no}R)")


def _flush_prediction_log(venue, race_date, entries):
    """
    複数レース分のエントリをまとめて1回でファイルに書き込む（高速化版）。
    entries: list of (race_no, bet_suggestions, race_judgment)
    """
    if not entries:
        return
    logs_dir = pathlib.Path(__file__).parent.parent / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    date_str = str(race_date).replace("/", "-")[:10]
    log_path = logs_dir / f"{date_str}_{venue}.json"

    # 既存ログを1回だけ読み込む
    log_data = {"venue": venue, "date": date_str, "races": []}
    if log_path.exists():
        try:
            with open(log_path, encoding="utf-8") as f:
                log_data = json.load(f)
        except Exception:
            pass

    # 既存エントリをrace_noをキーにした辞書に変換
    existing_map = {str(e.get("race_no")): e for e in log_data.get("races", [])}

    new_races = []
    for race_no, bet_suggestions, race_judgment in entries:
        new_entry = _build_prediction_entry(race_no, bet_suggestions, race_judgment)
        # 既存の結果欄を引き継ぐ
        ex = existing_map.get(str(race_no), {})
        for k in ("result_1st", "result_2nd", "result_3rd", "hit", "dividend"):
            if ex.get(k) is not None:
                new_entry[k] = ex[k]
        new_races.append(new_entry)
        print(f"  📝 予想ログ蓄積: {race_no}R")

    log_data["races"] = new_races
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(log_data, f, ensure_ascii=False, indent=2)
    print(f"  💾 予想ログ一括保存: {log_path.name} ({len(new_races)}R分)")


# ============================================================
# 回収率バックテスト（ログファイルから集計）
# ============================================================
def calc_roi_from_logs(logs_dir=None, strategy_filter=None):
    """
    logs/ フォルダの予想ログ（JSON）から回収率バックテストを実行する。

    レース後に手動記入した result_1st/result_2nd/result_3rd/hit/dividend を読み込み、
    戦略タイプ・ランク別の回収率（ROI）を集計して返す。

    【使い方】
        from load_race import calc_roi_from_logs
        summary = calc_roi_from_logs()
        print(summary)

    【出力形式】
        {
          "total_bets":    全ベット点数合計,
          "total_hits":    3連単的中数,
          "total_cost":    総購入金額（100円/点換算）,
          "total_payout":  総払戻金額,
          "roi":           回収率 (払戻/購入, 例: 1.23 = 123%),
          "hit_rate":      3連単的中率 (0〜1),
          "by_rank":       ランク別集計 {"S": {...}, "A": {...}, ...},
          "by_venue":      会場別集計,
          "skip_races":    見送り推奨（ランクD）だったレース数,
          "missing_logs":  結果未記入のレース数（dividend=None）,
          "details":       全レース詳細リスト,
        }

    Parameters
    ----------
    logs_dir       : Path or str  ログフォルダ（省略時: ../logs/）
    strategy_filter: str or None  "全速型"等で絞り込み（Noneは全戦略）
    """
    if logs_dir is None:
        logs_dir = pathlib.Path(__file__).parent.parent / "logs"
    logs_dir = pathlib.Path(logs_dir)

    if not logs_dir.exists():
        print(f"  ⚠️  logsフォルダが存在しません: {logs_dir}")
        return None

    log_files = sorted(logs_dir.glob("*.json"))
    if not log_files:
        print(f"  ⚠️  ログファイルが見つかりません: {logs_dir}")
        return None

    # 集計用変数
    total_bets    = 0
    total_hits    = 0
    total_cost    = 0      # 100円/点換算
    total_payout  = 0
    missing_logs  = 0
    skip_races    = 0
    details       = []

    by_rank   = {r: {"bets": 0, "hits": 0, "cost": 0, "payout": 0, "races": 0}
                 for r in ("S", "A", "B", "C", "D", "-")}
    by_venue  = {}

    for log_file in log_files:
        try:
            with open(log_file, encoding="utf-8") as f:
                log_data = json.load(f)
        except Exception as e:
            print(f"  ⚠️  ログ読み込みエラー: {log_file.name} ({e})")
            continue

        venue    = log_data.get("venue", "不明")
        date_str = log_data.get("date", "")

        for entry in log_data.get("races", []):
            race_no    = entry.get("race_no", "?")
            buy_list   = entry.get("buy_list", [])
            point_count = entry.get("point_count", len(buy_list))
            hit        = entry.get("hit")         # True/False/None
            dividend   = entry.get("dividend")    # 払戻金額（100円単位）/ None
            rank       = entry.get("rank", "-")   # ランク（S/A/B/C/D）

            # 結果未記入のレースはカウントのみ
            if dividend is None:
                missing_logs += 1
                continue

            cost   = point_count * 100  # 100円/点
            payout = int(dividend) if hit else 0

            total_bets   += point_count
            total_cost   += cost
            total_payout += payout
            if hit:
                total_hits += 1

            # 見送りレース（ランクD or 買い目0点）をカウント
            if rank == "D" or point_count == 0:
                skip_races += 1

            # ランク別
            r_key = rank if rank in by_rank else "-"
            by_rank[r_key]["bets"]   += point_count
            by_rank[r_key]["hits"]   += (1 if hit else 0)
            by_rank[r_key]["cost"]   += cost
            by_rank[r_key]["payout"] += payout
            by_rank[r_key]["races"]  += 1

            # 会場別
            if venue not in by_venue:
                by_venue[venue] = {"bets": 0, "hits": 0, "cost": 0, "payout": 0, "races": 0}
            by_venue[venue]["bets"]   += point_count
            by_venue[venue]["hits"]   += (1 if hit else 0)
            by_venue[venue]["cost"]   += cost
            by_venue[venue]["payout"] += payout
            by_venue[venue]["races"]  += 1

            details.append({
                "date":       date_str,
                "venue":      venue,
                "race_no":    race_no,
                "rank":       rank,
                "buy_list":   buy_list,
                "point_count": point_count,
                "hit":        hit,
                "dividend":   dividend,
                "cost":       cost,
                "payout":     payout,
                "roi":        round(payout / cost, 3) if cost > 0 else 0,
            })

    if total_cost == 0:
        print("  ⚠️  結果が記入されたレースが見つかりません。")
        print("  　  logs/*.json の result_1st/result_2nd/result_3rd/hit/dividend を記入してください。")
        return None

    roi = round(total_payout / total_cost, 4) if total_cost > 0 else 0.0
    hit_rate = round(total_hits / max(len(details), 1), 4)

    # ランク別ROI計算
    for rk_data in by_rank.values():
        rk_data["roi"] = round(rk_data["payout"] / max(rk_data["cost"], 1), 4)
        rk_data["hit_rate"] = round(rk_data["hits"] / max(rk_data["races"], 1), 4)

    # 会場別ROI計算
    for vk_data in by_venue.values():
        vk_data["roi"] = round(vk_data["payout"] / max(vk_data["cost"], 1), 4)
        vk_data["hit_rate"] = round(vk_data["hits"] / max(vk_data["races"], 1), 4)

    summary = {
        "total_bets":   total_bets,
        "total_hits":   total_hits,
        "total_cost":   total_cost,
        "total_payout": total_payout,
        "roi":          roi,
        "roi_pct":      f"{roi*100:.1f}%",
        "hit_rate":     hit_rate,
        "hit_rate_pct": f"{hit_rate*100:.1f}%",
        "total_races":  len(details),
        "missing_logs": missing_logs,
        "skip_races":   skip_races,
        "by_rank":      by_rank,
        "by_venue":     by_venue,
        "details":      details,
    }

    # ── コンソール出力 ──
    sep()
    print("  📊 回収率バックテスト結果")
    sep()
    print(f"  対象レース数   : {len(details)}")
    print(f"  総ベット点数   : {total_bets}")
    print(f"  総購入金額     : {total_cost:,}円")
    print(f"  総払戻金額     : {total_payout:,}円")
    print(f"  3連単的中数    : {total_hits}（的中率 {hit_rate*100:.1f}%）")
    print(f"  回収率         : {roi*100:.1f}%  ({'✅プラス' if roi >= 1.0 else '❌マイナス'})")
    print()
    print("  ランク別:")
    for rk, rk_data in by_rank.items():
        if rk_data["races"] == 0:
            continue
        print(f"    [{rk}] {rk_data['races']:3}レース | "
              f"的中{rk_data['hit_rate']*100:.0f}% | "
              f"ROI {rk_data['roi']*100:.1f}% | "
              f"{rk_data['cost']:,}→{rk_data['payout']:,}円")
    print()
    print("  会場別（ROI上位）:")
    venue_sorted = sorted(by_venue.items(), key=lambda x: x[1]["roi"], reverse=True)
    for vn, vk_data in venue_sorted[:5]:
        print(f"    {vn:6} | {vk_data['races']:3}レース | "
              f"ROI {vk_data['roi']*100:.1f}% | "
              f"的中{vk_data['hit_rate']*100:.0f}%")
    sep()

    return summary



def main():
    parser = argparse.ArgumentParser(description="ボートリサーチ新聞 全レース一括書き込み")
    parser.add_argument("--venue",      type=str, default=None, help="会場名 (例: 大村)")
    parser.add_argument("--race",       type=int, default=None, help="レース番号 (省略時: 全レース)")
    parser.add_argument("--date",       type=str, default=None, help="日付 (例: 2026-02-15, 省略時: 最新CSV)")
    parser.add_argument("--newspaper",  action="store_true",    help="新聞作成をスキップせず実行 (fill_newspaper.py)")
    parser.add_argument("--png",        action="store_true",    help="PNG発行をスキップせず実行 (xlsx_to_png_interactive.py)")
    parser.add_argument("--grade",      type=str, default=None,
                        choices=["一般", "G1", "G2", "G3"],
                        help="レースグレード (例: G1 / SG)。省略時は対話式メニューで選択")
    args = parser.parse_args()

    sep()
    print("  ボートリサーチ新聞 全レース一括書き込み")
    print(f"  バージョン: {_SCRIPT_VERSION}")
    sep()

    # ════════════════════════════════════════════════════════════════════
    # 📦 LZH → CSV 自動変換（scripts/ フォルダの LZH を解凍してCSV化）
    #     - scripts/ フォルダに .lzh/.LZH ファイルがあれば自動処理
    #     - lzh_to_csv.py の変換ロジックをインラインで実行
    #     - 変換後の LZH は scripts/lzh_archive/ に退避（二重処理防止）
    # ════════════════════════════════════════════════════════════════════
    _lzh_files = (
        glob.glob(str(pathlib.Path(__file__).parent / "*.lzh")) +
        glob.glob(str(pathlib.Path(__file__).parent / "*.LZH"))
    )
    if _lzh_files:
        sep("-")
        print(f"  📦 LZHファイルを検出 ({len(_lzh_files)}件) → CSV変換を開始します")
        sep("-")
        _lzh_to_csv_script = pathlib.Path(__file__).parent / "lzh_to_csv.py"
        if not _lzh_to_csv_script.exists():
            print(f"  ⚠️  lzh_to_csv.py が見つかりません。LZH変換をスキップします。")
            print(f"     ({_lzh_to_csv_script})")
        else:
            try:
                _lzh_result = subprocess.run(
                    [sys.executable, str(_lzh_to_csv_script)],
                    capture_output=False,
                    check=False,
                    cwd=str(pathlib.Path(__file__).parent),  # scripts/ をカレントに
                )
                if _lzh_result.returncode == 0:
                    print(f"  ✅ LZH→CSV変換 完了")
                    # 変換済み LZH を lzh_archive/ フォルダに退避（二重変換防止）
                    _archive_dir = pathlib.Path(__file__).parent / "lzh_archive"
                    _archive_dir.mkdir(exist_ok=True)
                    for _lf in _lzh_files:
                        try:
                            if not pathlib.Path(_lf).exists():
                                # lzh_to_csv.py が内部で削除済みのケース → スキップ
                                print(f"  🗂️  退避スキップ（変換済みで削除済み）: {pathlib.Path(_lf).name}")
                                continue
                            shutil.move(_lf, str(_archive_dir / pathlib.Path(_lf).name))
                            print(f"  🗂️  退避: {pathlib.Path(_lf).name} → lzh_archive/")
                        except Exception as _mv_e:
                            print(f"  ⚠️  LZH退避失敗（続行します）: {_mv_e}")
                else:
                    print(f"  ⚠️  lzh_to_csv.py が終了コード {_lzh_result.returncode} で終了しました")
                    print(f"  ⚠️  CSV変換に失敗しましたが、既存CSVがあれば処理を続行します")
            except Exception as _lzh_e:
                print(f"  ❌ LZH→CSV変換でエラーが発生しました: {_lzh_e}")
                print(f"  ⚠️  既存CSVがあれば処理を続行します")
        sep("-")
        print()
    # ════════════════════════════════════════════════════════════════════

    # ════════════════════════════════════════════════════════════════════
    # ⚠️  Excel起動チェック（処理開始前）
    #     Excelが開いたまま処理すると最後の wb.save() で PermissionError になる。
    #     処理完了後に「Excelを閉じてください」と言われる時間ロスを防ぐため、
    #     最初にチェックして即座に警告する。
    # ════════════════════════════════════════════════════════════════════
    if EXCEL_FILE.exists():
        # Windows環境では .bak.xlsx への書き込みテストで開き確認を行う
        _lock_check_path = EXCEL_FILE.with_suffix(".~lock.xlsx")
        _is_excel_open = False

        # 方法①: ロックファイルの存在チェック（LibreOffice / Excel共通）
        if _lock_check_path.exists():
            _is_excel_open = True

        # 方法②: Windowsの場合は psutil でプロセスチェック
        if not _is_excel_open:
            try:
                import psutil
                excel_procs = [p for p in psutil.process_iter(["name"])
                               if p.info["name"] and "EXCEL" in p.info["name"].upper()]
                if excel_procs:
                    # プロセスは存在するが、対象ファイルを開いているかはファイル書き込みテストで確認
                    try:
                        with open(str(EXCEL_FILE), "a+b"):
                            pass
                    except PermissionError:
                        _is_excel_open = True
            except ImportError:
                pass  # psutil なし → 方法③へ

        # 方法③: ファイルへの書き込みテスト（最も確実）
        if not _is_excel_open:
            try:
                with open(str(EXCEL_FILE), "a+b"):
                    pass
            except PermissionError:
                _is_excel_open = True

        if _is_excel_open:
            print()
            print("  ╔══════════════════════════════════════════════════════╗")
            print("  ║  ❌  Excelが開いています！                            ║")
            print("  ║                                                      ║")
            print(f"  ║  📄 {str(EXCEL_FILE.name)[:48]:<48} ║")
            print("  ║                                                      ║")
            print("  ║  Excelを閉じてから、このスクリプトを再実行してください  ║")
            print("  ╚══════════════════════════════════════════════════════╝")
            print()
            return
        else:
            print(f"  ✅ Excel起動チェック: 閉じていることを確認 ({EXCEL_FILE.name})")
    # ════════════════════════════════════════════════════════════════════

    # 会場名
    venue = args.venue
    if not venue:
        # csv_outputから候補を収集
        files = sorted(glob.glob(str(CSV_DIR / "*.csv")))
        candidates_venue = []
        seen = set()
        for f in reversed(files):  # 新しい順
            v = os.path.basename(f).split("_")[0]
            if v and v not in seen:
                candidates_venue.append(v)
                seen.add(v)

        if not candidates_venue:
            print("❌ --venue を指定するか、csv_output/ にCSVを置いてください")
            return

        if len(candidates_venue) == 1:
            venue = candidates_venue[0]
            print(f"  🔍 会場を自動検出: {venue}")
        else:
            print()
            print("  ┌─────────────────────────────────────┐")
            print("  │  📍 会場を選択してください            │")
            print("  └─────────────────────────────────────┘")
            for i, v in enumerate(candidates_venue, 1):
                print(f"    {i}. {v}")
            while True:
                sel = input(f"\n  番号を入力 (1〜{len(candidates_venue)}): ").strip()
                if sel.isdigit() and 1 <= int(sel) <= len(candidates_venue):
                    venue = candidates_venue[int(sel) - 1]
                    break
                print("  ⚠️  正しい番号を入力してください")

    print(f"  🏁 会場: {venue}")

    # ── グレード選択 ──────────────────────────────────────────────────────
    # --grade 引数が指定されていれば即確定、なければ対話式メニュー
    _GRADE_CHOICES = ["一般", "G1", "G2", "G3", "SG"]
    if args.grade:
        race_grade = args.grade
    else:
        print()
        print("  ┌─────────────────────────────────────┐")
        print("  │  🏆 レースグレードを選択してください  │")
        print("  └─────────────────────────────────────┘")
        _grade_labels = {
            "一般": "一般戦（デフォルト）",
            "G1":   "G1 / SG（最高峰グレード・共通マスタ）",
            "G2":   "G2 / G3",
        }
        _grade_menu = ["一般", "G1", "G2"]   # 3択（SG=G1、G3=G2）
        for i, g in enumerate(_grade_menu, 1):
            suffix = " ← G3はこちら" if g == "G2" else (" ← SGもこちら" if g == "G1" else "")
            print(f"    {i}. {_grade_labels[g]}{suffix}")
        while True:
            sel = input(f"\n  番号を入力 (1〜{len(_grade_menu)}, Enter=1): ").strip()
            if sel == "":
                race_grade = "一般"
                break
            if sel.isdigit() and 1 <= int(sel) <= len(_grade_menu):
                race_grade = _grade_menu[int(sel) - 1]
                break
            print("  ⚠️  正しい番号を入力してください")
    print(f"  🏆 グレード: {race_grade}")

    # ── グレードレース選択時：マスタを毎回自動集計 ──────────────────────────
    # 一般戦以外（G1/SG・G2/G3）を選んだ場合、そのまま update_master.py を実行して
    # グレード別マスタCSVを最新化してからマスタ読み込みに進む。
    # ・一般戦は従来通りスキップ（毎日 update_master.py を別途実行する運用）
    # ・G1/SG・G2/G3はレース前に必ず最新データで集計し直す
    if race_grade != "一般":
        print()
        print(f"  ╔══════════════════════════════════════════════════════╗")
        print(f"  ║  📊 {race_grade}用マスタを集計します（毎回実行）              ║")
        print(f"  ║     update_master.py --grade {race_grade} を実行中...           ║")
        print(f"  ╚══════════════════════════════════════════════════════╝")
        print()
        _um_script = pathlib.Path(__file__).parent / "update_master.py"
        if not _um_script.exists():
            print(f"  ❌ update_master.py が見つかりません: {_um_script}")
            print(f"  ⚠️  マスタなしで続行します（一般戦マスタにフォールバック）")
        else:
            # SG は G1 マスタと共用なので --grade G1 を渡す
            _grade_arg = "G1" if race_grade == "SG" else race_grade
            _um_result = subprocess.run(
                [sys.executable, str(_um_script), "--grade", _grade_arg],
                check=False,
                cwd=str(pathlib.Path(__file__).parent),
            )
            if _um_result.returncode == 0:
                print()
                print(f"  ✅ {race_grade}用マスタ集計完了")
            else:
                print()
                print(f"  ⚠️  マスタ集計が異常終了しました（returncode={_um_result.returncode}）")
                print(f"  ⚠️  一般戦マスタにフォールバックして続行します")
        print()

    # ── 新聞作成・PNG発行 の選択 ──────────────────────────────────────────
    # --newspaper / --png フラグがあればそちら優先（bat経由など）、なければ対話式
    if args.newspaper:
        _run_newspaper = True
        _run_png       = args.png
    else:
        print()
        print("  ┌─────────────────────────────────────┐")
        print("  │  📰 新聞を作成しますか？              │")
        print("  │     (fill_newspaper.py を実行)        │")
        print("  └─────────────────────────────────────┘")
        _ans_newspaper = input("  [y/N] > ").strip().lower()
        _run_newspaper = _ans_newspaper in ("y", "yes")

        if _run_newspaper:
            print()
            print("  ┌─────────────────────────────────────┐")
            print("  │  🖼️  新聞をPNG発行しますか？          │")
            print("  │     (xlsx_to_png を実行)              │")
            print("  └─────────────────────────────────────┘")
            _ans_png = input("  [y/N] > ").strip().lower()
            _run_png = _ans_png in ("y", "yes")
        else:
            _run_png = False

    print()

    # マスタExcelを開く（ボートリサーチ_マスタ.xlsx）
    if not MASTER_FILE.exists():
        print(f"❌ マスタファイルが見つかりません: {MASTER_FILE}")
        return
    print(f"  📊 マスタデータ読み込み中...")
    try:
        wb_master = load_workbook(str(MASTER_FILE))
    except Exception as e:
        print(f"❌ マスタExcelを開けませんでした: {e}")
        print("   Excelが開いている場合は閉じてから再実行してください")
        return

    # マスタ読み込み
    course_master, player_master, ininage_master, venue_stats_master, venue_course_master, tenkai_venue_master, tenkai_national_master = load_masters(wb_master, race_grade)

    # 新聞出力用Excelを開く（ボートリサーチ新聞_軽量版.xlsx）
    if not EXCEL_FILE.exists():
        print(f"❌ Excelファイルが見つかりません: {EXCEL_FILE}")
        return
    try:
        wb = load_workbook(str(EXCEL_FILE))
    except Exception as e:
        print(f"❌ Excelを開けませんでした: {e}")
        print("   Excelが開いている場合は閉じてから再実行してください")
        return

    # CSV読み込み
    print(f"  📂 CSVを読み込み中...")
    df, race_date = load_csv(venue, args.race, args.date)
    if df is None or len(df) == 0:
        print(f"❌ {venue}のCSVが見つかりません: {CSV_DIR / venue}*.csv")
        return

    # モーターデータを出走表CSVから直接取得（scrape_motor.py 不要）
    motor_df = load_motor_csv(venue, None, race_df=df) if JIZEN_AVAILABLE else None
    if motor_df is None and JIZEN_AVAILABLE:
        print("  ⚠️  モーターデータなし。機力評価は '-' で出力します。")

    # レース番号一覧
    race_col = next((c for c in df.columns if "レース" in c or c == "R"), None)
    if race_col:
        race_nos = sorted(df[race_col].unique(), key=lambda x: int(x) if str(x).isdigit() else 99)
    else:
        race_nos = [str(args.race)] if args.race else ["1"]
    
    print(f"  📋 対象レース: {list(race_nos)}")
    print(f"  ✏️  Excel書き込み中: {EXCEL_FILE.name}")
    print()

    # 各レース書き込み
    tmp_image_paths = []
    all_race_data   = []  # 数値シート用にレースデータを蓄積
    _log_entries    = []  # 予想ログ一括書き込み用（12R分をまとめてflush）

    for race_no in race_nos:
        rno_int = int(race_no) if str(race_no).isdigit() else 0
        if args.race and rno_int != args.race:
            continue
        
        # このレースの選手データ取得
        if race_col:
            race_df = df[df[race_col].astype(str) == str(race_no)]
        else:
            race_df = df
        
        players = []
        for _, row in race_df.iterrows():
            players.append(dict(row))
        
        if not players:
            continue
        
        # 締め切り時刻（同レース全選手共通なので先頭行から取得）
        # 列名ゆれ（締切時刻 / 締切 / 締め切り時刻）に対応
        _deadline_raw = (
            players[0].get("締切時刻") or
            players[0].get("締切") or
            players[0].get("締め切り時刻") or ""
        )
        deadline = str(_deadline_raw).strip()
        deadline = None if deadline in ("", "None", "nan") else deadline
        
        # 指数計算（改善①②③: 会場別コースマスタ・シナリオ確率・動的ハイブリッド係数）
        results, slit, venue_stats, frame_2nd, race_judgment, bet_suggestions = calc_race_indices(
            venue, race_no, players, course_master, player_master, ininage_master,
            venue_stats_master, venue_course_master
        )
        
        # 倶楽部流 事前評価用メンバーデータ組み立て
        jizen_members = None
        jizen_eval_result = None
        if JIZEN_AVAILABLE:
            try:
                jizen_members = build_jizen_members(
                    results, course_master, player_master, motor_df, race_no
                )
                if jizen_members:
                    jizen_eval_result = evaluate_all(jizen_members)
            except Exception as e:
                print(f"  ⚠️  事前評価計算エラー ({race_no}R): {e}")

        # ── 印・買い目の確定（jizen有無に関わらず必ず実行）────────────────
        # ① first_prob_map確定のため bet_suggestions を一度計算
        # ② first_prob_map を使って6視点印スコアを確定（_apply_jizen_honmei）
        # ③ 確定した印を honmei_map に変換して買い目を再生成（印と完全連動）
        # ④ 最終 s1_prob で ryotate を再計算し、3択判定と確率の一貫性を保証
        try:
            bet_suggestions = _suggest_3rentan(
                results, race_judgment, jizen_eval=jizen_eval_result,
                tenkai_venue=tenkai_venue_master, tenkai_national=tenkai_national_master, venue=venue
            )
            _first_prob_map  = bet_suggestions.get("first_prob_map", {})
            _tobi_prob_final = (race_judgment.get("ryotate", {})
                                .get("tobi_score", 0) or 0)
            # venue_stats を渡して6人相互作用モデルが会場特性（決まり手比率等）を参照できるようにする
            _venue_stats_for_honmei = _calc_venue_stats(venue_stats_master, venue)
            _apply_jizen_honmei(results, _tobi_prob_final, jizen_eval_result,
                                first_prob_map=_first_prob_map,
                                venue_stats=_venue_stats_for_honmei,
                                race_judgment=race_judgment)  # トップレベル関数
            # 1号艇の「逃◎/逃○/逃△/逃×」を honmei_map に格納する際の変換。
            # 【重要】逃◎を攻め◎と同じ「◎」で格納すると honmei_scenario の逆引きで
            # 「◎が2艇存在」→後勝ちで2〜6号艇の攻め◎が本命扱いになり飛び軸に倒れる。
            # 1号艇には専用キー「逃」を使い攻め◎と衝突させない。
            # _suggest_3rentan v8.0 は s1_prob バイパスで1号艇の軸判断を行うため
            # honmei_map に1号艇の◎を入れなくても逃げ軸は正しく生成される。
            _nige_to_honmei = {"逃◎": "逃", "逃○": "逃○", "逃△": "逃△", "逃×": " "}
            _honmei_map = {}
            for r in results:
                w = str(r["waku"])
                h = r.get("honmei", " ")
                if w == "1":
                    _honmei_map[w] = _nige_to_honmei.get(h, " ")
                else:
                    _honmei_map[w] = h
            bet_suggestions = _suggest_3rentan(
                results, race_judgment,
                jizen_eval=jizen_eval_result,
                honmei_map=_honmei_map,
                tenkai_venue=tenkai_venue_master, tenkai_national=tenkai_national_master, venue=venue,
            )

            # ④ 印確定・買い目確定後の最終 s1_prob で ryotate を再計算
            #    これにより「3択判定の%」と「展開シナリオ確率」が同一モデルに統一される
            _s1_prob_final = bet_suggestions.get("s1_prob")
            if _s1_prob_final is not None:
                _tobi_scenario = race_judgment.get("ryotate", {})
                # tobi_scenario は calc_race_indices 内で生成した tobi オブジェクトが
                # race_judgment["ryotate"] に格納されているが、tobi_prob は ryotate から取れる
                _tobi_for_ryotate = {
                    "tobi_prob":   race_judgment["ryotate"].get("tobi_score", 30),
                    "main_threat": race_judgment["ryotate"].get("main_threat",
                                   bet_suggestions.get("fly_axes", [None])[0] or "-"),
                    "tobi_type":   race_judgment["ryotate"].get("tobi_type", "不明"),
                }
                _ryotate_final = _judge_ryotate(
                    race_judgment, _tobi_for_ryotate, {},
                    s1_prob=_s1_prob_final
                )
                race_judgment["ryotate"] = _ryotate_final

                # ◎艇番とs1_prob最大艇の乖離チェックを印確定後に最終更新
                _fpm = bet_suggestions.get("first_prob_map", {})
                if _fpm:
                    _top_waku = max(_fpm, key=_fpm.get)
                    _honmei_w = next(
                        (str(r["waku"]) for r in results
                         if r.get("honmei") == "◎" and r["waku"] != "1"), None
                    )
                    if _honmei_w and _honmei_w != _top_waku:
                        race_judgment["honmei_prob_mismatch"] = True
                        race_judgment["honmei_prob_mismatch_detail"] = (
                            f"◎={_honmei_w}号艇 vs 確率最大={_top_waku}号艇"
                            f"({_fpm.get(_top_waku, 0)*100:.1f}%)"
                        )
                    else:
                        race_judgment["honmei_prob_mismatch"] = False
                        race_judgment["honmei_prob_mismatch_detail"] = ""

        except Exception as e:
            print(f"  ⚠️  印・買い目確定エラー ({race_no}R): {e}")

        # 数値シート用にデータを蓄積
        all_race_data.append({
            "race_no":         race_no,
            "venue":           venue,
            "race_date":       race_date,
            "results":         results,
            "slit":            slit,
            "venue_stats":     venue_stats,
            "frame_2nd":       frame_2nd,
            "deadline":        deadline,
            "jizen_members":   jizen_members,
            "jizen_eval":      jizen_eval_result,
            "player_master":   player_master,
            "tmp_image_paths": tmp_image_paths,
            "race_judgment":   race_judgment,
            "bet_suggestions": bet_suggestions,
        })

        # ── EV計算（URL優先 → Excelフォールバック） ────────────────────────
        ev_suggestion = None
        actual_odds   = None
        try:
            # ACTUAL_ODDS_URL が明示設定されている場合のみ使用
            # （URL自動生成は無効化：オッズ取得は apply_ev.py に分離済み）
            _odds_url = ACTUAL_ODDS_URL

            if _odds_url:
                # 方法A: URLから実際のオッズをスクレイピング
                try:
                    from fetch_odds import fetch_odds_from_url_bs4, fetch_odds_from_url
                    actual_odds = (
                        fetch_odds_from_url_bs4(_odds_url) or
                        fetch_odds_from_url(_odds_url)
                    )
                    if actual_odds:
                        print(f"    🌐 実オッズ取得完了 ({len(actual_odds)}件)")
                    else:
                        print(f"    ⚠️  実オッズ取得失敗 → Excelにフォールバック")
                except ImportError:
                    print(f"    ⚠️  fetch_odds.py が見つかりません → Excelにフォールバック")

            if actual_odds is None and ODDS_FILEPATH.exists():
                # 方法B: 理論オッズExcel
                actual_odds = load_actual_odds_from_excel(str(ODDS_FILEPATH))
                print(f"    📄 理論オッズExcel使用")

            if actual_odds:
                combos        = _calc_3rentan_probs_v2(results)
                ev_list       = calc_ev_from_actual_odds(combos, actual_odds)
                ev_suggestion = suggest_by_ev(ev_list, min_ev=0.0, max_bets=8)
                skip_str = "⛔見送り推奨" if ev_suggestion["skip"] else f"✅{ev_suggestion['total_bets']}点"
                print(f"    📊 EV計算: {skip_str} {ev_suggestion['reason']}")
                bet_suggestions["ev_suggestion"] = ev_suggestion
        except Exception as e:
            print(f"  ⚠️  EV計算エラー ({race_no}R): {e}")
        # ─────────────────────────────────────────────────────────────────────

        # 予想ログ蓄積（ループ後に _flush_prediction_log で一括書き込み）
        _log_entries.append((race_no, bet_suggestions, race_judgment))

        # ── 見送り判定ログ ──────────────────────────────────────────────────
        if bet_suggestions.get("skip"):
            print(f"  {bet_suggestions.get('skip_reason', '⛔見送り')} ({race_no}R)")

        # ── 展示前フォーメーション コンソール出力 ──────────────────────────
        try:
            jf = bet_suggestions.get("jizen_formation", {})
            if jf and jf.get("formation"):
                jf_axes  = jf.get("axes", [])
                jf_himo  = jf.get("himo", [])
                jf_pts   = jf.get("point_count", 0)
                jf_form  = jf.get("formation", [])
                jf_tenk  = jf.get("tenkai_priority", [])
                jf_aisho = jf.get("aisho_axes", [])
                after_n  = max(jf_pts // 2, 6)

                # 1着軸に事前評価記号を付けて表示
                _in_sym  = (jizen_eval_result or {}).get("in_nige", [""] * 6)
                _a_sym   = (jizen_eval_result or {}).get("aisho",   [""] * 6)
                def _axis_label(w):
                    idx   = int(w) - 1
                    marks = []
                    if idx < len(_in_sym) and _in_sym[idx]:
                        marks.append(_in_sym[idx])
                    if idx < len(_a_sym) and _a_sym[idx]:
                        marks.append(_a_sym[idx])
                    return f"{w}号[{'/'.join(marks)}]" if marks else f"{w}号"

                axes_str = "  ".join(_axis_label(w) for w in jf_axes)
                himo_str = "  ".join(
                    f"{w}号{'[展]' if w in jf_tenk else ''}" for w in jf_himo
                )
                # 組み合わせを5点ずつ区切って表示
                form_lines = []
                for ci in range(0, len(jf_form), 5):
                    form_lines.append("    " + "  ".join(jf_form[ci:ci+5]))

                print()
                print(f"  ┌─ 📋 {race_no}R 展示前フォーメーション ({jf_pts}点) ─────────────")
                print(f"  │ 【1着軸】 {axes_str}")
                print(f"  │ 【ヒ モ】 {himo_str}")
                print(f"  │ 【組合せ】")
                for fl in form_lines:
                    print(f"  │{fl}")
                if jf_aisho:
                    print(f"  │ ②相性◎○: {'/'.join(jf_aisho)}号艇 → 展示タイム良ければ軸キープ")
                if jf_tenk:
                    print(f"  │ ⑤展開◎○: {'/'.join(jf_tenk)}号艇 → ヒモ優先キープ")
                print(f"  │ 展示後: タイム上位で軸を1〜2艇に絞り → 目安{after_n}点に圧縮")
                print(f"  └────────────────────────────────────────────────────")
        except Exception:
            pass  # 表示失敗は無視して続行
        # ─────────────────────────────────────────────────────────────────────

        print(f"  {race_no}R 処理中... → 完了")

    # 予想ログ一括書き込み（12R分を1回のファイルI/Oで保存）
    if _log_entries:
        _flush_prediction_log(venue, race_date, _log_entries)

    # 数値シート書き込み
    if all_race_data:
        write_numeric_sheet(wb, all_race_data, course_master, venue_course_master)

    # Excelを保存
    print()
    print(f"  💾 保存中...")

    # 保存前に自動バックアップ（前回の正常ファイルを .bak.xlsx として退避）
    bak_file = EXCEL_FILE.with_suffix(".bak.xlsx")
    try:
        if EXCEL_FILE.exists():
            shutil.copy2(str(EXCEL_FILE), str(bak_file))
            print(f"  🗂️  バックアップ作成: {bak_file.name}")
    except Exception as e:
        print(f"  ⚠️  バックアップ失敗（続行します）: {e}")

    try:
        wb.save(str(EXCEL_FILE))
        print(f"  ✅ 完了！「{venue}_数値」シートを確認してください")
    except PermissionError:
        print("❌ 保存エラー: Excelが開いています。閉じてから再実行してください")
        return

    # ── 新聞作成（fill_newspaper.py）─────────────────────────────────────
    _newspaper_ok = False
    if _run_newspaper:
        sep()
        print("  📰 新聞作成を開始します（fill_newspaper.py）")
        sep()
        _newspaper_script = pathlib.Path(__file__).parent / "fill_newspaper.py"
        if not _newspaper_script.exists():
            print(f"  ❌ fill_newspaper.py が見つかりません: {_newspaper_script}")
        else:
            try:
                _result = subprocess.run(
                    [sys.executable, str(_newspaper_script), "--venue", venue],
                    check=False
                )
                if _result.returncode == 0:
                    _newspaper_ok = True
                else:
                    print(f"  ⚠️  fill_newspaper.py が終了コード {_result.returncode} で終了しました")
                    print(f"  ⚠️  新聞作成に失敗したため PNG 発行をスキップします")
            except Exception as _e:
                print(f"  ❌ fill_newspaper.py の実行に失敗しました: {_e}")

    # ── PNG発行（xlsx_to_png_interactive.py）──────────────────────────────
    if _run_png and _newspaper_ok:
        sep()
        print("  🖼️  PNG発行を開始します（xlsx_to_png_interactive.py）")
        sep()
        _png_script = pathlib.Path(__file__).parent / "xlsx_to_png_interactive.py"
        if not _png_script.exists():
            print(f"  ❌ xlsx_to_png_interactive.py が見つかりません: {_png_script}")
        else:
            try:
                _result = subprocess.run(
                    [sys.executable, str(_png_script)],
                    check=False
                )
                if _result.returncode != 0:
                    print(f"  ⚠️  xlsx_to_png_interactive.py が終了コード {_result.returncode} で終了しました")
            except Exception as _e:
                print(f"  ❌ xlsx_to_png_interactive.py の実行に失敗しました: {_e}")

    # ── 完了通知（Windowsトースト）────────────────────────────────────────
    try:
        from win10toast import ToastNotifier
        _msg_parts = [f"{venue} 処理完了"]
        if _run_newspaper:
            _msg_parts.append("新聞作成 ✅")
        if _run_png:
            _msg_parts.append("PNG発行 ✅")
        ToastNotifier().show_toast(
            "ボートリサーチ新聞",
            " / ".join(_msg_parts),
            duration=8,
            threaded=True,
        )
    except ImportError:
        pass  # win10toast未インストールの場合はスキップ
    except Exception:
        pass

    sep()

if __name__ == "__main__":
    main()