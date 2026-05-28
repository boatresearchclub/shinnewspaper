"""
master_data.py
==============
ボートリサーチ_マスタ.xlsx から推薦ロジックに必要な全データを読み込む。

ロード対象シート:
    展開別残存_全国   → 決まり手×1着コース → 各コースの2着率・3着率
    展開別残存_会場別 → 同上（会場特化版）
    イン逃げ分析      → 会場別・逃げ時の枠別2着率・3着以内率
    会場統計          → 会場別荒れスコア・レース番号別荒れ・コース別1着率
    選手指数マスタ    → フォーム指数・時系列補正1着率・FLY影響度等
    会場別コースマスタ → 選手×会場×コースの時系列補正1着率

全データは初回ロード時にキャッシュ。
"""
from __future__ import annotations
import os
from functools import lru_cache
from collections import defaultdict
import openpyxl

_XLSX_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "ボートリサーチ_マスタ.xlsx"
)

def _find_xlsx():
    """
    scripts/bet_engine/ → scripts/ → プロジェクトルート/ の順に探す。
    （最大3階層上まで検索）
    """
    here = os.path.dirname(os.path.abspath(__file__))
    base = here
    for _ in range(3):
        p = os.path.join(base, "ボートリサーチ_マスタ.xlsx")
        if os.path.exists(p):
            return p
        base = os.path.dirname(base)
    raise FileNotFoundError(
        "ボートリサーチ_マスタ.xlsx が見つかりません\n"
        f"  探した場所: {here} およびその上位3フォルダ\n"
        "  プロジェクトルートに ボートリサーチ_マスタ.xlsx を置いてください。"
    )

@lru_cache(maxsize=1)
def _load_all() -> dict:
    path = _find_xlsx()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    result = {
        "tenkai_zanson_zenkoku":  _load_tenkai_zanson(wb, "展開別残存_全国"),
        "tenkai_zanson_kaijo":    _load_tenkai_zanson_kaijo(wb, "展開別残存_会場別"),
        "innige_bunseki":         _load_innige_bunseki(wb),
        "kaijo_stats":            _load_kaijo_stats(wb),
        "senshu_index":           _load_senshu_index(wb),
        "venue_course_master":    _load_venue_course_master(wb),
        "course_master":          _load_course_master(wb),
    }
    wb.close()
    return result

# ── 展開別残存_全国 ────────────────────────────────────────────
def _load_tenkai_zanson(wb, sheet_name):
    """
    returns: {kimete: {first_course_str: {course_str: {r2, r3, r123, n}}}}
    信頼度 < 0.5 は除外（サンプル不足）
    """
    d = defaultdict(lambda: defaultdict(dict))
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue
        kimete = str(row[1])
        c1     = str(row[2])
        n      = row[3]
        trust  = row[4]
        course = str(row[5])
        r2, r3, r123 = row[6], row[7], row[8]
        if trust is not None and float(trust) < 0.5:
            continue
        if r2 is not None:
            d[kimete][c1][course] = {
                "r2": float(r2), "r3": float(r3 or 0),
                "r123": float(r123 or 0), "n": int(n or 0)
            }
    return dict(d)

# ── 展開別残存_会場別 ─────────────────────────────────────────
def _load_tenkai_zanson_kaijo(wb, sheet_name):
    """
    returns: {venue: {kimete: {first_course_str: {course_str: {r2, r3, r123, n}}}}}
    """
    d = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 列構成: col0=None(空), col1=会場名, col2=決まり手, col3=1着コース,
        #         col4=レース数, col5=信頼度, col6=進入コース, col7=2着率, col8=3着率, col9=3着以内率
        if not row[1]: continue
        venue  = str(row[1])
        kimete = str(row[2])
        c1     = str(row[3])
        n      = row[4]
        trust  = row[5]
        course = str(row[6])
        r2, r3, r123 = row[7], row[8], row[9]
        if trust is not None and float(trust) < 0.5:
            continue
        if r2 is not None:
            d[venue][kimete][c1][course] = {
                "r2": float(r2), "r3": float(r3 or 0),
                "r123": float(r123 or 0), "n": int(n or 0)
            }
    return dict(d)

# ── イン逃げ分析 ──────────────────────────────────────────────
def _load_innige_bunseki(wb):
    """
    returns: {venue: {course_str: {r2, r3w}}}
    1号艇逃げ時の各コース2着率・3着以内率（会場別）
    """
    d = {}
    ws = wb['イン逃げ分析']
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        venue = str(row[0])
        d[venue] = {}
        for c in range(1, 7):
            r2  = row[2 + c - 1]
            r3w = row[8 + c - 1]
            if r2 is not None:
                d[venue][str(c)] = {"r2": float(r2), "r3w": float(r3w or 0)}
    return d

# ── 会場統計 ─────────────────────────────────────────────────
def _load_kaijo_stats(wb):
    """
    returns: {venue: {
        are_score: float,
        r_are: {race_no_str: float},
        c_win1: {course_str: float},
        in_nige_rate: float,
    }}
    """
    d = {}
    ws = wb['会場統計']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    for row in rows[1:]:
        if not row[0]: continue
        venue = str(row[0])
        d[venue] = {
            "are_score":    float(row[8])  if row[8]  is not None else 50.0,
            "in_nige_rate": float(row[2])  if row[2]  is not None else 0.55,
            "r_are":        {str(r): float(row[21+r-1]) for r in range(1,13)
                             if row[21+r-1] is not None},
            "c_win1":       {str(c): float(row[9+c-1])  for c in range(1,7)
                             if row[9+c-1]  is not None},
        }
    return d

# ── 選手指数マスタ ────────────────────────────────────────────
def _load_senshu_index(wb):
    """
    returns: {senshu_name: {
        form_index, win1_recent3, win1_recent5, win3_recent5,
        st_stable, fly_count, fly_days, fly_impact,
        time_win1, time_win3,
    }}
    ヘッダー行は行2（行1はタイトル）
    """
    d = {}
    ws = wb['選手指数マスタ']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    hdr  = rows[0]  # 列名
    col  = {str(v).strip().replace("\n",""): i for i, v in enumerate(hdr) if v}

    def _f(row, key, default=None):
        i = col.get(key)
        if i is None: return default
        v = row[i]
        return float(v) if v is not None else default

    for row in rows[1:]:
        if not row[1]: continue
        name = str(row[1]).replace(" ","").replace("\u3000","")
        d[name] = {
            "form_index":   _f(row, "フォーム指数"),
            "win1_r3":      _f(row, "直近3走1着率"),
            "win1_r5":      _f(row, "直近5走1着率"),
            "win3_r5":      _f(row, "直近5走3連対率"),
            "st_stable":    _f(row, "ST安定スコア"),
            "fly_count":    _f(row, "FLY数", 0),
            "fly_days":     _f(row, "FLY経過日数"),
            "fly_impact":   _f(row, "FLY影響度"),      # FLY後ST変化量の正規化
            "fly_st_delta": _f(row, "FLY後ST変化量"),  # 実値
            "time_win1":    _f(row, "時系列補正1着率"),  # ← ない列名かも
        }
    return d

# ── コース別マスタ（📊コース別マスタ）────────────────────────
def _load_course_master(wb):
    """
    📊コース別マスタ → {(選手名, コース番号str): {差し%, まくり%, 平均ST, ...}}
    load_race.pyの course_master と同じ形式。
    """
    d = {}
    ws = wb['📊コース別マスタ']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    hdr  = rows[0]
    col  = {str(v).strip().replace("\n", ""): i for i, v in enumerate(hdr) if v}

    def _f(row, key, default=None):
        i = col.get(key)
        if i is None: return default
        v = row[i]
        return float(v) if v is not None else default

    for row in rows[1:]:
        if not row[0] or row[1] is None: continue
        name   = str(row[0]).replace(" ", "").replace("\u3000", "")
        course = str(int(row[1]))
        d[(name, course)] = {
            "差し%":       _f(row, "差し%",       0.0),
            "まくり%":     _f(row, "まくり%",     0.0),
            "まくり差し%": _f(row, "まくり差し%", 0.0),
            "逃げ%":       _f(row, "逃げ%",       0.0),
            "抜き%":       _f(row, "抜き%",       0.0),
            "恵まれ%":     _f(row, "恵まれ%",     0.0),
            "コース別平均ST":  _f(row, "コース別平均ST"),
            "差され%":     _f(row, "差され%",     0.0),
            "捲られ%":     _f(row, "捲られ%",     0.0),
            "捲り差され%": _f(row, "捲り差され%", 0.0),
            "1着率":       _f(row, "1着率"),
            "3連対率":     _f(row, "3連対率"),
            "C1敗戦数":    _f(row, "C1敗戦数"),
            "コース別ST順位": _f(row, "コース別ST順位"),
            "時系列補正1着率":   _f(row, "時系列補正1着率"),
            "時系列補正3連対率": _f(row, "時系列補正3連対率"),
            "時系列有効走数":    _f(row, "時系列有効走数"),
        }
    return d


# ── 会場別コースマスタ ────────────────────────────────────────
def _load_venue_course_master(wb):
    """
    returns: {senshu_name: {venue: {course_str: {time_win1, time_win3, trust}}}}
    """
    d = defaultdict(lambda: defaultdict(dict))
    ws = wb['会場別コースマスタ']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    hdr  = rows[0]
    col  = {str(v).strip().replace("\n",""): i for i, v in enumerate(hdr) if v}

    def _f(row, key, default=None):
        i = col.get(key)
        if i is None: return default
        v = row[i]
        return float(v) if v is not None else default

    for row in rows[1:]:
        if not row[0] or not row[1] or row[2] is None: continue
        name  = str(row[0]).replace(" ","").replace("\u3000","")
        venue = str(row[1])
        course= str(int(row[2]))
        d[name][venue][course] = {
            "time_win1":  _f(row, "時系列補正1着率"),
            "time_win3":  _f(row, "時系列補正3連対率"),
            "trust":      _f(row, "信頼度", 0),
            "win1":       _f(row, "1着率"),
            "win3":       _f(row, "3連対率"),
        }
    return dict(d)

# ── 公開API ──────────────────────────────────────────────────

def get_course_master(senshu_name: str, course: int) -> dict | None:
    """選手×コースのコース別マスタデータを返す。"""
    name = senshu_name.replace(" ","").replace("\u3000","")
    return _load_all()["course_master"].get((name, str(course)))



def get_tenkai_zanson(kimete: str, first_course: int,
                      target_course: int, venue: str = "") -> dict | None:
    """
    展開別残存率を返す。会場別 → 全国フォールバック。
    kimete: "逃げ" / "差し" / "まくり" / "まくり差し" / "抜き" / "恵まれ"
    returns: {"r2": float, "r3": float, "r123": float} or None
    """
    data = _load_all()
    c1 = str(first_course)
    tc = str(target_course)

    # 会場別優先
    if venue:
        vd = data["tenkai_zanson_kaijo"].get(venue, {})
        res = vd.get(kimete, {}).get(c1, {}).get(tc)
        if res:
            return res

    # 全国フォールバック
    return data["tenkai_zanson_zenkoku"].get(kimete, {}).get(c1, {}).get(tc)


def get_innige_bunseki(venue: str, course: int) -> dict | None:
    """
    イン逃げ時の当該コースの2着率・3着以内率を返す（会場別）。
    returns: {"r2": float, "r3w": float} or None
    """
    data = _load_all()
    return data["innige_bunseki"].get(venue, {}).get(str(course))


def get_kaijo_stats(venue: str) -> dict | None:
    return _load_all()["kaijo_stats"].get(venue)


def get_race_are_score(venue: str, race_no: int) -> float:
    """レース番号別荒れスコアを返す（なければ会場平均）。"""
    ks = get_kaijo_stats(venue)
    if not ks:
        return 50.0
    return ks["r_are"].get(str(race_no), ks["are_score"])


def get_senshu_index(senshu_name: str) -> dict | None:
    name = senshu_name.replace(" ","").replace("\u3000","")
    return _load_all()["senshu_index"].get(name)


def get_venue_course_rate(senshu_name: str, venue: str, course: int) -> dict | None:
    """選手×会場×コースの時系列補正着率を返す。"""
    name = senshu_name.replace(" ","").replace("\u3000","")
    return _load_all()["venue_course_master"].get(name, {}).get(venue, {}).get(str(course))


# ── 動作確認 ─────────────────────────────────────────────────
if __name__ == "__main__":
    print("=== 展開別残存（逃げ×1号艇→各コース）===")
    for c in range(2, 7):
        r = get_tenkai_zanson("逃げ", 1, c)
        if r:
            print(f"  {c}号艇: 2着率={r['r2']:.3f} 3着率={r['r3']:.3f}")

    print("\n=== 展開別残存（まくり差し×3号艇→各コース）===")
    for c in [1,2,4,5,6]:
        r = get_tenkai_zanson("まくり差し", 3, c)
        if r:
            print(f"  {c}号艇: 2着率={r['r2']:.3f} 3着率={r['r3']:.3f}")

    print("\n=== イン逃げ分析（大村）===")
    for c in range(1, 7):
        r = get_innige_bunseki("大村", c)
        if r:
            print(f"  {c}号艇: 2着率={r['r2']:.3f} 3着以内率={r['r3w']:.3f}")

    print("\n=== 会場統計（大村）===")
    ks = get_kaijo_stats("大村")
    if ks:
        print(f"  荒れスコア={ks['are_score']}")
        print(f"  R別荒れ(1〜6R): {[ks['r_are'].get(str(r)) for r in range(1,7)]}")
        print(f"  コース別1着率: {ks['c_win1']}")

    print("\n=== 選手指数マスタ（サンプル）===")
    data = _load_all()
    sample = list(data["senshu_index"].items())[:3]
    for name, idx in sample:
        print(f"  {name}: form={idx['form_index']} fly_impact={idx['fly_impact']} time_win1={idx['time_win1']}")

    print("\n✅ master_data.py 動作確認完了")
