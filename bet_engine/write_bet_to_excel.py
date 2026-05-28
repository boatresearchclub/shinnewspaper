"""
write_bet_to_excel.py
=====================
recommend_bet.recommend() の結果を「各会場_数値」シートに書き込む。

【各会場_数値 シートの想定構造】
    行1: ヘッダー行
    行2〜: レースデータ（1レース1行）

    必須列（既存）:
        "会場名"   : 会場名
        "レース番号": 1〜12

    追加列（このスクリプトが付与）:
        "レース考察"   : 展開分析テキスト
        "買い目ロジック": パターン・軸艇・スコア詳細
        "買い目リスト" : 3連単の買い目（改行区切り）
        "買い目点数"   : int
        "信頼度"       : "高" / "中" / "低"

【使い方（load_race.py からの呼び出し想定）】
    from write_bet_to_excel import write_bets_to_sheet

    # 全レースの推薦結果を集めた辞書
    bet_results = {
        ("大村", 1): recommend_result_dict,
        ("大村", 2): recommend_result_dict,
        ...
    }
    write_bets_to_sheet(wb, "各会場_数値", bet_results)
    wb.save(EXCEL_PATH)

【単体実行】
    python write_bet_to_excel.py
    → ダミーデータでサンプルExcelを生成して動作確認
"""
from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── スタイル定義 ──────────────────────────────────────────────
_thin = Side(border_style="thin", color="1A2840")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

HDR_FILLS = {
    "レース考察":     PatternFill("solid", start_color="0D1F38", end_color="0D1F38"),
    "買い目ロジック": PatternFill("solid", start_color="102010", end_color="102010"),
    "買い目リスト":   PatternFill("solid", start_color="201808", end_color="201808"),
    "買い目点数":     PatternFill("solid", start_color="18101E", end_color="18101E"),
    "信頼度":         PatternFill("solid", start_color="1E0810", end_color="1E0810"),
}
HDR_FONT = Font(bold=True, color="D8E8FF", size=10)

CONF_STYLE = {
    "高": {"bg": "0A2A0A", "fg": "60E860", "bold": True},
    "中": {"bg": "2A2600", "fg": "D0C030", "bold": True},
    "低": {"bg": "2A0E00", "fg": "E06030", "bold": True},
}

CELL_FONT  = Font(color="B8C8D8", size=9)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CTR_ALIGN  = Alignment(horizontal="center", vertical="center")

NEW_COLS = ["レース考察", "買い目ロジック", "買い目リスト", "買い目点数", "信頼度"]
COL_WIDTHS = {
    "レース考察":     36,
    "買い目ロジック": 30,
    "買い目リスト":   20,
    "買い目点数":     8,
    "信頼度":         8,
}

# ─── メイン関数 ───────────────────────────────────────────────
def write_bets_to_sheet(
    wb:          openpyxl.Workbook,
    sheet_name:  str,
    bet_results: dict[tuple, dict],
    venue_col:   str = "A",
    race_col:    str = "B",
) -> None:
    """
    bet_results: {(venue_str, race_no_int): recommend()の戻り値}
    """
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠ シート '{sheet_name}' が見つかりません")
        return
    ws = wb[sheet_name]

    # ── 既存ヘッダーから列マッピングを取得 ──────────────────────
    header_row  = 1
    col_map: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    venue_cidx = col_map.get(venue_col) or _letter_to_idx(venue_col)
    race_cidx  = col_map.get(race_col)  or _letter_to_idx(race_col)

    # ── 追加列の配置（既存末尾の後ろ）──────────────────────────
    start_col = ws.max_column + 1
    # すでに追加列があれば上書き
    for name in NEW_COLS:
        if name in col_map:
            start_col = min(start_col, col_map[name])
            break

    add_col_map: dict[str, int] = {}
    for i, name in enumerate(NEW_COLS):
        cidx = col_map.get(name, start_col + i)
        add_col_map[name] = cidx

    # ── ヘッダー行に追加列名を書き込む ──────────────────────────
    for name, cidx in add_col_map.items():
        cell = ws.cell(row=header_row, column=cidx, value=name)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILLS.get(name, HDR_FILLS["レース考察"])
        cell.alignment = CTR_ALIGN
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(cidx)].width = COL_WIDTHS.get(name, 15)

    # ── データ行に書き込む ───────────────────────────────────────
    written = 0
    for row in range(header_row + 1, ws.max_row + 1):
        venue   = ws.cell(row=row, column=venue_cidx).value
        race_no = ws.cell(row=row, column=race_cidx).value
        if not venue:
            continue
        try:
            rno = int(race_no)
        except (TypeError, ValueError):
            continue

        key = (str(venue), rno)
        bd  = bet_results.get(key)
        if not bd:
            continue

        ec    = bd.get("excel_cells", {})
        conf  = ec.get("信頼度", "低")
        lines = max(
            ec.get("レース考察", "").count("\n") + 1,
            ec.get("買い目ロジック", "").count("\n") + 1,
            ec.get("買い目リスト",   "").count("\n") + 1,
        )

        for name, cidx in add_col_map.items():
            val = ec.get(name, "")
            cell = ws.cell(row=row, column=cidx, value=val)
            cell.border = BORDER

            if name == "信頼度":
                st = CONF_STYLE.get(conf, {"bg": "1A1A1A", "fg": "888888", "bold": False})
                cell.font      = Font(bold=st["bold"], color=st["fg"], size=10)
                cell.fill      = PatternFill("solid", start_color=st["bg"], end_color=st["bg"])
                cell.alignment = CTR_ALIGN
            elif name == "買い目点数":
                cell.font      = Font(color="A0C0E0", size=10, bold=True)
                cell.alignment = CTR_ALIGN
            else:
                cell.font      = CELL_FONT
                cell.alignment = WRAP_ALIGN

        ws.row_dimensions[row].height = max(15, lines * 13)
        written += 1

    print(f"  ✅ '{sheet_name}': {written} レース分の買い目を書き込みました")


def _letter_to_idx(col: str) -> int:
    """'A'→1, 'B'→2 など"""
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(col)


# ─── 単体テスト（ダミーExcel生成）────────────────────────────
if __name__ == "__main__":
    import sys, os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from evaluate_jizen import evaluate_all
    from recommend_bet import recommend

    _sd = {"st_stable_score": 70.0, "fly_count": 0, "fly_days": None,
           "late_count": 0, "st_count": 50}

    MEMBERS = {
        # レース1: インパターン
        ("大村", 1): [
            {"rate_1st_c1":0.70,"st_rank_c1":1.0,"star_rate":False,"nigé_rate":0.82,
             "attack_rate":0.0,"sashi_rate":0.0,"makuri_rate":0.0,"makuri_zashi_rate":0.0,
             "avg_st_self":0.13,"lose_sashi_rate":0.08,"lose_makuri_rate":0.05,
             "lose_rate_reliable":True,"motor_2rate":50.0,"diversity_rate":0.0,
             "jizaisei_rate":0.05,"star_kimete":False,"rate_3ren":0.85,"course_int":1,
             **{**_sd,"st_stable_score":90.0}},
            {"rate_1st_c1":0.08,"st_rank_c1":5.0,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.20,"sashi_rate":0.10,"makuri_rate":0.05,"makuri_zashi_rate":0.05,
             "avg_st_self":0.22,"motor_2rate":35.0,"diversity_rate":0.15,"jizaisei_rate":0.10,
             "star_kimete":False,"rate_3ren":0.35,"course_int":2,**{**_sd,"fly_count":1,"fly_days":75.0}},
            {"rate_1st_c1":0.08,"st_rank_c1":4.0,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.45,"sashi_rate":0.10,"makuri_rate":0.20,"makuri_zashi_rate":0.15,
             "avg_st_self":0.16,"motor_2rate":40.0,"diversity_rate":0.40,"jizaisei_rate":0.18,
             "star_kimete":False,"rate_3ren":0.50,"course_int":3,**_sd},
            {"rate_1st_c1":0.10,"st_rank_c1":3.0,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.55,"sashi_rate":0.05,"makuri_rate":0.40,"makuri_zashi_rate":0.15,
             "avg_st_self":0.15,"motor_2rate":48.0,"diversity_rate":0.45,"jizaisei_rate":0.22,
             "star_kimete":False,"rate_3ren":0.52,"course_int":4,**{**_sd,"st_stable_score":85.0}},
            {"rate_1st_c1":0.07,"st_rank_c1":5.5,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.30,"sashi_rate":0.05,"makuri_rate":0.25,"makuri_zashi_rate":0.10,
             "avg_st_self":None,"motor_2rate":None,"diversity_rate":0.20,"jizaisei_rate":0.10,
             "star_kimete":False,"rate_3ren":0.40,"course_int":5,
             **{**_sd,"st_stable_score":55.0,"late_count":4}},
            {"rate_1st_c1":0.05,"st_rank_c1":6.0,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.20,"sashi_rate":0.05,"makuri_rate":0.15,"makuri_zashi_rate":0.10,
             "avg_st_self":0.21,"motor_2rate":28.0,"diversity_rate":0.15,"jizaisei_rate":0.08,
             "star_kimete":True,"rate_3ren":0.30,"course_int":6,
             **{**_sd,"st_stable_score":40.0,"fly_count":2,"fly_days":50.0}},
        ],
        # レース2: 差しパターン
        ("大村", 2): [
            {"rate_1st_c1":0.55,"st_rank_c1":3.0,"star_rate":False,"nigé_rate":0.72,
             "attack_rate":0.0,"sashi_rate":0.0,"makuri_rate":0.0,"makuri_zashi_rate":0.0,
             "avg_st_self":0.18,"lose_sashi_rate":0.18,"lose_makuri_rate":0.08,
             "lose_rate_reliable":True,"motor_2rate":45.0,"diversity_rate":0.0,
             "jizaisei_rate":0.05,"star_kimete":False,"rate_3ren":0.75,"course_int":1,**_sd},
            {"rate_1st_c1":0.10,"st_rank_c1":1.5,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.65,"sashi_rate":0.55,"makuri_rate":0.05,"makuri_zashi_rate":0.05,
             "avg_st_self":0.12,"motor_2rate":55.0,"diversity_rate":0.60,"jizaisei_rate":0.25,
             "star_kimete":False,"rate_3ren":0.60,"course_int":2,**_sd},
            {"rate_1st_c1":0.08,"st_rank_c1":4.0,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.55,"sashi_rate":0.10,"makuri_rate":0.20,"makuri_zashi_rate":0.25,
             "avg_st_self":0.15,"motor_2rate":30.0,"diversity_rate":0.50,"jizaisei_rate":0.20,
             "star_kimete":False,"rate_3ren":0.55,"course_int":3,**_sd},
            {"rate_1st_c1":0.12,"st_rank_c1":2.5,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.30,"sashi_rate":0.05,"makuri_rate":0.45,"makuri_zashi_rate":0.35,
             "avg_st_self":0.13,"motor_2rate":35.0,"diversity_rate":0.45,"jizaisei_rate":0.20,
             "star_kimete":False,"rate_3ren":0.50,"course_int":4,**_sd},
            {"rate_1st_c1":0.09,"st_rank_c1":5.0,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.20,"sashi_rate":0.05,"makuri_rate":0.25,"makuri_zashi_rate":0.20,
             "avg_st_self":None,"motor_2rate":None,"diversity_rate":0.20,"jizaisei_rate":0.10,
             "star_kimete":False,"rate_3ren":0.45,"course_int":5,**_sd},
            {"rate_1st_c1":0.06,"st_rank_c1":6.0,"star_rate":False,"nigé_rate":0.0,
             "attack_rate":0.15,"sashi_rate":0.05,"makuri_rate":0.20,"makuri_zashi_rate":0.15,
             "avg_st_self":0.22,"motor_2rate":20.0,"diversity_rate":0.15,"jizaisei_rate":0.08,
             "star_kimete":True,"rate_3ren":0.30,"course_int":6,**_sd},
        ],
    }

    # ── ダミーExcel作成 ──────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "各会場_数値"

    # ヘッダー
    headers = ["会場名", "レース番号", "1着(1)", "2着(1)", "3着(1)",
               "逃げ評価", "相性(2)", "相性(3)", "相性(4)", "機力(1)",
               "安定(1)", "展開(4)", "展開(5)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    # データ行 + 買い目計算
    bet_results = {}
    row = 2
    for (venue, rno), members in MEMBERS.items():
        ev = evaluate_all(members)
        res = recommend(members, ev, venue=venue, race_no=rno)
        bet_results[(venue, rno)] = res

        ws.cell(row=row, column=1, value=venue)
        ws.cell(row=row, column=2, value=rno)
        # 評価記号サンプル
        for ci, sym in enumerate(ev["in_nige"][:1], 6):
            ws.cell(row=row, column=ci, value=sym)
        for ci, sym in enumerate(ev["aisho"][1:4], 7):
            ws.cell(row=row, column=ci, value=sym)
        row += 1

    # 買い目列を追加
    write_bets_to_sheet(wb, "各会場_数値", bet_results,
                        venue_col="A", race_col="B")

    out = "/mnt/user-data/outputs/bet_sample.xlsx"
    wb.save(out)
    print(f"\n  📁 出力: {out}")

    # 内容確認
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2["各会場_数値"]
    print("\n  ヘッダー行:")
    for cell in ws2[1]:
        if cell.value:
            print(f"    col{cell.column}({get_column_letter(cell.column)}): {cell.value}")
    print("\n  データ確認（行2・行3 の追加列）:")
    # 末尾5列のデータを確認
    last = ws2.max_column
    for r in [2, 3]:
        venue = ws2.cell(row=r, column=1).value
        rno   = ws2.cell(row=r, column=2).value
        conf  = ws2.cell(row=r, column=last).value
        cnt   = ws2.cell(row=r, column=last-1).value
        first_bet = ws2.cell(row=r, column=last-2).value
        if first_bet:
            first_bet = first_bet.split("\n")[0]
        print(f"    R{r} {venue} R{rno}: 信頼度={conf} 点数={cnt} 最初の買い目={first_bet}")
