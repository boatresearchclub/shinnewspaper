#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_bet_engine.py
=================
既存の load_race.py が生成した CSV を読み込み、
新エンジン（recommend_bet.py v7）で3連単買い目を計算して
独立した Excel ファイルに出力する。

【使い方】
  # 最新のCSVを自動検出
  python scripts/bet_engine/run_bet_engine.py

  # 会場指定
  python scripts/bet_engine/run_bet_engine.py --venue 大村

  # 日付指定
  python scripts/bet_engine/run_bet_engine.py --venue 大村 --date 2026-03-20

  # レース番号指定（単発テスト）
  python scripts/bet_engine/run_bet_engine.py --venue 大村 --race 5

【出力先】
  bet_engine_output/ フォルダ（既存システムとは完全別ファイル）
  例: bet_engine_output/大村_2026-03-20_買い目.xlsx

【フォルダ構成】
  scripts/
    bet_engine/           ← このフォルダ（既存から完全分離）
      run_bet_engine.py   ← エントリポイント（このファイル）
      recommend_bet.py    ← 確率モデルエンジン v7
      master_data.py      ← マスタ.xlsx 読み込み
      accuracy_data.py    ← accuracy_data.json 読み込み
  bet_engine_output/      ← 出力先（既存の出力と分離）
  ボートリサーチ_マスタ.xlsx  ← 共有（上書きなし・読み取りのみ）
  accuracy_data.json       ← 共有（上書きなし・読み取りのみ）
"""
import argparse
import glob
import os
import pathlib
import re
import sys
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ── パス設定 ──────────────────────────────────────────────────
_HERE       = pathlib.Path(__file__).parent          # scripts/bet_engine/
_SCRIPTS    = _HERE.parent                           # scripts/
_BASE       = _SCRIPTS.parent                        # プロジェクトルート
_CSV_DIR    = _SCRIPTS / "csv_output"
_OUTPUT_DIR = _BASE / "bet_engine_output"
_MASTER_XLS = _BASE / "ボートリサーチ_マスタ.xlsx"
_ACC_JSON   = _SCRIPTS / "accuracy_data.json"

# bet_engine/ 自身をパスに追加
sys.path.insert(0, str(_HERE))
# evaluate_jizen.py は scripts/ にある
sys.path.insert(0, str(_SCRIPTS))

# ── 依存モジュール読み込み ─────────────────────────────────────
try:
    from recommend_bet import recommend
    from master_data import _find_xlsx
    print("✅ recommend_bet / master_data 読み込み完了")
except ImportError as e:
    print(f"❌ モジュール読み込み失敗: {e}")
    sys.exit(1)

try:
    from evaluate_jizen import evaluate_all, calculate_diversity_rate
    JIZEN_OK = True
    print("✅ evaluate_jizen 読み込み完了")
except ImportError:
    JIZEN_OK = False
    print("⚠️  evaluate_jizen.py が見つかりません（事前評価スキップ）")

try:
    from master_data import _load_all as _load_master_data
    # マスタの選手指数・会場別コースをキャッシュ
    _MD = _load_master_data()
    MASTER_OK = True
    print("✅ ボートリサーチ_マスタ.xlsx 読み込み完了")
except Exception as e:
    MASTER_OK = False
    print(f"⚠️  マスタ読み込み失敗: {e}")


# ============================================================
# CSV 読み込み
# ============================================================
def load_csv(venue: str, date_str: str = "", race_no: str = "") -> tuple[pd.DataFrame | None, str, str]:
    """
    csv_output/ から対象CSVを読み込み (df, race_date, venue) を返す。
    会場名は CSV ファイル名か df の列から自動検出。
    """
    search_dirs = []
    if venue:
        search_dirs += [_CSV_DIR / venue, _CSV_DIR]
    else:
        search_dirs += [_CSV_DIR]

    def _find(pattern):
        for d in search_dirs:
            files = sorted(glob.glob(str(d / pattern)))
            if files: return files
        return []

    if date_str and venue:
        files = _find(f"{venue}_{date_str}.csv")
    elif venue:
        files = _find(f"{venue}_*.csv")
    else:
        all_files = sorted(glob.glob(str(_CSV_DIR / "**" / "*.csv"), recursive=True))
        files = [all_files[-1]] if all_files else []

    if not files:
        print(f"❌ CSVが見つかりません (venue={venue or '自動'}, date={date_str or '最新'})")
        return None, "", venue

    csv_path = files[-1]
    print(f"📂 CSV: {csv_path}")

    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
    df.columns = df.columns.str.strip()
    for col in df.columns:
        df[col] = df[col].astype(str).str.lstrip("'").str.strip()

    # 日付取得
    race_date = ""
    if "日付" in df.columns:
        dv = df["日付"].dropna()
        dv = dv[dv.str.strip() != ""]
        if len(dv) > 0:
            raw = dv.iloc[0].strip()
            m = re.sub(r"(\d{4})[/\-](\d{2})[/\-](\d{2}).*", r"\1-\2-\3", raw)
            if re.match(r"\d{4}-\d{2}-\d{2}", m):
                race_date = m
    if not race_date:
        m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(csv_path))
        race_date = m.group(1) if m else datetime.today().strftime("%Y-%m-%d")

    # 会場名取得
    if not venue:
        if "会場" in df.columns:
            venue = df["会場"].iloc[0].strip()
        else:
            venue = os.path.basename(csv_path).split("_")[0]

    # レース番号フィルタ
    if race_no:
        for col in ["レース番号", "R", "レース", "race_no"]:
            if col in df.columns:
                df = df[df[col].astype(str) == str(race_no)]
                break

    return df, race_date, venue


# ============================================================
# CSV → jizen_members 変換
# ============================================================
def _safe(v, d=0.0):
    try:
        s = str(v).replace("%","").strip()
        if s in ("", "None", "nan", "-", "★"): return d
        return float(s)
    except: return d

def _safe_n(v):
    try:
        s = str(v).replace("%","").strip()
        if s in ("", "None", "nan", "-", "★"): return None
        return float(s)
    except: return None

def build_jizen_members_simple(players: list[dict], venue: str, race_no,
                               motor_data: dict = None) -> list[dict]:
    """
    CSV の1レース分の選手リスト → recommend_bet.recommend() に渡すmembersを構築。
    load_race.py の build_jizen_members() の独立版。

    【修正】差し%・まくり%・平均STなどをCSV列ではなく
    📊コース別マスタ（ボートリサーチ_マスタ.xlsx）から取得する。
    CSVにはこれらの列が存在しないため。
    """
    if not MASTER_OK:
        return []

    _CM  = _MD.get("senshu_index", {})       # {選手名: {指数dict}}
    _VC  = _MD.get("venue_course_master", {}) # {選手名: {会場: {コース: {...}}}}
    _CCM = _MD.get("course_master", {})       # {(選手名, コース): {差し%, まくり%, ...}}

    members = []
    # 1号艇の逃げ率（相性計算用）
    nige_rate_1 = 0.7  # デフォルト

    for i, p in enumerate(players[:6]):
        name_raw = str(p.get("選手名","") or p.get("name","")).strip()
        name_raw = re.sub(r'\s*\d+\s*$', '', name_raw).strip()
        name = name_raw.replace("　","").replace(" ","")

        waku   = str(p.get("枠", p.get("艇番", str(i+1)))).strip()
        course_raw = str(p.get("想定コース","") or p.get("コース","")).strip()
        course = course_raw if course_raw not in ("","nan","None") else waku
        try: course_int = int(course)
        except: course_int = i+1

        # モーター2連対率
        motor_2rate = None
        if motor_data:
            motor_2rate = motor_data.get((int(waku) if waku.isdigit() else i+1),)
        if motor_2rate is None:
            raw_m = p.get("モーター2連率") or p.get("M2率") or p.get("モータ2連") or p.get("motor_2rate")
            motor_2rate = _safe_n(raw_m)

        # 選手指数マスタから取得（フォーム・FLY等）
        pm = _CM.get(name, {})
        # 会場別コースマスタから取得
        vc = _VC.get(name, {}).get(venue, {}).get(str(course_int), {})
        # コース別マスタ（このコース / 1コース）
        cm_c  = _CCM.get((name, str(course_int)), {})
        cm_c1 = _CCM.get((name, '1'), {})

        # rate_1st_c1（1コース時系列補正1着率 → コース別マスタC1から）
        ts1 = _safe_n(cm_c1.get("時系列補正1着率"))
        fl1 = _safe_n(cm_c1.get("1着率"))
        rate_1st_c1 = ts1 if ts1 is not None else (fl1 if fl1 is not None else 0.0)

        # ST関連 → コース別マスタから取得
        avg_st_self    = _safe_n(cm_c.get("コース別平均ST"))
        st_rank_c1     = _safe_n(cm_c1.get("コース別ST順位"))
        course_st_rank = _safe_n(cm_c.get("コース別ST順位"))
        if avg_st_self is None and course_st_rank is not None:
            avg_st_self = round(0.12 + (course_st_rank - 1) / 5 * 0.08, 4)

        # 決まり手・ST → 📊コース別マスタから取得（CSVには存在しない）
        cm = _CCM.get((name, str(course_int)), {})

        sashi_rate        = _safe(cm.get("差し%"),       0.0)
        makuri_rate       = _safe(cm.get("まくり%"),     0.0)
        makuri_zashi_rate = _safe(cm.get("まくり差し%"), 0.0)
        nige_rate_self    = _safe(cm.get("逃げ%"),       0.0)
        attack_rate       = sashi_rate + makuri_rate + makuri_zashi_rate

        # コース別平均ST（コース別マスタ優先、なければ選手指数マスタから推定）
        st_from_cm = _safe_n(cm.get("コース別平均ST"))
        if st_from_cm is not None:
            avg_st_self = st_from_cm
        # （avg_st_selfはST関連ブロックで既にセット済みの場合はそちら優先）

        # 1号艇被決まり手
        lose_sashi_rate  = None
        lose_makuri_rate = None
        lose_rate_reliable = False
        if i == 0:
            lose_sashi_rate  = _safe_n(cm.get("差され%"))
            lose_makuri_rate = _safe_n(cm.get("捲られ%"))
            c1_lose = _safe_n(cm.get("C1敗戦数"))
            lose_rate_reliable = (c1_lose is not None and c1_lose >= 10)
            nige_rate_1 = nige_rate_self if nige_rate_self > 0 else 0.7

        # rate_3ren（コース別3連対率）→ コース別マスタ優先
        rate_3ren = _safe(cm.get("3連対率"), 0.0)
        if rate_3ren == 0.0:
            rate_3ren = _safe(p.get("3連対率") or p.get("コース別3連対率"), 0.0)

        # FLY・安定性 → senshu_indexの実際のキーを使用
        fly_count  = int(_safe(pm.get("fly_count"), 0))
        fly_days   = _safe_n(pm.get("fly_days"))
        late_count = 0  # senshu_indexに出遅れキーなし
        st_count   = 50  # デフォルト
        st_stable  = _safe_n(pm.get("st_stable"))

        # 星マーク
        star_rate   = str(p.get("★1着率","")).strip() == "★"
        star_kimete = str(p.get("★決手","")).strip() == "★"

        # 展示タイム偏差値（展示後のみ有効）
        tenji_hensa = _safe_n(p.get("展示偏差値") or p.get("tenji_hensa"))

        # diversity_rate（自在性）
        diversity_rate = _safe(pm.get("自在性1着率") or pm.get("自在性\n1着率"), 0.0)
        jizaisei_rate  = _safe(pm.get("自在性加重1着率") or pm.get("自在性\n加重1着率"), 0.0)

        members.append({
            # イン逃げ
            "rate_1st_c1":   rate_1st_c1,
            "st_rank_c1":    st_rank_c1,
            "star_rate":     star_rate,
            # 相性
            "nigé_rate":          nige_rate_1,
            "attack_rate":        attack_rate,
            "sashi_rate":         sashi_rate,
            "makuri_rate":        makuri_rate,
            "makuri_zashi_rate":  makuri_zashi_rate,
            "avg_st_self":        avg_st_self,
            "lose_sashi_rate":    lose_sashi_rate,
            "lose_makuri_rate":   lose_makuri_rate,
            "lose_rate_reliable": lose_rate_reliable,
            # 機力
            "motor_2rate":   motor_2rate,
            # 自在性
            "diversity_rate": diversity_rate,
            "jizaisei_rate":  jizaisei_rate,
            "star_kimete":    star_kimete,
            # 安定性
            "st_stable_score": st_stable,
            "fly_count":       fly_count,
            "fly_days":        fly_days,
            "late_count":      late_count,
            "st_count":        st_count,
            # 展開
            "rate_3ren":     rate_3ren,
            # 買い目エンジン用
            "senshu_name":   name,
            "course_int":    course_int,
            "tenji_hensa":   tenji_hensa,
            # 表示用
            "_name_display": name_raw,
            "_waku":         waku,
            "_course_str":   course,
        })

    return members


# ============================================================
# Excel 出力
# ============================================================
_BOAT_FILL = {1:"FFFFFFFF",2:"FF1A1A1A",3:"FFCC0000",4:"FF2255CC",5:"FFDDCC00",6:"FF115511"}
_BOAT_FONT = {1:"FF000000",2:"FFFFFFFF",3:"FFFFFFFF",4:"FFFFFFFF",5:"FF000000",6:"FFFFFFFF"}

def _sf(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _fn(bold=True, size=9, color="FF000000"):
    return Font(name="Meiryo UI", size=size, bold=bold, color=color)

def _al(h="center", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

_thin = Side(style="thin", color="FFCCCCCC")
_bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def write_bet_sheet(wb: openpyxl.Workbook, sheet_name: str,
                    all_race_results: list[dict], venue: str, race_date: str):
    """
    買い目シートを新規作成して書き込む。
    縦: 行ラベル, 横: 1R〜12R
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    race_nos = [r["race_no"] for r in all_race_results]
    n = len(race_nos)

    # 列幅
    ws.column_dimensions["A"].width = 14.0
    ws.column_dimensions["B"].width = 22.0
    ws.column_dimensions["C"].width = 7.0
    for i in range(n):
        ws.column_dimensions[get_column_letter(4+i)].width = 18.0

    # ── Row1: タイトル ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3+n)
    ws.cell(1,1).value = f"🎯 新エンジン買い目推薦  【{venue}】  {race_date}  ※テスト稼働中"
    ws.cell(1,1).font  = _fn(bold=True, size=12, color="FFFFFFFF")
    ws.cell(1,1).fill  = _sf("FF0D1F38")
    ws.cell(1,1).alignment = _al("left")
    ws.row_dimensions[1].height = 22.0

    # ── Row2: ヘッダ ──
    HDR_FILL = _sf("FF1F4E79")
    HDR_FONT = _fn(bold=True, color="FFFFFFFF")
    for col, txt in [(1,"分類"),(2,"項目"),(3,"艇")]:
        c = ws.cell(2, col)
        c.value, c.fill, c.font, c.alignment, c.border = txt, HDR_FILL, HDR_FONT, _al(), _bdr
    for i, rno in enumerate(race_nos):
        dl = all_race_results[i].get("deadline","")
        txt = f"{rno}R\n{dl}" if dl else f"{rno}R"
        c = ws.cell(2, 4+i)
        c.value, c.fill, c.font, c.alignment, c.border = txt, HDR_FILL, HDR_FONT, _al(wrap=True), _bdr
    ws.row_dimensions[2].height = 28.0

    row = 3

    # ── セクション書き込みヘルパー ──
    def sec_header(label, fill_hex):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3+n)
        c = ws.cell(row, 1)
        c.value, c.fill, c.font, c.alignment = label, _sf(fill_hex), _fn(bold=True, color="FFFFFFFF"), _al("left")
        ws.row_dimensions[row].height = 13.0
        row += 1

    def waku_block(cat, item, cat_hex, item_hex, data_by_waku, font_ovr=None, fill_ovr=None):
        """6艇分のデータブロック"""
        nonlocal row
        for w in range(1, 7):
            c1 = ws.cell(row+w-1, 1)
            c1.fill = _sf(cat_hex)
            c1.alignment = _al()
            c2 = ws.cell(row+w-1, 2)
            c2.fill = _sf(item_hex)
            c2.alignment = _al("left")
            if w == 1:
                c1.value = cat
                c1.font  = _fn(bold=True, color="FFFFFFFF")
                c2.value = item
                c2.font  = _fn(bold=True, color="FF000000")
            c3 = ws.cell(row+w-1, 3)
            c3.value, c3.fill = str(w), _sf(_BOAT_FILL[w])
            c3.font, c3.alignment, c3.border = _fn(bold=True, color=_BOAT_FONT[w]), _al(), _bdr
            vals = data_by_waku.get(w, [None]*n)
            for j, v in enumerate(vals):
                cell = ws.cell(row+w-1, 4+j)
                cell.value = v
                cell.font  = font_ovr[w][j] if font_ovr else _fn(bold=False)
                if fill_ovr: cell.fill = fill_ovr[w][j]
                cell.alignment = _al()
                cell.border = _bdr
            ws.row_dimensions[row+w-1].height = 15.0
        row += 6

    # ─────────────────────────────────────────────────────────
    # ▼ Step1: 確率分析
    # ─────────────────────────────────────────────────────────
    FILL_PROB = "FF1A3A5C"
    FILL_PROB_ITEM = "FFD9E1F2"
    FILL_PROB_SEC  = "FF203864"

    sec_header("▼ Step1: 確率分析（P(1着) / 展開パターン）", FILL_PROB_SEC)

    # 展開パターン行（1行）
    ws.cell(row, 1).value = "推薦"
    ws.cell(row, 1).fill  = _sf(FILL_PROB)
    ws.cell(row, 1).font  = _fn(bold=True, color="FFFFFFFF")
    ws.cell(row, 1).alignment = _al()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row, 2).value = "展開パターン / 信頼度"
    ws.cell(row, 2).fill  = _sf(FILL_PROB)
    ws.cell(row, 2).font  = _fn(bold=True, color="FFFFFFFF")
    ws.cell(row, 2).alignment = _al()
    _PAT_FILL = {
        "nige":   "FF0A2A0A", "sashi": "FF1A1A00",
        "tenkai": "FF1A0A00", "mixed": "FF15101A",
    }
    _PAT_FONT = {
        "nige":   "FF60E860", "sashi": "FFD0C030",
        "tenkai": "FFE06030", "mixed": "FFB090E0",
    }
    _CONF_FILL = {"高":"FF0A2A0A","中":"FF2A2600","低":"FF2A0E00"}
    _CONF_FONT = {"高":"FF60E860","中":"FFD0C030","低":"FFE06030"}
    max_h = 3
    for i, rr in enumerate(all_race_results):
        res = rr.get("bet_result")
        if not res:
            ws.cell(row, 4+i).value = "データなし"
            ws.cell(row, 4+i).fill  = _sf("FF1A1A1A")
            ws.cell(row, 4+i).font  = _fn(bold=False, size=8, color="FF606060")
            ws.cell(row, 4+i).alignment = _al()
            ws.cell(row, 4+i).border = _bdr
            continue
        pat  = res.get("pattern","mixed")
        conf = res.get("confidence","低")
        cnt  = res.get("count",0)
        txt  = f"【{res.get('pattern_jp','')}】\n信頼度:{conf}  {cnt}点\n{rr.get('venue_are','')}"
        max_h = max(max_h, txt.count("\n")+1)
        c = ws.cell(row, 4+i)
        c.value, c.fill = txt, _sf(_PAT_FILL.get(pat,"FF151515"))
        c.font = _fn(bold=True, size=9, color=_PAT_FONT.get(pat,"FF888888"))
        c.alignment = _al(h="left", wrap=True)
        c.border = _bdr
    ws.row_dimensions[row].height = max(14.0*max_h, 46.0)
    row += 1

    # P(1着) 6艇分
    P1_FILL  = "FF0D1830"
    P1_IFILL = "FF162540"
    _p1_data = {w: [] for w in range(1,7)}
    for rr in all_race_results:
        res = rr.get("bet_result")
        p1  = res.get("p1",[]) if res else []
        members = rr.get("members",[])
        for w in range(1,7):
            idx = next((j for j,m in enumerate(members)
                        if int(m.get("course_int",j+1))==w), w-1)
            v = p1[idx] if res and idx < len(p1) else None
            _p1_data[w].append(f"{v*100:.1f}%" if v is not None else "-")
    waku_block("推薦", "P(1着)", P1_FILL, P1_IFILL, _p1_data)

    # ─────────────────────────────────────────────────────────
    # ▼ Step2: 考察テキスト
    # ─────────────────────────────────────────────────────────
    KS_HDR  = "FF203864"
    KS_BODY = "FF0D1830"

    sec_header("▼ Step2: 展開考察（残存率 × 攻め力 × ST差 × 近況）", KS_HDR)

    ws.cell(row,1).value = "考察"
    ws.cell(row,1).fill  = _sf(KS_HDR)
    ws.cell(row,1).font  = _fn(bold=True, color="FFFFFFFF")
    ws.cell(row,1).alignment = _al()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row,2).value = "確率上位5点の根拠"
    ws.cell(row,2).fill  = _sf(KS_HDR)
    ws.cell(row,2).font  = _fn(bold=True, color="FFFFFFFF")
    ws.cell(row,2).alignment = _al()
    max_h2 = 3
    for i, rr in enumerate(all_race_results):
        res = rr.get("bet_result")
        txt = res.get("kousatsu","") if res else "-"
        max_h2 = max(max_h2, txt.count("\n")+1)
        c = ws.cell(row, 4+i)
        c.value, c.fill = txt, _sf(KS_BODY)
        c.font = _fn(bold=False, size=8, color="FFB0C8E0")
        c.alignment = _al(h="left", wrap=True)
        c.border = _bdr
    ws.row_dimensions[row].height = max(12.0*max_h2, 80.0)
    row += 1

    # ─────────────────────────────────────────────────────────
    # ▼ Step3: 3連単買い目
    # ─────────────────────────────────────────────────────────
    BET_HDR  = "FF0D1F38"
    BET_BODY = "FF0D1830"

    sec_header("▼ Step3: 3連単買い目リスト（累積70%カット）", BET_HDR)

    ws.cell(row,1).value = "買い目"
    ws.cell(row,1).fill  = _sf(BET_HDR)
    ws.cell(row,1).font  = _fn(bold=True, color="FFFFFFFF")
    ws.cell(row,1).alignment = _al()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row,2).value = "3連単  確率順"
    ws.cell(row,2).fill  = _sf(BET_HDR)
    ws.cell(row,2).font  = _fn(bold=True, color="FFFFFFFF")
    ws.cell(row,2).alignment = _al()
    max_bets = 6
    for i, rr in enumerate(all_race_results):
        res  = rr.get("bet_result")
        bets = res.get("bets",[]) if res else []
        combos = res.get("combos",[]) if res else []
        txt  = "\n".join(
            f"{c['bet']}  {c['prob']*100:.2f}%"
            for c in combos
        ) if combos else "\n".join(bets)
        max_bets = max(max_bets, len(bets))
        c = ws.cell(row, 4+i)
        c.value, c.fill = txt, _sf(BET_BODY)
        c.font = _fn(bold=True, size=10, color="FFFFD080")
        c.alignment = _al(h="left", wrap=True)
        c.border = _bdr
    ws.row_dimensions[row].height = max(14.0*max_bets, 80.0)
    row += 1

    # ─────────────────────────────────────────────────────────
    # ▼ 選手情報（攻め力・ST参照）
    # ─────────────────────────────────────────────────────────
    INF_SEC  = "FF1F4E79"
    INF_ITEM = "FFD9E1F2"

    sec_header("▼ 選手情報（攻め力・STは今回メンバーで相互計算に使用）", INF_SEC)

    # 選手名
    _nm = {w:[] for w in range(1,7)}
    for rr in all_race_results:
        ms = rr.get("members",[])
        for w in range(1,7):
            m = next((x for x in ms if int(x.get("course_int",0))==w), None)
            _nm[w].append(m.get("_name_display","") if m else "-")
    waku_block("選手", "選手名", INF_SEC, INF_ITEM, _nm)

    # 攻め力
    _atk = {w:[] for w in range(1,7)}
    for rr in all_race_results:
        ms = rr.get("members",[])
        for w in range(1,7):
            m = next((x for x in ms if int(x.get("course_int",0))==w), None)
            if m:
                c = int(m.get("course_int",w))
                if c == 2:
                    v = float(m.get("sashi_rate") or 0)
                else:
                    v = float(m.get("makuri_rate") or 0) + float(m.get("makuri_zashi_rate") or 0)
                _atk[w].append(f"{v:.2f}")
            else:
                _atk[w].append("-")
    waku_block("選手", "攻め力", INF_SEC, INF_ITEM, _atk)

    # 平均ST
    _st = {w:[] for w in range(1,7)}
    for rr in all_race_results:
        ms = rr.get("members",[])
        for w in range(1,7):
            m = next((x for x in ms if int(x.get("course_int",0))==w), None)
            v = m.get("avg_st_self") if m else None
            _st[w].append(f"{v:.3f}" if v is not None else "-")
    waku_block("選手", "平均ST(秒)", INF_SEC, INF_ITEM, _st)

    # motor
    _mot = {w:[] for w in range(1,7)}
    for rr in all_race_results:
        ms = rr.get("members",[])
        for w in range(1,7):
            m = next((x for x in ms if int(x.get("course_int",0))==w), None)
            v = m.get("motor_2rate") if m else None
            _mot[w].append(f"{v:.1f}%" if v is not None else "-")
    waku_block("選手", "モーター2連", INF_SEC, INF_ITEM, _mot)

    # 注記
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3+n)
    ws.cell(row,1).value = (
        "※ 本シートは新エンジン（recommend_bet v7）による独立出力です。"
        "既存の新聞シートとは別のロジックで計算しています。"
    )
    ws.cell(row,1).font = _fn(bold=False, size=8, color="FF808080")
    ws.cell(row,1).alignment = _al("left")
    ws.row_dimensions[row].height = 13.0


# ============================================================
# メイン処理
# ============================================================
def run(venue: str = "", date_str: str = "", race_no: str = "",
        output_path: str = ""):
    # CSV読み込み
    df, race_date, venue = load_csv(venue, date_str, race_no)
    if df is None or df.empty:
        print("❌ データが空です")
        return

    print(f"\n📊 会場: {venue}  日付: {race_date}")

    # レース番号一覧
    race_col = None
    for col in ["レース番号","R","レース","race_no"]:
        if col in df.columns:
            race_col = col; break

    if race_col:
        race_nos = sorted(
                [v for v in df[race_col].astype(str).unique()
                 if v not in ("","nan","None")],
                key=lambda x: int(x) if x.isdigit() else 99)
    else:
        race_nos = ["1"]

    # 全レース処理
    all_results = []
    for rno in race_nos:
        print(f"\n  {rno}R 処理中...")
        if race_col:
            race_df = df[df[race_col].astype(str) == str(rno)]
        else:
            race_df = df

        players = race_df.to_dict(orient="records")
        if not players:
            print(f"  ⚠️  {rno}R データなし")
            continue

        # 締切時刻
        dl_raw = (players[0].get("締切時刻") or players[0].get("締切") or
                  players[0].get("締め切り時刻") or "")
        deadline = str(dl_raw).strip()
        deadline = None if deadline in ("","None","nan") else deadline

        # jizen_members 構築
        members = build_jizen_members_simple(players, venue, rno)
        if not members:
            print(f"  ⚠️  {rno}R メンバー構築失敗")
            continue

        # 事前評価
        ev_result = {}
        if JIZEN_OK and members:
            try:
                ev_result = evaluate_all(members)
            except Exception as e:
                print(f"  ⚠️  evaluate_all エラー: {e}")

        # 買い目推薦
        bet_result = None
        try:
            bet_result = recommend(members, ev_result, venue=venue, race_no=int(rno))
            print(f"  ✅ {rno}R: {bet_result['pattern_jp']} {bet_result['count']}点 "
                  f"信頼度:{bet_result['confidence']}")
            print(f"     買い目: {' / '.join(bet_result['bets'][:5])}{'...' if len(bet_result['bets'])>5 else ''}")
        except Exception as e:
            print(f"  ❌ {rno}R 買い目エラー: {e}")
            import traceback; traceback.print_exc()

        # 会場荒れスコア
        from master_data import get_race_are_score
        try:
            are = get_race_are_score(venue, int(rno))
            venue_are = f"荒れ{are:.0f}pt"
        except:
            venue_are = ""

        all_results.append({
            "race_no":    rno,
            "deadline":   deadline,
            "members":    members,
            "ev_result":  ev_result,
            "bet_result": bet_result,
            "venue_are":  venue_are,
        })

    if not all_results:
        print("❌ 有効なレースデータがありません")
        return

    # Excel出力
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not output_path:
        output_path = str(_OUTPUT_DIR / f"{venue}_{race_date}_買い目.xlsx")

    # 既存ファイルがあれば開く、なければ新規
    if pathlib.Path(output_path).exists():
        wb = openpyxl.load_workbook(output_path)
        print(f"\n📄 既存ファイルに追記: {output_path}")
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        print(f"\n📄 新規ファイル: {output_path}")

    sheet_name = f"{venue}_買い目"
    write_bet_sheet(wb, sheet_name, all_results, venue, race_date)
    wb.save(output_path)
    print(f"\n✅ 保存完了: {output_path}")
    print(f"   シート: {sheet_name}")


# ============================================================
# CLI エントリポイント
# ============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="新エンジン 3連単買い目出力")
    parser.add_argument("--venue", default="", help="会場名 例: 大村")
    parser.add_argument("--date",  default="", help="日付 例: 2026-03-20")
    parser.add_argument("--race",  default="", help="レース番号（省略で全レース）")
    parser.add_argument("--output",default="", help="出力Excelパス（省略で自動）")
    args = parser.parse_args()
    run(venue=args.venue, date_str=args.date, race_no=args.race, output_path=args.output)
