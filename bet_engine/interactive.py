#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
interactive.py — 対話式 買い目推薦
====================================
python scripts/bet_engine/interactive.py
"""
import glob, os, pathlib, re, sys, logging, traceback
from datetime import datetime

_HERE    = pathlib.Path(__file__).parent
_SCRIPTS = _HERE.parent
_BASE    = _SCRIPTS.parent
_CSV_DIR = _SCRIPTS / "csv_output"

sys.path.insert(0, str(_HERE))
sys.path.insert(0, str(_SCRIPTS))

# ── エラーログ設定 ────────────────────────────────────────────
_LOG_PATH = _BASE / "bet_engine_error.log"
logging.basicConfig(
    filename=str(_LOG_PATH),
    level=logging.ERROR,
    format="%(asctime)s\n%(message)s\n" + "─"*60,
    encoding="utf-8",
)

def _log_error(label, e, tb):
    logging.error(f"{label}\n{e}\n{tb}")

def _show_error(label, e, tb):
    """エラーを画面に表示してEnter待ちで止める"""
    print()
    print(RD + "="*60 + R)
    print(RD + f"  ❌ {label}" + R)
    print(RD + f"  {e}" + R)
    print(RD + "─"*60 + R)
    for line in tb.strip().split("\n"):
        print(RD + f"  {line}" + R)
    print(RD + "="*60 + R)
    print(GY + f"  ログ: {_LOG_PATH}" + R)
    input(f"\n{CY}  Enterキーを押して続ける...{R}")

# ── 依存モジュール ────────────────────────────────────────────
try:
    from run_bet_engine import (
        build_jizen_members_simple, write_bet_sheet, _OUTPUT_DIR,
    )
    from recommend_bet import recommend
    from master_data import get_race_are_score
    print("✅ エンジンモジュール 読み込み完了")
except Exception as e:
    tb = traceback.format_exc()
    print(f"❌ モジュール読み込み失敗: {e}")
    print(tb)
    input("Enterキーを押して終了...")
    sys.exit(1)

try:
    from evaluate_jizen import evaluate_all
    JIZEN_OK = True
    print("✅ evaluate_jizen 読み込み完了")
except ImportError:
    JIZEN_OK = False
    print("⚠️  evaluate_jizen なし（評価スキップ）")

# ── 全24会場 ──────────────────────────────────────────────────
ALL_VENUES = [
    "桐生", "戸田", "江戸川", "平和島", "多摩川", "浜名湖",
    "蒲郡", "常滑", "津",    "三国",   "びわこ", "住之江",
    "尼崎", "鳴門", "丸亀",  "児島",   "宮島",   "徳山",
    "下関", "若松", "芦屋",  "福岡",   "唐津",   "大村",
]

# ── カラー ────────────────────────────────────────────────────
R  = "\033[0m"
B  = "\033[1m"
CY = "\033[96m"
GR = "\033[92m"
YL = "\033[93m"
RD = "\033[91m"
GY = "\033[90m"
MG = "\033[95m"

def ask(prompt, default=""):
    suffix = f" [{default}]" if default else ""
    try:
        v = input(f"\n{CY}{B}  {prompt}{suffix}{R}{CY} > {R}").strip()
        return v if v else default
    except (KeyboardInterrupt, EOFError):
        print(); return "q"

def hr(): print(GY + "─" * 60 + R)

# ── csv_output の会場・日付情報取得 ──────────────────────────
def get_venue_csv_info():
    info = {}
    if not _CSV_DIR.exists():
        return info
    for d in sorted(_CSV_DIR.iterdir()):
        if d.is_dir() and d.name in ALL_VENUES:
            csvs = sorted(glob.glob(str(d / "*.csv")))
            if csvs:
                dates = []
                for f in csvs:
                    m = re.search(r"(\d{4}-\d{2}-\d{2})", pathlib.Path(f).stem)
                    if m:
                        dates.append(m.group(1))
                if dates:
                    info[d.name] = sorted(set(dates), reverse=True)
    for f in glob.glob(str(_CSV_DIR / "*.csv")):
        p = pathlib.Path(f)
        name = p.stem.split("_")[0]
        if name in ALL_VENUES:
            m = re.search(r"(\d{4}-\d{2}-\d{2})", p.stem)
            if m:
                date = m.group(1)
                if name not in info:
                    info[name] = []
                if date not in info[name]:
                    info[name].append(date)
                    info[name].sort(reverse=True)
    return info

def find_csv(venue, date_str):
    dirs = [_CSV_DIR / venue, _CSV_DIR]
    for d in dirs:
        pattern = f"{venue}_{date_str}.csv" if date_str else f"{venue}_*.csv"
        hits = sorted(glob.glob(str(d / pattern)))
        if hits:
            return pathlib.Path(hits[-1])
    return None

# ── 会場一覧表示 ─────────────────────────────────────────────
def show_venues(venue_info):
    today = datetime.today().strftime("%Y-%m-%d")
    print(f"\n  {B}会場一覧{R}  {GY}（◎=当日CSV  ○=過去CSV  －=なし）{R}\n")
    cols = 6
    for row_i in range(4):
        line = "  "
        for col_i in range(cols):
            idx = row_i * cols + col_i
            if idx >= len(ALL_VENUES): break
            v = ALL_VENUES[idx]
            has_today = v in venue_info and today in venue_info[v]
            has_csv   = v in venue_info
            mark = (GR+"◎"+R) if has_today else ((YL+"○"+R) if has_csv else (GY+"－"+R))
            line += f"{CY}{idx+1:2d}{R}.{mark}{v:<4}  "
        print(line)
    print()

# ── 日付選択 ─────────────────────────────────────────────────
def select_date(venue, venue_info):
    today = datetime.today().strftime("%Y-%m-%d")
    dates = venue_info.get(venue, [])

    if not dates:
        print(YL + f"\n  ⚠ {venue}のCSVが見つかりません" + R)
        d = ask("日付を手入力 (例: 2026-03-20)")
        if d == "q": return "q"
        if re.match(r"^\d{8}$", d):
            d = f"{d[:4]}-{d[4:6]}-{d[6:]}"
        return d

    if today in dates:
        print(GR + f"\n  ✅ 当日({today})のCSVがあります → 自動選択" + R)
        return today

    print(f"\n  {B}利用可能な日付{R}  {GY}（{venue}）{R}")
    for i, d in enumerate(dates[:10], 1):
        days_ago = (datetime.today() - datetime.strptime(d, "%Y-%m-%d")).days
        ago_str = f"{days_ago}日前"
        print(f"    {CY}{i}{R}. {d}  {GY}({ago_str}){R}")

    v = ask(f"番号で選択 (1〜{min(len(dates),10)})  または日付を直接入力")
    if v == "q": return "q"
    if v.isdigit() and 1 <= int(v) <= len(dates):
        return dates[int(v)-1]
    if re.match(r"^\d{4}-\d{2}-\d{2}$", v): return v
    if re.match(r"^\d{8}$", v): return f"{v[:4]}-{v[4:6]}-{v[6:]}"
    print(RD + "  無効な入力です" + R)
    return None

# ── 1レース計算・表示 ─────────────────────────────────────────
PAT_COLOR  = {"nige": GR, "sashi": YL, "tenkai": MG, "mixed": CY}
CONF_COLOR = {"高": GR, "中": YL, "低": RD}

def run_race(df, venue, race_date, rno):
    import pandas as pd

    race_df = df
    for col in ["レース番号", "R", "レース", "race_no"]:
        if col in df.columns:
            race_df = df[df[col].astype(str) == str(rno)]
            break

    players = race_df.to_dict(orient="records")
    if not players:
        print(RD + f"  ❌ {rno}R のデータが見つかりません" + R)
        return None

    dl = str(players[0].get("締切時刻") or players[0].get("締切") or "").strip()
    dl = None if dl in ("", "None", "nan") else dl

    # メンバー構築
    try:
        members = build_jizen_members_simple(players, venue, rno)
    except Exception as e:
        tb = traceback.format_exc()
        _log_error(f"{venue} {rno}R build_jizen_members_simple", e, tb)
        _show_error(f"{venue} {rno}R メンバー構築エラー", e, tb)
        return None

    if not members:
        print(RD + f"  ❌ {rno}R メンバーデータ不足" + R)
        return None

    # 事前評価
    ev = {}
    if JIZEN_OK:
        try:
            ev = evaluate_all(members)
        except Exception as e:
            tb = traceback.format_exc()
            _log_error(f"{venue} {rno}R evaluate_all", e, tb)
            _show_error(f"{venue} {rno}R 事前評価エラー", e, tb)
            # evaluate_all失敗でも計算継続

    # 買い目推薦
    try:
        res = recommend(members, ev, venue=venue, race_no=int(rno))
    except Exception as e:
        tb = traceback.format_exc()
        _log_error(f"{venue} {rno}R recommend", e, tb)
        _show_error(f"{venue} {rno}R 計算エラー", e, tb)
        return None

    # 荒れスコア
    try:
        are = f"荒れ{get_race_are_score(venue, int(rno)):.0f}pt"
    except:
        are = ""

    # ── 表示 ──
    hr()
    pat  = res["pattern"]
    conf = res["confidence"]
    cnt  = res["count"]
    pc   = PAT_COLOR.get(pat, CY)
    cc   = CONF_COLOR.get(conf, GY)

    print(f"\n  {B}{CY}{venue}  {rno}R{R}" + (f"  {GY}締切:{dl}{R}" if dl else ""))
    print(f"  {pc}{B}{res['pattern_jp']}{R}  {cc}{B}信頼度:{conf}{R}  {cnt}点  {GY}{are}{R}")

    # 注目艇
    p1   = res.get("p1", [])
    nige = ev.get("in_nige", [""]*6)
    aisho= ev.get("aisho",   [""]*6)
    kiry = ev.get("kiryoku", [""]*6)
    top  = sorted(range(len(p1)), key=lambda i: -p1[i])[:3]
    print(f"\n  {B}【注目艇】{R}")
    for i in top:
        m   = members[i] if i < len(members) else {}
        cno = int(m.get("course_int", i+1))
        nm  = m.get("_name_display", "")
        ns  = nige[i].replace("⚠️","").replace("⚠","").strip()  if i < len(nige)  else ""
        as_ = aisho[i].replace("⚠️","").replace("⚠","").strip() if i < len(aisho) else ""
        ks  = kiry[i].replace("⚠️","").replace("⚠","").strip()  if i < len(kiry)  else ""
        sa  = float(m.get("sashi_rate") or 0)
        mk  = float(m.get("makuri_rate") or 0)
        mz  = float(m.get("makuri_zashi_rate") or 0)
        atk = sa + mk + mz
        st  = m.get("avg_st_self")
        print(f"    {B}{cno}号艇{R}  {nm:<8}"
              f"  {YL}P1={p1[i]*100:.1f}%{R}"
              f"  逃{ns or '－'}相{as_ or '－'}機{ks or '－'}"
              f"  攻め{atk:.2f}  ST{f'{st:.3f}' if st else '?'}")

    # 買い目
    print(f"\n  {B}【3連単買い目  {cnt}点】{R}")
    combos = res.get("combos", [])
    for ci, combo in enumerate(combos, 1):
        prob = combo["prob"] * 100
        bar  = GR + "█" * max(1, int(prob * 3)) + R
        fail = GY + " [崩れ]" + R if combo.get("_fail_scenario") else ""
        print(f"    {ci:2d}.  {GR}{B}{combo['bet']}{R}   {YL}{prob:5.2f}%{R}  {bar}{fail}")

    # 考察（上位6行）
    kousatsu = res.get("kousatsu", "")
    lines = [l for l in kousatsu.split("\n") if l.strip()][:6]
    print(f"\n  {GY}")
    for l in lines:
        print(f"  {l}")
    print(R, end="")

    return {
        "race_no":    rno,
        "deadline":   dl,
        "members":    members,
        "ev_result":  ev,
        "bet_result": res,
        "venue_are":  are,
    }

# ── メインループ ───────────────────────────────────────────────
def main():
    print(f"\n{CY}{'='*60}")
    print(f"  🎯 新エンジン 対話式 買い目推薦")
    print(f"  recommend_bet v7  |  テスト稼働中")
    print(f"{'='*60}{R}")
    print(GY + "  q で終了  /  Enter でデフォルト値を使用" + R)
    print(GY + f"  エラーログ: {_LOG_PATH}" + R)

    while True:
        print()

        # ── Step1: 会場選択 ──────────────────────────────────────
        try:
            venue_info = get_venue_csv_info()
        except Exception as e:
            tb = traceback.format_exc()
            _log_error("get_venue_csv_info", e, tb)
            _show_error("会場情報取得エラー", e, tb)
            venue_info = {}

        show_venues(venue_info)

        v_input = ask("会場番号を入力 (1〜24)")
        if v_input == "q": break
        if not v_input.isdigit() or not (1 <= int(v_input) <= 24):
            print(RD + "  1〜24の番号を入力してください" + R)
            continue
        venue = ALL_VENUES[int(v_input) - 1]
        print(GR + f"  → {venue}" + R)

        # ── Step2: 日付選択 ──────────────────────────────────────
        date_str = select_date(venue, venue_info)
        if date_str == "q": break
        if not date_str: continue

        csv_path = find_csv(venue, date_str)
        if not csv_path:
            csv_path = find_csv(venue, "")
            if not csv_path:
                print(RD + f"  ❌ CSVが見つかりません: {_CSV_DIR}/{venue}/" + R)
                continue
            print(YL + f"  ⚠ {date_str}のCSVなし → 最新を使用: {csv_path.name}" + R)

        # CSV読み込み
        try:
            import pandas as pd
            df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
            df.columns = df.columns.str.strip()
            for col in df.columns:
                df[col] = df[col].astype(str).str.lstrip("'").str.strip()
        except Exception as e:
            tb = traceback.format_exc()
            _log_error(f"CSV読み込み: {csv_path}", e, tb)
            _show_error(f"CSV読み込みエラー: {csv_path.name}", e, tb)
            continue

        # 日付をCSVから取得
        race_date = date_str
        if "日付" in df.columns:
            dv = df["日付"].dropna()
            dv = dv[dv.str.strip() != ""]
            if len(dv):
                raw = dv.iloc[0].strip()
                m = re.sub(r"(\d{4})[/\-](\d{2})[/\-](\d{2}).*", r"\1-\2-\3", raw)
                if re.match(r"\d{4}-\d{2}-\d{2}", m):
                    race_date = m

        # レース番号一覧
        race_nos = []
        for col in ["レース番号", "R", "レース", "race_no"]:
            if col in df.columns:
                race_nos = sorted(
                    [v for v in df[col].astype(str).unique().tolist()
                     if v not in ("", "nan", "None")],
                    key=lambda x: int(x) if x.isdigit() else 99
                )
                break
        if not race_nos:
            print(GY + f"  ⚠ レース番号列が見つかりません。列名: {list(df.columns)}" + R)
            race_nos = ["1"]

        print(GR + f"  ✅ {venue}  {race_date}  ({len(race_nos)}レース: {race_nos[0]}R〜{race_nos[-1]}R)" + R)

        # ── Step3: レース番号選択 ─────────────────────────────────
        today = datetime.today().strftime("%Y-%m-%d")
        is_today = (race_date == today)

        if is_today:
            print(f"\n  {B}レース選択{R}")
            print(f"    {CY}0{R}. 全レース一括")
            for rno in race_nos:
                print(f"    {CY}{rno:>2}{R}. {rno}R")
            race_input = ask("レース番号 (0=全レース)")
        else:
            race_input = ask("レース番号 (Enter=全レース)", "")

        if race_input == "q": break

        if race_input in ("", "0"):
            target_nos = race_nos
        elif race_input.isdigit():
            if race_input not in race_nos:
                print(RD + f"  ❌ {race_input}R はCSVにありません" + R)
                continue
            target_nos = [race_input]
        else:
            print(RD + "  数字で入力してください" + R)
            continue

        # ── Step4: 計算・表示 ────────────────────────────────────
        print(GY + f"\n  {venue} {' / '.join(f'{r}R' for r in target_nos)} を計算中..." + R)

        all_results = []
        for rno in target_nos:
            try:
                rr = run_race(df, venue, race_date, rno)
                if rr:
                    all_results.append(rr)
            except Exception as e:
                tb = traceback.format_exc()
                _log_error(f"{venue} {rno}R 予期しないエラー", e, tb)
                _show_error(f"{venue} {rno}R 予期しないエラー", e, tb)

        if not all_results:
            continue

        # ── Step5: Excel保存 ──────────────────────────────────────
        print()
        save = ask("Excelに保存しますか？ (y/n)", "y")
        if save.lower() in ("y", "yes", ""):
            import openpyxl as ox
            import os

            # 保存先の候補を順番に試す
            filename = f"{venue}_{race_date}_買い目.xlsx"
            save_candidates = [
                _OUTPUT_DIR / filename,
                pathlib.Path.home() / "Desktop" / filename,
                pathlib.Path.home() / "Documents" / filename,
                pathlib.Path(os.environ.get("USERPROFILE", str(pathlib.Path.home()))) / "Desktop" / filename,
            ]

            # bet_engine_output フォルダを作成
            try:
                _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass

            out = save_candidates[0]
            wb  = ox.load_workbook(str(out)) if out.exists() else ox.Workbook()
            if "Sheet" in wb.sheetnames: del wb["Sheet"]
            try:
                print(GY + f"  シート作成中..." + R)
                write_bet_sheet(wb, f"{venue}_買い目", all_results, venue, race_date)
                print(GY + f"  ファイル保存中: {out}" + R)
                wb.save(str(out))
                print(GR + B + f"\n  ✅ 保存: {out}" + R)
            except PermissionError as e:
                print(RD + "="*60 + R)
                print(RD + "  ❌ 保存エラー: ファイルが開いているか書き込み禁止です" + R)
                print(RD + f"  Excelで {out.name} を開いていたら閉じてください" + R)
                print(RD + "="*60 + R)
                print(YL + "  別の場所への保存を試みます..." + R)
                saved = False
                for alt_path in save_candidates[1:]:
                    try:
                        alt_path.parent.mkdir(parents=True, exist_ok=True)
                        wb2 = ox.Workbook()
                        if "Sheet" in wb2.sheetnames: del wb2["Sheet"]
                        write_bet_sheet(wb2, f"{venue}_買い目", all_results, venue, race_date)
                        wb2.save(str(alt_path))
                        print(GR + B + f"\n  ✅ 別の場所に保存しました: {alt_path}" + R)
                        saved = True
                        break
                    except Exception:
                        continue
                if not saved:
                    print(RD + "  ❌ どこにも保存できませんでした" + R)
                    print(RD + f"  元ファイル({out.name})を閉じてから再試行してください" + R)
                    input(f"\n{CY}  Enterキーを押して続ける...{R}")
            except Exception as e:
                tb = traceback.format_exc()
                _log_error(f"Excel保存: {out}", e, tb)
                _show_error("Excel保存エラー", e, tb)

        # ── 続けるか ──────────────────────────────────────────────
        hr()
        cont = ask("続けますか？ (y/n)", "y")
        if cont.lower() in ("n", "no", "q"):
            break

    print(CY + "\n  終了しました。\n" + R)

if __name__ == "__main__":
    main()
