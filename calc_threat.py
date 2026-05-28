# -*- coding: utf-8 -*-
"""
calc_threat.py
==============
【役割】
  ボートリサーチ新聞_軽量版.xlsx から各艇の選手情報・数値指標を読み込み、
  「1号艇視点での相対脅威スコア」を計算して蓄積CSVの 脅威合計 列を更新する。

【旧版との違い】
  旧版: 脅威合計 = 蓄積CSVに書いてある固定値をそのまま使うだけ
  新版: 脅威合計 = 会場×コース×選手ランク×決まり手×モータで相対計算

【脅威スコアの計算式】
  脅威スコア[艇i] =
      コース別3連対率[艇i]     # そのコースで連に絡む実績（核心）
    × 組ランク係数[艇i]        # A1=1.30, A2=1.10, B1=0.90, B2=0.70
    × まくり脅威係数[艇i]      # まくり%が高い → 1号艇の直接的脅威
    × モータ係数[艇i]          # モータ2連勝率が高い → 信頼度UP
    × 機力係数[艇i]            # A=1.20, B=1.00, C=0.85, D=0.70, E=0.50

  脅威合計[レース] = Σ 脅威スコア[艇2〜6]
  ※ 1号艇自身は除外（他艇が1号艇にとってどれだけ怖いかを測る）

【脅威スコアの目安】
  脅威合計 0〜12  → 低脅威: 1号艇逃げが決まりやすい
  脅威合計 12〜16 → 中脅威: 5・6号艇の1着をカット
  脅威合計 16〜20 → 高脅威: 4・5・6号艇の1着をカット
  脅威合計 20以上 → 超高脅威: 3〜6号艇の1着をカット

【使い方】
  # 全会場を更新（蓄積CSVに書き込む）
  python calc_threat.py

  # 特定会場のみ
  python calc_threat.py --venue びわこ

  # 確認のみ（書き込まない）
  python calc_threat.py --dry-run

  # Excelパスを明示
  python calc_threat.py --excel "C:/path/to/ボートリサーチ新聞_軽量版.xlsx"

【ループ学習との位置づけ】
  1. python calc_threat.py       ← このスクリプト（脅威合計を更新）
  2. python apply_correction.py  ← 買い目絞り込み
  3. python backtest_engine.py   ← バックテスト
  4. 2〜3を繰り返す
"""

import argparse
import glob
import re
from pathlib import Path

import pandas as pd

# ══════════════════════════════════════════════════════════
# パス定数
# ══════════════════════════════════════════════════════════
BASE_DIR      = Path(r"C:\Users\user\Desktop\データ収集")
CHIKUSEKI_DIR = BASE_DIR / "scripts" / "数値蓄積"
EXCEL_PATH    = BASE_DIR / "ボートリサーチ新聞_軽量版.xlsx"

# ══════════════════════════════════════════════════════════
# 係数定義
# ══════════════════════════════════════════════════════════

# 選手の組（クラス）別係数
RANK_COEFF = {
    "A1": 1.30,
    "A2": 1.10,
    "B1": 0.90,
    "B2": 0.70,
}

# 機力評価別係数（事前評価③）
KIKI_COEFF = {
    "A": 1.20,
    "B": 1.00,
    "C": 0.85,
    "D": 0.70,
    "E": 0.50,
}

# モータ2連の正規化係数（基準値=40%）
MOTOR_BASE = 40.0

# まくり%が高いほど1号艇への直接脅威（上限2.0倍）
MAKURI_MAX_COEFF = 2.0


def sep(c="=", w=55):
    print(c * w)


# ══════════════════════════════════════════════════════════
# Excelパーサー
# ══════════════════════════════════════════════════════════

def _pct_to_float(val) -> float:
    """'53.5%' → 0.535, 53.55 → 53.55（モータ2連はそのまま）"""
    if val is None:
        return 0.0
    s = str(val).strip().replace("%", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_venue_sheet(ws) -> dict:
    """
    会場シートを読んで {レース番号: {艇番: {指標: 値}}} の辞書を返す。

    返り値の構造:
        {
          1: {   # 1R
            1: { "1着率": 0.608, "3連対率": 0.857, "まくり%": 0.142,
                 "組": "A2", "モータ2連": 53.55, "機力": "A", ... },
            2: { ... },
            ...
          },
          2: { ... },
          ...
        }
    """
    rows = list(ws.iter_rows(values_only=True))

    # ── ヘッダ行（行2）からレース番号列を取得 ──────────────────────────
    header = rows[1]   # 0-indexed: rows[1] = 2行目
    # 列3以降が "1R\n08:47" のような形式
    race_cols = {}  # col_index → race_no
    for ci, cell in enumerate(header[3:], start=3):
        if cell is None:
            continue
        m = re.match(r"(\d+)R", str(cell))
        if m:
            race_cols[ci] = int(m.group(1))

    if not race_cols:
        return {}

    # ── データ行を解析 ──────────────────────────────────────────────────
    # 結果: data[race_no][boat] = {指標名: 値}
    data: dict = {}
    for rno in race_cols.values():
        data[rno] = {i: {} for i in range(1, 7)}

    current_category = ""   # "数値指標" / "決まり手" / "選手情報" / "事前評価"
    current_item     = ""   # "オリジナル1着率(%)" etc.

    for row in rows[2:]:    # 3行目以降
        cat  = row[0]
        item = row[1]
        boat = row[2]

        if cat is not None and str(cat).strip():
            current_category = str(cat).strip()
        if item is not None and str(item).strip():
            current_item = str(item).strip()

        if boat is None:
            continue
        try:
            boat_no = int(str(boat).strip())
        except Exception:
            continue

        if boat_no not in range(1, 7):
            continue

        for ci, rno in race_cols.items():
            if ci >= len(row):
                continue
            val = row[ci]

            key = _make_key(current_category, current_item)
            if key is None:
                continue

            data[rno][boat_no][key] = val

    return data


def _make_key(category: str, item: str):
    """(category, item) → 統一キー名。不要なものは None を返す。"""
    cat = category.strip()
    it  = item.strip()

    if cat == "数値指標":
        if "オリジナル1着率" in it:
            return "1着率"
        if "コース別3連対率" in it:
            return "3連対率"
        if "まくり%" in it or it == "まくり%":
            return "まくり%"
        if "2着優位度" in it:
            return "2着優位度"

    elif cat == "決まり手":
        if "まくり%" in it and "まくり差し" not in it:
            return "まくり%"
        if "まくり差し%" in it:
            return "まくり差し%"
        if "逃げ%" in it:
            return "逃げ%"
        if "差し%" in it and "まくり" not in it:
            return "差し%"

    elif cat == "選手情報":
        if it == "組":
            return "組"
        if "モータ2連" in it:
            return "モータ2連"

    elif cat == "事前評価":
        if "③機力" in it or it == "③機力":
            return "機力"

    return None


# ══════════════════════════════════════════════════════════
# 脅威スコア計算
# ══════════════════════════════════════════════════════════

def calc_threat_score(race_data: dict) -> float:
    """
    1レース分の艇データ {艇番: {指標: 値}} から脅威合計を計算する。

    1号艇以外（2〜6号艇）が1号艇にとってどれだけ脅威かを相対評価する。
    """
    # 1号艇自身の1着率（強ければ他艇の脅威は相対的に低下）
    boat1_win_rate = _pct_to_float(race_data.get(1, {}).get("1着率", 0)) / 100
    boat1_strength = max(0.1, boat1_win_rate)  # ゼロ除算防止

    threat_total = 0.0

    for boat_no in range(2, 7):
        bd = race_data.get(boat_no, {})

        # ① コース別3連対率（そのコースで連に絡む実績）
        renpuku_rate = _pct_to_float(bd.get("3連対率", 0)) / 100

        # ② 組ランク係数
        kumi = str(bd.get("組", "B1")).strip()
        rank_coeff = RANK_COEFF.get(kumi, 0.90)

        # ③ まくり脅威係数（まくり%が高い → 1号艇の直接的脅威）
        #    まくり + まくり差し を合算、30%超で最大係数
        makuri_pct = (
            _pct_to_float(bd.get("まくり%", 0)) +
            _pct_to_float(bd.get("まくり差し%", 0))
        )
        # 0%→1.0, 30%以上→MAKURI_MAX_COEFF に線形補間
        makuri_coeff = 1.0 + (MAKURI_MAX_COEFF - 1.0) * min(makuri_pct / 30.0, 1.0)

        # ④ モータ係数（モータ2連/基準値で正規化）
        motor = _pct_to_float(bd.get("モータ2連", MOTOR_BASE))
        motor_coeff = max(0.5, min(1.5, motor / MOTOR_BASE))

        # ⑤ 機力係数
        kiki  = str(bd.get("機力", "B")).strip()
        kiki_coeff = KIKI_COEFF.get(kiki, 1.0)

        # 艇iの脅威スコア
        score_i = (
            renpuku_rate
            * rank_coeff
            * makuri_coeff
            * motor_coeff
            * kiki_coeff
        )

        threat_total += score_i

    # 相対化: 1号艇が強いほど他艇の脅威は割り引く
    # スケール: 0〜30 の範囲に収めるよう * 30 / boat1_strength でスケーリング
    # ただし脅威スコアの絶対値も重要なので緩め（/2 程度）にする
    relative_threat = threat_total / (boat1_strength * 2)

    # 最終値は 0〜30 にクリップ
    return round(min(30.0, relative_threat), 2)


# ══════════════════════════════════════════════════════════
# Excelから全会場×全レースの脅威スコアを構築
# ══════════════════════════════════════════════════════════

def load_excel_threats(excel_path: Path) -> dict:
    """
    Excel全シートを読み込み、{会場名: {レース番号: 脅威スコア}} を返す。

    例:
        {
          "鳴門": {1: 8.3, 2: 14.1, ...},
          "びわこ": {1: 11.2, ...},
          ...
        }
    """
    from openpyxl import load_workbook

    print(f"  Excel読み込み中: {excel_path}")
    wb = load_workbook(str(excel_path), read_only=True)

    threat_db: dict = {}

    for shname in wb.sheetnames:
        # "_数値" で終わるシートのみ処理
        if not shname.endswith("_数値"):
            continue
        venue = shname.replace("_数値", "")

        ws = wb[shname]
        race_data = parse_venue_sheet(ws)

        if not race_data:
            print(f"    [{venue}] データなし（スキップ）")
            continue

        threat_db[venue] = {}
        for rno, bd in race_data.items():
            threat_db[venue][rno] = calc_threat_score(bd)

        r_str = ", ".join(
            f"{rno}R:{sc:.1f}"
            for rno, sc in sorted(threat_db[venue].items())
        )
        print(f"    [{venue}] {r_str}")

    wb.close()
    return threat_db


# ══════════════════════════════════════════════════════════
# 蓄積CSVへの書き込み
# ══════════════════════════════════════════════════════════

def update_chikuseki_csv(
    csv_path: Path,
    threat_db: dict,
    venue_name: str,
    dry_run: bool,
) -> dict:
    """
    蓄積CSVの 脅威合計 列を Excel から計算した値で上書きする。

    threat_db: {レース番号(int): 脅威スコア(float)}
    """
    try:
        df = pd.read_csv(str(csv_path), encoding="utf-8")
    except Exception as e:
        return {"venue": venue_name, "status": "スキップ", "reason": str(e)}

    if "レース番号" not in df.columns:
        return {"venue": venue_name, "status": "スキップ", "reason": "レース番号列なし"}

    if "脅威合計" not in df.columns:
        df["脅威合計"] = 0.0

    before = df["脅威合計"].copy()
    changed = 0

    for rno, threat in threat_db.items():
        mask = df["レース番号"].astype(int) == int(rno)
        if mask.any():
            df.loc[mask, "脅威合計"] = threat
            if (before[mask] != threat).any():
                changed += 1

    summary = {
        "venue":        venue_name,
        "status":       "更新" if changed > 0 else "変更なし",
        "races_updated": changed,
        "dry_run":      dry_run,
    }

    if not dry_run and changed > 0:
        df.to_csv(str(csv_path), index=False, encoding="utf-8")

    return summary


# ══════════════════════════════════════════════════════════
# main
# ══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Excelから脅威スコアを計算して蓄積CSVを更新"
    )
    parser.add_argument("--venue",    type=str,  default=None,
                        help="特定会場のみ更新 (例: びわこ)")
    parser.add_argument("--dry-run",  action="store_true",
                        help="確認のみ（蓄積CSVは変更しない）")
    parser.add_argument("--excel",    type=str,  default=None,
                        help="Excelパスを明示 (省略=デフォルト)")
    parser.add_argument("--chikuseki",type=str,  default=None,
                        help="蓄積CSVフォルダ (省略=デフォルト)")
    args = parser.parse_args()

    excel_path    = Path(args.excel)     if args.excel     else EXCEL_PATH
    chikuseki_dir = Path(args.chikuseki) if args.chikuseki else CHIKUSEKI_DIR

    print()
    sep()
    print("  脅威スコア計算スクリプト (calc_threat.py)")
    if args.dry_run:
        print("  ※ DRY-RUN モード（蓄積CSVは変更されません）")
    sep()

    # ── Step1: Excel読み込み ──────────────────────────────────────────
    print()
    print("[Step 1]  Excelから脅威スコアを計算中...")
    if not excel_path.exists():
        print(f"  [NG] Excelが見つかりません: {excel_path}")
        return

    threat_db = load_excel_threats(excel_path)
    print(f"  [OK] {len(threat_db)}会場 読み込み完了")

    # 会場フィルタ
    if args.venue:
        if args.venue not in threat_db:
            print(f"  [NG] 指定会場 '{args.venue}' はExcelに存在しません")
            print(f"       利用可能: {sorted(threat_db.keys())}")
            return
        threat_db = {args.venue: threat_db[args.venue]}

    # ── Step2: 蓄積CSV更新 ─────────────────────────────────────────────
    print()
    print("[Step 2]  蓄積CSV 更新中...")

    pattern   = f"{args.venue}.csv" if args.venue else "*.csv"
    csv_files = sorted(glob.glob(str(chikuseki_dir / pattern)))

    if not csv_files:
        print(f"  [NG] 蓄積CSVが見つかりません: {chikuseki_dir / pattern}")
        return

    results = []
    for csv_path in csv_files:
        venue_name = Path(csv_path).stem
        if venue_name not in threat_db:
            results.append({
                "venue": venue_name, "status": "スキップ",
                "reason": "Excelにシートなし",
            })
            continue

        result = update_chikuseki_csv(
            Path(csv_path),
            threat_db[venue_name],
            venue_name,
            args.dry_run,
        )
        results.append(result)

        status = result["status"]
        reason = result.get("reason", "")
        if reason:
            print(f"  [{venue_name}] {status} ← {reason}")
        else:
            updated = result.get("races_updated", 0)
            print(f"  [{venue_name}] {status}  ({updated}レース更新)")

    # ── サマリー ──────────────────────────────────────────────────────
    print()
    sep()
    total_updated = sum(r.get("races_updated", 0) for r in results)
    venues_upd    = sum(1 for r in results if r["status"] == "更新")
    venues_skip   = sum(1 for r in results if r["status"] == "スキップ")

    print(f"  処理会場数      : {len(results)}")
    print(f"  更新会場        : {venues_upd}")
    print(f"  スキップ        : {venues_skip}")
    print(f"  更新レース総数  : {total_updated}")

    if args.dry_run:
        print()
        print("  ※ DRY-RUN のため蓄積CSVは変更されていません。")
        print("  ※ 実際に反映するには --dry-run を外して再実行してください。")
    else:
        print()
        print("  [完了] 蓄積CSVの脅威合計を更新しました。")
        print("  次のステップ: python apply_correction.py --dry-run")

    sep()
    print()


if __name__ == "__main__":
    main()
