# -*- coding: utf-8 -*-
"""
load_race.py  ─  エントリポイント（main のみ）

【使い方】
  python scripts/load_race.py --venue 大村
  python scripts/load_race.py --venue 大村 --race 5
  python scripts/load_race.py --venue 大村 --date 2026-02-15
  python scripts/load_race.py --venue 大村 --date 2026-02-15 --race 5
  python scripts/load_race.py   # csv_output/ の最新ファイルを自動検出

【モジュール構成】
  lr_config.py   ─ パス定数 / グローバル設定
  lr_utils.py    ─ 汎用ユーティリティ / Excel書式ヘルパー
  lr_masters.py  ─ マスタ / CSV 読み込み
  lr_calc.py     ─ 指数計算 / 展開分析
  lr_probs.py    ─ 3連単確率計算 / EV計算
  lr_suggest.py  ─ 買い目提案 / 本命判定
  lr_excel.py    ─ Excel書き込み / 数値シート
  lr_log.py      ─ 予想ログ / ROIバックテスト
"""

import os, sys, glob, argparse, re, pathlib, shutil, platform, subprocess
from datetime import datetime

# ── サブモジュール一括インポート ──────────────────────────────────────────
from lr_config import (
    _SCRIPT_VERSION, EXCEL_FILE, MASTER_FILE, CSV_DIR,
    # ACTUAL_ODDS_URL, ODDS_FILEPATH は事前情報外（EV計算除去済み）のためインポート不要
    _GRADE_CSV_MAP,
    SHEET_OUTPUT, SHEET_SAMPLE,
)
from lr_utils import sep, safe_float
from lr_masters import load_masters, load_csv, load_motor_csv
from lr_calc import (
    calc_race_indices, _predict_first_turn, _build_conflict_map,
    _judge_race_type, _generate_tenkai_story, _generate_buy_hint,
    _judge_ryotate, _judge_himo_are,
)
from lr_probs import (
    _calc_3rentan_probs_v2, _should_skip_race,
    # calc_ev_from_actual_odds, suggest_by_ev, load_actual_odds_from_excel は
    # 事前情報外（オッズ依存）のため除去済み
)
from lr_suggest import (
    _suggest_3rentan, _calc_venue_stats, build_jizen_members,
    _apply_jizen_honmei,
)
from lr_excel import write_race_flat, write_numeric_sheet

from lr_log import _flush_prediction_log, calc_roi_from_logs

try:
    from export_indices_csv import append_all_races as _append_indices_csv
    _INDICES_CSV_AVAILABLE = True
except ImportError:
    _INDICES_CSV_AVAILABLE = False
    print('[!]  export_indices_csv.py が見つかりません。指数CSV出力はスキップされます。')


import openpyxl
from openpyxl import load_workbook

try:
    from calc_3rentan_probs_v2 import calc_3rentan_probs_v2 as _ext_calc_3rentan
    _EXT_CALC_AVAILABLE = True
except ImportError:
    _EXT_CALC_AVAILABLE = False

sys.path.insert(0, str(pathlib.Path(__file__).parent))
try:
    from evaluate_jizen import evaluate_all, evaluate_all_with_scores, calculate_diversity_rate
    JIZEN_AVAILABLE = True
except ImportError:
    JIZEN_AVAILABLE = False
    print('[!]  evaluate_jizen.py が見つかりません。事前評価はスキップされます。')

try:
    from honmei_scenario import integrate_with_suggest_3rentan
    HONMEI_SCENARIO_AVAILABLE = True
except ImportError:
    HONMEI_SCENARIO_AVAILABLE = False
    print('??  honmei_scenario.py が見つかりません。既存ロジックで動作します。')

try:
    from scenario_engine import build_scenarios as _build_scenarios_new
    SCENARIO_ENGINE_AVAILABLE = True
    print('[OK] scenario_engine.py 読込完了（シナリオ分岐型エンジン）')
except ImportError:
    SCENARIO_ENGINE_AVAILABLE = False
    print('??  scenario_engine.py が見つかりません。既存ロジックで動作します。')

# ── apply_correction（バックテスト学習による買い目補正）──────────────────
try:
    from apply_correction import apply_correction as _apply_correction
    APPLY_CORRECTION_AVAILABLE = True
    print('[OK] apply_correction.py 読込完了（バックテスト補正エンジン）')
except ImportError:
    APPLY_CORRECTION_AVAILABLE = False
    print('[!]  apply_correction.py が見つかりません。補正なしで動作します。')

# ════════════════════════════════════════════════════════════════════
# 数値蓄積並列ワーカー（トップレベル関数 ― pickle可能にするため必須）
# ════════════════════════════════════════════════════════════════════
def _collect_venue_worker(venue: str, dates: list, recalc: bool,
                          chikuseki_dir, masters_args: tuple) -> tuple:
    """
    1会場分の全日付を処理してCSVに書き込む。
    ProcessPoolExecutor から呼ばれるためトップレベルに定義。
    戻り値: (venue, saved_count, skipped_count, error_count)
    """
    import pathlib as _pl

    (course_master, player_master, ininage_master, venue_stats_master,
     venue_course_master, tenkai_venue_master, tenkai_national_master,
     st_kimete_master, kaiho_venue_master, kaiho_national_master,
     effective_grade) = masters_args

    _chikuseki_dir = _pl.Path(chikuseki_dir)
    _csv_out = _chikuseki_dir / f"{venue}.csv"

    # スキップ判定用の既存キーを1回だけ読む（recalcなら不要）
    _existing_keys = set()
    if not recalc:
        try:
            from export_indices_csv import get_existing_keys as _gek
            _existing_keys = _gek(_csv_out)
        except Exception:
            pass

    saved = skipped = errors = 0
    _venue_all_race_data = []

    for _proc_date in dates:
        if not recalc:
            _dup_key = (_proc_date, venue, "1", "1")
            if _dup_key in _existing_keys:
                skipped += 1
                continue

        df, race_date = load_csv(venue, None, _proc_date)
        if df is None or len(df) == 0:
            continue

        motor_df = load_motor_csv(venue, None, race_df=df) if JIZEN_AVAILABLE else None

        race_col = next((c for c in df.columns if "レース" in c or c == "R"), None)
        race_nos = (
            sorted(df[race_col].unique(), key=lambda x: int(x) if str(x).isdigit() else 99)
            if race_col else ["1"]
        )

        for race_no in race_nos:
            if race_col:
                race_df = df[df[race_col].astype(str) == str(race_no)]
            else:
                race_df = df
            players = [dict(row) for _, row in race_df.iterrows()]
            if not players:
                continue

            try:
                results, slit, venue_stats, frame_2nd, race_judgment, bet_suggestions = calc_race_indices(
                    venue, race_no, players, course_master, player_master, ininage_master,
                    venue_stats_master, venue_course_master,
                    tenkai_venue_master=tenkai_venue_master,
                    tenkai_national_master=tenkai_national_master,
                    st_kimete_master=st_kimete_master,
                    kaiho_venue_master=kaiho_venue_master,
                    kaiho_national_master=kaiho_national_master,
                    race_grade=effective_grade,
                )
            except Exception:
                errors += 1
                continue

            jizen_eval_result = None
            if JIZEN_AVAILABLE:
                try:
                    jizen_members = build_jizen_members(
                        results, course_master, player_master, motor_df, race_no
                    )
                    if jizen_members:
                        jizen_eval_result = evaluate_all_with_scores(jizen_members)
                except Exception:
                    pass

            # ── 買い目生成（--collect モードでも bet_suggestions を蓄積CSVに保存）──
            # 通常モードと同じ scenario_engine / _suggest_3rentan で買い目を生成する。
            # これにより backtest_engine.py が実買い目で正確に突合できる。
            _bet_suggestions_for_csv = bet_suggestions  # calc_race_indices の初期値
            try:
                if SCENARIO_ENGINE_AVAILABLE:
                    from lr_probs import _calc_3rentan_probs_v2
                    _combos_col = _calc_3rentan_probs_v2(
                        results,
                        venue_course_1c_rate=race_judgment.get("venue_c1_win_rate"),
                        race_judgment=race_judgment,
                        tenkai_national=tenkai_national_master,
                        tenkai_venue=tenkai_venue_master,
                        venue_stats=None,
                    )
                    _bet_suggestions_for_csv = _build_scenarios_new(
                        results             = results,
                        race_judgment       = race_judgment,
                        combos              = _combos_col,
                        ininage_master      = ininage_master,
                        tenkai_venue        = tenkai_venue_master,
                        tenkai_national     = tenkai_national_master,
                        venue               = venue,
                        venue_stats         = None,
                        kaiho_venue         = kaiho_venue_master,
                        kaiho_national      = kaiho_national_master,
                        max_bets            = 12,
                        jizen_eval          = jizen_eval_result,
                        venue_course_master = venue_course_master,
                    )
                else:
                    _bet_suggestions_for_csv = _suggest_3rentan(
                        results, race_judgment,
                        jizen_eval=jizen_eval_result,
                        tenkai_venue=tenkai_venue_master,
                        tenkai_national=tenkai_national_master,
                        venue=venue,
                    )
            except Exception as _be:
                import traceback
                print(f"  [!] 買い目生成エラー ({venue} {race_no}R): {_be}")
                traceback.print_exc()
                pass  # 買い目生成失敗時は初期値のまま続行

            # ── バックテスト学習補正（collectモード）──────────────────────────
            if APPLY_CORRECTION_AVAILABLE and not _bet_suggestions_for_csv.get("skip"):
                try:
                    _row_data_for_corr = players[0] if players else {}
                    _bet_suggestions_for_csv = _apply_correction(
                        _bet_suggestions_for_csv, race_judgment, _row_data_for_corr
                    )
                except Exception:
                    pass  # collectモードでは補正失敗時は無視して続行
            # ─────────────────────────────────────────────────────────────────────

            _venue_all_race_data.append({
                "race_no":        race_no,
                "venue":          venue,
                "race_date":      race_date,
                "results":        results,
                "jizen_eval":     jizen_eval_result,
                "race_judgment":  race_judgment,
                "bet_suggestions": _bet_suggestions_for_csv,  # ← 追加
            })
            saved += 1

    # 会場まとめてCSV書き込み（I/O回数を最小化）
    if _venue_all_race_data:
        try:
            from export_indices_csv import append_all_races as _append
            _append(_venue_all_race_data, output_path=_csv_out)
        except Exception as e:
            print(f"  [!]  {venue} CSV出力エラー: {e}")

    return (venue, saved, skipped, errors)


def main():
    parser = argparse.ArgumentParser(description="ボートリサーチ新聞 全レース一括書き込み")
    parser.add_argument("--venue",      type=str, default=None, help="会場名 (例: 大村)")
    parser.add_argument("--race",       type=int, default=None, help="レース番号 (省略時: 全レース)")
    parser.add_argument("--date",       type=str, default=None, help="日付 (例: 2026-02-15, 省略時: 最新CSV)")
    parser.add_argument("--newspaper",  action="store_true",    help="新聞作成をスキップせず実行 (fill_newspaper.py)")
    parser.add_argument("--png",        action="store_true",    help="PNG発行をスキップせず実行 (xlsx_to_png_interactive.py)")
    parser.add_argument("--all",        action="store_true",    help="全会場を一括処理 (fill_newspaper.py --all も連動)")
    parser.add_argument("--grade",      type=str, default=None,
                        choices=["一般", "G1", "G2", "G3", "SG"],
                        help="レースグレード (例: G1 / SG)。省略時は対話式メニューで選択")
    parser.add_argument("--collect",    action="store_true",
                        help="数値蓄積のみモード（Excel・新聞処理をスキップ）")
    parser.add_argument("--recalc",     action="store_true",
                        help="数値蓄積の強制再計算モード（既存データを上書き・スキップなし）")
    parser.add_argument("--date-from",  type=str, default=None, dest="date_from",
                        help="再計算開始日 (例: 2026-01-01)。--recalc と併用")
    parser.add_argument("--date-to",    type=str, default=None, dest="date_to",
                        help="再計算終了日 (例: 2026-04-01)。--recalc と併用")
    parser.add_argument("--workers",    type=int, default=6,
                        help="並列処理ワーカー数（デフォルト: 6）")
    args = parser.parse_args()

    sep()
    print("  ボートリサーチ新聞 全レース一括書き込み")
    print(f"  バージョン: {_SCRIPT_VERSION}")
    sep()

    # ════════════════════════════════════════════════════════════════════
    # 📦 LZH → CSV 自動変換（scripts/ フォルダの LZH を解凍してCSV化）
    #     - scripts/ フォルダに .lzh/.LZH ファイルがあれば自動処理
    #     - lzh_to_csv.py の変換ロジックをインラインで実行
    #     - 変換後の LZH は scripts/lzh_archive/ に退避（二重処理防止）
    # ════════════════════════════════════════════════════════════════════
    _lzh_files = (
        glob.glob(str(pathlib.Path(__file__).parent / "*.lzh")) +
        glob.glob(str(pathlib.Path(__file__).parent / "*.LZH"))
    )
    if _lzh_files:
        sep("-")
        print(f"  ? LZHファイルを検出 ({len(_lzh_files)}件) → CSV変換を開始します")
        sep("-")
        _lzh_to_csv_script = pathlib.Path(__file__).parent / "lzh_to_csv.py"
        if not _lzh_to_csv_script.exists():
            print(f"  [!]  lzh_to_csv.py が見つかりません。LZH変換をスキップします。")
            print(f"     ({_lzh_to_csv_script})")
        else:
            try:
                _lzh_result = subprocess.run(
                    [sys.executable, str(_lzh_to_csv_script)],
                    capture_output=False,
                    check=False,
                    cwd=str(pathlib.Path(__file__).parent),  # scripts/ をカレントに
                )
                if _lzh_result.returncode == 0:
                    print(f"  [OK] LZH→CSV変換 完了")
                    # 変換済み LZH を lzh_archive/ フォルダに退避（二重変換防止）
                    _archive_dir = pathlib.Path(__file__).parent / "lzh_archive"
                    _archive_dir.mkdir(exist_ok=True)
                    for _lf in _lzh_files:
                        try:
                            if not pathlib.Path(_lf).exists():
                                # lzh_to_csv.py が内部で削除済みのケース → スキップ
                                print(f"  [BAK]  退避スキップ（変換済みで削除済み）: {pathlib.Path(_lf).name}")
                                continue
                            shutil.move(_lf, str(_archive_dir / pathlib.Path(_lf).name))
                            print(f"  [BAK]  退避: {pathlib.Path(_lf).name} → lzh_archive/")
                        except Exception as _mv_e:
                            print(f"  [!]  LZH退避失敗（続行します）: {_mv_e}")
                else:
                    print(f"  [!]  lzh_to_csv.py が終了コード {_lzh_result.returncode} で終了しました")
                    print(f"  [!]  CSV変換に失敗しましたが、既存CSVがあれば処理を続行します")
            except Exception as _lzh_e:
                print(f"  [NG] LZH→CSV変換でエラーが発生しました: {_lzh_e}")
                print(f"  [!]  既存CSVがあれば処理を続行します")
        sep("-")
        print()
    # ════════════════════════════════════════════════════════════════════

    # ════════════════════════════════════════════════════════════════════
    # [!]  Excel起動チェック（処理開始前）
    #     Excelが開いたまま処理すると最後の wb.save() で PermissionError になる。
    #     処理完了後に「Excelを閉じてください」と言われる時間ロスを防ぐため、
    #     最初にチェックして即座に警告する。
    # ════════════════════════════════════════════════════════════════════
    if EXCEL_FILE.exists():
        # Windows環境では .bak.xlsx への書き込みテストで開き確認を行う
        _lock_check_path = EXCEL_FILE.with_suffix(".~lock.xlsx")
        _is_excel_open = False

        # 方法①: ロックファイルの存在チェック（LibreOffice / Excel共通）
        if _lock_check_path.exists():
            _is_excel_open = True

        # 方法②: Windowsの場合は psutil でプロセスチェック
        if not _is_excel_open:
            try:
                import psutil
                excel_procs = [p for p in psutil.process_iter(["name"])
                               if p.info["name"] and "EXCEL" in p.info["name"].upper()]
                if excel_procs:
                    # プロセスは存在するが、対象ファイルを開いているかはファイル書き込みテストで確認
                    try:
                        with open(str(EXCEL_FILE), "a+b"):
                            pass
                    except PermissionError:
                        _is_excel_open = True
            except ImportError:
                pass  # psutil なし → 方法③へ

        # 方法③: ファイルへの書き込みテスト（最も確実）
        if not _is_excel_open:
            try:
                with open(str(EXCEL_FILE), "a+b"):
                    pass
            except PermissionError:
                _is_excel_open = True

        if _is_excel_open:
            print()
            print("  ????????????????????????????????????????????????????????")
            print("  ?  [NG]  Excelが開いています！                            ?")
            print("  ?                                                      ?")
            print(f"  ?  ? {str(EXCEL_FILE.name)[:48]:<48} ?")
            print("  ?                                                      ?")
            print("  ?  Excelを閉じてから、このスクリプトを再実行してください  ?")
            print("  ????????????????????????????????????????????????????????")
            print()
            return
        else:
            print(f"  [OK] Excel起動チェック: 閉じていることを確認 ({EXCEL_FILE.name})")
    # ════════════════════════════════════════════════════════════════════

    # ── 会場リスト確定（--all / --venue / 対話式）──────────────────────────
    # csv_outputから候補を収集（フラット + サブフォルダ両対応）
    _files_flat = sorted(glob.glob(str(CSV_DIR / "*.csv")))
    _files_sub  = sorted(glob.glob(str(CSV_DIR / "**" / "*.csv"), recursive=True))
    _all_files  = sorted(set(_files_flat + _files_sub))
    candidates_venue = []
    _seen_v = set()
    for _f in reversed(_all_files):  # 新しい順
        _parts = pathlib.Path(_f).relative_to(CSV_DIR).parts
        _v = _parts[0] if len(_parts) >= 2 else os.path.basename(_f).split("_")[0]
        # 日付パターン（YYYY-MM-DD）を含まないファイルは会場CSVではないので除外
        # 例: indices_log.csv, indices_log_test.csv など
        if not re.search(r"\d{4}-\d{2}-\d{2}", os.path.basename(_f)):
            continue
        if _v and _v not in _seen_v:
            candidates_venue.append(_v)
            _seen_v.add(_v)

    if not candidates_venue:
        print("[NG] --venue を指定するか、csv_output/ にCSVを置いてください")
        return

    if args.all:
        # --all: csv_output にある全会場を処理
        venues_to_process = candidates_venue
        print(f"  🏟️  全会場一括処理モード: {', '.join(venues_to_process)}")
    elif args.venue:
        venues_to_process = [args.venue]
        print(f"  ? 会場: {args.venue}")
    elif len(candidates_venue) == 1:
        venues_to_process = [candidates_venue[0]]
        print(f"  ? 会場を自動検出: {venues_to_process[0]}")
    else:
        print()
        print("  ┌─────────────────────────────────────┐")
        print("  │  ? 会場を選択してください            │")
        print("  └─────────────────────────────────────┘")
        print(f"    0. 全会場（一括処理）")
        for i, v in enumerate(candidates_venue, 1):
            print(f"    {i}. {v}")
        while True:
            sel = input(f"\n  番号を入力 (0=全会場, 1〜{len(candidates_venue)}): ").strip()
            if sel == "0":
                venues_to_process = candidates_venue
                print(f"  🏟️  全会場一括処理モード: {', '.join(venues_to_process)}")
                break
            if sel.isdigit() and 1 <= int(sel) <= len(candidates_venue):
                venues_to_process = [candidates_venue[int(sel) - 1]]
                print(f"  ? 会場: {venues_to_process[0]}")
                break
            print("  [!]  正しい番号を入力してください")

    # 後続処理との互換性のため venue を先頭会場で初期化
    venue = venues_to_process[0]
    # 対話式「0=全会場」選択時も --all と同じ挙動にするフラグ
    _process_all = args.all or (len(venues_to_process) > 1)

    # ════════════════════════════════════════════════════════════════════
    # モード選択（通常処理 / 数値蓄積のみ / Excelのみ書き込み）
    # ════════════════════════════════════════════════════════════════════
    _collect_only   = getattr(args, 'collect', False)
    _excel_only     = False  # 新モード: Excelのみ書き込み（新聞・PNG スキップ）
    if not _collect_only:
        print()
        print("  ┌─────────────────────────────────────┐")
        print("  │  実行モードを選択してください         │")
        print("  └─────────────────────────────────────┘")
        print("    1. 通常処理（Excel書き込み・新聞作成）")
        print("    2. 数値蓄積のみ（蓄積CSV）")
        print("    3. Excelのみ書き込み（新聞・PNG作成スキップ）")
        while True:
            _mode_sel = input("\n  番号を入力 (1〜3, Enter=1): ").strip()
            if _mode_sel in ("", "1"):
                _collect_only = False
                _excel_only   = False
                print("  ✓ モード: 通常処理")
                break
            if _mode_sel == "2":
                _collect_only = True
                _excel_only   = False
                print("  ✓ モード: 数値蓄積のみ")
                break
            if _mode_sel == "3":
                _collect_only = False
                _excel_only   = True
                print("  ✓ モード: Excelのみ書き込み（新聞・PNG スキップ）")
                break
            print("  [!]  1 〜 3 を入力してください")

    # ── 数値蓄積のみモード ────────────────────────────────────────────
    if _collect_only:
        _CHIKUSEKI_DIR = pathlib.Path(
            r"C:\Users\user\Desktop\データ収集\scripts\数値蓄積"
        )
        if not _CHIKUSEKI_DIR.exists():
            print(f"  [!]  数値蓄積フォルダが見つかりません: {_CHIKUSEKI_DIR}")
            return

        # --recalc フラグ確認
        _recalc_mode = getattr(args, 'recalc', False)
        _date_from   = getattr(args, 'date_from', None)
        _date_to     = getattr(args, 'date_to', None)
        _n_workers   = max(1, getattr(args, 'workers', 4))

        if _recalc_mode:
            _range_str = ""
            if _date_from and _date_to:
                _range_str = f" ({_date_from} 〜 {_date_to})"
            elif _date_from:
                _range_str = f" ({_date_from} 〜)"
            elif _date_to:
                _range_str = f" (〜 {_date_to})"
            print(f"  ⚡ 強制再計算モード{_range_str}  ワーカー数: {_n_workers}")

        # ── 処理対象の日付リストを会場ごとに個別収集 ─────────────────────
        # 全会場共通の日付リストを使うと「その会場にない日付」でNGログが大量発生するため
        # 会場ごとに自分のCSVが存在する日付だけを収集する
        venue_dates_map: dict = {}
        for _sv in venues_to_process:
            _dc = []
            _ds_seen = set()
            for _f in (sorted(glob.glob(str(CSV_DIR / f"{_sv}_*.csv"))) +
                       sorted(glob.glob(str(CSV_DIR / _sv / f"{_sv}_*.csv")))):
                _m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(_f))
                if _m and _m.group(1) not in _ds_seen:
                    _dc.append(_m.group(1)); _ds_seen.add(_m.group(1))
            _dc = sorted(_dc)
            if _date_from:
                _dc = [d for d in _dc if d >= _date_from]
            if _date_to:
                _dc = [d for d in _dc if d <= _date_to]
            venue_dates_map[_sv] = _dc

        _total_days = sum(len(v) for v in venue_dates_map.values())
        if _total_days == 0:
            print("  [!]  処理対象のCSVが見つかりません。")
            return
        _all_dates = sorted(set(d for v in venue_dates_map.values() for d in v))
        print(f"  📅 処理対象: {len(venues_to_process)}会場 / 延べ{_total_days}日分 ({_all_dates[0]} 〜 {_all_dates[-1]})")

        # マスタ読み込み（一般戦・1回だけ）
        if not MASTER_FILE.exists():
            print(f"  [NG] マスタファイルが見つかりません: {MASTER_FILE}")
            return
        try:
            wb_master = load_workbook(str(MASTER_FILE))
        except Exception as e:
            print(f"  [NG] マスタを開けませんでした: {e}")
            return
        (course_master, player_master, ininage_master, venue_stats_master,
         venue_course_master, tenkai_venue_master, tenkai_national_master,
         st_kimete_master, kaiho_venue_master, kaiho_national_master,
         effective_grade) = load_masters(wb_master, "一般")

        # ════════════════════════════════════════════════════════════════
        # 並列実行（会場単位）
        # ════════════════════════════════════════════════════════════════
        import concurrent.futures as _cf
        import time as _time

        # マスタをまとめて引数用タプルに（各ワーカーに渡す）
        _masters_args = (
            course_master, player_master, ininage_master, venue_stats_master,
            venue_course_master, tenkai_venue_master, tenkai_national_master,
            st_kimete_master, kaiho_venue_master, kaiho_national_master,
            effective_grade,
        )

        _t_start = _time.perf_counter()
        _total_saved = _total_skipped = _total_errors = 0

        # 会場が1つだけなら並列化不要（オーバーヘッドを避ける）
        if len(venues_to_process) == 1 or _n_workers == 1:
            for _v in venues_to_process:
                _v_dates = venue_dates_map.get(_v, [])
                if not _v_dates:
                    print(f"  [SKIP] {_v} 対象日付なし")
                    continue
                print(f"  [蓄積] {_v} 処理開始... ({len(_v_dates)}日分)")
                _, _s, _sk, _e = _collect_venue_worker(
                    _v, _v_dates, _recalc_mode, _CHIKUSEKI_DIR, _masters_args
                )
                _total_saved   += _s
                _total_skipped += _sk
                _total_errors  += _e
                print(f"  [OK]  {_v}  保存:{_s}R  スキップ:{_sk}日  エラー:{_e}R")
        else:
            # 複数会場を並列処理
            print(f"  🚀 {len(venues_to_process)}会場を {_n_workers}並列で処理します...")
            with _cf.ProcessPoolExecutor(max_workers=_n_workers) as executor:
                futures = {
                    executor.submit(
                        _collect_venue_worker,
                        _v, venue_dates_map.get(_v, []), _recalc_mode, _CHIKUSEKI_DIR, _masters_args
                    ): _v
                    for _v in venues_to_process
                    if venue_dates_map.get(_v)
                }
                for future in _cf.as_completed(futures):
                    try:
                        _v, _s, _sk, _e = future.result()
                        _total_saved   += _s
                        _total_skipped += _sk
                        _total_errors  += _e
                        print(f"  [OK]  {_v}  保存:{_s}R  スキップ:{_sk}日  エラー:{_e}R")
                    except Exception as _fe:
                        _v = futures[future]
                        print(f"  [NG]  {_v}  並列処理エラー: {_fe}")

        _elapsed = _time.perf_counter() - _t_start
        sep()
        print(f"  [完了] 数値蓄積 終了  ({_elapsed:.1f}秒)")
        print(f"         保存: {_total_saved}R  スキップ: {_total_skipped}日  エラー: {_total_errors}R")
        sep()
        return
    # ════════════════════════════════════════════════════════════════════

    # ── 日付選択（--date 未指定時のみ対話式）────────────────────────────────
    # 全会場一括の場合は全会場のCSVから日付候補をまとめて収集する
    if not args.date:
        _date_candidates = []
        _seen_dates = set()
        _scan_venues = venues_to_process if _process_all else [venue]
        for _sv in _scan_venues:
            _df_flat = sorted(glob.glob(str(CSV_DIR / f"{_sv}_*.csv")), reverse=True)
            _df_sub  = sorted(glob.glob(str(CSV_DIR / _sv / f"{_sv}_*.csv")), reverse=True)
            for _f in (_df_sub if _df_sub else _df_flat):
                _m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(_f))
                if _m:
                    _d = _m.group(1)
                    if _d not in _seen_dates:
                        _date_candidates.append(_d)
                        _seen_dates.add(_d)
        _date_candidates = sorted(_date_candidates, reverse=True)  # 新しい順に並べ直し

        if len(_date_candidates) > 1:
            print()
            print("  ┌─────────────────────────────────────┐")
            print("  │  📅 日付を選択してください            │")
            print("  └─────────────────────────────────────┘")
            for i, _d in enumerate(_date_candidates, 1):
                _latest = " ← 最新" if i == 1 else ""
                print(f"    {i}. {_d}{_latest}")
            print(f"    Enter = 最新({_date_candidates[0]})")
            while True:
                _sel = input(f"\n  番号を入力 (1〜{len(_date_candidates)}, Enter=最新): ").strip()
                if _sel == "":
                    args.date = _date_candidates[0]  # 最新日付を明示的に確定
                    print(f"  ✓ 日付: {args.date}（最新）")
                    break
                if _sel.isdigit() and 1 <= int(_sel) <= len(_date_candidates):
                    args.date = _date_candidates[int(_sel) - 1]
                    print(f"  ✓ 日付: {args.date}")
                    break
                if re.match(r"\d{4}-\d{2}-\d{2}", _sel):
                    args.date = _sel
                    print(f"  ✓ 日付: {args.date}")
                    break
                print("  [!]  正しい番号か日付(例: 2026-03-20)を入力してください")
        elif _date_candidates:
            args.date = _date_candidates[0]
            print(f"  ✓ 日付: {args.date}（自動選択）")

    # ── グレード選択 ──────────────────────────────────────────────────────
    # --grade 引数が指定されていれば即確定
    # --all のとき「一般」を自動選択（対話しない）
    # それ以外は対話式メニュー
    _GRADE_CHOICES = ["一般", "G1", "G2", "G3", "SG"]
    if args.grade:
        race_grade = args.grade
    elif _process_all:
        race_grade = "一般"
        print(f"  ? グレード: {race_grade}（全会場処理のため自動選択）")
    else:
        print()
        print("  ┌─────────────────────────────────────┐")
        print("  │  ? レースグレードを選択してください  │")
        print("  └─────────────────────────────────────┘")
        _grade_labels = {
            "一般": "一般戦（デフォルト）",
            "SG":   "SG（蒲郡SGクラシック等・専用パッチ適用）",
            "G1":   "G1",
            "G2":   "G2 / G3",
        }
        _grade_menu = ["一般", "SG", "G1", "G2"]   # 4択（SG独立）
        for i, g in enumerate(_grade_menu, 1):
            suffix = " ← G3はこちら" if g == "G2" else ""
            print(f"    {i}. {_grade_labels[g]}{suffix}")
        while True:
            sel = input(f"\n  番号を入力 (1〜{len(_grade_menu)}, Enter=1): ").strip()
            if sel == "":
                race_grade = "一般"
                break
            if sel.isdigit() and 1 <= int(sel) <= len(_grade_menu):
                race_grade = _grade_menu[int(sel) - 1]
                break
            print("  [!]  正しい番号を入力してください")
        print(f"  ? グレード: {race_grade}")

    # ── グレードレース選択時：マスタを毎回自動集計 ──────────────────────────
    # 一般戦以外（G1/SG・G2/G3）を選んだ場合、そのまま update_master.py を実行して
    # グレード別マスタCSVを最新化してからマスタ読み込みに進む。
    # ・一般戦は従来通りスキップ（毎日 update_master.py を別途実行する運用）
    #   → update_master.py 側で SG/G1〜G3 を完全除外した一般戦のみデータで集計する
    # ・G1/SG・G2/G3はレース前に必ず最新データで集計し直す
    #   → update_master.py 側で「一般戦＋グレードレース」のミックスで集計する
    if race_grade != "一般":
        print()
        print(f"  ????????????????????????????????????????????????????????")
        print(f"  ?  ? {race_grade}用マスタを集計します（毎回実行）              ?")
        print(f"  ?     update_master.py --grade {race_grade} を実行中...           ?")
        print(f"  ????????????????????????????????????????????????????????")
        print()
        _um_script = pathlib.Path(__file__).parent / "update_master.py"
        if not _um_script.exists():
            print(f"  [NG] update_master.py が見つかりません: {_um_script}")
            print(f"  [!]  マスタなしで続行します（一般戦マスタにフォールバック）")
        else:
            # SG は G1 マスタと共用なので --grade G1 を渡す
            _grade_arg = "G1" if race_grade == "SG" else race_grade
            _um_result = subprocess.run(
                [sys.executable, str(_um_script), "--grade", _grade_arg],
                check=False,
                cwd=str(pathlib.Path(__file__).parent),
            )
            if _um_result.returncode == 0:
                print()
                print(f"  [OK] {race_grade}用マスタ集計完了")
            else:
                print()
                print(f"  [!]  マスタ集計が異常終了しました（returncode={_um_result.returncode}）")
                print(f"  [!]  一般戦マスタにフォールバックして続行します")
        print()

    # ── 新聞作成・PNG発行 の選択 ──────────────────────────────────────────
    # --newspaper / --png フラグがあればそちら優先（bat経由など）
    # モード3（Excelのみ書き込み）のとき新聞・PNGは常にスキップ
    # --all のとき通常処理なら新聞・PNGも自動で有効化（対話しない）
    # それ以外は対話式
    if _excel_only:
        # モード3: Excelのみ書き込み → 新聞・PNG は常にスキップ
        _run_newspaper = False
        _run_png       = False
        print("  ✓ 新聞作成・PNG発行: スキップ（Excelのみ書き込みモード）")
    elif _process_all:
        _run_newspaper = True
        _run_png       = True
    elif args.newspaper:
        _run_newspaper = True
        _run_png       = args.png
    elif _collect_only:
        # モード2（数値蓄積のみ）: 新聞・PNG は常にスキップ（質問しない）
        _run_newspaper = False
        _run_png       = False
        print("  ✓ 新聞作成・PNG発行: スキップ（数値蓄積のみモード）")
    else:
        # モード1（通常処理）: 新聞・PNG は自動で有効（質問しない）
        _run_newspaper = True
        _run_png       = True
        print("  ✓ 新聞作成・PNG発行: 有効（通常処理モード）")

    print()

    # マスタExcelを開く（ボートリサーチ_マスタ.xlsx）
    if not MASTER_FILE.exists():
        print(f"[NG] マスタファイルが見つかりません: {MASTER_FILE}")
        return
    print(f"  ? マスタデータ読み込み中...")
    try:
        wb_master = load_workbook(str(MASTER_FILE))
    except Exception as e:
        print(f"[NG] マスタExcelを開けませんでした: {e}")
        print("   Excelが開いている場合は閉じてから再実行してください")
        return

    # マスタ読み込み
    (course_master, player_master, ininage_master, venue_stats_master,
     venue_course_master, tenkai_venue_master, tenkai_national_master,
     st_kimete_master, kaiho_venue_master, kaiho_national_master,
     effective_grade) = load_masters(wb_master, race_grade)

    # 新聞出力用Excelを開く（ボートリサーチ新聞_軽量版.xlsx）
    if not EXCEL_FILE.exists():
        print(f"[NG] Excelファイルが見つかりません: {EXCEL_FILE}")
        return
    try:
        wb = load_workbook(str(EXCEL_FILE))
    except Exception as e:
        print(f"[NG] Excelを開けませんでした: {e}")
        print("   Excelが開いている場合は閉じてから再実行してください")
        return

    # ════════════════════════════════════════════════════════════════════
    # 🔒 日付ロックファイルの書き出し
    #    fill_newspaper.py / xlsx_to_png_interactive.py は
    #    このファイルが存在し、日付が一致する場合のみ実行を許可する。
    # ════════════════════════════════════════════════════════════════════
    _lock_date = args.date  # --date 指定値（None の場合は後でループ内の race_date で更新）
    _lock_file = pathlib.Path(__file__).parent / ".race_date_lock"
    try:
        _lock_file.write_text(_lock_date or "", encoding="utf-8")
    except Exception as _lf_e:
        print(f"  [!]  日付ロックファイルの書き込みに失敗しました（続行します）: {_lf_e}")

    # ── 全会場ループ（--all / 単独会場 共通） ───────────────────────────────
    for venue in venues_to_process:
        if len(venues_to_process) > 1:
            sep()
            print(f"  🏟️  [{venue}] 処理開始")
            sep()

        # CSV読み込み
        print(f"  ? CSVを読み込み中...")
        df, race_date = load_csv(venue, args.race, args.date)
        if df is None or len(df) == 0:
            print(f"[NG] {venue}のCSVが見つかりません: {CSV_DIR / venue}*.csv")
            if len(venues_to_process) > 1:
                print(f"  ⚠️  [{venue}] スキップして次の会場へ")
                continue
            return

        # ロックファイルに確定した race_date を上書き（--date 未指定時は CSV から取得した日付を記録）
        if race_date and not _lock_date:
            try:
                _lock_file.write_text(race_date, encoding="utf-8")
                _lock_date = race_date
            except Exception:
                pass

        # モーターデータを出走表CSVから直接取得（scrape_motor.py 不要）
        motor_df = load_motor_csv(venue, None, race_df=df) if JIZEN_AVAILABLE else None
        if motor_df is None and JIZEN_AVAILABLE:
            print("  [!]  モーターデータなし。機力評価は '-' で出力します。")

        # レース番号一覧
        race_col = next((c for c in df.columns if "レース" in c or c == "R"), None)
        if race_col:
            race_nos = sorted(df[race_col].unique(), key=lambda x: int(x) if str(x).isdigit() else 99)
        else:
            race_nos = [str(args.race)] if args.race else ["1"]
        
        print(f"  ? 対象レース: {list(race_nos)}")
        print(f"  ??  Excel書き込み中: {EXCEL_FILE.name}")
        print()

        # 各レース書き込み
        tmp_image_paths = []
        all_race_data   = []  # 数値シート用にレースデータを蓄積
        _log_entries    = []  # 予想ログ一括書き込み用（12R分をまとめてflush）

        for race_no in race_nos:
            rno_int = int(race_no) if str(race_no).isdigit() else 0
            if args.race and rno_int != args.race:
                continue
        
            # このレースの選手データ取得
            if race_col:
                race_df = df[df[race_col].astype(str) == str(race_no)]
            else:
                race_df = df
        
            players = []
            for _, row in race_df.iterrows():
                players.append(dict(row))
        
            if not players:
                continue
        
            # 締め切り時刻（同レース全選手共通なので先頭行から取得）
            # 列名ゆれ（締切時刻 / 締切 / 締め切り時刻）に対応
            _deadline_raw = (
                players[0].get("締切時刻") or
                players[0].get("締切") or
                players[0].get("締め切り時刻") or ""
            )
            deadline = str(_deadline_raw).strip()
            deadline = None if deadline in ("", "None", "nan") else deadline
        
            # 指数計算（改善①②③: 会場別コースマスタ・シナリオ確率・動的ハイブリッド係数）
            results, slit, venue_stats, frame_2nd, race_judgment, bet_suggestions = calc_race_indices(
                venue, race_no, players, course_master, player_master, ininage_master,
                venue_stats_master, venue_course_master,
                tenkai_venue_master=tenkai_venue_master,
                tenkai_national_master=tenkai_national_master,
                st_kimete_master=st_kimete_master,
                kaiho_venue_master=kaiho_venue_master,
                kaiho_national_master=kaiho_national_master,
                race_grade=effective_grade,
            )
        
            # ── venue_stats に展開別残存マスタ由来の正確なコース別決まり手分布を注入 ──
            # _calc_venue_stats が返す kimete_by_course は会場統計の全体決まり手率の近似値。
            # 展開別残存マスタ（tenkai_venue_master）にはコース×決まり手のレース数が含まれるため
            # ここでコース別の正確な比率を計算して venue_stats を上書きする。
            # これにより scenario_engine._get_venue_kimete_dist が正確な会場固有データを参照できる。
            if venue_stats and tenkai_venue_master and venue:
                _kimete_by_course: dict[str, dict[str, float]] = {}
                for course_int in range(2, 7):
                    c_str = str(course_int)
                    _counts: dict[str, float] = {}
                    # tenkai_venue_master のキーは (会場名, 決まり手, 1着コース, 進入コース) など
                    # 1着コース = course_int のデータを集計（進入コース軸で決まり手を集計）
                    for key, row in tenkai_venue_master.items():
                        if not isinstance(key, tuple):
                            continue
                        # キー形式: (会場名, 決まり手, 1着コース, 進入コース) or 変形
                        try:
                            if len(key) >= 3 and str(key[0]) == venue and str(key[2]) == c_str:
                                kimete_k = str(key[1])
                                if kimete_k in ("恵まれ", "抜き"):
                                    continue
                                n = safe_float(row.get("レース数") or row.get("n_races")) or 0
                                _counts[kimete_k] = _counts.get(kimete_k, 0) + n
                        except (IndexError, TypeError):
                            continue
                    total = sum(_counts.values())
                    if total > 0:
                        _kimete_by_course[c_str] = {k: v / total for k, v in _counts.items()}
                if _kimete_by_course:
                    venue_stats["kimete_by_course"] = _kimete_by_course

            # ── 倶楽部流 事前評価用メンバーデータ組み立て ──────────────────
            jizen_members = None
            jizen_eval_result = None
            if JIZEN_AVAILABLE:
                try:
                    jizen_members = build_jizen_members(
                        results, course_master, player_master, motor_df, race_no
                    )
                    if jizen_members:
                        jizen_eval_result = evaluate_all_with_scores(jizen_members)
                except Exception as e:
                    print(f"  [!]  事前評価計算エラー ({race_no}R): {e}")

            # ── 印・買い目の確定（jizen有無に関わらず必ず実行）────────────────
            # ① first_prob_map確定のため bet_suggestions を一度計算
            # ② first_prob_map を使って6視点印スコアを確定（_apply_jizen_honmei）
            # ③ 確定した印を honmei_map に変換して買い目を再生成（印と完全連動）
            # ④ 最終 s1_prob で ryotate を再計算し、3択判定と確率の一貫性を保証
            try:
                # ── ①②: first_prob_map確定 → 印スコア確定 ───────────────────
                if SCENARIO_ENGINE_AVAILABLE:
                    _combos_main = _calc_3rentan_probs_v2(
                        results,
                        venue_course_1c_rate=race_judgment.get("venue_c1_win_rate"),
                        race_judgment=race_judgment,
                        tenkai_national=tenkai_national_master,
                        tenkai_venue=tenkai_venue_master,
                        venue_stats=venue_stats,
                    )
                    bet_suggestions = _build_scenarios_new(
                        results              = results,
                        race_judgment        = race_judgment,
                        combos               = _combos_main,
                        ininage_master       = ininage_master,
                        tenkai_venue         = tenkai_venue_master,
                        tenkai_national      = tenkai_national_master,
                        venue                = venue,
                        venue_stats          = venue_stats,
                        kaiho_venue          = kaiho_venue_master,
                        kaiho_national       = kaiho_national_master,
                        max_bets             = 12,
                        jizen_eval           = jizen_eval_result,
                        venue_course_master  = venue_course_master,
                    )
                else:
                    bet_suggestions = _suggest_3rentan(
                        results, race_judgment, jizen_eval=jizen_eval_result,
                        tenkai_venue=tenkai_venue_master, tenkai_national=tenkai_national_master, venue=venue
                    )
                _first_prob_map  = bet_suggestions.get("first_prob_map", {})
                _tobi_prob_final = (race_judgment.get("ryotate", {})
                                    .get("tobi_score", 0) or 0)
                # venue_stats を渡して6人相互作用モデルが会場特性（決まり手比率等）を参照できるようにする
                _venue_stats_for_honmei = _calc_venue_stats(venue_stats_master, venue)
                _apply_jizen_honmei(results, _tobi_prob_final, jizen_eval_result,
                                    first_prob_map=_first_prob_map,
                                    venue_stats=_venue_stats_for_honmei,
                                    race_judgment=race_judgment,
                                    st_kimete_master=st_kimete_master)  # トップレベル関数
                # 1号艇の「逃◎/逃○/逃△/逆×」を honmei_map に格納する際の変換。
                # 【重要】逃◎を攻め◎と同じ「◎」で格納すると honmei_scenario の逆引きで
                # 「◎が2艇存在」→後勝ちで2〜6号艇の攻め◎が本命扱いになり飛び軸に倒れる。
                # 1号艇には専用キー「逃」を使い攻め◎と衝突させない。
                # シナリオエンジンは s1_prob を直接使うため honmei_map への依存はないが
                # _apply_jizen_honmei との連携のため引き続き生成する。
                _nige_to_honmei = {"逃◎": "逃", "逃○": "逃○", "逃△": "逃△", "逃×": " "}
                _honmei_map = {}
                for r in results:
                    w = str(r["waku"])
                    h = r.get("honmei", " ")
                    if w == "1":
                        _honmei_map[w] = _nige_to_honmei.get(h, " ")
                    else:
                        _honmei_map[w] = h

                # ── ③: 印確定後に買い目を最終生成 ─────────────────────────────
                if SCENARIO_ENGINE_AVAILABLE:
                    bet_suggestions = _build_scenarios_new(
                        results              = results,
                        race_judgment        = race_judgment,
                        combos               = _combos_main,
                        ininage_master       = ininage_master,
                        tenkai_venue         = tenkai_venue_master,
                        tenkai_national      = tenkai_national_master,
                        venue                = venue,
                        venue_stats          = venue_stats,
                        kaiho_venue          = kaiho_venue_master,
                        kaiho_national       = kaiho_national_master,
                        max_bets             = 12,
                        jizen_eval           = jizen_eval_result,
                        venue_course_master  = venue_course_master,
                    )
                else:
                    bet_suggestions = _suggest_3rentan(
                        results, race_judgment,
                        jizen_eval=jizen_eval_result,
                        honmei_map=_honmei_map,
                        tenkai_venue=tenkai_venue_master, tenkai_national=tenkai_national_master, venue=venue,
                    )

                # ④ 印確定・買い目確定後の最終 s1_prob で ryotate を再計算
                #    これにより「3択判定の%」と「展開シナリオ確率」が同一モデルに統一される
                _s1_prob_final = bet_suggestions.get("s1_prob")
                if _s1_prob_final is not None:
                    _tobi_scenario = race_judgment.get("ryotate", {})
                    # tobi_scenario は calc_race_indices 内で生成した tobi オブジェクトが
                    # race_judgment["ryotate"] に格納されているが、tobi_prob は ryotate から取れる
                    _tobi_for_ryotate = {
                        "tobi_prob":   race_judgment["ryotate"].get("tobi_score", 30),
                        "main_threat": race_judgment["ryotate"].get("main_threat",
                                       bet_suggestions.get("fly_axes", [None])[0] or "-"),
                        "tobi_type":   race_judgment["ryotate"].get("tobi_type", "不明"),
                    }
                    _ryotate_final = _judge_ryotate(
                        race_judgment, _tobi_for_ryotate, {},
                        s1_prob=_s1_prob_final
                    )
                    race_judgment["ryotate"] = _ryotate_final

                    # ◎艇番とs1_prob最大艇の乖離チェックを印確定後に最終更新
                    _fpm = bet_suggestions.get("first_prob_map", {})
                    if _fpm:
                        _top_waku = max(_fpm, key=_fpm.get)
                        _honmei_w = next(
                            (str(r["waku"]) for r in results
                             if r.get("honmei") == "◎" and r["waku"] != "1"), None
                        )
                        if _honmei_w and _honmei_w != _top_waku:
                            race_judgment["honmei_prob_mismatch"] = True
                            race_judgment["honmei_prob_mismatch_detail"] = (
                                f"◎={_honmei_w}号艇 vs 確率最大={_top_waku}号艇"
                                f"({_fpm.get(_top_waku, 0)*100:.1f}%)"
                            )
                        else:
                            race_judgment["honmei_prob_mismatch"] = False
                            race_judgment["honmei_prob_mismatch_detail"] = ""

            except Exception as e:
                print(f"  [!]  印・買い目確定エラー ({race_no}R): {e}")

            # ── 3連対指数: 展開補正を後付け適用 ────────────────────────────
            # build_scenarios 完了後に sanren_idx を最終更新する（第4層展開ペナルティ）
            # cascade_scores は candidates の確率分布から逆算して適用。
            # lr_calc.py 側で sanren_idx_needs_tenkai_update=True がセットされている
            # 場合のみ実行（エラー時は既存の sanren_idx をそのまま使う）。
            if race_judgment.get("sanren_idx_needs_tenkai_update"):
                try:
                    from lr_sanren_idx import recalc_sanren_idx_with_tenkai
                    results = recalc_sanren_idx_with_tenkai(
                        results, venue,
                        venue_course_master, venue_stats_master,
                        bet_suggestions,
                    )
                except Exception as _sri_e:
                    print(f"  [!]  3連対指数展開補正エラー（スキップ）: {_sri_e}")

            # 数値シート用にデータを蓄積
            all_race_data.append({
                "race_no":         race_no,
                "venue":           venue,
                "race_date":       race_date,
                "results":         results,
                "slit":            slit,
                "venue_stats":     venue_stats,
                "frame_2nd":       frame_2nd,
                "deadline":        deadline,
                "jizen_members":   jizen_members,
                "jizen_eval":      jizen_eval_result,
                "player_master":   player_master,
                "tmp_image_paths": tmp_image_paths,
                "race_judgment":   race_judgment,
                "bet_suggestions": bet_suggestions,
            })

            # ── バックテスト学習補正（apply_correction）──────────────────────────
            # correction_table.json が存在する場合、学習済みの補正を買い目に適用する。
            # composite_score・buy_level（STRONG_BUY〜SKIP）が bet_suggestions に追加される。
            # correction_table.json がなければ補正なしで通過（安全設計）。
            if APPLY_CORRECTION_AVAILABLE and not bet_suggestions.get("skip"):
                try:
                    # players の先頭行を row_data として渡す（会場・日付・レース番号を含む）
                    _row_data_for_corr = players[0] if players else {}
                    bet_suggestions = _apply_correction(
                        bet_suggestions, race_judgment, _row_data_for_corr
                    )
                    _buy_level = bet_suggestions.get("buy_level", "")
                    if _buy_level:
                        print(f"    [補正] buy_level={_buy_level}  "
                              f"composite={bet_suggestions.get('composite_score', 0):.3f}")
                except Exception as _corr_e:
                    print(f"  [!]  apply_correction エラー ({race_no}R): {_corr_e}")
            # ─────────────────────────────────────────────────────────────────────

            # ── EV計算: 除去済み（オッズは締め切り後にしか確定しない事前情報外データ）──
            # ─────────────────────────────────────────────────────────────────────

            # 予想ログ蓄積（ループ後に _flush_prediction_log で一括書き込み）
            _log_entries.append((race_no, bet_suggestions, race_judgment))

            # ── 見送り判定ログ ──────────────────────────────────────────────────
            if bet_suggestions.get("skip"):
                print(f"  {bet_suggestions.get('skip_reason', '?見送り')} ({race_no}R)")

            # ── 展示前フォーメーション コンソール出力 ──────────────────────────
            try:
                jf = bet_suggestions.get("jizen_formation", {})
                if jf and jf.get("formation"):
                    jf_axes  = jf.get("axes", [])
                    jf_himo  = jf.get("himo", [])
                    jf_pts   = jf.get("point_count", 0)
                    jf_form  = jf.get("formation", [])
                    jf_tenk  = jf.get("tenkai_priority", [])
                    jf_aisho = jf.get("aisho_axes", [])

                    # 1着軸に事前評価記号を付けて表示
                    _in_sym  = (jizen_eval_result or {}).get("in_nige", [""] * 6)
                    _a_sym   = (jizen_eval_result or {}).get("aisho",   [""] * 6)
                    def _axis_label(w):
                        idx   = int(w) - 1
                        marks = []
                        if idx < len(_in_sym) and _in_sym[idx]:
                            marks.append(_in_sym[idx])
                        if idx < len(_a_sym) and _a_sym[idx]:
                            marks.append(_a_sym[idx])
                        return f"{w}号[{'/'.join(marks)}]" if marks else f"{w}号"

                    axes_str = "  ".join(_axis_label(w) for w in jf_axes)
                    himo_str = "  ".join(
                        f"{w}号" for w in jf_himo
                    )
                    # 組み合わせを5点ずつ区切って表示
                    form_lines = []
                    for ci in range(0, len(jf_form), 5):
                        form_lines.append("    " + "  ".join(jf_form[ci:ci+5]))

                    print()
                    print(f"  ┌─ ? {race_no}R 展示前フォーメーション ({jf_pts}点) ─────────────")
                    print(f"  │ 【1着軸】 {axes_str}")
                    print(f"  │ 【ヒ モ】 {himo_str}")
                    print(f"  │ 【組合せ】")
                    for fl in form_lines:
                        print(f"  │{fl}")
                    if jf_aisho:
                        print(f"  │ ②相性◎○: {'/'.join(jf_aisho)}号艇")
                    if jf_tenk:
                        print(f"  │ ⑤展開◎○: {'/'.join(jf_tenk)}号艇 → ヒモ優先")
                    print(f"  └────────────────────────────────────────────────────")
            except Exception:
                pass  # 表示失敗は無視して続行
            # ─────────────────────────────────────────────────────────────────────

            print(f"  {race_no}R 処理中... → 完了")

        # 予想ログ一括書き込み（12R分を1回のファイルI/Oで保存）
        if _log_entries:
            _flush_prediction_log(venue, race_date, _log_entries)

        # 数値シート書き込み
        if all_race_data:
            write_numeric_sheet(wb, all_race_data, course_master, venue_course_master)

        # ── 指数CSV蓄積（会場別ファイルへ自動振り分け）───────────────────
        # 保存先: C:\Users\user\Desktop\データ収集\scripts\数値蓄積\{会場名}.csv
        # フォルダが存在しない場合は既定パス（csv_output/indices_log.csv）にフォールバック
        if all_race_data and _INDICES_CSV_AVAILABLE:
            try:
                _CHIKUSEKI_DIR = pathlib.Path(
                    r"C:\Users\user\Desktop\データ収集\scripts\数値蓄積"
                )
                if _CHIKUSEKI_DIR.exists():
                    _csv_out = _CHIKUSEKI_DIR / f"{venue}.csv"
                else:
                    # フォルダが見つからない場合は既定パスへ（警告のみ）
                    print(f"  [!]  数値蓄積フォルダが見つかりません。既定パスへ出力します。")
                    print(f"       ({_CHIKUSEKI_DIR})")
                    _csv_out = None
                _append_indices_csv(all_race_data, output_path=_csv_out)
            except Exception as _csv_e:
                print(f'  [!]  指数CSV出力エラー（続行します）: {_csv_e}')
        # ────────────────────────────────────────────────────────────────

        # Excelを保存（会場ごとに保存して軽量版Excelにシートを積み上げる）
        print()
        print(f"  [SAVE] [{venue}] 保存中...")

        # 保存前に自動バックアップ（先頭会場のみ）
        if venue == venues_to_process[0]:
            bak_file = EXCEL_FILE.with_suffix(".bak.xlsx")
            try:
                if EXCEL_FILE.exists():
                    shutil.copy2(str(EXCEL_FILE), str(bak_file))
                    print(f"  [BAK]  バックアップ作成: {bak_file.name}")
            except Exception as e:
                print(f"  [!]  バックアップ失敗（続行します）: {e}")

        try:
            wb.save(str(EXCEL_FILE))
            print(f"  [OK] [{venue}] 完了！「{venue}_数値」シートを確認してください")
        except PermissionError:
            print("[NG] 保存エラー: Excelが開いています。閉じてから再実行してください")
            break  # 保存不可のときはループを抜ける

    # ── 全会場ループ完了後: 新聞作成（fill_newspaper.py）──────────────────
    # args.date は日付選択で必ず確定済み（Enter=最新 / 自動選択時も代入済み）
    _selected_date = args.date  # PNG発行ブロックでも参照するためここで確定
    _newspaper_ok = False
    if _run_newspaper:
        sep()
        print("  [NEWS] 新聞作成を開始します（fill_newspaper.py）")
        sep()
        _newspaper_script = pathlib.Path(__file__).parent / "fill_newspaper.py"
        if not _newspaper_script.exists():
            print(f"  [NG] fill_newspaper.py が見つかりません: {_newspaper_script}")
        else:
            try:
                # 全会場のとき fill_newspaper も --all で呼ぶ
                # 単独会場のときは --venue を渡す
                # --date を渡すことで選択した日付の会場のみ新聞を発行する
                if _process_all:
                    _news_args = ["--all"]
                else:
                    _news_args = ["--venue", venues_to_process[0]]
                # _selected_date は if _run_newspaper: の外で確定済み
                if _selected_date:
                    _news_args += ["--date", _selected_date]
                _result = subprocess.run(
                    [sys.executable, str(_newspaper_script)] + _news_args,
                    check=False
                )
                if _result.returncode == 0:
                    _newspaper_ok = True
                else:
                    print(f"  [!]  fill_newspaper.py が終了コード {_result.returncode} で終了しました")
                    print(f"  [!]  新聞作成に失敗したため PNG 発行をスキップします")
            except Exception as _e:
                print(f"  [NG] fill_newspaper.py の実行に失敗しました: {_e}")

    # ── PNG発行（xlsx_to_png_interactive.py）──────────────────────────────
    if _run_png and _newspaper_ok:
        sep()
        print("  [PNG]  PNG発行を開始します（xlsx_to_png_interactive.py）")
        sep()
        _png_script = pathlib.Path(__file__).parent / "xlsx_to_png_interactive.py"
        if not _png_script.exists():
            print(f"  [NG] xlsx_to_png_interactive.py が見つかりません: {_png_script}")
        else:
            try:
                _png_args = []
                # fill_newspaper.py に渡したのと同じ日付を引き継ぐ
                if _selected_date:
                    _png_args += ["--date", _selected_date]
                _result = subprocess.run(
                    [sys.executable, str(_png_script)] + _png_args,
                    check=False
                )
                if _result.returncode != 0:
                    print(f"  [!]  xlsx_to_png_interactive.py が終了コード {_result.returncode} で終了しました")
            except Exception as _e:
                print(f"  [NG] xlsx_to_png_interactive.py の実行に失敗しました: {_e}")

    # ── 完了通知（Windowsトースト）────────────────────────────────────────
    try:
        from win10toast import ToastNotifier
        _venue_str = "全会場" if _process_all else venues_to_process[0]
        _msg_parts = [f"{_venue_str} 処理完了"]
        if _run_newspaper:
            _msg_parts.append("新聞作成 [OK]")
        if _run_png:
            _msg_parts.append("PNG発行 [OK]")
        ToastNotifier().show_toast(
            "ボートリサーチ新聞",
            " / ".join(_msg_parts),
            duration=8,
            threaded=True,
        )
    except ImportError:
        pass  # win10toast未インストールの場合はスキップ
    except Exception:
        pass

    sep()

if __name__ == "__main__":
    main()
