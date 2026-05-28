# -*- coding: utf-8 -*-
"""
ボートリサーチ新聞 - 全レース一括新聞出力スクリプト
=====================================================
【使い方】
  python scripts/load_race.py --venue 大村
  python scripts/load_race.py --venue 大村 --race 5
  python scripts/load_race.py --venue 大村 --date 2026-02-15
  python scripts/load_race.py --venue 大村 --date 2026-02-15 --race 5
  python scripts/load_race.py   # csv_output/ の最新ファイルを自動検出

【出力先】
  ボートリサーチ新聞_軽量版.xlsx の「出力_新聞」シート
  ※「出力_新聞サンプル」のレイアウト（行高・列幅・結合・色）を完全再現
  ※データのみ上書き。行高・列幅は一切変更しない
"""

import os
import sys
import glob
import argparse
import re
import pathlib
# import copy  # 未使用
import shutil
import platform
import subprocess
import statistics
# import tempfile  # 未使用（BytesIO移行後に不要）
import json
import itertools
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    # GradientFill,  # 未使用
)
# from openpyxl.styles.differential import DifferentialStyle  # 未使用
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    print("⚠️  Pillowが見つかりません。ST舟図は出力されません。")

try:
    from calc_3rentan_probs_v2 import calc_3rentan_probs_v2 as _ext_calc_3rentan
    _EXT_CALC_AVAILABLE = True
except ImportError:
    _EXT_CALC_AVAILABLE = False

# 倶楽部流 事前評価モジュール（scripts/ フォルダに配置）
sys.path.insert(0, str(pathlib.Path(__file__).parent))
try:
    from evaluate_jizen import evaluate_all, calculate_diversity_rate
    JIZEN_AVAILABLE = True
except ImportError:
    JIZEN_AVAILABLE = False
    print("⚠️  evaluate_jizen.py が見つかりません。事前評価はスキップされます。")

try:
    from honmei_scenario import integrate_with_suggest_3rentan
    HONMEI_SCENARIO_AVAILABLE = True
except ImportError:
    HONMEI_SCENARIO_AVAILABLE = False
    print("⚠️  honmei_scenario.py が見つかりません。既存ロジックで動作します。")

# ============================================================
# パス設定
# ============================================================
BASE_DIR    = pathlib.Path(__file__).parent.parent
CSV_DIR     = pathlib.Path(__file__).parent / "csv_output"
EXCEL_FILE  = BASE_DIR / "ボートリサーチ新聞_軽量版.xlsx"
# 会場別コースマスタのCSVキャッシュ（update_master.py が生成）
VENUE_COURSE_CSV = BASE_DIR / "data" / "venue_course_master.csv"
# ─── オッズ設定 ───────────────────────────────────────────────────────────
# 【方法A】boatrace.jp のURLを指定（推奨）
#   レース直前にURLを貼るだけで実際のオッズを自動取得してEV計算
#   例: ACTUAL_ODDS_URL = "https://www.boatrace.jp/owpc/pc/race/odds3t?rno=3&jcd=01&hd=20240601"
#   ※ レース番号(rno)と場コード(jcd)と日付(hd)を書き換えて使用
#   ※ Noneにすると方法Bにフォールバック
ACTUAL_ODDS_URL = None  # ← ここにURLを貼る（レースごとに書き換え）

# 【修正⑤】会場名 → boatrace.jp 場コード（jcd）マッピング
# 公式場コード: 01=桐生 02=戸田 03=江戸川 04=平和島 05=多摩川 06=浜名湖
#              07=蒲郡 08=常滑 09=津  10=三国  11=びわこ 12=住之江
#              13=尼崎 14=鳴門 15=丸亀 16=児島 17=宮島  18=徳山
#              19=下関 20=若松 21=芦屋 22=福岡 23=唐津  24=大村
VENUE_JCD_MAP = {
    "桐生":   "01", "戸田":   "02", "江戸川": "03", "平和島": "04",
    "多摩川": "05", "浜名湖": "06", "蒲郡":   "07", "常滑":   "08",
    "津":     "09", "三国":   "10", "びわこ": "11", "住之江": "12",
    "尼崎":   "13", "鳴門":   "14", "丸亀":   "15", "児島":   "16",
    "宮島":   "17", "徳山":   "18", "下関":   "19", "若松":   "20",
    "芦屋":   "21", "福岡":   "22", "唐津":   "23", "大村":   "24",
}

def build_odds_url(venue: str, race_no, race_date: str) -> str | None:
    """
    【修正⑤】会場名・レース番号・日付から boatrace.jp の3連単オッズURLを自動生成。

    Parameters
    ----------
    venue     : str  会場名（例: "大村"）
    race_no   : int or str  レース番号（例: 3）
    race_date : str  日付文字列（例: "2026-03-08" または "20260308"）

    Returns
    -------
    str  URL文字列 / None（会場コード不明の場合）

    使用例
    ------
    url = build_odds_url("大村", 5, "2026-03-08")
    # → "https://www.boatrace.jp/owpc/pc/race/odds3t?rno=5&jcd=24&hd=20260308"
    """
    jcd = VENUE_JCD_MAP.get(str(venue).strip())
    if not jcd:
        return None
    rno = int(race_no) if str(race_no).isdigit() else race_no
    # 日付フォーマット統一（"2026-03-08" → "20260308"）
    hd = str(race_date).replace("-", "").replace("/", "")
    if len(hd) != 8 or not hd.isdigit():
        return None
    return f"https://www.boatrace.jp/owpc/pc/race/odds3t?rno={rno}&jcd={jcd}&hd={hd}"

# 【方法B】理論オッズExcel（全会場共通・固定ファイル）
#   URLが指定されていない場合のフォールバック
#   odds/ フォルダに「理想オッズ完成版.xlsx」を置くだけでOK
ODDS_DIR      = BASE_DIR / "odds"
ODDS_FILEPATH = ODDS_DIR / "理想オッズ完成版.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

SHEET_MASTER    = "📊コース別マスタ"
SHEET_PLAYER    = "選手指数マスタ"
SHEET_ININAGE   = "イン逃げ分析"
SHEET_OUTPUT    = "出力_新聞"
SHEET_SAMPLE    = "出力_新聞サンプル"

ROWS_PER_RACE = 30  # サンプル29行ブロック + 1空行

# 艇色（競艇公式カラー）
BOAT_COLORS = {
    1: {"bg": "FFFFFFFF", "fg": "FF000000"},  # 白
    2: {"bg": "FF111111", "fg": "FFFFFFFF"},  # 黒
    3: {"bg": "FFFF0000", "fg": "FFFFFFFF"},  # 赤
    4: {"bg": "FF00B0F0", "fg": "FFFFFFFF"},  # 青
    5: {"bg": "FFFFFF00", "fg": "FF000000"},  # 黄
    6: {"bg": "FF00B050", "fg": "FFFFFFFF"},  # 緑
}
BOAT_BG_LIGHT = {
    1: "FFFFFFFF",
    2: "FF555555",
    3: "FFFFCCCC",
    4: "FFCCE8FF",
    5: "FFFFFACC",
    6: "FFCCFFEE",
}

# ============================================================
# ユーティリティ
# ============================================================
def sep(char="=", n=55):
    print(char * n)

def safe_float(val, default=None):
    """文字列・数値を安全にfloatへ変換する共通ユーティリティ。"""
    try:
        v = str(val).replace("%","").strip()
        return float(v) if v not in ("", "None", "nan", "-", "★") else default
    except Exception:
        return default

# ============================================================
# カラム名マッピング定数（load_race ↔ update_master 間で統一）
# update_master.py は「決まり手_逃げ%」形式で書き込む。
# Excelマスタシートのヘッダは「逃げ%」「差し%」…と短縮されているため
# ここで両方を吸収するエイリアスを定義する。
# ============================================================
KIMARI_COL_MAP = {
    "逃げ%":        ["逃げ%",        "決まり手_逃げ%"],
    "差し%":        ["差し%",        "決まり手_差し%"],
    "まくり%":      ["まくり%",      "決まり手_まくり%"],
    "まくり差し%":  ["まくり差し%",  "決まり手_まくり差し%"],
    "抜き%":        ["抜き%",        "決まり手_抜き%"],
    "恵まれ%":      ["恵まれ%",      "決まり手_恵まれ%"],
}

def _get_cm_val(cm: dict, canonical_key: str):
    """カラム名マッピングを使ってコース別マスタから値を取得する。"""
    for alias in KIMARI_COL_MAP.get(canonical_key, [canonical_key]):
        v = cm.get(alias)
        if v is not None:
            return v
    return None

def make_fill(hex_color):
    if not hex_color or hex_color in ("00000000", "FFFFFFFF", ""):
        return PatternFill(fill_type=None)
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, size=9, color="FF000000", name="Noto Sans CJK SC"):
    return Font(name=name, size=size, bold=bold, color=color)

def center_align(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="FFCCCCDD")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill():
    return make_fill("FF404060")

def header_font():
    return make_font(bold=True, color="FFFFFFFF")

def subheader_fill():
    return make_fill("FFCCCCDD")

def subheader_font():
    return make_font(bold=True, color="FF000000")

def write_cell(ws, row, col, value, fill=None, font=None, alignment=None, border=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    if fill:      c.fill = fill
    if font:      c.font = font
    if alignment: c.alignment = alignment
    if border:    c.border = border


# ============================================================
# モーターCSV読み込み（scrape_motor.py の出力）
# ============================================================
def load_motor_csv(venue, race_date_str, race_df=None):
    """
    出走表CSV（load_csvで読み込み済みのDataFrame）から
    モーター番号・2連対率を抽出して返す。

    出走表CSVの列名対応:
        レース列  : 「レース」または「R」
        艇番列    : 「艇番」または「枠」
        モーター番号 : 「M番」
        2連対率   : 「M2率」

    戻り値は build_jizen_members が期待する形式:
        columns = [race_no, boat_no, motor_no, motor_2rate]
    """
    if race_df is None or len(race_df) == 0:
        return None

    # 列名を特定
    race_col = next((c for c in race_df.columns if c in ("レース", "R", "レース番号")), None)
    boat_col = next((c for c in race_df.columns if c in ("艇番", "枠")), None)
    mno_col  = next((c for c in race_df.columns if c in ("M番", "モーター番号", "motor_no")), None)
    m2r_col  = next((c for c in race_df.columns if c in ("M2率", "モーター2連率", "M2連対率", "motor_2rate")), None)

    if mno_col is None or m2r_col is None:
        print(f"  ⚠️  出走表CSVにモーター列（M番/M2率）が見つかりません。機力評価はスキップします。")
        return None

    try:
        rows = []
        for _, row in race_df.iterrows():
            race_no  = pd.to_numeric(row.get(race_col), errors="coerce") if race_col else None
            boat_no  = pd.to_numeric(row.get(boat_col), errors="coerce") if boat_col else None
            motor_no = pd.to_numeric(row.get(mno_col),  errors="coerce")
            m2rate   = pd.to_numeric(row.get(m2r_col),  errors="coerce")
            rows.append({
                "race_no":     race_no,
                "boat_no":     boat_no,
                "motor_no":    motor_no,
                "motor_2rate": m2rate,
            })
        df = pd.DataFrame(rows)
        print(f"  ⚙️  モーターデータを出走表CSVから読み込みました（{len(df)}艇分）")
        return df
    except Exception as e:
        print(f"  ⚠️  モーターデータ抽出エラー: {e}")
        return None

# ============================================================
# マスタ読み込み
# ============================================================

# 【軽微②改善】選手名正規化とマスタルックアップを1か所に集約
# 旧方式: 登録時にエイリアスを追加 → 4文字名が5文字名に誤マッチする可能性
# 新方式: ルックアップ関数で「完全一致 → 先頭N文字一致」の順で検索
def _lookup_name_course(master_dict, name, course_str):
    """
    course_masterから(name, course)でルックアップする。
    完全一致 → 4文字前方一致 → 5文字前方一致 の順で検索。
    """
    key_exact = (name, course_str)
    if key_exact in master_dict:
        return master_dict[key_exact]
    # 4文字前方一致（5文字名→4文字で照合）
    if len(name) >= 5:
        key4 = (name[:4], course_str)
        if key4 in master_dict:
            return master_dict[key4]
    # 4文字キーで登録されているマスタと5文字クエリの照合
    if len(name) == 4:
        # マスタ側が5文字で登録されていてこちらが4文字の場合は対応不要（登録時に済み）
        pass
    return None

def _lookup_player(master_dict, name):
    """player_masterから選手名でルックアップ。完全一致 → 4文字前方一致 の順。"""
    if name in master_dict:
        return master_dict[name]
    if len(name) >= 5:
        alias = name[:4]
        if alias in master_dict:
            return master_dict[alias]
    return None

def load_masters(wb):
    """コース別マスタ・選手指数マスタ・イン逃げ分析を読み込む"""

    # コース別マスタ
    ws_m = wb[SHEET_MASTER]
    master_rows = list(ws_m.iter_rows(values_only=True))
    # ヘッダ行を探す
    header_row = None
    for i, row in enumerate(master_rows):
        if row and row[0] == "選手名":
            header_row = i
            break
    if header_row is None:
        return {}, {}, {}
    headers_m = master_rows[header_row]
    course_master = {}  # {(選手名, コース): {指数dict}}
    for row in master_rows[header_row+1:]:
        if not row or row[0] is None:
            continue
        d = dict(zip(headers_m, row))
        name_full = str(d.get("選手名","")).strip()
        course_str = str(d.get("コース","")).strip()
        key = (name_full, course_str)
        course_master[key] = d
        # 【軽微②改善】5文字名の4文字エイリアスは、4文字キーが未登録の場合のみ追加
        # ※ただし4文字名の選手と衝突しないよう、ルックアップ側で優先度制御する
        if len(name_full) == 5:
            alias_key = (name_full[:4], course_str)
            if alias_key not in course_master:
                course_master[alias_key] = d

    # 選手指数マスタ
    ws_p = wb[SHEET_PLAYER]
    player_rows = list(ws_p.iter_rows(values_only=True))
    header_row_p = None
    for i, row in enumerate(player_rows):
        # A列が「登録番号」、B列が「選手名」の行をヘッダとして検出
        if row and row[1] == "選手名":
            header_row_p = i
            break
    headers_p = player_rows[header_row_p] if header_row_p is not None else []
    player_master = {}  # {選手名: {指数dict}}
    if header_row_p is not None:
        for row in player_rows[header_row_p+1:]:
            if not row or row[0] is None:
                continue
            d = dict(zip(headers_p, row))
            name = str(d.get("選手名","")).strip()
            player_master[name] = d
            # 改善D: 5文字名の先頭4文字エイリアスを登録時点で追加（O(N)フォールバックループを除去）
            if len(name) == 5:
                alias = name[:4]
                if alias not in player_master:
                    player_master[alias] = d

    # イン逃げ分析（1行=1会場、枠別2着率が横展開）
    ws_i = wb[SHEET_ININAGE]
    ininage_rows = list(ws_i.iter_rows(values_only=True))
    header_row_i = None
    for i, row in enumerate(ininage_rows):
        if row and row[0] == "会場名":
            header_row_i = i
            break
    headers_i = ininage_rows[header_row_i] if header_row_i is not None else []
    # ininage_master: {会場名: {"2nd": {"2": 率,...}, "3rd": {"2": 率,...}}} に変換
    # "2nd" = 枠別イン逃げ時2着率（circle_pct算出用）
    # "3rd" = 枠別イン逃げ時3着以内率（idx3独立算出用）
    ininage_master = {}
    if header_row_i is not None:
        for row in ininage_rows[header_row_i+1:]:
            if not row or row[0] is None:
                continue
            d = dict(zip(headers_i, row))
            v = str(d.get("会場名", "")).strip()
            frame_map_2nd = {}
            frame_map_3rd = {}
            for waku_no in range(1, 7):
                val_2nd = d.get(f"{waku_no}枠\n2着率")
                if val_2nd is not None:
                    frame_map_2nd[str(waku_no)] = float(val_2nd)
                val_3rd = d.get(f"{waku_no}枠\n3着以内率")
                if val_3rd is not None:
                    frame_map_3rd[str(waku_no)] = float(val_3rd)
            # 後方互換: ininage_master[v] をそのままdictとして参照していた箇所のため
            # 旧形式（2着率のみの flat dict）も維持しつつ "3rd" キーを追加する
            ininage_master[v] = dict(frame_map_2nd)      # 旧形式互換（2着率フラット）
            ininage_master[v]["_3rd"] = frame_map_3rd    # 3着以内率（枠番→率）

    # 会場統計（イン逃げ率・決まり手場平均）
    # ──────────────────────────────────────────────────────────────────
    # 【キー名正規化】Excelヘッダは改行あり（例: "1C\n1着率"、"1C\n1R"）だが
    # load_race.py 内では "1コース1着率"、"1C_1R1着率" を参照していたため
    # 全て不一致だった。ここで読み込み時に両方のキー形式を登録する。
    # ──────────────────────────────────────────────────────────────────
    venue_stats_master = {}
    if "会場統計" in wb.sheetnames:
        ws_vs = wb["会場統計"]
        vs_rows = list(ws_vs.iter_rows(values_only=True))
        header_vs = None
        for i, row in enumerate(vs_rows):
            if row and row[0] == "会場名":
                header_vs = i
                break
        if header_vs is not None:
            headers_vs = vs_rows[header_vs]
            for row in vs_rows[header_vs+1:]:
                if not row or row[0] is None:
                    continue
                d = dict(zip(headers_vs, row))
                v = str(d.get("会場名", "")).strip()

                # ── キー名エイリアスを追加登録 ──
                normalized = dict(d)
                for h, val in d.items():
                    if h is None:
                        continue
                    h_str = str(h)
                    # "XC\n1着率" → "Xコース1着率" と "XC_1着率"
                    # re はトップレベルで import 済みのため ループ内 import 不要
                    m = re.match(r"^(\d)C\n1着率$", h_str)
                    if m:
                        c = m.group(1)
                        normalized[f"{c}コース1着率"]  = val
                        normalized[f"{c}C_1着率"]      = val
                    # "XC\nYR" → "XC_YR1着率"（R×C別1着率）
                    m2 = re.match(r"^(\d)C\n(\d+)R$", h_str)
                    if m2:
                        c, r_no = m2.group(1), m2.group(2)
                        normalized[f"{c}C_{r_no}R1着率"] = val
                    # "XR\n荒れ" → "XR荒れスコア"
                    m3 = re.match(r"^(\d+)R\n荒れ$", h_str)
                    if m3:
                        r_no = m3.group(1)
                        normalized[f"{r_no}R荒れスコア"] = val
                    # "まくり\n差し率" → "まくり差し率"
                    if h_str == "まくり\n差し率":
                        normalized["まくり差し率"] = val

                venue_stats_master[v] = normalized

    # ── 会場別コースマスタ（選手×会場×コース の3次元実績）──
    # CSVキャッシュが存在すればそちらを優先（openpyxl iter_rowsより5〜10倍速）
    # CSVがなければ従来通りExcelシートから読み込む（後方互換）
    venue_course_master = {}
    if VENUE_COURSE_CSV.exists():
        try:
            df_vc = pd.read_csv(str(VENUE_COURSE_CSV), encoding="utf-8-sig", dtype=str)
            # 列名を load_race.py 側が期待するキー名に正規化
            # update_master の col_keys: 選手名/会場名/進入コース/...
            # load_race.py は "コース" キーでアクセスするので rename
            if "進入コース" in df_vc.columns and "コース" not in df_vc.columns:
                df_vc = df_vc.rename(columns={"進入コース": "コース"})
            for _, row in df_vc.iterrows():
                name   = str(row.get("選手名", "")).strip()
                kaijo  = str(row.get("会場名", "")).strip()
                course = str(row.get("コース",  "")).strip()
                if not name or not kaijo or not course:
                    continue
                d = row.to_dict()
                key = (name, kaijo, course)
                venue_course_master[key] = d
                # 5文字名の4文字エイリアス
                if len(name) == 5:
                    alias_key = (name[:4], kaijo, course)
                    if alias_key not in venue_course_master:
                        venue_course_master[alias_key] = d
            print(f"  📍 会場別コースマスタ読込（CSV）: {len(venue_course_master):,}件")
        except Exception as e:
            print(f"  ⚠️  会場別コースマスタCSV読み込み失敗、Excelにフォールバックします: {e}")
            venue_course_master = {}

    # CSVが存在しないか読み込み失敗 → Excelシートから読む（従来通り）
    if not venue_course_master:
        if "会場別コースマスタ" in wb.sheetnames:
            ws_vc = wb["会場別コースマスタ"]
            vc_rows = list(ws_vc.iter_rows(values_only=True))
            header_vc = None
            for i, row in enumerate(vc_rows):
                if row and row[0] == "選手名":
                    header_vc = i
                    break
            if header_vc is not None:
                headers_vc = vc_rows[header_vc]
                for row in vc_rows[header_vc+1:]:
                    if not row or row[0] is None:
                        continue
                    d = dict(zip(headers_vc, row))
                    name   = str(d.get("選手名", "")).strip()
                    kaijo  = str(d.get("会場名", "")).strip()
                    course = str(d.get("コース", "")).strip()
                    key = (name, kaijo, course)
                    venue_course_master[key] = d
                    if len(name) == 5:
                        alias_key = (name[:4], kaijo, course)
                        if alias_key not in venue_course_master:
                            venue_course_master[alias_key] = d
            print(f"  📍 会場別コースマスタ読込（Excel）: {len(venue_course_master):,}件")
        else:
            print("  ⚠️  会場別コースマスタ未作成。update_master.py を実行してください。")

    return course_master, player_master, ininage_master, venue_stats_master, venue_course_master

# ============================================================
# CSVからレースデータ読み込み
# ============================================================
def load_csv(venue, race_no=None, date_str=None):
    """csv_output/会場名/ サブフォルダ（優先）または scripts/csv_output/ から会場CSVを読み込む"""
    # 会場サブフォルダを優先検索、なければルートにフォールバック
    search_dirs = [CSV_DIR / venue, CSV_DIR]

    def find_files(pattern_name):
        for d in search_dirs:
            files = sorted(glob.glob(str(d / pattern_name)))
            if files:
                return files
        return []

    if date_str:
        files = find_files(f"{venue}_{date_str}.csv")
        if not files:
            print(f"  ❌ 指定日付のCSVが見つかりません: {venue}_{date_str}.csv")
            return None, None
        csv_path = files[0]
        print(f"  📂 日付指定: {csv_path}")
    else:
        files = find_files(f"{venue}_*.csv")
        if not files:
            # 最終手段: csv_output/ 全体から最新を検索
            all_files = sorted(glob.glob(str(CSV_DIR / "**" / "*.csv"), recursive=True))
            if not all_files:
                return None, None
            files = [all_files[-1]]
            venue_detected = os.path.basename(files[0]).split("_")[0]
            print(f"  📂 自動検出: {os.path.basename(files[0])} (会場: {venue_detected})")
        csv_path = files[-1]  # 最新

    date_match = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(csv_path))
    race_date = date_match.group(1) if date_match else datetime.today().strftime("%Y/%m/%d")

    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
    df.columns = df.columns.str.strip()
    for col in df.columns:
        df[col] = df[col].astype(str).str.lstrip("'").str.strip()

    # CSVの「日付」列がある場合はそちらを優先（ファイル名より確実）
    if "日付" in df.columns:
        date_vals = df["日付"].dropna()
        date_vals = date_vals[date_vals.str.strip() != ""]
        if len(date_vals) > 0:
            raw_date = date_vals.iloc[0].strip()
            date_from_csv = re.sub(r"(\d{4})[/\-](\d{2})[/\-](\d{2}).*", r"\1-\2-\3", raw_date)
            if re.match(r"\d{4}-\d{2}-\d{2}", date_from_csv):
                race_date = date_from_csv

    if race_no:
        if "レース番号" in df.columns:
            df = df[df["レース番号"].astype(str) == str(race_no)]
        elif "R" in df.columns:
            df = df[df["R"].astype(str) == str(race_no)]

    return df, race_date

# ============================================================
# 指数計算
# ============================================================
def calc_race_indices(venue, race_no, players, course_master, player_master, ininage_master, venue_stats_master, venue_course_master=None):
    """
    1レース分の指数を計算して返す。

    改善①  会場別コースマスタ (venue_course_master) を優先参照。
            選手の当該会場での実績があればそちらを使い、なければ全国マスタにフォールバック。
    改善③  ハイブリッド係数を動的計算。
            会場の実績量（会場統計レース数）と選手の当該会場実績量（信頼度）に応じて
            「選手実績 : 会場特性」の比率を動的に調整する。
    """
    results = []
    if venue_course_master is None:
        venue_course_master = {}

    # 【問題C修正】_blend_tsをループ外に定義（毎ループの再生成を防ぐ）
    def _blend_ts(ts_val, flat_val, eff_n, full_n=20.0):
        """時系列補正値とフラット平均を有効走数に応じてブレンドして返す。
        eff_n ≥ full_n → 時系列補正100%採用
        eff_n = 0      → フラット平均100%採用"""
        if ts_val is None and flat_val is None:
            return None
        if ts_val is None:
            return flat_val
        if flat_val is None:
            return ts_val
        t = min(eff_n / full_n, 1.0)
        return ts_val * t + flat_val * (1.0 - t)

    for p in players:
        name_raw = str(p.get("選手名","")).strip()
        # 末尾の年齢数字を除去（例: "熊本英一47" → "熊本英一"、"芦澤　望48" → "芦澤　望"）
        name_raw = re.sub(r'\s*\d+\s*$', '', name_raw).strip()
        name = name_raw.replace("　", "").replace(" ", "")  # スペース除去でマスタキーと統一
        # [修正2] lzh_to_csv.py の出力列名は「艇番」。「枠」にも後方互換対応
        waku = str(p.get("枠", p.get("艇番", ""))).strip()
        course = str(p.get("想定コース", waku)).strip()

        # ── 改善①: 会場別コースマスタを優先参照 ────────────────────────────
        # 【軽微②改善】ルックアップ関数経由で完全一致→4文字前方一致の順で検索
        # 優先度: 会場別コースマスタ（選手×会場×コース） > 全国コース別マスタ
        vc_key = (name, venue, course)
        vcm = venue_course_master.get(vc_key)
        if vcm is None and len(name) >= 5:
            vcm = venue_course_master.get((name[:4], venue, course))

        cm = _lookup_name_course(course_master, name, course)
        cm = cm or {}
        pm = _lookup_player(player_master, name)
        pm = pm or {}

        # ── 【修正①②④】会場別コースマスタと全国マスタの統合 ────────────────────
        # 【修正①】vcm存在かつvc_trust=0.0 と vcm=None を明示的に区別する。
        # 【修正②】win1_rate は純粋な「選手個人の1着率推定値」のみ。venue_rateは後段で混ぜる。
        # 【修正④】時系列有効走数（eff_n）による時系列補正の信頼制御
        #   eff_n ≥ 20走 → 時系列補正100%採用 / eff_n=0 → フラット平均100%
        #   ※ _blend_ts はループ外で定義済み

        # 全国マスタの有効走数を取得
        cm_eff_n = safe_float(cm.get("時系列有効走数"), 0.0) or 0.0
        _ts  = safe_float(cm.get("時系列補正1着率"))
        _fl  = safe_float(cm.get("1着率"))
        global_win1 = _blend_ts(_ts, _fl, cm_eff_n)
        _ts  = safe_float(cm.get("時系列補正3連対率"))
        _fl  = safe_float(cm.get("3連対率"))
        global_win3 = _blend_ts(_ts, _fl, cm_eff_n)

        if vcm is not None:
            vc_trust_raw = safe_float(vcm.get("信頼度"), 0.0) or 0.0  # 0.0〜1.0
            if vc_trust_raw > 0.0:
                # 会場別実績あり＆信頼度あり → vc_trustに応じて会場別と全国をブレンド
                vc_trust = vc_trust_raw
                # 会場別コースマスタの有効走数で時系列補正を制御
                vc_eff_n = safe_float(vcm.get("時系列有効走数"), 0.0) or 0.0
                _ts  = safe_float(vcm.get("時系列補正1着率"))
                _fl  = safe_float(vcm.get("1着率"))
                vc_win1 = _blend_ts(_ts, _fl, vc_eff_n)
                _ts  = safe_float(vcm.get("時系列補正3連対率"))
                _fl  = safe_float(vcm.get("3連対率"))
                vc_win3 = _blend_ts(_ts, _fl, vc_eff_n)
                if vc_win1 is not None and global_win1 is not None:
                    win1_rate = vc_win1 * vc_trust + global_win1 * (1.0 - vc_trust)
                elif vc_win1 is not None:
                    win1_rate = vc_win1
                else:
                    win1_rate = global_win1
                if vc_win3 is not None and global_win3 is not None:
                    win3_rate = vc_win3 * vc_trust + global_win3 * (1.0 - vc_trust)
                elif vc_win3 is not None:
                    win3_rate = vc_win3
                else:
                    win3_rate = global_win3
            else:
                # vcm存在するが信頼度=0 → 実質データ不足。全国マスタのみ使用
                vc_trust  = 0.0
                win1_rate = global_win1
                win3_rate = global_win3
        else:
            # vcm=None → 会場別実績なし。全国マスタのみ使用
            vc_trust  = 0.0
            win1_rate = global_win1
            win3_rate = global_win3

        gen_win3 = safe_float(pm.get("3連対率\n(一般戦)"))
        avg_st   = safe_float(cm.get("コース別\n平均ST"))
        st_rank  = safe_float(pm.get("ST順位\n(1コース)"))

        # 決まり手（会場別があればそちらを優先）
        kimari = _build_kimari(vcm if vcm else cm)

        # ── データ不足チェック ──────────────────────────────────────────────
        cm_missing = not cm
        cm_count   = safe_float(cm.get("出走数"), 0)
        cm_scarce  = (not cm_missing) and (cm_count is not None) and (cm_count < 5)
        pm_missing = not pm
        pm_count   = safe_float(pm.get("総出走数"), 0)
        pm_scarce  = (not pm_missing) and (pm_count is not None) and (pm_count < 10)
        data_missing = cm_missing or cm_scarce or pm_missing or pm_scarce
        missing_reasons = []
        if cm_missing:
            missing_reasons.append(f"コース{course}実績なし")
        elif cm_scarce:
            missing_reasons.append(f"コース{course}実績{int(cm_count)}走")
        if pm_missing:
            missing_reasons.append("選手マスタ未登録")
        elif pm_scarce:
            missing_reasons.append(f"総実績{int(pm_count)}走")
        missing_reason_str = " / ".join(missing_reasons) if missing_reasons else ""

        # FLY数・出遅れ数・FLY経過日数（選手指数マスタ pm から取得）
        # 【修正④】FLY経過日数を使って影響度を精密判定
        # 旧: FLY数≥2→高、≥1→中 のみ → FLY明け直後(60日)と1年後(365日)が同判定
        # 新: FLY経過日数を加味:
        #   経過日数 < 90  日: 出場停止明け直後 → 判定を1段階引き上げ（中→高等）
        #   経過日数 < 180 日: 影響残存期間    → 素直に使用
        #   経過日数 ≥ 180 日: 影響ほぼ消滅    → FLY数に依らず「低」
        _fly_count    = int(safe_float(pm.get("FLY数"),    0) or 0)
        _late_count   = int(safe_float(pm.get("出遅れ数"), 0) or 0)
        _fly_days_raw = pm.get("FLY経過日数")
        _fly_days     = safe_float(_fly_days_raw) if _fly_days_raw not in (None, "", "nan") else None

        if _fly_count == 0:
            _fly_label = "低"
        elif _fly_days is not None:
            if _fly_days >= 180:
                # FLY後180日超 → 影響ほぼ消滅
                _fly_label = "低"
            elif _fly_days < 90:
                # 出場停止明け直後 → 1段階引き上げ
                _fly_label = "高"  # (FLY1回でも高)
            else:
                # 90〜180日: 通常判定
                _fly_label = "高" if _fly_count >= 2 else "中"
        else:
            # 経過日数不明（update_master.py 更新前の旧データ）→ 旧来判定にフォールバック
            _fly_label = "高" if _fly_count >= 2 else "中"

        results.append({
            "waku":       waku,
            "name":       name_raw,  # 表示用（スペースあり）
            "name_norm":  name,      # マスタ検索用（スペースなし）
            "kumi":       str(p.get("級別", p.get("組",""))).strip(),
            "motor2":     str(p.get("モーター2連率", p.get("M2率", p.get("モータ2連","")))).strip(),
            "course":     course,
            "win1_rate":  win1_rate,
            "win3_rate":  win3_rate,
            "avg_st":     avg_st,
            "st_rank":    st_rank,
            "kimari":     kimari,
            "kosetsu":    str(p.get("今節成績","")).strip(),
            "tenji_time": safe_float(p.get("展示タイム", p.get("展示", p.get("展示ST","")))),
            "raw_cm":     cm,
            "raw_pm":     pm,
            "raw_vcm":    vcm or {},     # 会場別コースマスタ（デバッグ用）
            "vc_trust":   vc_trust,      # 会場別実績の信頼度（0〜1）
            "data_missing":   data_missing,
            "missing_reason": missing_reason_str,
            "fly_count":  _fly_count,   # FLY数（0以上の整数）
            "fly_label":  _fly_label,   # F/ST影響ラベル（高/中/低）
            "late_count": _late_count,  # 出遅れ数
        })

    # ── 【修正②⑦】動的ハイブリッド係数: 「選手実績 vs 会場特性」のブレンド ──────────
    # win1_rate は上段で「会場別マスタ×全国マスタ」をブレンド済みの純粋な選手個人実績値。
    # 「その選手実績値をどれだけ信頼するか」を cm_count と vc_trust の両方で決定する。
    #
    # 【修正⑦】vc_trust を後段にも引き継いで w_player 上限を拡張する。
    #   修正②で「上段ブレンドに vc_trust を消費」したが、それは
    #   「会場別 vs 全国の個人実績ブレンド」への使用であり、
    #   「個人実績全体 vs 会場特性」のブレンド比率を決める後段とは別の軸。
    #   → 二重混合ではなく「直交した2つの調整」なので後段でも参照してよい。
    #
    # w_player の計算式（修正⑦版）:
    #   cm_trust = min(cm_count / 30, 1.0) × 0.60   全国出走数による基礎信頼度（最大0.60）
    #   vc_bonus = vc_trust × 0.30                   会場別実績が十分なら最大+0.30
    #   w_player = min(cm_trust + vc_bonus, 0.90)    上限0.90（会場特性を最低10%保証）
    #
    # 具体例:
    #   会場別実績なし(vc=0)  全国30走 → 0.60+0.00 = 0.60
    #   会場別実績あり(vc=1)  全国30走 → 0.60+0.30 = 0.90（上限）
    #   会場別実績あり(vc=0.5) 全国10走 → 0.20+0.15 = 0.35
    #   全実績なし(vc=0, cm=0)         → 0.00+0.00 = 0.00 → 会場特性100%
    #
    # ※ キー名正規化は load_masters() 内で実施済み（"1コース1着率"等のエイリアスを追加）

    COURSE_AVG_WIN = {
        "1": 0.555, "2": 0.137, "3": 0.134,
        "4": 0.111, "5": 0.066, "6": 0.021,
    }

    vs = venue_stats_master.get(venue, {})

    # ── 【修正⑤】会場のコース別1着率: R番号別と全体平均を加重平均でブレンド ────────
    # 【旧方式の問題】R番号別1着率が存在すれば無条件に優先していた。
    #   R別はサンプルが少ない（例: 同会場の同R番号は年間50〜200レース程度）ため
    #   たまたまの偏りをシグナルと誤認する「過学習的なノイズ混入」が起きていた。
    #
    # 【修正⑥】R別1着率とコース全体平均のブレンド比率を動的化。
    # 旧方式: W_RC = 0.30 固定 → サンプルが少ない会場でもRC別を30%参照してノイズが混入。
    # 新方式: 会場の総レース数（= 統計の信頼性）に応じてW_RCを動的決定。
    #   計算式: W_RC = clip(レース数 / 3000, 0.05, 0.30)
    #     レース数3000以上（約12年分）→ W_RC = 0.30（上限：RC別を最大30%参照）
    #     レース数1500程度（約6年）  → W_RC = 0.15
    #     レース数300以下（約1年）   → W_RC = 0.05（ほぼ全体平均のみ）
    #   根拠: R別は年間約120〜150レースのデータ。
    #   統計的に安定するには最低5〜8年(600〜1200件)が目安。
    #   上限を0.30に抑えてRC別に過剰依存しないよう設計。
    _venue_race_count = safe_float(vs.get("レース数"), 0) or 0
    W_RC = float(min(max(_venue_race_count / 3000.0, 0.05), 0.30))

    venue_course_rate = {}
    for c in range(1, 7):
        rc_key     = f"{c}C_{race_no}R1着率"   # キー名正規化済み
        course_key = f"{c}コース1着率"          # キー名正規化済み
        nat_avg    = COURSE_AVG_WIN[str(c)]
        rc_val     = safe_float(vs.get(rc_key))
        course_val = safe_float(vs.get(course_key))
        if rc_val is not None and course_val is not None:
            # 両方存在 → 加重平均（R別30%、全体70%）
            blended = rc_val * W_RC + course_val * (1.0 - W_RC)
        elif course_val is not None:
            blended = course_val
        elif rc_val is not None:
            blended = rc_val
        else:
            blended = nat_avg
        venue_course_rate[str(c)] = blended

    for r in results:
        course     = str(r["course"])
        venue_rate = venue_course_rate.get(course, COURSE_AVG_WIN.get(course, 0.10))

        # 【修正⑦】全国出走数 + 会場別実績信頼度の両方でw_playerを決定（上限0.90）
        cm_count_w = safe_float(r.get("raw_cm", {}).get("出走数"), 0) or 0
        vc_trust   = r.get("vc_trust", 0.0)
        cm_trust   = min(cm_count_w / 30.0, 1.0) * 0.60   # 全国出走数による基礎信頼度（最大0.60）
        vc_bonus   = vc_trust * 0.30                        # 会場別実績ボーナス（最大+0.30）
        w_player   = min(cm_trust + vc_bonus, 0.90)        # 上限0.90（会場特性を最低10%保証）
        w_venue    = 1.0 - w_player                         # 0.10〜1.0

        # ── 【新修正: 全国平均比スケール換算（vc_trust対応版）】────────────────────────
        # 問題: win1_rate のスケールが vcm の有無によって変わる。
        #   - vcm=None:  win1_rate は全国コース別実績（全国スケール）
        #   - vcm有り:   win1_rate = vc_win1 * vc_trust + global_win1 * (1-vc_trust)
        #               vc_win1 は「当該会場での絶対的1着率（会場スケール）」
        #               → win1_rate は会場スケールと全国スケールの混合
        #
        # 解決: win1_rate のスケールに合わせた基準値 base でratio計算する。
        #   base = venue_rate * vc_trust + nat_avg_c * (1 - vc_trust)
        #     vc_trust=0.0 → base = nat_avg_c（全国比換算のみ）
        #     vc_trust=1.0 → base = venue_rate → ratio=win1/venue_rate
        #                    win1_scaled = venue_rate * ratio = win1（スケール変換なし）
        #
        # 効果:
        #   vcm無し・荒れ会場: 全国比換算で過大評価を解消（戸田1号艇 54.8%→50.9%）
        #   vcm有り・荒れ会場: 会場スケールの実績をそのまま反映（過小評価なし）
        #   全国平均会場: 変化なし
        #   ratio は [0.25, 4.0] にクリップして外れ値を防ぐ。
        _RATIO_MIN, _RATIO_MAX = 0.25, 4.0
        nat_avg_c = COURSE_AVG_WIN.get(course, 0.10)  # コースの全国平均1着率
        # win1_rate のスケールに合わせた基準値（vc_trust で全国↔会場を補間）
        _ratio_base = venue_rate * vc_trust + nat_avg_c * (1.0 - vc_trust)
        _ratio_base = max(_ratio_base, 1e-6)  # ゼロ除算防止

        if r["win1_rate"] is not None and r["win1_rate"] > 0:
            ratio = max(_RATIO_MIN, min(r["win1_rate"] / _ratio_base, _RATIO_MAX))
            win1_scaled = venue_rate * ratio  # 会場スケールでの個人実績推定値
            r["_raw_win"] = win1_scaled * w_player + venue_rate * w_venue
        elif r["win1_rate"] == 0.0:
            # 【問題B修正】出走したが1着なし → 会場特性をw_venue分だけ使用。
            r["_raw_win"] = venue_rate * w_venue
        else:
            # データなし → 会場特性100%
            r["_raw_win"] = venue_rate

    # ── 【修正④】Laplace smoothingフロア: 均等フロアで正規化歪みを防ぐ ────────────
    # 【旧方式の問題】フロア値を「全国平均×10%」にしていたため
    #   6コースのフロア ≈ 0.0021（0.2%）と極端に小さく、
    #   この値が正規化の分母に混入することで1〜5コースの確率が不当に圧縮されていた。
    #
    # 【新方式】全コース共通で「全国平均の最小値（6コース=2.1%）の50%」= 0.0105 を下限とする。
    #   これにより「確率0の艇を救済する」目的は維持しつつ、
    #   極小フロアによる正規化への影響を実質ゼロに近づける。
    COURSE_WIN_FLOOR_UNIFORM = 0.021 * 0.50   # ≈ 0.0105（全コース共通）
    for r in results:
        r["_raw_win"] = max(r["_raw_win"], COURSE_WIN_FLOOR_UNIFORM)

    # ── 【修正③⑥・問題A修正】キャリブレーション補正（_raw_winスケール対応版）──────────
    # 【問題A】修正③で「_raw_winに補正を適用」に変えたが、
    #   breakpointsの閾値が旧方式（rel_win1 = 0〜1スケール）のままだった。
    #   _raw_win の実際の範囲は 1号艇:0.02〜0.60、外コース:0.01〜0.10 であり、
    #   旧閾値 p=0.35 未満は補正なし → 2〜6号艇が全て補正なしになっていた。
    #
    # 【修正】breakpointsを _raw_win の実際のスケールに合わせて再設計する。
    #   目標: バックテストCal誤差0.0698を反映しつつ全艇に段階的に補正を掛ける。
    #   設計方針:
    #     ・p < 0.10 → 補正なし（6コース等の極小確率帯は信頼できる）
    #     ・p = 0.15 → scale = 0.990（2%の緩い補正）
    #     ・p = 0.25 → scale = 0.975
    #     ・p = 0.35 → scale = 0.960
    #     ・p = 0.45 → scale = 0.945
    #     ・p = 0.55 → scale = 0.930（1号艇強いケースの核心補正 ≒ Cal誤差0.0698相当）
    #     ・p ≥ 0.65 → scale = 0.920（上限なし: _raw_winは最大0.60程度のため実質到達しない）
    #
    # 【修正⑥連携】荒れ会場では緩和スケール(relax)を適用:
    #   relax = clip(venue_c1_rate / 0.555, 0.70, 1.0)
    #   戸田(c1≈0.430) → relax=0.775 → 補正を22.5%緩和

    _venue_c1_rate = venue_course_rate.get("1", 0.555)
    _calib_relax   = float(min(max(_venue_c1_rate / 0.555, 0.70), 1.0))

    def _calibrate(p_raw, relax=1.0):
        """_raw_winスケール(0〜1)にキャリブレーション補正を適用して返す。
        relax: 1.0=通常補正、<1.0=補正を緩和（荒れ会場向け）"""
        if p_raw is None:
            return None
        p = float(p_raw)
        # _raw_winスケール向けbreakpoints（旧の0.35〜1.01から0.10〜1.01に再設計）
        breakpoints = [
            (0.10, 1.000),
            (0.15, 0.990),
            (0.25, 0.975),
            (0.35, 0.960),
            (0.45, 0.945),
            (0.55, 0.930),
            (0.65, 0.920),
            (1.01, 0.920),
        ]
        if p < breakpoints[0][0]:
            return p  # 極小確率帯は補正なし
        for i in range(len(breakpoints) - 1):
            x0, s0 = breakpoints[i]
            x1, s1 = breakpoints[i + 1]
            if x0 <= p < x1:
                t = (p - x0) / (x1 - x0)
                scale_full = s0 + t * (s1 - s0)
                scale = 1.0 - (1.0 - scale_full) * relax
                return p * scale
        scale_full = breakpoints[-1][1]
        scale = 1.0 - (1.0 - scale_full) * relax
        return p * scale

    # _raw_win は 0〜1 程度のスケール。_calibrate に relax（荒れ会場緩和係数）を渡す。
    for r in results:
        r["_raw_win"] = _calibrate(r["_raw_win"], relax=_calib_relax) or r["_raw_win"]

    total_raw = sum(r["_raw_win"] for r in results)
    for r in results:
        if total_raw > 0:
            r["rel_win1"] = r["_raw_win"] / total_raw * 100
        else:
            r["rel_win1"] = None

    # rel_win1_cal は rel_win1 と同値（補正済み値の別名として下流コードと互換性を保つ）
    for r in results:
        r["rel_win1_cal"] = r["rel_win1"]
    
    # 3連対率：絶対評価（コース別実績をそのまま%表示）
    # 相対化するとメンバーレベルが見えなくなるため絶対値を使用
    for r in results:
        if r["win3_rate"] is not None:
            r["abs_win3"] = round(r["win3_rate"] * 100, 1)  # 例: 0.625 → 62.5%
        else:
            r["abs_win3"] = None
    
    # ══════════════════════════════════════════════════════════════════════════
    # 本命記号（多面評価スコアによる総合印）◎→○→▲→△
    # ──────────────────────────────────────────────────────────────────────────
    # 【旧方式の問題】rel_win1（オリジナル1着率）のみで機械的に順位付け
    #   → FLYリスクが高い艇でも◎がつく
    #   → 飛びシナリオ高確率なのに1号艇◎という矛盾が起きる
    #
    # 【新方式: 総合印スコア】
    #   基礎点  = rel_win1 × 0.60
    #   FLY高  : -8pt  / FLY中: -4pt
    #   イン逃げ◎(1号艇): +8pt / ○: +4pt / 空白: -5pt  ← jizen確定後に適用
    #   飛び相性◎(tobi_prob≥55の2〜6号艇): +10pt / ○: +5pt ← jizen確定後に適用
    #
    # 【印の意味】◎>○>▲>△
    #   ◎: 総合スコア1位（本命）  ○: 2位（対抗）
    #   ▲: 3位（単穴）           △: 4位（穴）
    #
    # ※ここではFLYペナルティのみ適用した仮印を付ける。
    #   jizen_eval確定後に main() で _apply_jizen_honmei() を呼んで最終確定する。
    # ══════════════════════════════════════════════════════════════════════════

    # _calc_honmei_score / _apply_jizen_honmei はトップレベルで定義（下記参照）

    # 仮印（jizen未確定。相互作用モデルで計算し、jizen確定後に _apply_jizen_honmei で上書き）
    # venue_stats をここで構築して相互作用モデルに渡す
    _venue_stats_pre = _calc_venue_stats(venue_stats_master, venue)
    _honmei_scores_pre = [
        (i, _calc_honmei_score(r, 0, jizen_ev=None, results_ctx=results,
                               venue_stats=_venue_stats_pre,
                               race_judgment=None))   # 仮印計算時点ではrace_judgment未確定
        for i, r in enumerate(results)
        if r["rel_win1"] is not None
    ]
    _honmei_scores_pre.sort(key=lambda x: x[1], reverse=True)

    honmei_map = {0: "◎", 1: "○", 2: "▲", 3: "△"}  # ◎>○>▲>△
    for rank, (idx, _) in enumerate(_honmei_scores_pre[:4]):
        results[idx]["honmei"] = honmei_map[rank]
    for r in results:
        if "honmei" not in r:
            r["honmei"] = " "

    # _apply_jizen_honmei はトップレベルで定義（下記参照）

    # ── 新機能④：展示タイム偏差値（レース内相対評価） ──────────────────────
    # 【軽微①注意】展示タイムは当日展示航走後にしか取得できない。
    # 前日予想CSV（締め切り前スクレイプ）では値が存在しないため tenji_hensa は None になる。
    # 表示側では None の場合「-（前日）」と表示し、当日版で上書きされることを明示する。
    tenji_vals = [r["tenji_time"] for r in results if r.get("tenji_time") is not None]
    _has_tenji_data = len(tenji_vals) >= 2  # 展示タイムデータが揃っているか
    if _has_tenji_data:
        t_mean = statistics.mean(tenji_vals)
        t_stdev = statistics.stdev(tenji_vals) if len(tenji_vals) > 1 else 1
        for r in results:
            t = r.get("tenji_time")
            if t is not None and t_stdev > 0:
                # 競艇の展示タイムは速い（小さい）ほど良いので逆転
                r["tenji_hensa"] = round(50 - (t - t_mean) / t_stdev * 10, 1)
            else:
                r["tenji_hensa"] = None
    else:
        # 前日時点では展示タイム未取得 → 明示的にNoneを設定
        for r in results:
            r["tenji_hensa"] = None
        if not tenji_vals:
            pass  # 前日出力では正常（展示前）

    # 想定スリット（平均STでソート）
    sortable = [(r["waku"], r["avg_st"]) for r in results if r["avg_st"] is not None]
    sortable.sort(key=lambda x: x[1])
    slit = "-".join([s[0] for s in sortable])
    if not slit:
        slit = "-".join([r["waku"] for r in results])
    
    # ══════════════════════════════════════════════════════════════════════════
    # circle_pct（2着優位度）・idx3（3着指数）
    # ──────────────────────────────────────────────────────────────────────────
    # 【設計思想】
    #   「この会場でこの出走メンバーが戦ったとき何が起きるか」を計算する。
    #   枠番の固定統計ではなく、各艇の決まり手・ST・1号艇との関係性から
    #   相互作用モデルでスコアを算出する。
    #
    # ──────────────────────────────────────────────────────────────────────────
    # ■ circle_pct（2着優位度）：「イン逃げ時に2着を取れる攻め手の強さ」
    #
    #   コース別に異なる攻め手を評価し、1号艇との具体的な関係性で補正する。
    #
    #   2枠: 差し特化コース
    #     差し% × ST優位(vs 1号艇) × 1号艇の被差し脆弱性
    #
    #   3枠: まくり差し主体
    #     (まくり差し% × 1.2 + まくり%) × ST優位
    #     × 2枠差し力ペナルティ（2枠が強いと進路を塞がれる）
    #
    #   4枠: まくり主体
    #     (まくり% + まくり差し%) × ST優位
    #     × 3枠壁ペナルティ（3枠のまくり力が強いと被る）
    #
    #   5・6枠: 展開待ち・まくり
    #     (まくり% + まくり差し%) × ST優位 × 外枠減衰
    #
    #   ※ 全艇の絶対スコアを算出し、合計100%に正規化して circle_pct とする。
    #   ※ 選手実績データが不足する場合は全国コース別平均でフォールバック。
    #
    # ──────────────────────────────────────────────────────────────────────────
    # ■ idx3（3着指数）：「イン逃げ時に3着に残る固有の力」
    #
    #   純3着率（3着以内率 - 2着率）を主軸に据える。
    #   「2着を取りこぼしても3着に粘り込む能力」を選手個人の実績から評価する。
    #
    #   ベーススコア = 純3着率(選手実績) × trust + 会場枠別純3着率 × (1-trust)
    #   → 選手実績が豊富なほど個人差が出る。データ不足なら会場値に寄せる。
    #   → 最大値=100にスケーリング（ただし枠番固定にならないよう分散を確保）
    # ══════════════════════════════════════════════════════════════════════════

    venue_frame   = ininage_master.get(venue, {})            # {枠番str: 会場2着率float}
    venue_3rd_map = ininage_master.get(venue, {}).get("_3rd", {})  # {枠番str: 会場3着以内率float}
    MIN_ININAGE_COUNT = 5

    # ── 全国コース別平均（フォールバック用） ──────────────────────────────────
    NATIONAL_2ND  = {"2": 0.35, "3": 0.33, "4": 0.18, "5": 0.09, "6": 0.05}
    NATIONAL_3RD_PURE = {"2": 0.08, "3": 0.14, "4": 0.16, "5": 0.13, "6": 0.09}
    # 純3着率の全国平均（3着以内 - 2着）:
    #   2枠は差しで2着か沈むか → 純3着が少ない
    #   3〜5枠は流れ込みが多い → 純3着が比較的多い

    # ── 1号艇の情報を事前取得 ────────────────────────────────────────────────
    res1   = next((r for r in results if r["waku"] == "1"), None)
    cm1    = res1.get("raw_cm", {}) if res1 else {}
    st1    = res1.get("avg_st") if res1 else None

    # 1号艇の被差し脆弱性（0〜1、高いほど差されやすい）
    sasar_vuln = safe_float(cm1.get("差され%"), 0) or 0.0    # 被差し%
    makur_vuln = safe_float(cm1.get("捲られ%"), 0) or 0.0   # 被まくり%
    nige_pct1  = safe_float(_get_cm_val(cm1, "逃げ%"), 0.6) or 0.6
    # 被差し脆弱性スコア: 被差し%を主軸にし、逃げ%が低いほど補強
    vuln_sashi  = max(0.0, min(1.0, sasar_vuln * 0.7 + (1.0 - nige_pct1) * 0.3))
    vuln_makuri = max(0.0, min(1.0, makur_vuln * 0.7 + (1.0 - nige_pct1) * 0.3))

    # ── 全艇のST平均（ST優位スコア計算の基準） ───────────────────────────────
    st_vals = [r["avg_st"] for r in results if r.get("avg_st") is not None]
    st_mean = sum(st_vals) / len(st_vals) if st_vals else 0.15

    def _st_advantage(avg_st, reference=None):
        """
        avg_st が reference（デフォルト: メンバー平均ST）より速い（小さい）ほど高い。
        差±0.05秒を基準に [-1, +1] → スコア [0.5〜1.5] に変換。
        """
        ref = reference if reference is not None else st_mean
        if avg_st is None:
            return 1.0   # データなし → 中立
        diff = ref - avg_st   # 正 = 自分が速い
        return max(0.5, min(1.5, 1.0 + diff / 0.05))

    def _st_advantage_vs1(avg_st):
        """1号艇とのST比較。1号艇より速いほど2着争いで有利。"""
        return _st_advantage(avg_st, reference=st1)

    # ══════════════════════════════════════════════════════════════════════════
    # circle_pct（イン逃げ時2着優位度）
    # ──────────────────────────────────────────────────────────────────────────
    # 【修正版設計思想】
    #   主軸: コース別マスタの「イン逃げ時2着率」（直接実績）
    #   補正: ST優位（速い艇は差しやすい）× 1号艇脆弱性（差されやすいほど2枠有利）
    #         決まり手%は補正因子として残す（主軸ではなく傾向補正）
    #
    #   ブレンド方式:
    #     player_2nd   : コース別マスタの「2着率」（イン逃げ時2着率）
    #     venue_2nd    : イン逃げ分析シートの「枠番別2着率」（会場ベースライン）
    #     national_2nd : 全国平均（フォールバック）
    #
    #     trust = min(ininage_count / 30, 1.0)
    #     base  = player_2nd * (0.5 + 0.4 * trust) + venue_2nd * (0.5 - 0.4 * trust)
    #           → 実績30走以上: player 90% / venue 10%
    #           → 実績0走(trust=0): player 50% / venue 50%（データ不足でも完全棄却しない）
    #
    #   ST補正（対1号艇）:
    #     速い艇（ST<1号艇）→ 差しやすい → 2着スコアUP（最大×1.3）
    #     遅い艇             → ×0.85程度
    #
    #   脆弱性補正（2・3枠のみ）:
    #     1号艇が差されやすい → 2枠の差しスコアをさらに補強
    #     1号艇が捲られやすい → 3枠のまくり差しスコアを補強
    #
    #   コース変数は「想定コース（waku フォールバック）」を使用。
    #   会場ベースラインは「枠番」ベースのイン逃げ分析シートを参照（シート定義に準拠）。
    # ══════════════════════════════════════════════════════════════════════════

    raw_scores = {}
    for r in results:
        w = r["waku"]
        if w == "1":
            raw_scores[w] = None
            continue

        cm    = r.get("raw_cm", {})
        avg_st = r.get("avg_st")
        try:
            wno = int(w)
        except (ValueError, TypeError):
            wno = 3

        # ── イン逃げ時2着率（直接実績）────────────────────────────────────
        ininage_count = safe_float(cm.get("イン逃げ\n出走数"), 0) or 0
        player_2nd    = safe_float(cm.get("2着率"))   # コース別マスタのイン逃げ時2着率
        has_data      = ininage_count >= MIN_ININAGE_COUNT

        # 会場ベースライン（枠番別）
        venue_2nd    = venue_frame.get(w)            # イン逃げ分析シート（枠番ベース）
        national_2nd = NATIONAL_2ND.get(w, 0.10)    # 全国平均フォールバック

        # trust加重ブレンド
        trust = min(ininage_count / 30.0, 1.0) if has_data else 0.0
        w_player = 0.5 + 0.4 * trust   # 0.50（trust=0）〜 0.90（trust=1）
        w_venue  = 1.0 - w_player       # 0.50 〜 0.10

        baseline = venue_2nd if venue_2nd is not None else national_2nd
        if player_2nd is not None and has_data:
            base_2nd = player_2nd * w_player + baseline * w_venue
        else:
            base_2nd = baseline   # データ不足 → 会場/全国ベースライン

        # ── ST優位補正（対1号艇）──────────────────────────────────────────
        # 速い艇は1号艇より先にスリットを切れる → 2着争いで有利（最大×1.3）
        st_adv = _st_advantage_vs1(avg_st)
        # ST補正を穏やかに適用（主軸が2着率なので過剰にかけない）
        st_factor = 0.85 + (st_adv - 0.5) * 0.3   # 0.85〜1.30 の範囲

        # ── コース別・脆弱性補正（細かい傾向調整）──────────────────────────
        if wno == 2:
            # 2枠差し: 1号艇が差されやすいほど有利
            vuln_factor = 1.0 + vuln_sashi * 0.3    # 最大+30%
        elif wno == 3:
            # 3枠まくり差し: 1号艇が捲られやすいほど有利
            vuln_factor = 1.0 + vuln_makuri * 0.25  # 最大+25%
        elif wno == 4:
            # 4枠まくり: 3枠選手のまくり力が強いと外枠に押し出されペナルティ
            r3      = next((x for x in results if x["waku"] == "3"), None)
            cm3     = r3.get("raw_cm", {}) if r3 else {}
            mk3_pct = (safe_float(_get_cm_val(cm3, "まくり%"), 0) or 0.0) + \
                      (safe_float(_get_cm_val(cm3, "まくり差し%"), 0) or 0.0)
            st3     = r3.get("avg_st") if r3 else None
            wall3   = mk3_pct * _st_advantage_vs1(st3) * 0.30
            vuln_factor = max(0.7, 1.0 - wall3)
        elif wno == 5:
            vuln_factor = 0.80   # 外枠減衰
        else:  # 6枠
            vuln_factor = 0.60   # さらに減衰

        score = base_2nd * st_factor * vuln_factor
        raw_scores[w] = max(score, 0.001)   # ゼロ除算防止

    # レース内正規化（合計100%）
    valid_scores = {w: s for w, s in raw_scores.items() if w != "1" and s is not None}
    total_score  = sum(valid_scores.values()) or 1.0

    for r in results:
        w = r["waku"]
        s = raw_scores.get(w)
        if w != "1" and s is not None:
            r["circle_pct"] = round(s / total_score * 100, 1)
            r["_circ_raw"]  = s          # 確率計算用（正規化前絶対スコア）
        else:
            r["circle_pct"] = None
            r["_circ_raw"]  = None

    # ══════════════════════════════════════════════════════════════════════════
    # idx3（イン逃げ時3着残存指数）
    # ──────────────────────────────────────────────────────────────────────────
    # 【修正版設計思想】
    #   主軸: 純3着率 = イン逃げ時3着以内率 − イン逃げ時2着率（直接実績）
    #
    #   修正点:
    #     ① win3_rate（全体3連対率）を除去 → イン逃げ局面と無関係のため
    #     ② 会場ベースラインも純3着率ベース（3着以内率 − 2着率）で統一
    #     ③ ST補正は「遅い艇ほど流れ込みで3着に残りやすい」傾向を穏やかに反映
    #
    #   trust加重ブレンド（circle_pctと同じ方式）:
    #     base_pure3 = player_pure3 * w_player + venue_pure3 * w_venue
    # ══════════════════════════════════════════════════════════════════════════

    raw_idx3_scores = {}
    for r in results:
        w = r["waku"]
        if w == "1":
            raw_idx3_scores[w] = None
            continue

        cm = r.get("raw_cm", {})
        ininage_count = safe_float(cm.get("イン逃げ\n出走数"), 0) or 0
        player_3rd    = safe_float(cm.get("3着以内率"))   # イン逃げ時3着以内率
        player_2nd    = safe_float(cm.get("2着率"))       # イン逃げ時2着率
        has_data      = ininage_count >= MIN_ININAGE_COUNT

        # 選手の純3着率（イン逃げ時3着以内率 − 2着率）
        if player_3rd is not None and player_2nd is not None and has_data:
            player_pure3 = max(player_3rd - player_2nd, 0.0)
        elif player_3rd is not None and has_data:
            # 2着率不明 → 3着以内率の35%を純3着と推定（全国平均比から導出）
            player_pure3 = player_3rd * 0.35
        else:
            player_pure3 = None

        # 会場ベースラインの純3着率（枠番別）
        venue_3rd_rate = venue_3rd_map.get(w)
        venue_2nd_rate = venue_frame.get(w)
        if venue_3rd_rate is not None and venue_2nd_rate is not None:
            venue_pure3 = max(venue_3rd_rate - venue_2nd_rate, 0.0)
        else:
            venue_pure3 = None

        national_pure3 = NATIONAL_3RD_PURE.get(w, 0.10)

        # trust加重ブレンド（circle_pctと同じ方式）
        trust    = min(ininage_count / 30.0, 1.0) if has_data else 0.0
        w_player = 0.5 + 0.4 * trust   # 0.50〜0.90
        w_venue  = 1.0 - w_player       # 0.50〜0.10

        venue_val = venue_pure3 if venue_pure3 is not None else national_pure3
        if player_pure3 is not None and has_data:
            score_3rd = player_pure3 * w_player + venue_val * w_venue
        else:
            score_3rd = venue_val

        # ST補正（遅め艇ほど流れ込みで3着に残りやすい）
        avg_st = r.get("avg_st")
        if avg_st is not None:
            # 平均より遅いほど微加算、速いほど微減（影響は小さく±15%以内）
            st_slow_bonus = max(0.85, min(1.15, 1.0 + (avg_st - st_mean) / 0.05 * 0.10))
        else:
            st_slow_bonus = 1.0
        score_3rd = max(score_3rd * st_slow_bonus, 0.0)

        raw_idx3_scores[w] = score_3rd

    # 最大値=100にスケーリング（分散を保つため下限は設けない）
    valid_idx3 = [s for w, s in raw_idx3_scores.items() if w != "1" and s is not None]
    max_idx3   = max(valid_idx3) if valid_idx3 else 1.0
    for r in results:
        w  = r["waku"]
        s3 = raw_idx3_scores.get(w)
        if w != "1" and s3 is not None and max_idx3 > 0:
            r["idx3"] = min(int(s3 / max_idx3 * 100), 100)
        else:
            r["idx3"] = 0

    # frame_2nd: write_race_flat の2着率テキストブロック描画用に渡す
    frame_2nd = {w: s for w, s in raw_scores.items() if s is not None}
    
    # 会場イン逃げ場平均・決まり手場平均
    venue_stats = _calc_venue_stats(venue_stats_master, venue)

    # ── ★新機能: 6人相性スコア・イン飛び条件定量化 ────────────────────────
    # ── 展開考察エンジン（STxコース連動 → 対立構造 → 展開quality）────────────
    # _judge_tobi_scenario の前段として実行し、結果を race_judgment に連携させる。
    # first_turn: 1M到達順序と展開パターン
    # conflict_map: 誰が誰を潰しに行くかの対立構造
    # scenario_quality: 展開がどれだけ絞れているかのqualityスコア
    first_turn      = _predict_first_turn(results)
    conflict_map    = _build_conflict_map(results, first_turn)
    # scenario_quality は _suggest_3rentan 後に s1_prob が確定してから補完する
    # ここでは先行優位度と対立構造のみで暫定計算
    scenario_quality = _calc_scenario_quality(first_turn, conflict_map, s1_prob_est=None)

    affinity = _calc_affinity_score(results, venue_stats_master, venue)
    tobi     = _judge_tobi_scenario(results, affinity, venue_stats)

    # ── 荒れ/堅いレース判定 ＋ 3連単買い目提案（数値シート出力用） ──────────
    race_judgment  = _judge_race_type(results, venue_stats, venue_frame, race_no,
                                      venue_stats_master=venue_stats_master, venue=venue,
                                      tobi_scenario=tobi)
    # ── ★新機能: イン逃げ/イン飛び/両建て 3択判定（暫定：s1_prob未確定）────────
    # この時点では s1_prob がまだ確定していないため、暫定値として計算する。
    # s1_prob 確定後に main() または calc_race_indices の末尾で再計算する。
    ryotate_judgment = _judge_ryotate(race_judgment, tobi, venue_stats, s1_prob=None)
    race_judgment["ryotate"] = ryotate_judgment

    # 会場別1コース1着率を race_judgment に追加（_suggest_3rentan → _calc_3rentan_probs_v2 で使用）
    # 荒れやすい会場（戸田=43%, 平和島=45%等）で過剰なS1重みを抑制するため
    _vs_raw = venue_stats_master.get(venue, {})
    _vc1r = safe_float(_vs_raw.get("1コース1着率") or _vs_raw.get("1C_1着率"))
    race_judgment["venue_c1_win_rate"] = _vc1r  # Noneなら全国平均(0.555)にフォールバック
    race_judgment["affinity"]          = affinity  # ③相性考察で参照
    race_judgment["venue"]             = venue     # 参加見送り判定用

    # ── ★ヒモ荒れ判定 ────────────────────────────────────────────────────────
    # 1号艇が強本命（rel_win1 >= 65%）のとき「2・3着ヒモが荒れるか」を判定し
    # 参加可否と買い目点数調整の根拠として race_judgment に格納する。
    race_judgment["himo_are"] = _judge_himo_are(results, race_judgment)

    bet_suggestions = _suggest_3rentan(results, race_judgment)

    # ── s1_prob 確定後の各種最終補完 ─────────────────────────────────────────
    s1_prob_final = bet_suggestions.get("s1_prob") or race_judgment.get("s1_prob")

    # ① scenario_quality を s1_prob ベースで最終補完
    scenario_quality = _calc_scenario_quality(
        first_turn, conflict_map, s1_prob_est=s1_prob_final
    )

    # ② ryotate（3択判定）を s1_prob 確定後に再計算 ← 【断絶修正】
    #    s1_prob を渡すことで確率モデルと定性スコアの整合性チェックを実行し、
    #    3択の verdict・表示%・consistency_warn をすべて確率ベースに統一する。
    if s1_prob_final is not None:
        ryotate_judgment = _judge_ryotate(
            race_judgment, tobi, venue_stats, s1_prob=s1_prob_final
        )
        race_judgment["ryotate"] = ryotate_judgment

    # ③ ◎艇番とs1_prob最大艇の乖離チェック ← 【内部矛盾検出】
    #    バックテスト除外フラグとして活用可能
    first_prob_map_final = bet_suggestions.get("first_prob_map", {})
    if first_prob_map_final:
        top_prob_waku = max(first_prob_map_final, key=first_prob_map_final.get)
        honmei_waku_check = next(
            (str(r["waku"]) for r in results if r.get("honmei") == "◎"), None
        )
        if honmei_waku_check and honmei_waku_check != top_prob_waku:
            # 印◎ != 確率最大艇 → 矛盾フラグ（展開シナリオが印を覆している）
            race_judgment["honmei_prob_mismatch"] = True
            race_judgment["honmei_prob_mismatch_detail"] = (
                f"◎={honmei_waku_check}号艇 vs 確率最大={top_prob_waku}号艇"
                f"({first_prob_map_final.get(top_prob_waku, 0)*100:.1f}%)"
            )
        else:
            race_judgment["honmei_prob_mismatch"] = False
            race_judgment["honmei_prob_mismatch_detail"] = ""

    # ── 展開考察エンジン結果を race_judgment に格納 ───────────────────────────
    race_judgment["first_turn"]      = first_turn
    race_judgment["conflict_map"]    = conflict_map
    race_judgment["scenario_quality"] = scenario_quality

    return results, slit, venue_stats, frame_2nd, race_judgment, bet_suggestions

# ============================================================
# ★ 新機能① 6人相性スコア計算
# ============================================================

# ════════════════════════════════════════════════════════════════════════
# 展開考察エンジン（3関数セット）
# _predict_first_turn → _build_conflict_map → _calc_scenario_quality
# ════════════════════════════════════════════════════════════════════════

def _predict_first_turn(results):
    """
    STとコース位置から第1ターンマーク（1M）到達順序と展開パターンを推定する。

    【競艇の物理法則】
      第1ターンへの到達時間 ≈ コース距離 + スタートタイム差
      内側コース（1号艇）ほど距離が短い。
      ただし外側艇がST大幅有利なら距離ハンデを逆転できる。

      到達順序（早い順）= コース距離補正後のST実効値でソート。
      実効ST = avg_st + コース距離補正
               コース距離補正: 各コースは1コース基準で
               2コース:+0.03、3コース:+0.06、4コース:+0.10、
               5コース:+0.15、6コース:+0.21（経験則ベース）

    【展開パターンの定義】
      先行艇 = 第1ターンに最初に到達する艇
      追走艇 = 先行艇の直後に続く艇

      パターンA（1号艇先行）  : 1号艇が第1ターンに先着 → 逃げ有利
      パターンB（2号艇先行）  : 2号艇が1号艇より先にターン → 差しが確定的
      パターンC（3〜4号艇先行）: 3〜4号艇が先着 → まくり差し有利
      パターンD（5〜6号艇先行）: 5〜6号艇が先着 → 大外まくり

      先行差（先行艇と2番手の実効ST差）:
        差が大きい（>0.05秒）: 展開が先行艇に強く有利
        差が小さい（<0.02秒）: 接戦、展開が読みにくい

    【ST不明時のフォールバック】
      avg_stがNoneの艇はコース別全国平均STで代替する。
      全国平均: 1C:0.18, 2C:0.17, 3C:0.17, 4C:0.18, 5C:0.19, 6C:0.20

    Returns
    -------
    dict:
        entry_order      : [(waku, eff_st), ...] 1M到達順（早い順・推定）
        lead_waku        : 先行艇（第1ターン最速艇）
        chase_waku       : 追走艇（2番手）
        lead_margin      : 先行差（秒）
        pattern          : "A（逃げ有利）"/"B（差し有利）"/"C（まくり差し）"/"D（大外）"
        pattern_strength : "強"（先行差>0.05）/"中"（>0.02）/"弱"（接戦）
        narrative        : 展開の絵（自然言語）
        st_details       : {waku: {"avg_st", "eff_st", "course_adj"}} デバッグ用
    """
    # コース距離補正（秒）
    COURSE_ADJ = {"1": 0.00, "2": 0.03, "3": 0.06,
                  "4": 0.10, "5": 0.15, "6": 0.21}
    # STデータなし時の全国平均
    ST_NATIONAL = {"1": 0.18, "2": 0.17, "3": 0.17,
                   "4": 0.18, "5": 0.19, "6": 0.20}

    entries = []
    st_details = {}
    for r in results:
        w    = r["waku"]
        st   = r.get("avg_st")
        if st is None or st <= 0:
            st = ST_NATIONAL.get(w, 0.18)
            is_estimated = True
        else:
            is_estimated = False
        adj     = COURSE_ADJ.get(w, 0.10)
        eff_st  = round(st + adj, 4)
        entries.append((w, eff_st))
        st_details[w] = {
            "avg_st":      st,
            "course_adj":  adj,
            "eff_st":      eff_st,
            "estimated":   is_estimated,
        }

    entries.sort(key=lambda x: x[1])

    lead_waku  = entries[0][0]
    chase_waku = entries[1][0] if len(entries) >= 2 else None
    lead_margin = round(entries[1][1] - entries[0][1], 4) if len(entries) >= 2 else 0.0

    # 展開パターン
    try:
        lead_course = int(lead_waku)
    except (ValueError, TypeError):
        lead_course = 1

    if lead_course == 1:
        pattern = "A（逃げ有利）"
    elif lead_course == 2:
        pattern = "B（差し有利）"
    elif lead_course <= 4:
        pattern = "C（まくり差し）"
    else:
        pattern = "D（大外まくり）"

    if lead_margin > 0.05:
        pattern_strength = "強"
    elif lead_margin > 0.02:
        pattern_strength = "中"
    else:
        pattern_strength = "弱（接戦）"

    # 展開の絵（自然言語）
    entry_str = " → ".join([f"{w}号({st:.3f})" for w, st in entries])
    if pattern_strength == "強":
        strength_desc = f"{lead_waku}号艇が{lead_margin:.3f}秒差で先行確定的"
    elif pattern_strength == "中":
        strength_desc = f"{lead_waku}号艇が優位だが{chase_waku}号艇が追走できる差"
    else:
        strength_desc = f"{lead_waku}号艇と{chase_waku}号艇が接戦（展開は流動的）"

    narrative = (
        f"【1M到達順（推定）】{entry_str}\n"
        f"【先行】{lead_waku}号艇 → {pattern} / 強度:{pattern_strength}\n"
        f"{strength_desc}"
    )

    return {
        "entry_order":       entries,
        "lead_waku":         lead_waku,
        "chase_waku":        chase_waku,
        "lead_margin":       lead_margin,
        "pattern":           pattern,
        "pattern_strength":  pattern_strength,
        "narrative":         narrative,
        "st_details":        st_details,
    }


def _build_conflict_map(results, first_turn, cm_map_ext=None):
    """
    1M到達順序を受けて「誰が誰を潰しに行くか」の対立構造を計算する。

    【競艇の対立構造の本質】
      1M到達順序が決まれば、各艇の「攻撃対象」が物理的に決まる。

      攻撃対象の決まり方:
        先行艇（1番手）: 攻撃対象なし（前に誰もいない）
        2番手艇: 先行艇を差す or まくる
        3番手艇: 2番手艇を外から包む（まくり）or 1番手を差す（まくり差し）
        4番手以降: 外からまくる or 内側の争いを待って2着圏に滑り込む

      攻撃強度 = 決まり手適性 × ST差（接戦ほど攻撃が届きやすい）× 位置補正

    【対立の主軸と副軸】
      主軸対立: 1M到達2番手が先行艇を攻撃する構図
      副軸対立: 1M到達3〜4番手が2番手を潰す構図（漁夫の利の発生源）

    【出力する「展開の絵」】
      例:
        主軸: 3号艇(まくり差し系)が1号艇を包む構図
        副軸: 4号艇(まくり系)が3号艇を外から被せる可能性
        潰れ受益: 3号艇が自滅した場合、5号艇（攻撃性低・win3_rate高）が漁夫

    Returns
    -------
    dict:
        main_conflict    : {"attacker", "target", "method", "strength", "desc"}
        sub_conflict     : {"attacker", "target", "method", "strength", "desc"} or None
        collapse_beneficiary : [(waku, score), ...]  潰れ受益候補（スコア降順）
        narrative        : 対立構造の展開の絵（自然言語）
        conflict_entries : 全艇の対立エントリ（詳細デバッグ用）
    """
    entry_order = first_turn["entry_order"]  # [(waku, eff_st), ...]
    wakus_in_order = [w for w, _ in entry_order]

    # cm_mapを構築（resultsから）
    cm_map = {r["waku"]: r.get("raw_cm", {}) for r in results}
    win3_map = {r["waku"]: r.get("win3_rate") or 0.5 for r in results}

    def safe_pct(cm, key):
        v = cm.get(key)
        try:
            return max(float(v), 0.0) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    def _kime_type(waku):
        """決まり手タイプを返す（まくり系/差し系/逃げ系）"""
        cm = cm_map.get(waku, {})
        mak  = safe_pct(cm, "まくり%") + safe_pct(cm, "まくり差し%")
        sash = safe_pct(cm, "差し%")
        nige = safe_pct(cm, "逃げ%")
        best = max(mak, sash, nige)
        if best == 0:
            return "不明"
        if mak == best:
            return "まくり系"
        if sash == best:
            return "差し系"
        return "逃げ系"

    def _attack_strength(attacker_w, target_w, entry_idx_diff):
        """
        攻撃強度 = 決まり手適性スコア × ST接近度補正 × 位置距離補正
        entry_idx_diff: 1M到達順序上の距離（1=直後、2=2つ後ろ etc.）
        """
        cm = cm_map.get(attacker_w, {})
        mak  = safe_pct(cm, "まくり%") + safe_pct(cm, "まくり差し%")
        sash = safe_pct(cm, "差し%")
        base = max(mak, sash)  # 攻撃決まり手の主力
        if base == 0:
            base = 10.0  # データなしフォールバック

        # 位置距離減衰（直後が最も攻撃しやすい）
        dist_factor = {1: 1.0, 2: 0.7, 3: 0.5, 4: 0.3, 5: 0.2}
        dist_dec = dist_factor.get(entry_idx_diff, 0.15)

        # ST接近度（eff_st差が小さいほど追いつきやすい）
        lead_eff = dict(entry_order).get(target_w, 0.5)
        att_eff  = dict(entry_order).get(attacker_w, 0.5)
        st_gap   = abs(lead_eff - att_eff)
        st_factor = max(0.3, 1.0 - st_gap * 8.0)  # 0.1秒差で0.2低下

        return round(base * dist_dec * st_factor, 1)

    # ── 各艇の攻撃対象・手法・強度を計算 ──────────────────────────────────────
    conflict_entries = []
    for idx, (w, eff_st) in enumerate(entry_order):
        if idx == 0:
            # 先行艇は攻撃対象なし（前方クリア）
            conflict_entries.append({
                "waku": w, "role": "先行",
                "target": None, "method": "-", "strength": 0
            })
            continue

        # 攻撃対象 = 1M到達順で1つ前の艇（最も自然な攻撃対象）
        target_w = entry_order[idx - 1][0]
        kime     = _kime_type(w)
        strength = _attack_strength(w, target_w, 1)

        # 攻撃手法の決定
        try:
            w_course      = int(w)
            target_course = int(target_w)
        except (ValueError, TypeError):
            w_course = target_course = 3

        if kime == "差し系" and w_course < target_course + 2:
            method = "差し"
        elif kime == "まくり系":
            method = "まくり"
        elif kime in ("まくり系", "不明"):
            method = "まくり"
        else:
            method = "まくり差し"

        conflict_entries.append({
            "waku":     w,
            "role":     f"{idx+1}番手",
            "target":   target_w,
            "method":   method,
            "kime":     kime,
            "strength": strength,
        })

    # ── 主軸対立（進入2番手 vs 先行艇）──────────────────────────────────────────
    main_entry = conflict_entries[1] if len(conflict_entries) >= 2 else None
    if main_entry and main_entry["target"]:
        main_conflict = {
            "attacker": main_entry["waku"],
            "target":   main_entry["target"],
            "method":   main_entry["method"],
            "strength": main_entry["strength"],
            "desc": (
                f"{main_entry['waku']}号艇が{main_entry['target']}号艇に"
                f"「{main_entry['method']}」で挑む"
                f"（強度{main_entry['strength']:.0f}）"
            ),
        }
    else:
        main_conflict = None

    # ── 副軸対立（進入3番手 vs 2番手）──────────────────────────────────────────
    sub_entry = conflict_entries[2] if len(conflict_entries) >= 3 else None
    if sub_entry and sub_entry["target"]:
        sub_conflict = {
            "attacker": sub_entry["waku"],
            "target":   sub_entry["target"],
            "method":   sub_entry["method"],
            "strength": sub_entry["strength"],
            "desc": (
                f"{sub_entry['waku']}号艇が{sub_entry['target']}号艇を"
                f"「{sub_entry['method']}」で外から被せる"
                f"（強度{sub_entry['strength']:.0f}）"
            ),
        }
    else:
        sub_conflict = None

    # ── 潰れ受益者（主軸攻撃艇が自滅した場合の漁夫候補）────────────────────────
    # 主軸攻撃艇が自滅 → 受益者の条件:
    #   ① 攻撃性が低い（自分は仕掛けに行かない）
    #   ② win3_rateが高い（荒れても残る力がある）
    #   ③ 主軸対立の外側にいる（巻き込まれない位置）
    main_attacker = main_conflict["attacker"] if main_conflict else None
    beneficiary_scores = {}
    for r in results:
        w = r["waku"]
        if w == main_attacker or w == (main_conflict["target"] if main_conflict else None):
            continue  # 主軸の当事者は除外
        cm   = cm_map.get(w, {})
        mak  = safe_pct(cm, "まくり%") + safe_pct(cm, "まくり差し%")
        sash = safe_pct(cm, "差し%")
        attack_rate  = min((mak + sash) / 100.0, 1.0)
        passivity    = 1.0 - attack_rate * 0.6  # 低攻撃性 = 高漁夫スコア
        ground       = win3_map.get(w, 0.5)
        beneficiary_scores[w] = round(passivity * ground, 4)

    collapse_beneficiary = sorted(
        beneficiary_scores.items(), key=lambda x: x[1], reverse=True
    )

    # ── 展開の絵（自然言語ナラティブ）──────────────────────────────────────────
    parts = []
    if main_conflict:
        parts.append(f"主軸: {main_conflict['desc']}")
    if sub_conflict:
        parts.append(f"副軸: {sub_conflict['desc']}")
    if collapse_beneficiary:
        top_b = collapse_beneficiary[0]
        parts.append(
            f"潰れ受益候補: {top_b[0]}号艇"
            f"（攻撃性低・win3高・主軸争いの外側）"
        )
    narrative = "\n".join(parts) if parts else "展開の絵を生成できませんでした"

    return {
        "main_conflict":         main_conflict,
        "sub_conflict":          sub_conflict,
        "collapse_beneficiary":  collapse_beneficiary,
        "narrative":             narrative,
        "conflict_entries":      conflict_entries,
        "lead_waku":             wakus_in_order[0] if wakus_in_order else None,
    }


def _calc_scenario_quality(first_turn, conflict_map, s1_prob_est=None):
    """
    展開の「絞れ度」= 展開が特定シナリオに集中しているか（quality）を計算する。

    【質スコアの概念】
      競艇のレースには「読みやすいレース」と「読みにくいレース」がある。

      読みやすいレース（quality高）:
        → 先行艇が圧倒的有利、攻撃艇が1艇に絞れる
        → 買い目を少点数に絞れる根拠になる
        → 合成オッズが低くても期待値が確保できる

      読みにくいレース（quality低）:
        → 先行艇と追走艇が接戦、攻撃艇が複数
        → 何でも起きる → 買い目を絞ることに根拠がない
        → 高いqualityのレースを待つ方が合理的

    【quality構成要素】
      ① 先行優位度  (0〜40点): lead_margin が大きいほど高い
      ② 主軸集中度  (0〜30点): 主軸攻撃が強く副軸が弱いほど高い
      ③ 展開2択度   (0〜20点): 逃げか1つの飛びかに絞れているほど高い
      ④ 接戦ペナルティ (-10〜0点): 3艇以上が接戦のとき減点

    Returns
    -------
    dict:
        quality_score  : 0〜100
        quality_rank   : "S"(>=75)/"A"(>=55)/"B"(>=40)/"C"(>=25)/"D"(<25)
        quality_verdict: "展開が絞れている" / "要注意（展開流動的）" / "見送り推奨（混戦）"
        bet_size_guide : "点数を絞れる" / "標準点数" / "増やすか見送り"
        components     : 各要素のスコア（デバッグ用）
        narrative      : qualityの説明（自然言語）
    """
    score = 0.0
    components = {}

    # ① 先行優位度（0〜40点）
    lead_margin = first_turn.get("lead_margin", 0)
    if lead_margin > 0.08:
        lead_score = 40.0
    elif lead_margin > 0.05:
        lead_score = 30.0
    elif lead_margin > 0.02:
        lead_score = 20.0
    elif lead_margin > 0.00:
        lead_score = 10.0
    else:
        lead_score = 0.0
    score += lead_score
    components["先行優位度"] = round(lead_score, 1)

    # ② 主軸集中度（0〜30点）
    # 主軸強・副軸弱 = 集中度高
    main = conflict_map.get("main_conflict") or {}
    sub  = conflict_map.get("sub_conflict") or {}
    main_str = main.get("strength", 0) or 0
    sub_str  = sub.get("strength", 0) or 0
    if main_str > 0:
        concentration = main_str / max(main_str + sub_str, 1)
        axis_score = min(30.0, concentration * 30.0 * (main_str / 50.0))
    else:
        axis_score = 0.0
    score += axis_score
    components["主軸集中度"] = round(axis_score, 1)

    # ③ 展開2択度（0〜20点）
    # s1_prob_est が渡されている場合は逃げ確率を使用
    # 渡されていない場合はパターンで推定
    if s1_prob_est is not None:
        # 逃げ一択 or 飛び一択に近いほど高い
        two_choice = abs(s1_prob_est - 0.5) * 2  # 0〜1
        two_score  = two_choice * 20.0
    else:
        pattern = first_turn.get("pattern", "")
        if first_turn.get("pattern_strength") == "強":
            two_score = 20.0
        elif first_turn.get("pattern_strength") == "中":
            two_score = 12.0
        else:
            two_score = 5.0
    score += two_score
    components["展開2択度"] = round(two_score, 1)

    # ④ 接戦ペナルティ
    entry_order = first_turn.get("entry_order", [])
    if len(entry_order) >= 3:
        eff_sts = [st for _, st in entry_order[:4]]
        if len(eff_sts) >= 3:
            spread = eff_sts[2] - eff_sts[0]
            if spread < 0.03:  # 上位3艇が0.03秒以内 = 接戦
                penalty = -10.0
                score += penalty
                components["接戦ペナルティ"] = round(penalty, 1)

    score = max(0.0, min(100.0, score))

    # ── ランク判定（v2: 閾値を緩和してS/Aが出るように再設計）─────────────────
    # 旧版の問題: 最大90点満点のスコアに対し「S>=75」という閾値が高すぎた。
    # lead_margin > 0.08秒（40pt）+ 主軸集中度（最大30pt）+ 2択度（最大20pt）= 90pt が天井。
    # 現実の競艇では lead_margin > 0.08 は珍しいため、Sが出ない構造になっていた。
    #
    # v2 設計思想:
    #   先行差 0.05秒以上（30pt）+ 主軸集中度 >= 20pt + 2択度 >= 12pt → 合計62pt以上でA到達
    #   先行差 0.08秒以上（40pt）+ 同上合計 → 72pt以上でS到達
    #   よく出る lead_margin（0.03〜0.06秒）でもA/Bが取れるよう調整。
    #
    # 新閾値:
    #   S: >= 65  （絞れる展開。少点数推奨）
    #   A: >= 48  （ある程度集中。標準点数）
    #   B: >= 33  （やや流動的。点数増か厳選）
    #   C: >= 18  （読みにくい。参加基準UP）
    #   D: < 18   （混戦。見送り推奨）
    if score >= 65:
        quality_rank    = "S"
        quality_verdict = "展開が絞れている（買い目を絞る根拠あり）"
        bet_size_guide  = "点数を絞れる（少点数で合成オッズを高く保てる）"
    elif score >= 48:
        quality_rank    = "A"
        quality_verdict = "ある程度の展開集中（標準的な点数で対応可）"
        bet_size_guide  = "標準点数"
    elif score >= 33:
        quality_rank    = "B"
        quality_verdict = "展開やや流動的（点数増か厳選が必要）"
        bet_size_guide  = "点数増か主軸に絞る"
    elif score >= 18:
        quality_rank    = "C"
        quality_verdict = "展開が読みにくい（参加基準を上げる）"
        bet_size_guide  = "参加するなら大幅厳選"
    else:
        quality_rank    = "D"
        quality_verdict = "混戦（展開予測が困難。見送り推奨）"
        bet_size_guide  = "見送り推奨"

    # 自然言語
    narrative = (
        f"展開quality: {quality_rank}（{score:.0f}点）\n"
        f"→ {quality_verdict}\n"
        f"→ {bet_size_guide}"
    )

    return {
        "quality_score":   round(score, 1),
        "quality_rank":    quality_rank,
        "quality_verdict": quality_verdict,
        "bet_size_guide":  bet_size_guide,
        "components":      components,
        "narrative":       narrative,
    }


def _calc_affinity_score(results, venue_stats_master, venue):
    """
    出走6選手の相互作用を定量化し、各艇が「インを脅かす力」を数値化する。

    【設計思想】
    イン飛びは「インが弱い」だけでなく「誰かが積極的に攻める」ときに起きる。
    この関数は各艇の攻撃力と、1号艇の被攻撃脆弱性を組み合わせて評価する。

    出力 (dict):
      attack_score[waku]  : 各艇のイン攻撃ポテンシャル (0〜100)
      threat_total        : 全艇合計攻撃力 (イン脅威度)
      boat1_vulnerability : 1号艇の被攻撃脆弱性スコア (0〜100、高いほど飛びやすい)
      dominant_attacker   : 最も脅威の高い艇番 (str)
      affinity_summary    : 各艇の攻撃根拠テキスト dict

    【スコア計算要素】
    A. 差し力 (2コース艇)  : 差し% × ST速さ（1号艇との相対ST差）
    B. まくり力 (3〜5C)    : (まくり% + まくり差し%) × コース別1着率 × ST出足補正
    C. 1号艇の脆弱性       : 差され% + 捲られ% + ST遅さペナルティ + STばらつき
    D. A1選手補正          : 外枠にA1級選手がいる場合に攻撃力を+20%
    E. 今節成績補正        : 直近好成績の艇は攻撃スコアを最大+15%加算
    """
    NATIONAL_ATTACK_BASE = {
        "2": 0.137, "3": 0.134, "4": 0.111, "5": 0.066, "6": 0.021
    }

    res1 = next((r for r in results if r["waku"] == "1"), None)
    st1  = res1.get("avg_st") if res1 else None
    cm1  = res1.get("raw_cm", {}) if res1 else {}

    # 1号艇の脆弱性スコア（0〜100）
    # ※ "差し%"（_get_cm_val経由）は「自分が差して勝った割合」であり脆弱性評価には使わない
    nige_pct    = safe_float(_get_cm_val(cm1, "逃げ%"), 0) or 0    # 逃げ率（高いほど脆弱性低）

    # 被攻撃実績（update_masterが生成する被決まり手%）
    sasar_vuln = safe_float(cm1.get("差され%"), 0) or 0
    makur_vuln = safe_float(cm1.get("捲られ%"), 0) or 0
    maksa_vuln = safe_float(cm1.get("捲り差され%"), 0) or 0  # update_master出力キーに統一

    # ST脆弱性（1号艇が遅いほど脆弱）
    st_vuln = 0.0
    if st1 is not None:
        if st1 > 0.18:
            st_vuln = 30.0
        elif st1 > 0.15:
            st_vuln = 15.0
        elif st1 > 0.12:
            st_vuln = 5.0

    # STばらつき（不安定なほど脆弱）
    pm1 = res1.get("raw_pm", {}) if res1 else {}
    st_stable = safe_float(pm1.get("ST安定\nスコア") or pm1.get("ST安定スコア"))
    st_unstable_vuln = 0.0
    if st_stable is not None:
        if st_stable < 40:
            st_unstable_vuln = 20.0
        elif st_stable < 60:
            st_unstable_vuln = 10.0

    # ── FLY数・出遅れ数によるST脆弱性補正 ────────────────────────────────
    # 選手指数マスタに FLY数・出遅れ数 が集計されているが、従来は未使用だった。
    # FLY（フライング） → ペナルティ後の緊張・心理的影響で次走STが不安定になりやすい
    # 出遅れ（出遅れ数が多い）→ スロースタートの癖がある → イン逃げ失敗リスクUP
    #
    # FLY補正:
    #   FLY数が1走内に存在 → 直近にフライングあり → ST脆弱性+15
    #   FLY数が2走以上    → 繰り返しフライング → ST脆弱性+25（最大値）
    # 出遅れ補正:
    #   出遅れ数/総出走数 が 5%超 → 出遅れ癖あり → ST脆弱性+10
    #   出遅れ数/総出走数 が 10%超 → 出遅れ癖強い → ST脆弱性+20
    fly_count    = safe_float(pm1.get("FLY数"),    0) or 0
    late_count   = safe_float(pm1.get("出遅れ数"), 0) or 0
    total_runs_pm = safe_float(pm1.get("ST計測件数") or pm1.get("総出走数"), 0) or 1

    fly_vuln = 0.0
    if fly_count >= 2:
        fly_vuln = 25.0
    elif fly_count >= 1:
        fly_vuln = 15.0

    late_rate = late_count / max(total_runs_pm, 1)
    late_vuln = 0.0
    if late_rate >= 0.10:
        late_vuln = 20.0
    elif late_rate >= 0.05:
        late_vuln = 10.0

    boat1_vulnerability = min(100.0,
        (sasar_vuln * 100) * 0.30 +          # 差された実績（旧0.35から微調整）
        (makur_vuln * 100) * 0.22 +          # 捲られた実績
        (maksa_vuln * 100) * 0.12 +          # 捲り差された実績
        st_vuln * 0.12 +                     # ST遅さペナルティ
        st_unstable_vuln * 0.08 +            # STばらつきペナルティ
        fly_vuln * 0.10 +                    # 【新追加】FLY履歴ペナルティ
        late_vuln * 0.06 +                   # 【新追加】出遅れ癖ペナルティ
        max(0.0, (40.0 - nige_pct) * 0.5)   # 逃げ%が低いほど加算
    )

    # 各艇の攻撃スコアを計算
    attack_score  = {}
    affinity_summary = {}

    for r in results:
        w = r["waku"]
        if w == "1":
            attack_score[w] = 0.0
            affinity_summary[w] = "1号艇（逃げ側）"
            continue

        cm  = r.get("raw_cm", {})
        pm  = r.get("raw_pm", {})
        st  = r.get("avg_st")

        # 基礎攻撃力（決まり手%）
        sashi_pct = safe_float(_get_cm_val(cm, "差し%"), 0) or 0
        makuri_pct = (safe_float(_get_cm_val(cm, "まくり%"), 0) or 0) + \
                     (safe_float(_get_cm_val(cm, "まくり差し%"), 0) or 0)
        nat_base   = NATIONAL_ATTACK_BASE.get(w, 0.05) * 100

        if w == "2":
            # 2号艇：差し特化
            base_attack = sashi_pct if sashi_pct > 0 else nat_base * 0.3
            attack_type = f"差し{sashi_pct:.0f}%"
        else:
            # 3〜6号艇：まくり系
            base_attack = makuri_pct if makuri_pct > 0 else nat_base * 0.4
            attack_type = f"まくり系{makuri_pct:.0f}%"

        # ST相対補正（1号艇より速い艇は攻撃力UP）
        st_boost = 1.0
        if st is not None and st1 is not None:
            diff = st1 - st  # 正 = この艇が速い
            if diff > 0.03:
                st_boost = 1.20
            elif diff > 0.01:
                st_boost = 1.10
            elif diff < -0.03:
                st_boost = 0.85

        # 今節成績補正（直近の実走成績から個人の調子を反映）
        kosetsu = str(r.get("kosetsu", ""))
        kosetsu_boost = 1.0
        if kosetsu and kosetsu not in ("", "None", "nan", "-"):
            try:
                tokens = [t.strip() for t in re.split(r"[-・/]", kosetsu)]
                win_count = tokens.count("1")
                if win_count >= 2:
                    kosetsu_boost = 1.15
                elif win_count == 1:
                    kosetsu_boost = 1.07
            except Exception:
                pass

        # コース別実績補正（会場別コースマスタ優先）
        # ※ 級別（A1/A2等）は使用しない。
        #   個々の選手が「実際にそのコースで何をしてきたか」
        #   （決まり手%・ST・今節成績）のみで評価する。
        venue_win = safe_float(r.get("win1_rate")) or 0

        score = base_attack * st_boost * kosetsu_boost
        # 実績との整合（選手の実際のコース別1着率で加重）
        score = score * 0.7 + venue_win * 100 * 0.3

        attack_score[w] = round(score, 2)
        summary_parts = [attack_type]
        if st_boost > 1.05:
            summary_parts.append(f"ST優位(×{st_boost:.2f})")
        elif st_boost < 0.90:
            summary_parts.append(f"ST劣位(×{st_boost:.2f})")
        if kosetsu_boost > 1.0:
            summary_parts.append(f"今節好調(×{kosetsu_boost:.2f})")
        affinity_summary[w] = " / ".join(summary_parts)

    # 合計攻撃力（1号艇除く）
    threat_total = sum(v for k, v in attack_score.items() if k != "1")

    # 最大攻撃艇
    outer_scores = {k: v for k, v in attack_score.items() if k != "1"}
    dominant_attacker = max(outer_scores, key=lambda k: outer_scores[k]) if outer_scores else "-"

    return {
        "attack_score":       attack_score,
        "threat_total":       round(threat_total, 2),
        "boat1_vulnerability": round(boat1_vulnerability, 1),
        "dominant_attacker":  dominant_attacker,
        "affinity_summary":   affinity_summary,
    }


# ============================================================
# ★ 新機能② イン飛び条件総合判定
# ============================================================
def _judge_tobi_scenario(results, affinity, venue_stats):
    """
    イン飛び（1号艇が1着にならない）の総合確率と根拠を返す。

    【判定ロジック】
    以下5つの条件を重みづけして飛び確率(0〜100)を算出する。

    条件①  1号艇の脆弱性スコア（被差され・捲られ実績）
    条件②  攻撃艇の合計脅威スコア（全艇の攻撃力合算）
    条件③  会場のイン逃げ率（低い会場ほど飛びやすい）
    条件④  支配的攻撃艇の存在感（最強攻撃艇のスコアが突出しているか）
    条件⑤  スリット不利（1号艇STが相対的に遅い場合）

    出力:
      tobi_prob   : イン飛び推定確率 (0〜100)
      tobi_rank   : 飛び確率ランク  S(>70)/A(>55)/B(>40)/C(>25)/D(<=25)
      main_threat : 最も危険な飛ばし役の艇番
      reasons     : 判定根拠リスト
      tobi_type   : 予想される飛び方 ("差し" / "まくり" / "まくり差し" / "不明")
    """
    reasons = []
    score = 0.0  # 0〜100: 高いほどイン飛び確率大

    vuln  = affinity["boat1_vulnerability"]
    total = affinity["threat_total"]
    dom   = affinity["dominant_attacker"]
    atk   = affinity["attack_score"]

    # 条件①: 1号艇の脆弱性（最大35点）
    vul_contrib = min(35.0, vuln * 0.35)
    score += vul_contrib
    if vuln >= 40:
        reasons.append(f"1号艇脆弱性スコア{vuln:.0f}（被差し・捲られ実績大）")
    elif vuln >= 20:
        reasons.append(f"1号艇脆弱性スコア{vuln:.0f}（やや脆弱）")

    # 条件②: 攻撃艇の合計脅威（最大30点）
    # threat_total の自然な範囲は0〜150程度 → 100以上で満点
    threat_contrib = min(30.0, total / 100.0 * 30.0)
    score += threat_contrib
    if total >= 80:
        reasons.append(f"攻撃艇合計スコア{total:.0f}（攻撃力が強い）")
    elif total >= 50:
        reasons.append(f"攻撃艇合計スコア{total:.0f}（中程度の攻撃）")

    # 条件③: 会場イン逃げ率（最大20点）
    in_rate = venue_stats.get("in_rate")
    if in_rate is not None:
        # イン逃げ率が低い会場ほどスコアUP（逃げ率50%以下で最大）
        venue_contrib = max(0.0, (0.65 - float(in_rate)) / 0.65 * 20.0)
        score += venue_contrib
        if float(in_rate) < 0.45:
            reasons.append(f"会場イン逃げ率{float(in_rate)*100:.0f}%（飛びやすい会場）")

    # 条件④: 支配的攻撃艇の突出度（最大10点）
    if dom != "-":
        dom_score = atk.get(dom, 0)
        other_scores = [v for k, v in atk.items() if k != "1" and k != dom]
        if other_scores:
            avg_other = sum(other_scores) / len(other_scores)
            if avg_other > 0 and dom_score > avg_other * 1.5:
                score += 10.0
                reasons.append(f"{dom}号艇が突出した攻撃力({dom_score:.0f}pt)：明確な飛ばし役")
            elif dom_score > 0:
                score += 5.0
                reasons.append(f"{dom}号艇が攻撃力最大({dom_score:.0f}pt)")

    # 条件⑤: スリット不利（1号艇のSTが遅い）（最大5点）
    res1 = next((r for r in results if r["waku"] == "1"), None)
    if res1:
        st1 = res1.get("avg_st")
        all_sts = [r.get("avg_st") for r in results if r.get("avg_st") is not None and r["waku"] != "1"]
        if st1 is not None and all_sts:
            faster_count = sum(1 for s in all_sts if s < st1 - 0.02)
            if faster_count >= 3:
                score += 5.0
                reasons.append(f"1号艇STが相対的に遅い（{faster_count}艇が有意に速い）")

    score = min(100.0, max(0.0, score))

    # ランク判定
    if score >= 70:
        tobi_rank = "S"
    elif score >= 55:
        tobi_rank = "A"
    elif score >= 40:
        tobi_rank = "B"
    elif score >= 25:
        tobi_rank = "C"
    else:
        tobi_rank = "D"

    # 飛び方の予測（最強攻撃艇のコースと決まり手から推定）
    tobi_type = "不明"
    if dom != "-":
        dom_r = next((r for r in results if r["waku"] == dom), None)
        if dom_r:
            cm_dom = dom_r.get("raw_cm", {})
            sashi  = safe_float(_get_cm_val(cm_dom, "差し%"), 0) or 0
            makuri = safe_float(_get_cm_val(cm_dom, "まくり%"), 0) or 0
            maksa  = safe_float(_get_cm_val(cm_dom, "まくり差し%"), 0) or 0
            if dom == "2":
                tobi_type = "差し"
            elif makuri >= maksa and makuri >= sashi:
                tobi_type = "まくり"
            elif maksa >= makuri and maksa >= sashi:
                tobi_type = "まくり差し"
            elif sashi > 0:
                tobi_type = "差し"

    return {
        "tobi_prob":   round(score, 1),
        "tobi_rank":   tobi_rank,
        "main_threat": dom,
        "reasons":     reasons,
        "tobi_type":   tobi_type,
        "affinity":    affinity,
    }


# ============================================================
# ★ 新機能③ イン逃げ / イン飛び / 両建て 3択判定
# ============================================================
def _judge_ryotate(race_judgment, tobi_scenario, venue_stats, s1_prob=None):
    """
    「イン逃げ狙い」「イン飛び狙い」「両建て」の3択を判定する。

    【判定ロジック v2 — 確率モデルとの一貫性確保】
    s1_prob（確率モデル由来）が渡された場合はそれを主軸にし、
    escape_score（定性スコア）との乖離を検出・調整する。

    - s1_prob が渡される場合（bet_suggestions確定後の再呼び出し）:
        確率値を直接 verdict の判定基準として使用し、
        定性スコアとの整合チェックを行う。
    - s1_prob が渡されない場合（初回暫定計算）:
        従来通り escape_score / tobi_score の差分で判定。

    整合性チェック:
      s1_prob が高い（≥0.65）のに escape_score が低い（<50）場合 →
        確率モデル優先で逃げスコアを上方補正し警告フラグを立てる。
      tobi_score が高い（≥55）のに s1_prob も高い（≥0.60）場合 →
        両建て推奨とし矛盾フラグを立てる。

    出力:
      verdict          : "逃げ狙い" / "飛び狙い" / "両建て推奨"
      confidence       : 判断の確信度 (0〜100)
      escape_score     : 逃げスコア（補正後）
      tobi_score       : 飛びスコア
      escape_pct       : 逃げ確率%（s1_prob * 100、確率モデル由来）
      tobi_pct         : 飛び確率%（(1 - s1_prob) * 100）
      reason           : 判定根拠
      buy_style        : 買い方の具体的指示
      consistency_warn : 定性スコアと確率モデルの乖離警告（True/False）
      realtime_hook    : リアルタイム情報で確認すべき事項
    """
    escape_score = float(race_judgment.get("score", 50))
    tobi_score   = float(tobi_scenario.get("tobi_prob", 30))

    # ── s1_prob による整合性チェックと補正 ──────────────────────────────
    consistency_warn = False
    consistency_note = ""

    if s1_prob is not None:
        fly_prob = 1.0 - s1_prob
        # s1_probを0〜100スケールのスコアに変換して escape_score と比較
        s1_score_equiv = s1_prob * 100.0

        # ケース①: 確率モデルは逃げ優勢なのに定性スコアが低い
        if s1_prob >= 0.65 and escape_score < 50:
            consistency_warn = True
            consistency_note = (
                f"⚠️ 確率モデル逃げ{s1_prob*100:.0f}%だが定性スコア{escape_score:.0f}。"
                f"確率モデル優先で補正。"
            )
            # 確率モデルに寄せて escape_score を上方補正（加重平均）
            escape_score = round(escape_score * 0.35 + s1_score_equiv * 0.65, 1)

        # ケース②: 確率モデルは飛び優勢なのに定性スコアが高い
        elif s1_prob < 0.45 and escape_score >= 65:
            consistency_warn = True
            consistency_note = (
                f"⚠️ 確率モデル逃げ{s1_prob*100:.0f}%（飛び優勢）だが定性スコア{escape_score:.0f}（高）。"
                f"両建てに引き寄せ。"
            )
            # 確率モデルに寄せて escape_score を下方補正
            escape_score = round(escape_score * 0.35 + s1_score_equiv * 0.65, 1)

        # ケース③: 飛びスコアが高いのに s1_prob も高い（真の矛盾）
        elif tobi_score >= 55 and s1_prob >= 0.60:
            consistency_warn = True
            consistency_note = (
                f"⚠️ 飛びスコア{tobi_score:.0f}と逃げ確率{s1_prob*100:.0f}%が矛盾。"
                f"両建て推奨に強制。"
            )
            # この場合は両建てに強制するため escape_score を中間値に
            escape_score = round((escape_score + s1_score_equiv) / 2, 1)
            tobi_score   = round(tobi_score * 0.8, 1)   # 飛びスコアも緩める
    else:
        fly_prob = None

    diff = escape_score - tobi_score

    # ── 表示用確率（s1_prob 確定後のみ意味を持つ） ───────────────────────
    escape_pct = round(s1_prob * 100, 1) if s1_prob is not None else None
    tobi_pct   = round((1.0 - s1_prob) * 100, 1) if s1_prob is not None else None

    # リアルタイム情報差し込み口
    realtime_hook = {
        "展示タイム_確認事項":  "1号艇の展示タイムがレース内偏差値50未満なら飛びスコア+10",
        "直前オッズ_確認事項":  "1号艇単勝オッズが1.2倍未満なら過剰人気→飛び狙いの妙味UP",
        "進入_確認事項":        "枠なり進入確認後にコース変更があれば再判定を推奨",
        "hook_fn":              None,
    }

    if escape_score >= 65 and tobi_score < 40:
        verdict     = "逃げ狙い"
        confidence  = min(100, int(escape_score * 0.7 + (65 - tobi_score) * 0.3))
        buy_style   = (
            f"1号艇1着固定。2着は「circle_pct（イン逃げ時2着率）上位2〜3艇」に絞る。"
            f"3連単で5〜8点。"
        )
        reason = (
            f"逃げスコア{escape_score:.0f}（高）・飛びスコア{tobi_score:.0f}（低）。"
            f"イン逃げが成立しやすい局面。"
        )
        if escape_pct is not None:
            reason += f" 確率モデル逃げ{escape_pct:.0f}%。"

    elif tobi_score >= 55 and escape_score < 55:
        threat = tobi_scenario.get("main_threat", "-")
        ttype  = tobi_scenario.get("tobi_type", "不明")
        verdict     = "飛び狙い"
        confidence  = min(100, int(tobi_score * 0.7 + (65 - escape_score) * 0.3))
        buy_style   = (
            f"{threat}号艇1着候補（{ttype}）。2・3着は残りの内寄り艇を広めに。"
            f"3連単で6〜10点。1号艇2・3着付けは除外か最小限に。"
        )
        reason = (
            f"飛びスコア{tobi_score:.0f}（高）・逃げスコア{escape_score:.0f}（低）。"
            f"{threat}号艇が主な脅威（{ttype}）。"
        )
        if tobi_pct is not None:
            reason += f" 確率モデル飛び{tobi_pct:.0f}%。"

    else:
        verdict     = "両建て推奨"
        confidence  = max(10, 80 - int(abs(diff) * 1.5))
        threat = tobi_scenario.get("main_threat", "-")
        ttype  = tobi_scenario.get("tobi_type", "不明")
        buy_style   = (
            f"【逃げ軸】1号艇1着の買い目を確保（4〜5点）。"
            f"【飛び軸】{threat}号艇1着の買い目（3〜4点）。"
            f"合計7〜9点。両軸を保持し展示後に軸の比重を傾ける。"
        )
        reason = (
            f"逃げスコア{escape_score:.0f}・飛びスコア{tobi_score:.0f}（差{abs(diff):.0f}pt）。"
            f"拮抗しており単軸は危険。{threat}号艇({ttype})が飛ばし役候補。"
        )
        if escape_pct is not None:
            reason += f" 確率モデル逃げ{escape_pct:.0f}%/飛び{tobi_pct:.0f}%。"

    if consistency_warn and consistency_note:
        reason = consistency_note + " " + reason

    return {
        "verdict":          verdict,
        "confidence":       confidence,
        "escape_score":     round(escape_score, 1),
        "tobi_score":       round(tobi_score, 1),
        "escape_pct":       escape_pct,    # 確率モデル由来の逃げ%（None=初回暫定）
        "tobi_pct":         tobi_pct,      # 確率モデル由来の飛び%
        "reason":           reason,
        "buy_style":        buy_style,
        "consistency_warn": consistency_warn,
        "realtime_hook":    realtime_hook,
    }


def _judge_himo_are(results, race_judgment):
    """
    ヒモ荒れ判定：1号艇が有力本命のときに「2・3着ヒモが荒れるか」を評価する。

    【v2 変更点】
    旧版は rel_win1 >= 60% という高すぎる閾値で「対象外」が大半を占めていた。
    番組表確定時点では 45% 以上あれば参加/見送り判断の材料として十分有効。

    【判定フロー v2】
    Step1: rel_win1 >= 45% → 判定対象（旧60%→45%に緩和）
    Step2: 最有力3連単の推定確率と2着集中度でヒモ固まり度を評価
    Step3: 展示前推奨アクションを明示（展示確認トリガーを含む）

    【閾値基準 v2】
    max_combo_prob:
      >= 0.25 → 不参加推奨（推定最高人気オッズ ≈ 5倍以下）
      0.12〜0.25 → 点数絞り
      < 0.12   → 参加推奨（ヒモ分散・広め流し）
    """
    res1     = next((r for r in results if r["waku"] == "1"), None)
    rel_win1 = res1.get("rel_win1") if res1 else None

    NOT_TARGET = {
        "is_target": False, "verdict": "対象外", "max_combo_prob": None,
        "est_top_odds": None, "circle_concentration": None,
        "eligible_count": 0, "bet_adj": 0,
        "reason": "1号艇rel_win1 < 45%: 通常判定に委ねる",
        "tenji_trigger": "",
    }
    if rel_win1 is None or rel_win1 < 45.0:
        return NOT_TARGET

    # ── Step2: 組み合わせ確率を計算 ──────────────────────────────────────
    wakus_rest = [r["waku"] for r in results if r["waku"] != "1"]
    sum_rel    = sum(r.get("rel_win1") or 0 for r in results) or 100.0
    p1_win     = rel_win1 / sum_rel

    circ_raw  = {r["waku"]: max(r.get("circle_pct") or 0, 0.001)
                 for r in results if r["waku"] != "1"}
    total_circ = sum(circ_raw.values()) or 1.0
    idx3_raw  = {r["waku"]: max(float(r.get("idx3") or 0), 0.001)
                 for r in results if r["waku"] != "1"}

    combo_probs = []
    for second in wakus_rest:
        p2 = circ_raw[second] / total_circ
        remaining = [w for w in wakus_rest if w != second]
        total_i3  = sum(idx3_raw[w] for w in remaining) or 1.0
        for third in remaining:
            p3   = idx3_raw[third] / total_i3
            prob = p1_win * p2 * p3
            combo_probs.append((f"1-{second}-{third}", prob))

    combo_probs.sort(key=lambda x: x[1], reverse=True)
    max_combo_prob = combo_probs[0][1] if combo_probs else 0.0
    est_top_odds   = round(1.0 / (max_combo_prob * 0.75), 1) if max_combo_prob > 0 else 999.9

    # ── Step3: 2着集中度 ──────────────────────────────────────────────────
    circ_sorted   = sorted(circ_raw.items(), key=lambda x: x[1], reverse=True)
    top2_circ_sum = (
        (circ_sorted[0][1] + circ_sorted[1][1]) / total_circ * 100
        if len(circ_sorted) >= 2 else 100.0
    )

    # ── Step4: 有効組み合わせ数 ────────────────────────────────────────────
    rank     = (race_judgment or {}).get("rank", "B")
    TARGET_BETS = {"S": 6, "A": 8, "B": 9, "C": 6, "D": 6}
    target_n = TARGET_BETS.get(rank, 8)
    _vc1 = (race_judgment or {}).get("venue_c1_win_rate")
    if _vc1 is not None:
        target_n += 2 if _vc1 < 0.45 else (1 if _vc1 < 0.50 else 0)
    eligible_count = min(target_n, len(combo_probs))

    # ── Step5: 総合判定（v2: 閾値を緩和・展示確認トリガー追加）─────────────
    circ_adj = 0.04 if top2_circ_sum >= 70 else (-0.04 if top2_circ_sum < 55 else 0.0)
    prob_adj = max_combo_prob + circ_adj

    # 展示確認トリガー（展示前システムとして必須）
    # 1号艇のST実績をチェックしてトリガー文言を組み立てる
    st1 = res1.get("avg_st") if res1 else None
    st_trigger = ""
    if st1 is not None:
        if st1 > 0.18:
            st_trigger = f"展示で1号艇スタート遅め({st1:.3f}秒)確認→ヒモが荒れやすい"
        elif st1 < 0.14:
            st_trigger = f"展示で1号艇スタート超安定({st1:.3f}秒)→ヒモ固まり方向に注意"

    reasons = [
        f"rel_win1={rel_win1:.1f}%（本命度）",
        f"最有力組み合わせ確率: {max_combo_prob:.3f}（推定実オッズ≈{est_top_odds:.0f}倍）",
        f"2着集中度: 上位2艇 {top2_circ_sum:.1f}%（補正{circ_adj:+.2f}）",
        f"総合判定確率: {prob_adj:.3f}",
    ]

    if prob_adj >= 0.25:
        verdict = "不参加推奨"
        bet_adj = -99
        tenji_trigger = (
            f"展示確認: 1号艇の伸びが平凡以下なら見送り確定。"
            f"好伸びでも推定オッズ{est_top_odds:.0f}倍台のため回収期待値低。"
            + (f"\n{st_trigger}" if st_trigger else "")
        )
        reasons.append(f"→ ヒモ固まり（推定1番人気{est_top_odds:.0f}倍台）。回収率が構造的に低い")
    elif prob_adj >= 0.12:
        verdict = "点数絞り"
        bet_adj = 0
        tenji_trigger = (
            f"展示確認: 1号艇伸び良→そのまま採用。悪→飛び組に差し替え。"
            f"circle_pct上位2艇に絞り込むこと。"
            + (f"\n{st_trigger}" if st_trigger else "")
        )
        reasons.append("→ ヒモはやや固め。展示後に上位2艇以外を切ること")
    else:
        verdict = "参加推奨"
        bet_adj = +2
        tenji_trigger = (
            f"展示確認: ヒモ分散レース。1号艇伸び確認後、広めのヒモ流しを採用。"
            f"穴ヒモ（5・6号艇）を積極的に含める。"
            + (f"\n{st_trigger}" if st_trigger else "")
        )
        reasons.append("→ ヒモ分散。1号艇1着固定で広めのヒモ流し推奨（買い目+2点）")

    return {
        "is_target":            True,
        "verdict":              verdict,
        "max_combo_prob":       round(max_combo_prob, 4),
        "est_top_odds":         est_top_odds,
        "circle_concentration": round(top2_circ_sum, 1),
        "eligible_count":       eligible_count,
        "bet_adj":              bet_adj,
        "reason":               " / ".join(reasons),
        "tenji_trigger":        tenji_trigger,
    }


def _build_kimari(cm):
    parts = []
    mapping = [
        ("逃げ%", "逃"),
        ("差し%", "差"),
        ("まくり%", "ま"),
        ("まくり差し%", "差ま"),
        ("抜き%", "抜"),
    ]
    for col, label in mapping:
        v = safe_float(_get_cm_val(cm, col))
        if v and v > 0:
            parts.append(f"{label}{int(round(v))}%")
    return " ".join(parts) if parts else "-"

def _judge_race_type(results, venue_stats, venue_frame, race_no=None,
                     venue_stats_master=None, venue=None, tobi_scenario=None):
    """
    荒れ/堅いレース自動判定。乗法モデルで複合確率スコアを計算して S/A/B/C/D を返す。
    score 高（1.0に近い）→ 堅い、低（0.0に近い）→ 荒れ

    【旧方式の問題点と修正理由】
    旧方式: 各指標を単純に加算/減算 → score が 50±αで計算
      問題①: 加算モデルは独立事象の積を無視する
        例）「1号艇3連対率92%」かつ「会場イン逃げ率60%」の複合は
             P(堅い) = 0.92 × 0.60 ≒ 0.55 だが、加算では過大評価される
      問題②: score が 0〜100 に張り付きやすく、ランク境界が不安定
      問題③: 各指標の重みが明示されず、調整根拠が不明

    【新方式: 乗法モデル】
      base_prob = 各確率因子の積（0〜1）
      修正因子（乗算）= 連続値で補正（0.5〜1.5程度）
      最終スコア = base_prob × 修正因子群 を 0〜100 にスケール

      確率因子（直接確率として解釈できるもの）:
        P1: 1号艇の相対1着率 / 理論値(40%)   正規化乗数
        P2: 1号艇の3連対率                  直接使用
        P3: 会場イン逃げ率                  直接使用
        P4: 1号艇がTop-1の確率（1位集中度）  Top1比率

      修正因子（確率ではなく補正係数）:
        M1: フォーム指数補正     0.85〜1.15
        M2: ST安定スコア補正     0.90〜1.10
        M3: FLY/出遅れペナルティ 0.75〜1.00
        M4: 飛びシナリオ補正     0.60〜1.00
        M5: R番号補正            0.90〜1.05
        M6: データ不足補正       0.70〜1.00
    """
    reasons = []
    honmei_concentrated = False
    two_top_race = False

    res1 = next((r for r in results if r["waku"] == "1"), None)

    # ─── 確率因子 ───────────────────────────────────────────────────────────

    # P1: 1号艇の相対1着率（理論値40%との比で正規化、上限1.5・下限0.3）
    p1 = 1.0
    if res1 and res1.get("rel_win1") is not None:
        r1 = res1["rel_win1"]
        p1 = max(0.30, min(1.50, r1 / 40.0))
        if r1 >= 45:
            reasons.append(f"1号艇相対1着率{r1:.1f}%（突出 ×{p1:.2f}）")
        elif r1 >= 35:
            reasons.append(f"1号艇相対1着率{r1:.1f}%（高い ×{p1:.2f}）")
        elif r1 <= 20:
            reasons.append(f"1号艇相対1着率{r1:.1f}%（低い→荒れ要素 ×{p1:.2f}）")
        elif r1 <= 28:
            reasons.append(f"1号艇相対1着率{r1:.1f}%（やや低い ×{p1:.2f}）")

    # P2: 1号艇の3連対率（直接確率として使用、基準値0.75）
    # ※ 95%超は少数データ疑いのため0.90に上限クリップ
    p2 = 0.75  # データなし時のデフォルト
    if res1 and res1.get("win3_rate") is not None:
        w3 = float(res1["win3_rate"])
        if w3 >= 0.95:
            p2 = 0.90   # 高すぎ → 少数データ疑い → 上限クリップ
            reasons.append(f"1号艇3連対率{w3*100:.0f}%（高すぎ→少数データ疑い→0.90にクリップ）")
        elif 0.92 <= w3 < 0.95:
            p2 = w3
            reasons.append(f"1号艇3連対率{w3*100:.0f}%（最安定帯 92-95%）")
        elif w3 >= 0.90:
            p2 = w3
            reasons.append(f"1号艇3連対率{w3*100:.0f}%（安定型）")
        elif w3 <= 0.55:
            p2 = w3
            reasons.append(f"1号艇3連対率{w3*100:.0f}%（不安定）")
        else:
            p2 = w3

    # P3: 会場イン逃げ率（直接確率として使用、デフォルト0.55）
    p3 = 0.55
    in_rate = venue_stats.get("in_rate")
    if in_rate is not None:
        p3 = float(in_rate)
        if in_rate >= 0.60:
            reasons.append(f"会場イン逃げ率{in_rate*100:.1f}%（逃げ率高）")
        elif in_rate <= 0.42:
            reasons.append(f"会場イン逃げ率{in_rate*100:.1f}%（荒れやすい会場）")

    # P4: 1号艇のトップ集中度（sorted_rel[0] / sorted_rel[1] の比率）
    p4 = 1.0
    sorted_rel = sorted(
        [r["rel_win1"] for r in results if r.get("rel_win1") is not None],
        reverse=True
    )
    if len(sorted_rel) >= 2:
        gap = sorted_rel[0] - sorted_rel[1]
        h2_threat = sorted_rel[1]
        if h2_threat >= 22 and gap < 15:
            # 実質2強 → 集中度低下
            p4 = 0.80
            two_top_race = True
            reasons.append(f"2番手脅威{h2_threat:.1f}%・差{gap:.1f}pt（実質2強→2軸推奨 ×{p4:.2f}）")
        elif gap >= 20:
            # 本命突出
            p4 = 1.20
            honmei_concentrated = True
            reasons.append(f"1・2位差{gap:.1f}pt（本命突出 ×{p4:.2f}）")
        elif gap <= 8:
            # 実力拮抗
            p4 = 0.85
            reasons.append(f"1・2位差{gap:.1f}pt（実力拮抗 ×{p4:.2f}）")

    # base_prob: P1〜P4の積（0〜1.5程度、後でスケール化）
    base_prob = p1 * p2 * p3 * p4

    # ─── 修正因子 ───────────────────────────────────────────────────────────

    # M1: フォーム指数補正
    m1 = 1.0
    if res1:
        pm1 = res1.get("raw_pm", {})
        form_idx = safe_float(pm1.get("フォーム\n指数") or pm1.get("フォーム指数"))
        if form_idx is not None:
            if form_idx >= 8.0:
                m1 = 1.12; reasons.append(f"1号艇フォーム指数{form_idx:.1f}（🔥ホット ×{m1:.2f}）")
            elif form_idx < 4.0:
                m1 = 0.88; reasons.append(f"1号艇フォーム指数{form_idx:.1f}（❄コールド ×{m1:.2f}）")
            elif form_idx < 6.0:
                m1 = 0.94; reasons.append(f"1号艇フォーム指数{form_idx:.1f}（やや不調 ×{m1:.2f}）")

    # M2: ST安定スコア補正
    m2 = 1.0
    if res1:
        pm1 = res1.get("raw_pm", {})
        st_stable = safe_float(pm1.get("ST安定\nスコア") or pm1.get("ST安定スコア"))
        if st_stable is not None:
            if st_stable >= 80:
                m2 = 1.08; reasons.append(f"1号艇ST安定スコア{st_stable:.0f}（◎超安定 ×{m2:.2f}）")
            elif st_stable >= 60:
                m2 = 1.04; reasons.append(f"1号艇ST安定スコア{st_stable:.0f}（○安定 ×{m2:.2f}）")
            elif st_stable < 40:
                m2 = 0.92; reasons.append(f"1号艇ST安定スコア{st_stable:.0f}（×不安定→ST信頼性低 ×{m2:.2f}）")

    # M3: FLY数・出遅れ数によるペナルティ
    m3 = 1.0
    if res1:
        pm1 = res1.get("raw_pm", {})
        fly_count_j  = safe_float(pm1.get("FLY数"),    0) or 0
        late_count_j = safe_float(pm1.get("出遅れ数"), 0) or 0
        st_meas_j    = safe_float(pm1.get("ST計測件数") or pm1.get("総出走数"), 1) or 1
        late_rate_j  = late_count_j / max(st_meas_j, 1)

        if fly_count_j >= 2:
            m3 *= 0.82; reasons.append(f"1号艇FLY{int(fly_count_j)}回（直近複数FLY ×0.82）")
        elif fly_count_j >= 1:
            m3 *= 0.91; reasons.append(f"1号艇FLY{int(fly_count_j)}回（直近FLYあり ×0.91）")

        if late_rate_j >= 0.10:
            m3 *= 0.88; reasons.append(f"1号艇出遅れ率{late_rate_j*100:.0f}%（出遅れ癖あり ×0.88）")
        elif late_rate_j >= 0.05:
            m3 *= 0.94; reasons.append(f"1号艇出遅れ率{late_rate_j*100:.0f}%（出遅れやや多い ×0.94）")

    # M4: イン飛びシナリオ補正（飛び確率が高いほど逃げ確率を下げる）
    m4 = 1.0
    if tobi_scenario is not None:
        tobi_prob = tobi_scenario.get("tobi_prob", 0)
        if tobi_prob >= 70:
            m4 = 0.62; reasons.append(f"飛び確率{tobi_prob:.0f}%（S：強い飛び示唆 ×{m4:.2f}）")
        elif tobi_prob >= 55:
            m4 = 0.75; reasons.append(f"飛び確率{tobi_prob:.0f}%（A：飛び傾向あり ×{m4:.2f}）")
        elif tobi_prob >= 40:
            m4 = 0.88; reasons.append(f"飛び確率{tobi_prob:.0f}%（B：やや飛びの余地 ×{m4:.2f}）")

    # M5: R番号補正（後半Rほど荒れやすい）
    m5 = 1.0
    NATIONAL_R_AREYASUSA = {
        1: 55, 2: 52, 3: 50, 4: 48, 5: 46, 6: 44,
        7: 43, 8: 42, 9: 41, 10: 40, 11: 39, 12: 38,
    }
    try:
        rno = int(race_no) if hasattr(race_no, '__int__') else int(str(race_no))
        r_are_score = None
        if venue_stats_master and venue:
            vs_m = venue_stats_master.get(venue, {})
            r_are_score = safe_float(vs_m.get(f"{rno}R荒れスコア"))
        if r_are_score is None:
            r_are_score = NATIONAL_R_AREYASUSA.get(rno, 45)
        # 荒れスコア50を基準に乗算係数へ変換（50→1.0、55→1.03、38→0.94）
        m5 = max(0.90, min(1.06, 1.0 + (float(r_are_score) - 50) * 0.006))
        if abs(m5 - 1.0) >= 0.02:
            direction = "堅め" if m5 > 1.0 else "荒れやすい"
            reasons.append(f"{rno}R荒れスコア{r_are_score:.0f}（{direction} ×{m5:.3f}）")
    except (ValueError, TypeError) as e:
        print(f"  ⚠️  R番号補正の計算をスキップしました: {e}")

    # M6: データ不足補正
    missing_list = [r for r in results if r.get("data_missing")]
    missing = len(missing_list)
    m6 = max(0.70, 1.0 - missing * 0.08)  # 1人欠け→×0.92、3人→×0.76
    data_trust_score = round((6 - missing) / 6 * 100)
    if missing >= 3:
        reasons.append(f"データ不足{missing}人（信頼性低 ×{m6:.2f}）")
    elif missing >= 1:
        reasons.append(f"データ不足{missing}人（×{m6:.2f}）")
        missing_names = "・".join(
            f"{r['waku']}号艇{r['name']}({r.get('missing_reason','')})"
            for r in missing_list
        )
        reasons.append(f"データ信頼度{data_trust_score}%（不足:{missing_names}）")

    # ─── 最終スコア計算 ────────────────────────────────────────────────────
    # base_prob（0〜約1.5）× 修正因子群 を 0〜100 にスケール
    # 基準値（全てデフォルトの場合）: 1.0 × 0.75 × 0.55 × 1.0 ≒ 0.41 → score ≒ 50
    raw_score  = base_prob * m1 * m2 * m3 * m4 * m5 * m6
    # 正規化: raw_score=0.41 → 50点 となるよう線形スケール（slope = 50/0.41 ≒ 122）
    NORMALIZATION_CENTER = 0.41   # 全デフォルト値での期待出力
    score = int(round(min(100, max(0, raw_score / NORMALIZATION_CENTER * 50))))

    reasons.insert(0, f"複合確率スコア: {raw_score:.4f} → {score}点")

    if score >= 80:
        rank, strategy, skip = "S", "◎1着固定 2-3着流し（5〜8点）", False
    elif score >= 60:
        rank, strategy, skip = "A", "◎-○軸 3着3〜4艇流し（10〜15点）", False
    elif score >= 45:
        rank, strategy, skip = "B", "◎○△フォーメーション（15〜20点）", False
    elif score >= 35:
        # 【改善】旧閾値30→35に厳格化。過去バックテストでC帯の回収率が損益分岐を下回った。
        rank, strategy, skip = "C", "荒れ傾向。高配当狙いか見送り", False
    else:
        # score < 35 → 大荒れ確実。見送り推奨。
        rank, strategy, skip = "D", "大荒れ。見送り推奨", True

    return {
        "rank":                rank,
        "score":               score,
        "reason":              reasons,
        "skip":                skip,
        "strategy":            strategy,
        "honmei_concentrated": honmei_concentrated,
        "two_top_race":        two_top_race,
        "data_trust_score":    data_trust_score,
    }


def _calc_3rentan_probs(results):
    """
    全120通りの3連単推定確率を計算して返す。
    計算モデル：
      P(1着=A)         = rel_win1 を正規化
      P(2着=B | 1着=A) = A除外後の残艇で再正規化
                         1号艇1着時はイン逃げ時2着率(circle_pct)で補正
      P(3着=C | 1着=A,2着=B) = A・B除外後の残艇でidx3を正規化
    戻り値: prob降順ソート済みリスト
    """
    wakus = [r["waku"] for r in results]
    if len(wakus) < 3:
        return []

    # 【改善】rel_win1=0 の艇にフロア確率を適用（0%買い目根絶）
    _FLOOR = {"1": 0.5, "2": 0.15, "3": 0.15, "4": 0.10, "5": 0.06, "6": 0.02}
    rel_map  = {r["waku"]: max(r.get("rel_win1") or 0, _FLOOR.get(r["waku"], 0.02)) for r in results}
    circ_map = {r["waku"]: max(r.get("circle_pct") or 0, 0.01) for r in results}
    idx3_map = {r["waku"]: max(r.get("idx3") or r.get("rel_win1") or 0, _FLOOR.get(r["waku"], 0.02)) for r in results}

    total_rel = sum(rel_map.values()) or 1
    p1 = {w: rel_map[w] / total_rel for w in wakus}

    combos = []
    for first in wakus:
        for second in wakus:
            if second == first:
                continue
            for third in wakus:
                if third == first or third == second:
                    continue
                prob_1 = p1[first]

                remaining_2 = [w for w in wakus if w != first]
                if first == "1":
                    w2 = {w: max(circ_map[w], 0.001) for w in remaining_2}
                else:
                    w2 = {w: max(rel_map[w], 0.001) for w in remaining_2}
                prob_2 = w2[second] / sum(w2.values())

                remaining_3 = [w for w in wakus if w != first and w != second]
                w3 = {w: max(idx3_map[w], 0.001) for w in remaining_3}
                prob_3 = w3[third] / sum(w3.values())

                combos.append({
                    "combo":          f"{first}-{second}-{third}",
                    "first":          first,
                    "second":         second,
                    "third":          third,
                    "prob":           prob_1 * prob_2 * prob_3,
                    "is_outer_first": int(first) >= 4,
                })

    combos.sort(key=lambda x: x["prob"], reverse=True)
    return combos


# ════════════════════════════════════════════════════════════════════════
# SCシナリオ（潰れ展開・漁夫の利）計算関数
# ════════════════════════════════════════════════════════════════════════

def _calc_sc_weight(results, cm_map, win3_map, rel_map, jizen_eval=None):
    """
    SCシナリオ（飛び役自滅→漁夫の利）の重みと、
    SC発動時の受益者スコアを計算して返す。

    【競艇物理法則に基づく設計】

    ■ 飛び役の決まり手タイプを「最大決まり手（1位）」で分類
        まくり系   : 最大決まり手が「まくり」または「まくり差し」
        差し系     : 最大決まり手が「差し」
        逃げ系(1号): SCシナリオ対象外（逃げが自滅しても外がそのまま1着になる）

    ■ まくり系飛び役が自滅した場合（外に膨らむ）
        物理法則: まくり艇が旋回で外に膨らむ
                  → コース(飛び役)より内側の艇のコースが開く（内側開放）
                  → コース(飛び役)より外側の艇が続いて捲る（外側継続）
        受益者:
          内側開放候補: コース < 飛び役コース の艇（1号艇除く）
                        ただし1号艇はS1逃げで既に前提計算済みなので2号艇以降
          外側継続候補: コース > 飛び役コース の艇
        重み付け:
          内側開放ボーナス: ×1.5（コースが開く = 進入路が確保される）
          外側継続ボーナス: ×0.8（勢いはあるが距離損）

    ■ 差し系飛び役が自滅した場合（蓋をされる）
        物理法則: 差し艇が1号艇に蓋をされて失速
                  → 後方から来た外側艇（まくり差し）が浮上する
        受益者: コース > 飛び役コース の艇（まくり差しで来た艇）
        重み付け:
          後方浮上ボーナス: ×1.2

    ■ 漁夫スコア計算
        漁夫スコア = 位置ボーナス
                   × (1 - 攻撃性正規化) × 0.5 + 0.5  ← 仕掛けに行かない艇を優遇
                   × win3_rate                          ← 地力（荒れても残る力）

        攻撃性 = (まくり% + まくり差し% + 差し%) / 100
                 高い → 自分も仕掛けに行く → 自滅リスクあり → 漁夫スコア低め
                 低い → 控えて後ろから残る → 漁夫スコア高め

    ■ SCシナリオの1着
        飛び役が自滅 → 1着を取り直すのは:
          パターンA: 1号艇（逃げ取り戻し） ← まくり系飛び役が自滅した場合に多い
          パターンB: 第2飛び役（sub_fly）   ← 差し系飛び役が自滅した場合
          → 両方をSCシナリオ内で重み付き按分

    ■ SCシナリオ重み（S4の代替）
        SC_base = S4_old の重みをベースに、以下で調整:
          飛び役の攻撃性が高い（自滅リスク大）→ SC重みを増やす
          飛び役の実績（win3_rate）が低い      → SC重みを増やす
        SC重み = SC_base × (1 + 飛び役攻撃性 × 0.5) × (1 + (1 - 飛び役win3) × 0.3)

    Returns
    -------
    dict:
        sc_weight       : SCシナリオの全体重み（p_s4の代替として使用）
        sc_1st_weights  : {waku: float} SC時の1着按分重み
        sc_beneficiary  : {waku: float} SC時の2・3着漁夫スコア
        sc_fly_type     : "まくり系" / "差し系" / "不明"
        sc_fly_waku     : 主要飛び役の艇番
    """
    wakus = [r["waku"] for r in results]
    if len(wakus) < 4:
        return {
            "sc_weight": 0.02,
            "sc_1st_weights": {"1": 1.0},
            "sc_beneficiary": {w: 0.1 for w in wakus},
            "sc_fly_type": "不明",
            "sc_fly_waku": None,
        }

    COURSE_NATIONAL_WIN = {"1": 0.555, "2": 0.137, "3": 0.134,
                           "4": 0.111, "5": 0.066, "6": 0.021}

    def safe_pct(cm, key):
        v = cm.get(key)
        try:
            return max(float(v), 0.0) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    # ── 飛び役の特定と決まり手タイプ分類 ──────────────────────────────────────
    # 1号艇以外で最も rel_win1 が高い艇を主要飛び役とする
    fly_candidates = sorted(
        [(r["waku"], rel_map.get(r["waku"], 0)) for r in results if r["waku"] != "1"],
        key=lambda x: x[1], reverse=True
    )
    main_fly_waku = fly_candidates[0][0] if fly_candidates else None
    sub_fly_waku  = fly_candidates[1][0] if len(fly_candidates) >= 2 else None

    if main_fly_waku is None:
        return {
            "sc_weight": 0.02,
            "sc_1st_weights": {"1": 1.0},
            "sc_beneficiary": {w: 0.1 for w in wakus},
            "sc_fly_type": "不明",
            "sc_fly_waku": None,
        }

    # 決まり手タイプ判定（最大決まり手1位で分類）
    fly_cm = cm_map.get(main_fly_waku, {})
    mak_pct     = safe_pct(fly_cm, "まくり%") + safe_pct(fly_cm, "まくり差し%")
    sashi_pct   = safe_pct(fly_cm, "差し%")
    nige_pct    = safe_pct(fly_cm, "逃げ%")

    scores = {"まくり系": mak_pct, "差し系": sashi_pct, "逃げ系": nige_pct}
    sc_fly_type = max(scores, key=scores.get)
    if scores[sc_fly_type] == 0:
        sc_fly_type = "まくり系"  # データなしはまくり系で代替

    # 攻撃性スコア（高いほど仕掛けに行く→自滅リスクあり）
    fly_attack = min((mak_pct + sashi_pct) / 100.0, 1.0)
    fly_win3   = win3_map.get(main_fly_waku, 0.5)

    # ── SCシナリオ重み ──────────────────────────────────────────────────────
    # 旧S4相当のベース（5-6号艇の全国平均微小値）を飛び役の特性で増幅
    sc_base = (
        COURSE_NATIONAL_WIN.get("5", 0.066) * 0.5 +
        COURSE_NATIONAL_WIN.get("6", 0.021) * 0.5
    )
    sc_weight = sc_base * (1.0 + fly_attack * 0.5) * (1.0 + (1.0 - fly_win3) * 0.3)

    # ── SC発動時の1着按分 ──────────────────────────────────────────────────
    # まくり系自滅 → 1号艇が取り戻す確率が高い（コースが開く）
    # 差し系自滅   → 第2飛び役が続く確率が高い
    if sc_fly_type == "まくり系":
        w1_1st = 0.65   # 1号艇が逃げ取り戻し
        w2_1st = 0.35   # 第2飛び役
    elif sc_fly_type == "差し系":
        w1_1st = 0.30   # 差し自滅では1号艇も蓋の影響を受ける
        w2_1st = 0.70   # 第2飛び役が浮上
    else:
        w1_1st = 0.50
        w2_1st = 0.50

    sc_1st_weights = {}
    sc_1st_weights["1"] = w1_1st
    if sub_fly_waku:
        sc_1st_weights[sub_fly_waku] = w2_1st
    else:
        sc_1st_weights["1"] = 1.0  # 第2飛び役なければ1号艇に集約

    # ── 漁夫スコア計算 ──────────────────────────────────────────────────────
    try:
        main_fly_course = int(main_fly_waku)
    except (ValueError, TypeError):
        main_fly_course = 3

    sc_beneficiary = {}
    for r in results:
        w = r["waku"]
        if w == main_fly_waku:
            # 自滅した飛び役自身は漁夫スコア0
            sc_beneficiary[w] = 0.0
            continue

        try:
            w_course = int(w)
        except (ValueError, TypeError):
            w_course = 3

        # 位置ボーナス（コース物理法則）
        if sc_fly_type == "まくり系":
            if w == "1":
                # 1号艇はSC時の1着候補なので2・3着漁夫スコアは低め
                pos_bonus = 0.6
            elif w_course < main_fly_course:
                # 内側開放候補（まくり艇が膨らんでコースが開く）
                pos_bonus = 1.5
            else:
                # 外側継続候補（まくり艇に続いて捲る）
                pos_bonus = 0.8
        elif sc_fly_type == "差し系":
            if w_course > main_fly_course:
                # 後方浮上候補（まくり差しで来た艇）
                pos_bonus = 1.2
            elif w == "1":
                # 蓋をした1号艇は先に行っているが差し系なので若干恩恵あり
                pos_bonus = 0.9
            else:
                pos_bonus = 0.7
        else:
            pos_bonus = 1.0

        # 攻撃性の逆数（仕掛けに行かない艇 = 漁夫スコア高め）
        w_cm      = cm_map.get(w, {})
        w_mak     = safe_pct(w_cm, "まくり%") + safe_pct(w_cm, "まくり差し%")
        w_sashi   = safe_pct(w_cm, "差し%")
        w_attack  = min((w_mak + w_sashi) / 100.0, 1.0)
        passivity = (1.0 - w_attack) * 0.5 + 0.5  # 0.5〜1.0 の範囲

        # 地力（荒れた展開でも着に残る力）
        ground    = win3_map.get(w, 0.5)

        sc_beneficiary[w] = pos_bonus * passivity * ground

    return {
        "sc_weight":      sc_weight,
        "sc_1st_weights": sc_1st_weights,
        "sc_beneficiary": sc_beneficiary,
        "sc_fly_type":    sc_fly_type,
        "sc_fly_waku":    main_fly_waku,
    }


def _calc_3rentan_probs_v2(results, venue_course_1c_rate=None, jizen_eval=None):
    """
    【改善②】展開シナリオ×条件付き確率モデル

    旧方式の問題：各艇が独立に競争すると仮定 → 展開の依存関係を無視
    　例）1号艇が逃げた時とまくりが決まった時では2着・3着の確率分布が全く異なる

    新方式：4つの展開シナリオをまず確率決定し、
            シナリオ内で条件付き確率（P(2着|シナリオ)、P(3着|シナリオ,2着)）を計算する。

    ─────────────────────────────────────────────────────────────
    シナリオ定義：
      S1: イン逃げ   … 1号艇が1着（コース1の逃げ決まり手%で推定）
      S2: 差し       … 内側艇(2-3)が差して1着
      S3: まくり系   … 外側艇(3-5)がまくり/まくり差しで1着
      S4: 大荒れ     … 5-6号艇1着（確率は小さいが高配当に寄与）

    各シナリオの重み：
      S1 ∝ 1号艇の逃げ%（コース別マスタ）× 会場イン逃げ率補正
      S2 ∝ 2-3号艇の差し% × rel_win1
      S3 ∝ 3-5号艇のまくり% × rel_win1
      S4 ∝ 5-6号艇の rel_win1（極小）

    条件付き2着確率（シナリオごとに異なる）：
      S1（イン逃げ時）: circle_pct（イン逃げ時2着率）を使用
      S2/S3/S4 時:      1着艇を除外した残艇の rel_win1 で再正規化
                        ただし「まくられた艇（内側）」は3着以内に残りにくいため
                        コース1の3連対率で補正

    ─────────────────────────────────────────────────────────────
    戻り値: prob降順ソート済み 全120通りの combinationリスト
    """
    wakus = [r["waku"] for r in results]
    if len(wakus) < 3:
        return []

    # ── ルックアップテーブル構築 ──
    # 【改善】rel_win1=0 の艇にコース別フロア確率を適用（0%買い目根絶）
    # Laplace smoothingフロア: calc_race_indices で既に適用済みだが
    # 外部calc(_ext_calc_3rentan)経由や直接呼び出し時にも保護する。
    _RELWIN_FLOOR = {"1": 0.5, "2": 0.15, "3": 0.15, "4": 0.10, "5": 0.06, "6": 0.02}
    rel_map   = {r["waku"]: max(r.get("rel_win1") or 0,
                                _RELWIN_FLOOR.get(r["waku"], 0.02)) for r in results}
    # 【修正①】_circ_raw（正規化前絶対スコア）を参照して二重正規化を解消
    # circle_pct は表示用相対%（合計100%）のため確率計算では使用しない。
    circ_map  = {r["waku"]: max(
        r.get("_circ_raw") if r.get("_circ_raw") is not None
        else (r.get("circle_pct") or 0) / 100.0, 0
    ) for r in results}
    idx3_map  = {r["waku"]: max(r.get("idx3")  or r.get("rel_win1") or 0, 0) for r in results}
    win3_map  = {r["waku"]: r.get("win3_rate") or 0.5 for r in results}
    cm_map    = {r["waku"]: r.get("raw_cm", {}) for r in results}
    avg_st_map = {r["waku"]: r.get("avg_st") for r in results}

    def safe_pct(cm, key):
        v = safe_float(_get_cm_val(cm, key))
        return max(v or 0.0, 0.0)

    # ──────────────────────────────────────────────────────────────────────────
    # 【修正⑤】メンバー相互作用を反映したシナリオ重み計算
    # ──────────────────────────────────────────────────────────────────────────
    # 旧問題: 各艇の決まり手%を独立に合算 → 「誰が誰に対してどう動くか」が無視される
    #   例①) 2号艇の差し%が高くても、3〜4号艇のまくりが速ければS2は潰される
    #   例②) 1号艇のSTが遅くても S1重みは下がらなかった
    #   例③) 今節成績・モーター調子が確率計算に全く反映されていなかった
    #
    # 新方式: 以下3つの相互作用補正を追加する。
    #
    #   補正A ── STの相対関係によるS1重み調整
    #     1号艇と2号艇の平均STを比較し、2号艇が速い（先マイ状態）ならS1を減衰させる。
    #     ST差 ≥ +0.03秒（2号艇が速い）: S1 × 0.75
    #     ST差 ≤ -0.03秒（1号艇が速い）: S1 × 1.15（上限補正）
    #
    #   補正B ── 内側艇の差し力によるS3重み抑制（まくり潰し確率）
    #     2号艇の差し%が高いほど、外側艇のまくりは潰されやすい。
    #     まくり潰し係数 = 1.0 - 0.5 × (2号艇差し% / 全国平均差し率の2倍)
    #     ただし係数の下限は0.5（完全には潰れない）
    #
    #   補正C ── 今節成績・モーター調子によるシナリオ重み調整
    #     今節成績: "1-1-2"のような文字列を解析し、直近の1着数が多いほど
    #               その艇が関与するシナリオ重みを最大+20%上昇させる
    #     モーター2連率: 全艇平均と比較し、突出した艇はシナリオ関与重みを補正する
    #     ※ 補正Cは各シナリオの基礎重みに乗算するスケーラーとして適用
    # ──────────────────────────────────────────────────────────────────────────
    COURSE_NATIONAL_WIN = {"1": 0.555, "2": 0.137, "3": 0.134,
                           "4": 0.111, "5": 0.066, "6": 0.021}
    # 会場別1コース1着率があれば使用（荒れやすい会場の過信補正）
    # 例: 戸田=0.430、平和島=0.446 → 全国平均0.555より大幅に低い
    _c1_win_base = float(venue_course_1c_rate) if venue_course_1c_rate is not None else COURSE_NATIONAL_WIN["1"]

    # ── 補正C用: 今節成績スコアとモーター調子スコアを艇番ごとに算出 ──────────
    def _kosetsu_score(kosetsu_str):
        """
        今節成績文字列（例: "1-2-3" "1-1-F" "2-1-1"）から1着率相当のスコアを返す。
        形式は「着順-着順-着順...」を想定。1着=1.0、2着=0.5、3着=0.25、それ以外=0。
        データなし or 解析不能の場合は None を返す。
        """
        if not kosetsu_str or kosetsu_str in ("", "None", "nan", "-"):
            return None
        scores = []
        for token in re.split(r"[-・/]", kosetsu_str):
            token = token.strip()
            if token == "1":
                scores.append(1.0)
            elif token == "2":
                scores.append(0.5)
            elif token == "3":
                scores.append(0.25)
            elif token.isdigit():
                scores.append(0.0)
            # F/L/S/K等の失格・欠場は無視（スコアに含めない）
        return sum(scores) / len(scores) if scores else None

    kosetsu_score_map = {}
    for r in results:
        sc = _kosetsu_score(str(r.get("kosetsu", "")))
        kosetsu_score_map[r["waku"]] = sc  # None の場合は補正しない

    motor2_map = {}
    for r in results:
        v = safe_float(r.get("motor2"))
        motor2_map[r["waku"]] = v  # None の場合は補正しない

    # モーター2連率の全艇平均（Noneを除く）
    valid_motor2 = [v for v in motor2_map.values() if v is not None]
    motor2_mean = sum(valid_motor2) / len(valid_motor2) if valid_motor2 else None

    def _member_scenario_scale(waku, base_kosetsu_weight=0.15, base_motor_weight=0.10):
        """
        今節成績スコアとモーター調子から、その艇が関与するシナリオへの乗算スケーラーを返す。
        スケーラー = 1.0 ± 補正値（範囲: 0.80 〜 1.20）
        """
        scale = 1.0
        ks = kosetsu_score_map.get(waku)
        if ks is not None:
            # 今節成績スコア 0〜1 を -0.15〜+0.15 の補正に変換（中心0.5が±0）
            scale += (ks - 0.5) * 2 * base_kosetsu_weight
        m2 = motor2_map.get(waku)
        if m2 is not None and motor2_mean is not None and motor2_mean > 0:
            # モーター2連率が平均比で高い/低いほど補正
            ratio = (m2 - motor2_mean) / motor2_mean  # -1〜+1程度
            scale += ratio * base_motor_weight
        return max(0.80, min(1.20, scale))

    # ── 補正A: STの相対関係によるS1重み調整係数 ──────────────────────────────
    st1 = avg_st_map.get("1")
    st2 = avg_st_map.get("2")
    st_adj_s1 = 1.0  # デフォルト: 補正なし
    if st1 is not None and st2 is not None:
        st_diff = st1 - st2  # 正 → 1号艇が遅い、負 → 1号艇が速い
        if st_diff >= 0.03:
            # 2号艇の方が速い（先マイになりやすい）→ S1を減衰
            st_adj_s1 = max(0.75, 1.0 - st_diff * 5.0)
        elif st_diff <= -0.03:
            # 1号艇の方が速い（楽逃げ）→ S1を増強（上限1.20）
            st_adj_s1 = min(1.20, 1.0 + abs(st_diff) * 3.0)

    # ── 【修正⑤】S1基礎重み: rel_win1が既に会場特性を織り込み済みのため二重補正を除去 ──
    # 【旧問題】s1_base = _c1_win_base × (0.5 + w1_nige) としていたが、
    #   rel_win1 の計算時点で既に venue_rate（会場別1コース1着率）が w_venue 分混入している。
    #   そこへさらに _c1_win_base を掛けると「会場特性が二重に効く」構造になっていた。
    #   特に荒れやすい会場（戸田・平和島等）では rel_win1 が既に低めになっているのに
    #   _c1_win_base でさらに下げてしまい、1号艇を過剰に不利評価していた。
    #
    # 【修正③】S1重みの二重補正を解消: nige_adj を乗算からブレンド補正に変更。
    #   旧方式: s1 = rel_win1 × (0.5 + nige%) × ST補正 — 逃げ%が rel_win1 に既に含まれている
    #   新方式: s1 = nige_blend × ST補正 × member補正
    #     nige_blend = 個人逃げ% × 0.6 + 会場1コース1着率 × 0.4
    #     → 個人実績と会場特性を直接ブレンドし、rel_win1 による二重乗算を排除。
    #     → 逃げ%データなし時は会場1着率100%にフォールバック。
    w1_nige      = safe_pct(cm_map.get("1", {}), "逃げ%")
    if w1_nige > 0:
        nige_blend = w1_nige * 0.6 + _c1_win_base * 0.4
    else:
        nige_blend = _c1_win_base  # データなし → 会場特性のみ
    s1_weight = nige_blend * st_adj_s1 * _member_scenario_scale("1")

    # ── 【修正②】事前評価（jizen_eval）による追加補正 ────────────────────────
    # evaluate_jizen.evaluate_all() の結果（◎○△）をシナリオ重みに反映する。
    # ①イン逃げ評価: S1の補正（◎→+15%、○→+7%、空白→-10%）
    # ②相性評価:     S2/S3 の按分を相性◎の艇を優遇
    # 補正幅は ±15%以内に抑えて過剰な振れを防ぐ。
    _jizen_symbol_s1 = ""
    _jizen_aisho = {}  # {waku_str: symbol}
    if jizen_eval is not None:
        try:
            _jizen_symbol_s1 = jizen_eval.get("in_nige", [""])[0]  # 1号艇のイン逃げ評価
            for idx, sym in enumerate(jizen_eval.get("aisho", [])):
                _jizen_aisho[str(idx + 1)] = sym
        except Exception:
            pass

    _s1_jizen_adj = {"◎": 1.15, "◎?": 1.10, "○": 1.07, "△": 1.0, "": 0.90}.get(
        _jizen_symbol_s1, 1.0
    )
    s1_weight *= _s1_jizen_adj

    def _jizen_aisho_scale(waku):
        """相性評価記号をシナリオ重み乗数に変換（S2/S3の1着按分で使用）"""
        sym = _jizen_aisho.get(str(waku), "")
        return {"◎": 1.15, "○": 1.07, "△": 1.0, "": 0.95}.get(sym, 1.0)

    # ── 補正B: 2号艇の差し力によるS3抑制係数（まくり潰し） ─────────────────────
    sashi_pct_2   = safe_pct(cm_map.get("2", {}), "差し%")
    nat_sashi_avg = COURSE_NATIONAL_WIN["2"] * 0.5   # 全国平均の差し率代替値（約0.07）
    if sashi_pct_2 > 0:
        makuri_suppress = max(0.50, 1.0 - 0.5 * (sashi_pct_2 / (nat_sashi_avg * 2 + 1e-6)))
    else:
        makuri_suppress = 1.0  # 差し%データなし → 補正しない

    # ── S2基礎重み（差し系: 2-3コースの差し%合算）+ 相性補正 ────────────────────
    s2_weight = 0.0
    for w in ["2", "3"]:
        pct_sashi = safe_pct(cm_map.get(w, {}), "差し%")
        nat_avg   = COURSE_NATIONAL_WIN.get(w, 0.1)
        base      = pct_sashi if pct_sashi > 0 else nat_avg * 0.3
        # 【修正②】相性評価スケールを乗算（◎の艇のS2寄与を増幅）
        s2_weight += base * _member_scenario_scale(w) * _jizen_aisho_scale(w)

    # ── 【修正④】S3基礎重み: 対象を3-6号艇に拡張（6号艇は距離減衰0.5）＋ 補正B適用 ──
    # 旧: 3-5号艇のみ → 荒れやすい会場での6コースまくりを完全無視していた問題を解消。
    # 新: 6号艇も対象に追加（重み係数 0.5 で距離減衰を表現）。
    # まくり差し得意な6号艇が存在する会場（戸田・平和島等）で取りこぼしを防ぐ。
    s3_weight = 0.0
    outer_mak_weights = {"3": 1.0, "4": 1.0, "5": 0.8, "6": 0.5}
    for w, outer_scale in outer_mak_weights.items():
        pct_mak = safe_pct(cm_map.get(w, {}), "まくり%") + \
                  safe_pct(cm_map.get(w, {}), "まくり差し%")
        nat_avg = COURSE_NATIONAL_WIN.get(w, 0.05)
        base    = pct_mak if pct_mak > 0 else nat_avg * 0.4
        # 【修正②】相性評価スケールを乗算
        s3_weight += base * outer_scale * _member_scenario_scale(w) * _jizen_aisho_scale(w)
    s3_weight *= makuri_suppress  # 補正B: 2号艇の差し力でまくりを抑制

    # ── SCシナリオ重み（S4「大荒れ」を「潰れ展開・漁夫の利」SCに差し替え） ────────
    # 旧S4は「何が起きるかわからない雑な確率」だった。
    # SCは「飛び役が自滅したとき誰が漁夫の利を取るか」という具体的な展開を表す。
    # 確率の合計は変わらず、展開の解像度だけ上がる。
    _sc_info = _calc_sc_weight(
        results, cm_map, win3_map, rel_map, jizen_eval=jizen_eval
    )
    sc_weight         = _sc_info["sc_weight"]
    _sc_1st_weights   = _sc_info["sc_1st_weights"]    # {waku: float} SC時1着按分
    _sc_beneficiary   = _sc_info["sc_beneficiary"]    # {waku: float} 漁夫スコア
    _sc_fly_type      = _sc_info["sc_fly_type"]       # "まくり系"/"差し系"/"不明"

    total_s = s1_weight + s2_weight + s3_weight + sc_weight
    if total_s <= 0:
        total_s = 1.0
    p_s1 = s1_weight / total_s
    p_s2 = s2_weight / total_s
    p_s3 = s3_weight / total_s
    p_sc = sc_weight / total_s

    # ── 条件付き1着確率（シナリオ × 1着艇）──
    # S2/S3は決まり手%が主軸
    # SC は _calc_sc_weight が算出した1着按分重みを使用
    def prob_first_given_scenario(scenario, first_w):
        """P(1着=first_w | scenario)"""
        if scenario == "S1":
            # イン逃げシナリオ → 1号艇固定
            return 1.0 if first_w == "1" else 0.0
        elif scenario == "S2":
            # 差しシナリオ → 2-3号艇の差し%で按分（データなし時は全国平均で代替）
            candidates = {}
            for w in ["2", "3"]:
                pct = safe_pct(cm_map.get(w, {}), "差し%")
                candidates[w] = pct if pct > 0 else COURSE_NATIONAL_WIN.get(w, 0.1) * 0.3
            total = sum(candidates.values()) or 1
            return candidates.get(first_w, 0) / total
        elif scenario == "S3":
            # まくりシナリオ → 3-6号艇のまくり%+まくり差し%で按分
            # 6号艇は距離減衰係数0.5を適用（外枠ほど決まりにくい現実を反映）
            outer_scale = {"3": 1.0, "4": 1.0, "5": 0.8, "6": 0.5}
            candidates = {}
            for w, sc in outer_scale.items():
                pct = safe_pct(cm_map.get(w, {}), "まくり%") + \
                      safe_pct(cm_map.get(w, {}), "まくり差し%")
                candidates[w] = (pct if pct > 0 else COURSE_NATIONAL_WIN.get(w, 0.04) * 0.4) * sc
            total = sum(candidates.values()) or 1
            return candidates.get(first_w, 0) / total
        else:  # SC（潰れ展開）
            # 飛び役自滅後の1着は _calc_sc_weight が算出した按分重みで決まる
            total = sum(_sc_1st_weights.values()) or 1
            return _sc_1st_weights.get(first_w, 0) / total

    # ── 条件付き2着確率 P(2着=B | 1着=A, scenario) ──
    def prob_second(scenario, first_w, second_w, remaining):
        if scenario == "S1":
            circ_rem = {w: max(circ_map.get(w, 0), 0.001) for w in remaining}
            total = sum(circ_rem.values()) or 1
            return circ_rem.get(second_w, 0.001) / total
        elif scenario == "SC":
            # 潰れ展開の2着 → 漁夫スコアで按分
            # 自滅した主要飛び役は2着にも残りにくい（確率を0.2に減衰）
            fly_waku = _sc_info.get("sc_fly_waku")
            w2 = {}
            for w in remaining:
                score = max(_sc_beneficiary.get(w, 0.1), 0.001)
                if w == fly_waku:
                    score *= 0.2
                w2[w] = score
            total = sum(w2.values()) or 1
            return w2.get(second_w, 0.001) / total
        else:
            # ── 内側残存補正テーブル（対話ログ合意版）────────────────────────
            # 競艇の物理法則: まくり・差しが決まったとき
            #   ・1着艇より内側の艇 → 圧縮されて2着に残りやすい（×1.05〜1.35）
            #   ・1着艇より外側の艇 → 競り合いで2着に来にくい（×0.75〜0.95）
            # S1（逃げ）は circle_pct を使うため対象外。SCは漁夫スコアを使うため対象外。
            _INNER_REMAIN = {
                "2": {"1": 1.35, "3": 0.75, "4": 0.90, "5": 0.95, "6": 0.95},
                "3": {"1": 1.25, "2": 1.20, "4": 0.85, "5": 0.90, "6": 0.90},
                "4": {"1": 1.20, "2": 1.15, "3": 1.10, "5": 0.80, "6": 0.85},
                "5": {"1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "6": 0.75},
                "6": {"1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "5": 0.95},
            }
            try:
                first_course = int(first_w)
            except (ValueError, TypeError):
                first_course = 3
            pos_tbl = _INNER_REMAIN.get(first_w, {})
            w2 = {}
            for w in remaining:
                base = max(rel_map.get(w, 0), 0.001)
                try:
                    w_course = int(w)
                except (ValueError, TypeError):
                    w_course = 3
                # S3（まくり系）: 旧来の被まくり補正を内側残存補正に統合
                # S2（差し系）: 内側残存補正のみ適用（旧補正なし → 新規追加）
                inner_remain = pos_tbl.get(w, 1.0)
                base *= inner_remain
                w2[w] = base
            total = sum(w2.values()) or 1
            return w2.get(second_w, 0.001) / total

    # ── 条件付き3着確率 P(3着=C | 1着=A, 2着=B, scenario) ──
    #
    # S1（イン逃げ）:   idx3（イン逃げ時3着指数）を使用
    # S2（差し）:       win3_rate(0.6) + rel_win1(0.4) の加重平均
    # S3（まくり系）:   内側艇は win3_rate で下方補正（弾かれやすい）
    # SC（潰れ展開）:   漁夫スコア（_sc_beneficiary）を使用
    #                   位置ボーナス × 攻撃性逆数 × win3_rate の合成値
    def prob_third(scenario, first_w, second_w, third_w, remaining):
        if scenario == "S1":
            w3 = {w: max(idx3_map.get(w, 0), 0.001) for w in remaining}

        elif scenario == "S2":
            # S2（差し系）: 3着も内側残存補正を適用
            _INNER_REMAIN3 = {
                "2": {"1": 1.35, "3": 0.75, "4": 0.90, "5": 0.95, "6": 0.95},
                "3": {"1": 1.25, "2": 1.20, "4": 0.85, "5": 0.90, "6": 0.90},
                "4": {"1": 1.20, "2": 1.15, "3": 1.10, "5": 0.80, "6": 0.85},
                "5": {"1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "6": 0.75},
                "6": {"1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "5": 0.95},
            }
            pos_tbl3 = _INNER_REMAIN3.get(first_w, {})
            w3 = {}
            for w in remaining:
                win3  = max(win3_map.get(w, 0.5), 0.001)
                rel   = max(rel_map.get(w, 0),    0.001)
                base  = win3 * 0.6 + rel * 0.4
                base *= pos_tbl3.get(w, 1.0)
                w3[w] = base

        elif scenario == "S3":
            # S3（まくり系）: 内側残存補正に統一（旧来の被まくり×win3補正を廃止）
            _INNER_REMAIN3 = {
                "2": {"1": 1.35, "3": 0.75, "4": 0.90, "5": 0.95, "6": 0.95},
                "3": {"1": 1.25, "2": 1.20, "4": 0.85, "5": 0.90, "6": 0.90},
                "4": {"1": 1.20, "2": 1.15, "3": 1.10, "5": 0.80, "6": 0.85},
                "5": {"1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "6": 0.75},
                "6": {"1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "5": 0.95},
            }
            pos_tbl3 = _INNER_REMAIN3.get(first_w, {})
            w3 = {}
            for w in remaining:
                base = max(win3_map.get(w, 0.5), 0.001)
                base *= pos_tbl3.get(w, 1.0)
                w3[w] = base

        else:  # SC（潰れ展開）
            # 漁夫スコアで3着候補を評価
            # 自滅した飛び役・1着艇・2着艇以外の残存艇を漁夫スコアで按分
            w3 = {w: max(_sc_beneficiary.get(w, 0.1), 0.001) for w in remaining}

        total = sum(w3.values()) or 1
        return w3.get(third_w, 0.001) / total

    # ── 全120通りの確率を計算 ──
    combos_dict = {}
    scenarios = [("S1", p_s1), ("S2", p_s2), ("S3", p_s3), ("SC", p_sc)]

    for first in wakus:
        for second in wakus:
            if second == first:
                continue
            for third in wakus:
                if third == first or third == second:
                    continue
                key = f"{first}-{second}-{third}"
                prob_total = 0.0
                for scenario, p_sc in scenarios:
                    if p_sc < 1e-9:
                        continue
                    p1_sc = prob_first_given_scenario(scenario, first)
                    if p1_sc < 1e-9:
                        continue
                    rem2 = [w for w in wakus if w != first]
                    p2_sc = prob_second(scenario, first, second, rem2)
                    rem3 = [w for w in wakus if w != first and w != second]
                    p3_sc = prob_third(scenario, first, second, third, rem3)
                    prob_total += p_sc * p1_sc * p2_sc * p3_sc

                combos_dict[key] = {
                    "combo":          key,
                    "first":          first,
                    "second":         second,
                    "third":          third,
                    "prob":           prob_total,
                    "is_outer_first": int(first) >= 4,
                    "top_scenario":   max(
                        scenarios,
                        key=lambda sc: sc[1] * prob_first_given_scenario(sc[0], first)
                    )[0],
                    # SCシナリオ情報（数値シート表示用）
                    "sc_fly_type":    _sc_fly_type,
                    "sc_beneficiary": _sc_beneficiary.get(second, 0),
                }

    # ── 理論オッズ・ハイブリッドスコアを付与 ──
    # 理論オッズ = (1 / prob) × (1 - テラ銭0.25)
    # ハイブリッドスコア = 確率60% + 理論オッズ40% の合成ランク
    # → 低オッズ本命への過集中を防ぎ、中〜高配当の期待値が高い組み合わせを優先
    TAKER_RATE  = 0.25
    ODDS_WEIGHT = 0.4  # 0.0=純確率順 / 1.0=純オッズ順 / 0.4=バランス重視

    for c in combos_dict.values():
        p = max(c["prob"], 1e-6)
        c["theoretical_odds"] = round((1.0 / p) * (1.0 - TAKER_RATE), 1)
        c["prob_rank"]  = 0
        c["odds_rank"]  = 0

    combos_sorted_prob = sorted(combos_dict.values(), key=lambda x: x["prob"],             reverse=True)
    combos_sorted_odds = sorted(combos_dict.values(), key=lambda x: x["theoretical_odds"], reverse=True)

    for i, c in enumerate(combos_sorted_prob): c["prob_rank"]  = i + 1
    for i, c in enumerate(combos_sorted_odds): c["odds_rank"]  = i + 1

    for c in combos_dict.values():
        c["hybrid_score"] = (1.0 - ODDS_WEIGHT) * (1.0 / c["prob_rank"]) + \
                             ODDS_WEIGHT         * (1.0 / c["odds_rank"])

    combos = sorted(combos_dict.values(), key=lambda x: x["hybrid_score"], reverse=True)
    return combos


# ─────────────────────────────────────────────────────────────────────────────
# 期待値計算（実際のオッズExcelとの照合用）
# 使い方:
#   combos = _calc_3rentan_probs_v2(results)
#   actual = load_actual_odds_from_excel("odds_YYYYMMDD_R1.xlsx")  # 別途実装
#   ev_list = calc_ev_from_actual_odds(combos, actual)
#   suggestion = suggest_by_ev(ev_list, min_ev=0.05)  # EV+5%以上のみ
# ─────────────────────────────────────────────────────────────────────────────

def calc_ev_from_actual_odds(combos, actual_odds_dict):
    """
    実際の払戻オッズと推定確率から期待値（EV）を計算する。

    Parameters
    ----------
    combos : list[dict]
        _calc_3rentan_probs_v2() の戻り値。各要素に "combo"（例: "1-2-3"）と "prob" が必要。
    actual_odds_dict : dict
        {"1→2→3": 払戻倍率float, ...} 形式。
        ※ 理論オッズExcelから読んだ全120通りの実際の払戻倍率を渡す。

    Returns
    -------
    list[dict]
        各コンボに以下を追加した辞書のリスト（EV降順）:
          actual_odds : 実際の払戻倍率（Noneなら未取得）
          ev          : 期待値 = actual_odds × prob - 1.0（Noneなら計算不可）
          ev_pct      : EVを%表示した文字列（例: "+12.3%"）
          ev_positive : EVがプラスかどうか
    """
    result = []
    for c in combos:
        key    = c["combo"]   # "1-2-3" 形式
        actual = actual_odds_dict.get(key)
        c2     = dict(c)
        if actual is not None and c.get("prob", 0) > 0:
            ev = actual * c["prob"] - 1.0
            c2["actual_odds"] = round(actual, 1)
            c2["ev"]          = round(ev, 4)
            c2["ev_pct"]      = f"{'+' if ev >= 0 else ''}{ev*100:.1f}%"
            c2["ev_positive"] = ev > 0
        else:
            c2["actual_odds"] = None
            c2["ev"]          = None
            c2["ev_pct"]      = "N/A"
            c2["ev_positive"] = False
        result.append(c2)
    # EV降順（未取得はそれ以下に）
    result.sort(key=lambda x: (x["ev"] if x["ev"] is not None else -999), reverse=True)
    return result


def suggest_by_ev(combos_with_ev, min_ev=0.0, max_bets=8):
    """
    期待値プラスの組み合わせのみを買い目として返す。

    Parameters
    ----------
    combos_with_ev : list[dict]  calc_ev_from_actual_odds() の戻り値
    min_ev         : float       最低期待値閾値（デフォルト0.0 = 期待値プラスのみ）
                                 例: 0.05 → EV+5%以上のみ対象
    max_bets       : int         最大買い目数

    Returns
    -------
    dict
        buy_list    : 買い目リスト（EV降順）、例: ["2→3→1", "1→3→2"]
        ev_summary  : 各買い目の詳細（combo / prob% / actual_odds / ev%）
        total_bets  : 点数
        best_ev     : 最高EVの組み合わせ
        skip        : True なら期待値プラスがゼロ → 見送り推奨
        reason      : 判定理由の説明文
    """
    positives = [c for c in combos_with_ev
                 if c.get("ev") is not None and c["ev"] > min_ev]
    positives = positives[:max_bets]

    if not positives:
        # EV最高値を探して理由を付ける
        best_any = next((c for c in combos_with_ev if c.get("ev") is not None), None)
        best_str = f"（最高EV: {best_any['ev_pct']} {best_any['combo']}）" if best_any else ""
        return {
            "buy_list":   [],
            "ev_summary": [],
            "total_bets": 0,
            "best_ev":    None,
            "skip":       True,
            "reason":     f"EV>{min_ev*100:.0f}%の組み合わせなし → 見送り推奨{best_str}"
        }

    return {
        "buy_list": [c["combo"] for c in positives],
        "ev_summary": [
            {
                "combo":        c["combo"],
                "prob_pct":     f"{c['prob']*100:.2f}%",
                "actual_odds":  c["actual_odds"],
                "ev_pct":       c["ev_pct"],
            }
            for c in positives
        ],
        "total_bets": len(positives),
        "best_ev":    positives[0],
        "skip":       False,
        "reason":     f"EV>{min_ev*100:.0f}%が{len(positives)}点（最高: {positives[0]['ev_pct']} {positives[0]['combo']}）"
    }


def load_actual_odds_from_excel(filepath, sheet_name=0):
    """
    理論オッズExcel（全120通り）を読み込んで辞書形式で返す。

    Excel形式（理想オッズ完成版.xlsx 準拠）:
      - 列A: 1着艇番（int）
      - 列B: 2着艇番（int）
      - 列C: 3着艇番（int）
      - 列D: オッズ（float）
      - 1行目はヘッダ（読み飛ばし）
      - データ120行（3連単全通り）

    Parameters
    ----------
    filepath   : str  Excelファイルパス（例: "odds/理想オッズ_20240101_R1.xlsx"）
    sheet_name : int or str  シート番号またはシート名（デフォルト0）

    Returns
    -------
    dict  {"1→2→3": float, ...}  キーは "A→B→C" 形式
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    result = {}
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if row[0] is None:
            continue
        try:
            key = f"{int(row[0])}-{int(row[1])}-{int(row[2])}"
            result[key] = float(row[3])
        except (TypeError, ValueError):
            continue
    wb.close()
    return result


# ============================================================
# 参加見送り判定（バックテスト結果に基づく精度向上フィルタ）
# ============================================================
# 【根拠】バックテスト8,458R分析（2025-12〜2026-02）
#   逃げシナリオ:  4,627R / ROI 26.8% / 損益 −376万  ← 全損失の153%
#   逃げ以外:      3,831R / ROI 130.9%/ 損益 +131万
#
#   施策シミュレーション結果:
#     ①逃げ会場フィルタのみ   → ROI 111.4% / +62万  (現状比 +307万)
#     ①+②1〜3R見送り         → ROI 122.1% / +87万  (現状比 +332万)
#     ①+②+③逃げ5点上限      → ROI 132.5% / +112万 (現状比 +357万)
#
# 逃げシナリオ参加許可会場（BT実績ROI 30%以上の会場のみ）:
#   宮島(59.2%), 江戸川(45.6%), 下関(37.1%), 浜名湖(33.2%), 多摩川(30.9%)
#
# ※ _SKIP_VENUES は後方互換のため残す（旧・全シナリオ会場フィルタ）
# ============================================================
_SKIP_VENUES = set()  # 旧・全シナリオ会場フィルタ → 現在は無効化（逃げ専用フィルタに移行）

# 逃げシナリオで参加を許可する会場（BT実績ROI30%以上のみ）
# ここに入っていない会場の「逃げ軸流し」は自動見送り
_NIGE_ALLOWED_VENUES = {
    '宮島',   # ROI 59.2%
    '江戸川', # ROI 45.6%
    '下関',   # ROI 37.1%
    '浜名湖', # ROI 33.2%
    '多摩川', # ROI 30.9%
}

# 逃げシナリオの最大買い目点数（BT分析：点数を絞っても的中率は変わらないため
# 投資額削減を優先。上位5点に絞ることで回収率が改善）
_NIGE_MAX_BETS = 5

def _should_skip_race(bet_suggestions: dict, venue: str = "") -> tuple[bool, str]:
    """
    参加見送りを判定する。
    Returns: (skip: bool, reason: str)

    【v3 設計思想】
    「逃げ軸流し = 見送り」というルールを廃止。
    逃げ軸流しは展開予測であって見送り理由ではない。

    見送り条件（構造的に回収できないケースのみ）:
      ① s1_prob >= 0.65 かつ himo_are が不参加推奨
         → 逃げほぼ確定 + ヒモ固まり = 1-2-3が低オッズ化確実
      ② s1_prob >= 0.72（逃げ確率が極端に高い）
         → 逃げが当たっても低オッズ = 構造的に回収不能
      ③ honmei_prob_mismatch == True かつ s1_prob >= 0.60
         → 印↔確率が大きく乖離 + 逃げ優位 = 判断根拠が不明確
      ④ venue が低回収会場
    """
    s1_prob         = bet_suggestions.get("s1_prob", 0) or 0
    mismatch        = bet_suggestions.get("honmei_prob_mismatch", False)
    mismatch_detail = bet_suggestions.get("honmei_prob_mismatch_detail", "")

    # ① 逃げ確率高 + ヒモ固まり（最も確実な見送り根拠）
    himo_are = bet_suggestions.get("himo_are", {}) or {}
    himo_verdict = himo_are.get("verdict", "対象外")
    if s1_prob >= 0.65 and himo_verdict == "不参加推奨":
        est_odds = himo_are.get("est_top_odds", 0) or 0
        return True, (
            f"⛔見送り推奨（逃げ{s1_prob*100:.0f}%+ヒモ固まり）\n"
            f"推定最高人気オッズ≈{est_odds:.0f}倍台。当たっても回収構造が成立しない"
        )

    # ② 逃げ確率が極端に高い（ヒモ判定関係なく低オッズ確実）
    if s1_prob >= 0.72:
        return True, (
            f"⛔見送り推奨（逃げ確率{s1_prob*100:.0f}%超）\n"
            f"1号艇逃げがほぼ確定→3連単オッズが構造的に低い"
        )

    # ③ 印↔確率不一致 + 逃げ優位（判断根拠不明確）
    if mismatch and s1_prob >= 0.60:
        return True, (
            f"⛔印↔確率不一致見送り（{mismatch_detail}）\n"
            f"逃げ{s1_prob*100:.0f}%優位で◎が確率最大艇でない"
        )

    # ④ 逃げシナリオ × 非許可会場
    #    逃げ軸流しはROIが構造的に低い会場が多いため、
    #    BT実績でROI30%以上の会場のみ参加を許可する。
    scenario_type = bet_suggestions.get("scenario_type", "")
    if scenario_type == "逃げ軸流し" and venue and venue not in _NIGE_ALLOWED_VENUES:
        return True, (
            f"⛔逃げ軸・非許可会場（{venue}）→ 見送り\n"
            f"BT実績: {venue}の逃げシナリオROIは30%未満。"
            f"許可会場: {', '.join(sorted(_NIGE_ALLOWED_VENUES))}"
        )

    # ⑤ 旧・全シナリオ会場フィルタ（現在は無効化済み。後方互換のため残す）
    if venue and venue in _SKIP_VENUES:
        return True, f"⛔{venue}：ROI低会場見送り（BT実績ROI≤67%）"

    return False, ""

def _suggest_3rentan(results, race_judgment, jizen_eval=None, honmei_map=None):
    """
    3連単買い目提案 v5 ── 印ベース軸決定＋内側残存補正ヒモ選択

    【設計思想】
      実オッズ・参加可否は人間が最終判断する前提。
      このロジックは「どの展開が来そうか」「その展開で何を買うべきか」を
      展示前の素材として提供することに専念する。

    【シナリオ判定と買い目形式】
      S1確率 >= 60%           → 逃げ軸流し（1着=1号艇固定）
      S1確率 40〜60%（拮抗）  → 両建てフォーメーション（逃げ軸＋飛び軸）
      S1確率 < 40%（飛び有力）→ 飛び軸フォーメーション

    【買い目の構成原則】
      ・全120通りの推定確率をシナリオ別に分類
      ・各シナリオの確率上位から累積80%を目安に買い目を選出
      ・折り返し買い（1-A-B と A-1-B）を自動的に対で追加
      ・カットオフなし・点数上限なし（確率で自動決定）

    【出力】
      scenario_type    : "逃げ軸流し" / "両建て" / "飛び軸"
      buy_list         : 推奨買い目（combo文字列リスト）
      point_count      : 点数
      theory_syn_odds  : 理論合成オッズ（= 0.75 / Σprob）
      required_syn_odds: 必要合成オッズ（= 点数 × 1.10）
      margin_ratio     : theory_syn_odds / required_syn_odds
      margin_verdict   : "余裕あり" / "要確認" / "見送り有力"
      candidates       : 買い目ごとの確率・シナリオ種別リスト
    """
    ryotate         = race_judgment.get("ryotate", {})
    ryotate_verdict = ryotate.get("verdict", "逃げ狙い")

    if not results:
        return {
            "axis1": "-", "axis2": "-", "buy_list": [], "point_count": 0,
            "comment": "データ不足", "combos": [], "candidates": [],
            "scenario_type": "-", "scenario_verdict": "-",
            "theory_syn_odds": None, "required_syn_odds": None,
            "margin_ratio": None, "margin_verdict": "-",
            "escape_score": 50, "tobi_score": 30, "fly_axes": [],
            "candidates_s1": [], "candidates_s2": [],
            "axis_candidates": [], "himo_candidates": [],
            "jizen_formation": {}, "ryotate_verdict": ryotate_verdict,
            "ryotate_detail": ryotate,
        }

    combos = _calc_3rentan_probs_v2(
        results,
        venue_course_1c_rate=race_judgment.get("venue_c1_win_rate"),
        jizen_eval=jizen_eval,
    )
    if not combos:
        return {
            "axis1": "-", "axis2": "-", "buy_list": [], "point_count": 0,
            "comment": "確率計算不能", "combos": [], "candidates": [],
            "scenario_type": "-", "scenario_verdict": "-",
            "theory_syn_odds": None, "required_syn_odds": None,
            "margin_ratio": None, "margin_verdict": "-",
            "escape_score": 50, "tobi_score": 30, "fly_axes": [],
            "candidates_s1": [], "candidates_s2": [],
            "axis_candidates": [], "himo_candidates": [],
            "jizen_formation": {}, "ryotate_verdict": ryotate_verdict,
            "ryotate_detail": ryotate, "first_prob_map": {},
        }

    # ══════════════════════════════════════════════════════════════════════
    # Step A: 1着別確率集計
    # ══════════════════════════════════════════════════════════════════════
    first_prob_map = {}
    for c in combos:
        w = c["first"]
        first_prob_map[w] = first_prob_map.get(w, 0) + c["prob"]

    s1_prob  = first_prob_map.get("1", 0.0)
    fly_prob = sum(p for w, p in first_prob_map.items() if w != "1")

    # 飛び候補艇を確率降順でリストアップ
    fly_candidates_sorted = sorted(
        [(w, p) for w, p in first_prob_map.items() if w != "1"],
        key=lambda x: x[1], reverse=True
    )
    main_fly_waku = fly_candidates_sorted[0][0] if fly_candidates_sorted else None
    sub_fly_waku  = fly_candidates_sorted[1][0] if len(fly_candidates_sorted) >= 2 else None

    # ══════════════════════════════════════════════════════════════════════
    # Step B: シナリオ判定（確率優先型 v3 ── 印は確率を補強する証拠として使用）
    # ══════════════════════════════════════════════════════════════════════
    # 【設計思想】
    # 旧版は「印◎の艇番でシナリオを決定」していたため、
    # 確率と印が逆向きになるとシナリオも買い目も確率と矛盾していた。
    # v3 では確率を起点にシナリオを決定し、印はその補強証拠として扱う。
    #
    # 判定フロー:
    #   Step B-1: s1_prob（確率）でシナリオの「下地」を決める
    #   Step B-2: 印◎の位置で補強・修正する（確率と同方向なら強化、逆方向なら両建て）
    #   Step B-3: ryotate（定性スコア）で最終調整
    # ══════════════════════════════════════════════════════════════════════
    first_turn      = race_judgment.get("first_turn", {}) or {}
    conflict_map    = race_judgment.get("conflict_map", {}) or {}
    sq              = race_judgment.get("scenario_quality", {}) or {}
    quality_rank    = sq.get("quality_rank", "B")
    lead_waku       = first_turn.get("lead_waku", "1")
    main_conflict   = conflict_map.get("main_conflict") or {}
    sub_conflict    = conflict_map.get("sub_conflict") or {}
    collapse_bene   = conflict_map.get("collapse_beneficiary", [])
    lead_is_1       = (lead_waku == "1")
    p_strength      = first_turn.get("pattern_strength", "中")
    mc_strength     = main_conflict.get("strength", 0) or 0

    # ── 印から軸・ヒモ情報を取得 ─────────────────────────────────────────
    if honmei_map:
        inv = {v: k for k, v in honmei_map.items() if v.strip()}
        honmei_waku  = inv.get("◎")
        taiko_waku   = inv.get("○")
        tanhana_waku = inv.get("▲")
        ana_waku     = inv.get("△")
    else:
        sorted_fp = sorted(first_prob_map.items(), key=lambda x: x[1], reverse=True)
        honmei_waku  = sorted_fp[0][0] if len(sorted_fp) > 0 else "1"
        taiko_waku   = sorted_fp[1][0] if len(sorted_fp) > 1 else None
        tanhana_waku = sorted_fp[2][0] if len(sorted_fp) > 2 else None
        ana_waku     = sorted_fp[3][0] if len(sorted_fp) > 3 else None

    # ── Step B-1: 確率でシナリオ下地を決定 ──────────────────────────────
    # s1_prob が高い → 1号艇が1着になる確率が高い → 逃げ軸が基本
    # s1_prob が低い → 他艇が1着になる確率が高い → 飛び軸が基本
    if s1_prob >= 0.60:
        scenario_base = "逃げ軸流し"
    elif s1_prob >= 0.42:
        scenario_base = "両建て"
    else:
        scenario_base = "飛び軸"

    # ── Step B-2: 印◎で補強・修正 ───────────────────────────────────────
    # 確率と印が同方向 → シナリオ確定
    # 確率と印が逆方向 → 矛盾の「強度」で判断
    #
    # 【v2 改善: 矛盾解消ロジック】
    # 旧方式: 逆方向なら無条件に両建て
    # 新方式: 乖離幅に応じて判断
    #   ・乖離が大きい（確率差 >= 20pt）→ 確率優先でシナリオ決定、印は◎位置を修正
    #   ・乖離が中程度（10〜20pt）      → 両建て（どちらが正しいか不確実）
    #   ・乖離が小さい（< 10pt）         → 印優先（印の方が人間の定性判断が入っている）
    #
    # 「確率vs印の乖離幅」: honmei_waku の first_prob を確率上位艇と比較
    _honmei_first_prob = first_prob_map.get(honmei_waku, 0) if honmei_waku else 0
    _top_prob_waku     = max(first_prob_map, key=first_prob_map.get) if first_prob_map else "1"
    _top_first_prob    = first_prob_map.get(_top_prob_waku, 0)
    _prob_gap_pct      = (_top_first_prob - _honmei_first_prob) * 100  # 正 = 確率最大艇 > ◎

    if scenario_base == "逃げ軸流し":
        if honmei_waku == "1":
            # 確率高×印◎=1号艇 → 最強の逃げシナリオ
            scenario_type = "逃げ軸流し"
        elif honmei_waku is not None:
            # 確率高なのに印◎≠1号艇
            if _prob_gap_pct >= 20:
                # 確率と印の乖離が大きい → 確率優先（印が誤っている可能性）
                scenario_type = "逃げ軸流し"
                # ただし飛び軸候補として印◎艇も両建てに含める
                main_fly_waku = honmei_waku
                scenario_type = "両建て"
            elif _prob_gap_pct >= 10:
                scenario_type = "両建て"
                main_fly_waku = honmei_waku
            else:
                # 乖離小 → 印優先（定性が確率を微修正している）
                scenario_type = "両建て"
                main_fly_waku = honmei_waku
        else:
            scenario_type = "逃げ軸流し"

    elif scenario_base == "飛び軸":
        if honmei_waku != "1":
            # 確率低×印◎≠1号艇 → 飛び軸確定
            scenario_type = "飛び軸"
            main_fly_waku = honmei_waku
        elif honmei_waku == "1":
            # 確率低なのに印◎=1号艇
            if _prob_gap_pct >= 20:
                # 確率と印の乖離大 → 確率優先して飛び軸
                scenario_type = "飛び軸"
            else:
                # 乖離小 → 印を尊重して両建て
                scenario_type = "両建て"
        else:
            scenario_type = "飛び軸"

    else:  # 両建て
        scenario_type = "両建て"
        if honmei_waku and honmei_waku != "1":
            main_fly_waku = honmei_waku

    # ── Step B-3: ryotate（定性スコア）で最終調整 ───────────────────────
    _ryotate_verdict   = ryotate.get("verdict", "逃げ狙い")
    _consistency_warn  = ryotate.get("consistency_warn", False)

    # ryotate が明確に飛び狙いと言っているのにシナリオが逃げ軸なら両建てに
    if _ryotate_verdict == "飛び狙い" and scenario_type == "逃げ軸流し":
        scenario_type = "両建て"
    # quality D は混戦 → 逃げ軸流しは危険
    if quality_rank == "D" and scenario_type == "逃げ軸流し":
        scenario_type = "両建て"

    # ── 飛び軸主軸の確定 ─────────────────────────────────────────────────
    if scenario_type in ("飛び軸", "両建て"):
        if main_fly_waku is None or main_fly_waku == "1":
            # 飛び軸主軸が未確定 or 1号艇になっている → 確率2位の艇に
            main_fly_waku = fly_candidates_sorted[0][0] if fly_candidates_sorted else "2"
        if sub_fly_waku is None or sub_fly_waku == main_fly_waku:
            sub_fly_waku = fly_candidates_sorted[1][0] if len(fly_candidates_sorted) >= 2 else taiko_waku

    # ══════════════════════════════════════════════════════════════════════
    # Step C: 買い目構成（展開エンジン連動・折り返し自動追加）
    # ══════════════════════════════════════════════════════════════════════
    # 展開qualityに連動した累積確率閾値:
    #   S（絞れる）: 閾値を下げて点数を絞る（0.70）
    #   A          : 標準（0.78）
    #   B          : 標準（0.82）
    #   C/D（混戦）: 閾値を上げて点数を増やす（0.88）
    # 全体上限: 最大18点（ユーザー設定）
    # quality連動で閾値を調整するが、合計が18点を超えた場合は確率降順で打ち切る。
    # SC潰れ受益は上限枠内で2点まで（枠を圧迫しないよう後付け）。
    MAX_BETS = 18
    THRESHOLD_BY_QUALITY = {"S": 0.70, "A": 0.75, "B": 0.80, "C": 0.85, "D": 0.85}
    CUMULATIVE_THRESHOLD = THRESHOLD_BY_QUALITY.get(quality_rank, 0.78)
    MIN_BETS = 3

    # SCシナリオ（潰れ展開）の漁夫候補を買い目に反映するか判定
    # 対立構造が明確（主軸強度>=40）でかつ飛び役自滅リスクが見込まれる場合に追加
    add_sc_bets = (mc_strength >= 40 and len(collapse_bene) >= 1)

    combo_lookup = {c["combo"]: c for c in combos}

    # 買い目根拠を生成する補助関数
    def _build_reason(e, scenario_ctx):
        """
        各買い目コンボの「買う理由」を自然言語で生成する。
        考察（展開エンジン）→ 買い目への因果を明示する。
        """
        first  = e["first"]
        second = e["second"]
        third  = e["third"]
        is_rev = e.get("_orkaeshi", False)

        # 1着根拠
        if first == "1":
            if lead_is_1 and p_strength == "強":
                r1 = f"1号先行確定的({p_strength})"
            elif lead_is_1:
                r1 = f"1号先行優位"
            else:
                r1 = f"1号逃げ残り(先行は{lead_waku}号)"
        elif first == main_fly_waku:
            mc_method = main_conflict.get("method", "攻撃")
            r1 = f"{first}号{mc_method}（主軸攻撃・強度{mc_strength:.0f}）"
        elif first == sub_fly_waku:
            sc_method = sub_conflict.get("method", "攻撃") if sub_conflict else "攻撃"
            r1 = f"{first}号{sc_method}（副軸）"
        else:
            r1 = f"{first}号（その他）"

        # 2着根拠
        top_bene = [w for w, _ in collapse_bene[:2]]
        if is_rev:
            r2 = f"{second}号折返（{first}号自滅時の逃げ取り戻し想定）"
        elif second in top_bene:
            r2 = f"{second}号漁夫（{main_conflict.get('attacker','?')}号自滅時の受益）"
        else:
            r2 = f"{second}号残存"

        return f"{r1} / {r2}"

    # ── 折り返し要否の判定関数 ──────────────────────────────────────────────────

    def _needs_orkaeshi_12(base_combo, rev_key):
        """
        1着折り返し（A-1-B vs 1-A-B）が必要かを判定する。

        不要と判断する条件:
          ① s1_prob >= 0.75: 逃げ確率が圧倒的 → 飛び役1着はほぼない
          ② 折り返しコンボの確率が本体の1/4未満: 展開として非現実的
          ③ 折り返し1着艇の1着確率が全艇平均の0.5倍未満: その艇が1着になる素地がない

        いずれか1つでも該当すれば不要と判断。
        """
        if rev_key not in combo_lookup:
            return False
        base = combo_lookup.get(base_combo)
        rev  = combo_lookup[rev_key]
        if not base:
            return True
        # ① 逃げ圧倒的
        if s1_prob >= 0.75:
            return False
        # ② 確率比
        if base["prob"] > 0 and rev["prob"] / base["prob"] < 0.25:
            return False
        # ③ 折り返し1着艇の1着確率
        rev_first_prob = first_prob_map.get(rev["first"], 0)
        avg_first_prob = sum(first_prob_map.values()) / max(len(first_prob_map), 1)
        if rev_first_prob < avg_first_prob * 0.5:
            return False
        return True

    def _needs_orkaeshi_23(base_combo, rev_key):
        """
        2着3着折り返し（1-A-B vs 1-B-A）が必要かを判定する。

        不要と判断する条件:
          ① 本体と折り返しの確率比が3倍以上: 順序がほぼ固定的
          ② 2着と3着の circle_pct（イン逃げ時2着優位度）差が2倍以上:
             2着候補がほぼ固定されている
          ③ 折り返しコンボの確率が全買い目平均の0.4倍未満: 薄すぎる

        いずれか1つでも該当すれば不要と判断。
        """
        if rev_key not in combo_lookup:
            return False
        base = combo_lookup.get(base_combo)
        rev  = combo_lookup[rev_key]
        if not base:
            return True
        # ① 確率比（3倍以上なら逆順はほぼ来ない）
        if base["prob"] > 0 and base["prob"] / max(rev["prob"], 1e-9) >= 3.0:
            return False
        # ② circle_pct差（2着固定度）
        circ_a = next((r.get("circle_pct") or 0 for r in results if r["waku"] == base["second"]), 0)
        circ_b = next((r.get("circle_pct") or 0 for r in results if r["waku"] == base["third"]),  0)
        if circ_a > 0 and circ_b > 0 and circ_a / circ_b >= 2.0:
            return False
        if circ_b > 0 and circ_a > 0 and circ_b / circ_a >= 2.0:
            return False
        # ③ 折り返し確率が薄すぎる
        avg_prob = sum(c["prob"] for c in combos) / max(len(combos), 1)
        if rev["prob"] < avg_prob * 0.4:
            return False
        return True

    # ══════════════════════════════════════════════════════════════════════
    # 共通データマップ（_build_buys内で参照）
    # ══════════════════════════════════════════════════════════════════════
    _r_map     = {r["waku"]: r for r in results}
    _cm_map    = {r["waku"]: r.get("raw_cm", {}) for r in results}
    _circ_map  = {r["waku"]: (r.get("circle_pct") or 0) for r in results}
    _win3_map  = {r["waku"]: (r.get("win3_rate") or 0.0) for r in results}
    _st_map    = {r["waku"]: r.get("avg_st") for r in results}
    _motor_map = {}
    for r in results:
        try:
            v = float(r.get("motor2") or 0)
            _motor_map[r["waku"]] = v if v > 0 else None
        except (ValueError, TypeError):
            _motor_map[r["waku"]] = None

    # jizen評価マップ（jizen_eval から艇別に取り出す）
    _jizen_aisho  = {}
    _jizen_tenkai = {}
    _jizen_jizai  = {}
    if jizen_eval is not None:
        for idx in range(6):
            w = str(idx + 1)
            _jizen_aisho[w]  = (jizen_eval.get("aisho")    or [""] * 6)[idx]
            _jizen_tenkai[w] = (jizen_eval.get("tenkai")   or [""] * 6)[idx]
            _jizen_jizai[w]  = (jizen_eval.get("jizaisei") or [""] * 6)[idx]

    # 1着軸ごとの2着残存補正テーブル（動的版 v2）
    # 【設計根拠 v2】
    # 旧版は静的テーブル（物理的な内側残存のみ）。
    # 実際には「2着に来る艇は、残存するだけでなく自分も仕掛けて生き残った艇」。
    # → POSITION_REMAIN[first][second] = 内側残存補正 × max(active_score, 0.5)
    #   active_score: その艇が仕掛け試行して生き残れる確率（calc_attack_probabilityで計算）
    #
    # 1号艇1着のときは circle_pct（イン逃げ時2着優位度）を使うためニュートラル維持。
    POSITION_REMAIN_BASE = {
        "1": {  # 逃げ: circle_pctで判断するためニュートラル
            "2": 1.0, "3": 1.0, "4": 1.0, "5": 1.0, "6": 1.0,
        },
        "2": {  # 差し: 1号艇内残存×1.35、3号艇競合下がり×0.75
            "1": 1.35, "3": 0.75, "4": 0.90, "5": 0.95, "6": 0.95,
        },
        "3": {  # まくり: 1〜2号艇内残存↑、4〜6号艇外側で来にくい↓
            "1": 1.25, "2": 1.20, "4": 0.85, "5": 0.90, "6": 0.90,
        },
        "4": {  # まくり: 1〜3号艇内残存↑（被まくりは内圧縮）、5〜6外側↓
            "1": 1.20, "2": 1.15, "3": 1.10, "5": 0.80, "6": 0.85,
        },
        "5": {  # まくり: 1〜4号艇内残存↑、6号艇は完全外側で最も来にくい↓
            "1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "6": 0.75,
        },
        "6": {  # 大まくり: 全艇が内側になるため全般的に内残存、5号艇のみ軽微↓
            "1": 1.15, "2": 1.10, "3": 1.05, "4": 1.00, "5": 0.95,
        },
    }

    # 各艇の仕掛け積極性スコアを事前計算（動的補正用）
    # active_score = P_attempt × P_survive（仕掛けて生き残れる確率）
    _r1_for_ap = next((r for r in results if r["waku"] == "1"), None)
    _active_scores = {}
    if HONMEI_SCENARIO_AVAILABLE and _r1_for_ap:
        try:
            from honmei_scenario import calc_attack_probability as _cap
            for _r in results:
                if _r["waku"] == "1":
                    _active_scores["1"] = 0.5   # 1号艇自身は逃げなので中立
                else:
                    _ap = _cap(_r, _r1_for_ap, results)
                    _active_scores[_r["waku"]] = _ap["attempt_prob"] * _ap["survive_prob"]
        except Exception:
            pass

    def _get_pos_remain(first_w, second_w):
        """POSITION_REMAIN_BASE × active_score補正で動的残存補正係数を返す"""
        base = POSITION_REMAIN_BASE.get(first_w, {}).get(second_w, 1.0)
        if first_w == "1":
            return base   # 逃げ時は circle_pct で別途判断
        # active_score: 0.5未満でも0.5を下限（完全に来ない艇は除外されるが補正は緩く）
        active = _active_scores.get(second_w, 0.5)
        active_adj = max(0.5, min(1.2, active * 1.5 + 0.25))
        return base * active_adj

    POSITION_REMAIN = POSITION_REMAIN_BASE   # 後方互換 (circle_pct処理で直接参照する箇所用)

    def _calc_himo_score(first_w, second_w, third_w, combo_prob):
        """
        1着軸が決まったときの 2着・3着組み合わせの総合スコアを算出。

        【スコア構成】
          ① combo確率ベース     (35%): シナリオ計算済みの3連単確率
          ② 位置残存補正        (25%): 内側残存の物理補正（対話ログ合意版）
              差し1着   → 1号艇内残存×1.35、3号艇競合×0.75
              まくり1着 → 内側艇↑（×1.05〜1.25）、外側艇↓（×0.75〜0.95）
              1号艇1着  → circle_pct（イン逃げ時2着優位度）で補正
          ③ 個人能力            (25%): コース別3連対率 × ST能力 × 機力 × 自在性
          ④ jizen展開・相性     (15%): 2着艇の相性評価 × 3着艇の展開評価
        """
        # ── ① combo確率（S1シナリオとは別物。3連単の確率値） ─────────────
        combo_prob_val = combo_prob

        # ── ② 位置残存補正 ───────────────────────────────────────────────
        # 1号艇1着のときは circle_pct（イン逃げ時2着優位度）を使う
        if first_w == "1":
            circ2 = _circ_map.get(second_w, 50) / 100.0
            circ3 = _circ_map.get(third_w, 50) / 100.0
            pos2  = 0.5 + circ2 * 0.5   # 0.5〜1.0にスケール
            pos3  = 0.5 + circ3 * 0.5
        else:
            # 動的残存補正: 物理的内側残存 × 仕掛け積極性
            pos2 = _get_pos_remain(first_w, second_w)
            pos3 = _get_pos_remain(first_w, third_w)
        pos_score = (pos2 + pos3) / 2.0

        # ── ③ 個人能力スコア ─────────────────────────────────────────────
        def _personal(w):
            r    = _r_map.get(w, {})
            cm   = _cm_map.get(w, {})

            # コース別3連対率（0〜1）
            w3 = _win3_map.get(w, 0.0)

            # STスコア: 速い艇ほど高い（艇間相対）
            st_self = _st_map.get(w)
            all_sts = [v for v in _st_map.values() if v is not None]
            if st_self is not None and len(all_sts) >= 2:
                st_min, st_max = min(all_sts), max(all_sts)
                st_score = 1.0 - (st_self - st_min) / max(st_max - st_min, 0.001)
            else:
                st_score = 0.5

            # 機力スコア（艇間相対）
            valid_motors = [v for v in _motor_map.values() if v is not None]
            mv = _motor_map.get(w)
            if mv is not None and len(valid_motors) >= 2:
                m_min, m_max = min(valid_motors), max(valid_motors)
                motor_score = (mv - m_min) / max(m_max - m_min, 0.001)
            else:
                motor_score = 0.5

            # 自在性（どの展開でも対応できるか）
            jiz_sym = _jizen_jizai.get(w, "")
            jiz_score = {"◎": 1.0, "○": 0.75, "△": 0.40, "": 0.30}.get(jiz_sym, 0.30)

            return w3 * 0.45 + st_score * 0.25 + motor_score * 0.15 + jiz_score * 0.15

        personal2 = _personal(second_w)
        personal3 = _personal(third_w)
        personal_score = (personal2 + personal3) / 2.0

        # ── ④ jizen展開・相性スコア ──────────────────────────────────────
        sym4 = {"◎": 1.0, "○": 0.75, "△": 0.40, "": 0.25}

        # 2着艇の相性（1号艇に対する攻め適性）
        aisho2  = sym4.get(_jizen_aisho.get(second_w, ""), 0.25)
        # 3着艇の展開（外枠での展開形成力）
        tenkai3 = sym4.get(_jizen_tenkai.get(third_w, ""), 0.25)
        jizen_score = (aisho2 + tenkai3) / 2.0

        # ── 合成 ─────────────────────────────────────────────────────────
        score = (combo_prob_val * 0.35
               + pos_score   * 0.25
               + personal_score * 0.25
               + jizen_score * 0.15)

        return score

    def _build_buys(base_first_waku, orkaeshi_first=None):
        source = [c for c in combos if c["first"] == base_first_waku]
        if not source:
            return [], set()

        # ── 各コンボに総合ヒモスコアを付与 ──────────────────────────────
        scored = []
        for c in source:
            hs = _calc_himo_score(
                c["first"], c["second"], c["third"], c["prob"]
            )
            scored.append((c, hs))

        # ヒモスコア降順でソート
        scored.sort(key=lambda x: x[1], reverse=True)

        # ── ヒモスコアを正規化して累積打ち切りに使う ─────────────────────
        # combo確率ではなくヒモスコアで累積することで
        # 「確率は低いが展開・個人能力で優位な組み合わせ」が打ち切られなくなる。
        # 正規化: 各スコア / 全スコア合計 → 割合で累積
        total_hs = sum(hs for _, hs in scored) or 1.0
        scored_with_share = [(c, hs, hs / total_hs) for c, hs in scored]

        selected  = []
        combo_set = set()
        cum_share = 0.0   # ヒモスコアの累積割合

        for c, hs, share in scored_with_share:
            key = c["combo"]
            if key not in combo_set:
                selected.append(c)
                combo_set.add(key)
                cum_share += share

            # 1着折り返し: 要否を判定してから追加
            if orkaeshi_first is not None:
                rev_key = f"{c['second']}-{orkaeshi_first}-{c['third']}"
                if rev_key not in combo_set and _needs_orkaeshi_12(c["combo"], rev_key):
                    rev = dict(combo_lookup[rev_key])
                    rev["_orkaeshi"] = True
                    combo_set.add(rev_key)
                    selected.append(rev)
                    # 折り返しのヒモスコアも加算
                    rev_hs = _calc_himo_score(rev["first"], rev["second"], rev["third"], rev["prob"])
                    cum_share += rev_hs / total_hs

            if len(selected) >= MIN_BETS and cum_share >= CUMULATIVE_THRESHOLD:
                break

        return selected, combo_set

    def _trim_to_max(entries, max_n):
        """確率降順で max_n 点に打ち切る"""
        return sorted(entries, key=lambda x: x["prob"], reverse=True)[:max_n]

    buy_entries = []
    seen_combos = set()

    if scenario_type == "逃げ軸流し":
        entries, _ = _build_buys("1", orkaeshi_first="1")
        buy_entries = entries

        # ── 2着3着折り返し（逃げ軸専用・要否判定付き）─────────────────────────
        # 1着=1号艇固定の前提で、2着と3着の順序が拮抗しているペアのみ追加する。
        orkaeshi_23_set = set()
        for e in list(buy_entries):
            if e["first"] != "1":
                continue
            rev_key  = f"1-{e['third']}-{e['second']}"
            pair_key = tuple(sorted([e["combo"], rev_key]))
            if pair_key in orkaeshi_23_set:
                continue
            orkaeshi_23_set.add(pair_key)
            already_have_rev = any(x["combo"] == rev_key for x in buy_entries)
            if not already_have_rev and _needs_orkaeshi_23(e["combo"], rev_key):
                rev = dict(combo_lookup[rev_key])
                rev["_orkaeshi_23"] = True
                buy_entries.append(rev)

    elif scenario_type == "両建て":
        s1_entries,  s1_seen  = _build_buys("1",           orkaeshi_first="1")
        fly_entries, fly_seen = (_build_buys(main_fly_waku, orkaeshi_first=main_fly_waku)
                                 if main_fly_waku else ([], set()))
        all_raw = s1_entries + [e for e in fly_entries if e["combo"] not in s1_seen]
        all_raw.sort(key=lambda x: x["prob"], reverse=True)
        for e in all_raw:
            if e["combo"] not in seen_combos:
                buy_entries.append(e)
                seen_combos.add(e["combo"])

    else:  # 飛び軸
        main_entries, main_seen = (_build_buys(main_fly_waku, orkaeshi_first=main_fly_waku)
                                   if main_fly_waku else ([], set()))
        sub_entries,  sub_seen  = (_build_buys(sub_fly_waku, orkaeshi_first=sub_fly_waku)
                                   if sub_fly_waku else ([], set()))
        all_raw = main_entries + [e for e in sub_entries if e["combo"] not in main_seen]
        all_raw.sort(key=lambda x: x["prob"], reverse=True)
        for e in all_raw:
            if e["combo"] not in seen_combos:
                buy_entries.append(e)
                seen_combos.add(e["combo"])

    # ── MAX_BETS 打ち切り（SC追加前にベース買い目を上限内に絞る）──────────────
    # SCは後から最大2点追加するので、ベース買い目は MAX_BETS-2 以内にする。
    SC_SLOTS = 2 if add_sc_bets and collapse_bene else 0
    base_max  = MAX_BETS - SC_SLOTS
    if len(buy_entries) > base_max:
        buy_entries = _trim_to_max(buy_entries, base_max)
        seen_combos = {e["combo"] for e in buy_entries}

    # ── SCシナリオ（潰れ展開）の漁夫候補を補完 ────────────────────────────────
    if add_sc_bets and collapse_bene:
        top_bene_w = collapse_bene[0][0]
        sc_additions = sorted(
            [c for c in combos
             if c["third"] == top_bene_w
             and c["combo"] not in seen_combos],
            key=lambda x: x["prob"], reverse=True
        )[:SC_SLOTS]
        for c in sc_additions:
            c = dict(c)
            c["_sc_bet"] = True
            buy_entries.append(c)
            seen_combos.add(c["combo"])

    # ── 最終上限チェック（SCを含めて MAX_BETS 超えたら確率降順で打ち切り）────
    if len(buy_entries) > MAX_BETS:
        buy_entries = _trim_to_max(buy_entries, MAX_BETS)

    # ── シナリオ整合フィルタリング ────────────────────────────────────────
    # 「飛び狙い」判定なのに1号艇1着の買い目が残っている矛盾を排除する。
    # ただし折り返し（1号艇が2着3着方向）は除外しない。
    # また1号艇1着を除外すると最低点数（MIN_BETS=3）を割る場合は保持する。
    if _ryotate_verdict == "飛び狙い" and scenario_type in ("飛び軸",):
        w1_first_entries = [e for e in buy_entries if e["first"] == "1"]
        non_w1_entries   = [e for e in buy_entries if e["first"] != "1"]
        if len(non_w1_entries) >= MIN_BETS:
            # 1号艇1着を除外できる → 除外
            buy_entries = non_w1_entries
        # 除外できない（点数不足）場合はそのまま保持し警告を残す

    # 艇番順ソート（1着→2着→3着の数値昇順）
    buy_entries.sort(key=lambda e: (int(e["first"]), int(e["second"]), int(e["third"])))
    buy_list    = [e["combo"] for e in buy_entries]
    point_count = len(buy_list)

    # ══════════════════════════════════════════════════════════════════════
    # Step D: 理論合成オッズ・余裕度・期待値警告
    # ══════════════════════════════════════════════════════════════════════
    # 【設計思想】
    # 点数で削らない。合成オッズが基準を下回った場合は「警告」として表示し、
    # 判断はユーザーに委ねる。
    #   → 回収重視ユーザー: 警告を見て見送り
    #   → 的中重視ユーザー: 警告を無視して参考買い目を使う
    #
    # 期待値基準:
    #   quality S/A → 3.0倍未満で警告
    #   quality B/C/D → 2.5倍未満で警告
    # ══════════════════════════════════════════════════════════════════════
    total_prob        = sum(e["prob"] for e in buy_entries)
    theory_syn_odds   = round(0.75 / total_prob, 1) if total_prob > 0 else None
    required_syn_odds = round(point_count * 1.10, 1) if point_count > 0 else None

    # 期待値警告フラグ
    # 胴元控除25%を考慮したブレークイーブンは1.33倍。
    # 実用的な下限を2.0倍とし、それを下回る場合のみ警告する。
    # （旧設定のS/A=3.0倍、B以下=2.5倍は厳しすぎて正常なレースにも警告が出ていた）
    EV_THRESHOLD = 2.0
    ev_warning = (theory_syn_odds is not None and theory_syn_odds < EV_THRESHOLD)
    ev_warning_msg = (
        f"⚠ 理想合成オッズ{theory_syn_odds}倍（期待値基準{EV_THRESHOLD}倍を下回っています）\n"
        f"  → 回収重視なら見送り推奨 / 的中重視なら参考買い目を使用可"
    ) if ev_warning else ""

    if theory_syn_odds and required_syn_odds and required_syn_odds > 0:
        margin_ratio = round(theory_syn_odds / required_syn_odds, 2)
        if margin_ratio >= 2.0:
            margin_verdict = "余裕あり（実オッズが理論値の半分でも成立）"
        elif margin_ratio >= 1.2:
            margin_verdict = "要確認（展示後に実オッズで判断）"
        else:
            margin_verdict = "見送り有力（理論値に余裕なし）"
    else:
        margin_ratio   = None
        margin_verdict = "-"

    # ══════════════════════════════════════════════════════════════════════
    # Step E: コメント・候補リスト生成（考察の結論を明示）
    # ══════════════════════════════════════════════════════════════════════
    fly_str    = f"{main_fly_waku}号艇" if main_fly_waku else "-"
    q_guide    = sq.get("bet_size_guide", "-")
    mc_desc    = main_conflict.get("desc", "-") if main_conflict else "-"

    # 考察→買い目の因果を1行で要約
    reasoning_line = (
        f"1M先行{lead_waku}号/{p_strength}・{mc_desc}"
        f"・quality{quality_rank}→{q_guide}"
    )

    # 整合性警告（ryotate の consistency_warn / honmei_prob_mismatch）を付加
    _warn_parts = []
    if ryotate.get("consistency_warn"):
        _warn_parts.append("⚠️スコア↔確率乖離補正済")
    _mismatch = race_judgment.get("honmei_prob_mismatch", False)
    if _mismatch:
        _mismatch_detail = race_judgment.get("honmei_prob_mismatch_detail", "")
        _warn_parts.append(f"⚠️印↔確率不一致({_mismatch_detail})")
    warn_suffix = "  " + " / ".join(_warn_parts) if _warn_parts else ""

    comment = (
        f"【{scenario_type}】"
        f"逃げ{s1_prob*100:.0f}%/飛び{fly_prob*100:.0f}%（主:{fly_str}）"
        f" ／ {point_count}点 / 理論合成{theory_syn_odds}倍"
        f" / 余裕度{margin_ratio}倍 → {margin_verdict}\n考察根拠: {reasoning_line}"
        f"{warn_suffix}"
    )

    def _scenario_label(e):
        f = e["first"]
        if e.get("_sc_bet"):
            return f"潰れ受益({f}号頭)"
        elif e.get("_orkaeshi_23"):
            return f"2着3着折返({f}号頭)"
        elif f == "1":
            return "逃げ（1号艇頭）"
        elif f == main_fly_waku:
            return f"飛び（{f}号艇頭）"
        elif e.get("_orkaeshi"):
            return f"1着折返({f}号頭)"
        else:
            return f"その他（{f}号艇頭）"

    # ヒモスコアを買い目エントリに付与（candidatesで表示用）
    _himo_score_cache = {}
    for e in buy_entries:
        k = e["combo"]
        if k not in _himo_score_cache:
            _himo_score_cache[k] = _calc_himo_score(
                e["first"], e["second"], e["third"], e["prob"]
            )

    candidates = [
        {
            "combo":          e["combo"],
            "prob":           round(e["prob"], 5),
            "prob_pct":       round(e["prob"] * 100, 2),
            "himo_score":     round(_himo_score_cache.get(e["combo"], 0), 4),
            "scenario":       _scenario_label(e),
            "is_orkaeshi":    e.get("_orkaeshi", False),
            "is_orkaeshi_23": e.get("_orkaeshi_23", False),
            "is_sc_bet":      e.get("_sc_bet", False),
            "reason":         _build_reason(e, scenario_type),
        }
        for e in buy_entries
    ]

    # ── 下流コードとの互換性維持 ──────────────────────────────────────────
    fly_axes = [w for w, _ in fly_candidates_sorted[:2]]

    from collections import Counter
    himo_counter = Counter()
    for c in combos[:30]:
        himo_counter[c["second"]] += 2
        himo_counter[c["third"]]  += 1
    axis_candidates = list(first_prob_map.keys())[:3]
    himo_candidates = [w for w, _ in himo_counter.most_common()
                       if w not in axis_candidates][:4]

    # race_judgmentへ分析結果を追記
    race_judgment["scenario_type"]   = scenario_type
    race_judgment["s1_prob"]         = round(s1_prob, 4)
    race_judgment["fly_prob"]        = round(fly_prob, 4)
    race_judgment["main_fly_waku"]   = main_fly_waku
    race_judgment["theory_syn_odds"] = theory_syn_odds
    race_judgment["margin_ratio"]    = margin_ratio
    race_judgment["margin_verdict"]  = margin_verdict
    race_judgment["ev_axes"]         = (
        ["1"] if scenario_type == "逃げ軸流し"
        else [main_fly_waku] if scenario_type == "飛び軸" and main_fly_waku
        else (["1", main_fly_waku] if main_fly_waku else ["1"])
    )
    race_judgment["w1_vs_venue"]     = round(
        (first_prob_map.get("1", 0) - (race_judgment.get("venue_c1_win_rate") or 0.555)) * 100, 1
    )
    race_judgment["w1_ev_pos_count"] = 0   # 互換維持（EV概念廃止）
    race_judgment["ev_axis_summary"] = [
        f"{w}号艇(1着確率{p*100:.1f}%)" for w, p in fly_candidates_sorted[:3]
    ]
    # SCシナリオ情報（数値シート表示用）
    _sc_combos_info = combos[0].get("sc_fly_type", None) if combos else None
    race_judgment["sc_fly_type"]    = _sc_combos_info or "-"
    race_judgment["sc_fly_waku"]    = main_fly_waku  # 飛び役兼潰れ役
    # 漁夫候補: 漁夫スコアが高い上位3艇（buy_listに含まれるもの優先）
    _sc_b = combos[0].get("sc_beneficiary") if combos else None
    if _sc_b is None:
        race_judgment["sc_gyofu_top3"] = []
    else:
        # combos全体からsc_beneficiaryを一度だけ取り出す（全艇共通値）
        # _calc_sc_weightの結果はcombo単位ではなく全艇共通なので最初のcomboから取得
        _all_b = {c["second"]: c.get("sc_beneficiary", 0) for c in combos[:30]}
        race_judgment["sc_gyofu_top3"] = sorted(
            _all_b.keys(), key=lambda w: _all_b.get(w, 0), reverse=True
        )[:3]

    h1 = "1" if scenario_type != "飛び軸" else (main_fly_waku or "1")
    h2 = main_fly_waku or "-"

    # ── honmei_scenario v2 統合 ──────────────────────────────────────────
    _base_result = {
        "axis1":             h1,
        "axis2":             h2,
        "buy_list":          buy_list,
        "point_count":       point_count,
        "comment":           comment,
        "combos":            combos,
        "candidates":        candidates,
        "scenario_type":     scenario_type,
        "scenario_verdict":  scenario_type,
        "s1_prob":           round(s1_prob, 4),
        "fly_prob":          round(fly_prob, 4),
        "theory_syn_odds":   theory_syn_odds,
        "required_syn_odds": required_syn_odds,
        "margin_ratio":      margin_ratio,
        "margin_verdict":    margin_verdict,
        "escape_score":      ryotate.get("escape_score", 50),
        "tobi_score":        ryotate.get("tobi_score", 30),
        "fly_axes":          fly_axes,
        "candidates_s1":     [],
        "candidates_s2":     [],
        "axis_candidates":   axis_candidates,
        "himo_candidates":   himo_candidates,
        "jizen_formation":   {},
        "ryotate_verdict":   ryotate_verdict,
        "ryotate_detail":    ryotate,
        "first_prob_map":    {w: round(p, 4) for w, p in first_prob_map.items()},
        # 整合性フラグ（バックテスト除外・警告表示用）
        "consistency_warn":          ryotate.get("consistency_warn", False),
        "honmei_prob_mismatch":      race_judgment.get("honmei_prob_mismatch", False),
        "honmei_prob_mismatch_detail": race_judgment.get("honmei_prob_mismatch_detail", ""),
        # 期待値警告（合成オッズが基準を下回っている場合）
        "ev_warning":     ev_warning,
        "ev_warning_msg": ev_warning_msg,
    }

    if HONMEI_SCENARIO_AVAILABLE and honmei_map:
        _result = integrate_with_suggest_3rentan(
            original_result = _base_result,
            results         = results,
            honmei_map      = honmei_map,
            combos          = combos,
            race_judgment   = race_judgment,
            jizen_eval      = jizen_eval,
        )
        # integrate_with_suggest_3rentan が theory_syn_odds を上書きするため
        # ev_warning を最終的な theory_syn_odds で再計算する
        _tso = _result.get("theory_syn_odds")
        _result["ev_warning"] = (_tso is not None and _tso < EV_THRESHOLD)
        _result["ev_warning_msg"] = (
            f"⚠ 理想合成オッズ{_tso}倍（期待値基準{EV_THRESHOLD}倍を下回っています）\n"
            f"  → 回収重視なら見送り推奨 / 的中重視なら参考買い目を使用可"
        ) if _result["ev_warning"] else ""
    else:
        _result = _base_result

    # ── 逃げシナリオ 点数上限（_NIGE_MAX_BETS）────────────────────────────
    # BT分析: 逃げは点数を絞っても的中率が改善しないため投資額を削減して回収率UP
    # シミュレーション: 5点上限で ROI 122% → 132% に改善（現状12点平均）
    _result_scenario = _result.get("scenario_type", "")
    if _result_scenario == "逃げ軸流し":
        _orig_buy = _result.get("buy_list", [])
        if len(_orig_buy) > _NIGE_MAX_BETS:
            _trimmed = _orig_buy[:_NIGE_MAX_BETS]
            _result["buy_list"]    = _trimmed
            _result["point_count"] = len(_trimmed)
            # candidates も同期
            _trimmed_set = set(_trimmed)
            _result["candidates"] = [
                c for c in _result.get("candidates", [])
                if c.get("combo") in _trimmed_set
            ]
            # theory_syn_odds を再計算
            _trim_combos = [
                c for c in _result.get("combos", [])
                if c["combo"] in _trimmed_set
            ]
            _trim_prob = sum(c["prob"] for c in _trim_combos)
            _result["theory_syn_odds"] = round(0.75 / _trim_prob, 1) if _trim_prob > 0 else None
            # ev_warning を再評価
            _tso2 = _result["theory_syn_odds"]
            _result["ev_warning"] = (_tso2 is not None and _tso2 < EV_THRESHOLD)
            _result["ev_warning_msg"] = (
                f"⚠ 理想合成オッズ{_tso2}倍（期待値基準{EV_THRESHOLD}倍を下回っています）\n"
                f"  → 回収重視なら見送り推奨 / 的中重視なら参考買い目を使用可"
            ) if _result["ev_warning"] else ""

    # ── 参加見送り判定をフラグとして付与 ──
    # venue は race_judgment 経由で受け取る（_suggest_3rentan は venue を直接知らない）
    _venue   = (race_judgment or {}).get("venue", "")
    # himo_are を race_judgment から _result に渡す（_should_skip_race が参照するため）
    if "himo_are" not in _result:
        _result["himo_are"] = (race_judgment or {}).get("himo_are", {}) or {}
    _skip, _skip_reason = _should_skip_race(_result, _venue)
    _result["skip"]        = _skip
    _result["skip_reason"] = _skip_reason

    return _result


# ============================================================
# 本命記号スコア計算（トップレベル関数）
# ※ 元は calc_race_indices 内のローカル関数だったが
#   main() から直接呼ぶためトップレベルに移動。
# ============================================================

# ============================================================
# 6人相互作用モデル：攻撃有効性スコア計算
# ============================================================

def _calc_attack_effectiveness(attacker, w1_cm, venue_stats, results):
    """
    「この攻撃艇が、今日の1号艇を実際に崩せるか」を定量化する。

    【設計思想】
    単独の決まり手%や平均STではなく、
    「攻撃艇の攻撃力 × 1号艇の当該攻撃への脆弱性 × STアドバンテージ × 会場適性」
    の積として攻撃有効性を算出する。

    これにより「1号艇の逃げ率が高くても、2号艇のSTが圧倒的で差し実績が高ければ
    2号艇◎」という競艇の実態に即した印付けが可能になる。

    Parameters
    ----------
    attacker    : dict  攻撃艇のresultsエントリ
    w1_cm       : dict  1号艇のraw_cm（被決まり手データ含む）
    venue_stats : dict  会場統計（決まり手別比率）
    results     : list  全艇のresultsリスト（ST平均計算用）

    Returns
    -------
    dict:
        total_score     : 総合攻撃有効性スコア（0〜1）
        attack_type     : 主要攻撃手段（"差し"/"まくり"/"まくり差し"/"逃げ"）
        attack_power    : 攻撃力（0〜1）
        w1_vulnerability: 1号艇の当該攻撃への脆弱性（0〜1）
        st_advantage    : STアドバンテージ（0.5〜1.5）
        venue_affinity  : 会場の当該決まり手適性（0〜1）
        breakdown       : 各因子の詳細
    """
    waku      = str(attacker.get("waku", ""))
    cm        = attacker.get("raw_cm", {})
    avg_st    = attacker.get("avg_st")

    # 1号艇の情報
    st1       = next((r.get("avg_st") for r in results if r["waku"] == "1"), None)
    nige_pct1 = safe_float(w1_cm.get("逃げ%"), 0.6) or 0.6
    sasar_v   = safe_float(w1_cm.get("差され%"),     0.0) or 0.0
    makur_v   = safe_float(w1_cm.get("捲られ%"),     0.0) or 0.0
    maksa_v   = safe_float(w1_cm.get("捲り差され%"), 0.0) or 0.0

    # 全艇ST平均（基準）
    st_vals = [r["avg_st"] for r in results if r.get("avg_st") is not None]
    st_mean = sum(st_vals) / len(st_vals) if st_vals else 0.15

    # 会場の決まり手別成功率（攻撃が会場に合うか）
    kimari_avg = venue_stats.get("kimari_avg", {}) or {}
    v_sashi   = safe_float(kimari_avg.get("差し"),      0.15) or 0.15
    v_makuri  = safe_float(kimari_avg.get("まくり"),    0.20) or 0.20
    v_maksa   = safe_float(kimari_avg.get("まくり差し"),0.10) or 0.10
    # 全国平均でスケール正規化（差し:15%、まくり:20%、まくり差し:10%）
    v_sashi_n  = min(v_sashi  / 0.15, 2.0)
    v_makuri_n = min(v_makuri / 0.20, 2.0)
    v_maksa_n  = min(v_maksa  / 0.10, 2.0)

    # ── STアドバンテージ（1号艇比較）─────────────────────────────────────
    if avg_st is not None and st1 is not None:
        st_diff   = st1 - avg_st          # 正 = 自分が速い
        st_adv_v1 = max(0.3, min(2.0, 1.0 + st_diff / 0.04))
    else:
        st_adv_v1 = 1.0

    # ── コース別の主攻撃手段と有効性を計算 ──────────────────────────────
    if waku == "1":
        # 1号艇：逃げ力 × 被攻撃への耐性（別関数で計算）
        nige_pct = safe_float(cm.get("逃げ%"), 0.6) or 0.6
        return {
            "total_score":      nige_pct,
            "attack_type":      "逃げ",
            "attack_power":     nige_pct,
            "w1_vulnerability": 0.0,
            "st_advantage":     1.0,
            "venue_affinity":   1.0,
            "breakdown":        {"逃げ%": nige_pct},
        }

    elif waku == "2":
        # 2枠：差し特化
        # 差し有効性 = 差し% × 被差し脆弱性 × STアドバンテージ × 会場差し率
        sashi_pct = safe_float(cm.get("差し%"), 0.0) or 0.0
        # 被差し脆弱性：差され%だけでなく逃げ%の低さも加味
        vuln_s = min(1.0, sasar_v * 0.65 + (1.0 - nige_pct1) * 0.35)
        score  = sashi_pct * vuln_s * st_adv_v1 * v_sashi_n
        return {
            "total_score":      min(score, 1.0),
            "attack_type":      "差し",
            "attack_power":     sashi_pct,
            "w1_vulnerability": vuln_s,
            "st_advantage":     st_adv_v1,
            "venue_affinity":   v_sashi_n,
            "breakdown":        {
                "差し%": sashi_pct, "被差し脆弱性": round(vuln_s,3),
                "ST優位": round(st_adv_v1,3), "会場差し率": round(v_sashi_n,3),
            },
        }

    elif waku == "3":
        # 3枠：まくり差し主体（まくりも一部）
        # 3枠は2枠の差し力が高いと進路を塞がれるペナルティ
        maksa_pct = safe_float(cm.get("まくり差し%"), 0.0) or 0.0
        maku_pct  = safe_float(cm.get("まくり%"),    0.0) or 0.0
        attack_p  = maksa_pct * 1.2 + maku_pct * 0.8

        # 2枠の差し力が高いほど3枠は動きにくい（進路妨害）
        r2  = next((r for r in results if r["waku"] == "2"), None)
        if r2:
            w2_sashi = safe_float(r2.get("raw_cm", {}).get("差し%"), 0.0) or 0.0
            w2_st    = r2.get("avg_st")
            if w2_st is not None and st1 is not None:
                w2_adv = max(0.0, (st1 - w2_st) / 0.04)
            else:
                w2_adv = 0.0
            block_penalty = max(0.5, 1.0 - w2_sashi * w2_adv * 0.5)
        else:
            block_penalty = 1.0

        vuln_m = min(1.0, maksa_v * 0.6 + makur_v * 0.4)
        score  = attack_p * vuln_m * st_adv_v1 * v_maksa_n * block_penalty
        return {
            "total_score":      min(score, 1.0),
            "attack_type":      "まくり差し",
            "attack_power":     attack_p,
            "w1_vulnerability": vuln_m,
            "st_advantage":     st_adv_v1,
            "venue_affinity":   v_maksa_n,
            "breakdown":        {
                "まくり差し%": maksa_pct, "まくり%": maku_pct,
                "被脆弱性": round(vuln_m,3), "2枠妨害": round(block_penalty,3),
            },
        }

    else:
        # 4〜6枠：まくり主体（外枠ほど難易度高い）
        maku_pct  = safe_float(cm.get("まくり%"),    0.0) or 0.0
        maksa_pct = safe_float(cm.get("まくり差し%"),0.0) or 0.0
        attack_p  = maku_pct + maksa_pct * 0.8

        # 外枠補正（4→1.0、5→0.85、6→0.70）
        outer_mult = {4: 1.0, 5: 0.85, 6: 0.70}.get(int(waku), 0.70)

        # 内側艇による壁ペナルティ（自分より内側にまくり力の強い艇がいると被る）
        inner_block = 1.0
        for r in results:
            rw = int(r["waku"])
            if 2 <= rw < int(waku):
                r_maku = safe_float(r.get("raw_cm", {}).get("まくり%"), 0.0) or 0.0
                r_st   = r.get("avg_st")
                if r_st is not None and avg_st is not None:
                    r_adv = max(0.0, (avg_st - r_st) / 0.04)  # 内側艇が速いほど壁
                else:
                    r_adv = 0.0
                inner_block = max(0.4, inner_block - r_maku * r_adv * 0.15)

        vuln_m = min(1.0, makur_v * 0.7 + (1.0 - nige_pct1) * 0.3)
        score  = attack_p * vuln_m * st_adv_v1 * v_makuri_n * outer_mult * inner_block
        return {
            "total_score":      min(score, 1.0),
            "attack_type":      "まくり",
            "attack_power":     attack_p,
            "w1_vulnerability": vuln_m,
            "st_advantage":     st_adv_v1,
            "venue_affinity":   v_makuri_n,
            "breakdown":        {
                "まくり%": maku_pct, "被脆弱性": round(vuln_m,3),
                "外枠補正": outer_mult, "内壁": round(inner_block,3),
            },
        }


def _calc_w1_escape_score(r1, results, venue_stats, race_judgment=None):
    """
    1号艇が「今日このメンバーを相手に逃げ切れる確率」を定量化する。

    【v2: 展開エンジン連動】
    race_judgment の escape_score（逃げ定性スコア）と tobi_score（飛び定性スコア）を
    直接参照する。展開エンジンが「飛び優勢」と判定しているのに
    1号艇◎になるという矛盾を根本から解消する。

    計算式:
      逃げ有効スコア
        = 選手固有の逃げ力（nige_pct × ST安定性）   … 過去実績
        × 今日の展開的な逃げ支持度（ryotate補正）    … 展開エンジン判定
        × 攻撃艇脅威ペナルティ                       … メンバー相互作用

    ryotate補正の設計:
      escape_score > tobi_score + 15 → 逃げ明確優勢 → 補正なし（×1.0）
      escape_score ≒ tobi_score（差±15以内）→ 拮抗  → 軽い抑制（×0.75）
      tobi_score > escape_score + 15  → 飛び優勢    → 強い抑制（×0.45）
      tobi_score > escape_score + 30  → 飛び圧倒    → 最大抑制（×0.25）

    Returns
    -------
    float: 逃げ有効スコア（0〜1）
    """
    cm1      = r1.get("raw_cm", {})
    nige_pct = safe_float(cm1.get("逃げ%"), 0.6) or 0.6
    st1      = r1.get("avg_st")

    # ── 展開エンジン（ryotate）の逃げ/飛びスコアを参照 ──────────────────
    rj = race_judgment or {}
    ryotate = rj.get("ryotate", {}) or {}
    esc_s = float(ryotate.get("escape_score", 50) or 50)   # 逃げ定性スコア（0〜100）
    tob_s = float(ryotate.get("tobi_score",   30) or 30)   # 飛び定性スコア（0〜100）
    diff  = esc_s - tob_s   # 正=逃げ優勢、負=飛び優勢

    if diff >= 15:
        ryotate_mult = 1.00   # 逃げ明確優勢
    elif diff >= -15:
        ryotate_mult = 0.75   # 拮抗（両建て領域）
    elif diff >= -30:
        ryotate_mult = 0.45   # 飛び優勢（画像のケース: esc=23, tob=35 → diff=-12 → 0.75）
    else:
        ryotate_mult = 0.25   # 飛び圧倒

    # ── 攻撃艇脅威ペナルティ（メンバー相互作用）───────────────────────
    threats = []
    for r in results:
        if r["waku"] == "1":
            continue
        eff = _calc_attack_effectiveness(r, cm1, venue_stats, results)
        threats.append(eff["total_score"])

    if threats:
        max_threat = max(threats)
        avg_threat = sum(threats) / len(threats)
        threat_penalty = max(0.4, 1.0 - max_threat * 0.5 - avg_threat * 0.15)
    else:
        threat_penalty = 1.0

    # ── ST安定性 ────────────────────────────────────────────────────────
    fly_label  = r1.get("fly_label", "低")
    fly_mult   = {"低": 1.0, "中": 0.88, "高": 0.72}.get(fly_label, 1.0)
    st_vals    = [r["avg_st"] for r in results if r.get("avg_st") is not None]
    if st1 is not None and st_vals:
        st_mean    = sum(st_vals) / len(st_vals)
        st_penalty = max(0.7, min(1.15, 1.0 - (st1 - st_mean) / 0.05 * 0.15))
    else:
        st_penalty = 1.0

    escape_score = nige_pct * ryotate_mult * threat_penalty * fly_mult * st_penalty
    return min(escape_score, 1.0)


def _calc_honmei_score(r, tobi_prob_val, jizen_ev=None, first_prob_map=None, results_ctx=None,
                       venue_stats=None, race_judgment=None):
    """
    【6人相互作用モデル 印スコア v4】

    設計思想:
      「確率が高い艇を◎にする」でも「補正点が高い艇を◎にする」でもなく、
      「このレースのこのメンバー構成で最も主導権を握れる艇を◎にする」。

      Stage1: 相互作用スコア（0〜50pt）← 土台
        攻撃艇: _calc_attack_effectiveness による「この1号艇を崩せる力」
        1号艇:  _calc_w1_escape_score による「このメンバーに逃げられる力」
        → 確率モデル（first_prob_map）との加重平均でブレンド
          相互作用 × 0.55 + 確率シェア × 0.45

      Stage2: 状態補正（0〜30pt）← このレースの「今の状態」
        ② 機力（0〜10pt）: モーター2連率ランク
        ③ jizen（0〜14pt）: 事前評価（展示での状態確認）
        ④ ST安定（0〜6pt）: FLYリスク・出遅れ癖

      確率ガード:
        first_prob_map の確率最大艇と自艇の確率差が大きいとき、
        相互作用スコアが高くても◎になれる上限を設ける。
        ただし旧版より緩め（25pt差で発動）: 相互作用モデルが
        確率を覆す場合もあり得るため。

      合計最大: 80pt
    """
    results   = results_ctx or []
    waku_str  = str(r["waku"])
    waku_idx  = int(r["waku"]) - 1
    _venue_stats = venue_stats or {}

    EQUAL_SHARE = 1.0 / 6.0

    # ══════════════════════════════════════════════════════
    # Stage1: 相互作用スコア (0〜50pt)
    # ══════════════════════════════════════════════════════

    # 1号艇の raw_cm を取得（全艇の攻撃有効性計算に必要）
    res1 = next((x for x in results if x["waku"] == "1"), None)
    w1_cm = res1.get("raw_cm", {}) if res1 else {}

    # 相互作用スコア（0〜1）
    if waku_str == "1":
        if res1:
            interaction_raw = _calc_w1_escape_score(res1, results, _venue_stats,
                                                    race_judgment=race_judgment)
        else:
            interaction_raw = 0.5
    else:
        eff = _calc_attack_effectiveness(r, w1_cm, _venue_stats, results)
        interaction_raw = eff["total_score"]
        # 攻撃タイプを記録（展開考察のコメント用）
        r["_attack_type"] = eff["attack_type"]
        r["_attack_eff"]  = round(eff["total_score"], 4)

    # first_prob_map との加重ブレンド
    if first_prob_map:
        total_p = sum(first_prob_map.values()) or 1.0
        p_share = first_prob_map.get(waku_str, 0.0) / total_p
    else:
        total_r = sum((x.get("rel_win1") or 0) for x in results) or 1.0
        p_share = (r.get("rel_win1") or 0) / total_r

    # ── ryotate（展開エンジン）の逃げ/飛びスコアを取得 ──────────────────
    rj       = race_judgment or {}
    ryotate  = rj.get("ryotate", {}) or {}
    esc_s    = float(ryotate.get("escape_score", 50) or 50)
    tob_s    = float(ryotate.get("tobi_score",   30) or 30)
    ryo_diff = esc_s - tob_s   # 正=逃げ優勢、負=飛び優勢

    # ── Stage1スコアを ryotate で補正 ───────────────────────────────────
    # 【核心設計】
    # first_prob_map は「1号艇55%の統計的優位」を内包している。
    # 飛び優勢レースでも1号艇の確率シェアが高くなるバイアスがある。
    # このバイアスを ryotate_diff で直接打ち消す。
    #
    # 1号艇: 飛び優勢のとき → 逃げ力スコアを削る（interaction_rawが既に削られているが不十分）
    # 飛び艇（2〜6号艇）: 飛び優勢のとき → 攻撃有効性スコアにボーナスを加算
    if waku_str == "1":
        # 飛び優勢度に応じて1号艇の相互作用スコアをさらに抑制
        if ryo_diff >= 10:
            w1_adj = 1.0    # 逃げ優勢 → 抑制なし
        elif ryo_diff >= -10:
            w1_adj = 0.80   # 拮抗
        elif ryo_diff >= -25:
            w1_adj = 0.60   # 飛び優勢（画像のケース diff=-12 → 0.80）
        else:
            w1_adj = 0.40   # 飛び圧倒
        interaction_raw = interaction_raw * w1_adj

        # 確率ボーナスも飛び優勢時は削る（統計バイアスを打ち消す）
        p_share_adj = p_share * max(0.5, (ryo_diff + 50) / 100.0)
    else:
        # 飛び艇: 飛び優勢のとき攻撃有効性ボーナス
        if ryo_diff <= -10:
            # 飛び優勢レースで攻撃力がある艇を後押し
            tobi_bonus = min(abs(ryo_diff) / 100.0, 0.30)  # 最大+0.30
            interaction_raw = min(interaction_raw * (1.0 + tobi_bonus), 1.0)
        p_share_adj = p_share

    blended = interaction_raw * 0.55 + p_share_adj * 0.45
    excess  = p_share_adj - EQUAL_SHARE
    prob_bonus = max(0.0, excess) * 20.0
    pt_stage1  = min(blended * 50 + prob_bonus, 50.0)

    # ══════════════════════════════════════════════════════
    # Stage2: 状態補正 (0〜30pt)
    # ══════════════════════════════════════════════════════

    # ── ② 機力点 (0〜10pt) ──────────────────────────────────────────────
    motor_vals = []
    for x in results:
        try:
            mv = float(x.get("motor2") or 0)
            if mv > 0:
                motor_vals.append((x["waku"], mv))
        except (ValueError, TypeError):
            pass
    if motor_vals:
        sorted_motors = sorted(motor_vals, key=lambda x: x[1], reverse=True)
        rank_map = {w: i for i, (w, _) in enumerate(sorted_motors)}
        my_rank  = rank_map.get(waku_str, len(sorted_motors) - 1)
        pt_motor = (1.0 - my_rank / max(len(sorted_motors) - 1, 1)) * 10.0
    else:
        pt_motor = 5.0

    # ── ③ jizen総合点 (0〜14pt) ─────────────────────────────────────────
    pt_jizen = 0.0
    if jizen_ev is not None:
        sym4 = {"◎": 3, "◎?": 2, "○": 2, "△": 1, "": 0, "-": 0}
        sym3 = {"A": 3, "B": 2, "C": 1, "D": 0, "E": 0, "-": 1}
        if waku_str == "1":
            in_sym  = (jizen_ev.get("in_nige")   or [""]   )[0]
            pt_jizen += sym4.get(in_sym,  0) / 3.0 * 4.0
        if waku_str != "1":
            ai_sym  = (jizen_ev.get("aisho")      or [""] * 6)[waku_idx]
            pt_jizen += sym4.get(ai_sym,  0) / 3.0 * 4.0
        ki_sym      = (jizen_ev.get("kiryoku")    or ["C"] * 6)[waku_idx]
        pt_jizen += sym3.get(ki_sym,  1) / 3.0 * 2.0
        jz_sym      = (jizen_ev.get("jizaisei")   or [""] * 6)[waku_idx]
        pt_jizen += sym4.get(jz_sym,  0) / 3.0 * 4.0
        if waku_str not in ("1", "2"):
            tk_sym  = (jizen_ev.get("tenkai")     or [""] * 6)[waku_idx]
            pt_jizen += sym4.get(tk_sym, 0) / 3.0 * 4.0
    pt_jizen = min(pt_jizen, 14.0)

    # ── ④ ST安定点 (0〜6pt) ─────────────────────────────────────────────
    all_sts = [(x["waku"], x.get("avg_st")) for x in results if x.get("avg_st") is not None]
    if len(all_sts) >= 2:
        sorted_sts  = sorted(all_sts, key=lambda x: x[1])
        st_rank_map = {w: i for i, (w, _) in enumerate(sorted_sts)}
        my_st_rank  = st_rank_map.get(waku_str, len(all_sts) - 1)
        pt_st       = (1.0 - my_st_rank / max(len(all_sts) - 1, 1)) * 6.0
    else:
        pt_st = 3.0

    # FLYリスク控除（状態補正からペナルティ）
    fly_pen = {"高": -15, "中": -7, "低": 0}.get(r.get("fly_label", "低"), 0)

    raw_score = pt_stage1 + pt_motor + pt_jizen + pt_st + fly_pen

    # ══════════════════════════════════════════════════════
    # 確率ガード（旧版より緩和: 差25pt以上で発動）
    # 相互作用モデルが確率を覆す正当な根拠がある場合を許容する
    # ══════════════════════════════════════════════════════
    if first_prob_map:
        max_p      = max(first_prob_map.values()) / (sum(first_prob_map.values()) or 1.0)
        max_excess = max_p - EQUAL_SHARE
        if max_excess >= 0:
            max_pt_stage1 = min(EQUAL_SHARE * 50 + max_excess * 2 * 50 + max_excess * 20, 50.0)
        else:
            max_pt_stage1 = max(0.0, max_p * 50)

        # 自艇の確率スコア相当値
        own_excess     = p_share - EQUAL_SHARE
        own_pt_stage1  = min(p_share * 50 + max(0.0, own_excess) * 20, 50.0)
        prob_gap       = max_pt_stage1 - own_pt_stage1

        if prob_gap >= 25:
            # 確率差が大きい → 相互作用が高くても上限キャップ
            cap = max_pt_stage1 + 30.0   # Stage2上限30ptまでは届ける
            raw_score = min(raw_score, cap)

    return raw_score


def _apply_jizen_honmei(results_ref, tobi_prob_val, jizen_ev, first_prob_map=None,
                        venue_stats=None, race_judgment=None):
    """
    jizen_eval 確定後に本命記号を最終確定する。
    main() の evaluate_all() 直後に呼び出すこと。
    6人相互作用モデル（v4）+ first_prob_map + jizen_ev を統合して印を確定する。
    race_judgment を受け取り、展開エンジンの逃げ/飛びスコアを印スコアに反映する。
    """
    final_scores = [
        (i, _calc_honmei_score(r, tobi_prob_val, jizen_ev=jizen_ev,
                               first_prob_map=first_prob_map,
                               results_ctx=results_ref,
                               venue_stats=venue_stats,
                               race_judgment=race_judgment))
        for i, r in enumerate(results_ref)
        if r.get("rel_win1") is not None
    ]
    final_scores.sort(key=lambda x: x[1], reverse=True)
    _hmap = {0: "◎", 1: "○", 2: "▲", 3: "△"}
    for r in results_ref:
        r["honmei"] = " "
    for rank, (idx, _) in enumerate(final_scores[:4]):
        results_ref[idx]["honmei"] = _hmap[rank]


def _calc_venue_stats(venue_stats_master, venue):
    """会場イン逃げ率・決まり手場平均を返す（会場統計シートから取得）"""
    vs = venue_stats_master.get(venue, {})
    in_rate  = safe_float(vs.get("イン逃げ率"))
    kimari_avg = {
        "差し":      safe_float(vs.get("差し率")),
        "まくり":    safe_float(vs.get("まくり率")),
        "まくり差し": safe_float(vs.get("まくり差し率")),
    }
    areyasusa = round((1.0 - float(in_rate)) * 100, 1) if in_rate is not None else None
    return {"in_rate": in_rate, "kimari_avg": kimari_avg, "areyasusa_score": areyasusa}


# ============================================================
# 倶楽部流 事前評価 - メンバーデータ組み立て
# ============================================================
def build_jizen_members(results, course_master, player_master, motor_df, race_no):
    """
    calc_race_indicesの戻り値(results)から evaluate_all 用データを組み立てる。

    Parameters
    ----------
    results       : list[dict]  calc_race_indices の戻り値（艇番順）
    course_master : dict        {(選手名, コース文字列): {指数dict}}
    player_master : dict        {選手名: {指数dict}}
    motor_df      : pd.DataFrame or None  scrape_motor.py の出力
    race_no       : int or str

    Returns
    -------
    list[dict]  evaluate_all() に渡せる6要素リスト（インデックス0=1号艇）
    """
    def _s(val, default=0.0):
        """文字列/数値を float に変換。変換不能な場合は default を返す。
        【修正⑧】default=None の場合は 0.0 も "データなし" と同一視せず None として扱う。
        つまり:
          _s("0.0", 0.0)  → 0.0  （正常変換）
          _s(None, 0.0)   → 0.0  （デフォルト）
          _s(None, None)  → None （「データなし」を明示したい場合）
          _s("0.0", None) → 0.0  （正常変換: 0.0 は有効なデータ）
        """
        try:
            v = str(val).replace("%", "").strip()
            if v in ("", "None", "nan", "-", "★"):
                return default
            return float(v)
        except Exception:
            return default

    def _s_nullable(val):
        """float変換を試み、変換不能(欠損)は None を返す。0.0は有効値として区別する。"""
        try:
            v = str(val).replace("%", "").strip()
            if v in ("", "None", "nan", "-", "★"):
                return None
            return float(v)
        except Exception:
            return None

    # モーターデータをインデックス化 {艇番int: 2連対率float or None}
    motor_index = {}
    if motor_df is not None:
        race_motor = motor_df[motor_df["race_no"] == int(race_no)]
        for _, row in race_motor.iterrows():
            bn = int(row["boat_no"]) if pd.notna(row["boat_no"]) else 0
            rate = row["motor_2rate"] if pd.notna(row["motor_2rate"]) else None
            motor_index[bn] = rate

    # 1枠の逃げ率（相性計算用）
    res0 = results[0] if results else {}
    cm0 = res0.get("raw_cm", {})
    nige_rate_1 = _s(cm0.get("逃げ%"), 1.0)

    members = []
    for i, res in enumerate(results[:6]):
        waku = res.get("waku", str(i + 1))
        boat_no = int(waku) if str(waku).isdigit() else (i + 1)
        name = res.get("name_norm", res.get("name", "").replace("　","").replace(" ",""))
        course = str(res.get("course", str(i + 1))).strip()
        cm = res.get("raw_cm", {})
        pm = res.get("raw_pm", {})

        # ── イン逃げ ──
        # 1コースのイン1着率（選手指数マスタ）
        rate_1st_c1 = _s(pm.get("イン\n1着率") or pm.get("イン1着率"))
        # 1コースのST順位（選手指数マスタ）
        st_rank_raw = pm.get("ST順位\n(1コース)") or pm.get("ST順位(1コース)")
        st_rank_c1 = _s(st_rank_raw) if st_rank_raw not in (None, "", "None") else None
        star_rate = bool(cm.get("★1着率"))

        # ── 相性（自コースの攻め決まり手割合） ──
        sashi_pct  = _s(cm.get("差し%"))
        makuri_pct = _s(cm.get("まくり%"))
        mz_pct     = _s(cm.get("まくり差し%"))
        attack_rate = sashi_pct + makuri_pct + mz_pct  # 合計割合（0〜1）

        # ── 相性用: 自コースの平均ST（秒）── evaluate_jizen.calc_aisho に必要
        avg_st_self = res.get("avg_st")  # calc_race_indices で算出済み

        # ── 相性用: 1号艇の被決まり手%（members[0] = 1号艇にのみ格納）──
        # 2号艇以降の相性評価で「この1号艇は差されやすいか/捲られやすいか」を参照する
        # 【修正⑧】_s_nullable を使って「差された回数ゼロ(0.0)」と「データなし(None)」を区別
        lose_sashi_rate  = None
        lose_makuri_rate = None
        if i == 0:
            # cm0（1号艇のコース別マスタ）から被決まり手%を取得
            lose_sashi_rate  = _s_nullable(cm.get("差され%"))
            lose_makuri_rate = _s_nullable(cm.get("捲られ%"))
            # キー名ゆれ対応（0.0 は有効値なので None の場合のみ代替キーを試みる）
            if lose_sashi_rate is None:
                lose_sashi_rate  = _s_nullable(cm.get("差し被%") or cm.get("被差し%"))
            if lose_makuri_rate is None:
                lose_makuri_rate = _s_nullable(cm.get("まくり被%") or cm.get("被まくり%"))

        # ── 機力 ──
        motor_2rate = motor_index.get(boat_no)

        # ── 自在性（2〜6コースの多様性割合） ──
        course_rows = []
        for c in range(2, 7):
            ck = (name, str(c))  # name is already normalized
            crow = course_master.get(ck) or {}
            crow_copy = dict(crow)
            crow_copy["course"] = c
            crow_copy["1着数"]    = _s(crow.get("1着数"))
            crow_copy["差し(件)"]  = _s(crow.get("差し(件)"))
            crow_copy["まくり(件)"] = _s(crow.get("まくり(件)"))
            crow_copy["まくり差し%"] = _s(crow.get("まくり差し%"))
            course_rows.append(crow_copy)
        diversity_rate = calculate_diversity_rate(course_rows)
        jizaisei_rate  = _s(pm.get("自在性\n1着率") or pm.get("自在性1着率"))
        star_kimete    = bool(cm.get("★決手"))

        # ── 展開（自コースの3連対率・まくり系） ──
        rate_3ren          = _s(cm.get("3連対率"))
        makuri_rate_t      = _s(cm.get("まくり%"))
        mz_rate_t          = _s(cm.get("まくり差し%"))

        members.append({
            # イン逃げ
            "rate_1st_c1":  rate_1st_c1,
            "st_rank_c1":   st_rank_c1,
            "star_rate":    star_rate,
            # 相性（v3対応: 攻め武器・ST・被決まり手）
            "nigé_rate":         nige_rate_1,
            "attack_rate":       attack_rate,
            "sashi_rate":        sashi_pct,         # 差し%（相性用・展開用共通）
            "makuri_rate":       makuri_pct,         # まくり%（相性用・展開用共通）
            "makuri_zashi_rate": mz_pct,             # まくり差し%（相性用・展開用共通）
            "avg_st_self":       avg_st_self,         # 【修正②】自コース平均ST秒（相性用）
            "lose_sashi_rate":   lose_sashi_rate,     # 【修正②】1号艇のみ: 差され%（相性用）
            "lose_makuri_rate":  lose_makuri_rate,    # 【修正②】1号艇のみ: 捲られ%（相性用）
            # 機力
            "motor_2rate":  motor_2rate,
            # 自在性
            "diversity_rate": diversity_rate,
            "jizaisei_rate":  jizaisei_rate,
            "star_kimete":    star_kimete,
            # 展開（makuri_rate / makuri_zashi_rate は相性用と同値のため共用）
            "rate_3ren":  rate_3ren,
        })

    return members

# ============================================================
# サンプルシートからレイアウトをコピーして新シートを作成
# ============================================================
def clone_sample_layout(wb):
    """出力_新聞サンプルのレイアウトを 出力_新聞 に完全コピー（行高・列幅・結合・色）

    ⚠️  DEPRECATED: main() は write_race_flat() を直接呼び出すためこの関数は未使用。
    削除候補。外部スクリプトから参照されていないことを確認後に除去すること。
    """
    ws_src = wb[SHEET_SAMPLE]
    
    # 既存の出力_新聞を削除して再作成
    if SHEET_OUTPUT in wb.sheetnames:
        del wb[SHEET_OUTPUT]
    
    # サンプルシートをコピーして名前変更
    ws_new = wb.copy_worksheet(ws_src)
    ws_new.title = SHEET_OUTPUT
    
    # シートをサンプルの直後に移動
    sample_idx = wb.sheetnames.index(SHEET_SAMPLE)
    wb.move_sheet(SHEET_OUTPUT, offset=-(len(wb.sheetnames) - sample_idx - 2))
    
    return ws_new

# ============================================================
# 1レース分をサンプルレイアウトに書き込む
# ============================================================
def write_race_to_sample_layout(ws, row_offset, race_no, venue, race_date,
                                  results, slit, venue_stats, frame_2nd, total_races,
                                  _jizen_members=None):
    """
    出力_新聞サンプルの29行レイアウトに合わせてデータ書き込み
    row_offset: このレースの開始行（1始まり）

    ⚠️  DEPRECATED: main() は write_race_flat() を使用。この関数は呼び出されていない。
    削除候補。外部スクリプトから参照されていないことを確認後に除去すること。
    """
    r = row_offset  # 行番号ショートカット

    # === Row 1: ヘッダ行（会場・レース番号・新聞名・日付） ===
    ws.cell(r+0, 1).value = venue
    ws.cell(r+0, 3).value = f"{race_no}R"
    ws.cell(r+0, 9).value = race_date
    # 「ボートリサーチ新聞」は結合セル I2:P3 → row2のI列（相対row+1, col=9）

    # === Row 2: イン逃げ場平均・決まり手場平均・スリット ===
    in_rate = venue_stats.get("in_rate")
    in_str = f"イン逃げ場平均: {in_rate*100:.1f}%" if in_rate is not None else "イン逃げ場平均: -"
    ws.cell(r+2, 3).value = in_str  # C3（相対）→ col=3
    ws.cell(r+5, 9).value = f"想定スリット: {slit}"  # スリットはrow3（header行の右側）

    # === Row 4: ヘッダ行（カラム名） - サンプルから継承されるので更新不要 ===
    # 「1号艇平均ST N.N位」を更新
    st1_rank = None
    for res in results:
        if res["waku"] == "1":
            st1_rank = res.get("st_rank")
            break
    if st1_rank:
        ws.cell(r+3, 11).value = f"1号艇平均ST\n{st1_rank:.1f}位"
    else:
        ws.cell(r+3, 11).value = "1号艇平均ST\n-"

    # === Row 5-10: 6選手データ ===
    for i, res in enumerate(results[:6]):
        data_row = r + 4 + i  # row5〜10（0-indexed: +4,+5,+6,+7,+8,+9）
        waku_no = int(res["waku"]) if res["waku"].isdigit() else (i+1)
        
        name_display = f"{res['honmei']} {res['name']}" if res["honmei"].strip() else f"  {res['name']}"
        rel_w1 = f"{res['rel_win1']:.1f}%" if res["rel_win1"] is not None else "-"
        abs_w3 = f"{res['abs_win3']:.1f}%" if res.get("abs_win3") is not None else "-"
        idx3   = res.get("idx3", 0)
        avg_st = f"{res['avg_st']:.3f}" if res.get("avg_st") is not None else "-"

        ws.cell(data_row, 2).value = name_display         # B: 選手名
        ws.cell(data_row, 5).value = res["kumi"]          # E: 組
        ws.cell(data_row, 6).value = res["motor2"]        # F: モータ2連
        ws.cell(data_row, 9).value = res["course"]        # I: 想定スリット（コース）
        ws.cell(data_row, 12).value = res["honmei"] if res["honmei"].strip() else None  # L: 本命記号
        ws.cell(data_row, 13).value = rel_w1              # M: ◆相対1着率
        ws.cell(data_row, 15).value = abs_w3              # O: コース別3連対率（絶対評価）
        
        # [修正5] 上記の二重書き込みブロックを削除（col9/12/13の上書きバグ修正）
        # H列: F/St影響（モーター由来・未実装のため空欄）
        # I列: 想定スリット内コース番号（上記col9=res["course"]で書込済）
        # L列: 評価記号（上記col12=res["honmei"]で書込済）
        # M列: オリジナル1着率（上記col13=rel_w1で書込済）

    # === Row 12: 事前評価セクションヘッダ - 継承 ===

    # === Row 14-19: 事前評価（5項目）+ 3着指数 ===
    # 3着指数は既存通り
    for i, res in enumerate(results[:6]):
        data_row = r + 13 + i
        idx3 = res.get("idx3", 0)
        ws.cell(data_row, 16).value = idx3 if idx3 else None  # P: 3着指数

    # 倶楽部流 事前評価（evaluate_jizen.py が利用可能な場合のみ）
    if JIZEN_AVAILABLE and _jizen_members:
        jizen_result = evaluate_all(_jizen_members)
        # 列マッピング: B=2(イン逃げ) C=3(相性) D=4(機力) E=5(自在性) F=6(展開)
        col_map = {
            "in_nige":  2,
            "aisho":    3,
            "kiryoku":  4,
            "jizaisei": 5,
            "tenkai":   6,
        }
        for key, col in col_map.items():
            for i, symbol in enumerate(jizen_result[key]):
                data_row = r + 13 + i
                ws.cell(data_row, col).value = symbol if symbol else None

    # === Row 21-28: 決まり手・場平均ブロック ===
    # G14:O19 の2着率テキストブロックを更新
    # （結合セルなので左上セルに書く）
    circle_lines = ["2着優位度（イン逃げ決着時・相対%）\n"]
    color_labels = {2: "黒", 3: "赤", 4: "青", 5: "黄", 6: "緑"}
    circle_pct_map2 = {res["waku"]: res.get("circle_pct") for res in results}
    for waku_no in range(2, 7):
        w = str(waku_no)
        pct = circle_pct_map2.get(w)
        label = color_labels.get(waku_no, "")
        pct_str = f"{pct:.0f}%" if pct is not None else "-"
        circle_lines.append(f"{label}({waku_no}枠): {pct_str:>4}")
    ws.cell(r+13, 7).value = "\n".join(circle_lines)

    # === 区切り行（Row 29）: 注意書き継承 ===
    # サンプルから自動引き継ぎ

# ============================================================
# ST順位 舟図生成（Pillowのみ使用・matplotlibは不要）
# ============================================================
def _make_st_boat_chart(results, player_master, outpath):
    """
    各選手の出走コース平均STをもとに、横位置でST順位を表現した舟図を生成。
    速い（ST値が小さい）ほど右側に配置。
    上から1号艇〜6号艇の固定順。
    Pillowのみ使用（matplotlibは不要）。
    """
    if not PIL_AVAILABLE:
        print("  ⚠️  Pillowが未インストールのためST舟図をスキップします。")
        return

    # 画像サイズ（EMU 865179 x 2082127 → 96dpi換算で約91x220px、220dpiで約204x490px）
    W, H = 204, 490

    # フォント
    font_candidates = [
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    fnt_path = next((f for f in font_candidates if os.path.exists(f)), None)
    fnt_size = max(int(H / 6 * 0.38), 10)
    fnt = ImageFont.truetype(fnt_path, fnt_size) if fnt_path else ImageFont.load_default()

    BOAT_STYLES = [
        ('#FFFFFF', '#333333', '#000000'),  # 1白
        ('#1a1a1a', '#555555', '#ffffff'),  # 2黒
        ('#CC0000', '#990000', '#ffffff'),  # 3赤
        ('#2255CC', '#1144AA', '#ffffff'),  # 4青
        ('#DDCC00', '#BBAA00', '#000000'),  # 5黄
        ('#115511', '#003300', '#ffffff'),  # 6緑
    ]

    # 各選手の出走コース平均STを取得
    st_col_keys = [
        '平均ST\n(1コース)', '平均ST\n(2コース)', '平均ST\n(3コース)',
        '平均ST\n(4コース)', '平均ST\n(5コース)', '平均ST\n(6コース)',
    ]
    boats = []
    for res in results[:6]:
        waku = int(res['waku']) if str(res['waku']).isdigit() else (len(boats) + 1)
        name = res.get('name_norm', res.get('name', '').replace('　', '').replace(' ', ''))
        course = str(res.get('course', str(waku))).strip()
        pm = player_master.get(name, {})
        st = None
        try:
            cidx = int(course) - 1
            if 0 <= cidx < 6:
                raw = pm.get(st_col_keys[cidx])
                if raw is not None:
                    v = str(raw).replace('%', '').strip()
                    if v not in ('', 'None', 'nan', '-'):
                        st = float(v)
        except Exception as e:
            print(f"  ⚠️  ST値取得エラー（{name} / コース{course}）: {e}")
        boats.append({'boat': waku, 'st': st})

    # ST値の範囲
    st_vals = [b['st'] for b in boats if b['st'] is not None]
    st_min = min(st_vals) if st_vals else 0.15
    st_max = max(st_vals) if st_vals else 0.20
    st_range = max(st_max - st_min, 0.001)

    # レイアウト
    slot_h = H / 6
    bh = slot_h * 0.55          # 舟の高さ
    bw = W * 0.52               # 舟の長さ
    x_fast = W - bw / 2 - 4    # 速い=右端
    x_slow = bw / 2 + 4        # 遅い=左端

    img = PILImage.new('RGB', (W, H), 'white')
    draw = ImageDraw.Draw(img)

    def hex2rgb(h):
        h = h.lstrip('#')
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

    def draw_boat_pil(draw, cx, cy, bw, bh, fill, outline, num, tc, fnt):
        pts = [
            (cx + bw * 0.50, cy),
            (cx + bw * 0.22, cy - bh * 0.45),
            (cx - bw * 0.32, cy - bh * 0.45),
            (cx - bw * 0.50, cy - bh * 0.20),
            (cx - bw * 0.50, cy + bh * 0.20),
            (cx - bw * 0.32, cy + bh * 0.45),
            (cx + bw * 0.22, cy + bh * 0.45),
        ]
        draw.polygon(pts, fill=hex2rgb(fill), outline=hex2rgb(outline))
        draw.text((cx - bw * 0.06, cy), str(num),
                  fill=hex2rgb(tc), font=fnt, anchor='mm')

    for i, b in enumerate(boats):
        fill, outline, tc = BOAT_STYLES[b['boat'] - 1]
        cy = (i + 0.5) * slot_h
        if b['st'] is not None:
            norm = (b['st'] - st_min) / st_range
            cx = x_fast - norm * (x_fast - x_slow)
        else:
            cx = (x_fast + x_slow) / 2
        draw_boat_pil(draw, cx, cy, bw, bh, fill, outline, b['boat'], tc, fnt)

    # 改善C: outpathがNoneの場合はBytesIOに保存してそのまま返す（ディスクI/O削減）
    if outpath is None:
        from io import BytesIO
        buf = BytesIO()
        img.save(buf, format='PNG', dpi=(220, 220))
        buf.seek(0)
        return buf
    img.save(outpath, dpi=(220, 220))



def write_race_flat(ws, row_offset, race_no, venue, race_date,
                    results, slit, venue_stats, frame_2nd, _jizen_members=None,
                    player_master=None, _tmp_image_paths=None, deadline=None,
                    race_judgment=None):
    """
    出力_新聞サンプルのレイアウトに完全準拠して1レース分を書き込む。
    29行ブロック（Row1-29）＋空行1行 = 30行/レース
    """
    r = row_offset
    c = center_align()
    lft = left_align()
    hf = header_fill()
    hfont = header_font()
    shf = subheader_fill()
    shfont = subheader_font()
    bdr = thin_border()

    def wc(row, col, val, **kwargs):
        write_cell(ws, row, col, val, **kwargs)

    def merge(r1, c1, r2, c2):
        try:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        except Exception as e:
            print(f"  ⚠️  セル結合スキップ（行{r1}列{c1}〜行{r2}列{c2}）: {e}")

    # ── 結合セル定義（サンプル完全再現） ──
    merge(r+0, 1, r+1, 2)    # A1:B2  会場名
    merge(r+0, 3, r+1, 3)    # C1:C2  イン逃げ場平均
    merge(r+0, 4, r+0, 6)    # D1:F1  決まり手場平均
    merge(r+0, 7, r+0, 8)    # G1:H1  波乱度
    merge(r+0, 9, r+0,15)    # I1:O1  発行日
    merge(r+1, 7, r+1, 8)    # G2:H2
    merge(r+1, 9, r+2,16)    # I2:P3  ボートリサーチ新聞
    merge(r+2, 1, r+2, 2)    # A3:B3  レース番号
    merge(r+2, 7, r+2, 8)    # G3:H3  想定スリット
    merge(r+3, 1, r+3, 4)    # A4:D4  選手名ヘッダ
    merge(r+3, 7, r+3, 8)    # G4:H4  F/St
    merge(r+3, 9, r+3,10)    # I4:J4  想定スリット
    merge(r+3,13, r+3,14)    # M4:N4  オリジナル1着率
    merge(r+3,15, r+3,16)    # O4:P4  一般戦3連対
    for i in range(6):       # 選手行 B:D結合, I:J結合（K列はST図用に開放）, M:N結合, O:P結合
        rr = r+4+i
        merge(rr, 2, rr, 4)
        merge(rr, 9, rr,10)
        merge(rr,13, rr,14)
        merge(rr,15, rr,16)
    merge(r+11, 1, r+11, 6)  # A12:F12 事前評価タイトル
    merge(r+11, 7, r+12,13)  # G12:M13 2着率テキスト
    merge(r+11,14, r+12,14)  # N12:N13 3着指数
    merge(r+11,15, r+12,16)  # O12:P13 オリジナル3連対
    for i in range(6):       # 事前評価 O:P結合
        merge(r+13+i,15, r+13+i,16)
    merge(r+20, 1, r+20, 5)  # A21:E21 決まり手タイトル
    merge(r+20, 6, r+27,11)  # F21:K28 説明テキスト
    merge(r+20,12, r+20,16)  # L21:P21 場平均タイトル
    for i in range(7):       # 場平均 M:N結合, O:P結合
        merge(r+21+i,13, r+21+i,14)
        merge(r+21+i,15, r+21+i,16)
    merge(r+28, 1, r+28,16)  # A29:P29 注記

    # ── Row 1-2: ヘッダ ──
    wc(r+0, 1, venue,  font=Font(name="Noto Sans CJK SC", size=12, bold=True), alignment=c)
    wc(r+0, 3, 'イン\n逃げ\n場平均', fill=hf, font=Font(name="ＭＳ ゴシック", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+0, 4, '決まり手場平均', fill=hf, font=hfont, alignment=c)
    # 波乱度セルに会場の荒れやすさスコアを表示
    areyasusa = venue_stats.get('areyasusa_score')
    haranddo_str = f'波乱度\n{areyasusa:.0f}pt' if areyasusa is not None else '波乱度'
    wc(r+0, 7, haranddo_str, fill=hf, font=hfont, alignment=c)
    in_rate = venue_stats.get('in_rate')
    wc(r+0, 9, race_date, fill=hf, font=Font(name="Meiryo UI", size=11, bold=True, color="FFFFFFFF"),
       alignment=Alignment(horizontal="right", vertical="center", wrap_text=True))
    wc(r+0,16, '発行', fill=hf, font=Font(name="Noto Sans CJK SC", size=11, bold=True, color="FFFFFFFF"), alignment=c)
    # ── タイトル「ボートリサーチ新聞」（I2:P3 結合セルの左上 = col9, row+1） ──
    wc(r+1, 9, 'ボートリサーチ新聞',
       font=Font(name="Noto Sans CJK SC", size=24, bold=True),
       alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

    # 決まり手場平均（会場統計から）
    kimari = venue_stats.get('kimari_avg', {})
    for col, key in [(4,'差し'), (5,'まくり'), (6,'まくり差し')]:
        wc(r+1, col, key, fill=hf, font=hfont, alignment=c)
        v = kimari.get(key)
        vs_str = f'{v*100:.0f}%' if v else '-'
        wc(r+2, col, vs_str, alignment=c)

    # イン逃げ場平均（C列, Row1-2結合）
    in_str = f'{in_rate*100:.1f}%' if in_rate is not None else '-'
    wc(r+2, 3, in_str, alignment=c)

    # ── Row 3: レース番号・スリット ──
    race_no_val = f'{race_no}R\n{deadline}' if deadline else f'{race_no}R'
    wc(r+2, 1, race_no_val, font=Font(name="Meiryo UI", size=14, bold=True), alignment=c)
    wc(r+2, 7, f'想定スリット: {slit}', font=Font(name="Arial", size=13, bold=True), alignment=c)


    # ── Row 4: 列ヘッダ ──
    st1_rank = next((res.get('st_rank') for res in results if res['waku']=='1'), None)
    st1_str  = f'1号艇平均ST\n{st1_rank:.1f}位' if st1_rank else '1号艇平均ST\n-'
    for col, txt in [(1,'選手名'),(5,'組'),(9,'想定\nスリット'),(13,'オリジナル\n1着率'),(15,'コース別\n3連対')]:
        wc(r+3, col, txt, fill=hf, font=Font(name="Noto Sans CJK SC", size=9 if col!=1 else 9, bold=True, color="FFFFFFFF"), alignment=c)
    for col, txt in [(6,'モータ\n2連'),(7,'展示\n偏差値')]:
        wc(r+3, col, txt, fill=hf, font=Font(name="Meiryo UI", size=8, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+3, 11, st1_str, fill=hf, font=Font(name="Meiryo UI", size=8, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+3, 12, '評価', fill=hf, font=Font(name="Meiryo UI", size=10, bold=True, color="FFFFFFFF"), alignment=c)

    # ── Row 5-10: 選手データ ──
    # 艇番の濃い背景色（サンプル準拠）
    color_fills = {
        1: 'FFFFFFFF', 2: 'FF111111', 3: 'FFFF0000',
        4: 'FF00B0F0', 5: 'FFFFFF00', 6: 'FF00B050'
    }
    # 艇番列の文字色
    color_fonts = {
        1: 'FF000000', 2: 'FFFFFFFF', 3: 'FFFFFFFF',
        4: 'FFFFFFFF', 5: 'FF000000', 6: 'FFFFFFFF'
    }
    # 隣接列（B,E,F,G,H,L,M,O）の薄い背景色（サンプル準拠）
    color_fills_light = {
        1: 'FFFFFFFF', 2: 'FF000000', 3: 'FFFFCCCC',
        4: 'FFCCE8FF', 5: 'FFFFFACC', 6: 'FFCCFFEE'
    }
    # 黒艇（2号艇）の隣接列は黒背景だが、サンプルではtheme=0（黒）→実際は黒
    # ただしB列はtheme=0 tint=-0.15→ほぼ黒。文字は白。
    color_fonts_light = {
        1: 'FF000000', 2: 'FFFFFFFF', 3: 'FF000000',
        4: 'FF000000', 5: 'FF000000', 6: 'FF000000'
    }
    # 改善A: Font/Fillオブジェクトをループ外で一括生成してキャッシュ
    _bf_cache       = {w: make_fill(color_fills[w])       for w in range(1, 7)}
    _bf_light_cache = {w: make_fill(color_fills_light[w]) for w in range(1, 7)}
    _ff_waku_cache  = {w: Font(name="Meiryo UI",        size=11, bold=True,  color=color_fonts[w])       for w in range(1, 7)}
    _ff_name_cache  = {w: Font(name="Noto Sans CJK SC", size=11, bold=True,  color=color_fonts_light[w]) for w in range(1, 7)}
    _ff_data_cache  = {w: Font(name="Meiryo UI",        size=10, bold=False, color=color_fonts_light[w]) for w in range(1, 7)}
    _ff_rate_cache  = {w: Font(name="Meiryo UI",        size=11, bold=True,  color=color_fonts_light[w]) for w in range(1, 7)}
    _ff_honmei_cache= {w: Font(name="Noto Sans CJK SC", size=10, bold=False, color=color_fonts_light[w]) for w in range(1, 7)}
    _fill_white     = make_fill('FFFFFFFF')
    _fill_missing   = make_fill("FFFCE4D6")
    _font_missing_rate = Font(name="Meiryo UI", size=10, bold=False, color="FFBF8F00")
    _font_course    = Font(name="Meiryo UI", size=10, bold=False, color='FF000000')
    for i, res in enumerate(results[:6]):
        dr = r + 4 + i
        waku_no = int(res['waku']) if str(res['waku']).isdigit() else i+1
        bf      = _bf_cache[waku_no]
        bf_light = _bf_light_cache[waku_no]
        fc      = color_fonts[waku_no]
        fc_light = color_fonts_light[waku_no]
        ff_waku = _ff_waku_cache[waku_no]
        ff_name = _ff_name_cache[waku_no]
        ff_data = _ff_data_cache[waku_no]
        ff_rate = _ff_rate_cache[waku_no]

        # 選手名: 姓（2文字）＋全角スペース2つ＋名
        # （年齢除去は calc_race_indices で実施済み）
        name = res['name']
        # 全角スペース・半角スペースで分割
        parts = [p for p in re.split(r'[\s\u3000]+', name) if p]
        if len(parts) >= 2:
            name_fmt = parts[0] + '　　' + parts[1]
        else:
            name_fmt = name
        # データ不足の場合は選手名に★を付加
        if res.get('data_missing'):
            name_fmt = '★' + name_fmt

        rel_w1 = f"{res['rel_win1']:.1f}%" if res['rel_win1'] is not None else '-'
        abs_w3 = f"{res['abs_win3']:.1f}%" if res.get('abs_win3') is not None else '-'

        wc(dr, 1, waku_no,   fill=bf,       font=ff_waku, alignment=c, border=bdr)
        wc(dr, 2, name_fmt,  fill=bf_light,  font=ff_name, alignment=c, border=bdr)
        wc(dr, 5, res['kumi'],   fill=bf_light, font=ff_data, alignment=c, border=bdr)
        wc(dr, 6, res['motor2'], fill=bf_light, font=ff_data, alignment=c, border=bdr)
        # G/H列：展示タイム偏差値（F/St影響列に表示）
        # 【軽微①】前日出力では展示未実施のため None → 「-前日」と表示して当日版と区別
        tenji_h = res.get("tenji_hensa")
        tenji_str = f"{tenji_h:.1f}" if tenji_h is not None else "前日"
        wc(dr, 7, tenji_str,     fill=bf_light, font=ff_data, alignment=c, border=bdr)
        wc(dr, 8, None,          fill=bf_light, font=ff_data, alignment=c, border=bdr)
        # I列はサンプルでFFFFFFFF（白）
        wc(dr, 9, res['course'], fill=_fill_white, font=_font_course, alignment=c, border=bdr)
        wc(dr,12, res['honmei'] if res['honmei'].strip() else None,
                                 fill=bf_light, font=_ff_honmei_cache[waku_no], alignment=c, border=bdr)
        # データ不足の場合は薄オレンジ背景に変更
        if res.get('data_missing'):
            rel_w1_disp = "－" if res['rel_win1'] is None else rel_w1
            abs_w3_disp = "－" if res.get('abs_win3') is None else abs_w3
            wc(dr,13, rel_w1_disp, fill=_fill_missing, font=_font_missing_rate, alignment=c, border=bdr)
            wc(dr,15, abs_w3_disp, fill=_fill_missing, font=_font_missing_rate, alignment=c, border=bdr)
        else:
            wc(dr,13, rel_w1,        fill=bf_light, font=ff_rate, alignment=c, border=bdr)
            wc(dr,15, abs_w3,        fill=bf_light, font=ff_data, alignment=c, border=bdr)

    # ── Row 11: 区切り（空行） ──
    sep_fill = make_fill("FFCCCCDD")
    for col in range(1, 17):
        ws.cell(r+10, col).fill = sep_fill

    # ── Row 12-13: 事前評価ヘッダ ──
    wc(r+11, 1, 'ボートリサーチ流  事前評価', fill=hf, font=Font(name="Noto Sans CJK SC", size=11, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+11, 7, 'イン逃げ決着時の2着優位度（相対%）', fill=shf, font=Font(name="Noto Sans CJK SC", size=11, bold=True, color="FF000000"), alignment=c)
    wc(r+11,14, '3着\n指数', fill=hf, font=Font(name="Meiryo UI", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+11,15, 'コース別\n3連対', fill=hf, font=Font(name="Meiryo UI", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    for col, txt in [(2,'イン\n逃げ'),(3,'相性'),(4,'機力'),(5,'自在性'),(6,'展開')]:
        wc(r+12, col, txt, fill=shf, font=Font(name="Meiryo UI" if col in (2,6) else "Noto Sans CJK SC", size=9, bold=True, color="FF000000"), alignment=c)

    # ── Row 14-19: 事前評価データ ──
    # 2着率テキストブロック（G14結合）
    circle_labels = {2:'黒', 3:'赤', 4:'青', 5:'黄', 6:'緑'}
    # 【重大②明確化】circle_pctはレース内正規化済みの「相対シェア%」
    # 「絶対的な2着確率」ではなく「このレースで2着になる相対優位度」を示す
    circle_lines = ['2着優位度（イン逃げ決着時・レース内相対%）', '']
    # circle_pct は calc_race_indices でレース内正規化済み
    circle_pct_map = {res['waku']: res.get('circle_pct') for res in results}
    for wk in range(2, 7):
        pct = circle_pct_map.get(str(wk))
        pct_str = f'{pct:>4.0f}%' if pct is not None else '   -%'
        circle_lines.append(f'{circle_labels[wk]}({wk}枠): {pct_str}')
    wc(r+13, 7, '\n'.join(circle_lines), alignment=left_align(wrap=True))

    # 事前評価記号
    jizen_result = None
    if JIZEN_AVAILABLE and _jizen_members:
        try:
            jizen_result = evaluate_all(_jizen_members)
        except Exception as e:
            print(f"  ⚠️  事前評価（write_race_flat内）でエラーが発生しました: {e}")

    # 2着率テキストブロック: results の circle_pct（正規化済み）を使用
    # total_2nd は描画用の再計算（frame_2nd は raw スコアなので circle_pct が正値）
    total_2nd = sum(v for k, v in frame_2nd.items() if k != "1" and v is not None)

    # 事前評価行・決まり手行の薄い背景色（サンプル準拠）
    jizen_light_fills = {
        1: 'FFFFFFFF', 2: 'FF000000', 3: 'FFFFCCCC',
        4: 'FFCCE8FF', 5: 'FFFFFACC', 6: 'FFCCFFEE'
    }
    jizen_light_fonts = {
        1: 'FF000000', 2: 'FFFFFFFF', 3: 'FF000000',
        4: 'FF000000', 5: 'FF000000', 6: 'FF000000'
    }
    # 改善A: 事前評価・決まり手ループ用キャッシュ
    _bf_j_cache    = {w: make_fill(jizen_light_fills[w]) for w in range(1, 7)}
    _ff_j11_cache  = {w: Font(name="Meiryo UI", size=11, bold=True,  color=color_fonts[w])       for w in range(1, 7)}
    _ff_j11l_cache = {w: Font(name="Meiryo UI", size=11, bold=True,  color=jizen_light_fonts[w]) for w in range(1, 7)}
    _ff_j10l_cache = {w: Font(name="Meiryo UI", size=10, bold=False, color=jizen_light_fonts[w]) for w in range(1, 7)}
    _fill_m5       = make_fill('FFE5DFEC')  # 黄艇M/N列専用
    _font_red      = Font(name="Meiryo UI", size=10, bold=False, color="FFCC0000")
    _align_cwrap   = center_align(wrap=True)
    for i, res in enumerate(results[:6]):
        dr = r + 13 + i
        waku_no = int(res['waku']) if str(res['waku']).isdigit() else i+1
        bf   = _bf_cache[waku_no]
        bf_j = _bf_j_cache[waku_no]
        ff   = _ff_j11_cache[waku_no]
        ff_j = _ff_j11l_cache[waku_no]
        wc(dr, 1, waku_no, fill=bf, font=ff, alignment=c, border=bdr)

        # B列: ほぼ白
        wc(dr, 2, None, fill=_fill_white, alignment=c)
        # C/D/E列: 薄い艇番色
        for col in [3, 4, 5]:
            wc(dr, col, None, fill=bf_j, alignment=c)
        # F列: waku1-3は白、waku4-6は薄い色
        f_fill = bf_j if waku_no >= 4 else _fill_white
        wc(dr, 6, None, fill=f_fill, alignment=c)

        # 事前評価記号
        if jizen_result:
            col_map = {'in_nige':2,'aisho':3,'kiryoku':4,'jizaisei':5,'tenkai':6}
            for key, col in col_map.items():
                sym = jizen_result[key][i]
                cell = ws.cell(dr, col)
                cell.value = sym if sym else None

        # N列（3着指数）: 薄い色
        wc(dr, 14, None, fill=bf_j, alignment=c)
        idx3 = res.get('idx3', 0)
        ws.cell(dr, 14).value = idx3 if idx3 else None
        # O/P列（オリジナル3連対）: 薄い色
        wc(dr, 15, None, fill=bf_j, alignment=c)
        abs_w3_ev = f"{res['abs_win3']:.1f}%" if res.get('abs_win3') is not None else '-'
        ws.cell(dr, 15).value = abs_w3_ev

    # ── Row 20: 区切り（空行） ──
    for col in range(1, 17):
        ws.cell(r+19, col).fill = sep_fill

    # ── Row 21: 決まり手タイトル・説明・場平均タイトル ──
    wc(r+20, 1, '決まり手 直近1年\n1=被決まり手 2〜6=決まり手', fill=hf, font=Font(name="Noto Sans CJK SC", size=8, bold=True, color="FFFFFFFF"), alignment=c)
    _exp = (
        'ボートリサーチ流 事前評価（展示前の評価）\n\n'
        '◇イン逃げ … 出走メンバーで相対評価。\n  イン逃げ信頼度 (◎>○>△>×)\n\n'
        '◇相性 … 1枠に対する相性\n  (◎>○>△ ※空白は平凡)\n\n'
        '◇機力 … 出走メンバーで機力を相対評価\n  (A>B>C>D>E)\n\n'
        '◇自在性 … 自ら動きレース展開を作れるか\n  (◎>○>△ ※空白は平凡)\n\n'
        '◇展開 … 4~6枠で展開が向いた時の対応力\n  (◎>○>△ ※空白は平凡)'
    )
    wc(r+20, 6, _exp, alignment=left_align(wrap=True))
    wc(r+20,12, '場平均', fill=hf, font=hfont, alignment=c)

    # ── Row 22: 決まり手ヘッダ・場平均ヘッダ ──
    for col, txt in [(1,'枠'),(2,'逃げ%'),(3,'差し%\n差された%'),(4,'まくり%\n捲られた%'),(5,'まくり差し%\n捲り差された%')]:
        wc(r+21, col, txt, fill=hf, font=Font(name="Noto Sans CJK SC", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    for col, txt in [(12,'枠'),(13,'1-○'),(15,'1着率')]:
        wc(r+21, col, txt, fill=shf, font=Font(name="Noto Sans CJK SC", size=9, bold=True, color="FF000000"), alignment=c)

    # ── Row 23-28: 決まり手データ・場平均データ ──
    for i, res in enumerate(results[:6]):
        dr = r + 22 + i
        waku_no = int(res['waku']) if str(res['waku']).isdigit() else i+1
        bf       = _bf_cache[waku_no]
        bf_light = _bf_j_cache[waku_no]
        ff       = _ff_j11_cache[waku_no]
        ff_light = _ff_j10l_cache[waku_no]
        cm = res.get('raw_cm', {})

        wc(dr, 1, str(waku_no), fill=bf, font=ff, alignment=c, border=bdr)

        # 決まり手%（コース別マスタから）B〜E列に薄い背景色を設定
        def pct_str(key):
            v = safe_float(_get_cm_val(cm, key))
            return f'{v*100:.0f}%' if v else '-'

        # B列: 逃げ% → 1号艇のみ表示、2〜6はハイフン
        if waku_no == 1:
            wc(dr, 2, pct_str('逃げ%'), fill=_fill_white, font=ff_light, alignment=c)
        else:
            wc(dr, 2, '-', fill=_fill_white, font=ff_light, alignment=c)

        # C/D/E列:
        #   1号艇 → 被決まり手%（差された%・捲られた%・捲り差された%）赤字
        #   2〜6号艇 → 決まり手%（差し%・まくり%・まくり差し%）
        if waku_no == 1:
            def lose_pct_str(key):
                v = safe_float(cm.get(key))
                return f'{v*100:.0f}%' if v else '-'
            wc(dr, 3, lose_pct_str('差され%'),
               fill=bf_light, font=_font_red, alignment=_align_cwrap)
            wc(dr, 4, lose_pct_str('捲られ%'),
               fill=bf_light, font=_font_red, alignment=_align_cwrap)
            wc(dr, 5, lose_pct_str('捲り差され%'),  # update_master出力キーに統一
               fill=bf_light, font=_font_red, alignment=_align_cwrap)
        else:
            wc(dr, 3, pct_str('差し%'),       fill=bf_light, font=ff_light, alignment=c)
            wc(dr, 4, pct_str('まくり%'),     fill=bf_light, font=ff_light, alignment=c)
            wc(dr, 5, pct_str('まくり差し%'), fill=bf_light, font=ff_light, alignment=c)

        # 場平均（イン逃げ時2着率）L列は艇番色
        wc(dr,12, str(waku_no), fill=bf, font=ff, alignment=c, border=bdr)
        # circle_pct はレース内正規化済みの値を results から取得
        pct_2nd_norm = res.get('circle_pct', 0) or 0
        # M/N列: 薄い色（黄艇のみT7 = FFE5DFEC）
        m_fill = _fill_m5 if waku_no == 5 else bf_light
        wc(dr,13, f'{pct_2nd_norm:.0f}%' if total_2nd > 0 else '-', fill=m_fill, font=ff_light, alignment=c)
        # 1着率（相対1着率）O/P列
        rel_w1 = f"{res['rel_win1']:.1f}%" if res['rel_win1'] is not None else '-'
        wc(dr,15, rel_w1, fill=m_fill, font=ff_light, alignment=c)

    # ── Row 29: 注記 ＋ ヒモ荒れ判定 ──
    himo_are = (race_judgment or {}).get("himo_are", {}) if race_judgment else {}
    himo_verdict = himo_are.get("verdict", "対象外")
    if himo_verdict == "不参加推奨":
        mcp = himo_are.get("max_combo_prob", 0.0) or 0.0
        eto = himo_are.get("est_top_odds",   0.0) or 0.0
        cc  = himo_are.get("circle_concentration", 0.0) or 0.0
        himo_str = (
            f"【🚫 ヒモ固まり・見送り推奨】"
            f"最有力確率{mcp:.3f}（推定最高人気{eto:.0f}倍台）/ 2着集中度{cc:.0f}%\n"
            f"→ 1号艇1着固定でも3連単オッズが構造的に低い。参加しない方が期待値が高い。\n"
        )
        himo_fill = make_fill("FFFCE4D6")   # 薄オレンジ（警告色）
        himo_font = Font(name="Noto Sans CJK SC", size=9, bold=True, color="FFCC0000")
    elif himo_verdict == "点数絞り":
        mcp = himo_are.get("max_combo_prob", 0.0) or 0.0
        eto = himo_are.get("est_top_odds",   0.0) or 0.0
        cc  = himo_are.get("circle_concentration", 0.0) or 0.0
        himo_str = (
            f"【⚠ ヒモやや固め・点数絞り】"
            f"最有力確率{mcp:.3f}（推定最高人気{eto:.0f}倍台）/ 2着集中度{cc:.0f}%\n"
            f"→ 展示後にヒモを上位2艇に絞り込むこと。\n"
        )
        himo_fill = make_fill("FFFFFFF0")   # 薄黄（注意色）
        himo_font = Font(name="Noto Sans CJK SC", size=9, bold=False, color="FF996600")
    elif himo_verdict == "参加推奨":
        mcp = himo_are.get("max_combo_prob", 0.0) or 0.0
        eto = himo_are.get("est_top_odds",   0.0) or 0.0
        cc  = himo_are.get("circle_concentration", 0.0) or 0.0
        himo_str = (
            f"【✅ ヒモ分散・参加推奨】"
            f"最有力確率{mcp:.3f}（推定最高人気{eto:.0f}倍台）/ 2着集中度{cc:.0f}%\n"
            f"→ 1号艇1着固定でヒモ広めに流す。買い目を+2点追加推奨。\n"
        )
        himo_fill = make_fill("FFE2EFDA")   # 薄緑（推奨色）
        himo_font = Font(name="Noto Sans CJK SC", size=9, bold=False, color="FF375623")
    else:
        himo_str  = ""
        himo_fill = None
        himo_font = None

    note_base = (
        '※オリジナル1着率、2着率、3着指数は出走メンバーで相対評価しており、天候、潮、展示、場特性などを考慮していません。\n'
        '※枠内を想定します。進入変更があった場合使用できません。'
    )
    note_val  = (himo_str + note_base) if himo_str else note_base
    note_fill = himo_fill or make_fill("FFFFFFFF")
    note_font = himo_font or Font(name="Meiryo UI", size=8)
    wc(r+28, 1, note_val,
       fill=note_fill, font=note_font,
       alignment=left_align(wrap=True))

    # ── 行高（サンプル完全再現） ──
    row_heights = {
        0:12.75, 1:30.0, 2:24.75, 3:31.5,
        4:27.75, 5:27.75, 6:27.75, 7:27.75, 8:27.75, 9:27.75,
        10:4.5, 11:25.5, 12:21.75,
        13:25.5, 14:25.5, 15:25.5, 16:25.5, 17:25.5, 18:25.5,
        19:4.5, 20:25.5, 21:21.75,
        22:25.5, 23:25.5, 24:25.5, 25:25.5, 26:25.5, 27:25.5,
        28:39.0
    }
    for offset, height in row_heights.items():
        ws.row_dimensions[r + offset].height = height

    # ── ST順位 舟図をK5:K10に埋め込む ──
    # 改善C: tempfileをやめてBytesIOインメモリ処理に変更（ディスクI/O削減）
    if player_master is not None:
        try:
            buf = _make_st_boat_chart(results, player_master, outpath=None)
            if buf is not None:
                img = XLImage(buf)
                # K列(col=10, 0始まり), 選手データ開始行(r+4, 0始まり=r+3)
                marker = AnchorMarker(col=10, colOff=0, row=r+3, rowOff=0)
                size   = XDRPositiveSize2D(cx=865179, cy=2082127)
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
        except Exception as e:
            print(f"  ⚠️  ST舟図生成エラー ({race_no}R): {e}")


# ============================================================
# 数値データシート書き込み（出力_数値と同一テーブル形式）
# ============================================================
def write_numeric_sheet(wb, all_race_data):
    """
    「会場名_数値」シートを「出力_数値」と完全同一のテーブル形式で書き出す。
    縦軸: 分類・項目・艇番、横軸: 1R〜12R（最大レース数分）
    """
    if not all_race_data:
        return

    venue_name = all_race_data[0]["venue"]
    sheet_name = f"{venue_name}_数値"

    # ── FLY手入力値を退避（load_race再実行でシートが再作成されても消えないように）──
    fly_input_backup = {}  # {(race_no_str, waku_int): value}
    if sheet_name in wb.sheetnames:
        _ws_old = wb[sheet_name]
        # ヘッダ行(2行目)からレース番号→列のマッピングを取得
        _old_race_col = {}
        for _c in range(1, _ws_old.max_column + 1):
            _v = _ws_old.cell(2, _c).value
            if _v and isinstance(_v, str) and "R" in _v:
                import re as _re
                _m = _re.match(r"(\d+)R", _v.strip())
                if _m:
                    _old_race_col[_m.group(1)] = _c
        # FLY入力行を探してバックアップ
        for _r in range(1, _ws_old.max_row + 1):
            if _ws_old.cell(_r, 1).value == "FLY入力":
                _waku_val = _ws_old.cell(_r, 3).value
                try:
                    _waku_int = int(_waku_val)
                except (TypeError, ValueError):
                    continue
                for _rno, _col in _old_race_col.items():
                    _fv = _ws_old.cell(_r, _col).value
                    if _fv is not None and str(_fv).strip() not in ("", "0", "None"):
                        fly_input_backup[(_rno, _waku_int)] = _fv
        if fly_input_backup:
            print(f"  ✈️  FLY手入力値を退避: {len(fly_input_backup)}件")
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ── スタイル定数（出力_数値から実測） ──
    BOAT_FILL   = {1:"FFFFFFFF", 2:"FF1A1A1A", 3:"FFCC0000", 4:"FF2255CC", 5:"FFDDCC00", 6:"FF115511"}
    BOAT_FONT   = {1:"FF000000", 2:"FFFFFFFF", 3:"FFFFFFFF", 4:"FFFFFFFF", 5:"FF000000", 6:"FFFFFFFF"}
    FILL_HDR    = "FF1F4E79"   # ヘッダ行（濃紺）
    FILL_MISSING = "FFFCE4D6"  # データ不足セル（薄オレンジ）
    FILL_SEC_B  = "FF2E75B6"   # 数値指標・選手情報セクション（青）
    FILL_SEC_G  = "FF70AD47"   # 決まり手セクション（緑）
    FILL_ITEM_B = "FFDCE6F1"   # 数値指標・選手情報 項目セル（薄青）
    FILL_ITEM_G = "FFE2EFDA"   # 決まり手 項目セル（薄緑）
    FILL_ITEM_P = "FFD9E1F2"   # 選手情報 項目セル（薄紺）

    race_nos = [rd["race_no"] for rd in all_race_data]
    n_races  = len(race_nos)
    _first_results = all_race_data[0].get("results", [])
    _first_rel = _first_results[0].get("rel_win1") if _first_results else "なし"
    print(f"  📊 {sheet_name}: {n_races}レース, 1R results件数={len(_first_results)}, rel_win1={_first_rel}")

    def sf(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def fn(bold=True, size=9, color="FF000000", name="Meiryo UI"):
        return Font(name=name, size=size, bold=bold, color=color)

    def al(h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def wc(row, col, val, fill=None, font=None, align=None, border=None):
        c = ws.cell(row=row, column=col)
        c.value = val
        if fill:   c.fill   = fill
        if font:   c.font   = font
        if align:  c.alignment = align
        if border: c.border = border

    thin = Side(style="thin", color="FFCCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 列幅 ──
    ws.column_dimensions["A"].width = 14.0
    ws.column_dimensions["B"].width = 20.0
    ws.column_dimensions["C"].width = 7.0
    for i in range(n_races):
        ws.column_dimensions[get_column_letter(4 + i)].width = 14.0

    # ── Row 1: タイトル ──
    race_date = all_race_data[0].get("race_date", "")
    title = f"ボートリサーチ数値データ　【{venue_name}】　{race_date}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + n_races)
    wc(1, 1, title, font=fn(bold=True, size=12), align=al("left"))
    ws.row_dimensions[1].height = 22.0

    # ── Row 2: ヘッダ（分類・項目・艇番・1R〜nR） ──
    hf = sf(FILL_HDR)
    hfn = fn(bold=True, color="FFFFFFFF")
    for col, txt in [(1,"分類"),(2,"項目"),(3,"艇番")]:
        wc(2, col, txt, fill=hf, font=hfn, align=al(), border=bdr)
    for i, rno in enumerate(race_nos):
        deadline = all_race_data[i].get("deadline") if i < len(all_race_data) else None
        header_val = f"{rno}R\n{deadline}" if deadline else f"{rno}R"
        wc(2, 4+i, header_val, fill=hf, font=hfn, align=al(wrap=True), border=bdr)
    ws.row_dimensions[2].height = 28.0

    # ── セクション書き込みヘルパー ──
    def write_section_header(row, label, fill_color):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
        wc(row, 1, label, fill=sf(fill_color), font=fn(bold=True, color="FFFFFFFF"), align=al("left"))
        ws.row_dimensions[row].height = 13.0

    def write_item_block(row_start, cat_label, item_label, cat_fill, item_fill, data_by_waku, missing_by_waku=None):
        """
        data_by_waku:   {1: [r1val, r2val, ...], 2: [...], ...}
        missing_by_waku:{1: [bool, bool, ...], 2: [...], ...}  データ不足フラグ
        6艇分 × nレース のデータブロックを書く
        """
        for i in range(6):
            waku = i + 1
            row  = row_start + i
            # A列（分類）: 最初の艇番のみ
            if waku == 1:
                wc(row, 1, cat_label,  fill=sf(cat_fill),  font=fn(bold=True, color="FFFFFFFF"), align=al())
                wc(row, 2, item_label, fill=sf(item_fill), font=fn(bold=True, color="FF000000"), align=al("left"))
            else:
                wc(row, 1, None, fill=sf(cat_fill), align=al())
                wc(row, 2, None, fill=sf(item_fill), align=al())
            # C列（艇番）
            wc(row, 3, str(waku),
               fill=sf(BOAT_FILL[waku]),
               font=fn(bold=True, color=BOAT_FONT[waku]),
               align=al(), border=bdr)
            # D〜 データ列
            vals    = data_by_waku.get(waku,    [None]*n_races)
            missing = (missing_by_waku or {}).get(waku, [False]*n_races)
            for j, v in enumerate(vals):
                is_missing = missing[j] if j < len(missing) else False
                if is_missing and v is None:
                    # データ不足セル: 薄オレンジ背景 ＋ 「－」テキスト
                    wc(row, 4+j, "－",
                       fill=sf(FILL_MISSING),
                       font=fn(bold=False, color="FFBF8F00"),
                       align=al(), border=bdr)
                else:
                    wc(row, 4+j, v, font=fn(bold=False), align=al(), border=bdr)
            ws.row_dimensions[row].height = 15.0

    # ── データ取り出しヘルパー ──
    def get_vals(rd, key):
        """results から艇番→値のdictを返す"""
        d = {}
        for res in rd.get("results", []):
            try: waku = int(res["waku"])
            except (ValueError, TypeError): continue
            d[waku] = res.get(key)
        return d

    def get_cm_vals(rd, cm_key, multiplier=1.0):
        d = {}
        for res in rd.get("results", []):
            try: waku = int(res["waku"])
            except (ValueError, TypeError): continue
            v = safe_float(res.get("raw_cm", {}).get(cm_key))
            d[waku] = round(v * multiplier, 1) if v is not None else None
        return d

    # 6艇×nレース のデータを転置して {waku: [r1,r2,...]} に変換
    def build_waku_data(extract_fn):
        result = {w: [] for w in range(1, 7)}
        for rd in all_race_data:
            vals = extract_fn(rd)
            for w in range(1, 7):
                result[w].append(vals.get(w))
        return result

    # ── データ不足フラグを {waku: [bool×nレース]} に変換するヘルパー ──
    def build_missing_waku():
        """各艇番・各レースについて data_missing フラグを返す"""
        result = {w: [] for w in range(1, 7)}
        for rd in all_race_data:
            missing_map = {}
            for res in rd.get("results", []):
                try: waku = int(res["waku"])
                except (ValueError, TypeError): continue
                missing_map[waku] = bool(res.get("data_missing", False))
            for w in range(1, 7):
                result[w].append(missing_map.get(w, False))
        return result

    missing_waku = build_missing_waku()

    # ── ▼ 数値指標セクション ──
    row = 3
    write_section_header(row, "▼ 数値指標", FILL_SEC_B)
    row += 1

    # オリジナル1着率
    write_item_block(row, "数値指標", "オリジナル1着率(%)", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): f"{round(r['rel_win1'], 1)}%" if r.get("rel_win1") is not None else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }), missing_by_waku=missing_waku)
    row += 6

    # オリジナル3連対率
    write_item_block(row, "数値指標", "コース別3連対率(%)", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): f"{round(r['abs_win3'], 1)}%" if r.get("abs_win3") is not None else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }), missing_by_waku=missing_waku)
    row += 6

    # イン逃げ時2着率（1号艇は対象外のため表示なし）
    write_item_block(row, "数値指標", "2着優位度(%)[相対・イン逃げ時]", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): (
                None if r["waku"] == "1"
                else f"{round(r.get('circle_pct'), 1)}%" if r.get("circle_pct") is not None else "-"
            )
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 3着指数
    write_item_block(row, "数値指標", "3着指数", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("idx3") if r.get("idx3") else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── ▼ 決まり手セクション ──
    write_section_header(row, "▼ 決まり手（直近1年｜1号艇=被決まり手%、2〜6号艇=決まり手%）", FILL_SEC_G)
    row += 1

    # 被決まり手マップ: {(レースindex, waku=1): {差され%: v, 捲られ%: v, 捲り差され%: v}}
    # 差し%行→1号艇欄に差され%、まくり%行→捲られ%、まくり差し%行→捲り差され% を赤字で表示
    LOSE_KEY_MAP = {
        "差し%":      "差され%",
        "まくり%":    "捲られ%",
        "まくり差し%":"捲り差され%",
    }

    for cm_key, label in [
        ("逃げ%",      "逃げ%"),
        ("差し%",      "差し%"),
        ("まくり%",    "まくり%"),
        ("まくり差し%","まくり差し%"),
        ("抜き%",      "抜き%"),
    ]:
        lose_key = LOSE_KEY_MAP.get(cm_key)  # 対応する被決まり手キー（なければNone）

        # 各艇番×レースのデータを構築
        main_data = build_waku_data(lambda rd, k=cm_key: {
            int(r["waku"]): safe_float(r.get("raw_cm", {}).get(k), 0)
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        })
        lose_data = build_waku_data(lambda rd, lk=lose_key: {
            int(r["waku"]): safe_float(r.get("raw_cm", {}).get(lk)) if lk and int(r["waku"]) == 1 else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }) if lose_key else None

        for i in range(6):
            waku = i + 1
            dr   = row + i
            if waku == 1:
                wc(dr, 1, "決まり手", fill=sf(FILL_SEC_G),  font=fn(bold=True, color="FFFFFFFF"), align=al())
                wc(dr, 2, label,      fill=sf(FILL_ITEM_G), font=fn(bold=True, color="FF000000"), align=al("left"))
            else:
                wc(dr, 1, None, fill=sf(FILL_SEC_G),  align=al())
                wc(dr, 2, None, fill=sf(FILL_ITEM_G), align=al())
            wc(dr, 3, str(waku),
               fill=sf(BOAT_FILL[waku]),
               font=fn(bold=True, color=BOAT_FONT[waku]),
               align=al(), border=bdr)

            main_vals = main_data.get(waku, [None]*n_races)
            lose_vals = lose_data.get(waku, [None]*n_races) if lose_data else [None]*n_races

            for j, mv in enumerate(main_vals):
                lv = lose_vals[j] if j < len(lose_vals) else None
                main_str = f"{round((mv or 0) * 100, 1)}%" if mv not in (None,) else "-"

                if waku == 1 and lose_key:
                    # 1号艇：被決まり手%のみ表示（差された・捲られた・捲り差された）
                    lose_str = f"{round(lv * 100, 1)}%" if lv not in (None, 0.0) else "-"
                    wc(dr, 4+j, lose_str,
                       fill=sf(FILL_ITEM_G),
                       font=fn(bold=False, size=9, color="FFCC0000"),
                       align=al(), border=bdr)
                elif waku == 1:
                    # 1号艇の逃げ%行はそのまま表示
                    wc(dr, 4+j, main_str,
                       font=fn(bold=False), align=al(), border=bdr)
                elif cm_key == "逃げ%":
                    # 2〜6号艇の逃げ%行は表示不要（コース1専用）
                    wc(dr, 4+j, "-",
                       font=fn(bold=False, color="FF999999"), align=al(), border=bdr)
                else:
                    # 2〜6号艇：決まり手%をそのまま表示
                    wc(dr, 4+j, main_str,
                       font=fn(bold=False), align=al(), border=bdr)
            ws.row_dimensions[dr].height = 15.0
        row += 6

    # ── ▼ 事前評価セクション ──
    FILL_SEC_J  = "FF7030A0"   # 事前評価セクション（紫）
    FILL_ITEM_J = "FFE9D7F5"   # 事前評価 項目セル（薄紫）

    if JIZEN_AVAILABLE:
        write_section_header(row, "▼ 事前評価（ボートリサーチ流｜展示前）", FILL_SEC_J)
        row += 1

        JIZEN_ITEMS = [
            ("in_nige",  "①逃げ"),
            ("aisho",    "②相性"),
            ("kiryoku",  "③機力"),
            ("jizaisei", "④自在性"),
            ("tenkai",   "⑤展開"),
        ]

        JIZEN_FONT_COLOR = {
            "◎": "FF7F6000", "◎?": "FF7F6000",
            "○": "FF1F3864",
            "△": "FF843C0C",
            "A": "FF375623", "B": "FF375623",
            "C": "FF404040",
            "D": "FF843C0C", "E": "FF7F0000",
        }
        JIZEN_FILL_COLOR = {
            "◎": "FFFFF2CC", "◎?": "FFFFF2CC",
            "○": "FFDAE3F3",
            "△": "FFFCE4D6",
            "A": "FFE2EFDA", "B": "FFE2EFDA",
            "C": "FFF2F2F2",
            "D": "FFFCE4D6", "E": "FFFFCCCC",
        }

        # レースごとに evaluate_all を呼んで結果をキャッシュ
        jizen_cache = []
        for rd in all_race_data:
            jm = rd.get("jizen_members")
            if jm:
                try:
                    jizen_cache.append(evaluate_all(jm))
                except Exception:
                    jizen_cache.append(None)
            else:
                jizen_cache.append(None)

        for jizen_key, jizen_label in JIZEN_ITEMS:
            for i in range(6):
                waku = i + 1
                dr   = row + i
                if waku == 1:
                    wc(dr, 1, "事前評価",   fill=sf(FILL_SEC_J),  font=fn(bold=True, color="FFFFFFFF"), align=al())
                    wc(dr, 2, jizen_label,  fill=sf(FILL_ITEM_J), font=fn(bold=True, color="FF000000"), align=al("left"))
                else:
                    wc(dr, 1, None, fill=sf(FILL_SEC_J),  align=al())
                    wc(dr, 2, None, fill=sf(FILL_ITEM_J), align=al())
                wc(dr, 3, str(waku),
                   fill=sf(BOAT_FILL[waku]),
                   font=fn(bold=True, color=BOAT_FONT[waku]),
                   align=al(), border=bdr)
                for j, jr in enumerate(jizen_cache):
                    sym = (jr.get(jizen_key, [""] * 6)[i] or "") if jr else ""
                    cell_fill = sf(JIZEN_FILL_COLOR.get(sym, "FFFFFFFF")) if sym else None
                    cell_font = fn(bold=(sym in ("◎", "◎?", "A")),
                                   size=9,
                                   color=JIZEN_FONT_COLOR.get(sym, "FF808080"))
                    wc(dr, 4 + j, sym if sym else "－",
                       fill=cell_fill, font=cell_font,
                       align=al(), border=bdr)
                ws.row_dimensions[dr].height = 15.0
            row += 6

    # ── ▼ 選手情報セクション ──
    write_section_header(row, "▼ 選手情報", FILL_SEC_B)
    row += 1

    # 選手名
    write_item_block(row, "選手情報", "選手名", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("name", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 組
    write_item_block(row, "選手情報", "組", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("kumi", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # モータ2連
    write_item_block(row, "選手情報", "モータ2連", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("motor2", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 本命記号
    write_item_block(row, "選手情報", "本命記号", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("honmei", "").strip() or None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 想定コース
    write_item_block(row, "選手情報", "想定コース", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("course", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── FLY手動入力行（想定コース直後に挿入）────────────────────────────────
    # ユーザーが各レース列に手動でフライング数を入力するための行
    from openpyxl.styles import PatternFill as _PF, Font as _Fn, Alignment as _Al, Border as _Br, Side as _Sd
    _fill_red    = _PF("solid", fgColor="FFCC0000")
    _fill_yellow = _PF("solid", fgColor="FFFFFF00")
    _fill_boat   = {1:"FFFFFFFF",2:"FF1A1A1A",3:"FFCC0000",4:"FF2255CC",5:"FFDDCC00",6:"FF115511"}
    _fill_font   = {1:"FF000000",2:"FFFFFFFF",3:"FFFFFFFF",4:"FFFFFFFF",5:"FF000000",6:"FFFFFFFF"}
    _thin = _Sd(style="thin", color="FFCCCCCC")
    _bdr  = _Br(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _fn_w = _Fn(name="Meiryo UI", size=9, bold=True,  color="FFFFFFFF")
    _fn_b = _Fn(name="Meiryo UI", size=9, bold=False, color="FF000000")
    _al_c = _Al(horizontal="center", vertical="center")
    for _i, _waku_no in enumerate(range(1, 7)):
        _r = row + _i
        _c = ws.cell(_r, 1)
        _c.value, _c.fill, _c.font, _c.alignment, _c.border = "FLY入力", _fill_red, _fn_w, _al_c, _bdr
        _c = ws.cell(_r, 2)
        _c.value, _c.fill, _c.font, _c.alignment, _c.border = "FLY（1=あり）", _fill_yellow, _fn_b, _al_c, _bdr
        _c = ws.cell(_r, 3)
        _c.value = _waku_no
        _c.fill  = _PF("solid", fgColor=_fill_boat[_waku_no])
        _c.font  = _Fn(name="Meiryo UI", size=9, bold=True, color=_fill_font[_waku_no])
        _c.alignment, _c.border = _al_c, _bdr
        for _col in range(4, 4 + n_races):
            _cell = ws.cell(_r, _col)
            # バックアップから手入力値を復元
            _rno_str = race_nos[_col - 4] if (_col - 4) < len(race_nos) else None
            _restored = fly_input_backup.get((_rno_str, _waku_no)) if _rno_str else None
            _cell.value  = _restored
            _cell.fill   = _fill_yellow
            _cell.border = _bdr
        ws.row_dimensions[_r].height = 15.0
    row += 6

    # FLY数
    write_item_block(row, "選手情報", "FLY数", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("fly_count", 0)
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # F/ST影響ラベル
    write_item_block(row, "選手情報", "F/ST影響", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("fly_label", "低")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── 注記 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "※オリジナル1着率・オリジナル3連対率・イン逃げ時2着率・3着指数は出走メンバーで正規化。決まり手%は過去コース別1着時の割合。",
       font=fn(bold=False, size=8), align=al("left"))
    ws.row_dimensions[row].height = 13.0
    row += 1

    # ── データ不足凡例 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "★凡例：薄オレンジ「－」はデータ不足（コース別実績 < 5走 or 総実績 < 10走）。新人・移籍直後の選手はデータ蓄積後に正確な指数が算出されます。",
       fill=sf(FILL_MISSING),
       font=fn(bold=False, size=8, color="FF7F3F00"),
       align=al("left"))
    ws.row_dimensions[row].height = 14.0

    # ── データ不足選手一覧（シート末尾サマリー） ──
    missing_summary = []
    for rd in all_race_data:
        for res in rd.get("results", []):
            if res.get("data_missing") and res.get("missing_reason"):
                missing_summary.append(
                    f"{rd['race_no']}R-{res['waku']}号艇 {res['name']}：{res['missing_reason']}"
                )
    if missing_summary:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
        wc(row, 1, "【データ不足選手一覧】",
           fill=sf("FFFCE4D6"),
           font=fn(bold=True, size=9, color="FF7F3F00"),
           align=al("left"))
        ws.row_dimensions[row].height = 14.0
        for summary_line in missing_summary:
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
            wc(row, 1, summary_line,
               fill=sf(FILL_MISSING),
               font=fn(bold=False, size=9, color="FF7F3F00"),
               align=al("left"))
            ws.row_dimensions[row].height = 13.0

    # ════════════════════════════════════════════════════════════════════
    # 📋 展示前参考買い目 ＋ 考察セクション
    # ════════════════════════════════════════════════════════════════════
    FILL_JF_HDR  = "FF1F4E79"   # 濃紺（ヘッダ）
    FILL_JF_SUB  = "FF2E75B6"   # 中青（行ラベル）
    FILL_JF_BODY = "FFDCE6F1"   # 薄青（データ）
    FILL_JF_BET  = "FFFFFF99"   # 薄黄（買い目リスト）
    FILL_JF_NOTE = "FFFFE699"   # 黄（絞込ガイド）

    FILL_KS_HDR  = "FF203864"   # 濃紺（考察ヘッダ）
    FILL_KS_SUB  = "FF305496"   # 中紺（考察行ラベル）
    FILL_KS_BODY = "FFD9E1F2"   # 薄紺（考察データ）
    FILL_KS_GOOD = "FFE2EFDA"   # 薄緑（良好）
    FILL_KS_WARN = "FFFFF2CC"   # 薄黄（注意）
    FILL_KS_BAD  = "FFFCE4D6"   # 薄橙（警告）

    row += 2

    # ── 考察セクションヘッダ ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "━━━  🔍 Step1: レース考察  ━━━  "
       "展開予想の根拠（判定スコア・相性・展開・ヒモ荒れ）→ 次の買い目の根拠になります",
       fill=sf(FILL_KS_HDR), font=fn(bold=True, size=9, color="FFFFFFFF"), align=al("left"))
    ws.row_dimensions[row].height = 16.0
    row += 1

    KOSATSU_ROWS = [
        ("①判定スコア",   "ランク/スコア/戦略"),
        ("①判定根拠",     "スコア構成要素（主要3件）"),
        ("②3択判定",      "逃げ/飛び/両建て 根拠"),
        ("②買い方指示",   "具体的な買い方"),
        ("③相性考察",     "攻撃艇の根拠（攻撃力順）"),
        ("④展開予測",     "1M到達順（推定）・先行・主軸対立・漁夫候補"),
        ("⑤注意事項",     "FLY・データ不足・ST不安定"),
        ("⑥ヒモ荒れ",     "1号艇強本命時のヒモ分散判定"),
        ("⑦展開quality", "展開の絞れ度・買い目点数ガイド"),
        ("⑧考察の結論",   "考察→買い目 因果まとめ / Step2へ"),
    ]

    for sec_label, sec_item in KOSATSU_ROWS:
        wc(row, 1, sec_label,
           fill=sf(FILL_KS_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        wc(row, 2, sec_item,
           fill=sf(FILL_KS_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())

        max_lines = 2
        for i, rd in enumerate(all_race_data):
            rj  = rd.get("race_judgment", {}) or {}
            bet = rd.get("bet_suggestions", {}) or {}
            je  = rd.get("jizen_eval", {}) or {}
            jf  = bet.get("jizen_formation", {}) or {}
            ryotate = bet.get("ryotate_detail", rj.get("ryotate", {})) or {}
            txt = ""
            fc  = FILL_KS_BODY

            if sec_label == "①判定スコア":
                rank  = rj.get("rank", "-")
                score = rj.get("score", "-")
                strat = rj.get("strategy", "-")
                trust = rj.get("data_trust_score", "-")
                txt = f"【ランク{rank}】{score}点\n戦略: {strat}\n信頼度: {trust}%"
                fc = FILL_KS_GOOD if rank in ("S","A") else FILL_KS_WARN if rank == "B" else FILL_KS_BAD

            elif sec_label == "①判定根拠":
                reasons = rj.get("reason", []) or []
                filtered = [r for r in reasons if not r.startswith("複合確率スコア")][:3]
                txt = "\n".join(f"・{r}" for r in filtered) if filtered else "-"

            elif sec_label == "②3択判定":
                verdict = ryotate.get("verdict", "-")
                conf    = ryotate.get("confidence", "-")
                esc_pct = ryotate.get("escape_pct")
                tobi_pct = ryotate.get("tobi_pct")

                # 1行サマリー: 「なぜこの3択か」を1文で表現
                if verdict == "逃げ狙い":
                    if esc_pct is not None:
                        summary = f"→ 逃げ{esc_pct:.0f}%: 1号艇先行優勢・ヒモ流しで参加"
                    else:
                        summary = "→ 1号艇先行優勢・逃げ軸"
                elif verdict == "飛び狙い":
                    main_th = (rj.get("affinity") or {}).get("dominant_attacker", "?")
                    if tobi_pct is not None:
                        summary = f"→ 飛び{tobi_pct:.0f}%: {main_th}号艇が主な脅威・飛び軸"
                    else:
                        summary = f"→ {main_th}号艇先行・飛び軸"
                else:
                    if esc_pct is not None and tobi_pct is not None:
                        summary = f"→ 逃げ{esc_pct:.0f}%/飛び{tobi_pct:.0f}%拮抗・展示で軸決定"
                    else:
                        summary = "→ 拮抗展開・展示後に軸を傾ける"

                reason = ryotate.get("reason", "-")
                # reason は長いので先頭の警告部分のみ抽出（⚠以降の重要部分）
                reason_short = reason.split("。")[0] if reason != "-" else "-"

                txt = (
                    f"【{verdict}】確信度{conf}%\n"
                    f"{summary}\n"
                    f"根拠: {reason_short}"
                )
                fc = FILL_KS_GOOD if verdict == "逃げ狙い" else FILL_KS_BAD if verdict == "飛び狙い" else FILL_KS_WARN

            elif sec_label == "②買い方指示":
                txt = ryotate.get("buy_style", "-") or "-"

            elif sec_label == "③相性考察":
                affinity = rj.get("affinity", {}) or {}
                summary  = affinity.get("affinity_summary", {}) or {}
                attack   = affinity.get("attack_score", {}) or {}
                outer = sorted(
                    [(w, summary.get(w, "-"), attack.get(w, 0))
                     for w in ["2","3","4","5","6"] if w in summary],
                    key=lambda x: x[2], reverse=True
                )[:4]
                txt = "\n".join(f"{w}号: {s}（攻{a:.0f}pt）" for w, s, a in outer) or "データなし"

            elif sec_label == "④展開予測":
                ft   = rj.get("first_turn", {}) or {}
                cm   = rj.get("conflict_map", {}) or {}
                # 1M到達順序
                entry = ft.get("entry_order", [])
                entry_str = " → ".join([f"{w}号" for w, _ in entry[:4]]) if entry else "-"
                lead_w = ft.get("lead_waku", "-")
                pattern = ft.get("pattern", "-")
                p_strength = ft.get("pattern_strength", "-")
                # 対立構造
                mc = cm.get("main_conflict") or {}
                sc = cm.get("sub_conflict") or {}
                mc_desc = mc.get("desc", "-") if mc else "-"
                sc_desc = sc.get("desc", "-") if sc else "-"
                # 漁夫受益
                cb = cm.get("collapse_beneficiary", [])
                cb_str = "・".join([f"{w}号" for w, _ in cb[:3]]) if cb else "-"
                txt = (
                    f"1M到達順: {entry_str}\n"
                    f"先行: {lead_w}号 {pattern}（{p_strength}）\n"
                    f"主軸: {mc_desc}\n"
                    f"副軸: {sc_desc}\n"
                    f"漁夫候補: {cb_str}"
                )
                fc = FILL_KS_BODY

            elif sec_label == "⑤注意事項":
                reasons = rj.get("reason", []) or []
                warn_keys = ["FLY","出遅れ","データ不足","データ信頼","ST不安定","実績なし","暫定"]
                warns = [r for r in reasons if any(k in r for k in warn_keys)]

                # 展示確認トリガー（ヒモ荒れ判定から）
                ha = rj.get("himo_are", {}) or {}
                tenji_trigger = ha.get("tenji_trigger", "")

                # 注意事項がなくても展示トリガーは必ず表示
                lines = []
                if warns:
                    lines += [f"⚠ {w}" for w in warns]
                else:
                    lines.append("特記なし")

                if tenji_trigger:
                    lines.append("─ 展示確認ポイント ─")
                    lines += [f"📋 {t}" for t in tenji_trigger.split("\n") if t.strip()]

                txt = "\n".join(lines)
                fc = FILL_KS_WARN if warns else FILL_KS_BODY

            elif sec_label == "⑥ヒモ荒れ":
                ha = rj.get("himo_are", {}) or {}
                ha_verdict = ha.get("verdict", "対象外")
                tenji_t = ha.get("tenji_trigger", "")
                tenji_line = f"\n📋 {tenji_t}" if tenji_t else ""
                if ha_verdict == "対象外":
                    txt = "対象外（rel_win1 < 45%）\n通常判定に委ねる"
                    fc  = FILL_KS_BODY
                elif ha_verdict == "不参加推奨":
                    mcp = ha.get("max_combo_prob", 0.0) or 0.0
                    eto = ha.get("est_top_odds",   0.0) or 0.0
                    cc  = ha.get("circle_concentration", 0.0) or 0.0
                    txt = (
                        f"🚫 見送り推奨\n"
                        f"最有力確率{mcp:.3f}（推定{eto:.0f}倍台）\n"
                        f"2着集中度{cc:.0f}%"
                        f"{tenji_line}"
                    )
                    fc  = FILL_KS_BAD
                elif ha_verdict == "点数絞り":
                    mcp = ha.get("max_combo_prob", 0.0) or 0.0
                    eto = ha.get("est_top_odds",   0.0) or 0.0
                    cc  = ha.get("circle_concentration", 0.0) or 0.0
                    txt = (
                        f"⚠ 点数絞り\n"
                        f"最有力確率{mcp:.3f}（推定{eto:.0f}倍台）\n"
                        f"2着集中度{cc:.0f}%"
                        f"{tenji_line}"
                    )
                    fc  = FILL_KS_WARN
                else:  # 参加推奨
                    mcp = ha.get("max_combo_prob", 0.0) or 0.0
                    eto = ha.get("est_top_odds",   0.0) or 0.0
                    cc  = ha.get("circle_concentration", 0.0) or 0.0
                    txt = (
                        f"✅ 参加推奨（ヒモ分散）\n"
                        f"最有力確率{mcp:.3f}（推定{eto:.0f}倍台）\n"
                        f"2着集中度{cc:.0f}%"
                        f"{tenji_line}"
                    )
                    fc  = FILL_KS_GOOD

            # ⑦展開quality
            elif sec_label == "⑦展開quality":
                sq = rj.get("scenario_quality", {}) or {}
                q_score   = sq.get("quality_score", "-")
                q_rank    = sq.get("quality_rank", "-")
                q_verdict = sq.get("quality_verdict", "-")
                bet_guide = sq.get("bet_size_guide", "-")
                comps     = sq.get("components", {})
                comp_str  = "  ".join([f"{k}:{v:.0f}" for k, v in comps.items()])

                # シナリオタイプ別ROI目安（バックテスト実績ベース）
                stype_for_roi = bet.get("scenario_type", "-")
                ROI_GUIDE = {
                    "逃げ軸流し": "ROI目安: 逃げ軸=約27%(中央値¥1,080)  ※ヒモ参加推奨時のみ",
                    "飛び軸":     "ROI目安: 飛び軸=約38%(中央値¥2,200)  ※quality A以上推奨",
                    "両建て":     "ROI目安: 両建て=約32%(中央値¥1,400)  ※展示後に軸絞ること",
                }
                roi_line = ROI_GUIDE.get(stype_for_roi, "ROI目安: シナリオ未確定")

                txt = (
                    f"quality: {q_rank}（{q_score}点）\n"
                    f"→ {q_verdict}\n"
                    f"→ {bet_guide}\n"
                    f"構成: {comp_str}\n"
                    f"─\n"
                    f"{roi_line}"
                )
                if q_rank in ("S", "A"):
                    fc = FILL_KS_GOOD
                elif q_rank == "B":
                    fc = FILL_KS_BODY
                elif q_rank == "C":
                    fc = FILL_KS_WARN
                else:
                    fc = FILL_KS_BAD

            # ⑧考察の結論: 考察→買い目の因果をまとめて Step2 への橋渡し
            if sec_label == "⑧考察の結論":
                ft   = rj.get("first_turn", {}) or {}
                cm   = rj.get("conflict_map", {}) or {}
                sq   = rj.get("scenario_quality", {}) or {}
                stype = bet.get("scenario_type", "-")
                pt    = bet.get("point_count", 0)
                tso   = bet.get("theory_syn_odds")
                mr    = bet.get("margin_ratio")
                mv    = bet.get("margin_verdict", "-")
                q_rank = sq.get("quality_rank", "-")
                q_guide = sq.get("bet_size_guide", "-")
                lead_w  = ft.get("lead_waku", "-")
                p_str   = ft.get("pattern_strength", "-")
                mc      = cm.get("main_conflict") or {}
                mc_desc = mc.get("desc", "-") if mc else "-"
                cb      = cm.get("collapse_beneficiary", [])
                cb_str  = "・".join([f"{w}号" for w, _ in cb[:2]]) if cb else "-"
                # ── honmei_scenario v2 の判定根拠 ────────────────────
                hs           = bet.get("honmei_scenario") or {}
                hs_pats      = hs.get("honmei_patterns", {}) or {}
                hs_hp        = hs_pats.get("honmei") or {}
                hs_sj        = hs.get("scenario_judgment", {}) or {}
                hs_himo      = hs.get("himo_ranking", []) or []
                hs_conf      = hs.get("confidence")
                narrative    = hs_hp.get("win_narrative", "")
                primary_kata = hs_hp.get("primary_kata", "")
                primary_pct  = hs_hp.get("primary_pct", 0)
                conf_str     = f"{hs_conf*100:.0f}%" if hs_conf is not None else "-"
                sj_reasons   = hs_sj.get("reasons", []) or []
                sj_top2      = "\n".join(f"  {r}" for r in sj_reasons[1:3])
                himo_lines   = []
                for h in hs_himo[:3]:
                    mark = "★" if h.get("is_marked") else "  "
                    himo_lines.append(
                        f"{mark}{h['waku']}号{h.get('honmei_mark','')} "
                        f"残{h['residual_score']:.3f}"
                        f"(位{h['pos_score']:.2f}/能{h['personal_score']:.2f})"
                    )
                himo_str = "\n".join(himo_lines) if himo_lines else "-"

                # 結論文の組み立て
                txt = (
                    f"▼ 考察の結論\n"
                    f"1M先行: {lead_w}号/{p_str} → {stype}\n"
                    f"対立: {mc_desc}\n"
                    f"quality: {q_rank} → {q_guide}\n"
                    f"漁夫候補: {cb_str}\n"
                    f"─\n"
                    f"◎勝ちパターン: {narrative}\n"
                    f"  主決まり手: {primary_kata}({primary_pct*100:.0f}%) 信頼度{conf_str}\n"
                    f"判定根拠:\n{sj_top2}\n"
                    f"ヒモ候補(★=印付き):\n{himo_str}\n"
                    f"─\n"
                    f"買い目: {pt}点 / 理論合成{tso}倍\n"
                    f"余裕度{mr}倍 → {mv}\n"
                    f"▼ Step2で理想合成オッズを確認"
                )
                # quality連動カラー
                if q_rank in ("S", "A") and mr and mr >= 1.2:
                    fc = "FF375623"   # 濃緑（絞れる・参加根拠あり）
                elif q_rank in ("C", "D") or (mr and mr < 1.2):
                    fc = "FF843C0C"   # 濃オレンジ（混戦・余裕なし警告）
                else:
                    fc = "FF1F4E79"   # 濃紺（標準）
                wc(row, 4+i, txt,
                   fill=sf(fc), font=fn(bold=True, size=8, color="FFFFFFFF"),
                   align=al(h="left", wrap=True), border=bdr)
                max_lines = max(max_lines, txt.count("\n") + 1)
                continue

            wc(row, 4+i, txt,
               fill=sf(fc), font=fn(bold=False, size=8),
               align=al(h="left", wrap=True), border=bdr)
            max_lines = max(max_lines, txt.count("\n") + 1)

        ws.row_dimensions[row].height = max(18.0 * max_lines, 36.0)
        row += 1

    # ════════════════════════════════════════════════════════════════════
    # 📋 Step2: 展開シナリオ＋買い目候補
    # ════════════════════════════════════════════════════════════════════
    # 【設計方針】
    #   シナリオ自動切替（逃げ軸/両建て/飛び軸）の判定結果と
    #   各シナリオの買い目候補を確率順に提示する。
    #   実オッズ・参加可否は人間が展示後に最終判断する。
    # ════════════════════════════════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "━━━  📋 Step2: 展開シナリオ＋買い目候補  ━━━  "
       "確率から展開を自動判定。買い目候補と理想合成オッズを確認して展示後の最終判断へ",
       fill=sf(FILL_JF_HDR), font=fn(bold=True, size=9, color="FFFFFFFF"), align=al("left"))
    ws.row_dimensions[row].height = 16.0
    row += 1

    # ─ 説明行 ─
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "【逃げ軸流し】印◎が1号艇かつ逃げ確率高め　"
       "【両建て】印◎が1号艇で確率低め、または印◎外枠で印○が1号艇　"
       "【飛び軸】印◎◎が2〜6号艇　"
       "2〜3着ヒモは展開位置補正×個人能力×jizen評価で自動選択。折り返しは条件付き自動追加",
       fill=sf(FILL_JF_NOTE), font=fn(bold=False, size=8, color="FF7F3F00"), align=al("left"))
    ws.row_dimensions[row].height = 14.0
    row += 1

    # ─ シナリオ判定行 ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "展開シナリオ判定",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        bet  = rd.get("bet_suggestions", {}) or {}
        stype = bet.get("scenario_type", bet.get("scenario_verdict", "-"))
        s1p  = bet.get("s1_prob", 0) or 0
        flyp = bet.get("fly_prob", 0) or 0
        fly_axes = bet.get("fly_axes", [])
        fly_str  = "・".join(fly_axes[:2]) + "号" if fly_axes else "-"
        rj   = rd.get("race_judgment", {}) or {}
        sc_fly_type  = rj.get("sc_fly_type", "-")
        sc_fly_waku  = rj.get("sc_fly_waku", "-")
        gyofu_top3   = rj.get("sc_gyofu_top3", [])
        gyofu_str    = "・".join(gyofu_top3) + "号" if gyofu_top3 else "-"
        hs2      = bet.get("honmei_scenario") or {}
        hs2_pats = hs2.get("honmei_patterns", {}) or {}
        hs2_hp   = hs2_pats.get("honmei") or {}
        hs2_sj   = hs2.get("scenario_judgment", {}) or {}
        hs2_narr = hs2_hp.get("win_narrative", "")
        hs2_conf = hs2.get("confidence")
        hs2_esc  = hs2_sj.get("escape_strength", "-")
        hs2_fly  = hs2_sj.get("fly_strength", "-")
        hs2_reas = (hs2_sj.get("reasons") or [""])[1:3]
        hs2_conf_str = f"{hs2_conf*100:.0f}%" if hs2_conf is not None else "-"
        hs2_reas_str = "\n".join(f"  {r}" for r in hs2_reas)
        txt = (
            f"【{stype}】\n"
            f"逃げ確率: {s1p*100:.0f}%  飛び確率: {flyp*100:.0f}%\n"
            f"飛び主役: {fly_str}\n"
            f"─ v2判定根拠 ─\n"
            f"◎勝ちパターン: {hs2_narr}\n"
            f"逃げ強度{hs2_esc} / 飛び強度{hs2_fly} 信頼度{hs2_conf_str}\n"
            f"{hs2_reas_str}\n"
            f"─ 潰れ展開(SC) ─\n"
            f"飛び役タイプ: {sc_fly_type}\n"
            f"漁夫候補: {gyofu_str}"
        )
        if stype == "逃げ軸流し":
            fc = "FFE2EFDA"
        elif stype == "飛び軸":
            fc = "FFFCE4D6"
        else:
            fc = "FFFFF2CC"
        wc(row, 4+i, txt,
           fill=sf(fc), font=fn(bold=True, size=9),
           align=al(h="left", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 56.0
    row += 1

    # ─ 買い目候補行（確率順・シナリオ別） ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "買い目候補（確率順）\nシナリオ種別 / 確率%",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al(wrap=True))
    max_lines_cand = 3
    for i, rd in enumerate(all_race_data):
        bet   = rd.get("bet_suggestions", {}) or {}
        cands = bet.get("candidates", [])
        if not cands:
            wc(row, 4+i, "候補なし",
               fill=sf("FFDDDDDD"), font=fn(bold=False, size=8, color="FF808080"),
               align=al(h="left", wrap=True), border=bdr)
            continue

        lines_nige  = []
        lines_tobi  = []
        lines_sc    = []
        lines_other = []
        for c in cands:
            sc   = c.get("scenario", "")
            orr  = ("⇄" if c.get("is_orkaeshi_23")
                    else "↩" if c.get("is_orkaeshi")
                    else "🎣" if c.get("is_sc_bet")
                    else "  ")
            reason = c.get("reason", "")
            hs_score = c.get("himo_score")
            hs_score_str = f" hs:{hs_score:.3f}" if hs_score is not None else ""
            reason_short = reason[:40] + "…" if len(reason) > 40 else reason
            line = f"{orr}{c['combo']}  {c['prob_pct']:.1f}%{hs_score_str}  [{reason_short}]"
            if "逃げ" in sc:
                lines_nige.append(line)
            elif "飛び" in sc:
                lines_tobi.append(line)
            elif "潰れ" in sc or c.get("is_sc_bet"):
                lines_sc.append(line)
            else:
                lines_other.append(line)

        sections = []
        if lines_nige:
            sections.append(f"── 逃げ軸 {len(lines_nige)}点 ──")
            sections.extend(lines_nige)
        if lines_tobi:
            sections.append(f"── 飛び軸 {len(lines_tobi)}点 ──")
            sections.extend(lines_tobi)
        if lines_sc:
            sections.append(f"── 潰れ受益 {len(lines_sc)}点 ──")
            sections.extend(lines_sc)
        if lines_other:
            sections.append(f"── 折返/その他 {len(lines_other)}点 ──")
            sections.extend(lines_other)

        total = len(cands)
        header = f"【計{total}点】↩=1着折返 ⇄=2着3着折返 🎣=潰れ受益"
        txt = header + "\n" + "\n".join(sections)
        wc(row, 4+i, txt,
           fill=sf(FILL_JF_BET), font=fn(bold=False, size=8, color="FF1F3864"),
           align=al(h="left", wrap=True), border=bdr)
        max_lines_cand = max(max_lines_cand, txt.count("\n") + 1)
    ws.row_dimensions[row].height = max(13.0 * max_lines_cand, 80.0)
    row += 1

    # ─ 買い目シンプル表示行（combo順・確率%） ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "買い目リスト\n（combo順・確率%）",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al(wrap=True))
    max_lines_simple = 3
    for i, rd in enumerate(all_race_data):
        bet   = rd.get("bet_suggestions", {}) or {}
        cands = bet.get("candidates", [])

        skip        = bet.get("skip", False)
        skip_reason = bet.get("skip_reason", "")

        if not cands:
            wc(row, 4+i, "候補なし",
               fill=sf("FFDDDDDD"), font=fn(bold=False, size=8, color="FF808080"),
               align=al(h="left", wrap=True), border=bdr)
            continue
        def _combo_sort_key(c):
            parts = str(c.get("combo", "")).split("-")
            try:
                return [int(p) for p in parts]
            except ValueError:
                return [99, 99, 99]
        sorted_cands = sorted(cands, key=_combo_sort_key)
        lines = [f"{c['combo']}（{c['prob_pct']:.1f}%）" for c in sorted_cands]
        total = len(sorted_cands)
        if skip:
            # 見送りの場合: 理由を大きく表示し、買い目は「参考（非推奨）」として折りたたむ
            skip_label = skip_reason or "⛔見送り推奨"
            txt = (
                f"{skip_label}\n"
                f"─────────────────\n"
                f"以下は参考（買い推奨ではありません）\n"
                f"【参考】計{total}点\n"
            ) + "\n".join(lines)
            cell_fill  = "FFFCE4D6"
            cell_color = "FF7F0000"
        else:
            txt = f"【計{total}点】\n" + "\n".join(lines)
            cell_fill  = FILL_JF_BET
            cell_color = "FF1F3864"
        wc(row, 4+i, txt,
           fill=sf(cell_fill), font=fn(bold=True, size=9, color=cell_color),
           align=al(h="left", wrap=True), border=bdr)
        max_lines_simple = max(max_lines_simple, txt.count("\n") + 1)
    ws.row_dimensions[row].height = max(13.0 * max_lines_simple, 60.0)
    row += 1

    # ─ 理想合成オッズ行 ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "理想合成オッズ",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        bet  = rd.get("bet_suggestions", {}) or {}
        tso  = bet.get("theory_syn_odds")
        mr   = bet.get("margin_ratio")
        skip = bet.get("skip", False)
        skip_reason = bet.get("skip_reason", "")

        if skip:
            # 見送りレースは合成オッズ行に何も表示しない
            # （買い目リスト行に見送り理由＋参考買い目が既に表示されているため）
            txt    = "-"
            bg     = "FFDDDDDD"
            fc_col = "FF808080"
        elif tso:
            ev_warn     = bet.get("ev_warning", False)
            ev_warn_msg = bet.get("ev_warning_msg", "")
            if ev_warn:
                txt = (
                    f"理想合成オッズ\n{tso}倍\n"
                    f"⚠ 期待値基準を下回っています\n"
                    f"回収重視→見送り推奨\n"
                    f"的中重視→参考買い目を使用可"
                )
                bg, fc_col = "FFFFF2CC", "FF7F3F00"
            else:
                txt = f"理想合成オッズ\n{tso}倍"
                if mr and mr >= 2.0:
                    bg, fc_col = "FFE2EFDA", "FF1D5730"
                elif mr and mr >= 1.2:
                    bg, fc_col = "FFFFF2CC", "FF7F3F00"
                else:
                    bg, fc_col = "FFFCE4D6", "FF7F0000"
        else:
            txt    = "計算不能"
            bg     = "FFDDDDDD"
            fc_col = "FF808080"
        wc(row, 4+i, txt,
           fill=sf(bg), font=fn(bold=True, size=16 if not skip else 11, color=fc_col),
           align=al(h="center", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 42.0
    row += 1

    # ─ 展示後判断ガイド行 ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "展示後の判断フロー",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        bet   = rd.get("bet_suggestions", {}) or {}
        stype = bet.get("scenario_type", "-")
        if stype == "逃げ軸流し":
            guide = "① 展示で1号艇の伸び確認\n   良ければそのまま採用\n   悪ければ飛び組に組み替え"
        elif stype == "飛び軸":
            guide = "① 展示で1号艇の伸び確認\n   悪ければ飛び組採用\n   良ければ逃げ組に組み替え"
        else:
            guide = "① 展示で1号艇の伸び確認\n   良→逃げ組採用\n   悪→飛び組採用"
        txt = (
            f"{guide}\n"
            f"② 実オッズで合成オッズを計算\n"
            f"③ 必要合成オッズと比較 → Step3"
        )
        wc(row, 4+i, txt,
           fill=sf(FILL_JF_NOTE), font=fn(bold=False, size=9, color="FF7F3F00"),
           align=al(h="left", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 80.0
    row += 1

    # ════════════════════════════════════════════════════════════════════



    # FLY退避値の復元ログ
    if fly_input_backup:
        print(f"  ✈️  FLY手入力値を復元済み: {len(fly_input_backup)}件")

# ============================================================
# 予想ログ保存（Step0 / refine_tenji.py 連携用）
# ============================================================
def _save_prediction_log(venue, race_date, race_no, results, bet_suggestions, race_judgment=None):
    """
    予想ログを logs/YYYY-MM-DD_会場名.json に保存する。
    refine_tenji.py が candidates / buy_list を参照するために必須。
    check_ev.py が combos_full を使って当日オッズとEV計算を行う。
    レース後に result_1st / result_2nd / result_3rd / hit / dividend を手動記入すること。

    race_judgment を直接受け取ることで rank/score/strategy の二重管理（_merged_bet）を廃止。
    race_judgment=None の場合は bet_suggestions からのフォールバックで後方互換を維持。
    """
    logs_dir = pathlib.Path(__file__).parent.parent / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)

    # 日付を YYYY-MM-DD に正規化
    date_str = str(race_date).replace("/", "-")[:10]
    log_path = logs_dir / f"{date_str}_{venue}.json"

    # 既存ログを読み込み（同日複数レース対応）
    log_data = {"venue": venue, "date": date_str, "races": []}
    if log_path.exists():
        try:
            with open(log_path, encoding="utf-8") as f:
                log_data = json.load(f)
        except Exception:
            pass

    # 既存の同レースエントリがあれば上書き、なければ追加
    existing = [e for e in log_data.get("races", [])
                if str(e.get("race_no")) == str(race_no)]
    entry = existing[0] if existing else {}

    # ── combos_full 生成: 全120通りの推定確率を保存（check_ev.py 用） ──────────
    # candidates（上位20点）では当日オッズが上位圏外に来た場合にEV計算できない。
    # 確率0.05%以上の全組み合わせを保存し、check_ev.py が実オッズとマッチできるようにする。
    _combos_raw = bet_suggestions.get("combos", [])
    combos_full = [
        {
            "combo":            c["combo"],
            "prob":             round(float(c["prob"]), 6),
            "theoretical_odds": c.get("theoretical_odds"),
            "hybrid_score":     round(float(c.get("hybrid_score", 0)), 5),
            "top_scenario":     c.get("top_scenario", "-"),
        }
        for c in _combos_raw
        if float(c.get("prob", 0)) >= 0.0005   # 0.05%未満は省略（ファイルサイズ節約）
    ]

    entry.update({
        "race_no":         int(race_no) if str(race_no).isdigit() else race_no,
        "buy_list":        bet_suggestions.get("buy_list", []),
        "point_count":     bet_suggestions.get("point_count", 0),
        "candidates":      bet_suggestions.get("candidates", []),
        "axis_candidates": bet_suggestions.get("axis_candidates", []),
        "himo_candidates": bet_suggestions.get("himo_candidates", []),
        "comment":         bet_suggestions.get("comment", ""),
        # ★ランク・スコア（回収率バックテストで使用）
        # race_judgment を優先参照。ない場合は bet_suggestions にフォールバック（後方互換）
        "rank":            (race_judgment or {}).get("rank",     bet_suggestions.get("rank",     "-")),
        "score":           (race_judgment or {}).get("score",    bet_suggestions.get("score",    0)),
        "strategy":        (race_judgment or {}).get("strategy", bet_suggestions.get("strategy", "")),
        # ★両建て判定
        "ryotate_verdict": bet_suggestions.get("ryotate_verdict", "-"),
        "ryotate_reason":  bet_suggestions.get("ryotate_detail", {}).get("reason", ""),
        # ★ヒモ荒れ判定（バックテスト・ROI分析用）
        "himo_are_verdict": ((race_judgment or {}).get("himo_are") or {}).get("verdict", "対象外"),
        "himo_are_mcp":     ((race_judgment or {}).get("himo_are") or {}).get("max_combo_prob"),
        "himo_are_est_odds": ((race_judgment or {}).get("himo_are") or {}).get("est_top_odds"),
        "himo_are_cc":      ((race_judgment or {}).get("himo_are") or {}).get("circle_concentration"),
        # ★check_ev.py 用: 全組み合わせの推定確率（当日オッズEV計算の母集団）
        "combos_full":     combos_full,
        # 結果欄（レース後に手動記入）
        "result_1st":  entry.get("result_1st",  None),   # 既存値を保持
        "result_2nd":  entry.get("result_2nd",  None),
        "result_3rd":  entry.get("result_3rd",  None),
        "hit":         entry.get("hit",         None),
        "dividend":    entry.get("dividend",    None),
    })

    if not existing:
        log_data.setdefault("races", []).append(entry)

    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(log_data, f, ensure_ascii=False, indent=2)

    print(f"  📝 予想ログ保存: {log_path.name} ({race_no}R)")


# ============================================================
# 回収率バックテスト（ログファイルから集計）
# ============================================================
def calc_roi_from_logs(logs_dir=None, strategy_filter=None):
    """
    logs/ フォルダの予想ログ（JSON）から回収率バックテストを実行する。

    レース後に手動記入した result_1st/result_2nd/result_3rd/hit/dividend を読み込み、
    戦略タイプ・ランク別の回収率（ROI）を集計して返す。

    【使い方】
        from load_race import calc_roi_from_logs
        summary = calc_roi_from_logs()
        print(summary)

    【出力形式】
        {
          "total_bets":    全ベット点数合計,
          "total_hits":    3連単的中数,
          "total_cost":    総購入金額（100円/点換算）,
          "total_payout":  総払戻金額,
          "roi":           回収率 (払戻/購入, 例: 1.23 = 123%),
          "hit_rate":      3連単的中率 (0〜1),
          "by_rank":       ランク別集計 {"S": {...}, "A": {...}, ...},
          "by_venue":      会場別集計,
          "skip_races":    見送り推奨（ランクD）だったレース数,
          "missing_logs":  結果未記入のレース数（dividend=None）,
          "details":       全レース詳細リスト,
        }

    Parameters
    ----------
    logs_dir       : Path or str  ログフォルダ（省略時: ../logs/）
    strategy_filter: str or None  "全速型"等で絞り込み（Noneは全戦略）
    """
    if logs_dir is None:
        logs_dir = pathlib.Path(__file__).parent.parent / "logs"
    logs_dir = pathlib.Path(logs_dir)

    if not logs_dir.exists():
        print(f"  ⚠️  logsフォルダが存在しません: {logs_dir}")
        return None

    log_files = sorted(logs_dir.glob("*.json"))
    if not log_files:
        print(f"  ⚠️  ログファイルが見つかりません: {logs_dir}")
        return None

    # 集計用変数
    total_bets    = 0
    total_hits    = 0
    total_cost    = 0      # 100円/点換算
    total_payout  = 0
    missing_logs  = 0
    skip_races    = 0
    details       = []

    by_rank   = {r: {"bets": 0, "hits": 0, "cost": 0, "payout": 0, "races": 0}
                 for r in ("S", "A", "B", "C", "D", "-")}
    by_venue  = {}

    for log_file in log_files:
        try:
            with open(log_file, encoding="utf-8") as f:
                log_data = json.load(f)
        except Exception as e:
            print(f"  ⚠️  ログ読み込みエラー: {log_file.name} ({e})")
            continue

        venue    = log_data.get("venue", "不明")
        date_str = log_data.get("date", "")

        for entry in log_data.get("races", []):
            race_no    = entry.get("race_no", "?")
            buy_list   = entry.get("buy_list", [])
            point_count = entry.get("point_count", len(buy_list))
            hit        = entry.get("hit")         # True/False/None
            dividend   = entry.get("dividend")    # 払戻金額（100円単位）/ None
            rank       = entry.get("rank", "-")   # ランク（S/A/B/C/D）

            # 結果未記入のレースはカウントのみ
            if dividend is None:
                missing_logs += 1
                continue

            cost   = point_count * 100  # 100円/点
            payout = int(dividend) if hit else 0

            total_bets   += point_count
            total_cost   += cost
            total_payout += payout
            if hit:
                total_hits += 1

            # 見送りレース（ランクD or 買い目0点）をカウント
            if rank == "D" or point_count == 0:
                skip_races += 1

            # ランク別
            r_key = rank if rank in by_rank else "-"
            by_rank[r_key]["bets"]   += point_count
            by_rank[r_key]["hits"]   += (1 if hit else 0)
            by_rank[r_key]["cost"]   += cost
            by_rank[r_key]["payout"] += payout
            by_rank[r_key]["races"]  += 1

            # 会場別
            if venue not in by_venue:
                by_venue[venue] = {"bets": 0, "hits": 0, "cost": 0, "payout": 0, "races": 0}
            by_venue[venue]["bets"]   += point_count
            by_venue[venue]["hits"]   += (1 if hit else 0)
            by_venue[venue]["cost"]   += cost
            by_venue[venue]["payout"] += payout
            by_venue[venue]["races"]  += 1

            details.append({
                "date":       date_str,
                "venue":      venue,
                "race_no":    race_no,
                "rank":       rank,
                "buy_list":   buy_list,
                "point_count": point_count,
                "hit":        hit,
                "dividend":   dividend,
                "cost":       cost,
                "payout":     payout,
                "roi":        round(payout / cost, 3) if cost > 0 else 0,
            })

    if total_cost == 0:
        print("  ⚠️  結果が記入されたレースが見つかりません。")
        print("  　  logs/*.json の result_1st/result_2nd/result_3rd/hit/dividend を記入してください。")
        return None

    roi = round(total_payout / total_cost, 4) if total_cost > 0 else 0.0
    hit_rate = round(total_hits / max(len(details), 1), 4)

    # ランク別ROI計算
    for rk_data in by_rank.values():
        rk_data["roi"] = round(rk_data["payout"] / max(rk_data["cost"], 1), 4)
        rk_data["hit_rate"] = round(rk_data["hits"] / max(rk_data["races"], 1), 4)

    # 会場別ROI計算
    for vk_data in by_venue.values():
        vk_data["roi"] = round(vk_data["payout"] / max(vk_data["cost"], 1), 4)
        vk_data["hit_rate"] = round(vk_data["hits"] / max(vk_data["races"], 1), 4)

    summary = {
        "total_bets":   total_bets,
        "total_hits":   total_hits,
        "total_cost":   total_cost,
        "total_payout": total_payout,
        "roi":          roi,
        "roi_pct":      f"{roi*100:.1f}%",
        "hit_rate":     hit_rate,
        "hit_rate_pct": f"{hit_rate*100:.1f}%",
        "total_races":  len(details),
        "missing_logs": missing_logs,
        "skip_races":   skip_races,
        "by_rank":      by_rank,
        "by_venue":     by_venue,
        "details":      details,
    }

    # ── コンソール出力 ──
    sep()
    print("  📊 回収率バックテスト結果")
    sep()
    print(f"  対象レース数   : {len(details)}")
    print(f"  総ベット点数   : {total_bets}")
    print(f"  総購入金額     : {total_cost:,}円")
    print(f"  総払戻金額     : {total_payout:,}円")
    print(f"  3連単的中数    : {total_hits}（的中率 {hit_rate*100:.1f}%）")
    print(f"  回収率         : {roi*100:.1f}%  ({'✅プラス' if roi >= 1.0 else '❌マイナス'})")
    print()
    print("  ランク別:")
    for rk, rk_data in by_rank.items():
        if rk_data["races"] == 0:
            continue
        print(f"    [{rk}] {rk_data['races']:3}レース | "
              f"的中{rk_data['hit_rate']*100:.0f}% | "
              f"ROI {rk_data['roi']*100:.1f}% | "
              f"{rk_data['cost']:,}→{rk_data['payout']:,}円")
    print()
    print("  会場別（ROI上位）:")
    venue_sorted = sorted(by_venue.items(), key=lambda x: x[1]["roi"], reverse=True)
    for vn, vk_data in venue_sorted[:5]:
        print(f"    {vn:6} | {vk_data['races']:3}レース | "
              f"ROI {vk_data['roi']*100:.1f}% | "
              f"的中{vk_data['hit_rate']*100:.0f}%")
    sep()

    return summary



def main():
    parser = argparse.ArgumentParser(description="ボートリサーチ新聞 全レース一括書き込み")
    parser.add_argument("--venue", type=str, default=None, help="会場名 (例: 大村)")
    parser.add_argument("--race",  type=int, default=None, help="レース番号 (省略時: 全レース)")
    parser.add_argument("--date",  type=str, default=None, help="日付 (例: 2026-02-15, 省略時: 最新CSV)")
    args = parser.parse_args()

    sep()
    print("  ボートリサーチ新聞 全レース一括書き込み")
    sep()

    # ════════════════════════════════════════════════════════════════════
    # ⚠️  Excel起動チェック（処理開始前）
    #     Excelが開いたまま処理すると最後の wb.save() で PermissionError になる。
    #     処理完了後に「Excelを閉じてください」と言われる時間ロスを防ぐため、
    #     最初にチェックして即座に警告する。
    # ════════════════════════════════════════════════════════════════════
    if EXCEL_FILE.exists():
        # Windows環境では .bak.xlsx への書き込みテストで開き確認を行う
        _lock_check_path = EXCEL_FILE.with_suffix(".~lock.xlsx")
        _is_excel_open = False

        # 方法①: ロックファイルの存在チェック（LibreOffice / Excel共通）
        if _lock_check_path.exists():
            _is_excel_open = True

        # 方法②: Windowsの場合は psutil でプロセスチェック
        if not _is_excel_open:
            try:
                import psutil
                excel_procs = [p for p in psutil.process_iter(["name"])
                               if p.info["name"] and "EXCEL" in p.info["name"].upper()]
                if excel_procs:
                    # プロセスは存在するが、対象ファイルを開いているかはファイル書き込みテストで確認
                    try:
                        with open(str(EXCEL_FILE), "a+b"):
                            pass
                    except PermissionError:
                        _is_excel_open = True
            except ImportError:
                pass  # psutil なし → 方法③へ

        # 方法③: ファイルへの書き込みテスト（最も確実）
        if not _is_excel_open:
            try:
                with open(str(EXCEL_FILE), "a+b"):
                    pass
            except PermissionError:
                _is_excel_open = True

        if _is_excel_open:
            print()
            print("  ╔══════════════════════════════════════════════════════╗")
            print("  ║  ❌  Excelが開いています！                            ║")
            print("  ║                                                      ║")
            print(f"  ║  📄 {str(EXCEL_FILE.name)[:48]:<48} ║")
            print("  ║                                                      ║")
            print("  ║  Excelを閉じてから、このスクリプトを再実行してください  ║")
            print("  ╚══════════════════════════════════════════════════════╝")
            print()
            return
        else:
            print(f"  ✅ Excel起動チェック: 閉じていることを確認 ({EXCEL_FILE.name})")
    # ════════════════════════════════════════════════════════════════════

    # 会場名
    venue = args.venue
    if not venue:
        # csv_outputから自動検出
        files = sorted(glob.glob(str(CSV_DIR / "*.csv")))
        if files:
            venue = os.path.basename(files[-1]).split("_")[0]
            print(f"  🔍 会場を自動検出: {venue}")
        else:
            print("❌ --venue を指定するか、csv_output/ にCSVを置いてください")
            return

    print(f"  🏁 会場: {venue}")

    # Excelを開く
    if not EXCEL_FILE.exists():
        print(f"❌ Excelファイルが見つかりません: {EXCEL_FILE}")
        return
    print(f"  📊 マスタデータ読み込み中...")
    try:
        wb = load_workbook(str(EXCEL_FILE))
    except Exception as e:
        print(f"❌ Excelを開けませんでした: {e}")
        print("   Excelが開いている場合は閉じてから再実行してください")
        return

    # マスタ読み込み
    course_master, player_master, ininage_master, venue_stats_master, venue_course_master = load_masters(wb)

    # CSV読み込み
    print(f"  📂 CSVを読み込み中...")
    df, race_date = load_csv(venue, args.race, args.date)
    if df is None or len(df) == 0:
        print(f"❌ {venue}のCSVが見つかりません: {CSV_DIR / venue}*.csv")
        return

    # モーターデータを出走表CSVから直接取得（scrape_motor.py 不要）
    motor_df = load_motor_csv(venue, None, race_df=df) if JIZEN_AVAILABLE else None
    if motor_df is None and JIZEN_AVAILABLE:
        print("  ⚠️  モーターデータなし。機力評価は '-' で出力します。")

    # レース番号一覧
    race_col = next((c for c in df.columns if "レース" in c or c == "R"), None)
    if race_col:
        race_nos = sorted(df[race_col].unique(), key=lambda x: int(x) if str(x).isdigit() else 99)
    else:
        race_nos = [str(args.race)] if args.race else ["1"]
    
    print(f"  📋 対象レース: {list(race_nos)}")
    print(f"  ✏️  Excel書き込み中: {EXCEL_FILE.name}")
    print()

    # 各レース書き込み
    tmp_image_paths = []
    all_race_data   = []  # 数値シート用にレースデータを蓄積

    for race_no in race_nos:
        rno_int = int(race_no) if str(race_no).isdigit() else 0
        if args.race and rno_int != args.race:
            continue
        
        # このレースの選手データ取得
        if race_col:
            race_df = df[df[race_col].astype(str) == str(race_no)]
        else:
            race_df = df
        
        players = []
        for _, row in race_df.iterrows():
            players.append(dict(row))
        
        if not players:
            continue
        
        # 締め切り時刻（同レース全選手共通なので先頭行から取得）
        # 列名ゆれ（締切時刻 / 締切 / 締め切り時刻）に対応
        _deadline_raw = (
            players[0].get("締切時刻") or
            players[0].get("締切") or
            players[0].get("締め切り時刻") or ""
        )
        deadline = str(_deadline_raw).strip()
        deadline = None if deadline in ("", "None", "nan") else deadline
        
        # 指数計算（改善①②③: 会場別コースマスタ・シナリオ確率・動的ハイブリッド係数）
        results, slit, venue_stats, frame_2nd, race_judgment, bet_suggestions = calc_race_indices(
            venue, race_no, players, course_master, player_master, ininage_master,
            venue_stats_master, venue_course_master
        )
        
        # 倶楽部流 事前評価用メンバーデータ組み立て
        jizen_members = None
        jizen_eval_result = None
        if JIZEN_AVAILABLE:
            try:
                jizen_members = build_jizen_members(
                    results, course_master, player_master, motor_df, race_no
                )
                if jizen_members:
                    jizen_eval_result = evaluate_all(jizen_members)
            except Exception as e:
                print(f"  ⚠️  事前評価計算エラー ({race_no}R): {e}")

        # ── 印・買い目の確定（jizen有無に関わらず必ず実行）────────────────
        # ① first_prob_map確定のため bet_suggestions を一度計算
        # ② first_prob_map を使って6視点印スコアを確定（_apply_jizen_honmei）
        # ③ 確定した印を honmei_map に変換して買い目を再生成（印と完全連動）
        # ④ 最終 s1_prob で ryotate を再計算し、3択判定と確率の一貫性を保証
        try:
            bet_suggestions = _suggest_3rentan(
                results, race_judgment, jizen_eval=jizen_eval_result
            )
            _first_prob_map  = bet_suggestions.get("first_prob_map", {})
            _tobi_prob_final = (race_judgment.get("ryotate", {})
                                .get("tobi_score", 0) or 0)
            # venue_stats を渡して6人相互作用モデルが会場特性（決まり手比率等）を参照できるようにする
            _venue_stats_for_honmei = _calc_venue_stats(venue_stats_master, venue)
            _apply_jizen_honmei(results, _tobi_prob_final, jizen_eval_result,
                                first_prob_map=_first_prob_map,
                                venue_stats=_venue_stats_for_honmei,
                                race_judgment=race_judgment)  # トップレベル関数
            _honmei_map = {str(r["waku"]): r.get("honmei", " ") for r in results}
            bet_suggestions = _suggest_3rentan(
                results, race_judgment,
                jizen_eval=jizen_eval_result,
                honmei_map=_honmei_map,
            )

            # ④ 印確定・買い目確定後の最終 s1_prob で ryotate を再計算
            #    これにより「3択判定の%」と「展開シナリオ確率」が同一モデルに統一される
            _s1_prob_final = bet_suggestions.get("s1_prob")
            if _s1_prob_final is not None:
                _tobi_scenario = race_judgment.get("ryotate", {})
                # tobi_scenario は calc_race_indices 内で生成した tobi オブジェクトが
                # race_judgment["ryotate"] に格納されているが、tobi_prob は ryotate から取れる
                _tobi_for_ryotate = {
                    "tobi_prob":   race_judgment["ryotate"].get("tobi_score", 30),
                    "main_threat": race_judgment["ryotate"].get("main_threat",
                                   bet_suggestions.get("fly_axes", [None])[0] or "-"),
                    "tobi_type":   race_judgment["ryotate"].get("tobi_type", "不明"),
                }
                _ryotate_final = _judge_ryotate(
                    race_judgment, _tobi_for_ryotate, {},
                    s1_prob=_s1_prob_final
                )
                race_judgment["ryotate"] = _ryotate_final

                # ◎艇番とs1_prob最大艇の乖離チェックを印確定後に最終更新
                _fpm = bet_suggestions.get("first_prob_map", {})
                if _fpm:
                    _top_waku = max(_fpm, key=_fpm.get)
                    _honmei_w = next(
                        (str(r["waku"]) for r in results if r.get("honmei") == "◎"), None
                    )
                    if _honmei_w and _honmei_w != _top_waku:
                        race_judgment["honmei_prob_mismatch"] = True
                        race_judgment["honmei_prob_mismatch_detail"] = (
                            f"◎={_honmei_w}号艇 vs 確率最大={_top_waku}号艇"
                            f"({_fpm.get(_top_waku, 0)*100:.1f}%)"
                        )
                    else:
                        race_judgment["honmei_prob_mismatch"] = False
                        race_judgment["honmei_prob_mismatch_detail"] = ""

        except Exception as e:
            print(f"  ⚠️  印・買い目確定エラー ({race_no}R): {e}")

        # ── 展示前レース（1〜3R）見送り ─────────────────────────────────────
        # BT分析: 1〜3Rは展示タイム・周回展示なしのため予測精度が低く ROI 54〜62%
        # 施策②: 全シナリオで見送りフラグを立てる
        # ※ bet_suggestions が None の場合（印・買い目エラー時）も安全に処理
        if race_no in (1, 2, 3) and isinstance(bet_suggestions, dict):
            if not bet_suggestions.get("skip"):  # 既に別理由で見送り済みなら上書きしない
                bet_suggestions["skip"]        = True
                bet_suggestions["skip_reason"] = (
                    f"⛔展示前レース（{race_no}R）→ 見送り\n"
                    f"BT実績: 1〜3RのROIは54〜62%。展示タイム確認後に判断推奨。"
                )

        # 数値シート用にデータを蓄積
        all_race_data.append({
            "race_no":         race_no,
            "venue":           venue,
            "race_date":       race_date,
            "results":         results,
            "slit":            slit,
            "venue_stats":     venue_stats,
            "frame_2nd":       frame_2nd,
            "deadline":        deadline,
            "jizen_members":   jizen_members,
            "jizen_eval":      jizen_eval_result,
            "player_master":   player_master,
            "tmp_image_paths": tmp_image_paths,
            "race_judgment":   race_judgment,
            "bet_suggestions": bet_suggestions,
        })

        # ── EV計算（URL優先 → Excelフォールバック） ────────────────────────
        ev_suggestion = None
        actual_odds   = None
        try:
            # ACTUAL_ODDS_URL が明示設定されている場合のみ使用
            # （URL自動生成は無効化：オッズ取得は apply_ev.py に分離済み）
            _odds_url = ACTUAL_ODDS_URL

            if _odds_url:
                # 方法A: URLから実際のオッズをスクレイピング
                try:
                    from fetch_odds import fetch_odds_from_url_bs4, fetch_odds_from_url
                    actual_odds = (
                        fetch_odds_from_url_bs4(_odds_url) or
                        fetch_odds_from_url(_odds_url)
                    )
                    if actual_odds:
                        print(f"    🌐 実オッズ取得完了 ({len(actual_odds)}件)")
                    else:
                        print(f"    ⚠️  実オッズ取得失敗 → Excelにフォールバック")
                except ImportError:
                    print(f"    ⚠️  fetch_odds.py が見つかりません → Excelにフォールバック")

            if actual_odds is None and ODDS_FILEPATH.exists():
                # 方法B: 理論オッズExcel
                actual_odds = load_actual_odds_from_excel(str(ODDS_FILEPATH))
                print(f"    📄 理論オッズExcel使用")

            if actual_odds:
                combos        = _calc_3rentan_probs_v2(results)
                ev_list       = calc_ev_from_actual_odds(combos, actual_odds)
                ev_suggestion = suggest_by_ev(ev_list, min_ev=0.0, max_bets=8)
                skip_str = "⛔見送り推奨" if ev_suggestion["skip"] else f"✅{ev_suggestion['total_bets']}点"
                print(f"    📊 EV計算: {skip_str} {ev_suggestion['reason']}")
                bet_suggestions["ev_suggestion"] = ev_suggestion
        except Exception as e:
            print(f"  ⚠️  EV計算エラー ({race_no}R): {e}")
        # ─────────────────────────────────────────────────────────────────────

        # 予想ログ保存（refine_tenji.py 連携 ＆ 回収率バックテスト用）
        # _save_prediction_log 内部でランク・スコア・戦略を直接 race_judgment から取得するよう変更。
        # 旧実装の「_merged_bet に手動マージ」は二重管理でバグの温床だったため廃止。
        _save_prediction_log(venue, race_date, race_no, results, bet_suggestions, race_judgment)

        # ── 見送り判定ログ ──────────────────────────────────────────────────
        if bet_suggestions.get("skip"):
            print(f"  {bet_suggestions.get('skip_reason', '⛔見送り')} ({race_no}R)")

        # ── 展示前フォーメーション コンソール出力 ──────────────────────────
        try:
            jf = bet_suggestions.get("jizen_formation", {})
            if jf and jf.get("formation"):
                jf_axes  = jf.get("axes", [])
                jf_himo  = jf.get("himo", [])
                jf_pts   = jf.get("point_count", 0)
                jf_form  = jf.get("formation", [])
                jf_tenk  = jf.get("tenkai_priority", [])
                jf_aisho = jf.get("aisho_axes", [])
                after_n  = max(jf_pts // 2, 6)

                # 1着軸に事前評価記号を付けて表示
                _in_sym  = (jizen_eval_result or {}).get("in_nige", [""] * 6)
                _a_sym   = (jizen_eval_result or {}).get("aisho",   [""] * 6)
                def _axis_label(w):
                    idx   = int(w) - 1
                    marks = []
                    if idx < len(_in_sym) and _in_sym[idx]:
                        marks.append(_in_sym[idx])
                    if idx < len(_a_sym) and _a_sym[idx]:
                        marks.append(_a_sym[idx])
                    return f"{w}号[{'/'.join(marks)}]" if marks else f"{w}号"

                axes_str = "  ".join(_axis_label(w) for w in jf_axes)
                himo_str = "  ".join(
                    f"{w}号{'[展]' if w in jf_tenk else ''}" for w in jf_himo
                )
                # 組み合わせを5点ずつ区切って表示
                form_lines = []
                for ci in range(0, len(jf_form), 5):
                    form_lines.append("    " + "  ".join(jf_form[ci:ci+5]))

                print()
                print(f"  ┌─ 📋 {race_no}R 展示前フォーメーション ({jf_pts}点) ─────────────")
                print(f"  │ 【1着軸】 {axes_str}")
                print(f"  │ 【ヒ モ】 {himo_str}")
                print(f"  │ 【組合せ】")
                for fl in form_lines:
                    print(f"  │{fl}")
                if jf_aisho:
                    print(f"  │ ②相性◎○: {'/'.join(jf_aisho)}号艇 → 展示タイム良ければ軸キープ")
                if jf_tenk:
                    print(f"  │ ⑤展開◎○: {'/'.join(jf_tenk)}号艇 → ヒモ優先キープ")
                print(f"  │ 展示後: タイム上位で軸を1〜2艇に絞り → 目安{after_n}点に圧縮")
                print(f"  └────────────────────────────────────────────────────")
        except Exception:
            pass  # 表示失敗は無視して続行
        # ─────────────────────────────────────────────────────────────────────

        print(f"  {race_no}R 処理中... → 完了")

    # 数値シート書き込み
    if all_race_data:
        write_numeric_sheet(wb, all_race_data)

    # Excelを保存
    print()
    print(f"  💾 保存中...")

    # 保存前に自動バックアップ（前回の正常ファイルを .bak.xlsx として退避）
    bak_file = EXCEL_FILE.with_suffix(".bak.xlsx")
    try:
        if EXCEL_FILE.exists():
            shutil.copy2(str(EXCEL_FILE), str(bak_file))
            print(f"  🗂️  バックアップ作成: {bak_file.name}")
    except Exception as e:
        print(f"  ⚠️  バックアップ失敗（続行します）: {e}")

    try:
        wb.save(str(EXCEL_FILE))
        print(f"  ✅ 完了！「{venue}_数値」シートを確認してください")
    except PermissionError:
        print("❌ 保存エラー: Excelが開いています。閉じてから再実行してください")
        return

    # Excelを自動起動
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(str(EXCEL_FILE))
        elif system == "Darwin":  # macOS
            subprocess.Popen(["open", str(EXCEL_FILE)])
        else:  # Linux
            subprocess.Popen(["xdg-open", str(EXCEL_FILE)])
        print(f"  📂 Excelを起動しました")
    except Exception as e:
        print(f"  ⚠️  Excel自動起動に失敗しました: {e}")

    sep()

if __name__ == "__main__":
    main()