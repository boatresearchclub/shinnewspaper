"""
BoatAI ローカルサーバー
-----------------------
起動方法: python server.py
アクセス: http://localhost:8080  (スマホからは http://PCのIPアドレス:8080)

必要ライブラリ: pip install watchdog openpyxl pandas
"""

import os, sys, json, re, time, threading, glob
from pathlib import Path
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import urlparse, parse_qs

# ── 設定 ───────────────────────────────────────────────
CSV_DIR    = r"C:\Users\user\Desktop\データ収集\scripts\csv_output"
MASTER_XLSX = r"C:\Users\user\Desktop\データ収集\ボートリサーチ_マスタ.xlsx"  # ← 実際のパスに変更
PORT       = 8080
# ────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
CACHE = {}          # venue_date -> processed data
CURRENT = {}        # 最新ロード済みデータ
LOCK = threading.Lock()
WATCH_STOP = threading.Event()  # セットされたら監視スレッドが停止する

# ── マスタ読み込み ──────────────────────────────────────
def load_master():
    """会場統計・イン逃げ分析・コース別マスタを読み込む"""
    if not Path(MASTER_XLSX).exists():
        print(f"[WARN] マスタファイルが見つかりません: {MASTER_XLSX}")
        return {}, {}, {}

    from openpyxl import load_workbook
    wb = load_workbook(MASTER_XLSX, read_only=True)

    # 会場統計
    venue_stats = {}
    ws = wb['会場統計']
    h3 = None
    for i, row in enumerate(ws.iter_rows(max_row=60, values_only=True)):
        if i == 2: h3 = [str(c).replace('\n','_') if c else '' for c in row]
        if i > 2 and row[0]:
            venue_stats[row[0]] = dict(zip(h3, row))

    # イン逃げ分析
    inn_analysis = {}
    ws4 = wb['イン逃げ分析']
    h4 = None
    for i, row in enumerate(ws4.iter_rows(max_row=40, values_only=True)):
        if i == 1: h4 = [str(c).replace('\n','_') if c else '' for c in row]
        if i > 1 and row[0]:
            inn_analysis[row[0]] = dict(zip(h4, row))

    # コース別マスタ
    course_master = {}
    ws2 = wb['📊コース別マスタ']
    for i, row in enumerate(ws2.iter_rows(max_row=200000, values_only=True)):
        if i <= 1: continue
        if row[0] is None: continue
        n = str(row[0]); c = row[1]
        if n not in course_master: course_master[n] = {}
        course_master[n][c] = {'rate1': row[4], 'rate3': row[6], 'st': row[9]}

    # 会場別コースマスタ
    venue_master = {}
    ws3 = wb['会場別コースマスタ']
    for i, row in enumerate(ws3.iter_rows(max_row=100000, values_only=True)):
        if i <= 1: continue
        if not row[1]: continue
        v = row[1]; n = str(row[0]); c = row[2]
        if v not in venue_master: venue_master[v] = {}
        if n not in venue_master[v]: venue_master[v][n] = {}
        venue_master[v][n][c] = {'rate1': row[5], 'rate3': row[7], 'st': row[8], 'ts_rate1': row[16]}

    print(f"[OK] マスタ読み込み完了 — 会場統計:{len(venue_stats)}件, コース別:{len(course_master)}選手")
    return venue_stats, inn_analysis, course_master, venue_master

# マスタを起動時に一度だけロード
print("[BOOT] マスタデータ読み込み中...")
try:
    VENUE_STATS, INN_ANALYSIS, COURSE_MASTER, VENUE_MASTER = load_master()
except Exception as e:
    print(f"[WARN] マスタ読み込みエラー: {e} — フォールバックモードで動作します")
    VENUE_STATS, INN_ANALYSIS, COURSE_MASTER, VENUE_MASTER = {}, {}, {}, {}

# ── CSV処理 ────────────────────────────────────────────
GRADE_SCORE = {'A1':1.0,'A2':0.82,'B1':0.65,'B2':0.50}
COURSE_BASE = {1:0.4734, 2:0.15, 3:0.1555, 4:0.1328, 5:0.0742, 6:0.0141}

def clean_name(s):
    return re.sub(r'\d+$', '', str(s)).strip()

def parse_form(s):
    s = re.sub(r'[^1-6SF]', '', str(s))
    if not s: return 0.5
    chars = s[-4:]
    score = sum({'1':1.0,'2':0.7,'3':0.5,'F':-0.3,'S':-0.3}.get(c,0.1) for c in chars)
    return score / len(chars)

def build_inn_data(venue):
    s = VENUE_STATS.get(venue, {})
    inn = INN_ANALYSIS.get(venue, {})
    race_arek = {}
    for k, v in s.items():
        m = re.match(r'(\d+)R_荒れ', str(k))
        if m and v:
            race_arek[int(m.group(1))] = round(float(v), 1)
    return {
        'inn_rate':    round(float(s.get('イン逃げ率') or s.get('逃げ率') or 0.47), 4),
        'arek_score':  round(float(s.get('荒れ_スコア') or 54), 1),
        'course_rates':[0,
            round(float(s.get('1C_1着率') or COURSE_BASE[1]), 3),
            round(float(s.get('2C_1着率') or COURSE_BASE[2]), 3),
            round(float(s.get('3C_1着率') or COURSE_BASE[3]), 3),
            round(float(s.get('4C_1着率') or COURSE_BASE[4]), 3),
            round(float(s.get('5C_1着率') or COURSE_BASE[5]), 3),
            round(float(s.get('6C_1着率') or COURSE_BASE[6]), 3),
        ],
        'inn_2place': [0, 0,
            round(float(inn.get('2枠_2着率') or 0.32), 3),
            round(float(inn.get('3枠_2着率') or 0.27), 3),
            round(float(inn.get('4枠_2着率') or 0.21), 3),
            round(float(inn.get('5枠_2着率') or 0.13), 3),
            round(float(inn.get('6枠_2着率') or 0.07), 3),
        ],
        'race_arek': race_arek,
    }

def process_csv(filepath):
    """CSVを読み込んでスコアリングし、JSON形式で返す"""
    import pandas as pd
    fpath = Path(filepath)

    # ファイル名から会場・日付を取得 (例: 鳴門_2026-04-25.csv)
    m = re.match(r'(.+?)_(\d{4}-\d{2}-\d{2})', fpath.stem)
    venue = m.group(1) if m else '不明'
    date  = m.group(2) if m else ''

    try:
        df = pd.read_csv(filepath, encoding='utf-8')
    except Exception:
        df = pd.read_csv(filepath, encoding='shift_jis')

    df['選手名_clean'] = df['選手名'].apply(clean_name)
    inn_data = build_inn_data(venue)
    course_rates = inn_data['course_rates']
    race_arek_map = inn_data.get('race_arek', {})

    vm_venue = VENUE_MASTER.get(venue, {})

    races = {}
    for _, grp in df.groupby('レース'):
        rno = int(grp['レース'].iloc[0])
        boats = []
        for _, row in grp.iterrows():
            name  = row['選手名_clean']
            boat  = int(row['艇番'])
            grade = str(row['級別'])
            wr    = float(row['全国勝率']) if pd.notna(row.get('全国勝率')) else 0
            lr    = float(row['当地勝率']) if pd.notna(row.get('当地勝率')) else 0
            m2    = float(row['M2率'])     if pd.notna(row.get('M2率'))     else 0
            b2    = float(row['B2率'])     if pd.notna(row.get('B2率'))     else 0
            hay   = float(row['早見'])     if pd.notna(row.get('早見'))     else None
            res   = str(row.get('今節成績','')).replace(' ','')
            time_str = str(row.get('締切時刻',''))

            vm = vm_venue.get(name, {}).get(boat, {})
            cm = COURSE_MASTER.get(name, {}).get(boat, {})

            if vm and vm.get('ts_rate1') and float(vm['ts_rate1'] or 0) > 0:
                pr = float(vm['ts_rate1']); dq = 'venue'
            elif vm and vm.get('rate1') and float(vm['rate1'] or 0) > 0:
                pr = float(vm['rate1']); dq = 'venue_raw'
            elif cm and cm.get('rate1') and float(cm['rate1'] or 0) > 0:
                pr = float(cm['rate1']); dq = 'course'
            else:
                pr = wr / 6.0 if wr > 0 else course_rates[boat]; dq = 'fallback'

            sc = pr
            sc += (GRADE_SCORE.get(grade, 0.65) - 0.65) * 0.10
            sc += (m2 - 30) / 100 * 0.12
            sc += (b2 - 30) / 100 * 0.06
            sc += (parse_form(res) - 0.5) * 0.08
            sc += (lr - wr) / 10 * 0.05 if lr > 0 else 0
            if hay: sc += (hay - 6) * 0.01
            sc = max(0.01, sc)

            boats.append({
                'boat': boat, 'name': name, 'grade': grade,
                'win_rate': round(wr, 2), 'local_rate': round(lr, 2),
                'motor2': round(m2, 1), 'boat2': round(b2, 1),
                'results': res, 'hayami': round(hay,1) if hay else None,
                'score': round(sc, 4), 'dq': dq, 'time': time_str, 'prob': 0
            })

        total = sum(b['score'] for b in boats)
        for b in boats:
            b['prob'] = round(b['score'] / total, 4)
        boats.sort(key=lambda x: -x['prob'])

        races[rno] = {
            'arek': race_arek_map.get(rno, inn_data['arek_score']),
            'time': boats[0]['time'] if boats else '',
            'boats': boats
        }

    result = {
        'venue': venue, 'date': date,
        'inn_data': inn_data,
        'races': races,
        'loaded_at': time.strftime('%H:%M:%S'),
        'source_file': fpath.name,
    }
    print(f"[OK] 処理完了: {fpath.name} — {venue} {date} ({len(races)}R)")
    return result

# ── フォルダ監視 ──────────────────────────────────────
def is_today_file(filepath):
    """ファイル名またはファイルの更新日時が今日かどうか判定する"""
    today = time.strftime('%Y-%m-%d')
    stem = Path(filepath).stem
    # ファイル名に今日の日付が含まれているか (例: 鳴門_2026-04-25.csv)
    if today in stem:
        return True
    # ファイル名で判定できない場合は更新日時で確認
    mtime = time.strftime('%Y-%m-%d', time.localtime(os.path.getmtime(filepath)))
    return mtime == today

def scan_csv_dir():
    """CSVフォルダをスキャンして最新ファイルを処理"""
    pattern = os.path.join(CSV_DIR, '*.csv')
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not files:
        return None

    latest = files[0]
    key = f"{latest}_{os.path.getmtime(latest)}"

    with LOCK:
        if key not in CACHE:
            print(f"[SCAN] 新しいCSV検出: {Path(latest).name}")
            try:
                data = process_csv(latest)
                CACHE[key] = data
                CURRENT['data'] = data
                CURRENT['key'] = key
                CURRENT['all_files'] = [Path(f).name for f in files[:20]]

                # 当日データを取り込んだら監視を停止
                if is_today_file(latest):
                    WATCH_STOP.set()
                    print(f"[WATCH] 当日データ({data.get('date','')})を取り込みました — 監視を停止します")
                    print(f"[WATCH] 再開するには /api/reload を呼び出してください")

            except Exception as e:
                print(f"[ERROR] CSV処理失敗: {e}")
                return None
        elif CURRENT.get('key') != key:
            CURRENT['data'] = CACHE[key]
            CURRENT['key'] = key

    return CURRENT.get('data')

def watcher_thread():
    """バックグラウンドでフォルダを監視。当日データ取り込み後は停止する"""
    print(f"[WATCH] 監視開始: {CSV_DIR}")
    while not WATCH_STOP.is_set():
        try:
            scan_csv_dir()
        except Exception as e:
            print(f"[WATCH ERROR] {e}")
        time.sleep(5)  # 5秒ごとにチェック
    print(f"[WATCH] 監視スレッド終了")

# ── HTTPサーバー ─────────────────────────────────────
class Handler(SimpleHTTPRequestHandler):
    def log_message(self, format, *args):
        pass  # 静かにする

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        # /api/data — 最新予想データ(JSON)
        if path == '/api/data':
            self.send_json(CURRENT.get('data'))

        # /api/files — CSVファイル一覧（同会場は最新日付のみ）
        elif path == '/api/files':
            files = sorted(glob.glob(os.path.join(CSV_DIR, '*.csv')),
                           key=os.path.getmtime, reverse=True)
            # 同じ会場名で複数ファイルがある場合、最新日付のみを残す
            seen_venues = {}
            for f in files:
                stem = Path(f).stem
                m = re.match(r'(.+?)[\s_](\d{4}-\d{2}-\d{2})', stem)
                if m:
                    venue_name = m.group(1)
                    date_str   = m.group(2)
                    if venue_name not in seen_venues or date_str > seen_venues[venue_name][1]:
                        seen_venues[venue_name] = (f, date_str)
                else:
                    seen_venues[stem] = (f, '')
            # 更新日時順に並べ直して返す
            deduped = sorted(seen_venues.values(), key=lambda x: os.path.getmtime(x[0]), reverse=True)
            names = [Path(f).name for f, _ in deduped[:24]]
            self.send_json({'files': names})

        # /api/status — 監視状態を確認
        elif path == '/api/status':
            self.send_json({
                'watching': not WATCH_STOP.is_set(),
                'loaded': bool(CURRENT.get('data')),
                'venue': CURRENT['data'].get('venue','') if CURRENT.get('data') else '',
                'date':  CURRENT['data'].get('date','')  if CURRENT.get('data') else '',
            })

        # /api/reload — 監視を手動で再開する
        elif path == '/api/reload':
            if WATCH_STOP.is_set():
                WATCH_STOP.clear()
                t = threading.Thread(target=watcher_thread, daemon=True)
                t.start()
                print(f"[WATCH] 監視を再開しました")
                self.send_json({'ok': True, 'message': '監視を再開しました'})
            else:
                self.send_json({'ok': True, 'message': 'すでに監視中です'})

        # /api/load?file=xxx.csv — 指定ファイルを読み込む
        elif path == '/api/load':
            qs = parse_qs(parsed.query)
            fname = qs.get('file', [''])[0]
            fpath = os.path.join(CSV_DIR, fname)
            if fname and os.path.exists(fpath):
                try:
                    data = process_csv(fpath)
                    key = f"{fpath}_{os.path.getmtime(fpath)}"
                    with LOCK:
                        CACHE[key] = data
                        CURRENT['data'] = data
                        CURRENT['key'] = key
                    self.send_json({'ok': True, 'data': data})
                except Exception as e:
                    self.send_json({'ok': False, 'error': str(e)})
            else:
                self.send_json({'ok': False, 'error': 'file not found'})

        # / — index.html
        elif path == '/' or path == '/index.html':
            html_path = SCRIPT_DIR / 'index.html'
            if html_path.exists():
                content = html_path.read_bytes()
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.send_header('Content-Length', len(content))
                self.end_headers()
                self.wfile.write(content)
            else:
                self.send_error(404, 'index.html not found')
        else:
            self.send_error(404)

    def send_json(self, data):
        body = json.dumps(data, ensure_ascii=False, default=str).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Length', len(body))
        self.end_headers()
        self.wfile.write(body)

# ── メイン ───────────────────────────────────────────
if __name__ == '__main__':
    # 初回スキャン
    print("[BOOT] CSVフォルダ初回スキャン...")
    scan_csv_dir()
    if CURRENT.get('data'):
        d = CURRENT['data']
        print(f"[OK] 初期データ: {d['venue']} {d['date']} ({len(d['races'])}R)")
    else:
        print(f"[INFO] CSVが見つかりません。フォルダに配置されると自動読み込みされます")
        print(f"       監視フォルダ: {CSV_DIR}")

    # 監視スレッド起動（当日データ取り込み済みなら起動しない）
    if not WATCH_STOP.is_set():
        t = threading.Thread(target=watcher_thread, daemon=True)
        t.start()
    else:
        print(f"[WATCH] 当日データ取り込み済みのため監視スレッドは起動しません")

    # ローカルIPを表示
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except:
        local_ip = "127.0.0.1"

    print(f"\n{'='*50}")
    print(f"  BoatAI サーバー起動")
    print(f"  PC:     http://localhost:{PORT}")
    print(f"  スマホ: http://{local_ip}:{PORT}")
    print(f"{'='*50}\n")
    print("  Ctrl+C で停止\n")

    server = HTTPServer(('0.0.0.0', PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[STOP] サーバーを停止しました")
