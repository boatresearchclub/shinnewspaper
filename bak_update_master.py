r"""
update_master.py
================
data/raw/ の全CSVを読み込み、ボートリサーチ_マスタ.xlsx の
マスタシートを一括更新するスクリプト。

【使い方】
    cd データ収集
    python scripts\update_master.py

【対応CSVファイル名】
    日別：  2026_02_28.csv  （年_月_日）
    月別：  2026_02.csv     （年_月）
    混在OK：両方が raw フォルダにあっても自動でまとめて読み込む

【フォルダ構成】
    データ収集\
      +-- data\
      |     +-- raw\
      |         +-- 2026_02_01.csv   ← 日別でもOK
      |         +-- 2026_02_02.csv
      |         +-- （月まとめ 2026_02.csv でもOK）
      +-- scripts\
      |     +-- update_master.py
      +-- ボートリサーチ_マスタ.xlsx

【更新されるシート】
    📊コース別マスタ  選手×コース別の全指数
    選手指数マスタ    選手全体の集計
    会場統計          会場別の決まり手・イン逃げ率
    イン逃げ分析      会場別のイン逃げ時2・3着率
"""

import os, glob, re
import math
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# -- パス設定 -----------------------------------------------------------------
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR    = os.path.join(BASE_DIR, "data", "raw")
EXCEL_PATH = os.path.join(BASE_DIR, "ボートリサーチ_マスタ.xlsx")
# 会場別コースマスタのCSVキャッシュ（load_race.py が高速読み込みに使用）
VENUE_COURSE_CSV = os.path.join(BASE_DIR, "data", "venue_course_master.csv")
# 【⑤追加】会場別コース距離補正値CSVキャッシュ（load_race.py の _predict_first_turn で使用）
VENUE_COURSE_ADJ_CSV = os.path.join(BASE_DIR, "data", "venue_course_adj.csv")

# ※ EXCLUDE_RACE_TYPES は現在未使用（実際の除外は load_raw_results 内の
#    EXCLUDE_GRADES / EXCLUDE_RACE_TYPES_FINAL で実施）。
#    将来レース種別フィルタが必要になった場合の参照用として残す。
# EXCLUDE_RACE_TYPES = {
#     "グランプリ", "グランプリ優勝戦", "グランドチャンピオン", "オールスター",
#     "メモリアル", "チャレンジカップ", "MB大賞",
#     "優勝戦", "準優勝戦", "準決勝戦", "準々決勝戦", "１回戦",
#     "選抜戦", "特別選抜戦", "記者選抜戦", "ドリーム戦",
#     "特選", "特賞", "予選特賞", "予選特選", "戸田特賞",
#     "一般特選", "団体・予選",
#     "準優進出戦", "シリーズ進出戦", "順位決定戦",
#     "ヴィーナス戦", "レディース戦", "女子リーグ",
#     "飛車角戦", "金銀戦", "王将戦", "全日本王座",
#     "ランチタイム", "ウインウイン", "朝トコ小判Ｒ", "昼トコ小判Ｒ",
# }

# -- 暫定★フラグの閾値 --------------------------------------------------------
THRESH_ICHI_RATE  = 20   # 1着率 (コース別出走数)
THRESH_ST         = 10   # 平均ST (コース別出走数)
THRESH_KIMETE     = 5    # 決まり手割合 (コース別1着数)

# -----------------------------------------------------------------------------
# 1. RAW_results を全CSVから読み込む
# -----------------------------------------------------------------------------
def load_raw_results():
    # *_results.csv のみ対象（payouts.csv などを除外）
    csv_files = sorted(glob.glob(os.path.join(RAW_DIR, "*_results.csv")))
    # _results.csv が1件もなければ旧形式（日別・月別）にフォールバック
    if not csv_files:
        csv_files = sorted(glob.glob(os.path.join(RAW_DIR, "2*.csv")))
        csv_files = [f for f in csv_files if "_payouts" not in f and "grade_master" not in f]
    if not csv_files:
        print(f"\n【エラー】CSVが見つかりません。")
        print(f"  探した場所: {RAW_DIR}")
        print(f"  フォルダの中身:")
        all_files = glob.glob(os.path.join(RAW_DIR, "*"))
        if all_files:
            for f in all_files:
                print(f"    {os.path.basename(f)}")
        else:
            print(f"    （フォルダが空か、フォルダ自体が存在しません）")
        print(f"\n  ファイル名は 2026_02_28.csv（日別）か 2026_02.csv（月別）のように")
        print(f"  先頭が 2 で始まる .csv ファイルを data/raw/ フォルダに置いてください。")
        raise FileNotFoundError(f"CSVなし: {RAW_DIR}")

    daily   = [f for f in csv_files if re.search(r'2\d{3}_\d{2}_\d{2}\.csv$', os.path.basename(f))]
    monthly = [f for f in csv_files if re.search(r'2\d{3}_\d{2}\.csv$',       os.path.basename(f))]
    other   = [f for f in csv_files if f not in daily and f not in monthly]
    if daily:
        print(f"  日別CSV: {len(daily)}ファイル")
    if monthly:
        print(f"  月別CSV: {len(monthly)}ファイル")
    if other:
        print(f"  その他CSV: {len(other)}ファイル")

    dfs = []
    for f in csv_files:
        try:
            df = pd.read_csv(f, dtype=str)
            dfs.append(df)
            print(f"  読込: {os.path.basename(f)}  ({len(df):,}行)")
        except Exception as e:
            print(f"  ⚠ スキップ: {os.path.basename(f)}  ({e})")

    if not dfs:
        raise ValueError("読み込めるCSVがありませんでした。")

    raw = pd.concat(dfs, ignore_index=True)
    # 重複行を除去（同じ日のCSVを誤って二重に置いた場合の対策）
    raw = raw.drop_duplicates()
    print(f"  合計: {len(raw):,}行")

    # 型変換
    raw["日付"]           = pd.to_datetime(raw["日付"], errors="coerce")
    raw["着順"]           = pd.to_numeric(raw["着順"], errors="coerce")
    raw["進入コース"]     = pd.to_numeric(raw["進入コース"], errors="coerce")
    raw["スタートタイミング"] = pd.to_numeric(raw["スタートタイミング"], errors="coerce")
    raw["レース番号"]     = pd.to_numeric(raw["レース番号"], errors="coerce")

    # 選手名のスペース正規化（「田中 和也」→「田中和也」）
    raw["選手名"] = raw["選手名"].str.replace(r"\s+", "", regex=True)

# グレードマスタを読み込んでグレード列を付与
    grade_csv = os.path.join(RAW_DIR, "grade_master.csv")
    if os.path.exists(grade_csv):
        gdf = pd.read_csv(grade_csv, dtype=str)
        gdf["開始日"] = pd.to_datetime(gdf["開始日"], errors="coerce")
        gdf["終了日"] = pd.to_datetime(gdf["終了日"], errors="coerce")

        print("  グレードマスタ適用中...")
        # merge ベースで会場名×日付範囲を突合（row-by-row apply より大幅に高速）
        raw["_idx"] = range(len(raw))
        raw_small = raw[["_idx", "日付", "会場名"]].copy()
        merged_g = raw_small.merge(gdf, on="会場名", how="left")
        merged_g = merged_g[
            (merged_g["開始日"].isna() | (merged_g["開始日"] <= merged_g["日付"])) &
            (merged_g["終了日"].isna() | (merged_g["日付"] <= merged_g["終了日"]))
        ]
        merged_g = merged_g.drop_duplicates(subset=["_idx"], keep="first")[["_idx", "グレード"]]
        raw = raw.merge(merged_g, on="_idx", how="left")
        raw["グレード"] = raw["グレード"].fillna("一般")
        raw.drop(columns=["_idx"], inplace=True)
        print(f"  グレード分布:\n{raw['グレード'].value_counts()}")
    else:
        print("  [INFO] grade_master.csv なし → 全レースを一般戦扱い")
        print("         make_grade_master.py を実行して data/raw/ に置いてください。")
        raw["グレード"] = "一般"

    # グレードで除外（SG・G1・G2・G3・女子戦）
    EXCLUDE_GRADES = {"SG", "G1", "G2", "G3", "女子戦", "ルーキーS", "マスターズL"}
    ippan = raw[~raw["グレード"].isin(EXCLUDE_GRADES)].copy()

    # 準優勝戦・優勝戦も除外
    EXCLUDE_RACE_TYPES_FINAL = {"準優勝戦", "優勝戦", "順位決定戦", "Ｓ戦優勝戦", "賞金女王決定"}
    ippan = ippan[~ippan["レース種別"].isin(EXCLUDE_RACE_TYPES_FINAL)].copy()

    print(f"  一般戦のみ: {len(ippan):,}行")

    # ── 進入変更レースの除外 ──────────────────────────────────────────────────
    # 進入変更レースは「枠なり進入」前提のマスタ統計を歪める。
    # 「進入コース ≠ 艇番」の艇が1艇でもいるレースを丸ごと除外する。
    #
    # 除外方法: 同一レース（日付×会場名×レース番号）内に
    #   進入コース != 艇番 の行が存在するレースIDを特定 → そのレース全行を除外
    #
    # 列名: CSVの「進入コース」と「艇番」を使用。
    # どちらかの列が存在しない場合は除外処理をスキップ（後方互換）。
    _has_nyujo  = "進入コース" in ippan.columns
    _has_waku   = "艇番" in ippan.columns or "枠" in ippan.columns
    _waku_col   = "艇番" if "艇番" in ippan.columns else "枠"

    if _has_nyujo and _has_waku:
        _nyujo  = pd.to_numeric(ippan["進入コース"], errors="coerce")
        _waku   = pd.to_numeric(ippan[_waku_col],   errors="coerce")

        # 進入コース ≠ 艇番 の行が含まれるレースIDを抽出
        # レースIDを (日付, 会場名, レース番号) の文字列キーで表現してメモリ節約
        _changed_mask = _nyujo != _waku
        _changed_keys = (
            ippan.loc[_changed_mask, "日付"].astype(str) + "_" +
            ippan.loc[_changed_mask, "会場名"].astype(str) + "_" +
            ippan.loc[_changed_mask, "レース番号"].astype(str)
        ).unique()

        before = len(ippan)
        if len(_changed_keys) > 0:
            _all_keys = (
                ippan["日付"].astype(str) + "_" +
                ippan["会場名"].astype(str) + "_" +
                ippan["レース番号"].astype(str)
            )
            ippan = ippan[~_all_keys.isin(_changed_keys)].copy()
            removed_races = len(_changed_keys)
            removed_rows  = before - len(ippan)
            print(f"  進入変更レース除外: {removed_races:,}レース / {removed_rows:,}行削除")
        else:
            print(f"  進入変更レース除外: 対象なし（全レース枠なり進入）")
    else:
        print(f"  進入変更レース除外: スキップ（列なし: 進入コース={_has_nyujo}, 艇番={_has_waku}）")

    print(f"  最終: {len(ippan):,}行")
    return raw, ippan


# -----------------------------------------------------------------------------
# 2. コース別マスタ集計
# -----------------------------------------------------------------------------
def calc_course_master(ippan: pd.DataFrame) -> pd.DataFrame:
    df = ippan.copy()
    g = df.groupby(["選手名", "進入コース"])

    # ── 決まり手集計用：直近1年フィルタ ──────────────────────────────────────
    # 決まり手%・被決まり手%のみ直近365日に限定（出走数・1着率等は全期間のまま）
    cutoff_date = df["日付"].max() - pd.Timedelta(days=365)
    df_1y = df[df["日付"] >= cutoff_date].copy()
    print(f"  📅 決まり手集計期間: {cutoff_date.date()} 〜 {df['日付'].max().date()} （直近1年）")

    # 基本統計（コース別平均STを含む）
    base = g.agg(
        出走数=("着順", "count"),
        コース別平均ST=("スタートタイミング", "mean"),
        コース別ST件数=("スタートタイミング", "count"),
        コース別最速ST=("スタートタイミング", "min"),
        コース別最遅ST=("スタートタイミング", "max"),
    ).reset_index()
    base["1着数"]   = g["着順"].apply(lambda x: (x == 1).sum()).values
    base["3連対数"] = g["着順"].apply(lambda x: (x <= 3).sum()).values

    base["1着率"]   = base["1着数"]   / base["出走数"]
    base["3連対率"] = base["3連対数"] / base["出走数"]

    # ──────────────────────────────────────────────────────────────────────────
    # 【重大①改善】時系列減衰補正（指数移動平均）による成長・衰え反映
    # ──────────────────────────────────────────────────────────────────────────
    # 問題: 全期間フラットな平均では、成長中の若手や衰退中のベテランを正しく評価できない
    # 解決: 各走の日付から「最新日」との経過日数に応じて指数減衰ウェイトを付与する
    #
    # 減衰式: weight = exp(-λ × 経過日数)
    #   λ = ln(2) / HALF_LIFE_DAYS  → HALF_LIFE_DAYS 日前の実績は現在の半分の重みになる
    #   HALF_LIFE_DAYS = 90 日（約3ヶ月）がバックテスト最適値（的中率0.5352）
    #   ※ backtest_20260306_113802.xlsx のサマリーセクションで確認済み
    #     半減期30日:0.5251 / 60日:0.5323 / 90日:0.5352(◆最適) / 180日:0.5346 / 270日:0.5341
    #   270日に設定していたのは誤りで、バックテスト結果を実装に反映していなかった。
    #   90日にすることで直近3ヶ月の成績を最重視し、選手の「今の調子」を正しく反映する。
    #   ・1ヶ月前の走: weight ≈ 0.79
    #   ・3ヶ月前の走: weight = 0.50
    #   ・6ヶ月前の走: weight ≈ 0.25（古い実績は自然に薄まる）
    #
    # 出力: 「時系列補正1着率」「時系列補正3連対率」を追加カラムとして付与
    # load_race.py 側でこのカラムがあれば優先使用し、なければ通常1着率にフォールバック
    # ──────────────────────────────────────────────────────────────────────────
    HALF_LIFE_DAYS = 90  # 半減期(日)。バックテスト最適値(的中率0.5352)。
    # ※ venue_course_master(会場別コースマスタ)側も90日で統一済み。
    # 旧値730日はほぼ全期間フラット平均と等価になり「時系列補正」が機能しなかった。
    import math
    _lambda = math.log(2) / HALF_LIFE_DAYS

    latest_date = df["日付"].max()

    # 【高速化】groupby.apply廃止 → 完全ベクトル化（全行一括計算）
    df_ew = df.dropna(subset=["日付", "着順"]).copy()
    elapsed_days_cm = (latest_date - df_ew["日付"]).dt.days.clip(lower=0)
    df_ew["_w"]  = np.exp(-_lambda * elapsed_days_cm.values)
    df_ew["_w1"] = df_ew["_w"] * (df_ew["着順"] == 1).astype(float)
    df_ew["_w3"] = df_ew["_w"] * (df_ew["着順"] <= 3).astype(float)
    df_ew["_w2"] = df_ew["_w"] * df_ew["_w"]  # ウェイトの二乗（実効走数計算用）

    ewma_agg_cm = df_ew.groupby(["選手名", "進入コース"])[["_w", "_w1", "_w3", "_w2"]].sum().reset_index()
    w_safe_cm = ewma_agg_cm["_w"].replace(0, np.nan)
    ewma_agg_cm["時系列補正1着率"]  = (ewma_agg_cm["_w1"] / w_safe_cm).round(4)
    ewma_agg_cm["時系列補正3連対率"] = (ewma_agg_cm["_w3"] / w_safe_cm).round(4)
    # 有効走数 = Σwi^2の逆数（実効サンプル数）
    ewma_agg_cm["時系列有効走数"]   = (ewma_agg_cm["_w"] ** 2 / ewma_agg_cm["_w2"].replace(0, np.nan)).round(1)
    ewma_agg_cm = ewma_agg_cm.drop(columns=["_w", "_w1", "_w3", "_w2"])
    base = base.merge(ewma_agg_cm, on=["選手名", "進入コース"], how="left")
    print(f"  📈 時系列減衰補正（半減期{HALF_LIFE_DAYS}日）: 完了")

    # 決まり手（1着時の決まり手分布）— 直近1年
    # ・逃げ%：コース1を含む全コースで集計（コース1にとって本来の決め技）
    # ・差し/まくり/まくり差し/抜き/恵まれ：コース2〜6のみ（コース1では使わない技）
    if "決まり手" in df_1y.columns:
        kimete_cats = ["逃げ", "差し", "まくり", "まくり差し", "抜き", "恵まれ"]

        # 逃げ%：全コース対象
        ichi_1y_all = df_1y[df_1y["着順"] == 1].copy()
        km_all = ichi_1y_all.groupby(["選手名", "進入コース"])["決まり手"].value_counts().unstack(fill_value=0)
        for c in kimete_cats:
            if c not in km_all.columns:
                km_all[c] = 0
        km_all = km_all[["逃げ"]].reset_index()
        km_all.columns = ["選手名", "進入コース", "決まり手_逃げ_件数"]
        km_base_all = df_1y.groupby(["選手名", "進入コース"]).size().reset_index(name="出走数_1y")
        km_all = km_all.merge(km_base_all, on=["選手名", "進入コース"], how="left")
        km_all["決まり手_逃げ%"] = km_all["決まり手_逃げ_件数"] / km_all["出走数_1y"]
        km_all = km_all.drop(columns=["出走数_1y"])

        # 差し/まくり/まくり差し/抜き/恵まれ：コース2〜6のみ
        ichi_1y_26 = df_1y[(df_1y["着順"] == 1) & (df_1y["進入コース"] != 1)].copy()
        km_26 = ichi_1y_26.groupby(["選手名", "進入コース"])["決まり手"].value_counts().unstack(fill_value=0)
        attack_cats = ["差し", "まくり", "まくり差し", "抜き", "恵まれ"]
        for c in attack_cats:
            if c not in km_26.columns:
                km_26[c] = 0
        km_26 = km_26[attack_cats].reset_index()
        km_26.columns = ["選手名", "進入コース"] + [f"決まり手_{c}_件数" for c in attack_cats]
        km_base_26 = df_1y[df_1y["進入コース"] != 1].groupby(["選手名", "進入コース"]).size().reset_index(name="出走数_1y")
        km_26 = km_26.merge(km_base_26, on=["選手名", "進入コース"], how="left")
        for c in attack_cats:
            km_26[f"決まり手_{c}%"] = km_26[f"決まり手_{c}_件数"] / km_26["出走数_1y"]
        km_26 = km_26.drop(columns=["出走数_1y"])

        # マージ
        km_merged = km_all.merge(km_26, on=["選手名", "進入コース"], how="outer")
        # 件数列を全部揃える
        for c in kimete_cats:
            col = f"決まり手_{c}_件数"
            if col not in km_merged.columns:
                km_merged[col] = 0
        base = base.merge(km_merged, on=["選手名", "進入コース"], how="left")

    # イン逃げ時成績（1コースが逃げで1着のレース）
    in_nige = df[(df["進入コース"] == 1) & (df["着順"] == 1)]
    if "決まり手" in df.columns:
        in_nige = in_nige[in_nige["決まり手"] == "逃げ"]
    in_nige_keys = in_nige[["日付", "会場名", "レース番号"]].drop_duplicates()
    in_nige_races = df.merge(in_nige_keys, on=["日付", "会場名", "レース番号"])

    # 1コース選手を除いた2・3着集計
    in_nige_non1 = in_nige_races[in_nige_races["進入コース"] != 1]
    in_nige_agg = in_nige_non1.groupby(["選手名", "進入コース"]).apply(lambda x: pd.Series({
        "イン逃げ出走数":   len(x),
        "イン逃げ2着数":    (x["着順"] == 2).sum(),
        "イン逃げ3着以内数": (x["着順"] <= 3).sum(),
    }), include_groups=False).reset_index()
    in_nige_agg["イン逃げ2着率"]    = in_nige_agg["イン逃げ2着数"]    / in_nige_agg["イン逃げ出走数"]
    in_nige_agg["イン逃げ3着以内率"] = in_nige_agg["イン逃げ3着以内数"] / in_nige_agg["イン逃げ出走数"]

    base = base.merge(in_nige_agg, on=["選手名", "進入コース"], how="left")

    # 1コース出走数
    c1 = df[df["進入コース"] == 1].groupby("選手名").agg(
        C1出走数=("スタートタイミング", "count"),
    ).reset_index()
    base = base.merge(c1, on="選手名", how="left")

    # コース別ST順位: 同レース内で何番目に早いSTか（選手×コース別の平均）
    df_st = df[["日付", "会場名", "レース番号", "選手名", "進入コース", "スタートタイミング"]].dropna().copy()
    df_st["ST順位"] = df_st.groupby(["日付", "会場名", "レース番号"])["スタートタイミング"].rank()
    course_rank = df_st.groupby(["選手名", "進入コース"])["ST順位"].mean().reset_index()
    course_rank.columns = ["選手名", "進入コース", "コース別ST順位"]
    base = base.merge(course_rank, on=["選手名", "進入コース"], how="left")

    # ── 被決まり手集計（直近1年・コース1出走全体が分母）＋ 【修正③】直近3ヶ月重み付け ──
    # 差され% = コース1で差されて負けた件数 ÷ コース1の出走数全体
    # まくられ% = コース1でまくられて負けた件数 ÷ コース1の出走数全体
    # 捲り差され% = コース1でまくり差されて負けた件数 ÷ コース1の出走数全体
    #
    # 【修正③】直近3ヶ月（90日）の被決まり手%と直近1年の被決まり手%を加重平均。
    #   ブレンド = 直近3ヶ月 × 0.6 + 直近1年 × 0.4
    #   根拠: 弱点は直近の実走データにより強く反映されるべき。
    #   例: かつて差され率が高かった選手が最近STを改善した場合、1年平均では過大評価してしまう。
    #   直近3ヶ月のサンプルが3件未満の場合は1年平均のみを使用（ノイズ防止）。
    if "決まり手" in df_1y.columns:
        # ── 1年分の集計 ─────────────────────────────────────────────────────
        c1_all = df_1y[df_1y["進入コース"] == 1][["日付", "会場名", "レース番号", "選手名", "着順"]].rename(columns={"選手名": "C1選手名"})

        if len(c1_all) > 0:
            # C1出走数（分母）
            c1_total = c1_all.groupby("C1選手名").size().reset_index(name="C1出走数_1y")

            # コース1で負けたレースの1着選手の決まり手を取得
            c1_lose_keys = c1_all[c1_all["着順"] != 1][["日付", "会場名", "レース番号", "C1選手名"]]
            winners = df_1y[df_1y["着順"] == 1][["日付", "会場名", "レース番号", "決まり手"]].copy()
            c1_lose_merged = c1_lose_keys.merge(
                winners, on=["日付", "会場名", "レース番号"], how="left"
            )
            # 被決まり手件数を集計
            lose_agg = c1_lose_merged.groupby("C1選手名").apply(lambda x: pd.Series({
                "C1敗戦数":       len(x),
                "差され件数":     (x["決まり手"] == "差し").sum(),
                "捲られ件数":     (x["決まり手"] == "まくり").sum(),
                "捲り差され件数": (x["決まり手"] == "まくり差し").sum(),
            }), include_groups=False).reset_index()
            lose_agg.rename(columns={"C1選手名": "選手名"}, inplace=True)

            # C1出走数全体を分母にして割合を算出（1年分）
            lose_agg = lose_agg.merge(c1_total.rename(columns={"C1選手名": "選手名"}), on="選手名", how="left")
            lose_agg["差され%_1y"]     = lose_agg["差され件数"]     / lose_agg["C1出走数_1y"]
            lose_agg["捲られ%_1y"]     = lose_agg["捲られ件数"]     / lose_agg["C1出走数_1y"]
            lose_agg["捲り差され%_1y"] = lose_agg["捲り差され件数"] / lose_agg["C1出走数_1y"]

            # ── 【修正③】直近3ヶ月分の集計 ─────────────────────────────────
            cutoff_3m = df["日付"].max() - pd.Timedelta(days=90)
            df_3m = df[df["日付"] >= cutoff_3m].copy()
            c1_all_3m = df_3m[df_3m["進入コース"] == 1][
                ["日付", "会場名", "レース番号", "選手名", "着順"]
            ].rename(columns={"選手名": "C1選手名"})

            lose_agg_3m = None
            if len(c1_all_3m) > 0 and "決まり手" in df_3m.columns:
                c1_total_3m  = c1_all_3m.groupby("C1選手名").size().reset_index(name="C1出走数_3m")
                c1_lose_3m   = c1_all_3m[c1_all_3m["着順"] != 1][["日付", "会場名", "レース番号", "C1選手名"]]
                winners_3m   = df_3m[df_3m["着順"] == 1][["日付", "会場名", "レース番号", "決まり手"]].copy()
                merged_3m    = c1_lose_3m.merge(winners_3m, on=["日付", "会場名", "レース番号"], how="left")
                _agg_3m = merged_3m.groupby("C1選手名").apply(lambda x: pd.Series({
                    "差され件数_3m":     (x["決まり手"] == "差し").sum(),
                    "捲られ件数_3m":     (x["決まり手"] == "まくり").sum(),
                    "捲り差され件数_3m": (x["決まり手"] == "まくり差し").sum(),
                }), include_groups=False).reset_index()
                _agg_3m.rename(columns={"C1選手名": "選手名"}, inplace=True)
                _agg_3m = _agg_3m.merge(c1_total_3m.rename(columns={"C1選手名": "選手名"}), on="選手名", how="left")
                # 直近3ヶ月のサンプルが3件以上の選手のみ計算（ノイズ防止）
                _agg_3m = _agg_3m[_agg_3m["C1出走数_3m"] >= 3]
                if len(_agg_3m) > 0:
                    _agg_3m["差され%_3m"]     = _agg_3m["差され件数_3m"]     / _agg_3m["C1出走数_3m"]
                    _agg_3m["捲られ%_3m"]     = _agg_3m["捲られ件数_3m"]     / _agg_3m["C1出走数_3m"]
                    _agg_3m["捲り差され%_3m"] = _agg_3m["捲り差され件数_3m"] / _agg_3m["C1出走数_3m"]
                    lose_agg_3m = _agg_3m[["選手名", "差され%_3m", "捲られ%_3m", "捲り差され%_3m"]]

            # ── 【修正③】1年集計と3ヶ月集計をブレンド ─────────────────────
            # ブレンド = 直近3ヶ月 × W_3M + 直近1年 × (1 - W_3M)
            W_3M = 0.60  # 直近3ヶ月への重み
            if lose_agg_3m is not None:
                lose_merged = lose_agg.merge(lose_agg_3m, on="選手名", how="left")
                for col_base, col_3m in [
                    ("差され%_1y", "差され%_3m"),
                    ("捲られ%_1y", "捲られ%_3m"),
                    ("捲り差され%_1y", "捲り差され%_3m"),
                ]:
                    col_out = col_base.replace("_1y", "")
                    has_3m = lose_merged[col_3m].notna()
                    lose_merged[col_out] = np.where(
                        has_3m,
                        lose_merged[col_3m] * W_3M + lose_merged[col_base] * (1 - W_3M),
                        lose_merged[col_base]  # 3ヶ月データなし → 1年のみ
                    )
                lose_agg_final = lose_merged
            else:
                lose_agg_final = lose_agg.copy()
                lose_agg_final["差され%"]     = lose_agg_final["差され%_1y"]
                lose_agg_final["捲られ%"]     = lose_agg_final["捲られ%_1y"]
                lose_agg_final["捲り差され%"] = lose_agg_final["捲り差され%_1y"]

            # 選手名でマージ（コース≠1の行はNaN、問題なし）
            base = base.merge(
                lose_agg_final[["選手名", "C1敗戦数", "差され%", "捲られ%", "捲り差され%"]],
                on="選手名", how="left"
            )
        else:
            base["C1敗戦数"]    = np.nan
            base["差され%"]     = np.nan
            base["捲られ%"]     = np.nan
            base["捲り差され%"] = np.nan
    else:
        base["C1敗戦数"]    = np.nan
        base["差され%"]     = np.nan
        base["捲られ%"]     = np.nan
        base["捲り差され%"] = np.nan

    # ──────────────────────────────────────────────────────────────────────────
    # 【重大①改善】暫定★フラグ：コース別に必要サンプル数を変更
    # ──────────────────────────────────────────────────────────────────────────
    # 旧: 全コース一律20走未満で★フラグ → 外コース（5・6C）は30走あっても信頼不足
    # 新: コース別難易度（出現頻度の逆数）に応じて閾値を設定
    #   1C: 最低50走 ← 出現頻度が高く・的中に直結するため閾値を厳しく
    #   2C: 最低60走
    #   3C: 最低70走
    #   4C: 最低80走
    #   5C: 最低100走 ← 出現頻度が低く少数データで偏りやすい
    #   6C: 最低120走 ← 最外枠で最も出現頻度が低い
    THRESH_BY_COURSE = {1: 50, 2: 60, 3: 70, 4: 80, 5: 100, 6: 120}
    # apply廃止 → ベクトル化
    thresh_vec = base["進入コース"].astype(int).map(THRESH_BY_COURSE).fillna(50)
    base["★1着率"] = base["出走数"] < thresh_vec
    base["★ST"]   = base["出走数"] < THRESH_ST
    base["★決手"] = base["1着数"] < THRESH_KIMETE

    # 【重大①】時系列補正値が計算できていない行は通常1着率でフォールバック
    base["時系列補正1着率"]  = base["時系列補正1着率"].where(
        base["時系列補正1着率"].notna(), base["1着率"]
    )
    base["時系列補正3連対率"] = base["時系列補正3連対率"].where(
        base["時系列補正3連対率"].notna(), base["3連対率"]
    )

    base = base.sort_values(["選手名", "進入コース"]).reset_index(drop=True)
    return base


# -----------------------------------------------------------------------------
# 3. 選手指数マスタ集計
# -----------------------------------------------------------------------------
def calc_senshu_master(raw: pd.DataFrame, ippan: pd.DataFrame) -> pd.DataFrame:
    g_all  = raw.groupby(["登録番号", "選手名"])
    g_ipp  = ippan.groupby(["登録番号", "選手名"])

    def safe_agg(g, col, func):
        try:
            return g[col].agg(func)
        except Exception:
            return pd.Series(dtype=float)

    total = g_all.apply(lambda x: pd.Series({
        "総出走数":    len(x),
        "1着率(全体)": (x["着順"] == 1).sum() / max(len(x), 1),
        "3連対率(全体)": (x["着順"] <= 3).sum() / max(len(x), 1),
        "平均ST(全体)": x["スタートタイミング"].mean(),
    }), include_groups=False).reset_index()

    ipp = g_ipp.apply(lambda x: pd.Series({
        "一般戦出走数":    len(x),
        "1着率(一般戦)":   (x["着順"] == 1).sum() / max(len(x), 1),
        "3連対率(一般戦)": (x["着順"] <= 3).sum() / max(len(x), 1),
    }), include_groups=False).reset_index()

    # 1コース実績
    c1 = ippan[ippan["進入コース"] == 1].groupby(["登録番号", "選手名"]).apply(lambda x: pd.Series({
        "イン出走数": len(x),
        "イン1着率":  (x["着順"] == 1).sum() / max(len(x), 1),
        # 平均ST(1コース)はcourse_st_pivotで統一するためここでは除外
    }), include_groups=False).reset_index()

    # コース別平均ST順位（1〜6コース全て）
    df_st = ippan[["日付", "会場名", "レース番号", "登録番号", "選手名", "進入コース", "スタートタイミング"]].dropna().copy()
    df_st["ST順位"] = df_st.groupby(["日付", "会場名", "レース番号"])["スタートタイミング"].rank()
    st_rank_pivot = (
        df_st.groupby(["登録番号", "選手名", "進入コース"])["ST順位"]
        .mean()
        .unstack(level="進入コース")
    )
    for c in range(1, 7):
        if c not in st_rank_pivot.columns:
            st_rank_pivot[c] = np.nan
    st_rank_pivot = st_rank_pivot[[c for c in range(1, 7)]].copy()
    st_rank_pivot.columns = [f"ST順位\n({c}コース)" for c in range(1, 7)]  # Excelヘッダと統一
    st_rank_pivot = st_rank_pivot.reset_index()

    # コース別平均ST（1〜6コース全て）
    course_st = ippan[["登録番号", "選手名", "進入コース", "スタートタイミング"]].dropna()
    course_st_pivot = (
        course_st.groupby(["登録番号", "選手名", "進入コース"])["スタートタイミング"]
        .mean()
        .unstack(level="進入コース")
    )
    # 1〜6コースの列を確保（データがないコースはNaN）
    for c in range(1, 7):
        if c not in course_st_pivot.columns:
            course_st_pivot[c] = np.nan
    course_st_pivot = course_st_pivot[[c for c in range(1, 7)]].copy()
    course_st_pivot.columns = [f"平均ST({c}コース)" for c in range(1, 7)]
    course_st_pivot = course_st_pivot.reset_index()

    merged = total.merge(ipp, on=["登録番号", "選手名"], how="outer")
    merged = merged.merge(c1, on=["登録番号", "選手名"], how="left")
    merged = merged.merge(st_rank_pivot, on=["登録番号", "選手名"], how="left")
    merged = merged.merge(course_st_pivot, on=["登録番号", "選手名"], how="left")

    # [改善⑥] 自在性：4-6コースを一括ではなくコース別難易度重み付けに変更
    # 旧：4〜6コース全レースをひとまとめで割り算 → 4コースと6コースの難易度差を無視
    # 新：コース別に難易度係数を掛けた重み付き1着率を算出
    # 【高速化】自在性: groupby.apply廃止 → 完全ベクトル化
    COURSE_DIFF_WEIGHT = {4: 1.0, 5: 1.15, 6: 1.35}
    outer = ippan[ippan["進入コース"] >= 4].copy()
    outer["_is1"]      = (outer["着順"] == 1).astype(float)
    outer["_diff"]     = outer["進入コース"].map(COURSE_DIFF_WEIGHT).fillna(1.0)
    outer["_w1_diff"]  = outer["_is1"] * outer["_diff"]
    outer["_cnt_diff"] = outer["_diff"]

    zizai_agg = outer.groupby(["登録番号", "選手名"]).agg(
        _total_cnt=("着順",     "count"),
        _win_cnt  =("_is1",     "sum"),
        _w1d      =("_w1_diff", "sum"),
        _cnt_d    =("_cnt_diff","sum"),
    ).reset_index()
    zizai_agg["自在性1着率"]        = (zizai_agg["_win_cnt"] / zizai_agg["_total_cnt"].replace(0, np.nan)).round(4).fillna(0.0)
    zizai_agg["自在性加重1着率"]    = (zizai_agg["_w1d"]     / zizai_agg["_cnt_d"].replace(0, np.nan)).round(4).fillna(0.0)
    zizai_agg["自在性出走数(4-6C)"] = zizai_agg["_total_cnt"]
    zizai_agg = zizai_agg[["登録番号", "選手名", "自在性1着率", "自在性加重1着率", "自在性出走数(4-6C)"]]
    merged = merged.merge(zizai_agg, on=["登録番号", "選手名"], how="left")
    for col in ["自在性1着率", "自在性加重1着率", "自在性出走数(4-6C)"]:
        merged[col] = merged[col].fillna(0)

    # [改善④] フォーム指数：コース考慮・着順点数制に刷新（着順点数を1着重視型に変更）
    # 旧方式の問題点：
    #   1. コースが違う走をすべて同列扱い（1コース1着≠6コース1着）
    #   2. 重みが単純な1/(i+1)でコース難易度補正なし
    #   3. 2着の点数が高すぎた（1着=10, 2着=7はフラットすぎる）
    # 新方式：
    #   - 着順を点数化（1着=10, 2着=2, 3着=1, 4〜6着=0）← 競艇は1着がすべて
    #   - 線形ウェイト（直近1走が最大10、10走前が1）で加重平均
    #   - コース別難易度係数でスコアを補正（6コース1着は1コース1着より高評価）
    #   - 補助指標：直近3走・5走の1着率、直近5走の3連対率も追加
    COURSE_DIFFICULTY = {1: 1.0, 2: 1.23, 3: 1.24, 4: 1.28, 5: 1.48, 6: 2.5}
    # バックテスト最適係数（的中率0.2515）。旧値: {2:1.8, 3:2.2, 4:2.4, 5:2.6, 6:2.8}
    # 【中程度②改善】1着にのみ高い点を与える。2着以下はほぼ無価値（回収率的に）
    RANK_POINTS = {1: 10, 2: 2, 3: 1, 4: 0, 5: 0, 6: 0}

    # フォーム指数用テーブルを事前にnumpy配列化（apply内のdict lookupを高速化）
    _rank_pts_arr  = np.array([0, 10, 2, 1, 0, 0, 0])  # index=着順(0-6), 0は欠損用
    _course_diff_arr = np.array([1.0, 1.0, 1.8, 2.2, 2.4, 2.6, 2.8])  # index=コース(0-6)

    def calc_form_index(grp):
        g = grp.sort_values("日付", ascending=False).head(10)
        n = len(g)
        if n == 0:
            return pd.Series({
                "フォーム指数": np.nan, "直近10走平均着順": np.nan,
                "直近3走1着率": np.nan, "直近5走1着率": np.nan, "直近5走3連対率": np.nan,
            })
        # 線形ウェイト（直近=最大、古い=最小）
        weights  = np.arange(n, 0, -1, dtype=float)  # [n, n-1, ..., 1]
        w_sum    = weights.sum()
        orders   = g["着順"].fillna(6).astype(int).clip(1, 6).values
        courses  = g["進入コース"].fillna(1).astype(int).clip(1, 6).values

        # numpy配列ルックアップ（Pythonループなし）
        pts_arr  = _rank_pts_arr[orders]   # 各走の着順点数
        diff_arr = _course_diff_arr[courses]  # 各走のコース難易度
        # 1着（pts==10）のみコース補正、それ以外は補正なし
        corrected = np.where(pts_arr == 10, pts_arr * diff_arr, pts_arr.astype(float))
        form_idx  = round(float(np.dot(corrected, weights) / w_sum), 2)

        valid_orders = orders[orders <= 6]
        simple_avg   = round(float(valid_orders.mean()), 3) if len(valid_orders) > 0 else np.nan

        r3_win = round(float((orders[:3] == 1).sum() / max(min(n, 3), 1)), 4)
        r5_win = round(float((orders[:5] == 1).sum() / max(min(n, 5), 1)), 4)
        r5_ren = round(float((orders[:5] <= 3).sum() / max(min(n, 5), 1)), 4)

        return pd.Series({
            "フォーム指数":      form_idx,
            "直近10走平均着順":  simple_avg,
            "直近3走1着率":      r3_win,
            "直近5走1着率":      r5_win,
            "直近5走3連対率":    r5_ren,
        })

    form = ippan.groupby(["登録番号", "選手名"]).apply(calc_form_index, include_groups=False).reset_index()
    merged = merged.merge(form, on=["登録番号", "選手名"], how="left")

    # [改善⑤] ST安定スコア：係数根拠を明示・FLY/出遅れカウント追加
    # 旧：全コース混合の標準偏差 × 1000（係数に統計的根拠なし）
    # 新：実測分布に基づく区間線形補間スコア（0.03→95点、0.05→80点、0.08→50点）
    def _st_score_from_std(std_val):
        """標準偏差[秒] → 安定スコア[0-100]（実測分布基準・区間線形補間）"""
        if pd.isna(std_val):
            return np.nan
        v = float(std_val)
        breakpoints = [(0.00, 100.0), (0.03, 95.0), (0.05, 80.0),
                       (0.08, 50.0),  (0.12, 10.0), (0.20, 0.0)]
        for i in range(len(breakpoints) - 1):
            x0, y0 = breakpoints[i]
            x1, y1 = breakpoints[i + 1]
            if x0 <= v <= x1:
                t = (v - x0) / (x1 - x0)
                return round(y0 + t * (y1 - y0), 1)
        return 0.0

    # 【高速化】ST安定スコア: apply廃止 → ベクトル化
    _st = ippan.copy()
    _st["_st_valid"] = _st["スタートタイミング"].notna()
    _st["_fly"]      = (_st["スタートタイミング"] < 0).astype(float)
    _st["_late"]     = (_st["スタートタイミング"] > 0.18).astype(float)
    st_agg = _st.groupby(["登録番号", "選手名"]).agg(
        ST標準偏差=("スタートタイミング", "std"),
        FLY数     =("_fly",              "sum"),
        出遅れ数  =("_late",             "sum"),
        ST計測件数=("_st_valid",         "sum"),
    ).reset_index()
    st_stability = st_agg
    st_stability["ST安定スコア"] = st_stability["ST標準偏差"].apply(_st_score_from_std)

    # ── 【修正④】FLY経過日数の計算 ─────────────────────────────────────────────
    # 「最後にFLYした日から現在(データ最終日)まで何日経過しているか」を算出する。
    # 旧: FLY数の合計のみ → FLY明け1節目なのか3節目なのか判別不能
    # 新: 最終FLY日付を追加 → load_race.py で「FLY経過日数 × FLY影響係数」を計算できる
    #   - FLY後は通常60日間の出場停止（節ごとに異なる場合あり）
    #   - 出場停止明け直後（経過日数60〜90日）は慎重ST傾向 → ST安定スコアを若干割引
    #   - 経過日数が長い（180日以上）はFLY影響ほぼ消滅
    _latest_date = ippan["日付"].max()
    _fly_df = _st[_st["_fly"] > 0.5][["登録番号", "選手名", "日付"]].copy()
    if len(_fly_df) > 0:
        _last_fly = _fly_df.groupby(["登録番号", "選手名"])["日付"].max().reset_index()
        _last_fly.columns = ["登録番号", "選手名", "最終FLY日"]
        _last_fly["FLY経過日数"] = (_latest_date - _last_fly["最終FLY日"]).dt.days
        st_stability = st_stability.merge(_last_fly[["登録番号", "選手名", "最終FLY日", "FLY経過日数"]],
                                          on=["登録番号", "選手名"], how="left")
    else:
        st_stability["最終FLY日"]  = pd.NaT
        st_stability["FLY経過日数"] = np.nan

    # ── 【新規】FLY前後ST比較集計 ────────────────────────────────────────────────
    # 目的:
    #   「いつもは平均0.15だがFLY後は0.25になっている」という
    #   FLYの心理的影響（慎重ST傾向）を数値で可視化する。
    #
    # 計算方法:
    #   FLY前ST平均 = 最終FLY日より前の直近20走の平均ST
    #                （FLYした走自体は除外。FLYなし選手はNaN）
    #   FLY後ST平均 = 最終FLY日以降の全走の平均ST
    #                （出場停止明け〜現在。FLYした走自体は除外）
    #   ST変化量    = FLY後ST平均 − FLY前ST平均
    #                （正 = 遅くなった = FLY影響あり / 負 = 変化なし or 改善）
    #   ST影響度    = ST変化量をカテゴリ化
    #                「大」: +0.03秒以上（明確に遅くなった）
    #                「中」: +0.01〜0.03秒
    #                「小」: −0.01〜+0.01秒（誤差範囲）
    #                「なし」: −0.01秒以下（改善 or 変化なし）
    #                NaN: FLYなし選手（新聞では空白表示）
    #
    # 注意:
    #   ・FLY走自体（スタートタイミング < 0）はST平均の計算から除外する。
    #     FLY走は-0.xx秒になるため含めると平均が大きく歪む。
    #   ・FLY前の直近20走に限定するのは、古すぎる実績（1年以上前）と
    #     比較しても現在の「素のST」を正確に反映しないため。
    #   ・FLY後の走数が2未満の場合はNaN（サンプル不足）。
    # ─────────────────────────────────────────────────────────────────────────────
    FLY_PRE_WINDOW = 20   # FLY前の参照走数

    # FLYした選手の最終FLY日を取得（_fly_dfは上で計算済み）
    fly_compare_rows = []
    if len(_fly_df) > 0:
        # _last_fly: 登録番号・選手名・最終FLY日
        for _, lf_row in _last_fly.iterrows():
            reg  = lf_row["登録番号"]
            name = lf_row["選手名"]
            fly_date = lf_row["最終FLY日"]

            # その選手の全走データ（FLY走=ST<0は除外）
            mask_player = (ippan["登録番号"] == reg) | (ippan["選手名"] == name)
            p_df = ippan[mask_player].copy()
            p_df = p_df[p_df["スタートタイミング"].notna()]
            p_df = p_df[p_df["スタートタイミング"] >= 0]  # FLY走除外

            # FLY前：最終FLY日より前 × 直近20走
            pre = (p_df[p_df["日付"] < fly_date]
                   .sort_values("日付", ascending=False)
                   .head(FLY_PRE_WINDOW))
            st_pre = pre["スタートタイミング"].mean() if len(pre) >= 3 else np.nan

            # FLY後：最終FLY日以降
            post = p_df[p_df["日付"] >= fly_date]
            st_post = post["スタートタイミング"].mean() if len(post) >= 2 else np.nan

            # ST変化量とカテゴリ
            if pd.notna(st_pre) and pd.notna(st_post):
                st_diff = round(st_post - st_pre, 4)
                if st_diff >= 0.03:
                    st_impact = "大"
                elif st_diff >= 0.01:
                    st_impact = "中"
                elif st_diff >= -0.01:
                    st_impact = "小"
                else:
                    st_impact = "なし"
            else:
                st_diff   = np.nan
                st_impact = np.nan

            fly_compare_rows.append({
                "登録番号":       reg,
                "選手名":         name,
                "FLY前ST平均":    round(st_pre,  4) if pd.notna(st_pre)  else np.nan,
                "FLY後ST平均":    round(st_post, 4) if pd.notna(st_post) else np.nan,
                "FLY後ST変化量":  st_diff,
                "FLY影響度":      st_impact,
            })

    if fly_compare_rows:
        df_fly_compare = pd.DataFrame(fly_compare_rows)
        merged = merged.merge(
            df_fly_compare[["登録番号", "選手名",
                            "FLY前ST平均", "FLY後ST平均", "FLY後ST変化量", "FLY影響度"]],
            on=["登録番号", "選手名"], how="left"
        )
        n_fly = df_fly_compare["FLY影響度"].notna().sum()
        print(f"  ✅ FLY前後ST比較: {n_fly}選手で算出完了")
    else:
        merged["FLY前ST平均"]   = np.nan
        merged["FLY後ST平均"]   = np.nan
        merged["FLY後ST変化量"] = np.nan
        merged["FLY影響度"]     = np.nan
        print("  [INFO] FLY選手なし → FLY前後ST比較はすべてNaN")

    merged = merged.merge(
        st_stability[["登録番号", "選手名", "ST標準偏差", "ST安定スコア", "FLY数", "出遅れ数",
                       "ST計測件数", "最終FLY日", "FLY経過日数"]],
        on=["登録番号", "選手名"], how="left"
    )

    merged = merged.sort_values(["登録番号"]).reset_index(drop=True)
    return merged


# -----------------------------------------------------------------------------
# 4. 会場統計集計
# -----------------------------------------------------------------------------
def calc_kaijo_stats(ippan: pd.DataFrame) -> pd.DataFrame:
    # ── 直近1年フィルタ ──────────────────────────────────────────────────────
    latest_date = ippan["日付"].max()
    cutoff_date = latest_date - pd.Timedelta(days=365)
    ippan_1y = ippan[ippan["日付"] >= cutoff_date].copy()
    print(f"  📅 会場統計集計期間: {cutoff_date.date()} 〜 {latest_date.date()} （直近1年）")
    print(f"     フィルタ前: {len(ippan):,}行 → フィルタ後: {len(ippan_1y):,}行")

    # 集計期間をメタ情報として保持（write_kaijo_stats で使用）
    calc_kaijo_stats._period = (cutoff_date.date(), latest_date.date())

    # 1着選手のみ
    ichi = ippan_1y[ippan_1y["着順"] == 1].copy()

    # ── 基本集計: groupby().apply() → vectorized演算に置き換え ──────────
    # 各決まり手フラグをboolカラムで作成してgroupby().sum()で一括集計
    ichi["_in_nige"] = (ichi["進入コース"] == 1) & (ichi["決まり手"] == "逃げ")
    for kimari in ["逃げ", "差し", "まくり", "まくり差し", "抜き"]:
        ichi[f"_is_{kimari}"] = (ichi["決まり手"] == kimari)

    g = ichi.groupby("会場名")
    レース数  = g.size().rename("レース数")
    in_nige   = g["_in_nige"].sum()
    flag_cols = {k: g[f"_is_{k}"].sum() for k in ["逃げ", "差し", "まくり", "まくり差し", "抜き"]}

    stats = pd.DataFrame({"レース数": レース数}).reset_index()
    # in_nige / flag_cols は会場名インデックスのまま。
    # stats は reset_index() 済みの整数インデックスなので
    # .values で numpy配列に変換してアライメントずれを防ぐ。
    stats["イン逃げ率"]   = (in_nige.values / stats["レース数"].values).round(4)
    for kimari, s in flag_cols.items():
        stats[f"{kimari}率"] = (s.values / stats["レース数"].values).round(4)

    # 荒れやすさスコア（vectorized）
    stats["荒れやすさスコア"] = ((1.0 - stats["イン逃げ率"]) * 100).round(1)

    # ── ★追加①：会場×コース別 1着率 ── ループmerge → pivot_tableで一括化 ──
    COURSE_NATIONAL_AVG = {1: 0.555, 2: 0.137, 3: 0.134, 4: 0.111, 5: 0.066, 6: 0.021}

    # コース別件数を一括pivotして取得
    course_pivot = ichi.groupby(["会場名", "進入コース"]).size().unstack(fill_value=0)
    course_pivot.columns = [int(c) for c in course_pivot.columns]
    # 欠損コースを補完
    for c in range(1, 7):
        if c not in course_pivot.columns:
            course_pivot[c] = 0
    course_pivot = course_pivot[[1,2,3,4,5,6]].reset_index()

    stats = stats.merge(course_pivot, on="会場名", how="left")
    for course in range(1, 7):
        stats[f"{course}コース1着率"] = (
            stats[course].fillna(0) / stats["レース数"]
        ).round(4)
        stats[f"{course}コース1着率_全国比"] = (
            stats[f"{course}コース1着率"] - COURSE_NATIONAL_AVG[course]
        ).round(4)
    stats = stats.drop(columns=[1,2,3,4,5,6])

    # ── ★追加②：R番号×コース別 1着率 ── 6回ループmerge → 一括pivot ──
    race_course_win = ichi.groupby(["会場名", "レース番号", "進入コース"]).size().reset_index(name="1着数")
    race_total      = ichi.groupby(["会場名", "レース番号"]).size().reset_index(name="総1着数")
    race_course_win = race_course_win.merge(race_total, on=["会場名", "レース番号"])
    race_course_win["1着率"] = (race_course_win["1着数"] / race_course_win["総1着数"]).round(4)

    # 一括pivot（コース×R番号を複合インデックスで展開）
    rc_pivot = race_course_win.pivot_table(
        index="会場名", columns=["進入コース", "レース番号"], values="1着率", aggfunc="mean"
    )
    rc_pivot.columns = [f"{int(c)}C_{int(r)}R1着率" for c, r in rc_pivot.columns]
    rc_pivot = rc_pivot.reset_index()
    stats = stats.merge(rc_pivot, on="会場名", how="left")

    # ── レース番号別 荒れやすさ ── apply → vectorized ──
    rg = ichi.groupby(["会場名", "レース番号"])
    race_cnt  = rg.size().rename("該当レース数")
    c1_cnt    = ichi[ichi["進入コース"] == 1].groupby(["会場名", "レース番号"]).size().rename("_c1cnt")
    race_vol  = pd.DataFrame({"該当レース数": race_cnt}).join(c1_cnt, how="left").fillna(0).reset_index()
    race_vol["R別荒れスコア"] = ((1 - race_vol["_c1cnt"] / race_vol["該当レース数"]) * 100).round(1)

    if not race_vol.empty:
        pivot_vol = race_vol.pivot_table(
            index="会場名", columns="レース番号", values="R別荒れスコア", aggfunc="mean"
        )
        pivot_vol.columns = [f"{int(c)}R荒れスコア" for c in pivot_vol.columns]
        stats = stats.merge(pivot_vol.reset_index(), on="会場名", how="left")

    # 作業用カラム削除
    stats = stats.drop(columns=[c for c in stats.columns if c.startswith("_")], errors="ignore")

    return stats


# -----------------------------------------------------------------------------
# 5. イン逃げ分析集計
# -----------------------------------------------------------------------------
def calc_in_nige_analysis(ippan: pd.DataFrame) -> pd.DataFrame:
    # ── 直近1年フィルタ ──────────────────────────────────────────────────────
    latest_date = ippan["日付"].max()
    cutoff_date = latest_date - pd.Timedelta(days=365)
    ippan_1y = ippan[ippan["日付"] >= cutoff_date].copy()
    print(f"  📅 イン逃げ分析集計期間: {cutoff_date.date()} 〜 {latest_date.date()} （直近1年）")
    print(f"     フィルタ前: {len(ippan):,}行 → フィルタ後: {len(ippan_1y):,}行")

    # 集計期間をメタ情報として保持（write_in_nige で使用）
    calc_in_nige_analysis._period = (cutoff_date.date(), latest_date.date())

    in_nige_keys = ippan_1y[
        (ippan_1y["進入コース"] == 1) & (ippan_1y["着順"] == 1) & (ippan_1y["決まり手"] == "逃げ")
    ][["日付", "会場名", "レース番号"]].drop_duplicates()

    in_nige_races = ippan_1y.merge(in_nige_keys, on=["日付", "会場名", "レース番号"])
    in_nige_non1  = in_nige_races[in_nige_races["進入コース"] != 1].copy()

    # イン逃げ回数（会場別）
    race_count_s = in_nige_keys.groupby("会場名").size().rename("イン逃げ回数")

    # 2着・3着以内フラグをbool化してgroupby→一括pivotで処理（Pythonループ廃止）
    in_nige_non1["_is2"] = (in_nige_non1["着順"] == 2)
    in_nige_non1["_is3"] = (in_nige_non1["着順"] <= 3)

    pivot2 = in_nige_non1.groupby(["会場名", "進入コース"])["_is2"].sum().unstack(fill_value=0)
    pivot3 = in_nige_non1.groupby(["会場名", "進入コース"])["_is3"].sum().unstack(fill_value=0)

    result = pd.DataFrame({"イン逃げ回数": race_count_s}).reset_index()
    for course in range(1, 7):
        cnt2 = pivot2[course] if course in pivot2.columns else pd.Series(0, index=pivot2.index)
        cnt3 = pivot3[course] if course in pivot3.columns else pd.Series(0, index=pivot3.index)
        result = result.merge(
            cnt2.rename(f"_2_{course}").reset_index(), on="会場名", how="left"
        ).merge(
            cnt3.rename(f"_3_{course}").reset_index(), on="会場名", how="left"
        )
        result[f"_2_{course}"] = result[f"_2_{course}"].fillna(0)
        result[f"_3_{course}"] = result[f"_3_{course}"].fillna(0)
        result[f"{course}枠2着率"]    = (result[f"_2_{course}"] / result["イン逃げ回数"].clip(lower=1)).round(4)
        result[f"{course}枠3着以内率"] = (result[f"_3_{course}"] / result["イン逃げ回数"].clip(lower=1)).round(4)
        result = result.drop(columns=[f"_2_{course}", f"_3_{course}"])

    return result.sort_values("会場名").reset_index(drop=True)


# -----------------------------------------------------------------------------
# 6-0. 展開別残存マスタ集計（決まり手×1着コース → 各コース2・3着残存率）
# -----------------------------------------------------------------------------
TENKAI_VENUE_CSV    = os.path.join(BASE_DIR, "data", "tenkai_survival_venue.csv")
TENKAI_NATIONAL_CSV = os.path.join(BASE_DIR, "data", "tenkai_survival_national.csv")

# 信頼度計算の基準レース数（これ以上で信頼度1.0）
TRUST_THRESHOLD = 150


def calc_tenkai_survival_master(ippan: pd.DataFrame):
    """
    「決まり手×1着コース」のシナリオ別に各コースの2・3着残存率を集計する。

    Returns
    -------
    df_venue    : 会場別集計 DataFrame
    df_national : 全国集計 DataFrame
    """
    latest_date = ippan["日付"].max()
    cutoff_date = latest_date - pd.Timedelta(days=365)
    df = ippan[ippan["日付"] >= cutoff_date].copy()

    print(f"  📅 展開別残存マスタ集計期間: {cutoff_date.date()} 〜 {latest_date.date()}")
    print(f"     対象行数: {len(df):,}行")

    # 1着行（決まり手・1着コースを取得）
    df1 = df[df["着順"] == 1][["日付", "会場名", "レース番号", "進入コース", "決まり手"]].copy()
    df1 = df1.rename(columns={"進入コース": "1着コース"})
    df1["1着コース"] = df1["1着コース"].astype(str)

    # 全艇行に1着情報（決まり手・1着コース）をマージ
    # ※ df側にも「決まり手」列があるため suffixes で区別し、1着側を使う
    df_all = df.merge(
        df1[["日付", "会場名", "レース番号", "決まり手", "1着コース"]],
        on=["日付", "会場名", "レース番号"],
        suffixes=("_艇", ""),   # 1着側の決まり手を無suffix、艇側を_艇suffix
    )
    df_non1 = df_all[df_all["着順"] != 1].copy()
    df_non1["進入コース"] = df_non1["進入コース"].astype(str)
    df_non1["_is2"]     = (df_non1["着順"] == 2)
    df_non1["_is3only"] = (df_non1["着順"] == 3)
    df_non1["_is3"]     = (df_non1["着順"] <= 3)

    # 複合キーでレース数を数える
    # df1は1着行のみなので 日付+会場名+レース番号 の組み合わせ = レース数
    df1["_rkey"] = df1["日付"].astype(str) + "_" + df1["会場名"] + "_" + df1["レース番号"].astype(str)
    race_count_key = (
        df1.groupby(["会場名", "決まり手", "1着コース"])["_rkey"]
        .nunique().rename("レース数").reset_index()
    )
    race_count_nat = (
        df1.groupby(["決まり手", "1着コース"])["_rkey"]
        .nunique().rename("レース数").reset_index()
    )

    def _build(group_keys, race_count_df, label):
        """group_keys: グループ列リスト（会場名を含む or 含まない）"""
        grp = df_non1.groupby(group_keys + ["進入コース"])

        # 2着・3着ちょうど・3着以内カウント
        cnt2     = grp["_is2"].sum().rename("cnt2").reset_index()
        cnt3only = grp["_is3only"].sum().rename("cnt3only").reset_index()
        cnt3     = grp["_is3"].sum().rename("cnt3").reset_index()
        merged = cnt2.merge(cnt3only, on=group_keys + ["進入コース"])
        merged = merged.merge(cnt3,   on=group_keys + ["進入コース"])

        # レース数マージ
        merged = merged.merge(race_count_df, on=group_keys, how="left")
        merged["レース数"]   = merged["レース数"].fillna(0).astype(int)
        merged["2着率"]      = (merged["cnt2"]     / merged["レース数"].clip(lower=1)).round(4)
        merged["3着率"]      = (merged["cnt3only"] / merged["レース数"].clip(lower=1)).round(4)
        merged["3着以内率"]  = (merged["cnt3"]     / merged["レース数"].clip(lower=1)).round(4)
        merged["信頼度"]     = (merged["レース数"] / TRUST_THRESHOLD).clip(upper=1.0).round(3)

        return merged.drop(columns=["cnt2", "cnt3only", "cnt3"])

    df_venue    = _build(["会場名", "決まり手", "1着コース"], race_count_key,    "会場別")
    df_national = _build(["決まり手", "1着コース"],           race_count_nat,  "全国")

    print(f"  ✅ 会場別: {len(df_venue)}行  全国集計: {len(df_national)}行")

    # CSVに保存
    os.makedirs(os.path.join(BASE_DIR, "data"), exist_ok=True)
    df_venue.to_csv(TENKAI_VENUE_CSV, index=False, encoding="utf-8-sig")
    df_national.to_csv(TENKAI_NATIONAL_CSV, index=False, encoding="utf-8-sig")
    print(f"  💾 CSV保存: {os.path.basename(TENKAI_VENUE_CSV)}, {os.path.basename(TENKAI_NATIONAL_CSV)}")

    return df_venue, df_national


def write_tenkai_survival(wb, df_venue: "pd.DataFrame", df_national: "pd.DataFrame"):
    """展開別残存マスタをExcelに書き込む（2シート）"""
    # 決まり手ごとの背景色
    KIMETE_COLORS = {
        "逃げ":     "DEEAF1",
        "差し":     "E2EFDA",
        "まくり":   "FFF2CC",
        "まくり差し": "FCE4D6",
        "抜き":     "F2F2F2",
    }
    H_FILL = PatternFill("solid", start_color="1F4E79")
    H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    D_FONT = Font(name="Arial", size=9)
    AL_C   = Alignment(horizontal="center", vertical="center")
    AL_L   = Alignment(horizontal="left",   vertical="center")

    def _thin():
        s = Side(border_style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def _write_sheet(ws, df, group_keys):
        ws.sheet_view.showGridLines = False
        headers = group_keys + ["レース数", "信頼度", "進入コース", "2着率", "3着率", "3着以内率"]
        ws.append([""] + headers)
        for ci, h in enumerate(headers, 2):
            c = ws.cell(1, ci)
            c.value = h
            c.fill  = H_FILL
            c.font  = H_FONT
            c.alignment = AL_C
            c.border = _thin()

        # データ一括append後、スタイルは決まり手ごとに1回だけ適用（高速化）
        rows_data = []
        for row in df.to_dict(orient="records"):
            kimete = str(row.get("決まり手", ""))
            vals   = [""] + [row.get(k, "") for k in headers]
            rows_data.append((kimete, vals))

        for kimete, vals in rows_data:
            ws.append(vals)

        # スタイル適用：append後に行番号を使って一括処理
        start_row = 2  # ヘッダ1行 + データ開始
        for row_offset, (kimete, vals) in enumerate(rows_data):
            r   = start_row + row_offset
            bg  = PatternFill("solid", start_color=KIMETE_COLORS.get(kimete, "F2F2F2"))
            for ci in range(2, len(headers) + 2):
                c           = ws.cell(r, ci)
                c.font      = D_FONT
                c.fill      = bg
                c.alignment = AL_C if ci > 2 else AL_L
                c.border    = _thin()

        # 列幅
        col_widths = [3] + [10] * len(group_keys) + [8, 7, 8, 7, 7, 8]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # %フォーマット適用（2着率・3着率・3着以内率）
        # 列構成: 1(空白) + len(group_keys) + レース数(1) + 信頼度(1) + 進入コース(1) + 2着率 ...
        # 2着率の列番号 = 1(空白) + len(group_keys) + 3(レース数・信頼度・進入コース) + 1
        pct_start = 1 + len(group_keys) + 3 + 1
        TENKAI_PCT_COLS = {pct_start, pct_start + 1, pct_start + 2}
        start_row_t = 2  # ヘッダ1行 + データ開始
        for row_offset, (kimete, vals) in enumerate(rows_data):
            r = start_row_t + row_offset
            for ci in TENKAI_PCT_COLS:
                v = ws.cell(r, ci).value
                if isinstance(v, float):
                    ws.cell(r, ci).number_format = "0.0%"

    # 会場別シート
    sname_v = "展開別残存_会場別"
    if sname_v in wb.sheetnames:
        del wb[sname_v]
    ws_v = wb.create_sheet(sname_v)
    _write_sheet(ws_v, df_venue, ["会場名", "決まり手", "1着コース"])

    # 全国シート
    sname_n = "展開別残存_全国"
    if sname_n in wb.sheetnames:
        del wb[sname_n]
    ws_n = wb.create_sheet(sname_n)
    _write_sheet(ws_n, df_national, ["決まり手", "1着コース"])

    print(f"  ✅ Excelシート書き込み完了: {sname_v}, {sname_n}")


# -----------------------------------------------------------------------------
# 6. 会場別コースマスタ集計（改善①：選手×会場×コース の3次元実績）
# -----------------------------------------------------------------------------
def calc_venue_course_master(ippan: pd.DataFrame) -> pd.DataFrame:
    """
    選手×会場×コース の3次元で1着率・3連対率・決まり手%・平均STを集計。
    全国混合のコース別マスタより精度が高い「会場特化型」選手評価を提供する。

    【重大①改善】信頼度スコアをコース別最低サンプル数で計算
      コース別に必要走数が異なる（外コースほど出現頻度が低く多くのサンプルが必要）
      1C: 50走で信頼度1.0  / 6C: 120走で信頼度1.0
      加えて時系列減衰補正済みの1着率・3連対率を追加

    時系列補正:
      半減期90日の指数減衰ウェイトで直近の走を重視（バックテスト最適値）
      成長中の選手→直近が優れるため時系列補正値 > 全期間平均
      衰退中の選手→直近が劣るため時系列補正値 < 全期間平均
    """
    # コース別最低信頼走数（これ以上あれば信頼度1.0）
    TRUST_RUNS_BY_COURSE = {1: 50, 2: 60, 3: 70, 4: 80, 5: 100, 6: 120}
    MIN_RUNS = 3  # これ未満は除外
    KEY = ["選手名", "会場名", "進入コース"]

    # import math はトップレベルに移動済み
    HALF_LIFE_DAYS = 90  # バックテスト最適値(的中率0.5352)。旧値:270(誤設定)
    _lambda = math.log(2) / HALF_LIFE_DAYS
    latest_date = ippan["日付"].max()

    df = ippan.copy()

    # ── 基本集計: apply(lambda) → vectorized ──────────────────────────────
    # 1着・3連対フラグをbool列で事前作成 → groupby().sum()で一括集計
    df["_is1"]  = (df["着順"] == 1)
    df["_is3"]  = (df["着順"] <= 3)

    g = df.groupby(KEY)
    base = g.agg(
        出走数=("着順",           "count"),
        平均ST=("スタートタイミング", "mean"),
        **{"1着数":  ("_is1", "sum"),
           "3連対数": ("_is3", "sum")},
    ).reset_index()

    base["1着率"]   = (base["1着数"]   / base["出走数"]).round(4)
    base["3連対率"] = (base["3連対数"] / base["出走数"]).round(4)

    # ──────────────────────────────────────────────────────────────────────────
    # 【信頼度 改善版】Wilson信頼区間ベースの信頼度スコア
    # ──────────────────────────────────────────────────────────────────────────
    # 旧方式の問題:
    #   「出走数 / 閾値」の単純割り算では、p(1-p)の分散特性を無視している。
    #   例) n=50, p=0.50 と n=50, p=0.05 を同じ信頼度1.0にしてしまう。
    #   二項比率の推定精度は分散 p*(1-p)/n に依存し、p≒0.5が最も不確かで、
    #   p≒0 or 1 に近いほど少ないサンプルで精度が出る。
    #
    # 新方式: Wilson信頼区間の半幅で信頼度を計算
    #   half_width = z * sqrt(p*(1-p)/n + z²/(4n²)) / (1 + z²/n)
    #   z = 1.96（95%信頼区間）
    #   許容誤差 = コース別全国平均1着率 × TOLERANCE_RATIO
    #     例: 1コース → 0.555 × 0.20 = ±0.111 以内なら信頼度1.0
    #         6コース → 0.021 × 0.20 = ±0.004 以内なら信頼度1.0
    #   信頼度 = clip(1.0 - half_width/許容誤差, 0, 1)
    #
    # 効果:
    #   ・p≒0.5（中間コース）は分散が大きいため多くのサンプルが必要 → 信頼度が上がりにくい
    #   ・p≒0（外コース）は分散が小さいため少ないサンプルでも信頼度が確保できる
    #   ・統計的に意味のある信頼度として load_race.py の動的ハイブリッド係数に使われる
    # ──────────────────────────────────────────────────────────────────────────
    NATIONAL_AVG_COURSE = {1: 0.555, 2: 0.137, 3: 0.134, 4: 0.111, 5: 0.066, 6: 0.021}
    WILSON_Z = 1.96  # 95%信頼区間

    # ── 【修正③】コース別許容誤差比率（外コースほど広く設定）────────────────────
    # 【旧問題】全コース一律 TOLERANCE_RATIO=0.20 では
    #   6コースの許容誤差 = 0.021 × 0.20 = ±0.004 と極端に狭く、
    #   6コースの実際の分布（2〜8%）はほぼ全員が許容誤差を超えるため
    #   vc_trust が常に低くなり個人差が会場特性に飲み込まれていた。
    #
    # 【新方式】コース別に「統計的に意味のある個人差」を許容誤差として設定する。
    #   1コース: ±15%（0.555 × 0.15 = ±0.083）← 出現頻度高く精度要求も高い
    #   2コース: ±20%（0.137 × 0.20 = ±0.027）
    #   3コース: ±25%（0.134 × 0.25 = ±0.034）
    #   4コース: ±30%（0.111 × 0.30 = ±0.033）
    #   5コース: ±40%（0.066 × 0.40 = ±0.026）
    #   6コース: ±60%（0.021 × 0.60 = ±0.013）← 外コースは個人差を広く許容
    TOLERANCE_BY_COURSE = {1: 0.15, 2: 0.20, 3: 0.25, 4: 0.30, 5: 0.40, 6: 0.60}

    n_arr  = base["出走数"].values.astype(float)
    p_arr  = base["1着率"].values.clip(1e-6, 1 - 1e-6)
    z      = WILSON_Z
    z2     = z ** 2

    # Wilson信頼区間の半幅（ベクトル化）
    hw_num   = z * np.sqrt(p_arr * (1 - p_arr) / n_arr + z2 / (4 * n_arr ** 2))
    hw_denom = 1 + z2 / n_arr
    half_width = hw_num / hw_denom

    # コース別許容誤差（コース別比率 × 全国平均）
    course_avg_arr    = base["進入コース"].astype(int).map(NATIONAL_AVG_COURSE).fillna(0.1).values
    tolerance_ratio_arr = base["進入コース"].astype(int).map(TOLERANCE_BY_COURSE).fillna(0.20).values
    tolerance_arr     = course_avg_arr * tolerance_ratio_arr

    # 信頼度: half_widthが許容誤差に収まるほど1.0に近づく
    raw_trust = 1.0 - (half_width / tolerance_arr)
    base["信頼度"] = np.clip(raw_trust, 0.0, 1.0).round(3)

    # 【重大①・高速化】時系列減衰補正 — groupby.apply廃止 → 完全ベクトル化
    # 問題: groupby(KEY).apply(func) は選手×会場×コースの数万グループ×Pythonループで極遅
    # 解決: 全行のウェイトを一括計算 → groupby.sum() のみで完結（Pythonループゼロ）
    df_ew = df.dropna(subset=["日付", "着順"]).copy()
    elapsed_days = (latest_date - df_ew["日付"]).dt.days.clip(lower=0)
    df_ew["_w"]  = np.exp(-_lambda * elapsed_days.values)
    df_ew["_w1"] = df_ew["_w"] * (df_ew["着順"] == 1).astype(float)
    df_ew["_w3"] = df_ew["_w"] * (df_ew["着順"] <= 3).astype(float)
    df_ew["_w2"] = df_ew["_w"] * df_ew["_w"]  # ウェイトの二乗（実効走数計算用）

    ewma_agg = df_ew.groupby(KEY)[["_w", "_w1", "_w3", "_w2"]].sum().reset_index()
    w_safe = ewma_agg["_w"].replace(0, np.nan)
    ewma_agg["時系列補正1着率"]  = (ewma_agg["_w1"] / w_safe).round(4)
    ewma_agg["時系列補正3連対率"] = (ewma_agg["_w3"] / w_safe).round(4)
    # 【修正④連携】有効走数を会場別コースマスタにも追加（load_race.py側の④修正で使用）
    ewma_agg["時系列有効走数"]   = (ewma_agg["_w"] ** 2 / ewma_agg["_w2"].replace(0, np.nan)).round(1)
    ewma_agg = ewma_agg.drop(columns=["_w", "_w1", "_w3", "_w2"])
    base = base.merge(ewma_agg, on=KEY, how="left")
    # フォールバック: 時系列補正が計算できない行は通常1着率を使用
    base["時系列補正1着率"]  = base["時系列補正1着率"].where(
        base["時系列補正1着率"].notna(), base["1着率"])
    base["時系列補正3連対率"] = base["時系列補正3連対率"].where(
        base["時系列補正3連対率"].notna(), base["3連対率"])

    # ── 決まり手（直近1年・1着時のみ） ───────────────────────────────────
    cutoff = df["日付"].max() - pd.Timedelta(days=365)
    df_1y  = df[df["日付"] >= cutoff]

    if "決まり手" in df_1y.columns:
        ichi_1y = df_1y[df_1y["着順"] == 1]
        km = ichi_1y.groupby(KEY)["決まり手"] \
                    .value_counts().unstack(fill_value=0)
        for cat in ["逃げ", "差し", "まくり", "まくり差し", "抜き"]:
            if cat not in km.columns:
                km[cat] = 0
        km = km[["逃げ", "差し", "まくり", "まくり差し", "抜き"]].reset_index()

        ichi_1y_cnt = ichi_1y.groupby(KEY).size().reset_index(name="1着数_1y")
        km = km.merge(ichi_1y_cnt, on=KEY, how="left")
        denom = km["1着数_1y"].fillna(1).replace(0, 1)
        for cat in ["逃げ", "差し", "まくり", "まくり差し", "抜き"]:
            km[f"{cat}%"] = (km[cat] / denom).round(4)
        km = km.drop(columns=["逃げ", "差し", "まくり", "まくり差し", "抜き", "1着数_1y"])
        base = base.merge(km, on=KEY, how="left")

    # 最低出走数フィルタ
    base = base[base["出走数"] >= MIN_RUNS].copy()

    # ── 全国比: apply(lambda) → map + vectorized ──────────────────────────
    NATIONAL_AVG = {1: 0.555, 2: 0.137, 3: 0.134, 4: 0.111, 5: 0.066, 6: 0.021}
    base["1着率_全国比"] = (
        base["1着率"] - base["進入コース"].astype(int).map(NATIONAL_AVG).fillna(0.1)
    ).round(4)

    # 作業用列削除
    base = base.drop(columns=["_is1", "_is3"], errors="ignore")
    base = base.sort_values(KEY).reset_index(drop=True)
    print(f"  会場別コースマスタ: {len(base):,}件 集計完了")
    return base


def write_venue_course_master(wb, df: pd.DataFrame):
    """「会場別コースマスタ」シートへ書き込む"""
    sname = "会場別コースマスタ"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname)

    # ── スタイル ──
    H_FILL  = PatternFill("solid", start_color="1F4E79")
    H_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    S_FILL  = PatternFill("solid", start_color="2E75B6")
    S_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    D_FONT  = Font(name="Arial", size=9)
    POS_FILL = PatternFill("solid", start_color="E2EFDA")  # 全国比プラス（緑）
    NEG_FILL = PatternFill("solid", start_color="FCE4D6")  # 全国比マイナス（橙）
    AL_C    = Alignment(horizontal="center", vertical="center")
    AL_L    = Alignment(horizontal="left",   vertical="center")

    def bdr():
        s = Side(border_style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── タイトル行 ──
    ws.row_dimensions[1].height = 16
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    c = ws.cell(1, 1, "■ 会場別コースマスタ（選手×会場×コース の3次元実績）"
                      "  ※load_race.py の動的ハイブリッド係数に使用")
    c.fill = PatternFill("solid", start_color="E2EFDA")
    c.font = Font(name="Arial", size=9, italic=True)

    # ── ヘッダ行 ──
    headers = [
        "選手名", "会場名", "コース", "出走数", "1着数", "1着率",
        "3連対数", "3連対率", "平均ST", "信頼度",
        "逃げ%", "差し%", "まくり%", "まくり差し%", "抜き%",
        "1着率_全国比",
        "時系列補正1着率", "時系列補正3連対率", "時系列有効走数",  # 【重大①・修正④追加】
    ]
    col_widths = [16, 8, 6, 7, 7, 8, 7, 8, 8, 8, 8, 8, 8, 10, 8, 10, 10, 10, 8]
    col_keys   = [
        "選手名", "会場名", "進入コース", "出走数", "1着数", "1着率",
        "3連対数", "3連対率", "平均ST", "信頼度",
        "逃げ%", "差し%", "まくり%", "まくり差し%", "抜き%",
        "1着率_全国比",
        "時系列補正1着率", "時系列補正3連対率", "時系列有効走数",  # 【重大①・修正④追加】
    ]

    ws.row_dimensions[2].height = 22
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(2, ci, h)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── データ行（高速版）──
    # border/font/alignmentは省略（デフォルト）。%書式と条件付き背景のみ個別処理
    PCT_COLS = {6, 8, 11, 12, 13, 14, 15, 16, 17, 18}  # 1-indexed（%表示列）

    def _vc_clean(v):
        if isinstance(v, float):
            return round(v, 4) if pd.notna(v) else ""
        elif not isinstance(v, (str, int)) and pd.isna(v) if hasattr(pd, "isna") else False:
            return ""
        return v

    # to_dict(orient="records") で一括変換（iterrows より高速）
    data_rows_vc = []
    for row in df.to_dict(orient="records"):
        vals = [_vc_clean(row.get(key, "")) for key in col_keys]
        data_rows_vc.append(vals)

    # 一括append
    for vals in data_rows_vc:
        ws.append(vals)

    # %書式 + 全国比の条件付き背景のみ個別処理
    start_row = 3  # タイトル1行 + ヘッダ1行 + データ開始
    col16_idx  = col_keys.index("1着率_全国比") + 1  # 1-indexed
    for row_offset, vals in enumerate(data_rows_vc):
        r = start_row + row_offset
        for ci, v in enumerate(vals, 1):
            if ci in PCT_COLS and isinstance(v, float):
                ws.cell(r, ci).number_format = "0.0%"
        # 全国比カラムの条件付き色付け
        v16 = vals[col16_idx - 1]
        if isinstance(v16, float) and v16 != "":
            ws.cell(r, col16_idx).fill = POS_FILL if v16 > 0 else NEG_FILL

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(df)+2}"
    print(f"  会場別コースマスタ: {len(df):,}行 書き込み完了")

    # CSVキャッシュに書き出す（load_race.py が高速読み込みに使用）
    # Excelのイテレートより pd.read_csv の方が5〜10倍速いため
    try:
        os.makedirs(os.path.dirname(VENUE_COURSE_CSV), exist_ok=True)
        # col_keysと同じ列名でCSV保存（load_race.py側でそのまま読めるよう統一）
        save_cols = [c for c in col_keys if c in df.columns]
        df[save_cols].to_csv(VENUE_COURSE_CSV, index=False, encoding="utf-8-sig")
        print(f"  会場別コースマスタCSV: {VENUE_COURSE_CSV} に書き出し完了")
    except Exception as e:
        print(f"  ⚠️  会場別コースマスタCSV書き出し失敗（Excelは正常更新済み）: {e}")


# -----------------------------------------------------------------------------
# 7. Excelシートへ書き込む
# -----------------------------------------------------------------------------
HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
SUBHEAD_FILL  = PatternFill("solid", start_color="2E75B6")
SUBHEAD_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
UPDATE_FILL   = PatternFill("solid", start_color="E2EFDA")
DATA_FONT     = Font(name="Arial", size=9)
STAR_FILL     = PatternFill("solid", start_color="FFF2CC")
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")

def thin_border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_course_master(wb, df: pd.DataFrame, n_raw: int, n_ippan: int, n_senshu: int):
    sname = "📊コース別マスタ"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname)

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    kimete_cols = ["決まり手_逃げ%", "決まり手_差し%", "決まり手_まくり%", "決まり手_まくり差し%", "決まり手_抜き%", "決まり手_恵まれ%"]
    kimete_ken  = ["決まり手_逃げ_件数", "決まり手_差し_件数", "決まり手_まくり_件数", "決まり手_まくり差し_件数", "決まり手_抜き_件数", "決まり手_恵まれ_件数"]

    # -- 行1: 更新情報 -----------------------------------------------------
    ws.row_dimensions[1].height = 18
    c = ws.cell(1, 1, f"最終更新: {now_str}  ／  RAW: {n_raw:,}行  ／  一般戦: {n_ippan:,}行  ／  選手×コース: {len(df):,}件")
    c.fill = UPDATE_FILL
    c.font = Font(name="Arial", size=9, italic=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=33)

    # -- 行2: グループヘッダー ---------------------------------------------
    ws.row_dimensions[2].height = 16
    groups = [
        (1, 3, "基本情報"),
        (4, 9, "1着・3連対"),
        (10, 13, "コース別ST"),
        (14, 19, "決まり手（1着時）"),
        (20, 25, "イン逃げ時成績"),
        (26, 28, "1コースST"),
        (29, 31, "暫定★"),
        (32, 34, "決まり手(件数)"),
        (35, 38, "1C敗戦パターン"),
        (39, 41, "時系列補正"),  # 【重大①・修正④追加】
    ]
    for s, e, label in groups:
        ws.cell(2, s, label).fill = HEADER_FILL
        ws.cell(2, s).font = HEADER_FONT
        ws.cell(2, s).alignment = CENTER
        if s != e:
            ws.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)

    # -- 行3: 列ヘッダー --------------------------------------------------
    ws.row_dimensions[3].height = 28
    headers = [
        "選手名", "コース", "出走数",
        "1着数", "1着率", "3連対数", "3連対率", "1着率\n相対", "3連対率\n相対",
        "コース別\n平均ST", "ST件数", "最速ST", "最遅ST",
        "逃げ%", "差し%", "まくり%", "まくり差し%", "抜き%", "恵まれ%",
        "イン逃げ\n出走数", "2着数", "2着率", "3着以内数", "3着以内率", "2着率\n相対",
        "1C\n出走数", "コース別\n平均ST", "コース別\nST順位",
        "★1着率", "★ST", "★決手",
        "逃げ(件)", "差し(件)", "まくり(件)",
        "C1敗戦数", "差され%", "捲られ%", "捲り差され%",
        "時系列補正\n1着率", "時系列補正\n3連対率", "時系列\n有効走数",  # 【重大①・修正④追加】
    ]
    col_widths = [
        18, 6, 8,
        8, 8, 8, 8, 8, 8,
        9, 7, 8, 8,
        8, 8, 8, 10, 8, 8,
        8, 8, 8, 8, 8, 8,
        8, 8, 8,
        7, 7, 7,
        8, 8, 8,
        8, 8, 8, 10,
        10, 10, 8,  # 【重大①・修正④追加】
    ]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(3, ci, h)
        cell.fill = SUBHEAD_FILL
        cell.font = SUBHEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # -- 相対評価の計算（コースごとに正規化） ----------------------------
    for course in range(1, 7):
        mask = df["進入コース"] == course
        for col in ["1着率", "3連対率"]:
            total = df.loc[mask, col].sum()
            if total > 0:
                df.loc[mask, f"{col}_相対"] = df.loc[mask, col] / total
            else:
                df.loc[mask, f"{col}_相対"] = np.nan

    # イン逃げ2着率の相対評価
    for course in range(2, 7):
        mask = df["進入コース"] == course
        total = df.loc[mask, "イン逃げ2着率"].sum() if "イン逃げ2着率" in df.columns else 0
        if total > 0:
            df.loc[mask, "2着率_相対"] = df.loc[mask, "イン逃げ2着率"] / total
        else:
            df.loc[mask, "2着率_相対"] = np.nan

    # ──────────────────────────────────────────────────────────────────────────
    # 高速書き込み: openpyxlの最大ボトルネック「セルごとスタイル適用」を最小化
    # ──────────────────────────────────────────────────────────────────────────
    # ❌ 旧方式: for row: ws.append() → for ci: cell.font/border/alignment  (行×列 = 数万回)
    # ✅ 新方式:
    #   1. ws.append() でデータだけ一括書き込み（書式なし）
    #   2. 列スタイルをColumnDimensionに設定（全行一括）
    #   3. 個別書式（★背景色・%フォーマット）は該当セルのみ
    # border/font/alignmentは列単位の共通スタイルをnamed_styleで一括付与する代わりに
    # ws.sheet_format と ws.column_dimensions で処理する
    #
    # ★行（is_star）とパーセント書式のみ個別セル処理（全体の数%以下）
    # ──────────────────────────────────────────────────────────────────────────

    # パーセント表示列（1-indexed）
    PCT_COLS_IDX = {5, 7, 8, 9, 14, 15, 16, 17, 18, 19, 22, 24, 25, 36, 37, 38, 39, 40}

    def _clean(v):
        if isinstance(v, float):
            return round(v, 4) if not np.isnan(v) else ""
        if isinstance(v, (int, np.integer)):
            return int(v)
        if v is None or (not isinstance(v, str) and pd.isna(v)):
            return ""
        return v

    _bdr  = thin_border()
    _star_fill = STAR_FILL
    _d_font    = DATA_FONT
    _center    = CENTER
    _left      = LEFT

    # ── 全行をappendで一括書き込み（to_dict で iterrows より高速に変換）──
    data_rows = []
    star_rows = []  # ★フラグ行の行インデックスを記録（後で個別処理）
    for ri_idx, row in enumerate(df.to_dict(orient="records")):
        is_star = bool(row.get("★1着率", False))
        vals = [
            row["選手名"],
            _clean(row["進入コース"]),
            _clean(row["出走数"]),
            _clean(row["1着数"]),
            _clean(row["1着率"]),
            _clean(row["3連対数"]),
            _clean(row["3連対率"]),
            _clean(row.get("1着率_相対")),
            _clean(row.get("3連対率_相対")),
            _clean(row.get("コース別平均ST")),
            _clean(row.get("コース別ST件数")),
            _clean(row.get("コース別最速ST")),
            _clean(row.get("コース別最遅ST")),
            _clean(row.get("決まり手_逃げ%")),
            _clean(row.get("決まり手_差し%")),
            _clean(row.get("決まり手_まくり%")),
            _clean(row.get("決まり手_まくり差し%")),
            _clean(row.get("決まり手_抜き%")),
            _clean(row.get("決まり手_恵まれ%")),
            _clean(row.get("イン逃げ出走数")),
            _clean(row.get("イン逃げ2着数")),
            _clean(row.get("イン逃げ2着率")),
            _clean(row.get("イン逃げ3着以内数")),
            _clean(row.get("イン逃げ3着以内率")),
            _clean(row.get("2着率_相対")),
            _clean(row.get("C1出走数")),
            _clean(row.get("コース別平均ST")),
            _clean(row.get("コース別ST順位")),
            "★" if row.get("★1着率") else "",
            "★" if row.get("★ST")    else "",
            "★" if row.get("★決手")  else "",
            _clean(row.get("決まり手_逃げ_件数")),
            _clean(row.get("決まり手_差し_件数")),
            _clean(row.get("決まり手_まくり_件数")),
            _clean(row.get("C1敗戦数")),
            _clean(row.get("差され%")),
            _clean(row.get("捲られ%")),
            _clean(row.get("捲り差され%")),
            _clean(row.get("時系列補正1着率")),
            _clean(row.get("時系列補正3連対率")),
            _clean(row.get("時系列有効走数")),
        ]
        data_rows.append((vals, is_star))

    # 一括append（書式なし）
    for vals, _ in data_rows:
        ws.append(vals)

    # ── 必要最小限のセルのみ個別スタイル適用 ──
    n_cols = 41  # 時系列有効走数追加で40→41列
    start_data_row = 4  # ヘッダ3行 + データ開始

    # パーセント書式 + ★背景色のみ個別処理（border/font/alignmentは省略 → デフォルト）
    for row_offset, (vals, is_star) in enumerate(data_rows):
        r = start_data_row + row_offset
        for ci, v in enumerate(vals, 1):
            if ci in PCT_COLS_IDX and isinstance(v, float):
                ws.cell(r, ci).number_format = "0.0%"
            if is_star:
                ws.cell(r, ci).fill = _star_fill

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(41)}{len(df)+3}"
    print(f"  📊コース別マスタ: {len(df):,}行 書き込み完了")


def write_senshu_master(wb, df: pd.DataFrame):
    sname = "選手指数マスタ"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname)

    ws.row_dimensions[1].height = 22
    ws.cell(1, 1, "■ 選手別独自指数マスタ（データ蓄積で自動精度向上）").font = Font(bold=True, name="Arial", size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=28)

    headers = [
        "登録番号", "選手名", "総出走数", "1着率\n(全体)", "3連対率\n(全体)",
        "1着率\n(一般戦)", "3連対率\n(一般戦)", "一般戦\n出走数",
        "イン\n出走数", "イン\n1着率", "自在性\n1着率", "自在性\n加重1着率", "自在性\n(4-6C出走数)",
        "平均ST\n(全体)",
        "平均ST\n(1コース)", "平均ST\n(2コース)", "平均ST\n(3コース)",
        "平均ST\n(4コース)", "平均ST\n(5コース)", "平均ST\n(6コース)",
        "ST順位\n(1コース)", "ST順位\n(2コース)", "ST順位\n(3コース)",
        "ST順位\n(4コース)", "ST順位\n(5コース)", "ST順位\n(6コース)",
        "フォーム\n指数", "直近10走\n平均着順", "直近3走\n1着率", "直近5走\n1着率", "直近5走\n3連対率",
        "ST安定\nスコア", "ST\n標準偏差", "FLY数", "出遅れ数", "ST\n計測件数",
        # 【修正④】FLY経過日数
        "FLY経過\n日数",
        # 【新規】FLY前後ST比較
        "FLY前\nST平均", "FLY後\nST平均", "FLY後\nST変化量", "FLY\n影響度",
    ]
    col_map = [
        "登録番号", "選手名", "総出走数", "1着率(全体)", "3連対率(全体)",
        "1着率(一般戦)", "3連対率(一般戦)", "一般戦出走数",
        "イン出走数", "イン1着率", "自在性1着率", "自在性加重1着率", "自在性出走数(4-6C)",
        "平均ST(全体)",
        "平均ST(1コース)", "平均ST(2コース)", "平均ST(3コース)",
        "平均ST(4コース)", "平均ST(5コース)", "平均ST(6コース)",
        "ST順位\n(1コース)", "ST順位\n(2コース)", "ST順位\n(3コース)",
        "ST順位\n(4コース)", "ST順位\n(5コース)", "ST順位\n(6コース)",  # Excelヘッダと統一
        "フォーム指数", "直近10走平均着順", "直近3走1着率", "直近5走1着率", "直近5走3連対率",
        "ST安定スコア", "ST標準偏差", "FLY数", "出遅れ数", "ST計測件数",
        # 【修正④】FLY経過日数
        "FLY経過日数",
        # 【新規】FLY前後ST比較
        "FLY前ST平均", "FLY後ST平均", "FLY後ST変化量", "FLY影響度",
    ]
    col_widths = [10, 16, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9,
                  10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10,
                  10, 10, 10, 10, 10, 10, 10, 8, 8, 9,
                  9,                          # FLY経過日数
                  9, 9, 9, 8]                 # FLY前後ST比較

    ws.row_dimensions[2].height = 30
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(2, ci, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    _bdr = thin_border()
    # to_dict(orient="records") で一括変換（iterrows より大幅に高速）
    senshu_rows = []
    for row in df.to_dict(orient="records"):
        vals = []
        for col in col_map:
            v = row.get(col, "")
            if col == "登録番号":
                vals.append(str(row.get("登録番号", "")) if v is not None and str(v) != "nan" else "")
            elif isinstance(v, float):
                import math
                vals.append(round(v, 4) if not math.isnan(v) else "")
            elif v is None or (not isinstance(v, str) and pd.isna(v) if hasattr(pd, "isna") else False):
                vals.append("")
            else:
                vals.append(v)
        senshu_rows.append(vals)

    for vals in senshu_rows:
        ws.append(vals)

    # %フォーマット適用（1-indexed）
    # 4:1着率(全体) 5:3連対率(全体) 6:1着率(一般戦) 7:3連対率(一般戦)
    # 10:イン1着率 11:自在性1着率 12:自在性加重1着率
    # 29:直近3走1着率 30:直近5走1着率 31:直近5走3連対率
    SENSHU_PCT_COLS = {4, 5, 6, 7, 10, 11, 12, 29, 30, 31}
    start_row_s = 3  # タイトル1行 + ヘッダ1行 + データ開始
    for row_offset, vals in enumerate(senshu_rows):
        r = start_row_s + row_offset
        for ci, v in enumerate(vals, 1):
            if ci in SENSHU_PCT_COLS and isinstance(v, float):
                ws.cell(r, ci).number_format = "0.0%"

    n_cols = len(headers)
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{len(df)+2}"
    print(f"  選手指数マスタ: {len(df):,}行 書き込み完了")


def write_kaijo_stats(wb, df: pd.DataFrame):
    sname = "会場統計"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname)

    # 集計期間を取得（calc_kaijo_stats が設定）
    _period = getattr(calc_kaijo_stats, "_period", None)
    _period_str = f"  集計期間: {_period[0]} 〜 {_period[1]}（直近1年）" if _period else ""
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    title_text = f"■ 会場別 統計データ（決まり手・イン逃げ率・コース別1着率）  最終更新: {now_str}{_period_str}"
    ws.cell(1, 1, title_text).font = Font(bold=True, name="Arial", size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.row_dimensions[1].height = 18

    # ── 動的列を収集 ──
    r_cols        = sorted([c for c in df.columns if c.endswith("R荒れスコア")],
                           key=lambda x: int(x.replace("R荒れスコア", "")))
    course_cols   = [f"{c}コース1着率"      for c in range(1, 7)]
    course_diff   = [f"{c}コース1着率_全国比" for c in range(1, 7)]
    # R×C列：1C_1R〜6C_12R（存在するもののみ）
    rc_cols = sorted([c for c in df.columns if len(c) >= 8 and c[1] == "C" and c.endswith("R1着率")],
                     key=lambda x: (int(x[0]), int(x.split("_")[1].replace("R1着率", ""))))

    # ── セクション別にグループヘッダーを作る ──
    BASE_COLS    = ["会場名", "レース数", "イン逃げ率", "逃げ率", "差し率",
                    "まくり率", "まくり差し率", "抜き率", "荒れやすさスコア"]
    all_cols     = BASE_COLS + course_cols + course_diff + r_cols + rc_cols

    FILL_SEC1 = PatternFill("solid", start_color="1F4E79")  # 基本情報（濃紺）
    FILL_SEC2 = PatternFill("solid", start_color="375623")  # コース別1着率（緑）
    FILL_SEC3 = PatternFill("solid", start_color="7F3F00")  # 全国比（茶）
    FILL_SEC4 = PatternFill("solid", start_color="4472C4")  # R別荒れ（青）
    FILL_SEC5 = PatternFill("solid", start_color="833C00")  # R×C（橙）
    FONT_SEC  = Font(bold=True, color="FFFFFF", name="Arial", size=8)

    # 行2：セクションヘッダー
    ws.row_dimensions[2].height = 14
    sections = [
        (1,          len(BASE_COLS),                              "基本情報",         FILL_SEC1),
        (len(BASE_COLS)+1, len(BASE_COLS)+6,                      "コース別1着率",    FILL_SEC2),
        (len(BASE_COLS)+7, len(BASE_COLS)+12,                     "全国平均比",       FILL_SEC3),
        (len(BASE_COLS)+13, len(BASE_COLS)+12+len(r_cols),        "R別荒れスコア",    FILL_SEC4),
        (len(BASE_COLS)+13+len(r_cols), len(all_cols),            "R×コース別1着率",  FILL_SEC5),
    ]
    for cs, ce, label, fill in sections:
        if cs > len(all_cols): continue
        ce = min(ce, len(all_cols))
        ws.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=ce)
        cell = ws.cell(2, cs, label)
        cell.fill = fill; cell.font = FONT_SEC; cell.alignment = CENTER

    # 行3：列ヘッダー
    ws.row_dimensions[3].height = 28
    header_labels = {
        "会場名": "会場名", "レース数": "レース数", "イン逃げ率": "イン逃げ率",
        "逃げ率": "逃げ率", "差し率": "差し率", "まくり率": "まくり率",
        "まくり差し率": "まくり\n差し率", "抜き率": "抜き率",
        "荒れやすさスコア": "荒れ\nスコア",
    }
    for c in range(1, 7):
        header_labels[f"{c}コース1着率"]       = f"{c}C\n1着率"
        header_labels[f"{c}コース1着率_全国比"] = f"{c}C\n全国比"
    for col in r_cols:
        header_labels[col] = col.replace("R荒れスコア", "R\n荒れ")
    for col in rc_cols:
        # "1C_7R1着率" → "1C\n7R"
        header_labels[col] = col.replace("R1着率", "R").replace("_", "\n")

    col_widths = {
        "会場名": 12, "レース数": 8, "イン逃げ率": 9, "逃げ率": 8, "差し率": 8,
        "まくり率": 8, "まくり差し率": 9, "抜き率": 8, "荒れやすさスコア": 8,
    }

    for ci, col in enumerate(all_cols, 1):
        label = header_labels.get(col, col)
        cell = ws.cell(3, ci, label)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        w = col_widths.get(col, 7)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── データ行（高速版）──
    FILL_DIFF_POS = PatternFill("solid", start_color="E2EFDA")  # 全国比+（緑薄）
    FILL_DIFF_NEG = PatternFill("solid", start_color="FCE4D6")  # 全国比-（橙薄）
    _bdr = thin_border()
    # パーセント書式が必要な列インデックス（1-indexed）を事前計算
    pct_col_idx  = {ci for ci, col in enumerate(all_cols, 1) if "率" in col and col != "レース数"}
    diff_col_idx = {ci for ci, col in enumerate(all_cols, 1) if "全国比" in col}

    kaijo_rows = []
    for row in df.to_dict(orient="records"):
        vals = []
        for col in all_cols:
            v = row.get(col, "")
            if isinstance(v, float):
                import math
                v = round(v, 4) if not math.isnan(v) else ""
            vals.append(v)
        kaijo_rows.append(vals)

    for vals in kaijo_rows:
        ws.append(vals)

    # %書式・条件付き色のみ個別処理
    start_row = 4  # タイトル2行 + ヘッダ1行
    for row_offset, vals in enumerate(kaijo_rows):
        r = start_row + row_offset
        for ci, v in enumerate(vals, 1):
            if ci in pct_col_idx and isinstance(v, float):
                ws.cell(r, ci).number_format = "0.0%"
            if ci in diff_col_idx and isinstance(v, float) and v != "":
                ws.cell(r, ci).fill = FILL_DIFF_POS if v > 0 else FILL_DIFF_NEG

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(all_cols))}3"
    print(f"  会場統計: {len(df):,}行・{len(all_cols)}列 書き込み完了")


def write_in_nige(wb, df: pd.DataFrame):
    sname = "イン逃げ分析"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname)

    # 集計期間を取得（calc_in_nige_analysis が設定）
    _period = getattr(calc_in_nige_analysis, "_period", None)
    _period_str = f"  集計期間: {_period[0]} 〜 {_period[1]}（直近1年）" if _period else ""
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    title_text = f"■ イン逃げ時 2着・3着 枠別分布（会場別）  最終更新: {now_str}{_period_str}"
    ws.cell(1, 1, title_text).font = Font(bold=True, name="Arial", size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws.row_dimensions[1].height = 18

    hdrs = ["会場名", "イン逃げ回数"] + \
           [f"{c}枠\n2着率" for c in range(1,7)] + \
           [f"{c}枠\n3着以内率" for c in range(1,7)]
    cols = ["会場名", "イン逃げ回数"] + \
           [f"{c}枠2着率" for c in range(1,7)] + \
           [f"{c}枠3着以内率" for c in range(1,7)]

    ws.row_dimensions[2].height = 28
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(2, ci, h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.column_dimensions["A"].width = 12

    _bdr = thin_border()
    # to_dict(orient="records") で一括変換（iterrows より高速）
    in_nige_rows = []
    for row in df.to_dict(orient="records"):
        vals = []
        for col in cols:
            v = row.get(col, "")
            if isinstance(v, float):
                import math
                v = round(v, 4) if not math.isnan(v) else ""
            vals.append(v)
        in_nige_rows.append(vals)

    for vals in in_nige_rows:
        ws.append(vals)

    # %フォーマット適用（3列目以降がすべて率列）
    # cols: 会場名(1), イン逃げ回数(2), 1枠2着率(3)〜6枠2着率(8), 1枠3着以内率(9)〜6枠3着以内率(14)
    IN_NIGE_PCT_COLS = set(range(3, 15))  # 3〜14列
    start_row_in = 3  # タイトル1行 + ヘッダ1行 + データ開始
    for row_offset, vals in enumerate(in_nige_rows):
        r = start_row_in + row_offset
        for ci, v in enumerate(vals, 1):
            if ci in IN_NIGE_PCT_COLS and isinstance(v, float):
                ws.cell(r, ci).number_format = "0.0%"

    print(f"  イン逃げ分析: {len(df):,}行 書き込み完了")


def write_guide(wb):
    """📖指数ガイドシートを作成（新規追加指数の説明）"""
    sname = "📖指数ガイド"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname)

    # ── カラー & スタイル定義 ──────────────────────────────────
    C_NAVY   = "1F4E79"
    C_BLUE   = "2E75B6"
    C_LBLUE  = "D6E4F0"
    C_WHITE  = "FFFFFF"
    C_YELLOW = "FFF3CD"
    C_GREEN  = "E2EFDA"
    C_RED_BG = "FCE4D6"
    C_GRAY   = "F2F2F2"

    FONT_H1    = Font(name="Arial", size=12, bold=True,  color=C_WHITE)
    FONT_H2    = Font(name="Arial", size=10, bold=True,  color=C_WHITE)
    FONT_LABEL = Font(name="Arial", size=9,  bold=True,  color="000000")
    FONT_BODY  = Font(name="Arial", size=9,  bold=False, color="333333")
    FONT_SMALL = Font(name="Arial", size=8,  bold=False, color="666666")
    FONT_BADGE = Font(name="Arial", size=9,  bold=True,  color=C_WHITE)
    FONT_WARN  = Font(name="Arial", size=9,  bold=True,  color="CC0000")

    FILL_NAVY   = PatternFill("solid", start_color=C_NAVY)
    FILL_BLUE   = PatternFill("solid", start_color=C_BLUE)
    FILL_LBLUE  = PatternFill("solid", start_color=C_LBLUE)
    FILL_SILVER = PatternFill("solid", start_color="D9D9D9")
    FILL_WHITE  = PatternFill("solid", start_color=C_WHITE)
    FILL_YELLOW = PatternFill("solid", start_color=C_YELLOW)
    FILL_GREEN  = PatternFill("solid", start_color=C_GREEN)
    FILL_RED    = PatternFill("solid", start_color=C_RED_BG)
    FILL_GRAY   = PatternFill("solid", start_color=C_GRAY)

    AL_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AL_LT = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    def bdr():
        s = Side(border_style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def wc(row, col, val="", fill=None, font=None, align=None):
        c = ws.cell(row, col, val)
        if fill:  c.fill      = fill
        if font:  c.font      = font
        if align: c.alignment = align
        c.border = bdr()
        return c

    def mg(r1, c1, r2, c2):
        try: ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        except Exception as e: print(f"  ⚠️  セル結合スキップ（行{r1}列{c1}〜行{r2}列{c2}）: {e}")

    def sec(row, cs, ce, title):
        mg(row, cs, row, ce)
        wc(row, cs, title, fill=FILL_NAVY, font=FONT_H1, align=AL_C)
        ws.row_dimensions[row].height = 22

    def sub(row, cs, ce, title):
        mg(row, cs, row, ce)
        wc(row, cs, title, fill=FILL_BLUE, font=FONT_H2, align=AL_C)
        ws.row_dimensions[row].height = 18

    def hdr(row, items):
        for col, val in items:
            wc(row, col, val, fill=FILL_SILVER, font=FONT_LABEL, align=AL_C)
        ws.row_dimensions[row].height = 18

    def blk(row, cs, ce, text, fill=None, h=15):
        mg(row, cs, row, ce)
        wc(row, cs, text, fill=fill or FILL_WHITE, font=FONT_BODY, align=AL_LT)
        ws.row_dimensions[row].height = h

    # ── 列幅 ──────────────────────────────────────────────────
    for col, w in [("A",3),("B",15),("C",26),("D",16),("E",14),("F",20),
                   ("G",3), ("H",15),("I",26),("J",16),("K",3)]:
        ws.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════
    # タイトルバナー（行2）
    # ════════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 8

    mg(2, 2, 2, 10)
    wc(2, 2, "📖  ボートリサーチ新聞  ／  新規追加 指数ガイド",
       fill=FILL_NAVY, font=Font(name="Arial", size=14, bold=True, color=C_WHITE), align=AL_C)
    mg(3, 2, 3, 10)
    wc(3, 2, "update_master.py 追加3指数 ＋ load_race.py 展示タイム偏差値・総合スコア 解説",
       fill=FILL_LBLUE, font=Font(name="Arial", size=9, color=C_NAVY), align=AL_C)

    # ════════════════════════════════════════════════════════════
    # 一覧表（行5〜）
    # ════════════════════════════════════════════════════════════
    R = 5
    sec(R, 2, 10, "📋  追加指数 一覧"); R += 1
    hdr(R, [(2,"指数名"),(3,"追加ファイル"),(4,"出力先"),(5,"目的"),(6,"判定例")])
    R += 1

    for name, f, out, pur, ex, bg in [
        ("📈 フォーム指数",    "update_master.py", "📈フォーム指数シート",   "直近好不調の数値化",  "🔥ホット〜❄コールド", "E2EFDA"),
        ("🎲 荒れやすさスコア","update_master.py", "🎲荒れ予報シート",       "レース波乱度の把握",  "🔵堅い〜🔴大荒れ",    "D6E4F0"),
        ("⚡ ST安定度",        "update_master.py", "⚡ST安定度シート",        "スタート信頼性の評価","◎超安定〜×不安定",   "FFF3CD"),
        ("📊 展示タイム偏差値","load_race.py",     "新聞 col7「展示偏差値」","当日機力の相対評価",  "偏差値50中心",        "F2F2F2"),
        ("🏆 総合スコア",      "load_race.py",     "新聞 col8「総合スコア」","全指数の統合評価",    "◎○△自動表示",       "F2F2F2"),
    ]:
        fill = PatternFill("solid", start_color=bg)
        wc(R, 2, name,        fill=fill, font=FONT_LABEL, align=AL_L)
        wc(R, 3, f,           fill=fill, font=FONT_BODY,  align=AL_L)
        wc(R, 4, out,         fill=fill, font=FONT_BODY,  align=AL_L)
        wc(R, 5, pur,         fill=fill, font=FONT_BODY,  align=AL_L)
        wc(R, 6, ex,          fill=fill, font=FONT_BODY,  align=AL_L)
        ws.row_dimensions[R].height = 16; R += 1

    ws.row_dimensions[R].height = 8; R += 1

    # ════════════════════════════════════════════════════════════
    # 左ブロック B〜F  ／  右ブロック H〜J
    # ════════════════════════════════════════════════════════════
    START = R

    # ─── 左: フォーム指数 ───────────────────────────────────
    sec(R, 2, 6, "📈  フォーム指数"); R += 1
    blk(R, 2, 6, "直近10走（一般戦）を加重平均で評価。最新走ほど重み大。好不調をリアルタイムに反映。", FILL_LBLUE, h=24); R += 1
    sub(R, 2, 6, "計算ロジック"); R += 1
    hdr(R, [(2,"項目"),(3,"内容")]); mg(R, 3, R, 6); R += 1
    for k, v in [("参照走数","直近10走（一般戦のみ）"),
                 ("重み付け","直近=最大（線形ウェイト n, n-1, ..., 1）"),
                 ("着順点数","1着=10 / 2着=7 / 3着=5 / 4着=3 / 5着=1 / 6着=0"),
                 ("コース難易度補正","1着時のみ適用: 1C×1.0, 2C×1.8, 3C×2.2, 4C×2.4, 5C×2.6, 6C×2.8"),
                 ("計算式",  "フォーム指数 = Σ(重み×難易度補正点数) ÷ Σ重み"),
                 ("補助指標","直近3走1着率 / 直近5走1着率 / 直近5走3連対率")]:
        wc(R, 2, k, fill=FILL_LBLUE, font=FONT_LABEL, align=AL_L)
        mg(R, 3, R, 6); wc(R, 3, v, fill=FILL_WHITE, font=FONT_BODY, align=AL_L)
        ws.row_dimensions[R].height = 15; R += 1
    sub(R, 2, 6, "判定ランク"); R += 1
    hdr(R, [(2,"ランク"),(3,"条件"),(4,"意味")]); mg(R, 4, R, 6); R += 1
    for lbl, cond, meaning, color in [
        ("🔥 ホット",   "8.0以上",      "絶好調。積極的に評価。",       "FF6600"),
        ("▲ 普通",     "6.0〜8.0未満", "平均的。他の指数と組合わせ。", "4472C4"),
        ("▽ やや不調", "4.0〜6.0未満", "やや不振。割引き要素。",       "767171"),
        ("❄ コールド", "4.0未満",      "直近不振。評価を下げる。",     "1F4E79"),
    ]:
        wc(R, 2, lbl,    fill=PatternFill("solid", start_color=color), font=FONT_BADGE, align=AL_C)
        wc(R, 3, cond,   fill=FILL_WHITE, font=FONT_BODY, align=AL_C)
        mg(R, 4, R, 6); wc(R, 4, meaning, fill=FILL_WHITE, font=FONT_BODY, align=AL_L)
        ws.row_dimensions[R].height = 15; R += 1
    sub(R, 2, 6, "活用ポイント"); R += 1
    for tip in ["・🔥ホット × コース1着率が高い → ◎最有力",
                "・❄コールドは実力者でも当日は割引き",
                "・直近5走1着率もあわせて確認"]:
        blk(R, 2, 6, tip, FILL_GREEN, h=15); R += 1
    L1_END = R

    # ─── 右: ST安定度 ──────────────────────────────────────
    R = START
    sec(R, 8, 10, "⚡  ST安定度"); R += 1
    blk(R, 8, 10, "コース別STの標準偏差でばらつきを判定。小さいほど安定。", FILL_LBLUE, h=24); R += 1
    sub(R, 8, 10, "計算ロジック"); R += 1
    hdr(R, [(8,"項目"),(9,"内容")]); mg(R, 9, R, 10); R += 1
    for k, v in [("集計単位","選手 × 一般戦全コース"),
                 ("ST標準偏差","ばらつき（秒）。小さい→安定"),
                 ("スコア基準","0.03以下→95点 / 0.05→80点 / 0.08→50点（区間線形補間）"),
                 ("FLY数","ST < 0 の回数（フライング）"),
                 ("出遅れ数","ST > 0.18秒 の回数"),
                 ("最低件数","5件未満は除外")]:
        wc(R, 8, k, fill=FILL_LBLUE, font=FONT_LABEL, align=AL_L)
        mg(R, 9, R, 10); wc(R, 9, v, fill=FILL_WHITE, font=FONT_BODY, align=AL_L)
        ws.row_dimensions[R].height = 15; R += 1
    sub(R, 8, 10, "安定ランク"); R += 1
    hdr(R, [(8,"ランク"),(9,"標準偏差")]); mg(R, 9, R, 10); R += 1
    for lbl, cond, color in [("◎ 超安定","0.03以下","375623"),
                              ("○ 安定",  "0.03〜0.05","4472C4"),
                              ("△ 普通",  "0.05〜0.08","7F6000"),
                              ("× 不安定","0.08超",   "C00000")]:
        wc(R, 8, lbl,  fill=PatternFill("solid", start_color=color), font=FONT_BADGE, align=AL_C)
        mg(R, 9, R, 10); wc(R, 9, cond, fill=FILL_WHITE, font=FONT_BODY, align=AL_C)
        ws.row_dimensions[R].height = 15; R += 1
    sub(R, 8, 10, "活用ポイント"); R += 1
    for tip in ["・◎超安定×平均ST速い→スリット信頼度UP",
                "・F（フライング）多い選手は要注意"]:
        blk(R, 8, 10, tip, FILL_GREEN, h=15); R += 1
    R_ST_END = R

    # 右下: 展示タイム偏差値
    ws.row_dimensions[R].height = 8; R += 1
    sec(R, 8, 10, "📊  展示タイム偏差値"); R += 1
    blk(R, 8, 10, "同レース内で相対評価。速い(小さい)ほど偏差値高。", FILL_LBLUE, h=20); R += 1
    sub(R, 8, 10, "計算式"); R += 1
    blk(R, 8, 10, "偏差値 = 50 + (平均−自分) ÷ 標準偏差 × 10", FILL_YELLOW, h=18); R += 1
    sub(R, 8, 10, "読み方"); R += 1
    hdr(R, [(8,"偏差値"),(9,"評価")]); mg(R, 9, R, 10); R += 1
    for lbl, meaning, color in [("60以上","◎ 最速クラス","375623"),
                                 ("55〜59","○ 平均より速い","4472C4"),
                                 ("45〜54","△ 平均的","767171"),
                                 ("44以下","× 平均より遅い","C00000")]:
        wc(R, 8, lbl,    fill=PatternFill("solid", start_color=color), font=FONT_BADGE, align=AL_C)
        mg(R, 9, R, 10); wc(R, 9, meaning, fill=FILL_WHITE, font=FONT_BODY, align=AL_L)
        ws.row_dimensions[R].height = 15; R += 1
    R_TENJI_END = R

    # ─── 左下: 荒れやすさスコア ─────────────────────────────
    R = L1_END
    ws.row_dimensions[R].height = 8; R += 1
    sec(R, 2, 6, "🎲  荒れやすさスコア"); R += 1
    blk(R, 2, 6, "会場×レース番号ごとの1コース1着率から算出。低いほど荒れやすい。", FILL_LBLUE, h=24); R += 1
    sub(R, 2, 6, "計算ロジック"); R += 1
    hdr(R, [(2,"項目"),(3,"内容")]); mg(R, 3, R, 6); R += 1
    for k, v in [("集計単位","会場名 × レース番号"),
                 ("1コース1着率","過去の該当レースでの1C1着割合"),
                 ("荒れやすさ","1 − 1コース1着率（0〜1）"),
                 ("攻め決まり手率","差し・まくり・まくり差しの割合"),
                 ("最低件数","5レース未満は除外")]:
        wc(R, 2, k, fill=FILL_LBLUE, font=FONT_LABEL, align=AL_L)
        mg(R, 3, R, 6); wc(R, 3, v, fill=FILL_WHITE, font=FONT_BODY, align=AL_L)
        ws.row_dimensions[R].height = 15; R += 1
    sub(R, 2, 6, "荒れ予報ランク"); R += 1
    hdr(R, [(2,"ランク"),(3,"1C1着率"),(4,"活用")]); mg(R, 4, R, 6); R += 1
    for lbl, cond, meaning, color in [
        ("🔵 堅い",     "60%以上",  "本命軸の流し",            "1F4E79"),
        ("🟡 中穴",    "45〜60%",  "2〜3着の枠に注目",        "7F6000"),
        ("🟠 荒れ気味", "30〜45%", "外枠・差しまくりを意識",  "C55A11"),
        ("🔴 大荒れ",   "30%未満",  "高配当狙い。1C軸は危険", "C00000"),
    ]:
        wc(R, 2, lbl,     fill=PatternFill("solid", start_color=color), font=FONT_BADGE, align=AL_C)
        wc(R, 3, cond,    fill=FILL_WHITE, font=FONT_BODY, align=AL_C)
        mg(R, 4, R, 6); wc(R, 4, meaning, fill=FILL_WHITE, font=FONT_BODY, align=AL_L)
        ws.row_dimensions[R].height = 15; R += 1
    sub(R, 2, 6, "活用ポイント"); R += 1
    for tip in ["・🔵堅い → 1C本命軸の流し",
                "・🔴大荒れ → 3連単高配当狙い",
                "・攻め決まり手率で「何で荒れるか」を判断"]:
        blk(R, 2, 6, tip, FILL_GREEN, h=15); R += 1
    L2_END = R

    # ════════════════════════════════════════════════════════════
    # 総合活用フロー（フルwidth）
    # ════════════════════════════════════════════════════════════
    FLOW_R = max(L2_END, R_TENJI_END) + 2
    ws.row_dimensions[FLOW_R - 1].height = 8

    sec(FLOW_R, 2, 10, "🏆  総合活用フロー（5ステップ）"); FLOW_R += 1
    hdr(FLOW_R, [(2,"ステップ"),(3,"内容")]); mg(FLOW_R, 3, FLOW_R, 10); FLOW_R += 1
    for (step, content), color in zip([
        ("① レース性格把握",   "🎲荒れ予報シートで会場×R番号の荒れランクを確認。本命か穴かの方針を決める。"),
        ("② 本命候補の絞込み", "コース別マスタの相対1着率 × 📈フォーム指数で「今好調な実力選手」を特定。"),
        ("③ ST信頼性の確認",   "⚡ST安定度で候補選手のSTばらつきを確認。◎超安定なら評価UP、×不安定なら割引。"),
        ("④ 当日機力の確認",   "展示後に新聞出力 col7「展示偏差値」をチェック。60以上なら機力も裏付けOK。"),
        ("⑤ 総合スコアで確認", "col8「総合スコア」の◎が①〜④の判断と一致していれば◎本命で自信を持って投票。"),
    ], ["1F4E79","2E75B6","375623","7F6000","C55A11"]):
        wc(FLOW_R, 2, step, fill=PatternFill("solid", start_color=color), font=FONT_BADGE, align=AL_C)
        mg(FLOW_R, 3, FLOW_R, 10)
        wc(FLOW_R, 3, content, fill=FILL_WHITE, font=FONT_BODY, align=AL_L)
        ws.row_dimensions[FLOW_R].height = 18; FLOW_R += 1

    # 注意事項
    FLOW_R += 1
    mg(FLOW_R, 2, FLOW_R, 10)
    wc(FLOW_R, 2, "⚠️  共通注意事項", fill=FILL_RED, font=FONT_WARN, align=AL_L)
    ws.row_dimensions[FLOW_R].height = 18; FLOW_R += 1
    for note in [
        "・全指数は過去データに基づく統計値です。データ蓄積量が少ない選手・会場では精度が下がります。",
        "・進入変更・天候・潮位・事故などリアルタイム情報は指数に反映されません。最終判断は現地情報と合わせてください。",
        "・本指数はあくまで予想参考情報です。投票・購入は自己責任でお願いします。",
    ]:
        mg(FLOW_R, 2, FLOW_R, 10)
        wc(FLOW_R, 2, note, fill=PatternFill("solid", start_color="FFF0F0"), font=FONT_SMALL, align=AL_LT)
        ws.row_dimensions[FLOW_R].height = 15; FLOW_R += 1

    ws.freeze_panes = "B6"
    ws.sheet_view.showGridLines = False
    print(f"  📖指数ガイド: シート作成完了")


# -----------------------------------------------------------------------------
# 気象統計マスタ集計・書き込み
# -----------------------------------------------------------------------------
def calc_kisho_stats(ippan: pd.DataFrame):
    """
    会場別×気象条件別（天候/風速/波高/風向）のインコース1着率・決まり手分布を集計。

    Returns
    -------
    df_all : 全条件縦持ち DataFrame（気象統計マスタシート用）
    df_wind_pivot  : 会場×風速ピボット
    df_wave_pivot  : 会場×波高ピボット
    df_dir_pivot   : 会場×風向ピボット
    df_tenkou      : 会場×天候詳細
    period         : (開始日, 終了日) tuple
    """
    latest_date = ippan["日付"].max()
    start_date  = ippan["日付"].min()
    period = (start_date.date(), latest_date.date())
    print(f"  📅 気象統計集計期間: {period[0]} 〜 {period[1]}")

    # レース単位に集約（気象は全艇共通）
    # load_raw_results が dtype=str で読み込むため、風速・波高を数値変換してから集約
    ippan = ippan.copy()
    ippan["風速"] = pd.to_numeric(ippan["風速"], errors="coerce")
    ippan["波高"] = pd.to_numeric(ippan["波高"], errors="coerce")

    race_df = ippan.groupby(
        ["日付", "会場名", "レース番号"]
    ).first().reset_index()[["日付", "会場名", "レース番号", "天候", "風向", "風速", "波高"]]

    win_df = ippan[ippan["着順"] == 1][
        ["日付", "会場名", "レース番号", "決まり手", "進入コース"]
    ].rename(columns={"決まり手": "決まり手_1着", "進入コース": "1着コース"})

    race_df = race_df.merge(win_df, on=["日付", "会場名", "レース番号"], how="left")
    race_df["イン1着"] = (race_df["1着コース"] == 1).astype(int)

    venues = sorted(race_df["会場名"].unique())

    def _stats(df):
        n = len(df)
        if n == 0:
            return None
        return [
            n,
            round(df["イン1着"].mean() * 100, 1),
            round((df["決まり手_1着"] == "逃げ").mean() * 100, 1),
            round((df["決まり手_1着"] == "差し").mean() * 100, 1),
            round((df["決まり手_1着"] == "まくり").mean() * 100, 1),
            round((df["決まり手_1着"] == "まくり差し").mean() * 100, 1),
            round((df["決まり手_1着"] == "抜き").mean() * 100, 1),
            "A" if n >= 50 else ("B" if n >= 20 else "C"),
        ]

    tenkou_order = ["晴", "曇り", "雨", "雪"]
    fukuko_order = ["北", "北東", "東", "南東", "南", "南西", "西", "北西", "無風"]

    # ── 縦持ち全データ ──────────────────────────────────────────────────────
    rows = []
    # 全国平均
    s = _stats(race_df)
    if s:
        rows.append(["【全国平均】", "全体", "全体"] + s)
    for t in tenkou_order:
        s = _stats(race_df[race_df["天候"] == t])
        if s:
            rows.append(["【全国平均】", "天候", t] + s)
    for spd in sorted(race_df["風速"].unique()):
        s = _stats(race_df[race_df["風速"] == spd])
        if s and s[0] >= 10:
            rows.append(["【全国平均】", "風速", f"{int(spd)}m"] + s)
    for wh in sorted(race_df["波高"].unique()):
        s = _stats(race_df[race_df["波高"] == wh])
        if s and s[0] >= 10:
            rows.append(["【全国平均】", "波高", f"{int(wh)}cm"] + s)
    for fk in fukuko_order:
        if fk in race_df["風向"].values:
            s = _stats(race_df[race_df["風向"] == fk])
            if s and s[0] >= 10:
                rows.append(["【全国平均】", "風向", fk] + s)
    # 会場別
    for venue in venues:
        vdf = race_df[race_df["会場名"] == venue]
        s = _stats(vdf)
        if s:
            rows.append([venue, "全体", "全体"] + s)
        for t in tenkou_order:
            d = vdf[vdf["天候"] == t]
            s = _stats(d)
            if s and s[0] >= 5:
                rows.append([venue, "天候", t] + s)
        for spd in sorted(vdf["風速"].unique()):
            d = vdf[vdf["風速"] == spd]
            s = _stats(d)
            if s and s[0] >= 5:
                rows.append([venue, "風速", f"{int(spd)}m"] + s)
        for wh in sorted(vdf["波高"].unique()):
            d = vdf[vdf["波高"] == wh]
            s = _stats(d)
            if s and s[0] >= 5:
                rows.append([venue, "波高", f"{int(wh)}cm"] + s)
        for fk in fukuko_order:
            d = vdf[vdf["風向"] == fk]
            if len(d) == 0:
                continue
            s = _stats(d)
            if s and s[0] >= 5:
                rows.append([venue, "風向", fk] + s)

    COLS = ["会場名", "条件種別", "条件値", "レース数",
            "イン1着率", "逃げ率", "差し率", "まくり率", "まくり差し率", "抜き率", "信頼度"]
    df_all = pd.DataFrame(rows, columns=COLS)

    # ── 風速ピボット ────────────────────────────────────────────────────────
    spd_vals = list(range(0, 11))
    nat_row_w = {"会場名": "【全国平均】", "レース数": len(race_df)}
    for spd in spd_vals:
        d = race_df[race_df["風速"] == spd]
        nat_row_w[f"{spd}m"] = round(d["イン1着"].mean() * 100, 1) if len(d) >= 5 else None
    nat_row_w["全体"] = round(race_df["イン1着"].mean() * 100, 1)

    wind_rows = [nat_row_w]
    for venue in venues:
        vdf = race_df[race_df["会場名"] == venue]
        row = {"会場名": venue, "レース数": len(vdf)}
        for spd in spd_vals:
            d = vdf[vdf["風速"] == spd]
            row[f"{spd}m"] = round(d["イン1着"].mean() * 100, 1) if len(d) >= 5 else None
        row["全体"] = round(vdf["イン1着"].mean() * 100, 1)
        wind_rows.append(row)
    df_wind_pivot = pd.DataFrame(wind_rows)

    # ── 波高ピボット ────────────────────────────────────────────────────────
    wh_vals = list(range(0, 11))
    nat_row_wv = {"会場名": "【全国平均】", "レース数": len(race_df)}
    for wh in wh_vals:
        d = race_df[race_df["波高"] == wh]
        nat_row_wv[f"{wh}cm"] = round(d["イン1着"].mean() * 100, 1) if len(d) >= 5 else None
    nat_row_wv["全体"] = round(race_df["イン1着"].mean() * 100, 1)

    wave_rows = [nat_row_wv]
    for venue in venues:
        vdf = race_df[race_df["会場名"] == venue]
        row = {"会場名": venue, "レース数": len(vdf)}
        for wh in wh_vals:
            d = vdf[vdf["波高"] == wh]
            row[f"{wh}cm"] = round(d["イン1着"].mean() * 100, 1) if len(d) >= 5 else None
        row["全体"] = round(vdf["イン1着"].mean() * 100, 1)
        wave_rows.append(row)
    df_wave_pivot = pd.DataFrame(wave_rows)

    # ── 風向ピボット ────────────────────────────────────────────────────────
    nat_row_d = {"会場名": "【全国平均】", "レース数": len(race_df)}
    for fk in fukuko_order:
        d = race_df[race_df["風向"] == fk]
        nat_row_d[fk] = round(d["イン1着"].mean() * 100, 1) if len(d) >= 5 else None
    nat_row_d["全体"] = round(race_df["イン1着"].mean() * 100, 1)

    dir_rows = [nat_row_d]
    for venue in venues:
        vdf = race_df[race_df["会場名"] == venue]
        row = {"会場名": venue, "レース数": len(vdf)}
        for fk in fukuko_order:
            d = vdf[vdf["風向"] == fk]
            row[fk] = round(d["イン1着"].mean() * 100, 1) if len(d) >= 5 else None
        row["全体"] = round(vdf["イン1着"].mean() * 100, 1)
        dir_rows.append(row)
    df_dir_pivot = pd.DataFrame(dir_rows)

    # ── 天候詳細（会場×天候×決まり手）────────────────────────────────────
    natl_total = round(race_df["イン1着"].mean() * 100, 1)
    natl_t = {}
    for t in tenkou_order:
        d = race_df[race_df["天候"] == t]
        natl_t[t] = round(d["イン1着"].mean() * 100, 1) if len(d) > 0 else None

    tenkou_rows = []
    for venue in venues:
        vdf = race_df[race_df["会場名"] == venue]
        venue_total = round(vdf["イン1着"].mean() * 100, 1)
        # 会場全体行
        tenkou_rows.append({
            "会場名": venue, "天候": "全体", "レース数": len(vdf),
            "イン1着率": venue_total,
            "逃げ率":    round((vdf["決まり手_1着"] == "逃げ").mean() * 100, 1),
            "差し率":    round((vdf["決まり手_1着"] == "差し").mean() * 100, 1),
            "まくり率":  round((vdf["決まり手_1着"] == "まくり").mean() * 100, 1),
            "まくり差し率": round((vdf["決まり手_1着"] == "まくり差し").mean() * 100, 1),
            "抜き率":    round((vdf["決まり手_1着"] == "抜き").mean() * 100, 1),
            "全国比":    round(venue_total - natl_total, 1),
        })
        for t in tenkou_order:
            d = vdf[vdf["天候"] == t]
            if len(d) < 5:
                continue
            in_r = round(d["イン1着"].mean() * 100, 1)
            tenkou_rows.append({
                "会場名": venue, "天候": t, "レース数": len(d),
                "イン1着率": in_r,
                "逃げ率":    round((d["決まり手_1着"] == "逃げ").mean() * 100, 1),
                "差し率":    round((d["決まり手_1着"] == "差し").mean() * 100, 1),
                "まくり率":  round((d["決まり手_1着"] == "まくり").mean() * 100, 1),
                "まくり差し率": round((d["決まり手_1着"] == "まくり差し").mean() * 100, 1),
                "抜き率":    round((d["決まり手_1着"] == "抜き").mean() * 100, 1),
                "全国比":    round(in_r - (natl_t.get(t) or natl_total), 1),
            })
    df_tenkou = pd.DataFrame(tenkou_rows)

    print(f"  気象統計マスタ: 縦持ち{len(df_all)}行 / 風速ピボット{len(df_wind_pivot)}行 / "
          f"波高ピボット{len(df_wave_pivot)}行 / 風向ピボット{len(df_dir_pivot)}行 / "
          f"天候詳細{len(df_tenkou)}行 集計完了")
    return df_all, df_wind_pivot, df_wave_pivot, df_dir_pivot, df_tenkou, period


def write_kisho_stats(wb, df_all, df_wind_pivot, df_wave_pivot, df_dir_pivot, df_tenkou, period):
    """気象統計5シートをExcelに書き込む"""
    from openpyxl.formatting.rule import ColorScaleRule

    period_str = f"集計期間: {period[0]} 〜 {period[1]}"

    H_FILL  = PatternFill("solid", start_color="1F4E79")
    H_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    D_FONT  = Font(name="Arial", size=9)
    AL_C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_L    = Alignment(horizontal="left",   vertical="center")
    FILL_NATL = PatternFill("solid", start_color="FFF2CC")   # 全国平均行：薄黄
    FILL_VTOT = PatternFill("solid", start_color="DAE3F3")   # 会場全体行：薄青
    COLOR_MAP = {
        "天候": "ECEFF8", "風速": "FFFFF0", "波高": "ECFFF0", "風向": "FFF0EC",
    }
    TRUST_COLOR = {"A": "FF00B050", "B": "FFFF9900", "C": "FFFF0000"}

    def _bdr():
        s = Side(border_style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    def _wc(ws, r, c, v, bg="FFFFFFFF", bold=False, fg="FF000000",
            sz=9, fmt=None, left=False, wrap=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.border    = _bdr()
        cell.alignment = Alignment(
            horizontal="left" if left else "center",
            vertical="center", wrap_text=wrap)
        if fmt:
            cell.number_format = fmt
        return cell

    def _title(ws, text, merge_cols):
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=merge_cols)
        cell = ws.cell(row=1, column=1, value=text)
        cell.font      = Font(name="Arial", bold=True, size=10, color="FF1F4E79")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

    def _color_scale(ws, cell_range):
        ws.conditional_formatting.add(cell_range, ColorScaleRule(
            start_type="num", start_value=20, start_color="FFF8696B",
            mid_type="num",   mid_value=55,   mid_color="FFFFEB84",
            end_type="num",   end_value=80,   end_color="FF63BE7B",
        ))

    # ── ① 気象統計マスタ ──────────────────────────────────────────────────
    sname1 = "気象統計マスタ"
    if sname1 in wb.sheetnames:
        del wb[sname1]
    ws1 = wb.create_sheet(sname1)

    _title(ws1, f"■ 気象統計マスタ（会場別×気象条件別）　※インコース1着率・決まり手分布　　{period_str}", 11)

    hdrs = ["会場名", "条件種別", "条件値", "レース数",
            "イン1着率", "逃げ率", "差し率", "まくり率", "まくり差し率", "抜き率", "信頼度"]
    cws1 = [10, 8, 8, 8, 9, 8, 8, 9, 11, 8, 6]
    for ci, (h, w) in enumerate(zip(hdrs, cws1), 1):
        _wc(ws1, 2, ci, h, bg="1F4E79", bold=True, fg="FFFFFFFF", wrap=True)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[2].height = 28

    for ri, row in enumerate(df_all.to_dict(orient="records"), 3):
        venue     = str(row["会場名"])
        cond_type = str(row["条件種別"])
        is_natl   = "全国平均" in venue
        is_total  = cond_type == "全体"
        bg = ("FFF2CC" if is_natl
              else "DAE3F3" if is_total
              else COLOR_MAP.get(cond_type, "FFFFFF"))

        for ci, key in enumerate(["会場名", "条件種別", "条件値", "レース数",
                                   "イン1着率", "逃げ率", "差し率",
                                   "まくり率", "まくり差し率", "抜き率", "信頼度"], 1):
            v = row.get(key, "")
            fmt = "0.0\"%\"" if ci in (5, 6, 7, 8, 9, 10) else ("#,##0" if ci == 4 else None)
            if key == "信頼度":
                c = _wc(ws1, ri, ci, v, bg=bg, bold=True)
                c.font = Font(name="Arial", bold=True,
                              color=TRUST_COLOR.get(str(v), "FF000000"), size=9)
            else:
                _wc(ws1, ri, ci, v, bg=bg, bold=is_total or is_natl,
                    fmt=fmt, left=(ci == 1))

    ws1.freeze_panes = "A3"
    last1 = 2 + len(df_all)
    _color_scale(ws1, f"E3:E{last1}")

    # ── ② 気象ピボット_風速 ───────────────────────────────────────────────
    sname2 = "気象ピボット_風速"
    if sname2 in wb.sheetnames:
        del wb[sname2]
    ws2 = wb.create_sheet(sname2)

    spd_cols = [f"{s}m" for s in range(0, 11)]
    all_cols2 = ["会場名", "レース数"] + spd_cols + ["全体"]
    _title(ws2, f"■ 会場別 風速別インコース1着率（%）　― カラースケール：赤=低 → 緑=高　　{period_str}", len(all_cols2))
    for ci, (h, w) in enumerate(zip(all_cols2, [9, 7] + [6] * 11 + [6]), 1):
        _wc(ws2, 2, ci, h, bg="1F4E79", bold=True, fg="FFFFFFFF", wrap=True)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 24

    for ri, row in enumerate(df_wind_pivot.to_dict(orient="records"), 3):
        is_natl = "全国平均" in str(row.get("会場名", ""))
        bg = "FFF2CC" if is_natl else "FFFFFF"
        for ci, key in enumerate(all_cols2, 1):
            v = row.get(key, None)
            fmt = "0.0\"%\"" if ci > 2 else ("#,##0" if ci == 2 else None)
            _wc(ws2, ri, ci, v, bg=bg, bold=is_natl, fmt=fmt, left=(ci == 1))

    ws2.freeze_panes = "C3"
    last2 = 2 + len(df_wind_pivot)
    _color_scale(ws2, f"C3:{get_column_letter(len(all_cols2))}{last2}")

    # ── ③ 気象ピボット_波高 ───────────────────────────────────────────────
    sname3 = "気象ピボット_波高"
    if sname3 in wb.sheetnames:
        del wb[sname3]
    ws3 = wb.create_sheet(sname3)

    wh_cols = [f"{w}cm" for w in range(0, 11)]
    all_cols3 = ["会場名", "レース数"] + wh_cols + ["全体"]
    _title(ws3, f"■ 会場別 波高別インコース1着率（%）　― カラースケール：赤=低 → 緑=高　　{period_str}", len(all_cols3))
    for ci, (h, w) in enumerate(zip(all_cols3, [9, 7] + [6] * 11 + [6]), 1):
        _wc(ws3, 2, ci, h, bg="1F4E79", bold=True, fg="FFFFFFFF", wrap=True)
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[2].height = 24

    for ri, row in enumerate(df_wave_pivot.to_dict(orient="records"), 3):
        is_natl = "全国平均" in str(row.get("会場名", ""))
        bg = "FFF2CC" if is_natl else "FFFFFF"
        for ci, key in enumerate(all_cols3, 1):
            v = row.get(key, None)
            fmt = "0.0\"%\"" if ci > 2 else ("#,##0" if ci == 2 else None)
            _wc(ws3, ri, ci, v, bg=bg, bold=is_natl, fmt=fmt, left=(ci == 1))

    ws3.freeze_panes = "C3"
    last3 = 2 + len(df_wave_pivot)
    _color_scale(ws3, f"C3:{get_column_letter(len(all_cols3))}{last3}")

    # ── ④ 気象ピボット_風向 ───────────────────────────────────────────────
    sname4 = "気象ピボット_風向"
    if sname4 in wb.sheetnames:
        del wb[sname4]
    ws4 = wb.create_sheet(sname4)

    dir_cols = ["北", "北東", "東", "南東", "南", "南西", "西", "北西", "無風"]
    all_cols4 = ["会場名", "レース数"] + dir_cols + ["全体"]
    _title(ws4, f"■ 会場別 風向別インコース1着率（%）　― カラースケール：赤=低 → 緑=高　　{period_str}", len(all_cols4))
    for ci, (h, w) in enumerate(zip(all_cols4, [9, 7] + [7] * 9 + [6]), 1):
        _wc(ws4, 2, ci, h, bg="1F4E79", bold=True, fg="FFFFFFFF", wrap=True)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[2].height = 24

    for ri, row in enumerate(df_dir_pivot.to_dict(orient="records"), 3):
        is_natl = "全国平均" in str(row.get("会場名", ""))
        bg = "FFF2CC" if is_natl else "FFFFFF"
        for ci, key in enumerate(all_cols4, 1):
            v = row.get(key, None)
            fmt = "0.0\"%\"" if ci > 2 else ("#,##0" if ci == 2 else None)
            _wc(ws4, ri, ci, v, bg=bg, bold=is_natl, fmt=fmt, left=(ci == 1))

    ws4.freeze_panes = "C3"
    last4 = 2 + len(df_dir_pivot)
    _color_scale(ws4, f"C3:{get_column_letter(len(all_cols4))}{last4}")

    # ── ⑤ 気象ピボット_天候 ───────────────────────────────────────────────
    sname5 = "気象ピボット_天候"
    if sname5 in wb.sheetnames:
        del wb[sname5]
    ws5 = wb.create_sheet(sname5)

    _title(ws5, f"■ 会場別×天候別 詳細統計（イン1着率・決まり手分布）　　{period_str}", 10)

    hdrs5 = ["会場名", "天候", "レース数", "イン1着率", "逃げ率", "差し率",
             "まくり率", "まくり差し率", "抜き率", "全国比(±)"]
    cws5  = [9, 6, 7, 9, 8, 8, 9, 11, 8, 9]
    for ci, (h, w) in enumerate(zip(hdrs5, cws5), 1):
        _wc(ws5, 2, ci, h, bg="1F4E79", bold=True, fg="FFFFFFFF", wrap=True)
        ws5.column_dimensions[get_column_letter(ci)].width = w
    ws5.row_dimensions[2].height = 28

    for ri, row in enumerate(df_tenkou.to_dict(orient="records"), 3):
        is_total = str(row.get("天候", "")) == "全体"
        bg = "DAE3F3" if is_total else ("FFFFF0" if row.get("天候") in ("晴", "曇り") else "EFF8FF")
        diff = row.get("全国比", 0) or 0
        diff_col = "FF00B050" if diff > 0 else ("FFFF0000" if diff < 0 else "FF000000")

        keys = ["会場名", "天候", "レース数", "イン1着率", "逃げ率", "差し率",
                "まくり率", "まくり差し率", "抜き率"]
        for ci, key in enumerate(keys, 1):
            v = row.get(key, "")
            fmt = "0.0\"%\"" if ci in (4, 5, 6, 7, 8, 9) else ("#,##0" if ci == 3 else None)
            _wc(ws5, ri, ci, v, bg=bg, bold=is_total, fmt=fmt, left=(ci == 1))

        # 全国比列（±符号つき）
        sign = "+" if diff > 0 else ""
        c = _wc(ws5, ri, 10, f"{sign}{diff}", bg=bg)
        c.font = Font(name="Arial", bold=is_total, color=diff_col, size=9)

    ws5.freeze_panes = "A3"
    last5 = 2 + len(df_tenkou)
    _color_scale(ws5, f"D3:D{last5}")

    print(f"  気象統計5シート 書き込み完了: {sname1}, {sname2}, {sname3}, {sname4}, {sname5}")


# -----------------------------------------------------------------------------
# 風種マスタ集計・書き込み（追い風/向かい風/横風）
# -----------------------------------------------------------------------------

# 各会場のコース1直走路走行方向（スタート→1マーク方向）
COURSE1_DIRECTION = {
    "びわこ": "北",   "三国":   "南西",  "下関":  "北",   "丸亀":  "北西",
    "住之江": "北",   "児島":   "北",    "唐津":  "東",   "多摩川":"東",
    "大村":   "北西", "宮島":   "北東",  "尼崎":  "北",   "常滑":  "北",
    "平和島": "北",   "徳山":   "北東",  "戸田":  "北",   "桐生":  "南",
    "江戸川": "南",   "津":     "東",    "浜名湖":"北",   "福岡":  "北東",
    "芦屋":   "北東", "若松":   "北東",  "蒲郡":  "北",   "鳴門":  "北",
}
_DIR_DEG = {
    "北": 0, "北東": 45, "東": 90, "南東": 135,
    "南": 180, "南西": 225, "西": 270, "北西": 315,
}

def _wind_type(venue_dir: str, wind_dir: str) -> str:
    if wind_dir == "無風":
        return "無風"
    vd = _DIR_DEG.get(venue_dir)
    wd = _DIR_DEG.get(wind_dir)
    if vd is None or wd is None:
        return "不明"
    diff = abs((wd - vd + 360) % 360)
    if diff > 180:
        diff = 360 - diff
    if diff <= 45:
        return "追い風"
    elif diff >= 135:
        return "向かい風"
    return "横風"


def calc_kisho_fuushu(ippan: pd.DataFrame):
    """
    各会場のコース1走行方向を基準に風向を「追い風/向かい風/横風/無風」に分類し、
    インコース1着率・決まり手分布を集計する。

    Returns
    -------
    df_detail  : 会場別×風種別 縦持ち DataFrame
    df_compare : 会場×風種 横並びピボット DataFrame
    period     : (開始日, 終了日) tuple
    """
    latest_date = ippan["日付"].max()
    start_date  = ippan["日付"].min()
    period = (start_date.date(), latest_date.date())
    print(f"  📅 風種マスタ集計期間: {period[0]} 〜 {period[1]}")

    # 数値変換（load_raw_results が dtype=str で読み込むため）
    ippan = ippan.copy()
    ippan["風速"] = pd.to_numeric(ippan["風速"], errors="coerce")
    ippan["波高"] = pd.to_numeric(ippan["波高"], errors="coerce")

    # レース単位集約
    race_df = ippan.groupby(
        ["日付", "会場名", "レース番号"]
    ).first().reset_index()[["日付", "会場名", "レース番号", "天候", "風向", "風速", "波高"]]

    win_df = ippan[ippan["着順"] == 1][
        ["日付", "会場名", "レース番号", "決まり手", "進入コース"]
    ].rename(columns={"決まり手": "決まり手_1着", "進入コース": "1着コース"})

    race_df = race_df.merge(win_df, on=["日付", "会場名", "レース番号"], how="left")
    race_df["イン1着"]  = (race_df["1着コース"] == 1).astype(int)
    race_df["走行方向"] = race_df["会場名"].map(COURSE1_DIRECTION)
    race_df["風種"] = race_df.apply(
        lambda r: _wind_type(r["走行方向"], r["風向"])
        if pd.notna(r["走行方向"]) else "不明", axis=1
    )

    venues      = sorted(race_df["会場名"].unique())
    wind_order  = ["追い風", "向かい風", "横風", "無風"]
    natl_total  = round(race_df["イン1着"].mean() * 100, 1)
    natl_by_type = {}
    for wt in wind_order:
        d = race_df[race_df["風種"] == wt]
        natl_by_type[wt] = round(d["イン1着"].mean() * 100, 1) if len(d) > 0 else None

    def _s(d, col):
        return round((d["決まり手_1着"] == col).mean() * 100, 1) if len(d) > 0 else None

    def _fmt_diff(v):
        return f"+{v}" if v > 0 else str(v)

    # ── 縦持ち詳細 ────────────────────────────────────────────────────────
    rows = []
    for wt in wind_order:
        d = race_df[race_df["風種"] == wt]
        if len(d) < 5: continue
        in_r = round(d["イン1着"].mean() * 100, 1)
        rows.append(["【全国平均】", "-", wt, len(d), in_r,
                     _s(d,"逃げ"), _s(d,"差し"), _s(d,"まくり"), _s(d,"まくり差し"), _s(d,"抜き"),
                     "±0.0", "A"])

    for venue in venues:
        vdf  = race_df[race_df["会場名"] == venue]
        vdir = COURSE1_DIRECTION.get(venue, "-")
        vt   = round(vdf["イン1着"].mean() * 100, 1)
        trust0 = "A" if len(vdf) >= 50 else ("B" if len(vdf) >= 20 else "C")
        rows.append([venue, vdir, "全体", len(vdf), vt,
                     _s(vdf,"逃げ"), _s(vdf,"差し"), _s(vdf,"まくり"), _s(vdf,"まくり差し"), _s(vdf,"抜き"),
                     _fmt_diff(round(vt - natl_total, 1)), trust0])
        for wt in wind_order:
            d = vdf[vdf["風種"] == wt]
            if len(d) < 5: continue
            in_r  = round(d["イン1着"].mean() * 100, 1)
            diff  = round(in_r - (natl_by_type.get(wt) or natl_total), 1)
            trust = "A" if len(d) >= 50 else ("B" if len(d) >= 20 else "C")
            rows.append([venue, vdir, wt, len(d), in_r,
                         _s(d,"逃げ"), _s(d,"差し"), _s(d,"まくり"), _s(d,"まくり差し"), _s(d,"抜き"),
                         _fmt_diff(diff), trust])

    COLS = ["会場名", "走行方向", "風種", "レース数",
            "イン1着率", "逃げ率", "差し率", "まくり率", "まくり差し率", "抜き率",
            "全国比(±)", "信頼度"]
    df_detail = pd.DataFrame(rows, columns=COLS)

    # ── 横並び比較ピボット ────────────────────────────────────────────────
    compare_rows = []
    # 全国平均行
    row = {"会場名": "【全国平均】", "走行方向": "-",
           "全体イン1着率": natl_total}
    for wt in wind_order:
        d = race_df[race_df["風種"] == wt]
        row[f"{wt}_n"]    = len(d) if len(d) >= 5 else None
        row[f"{wt}_in"]   = round(d["イン1着"].mean() * 100, 1) if len(d) >= 5 else None
        row[f"{wt}_nige"] = _s(d, "逃げ") if len(d) >= 5 else None
        if wt != "無風":
            row[f"{wt}_maku"] = _s(d, "まくり") if len(d) >= 5 else None
    compare_rows.append(row)

    for venue in venues:
        vdf  = race_df[race_df["会場名"] == venue]
        vdir = COURSE1_DIRECTION.get(venue, "-")
        row  = {"会場名": venue, "走行方向": vdir,
                "全体イン1着率": round(vdf["イン1着"].mean() * 100, 1)}
        for wt in wind_order:
            d = vdf[vdf["風種"] == wt]
            row[f"{wt}_n"]    = len(d) if len(d) >= 5 else None
            row[f"{wt}_in"]   = round(d["イン1着"].mean() * 100, 1) if len(d) >= 5 else None
            row[f"{wt}_nige"] = _s(d, "逃げ") if len(d) >= 5 else None
            if wt != "無風":
                row[f"{wt}_maku"] = _s(d, "まくり") if len(d) >= 5 else None
        compare_rows.append(row)

    df_compare = pd.DataFrame(compare_rows)

    print(f"  風種マスタ: 縦持ち{len(df_detail)}行 / 比較ピボット{len(df_compare)}行 集計完了")
    return df_detail, df_compare, period


def write_kisho_fuushu(wb, df_detail, df_compare, period):
    """風種統計2シートをExcelに書き込む"""
    from openpyxl.formatting.rule import ColorScaleRule

    period_str = f"集計期間: {period[0]} 〜 {period[1]}"
    wind_order = ["追い風", "向かい風", "横風", "無風"]

    WIND_COLOR = {
        "追い風":   "DEEAF1",
        "向かい風": "FCE4D6",
        "横風":     "E2EFDA",
        "無風":     "F2F2F2",
    }
    WIND_HDR_COLOR = {
        "追い風":   "2E75B6",
        "向かい風": "C55A11",
        "横風":     "375623",
        "無風":     "595959",
    }

    def _bdr():
        s = Side(border_style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    def _wc(ws, r, c, v, bg="FFFFFF", bold=False, fg="000000",
            sz=9, fmt=None, left=False, wrap=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(name="メイリオ", bold=bold, color=fg, size=sz)
        cell.fill      = PatternFill("solid", start_color=bg, end_color=bg)
        cell.border    = _bdr()
        cell.alignment = Alignment(
            horizontal="left" if left else "center",
            vertical="center", wrap_text=wrap)
        if fmt:
            cell.number_format = fmt
        return cell

    def _color_scale(ws, cell_range):
        ws.conditional_formatting.add(cell_range, ColorScaleRule(
            start_type="num", start_value=20, start_color="FFF8696B",
            mid_type="num",   mid_value=55,   mid_color="FFFFEB84",
            end_type="num",   end_value=80,   end_color="FF63BE7B",
        ))

    TRUST_FG = {"A": "00B050", "B": "FF9900", "C": "FF0000"}

    # ── ① 気象ピボット_風種（縦持ち詳細）────────────────────────────────
    sname1 = "気象ピボット_風種"
    if sname1 in wb.sheetnames:
        del wb[sname1]
    ws1 = wb.create_sheet(sname1)

    ws1.merge_cells("A1:L1")
    ws1["A1"].value = (f"■ 会場別 風種別インコース1着率・決まり手分布"
                       f"（追い風/向かい風/横風）　　{period_str}")
    ws1["A1"].font      = Font(name="メイリオ", bold=True, size=10, color="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 22

    ws1.merge_cells("A2:L2")
    ws1["A2"].value = ("※ 追い風=コース1直走路と同方向　向かい風=逆方向　横風=直交方向"
                       "（各会場のコース1走行方向を基準に分類）")
    ws1["A2"].font      = Font(name="メイリオ", size=8, color="595959", italic=True)
    ws1["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[2].height = 16

    hdrs1 = ["会場名", "走行方向", "風種", "レース数", "イン1着率", "逃げ率", "差し率",
             "まくり率", "まくり差し率", "抜き率", "全国比(±)", "信頼度"]
    cws1  = [9, 7, 8, 7, 9, 8, 8, 9, 11, 8, 9, 6]
    for ci, (h, w) in enumerate(zip(hdrs1, cws1), 1):
        _wc(ws1, 3, ci, h, bg="1F4E79", bold=True, fg="FFFFFF", wrap=True)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[3].height = 28

    ri = 4
    for row in df_detail.to_dict(orient="records"):
        venue     = str(row["会場名"])
        wind_type = str(row["風種"])
        is_natl   = "全国平均" in venue
        is_total  = wind_type == "全体"
        bg = ("FFF2CC" if is_natl
              else "DAE3F3" if is_total
              else WIND_COLOR.get(wind_type, "FFFFFF"))

        _wc(ws1, ri, 1,  row["会場名"],       bg=bg, bold=is_total or is_natl, left=True)
        _wc(ws1, ri, 2,  row["走行方向"],      bg=bg, bold=is_total or is_natl)
        _wc(ws1, ri, 3,  row["風種"],          bg=bg, bold=is_total or is_natl)
        _wc(ws1, ri, 4,  row["レース数"],      bg=bg, bold=is_total or is_natl, fmt="#,##0")
        _wc(ws1, ri, 5,  row["イン1着率"],     bg=bg, bold=is_total or is_natl, fmt='0.0"%"')
        _wc(ws1, ri, 6,  row["逃げ率"],        bg=bg, bold=is_total or is_natl, fmt='0.0"%"')
        _wc(ws1, ri, 7,  row["差し率"],        bg=bg, bold=is_total or is_natl, fmt='0.0"%"')
        _wc(ws1, ri, 8,  row["まくり率"],      bg=bg, bold=is_total or is_natl, fmt='0.0"%"')
        _wc(ws1, ri, 9,  row["まくり差し率"],  bg=bg, bold=is_total or is_natl, fmt='0.0"%"')
        _wc(ws1, ri, 10, row["抜き率"],        bg=bg, bold=is_total or is_natl, fmt='0.0"%"')

        diff_val = str(row.get("全国比(±)", ""))
        diff_num = float(diff_val.replace("+", "")) if diff_val not in ("", "±0.0") else 0.0
        diff_col = "00B050" if diff_num > 0 else ("FF0000" if diff_num < 0 else "000000")
        c = _wc(ws1, ri, 11, diff_val, bg=bg, bold=is_total or is_natl)
        c.font = Font(name="メイリオ", bold=is_total or is_natl, color=diff_col, size=9)

        trust = str(row.get("信頼度", ""))
        c2 = _wc(ws1, ri, 12, trust, bg=bg, bold=True)
        c2.font = Font(name="メイリオ", bold=True,
                       color=TRUST_FG.get(trust, "000000"), size=9)
        ri += 1

    ws1.freeze_panes = "A4"
    _color_scale(ws1, f"E4:E{ri}")

    # ── ② 気象ピボット_風種比較（横並びピボット）──────────────────────────
    sname2 = "気象ピボット_風種比較"
    if sname2 in wb.sheetnames:
        del wb[sname2]
    ws2 = wb.create_sheet(sname2)

    ws2.merge_cells("A1:R1")
    ws2["A1"].value = (f"■ 会場別 追い風/向かい風/横風/無風 インコース1着率 横並び比較"
                       f"　　{period_str}")
    ws2["A1"].font      = Font(name="メイリオ", bold=True, size=10, color="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 22

    # 走行方向一覧の注釈行
    ws2.merge_cells("A2:R2")
    dir_note = "　".join([f"{v}:{d}方向" for v, d in sorted(COURSE1_DIRECTION.items())])
    ws2["A2"].value     = "走行方向一覧: " + dir_note
    ws2["A2"].font      = Font(name="メイリオ", size=7, color="595959", italic=True)
    ws2["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws2.row_dimensions[2].height = 28

    # 行3: 風種セクションヘッダー
    ws2.row_dimensions[3].height = 16
    for ci in range(1, 19):
        cell = ws2.cell(3, ci, "")
        col_wt = ({4:"追い風",5:"追い風",6:"追い風",7:"追い風",
                   8:"向かい風",9:"向かい風",10:"向かい風",11:"向かい風",
                   12:"横風",13:"横風",14:"横風",15:"横風",
                   16:"無風",17:"無風"}).get(ci)
        bg = WIND_HDR_COLOR.get(col_wt, "1F4E79") if col_wt else "1F4E79"
        cell.fill      = PatternFill("solid", start_color=bg, end_color=bg)
        cell.font      = Font(name="メイリオ", bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _bdr()
    for s, e, label in [(4,7,"← 追い風 →"),(8,11,"← 向かい風 →"),
                        (12,15,"← 横風 →"),(16,17,"無風")]:
        ws2.merge_cells(start_row=3, start_column=s, end_row=3, end_column=e)
        ws2.cell(3, s).value = label

    # 行4: 列ヘッダー
    hdrs2 = ["会場名", "走行方向", "全体\nイン1着率",
             "n", "イン\n1着率", "逃げ%", "まくり%",
             "n", "イン\n1着率", "逃げ%", "まくり%",
             "n", "イン\n1着率", "逃げ%", "まくり%",
             "n", "イン\n1着率"]
    cws2  = [9, 7, 9, 6, 9, 8, 9, 6, 9, 8, 9, 6, 9, 8, 9, 6, 9]
    ws2.row_dimensions[4].height = 28
    for ci, (h, w) in enumerate(zip(hdrs2, cws2), 1):
        _wc(ws2, 4, ci, h, bg="1F4E79", bold=True, fg="FFFFFF", wrap=True)
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ri2 = 5
    for row in df_compare.to_dict(orient="records"):
        is_natl = "全国平均" in str(row.get("会場名", ""))
        bg_base = "FFF2CC" if is_natl else "FFFFFF"
        _wc(ws2, ri2, 1, row["会場名"],        bg=bg_base, bold=is_natl, left=True)
        _wc(ws2, ri2, 2, row["走行方向"],       bg=bg_base, bold=is_natl)
        _wc(ws2, ri2, 3, row["全体イン1着率"],  bg=bg_base, bold=is_natl, fmt='0.0"%"')
        ci_off = 4
        for wt in wind_order:
            bg = "FFF2CC" if is_natl else WIND_COLOR.get(wt, "F2F2F2")
            n_val    = row.get(f"{wt}_n")
            in_val   = row.get(f"{wt}_in")
            nige_val = row.get(f"{wt}_nige")
            maku_val = row.get(f"{wt}_maku") if wt != "無風" else None
            has_data = n_val is not None
            empty_bg = "F0F0F0"

            _wc(ws2, ri2, ci_off,   n_val    if has_data else None,
                bg=bg if has_data else empty_bg, bold=is_natl, fmt="#,##0")
            _wc(ws2, ri2, ci_off+1, in_val   if has_data else None,
                bg=bg if has_data else empty_bg, bold=is_natl, fmt='0.0"%"')
            _wc(ws2, ri2, ci_off+2, nige_val if has_data else None,
                bg=bg if has_data else empty_bg, bold=is_natl, fmt='0.0"%"')
            if wt != "無風":
                _wc(ws2, ri2, ci_off+3, maku_val if has_data else None,
                    bg=bg if has_data else empty_bg, bold=is_natl, fmt='0.0"%"')
            ci_off += 4 if wt != "無風" else 2
        ri2 += 1

    ws2.freeze_panes = "A5"
    for col_letter in ["E", "I", "M", "Q"]:
        _color_scale(ws2, f"{col_letter}5:{col_letter}{ri2}")

    print(f"  風種統計2シート 書き込み完了: {sname1}, {sname2}")


# -----------------------------------------------------------------------------
# 【⑤追加】会場別コース距離補正値集計
# -----------------------------------------------------------------------------
def calc_venue_course_adj(ippan: pd.DataFrame) -> pd.DataFrame:
    """
    【⑤改善】_predict_first_turn の固定コース距離補正を会場別実測値に置き換えるための
    集計関数。

    【設計思想】
    現状の _predict_first_turn は全会場共通の固定補正値を使用している。
        2C: +0.03秒 / 3C: +0.06秒 / 4C: +0.10秒 / 5C: +0.15秒 / 6C: +0.21秒
    しかし鳴門・江戸川・戸田のような特殊なコース形状の会場では
    この固定値が実態と大きく乖離し、1M到達順予測の精度を下げる。

    【計算方法】
    同一レース内で「1着艇がどのコースから出たか」の分布を使って
    各コースの「実際の1M先着有利度」を全国平均との差として算出する。

    具体的には:
        全国平均1着率（COURSE_NATIONAL_WIN）をベースラインとし、
        会場別コース1着率との差から「相対的な距離有利/不利」を秒換算する。
        1C差し = 会場1着率差 × 換算係数(0.10秒/0.10差)

    出力列:
        会場名, 2C補正, 3C補正, 4C補正, 5C補正, 6C補正, レース数, 信頼度

    信頼度:
        レース数 >= 500 → 1.0 / >= 200 → 0.6 / >= 100 → 0.3 / 未満 → 0.0（全国固定値使用）

    Returns
    -------
    df : pd.DataFrame  会場別コース距離補正値テーブル
    """
    COURSE_NATIONAL_WIN = {1: 0.555, 2: 0.137, 3: 0.134, 4: 0.111, 5: 0.066, 6: 0.021}
    # 1コース1着率の差を秒換算する係数（経験則：差0.10 ≒ 0.05秒の距離差）
    SCALE_TO_SEC = 0.50  # 1着率差1.0 → 0.50秒補正
    # 固定値（全国デフォルト）
    FIXED_ADJ = {1: 0.00, 2: 0.03, 3: 0.06, 4: 0.10, 5: 0.15, 6: 0.21}

    latest_date = ippan["日付"].max()
    cutoff_date = latest_date - pd.Timedelta(days=365)
    df_1y = ippan[ippan["日付"] >= cutoff_date].copy()

    venues = sorted(df_1y["会場名"].dropna().unique())

    # 1着行のみ抽出
    ichi = df_1y[df_1y["着順"] == 1][["会場名", "進入コース"]].copy()

    rows = []
    for venue in venues:
        vdf = ichi[ichi["会場名"] == venue]
        n_race = len(vdf)

        # 信頼度
        if n_race >= 500:
            trust = 1.0
        elif n_race >= 200:
            trust = 0.6
        elif n_race >= 100:
            trust = 0.3
        else:
            trust = 0.0

        row = {"会場名": venue, "レース数": n_race, "信頼度": round(trust, 2)}

        for c in range(2, 7):
            if n_race < 100:
                # サンプル不足 → 固定値をそのまま使用
                row[f"{c}C補正"] = FIXED_ADJ[c]
            else:
                # 会場の実際のコース別1着率を計算
                venue_c_rate = len(vdf[vdf["進入コース"] == c]) / n_race
                nat_c_rate   = COURSE_NATIONAL_WIN[c]
                # 全国比から実効距離差を推定
                rate_diff    = venue_c_rate - nat_c_rate
                # 固定値を基準に差分を加減算（外コースほど距離差が大きいため係数を調整）
                course_scale = SCALE_TO_SEC * (c / 3.5)  # コース番号が大きいほど影響大
                adj = FIXED_ADJ[c] - rate_diff * course_scale
                # 物理的な下限・上限（逆転しないよう制約）
                adj = max(FIXED_ADJ[c] * 0.4, min(FIXED_ADJ[c] * 1.8, adj))
                row[f"{c}C補正"] = round(adj, 4)

        rows.append(row)

    df = pd.DataFrame(rows)
    cols = ["会場名", "レース数", "信頼度"] + [f"{c}C補正" for c in range(2, 7)]
    df = df[cols].sort_values("会場名").reset_index(drop=True)

    # CSVに保存
    os.makedirs(os.path.join(BASE_DIR, "data"), exist_ok=True)
    df.to_csv(VENUE_COURSE_ADJ_CSV, index=False, encoding="utf-8-sig")
    print(f"  💾 会場別コース距離補正CSV: {os.path.basename(VENUE_COURSE_ADJ_CSV)} ({len(df)}会場)")

    return df


# -----------------------------------------------------------------------------
# メイン
# -----------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("  ボートリサーチ新聞 マスタ更新スクリプト")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excelファイルが見つかりません: {EXCEL_PATH}")
        return

    # 1. データ読み込み
    print("\n[1/5] CSVデータ読み込み中...")
    raw, ippan = load_raw_results()

    # ================================================================
    # 詳細データ診断
    # ================================================================
    _EXCL_DEBUG = {"準優勝戦", "優勝戦", "順位決定戦", "Ｓ戦優勝戦", "賞金女王決定"}
    print("\n" + "=" * 60)
    print("  📊 データ診断レポート")
    print("=" * 60)

    # ① 基本件数
    print(f"\n【① 基本件数】")
    print(f"  RAW 総行数  : {len(raw):,} 行")
    print(f"  一般戦 行数 : {len(ippan):,} 行")
    if len(raw) > 0:
        _ratio = len(ippan) / len(raw) * 100
        print(f"  一般戦 割合 : {_ratio:.1f}%")
        if _ratio > 90:
            print("  ⚠ 90%超 → grade_master.csv 未適用の可能性大！")
        elif _ratio < 30:
            print("  ⚠ 30%未満 → 除外条件が厳しすぎる可能性")

    # ② グレード分布
    print(f"\n【② グレード分布】")
    if "グレード" in raw.columns:
        for _g, _cnt in raw["グレード"].value_counts().items():
            _mark = "✅ 残す" if _g == "一般" else "🚫 除外"
            print(f"  {_mark} | {str(_g):<12}: {_cnt:,} 行")
        if set(raw["グレード"].unique()) == {"一般"}:
            print("\n  ⚠⚠⚠ 全レースが「一般」！grade_master.csv が未適用です。")
            print("       make_grade_master_auto.py を実行してください。")

    # ③ 日付範囲・月別件数
    print(f"\n【③ 日付範囲と月別件数】")
    if "日付" in raw.columns:
        _vd = raw["日付"].dropna()
        if len(_vd) > 0:
            print(f"  最古: {_vd.min().date()}  最新: {_vd.max().date()}")
            print(f"  NaN : {raw['日付'].isna().sum():,} 行")
            raw["_月"] = raw["日付"].dt.to_period("M")
            for _p, _cnt in raw.groupby("_月").size().items():
                _ic = len(ippan[ippan["日付"].dt.to_period("M") == _p]) if len(ippan) > 0 else 0
                _bar = "█" * min(_cnt // 500, 20)
                _warn = " ⚠少" if 0 < _ic < 3000 else ""
                print(f"    {_p}: RAW {_cnt:5,}行  一般戦 {_ic:5,}行{_warn} {_bar}")
            raw.drop(columns=["_月"], inplace=True, errors="ignore")

    # ④ 進入コース
    print(f"\n【④ 進入コース分布】")
    _nan_c = raw["進入コース"].isna().sum()
    print(f"  NaN: {_nan_c:,} / {len(raw):,}")
    print(f"  分布: {dict(raw['進入コース'].value_counts().sort_index())}")

    # ⑤ レース種別
    print(f"\n【⑤ レース種別 TOP30 (🚫=除外対象)】")
    if "レース種別" in raw.columns:
        for _rc, _cnt in raw["レース種別"].value_counts().head(30).items():
            _m = "🚫" if _rc in _EXCL_DEBUG else "  "
            print(f"  {_m} {str(_rc):<22}: {_cnt:,}")

    # ⑥ 選手別出走数
    print(f"\n【⑥ 選手別出走数サマリー（一般戦）】")
    if len(ippan) > 0 and "選手名" in ippan.columns:
        _sc = ippan.groupby("選手名").size().describe()
        print(f"  選手数: {ippan['選手名'].nunique():,}人  "
              f"平均: {_sc['mean']:.1f}走  中央値: {_sc['50%']:.0f}走  "
              f"最大: {_sc['max']:.0f}走  最小: {_sc['min']:.0f}走")
        _low = (ippan.groupby("選手名").size() < 10).sum()
        if _low:
            print(f"  ⚠ 10走未満の選手: {_low}人（★フラグ多数立つ可能性）")

    print("\n" + "=" * 60 + "\n")

    # 2. 集計
    print("\n[2/5] コース別マスタ集計中...")
    df_course = calc_course_master(ippan)

    print("\n[3/5] 選手指数マスタ集計中...")
    df_senshu = calc_senshu_master(raw, ippan)

    print("\n[4/5] 会場統計・イン逃げ分析・会場別コースマスタ・展開別残存マスタ集計中...")
    df_kaijo        = calc_kaijo_stats(ippan)
    df_in           = calc_in_nige_analysis(ippan)
    df_venue_course = calc_venue_course_master(ippan)
    df_tenkai_venue, df_tenkai_national = calc_tenkai_survival_master(ippan)

    # 【⑤追加】会場別コース距離補正値集計（load_race.py の _predict_first_turn で使用）
    print("\n[⑤] 会場別コース距離補正値集計中...")
    calc_venue_course_adj(ippan)

    # 5. Excel書き込み
    print(f"\n[5/5] Excel書き込み中: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    write_course_master(wb, df_course, len(raw), len(ippan), len(df_senshu))
    write_senshu_master(wb, df_senshu)
    write_kaijo_stats(wb, df_kaijo)
    write_in_nige(wb, df_in)
    write_venue_course_master(wb, df_venue_course)
    write_tenkai_survival(wb, df_tenkai_venue, df_tenkai_national)
    write_guide(wb)

    print("\n[気象統計] 集計・書き込み中...")
    (df_kisho_all, df_wind_pivot, df_wave_pivot,
     df_dir_pivot, df_tenkou, kisho_period) = calc_kisho_stats(ippan)
    write_kisho_stats(wb, df_kisho_all, df_wind_pivot, df_wave_pivot,
                      df_dir_pivot, df_tenkou, kisho_period)

    print("\n[風種統計] 集計・書き込み中...")
    df_fuushu_detail, df_fuushu_compare, fuushu_period = calc_kisho_fuushu(ippan)
    write_kisho_fuushu(wb, df_fuushu_detail, df_fuushu_compare, fuushu_period)

    # シート順序を整える
    desired_order = [
        "📊コース別マスタ", "選手指数マスタ", "会場別コースマスタ",
        "会場統計", "イン逃げ分析", "展開別残存_会場別", "展開別残存_全国",
        "気象統計マスタ", "気象ピボット_風速", "気象ピボット_波高",
        "気象ピボット_風向", "気象ピボット_天候",
        "気象ピボット_風種", "気象ピボット_風種比較",
        "選手別当日指数", "📋本日入力", "🗞️新聞出力", "📖指数ガイド", "📖使い方ガイド"
    ]
    existing = [s for s in desired_order if s in wb.sheetnames]
    others   = [s for s in wb.sheetnames if s not in desired_order]
    for i, name in enumerate(existing + others):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    try:
        wb.save(EXCEL_PATH)
    except PermissionError:
        print()
        print("❌ 保存エラー: Excelファイルが開いています。")
        print(f"   👉 '{os.path.basename(EXCEL_PATH)}' を閉じてから、もう一度実行してください。")
        print("=" * 60)
        return

    size = os.path.getsize(EXCEL_PATH) / 1024 / 1024
    print(f"\n✅ 完了！  ファイルサイズ: {size:.1f} MB")
    print(f"   保存先: {EXCEL_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()