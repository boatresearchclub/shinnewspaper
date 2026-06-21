"""
apply_ev.py  ─  展示後 実オッズ照合・EV計算・参戦判断 更新スクリプト
=======================================================================
使い方:
    python apply_ev.py --venue 大村 --date 2026-03-09
    python apply_ev.py --venue 大村 --date 2026-03-09 --race 5   # 特定レースのみ

処理フロー:
    1. logs/YYYY-MM-DD_会場名.json（load_race.py が生成）を読み込む
    2. boatrace.jp から各レースの実オッズをスクレイピング
    3. combos_full × 実オッズ → EV計算
    4. 参戦判断（buy_list更新）を logs/ に書き戻す
    5. ボートリサーチ新聞_軽量版.xlsx の「会場名_数値」シートの
       「★参戦判断（EV）」行を上書き更新

毎日実行.bat との関係:
    load_race.py  → 理論オッズで買い目確定（展示前）
    apply_ev.py   → 展示後に手動 or bat で実行してEV反映・参戦可否を確定
"""

import argparse
import json
import pathlib
import sys
import time
import re

# ── 同ディレクトリの load_race.py から定数・関数を再利用 ──────────────────
sys.path.insert(0, str(pathlib.Path(__file__).parent))

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

try:
    import requests
    from bs4 import BeautifulSoup
    SCRAPE_AVAILABLE = True
except ImportError:
    SCRAPE_AVAILABLE = False
    print("⚠️  requests/beautifulsoup4 がなければスクレイピング不可（pip install requests beautifulsoup4）")

# ── パス定数 ──────────────────────────────────────────────────────────────
BASE_DIR   = pathlib.Path(__file__).parent.parent
EXCEL_FILE = BASE_DIR / "ボートリサーチ新聞_軽量版.xlsx"
LOGS_DIR   = BASE_DIR / "logs"

# 会場コードマップ（load_race.py と同一）
VENUE_JCD_MAP = {
    "桐生":   "01", "戸田":   "02", "江戸川": "03", "平和島": "04",
    "多摩川": "05", "浜名湖": "06", "蒲郡":   "07", "常滑":   "08",
    "津":     "09", "三国":   "10", "びわこ": "11", "住之江": "12",
    "尼崎":   "13", "鳴門":   "14", "丸亀":   "15", "児島":   "16",
    "宮島":   "17", "徳山":   "18", "下関":   "19", "若松":   "20",
    "芦屋":   "21", "福岡":   "22", "唐津":   "23", "大村":   "24",
}

# ── EV計算（load_race.py と同一ロジック） ────────────────────────────────
def calc_ev(combos_full: list[dict], actual_odds: dict) -> list[dict]:
    """combos_full × 実オッズ → EV付きリスト（EV降順）"""
    result = []
    for c in combos_full:
        key    = c["combo"]
        actual = actual_odds.get(key)
        c2     = dict(c)
        if actual is not None and c.get("prob", 0) > 0:
            ev = actual * c["prob"] - 1.0
            c2["actual_odds"] = round(actual, 1)
            c2["ev"]          = round(ev, 4)
            c2["ev_pct"]      = f"{'+' if ev >= 0 else ''}{ev*100:.1f}%"
            c2["ev_positive"] = ev > 0
        else:
            c2["actual_odds"] = None
            c2["ev"]          = None
            c2["ev_pct"]      = "N/A"
            c2["ev_positive"] = False
        result.append(c2)
    result.sort(key=lambda x: (x["ev"] if x["ev"] is not None else -999), reverse=True)
    return result


def suggest_by_ev(combos_with_ev: list[dict], min_ev: float = 0.0, max_bets: int = 8) -> dict:
    """EV+の組み合わせから参戦判断を生成"""
    positives = [c for c in combos_with_ev
                 if c.get("ev") is not None and c["ev"] > min_ev][:max_bets]

    if not positives:
        best = next((c for c in combos_with_ev if c.get("ev") is not None), None)
        best_str = f"（最高EV: {best['ev_pct']} {best['combo']}）" if best else ""
        return {
            "buy_list":   [],
            "ev_summary": [],
            "total_bets": 0,
            "best_ev":    None,
            "skip":       True,
            "reason":     f"EV>{min_ev*100:.0f}%の組み合わせなし → 見送り推奨{best_str}",
        }

    return {
        "buy_list": [c["combo"] for c in positives],
        "ev_summary": [
            {
                "combo":       c["combo"],
                "prob_pct":    f"{c['prob']*100:.2f}%",
                "actual_odds": c["actual_odds"],
                "ev_pct":      c["ev_pct"],
            }
            for c in positives
        ],
        "total_bets": len(positives),
        "best_ev":    positives[0],
        "skip":       False,
        "reason":     f"EV>{min_ev*100:.0f}%が{len(positives)}点（最高: {positives[0]['ev_pct']} {positives[0]['combo']}）",
    }


# ── 実オッズ取得 ─────────────────────────────────────────────────────────
def build_odds_url(venue: str, race_no, race_date: str) -> str | None:
    jcd = VENUE_JCD_MAP.get(str(venue).strip())
    if not jcd:
        return None
    rno = int(race_no) if str(race_no).isdigit() else race_no
    hd  = str(race_date).replace("-", "").replace("/", "")
    if len(hd) != 8 or not hd.isdigit():
        return None
    return f"https://www.boatrace.jp/owpc/pc/race/odds3t?rno={rno}&jcd={jcd}&hd={hd}"


def fetch_actual_odds(url: str, retries: int = 3) -> dict | None:
    """
    boatrace.jp の3連単オッズページをスクレイピングして
    {"1-2-3": 倍率float, ...} 形式で返す。
    """
    if not SCRAPE_AVAILABLE:
        return None

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    }

    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=headers, timeout=15)
            resp.raise_for_status()
            break
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
                continue
            print(f"  ❌ オッズ取得失敗: {e}")
            return None

    soup = BeautifulSoup(resp.content, "html.parser")
    odds_dict = {}

    # boatrace.jp の3連単テーブル構造を解析
    # テーブルは .table1 クラスの <table> に格納
    tables = soup.select("table.is-p3-0")
    if not tables:
        tables = soup.select("table")

    for table in tables:
        rows = table.find_all("tr")
        for tr in rows:
            cells = tr.find_all("td")
            if len(cells) < 2:
                continue
            # セルテキストから艇番・オッズを抽出
            for i, td in enumerate(cells):
                text = td.get_text(strip=True)
                # "1-2-3" 形式のコンボを探す
                m = re.match(r"^(\d)-(\d)-(\d)$", text)
                if m:
                    combo = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
                    # 次のセルがオッズ
                    if i + 1 < len(cells):
                        odds_text = cells[i + 1].get_text(strip=True).replace(",", "")
                        try:
                            odds_dict[combo] = float(odds_text)
                        except ValueError:
                            pass  # "欠場" などはスキップ

    if odds_dict:
        print(f"    ✅ 実オッズ取得: {len(odds_dict)}通り")
        return odds_dict

    # フォールバック: より汎用的なパース（数値パターン）
    all_text = soup.get_text()
    pattern  = re.compile(r"(\d)-(\d)-(\d)\s+([\d,]+\.?\d*)")
    for m in pattern.finditer(all_text):
        combo = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        try:
            odds_dict[combo] = float(m.group(4).replace(",", ""))
        except ValueError:
            pass

    if odds_dict:
        print(f"    ✅ 実オッズ取得（フォールバック）: {len(odds_dict)}通り")
        return odds_dict

    print("    ⚠️  実オッズを解析できませんでした")
    return None


# ── Excel 参戦判断行の更新 ────────────────────────────────────────────────
def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _make_font(bold=False, size=9, color="FF000000") -> Font:
    return Font(bold=bold, size=size, color=color)

def _make_align(wrap=True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _thin_border() -> Border:
    s = Side(style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)


def update_excel_ev_row(excel_path: str, sheet_name: str,
                        race_results: dict) -> bool:
    """
    「★参戦判断（EV）」行を見つけて各レース列のセルを上書きする。

    Parameters
    ----------
    excel_path   : str   Excelファイルパス
    sheet_name   : str   対象シート名（例: "大村_数値"）
    race_results : dict  {race_no_str: ev_suggestion_dict, ...}
    """
    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        print(f"  ❌ Excel を開けませんでした: {e}")
        return False

    if sheet_name not in wb.sheetnames:
        print(f"  ❌ シート '{sheet_name}' が見つかりません")
        wb.close()
        return False

    ws = wb[sheet_name]

    # ── ヘッダ行（「1R」「2R」... が書いてある行）からレース番号→列マップを作成 ──
    race_col_map = {}   # {"1": 列番号int, "2": 列番号int, ...}
    header_row   = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and re.match(r"^\d+R$", v.strip()):
                race_col_map[v.strip().replace("R", "")] = c
                header_row = r
        if race_col_map:
            break

    if not race_col_map:
        print("  ❌ レース番号ヘッダが見つかりませんでした")
        wb.close()
        return False

    print(f"  📍 ヘッダ行: {header_row}行目 / 検出レース: {sorted(race_col_map.keys(), key=int)}")

    # ── 「★参戦判断（EV）」行を検索 ──
    ev_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value  # B列に「★参戦判断（EV）」が入っている
        if v and "参戦判断" in str(v):
            ev_row = r
            break

    if ev_row is None:
        print("  ❌ 「★参戦判断（EV）」行が見つかりません")
        wb.close()
        return False

    print(f"  📍 参戦判断行: {ev_row}行目")

    # ── 各レースのEV結果をセルに書き込み ──
    FILL_GREEN  = _make_fill("FF00B050")  # 参戦OK（緑）
    FILL_RED    = _make_fill("FFFF0000")  # 見送り（赤）
    FONT_WHITE  = _make_font(bold=True, size=9, color="FFFFFFFF")
    FONT_BLACK  = _make_font(bold=False, size=8, color="FF000000")
    bdr         = _thin_border()
    align       = _make_align(wrap=True)

    updated = 0
    for race_no_str, ev_sug in race_results.items():
        col = race_col_map.get(str(race_no_str))
        if col is None:
            print(f"  ⚠️  {race_no_str}R の列が見つかりません（スキップ）")
            continue

        skip = ev_sug.get("skip", True)
        if skip:
            label = f"⛔ 見送り\n{ev_sug.get('reason', '')}"
            fill  = FILL_RED
            font  = FONT_WHITE
        else:
            bl    = ev_sug.get("buy_list", [])
            bets  = ev_sug.get("total_bets", 0)
            best  = ev_sug.get("best_ev") or {}
            label = (
                f"✅ 参戦 {bets}点\n"
                f"最高EV: {best.get('ev_pct', '-')} {best.get('combo', '')}\n"
                + "\n".join(bl[:8])
                + (f"\n…他{bets-8}点" if bets > 8 else "")
            )
            fill  = FILL_GREEN
            font  = FONT_WHITE

        cell            = ws.cell(ev_row, col)
        cell.value      = label
        cell.fill       = fill
        cell.font       = font
        cell.alignment  = align
        cell.border     = bdr
        updated += 1

    ws.row_dimensions[ev_row].height = 60.0

    try:
        wb.save(excel_path)
        print(f"  💾 Excel 保存完了: {excel_path} （{updated}レース更新）")
        return True
    except Exception as e:
        print(f"  ❌ Excel 保存失敗: {e}")
        return False


# ── ログ更新 ─────────────────────────────────────────────────────────────
def update_log_ev(log_path: pathlib.Path, race_no: str, ev_suggestion: dict,
                  actual_odds: dict):
    """予想ログ JSON の該当レースに ev_suggestion を書き戻す"""
    try:
        with open(log_path, encoding="utf-8") as f:
            log_data = json.load(f)
    except Exception as e:
        print(f"  ⚠️  ログ読み込み失敗: {e}")
        return

    for entry in log_data.get("races", []):
        if str(entry.get("race_no")) == str(race_no):
            entry["ev_suggestion"]   = ev_suggestion
            entry["actual_odds_top"] = [
                {"combo": c["combo"], "odds": c.get("actual_odds"), "ev": c.get("ev_pct")}
                for c in sorted(
                    [c2 for c2 in [
                        dict(cf, actual_odds=actual_odds.get(cf["combo"]),
                             ev_pct=(f"{(actual_odds.get(cf['combo'],0)*cf['prob']-1)*100:+.1f}%"
                                     if actual_odds.get(cf['combo']) else "N/A"))
                        for cf in entry.get("combos_full", [])
                    ] if c2.get("actual_odds")],
                    key=lambda x: (actual_odds.get(x["combo"], 0) * x.get("prob", 0) - 1),
                    reverse=True
                )[:15]
            ]
            break

    try:
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump(log_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"  ⚠️  ログ書き戻し失敗: {e}")


# ── メイン ───────────────────────────────────────────────────────────────
def sep(char="─", width=60):
    print(char * width)


def main():
    parser = argparse.ArgumentParser(
        description="展示後 実オッズ取得・EV計算・参戦判断更新"
    )
    parser.add_argument("--venue",   required=True,  help="会場名（例: 大村）")
    parser.add_argument("--date",    required=True,  help="日付（例: 2026-03-09）")
    parser.add_argument("--race",    type=int, default=None, help="対象レース番号（省略時: 全レース）")
    parser.add_argument("--min-ev",  type=float, default=0.0,
                        help="参戦EV閾値（デフォルト0.0 = EV+のみ）")
    parser.add_argument("--max-bets",type=int, default=8,
                        help="最大買い目数（デフォルト8）")
    parser.add_argument("--dry-run", action="store_true",
                        help="Excel/ログ書き込みを行わず結果のみ表示")
    args = parser.parse_args()

    sep()
    print("  apply_ev.py  ─  展示後EV計算・参戦判断更新")
    sep()

    # ── ログ読み込み ──────────────────────────────────────────────────────
    date_str = args.date.replace("/", "-")[:10]
    log_path = LOGS_DIR / f"{date_str}_{args.venue}.json"

    if not log_path.exists():
        print(f"❌ 予想ログが見つかりません: {log_path}")
        print("   先に load_race.py を実行してください")
        return

    with open(log_path, encoding="utf-8") as f:
        log_data = json.load(f)

    races = log_data.get("races", [])
    if args.race:
        races = [r for r in races if str(r.get("race_no")) == str(args.race)]
        if not races:
            print(f"❌ {args.race}R のログが見つかりません")
            return

    print(f"  🏁 会場: {args.venue}  日付: {date_str}")
    print(f"  📋 対象レース: {[r['race_no'] for r in races]}")
    print()

    # ── 各レース EV計算 ──────────────────────────────────────────────────
    excel_results = {}   # {race_no_str: ev_suggestion}

    for entry in races:
        race_no     = str(entry.get("race_no"))
        combos_full = entry.get("combos_full", [])

        print(f"  ── {race_no}R ─────────────────────────────────")

        if not combos_full:
            print(f"  ⚠️  combos_full が空です（load_race.py の再実行が必要）")
            continue

        # 実オッズ取得
        url = build_odds_url(args.venue, race_no, date_str)
        actual_odds = {}

        if url:
            print(f"  🔗 URL: {url}")
            actual_odds = fetch_actual_odds(url) or {}
        else:
            print(f"  ⚠️  URLを生成できません（会場コード未登録?）")

        if not actual_odds:
            print(f"  ❌ 実オッズ取得失敗 → このレースはスキップ")
            excel_results[race_no] = {
                "skip":   True,
                "reason": "実オッズ取得失敗",
                "buy_list": [],
                "total_bets": 0,
                "best_ev": None,
            }
            continue

        # EV計算
        ev_list    = calc_ev(combos_full, actual_odds)
        ev_sug     = suggest_by_ev(ev_list, min_ev=args.min_ev, max_bets=args.max_bets)

        # 結果表示
        if ev_sug["skip"]:
            print(f"  ⛔ 見送り推奨  {ev_sug['reason']}")
        else:
            print(f"  ✅ 参戦推奨  {ev_sug['total_bets']}点  {ev_sug['reason']}")
            for s in ev_sug["ev_summary"][:5]:
                print(f"     {s['combo']}  EV:{s['ev_pct']}  実{s['actual_odds']}倍  推定{s['prob_pct']}")

        excel_results[race_no] = ev_sug

        # ログ更新
        if not args.dry_run:
            update_log_ev(log_path, race_no, ev_sug, actual_odds)

        print()
        time.sleep(1.0)   # 連続アクセス抑制

    # ── Excel 更新 ────────────────────────────────────────────────────────
    if not excel_results:
        print("⚠️  更新するレースがありませんでした")
        return

    sep()
    if args.dry_run:
        print("  [dry-run] Excel/ログへの書き込みをスキップしました")
        return

    sheet_name = f"{args.venue}_数値"
    print(f"  📊 Excel 更新中: {EXCEL_FILE.name}  シート: {sheet_name}")
    update_excel_ev_row(str(EXCEL_FILE), sheet_name, excel_results)

    sep()
    print("  ✅ apply_ev.py 完了")
    sep()


if __name__ == "__main__":
    main()
