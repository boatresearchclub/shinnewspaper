#!/usr/bin/env python3
"""
Excel -> PNG converter (Windows)
CopyPicture方式: 印刷設定に依存せず、画面表示そのままをPNG化する
図形・オートシェイプのずれを防ぐ
A4 1ページ目のみ出力
"""

import sys
import time
from pathlib import Path


# ============================================================
# [lock] 日付ロックガード
#    load_race.py → fill_newspaper.py の流れで書き出された
#    .race_date_lock を確認し、日付が一致しない場合は終了する。
# ============================================================
def _check_date_lock(args_date: str | None) -> None:
    _lock_file = Path(__file__).parent / ".race_date_lock"
    if not _lock_file.exists():
        print("[NG] [GUARD] .race_date_lock が見つかりません。")
        print("   load_race.py を経由して実行してください。")
        sys.exit(1)

    _locked_date = _lock_file.read_text(encoding="utf-8").strip()

    # ロックファイルが空は通す
    if not _locked_date:
        return

    # --date 引数がある場合は一致チェック
    if args_date and args_date != _locked_date:
        print(f"[NG] [GUARD] 日付不一致: 指定日付={args_date} / ロック日付={_locked_date}")
        print("   load_race.py で処理した日付と一致しません。")
        sys.exit(1)


def get_sheet_names(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def get_newspaper_date(xlsx_path, sheet_name):
    """
    各シートのI1セルから新聞の日付を読み取り、YYYYMMDD形式の文字列で返す。
    日付が取得できない場合は今日の日付を返す。
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        cell_value = ws.cell(row=1, column=9).value  # I1セル
        wb.close()
        if cell_value is None:
            raise ValueError("I1セルが空です")
        if hasattr(cell_value, 'strftime'):
            return cell_value.strftime("%Y%m%d")
        # 文字列の場合はパース
        from datetime import datetime
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(str(cell_value), fmt).strftime("%Y%m%d")
            except ValueError:
                continue
        raise ValueError(f"日付形式が不明: {cell_value}")
    except Exception as e:
        print(f"  [!] 日付取得エラー ({e}) → 今日の日付を使用")
        return time.strftime("%Y%m%d")


def get_page1_range(ws, xl):
    """
    印刷範囲（PrintArea）を優先して使用し、
    未設定の場合は改ページ位置から1ページ目の範囲を返す。
    """
    # (1) PrintAreaが設定されていればそれを使う（最も確実）
    print_area = ws.PageSetup.PrintArea
    if print_area:
        rng = ws.Range(print_area)
        print(f"  PrintArea使用: {print_area}")
        return rng

    # (2) 改ページ位置から1ページ目を算出
    h_breaks = ws.HPageBreaks
    last_row = (h_breaks.Item(1).Location.Row - 1
                if h_breaks.Count > 0
                else ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1)

    v_breaks = ws.VPageBreaks
    last_col = (v_breaks.Item(1).Location.Column - 1
                if v_breaks.Count > 0
                else ws.UsedRange.Column + ws.UsedRange.Columns.Count - 1)

    first_row = ws.UsedRange.Row
    first_col = ws.UsedRange.Column

    return ws.Range(
        ws.Cells(first_row, first_col),
        ws.Cells(last_row, last_col)
    )


def convert_with_copypicture(xlsx_path, sheet_names, out_dir, dpi=150):
    import win32com.client
    import win32clipboard
    import win32con
    import pythoncom
    from PIL import Image
    import io

    pythoncom.CoInitialize()

    # 常に新規Excelプロセスを起動（既存プロセスのBOOK1等が干渉しないよう）
    xl_created_here = True
    xl = win32com.client.Dispatch("Excel.Application")
    print("  Excelを新規起動しました")

    xl.Visible = True   # CopyPictureはVisible=Trueが必要
    xl.DisplayAlerts = False

    try:
        wb = xl.Workbooks.Open(str(xlsx_path))

        for sheet_name in sheet_names:
            ws = wb.Sheets(sheet_name)
            ws.Activate()

            # A1セルを画面左上に表示（スクロール位置をリセット）
            xl.ActiveWindow.ScrollRow = 1
            xl.ActiveWindow.ScrollColumn = 1

            # 1ページ目の範囲を取得
            page1 = get_page1_range(ws, xl)
            print(f"  1ページ目の範囲: {page1.Address}")

            # 画面表示でクリップボードにコピー
            xl.ActiveWindow.WindowState = -4137  # xlMaximized
            xl.ActiveWindow.ScrollRow    = 1
            xl.ActiveWindow.ScrollColumn = 1
            time.sleep(0.3)

            page1.CopyPicture(
                Appearance=1,   # xlScreen
                Format=2        # xlBitmap
            )

            # クリップボードから画像を取得（占有中の場合はリトライ）
            time.sleep(0.5)
            data = None
            for attempt in range(10):
                try:
                    win32clipboard.OpenClipboard(0)
                    try:
                        data = win32clipboard.GetClipboardData(win32con.CF_DIB)
                    finally:
                        win32clipboard.CloseClipboard()
                    break
                except Exception as e:
                    print(f"  クリップボード待機中... ({attempt + 1}/10): {e}")
                    time.sleep(0.5)
            if data is None:
                raise RuntimeError("クリップボードからデータを取得できませんでした（10回リトライ後）")

            img = Image.open(io.BytesIO(_dib_to_bmp(data)))

            # DPIリサンプル（元画像は96DPI相当なのでdpi/96倍に拡大）
            if dpi != 96:
                w, h = img.size
                scale = dpi / 96
                img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)

            # 右・下の白余白をtrimしてからpadding追加
            import numpy as np
            from PIL import ImageOps
            arr = np.array(img.convert("RGB"))
            h_img, w_img = arr.shape[:2]
            # 下端: 閾値240（真っ白のみ除去）で最後の非白行を検出
            non_white_rows = ~((arr[:,:,0] > 240) & (arr[:,:,1] > 240) & (arr[:,:,2] > 240))
            rows = np.any(non_white_rows, axis=1)
            bottom = int(np.where(rows)[0][-1]) if np.any(rows) else h_img - 1
            # 右端: 閾値240で最後の非白列を検出
            cols = np.any(non_white_rows, axis=0)
            right = int(np.where(cols)[0][-1]) if np.any(cols) else w_img - 1
            print(f"  [trim] bottom={bottom}/{h_img-1}  right={right}/{w_img-1}")
            img = img.crop((0, 0, right + 1, bottom + 1))
            # 四方均等padding
            img = ImageOps.expand(img, border=15, fill=(255, 255, 255))

            safe_name = "".join(c for c in sheet_name if c not in r'\/:*?"<>|')
            news_date = get_newspaper_date(xlsx_path, sheet_name)
            png_path = out_dir / f"{safe_name}_{news_date}.png"

            # 既存ファイルがロックされている場合はタイムスタンプ付き別名で保存
            if png_path.exists():
                try:
                    png_path.unlink()
                except PermissionError:
                    ts = time.strftime("%Y%m%d_%H%M%S")
                    png_path = out_dir / f"{safe_name}_{news_date}_{ts}.png"
                    print(f"  [!] ファイルがロック中のため別名で保存: {png_path.name}")

            try:
                img.save(str(png_path), "PNG", dpi=(dpi, dpi))
                print(f"  Saved: {png_path.name}")
            except Exception as e:
                ts = time.strftime("%Y%m%d_%H%M%S")
                fallback_path = out_dir / f"{safe_name}_{news_date}_{ts}.png"
                img.save(str(fallback_path), "PNG", dpi=(dpi, dpi))
                print(f"  [!] 保存エラー ({e}) → 別名で保存: {fallback_path.name}")

        wb.Close(False)
    finally:
        # このスクリプトが新規起動したExcelのみ終了する
        # 既存のExcelプロセスはそのまま残す
        if xl_created_here:
            xl.Quit()
        pythoncom.CoUninitialize()


def trim_whitespace(img, padding=40):
    """
    四辺の白余白を検出してcropし、均等なpaddingを付けて返す。
    右端のみ罫線検出（縦方向連続非白列）でカット。
    上・左・下は非白ピクセルの端を検出してcrop。
    """
    import numpy as np
    from PIL import Image

    arr = np.array(img.convert("RGB"))
    img_h, img_w = arr.shape[:2]

    # 非白ピクセル判定（グレー罫線も拾うため閾値220）
    non_white = ~((arr[:,:,0] > 235) & (arr[:,:,1] > 235) & (arr[:,:,2] > 235))

    if not np.any(non_white):
        return img

    col_counts = np.sum(non_white, axis=0)
    row_counts = np.sum(non_white, axis=1)

    # 上・下・左: 非白ピクセルが存在する最初/最後の行・列
    top    = int(np.where(row_counts > 0)[0][0])
    bottom = int(np.where(row_counts > 0)[0][-1])
    left   = int(np.where(col_counts > 0)[0][0])

    # 右端罫線: 非白ピクセルが高さの一定割合以上の列を閾値を下げながら検索
    right_border = None
    for ratio in (0.90, 0.80, 0.70, 0.60, 0.50):
        cols = np.where(col_counts >= img_h * ratio)[0]
        if len(cols) > 0:
            right_border = int(cols[-1])
            break
    if right_border is None:
        occupied = np.where(col_counts > img_h * 0.05)[0]
        right_border = int(occupied[-1]) if len(occupied) > 0 else img_w - 1

    print(f"  [trim] img={img_w}x{img_h}  top={top} bottom={bottom} left={left} right={right_border}")

    cropped = img.crop((left, top, right_border + 1, bottom + 1))
    cw, ch = cropped.size
    canvas = Image.new("RGB", (cw + padding * 2, ch + padding * 2), (255, 255, 255))
    canvas.paste(cropped, (padding, padding))
    return canvas


def _dib_to_bmp(dib_data):
    """DIBデータをBMPに変換（ヘッダサイズを動的取得）
    
    DIBヘッダのサイズは環境・Excelバージョンによって異なる:
      BITMAPINFOHEADER  = 40バイト（旧来）→ オフセット54
      BITMAPV4HEADER    = 108バイト        → オフセット122
      BITMAPV5HEADER    = 124バイト        → オフセット138
    固定値54を使うと後者の場合にピクセルデータ位置がずれ、
    Pillowが真っ白画像として読み込んでしまう。
    """
    import struct
    # DIBヘッダの先頭4バイトからヘッダサイズを動的取得
    dib_header_size = struct.unpack_from('<I', dib_data, 0)[0]
    pixel_offset = 14 + dib_header_size  # ファイルヘッダ(14) + DIBヘッダ(可変)

    bmp_header = b'BM'
    file_size = 14 + len(dib_data)
    bmp_header += struct.pack('<I', file_size)
    bmp_header += b'\x00\x00\x00\x00'
    bmp_header += struct.pack('<I', pixel_offset)
    return bmp_header + dib_data


def ensure_deps():
    import subprocess
    pkgs = []
    try:
        import win32com.client
    except ImportError:
        pkgs.append("pywin32")
    try:
        import PIL
    except ImportError:
        pkgs.append("Pillow")
    if pkgs:
        print(f"  Installing: {', '.join(pkgs)} ...")
        subprocess.run([sys.executable, "-m", "pip", "install"] + pkgs, check=True)


# デフォルトの処理対象ファイル
DEFAULT_XLSX = Path(r"C:\Users\user\Desktop\データ収集\新聞出力_filled.xlsx")


def main():
    print("=" * 50)
    print("  Excel -> PNG converter  (CopyPicture方式)")
    print("  ※図形・オートシェイプのずれを防ぎます")
    print("  ※A4サイズ・1ページ目のみ出力")
    print("=" * 50)
    print()

    ensure_deps()

    # --file / --date 引数があればそれを使用
    import argparse
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--file", type=str, default=None)
    parser.add_argument("--date", type=str, default=None,
                        help="出力フォルダの日付 (例: 2026-03-31)。省略時は今日の日付")
    args, _ = parser.parse_known_args()

    # [lock] 日付ロックチェック（load_race.py 経由でない直接実行を防ぐ）
    _check_date_lock(args.date)

    if args.file:
        # 引数で明示指定
        xlsx_path = Path(args.file).resolve()
    elif DEFAULT_XLSX.exists():
        # デフォルトパスが存在すれば自動使用
        xlsx_path = DEFAULT_XLSX
        print(f"  対象ファイル: {xlsx_path}")
    else:
        # どちらもなければ対話式
        while True:
            try:
                xlsx_input = input("Excelファイルのパス (ドラッグ&ドロップ可): ").strip().strip("'\"")
            except (KeyboardInterrupt, EOFError):
                print("\nキャンセルしました。")
                sys.exit(0)
            if not xlsx_input:
                continue
            xlsx_path = Path(xlsx_input).resolve()
            if xlsx_path.exists() and xlsx_path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
                break
            print(f"  ファイルが見つかりません: {xlsx_path}\n")

    if not xlsx_path.exists():
        print(f"[NG] ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    try:
        sheet_names = get_sheet_names(xlsx_path)
    except Exception as e:
        print(f"Excelの読み込みエラー: {e}")
        sys.exit(1)

    print(f"\n  シート ({len(sheet_names)}): {', '.join(sheet_names)}")

    # --date が指定されている場合は、そのシートのI1セル日付と照合してフィルタする
    if args.date:
        filtered = []
        for sn in sheet_names:
            news_date_str = get_newspaper_date(xlsx_path, sn)  # YYYYMMDD形式
            sheet_date_hyphen = (
                f"{news_date_str[:4]}-{news_date_str[4:6]}-{news_date_str[6:8]}"
                if len(news_date_str) == 8 else ""
            )
            if sheet_date_hyphen == args.date:
                filtered.append(sn)
            else:
                print(f"  [スキップ]  シート [{sn}] 日付({sheet_date_hyphen}) ≠ 指定日付({args.date}) → スキップ")
        if not filtered:
            print(f"[NG] 指定日付({args.date})に一致するシートが見つかりません。")
            sys.exit(1)
        sheet_names = filtered
        print(f"  [OK] 対象シート ({len(sheet_names)}): {', '.join(sheet_names)}")

    # 日付フォルダに保存（例: ボートリサーチ新聞_出力\2026-03-20\）
    # --date 指定があればその日付、なければ今日の日付をフォルダ名に使う
    out_date = args.date if args.date else time.strftime("%Y-%m-%d")
    out_dir = Path(r"C:\Users\user\Pictures\ボートリサーチ新聞\ボートリサーチ新聞_出力") / out_date
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"  出力先: {out_dir}")
    print(f"  DPI: 150")
    print()
    print("変換中... しばらくお待ちください")
    print("(Excelが一時的に起動します)")
    print()

    try:
        convert_with_copypicture(xlsx_path, sheet_names, out_dir, dpi=150)
    except Exception as e:
        print(f"\nエラー: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    saved = list(out_dir.glob("*.png"))
    print(f"\n完了！ {len(sheet_names)}枚のPNGを保存しました (出力先に計{len(saved)}枚):")
    print(f"  {out_dir}")


if __name__ == "__main__":
    main()
