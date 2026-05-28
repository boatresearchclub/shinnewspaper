# -*- coding: utf-8 -*-
"""
learn_bias.py
=============
backtest_engine.py が生成した突合結果キャッシュ（.bt_result_cache.csv）を読み込み、
指標別のバイアスを深掘り分析して correction_table.json を高精度に更新する。

【設計思想】
  backtest_engine.py → 基本的中率・ROI を計算する「事実記録係」
  learn_bias.py      → そこから「なぜ当たる/外れるか」を学習する「分析官」

  Layer 2: 指標別精度分析
    - レーススコア最適閾値の自動探索（グリッドサーチ）
    - 複合条件分析（ランク×戦略、スコア帯×脅威帯）
    - 会場固有バイアスの抽出（びわこ等の個性）
    - 時系列トレンド検出（最近3ヶ月のモデル劣化/改善）
    - 見送り品質の精緻評価

  Layer 3: 高精度補正テーブル生成
    - 基本重み（backtest_engine互換）に加え以下を追加：
      * composite_weight  複合条件重み（ランク×戦略）
      * score_threshold   最適化済みスコア閾値
      * venue_bias        会場固有補正（基準差分）
      * trend_correction  直近トレンド補正係数
      * skip_quality      見送り品質スコア（0〜100）
      * roi_band_matrix   ROI帯ごとの期待値マップ

【使い方】
  python learn_bias.py                    # 全会場・全データ
  python learn_bias.py --venue びわこ     # 特定会場のみ
  python learn_bias.py --months 3         # 直近3ヶ月のみ
  python learn_bias.py --min_samples 5    # 集計最低サンプル数変更
  python learn_bias.py --no-merge         # correction_table.json を上書きしない
  python learn_bias.py --report           # learn_bias_report.xlsx を出力

【ファイル構成（backtest_engine と同じフォルダに置く）】
  scripts/
    backtest_engine.py
    learn_bias.py        ← このファイル
    correction_table.json   ← backtest_engine が生成、learn_bias が高度化
    learn_bias_detail.json  ← learn_bias の詳細分析出力（新規生成）
    learn_bias_report.xlsx  ← --report オプションで生成
    .bt_result_cache.csv    ← backtest_engine の突合結果キャッシュ（入力）
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

# ══════════════════════════════════════════════════════════
# パス定数（backtest_engine.py と同じディレクトリ想定）
# ══════════════════════════════════════════════════════════
_HERE            = Path(__file__).parent
BT_CACHE_CSV     = _HERE / ".bt_result_cache.csv"
CORRECTION_JSON  = _HERE / "correction_table.json"
DETAIL_JSON      = _HERE / "learn_bias_detail.json"
REPORT_XLSX      = _HERE / "learn_bias_report.xlsx"


# ══════════════════════════════════════════════════════════
# ユーティリティ
# ══════════════════════════════════════════════════════════

def sep(c="=", w=55):
    print(c * w)


def _load_json(path: Path) -> dict:
    if path.exists():
        with open(str(path), encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_json(path: Path, data: dict):
    with open(str(path), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  [保存] {path.name}")


def _fix_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    """キャッシュCSVの型を修正（backtest_engine._fix_bt_dtypes と同等）"""
    if df.empty:
        return df
    for col in ["結果データあり"]:
        if col in df.columns:
            df[col] = df[col].map(
                lambda v: True  if str(v).strip().lower() == "true"
                          else False if str(v).strip().lower() == "false"
                          else None
            )
    if "的中" in df.columns:
        df["的中"] = df["的中"].map(
            lambda v: True  if str(v).strip().lower() == "true"
                      else False if str(v).strip().lower() == "false"
                      else None
        )
    for col in ["払戻金", "投資額", "回収額", "収支", "生成点数", "見送り推奨",
                "レーススコア", "脅威合計", "1号艇脆弱性", "最大脅威艇"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def _stats(df: pd.DataFrame, label: str = "") -> dict:
    """基本統計を計算"""
    n = len(df)
    if n == 0:
        return {"レース数": 0, "的中数": 0, "的中率": 0.0,
                "総投資": 0, "総回収": 0, "ROI": 0.0,
                "平均払戻": 0.0, "信頼度": "低（N=0）"}
    hits = df["的中"].sum()
    inv  = df["投資額"].sum()
    ret  = df["回収額"].sum()
    hit_rate = hits / n * 100
    roi      = (ret - inv) / inv * 100 if inv > 0 else 0.0
    # ベルヌーイ信頼区間 95%（Wilson score）
    p   = hits / n
    z   = 1.96
    lo  = (p + z**2/(2*n) - z * (p*(1-p)/n + z**2/(4*n**2))**0.5) / (1 + z**2/n)
    hi  = (p + z**2/(2*n) + z * (p*(1-p)/n + z**2/(4*n**2))**0.5) / (1 + z**2/n)
    confidence = "高(N≥50)" if n >= 50 else "中(N≥20)" if n >= 20 else "低(N<20)"
    return {
        "レース数": n,
        "的中数": int(hits),
        "的中率": round(hit_rate, 2),
        "95%CI": f"{lo*100:.1f}〜{hi*100:.1f}",
        "総投資": int(inv),
        "総回収": int(ret),
        "収支": int(ret - inv),
        "ROI": round(roi, 2),
        "平均払戻": round(ret / hits, 0) if hits > 0 else 0.0,
        "信頼度": confidence,
        **({"label": label} if label else {}),
    }


def _weight(hr: float, base_hr: float,
            lo: float = 0.3, hi: float = 3.0) -> float:
    """的中率から補正重みを計算（0.3〜3.0にクリップ）"""
    if base_hr <= 0:
        return 1.0
    return round(max(lo, min(hi, hr / base_hr)), 3)


# ══════════════════════════════════════════════════════════
# Step 0: データ読み込み
# ══════════════════════════════════════════════════════════

def load_cache(cache_path: Path, venue: str | None, months: int | None) -> pd.DataFrame:
    """突合結果キャッシュを読み込み、フィルタを適用"""
    if not cache_path.exists():
        print(f"  [NG] 突合キャッシュが見つかりません: {cache_path}")
        print("       先に backtest_engine.py を実行してください。")
        sys.exit(1)

    print(f"  読込: {cache_path.name}")
    df = pd.read_csv(str(cache_path), encoding="utf-8-sig", dtype=str)
    df = _fix_dtypes(df)

    # 結果データあり＆購入レースのみ
    df = df[df["結果データあり"] == True].copy()
    df_buy = df[df["見送り推奨"] == 0].copy()

    print(f"  総レース数（結果あり）: {len(df)}R  うち購入: {len(df_buy)}R")

    if venue:
        df_buy = df_buy[df_buy["会場"] == venue]
        print(f"  会場フィルタ[{venue}]: {len(df_buy)}R")

    if months and months > 0:
        df_buy["日付"] = pd.to_datetime(df_buy["日付"], errors="coerce")
        cutoff = pd.Timestamp.now() - pd.DateOffset(months=months)
        df_buy = df_buy[df_buy["日付"] >= cutoff]
        print(f"  直近{months}ヶ月フィルタ: {len(df_buy)}R")

    # 見送りデータも別途返す（見送り品質評価用）
    df_skip = df[df["見送り推奨"] == 1].copy()

    return df_buy, df_skip, df


# ══════════════════════════════════════════════════════════
# Step 1: レーススコア最適閾値グリッドサーチ
# ══════════════════════════════════════════════════════════

def search_score_threshold(df: pd.DataFrame, min_samples: int = 10) -> dict:
    """
    レーススコア閾値をグリッドサーチして最適値を探索。

    評価指標：
      - 閾値以上のレースの的中率が最大になる閾値
      - かつROI > -10% を満たすこと
      - かつサンプル数 >= min_samples

    戻り値例:
      {
        "optimal_threshold": 50,
        "optimal_hit_rate": 35.2,
        "optimal_roi": +8.3,
        "optimal_n": 42,
        "grid": { "0": {...}, "10": {...}, ... }
      }
    """
    results = {}
    best = {"threshold": 0, "hit_rate": 0.0, "roi": 0.0, "n": 0, "score": -999}

    scores = sorted(df["レーススコア"].dropna().unique())
    # グリッドは5刻み（データの分布に合わせて調整）
    candidates = list(range(0, 101, 5))

    for th in candidates:
        sub = df[df["レーススコア"] >= th]
        if len(sub) < min_samples:
            continue
        s = _stats(sub)
        hr  = s["的中率"]
        roi = s["ROI"]
        n   = s["レース数"]
        # 複合スコア: 的中率 × ROI補正 × サンプル信頼性
        reliability = min(1.0, n / 50)
        composite   = hr * reliability * (1 + max(0, roi) / 100)

        results[str(th)] = {
            "閾値": th,
            "レース数": n,
            "的中率": hr,
            "ROI": roi,
            "複合スコア": round(composite, 3),
        }
        if composite > best["score"] and roi > -10 and n >= min_samples:
            best = {"threshold": th, "hit_rate": hr,
                    "roi": roi, "n": n, "score": composite}

    print(f"  スコア最適閾値: {best['threshold']}点  "
          f"(的中率 {best['hit_rate']:.1f}%  ROI {best['roi']:+.1f}%  {best['n']}R)")
    return {
        "optimal_threshold": best["threshold"],
        "optimal_hit_rate":  best["hit_rate"],
        "optimal_roi":       best["roi"],
        "optimal_n":         best["n"],
        "grid": results,
    }


# ══════════════════════════════════════════════════════════
# Step 2: 複合条件分析（クロス集計）
# ══════════════════════════════════════════════════════════

def analyze_composite(df: pd.DataFrame, min_samples: int = 5) -> dict:
    """
    ランク×戦略、スコア帯×脅威帯 のクロス集計で複合重みを算出。
    サンプル不足は全体平均でスムージング。
    """
    base_hr = _stats(df)["的中率"]
    result  = {}

    # ── 2-1: レースランク × 戦略 ─────────────────────────────────────
    rank_strat = {}
    for rank in df["レースランク"].dropna().unique():
        for strat in df["戦略"].dropna().unique():
            sub = df[(df["レースランク"] == rank) & (df["戦略"] == strat)]
            if len(sub) < min_samples:
                continue
            s = _stats(sub)
            key = f"{rank}×{strat[:20]}"
            rank_strat[key] = {
                **s,
                "weight": _weight(s["的中率"], base_hr),
            }
    result["rank_x_strategy"] = rank_strat

    # ── 2-2: スコア帯 × 脅威帯 ───────────────────────────────────────
    score_bands  = [(0,30),(30,50),(50,70),(70,100),(100,999)]
    threat_bands = [(0,12),(12,16),(16,20),(20,999)]
    score_threat = {}
    for (slo,shi) in score_bands:
        for (tlo,thi) in threat_bands:
            sub = df[
                (df["レーススコア"] >= slo) & (df["レーススコア"] < shi) &
                (df["脅威合計"]   >= tlo) & (df["脅威合計"]   < thi)
            ]
            if len(sub) < min_samples:
                continue
            s = _stats(sub)
            key = f"S{slo}〜{shi}×T{tlo}〜{thi}"
            score_threat[key] = {
                **s,
                "weight": _weight(s["的中率"], base_hr),
            }
    result["score_x_threat"] = score_threat

    # ── 2-3: 1号艇脆弱性 × レースランク ──────────────────────────────
    vuln_bands = [(0,25),(25,30),(30,35),(35,100)]
    vuln_rank  = {}
    for (vlo,vhi) in vuln_bands:
        for rank in ["S","A","B","C","D"]:
            sub = df[
                (df["1号艇脆弱性"] >= vlo) & (df["1号艇脆弱性"] < vhi) &
                (df["レースランク"] == rank)
            ]
            if len(sub) < min_samples:
                continue
            s = _stats(sub)
            key = f"V{vlo}〜{vhi}×{rank}"
            vuln_rank[key] = {
                **s,
                "weight": _weight(s["的中率"], base_hr),
            }
    result["vuln_x_rank"] = vuln_rank

    print(f"  複合条件: rank×strategy {len(rank_strat)}件 / "
          f"score×threat {len(score_threat)}件 / "
          f"vuln×rank {len(vuln_rank)}件")
    return result


# ══════════════════════════════════════════════════════════
# Step 3: 会場固有バイアス
# ══════════════════════════════════════════════════════════

def analyze_venue_bias(df: pd.DataFrame, min_samples: int = 10) -> dict:
    """
    会場ごとの的中率・ROIを全体平均との差分（バイアス）として数値化。
    """
    base = _stats(df)
    base_hr  = base["的中率"]
    base_roi = base["ROI"]
    result   = {}

    for venue, g in df.groupby("会場"):
        if len(g) < min_samples:
            continue
        s = _stats(g)
        result[venue] = {
            **s,
            "hit_rate_bias": round(s["的中率"] - base_hr,  2),   # 正 = 平均より当たりやすい
            "roi_bias":      round(s["ROI"]     - base_roi, 2),   # 正 = 平均よりROI高い
            "weight":        _weight(s["的中率"], base_hr),
        }
        bias_sign = "↑" if s["的中率"] > base_hr else "↓"
        print(f"    [{venue}]  的中率 {s['的中率']:.1f}% "
              f"({bias_sign}{abs(s['的中率']-base_hr):.1f}pp)  "
              f"ROI {s['ROI']:+.1f}%  ({s['レース数']}R)")

    return result


# ══════════════════════════════════════════════════════════
# Step 4: 時系列トレンド分析
# ══════════════════════════════════════════════════════════

def analyze_trend(df: pd.DataFrame, min_samples: int = 5) -> dict:
    """
    月別の的中率・ROIを計算し、直近トレンドを検出。

    - 直近3ヶ月の的中率 vs 全体的中率 → 劣化/改善を数値化
    - trend_coeff: 的中率の線形回帰傾き（月当たりのpp変化）
    - recent_weight: 直近3ヶ月の的中率ベース補正係数
    """
    if "日付" not in df.columns:
        return {}

    df = df.copy()
    df["日付"] = pd.to_datetime(df["日付"], errors="coerce")
    df["月"]   = df["日付"].dt.to_period("M").astype(str)

    monthly = []
    for month, g in df.groupby("月"):
        if len(g) < min_samples:
            continue
        s = _stats(g)
        monthly.append({"月": month, **s})

    if len(monthly) < 2:
        return {"monthly": monthly, "trend_coeff": 0.0, "recent_weight": 1.0}

    mdf = pd.DataFrame(monthly).sort_values("月")

    # 線形回帰（月インデックス vs 的中率）
    x = np.arange(len(mdf))
    y = mdf["的中率"].values
    if len(x) >= 2:
        coeff = float(np.polyfit(x, y, 1)[0])   # pp/月
    else:
        coeff = 0.0

    # 直近3ヶ月の的中率
    recent3  = mdf.tail(3)
    recent_hr = recent3["的中率"].mean()
    base_hr   = _stats(df)["的中率"]
    recent_w  = _weight(recent_hr, base_hr) if base_hr > 0 else 1.0

    trend_label = (
        "改善傾向" if coeff > 0.3 else
        "劣化傾向" if coeff < -0.3 else
        "横ばい"
    )
    print(f"  トレンド: {trend_label}  傾き {coeff:+.2f} pp/月  "
          f"直近3M的中率 {recent_hr:.1f}%  直近重み {recent_w:.3f}")

    return {
        "monthly": mdf.to_dict(orient="records"),
        "trend_coeff":    round(coeff, 3),
        "trend_label":    trend_label,
        "recent_hit_rate": round(recent_hr, 2),
        "recent_weight":   round(recent_w, 3),
    }


# ══════════════════════════════════════════════════════════
# Step 5: 見送り品質評価
# ══════════════════════════════════════════════════════════

def evaluate_skip_quality(df_skip: pd.DataFrame, df_buy: pd.DataFrame) -> dict:
    """
    見送り推奨レースの「品質」を評価。

    品質指標：
      - 見送り正解率（高払戻レースを回避できたか）: 高いほど良い
      - 見送り損失率（低倍率的中を見逃したか）   : 低いほど良い
      - 見送りによるROI改善効果                  : 正なら見送りが効いている

    skip_quality_score (0〜100):
      正解率 × 0.5 + (1 - 損失率) × 0.3 + ROI改善効果 × 0.2
    """
    if len(df_skip) == 0:
        return {"message": "見送りレースなし"}

    # 高倍率（3000円以上）= 荒れレース
    n_skip   = len(df_skip)
    n_rough  = int((df_skip["払戻金"] >= 3000).sum())  # 荒れ回避（正解）
    n_miss   = int((df_skip["払戻金"] < 3000).sum())   # 堅い見送り（損失）

    skip_acc  = round(n_rough / n_skip * 100, 1) if n_skip else 0  # 正解率
    miss_rate = round(n_miss  / n_skip * 100, 1) if n_skip else 0  # 損失率

    # 見送りなし仮想ROI（もし全部買っていたら）
    buy_inv_hyp = n_skip * 100  # 仮に1点100円と仮定
    buy_ret_hyp = int(df_skip[df_skip["払戻金"] < 3000]["払戻金"].sum() * 0.1)  # 仮定
    roi_skip_effect = 0.0
    if len(df_buy) > 0:
        base_roi = _stats(df_buy)["ROI"]
        # 見送り分を追加購入した場合のROI変化（近似）
        skip_hit_rate = df_skip["的中"].sum() / n_skip * 100 if n_skip else 0
        roi_skip_effect = round(base_roi - skip_hit_rate, 2)

    # 品質スコア
    quality = (
        skip_acc  * 0.5 +
        (100 - miss_rate) * 0.3 +
        max(0, min(100, roi_skip_effect)) * 0.2
    )
    quality = round(max(0, min(100, quality)), 1)

    print(f"  見送り品質: スコア {quality:.0f}/100  "
          f"正解率 {skip_acc:.1f}%  損失率 {miss_rate:.1f}%  ({n_skip}R)")

    return {
        "見送りレース数":      n_skip,
        "荒れ回避数(正解)":    n_rough,
        "堅い見送り数(損失)":  n_miss,
        "見送り正解率":        skip_acc,
        "見送り損失率":        miss_rate,
        "見送り品質スコア":    quality,
        "skip_quality_score": quality,
    }


# ══════════════════════════════════════════════════════════
# Step 6: ROI帯マトリックス
# ══════════════════════════════════════════════════════════

def build_roi_band_matrix(df: pd.DataFrame, min_samples: int = 5) -> dict:
    """
    スコア帯 × ランク の ROI を2Dマトリックスで作成。
    「どのゾーンで買えばプラス収支か」を一覧化する。
    """
    score_bands = [(0,30,"低"),(30,50,"中低"),(50,70,"中"),(70,100,"高"),(100,999,"最高")]
    ranks       = ["S","A","B","C","D"]
    matrix      = {}

    for (slo, shi, slabel) in score_bands:
        row = {}
        for rank in ranks:
            sub = df[
                (df["レーススコア"] >= slo) & (df["レーススコア"] < shi) &
                (df["レースランク"] == rank)
            ]
            if len(sub) < min_samples:
                row[rank] = None
                continue
            s = _stats(sub)
            row[rank] = {
                "n":       s["レース数"],
                "hit_rate": s["的中率"],
                "roi":     s["ROI"],
                "signal":  "◎" if s["ROI"] > 0 and s["的中率"] > 30
                            else "○" if s["ROI"] > -10
                            else "×",
            }
        matrix[f"S{slo}〜{shi}({slabel})"] = row

    return matrix


# ══════════════════════════════════════════════════════════
# Step 7: 補正テーブルに学習結果をマージ
# ══════════════════════════════════════════════════════════

def merge_into_correction_table(
    existing_table: dict,
    score_analysis: dict,
    composite:      dict,
    venue_bias:     dict,
    trend:          dict,
    skip_quality:   dict,
    roi_matrix:     dict,
    df_buy:         pd.DataFrame,
) -> dict:
    """
    既存の correction_table.json に learn_bias の分析結果を追加/上書きする。

    上書きするキー:
      score_threshold   → グリッドサーチ最適値に更新
      venue_weight      → バイアス込みの精密重みに更新

    追加するキー（新規）:
      composite_weight  → ランク×戦略の複合重み
      score_x_threat    → スコア帯×脅威帯の重み
      vuln_x_rank       → 脆弱性×ランクの重み
      venue_bias        → 会場バイアス詳細
      trend             → 時系列トレンド情報
      skip_quality      → 見送り品質評価
      roi_band_matrix   → ROI帯マトリックス
      learn_bias_at     → 最終学習日時
    """
    merged = dict(existing_table)  # コピー

    # ── スコア閾値更新 ─────────────────────────────────────────────
    opt_th = score_analysis.get("optimal_threshold", 0)
    if opt_th > 0:
        merged["score_threshold"] = opt_th
        merged["score_threshold_source"] = "learn_bias_grid_search"

    # ── 会場重み更新（精密版）────────────────────────────────────────
    if venue_bias:
        merged["venue_weight"] = {
            k: v["weight"] for k, v in venue_bias.items()
        }

    # ── 複合条件重み（新規追加）──────────────────────────────────────
    # ランク×戦略
    rs_weights = {}
    for key, v in composite.get("rank_x_strategy", {}).items():
        rs_weights[key] = v.get("weight", 1.0)
    if rs_weights:
        merged["composite_weight"] = rs_weights

    # スコア帯×脅威帯
    st_weights = {}
    for key, v in composite.get("score_x_threat", {}).items():
        st_weights[key] = v.get("weight", 1.0)
    if st_weights:
        merged["score_x_threat_weight"] = st_weights

    # 脆弱性×ランク
    vr_weights = {}
    for key, v in composite.get("vuln_x_rank", {}).items():
        vr_weights[key] = v.get("weight", 1.0)
    if vr_weights:
        merged["vuln_x_rank_weight"] = vr_weights

    # ── トレンド補正 ──────────────────────────────────────────────
    if trend:
        merged["trend"] = {
            "trend_coeff":    trend.get("trend_coeff", 0.0),
            "trend_label":    trend.get("trend_label", ""),
            "recent_weight":  trend.get("recent_weight", 1.0),
            "recent_hit_rate": trend.get("recent_hit_rate", 0.0),
        }

    # ── 見送り品質 ──────────────────────────────────────────────────
    if skip_quality:
        merged["skip_quality"] = skip_quality

    # ── ROIマトリックス ─────────────────────────────────────────────
    if roi_matrix:
        merged["roi_band_matrix"] = roi_matrix

    # ── 会場バイアス詳細 ────────────────────────────────────────────
    if venue_bias:
        merged["venue_bias"] = {
            k: {
                "hit_rate":      v["的中率"],
                "hit_rate_bias": v["hit_rate_bias"],
                "roi":           v["ROI"],
                "roi_bias":      v["roi_bias"],
                "n":             v["レース数"],
            }
            for k, v in venue_bias.items()
        }

    # ── メタ情報 ─────────────────────────────────────────────────────
    merged["learn_bias_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    base = _stats(df_buy)
    merged["learn_bias_base_hit_rate"] = base["的中率"]
    merged["learn_bias_base_roi"]      = base["ROI"]
    merged["learn_bias_n"]             = base["レース数"]

    return merged


# ══════════════════════════════════════════════════════════
# Step 8: Excelレポート出力
# ══════════════════════════════════════════════════════════

def write_report_excel(
    df_buy:        pd.DataFrame,
    score_analysis: dict,
    composite:      dict,
    venue_bias:     dict,
    trend:          dict,
    skip_quality:   dict,
    roi_matrix:     dict,
    output_path:    Path,
):
    """learn_bias の詳細分析をExcelに出力"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        print("  [!] pip install openpyxl が必要です（レポートスキップ）")
        return

    HDR_FILL = PatternFill("solid", fgColor="2E4057")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    POS_FILL = PatternFill("solid", fgColor="C6EFCE")
    NEG_FILL = PatternFill("solid", fgColor="FCE4D6")
    SEC_FILL = PatternFill("solid", fgColor="D9E1F2")
    THIN     = Side(style="thin", color="CCCCCC")
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    ws = wb.active
    wb.remove(ws)

    def _hdr(ws, cols):
        row = []
        for c in cols:
            cell = ws.cell(row=ws.max_row + 1 if ws.max_row else 1, column=1)
            cell = ws.cell(row=1, column=1)
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    def _row(ws, vals, fill=None):
        ws.append(vals)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
        for cell in ws[ws.max_row]:
            cell.border = BORDER

    # ── シート1: スコア閾値グリッド ──────────────────────────────────
    ws1 = wb.create_sheet("スコア閾値グリッド")
    ws1.append(["スコア閾値グリッドサーチ結果"])
    ws1.append([f"最適閾値: {score_analysis.get('optimal_threshold')}点  "
                f"的中率: {score_analysis.get('optimal_hit_rate')}%  "
                f"ROI: {score_analysis.get('optimal_roi'):+.1f}%"])
    ws1.append([])
    ws1.append(["閾値", "レース数", "的中率(%)", "ROI(%)", "複合スコア"])
    for cell in ws1[ws1.max_row]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
    opt = score_analysis.get("optimal_threshold", -1)
    for th_str, d in sorted(score_analysis.get("grid", {}).items(), key=lambda x: int(x[0])):
        th = int(th_str)
        fill = POS_FILL if th == opt else None
        row = [th, d["レース数"], d["的中率"], d["ROI"], d["複合スコア"]]
        ws1.append(row)
        if fill:
            for cell in ws1[ws1.max_row]:
                cell.fill = fill
        for cell in ws1[ws1.max_row]:
            cell.border = BORDER

    # ── シート2: 複合条件（ランク×戦略）──────────────────────────────
    ws2 = wb.create_sheet("複合条件_ランク×戦略")
    ws2.append(["ランク×戦略 複合条件分析"])
    ws2.append([])
    ws2.append(["条件", "レース数", "的中率(%)", "ROI(%)", "重み", "信頼度"])
    for cell in ws2[ws2.max_row]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
    for key, d in sorted(composite.get("rank_x_strategy", {}).items()):
        w = d.get("weight", 1.0)
        fill = POS_FILL if w >= 1.3 else NEG_FILL if w <= 0.7 else None
        ws2.append([key, d["レース数"], d["的中率"], d["ROI"], w, d["信頼度"]])
        if fill:
            for cell in ws2[ws2.max_row]:
                cell.fill = fill
        for cell in ws2[ws2.max_row]:
            cell.border = BORDER

    # ── シート3: 会場バイアス ────────────────────────────────────────
    ws3 = wb.create_sheet("会場バイアス")
    ws3.append(["会場バイアス分析（全体平均との差分）"])
    ws3.append([])
    ws3.append(["会場", "レース数", "的中率(%)", "全体差(pp)", "ROI(%)", "ROI差(pp)", "重み"])
    for cell in ws3[ws3.max_row]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
    for venue, d in sorted(venue_bias.items(), key=lambda x: -x[1]["hit_rate_bias"]):
        bias = d["hit_rate_bias"]
        fill = POS_FILL if bias > 2 else NEG_FILL if bias < -2 else None
        ws3.append([venue, d["レース数"], d["的中率"],
                    bias, d["ROI"], d["roi_bias"], d["weight"]])
        if fill:
            for cell in ws3[ws3.max_row]:
                cell.fill = fill
        for cell in ws3[ws3.max_row]:
            cell.border = BORDER

    # ── シート4: 月別トレンド ────────────────────────────────────────
    ws4 = wb.create_sheet("月別トレンド")
    ws4.append(["月別トレンド分析"])
    ws4.append([f"トレンド: {trend.get('trend_label','')}  "
                f"傾き: {trend.get('trend_coeff',0):+.2f} pp/月  "
                f"直近重み: {trend.get('recent_weight',1.0):.3f}"])
    ws4.append([])
    ws4.append(["月", "レース数", "的中率(%)", "ROI(%)", "収支(円)"])
    for cell in ws4[ws4.max_row]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
    for m in trend.get("monthly", []):
        roi = m.get("ROI", 0)
        fill = POS_FILL if roi >= 0 else NEG_FILL
        ws4.append([m.get("月"), m.get("レース数"), m.get("的中率"), roi, m.get("収支")])
        for cell in ws4[ws4.max_row]:
            cell.fill = fill; cell.border = BORDER

    # ── シート5: ROIマトリックス ─────────────────────────────────────
    ws5 = wb.create_sheet("ROIマトリックス")
    ws5.append(["ROI帯マトリックス（スコア帯 × ランク）"])
    ws5.append(["☆ ◎ = ROI+・的中率30%超  ○ = ROI-10%以内  × = 不振ゾーン"])
    ws5.append([])
    ranks = ["S","A","B","C","D"]
    ws5.append(["スコア帯"] + ranks)
    for cell in ws5[ws5.max_row]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
    for band, row_data in roi_matrix.items():
        row = [band]
        for rank in ranks:
            d = row_data.get(rank)
            if d is None:
                row.append("―")
            else:
                row.append(f"{d['signal']} {d['hit_rate']:.0f}% ROI{d['roi']:+.0f}% ({d['n']}R)")
        ws5.append(row)
        for cell in ws5[ws5.max_row]:
            cell.border = BORDER

    wb.save(str(output_path))
    print(f"  [OK] レポート出力: {output_path}")


# ══════════════════════════════════════════════════════════
# コンソールサマリー出力
# ══════════════════════════════════════════════════════════

def print_summary(
    df_buy: pd.DataFrame,
    score_analysis: dict,
    venue_bias: dict,
    trend: dict,
    skip_quality: dict,
    merged_table: dict,
):
    sep()
    print("  learn_bias.py  分析サマリー")
    sep()
    base = _stats(df_buy)
    print(f"""
  ┌──────────────────────────────────────────────┐
  │  分析対象                                      │
  ├──────────────────────────────────────────────┤
  │  購入レース数 : {base['レース数']:>6} R                      │
  │  的中数       : {base['的中数']:>6} R                      │
  │  的中率       : {base['的中率']:>6.1f} %                    │
  │  ROI          : {base['ROI']:>+7.1f} %                    │
  └──────────────────────────────────────────────┘""")

    opt_th = score_analysis.get("optimal_threshold", 0)
    opt_hr = score_analysis.get("optimal_hit_rate", 0)
    print(f"\n  【スコア閾値】  最適: {opt_th}点  (的中率 {opt_hr:.1f}%)")

    if trend:
        print(f"  【トレンド】   {trend.get('trend_label','')}  "
              f"傾き {trend.get('trend_coeff',0):+.2f} pp/月  "
              f"直近3M {trend.get('recent_hit_rate',0):.1f}%")

    sq = skip_quality.get("skip_quality_score", 0)
    print(f"  【見送り品質】  スコア {sq:.0f}/100")

    # 高重み条件 TOP5
    cw = merged_table.get("composite_weight", {})
    if cw:
        top5 = sorted(cw.items(), key=lambda x: -x[1])[:5]
        print("\n  【複合条件 高重みTOP5】")
        for k, w in top5:
            bar = "█" * int(w * 5)
            print(f"    {k:<35}  重み {w:.2f}  {bar}")

    print()


# ══════════════════════════════════════════════════════════
# main
# ══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="競艇バイアス学習エンジン learn_bias.py"
    )
    parser.add_argument("--venue",       type=str,  default=None,
                        help="会場名フィルタ（例: びわこ）")
    parser.add_argument("--months",      type=int,  default=None,
                        help="直近N ヶ月のみ分析（例: 3）")
    parser.add_argument("--min_samples", type=int,  default=5,
                        help="集計最低サンプル数（デフォルト: 5）")
    parser.add_argument("--no-merge",    action="store_true",
                        help="correction_table.json を更新しない（dry-run）")
    parser.add_argument("--report",      action="store_true",
                        help="learn_bias_report.xlsx を出力する")
    parser.add_argument("--cache",       type=str,  default=None,
                        help="突合キャッシュCSVのパスを指定（省略=デフォルト）")
    parser.add_argument("--out_correction", type=str, default=None,
                        help="補正テーブルの出力先（省略=デフォルト）")
    args = parser.parse_args()

    cache_path      = Path(args.cache)          if args.cache           else BT_CACHE_CSV
    correction_path = Path(args.out_correction) if args.out_correction  else CORRECTION_JSON

    print()
    sep()
    print("  learn_bias.py  バイアス学習エンジン")
    sep()

    # ── Step 0: データ読み込み ────────────────────────────────────────
    print("\n[Step 0]  データ読み込み中...")
    df_buy, df_skip, df_all = load_cache(
        cache_path, venue=args.venue, months=args.months
    )
    if len(df_buy) < args.min_samples:
        print(f"  [NG] 購入レースが{args.min_samples}件未満です ({len(df_buy)}R)。")
        print("       backtest_engine.py を実行してから再試行してください。")
        sys.exit(1)

    # ── Step 1: スコア閾値グリッドサーチ ─────────────────────────────
    print("\n[Step 1]  レーススコア最適閾値を探索中...")
    score_analysis = search_score_threshold(df_buy, min_samples=args.min_samples)

    # ── Step 2: 複合条件分析 ──────────────────────────────────────────
    print("\n[Step 2]  複合条件分析中...")
    composite = analyze_composite(df_buy, min_samples=args.min_samples)

    # ── Step 3: 会場バイアス ──────────────────────────────────────────
    print("\n[Step 3]  会場バイアス分析中...")
    venue_bias = analyze_venue_bias(df_buy, min_samples=args.min_samples)

    # ── Step 4: 時系列トレンド ────────────────────────────────────────
    print("\n[Step 4]  時系列トレンド分析中...")
    trend = analyze_trend(df_buy, min_samples=args.min_samples)

    # ── Step 5: 見送り品質 ───────────────────────────────────────────
    print("\n[Step 5]  見送り品質評価中...")
    skip_quality = evaluate_skip_quality(df_skip, df_buy)

    # ── Step 6: ROIマトリックス ───────────────────────────────────────
    print("\n[Step 6]  ROI帯マトリックス構築中...")
    roi_matrix = build_roi_band_matrix(df_buy, min_samples=args.min_samples)

    # ── Step 7: 補正テーブルマージ ───────────────────────────────────
    print("\n[Step 7]  補正テーブルに学習結果をマージ中...")
    existing_table = _load_json(correction_path)
    if not existing_table:
        print("  [!] correction_table.json が見つかりません。")
        print("      backtest_engine.py を先に実行するか、空テーブルから開始します。")
        existing_table = {}

    merged = merge_into_correction_table(
        existing_table, score_analysis, composite,
        venue_bias, trend, skip_quality, roi_matrix, df_buy
    )

    # ── Step 8: 保存 ─────────────────────────────────────────────────
    print("\n[Step 8]  結果を保存中...")
    if not args.no_merge:
        _save_json(correction_path, merged)
        print(f"  [OK] correction_table.json 更新: {correction_path}")
    else:
        print("  [--no-merge] correction_table.json は更新しませんでした")

    # 詳細JSON
    detail = {
        "generated_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
        "score_analysis": score_analysis,
        "composite":      composite,
        "venue_bias":     venue_bias,
        "trend":          trend,
        "skip_quality":   skip_quality,
        "roi_band_matrix": roi_matrix,
    }
    _save_json(DETAIL_JSON, detail)

    # ── Step 9: Excelレポート ─────────────────────────────────────────
    if args.report:
        print("\n[Step 9]  Excelレポート出力中...")
        write_report_excel(
            df_buy, score_analysis, composite, venue_bias,
            trend, skip_quality, roi_matrix, REPORT_XLSX
        )

    # ── サマリー出力 ──────────────────────────────────────────────────
    print_summary(df_buy, score_analysis, venue_bias, trend, skip_quality, merged)

    sep()
    print("  完了！")
    sep()
    print(f"  補正テーブル  : {correction_path}")
    print(f"  詳細分析JSON  : {DETAIL_JSON}")
    if args.report:
        print(f"  レポート      : {REPORT_XLSX}")
    print()
    print("  【運用フロー】")
    print("    1. 月次: python backtest_engine.py")
    print("    2. 月次: python learn_bias.py --report")
    print("    3. load_race.py が correction_table.json を自動参照")
    print()


if __name__ == "__main__":
    main()
