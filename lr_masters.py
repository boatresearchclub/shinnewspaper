# -*- coding: utf-8 -*-
"""
lr_masters.py  ─  マスタ / CSV 読み込み
分割元: load_race.py
"""
import os, sys, glob, re, pathlib
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from lr_config import (
    CSV_DIR, MASTER_FILE, SHEET_MASTER, SHEET_PLAYER, SHEET_ININAGE,
    _GRADE_CSV_MAP, ST_KIMETE_CSV, KAIHO_VENUE_CSV, KAIHO_NATIONAL_CSV,
    VENUE_COURSE_CSV,
)
from lr_utils import safe_float

def load_motor_csv(venue, race_date_str, race_df=None):
    """
    出走表CSV（load_csvで読み込み済みのDataFrame）から
    モーター番号・2連対率を抽出して返す。

    出走表CSVの列名対応:
        レース列  : 「レース」または「R」
        艇番列    : 「艇番」または「枠」
        モーター番号 : 「M番」
        2連対率   : 「M2率」

    戻り値は build_jizen_members が期待する形式:
        columns = [race_no, boat_no, motor_no, motor_2rate]
    """
    if race_df is None or len(race_df) == 0:
        return None

    # 列名を特定
    race_col = next((c for c in race_df.columns if c in ("レース", "R", "レース番号")), None)
    boat_col = next((c for c in race_df.columns if c in ("艇番", "枠")), None)
    mno_col  = next((c for c in race_df.columns if c in ("M番", "モーター番号", "motor_no")), None)
    m2r_col  = next((c for c in race_df.columns if c in ("M2率", "モーター2連率", "M2連対率", "motor_2rate")), None)

    if mno_col is None or m2r_col is None:
        print(f"  [!]  出走表CSVにモーター列（M番/M2率）が見つかりません。機力評価はスキップします。")
        return None

    try:
        rows = []
        for _, row in race_df.iterrows():
            race_no  = pd.to_numeric(row.get(race_col), errors="coerce") if race_col else None
            boat_no  = pd.to_numeric(row.get(boat_col), errors="coerce") if boat_col else None
            motor_no = pd.to_numeric(row.get(mno_col),  errors="coerce")
            m2rate   = pd.to_numeric(row.get(m2r_col),  errors="coerce")
            rows.append({
                "race_no":     race_no,
                "boat_no":     boat_no,
                "motor_no":    motor_no,
                "motor_2rate": m2rate,
            })
        df = pd.DataFrame(rows)
        print(f"  ??  モーターデータを出走表CSVから読み込みました（{len(df)}艇分）")
        return df
    except Exception as e:
        print(f"  [!]  モーターデータ抽出エラー: {e}")
        return None

# ============================================================
# マスタ読み込み
# ============================================================

# 【軽微②改善】選手名正規化とマスタルックアップを1か所に集約
# 旧方式: 登録時にエイリアスを追加 → 4文字名が5文字名に誤マッチする可能性
# 新方式: ルックアップ関数で「完全一致 → 先頭N文字一致」の順で検索
def _lookup_name_course(master_dict, name, course_str):
    """
    course_masterから(name, course)でルックアップする。
    完全一致 → 4文字前方一致(クエリ5字→マスタ4字) → 5文字前方一致(クエリ4字→マスタ5字) の順で検索。

    【修正②】逆方向マッチを追加。
    旧: クエリが5字のとき先頭4字でマスタを引く方向のみ対応。
    新: クエリが4字のとき、マスタ側に5字キーしかない場合も先頭4字で照合する。
        ※ マスタ登録時にエイリアスキー(name[:4], course)を追加しているため
           実質は本検索で拾えるが、旧データ等でエイリアス未登録の場合の安全網。
    """
    key_exact = (name, course_str)
    if key_exact in master_dict:
        return master_dict[key_exact]
    # 【方向A】クエリが5字以上 → 先頭4字でマスタを引く（旧来の動作）
    if len(name) >= 5:
        key4 = (name[:4], course_str)
        if key4 in master_dict:
            return master_dict[key4]
    # 【方向B・修正②追加】クエリが4字 → マスタ側が5字以上で先頭4字一致するものを探す
    # エイリアス未登録の旧マスタデータでも正しく引けるようにする
    if len(name) == 4:
        for (mname, mcourse), mval in master_dict.items():
            if mcourse == course_str and len(mname) >= 5 and mname[:4] == name:
                return mval
    return None

def _lookup_player(master_dict, name):
    """player_masterから選手名でルックアップ。完全一致 → 4文字前方一致(双方向) の順。

    【修正②】逆方向マッチを追加（クエリ4字・マスタ5字以上のケース）。
    """
    if name in master_dict:
        return master_dict[name]
    # 【方向A】クエリが5字以上 → 先頭4字エイリアスで引く
    if len(name) >= 5:
        alias = name[:4]
        if alias in master_dict:
            return master_dict[alias]
    # 【方向B・修正②追加】クエリが4字 → マスタ側が5字以上で先頭4字一致するものを探す
    if len(name) == 4:
        for mname, mval in master_dict.items():
            if isinstance(mname, str) and len(mname) >= 5 and mname[:4] == name:
                return mval
    return None

def load_masters(wb, race_grade: str = "一般"):
    """コース別マスタ・選手指数マスタ・イン逃げ分析を読み込む

    Parameters
    ----------
    wb          : openpyxl Workbook（ボートリサーチ_マスタ.xlsx）
    race_grade  : "一般" / "G1" / "G2" / "G3" / "SG"
                  グレードに応じて参照するCSV・シートを切り替える。
                  対応CSVが存在しない場合は自動的に一般戦マスタにフォールバック。
    """
    # グレード → CSV/シートマップを解決（未知グレードは一般扱い）
    _gmap = _GRADE_CSV_MAP.get(race_grade, _GRADE_CSV_MAP["一般"])
    _gmap_ippan = _GRADE_CSV_MAP["一般"]  # フォールバック用

    # 【修正④】G3はG2マスタを流用する旨をログに明示
    if _gmap.get("_uses_g2_master"):
        print(f"  [i]  グレード「{race_grade}」はG2マスタを流用します（G3専用マスタは未作成）")

    # SG専用CSVが未生成の場合はG1マスタにフォールバック
    _effective_grade = race_grade
    if race_grade == "SG":
        _sg_vc = _gmap["venue_course"]
        if isinstance(_sg_vc, pathlib.Path) and not _sg_vc.exists():
            print("  [!]  SG専用マスタ未生成 → G1マスタにフォールバックします")
            _gmap = _GRADE_CSV_MAP["G1"]
            _effective_grade = "G1"

    def _resolve_csv(key):
        """グレード別パスが存在すればそれを、なければ一般戦パスを返す"""
        p = _gmap[key]
        if isinstance(p, pathlib.Path) and not p.exists():
            p_fallback = _gmap_ippan[key]
            print(f"  [!]  {race_grade}用マスタなし → 一般戦マスタを使用: {p.name}")
            return p_fallback
        return p

    _vc_csv_path       = _resolve_csv("venue_course")
    _tenkai_venue_path = _resolve_csv("tenkai_venue")
    _tenkai_nat_path   = _resolve_csv("tenkai_national")

    # コース別マスタシート名（グレード別シートがなければ一般戦シートにフォールバック）
    _sheet_master  = _gmap["sheet_master"]
    _sheet_ininage = _gmap["sheet_ininage"]
    if _sheet_master not in wb.sheetnames:
        print(f"  [!]  シート「{_sheet_master}」なし → 一般戦シートを使用")
        _sheet_master = SHEET_MASTER
    if _sheet_ininage not in wb.sheetnames:
        _sheet_ininage = SHEET_ININAGE

    # コース別マスタ
    ws_m = wb[_sheet_master]
    master_rows = list(ws_m.iter_rows(values_only=True))
    # ヘッダ行を探す
    header_row = None
    for i, row in enumerate(master_rows):
        if row and row[0] == "選手名":
            header_row = i
            break
    if header_row is None:
        return {}, {}, {}
    headers_m = master_rows[header_row]
    course_master = {}  # {(選手名, コース): {指数dict}}
    for row in master_rows[header_row+1:]:
        if not row or row[0] is None:
            continue
        d = dict(zip(headers_m, row))
        name_full = str(d.get("選手名","")).strip()
        course_str = str(d.get("コース","")).strip()
        key = (name_full, course_str)
        course_master[key] = d
        # 【軽微②改善】5文字名の4文字エイリアスは、4文字キーが未登録の場合のみ追加
        # 【修正】エイリアスdictの「選手名」はフルネームのまま保持する。
        # （例: 大豆生田蒼 → エイリアスキー(大豆生田, course)でも選手名は大豆生田蒼）
        if len(name_full) >= 5:
            alias_key = (name_full[:4], course_str)
            if alias_key not in course_master:
                alias_d = dict(d)
                alias_d["選手名"] = name_full  # フルネームを明示保持
                course_master[alias_key] = alias_d

    # 選手指数マスタ
    ws_p = wb[SHEET_PLAYER]
    player_rows = list(ws_p.iter_rows(values_only=True))
    header_row_p = None
    for i, row in enumerate(player_rows):
        # A列が「登録番号」、B列が「選手名」の行をヘッダとして検出
        if row and row[1] == "選手名":
            header_row_p = i
            break
    headers_p = player_rows[header_row_p] if header_row_p is not None else []
    player_master = {}  # {選手名: {指数dict}}
    if header_row_p is not None:
        for row in player_rows[header_row_p+1:]:
            if not row or row[0] is None:
                continue
            d = dict(zip(headers_p, row))
            name = str(d.get("選手名","")).strip()
            # ── コース別ST順位をd に追加（U=21列〜Z=26列、fill_newspaper.pyと同一ロジック）──
            # ヘッダ名に依存せず列番号で直接取得し、固定キー "ST順位_コースN" で格納する。
            for _course, _col in enumerate(range(21, 27), start=1):
                if _col - 1 < len(row):
                    _v = row[_col - 1]
                    if _v is not None:
                        try:
                            d[f"ST順位_コース{_course}"] = float(_v)
                        except (TypeError, ValueError):
                            pass
            player_master[name] = d
            # 改善D: 5文字名の先頭4文字エイリアスを登録時点で追加（O(N)フォールバックループを除去）
            # 【修正】エイリアスdictの「選手名」はフルネームのまま保持する。
            # （例: 大豆生田蒼 → エイリアスキー(大豆生田)でも選手名は大豆生田蒼）
            if len(name) >= 5:
                alias = name[:4]
                if alias not in player_master:
                    alias_d = dict(d)
                    alias_d["選手名"] = name  # フルネームを明示保持
                    player_master[alias] = alias_d

    # イン逃げ分析（1行=1会場、枠別2着率が横展開）
    ws_i = wb[_sheet_ininage]
    ininage_rows = list(ws_i.iter_rows(values_only=True))
    header_row_i = None
    for i, row in enumerate(ininage_rows):
        if row and row[0] == "会場名":
            header_row_i = i
            break
    headers_i = ininage_rows[header_row_i] if header_row_i is not None else []
    # ininage_master: {会場名: {"2nd": {"2": 率,...}, "3rd": {"2": 率,...}}} に変換
    # "2nd" = 枠別イン逃げ時2着率（circle_pct算出用）
    # "3rd" = 枠別イン逃げ時3着以内率（idx3独立算出用）
    ininage_master = {}
    if header_row_i is not None:
        for row in ininage_rows[header_row_i+1:]:
            if not row or row[0] is None:
                continue
            d = dict(zip(headers_i, row))
            v = str(d.get("会場名", "")).strip()
            frame_map_2nd = {}
            frame_map_3rd = {}
            for waku_no in range(1, 7):
                val_2nd = d.get(f"{waku_no}枠\n2着率")
                if val_2nd is not None:
                    frame_map_2nd[str(waku_no)] = float(val_2nd)
                val_3rd = d.get(f"{waku_no}枠\n3着以内率")
                if val_3rd is not None:
                    frame_map_3rd[str(waku_no)] = float(val_3rd)
            # 後方互換: ininage_master[v] をそのままdictとして参照していた箇所のため
            # 旧形式（2着率のみの flat dict）も維持しつつ "3rd" キーを追加する
            ininage_master[v] = dict(frame_map_2nd)      # 旧形式互換（2着率フラット）
            ininage_master[v]["_3rd"] = frame_map_3rd    # 3着以内率（枠番→率）

    # 会場統計（イン逃げ率・決まり手場平均）
    # ──────────────────────────────────────────────────────────────────
    # 【キー名正規化】Excelヘッダは改行あり（例: "1C\n1着率"、"1C\n1R"）だが
    # load_race.py 内では "1コース1着率"、"1C_1R1着率" を参照していたため
    # 全て不一致だった。ここで読み込み時に両方のキー形式を登録する。
    # ──────────────────────────────────────────────────────────────────
    venue_stats_master = {}
    if "会場統計" in wb.sheetnames:
        ws_vs = wb["会場統計"]
        vs_rows = list(ws_vs.iter_rows(values_only=True))
        header_vs = None
        for i, row in enumerate(vs_rows):
            if row and row[0] == "会場名":
                header_vs = i
                break
        if header_vs is not None:
            headers_vs = vs_rows[header_vs]
            for row in vs_rows[header_vs+1:]:
                if not row or row[0] is None:
                    continue
                d = dict(zip(headers_vs, row))
                v = str(d.get("会場名", "")).strip()

                # ── キー名エイリアスを追加登録 ──
                normalized = dict(d)
                for h, val in d.items():
                    if h is None:
                        continue
                    h_str = str(h)
                    # "XC\n1着率" → "Xコース1着率" と "XC_1着率"
                    # re はトップレベルで import 済みのため ループ内 import 不要
                    m = re.match(r"^(\d)C\n1着率$", h_str)
                    if m:
                        c = m.group(1)
                        normalized[f"{c}コース1着率"]  = val
                        normalized[f"{c}C_1着率"]      = val
                    # "XC\nYR" → "XC_YR1着率"（R×C別1着率）
                    m2 = re.match(r"^(\d)C\n(\d+)R$", h_str)
                    if m2:
                        c, r_no = m2.group(1), m2.group(2)
                        normalized[f"{c}C_{r_no}R1着率"] = val
                    # "XR\n荒れ" → "XR荒れスコア"
                    m3 = re.match(r"^(\d+)R\n荒れ$", h_str)
                    if m3:
                        r_no = m3.group(1)
                        normalized[f"{r_no}R荒れスコア"] = val
                    # "まくり\n差し率" → "まくり差し率"
                    if h_str == "まくり\n差し率":
                        normalized["まくり差し率"] = val

                venue_stats_master[v] = normalized

    # ── 会場別コースマスタ（選手×会場×コース の3次元実績）──
    # CSVキャッシュが存在すればそちらを優先（openpyxl iter_rowsより5〜10倍速）
    # CSVがなければ従来通りExcelシートから読み込む（後方互換）
    venue_course_master = {}
    if _vc_csv_path.exists():
        try:
            df_vc = pd.read_parquet(str(_vc_csv_path))
            df_vc = df_vc.astype(str).replace("nan", "")
            # 列名を load_race.py 側が期待するキー名に正規化
            if "進入コース" in df_vc.columns and "コース" not in df_vc.columns:
                df_vc = df_vc.rename(columns={"進入コース": "コース"})
            # iterrows()は91,491行で遅いため to_dict(orient="records") で一括変換
            for row in df_vc.to_dict(orient="records"):
                name   = str(row.get("選手名", "") or "").strip()
                kaijo  = str(row.get("会場名", "") or "").strip()
                course = str(row.get("コース",  "") or "").strip()
                if not name or not kaijo or not course:
                    continue
                key = (name, kaijo, course)
                venue_course_master[key] = row
                # 5文字名の4文字エイリアス
                # 【修正】エイリアスdictの「選手名」はフルネームのまま保持する
                if len(name) >= 5:
                    alias_key = (name[:4], kaijo, course)
                    if alias_key not in venue_course_master:
                        alias_row = dict(row)
                        alias_row["選手名"] = name  # フルネームを明示保持
                        venue_course_master[alias_key] = alias_row
            print(f"  ? 会場別コースマスタ読込（Parquet）: {len(venue_course_master):,}件")
        except Exception as e:
            print(f"  [!]  会場別コースマスタParquet読み込み失敗、Excelにフォールバックします: {e}")
            venue_course_master = {}

    # CSVが存在しないか読み込み失敗 → Excelシートから読む（従来通り）
    if not venue_course_master:
        if "会場別コースマスタ" in wb.sheetnames:
            ws_vc = wb["会場別コースマスタ"]
            vc_rows = list(ws_vc.iter_rows(values_only=True))
            header_vc = None
            for i, row in enumerate(vc_rows):
                if row and row[0] == "選手名":
                    header_vc = i
                    break
            if header_vc is not None:
                headers_vc = vc_rows[header_vc]
                for row in vc_rows[header_vc+1:]:
                    if not row or row[0] is None:
                        continue
                    d = dict(zip(headers_vc, row))
                    name   = str(d.get("選手名", "")).strip()
                    kaijo  = str(d.get("会場名", "")).strip()
                    course = str(d.get("コース", "")).strip()
                    key = (name, kaijo, course)
                    venue_course_master[key] = d
                    if len(name) >= 5:
                        alias_key = (name[:4], kaijo, course)
                        if alias_key not in venue_course_master:
                            alias_d = dict(d)
                            alias_d["選手名"] = name  # フルネームを明示保持
                            venue_course_master[alias_key] = alias_d
            print(f"  ? 会場別コースマスタ読込（Excel）: {len(venue_course_master):,}件")
        else:
            print("  [!]  会場別コースマスタ未作成。update_master.py を実行してください。")


    # ── 展開別残存マスタ読み込み ─────────────────────────────────────────────
    # 会場別: キー=(会場名, 決まり手, 1着コース, 進入コース) → row_dict
    # 全国版: キー=(決まり手, 1着コース, 進入コース)         → row_dict
    # ※ 旧設計では (決まり手, 1着コース) をキーとしていたため、同一シナリオの
    #   進入コース別行（5行）が最後の1行しか残らなかった。キーに進入コースを
    #   追加して全行を保持するよう修正（v6）。
    tenkai_venue_master    = {}
    tenkai_national_master = {}
    for csv_path, master_dict, key_cols in [
        (_tenkai_venue_path,  tenkai_venue_master,    ["会場名", "決まり手", "1着コース", "進入コース"]),
        (_tenkai_nat_path,    tenkai_national_master, ["決まり手", "1着コース", "進入コース"]),
    ]:
        if csv_path.exists():
            try:
                _df = pd.read_parquet(str(csv_path))
                _df = _df.astype(str).replace("nan", "")
                # iterrows()は遅いため to_dict(orient="records") で一括変換
                for row in _df.to_dict(orient="records"):
                    key = tuple(str(row.get(c, "") or "").strip() for c in key_cols)
                    master_dict[key] = row
                print(f"  ? 展開別残存マスタ読込: {csv_path.name} {len(master_dict):,}件")
            except Exception as e:
                print(f"  [!]  展開別残存マスタ読み込み失敗 ({csv_path.name}): {e}")

    # ── CSVが存在しない場合はExcelシートから直接読み込む（v6フォールバック）──
    # update_master.py 未実行の環境でもマスタが使えるよう、Excelシートを優先参照する。
    # CSVが既に読み込まれている場合はスキップ。
    _TENKAI_SHEET_MAP = [
        ("展開別残存_全国",  tenkai_national_master, ["決まり手", "1着コース", "進入コース"]),
        ("展開別残存_会場別", tenkai_venue_master,    ["会場名", "決まり手", "1着コース", "進入コース"]),
    ]
    for sheet_name, master_dict, key_cols in _TENKAI_SHEET_MAP:
        if master_dict:
            continue  # CSV読み込み済みはスキップ
        if sheet_name not in wb.sheetnames:
            continue
        try:
            ws_tk = wb[sheet_name]
            tk_rows = list(ws_tk.iter_rows(values_only=True))
            # ヘッダ行を探す（最初の非Noneセルが「決まり手」か「会場名」の行）
            hdr_idx = None
            for i, row in enumerate(tk_rows):
                cells = [c for c in row if c is not None]
                if cells and str(cells[0]).strip() in ("決まり手", "会場名"):
                    hdr_idx = i
                    break
            if hdr_idx is None:
                continue
            headers_tk = [str(h).strip() if h is not None else "" for h in tk_rows[hdr_idx]]
            for row in tk_rows[hdr_idx + 1:]:
                if not row or all(c is None for c in row):
                    continue
                d = {headers_tk[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
                # 必須キー列が欠けている行はスキップ
                if any(not d.get(c) for c in key_cols):
                    continue
                key = tuple(d.get(c, "") for c in key_cols)
                master_dict[key] = d
            print(f"  ? 展開別残存マスタ(Excel): {sheet_name} {len(master_dict):,}件")
        except Exception as e:
            print(f"  [!]  展開別残存マスタ(Excel)読み込み失敗 ({sheet_name}): {e}")

    # ── ② ST差×決まり手 閾値マスタ読み込み ──────────────────────────────────
    # キー: (進入コース_str, 決まり手, ST差帯) → row_dict
    # update_master.py 実行後に data/st_kimete_threshold.csv が生成される。
    # 未生成の場合は空dictのままにして呼び出し側でフォールバックする。
    st_kimete_master: dict = {}
    if ST_KIMETE_CSV.exists():
        try:
            _df_stk = pd.read_parquet(str(ST_KIMETE_CSV))
            _df_stk = _df_stk.astype(str).replace("nan", "")
            for row in _df_stk.to_dict(orient="records"):
                _c   = str(row.get("進入コース", "") or "").strip()
                _k   = str(row.get("決まり手",   "") or "").strip()
                _b   = str(row.get("ST差帯",     "") or "").strip()
                if _c and _k and _b:
                    st_kimete_master[(_c, _k, _b)] = row
            print(f"  ✓ ST×決まり手閾値マスタ読込: {len(st_kimete_master):,}件")
        except Exception as e:
            print(f"  [!]  ST×決まり手閾値マスタ読み込み失敗: {e}")
    else:
        print("  [!]  ST×決まり手閾値マスタ未生成 (update_master.py を実行してください)")

    # ── ③ コース開放連鎖マスタ読み込み ────────────────────────────────────────
    # キー: (会場名, 攻撃コース_str, 決まり手, 進入コース_str) → row_dict  [会場別]
    #       (攻撃コース_str, 決まり手, 進入コース_str)           → row_dict  [全国]
    kaiho_venue_master:    dict = {}
    kaiho_national_master: dict = {}
    for _csv_path, _master, _key_cols in [
        (KAIHO_VENUE_CSV,    kaiho_venue_master,    ["会場名", "攻撃コース", "決まり手", "進入コース"]),
        (KAIHO_NATIONAL_CSV, kaiho_national_master, ["攻撃コース", "決まり手", "進入コース"]),
    ]:
        if _csv_path.exists():
            try:
                _df_kh = pd.read_parquet(str(_csv_path))
                _df_kh = _df_kh.astype(str).replace("nan", "")
                for row in _df_kh.to_dict(orient="records"):
                    _key = tuple(str(row.get(c, "") or "").strip() for c in _key_cols)
                    if all(_key):
                        _master[_key] = row
                print(f"  ✓ コース開放連鎖マスタ読込 ({_csv_path.name}): {len(_master):,}件")
            except Exception as e:
                print(f"  [!]  コース開放連鎖マスタ読み込み失敗 ({_csv_path.name}): {e}")
        else:
            print(f"  [!]  コース開放連鎖マスタ未生成 ({_csv_path.name})")

    # ── グレード別マスタのブレンド処理 ────────────────────────────────────────
    # race_grade が G1/SG または G2/G3 の場合、一般戦マスタとグレードマスタを
    # G1出走数に応じた比率でブレンドして course_master / venue_course_master を補正する。
    #
    # ブレンド比率（G1/G2出走数ベース）:
    #   50走以上  → 一般0.6 : グレード0.4
    #   20〜49走  → 一般0.7 : グレード0.3
    #   20走未満  → 一般0.9 : グレード0.1
    #   データなし → 一般1.0 : グレード0.0（一般戦マスタのみ）
    if race_grade != "一般":
        # ══════════════════════════════════════════════════════════════
        # 指数ごとのブレンド方針
        # ══════════════════════════════════════════════════════════════
        #
        # [NG] ブレンドしない（グレードマスタが閾値以上あればグレード優先、なければ一般戦をそのまま使う）
        #   1C1着率・3連対率・イン逃げ率・決まり手%
        #   → 相手レベルで意味が全く変わる。一般戦の「逃げ率」とG1の「逃げ率」は別物。
        #
        # [OK] ブレンドしてよい（体の癖・直近調子など相手によらない指数）
        #   STタイミング・フォーム指数・時系列補正値（直近調子）
        #   → 選手固有の傾向なのでグレードに関係なく参考になる。
        #
        # 閾値: グレード出走数 >= 30走 → グレードマスタをそのまま使用
        #        グレード出走数 <  30走 → 一般戦マスタをそのまま使用（補正なし）
        #                                  ＋「グレード初出場注意」フラグ
        # ══════════════════════════════════════════════════════════════

        # ブレンドしてよいキー（STのみ。フォーム指数は player_master 側なので除外）
        _BLEND_KEYS_COURSE = [
            "コース別平均ST",
            "時系列補正1着率", "時系列補正3連対率",  # 直近調子は参考にする
        ]
        _BLEND_KEYS_VC = [
            "平均ST",
            "時系列補正1着率", "時系列補正3連対率",
        ]

        # グレードマスタを優先使用する閾値（これ以上の出走数があればグレードデータを信頼）
        _GRADE_TRUST_THRESHOLD = 30

        def _blend_ratio(n_grade):
            """
            G1出走数 → (一般戦比率, グレード比率)
            ブレンド対象は ST・時系列補正のみ。
            閾値未満は一般戦100%（グレードデータ不足）。
            閾値以上はST等をわずかにグレード側に寄せる。
            """
            try:
                n = int(float(n_grade)) if n_grade not in (None, "", "nan") else 0
            except (ValueError, TypeError):
                n = 0
            if n >= _GRADE_TRUST_THRESHOLD:
                return 0.4, 0.6   # ST等：一般4割・グレード6割
            else:
                return 1.0, 0.0   # データ不足：一般戦100%

        def _grade_trusted(n_grade):
            """グレードマスタを信頼できるか（閾値以上の出走数があるか）"""
            try:
                n = int(float(n_grade)) if n_grade not in (None, "", "nan") else 0
            except (ValueError, TypeError):
                n = 0
            return n >= _GRADE_TRUST_THRESHOLD

        def _safe_float(v):
            try:
                return float(v) if v not in (None, "", "nan") else None
            except (ValueError, TypeError):
                return None

        def _blend_val(v_ippan, v_grade, w_ippan, w_grade):
            """2値をブレンド。どちらかがNoneなら存在する方を使う"""
            vi = _safe_float(v_ippan)
            vg = _safe_float(v_grade)
            if vi is not None and vg is not None:
                return round(vi * w_ippan + vg * w_grade, 4)
            return vi if vi is not None else vg

        # ── 一般戦コース別マスタを追加読み込み（ブレンド用） ──────────────
        _ippan_course_master = {}
        try:
            _ws_ippan = wb[SHEET_MASTER]
            _rows_i   = list(_ws_ippan.iter_rows(values_only=True))
            _hdr_i    = next((i for i, r in enumerate(_rows_i) if r and r[0] == "選手名"), None)
            if _hdr_i is not None:
                _hdrs_i = _rows_i[_hdr_i]
                for _row in _rows_i[_hdr_i + 1:]:
                    if not _row or _row[0] is None:
                        continue
                    _d = dict(zip(_hdrs_i, _row))
                    _nm = str(_d.get("選手名", "")).strip()
                    _cs = str(_d.get("コース",  "")).strip()
                    _ippan_course_master[(_nm, _cs)] = _d
                    if len(_nm) >= 5:
                        _ak = (_nm[:4], _cs)
                        if _ak not in _ippan_course_master:
                            _ippan_course_master[_ak] = _d
            print(f"  ? 一般戦コース別マスタ（ブレンド用）: {len(_ippan_course_master):,}件")
        except Exception as _e:
            print(f"  [!]  一般戦マスタ読み込み失敗（ブレンドなしで続行）: {_e}")

        # ── 一般戦 venue_course_master をCSVから追加読み込み ──────────────
        _ippan_vc_master = {}
        try:
            if VENUE_COURSE_CSV.exists():
                _df_vc_i = pd.read_parquet(str(VENUE_COURSE_CSV))
                _df_vc_i = _df_vc_i.astype(str).replace("nan", "")
                if "進入コース" in _df_vc_i.columns and "コース" not in _df_vc_i.columns:
                    _df_vc_i = _df_vc_i.rename(columns={"進入コース": "コース"})
                for _row in _df_vc_i.to_dict(orient="records"):
                    _nm  = str(_row.get("選手名", "") or "").strip()
                    _kj  = str(_row.get("会場名", "") or "").strip()
                    _cs  = str(_row.get("コース",  "") or "").strip()
                    if not _nm or not _kj or not _cs:
                        continue
                    _ippan_vc_master[(_nm, _kj, _cs)] = _row
                    if len(_nm) >= 5:
                        _ak = (_nm[:4], _kj, _cs)
                        if _ak not in _ippan_vc_master:
                            _ippan_vc_master[_ak] = _row
            print(f"  ? 一般戦会場別コースマスタ（ブレンド用）: {len(_ippan_vc_master):,}件")
        except Exception as _e:
            print(f"  [!]  一般戦会場別コースマスタ読み込み失敗: {_e}")

        # ── course_master を閾値切り替え方式で更新 ─────────────────────────
        # 閾値以上（グレード信頼）:
        #   [NG] 混ぜない指数（1着率・3連対率・決まり手%・イン逃げ率）→ グレードマスタをそのまま使用
        #   [OK] 混ぜてよい指数（ST・時系列補正）→ 一般戦とグレードをブレンド
        # 閾値未満（データ不足）:
        #   全指数 → 一般戦マスタで上書き ＋「グレード初出場注意」フラグを付与
        _blended_count = 0
        _ippan_fallback_count = 0
        for _key, _gd in course_master.items():
            _id = _ippan_course_master.get(_key)
            _n_grade = _safe_float(_gd.get("出走数"))
            _trusted = _grade_trusted(_n_grade)

            if not _trusted:
                # データ不足 → 一般戦マスタで完全上書き（フラグ付き）
                if _id is not None:
                    _fallback = dict(_id)
                    _fallback["_grade_data_shortage"] = True   # 予想時に注意フラグとして参照可
                    course_master[_key] = _fallback
                    _ippan_fallback_count += 1
                # 一般戦データもなければグレードマスタのまま（初出場選手等）
                continue

            # 閾値以上: STのみブレンド、それ以外はグレードマスタ優先
            if _id is None:
                continue  # 一般戦データなし → グレードマスタのまま
            _wi, _wg = _blend_ratio(_n_grade)
            _blended = dict(_gd)  # グレードマスタをベースにコピー
            for _k in _BLEND_KEYS_COURSE:
                _bv = _blend_val(_id.get(_k), _gd.get(_k), _wi, _wg)
                if _bv is not None:
                    _blended[_k] = _bv
            course_master[_key] = _blended
            _blended_count += 1

        print(f"  ? course_master: グレード優先={_blended_count:,}件 / 一般戦フォールバック={_ippan_fallback_count:,}件")

        # ── venue_course_master を閾値切り替え方式で更新 ──────────────────
        _blended_vc = 0
        _ippan_fb_vc = 0
        for _key, _gd in venue_course_master.items():
            _id = _ippan_vc_master.get(_key)
            _n_grade = _safe_float(_gd.get("出走数"))
            _trusted = _grade_trusted(_n_grade)

            if not _trusted:
                if _id is not None:
                    _fallback = dict(_id)
                    _fallback["_grade_data_shortage"] = True
                    venue_course_master[_key] = _fallback
                    _ippan_fb_vc += 1
                continue

            if _id is None:
                continue
            _wi, _wg = _blend_ratio(_n_grade)
            _blended = dict(_gd)
            for _k in _BLEND_KEYS_VC:
                _bv = _blend_val(_id.get(_k), _gd.get(_k), _wi, _wg)
                if _bv is not None:
                    _blended[_k] = _bv
            venue_course_master[_key] = _blended
            _blended_vc += 1

        print(f"  ? venue_course_master: グレード優先={_blended_vc:,}件 / 一般戦フォールバック={_ippan_fb_vc:,}件")

        # ── グレードマスタにしか存在しない選手はそのまま残す（フォールバック済み）──
        # ── 一般戦にしかいない選手も course_master に追加（出場しない可能性高いが念のため）──
        for _key, _id in _ippan_course_master.items():
            if _key not in course_master:
                course_master[_key] = _id
        for _key, _id in _ippan_vc_master.items():
            if _key not in venue_course_master:
                venue_course_master[_key] = _id

    return (course_master, player_master, ininage_master, venue_stats_master,
            venue_course_master, tenkai_venue_master, tenkai_national_master,
            st_kimete_master, kaiho_venue_master, kaiho_national_master,
            _effective_grade)

# ============================================================
# CSVからレースデータ読み込み
# ============================================================
def load_csv(venue, race_no=None, date_str=None):
    """csv_output/会場名/ サブフォルダ（優先）または scripts/csv_output/ から会場CSVを読み込む"""
    # 会場サブフォルダを優先検索、なければルートにフォールバック
    search_dirs = [CSV_DIR / venue, CSV_DIR]

    def find_files(pattern_name):
        for d in search_dirs:
            files = sorted(glob.glob(str(d / pattern_name)))
            if files:
                return files
        return []

    if date_str:
        files = find_files(f"{venue}_{date_str}.csv")
        if not files:
            print(f"  [NG] 指定日付のCSVが見つかりません: {venue}_{date_str}.csv")
            return None, None
        csv_path = files[0]
        print(f"  ? 日付指定: {csv_path}")
    else:
        files = find_files(f"{venue}_*.csv")
        if not files:
            # 最終手段: csv_output/ 全体から最新を検索
            all_files = sorted(glob.glob(str(CSV_DIR / "**" / "*.csv"), recursive=True))
            if not all_files:
                return None, None
            files = [all_files[-1]]
            venue_detected = os.path.basename(files[0]).split("_")[0]
            print(f"  ? 自動検出: {os.path.basename(files[0])} (会場: {venue_detected})")
        csv_path = files[-1]  # 最新

    date_match = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(csv_path))
    race_date = date_match.group(1) if date_match else datetime.today().strftime("%Y/%m/%d")

    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
    df.columns = df.columns.str.strip()
    for col in df.columns:
        df[col] = df[col].astype(str).str.lstrip("'").str.strip()

    # CSVの「日付」列がある場合はそちらを優先（ファイル名より確実）
    if "日付" in df.columns:
        date_vals = df["日付"].dropna()
        date_vals = date_vals[date_vals.str.strip() != ""]
        if len(date_vals) > 0:
            raw_date = date_vals.iloc[0].strip()
            date_from_csv = re.sub(r"(\d{4})[/\-](\d{2})[/\-](\d{2}).*", r"\1-\2-\3", raw_date)
            if re.match(r"\d{4}-\d{2}-\d{2}", date_from_csv):
                race_date = date_from_csv

    if race_no:
        if "レース番号" in df.columns:
            df = df[df["レース番号"].astype(str) == str(race_no)]
        elif "R" in df.columns:
            df = df[df["R"].astype(str) == str(race_no)]

    return df, race_date

# ============================================================
# 指数計算
# ============================================================
