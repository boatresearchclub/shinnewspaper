# -*- coding: utf-8 -*-
"""
lr_excel.py  ─  Excel書き込み / ST舟図 / 数値シート
分割元: load_race.py
※ clone_sample_layout / write_race_to_sample_layout (DEPRECATED) は削除済み。
"""
import re, sys, glob, pathlib
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
from lr_config import SHEET_OUTPUT, SHEET_SAMPLE, ROWS_PER_RACE, BOAT_COLORS, BOAT_BG_LIGHT, BASE_DIR
from lr_utils import safe_float, make_fill, make_font, center_align, left_align, thin_border, write_cell
from lr_calc import _generate_tenkai_story, _generate_buy_hint

try:
    from evaluate_jizen import evaluate_all, evaluate_all_with_scores, calculate_diversity_rate
    JIZEN_AVAILABLE = True
except ImportError:
    JIZEN_AVAILABLE = False

# ──────────────────────────────────────────────────────────────────────
# 【削除済み】clone_sample_layout      (DEPRECATED・main未使用)
# 【削除済み】write_race_to_sample_layout (DEPRECATED・main未使用)
# ──────────────────────────────────────────────────────────────────────

def _make_st_boat_chart(results, player_master, outpath):
    """
    各選手の出走コース平均STをもとに、横位置でST順位を表現した舟図を生成。
    速い（ST値が小さい）ほど右側に配置。
    上から1号艇〜6号艇の固定順。
    Pillowのみ使用（matplotlibは不要）。
    """
    if not PIL_AVAILABLE:
        print("  [!]  Pillowが未インストールのためST舟図をスキップします。")
        return

    # 画像サイズ（EMU 865179 x 2082127 → 96dpi換算で約91x220px、220dpiで約204x490px）
    W, H = 204, 490

    # フォント
    font_candidates = [
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    fnt_path = next((f for f in font_candidates if os.path.exists(f)), None)
    fnt_size = max(int(H / 6 * 0.38), 10)
    fnt = ImageFont.truetype(fnt_path, fnt_size) if fnt_path else ImageFont.load_default()

    BOAT_STYLES = [
        ('#FFFFFF', '#333333', '#000000'),  # 1白
        ('#1a1a1a', '#555555', '#ffffff'),  # 2黒
        ('#CC0000', '#990000', '#ffffff'),  # 3赤
        ('#2255CC', '#1144AA', '#ffffff'),  # 4青
        ('#DDCC00', '#BBAA00', '#000000'),  # 5黄
        ('#115511', '#003300', '#ffffff'),  # 6緑
    ]

    # 各選手の出走コース平均STを取得
    st_col_keys = [
        '平均ST\n(1コース)', '平均ST\n(2コース)', '平均ST\n(3コース)',
        '平均ST\n(4コース)', '平均ST\n(5コース)', '平均ST\n(6コース)',
    ]
    boats = []
    for res in results[:6]:
        waku = int(res['waku']) if str(res['waku']).isdigit() else (len(boats) + 1)
        name = res.get('name_norm', res.get('name', '').replace('　', '').replace(' ', ''))
        course = str(res.get('course', str(waku))).strip()
        pm = player_master.get(name, {})
        st = None
        try:
            cidx = int(course) - 1
            if 0 <= cidx < 6:
                raw = pm.get(st_col_keys[cidx])
                if raw is not None:
                    v = str(raw).replace('%', '').strip()
                    if v not in ('', 'None', 'nan', '-'):
                        st = float(v)
        except Exception as e:
            print(f"  [!]  ST値取得エラー（{name} / コース{course}）: {e}")
        boats.append({'boat': waku, 'st': st})

    # ST値の範囲
    st_vals = [b['st'] for b in boats if b['st'] is not None]
    st_min = min(st_vals) if st_vals else 0.15
    st_max = max(st_vals) if st_vals else 0.20
    st_range = max(st_max - st_min, 0.001)

    # レイアウト
    slot_h = H / 6
    bh = slot_h * 0.55          # 舟の高さ
    bw = W * 0.52               # 舟の長さ
    x_fast = W - bw / 2 - 4    # 速い=右端
    x_slow = bw / 2 + 4        # 遅い=左端

    img = PILImage.new('RGB', (W, H), 'white')
    draw = ImageDraw.Draw(img)

    def hex2rgb(h):
        h = h.lstrip('#')
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

    def draw_boat_pil(draw, cx, cy, bw, bh, fill, outline, num, tc, fnt):
        pts = [
            (cx + bw * 0.50, cy),
            (cx + bw * 0.22, cy - bh * 0.45),
            (cx - bw * 0.32, cy - bh * 0.45),
            (cx - bw * 0.50, cy - bh * 0.20),
            (cx - bw * 0.50, cy + bh * 0.20),
            (cx - bw * 0.32, cy + bh * 0.45),
            (cx + bw * 0.22, cy + bh * 0.45),
        ]
        draw.polygon(pts, fill=hex2rgb(fill), outline=hex2rgb(outline))
        draw.text((cx - bw * 0.06, cy), str(num),
                  fill=hex2rgb(tc), font=fnt, anchor='mm')

    for i, b in enumerate(boats):
        fill, outline, tc = BOAT_STYLES[b['boat'] - 1]
        cy = (i + 0.5) * slot_h
        if b['st'] is not None:
            norm = (b['st'] - st_min) / st_range
            cx = x_fast - norm * (x_fast - x_slow)
        else:
            cx = (x_fast + x_slow) / 2
        draw_boat_pil(draw, cx, cy, bw, bh, fill, outline, b['boat'], tc, fnt)

    # 改善C: outpathがNoneの場合はBytesIOに保存してそのまま返す（ディスクI/O削減）
    if outpath is None:
        from io import BytesIO
        buf = BytesIO()
        img.save(buf, format='PNG', dpi=(220, 220))
        buf.seek(0)
        return buf
    img.save(outpath, dpi=(220, 220))



def write_race_flat(ws, row_offset, race_no, venue, race_date,
                    results, slit, venue_stats, frame_2nd, _jizen_members=None,
                    player_master=None, _tmp_image_paths=None, deadline=None,
                    race_judgment=None, bet_suggestions=None):
    """
    出力_新聞サンプルのレイアウトに完全準拠して1レース分を書き込む。
    29行ブロック（Row1-29）＋空行1行 = 30行/レース
    ＋ 考察・買い目ブロック（引き継ぎ書 第10章フォーマット）
    """
    r = row_offset
    c = center_align()
    lft = left_align()
    hf = header_fill()
    hfont = header_font()
    shf = subheader_fill()
    shfont = subheader_font()
    bdr = thin_border()

    def wc(row, col, val, **kwargs):
        write_cell(ws, row, col, val, **kwargs)

    def merge(r1, c1, r2, c2):
        try:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        except Exception as e:
            print(f"  [!]  セル結合スキップ（行{r1}列{c1}〜行{r2}列{c2}）: {e}")

    # ── 結合セル定義（サンプル完全再現） ──
    merge(r+0, 1, r+1, 2)    # A1:B2  会場名
    merge(r+0, 3, r+1, 3)    # C1:C2  イン逃げ場平均
    merge(r+0, 4, r+0, 6)    # D1:F1  決まり手場平均
    merge(r+0, 7, r+0, 8)    # G1:H1  波乱度
    merge(r+0, 9, r+0,15)    # I1:O1  発行日
    merge(r+1, 7, r+1, 8)    # G2:H2
    merge(r+1, 9, r+2,16)    # I2:P3  ボートリサーチ新聞
    merge(r+2, 1, r+2, 2)    # A3:B3  レース番号
    merge(r+2, 7, r+2, 8)    # G3:H3  想定スリット
    merge(r+3, 1, r+3, 4)    # A4:D4  選手名ヘッダ
    merge(r+3, 7, r+3, 8)    # G4:H4  F/St
    merge(r+3, 9, r+3,10)    # I4:J4  想定スリット
    merge(r+3,13, r+3,14)    # M4:N4  オリジナル1着率
    merge(r+3,15, r+3,16)    # O4:P4  一般戦3連対
    for i in range(6):       # 選手行 B:D結合, I:J結合（K列はST図用に開放）, M:N結合, O:P結合
        rr = r+4+i
        merge(rr, 2, rr, 4)
        merge(rr, 9, rr,10)
        merge(rr,13, rr,14)
        merge(rr,15, rr,16)
    merge(r+11, 1, r+11, 6)  # A12:F12 事前評価タイトル
    merge(r+11, 7, r+12,13)  # G12:M13 2着率テキスト
    merge(r+11,14, r+12,14)  # N12:N13 3着指数
    merge(r+11,15, r+12,16)  # O12:P13 オリジナル3連対
    for i in range(6):       # 事前評価 O:P結合
        merge(r+13+i,15, r+13+i,16)
    merge(r+20, 1, r+20, 5)  # A21:E21 決まり手タイトル
    merge(r+20, 6, r+27,11)  # F21:K28 説明テキスト
    merge(r+20,12, r+20,16)  # L21:P21 場平均タイトル
    for i in range(7):       # 場平均 M:N結合, O:P結合
        merge(r+21+i,13, r+21+i,14)
        merge(r+21+i,15, r+21+i,16)
    merge(r+28, 1, r+28,16)  # A29:P29 注記

    # ── Row 1-2: ヘッダ ──
    wc(r+0, 1, venue,  font=Font(name="Noto Sans CJK SC", size=12, bold=True), alignment=c)
    wc(r+0, 3, 'イン\n逃げ\n場平均', fill=hf, font=Font(name="ＭＳ ゴシック", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+0, 4, '決まり手場平均', fill=hf, font=hfont, alignment=c)
    # 波乱度セルに会場の荒れやすさスコアを表示
    areyasusa = venue_stats.get('areyasusa_score')
    haranddo_str = f'波乱度\n{areyasusa:.0f}pt' if areyasusa is not None else '波乱度'
    wc(r+0, 7, haranddo_str, fill=hf, font=hfont, alignment=c)
    in_rate = venue_stats.get('in_rate')
    wc(r+0, 9, race_date, fill=hf, font=Font(name="Meiryo UI", size=11, bold=True, color="FFFFFFFF"),
       alignment=Alignment(horizontal="right", vertical="center", wrap_text=True))
    wc(r+0,16, '発行', fill=hf, font=Font(name="Noto Sans CJK SC", size=11, bold=True, color="FFFFFFFF"), alignment=c)
    # ── タイトル「ボートリサーチ新聞」（I2:P3 結合セルの左上 = col9, row+1） ──
    wc(r+1, 9, 'ボートリサーチ新聞',
       font=Font(name="Noto Sans CJK SC", size=24, bold=True),
       alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

    # 決まり手場平均（会場統計から）
    kimari = venue_stats.get('kimari_avg', {})
    for col, key in [(4,'差し'), (5,'まくり'), (6,'まくり差し')]:
        wc(r+1, col, key, fill=hf, font=hfont, alignment=c)
        v = kimari.get(key)
        vs_str = f'{v*100:.0f}%' if v else '-'
        wc(r+2, col, vs_str, alignment=c)

    # イン逃げ場平均（C列, Row1-2結合）
    in_str = f'{in_rate*100:.1f}%' if in_rate is not None else '-'
    wc(r+2, 3, in_str, alignment=c)

    # ── Row 3: レース番号・スリット ──
    race_no_val = f'{race_no}R\n{deadline}' if deadline else f'{race_no}R'
    wc(r+2, 1, race_no_val, font=Font(name="Meiryo UI", size=14, bold=True), alignment=c)
    wc(r+2, 7, f'想定スリット: {slit}', font=Font(name="Arial", size=13, bold=True), alignment=c)


    # ── Row 4: 列ヘッダ ──
    st1_rank = next((res.get('st_rank') for res in results if res['waku']=='1'), None)
    st1_str  = f'1号艇平均ST\n{st1_rank:.1f}位' if st1_rank else '1号艇平均ST\n-'
    for col, txt in [(1,'選手名'),(5,'組'),(9,'想定\nスリット'),(13,'オリジナル\n1着率'),(15,'コース別\n3連対')]:
        wc(r+3, col, txt, fill=hf, font=Font(name="Noto Sans CJK SC", size=9 if col!=1 else 9, bold=True, color="FFFFFFFF"), alignment=c)
    for col, txt in [(6,'モータ\n2連'),(7,'展示\n偏差値')]:
        wc(r+3, col, txt, fill=hf, font=Font(name="Meiryo UI", size=8, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+3, 11, st1_str, fill=hf, font=Font(name="Meiryo UI", size=8, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+3, 12, '評価', fill=hf, font=Font(name="Meiryo UI", size=10, bold=True, color="FFFFFFFF"), alignment=c)

    # ── Row 5-10: 選手データ ──
    # 艇番の濃い背景色（サンプル準拠）
    color_fills = {
        1: 'FFFFFFFF', 2: 'FF111111', 3: 'FFFF0000',
        4: 'FF00B0F0', 5: 'FFFFFF00', 6: 'FF00B050'
    }
    # 艇番列の文字色
    color_fonts = {
        1: 'FF000000', 2: 'FFFFFFFF', 3: 'FFFFFFFF',
        4: 'FFFFFFFF', 5: 'FF000000', 6: 'FFFFFFFF'
    }
    # 隣接列（B,E,F,G,H,L,M,O）の薄い背景色（サンプル準拠）
    color_fills_light = {
        1: 'FFFFFFFF', 2: 'FF000000', 3: 'FFFFCCCC',
        4: 'FFCCE8FF', 5: 'FFFFFACC', 6: 'FFCCFFEE'
    }
    # 黒艇（2号艇）の隣接列は黒背景だが、サンプルではtheme=0（黒）→実際は黒
    # ただしB列はtheme=0 tint=-0.15→ほぼ黒。文字は白。
    color_fonts_light = {
        1: 'FF000000', 2: 'FFFFFFFF', 3: 'FF000000',
        4: 'FF000000', 5: 'FF000000', 6: 'FF000000'
    }
    # 改善A: Font/Fillオブジェクトをループ外で一括生成してキャッシュ
    _bf_cache       = {w: make_fill(color_fills[w])       for w in range(1, 7)}
    _bf_light_cache = {w: make_fill(color_fills_light[w]) for w in range(1, 7)}
    _ff_waku_cache  = {w: Font(name="Meiryo UI",        size=11, bold=True,  color=color_fonts[w])       for w in range(1, 7)}
    _ff_name_cache  = {w: Font(name="Noto Sans CJK SC", size=11, bold=True,  color=color_fonts_light[w]) for w in range(1, 7)}
    _ff_data_cache  = {w: Font(name="Meiryo UI",        size=10, bold=False, color=color_fonts_light[w]) for w in range(1, 7)}
    _ff_rate_cache  = {w: Font(name="Meiryo UI",        size=11, bold=True,  color=color_fonts_light[w]) for w in range(1, 7)}
    _ff_honmei_cache= {w: Font(name="Noto Sans CJK SC", size=10, bold=False, color=color_fonts_light[w]) for w in range(1, 7)}
    _fill_white     = make_fill('FFFFFFFF')
    _fill_missing   = make_fill("FFFCE4D6")
    _font_missing_rate = Font(name="Meiryo UI", size=10, bold=False, color="FFBF8F00")
    _font_course    = Font(name="Meiryo UI", size=10, bold=False, color='FF000000')
    for i, res in enumerate(results[:6]):
        dr = r + 4 + i
        waku_no = int(res['waku']) if str(res['waku']).isdigit() else i+1
        bf      = _bf_cache[waku_no]
        bf_light = _bf_light_cache[waku_no]
        fc      = color_fonts[waku_no]
        fc_light = color_fonts_light[waku_no]
        ff_waku = _ff_waku_cache[waku_no]
        ff_name = _ff_name_cache[waku_no]
        ff_data = _ff_data_cache[waku_no]
        ff_rate = _ff_rate_cache[waku_no]

        # 選手名: 姓（2文字）＋全角スペース2つ＋名
        # （年齢除去は calc_race_indices で実施済み）
        name = res['name']
        # 全角スペース・半角スペースで分割
        parts = [p for p in re.split(r'[\s\u3000]+', name) if p]
        if len(parts) >= 2:
            name_fmt = parts[0] + '　　' + parts[1]
        else:
            name_fmt = name
        # データ不足の場合は選手名に★を付加
        if res.get('data_missing'):
            name_fmt = '★' + name_fmt
        # 自コース出走数20件未満の場合は選手名の前に[!]を付加
        _jm = (_jizen_members[i] if _jizen_members and i < len(_jizen_members) else None)
        if _jm and _jm.get('star_rate'):
            name_fmt = '[!]' + name_fmt

        rel_w1 = f"{res['rel_win1']:.1f}%" if res['rel_win1'] is not None else '-'
        abs_w3 = f"{res['abs_win3']:.1f}%" if res.get('abs_win3') is not None else '-'

        wc(dr, 1, waku_no,   fill=bf,       font=ff_waku, alignment=c, border=bdr)
        wc(dr, 2, name_fmt,  fill=bf_light,  font=ff_name, alignment=c, border=bdr)
        wc(dr, 5, res['kumi'],   fill=bf_light, font=ff_data, alignment=c, border=bdr)
        wc(dr, 6, res['motor2'], fill=bf_light, font=ff_data, alignment=c, border=bdr)
        # G/H列：展示タイム偏差値（F/St影響列に表示）
        # 【軽微(1)】前日出力では展示未実施のため None → 「-前日」と表示して当日版と区別
        tenji_h = res.get("tenji_hensa")
        tenji_str = f"{tenji_h:.1f}" if tenji_h is not None else "前日"
        wc(dr, 7, tenji_str,     fill=bf_light, font=ff_data, alignment=c, border=bdr)
        wc(dr, 8, None,          fill=bf_light, font=ff_data, alignment=c, border=bdr)
        # I列はサンプルでFFFFFFFF（白）
        wc(dr, 9, res['course'], fill=_fill_white, font=_font_course, alignment=c, border=bdr)
        wc(dr,12, res['honmei'] if res['honmei'].strip() else None,
                                 fill=bf_light, font=_ff_honmei_cache[waku_no], alignment=c, border=bdr)
        # データ不足の場合は薄オレンジ背景に変更
        if res.get('data_missing'):
            rel_w1_disp = "－" if res['rel_win1'] is None else rel_w1
            abs_w3_disp = "－" if res.get('abs_win3') is None else abs_w3
            wc(dr,13, rel_w1_disp, fill=_fill_missing, font=_font_missing_rate, alignment=c, border=bdr)
            wc(dr,15, abs_w3_disp, fill=_fill_missing, font=_font_missing_rate, alignment=c, border=bdr)
        else:
            wc(dr,13, rel_w1,        fill=bf_light, font=ff_rate, alignment=c, border=bdr)
            wc(dr,15, abs_w3,        fill=bf_light, font=ff_data, alignment=c, border=bdr)

    # ── Row 11: 区切り（空行） ──
    sep_fill = make_fill("FFCCCCDD")
    for col in range(1, 17):
        ws.cell(r+10, col).fill = sep_fill

    # ── Row 12-13: 事前評価ヘッダ ──
    wc(r+11, 1, 'ボートリサーチ流  事前評価', fill=hf, font=Font(name="Noto Sans CJK SC", size=11, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+11, 7, 'イン逃げ決着時の2着優位度（相対%）', fill=shf, font=Font(name="Noto Sans CJK SC", size=11, bold=True, color="FF000000"), alignment=c)
    wc(r+11,14, '3着\n指数', fill=hf, font=Font(name="Meiryo UI", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    wc(r+11,15, 'コース別\n3連対', fill=hf, font=Font(name="Meiryo UI", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    for col, txt in [(2,'イン\n逃げ'),(3,'相性'),(4,'機力'),(5,'自在性'),(6,'展開')]:
        wc(r+12, col, txt, fill=shf, font=Font(name="Meiryo UI" if col in (2,6) else "Noto Sans CJK SC", size=9, bold=True, color="FF000000"), alignment=c)

    # ── Row 14-19: 事前評価データ ──
    # 2着率テキストブロック（G14結合）
    circle_labels = {2:'黒', 3:'赤', 4:'青', 5:'黄', 6:'緑'}
    # 【重大(2)明確化】circle_pctはレース内正規化済みの「相対シェア%」
    # 「絶対的な2着確率」ではなく「このレースで2着になる相対優位度」を示す
    circle_lines = ['2着優位度（イン逃げ決着時・レース内相対%）', '']
    # circle_pct は calc_race_indices でレース内正規化済み
    circle_pct_map = {res['waku']: res.get('circle_pct') for res in results}
    for wk in range(2, 7):
        pct = circle_pct_map.get(str(wk))
        pct_str = f'{pct:>4.0f}%' if pct is not None else '   -%'
        circle_lines.append(f'{circle_labels[wk]}({wk}枠): {pct_str}')
    wc(r+13, 7, '\n'.join(circle_lines), alignment=left_align(wrap=True))

    # 事前評価記号
    jizen_result = None
    if JIZEN_AVAILABLE and _jizen_members:
        try:
            jizen_result = evaluate_all(_jizen_members)
        except Exception as e:
            print(f"  [!]  事前評価（write_race_flat内）でエラーが発生しました: {e}")

    # 2着率テキストブロック: results の circle_pct（正規化済み）を使用
    # total_2nd は描画用の再計算（frame_2nd は raw スコアなので circle_pct が正値）
    total_2nd = sum(v for k, v in frame_2nd.items() if k != "1" and v is not None)

    # 事前評価行・決まり手行の薄い背景色（サンプル準拠）
    jizen_light_fills = {
        1: 'FFFFFFFF', 2: 'FF000000', 3: 'FFFFCCCC',
        4: 'FFCCE8FF', 5: 'FFFFFACC', 6: 'FFCCFFEE'
    }
    jizen_light_fonts = {
        1: 'FF000000', 2: 'FFFFFFFF', 3: 'FF000000',
        4: 'FF000000', 5: 'FF000000', 6: 'FF000000'
    }
    # 改善A: 事前評価・決まり手ループ用キャッシュ
    _bf_j_cache    = {w: make_fill(jizen_light_fills[w]) for w in range(1, 7)}
    _ff_j11_cache  = {w: Font(name="Meiryo UI", size=11, bold=True,  color=color_fonts[w])       for w in range(1, 7)}
    _ff_j11l_cache = {w: Font(name="Meiryo UI", size=11, bold=True,  color=jizen_light_fonts[w]) for w in range(1, 7)}
    _ff_j10l_cache = {w: Font(name="Meiryo UI", size=10, bold=False, color=jizen_light_fonts[w]) for w in range(1, 7)}
    _fill_m5       = make_fill('FFE5DFEC')  # 黄艇M/N列専用
    _font_red      = Font(name="Meiryo UI", size=10, bold=False, color="FFCC0000")
    _align_cwrap   = center_align(wrap=True)
    for i, res in enumerate(results[:6]):
        dr = r + 13 + i
        waku_no = int(res['waku']) if str(res['waku']).isdigit() else i+1
        bf   = _bf_cache[waku_no]
        bf_j = _bf_j_cache[waku_no]
        ff   = _ff_j11_cache[waku_no]
        ff_j = _ff_j11l_cache[waku_no]
        wc(dr, 1, waku_no, fill=bf, font=ff, alignment=c, border=bdr)

        # B列: ほぼ白
        wc(dr, 2, None, fill=_fill_white, alignment=c)
        # C/D/E列: 薄い艇番色
        for col in [3, 4, 5]:
            wc(dr, col, None, fill=bf_j, alignment=c)
        # F列: waku1-3は白、waku4-6は薄い色
        f_fill = bf_j if waku_no >= 4 else _fill_white
        wc(dr, 6, None, fill=f_fill, alignment=c)

        # 事前評価記号
        if jizen_result:
            col_map = {'in_nige':2,'aisho':3,'kiryoku':4,'jizaisei':5,'tenkai':6}
            for key, col in col_map.items():
                sym = jizen_result[key][i]
                cell = ws.cell(dr, col)
                cell.value = sym if sym else None

        # N列（3着指数）: 薄い色
        wc(dr, 14, None, fill=bf_j, alignment=c)
        idx3 = res.get('idx3', 0)
        ws.cell(dr, 14).value = idx3 if idx3 else None
        # O/P列（オリジナル3連対）: 薄い色
        wc(dr, 15, None, fill=bf_j, alignment=c)
        abs_w3_ev = f"{res['abs_win3']:.1f}%" if res.get('abs_win3') is not None else '-'
        ws.cell(dr, 15).value = abs_w3_ev

    # ── Row 20: 区切り（空行） ──
    for col in range(1, 17):
        ws.cell(r+19, col).fill = sep_fill

    # ── Row 21: 決まり手タイトル・説明・場平均タイトル ──
    wc(r+20, 1, '決まり手 直近1年\n1=被決まり手 2〜6=決まり手', fill=hf, font=Font(name="Noto Sans CJK SC", size=8, bold=True, color="FFFFFFFF"), alignment=c)
    _exp = (
        'ボートリサーチ流 事前評価（展示前の評価）\n\n'
        '◇イン逃げ … 出走メンバーで相対評価。\n  イン逃げ信頼度 (◎>○>△>×)\n\n'
        '◇相性 … 1枠に対する相性\n  (◎>○>△ ※空白は平凡)\n\n'
        '◇機力 … 出走メンバーで機力を相対評価\n  (A>B>C>D>E)\n\n'
        '◇自在性 … 自ら動きレース展開を作れるか\n  (◎>○>△ ※空白は平凡)\n\n'
        '◇展開 … 4~6枠で展開が向いた時の対応力\n  (◎>○>△ ※空白は平凡)'
    )
    wc(r+20, 6, _exp, alignment=left_align(wrap=True))
    wc(r+20,12, '場平均', fill=hf, font=hfont, alignment=c)

    # ── Row 22: 決まり手ヘッダ・場平均ヘッダ ──
    for col, txt in [(1,'枠'),(2,'逃げ%'),(3,'差し%\n差された%'),(4,'まくり%\n捲られた%'),(5,'まくり差し%\n捲り差された%')]:
        wc(r+21, col, txt, fill=hf, font=Font(name="Noto Sans CJK SC", size=9, bold=True, color="FFFFFFFF"), alignment=c)
    for col, txt in [(12,'枠'),(13,'1-○'),(15,'1着率')]:
        wc(r+21, col, txt, fill=shf, font=Font(name="Noto Sans CJK SC", size=9, bold=True, color="FF000000"), alignment=c)

    # ── Row 23-28: 決まり手データ・場平均データ ──
    for i, res in enumerate(results[:6]):
        dr = r + 22 + i
        waku_no = int(res['waku']) if str(res['waku']).isdigit() else i+1
        bf       = _bf_cache[waku_no]
        bf_light = _bf_j_cache[waku_no]
        ff       = _ff_j11_cache[waku_no]
        ff_light = _ff_j10l_cache[waku_no]
        cm = res.get('raw_cm', {})

        wc(dr, 1, str(waku_no), fill=bf, font=ff, alignment=c, border=bdr)

        # 決まり手%（コース別マスタから）B〜E列に薄い背景色を設定
        def pct_str(key):
            v = safe_float(_get_cm_val(cm, key))
            return f'{v*100:.0f}%' if v else '-'

        # B列: 逃げ% → 1号艇のみ表示、2〜6はハイフン
        if waku_no == 1:
            wc(dr, 2, pct_str('逃げ%'), fill=_fill_white, font=ff_light, alignment=c)
        else:
            wc(dr, 2, '-', fill=_fill_white, font=ff_light, alignment=c)

        # C/D/E列:
        #   1号艇 → 被決まり手%（差された%・捲られた%・捲り差された%）赤字
        #   2〜6号艇 → 決まり手%（差し%・まくり%・まくり差し%）
        if waku_no == 1:
            def lose_pct_str(key):
                v = safe_float(cm.get(key))
                return f'{v*100:.0f}%' if v else '-'
            wc(dr, 3, lose_pct_str('差され%'),
               fill=bf_light, font=_font_red, alignment=_align_cwrap)
            wc(dr, 4, lose_pct_str('捲られ%'),
               fill=bf_light, font=_font_red, alignment=_align_cwrap)
            wc(dr, 5, lose_pct_str('捲り差され%'),  # update_master出力キーに統一
               fill=bf_light, font=_font_red, alignment=_align_cwrap)
        else:
            wc(dr, 3, pct_str('差し%'),       fill=bf_light, font=ff_light, alignment=c)
            wc(dr, 4, pct_str('まくり%'),     fill=bf_light, font=ff_light, alignment=c)
            wc(dr, 5, pct_str('まくり差し%'), fill=bf_light, font=ff_light, alignment=c)

        # 場平均（イン逃げ時2着率）L列は艇番色
        wc(dr,12, str(waku_no), fill=bf, font=ff, alignment=c, border=bdr)
        # circle_pct はレース内正規化済みの値を results から取得
        pct_2nd_norm = res.get('circle_pct', 0) or 0
        # M/N列: 薄い色（黄艇のみT7 = FFE5DFEC）
        m_fill = _fill_m5 if waku_no == 5 else bf_light
        wc(dr,13, f'{pct_2nd_norm:.0f}%' if total_2nd > 0 else '-', fill=m_fill, font=ff_light, alignment=c)
        # 1着率（相対1着率）O/P列
        rel_w1 = f"{res['rel_win1']:.1f}%" if res['rel_win1'] is not None else '-'
        wc(dr,15, rel_w1, fill=m_fill, font=ff_light, alignment=c)

    # ── Row 29: 注記 ＋ ヒモ荒れ判定 ──
    himo_are = (race_judgment or {}).get("himo_are", {}) if race_judgment else {}
    himo_verdict = himo_are.get("verdict", "対象外")
    if himo_verdict == "不参加推奨":
        mcp = himo_are.get("max_combo_prob", 0.0) or 0.0
        eto = himo_are.get("est_top_odds",   0.0) or 0.0
        cc  = himo_are.get("circle_concentration", 0.0) or 0.0
        himo_str = (
            f"【? ヒモ固まり・見送り推奨】"
            f"最有力確率{mcp:.3f}（推定最高人気{eto:.0f}倍台）/ 2着集中度{cc:.0f}%\n"
            f"→ 1号艇1着固定でも3連単オッズが構造的に低い。参加しない方が期待値が高い。\n"
        )
        himo_fill = make_fill("FFFCE4D6")   # 薄オレンジ（警告色）
        himo_font = Font(name="Noto Sans CJK SC", size=9, bold=True, color="FFCC0000")
    elif himo_verdict == "点数絞り":
        mcp = himo_are.get("max_combo_prob", 0.0) or 0.0
        eto = himo_are.get("est_top_odds",   0.0) or 0.0
        cc  = himo_are.get("circle_concentration", 0.0) or 0.0
        himo_str = (
            f"【[!] ヒモやや固め・点数絞り】"
            f"最有力確率{mcp:.3f}（推定最高人気{eto:.0f}倍台）/ 2着集中度{cc:.0f}%\n"
            f"→ ヒモを上位2艇に絞り込むこと。\n"
        )
        himo_fill = make_fill("FFFFFFF0")   # 薄黄（注意色）
        himo_font = Font(name="Noto Sans CJK SC", size=9, bold=False, color="FF996600")
    elif himo_verdict == "参加推奨":
        mcp = himo_are.get("max_combo_prob", 0.0) or 0.0
        eto = himo_are.get("est_top_odds",   0.0) or 0.0
        cc  = himo_are.get("circle_concentration", 0.0) or 0.0
        himo_str = (
            f"【[OK] ヒモ分散・参加推奨】"
            f"最有力確率{mcp:.3f}（推定最高人気{eto:.0f}倍台）/ 2着集中度{cc:.0f}%\n"
            f"→ 1号艇1着固定でヒモ広めに流す。買い目を+2点追加推奨。\n"
        )
        himo_fill = make_fill("FFE2EFDA")   # 薄緑（推奨色）
        himo_font = Font(name="Noto Sans CJK SC", size=9, bold=False, color="FF375623")
    else:
        himo_str  = ""
        himo_fill = None
        himo_font = None

    note_base = (
        '※オリジナル1着率、2着率、3着指数は出走メンバーで相対評価しており、天候、潮、展示、場特性などを考慮していません。\n'
        '※枠内を想定します。進入変更があった場合使用できません。'
    )
    note_val  = (himo_str + note_base) if himo_str else note_base
    note_fill = himo_fill or make_fill("FFFFFFFF")
    note_font = himo_font or Font(name="Meiryo UI", size=8)
    wc(r+28, 1, note_val,
       fill=note_fill, font=note_font,
       alignment=left_align(wrap=True))

    # ── 行高（サンプル完全再現） ──
    row_heights = {
        0:12.75, 1:30.0, 2:24.75, 3:31.5,
        4:27.75, 5:27.75, 6:27.75, 7:27.75, 8:27.75, 9:27.75,
        10:4.5, 11:25.5, 12:21.75,
        13:25.5, 14:25.5, 15:25.5, 16:25.5, 17:25.5, 18:25.5,
        19:4.5, 20:25.5, 21:21.75,
        22:25.5, 23:25.5, 24:25.5, 25:25.5, 26:25.5, 27:25.5,
        28:39.0
    }
    for offset, height in row_heights.items():
        ws.row_dimensions[r + offset].height = height

    # ── ST順位 舟図をK5:K10に埋め込む ──
    # 改善C: tempfileをやめてBytesIOインメモリ処理に変更（ディスクI/O削減）
    if player_master is not None:
        try:
            buf = _make_st_boat_chart(results, player_master, outpath=None)
            if buf is not None:
                img = XLImage(buf)
                # K列(col=10, 0始まり), 選手データ開始行(r+4, 0始まり=r+3)
                marker = AnchorMarker(col=10, colOff=0, row=r+3, rowOff=0)
                size   = XDRPositiveSize2D(cx=865179, cy=2082127)
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
        except Exception as e:
            print(f"  [!]  ST舟図生成エラー ({race_no}R): {e}")

    # ======================================================================
    # 考察・買い目ブロック（引き継ぎ書 第10章フォーマット）
    # 29行ブロックの直後（r+29 以降）に追記する
    # ======================================================================
    bet = bet_suggestions or {}

    # スタイル定数
    def _sf(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def _fn(bold=True, size=9, color="FF000000"):
        return Font(name="Meiryo UI", size=size, bold=bold, color=color)
    def _al(h="left", wrap=True):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    _row = r + 29  # 29行ブロック直後から開始（r+28が最終行、r+29が空行、r+30から考察）
    _row += 1       # 空行1行あけてから考察開始

    skip        = bet.get("skip", False)
    skip_reason = bet.get("skip_reason", "")
    comment     = bet.get("comment", "")
    buy_list    = bet.get("buy_list", [])
    candidates  = bet.get("candidates", [])
    scenario_type = bet.get("scenario_type", "-")
    theory_syn    = bet.get("theory_syn_odds")
    point_count   = bet.get("point_count", len(buy_list))
    margin_verdict = bet.get("margin_verdict", "-")

    # ── 見送り / 参加 ヘッダ ────────────────────────────────────────────
    if skip:
        _hdr_text  = f"[NG] {race_no}R 見送り推奨"
        _hdr_fill  = _sf("FFFF9999")  # 薄赤
        _hdr_font  = _fn(bold=True, size=11, color="FF7F0000")
    else:
        _hdr_text  = f"[OK] {race_no}R 参加"
        _hdr_fill  = _sf("FFE2EFDA")  # 薄緑
        _hdr_font  = _fn(bold=True, size=11, color="FF375623")

    try:
        ws.merge_cells(start_row=_row, start_column=1, end_row=_row, end_column=16)
    except Exception:
        pass
    wc(_row, 1, _hdr_text,
       fill=_hdr_fill, font=_hdr_font,
       alignment=_al(h="center", wrap=False))
    ws.row_dimensions[_row].height = 20.0
    _row += 1

    # ── 見送り理由（見送り時のみ） ────────────────────────────────────────
    if skip and skip_reason:
        # ev_warning_msg（合成オッズ警告）は事前情報では判断不能なため除去
        _sr2_filtered = []
        _sr2_skip_next = False
        for _ln in skip_reason.splitlines():
            if _ln.strip().startswith("[!]") and (
                "合成オッズ" in _ln or "回収率" in _ln or "見送り推奨" in _ln
            ):
                _sr2_skip_next = True
                continue
            if _sr2_skip_next and _ln.strip().startswith("→"):
                _sr2_skip_next = False
                continue
            _sr2_skip_next = False
            _sr2_filtered.append(_ln)
        skip_reason = "\n".join(_sr2_filtered).strip()
        if skip_reason:
            try:
                ws.merge_cells(start_row=_row, start_column=1, end_row=_row, end_column=16)
            except Exception:
                pass
            wc(_row, 1, skip_reason,
               fill=_sf("FFFCE4D6"), font=_fn(bold=False, size=9, color="FF7F0000"),
               alignment=_al())
            ws.row_dimensions[_row].height = max(14.0 * (skip_reason.count("\n") + 1), 28.0)
            _row += 1

    # ── 考察テキスト（Step1〜6） ──────────────────────────────────────────
    # comment は lr_suggest.py の Step E で生成した引き継ぎ書フォーマットのテキスト
    if comment:
        # 【考察】と【買い目】でブロックを分割して書き込む
        _blocks = comment.split("\n\n【買い目】")
        _kosatsu_text = _blocks[0].replace("【考察】\n", "").strip() if _blocks else comment
        _kaimoku_header = _blocks[1].strip() if len(_blocks) > 1 else ""

        # 考察ヘッダ
        try:
            ws.merge_cells(start_row=_row, start_column=1, end_row=_row, end_column=16)
        except Exception:
            pass
        wc(_row, 1, "[考察] 考察",
           fill=_sf("FF203864"), font=_fn(bold=True, size=9, color="FFFFFFFF"),
           alignment=_al(h="center", wrap=False))
        ws.row_dimensions[_row].height = 16.0
        _row += 1

        # 考察本文（Step1〜6を1セルに収める）
        try:
            ws.merge_cells(start_row=_row, start_column=1, end_row=_row, end_column=16)
        except Exception:
            pass
        wc(_row, 1, _kosatsu_text,
           fill=_sf("FFD9E1F2"), font=_fn(bold=False, size=9, color="FF000000"),
           alignment=_al())
        _n_lines = max(_kosatsu_text.count("\n") + 1, 6)
        ws.row_dimensions[_row].height = max(14.0 * _n_lines, 80.0)
        _row += 1

    # ── 買い目ヘッダ ────────────────────────────────────────────────────
    try:
        ws.merge_cells(start_row=_row, start_column=1, end_row=_row, end_column=16)
    except Exception:
        pass
    _syn_str = f" / 理論合成{theory_syn}倍" if theory_syn else ""
    _buy_hdr = (
        f"[表] 買い目【{scenario_type}】 {point_count}点{_syn_str} → {margin_verdict}"
        if not skip else
        f"[表] 参考買い目（見送り推奨）{point_count}点"
    )
    wc(_row, 1, _buy_hdr,
       fill=_sf("FF1F4E79"), font=_fn(bold=True, size=9, color="FFFFFFFF"),
       alignment=_al(h="center", wrap=False))
    ws.row_dimensions[_row].height = 16.0
    _row += 1

    # ── 買い目リスト ─────────────────────────────────────────────────────
    if buy_list:
        # 圧縮表示（format_buy_list）を使って1行に並べる
        try:
            from lr_suggest import format_buy_list as _fbl
            _compressed = _fbl(buy_list)
        except Exception:
            _compressed = buy_list  # フォールバック: そのまま

        # 本線 / 押さえ を分けて表示
        _honsen  = [c for c in candidates if c.get("tier") == "本線"]
        _osaae   = [c for c in candidates if c.get("tier") == "押さえ"]

        def _write_buy_block(label, combos, fill_color, font_color):
            nonlocal _row
            if not combos:
                return
            try:
                ws.merge_cells(start_row=_row, start_column=1, end_row=_row, end_column=2)
            except Exception:
                pass
            wc(_row, 1, label,
               fill=_sf(fill_color), font=_fn(bold=True, size=9, color=font_color),
               alignment=_al(h="center", wrap=False))
            # 6点ずつ列に並べる（C列〜P列）
            for _ci, _combo in enumerate(combos[:14]):
                _col = 3 + _ci
                if _col > 16:
                    break
                wc(_row, _col, _combo,
                   fill=_sf("FFFFFF99"), font=_fn(bold=False, size=10, color="FF000000"),
                   alignment=_al(h="center", wrap=False))
            ws.row_dimensions[_row].height = 18.0
            _row += 1

        if _honsen:
            _h_combos = [c.get("combo","") for c in _honsen]
            _write_buy_block("本線", _h_combos, "FF305496", "FFFFFFFF")
        if _osaae:
            _o_combos = [c.get("combo","") for c in _osaae]
            _write_buy_block("押さえ", _o_combos, "FF70AD47", "FFFFFFFF")
        if not _honsen and not _osaae:
            # tierなし: 全件を本線扱いで表示
            _write_buy_block("買い目", buy_list[:14], "FF305496", "FFFFFFFF")

    elif skip:
        # 見送り時: 買い目なし
        try:
            ws.merge_cells(start_row=_row, start_column=1, end_row=_row, end_column=16)
        except Exception:
            pass
        wc(_row, 1, "（買い目なし）",
           fill=_sf("FFDDDDDD"), font=_fn(bold=False, size=9, color="FF808080"),
           alignment=_al(h="center", wrap=False))
        ws.row_dimensions[_row].height = 16.0
        _row += 1


# ============================================================
# 数値データシート書き込み（出力_数値と同一テーブル形式）
# ============================================================
def write_numeric_sheet(wb, all_race_data, course_master=None, venue_course_master=None, player_master=None):
    """
    「会場名_数値」シートを「出力_数値」と完全同一のテーブル形式で書き出す。
    縦軸: 分類・項目・艇番、横軸: 1R〜12R（最大レース数分）
    player_master: {正規化選手名: {'平均ST\n(Nコース)': float, ...}} ← STブロック書き込みに使用
    """
    if not all_race_data:
        return

    venue_name = all_race_data[0]["venue"]
    sheet_name = f"{venue_name}_数値"

    # ── FLY入力値の準備（flying_*.xlsx を優先、なければ手入力値を退避して復元）──
    fly_input_backup = {}  # {(race_no_str, waku_int): value}

    # まず flying_*.xlsx から自動取得を試みる
    _fly_auto = {}   # {(race_no_int, waku_int): 1}
    _venue_name_for_fly = all_race_data[0]["venue"] if all_race_data else ""
    _flying_path = None
    for _d in [pathlib.Path(__file__).parent, BASE_DIR, pathlib.Path(".")]:
        _cands = sorted(_d.glob("flying_*.xlsx"), reverse=True)
        if _cands:
            _flying_path = _cands[0]
            break
    if _flying_path:
        try:
            _wb_fly = openpyxl.load_workbook(str(_flying_path), read_only=True, data_only=True)
            _ws_fly = _wb_fly["フライング一覧"]
            for _frow in _ws_fly.iter_rows(min_row=2, values_only=True):
                _fcols = list(_frow) + [None] * 7
                # 列順: 会場, レース, 枠, 選手名, 登録/級別, フライング, 合計F数
                _fv, _fr, _fw, _, _, _fi, _ff_total = _fcols[:7]
                if _fv != _venue_name_for_fly or not _fr or not _fw or not _fi:
                    continue
                try:
                    # 合計F数列（G列）が取れればその値、なければ1をデフォルト
                    _f_count = int(_ff_total) if _ff_total is not None else 1
                    _f_count = max(1, _f_count)  # 最小1
                    _fly_auto[(int(_fr), int(str(_fw).strip()))] = _f_count
                except (TypeError, ValueError):
                    pass
            _wb_fly.close()
            if _fly_auto:
                print(f"  ??  FLY自動取得: {_venue_name_for_fly} {len(_fly_auto)}件 ({_flying_path.name})")
        except Exception as _e:
            print(f"  [!]  flying xlsx 読み込み失敗（手入力値を使用）: {_e}")

    if sheet_name in wb.sheetnames:
        _ws_old = wb[sheet_name]
        # ヘッダ行(2行目)からレース番号→列のマッピングを取得
        _old_race_col = {}
        for _c in range(1, _ws_old.max_column + 1):
            _v = _ws_old.cell(2, _c).value
            if _v and isinstance(_v, str) and "R" in _v:
                import re as _re
                _m = _re.match(r"(\d+)R", _v.strip())
                if _m:
                    _old_race_col[_m.group(1)] = _c
        # flying_*.xlsx がない場合のみ手入力値をバックアップ
        if not _fly_auto:
            for _r in range(1, _ws_old.max_row + 1):
                if _ws_old.cell(_r, 1).value == "FLY入力":
                    _waku_val = _ws_old.cell(_r, 3).value
                    try:
                        _waku_int = int(_waku_val)
                    except (TypeError, ValueError):
                        continue
                    for _rno, _col in _old_race_col.items():
                        _fv = _ws_old.cell(_r, _col).value
                        if _fv is not None and str(_fv).strip() not in ("", "0", "None"):
                            fly_input_backup[(_rno, _waku_int)] = _fv
            if fly_input_backup:
                print(f"  ??  FLY手入力値を退避: {len(fly_input_backup)}件")
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ── スタイル定数（出力_数値から実測） ──
    BOAT_FILL   = {1:"FFFFFFFF", 2:"FF1A1A1A", 3:"FFCC0000", 4:"FF2255CC", 5:"FFDDCC00", 6:"FF115511"}
    BOAT_FONT   = {1:"FF000000", 2:"FFFFFFFF", 3:"FFFFFFFF", 4:"FFFFFFFF", 5:"FF000000", 6:"FFFFFFFF"}
    FILL_HDR    = "FF1F4E79"   # ヘッダ行（濃紺）
    FILL_MISSING = "FFFCE4D6"  # データ不足セル（薄オレンジ）
    FILL_SEC_B  = "FF2E75B6"   # 数値指標・選手情報セクション（青）
    FILL_SEC_G  = "FF70AD47"   # 決まり手セクション（緑）
    FILL_ITEM_B = "FFDCE6F1"   # 数値指標・選手情報 項目セル（薄青）
    FILL_ITEM_G = "FFE2EFDA"   # 決まり手 項目セル（薄緑）
    FILL_ITEM_P = "FFD9E1F2"   # 選手情報 項目セル（薄紺）

    race_nos = [rd["race_no"] for rd in all_race_data]
    n_races  = len(race_nos)
    _first_results = all_race_data[0].get("results", [])
    _first_rel = _first_results[0].get("rel_win1") if _first_results else "なし"
    print(f"  ? {sheet_name}: {n_races}レース, 1R results件数={len(_first_results)}, rel_win1={_first_rel}")

    def sf(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def fn(bold=True, size=9, color="FF000000", name="Meiryo UI"):
        return Font(name=name, size=size, bold=bold, color=color)

    def al(h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def wc(row, col, val, fill=None, font=None, align=None, border=None):
        c = ws.cell(row=row, column=col)
        c.value = val
        if fill:   c.fill   = fill
        if font:   c.font   = font
        if align:  c.alignment = align
        if border: c.border = border

    thin = Side(style="thin", color="FFCCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 列幅 ──
    ws.column_dimensions["A"].width = 14.0
    ws.column_dimensions["B"].width = 20.0
    ws.column_dimensions["C"].width = 7.0
    for i in range(n_races):
        ws.column_dimensions[get_column_letter(4 + i)].width = 14.0

    # ── Row 1: タイトル ──
    race_date = all_race_data[0].get("race_date", "")
    title = f"ボートリサーチ数値データ　【{venue_name}】　{race_date}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + n_races)
    wc(1, 1, title, font=fn(bold=True, size=12), align=al("left"))
    ws.row_dimensions[1].height = 22.0

    # ── Row 2: ヘッダ（分類・項目・艇番・1R〜nR） ──
    hf = sf(FILL_HDR)
    hfn = fn(bold=True, color="FFFFFFFF")
    for col, txt in [(1,"分類"),(2,"項目"),(3,"艇番")]:
        wc(2, col, txt, fill=hf, font=hfn, align=al(), border=bdr)
    for i, rno in enumerate(race_nos):
        deadline = all_race_data[i].get("deadline") if i < len(all_race_data) else None
        header_val = f"{rno}R\n{deadline}" if deadline else f"{rno}R"
        wc(2, 4+i, header_val, fill=hf, font=hfn, align=al(wrap=True), border=bdr)
    ws.row_dimensions[2].height = 28.0

    # ── セクション書き込みヘルパー ──
    def write_section_header(row, label, fill_color):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
        wc(row, 1, label, fill=sf(fill_color), font=fn(bold=True, color="FFFFFFFF"), align=al("left"))
        ws.row_dimensions[row].height = 13.0

    def write_item_block(row_start, cat_label, item_label, cat_fill, item_fill, data_by_waku, missing_by_waku=None):
        """
        data_by_waku:   {1: [r1val, r2val, ...], 2: [...], ...}
        missing_by_waku:{1: [bool, bool, ...], 2: [...], ...}  データ不足フラグ
        6艇分 × nレース のデータブロックを書く
        """
        for i in range(6):
            waku = i + 1
            row  = row_start + i
            # A列（分類）: 最初の艇番のみ
            if waku == 1:
                wc(row, 1, cat_label,  fill=sf(cat_fill),  font=fn(bold=True, color="FFFFFFFF"), align=al())
                wc(row, 2, item_label, fill=sf(item_fill), font=fn(bold=True, color="FF000000"), align=al("left"))
            else:
                wc(row, 1, None, fill=sf(cat_fill), align=al())
                wc(row, 2, None, fill=sf(item_fill), align=al())
            # C列（艇番）
            wc(row, 3, str(waku),
               fill=sf(BOAT_FILL[waku]),
               font=fn(bold=True, color=BOAT_FONT[waku]),
               align=al(), border=bdr)
            # D〜 データ列
            vals    = data_by_waku.get(waku,    [None]*n_races)
            missing = (missing_by_waku or {}).get(waku, [False]*n_races)
            for j, v in enumerate(vals):
                is_missing = missing[j] if j < len(missing) else False
                if is_missing and v is None:
                    # データ不足セル: 薄オレンジ背景 ＋ 「－」テキスト
                    wc(row, 4+j, "－",
                       fill=sf(FILL_MISSING),
                       font=fn(bold=False, color="FFBF8F00"),
                       align=al(), border=bdr)
                else:
                    wc(row, 4+j, v, font=fn(bold=False), align=al(), border=bdr)
            ws.row_dimensions[row].height = 15.0

    # ── データ取り出しヘルパー ──
    def get_vals(rd, key):
        """results から艇番→値のdictを返す"""
        d = {}
        for res in rd.get("results", []):
            try: waku = int(res["waku"])
            except (ValueError, TypeError): continue
            d[waku] = res.get(key)
        return d

    def get_cm_vals(rd, cm_key, multiplier=1.0):
        d = {}
        for res in rd.get("results", []):
            try: waku = int(res["waku"])
            except (ValueError, TypeError): continue
            v = safe_float(res.get("raw_cm", {}).get(cm_key))
            d[waku] = round(v * multiplier, 1) if v is not None else None
        return d

    # 6艇×nレース のデータを転置して {waku: [r1,r2,...]} に変換
    def build_waku_data(extract_fn):
        result = {w: [] for w in range(1, 7)}
        for rd in all_race_data:
            vals = extract_fn(rd)
            for w in range(1, 7):
                result[w].append(vals.get(w))
        return result

    # ── データ不足フラグを {waku: [bool×nレース]} に変換するヘルパー ──
    def build_missing_waku():
        """各艇番・各レースについて data_missing フラグを返す"""
        result = {w: [] for w in range(1, 7)}
        for rd in all_race_data:
            missing_map = {}
            for res in rd.get("results", []):
                try: waku = int(res["waku"])
                except (ValueError, TypeError): continue
                missing_map[waku] = bool(res.get("data_missing", False))
            for w in range(1, 7):
                result[w].append(missing_map.get(w, False))
        return result

    missing_waku = build_missing_waku()

    # ── ▼ 数値指標セクション ──
    row = 3
    write_section_header(row, "▼ 数値指標", FILL_SEC_B)
    row += 1

    # オリジナル1着率
    write_item_block(row, "数値指標", "オリジナル1着率(%)", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): f"{round(r['rel_win1'], 1)}%" if r.get("rel_win1") is not None else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }), missing_by_waku=missing_waku)
    row += 6

    # オリジナル3連対率
    write_item_block(row, "数値指標", "コース別3連対率(%)", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): f"{round(r['abs_win3'], 1)}%" if r.get("abs_win3") is not None else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }), missing_by_waku=missing_waku)
    row += 6

    # ★追加: 3連対指数（会場×コース補正付き相対スコア 0〜100）
    # 色分け: 70以上=緑（有利）/ 40〜69=白（標準）/ 39以下=薄橙（不利）
    def _sanren_idx_fill(val):
        v = safe_float(val)
        if v is None:
            return FILL_MISSING
        if v >= 70:
            return "FFE2EFDA"   # 薄緑（有利）
        if v >= 40:
            return "FFFFFFFF"   # 白（標準）
        return "FFFCE4D6"       # 薄橙（不利）

    for _i in range(6):
        _waku = _i + 1
        _dr   = row + _i
        if _waku == 1:
            wc(_dr, 1, "数値指標",  fill=sf(FILL_SEC_B),  font=fn(bold=True, color="FFFFFFFF"), align=al())
            wc(_dr, 2, "3連対指数", fill=sf(FILL_ITEM_B), font=fn(bold=True, color="FF000000"), align=al("left"))
        else:
            wc(_dr, 1, None, fill=sf(FILL_SEC_B),  align=al())
            wc(_dr, 2, None, fill=sf(FILL_ITEM_B), align=al())
        wc(_dr, 3, str(_waku),
           fill=sf(BOAT_FILL[_waku]),
           font=fn(bold=True, color=BOAT_FONT[_waku]),
           align=al(), border=bdr)
        for _j, _rd in enumerate(all_race_data):
            _val = None
            for _r in _rd.get("results", []):
                if str(_r.get("waku")) == str(_waku):
                    _raw = _r.get("sanren_idx")
                    _val = f"{_raw:.0f}" if _raw is not None else None
                    break
            _fc = _sanren_idx_fill(_val)
            wc(_dr, 4 + _j, _val,
               fill=sf(_fc),
               font=fn(bold=(_val is not None and safe_float(_val, 0) >= 70), size=9),
               align=al(), border=bdr)
        ws.row_dimensions[_dr].height = 15.0
    row += 6
    # ★追加ここまで

    # イン逃げ時2着率（1号艇は対象外のため表示なし）
    write_item_block(row, "数値指標", "2着優位度(%)[相対・イン逃げ時]", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): (
                None if r["waku"] == "1"
                else f"{round(r.get('circle_pct'), 1)}%" if r.get("circle_pct") is not None else "-"
            )
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 3着指数
    write_item_block(row, "数値指標", "3着指数", FILL_SEC_B, FILL_ITEM_B,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("idx3") if r.get("idx3") else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── ▼ 決まり手セクション ──
    write_section_header(row, "▼ 決まり手（直近1年｜1号艇=被決まり手%、2〜6号艇=決まり手%）", FILL_SEC_G)
    row += 1

    # 被決まり手マップ: {(レースindex, waku=1): {差され%: v, 捲られ%: v, 捲り差され%: v}}
    # 差し%行→1号艇欄に差され%、まくり%行→捲られ%、まくり差し%行→捲り差され% を赤字で表示
    LOSE_KEY_MAP = {
        "差し%":      "差され%",
        "まくり%":    "捲られ%",
        "まくり差し%":"捲り差され%",
    }

    for cm_key, label in [
        ("逃げ%",      "逃げ%"),
        ("差し%",      "差し%"),
        ("まくり%",    "まくり%"),
        ("まくり差し%","まくり差し%"),
        ("抜き%",      "抜き%"),
    ]:
        lose_key = LOSE_KEY_MAP.get(cm_key)  # 対応する被決まり手キー（なければNone）

        # 各艇番×レースのデータを構築
        main_data = build_waku_data(lambda rd, k=cm_key: {
            int(r["waku"]): safe_float(r.get("raw_cm", {}).get(k), 0)
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        })
        lose_data = build_waku_data(lambda rd, lk=lose_key: {
            int(r["waku"]): safe_float(r.get("raw_cm", {}).get(lk)) if lk and int(r["waku"]) == 1 else None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }) if lose_key else None

        for i in range(6):
            waku = i + 1
            dr   = row + i
            if waku == 1:
                wc(dr, 1, "決まり手", fill=sf(FILL_SEC_G),  font=fn(bold=True, color="FFFFFFFF"), align=al())
                wc(dr, 2, label,      fill=sf(FILL_ITEM_G), font=fn(bold=True, color="FF000000"), align=al("left"))
            else:
                wc(dr, 1, None, fill=sf(FILL_SEC_G),  align=al())
                wc(dr, 2, None, fill=sf(FILL_ITEM_G), align=al())
            wc(dr, 3, str(waku),
               fill=sf(BOAT_FILL[waku]),
               font=fn(bold=True, color=BOAT_FONT[waku]),
               align=al(), border=bdr)

            main_vals = main_data.get(waku, [None]*n_races)
            lose_vals = lose_data.get(waku, [None]*n_races) if lose_data else [None]*n_races

            for j, mv in enumerate(main_vals):
                lv = lose_vals[j] if j < len(lose_vals) else None
                main_str = f"{round((mv or 0) * 100, 1)}%" if mv not in (None,) else "-"

                if waku == 1 and lose_key:
                    # 1号艇：被決まり手%のみ表示（差された・捲られた・捲り差された）
                    lose_str = f"{round(lv * 100, 1)}%" if lv not in (None, 0.0) else "-"
                    wc(dr, 4+j, lose_str,
                       fill=sf(FILL_ITEM_G),
                       font=fn(bold=False, size=9, color="FFCC0000"),
                       align=al(), border=bdr)
                elif waku == 1:
                    # 1号艇の逃げ%行はそのまま表示
                    wc(dr, 4+j, main_str,
                       font=fn(bold=False), align=al(), border=bdr)
                elif cm_key == "逃げ%":
                    # 2〜6号艇の逃げ%行は表示不要（コース1専用）
                    wc(dr, 4+j, "-",
                       font=fn(bold=False, color="FF999999"), align=al(), border=bdr)
                else:
                    # 2〜6号艇：決まり手%をそのまま表示
                    wc(dr, 4+j, main_str,
                       font=fn(bold=False), align=al(), border=bdr)
            ws.row_dimensions[dr].height = 15.0
        row += 6

    # ── ▼ 事前評価セクション ──
    FILL_SEC_J  = "FF7030A0"   # 事前評価セクション（紫）
    FILL_ITEM_J = "FFE9D7F5"   # 事前評価 項目セル（薄紫）

    if JIZEN_AVAILABLE:
        write_section_header(row, "▼ 事前評価（ボートリサーチ流｜展示前）", FILL_SEC_J)
        row += 1

        JIZEN_ITEMS = [
            ("in_nige",  "(1)逃げ"),
            ("aisho",    "(2)相性"),
            ("kiryoku",  "(3)機力"),
            ("tenkai",   "(4)展開"),
            ("jizaisei", "(5)S安定"),
        ]

        JIZEN_FONT_COLOR = {
            "◎": "FF7F6000", "◎?": "FF7F6000",
            "○": "FF1F3864",
            "△": "FF843C0C",
            "A": "FF375623", "B": "FF375623",
            "C": "FF404040",
            "D": "FF843C0C", "E": "FF7F0000",
        }
        JIZEN_FILL_COLOR = {
            "◎": "FFFFF2CC", "◎?": "FFFFF2CC",
            "○": "FFDAE3F3",
            "△": "FFFCE4D6",
            "A": "FFE2EFDA", "B": "FFE2EFDA",
            "C": "FFF2F2F2",
            "D": "FFFCE4D6", "E": "FFFFCCCC",
        }

        # レースごとに evaluate_all を呼んで結果をキャッシュ
        jizen_cache = []
        for rd in all_race_data:
            jm = rd.get("jizen_members")
            if jm:
                try:
                    jizen_cache.append(evaluate_all(jm))
                except Exception:
                    jizen_cache.append(None)
            else:
                jizen_cache.append(None)

        for jizen_key, jizen_label in JIZEN_ITEMS:
            for i in range(6):
                waku = i + 1
                dr   = row + i
                if waku == 1:
                    wc(dr, 1, "事前評価",   fill=sf(FILL_SEC_J),  font=fn(bold=True, color="FFFFFFFF"), align=al())
                    wc(dr, 2, jizen_label,  fill=sf(FILL_ITEM_J), font=fn(bold=True, color="FF000000"), align=al("left"))
                else:
                    wc(dr, 1, None, fill=sf(FILL_SEC_J),  align=al())
                    wc(dr, 2, None, fill=sf(FILL_ITEM_J), align=al())
                wc(dr, 3, str(waku),
                   fill=sf(BOAT_FILL[waku]),
                   font=fn(bold=True, color=BOAT_FONT[waku]),
                   align=al(), border=bdr)
                for j, jr in enumerate(jizen_cache):
                    sym = (jr.get(jizen_key, [""] * 6)[i] or "") if jr else ""
                    cell_fill = sf(JIZEN_FILL_COLOR.get(sym, "FFFFFFFF")) if sym else None
                    cell_font = fn(bold=(sym in ("◎", "◎?", "A")),
                                   size=9,
                                   color=JIZEN_FONT_COLOR.get(sym, "FF808080"))
                    # sym が空のとき None を書く（旧: "－"）
                    # "－"（全角）は fill_newspaper.py のフィルタ（半角"-"のみ）を
                    # すり抜けて新聞に転記され、印が出なくなるバグを修正
                    wc(dr, 4 + j, sym if sym else None,
                       fill=cell_fill, font=cell_font,
                       align=al(), border=bdr)
                ws.row_dimensions[dr].height = 15.0
            row += 6

    # ── ▼ 選手情報セクション ──
    write_section_header(row, "▼ 選手情報", FILL_SEC_B)
    row += 1

    # 選手名（マスタのフルネームで補完）
    # 公式サイトの選手名が4文字に切り詰められている場合（例: 安河内鈴 → 安河内鈴之介）、
    # course_master / venue_course_master のキーから前方一致でフルネームを補完する。
    def _resolve_full_name(name_raw):
        """4文字名をマスタのフルネームに補完して返す。見つからなければ元の名前を返す。"""
        if not name_raw:
            return name_raw
        norm = name_raw.replace("\u3000", "").replace(" ", "").strip()
        # 【修正】5文字以上は既にフルネームとみなし補完しない。
        # 姓4字+名1字の5文字名（例: 大豆生田蒼）が前方一致で誤補完されるのを防ぐ。
        if len(norm) >= 5:
            return name_raw
        # course_master: キーは (選手名, コース文字列)
        if course_master:
            for master_key in course_master:
                master_nm = master_key[0].replace("\u3000", "").replace(" ", "").strip()
                if master_nm.startswith(norm) and len(master_nm) > len(norm):
                    return master_key[0]
        # venue_course_master: キーは (選手名, 会場名, コース文字列) または (選手名, 会場名, int)
        if venue_course_master:
            for master_key in venue_course_master:
                master_nm = str(master_key[0]).replace("\u3000", "").replace(" ", "").strip()
                if master_nm.startswith(norm) and len(master_nm) > len(norm):
                    return master_key[0]
        return name_raw

    def _name_with_mark(r):
        """選手名にscarce_mark（同コース20走未満）の場合は先頭に※を付けて返す"""
        name = _resolve_full_name(r.get("name", ""))
        if r.get("scarce_mark"):
            name = "※" + name
        return name

    write_item_block(row, "選手情報", "選手名", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): _name_with_mark(r)
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 組
    write_item_block(row, "選手情報", "組", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("kumi", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # モータ2連
    write_item_block(row, "選手情報", "モータ2連", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("motor2", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 攻め記号（2〜6号艇）/ 逃げ評価（1号艇）
    # ※ item ラベルを "攻め記号" に統一: fill_newspaper.py が "選手情報_攻め記号" で参照するため
    write_item_block(row, "選手情報", "攻め記号", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("honmei", "").strip() or None
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # 想定コース
    write_item_block(row, "選手情報", "想定コース", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("course", "")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── FLY手動入力行（想定コース直後に挿入）────────────────────────────────
    # ユーザーが各レース列に手動でフライング数を入力するための行
    from openpyxl.styles import PatternFill as _PF, Font as _Fn, Alignment as _Al, Border as _Br, Side as _Sd
    _fill_red    = _PF("solid", fgColor="FFCC0000")
    _fill_yellow = _PF("solid", fgColor="FFFFFF00")
    _fill_boat   = {1:"FFFFFFFF",2:"FF1A1A1A",3:"FFCC0000",4:"FF2255CC",5:"FFDDCC00",6:"FF115511"}
    _fill_font   = {1:"FF000000",2:"FFFFFFFF",3:"FFFFFFFF",4:"FFFFFFFF",5:"FF000000",6:"FFFFFFFF"}
    _thin = _Sd(style="thin", color="FFCCCCCC")
    _bdr  = _Br(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _fn_w = _Fn(name="Meiryo UI", size=9, bold=True,  color="FFFFFFFF")
    _fn_b = _Fn(name="Meiryo UI", size=9, bold=False, color="FF000000")
    _al_c = _Al(horizontal="center", vertical="center")
    for _i, _waku_no in enumerate(range(1, 7)):
        _r = row + _i
        _c = ws.cell(_r, 1)
        _c.value, _c.fill, _c.font, _c.alignment, _c.border = "FLY入力", _fill_red, _fn_w, _al_c, _bdr
        _c = ws.cell(_r, 2)
        _c.value, _c.fill, _c.font, _c.alignment, _c.border = "FLY（1=あり）", _fill_yellow, _fn_b, _al_c, _bdr
        _c = ws.cell(_r, 3)
        _c.value = _waku_no
        _c.fill  = _PF("solid", fgColor=_fill_boat[_waku_no])
        _c.font  = _Fn(name="Meiryo UI", size=9, bold=True, color=_fill_font[_waku_no])
        _c.alignment, _c.border = _al_c, _bdr
        for _col in range(4, 4 + n_races):
            _cell = ws.cell(_r, _col)
            # flying_*.xlsx のデータを優先、なければバックアップ手入力値を復元
            _rno_str = race_nos[_col - 4] if (_col - 4) < len(race_nos) else None
            if _rno_str is not None and _fly_auto.get((int(_rno_str), _waku_no)):
                _cell.value = _fly_auto[(int(_rno_str), _waku_no)]  # F数そのまま（F2=2, F1=1）
            else:
                _restored = fly_input_backup.get((_rno_str, _waku_no)) if _rno_str else None
                _cell.value = _restored
            _cell.fill   = _fill_yellow
            _cell.border = _bdr
        ws.row_dimensions[_r].height = 15.0
    row += 6

    # FLY数
    write_item_block(row, "選手情報", "FLY数", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("fly_count", 0)
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # F/ST影響ラベル
    write_item_block(row, "選手情報", "F/ST影響", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): r.get("fly_label", "低")
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── コース別ST順位 ──
    # player_master の "ST順位_コースN" キー（lr_masters.py で列番号21〜26から取得済み）を使用。
    # fill_newspaper.py の read_numeric_sheet が "選手情報_コース別ST順位" キーで参照する。

    def _get_st_rank_for_result(res):
        """1艇分の想定コース×ST順位を返す。取得できなければ None。"""
        if not player_master:
            return None
        name_raw = res.get('name', '')
        norm = str(name_raw).replace('　', '').replace(' ', '').strip()
        pm = player_master.get(norm)
        if pm is None:
            for k in player_master:
                if norm and (norm in k or k in norm):
                    pm = player_master[k]
                    break
        if pm is None:
            return None
        try:
            course = int(str(res.get('course', res.get('waku', 1))).strip())
            v = pm.get(f"ST順位_コース{course}")
            if v is not None:
                return round(float(v), 2)
        except (TypeError, ValueError):
            pass
        return None

    write_item_block(row, "選手情報", "コース別ST順位", FILL_SEC_B, FILL_ITEM_P,
        build_waku_data(lambda rd: {
            int(r["waku"]): _get_st_rank_for_result(r)
            for r in rd.get("results", []) if str(r["waku"]).isdigit()
        }))
    row += 6

    # ── 注記 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "※オリジナル1着率・オリジナル3連対率・イン逃げ時2着率・3着指数は出走メンバーで正規化。決まり手%は過去コース別1着時の割合。",
       font=fn(bold=False, size=8), align=al("left"))
    ws.row_dimensions[row].height = 13.0
    row += 1

    # ── データ不足凡例 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "★凡例：薄オレンジ「－」はデータ不足（コース別実績 < 5走 or 総実績 < 10走）。新人・移籍直後の選手はデータ蓄積後に正確な指数が算出されます。",
       fill=sf(FILL_MISSING),
       font=fn(bold=False, size=8, color="FF7F3F00"),
       align=al("left"))
    ws.row_dimensions[row].height = 14.0

    # ── データ不足選手一覧（シート末尾サマリー） ──
    missing_summary = []
    for rd in all_race_data:
        for res in rd.get("results", []):
            if res.get("data_missing") and res.get("missing_reason"):
                missing_summary.append(
                    f"{rd['race_no']}R-{res['waku']}号艇 {res['name']}：{res['missing_reason']}"
                )
    if missing_summary:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
        wc(row, 1, "【データ不足選手一覧】",
           fill=sf("FFFCE4D6"),
           font=fn(bold=True, size=9, color="FF7F3F00"),
           align=al("left"))
        ws.row_dimensions[row].height = 14.0
        for summary_line in missing_summary:
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
            wc(row, 1, summary_line,
               fill=sf(FILL_MISSING),
               font=fn(bold=False, size=9, color="FF7F3F00"),
               align=al("left"))
            ws.row_dimensions[row].height = 13.0

    # ====================================================================
    # [表] 展示前参考買い目 ＋ 考察セクション
    # ====================================================================
    FILL_JF_HDR  = "FF1F4E79"   # 濃紺（ヘッダ）
    FILL_JF_SUB  = "FF2E75B6"   # 中青（行ラベル）
    FILL_JF_BODY = "FFDCE6F1"   # 薄青（データ）
    FILL_JF_BET  = "FFFFFF99"   # 薄黄（買い目リスト）
    FILL_JF_NOTE = "FFFFE699"   # 黄（絞込ガイド）

    FILL_KS_HDR  = "FF203864"   # 濃紺（考察ヘッダ）
    FILL_KS_SUB  = "FF305496"   # 中紺（考察行ラベル）
    FILL_KS_BODY = "FFD9E1F2"   # 薄紺（考察データ）
    FILL_KS_GOOD = "FFE2EFDA"   # 薄緑（良好）
    FILL_KS_WARN = "FFFFF2CC"   # 薄黄（注意）
    FILL_KS_BAD  = "FFFCE4D6"   # 薄橙（警告）

    row += 2

    # ── 考察セクションヘッダ ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "━━━  [考察] Step1: レース考察  ━━━  "
       "展開予想の根拠（判定スコア・相性・展開・ヒモ荒れ）→ 次の買い目の根拠になります",
       fill=sf(FILL_KS_HDR), font=fn(bold=True, size=9, color="FFFFFFFF"), align=al("left"))
    ws.row_dimensions[row].height = 16.0
    row += 1

    KOSATSU_ROWS = [
        # (1) 展開ストーリー（逃げ判定・主役・崩れ・買い目）
        ("[狙]展開ストーリー",  "(1)逃げ判定 (2)逃時2・3着 (3)主役 (4)主役時2・3着 (5)崩れ → 買い目"),
        # (2) 展開パターン（A/B/C/D）と方針
        ("[狙]展開パターン",  "A鉄板/B主役/C拮抗/D荒れ ＋ 方針"),
        # (3) 堅実度スコア（quality_score ベース）
        ("[狙]堅実度",  "S(>=75)/A(>=55)/B(>=45)/C(>=35)/D(<35)"),
    ]

    for sec_label, sec_item in KOSATSU_ROWS:
        wc(row, 1, sec_label,
           fill=sf(FILL_KS_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        wc(row, 2, sec_item,
           fill=sf(FILL_KS_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())

        max_lines = 2
        for i, rd in enumerate(all_race_data):
            rj  = rd.get("race_judgment", {}) or {}
            bet = rd.get("bet_suggestions", {}) or {}
            je  = rd.get("jizen_eval", {}) or {}
            jf  = bet.get("jizen_formation", {}) or {}
            ryotate = bet.get("ryotate_detail", rj.get("ryotate", {})) or {}
            txt = ""
            fc  = FILL_KS_BODY

            if sec_label == "[狙]展開ストーリー":
                # ── 7ステップ統合ストーリー ───────────────────────────────────
                try:
                    # scenario_engine が生成済みのストーリーを優先使用
                    # フォールバック: 旧 _generate_tenkai_story
                    _story_from_engine = bet.get("story", "")
                    if _story_from_engine and _story_from_engine.strip():
                        txt = _story_from_engine
                    else:
                        txt = _generate_tenkai_story(
                            results         = rd.get("results", []),
                            venue           = rd.get("venue", ""),
                            venue_stats     = rd.get("venue_stats", {}),
                            race_judgment   = rj,
                            bet_suggestions = bet,
                        )
                    if not txt or not txt.strip():
                        txt = "（データ不足のためストーリーを生成できませんでした）"
                except Exception as _e:
                    txt = f"（展開ストーリー生成エラー: {_e}）"

                # 濃紺背景・白文字（(8)考察の結論と同じスタイルだが行高を大きく取る）
                fc = "FF1A2E4A"
                wc(row, 4+i, txt,
                   fill=sf(fc), font=fn(bold=False, size=8, color="FFFFFFFF"),
                   align=al(h="left", wrap=True), border=bdr)
                max_lines = max(max_lines, txt.count("\n") + 1)
                continue

            elif sec_label == "[狙]展開パターン":
                # 層1先頭: 展開パターン（A/B/C/D）＋方針を最上部に表示
                tp   = bet.get("tenkai_pattern", rj.get("tenkai_pattern", "?"))
                tp_pol = bet.get("tenkai_pattern_policy", rj.get("tenkai_pattern_policy", "-"))
                rank = rj.get("rank", "-")
                _TP_FULL = {
                    "A": "[緑] A: 鉄板逃げ",
                    "B": "[赤] B: 主役展開",
                    "C": "[黄] C: 拮抗",
                    "D": "[紫] D: 荒れ",
                }
                tp_label_full = _TP_FULL.get(tp, f"[白] ?: {tp}")
                txt = f"【{tp_label_full}】\n方針: {tp_pol}\nランク: {rank}"
                fc = {
                    "A": FILL_KS_GOOD,
                    "B": FILL_KS_BAD,
                    "C": FILL_KS_WARN,
                    "D": "FFEDE7F6",
                }.get(tp, FILL_KS_BODY)

            elif sec_label == "[狙]堅実度":
                # 堅実度スコア（v3: quality_scoreをそのまま使用）
                # tenkai_patternは_calc_scenario_quality内で展開確定度（0〜30点）として
                # 組み込み済みのため、ここでの後付け補正は不要
                sq      = rj.get("scenario_quality", {}) or {}
                q_score = sq.get("quality_score")
                if isinstance(q_score, (int, float)):
                    k_score = float(q_score)
                else:
                    k_score = None
                def _kendai_rank_num(sc):
                    if sc is None:
                        return "堅実-"
                    if sc >= 80: return "堅実S"
                    if sc >= 60: return "堅実A"
                    if sc >= 40: return "堅実B"
                    if sc >= 20: return "堅実C"
                    return "堅実D"
                k_rank = _kendai_rank_num(k_score)
                k_score_str = f"{k_score:.0f}点" if k_score is not None else "-"
                txt = f"{k_score_str} / {k_rank}"
                # ランクに応じた背景色
                fc = (
                    FILL_KS_GOOD if k_rank in ("堅実S", "堅実A") else
                    FILL_KS_BODY if k_rank == "堅実B" else
                    FILL_KS_WARN if k_rank == "堅実C" else
                    FILL_KS_BAD
                )

            elif sec_label == "[狙]狙い目":
                # ── 個人攻撃有効性ベースの狙い目 ──────────────────────────────
                neraime_list = bet.get("neraime", []) or []
                neraime_2nd  = bet.get("neraime_2nd", []) or []
                neraime_top  = bet.get("neraime_top") or {}
                atk_eff_map  = bet.get("atk_eff_map", {}) or {}
                s1p_nr = bet.get("s1_prob", 0) or 0
                fp_map = bet.get("first_prob_map", {}) or {}

                # ── 全艇スコア行（共通）──
                atk_lines = []
                for _w, _s in sorted(atk_eff_map.items(), key=lambda x: x[1], reverse=True):
                    _p = fp_map.get(_w, 0)
                    _top_n = neraime_top.get("waku") if neraime_top else None
                    _marker = "★" if _w == _top_n else "  "
                    _level_tag = ""
                    for _n in neraime_list:
                        if _n["waku"] == _w:
                            _level_tag = f"[{_n.get('level','?')}]"
                            break
                    atk_lines.append(
                        f"{_marker}{_w}号: {_level_tag}有効{_s*100:.0f}%"
                        f" → 確率{_p*100:.1f}%"
                    )

                if neraime_top:
                    # 攻め型狙い目あり
                    nw    = neraime_top.get("waku", "-")
                    ns    = neraime_top.get("score", 0)
                    nlv   = neraime_top.get("level", "?")
                    nat   = neraime_top.get("attack_type", "-")
                    nr    = neraime_top.get("reason", "-")
                    np    = neraime_top.get("prob_after", 0)

                    # 残存型も合わせて表示
                    _2nd_lines = []
                    if neraime_2nd:
                        _2nd_lines.append("─ 逃げ時2着残存 ─")
                        for _n2 in neraime_2nd[:3]:
                            _2nd_lines.append(
                                f"  {_n2['waku']}号: {_n2['r2_rate']*100:.0f}%"
                                f"（マスタ）/ {_n2['r3i_rate']*100:.0f}%（3着以内）"
                            )

                    txt = (
                        f"? 攻め型: {nw}号艇【{nat}】[{nlv}]\n"
                        f"   有効性{ns*100:.0f}% → 補正後1着確率{np*100:.1f}%\n"
                        f"   {nr}\n"
                        f"─────────────────\n"
                        f"全艇スコア:\n"
                        + "\n".join(atk_lines)
                        + ("\n" + "\n".join(_2nd_lines) if _2nd_lines else "")
                    )
                    fc = (FILL_KS_BAD   if ns >= 0.35 else
                          FILL_KS_WARN  if ns >= 0.22 else
                          FILL_KS_BODY)

                elif neraime_2nd:
                    # 攻め型なし＋逃げ本命 → 残存型のみ
                    _2nd_top = neraime_2nd[0]
                    _2nd_lines = ["─ 逃げ時2着残存（展開別残存マスタ） ─"]
                    for _n2 in neraime_2nd[:4]:
                        _2nd_lines.append(
                            f"  {_n2['waku']}号: 2着{_n2['r2_rate']*100:.0f}%"
                            f" / 3着以内{_n2['r3i_rate']*100:.0f}%"
                        )
                    txt = (
                        f"? 残存型: {_2nd_top['waku']}号艇が2着有力\n"
                        f"   逃げ時2着率{_2nd_top['r2_rate']*100:.0f}%"
                        f"（1号艇1着確率{fp_map.get('1',0)*100:.1f}%）\n"
                        f"─────────────────\n"
                        f"全艇スコア:\n"
                        + "\n".join(atk_lines) + "\n"
                        + "\n".join(_2nd_lines)
                    )
                    fc = FILL_KS_GOOD   # 緑：逃げ安定・残存狙い

                else:
                    # 攻め型も残存型も該当なし
                    txt = (
                        f"? 狙い目: 明確な攻め手なし\n"
                        f"   （攻撃有効性15%未満・逃げ2着残存も集計中）\n"
                        f"─────────────────\n"
                        f"全艇スコア:\n"
                        + "\n".join(atk_lines)
                    )
                    fc = FILL_KS_BODY

            elif sec_label == "(1)判定スコア":
                rank  = rj.get("rank", "-")
                score = rj.get("score", "-")
                strat = rj.get("strategy", "-")
                trust = rj.get("data_trust_score", "-")
                txt = f"【ランク{rank}】{score}点\n戦略: {strat}\n信頼度: {trust}%"
                fc = FILL_KS_GOOD if rank in ("S","A") else FILL_KS_WARN if rank == "B" else FILL_KS_BAD

            elif sec_label == "(1)判定根拠":
                reasons = rj.get("reason", []) or []
                filtered = [r for r in reasons if not r.startswith("複合確率スコア")][:3]
                txt = "\n".join(f"・{r}" for r in filtered) if filtered else "-"

            elif sec_label == "(2)3択判定":
                verdict = ryotate.get("verdict", "-")
                conf    = ryotate.get("confidence", "-")
                esc_pct = ryotate.get("escape_pct")
                tobi_pct = ryotate.get("tobi_pct")

                # 1行サマリー: 「なぜこの3択か」を1文で表現
                if verdict == "逃げ狙い":
                    if esc_pct is not None:
                        summary = f"→ 逃げ{esc_pct:.0f}%: 1号艇先行優勢・ヒモ流しで参加"
                    else:
                        summary = "→ 1号艇先行優勢・逃げ軸"
                elif verdict == "飛び狙い":
                    main_th = (rj.get("affinity") or {}).get("dominant_attacker", "?")
                    if tobi_pct is not None:
                        summary = f"→ 飛び{tobi_pct:.0f}%: {main_th}号艇が主な脅威・飛び軸"
                    else:
                        summary = f"→ {main_th}号艇先行・飛び軸"
                else:
                    if esc_pct is not None and tobi_pct is not None:
                        summary = f"→ 逃げ{esc_pct:.0f}%/飛び{tobi_pct:.0f}%拮抗・展示で軸決定"
                    else:
                        summary = "→ 拮抗展開・軸を傾ける"

                reason = ryotate.get("reason", "-")
                # reason は長いので先頭の警告部分のみ抽出（[!]以降の重要部分）
                reason_short = reason.split("。")[0] if reason != "-" else "-"

                txt = (
                    f"【{verdict}】確信度{conf}%\n"
                    f"{summary}\n"
                    f"根拠: {reason_short}"
                )
                fc = FILL_KS_GOOD if verdict == "逃げ狙い" else FILL_KS_BAD if verdict == "飛び狙い" else FILL_KS_WARN

            elif sec_label == "(2)買い方指示":
                txt = ryotate.get("buy_style", "-") or "-"

            elif sec_label == "(3)相性考察":
                affinity = rj.get("affinity", {}) or {}
                summary  = affinity.get("affinity_summary", {}) or {}
                attack   = affinity.get("attack_score", {}) or {}
                outer = sorted(
                    [(w, summary.get(w, "-"), attack.get(w, 0))
                     for w in ["2","3","4","5","6"] if w in summary],
                    key=lambda x: x[2], reverse=True
                )[:4]
                txt = "\n".join(f"{w}号: {s}（攻{a:.0f}pt）" for w, s, a in outer) or "データなし"

            elif sec_label == "(4)展開予測":
                ft   = rj.get("first_turn", {}) or {}
                cm   = rj.get("conflict_map", {}) or {}
                w1e  = rj.get("w1_escape", {}) or {}
                mp   = rj.get("main_player", {}) or {}
                ef   = rj.get("escape_fallback", {}) or {}
                dh   = rj.get("dark_horse", {}) or {}
                tp   = bet.get("tenkai_pattern", rj.get("tenkai_pattern", "?"))

                # 展開パターンラベル
                _TP_LABEL = {
                    "A": "A: 鉄板逃げ",
                    "B": "B: 主役展開",
                    "C": "C: 拮抗",
                    "D": "D: 荒れ",
                }
                tp_label = _TP_LABEL.get(tp, f"?: {tp}")

                # (1)
                er   = w1e.get("escape_rank", "-")
                epct = w1e.get("escape_pct", "-")
                thr  = w1e.get("top_threat_waku", "-")
                tht  = w1e.get("top_threat_type", "-")
                # (2)
                mw   = mp.get("main_waku", "-")
                mt   = mp.get("main_type", "-")
                ms   = mp.get("main_score", 0) or 0
                sw   = mp.get("sub_waku")
                # (3)
                fbp  = ef.get("fallback_pct", "-")
                fbr  = ef.get("fallback_rank", "-")
                fbt  = ef.get("fly_type", "-")
                # (4)
                dh_ok  = dh.get("is_valid", False)
                dh_top = dh.get("top_waku", "-")
                dh_sc  = dh.get("top_score", 0) or 0
                dh_cands = dh.get("dark_horse_candidates", [])
                dh_str = "  ".join(
                    f"{w}号【{tag}】{s*100:.0f}%"
                    for w, s, tag in dh_cands[:2]
                ) if dh_cands else "なし"

                # 1M到達・対立構造
                entry = ft.get("entry_order", [])
                entry_str = "→".join([f"{w}号" for w, _ in entry[:4]]) if entry else "-"
                mc = cm.get("main_conflict") or {}
                mc_desc = mc.get("desc", "-") if mc else "-"
                cb = cm.get("collapse_beneficiary", [])
                cb_str = "・".join([f"{w}号" for w, _ in cb[:2]]) if cb else "-"

                sub_str = f"/{sw}号" if sw else ""
                dh_line = f"[OK]{dh_top}号{dh_sc*100:.0f}%" if dh_ok else "─"
                txt = (
                    f"?逃 {epct}【{er}】脅={thr}号({tht})\n"
                    f"?主 {mw}号【{mt}】{ms*100:.0f}%{sub_str}\n"
                    f"?残 {fbp}【{fbr}】  ?穴 {dh_line}\n"
                    f"─────────────────\n"
                    f"1M: {entry_str}\n"
                    f"主軸: {mc_desc} / 漁夫: {cb_str}"
                )
                # パターン別カラー
                fc = {
                    "A": FILL_KS_GOOD,
                    "B": FILL_KS_BAD,
                    "C": FILL_KS_WARN,
                    "D": "FFEDE7F6",  # 薄紫（荒れ）
                }.get(tp, FILL_KS_BODY)

            elif sec_label == "(5)注意事項":
                reasons = rj.get("reason", []) or []
                warn_keys = ["FLY","出遅れ","データ不足","データ信頼","ST不安定","実績なし","暫定"]
                warns = [r for r in reasons if any(k in r for k in warn_keys)]

                # 展示確認トリガー（ヒモ荒れ判定から）
                ha = rj.get("himo_are", {}) or {}
                tenji_trigger = ha.get("tenji_trigger", "")

                # 注意事項がなくても展示トリガーは必ず表示
                lines = []
                if warns:
                    lines += [f"[!] {w}" for w in warns]
                else:
                    lines.append("特記なし")

                if tenji_trigger:
                    lines.append("─ 展示確認ポイント ─")
                    lines += [f"[表] {t}" for t in tenji_trigger.split("\n") if t.strip()]

                txt = "\n".join(lines)
                fc = FILL_KS_WARN if warns else FILL_KS_BODY

            elif sec_label == "(6)ヒモ荒れ":
                ha = rj.get("himo_are", {}) or {}
                ha_verdict = ha.get("verdict", "対象外")
                tenji_t = ha.get("tenji_trigger", "")
                tenji_line = f"\n[表] {tenji_t}" if tenji_t else ""
                if ha_verdict == "対象外":
                    txt = "対象外（rel_win1 < 45%）\n通常判定に委ねる"
                    fc  = FILL_KS_BODY
                elif ha_verdict == "不参加推奨":
                    mcp = ha.get("max_combo_prob", 0.0) or 0.0
                    eto = ha.get("est_top_odds",   0.0) or 0.0
                    cc  = ha.get("circle_concentration", 0.0) or 0.0
                    txt = (
                        f"? 見送り推奨\n"
                        f"最有力確率{mcp:.3f}（推定{eto:.0f}倍台）\n"
                        f"2着集中度{cc:.0f}%"
                        f"{tenji_line}"
                    )
                    fc  = FILL_KS_BAD
                elif ha_verdict == "点数絞り":
                    mcp = ha.get("max_combo_prob", 0.0) or 0.0
                    eto = ha.get("est_top_odds",   0.0) or 0.0
                    cc  = ha.get("circle_concentration", 0.0) or 0.0
                    txt = (
                        f"[!] 点数絞り\n"
                        f"最有力確率{mcp:.3f}（推定{eto:.0f}倍台）\n"
                        f"2着集中度{cc:.0f}%"
                        f"{tenji_line}"
                    )
                    fc  = FILL_KS_WARN
                else:  # 参加推奨
                    mcp = ha.get("max_combo_prob", 0.0) or 0.0
                    eto = ha.get("est_top_odds",   0.0) or 0.0
                    cc  = ha.get("circle_concentration", 0.0) or 0.0
                    txt = (
                        f"[OK] 参加推奨（ヒモ分散）\n"
                        f"最有力確率{mcp:.3f}（推定{eto:.0f}倍台）\n"
                        f"2着集中度{cc:.0f}%"
                        f"{tenji_line}"
                    )
                    fc  = FILL_KS_GOOD

            # (7)展開quality
            elif sec_label == "(7)展開quality":
                sq = rj.get("scenario_quality", {}) or {}
                q_score   = sq.get("quality_score", "-")
                q_rank    = sq.get("quality_rank", "-")
                q_verdict = sq.get("quality_verdict", "-")
                bet_guide = sq.get("bet_size_guide", "-")
                comps     = sq.get("components", {})
                comp_str  = "  ".join([f"{k}:{v:.0f}" for k, v in comps.items()])

                # ── 堅実度スコア（v3: quality_scoreをそのまま使用）────────
                # tenkai_patternは_calc_scenario_quality内で展開確定度として組み込み済み
                def _kendai_rank_kd(sc):
                    if sc >= 80: return "堅実S"
                    if sc >= 60: return "堅実A"
                    if sc >= 40: return "堅実B"
                    if sc >= 20: return "堅実C"
                    return "堅実D"
                if isinstance(q_score, (int, float)):
                    kendai_str = f"{q_score:.0f}点 / {_kendai_rank_kd(q_score)}"
                else:
                    kendai_str = "- / 堅実-"

                # シナリオタイプ別ROI目安（バックテスト実績ベース）
                stype_for_roi = bet.get("scenario_type", "-")
                ROI_GUIDE = {
                    "逃げ軸流し": "ROI目安: 逃げ軸=約27%(中央値Y1,080)  ※ヒモ参加推奨時のみ",
                    "飛び軸":     "ROI目安: 飛び軸=約38%(中央値Y2,200)  ※quality A以上推奨",
                    "両建て":     "ROI目安: 両建て=約32%(中央値Y1,400)  ※軸絞ること",
                }
                roi_line = ROI_GUIDE.get(stype_for_roi, "ROI目安: シナリオ未確定")

                txt = (
                    f"【堅実度】{kendai_str}\n"
                    f"─\n"
                    f"quality: {q_rank}（{q_score}点）\n"
                    f"→ {q_verdict}\n"
                    f"→ {bet_guide}\n"
                    f"構成: {comp_str}\n"
                    f"─\n"
                    f"{roi_line}"
                )
                if q_rank in ("S", "A"):
                    fc = FILL_KS_GOOD
                elif q_rank == "B":
                    fc = FILL_KS_BODY
                elif q_rank == "C":
                    fc = FILL_KS_WARN
                else:
                    fc = FILL_KS_BAD

            # (8)考察の結論: 初心者向け短文（_generate_buy_hint → fill_newspaper.py kosatsu_raw に転記）
            if sec_label == "(8)考察の結論":
                try:
                    txt = _generate_buy_hint(
                        results         = rd.get("results", []),
                        venue           = rd.get("venue", ""),
                        venue_stats     = rd.get("venue_stats", {}),
                        race_judgment   = rj,
                        bet_suggestions = bet,
                    )
                    if not txt or not txt.strip():
                        txt = "（データ不足のため考察を生成できませんでした）"
                except Exception as _e:
                    txt = f"（考察生成エラー: {_e}）"
                fc = "FF1F4E79"   # 濃紺（固定）
                wc(row, 4+i, txt,
                   fill=sf(fc), font=fn(bold=False, size=8, color="FFFFFFFF"),
                   align=al(h="left", wrap=True), border=bdr)
                max_lines = max(max_lines, txt.count("\n") + 1)
                continue

            # (9)買い方ヒント: KOSATSU_ROWSから削除済み（デッドコード）
            # 必要な場合は KOSATSU_ROWS に ("(9)買い方ヒント", ...) を追加すれば復活可能
            if sec_label == "(9)買い方ヒント":
                try:
                    txt = _generate_buy_hint(
                        results   = rd.get("results", []),
                        venue     = rd.get("venue", ""),
                        venue_stats = rd.get("venue_stats", {}),
                        race_judgment  = rj,
                        bet_suggestions = bet,
                    )
                    if not txt or not txt.strip():
                        txt = "（データ不足：s1_prob未計算の可能性）"
                except Exception as _e:
                    txt = f"（買い方ヒント生成エラー: {_e}）"
                fc = "FF1A3A5C"   # 深紺（専用色）
                wc(row, 4+i, txt,
                   fill=sf(fc), font=fn(bold=False, size=8, color="FFFFFFFF"),
                   align=al(h="left", wrap=True), border=bdr)
                max_lines = max(max_lines, txt.count("\n") + 1)
                continue

            wc(row, 4+i, txt,
               fill=sf(fc), font=fn(bold=False, size=8),
               align=al(h="left", wrap=True), border=bdr)
            max_lines = max(max_lines, txt.count("\n") + 1)

        # 変更6: 行高を情報密度に合わせて最適化
        # 層1（先頭3行）は低め固定、層2（根拠詳細）は可変、層3（(8)結論）はコンパクト
        _ROW_HEIGHT_MAP = {
            "[狙]展開ストーリー": max(13.0 * max_lines, 160.0),  # 最優先・7ステップ分を確保
            "[狙]展開パターン": max(14.0 * max_lines, 36.0),
            "[狙]狙い目":       max(13.0 * max_lines, 80.0),
            "(4)展開予測":     max(14.0 * max_lines, 40.0),
            "(1)判定スコア":   max(14.0 * max_lines, 36.0),
            "(2)3択判定":      max(16.0 * max_lines, 36.0),
            "(3)相性考察":     max(14.0 * max_lines, 36.0),
            "(5)注意事項":     max(14.0 * max_lines, 28.0),
            "(6)ヒモ荒れ":     max(14.0 * max_lines, 28.0),
            "(7)展開quality":  max(14.0 * max_lines, 36.0),
            "(8)考察の結論":   max(14.0 * max_lines, 40.0),
        }
        ws.row_dimensions[row].height = _ROW_HEIGHT_MAP.get(
            sec_label, max(18.0 * max_lines, 36.0)
        )
        row += 1

    # ====================================================================
    # [考察] 新フォーマット：本命・対抗・選定理由・参考買い目
    # 引き継ぎ書フォーマットに準拠した出力確認用行
    # ====================================================================
    row += 1  # 空行1行あけ

    # セクションヘッダ
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "━━━  [考察] 本命・対抗・選定理由  ━━━  "
       "本命/対抗の選定根拠と参考買い目（本線/押さえ）",
       fill=sf("FF203864"), font=fn(bold=True, size=9, color="FFFFFFFF"), align=al("left"))
    ws.row_dimensions[row].height = 16.0
    row += 1

    # ラベル列
    wc(row, 1, "[考察]本命・対抗",
       fill=sf("FF305496"), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "本命/対抗/選定理由",
       fill=sf("FF305496"), font=fn(bold=True, color="FFFFFFFF"), align=al())

    _new_max_lines = 2
    for i, rd in enumerate(all_race_data):
        bet = rd.get("bet_suggestions", {}) or {}
        _comment = bet.get("comment", "") or ""

        # comment から【考察】ブロックだけ取り出す
        # フォーマット:
        #   【考察】
        #   本命：●号艇
        #   対抗：●号艇
        #   選定理由：...
        #   \n【参考買い目】\n...
        if _comment:
            _parts = _comment.split("\n\n【参考買い目】")
            _kosatsu_block = _parts[0].replace("【考察】\n", "").strip()
            # ev_warning_msg（合成オッズ警告）は事前情報では判断不能なため除去
            # "[!]" で始まり合成オッズ・回収率・見送り推奨を含む行と
            # その直後の "  →" 行をまとめて削除する
            _filtered_lines = []
            _skip_next = False
            for _ln in _kosatsu_block.splitlines():
                if _ln.strip().startswith("[!]") and (
                    "合成オッズ" in _ln or "回収率" in _ln or "見送り推奨" in _ln
                ):
                    _skip_next = True
                    continue
                if _skip_next and _ln.strip().startswith("→"):
                    _skip_next = False
                    continue
                _skip_next = False
                _filtered_lines.append(_ln)
            _kosatsu_block = "\n".join(_filtered_lines).strip()
        else:
            # comment未生成時は最低限の情報を組み立て
            _honmei_w = bet.get("axis1", "-")
            _taiko_w  = bet.get("axis2", "-")
            _s1p      = bet.get("s1_prob", 0) or 0
            _stype    = bet.get("scenario_type", "-")
            _kosatsu_block = (
                f"本命：{_honmei_w}号艇\n"
                f"対抗：{_taiko_w}号艇\n"
                f"選定理由：{_stype}（逃げ{_s1p*100:.0f}%）"
            )

        # 見送りの場合はスキップ理由を表示
        if bet.get("skip", False):
            _skip_reason = bet.get("skip_reason", "見送り")
            # ev_warning_msg（合成オッズ警告）を skip_reason からも除去
            _sr_filtered = []
            _sr_skip_next = False
            for _ln in _skip_reason.splitlines():
                if _ln.strip().startswith("[!]") and (
                    "合成オッズ" in _ln or "回収率" in _ln or "見送り推奨" in _ln
                ):
                    _sr_skip_next = True
                    continue
                if _sr_skip_next and _ln.strip().startswith("→"):
                    _sr_skip_next = False
                    continue
                _sr_skip_next = False
                _sr_filtered.append(_ln)
            _skip_reason = "\n".join(_sr_filtered).strip()
            _kosatsu_block = f"【見送り】\n{_skip_reason}" if _skip_reason else "【見送り】"
            _fc = "FFFCE4D6"  # 薄橙
        else:
            _fc = "FFD9E1F2"  # 薄紺

        wc(row, 4 + i, _kosatsu_block,
           fill=sf(_fc), font=fn(bold=False, size=8),
           align=al(h="left", wrap=True), border=bdr)
        _new_max_lines = max(_new_max_lines, _kosatsu_block.count("\n") + 1)

    ws.row_dimensions[row].height = max(14.0 * _new_max_lines, 60.0)
    row += 1

    # 参考買い目行（本線/押さえ）
    wc(row, 1, "[考察]参考買い目",
       fill=sf("FF305496"), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "本線 / 押さえ",
       fill=sf("FF305496"), font=fn(bold=True, color="FFFFFFFF"), align=al())

    _buy_max_lines = 1
    for i, rd in enumerate(all_race_data):
        bet = rd.get("bet_suggestions", {}) or {}
        _comment = bet.get("comment", "") or ""

        if _comment and "\n\n【参考買い目】" in _comment:
            _buy_block = _comment.split("\n\n【参考買い目】")[1].strip()
        else:
            # candidates から本線/押さえを組み立て
            _cands = bet.get("candidates", []) or []
            _honsen = [c.get("combo","") for c in _cands if c.get("tier") != "押さえ"]
            _osaae  = [c.get("combo","") for c in _cands if c.get("tier") == "押さえ"]
            if not _honsen and not _osaae:
                _bl = bet.get("buy_list", []) or []
                _honsen = _bl[:14]
            _buy_block = (
                f"本線：{'　'.join(_honsen) if _honsen else '（なし）'}\n"
                f"押さえ：{'　'.join(_osaae) if _osaae else 'なし'}"
            )

        if bet.get("skip", False):
            _buy_block = "（買い目なし）"
            _buy_fc = "FFDDDDDD"
        else:
            _buy_fc = "FFFFFF99"  # 薄黄

        wc(row, 4 + i, _buy_block,
           fill=sf(_buy_fc), font=fn(bold=False, size=9),
           align=al(h="left", wrap=True), border=bdr)
        _buy_max_lines = max(_buy_max_lines, _buy_block.count("\n") + 1)

    ws.row_dimensions[row].height = max(16.0 * _buy_max_lines, 36.0)
    row += 1

    row += 1  # 空行1行あけ

    # ====================================================================
    # [表] Step2: 展開シナリオ＋買い目候補
    # ====================================================================
    # 【設計方針】
    #   シナリオ自動切替（逃げ軸/両建て/飛び軸）の判定結果と
    #   各シナリオの買い目候補を確率順に提示する。
    #   実オッズ・参加可否は人間が最終判断する。
    # ====================================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "━━━  [表] Step2: 展開シナリオ＋買い目候補  ━━━  "
       "確率から展開を自動判定。買い目候補と理想合成オッズを確認して展示後の最終判断へ",
       fill=sf(FILL_JF_HDR), font=fn(bold=True, size=9, color="FFFFFFFF"), align=al("left"))
    ws.row_dimensions[row].height = 16.0
    row += 1

    # ─ 説明行 ─
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + n_races)
    wc(row, 1,
       "【逃げ軸流し】印◎が1号艇かつ逃げ確率高め　"
       "【両建て】印◎が1号艇で確率低め、または印◎外枠で印○が1号艇　"
       "【飛び軸】印◎◎が2〜6号艇　"
       "2〜3着ヒモは展開位置補正×個人能力×jizen評価で自動選択。折り返しは条件付き自動追加",
       fill=sf(FILL_JF_NOTE), font=fn(bold=False, size=8, color="FF7F3F00"), align=al("left"))
    ws.row_dimensions[row].height = 14.0
    row += 1

    # ─ シナリオ判定行 ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "展開シナリオ判定",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        bet  = rd.get("bet_suggestions", {}) or {}
        stype = bet.get("scenario_type", bet.get("scenario_verdict", "-"))
        s1p  = bet.get("s1_prob", 0) or 0
        flyp = bet.get("fly_prob", 0) or 0
        fly_axes = bet.get("fly_axes", [])
        fly_str  = "・".join(fly_axes[:2]) + "号" if fly_axes else "-"
        rj   = rd.get("race_judgment", {}) or {}
        sc_fly_type  = rj.get("sc_fly_type", "-")
        gyofu_top3   = rj.get("sc_gyofu_top3", [])
        gyofu_str    = "・".join(gyofu_top3) + "号" if gyofu_top3 else "-"

        # (1)〜(4)サマリー
        w1e  = rj.get("w1_escape", {}) or {}
        mp2  = rj.get("main_player", {}) or {}
        ef2  = rj.get("escape_fallback", {}) or {}
        dh2  = rj.get("dark_horse", {}) or {}
        tp2  = bet.get("tenkai_pattern", rj.get("tenkai_pattern", "?"))
        tp2_pol = bet.get("tenkai_pattern_policy", rj.get("tenkai_pattern_policy", "-"))
        _TP2_EMOJI = {"A": "[緑]", "B": "[赤]", "C": "[黄]", "D": "[紫]"}
        tp2_icon = _TP2_EMOJI.get(tp2, "[白]")

        er2   = w1e.get("escape_rank", "-")
        epct2 = w1e.get("escape_pct", "-")
        mw2   = mp2.get("main_waku", "-")
        mt2   = mp2.get("main_type", "-")
        ms2   = mp2.get("main_score", 0) or 0
        fbp2  = ef2.get("fallback_pct", "-")
        fbr2  = ef2.get("fallback_rank", "-")
        dh_ok2  = dh2.get("is_valid", False)
        dh_top2 = dh2.get("top_waku", "-")
        dh_sc2  = dh2.get("top_score", 0) or 0

        hs2      = bet.get("honmei_scenario") or {}
        hs2_pats = hs2.get("honmei_patterns", {}) or {}
        hs2_hp   = hs2_pats.get("honmei") or {}
        hs2_sj   = hs2.get("scenario_judgment", {}) or {}
        hs2_narr = hs2_hp.get("win_narrative", "")
        hs2_conf = hs2.get("confidence")
        hs2_conf_str = f"{hs2_conf*100:.0f}%" if hs2_conf is not None else "-"
        hs2_reas = (hs2_sj.get("reasons") or [""])[1:3]
        hs2_reas_str = "\n".join(f"  {r}" for r in hs2_reas)

        # 実際の買い目1着分布を集計してシナリオとの整合性を明示
        _cands2 = bet.get("candidates", [])
        from collections import Counter as _C2
        _fd2 = _C2(
            c["combo"].split("-")[0] for c in _cands2
            if c.get("combo") and not c.get("is_fallback_bet") and not c.get("is_dh_bet")
        )
        _w1c2  = _fd2.get("1", 0)
        _flyc2 = sum(v for k, v in _fd2.items() if k != "1")
        _topf2 = _fd2.most_common(1)[0][0] if _flyc2 > 0 else "-"
        _stn2  = bet.get("scenario_type_note", "")
        # 買い目軸サマリー（実態）
        if _w1c2 > 0 and _flyc2 > 0:
            _axis_summary = f"軸: 1号頭{_w1c2}点 / {_topf2}号頭{_flyc2}点"
        elif _w1c2 > 0:
            _axis_summary = f"軸: 1号頭{_w1c2}点（逃げ一本）"
        elif _flyc2 > 0:
            _axis_summary = f"軸: {_topf2}号頭{_flyc2}点（飛び一本）"
        else:
            _axis_summary = "軸: -"

        txt = (
            f"{tp2_icon}【展開:{tp2}】{stype}\n"
            f"逃げ{s1p*100:.0f}% / 飛び{flyp*100:.0f}%（主:{fly_str}）\n"
            f"─────────────────\n"
            f"? 逃げ力: {epct2}【{er2}】\n"
            f"? 主役:   {mw2}号【{mt2}】{ms2*100:.0f}%\n"
            f"? 残存:   {fbp2}【{fbr2}】\n"
            f"? 穴:     {'[OK]' + str(dh_top2) + '号 ' + f'{dh_sc2*100:.0f}%' if dh_ok2 else '─'}\n"
            f"─────────────────\n"
            f"買い目: {_axis_summary}\n"
            + (f"[!] {_stn2}\n" if _stn2 else "")
            + f"方針: {tp2_pol}\n"
            f"─ 潰れ展開(SC) ─\n"
            f"飛び役:{sc_fly_type} / 漁夫:{gyofu_str}\n"
            f"─ ◎勝ちパターン ─\n"
            f"{hs2_narr} 信頼度{hs2_conf_str}\n"
            f"{hs2_reas_str}"
        )
        # 変更4: tenkai_pattern（A/B/C/D）で色分け（Step1の展開パターン行と統一）
        fc = {
            "A": "FFE2EFDA",   # 緑（鉄板逃げ）
            "B": "FFFCE4D6",   # 橙（主役展開）
            "C": "FFFFF2CC",   # 黄（拮抗）
            "D": "FFEDE7F6",   # 薄紫（荒れ）
        }.get(tp2, "FFFFF2CC")
        wc(row, 4+i, txt,
           fill=sf(fc), font=fn(bold=True, size=9),
           align=al(h="left", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 56.0
    row += 1

    # ─ 変更5: 展示後チェックポイント行（買い目候補の直上） ─────────────────────
    FILL_CHECK = "FF4A235A"   # 深紫（展示後確認専用色）
    wc(row, 1, "展示後\n確認",
       fill=sf(FILL_CHECK), font=fn(bold=True, size=8, color="FFFFFFFF"), align=al(wrap=True))
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "展示後チェックポイント",
       fill=sf(FILL_CHECK), font=fn(bold=True, size=8, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        rj_c  = rd.get("race_judgment", {}) or {}
        bet_c = rd.get("bet_suggestions", {}) or {}
        tp_c  = bet_c.get("tenkai_pattern", rj_c.get("tenkai_pattern", "?"))
        mw_c  = (rj_c.get("main_player", {}) or {}).get("main_waku", "-")
        mt_c  = (rj_c.get("main_player", {}) or {}).get("main_type", "-")
        er_c  = (rj_c.get("w1_escape", {}) or {}).get("escape_rank", "-")
        ha_c  = (rj_c.get("himo_are", {}) or {}).get("tenji_trigger", "")
        _CHECK_MSG = {
            "A": f"1号艇ST確認（逃げ力【{er_c}】→ 出遅れなければ軸固定）",
            "B": f"{mw_c}号艇({mt_c})の展示タイム確認（主役スタート・伸び足が鍵）",
            "C": f"1号艇 vs {mw_c}号艇のST比較（拮抗→展示で軸を絞る）",
            "D": f"展示タイム全艇チェック（荒れ展開→突出艇を再確認）",
        }
        check_msg = _CHECK_MSG.get(tp_c, "展示タイム・ST確認後に軸決定")
        if ha_c:
            check_msg += f"\n[表] {ha_c.split(chr(10))[0]}"
        wc(row, 4+i, check_msg,
           fill=sf("FFFBEAFE"), font=fn(bold=False, size=8, color="FF4A235A"),
           align=al(h="left", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 36.0
    row += 1

    # ─ 買い目候補行（確率順・シナリオ別） ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "買い目候補（確率順）\nシナリオ種別 / 確率%",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al(wrap=True))
    max_lines_cand = 3
    for i, rd in enumerate(all_race_data):
        bet   = rd.get("bet_suggestions", {}) or {}
        rj_b  = rd.get("race_judgment",   {}) or {}
        cands = bet.get("candidates", [])
        # ── デバッグ: candidates の重複チェック ──
        _seen_debug = set()
        _dup_debug  = []
        for _c in cands:
            _ck = _c.get("combo", "")
            if _ck in _seen_debug:
                _dup_debug.append(_ck)
            _seen_debug.add(_ck)
        if _dup_debug:
            print(f"[DEBUG] candidates重複検出: {_dup_debug} (race={rd.get('race_no','-')})")
        # ────────────────────────────────────────
        if not cands:
            wc(row, 4+i, "候補なし",
               fill=sf("FFDDDDDD"), font=fn(bold=False, size=8, color="FF808080"),
               align=al(h="left", wrap=True), border=bdr)
            continue

        # ── 考察エンジン出力を収集（Step1→買い目の橋渡し用） ──────────────
        _tp_b    = bet.get("tenkai_pattern", rj_b.get("tenkai_pattern", "?"))
        _stype_b = bet.get("scenario_type", "-")
        _s1p_b   = bet.get("s1_prob", 0) or 0
        _flyp_b  = bet.get("fly_prob", 0) or 0
        _fly_ax  = bet.get("fly_axes", [])
        _mp_b    = rj_b.get("main_player", {}) or {}
        _mw_b    = _mp_b.get("main_waku", "-")
        _mt_b    = _mp_b.get("main_type", "-")
        _ms_b    = float(_mp_b.get("main_score", 0) or 0)
        _w1e_b   = rj_b.get("w1_escape", {}) or {}
        _er_b    = _w1e_b.get("escape_rank", "-")
        _ep_b    = _w1e_b.get("escape_pct", "-")
        _ef_b    = rj_b.get("escape_fallback", {}) or {}
        _fbr_b   = _ef_b.get("fallback_rank", "-")
        _fbp_b   = _ef_b.get("fallback_pct", "-")
        _dh_b    = rj_b.get("dark_horse", {}) or {}
        _dh_ok_b = _dh_b.get("is_valid", False)
        _dh_w_b  = _dh_b.get("top_waku", "-")

        # ── candidatesの実態から1着分布を集計（表示との乖離を防ぐ） ──────────
        # honmei_scenario等がbuy_listを差し替えた場合もここで実態を反映する
        _w1_count  = sum(1 for c in cands if c.get("combo","").split("-")[0] == "1"
                         and not c.get("is_fallback_bet") and not c.get("is_dh_bet"))
        _fly_count = sum(1 for c in cands if c.get("combo","").split("-")[0] != "1"
                         and not c.get("is_fallback_bet") and not c.get("is_dh_bet")
                         and not c.get("is_sc_bet"))
        # 実際の飛び軸（買い目に最も多く出てくる非1号頭の1着艇）
        from collections import Counter as _Counter
        _fly_first_dist = _Counter(
            c["combo"].split("-")[0] for c in cands
            if c.get("combo","").split("-")[0] != "1"
            and not c.get("is_fallback_bet") and not c.get("is_dh_bet")
        )
        _actual_fly_waku = _fly_first_dist.most_common(1)[0][0] if _fly_first_dist else _mw_b

        # 実態と考察の整合性チェック
        # scenario_typeが「逃げ軸流し」なのに1号頭買い目がゼロ → 矛盾を検出して表示
        _stype_display = _stype_b
        if _stype_b == "逃げ軸流し" and _w1_count == 0 and _fly_count > 0:
            _stype_display = f"[!]印◎軸（{_actual_fly_waku}号頭）※確率逃げ{_s1p_b*100:.0f}%"
        elif _stype_b == "逃げ軸流し" and _w1_count > 0:
            _stype_display = f"逃げ軸流し（1号頭{_w1_count}点+{_actual_fly_waku}号頭{_fly_count}点）"

        # 考察(1)(2)(3)(4) → 実際の買い目構成 の因果を1行に圧縮
        # ポイント: "実際の軸" を使う（main_fly_wakuではなく_actual_fly_wakuを使う）
        if _tp_b == "A" and _w1_count > 0:
            _bridge = f"(1)逃げ{_ep_b}【{_er_b}】× 逃げ{_s1p_b*100:.0f}% → 1号頭固定({_w1_count}点)"
        elif _w1_count == 0:
            # 1号頭が一切ない → 印◎軸が確率モデルを上回っている
            _honmei_w = next((r.get("waku") for r in rj_b.get("results",[]) if r.get("honmei") == "◎"), None) if rj_b.get("results") else None
            _h_label  = f"印◎{_honmei_w}号" if _honmei_w else f"印◎{_actual_fly_waku}号"
            _bridge = (
                f"{_h_label}頭軸（全{len(cands)}点）"
                f"  ※逃げ{_s1p_b*100:.0f}%だが印◎が{_actual_fly_waku}号 → 印優先"
            )
        elif _tp_b == "B":
            _fly_lbl = f"{_actual_fly_waku}号【{_mt_b}】{_ms_b*100:.0f}%"
            _bridge  = f"(2)主役{_fly_lbl}→{_actual_fly_waku}号頭{_fly_count}点  (3)残存{_fbp_b}【{_fbr_b}】{_w1_count}点"
        elif _tp_b == "C":
            _bridge = (
                f"逃げ{_s1p_b*100:.0f}%/飛び{_flyp_b*100:.0f}%拮抗"
                f" → 1号頭{_w1_count}点 vs {_actual_fly_waku}号頭{_fly_count}点"
            )
        elif _tp_b == "D":
            _dh_lbl = f"(4){_dh_w_b}号穴" if _dh_ok_b else "穴候補不明"
            _bridge = f"(1)逃げ力【{_er_b}】低・混戦 → 1号{_w1_count}点/{_actual_fly_waku}号{_fly_count}点 {_dh_lbl}"
        else:
            _bridge = f"1号頭{_w1_count}点 / {_actual_fly_waku}号頭{_fly_count}点"

        # セクションヘッダ用の考察橋渡しラベル（実際の軸を使う）
        _fly_w_lbl = f"{_actual_fly_waku}号{_mt_b}"
        _nige_why  = f"(1){_ep_b}【{_er_b}】逃げ{_s1p_b*100:.0f}%"
        _tobi_why  = f"(2){_fly_w_lbl} {_ms_b*100:.0f}%"
        _fb_why    = f"(3)残存{_fbp_b}【{_fbr_b}】1号2着狙い"
        _dh_why    = f"(4){_dh_w_b}号穴" if _dh_ok_b else "(4)穴"

        combo_map_local = {c["combo"]: c for c in cands}
        used_combos     = set()
        lines_nige      = []
        lines_tobi      = []
        lines_sc        = []
        lines_fallback  = []
        lines_dh        = []
        lines_other     = []

        def _append_to_section(line, sc, c):
            # フラグ優先（明示的に付与されたフラグで分類）
            if c.get("is_fallback_bet"):
                lines_fallback.append(line)
                return
            if c.get("is_dh_bet"):
                lines_dh.append(line)
                return
            if c.get("is_sc_bet") or "潰れ" in sc:
                lines_sc.append(line)
                return
            # first waku で正しく分類（scenarioテキストへの依存をやめる）
            # honmei_scenario版でscenario文字列が変わっても確実に機能する
            _first_w = c.get("combo", "-").split("-")[0] if c.get("combo") else sc
            if _first_w == "1":
                lines_nige.append(line)
            else:
                lines_tobi.append(line)

        # is_orkaeshi フラグ優先でペアを構築（フラグが消えた場合はフラグなし扱い）
        # integrate が candidates を差し替えても、直前の行6122でフラグを復元済み
        paired_cands = {}
        for c in cands:
            key = c["combo"]
            parts = key.split("-")
            if len(parts) != 3:
                continue
            f, s, t = parts
            if c.get("is_orkaeshi"):
                # 1着折り返し: A-B-C の本体は B-A-C
                base_key = f"{s}-{f}-{t}"
                if base_key in combo_map_local:
                    paired_cands[key]      = base_key
                    paired_cands[base_key] = key
            elif c.get("is_orkaeshi_23"):
                # 2着3着折り返し: A-B-C の本体は A-C-B
                base_key = f"{f}-{t}-{s}"
                if base_key in combo_map_local:
                    paired_cands[key]      = base_key
                    paired_cands[base_key] = key
        # ── デバッグ: paired_cands の内容を出力 ──
        if paired_cands:
            print(f"[DEBUG] paired_cands: {paired_cands} (race={rd.get('race_no','-')})")
        # ── デバッグ: cands の順番とフラグを出力 ──
        print(f"[DEBUG] cands order (race={rd.get('race_no','-')}): {[(c['combo'], c.get('is_orkaeshi'), c.get('is_orkaeshi_23')) for c in cands]}")
        # ── デバッグ: is_orkaeshiフラグが消えているcomboを検出 ──
        _orkaeshi_combos = {c["combo"] for c in cands if c.get("is_orkaeshi")}
        _orkaeshi_23_combos = {c["combo"] for c in cands if c.get("is_orkaeshi_23")}
        print(f"[DEBUG] is_orkaeshi={_orkaeshi_combos}, is_orkaeshi_23={_orkaeshi_23_combos} (race={rd.get('race_no','-')})")

        for c in cands:
            key = c["combo"]
            if key in used_combos:
                continue
            sc  = c.get("scenario", "")
            rsn = c.get("reason", "")
            # reason の1着根拠部分のみ抽出（「1号先行優位 / 2号残存」→「1号先行優位」）
            rsn_short = rsn.split(" / ")[0] if rsn else ""
            parts = key.split("-")

            if len(parts) != 3:
                used_combos.add(key)
                line = f"  {key}  {c['prob_pct']:.1f}%"
                if rsn_short:
                    line += f"\n    ← {rsn_short}"
                _append_to_section(line, sc, c)
                continue

            # 折り返しペアも単独行として個別表示（＝表記なし）
            used_combos.add(key)
            line = f"  {key}  {c['prob_pct']:.1f}%"
            if rsn_short:
                line += f"\n    ← {rsn_short}"
            _append_to_section(line, sc, c)

        # セクションヘッダに考察との橋渡しを付与
        sections = []
        if lines_nige:
            sections.append(f"── [緑]逃げ軸 {len(lines_nige)}点（{_nige_why}） ──")
            sections.extend(lines_nige)
        if lines_tobi:
            sections.append(f"── [赤]飛び軸 {len(lines_tobi)}点（{_tobi_why}） ──")
            sections.extend(lines_tobi)
        if lines_sc:
            sections.append(f"── [狙]潰れ受益 {len(lines_sc)}点 ──")
            sections.extend(lines_sc)
        if lines_fallback:
            sections.append(f"── (3)逃げ残存 {len(lines_fallback)}点（{_fb_why}） ──")
            sections.extend(lines_fallback)
        if lines_dh:
            sections.append(f"── (4)穴ヒモ {len(lines_dh)}点（{_dh_why}） ──")
            sections.extend(lines_dh)
        if lines_other:
            sections.append(f"── その他 {len(lines_other)}点 ──")
            sections.extend(lines_other)

        total  = len(cands)
        # 冒頭ブリッジ: 考察→シナリオ→買い目の一気通貫を1〜2行で表示
        header = (
            f"【計{total}点】{_stype_display}\n"
            f"? {_bridge}\n"
            f"＝折返 ?潰れ ?残存 ?穴"
        )
        txt = header + "\n" + "\n".join(sections)
        wc(row, 4+i, txt,
           fill=sf(FILL_JF_BET), font=fn(bold=False, size=8, color="FF1F3864"),
           align=al(h="left", wrap=True), border=bdr)
        max_lines_cand = max(max_lines_cand, txt.count("\n") + 1)
    ws.row_dimensions[row].height = max(13.0 * max_lines_cand, 80.0)
    row += 1

    # ─ 買い目シンプル表示行（combo順・確率%） ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "買い目リスト\n（combo順・確率%）",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al(wrap=True))
    max_lines_simple = 3
    for i, rd in enumerate(all_race_data):
        bet   = rd.get("bet_suggestions", {}) or {}
        cands = bet.get("candidates", [])

        skip        = bet.get("skip", False)
        skip_reason = bet.get("skip_reason", "")

        if not cands:
            wc(row, 4+i, "候補なし",
               fill=sf("FFDDDDDD"), font=fn(bold=False, size=8, color="FF808080"),
               align=al(h="left", wrap=True), border=bdr)
            continue
        def _combo_sort_key(c):
            parts = str(c.get("combo", "")).split("-")
            try:
                return [int(p) for p in parts]
            except ValueError:
                return [99, 99, 99]
        # 艇番若い順ソート
        sorted_cands = sorted(cands, key=lambda c: [int(x) for x in c.get("combo", "9-9-9").split("-")])
        # ── 折り返しペアをフラグベースで事前に対応付け ──────────────────────
        # is_orkaeshi フラグを持つ側が見つかった時点で両方向に登録する。
        # 「base_key not in paired」を外すことで、本体側が先に処理された場合でも
        # 折返側が後から正しく上書き登録される。
        combo_to_cand = {c["combo"]: c for c in sorted_cands}
        # is_orkaeshi フラグ優先でペアを構築（フラグ復元済みのため確実）
        paired = {}
        for c in sorted_cands:
            key = c["combo"]
            parts = key.split("-")
            if len(parts) != 3:
                continue
            f, s, t = parts
            if c.get("is_orkaeshi"):
                base_key = f"{s}-{f}-{t}"
                if base_key in combo_to_cand:
                    paired[key]      = base_key
                    paired[base_key] = key
            elif c.get("is_orkaeshi_23"):
                base_key = f"{f}-{t}-{s}"
                if base_key in combo_to_cand:
                    paired[key]      = base_key
                    paired[base_key] = key

        # ── ペア対応付けを使ってlines生成 ────────────────────────────────────
        used_simple = set()
        lines = []
        for c in sorted_cands:
            key = c["combo"]
            if key in used_simple:
                continue
            parts = key.split("-")
            if len(parts) != 3:
                used_simple.add(key)
                lines.append(f"{key}（{c['prob_pct']:.1f}%）")
                continue
            f, s, t = parts

            # 折り返しペアも単独行として個別表示（＝表記なし）
            used_simple.add(key)
            lines.append(f"{key}（{c['prob_pct']:.1f}%）")
        total = len(sorted_cands)
        eg       = bet.get("entry_grade") or {}
        eg_grade = eg.get("grade", "?")
        eg_bg, eg_fc = eg.get("fill", (FILL_JF_BET, "FF1F3864"))

        if skip:
            # 見送りでもcomboのみ表示（ヘッダーなし）
            txt        = "\n".join(lines)
            cell_fill  = "FFFCE4D6"
            cell_color = "FF7F0000"
        else:
            # グレード・ROIバッジなし。comboのみ表示
            txt        = "\n".join(lines)
            cell_fill  = eg_bg
            cell_color = eg_fc
        wc(row, 4+i, txt,
           fill=sf(cell_fill), font=fn(bold=True, size=9, color=cell_color),
           align=al(h="left", wrap=True), border=bdr)
        max_lines_simple = max(max_lines_simple, txt.count("\n") + 1)
    ws.row_dimensions[row].height = max(13.0 * max_lines_simple, 60.0)
    row += 1

    # ─ 理想合成オッズ行 ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "理想合成オッズ",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        bet  = rd.get("bet_suggestions", {}) or {}
        tso  = bet.get("theory_syn_odds")
        mr   = bet.get("margin_ratio")
        skip = bet.get("skip", False)
        skip_reason = bet.get("skip_reason", "")

        if skip:
            # 見送りレースは合成オッズ行に何も表示しない
            # （買い目リスト行に見送り理由＋参考買い目が既に表示されているため）
            txt    = "-"
            bg     = "FFDDDDDD"
            fc_col = "FF808080"
        elif tso:
            ev_warn     = bet.get("ev_warning", False)
            ev_warn_msg = bet.get("ev_warning_msg", "")
            if ev_warn:
                txt = (
                    f"理想合成オッズ\n{tso}倍\n"
                    f"[!] 期待値基準を下回っています\n"
                    f"回収重視→見送り推奨\n"
                    f"的中重視→参考買い目を使用可"
                )
                bg, fc_col = "FFFFF2CC", "FF7F3F00"
            else:
                # 【(9)】ケリー賭け比率を合成オッズ行に追記
                _kelly2 = bet.get("kelly") or {}
                _kelly_pct = _kelly2.get("kelly_pct", "")
                _kelly_label = _kelly2.get("kelly_label", "")
                _kelly_line = f"\n推奨賭け比率: {_kelly_pct}（{_kelly_label}）" if _kelly_pct else ""
                txt = f"理想合成オッズ\n{tso}倍{_kelly_line}"
                _eg2 = bet.get("entry_grade") or {}
                _g2  = _eg2.get("grade", "?")
                _b2, _f2 = _eg2.get("fill", ("", ""))
                if _g2 in ("S", "A") and _b2:
                    bg, fc_col = _b2, _f2
                elif _g2 == "B":
                    bg, fc_col = "FFFFF2CC", "FF7F3F00"
                elif _g2 in ("C", "D"):
                    bg, fc_col = "FFFCE4D6", "FF7F0000"
                elif mr and mr >= 2.0:
                    bg, fc_col = "FFE2EFDA", "FF1D5730"
                elif mr and mr >= 1.2:
                    bg, fc_col = "FFFFF2CC", "FF7F3F00"
                else:
                    bg, fc_col = "FFFCE4D6", "FF7F0000"
        else:
            txt    = "計算不能"
            bg     = "FFDDDDDD"
            fc_col = "FF808080"
        wc(row, 4+i, txt,
           fill=sf(bg), font=fn(bold=True, size=16 if not skip else 11, color=fc_col),
           align=al(h="center", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 42.0
    row += 1

    # ─ 展示後判断ガイド行 ─
    wc(row, 1, "展示前", fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    wc(row, 2, "展示後の判断フロー",
       fill=sf(FILL_JF_SUB), font=fn(bold=True, color="FFFFFFFF"), align=al())
    for i, rd in enumerate(all_race_data):
        bet   = rd.get("bet_suggestions", {}) or {}
        stype = bet.get("scenario_type", "-")
        if stype == "逃げ軸流し":
            guide = "(1) 展示で1号艇の伸び確認\n   良ければそのまま採用\n   悪ければ飛び組に組み替え"
        elif stype == "飛び軸":
            guide = "(1) 展示で1号艇の伸び確認\n   悪ければ飛び組採用\n   良ければ逃げ組に組み替え"
        else:
            guide = "(1) 展示で1号艇の伸び確認\n   良→逃げ組採用\n   悪→飛び組採用"
        txt = (
            f"{guide}\n"
            f"(2) 実オッズで合成オッズを計算\n"
            f"(3) 必要合成オッズと比較 → Step3"
        )
        wc(row, 4+i, txt,
           fill=sf(FILL_JF_NOTE), font=fn(bold=False, size=9, color="FF7F3F00"),
           align=al(h="left", wrap=True), border=bdr)
    ws.row_dimensions[row].height = 80.0
    row += 1

    # ====================================================================



    # FLY退避値の復元ログ
    if fly_input_backup:
        print(f"  ??  FLY手入力値を復元済み: {len(fly_input_backup)}件")

# ============================================================
# 予想ログ保存（Step0 / refine_tenji.py 連携用）
# ============================================================
