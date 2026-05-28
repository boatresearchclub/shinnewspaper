# -*- coding: utf-8 -*-
"""
backtest_engine.py  v4  ★修正版★
======================
【v4 修正内容】
  修正①: 補正係数テーブルを「補正前の生データ」から生成するように変更
          - generate_correction_table() に bt_raw（補正未適用の突合結果）を渡す
          - 蓄積CSVの buy_list は書き換えず、バックテスト内で仮想補正を評価専用に適用
          - これにより apply_correction → backtest → apply_correction... の
            自己参照ループ汚染が解消される

  修正②: EMA の起点を「補正前バックテスト」に固定
          - generate_correction_table() が毎回「素の統計値」を受け取るため
            ループを回しても誤差が累積しない

  修正③: _find_band の戻り値を統一
          - backtest 内部の脅威/脆弱性バンド集計用ヘルパを float 戻り値版に統一

【機能】
  1. 起動時に data_csv/ フォルダを自動スキャン
  2. 蓄積CSV（数値蓄積/*.csv）と結果CSV（data_csv/*_payouts.csv）を突合
  3. 新しい月のCSVが追加されていれば自動取り込み
  4. 精度分析 → correction_table.json（補正係数）を自動更新
  5. backtest_report.xlsx にレポート出力

【使い方】
  python backtest_engine.py              # 全会場・全月
  python backtest_engine.py --venue びわこ  # 特定会場のみ
  python backtest_engine.py --force      # キャッシュ無視して全再計算
"""

import os, sys, glob, json, argparse, itertools, re, shutil
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np

# ══════════════════════════════════════════════════════════
# パス定数（環境に合わせて変更）
# ══════════════════════════════════════════════════════════
BASE_DIR        = Path(r"C:\Users\user\Desktop\データ収集")
CHIKUSEKI_DIR   = BASE_DIR / "scripts" / "数値蓄積"
RESULTS_DIR     = BASE_DIR / "data_csv"
OUTPUT_EXCEL    = BASE_DIR / "scripts" / "backtest_report.xlsx"
CORRECTION_JSON = BASE_DIR / "scripts" / "correction_table.json"
CACHE_FILE      = BASE_DIR / "scripts" / ".backtest_cache.json"
BT_CACHE_CSV    = BASE_DIR / "scripts" / ".backtest_results_cache.csv"

# ★修正①: 補正前の突合結果キャッシュを別ファイルで保持する
#   - BT_RAW_CACHE_CSV: apply_correction を一切かけていない「生」の突合結果
#   - backtest_engine はここから補正係数を計算する
#   - apply_correction.py がこのファイルを書き換えることは絶対にない
BT_RAW_CACHE_CSV = BASE_DIR / "scripts" / ".backtest_raw_cache.csv"

MARK_PRIORITY = {"◎": 1, "○": 2, "▲": 3, "△": 4, "-": 9}


def sep(c="=", w=55):
    print(c * w)

def _load_json(path):
    if Path(path).exists():
        with open(str(path), encoding="utf-8") as f:
            return json.load(f)
    return {}

def _save_json(path, data):
    with open(str(path), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ══════════════════════════════════════════════════════════
# Step1: data_csv/ スキャン
# ══════════════════════════════════════════════════════════

def scan_available_months(results_dir):
    files = glob.glob(str(results_dir / "*_payouts.csv"))
    months = []
    for f in files:
        m = re.search(r"(\d{6})_payouts\.csv", Path(f).name)
        if m:
            months.append(m.group(1))
    return sorted(months)


def collect_all_results(results_dir, months):
    all_res, all_pay = [], []
    for ym in months:
        pay_f = results_dir / f"{ym}_payouts.csv"
        res_f = results_dir / f"{ym}_results.csv"
        if pay_f.exists():
            all_pay.append(pd.read_csv(str(pay_f), encoding="utf-8"))
        else:
            print(f"    [!] {pay_f.name} なし")
        if res_f.exists():
            all_res.append(pd.read_csv(str(res_f), encoding="utf-8"))
        else:
            print(f"    [!] {res_f.name} なし")
    res_df = pd.concat(all_res, ignore_index=True).drop_duplicates() if all_res else pd.DataFrame()
    pay_df = pd.concat(all_pay, ignore_index=True).drop_duplicates() if all_pay else pd.DataFrame()
    return res_df, pay_df


# ══════════════════════════════════════════════════════════
# Step2: 蓄積CSV読み込み
# ══════════════════════════════════════════════════════════

def load_chikuseki(chikuseki_dir, venue_filter=None):
    pattern = f"{venue_filter}.csv" if venue_filter else "*.csv"
    files = sorted(glob.glob(str(chikuseki_dir / pattern)))
    if not files:
        raise FileNotFoundError(f"蓄積CSVなし: {chikuseki_dir / pattern}")
    dfs = []
    for f in files:
        try:
            tmp = pd.read_csv(f, encoding="utf-8")
            tmp["_venue_file"] = Path(f).stem
            dfs.append(tmp)
            print(f"    {Path(f).name}: {len(tmp)}行")
        except Exception as e:
            print(f"    [!] スキップ: {Path(f).name} ({e})")
    df = pd.concat(dfs, ignore_index=True)
    if "会場" not in df.columns:
        df["会場"] = df["_venue_file"]
    print(f"  [OK] 蓄積CSV合計: {len(df)}行 / {df['会場'].nunique()}会場")
    return df


# ══════════════════════════════════════════════════════════
# Step3: 買い目生成（近似再現 fallback 用）
# ══════════════════════════════════════════════════════════

def build_bets(marks, strategy, skip):
    if skip == 1:
        return []
    sorted_waku = sorted(
        [w for w, m in marks.items() if m != "-"],
        key=lambda w: MARK_PRIORITY.get(marks.get(w, "-"), 9)
    )
    honmei  = [w for w in sorted_waku if marks.get(w) == "◎"]
    taikou  = [w for w in sorted_waku if marks.get(w) == "○"]
    sashi   = [w for w in sorted_waku if marks.get(w) == "▲"]
    renpuku = [w for w in sorted_waku if marks.get(w) == "△"]
    himo    = taikou + sashi + renpuku
    bets = []

    if "◎1着固定" in strategy or "5〜8点" in strategy:
        if honmei and himo:
            for second in himo:
                thirds = [w for w in sorted_waku if w not in honmei and w != second]
                for third in thirds[:3]:
                    bets.append(f"{honmei[0]}-{second}-{third}")

    elif "◎-○軸" in strategy or "10〜15点" in strategy:
        axes = honmei + taikou
        if len(axes) >= 2:
            a1, a2 = axes[0], axes[1]
            thirds = [w for w in sorted_waku if w not in [a1, a2]]
            for t in thirds[:4]:
                bets.append(f"{a1}-{a2}-{t}")
                bets.append(f"{a2}-{a1}-{t}")

    elif "◎○△フォーメーション" in strategy or "15〜20点" in strategy:
        top3 = (honmei + taikou + sashi + renpuku)[:3]
        if len(top3) >= 3:
            for perm in itertools.permutations(top3, 3):
                bets.append(f"{perm[0]}-{perm[1]}-{perm[2]}")

    elif "荒れ傾向" in strategy:
        if honmei and himo:
            others = [w for w in sorted_waku if w != honmei[0]]
            for second in others[:4]:
                for third in others:
                    if third != second:
                        bets.append(f"{honmei[0]}-{second}-{third}")

    seen, unique = set(), []
    for b in bets:
        if b not in seen:
            seen.add(b); unique.append(b)
    return unique


# ══════════════════════════════════════════════════════════
# ★修正③: _find_band を float 戻り値に統一（モジュールレベル）
#
# ・旧バージョンでは apply_correction.py に2種類の _find_band が存在し、
#   一方は float、もう一方は文字列キーを返していた。
# ・バックテスト内部での集計（analyze）は文字列キーを必要とするため
#   _find_band_key（文字列キー戻り）と _find_band_weight（float戻り）を分離。
# ══════════════════════════════════════════════════════════

def _find_band_key(value: float, band_dict: dict) -> str:
    """バンドキー（文字列）を返す。analyze() 内のグループ集計用。"""
    for band_key in band_dict:
        try:
            lo_str, hi_str = band_key.split("〜")
            if float(lo_str) <= value < float(hi_str):
                return band_key
        except Exception:
            continue
    return ""


def _find_band_weight(value: float, band_dict: dict) -> float:
    """バンドに対応する重み（float）を返す。スコア計算用。"""
    for band_key, w in band_dict.items():
        try:
            lo_str, hi_str = band_key.split("〜")
            if float(lo_str) <= value < float(hi_str):
                return float(w)
        except Exception:
            continue
    return 1.0


# ══════════════════════════════════════════════════════════
# Step4: 突合
# ══════════════════════════════════════════════════════════

def _build_result_index(res_df: pd.DataFrame, pay_df: pd.DataFrame) -> tuple[dict, dict]:
    top3_map = {}
    for (venue, date, rno), g in res_df.groupby(["会場名", "日付", "レース番号"]):
        top3 = (g.sort_values("着順")
                 .loc[g["着順"] <= 3, "艇番"]
                 .astype(str).tolist())
        top3_map[(venue, str(date), int(rno))] = top3

    pay3 = pay_df[pay_df["券種"] == "３連単"]
    payout_map = {}
    for _, r in pay3.iterrows():
        payout_map[(str(r["会場名"]), str(r["日付"]), int(r["レース番号"]))] = int(r["払戻金"])

    return top3_map, payout_map


def _has_real_bets(chikuseki_df):
    return "買い目" in chikuseki_df.columns


BT_CACHE_CSV  = Path(__file__).parent / ".bt_result_cache.csv"
BT_HASH_FILE  = Path(__file__).parent / ".bt_chikuseki_hash.json"
# ★修正①: 生データキャッシュのパスもスクリプト相対に統一
BT_RAW_CACHE_CSV = Path(__file__).parent / ".bt_raw_cache.csv"


def _calc_venue_hash(chikuseki_df: pd.DataFrame, venue: str) -> str:
    import hashlib
    venue_col = "会場" if "会場" in chikuseki_df.columns else "_venue_file"
    vdf = chikuseki_df[chikuseki_df[venue_col] == venue]
    if vdf.empty:
        return ""
    sensitive_cols = [c for c in [
        "買い目", "買い目_全", "シナリオ種別", "展開パターン",
        "見送りフラグ", "点数", "1号艇1着確率",
        "印", "レースランク", "レーススコア",
        "3連対指数", "オリジナル1着率", "脅威合計", "1号艇脆弱性",
    ] if c in vdf.columns]
    target = vdf[["日付", "レース番号", "枠番"] + sensitive_cols]
    buf = target.sort_values(["日付","レース番号","枠番"]).to_csv(index=False)
    return __import__("hashlib").md5(buf.encode()).hexdigest()[:16]


def _load_venue_hashes() -> dict:
    return _load_json(BT_HASH_FILE)

def _save_venue_hashes(hashes: dict):
    _save_json(BT_HASH_FILE, hashes)


def _fix_bt_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    for col in ["結果データあり"]:
        if col in df.columns:
            df[col] = df[col].map(
                lambda v: True  if str(v).strip().lower() == "true"
                          else False if str(v).strip().lower() == "false"
                          else None
            )
    if "的中" in df.columns:
        df["的中"] = df["的中"].map(
            lambda v: True  if str(v).strip().lower() == "true"
                      else False if str(v).strip().lower() == "false"
                      else None
        )
    for col in ["払戻金", "投資額", "回収額", "収支", "生成点数", "見送り推奨",
                "レーススコア", "脅威合計", "1号艇脆弱性", "最大脅威艇"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def _load_bt_cache() -> pd.DataFrame:
    if BT_CACHE_CSV.exists():
        try:
            df = pd.read_csv(str(BT_CACHE_CSV), encoding="utf-8-sig", dtype=str)
            return _fix_bt_dtypes(df)
        except Exception:
            pass
    return pd.DataFrame()


def _save_bt_cache(df: pd.DataFrame):
    try:
        df.to_csv(str(BT_CACHE_CSV), index=False, encoding="utf-8-sig")
    except Exception as e:
        print(f"  [!] キャッシュ保存失敗（続行）: {e}")


# ★修正①: 生データキャッシュの読み書き
def _load_raw_cache() -> pd.DataFrame:
    """補正未適用の突合結果キャッシュを読み込む"""
    if BT_RAW_CACHE_CSV.exists():
        try:
            df = pd.read_csv(str(BT_RAW_CACHE_CSV), encoding="utf-8-sig", dtype=str)
            return _fix_bt_dtypes(df)
        except Exception:
            pass
    return pd.DataFrame()


def _save_raw_cache(df: pd.DataFrame):
    """補正未適用の突合結果キャッシュを保存する"""
    try:
        df.to_csv(str(BT_RAW_CACHE_CSV), index=False, encoding="utf-8-sig")
    except Exception as e:
        print(f"  [!] 生データキャッシュ保存失敗（続行）: {e}")


def detect_changed_venues(chikuseki_df: pd.DataFrame,
                          new_months: list) -> tuple:
    venue_col  = "会場" if "会場" in chikuseki_df.columns else "_venue_file"
    all_venues = list(chikuseki_df[venue_col].unique())
    new_hashes = {v: _calc_venue_hash(chikuseki_df, v) for v in all_venues}
    old_hashes = _load_venue_hashes()
    new_ym_set = set(ym[:6] for ym in new_months)

    stale_venues = []
    fresh_venues = []

    for venue in all_venues:
        reasons = []
        if new_hashes[venue] != old_hashes.get(venue, ""):
            reasons.append("初回" if venue not in old_hashes else "指数/ロジック変更")
        if new_months:
            vdf = chikuseki_df[chikuseki_df[venue_col] == venue]
            venue_yms = set(
                str(d)[:7].replace("-", "") for d in vdf["日付"].unique()
            )
            hit_yms = venue_yms & new_ym_set
            if hit_yms:
                reasons.append(f"新規月({','.join(sorted(hit_yms))})")

        if reasons:
            stale_venues.append(venue)
            print(f"    [{venue}] 再処理: {' / '.join(reasons)}")
        else:
            fresh_venues.append(venue)
            print(f"    [{venue}] キャッシュ有効（スキップ）")

    return stale_venues, fresh_venues, new_hashes


def run_backtest(chikuseki_df: pd.DataFrame,
                 res_df: pd.DataFrame,
                 pay_df: pd.DataFrame,
                 new_months: list[str],
                 force: bool = False) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    ★修正①: 戻り値を (bt_corrected, bt_raw) のタプルに変更。

    bt_corrected : 蓄積CSVの買い目列をそのまま使った従来の突合結果（Excelレポート用）
    bt_raw       : 蓄積CSVの見送り推奨・買い目を「補正前オリジナル」として扱った突合結果
                   → generate_correction_table() に渡す（自己参照ループ防止）

    「補正前オリジナル」の定義:
      - 見送り推奨 = 0（全レース買い付けとして評価）
      - 買い目リスト = 蓄積CSVの「買い目_全」列があればそちら、なければ「買い目」列
        （apply_correction.py が書き換えるのは「買い目」列のみのため）
    """
    venue_col = "会場" if "会場" in chikuseki_df.columns else "_venue_file"
    has_real  = _has_real_bets(chikuseki_df)

    if has_real:
        print("  [OK] 実買い目カラムあり → scenario_engineの実出力で突合します")
    else:
        print("  [!] 実買い目カラムなし → 印+戦略から近似再現します（旧CSV）")

    top3_map, payout_map = _build_result_index(res_df, pay_df)

    new_dates: set[str] = set()
    for ym in new_months:
        y, m = int(ym[:4]), int(ym[4:])
        mask = (
            (res_df["日付"].str[:4].astype(int) == y) &
            (res_df["日付"].str[5:7].astype(int) == m)
        ) if len(res_df) else pd.Series([], dtype=bool)
        if mask.any():
            new_dates |= set(res_df.loc[mask, "日付"].unique())

    base_cols = ["日付","レース番号", venue_col,
                 "戦略","レースランク","レーススコア","見送り推奨",
                 "3択verdict","脅威合計","1号艇脆弱性","最大脅威艇"]
    if has_real:
        for col in ["買い目","買い目_全","買い目_シナリオ","点数","点数_全",
                    "1号艇1着確率","見送りフラグ","シナリオ種別","展開パターン",
                    "グレード"]:
            if col in chikuseki_df.columns:
                base_cols.append(col)

    race_df = (
        chikuseki_df.sort_values("枠番")
        .drop_duplicates(subset=["日付", "レース番号", venue_col])
        [[c for c in base_cols if c in chikuseki_df.columns]]
        .copy()
    )
    race_df = race_df.rename(columns={venue_col: "会場"})

    cached_bt  = pd.DataFrame() if force else _load_bt_cache()
    cached_raw = pd.DataFrame() if force else _load_raw_cache()

    if force:
        process_df   = race_df.copy()
        stale_venues = list(race_df["会場"].unique())
        fresh_venues = []
        new_hashes   = {v: _calc_venue_hash(chikuseki_df, v) for v in stale_venues}
        print(f"  [--force] 全{len(process_df)}R を再処理します")
    else:
        stale_venues, fresh_venues, new_hashes = detect_changed_venues(
            chikuseki_df, new_months
        )
        if stale_venues:
            process_df = race_df[race_df["会場"].isin(stale_venues)].copy()
            if not cached_bt.empty:
                cached_bt  = cached_bt[cached_bt["会場"].isin(fresh_venues)]
                cached_raw = cached_raw[cached_raw["会場"].isin(fresh_venues)] \
                             if not cached_raw.empty else pd.DataFrame()
            print(f"  再処理: {len(process_df)}R / キャッシュ流用: {len(cached_bt)}R")
        else:
            process_df = pd.DataFrame()
            print(f"  全会場キャッシュ有効 → 突合処理スキップ ({len(cached_bt)}R)")

    total  = len(process_df)
    new_bt = pd.DataFrame()
    new_raw= pd.DataFrame()

    if total > 0:
        df = process_df.copy()
        df["レース番号"] = df["レース番号"].astype(int)

        # ── 見送り判定（補正後・bt_corrected 用）──────────────────────
        if has_real and "見送りフラグ" in df.columns:
            df["_skip"] = (df["見送りフラグ"].astype(str) == "1").astype(int)
        else:
            df["_skip"] = pd.to_numeric(
                df.get("見送り推奨", 0), errors="coerce"
            ).fillna(0).astype(int)

        # ── 実結果・払戻（O(1)ルックアップ）──────────────────────────
        df["_key"]    = list(zip(df["会場"], df["日付"], df["レース番号"]))
        df["実結果"]  = df["_key"].map(
            lambda k: "-".join(top3_map[k]) if k in top3_map else "不明"
        )
        df["払戻金"]  = df["_key"].map(
            lambda k: payout_map.get(k, 0) if k in top3_map else 0
        )
        df["結果データあり"] = df["実結果"] != "不明"

        # ── 買い目取得（補正後・bt_corrected 用）──────────────────────
        if has_real and "買い目" in df.columns:
            df["買い目リスト"] = (
                df["買い目"].fillna("").astype(str)
                .where(lambda s: ~s.isin(["", "nan", "None"]), other="")
            )
            df["買い目ソース"] = "実買い目"
        else:
            _vcol = "会場" if "会場" in chikuseki_df.columns else "_venue_file"
            mark_index: dict = {}
            for _, ar in chikuseki_df.iterrows():
                k = (str(ar["日付"]), int(ar["レース番号"]), str(ar.get(_vcol, "")))
                w = str(int(ar["枠番"]))
                m = str(ar["印"]) if str(ar["印"]) not in ("nan","None","") else "-"
                mark_index.setdefault(k, {})[w] = m

            def _approx(row):
                k = (str(row["日付"]), int(row["レース番号"]), str(row["会場"]))
                bets = build_bets(mark_index.get(k, {}),
                                  str(row.get("戦略", "")), int(row["_skip"]))
                return "|".join(bets)

            df["買い目リスト"] = df.apply(_approx, axis=1)
            df["買い目ソース"] = "近似再現"

        # ── 点数・的中・収支（補正後）─────────────────────────────────
        df["生成点数"] = df["買い目リスト"].apply(
            lambda s: len([b for b in s.split("|") if b]) if s else 0
        )
        df["_bets_set"] = df["買い目リスト"].apply(
            lambda s: set(s.split("|")) - {""} if s else set()
        )
        df["的中"] = df.apply(
            lambda r: (r["実結果"] in r["_bets_set"]) if r["結果データあり"] else None,
            axis=1
        )
        df["的中買い目"] = df.apply(
            lambda r: r["実結果"] if r["的中"] is True else "", axis=1
        )
        df["投資額"] = df["生成点数"] * 100
        df["回収額"] = df.apply(
            lambda r: int(r["払戻金"]) if r["的中"] is True else 0, axis=1
        )
        df["収支"] = df["回収額"] - df["投資額"]

        for col in ["シナリオ種別", "展開パターン"]:
            if col not in df.columns:
                df[col] = ""

        new_bt = df[[
            "日付","会場","レース番号","戦略","レースランク","レーススコア",
            "_skip","3択verdict","シナリオ種別","展開パターン",
            "脅威合計","1号艇脆弱性","最大脅威艇",
            "買い目ソース","生成点数","買い目リスト",
            "実結果","的中","的中買い目","払戻金","投資額","回収額","収支",
            "結果データあり",
        ]].rename(columns={"_skip": "見送り推奨"}).copy()

        # ══════════════════════════════════════════════════════
        # ★修正①: bt_raw を生成
        #
        # 「補正前」の定義:
        #   - 見送り推奨 = 0 （全レース対象）
        #   - 買い目 = 「買い目_全」列（apply_correction が書き換えない元の全買い目）
        #             なければ「買い目」列をそのまま使用
        #
        # 目的: generate_correction_table() に渡す素の統計を汚染しない。
        # ══════════════════════════════════════════════════════
        raw = df.copy()

        # 見送り推奨を強制的に 0（全レース評価）
        raw["_raw_skip"] = 0

        # 買い目_全があればそちらを使う（補正前の全買い目候補）
        if "買い目_全" in raw.columns:
            raw["_raw_bets"] = raw["買い目_全"].fillna("").astype(str).where(
                lambda s: ~s.isin(["", "nan", "None"]), other=""
            )
            raw["買い目ソース"] = "実買い目_全"
        else:
            # フォールバック: 現在の「買い目リスト」をそのまま使う
            raw["_raw_bets"] = raw["買い目リスト"]

        raw["_raw_bets_set"] = raw["_raw_bets"].apply(
            lambda s: set(s.split("|")) - {""} if s else set()
        )
        raw["生成点数_raw"] = raw["_raw_bets"].apply(
            lambda s: len([b for b in s.split("|") if b]) if s else 0
        )
        raw["的中_raw"] = raw.apply(
            lambda r: (r["実結果"] in r["_raw_bets_set"]) if r["結果データあり"] else None,
            axis=1
        )
        raw["投資額_raw"] = raw["生成点数_raw"] * 100
        raw["回収額_raw"] = raw.apply(
            lambda r: int(r["払戻金"]) if r["的中_raw"] is True else 0, axis=1
        )
        raw["収支_raw"] = raw["回収額_raw"] - raw["投資額_raw"]

        # bt_raw は生成点数・的中・投資・回収を _raw 版で上書きしたコピー
        raw_out = raw.copy()
        raw_out["見送り推奨"] = 0  # 全レース対象
        raw_out["生成点数"]   = raw["生成点数_raw"]
        raw_out["買い目リスト"] = raw["_raw_bets"]
        raw_out["的中"]       = raw["的中_raw"]
        raw_out["的中買い目"] = raw.apply(
            lambda r: r["実結果"] if r["的中_raw"] is True else "", axis=1
        )
        raw_out["投資額"]     = raw["投資額_raw"]
        raw_out["回収額"]     = raw["回収額_raw"]
        raw_out["収支"]       = raw["収支_raw"]

        new_raw = raw_out[[
            "日付","会場","レース番号","戦略","レースランク","レーススコア",
            "見送り推奨","3択verdict","シナリオ種別","展開パターン",
            "脅威合計","1号艇脆弱性","最大脅威艇",
            "買い目ソース","生成点数","買い目リスト",
            "実結果","的中","的中買い目","払戻金","投資額","回収額","収支",
            "結果データあり",
        ]].copy()

        print(f"  [OK] {total}R 突合完了")

    # ── キャッシュ＋差分を結合 ──────────────────────────────────────
    def _merge(cached, new):
        if not cached.empty and not new.empty:
            return pd.concat([cached, new], ignore_index=True)
        elif not cached.empty:
            return cached.copy()
        else:
            return new.copy()

    bt  = _merge(cached_bt,  new_bt)
    raw = _merge(cached_raw, new_raw)

    if not new_bt.empty:
        _save_bt_cache(bt)
        _save_raw_cache(raw)   # ★修正①: 生データキャッシュも保存
        old_hashes = _load_venue_hashes()
        old_hashes.update(new_hashes)
        _save_venue_hashes(old_hashes)
        print(f"  [OK] キャッシュ更新: {len(bt)}R（再処理 {len(new_bt)}R）")
        print(f"       生データキャッシュ: {len(raw)}R")
        print(f"       ハッシュ更新会場: {sorted(new_hashes.keys())}")

    return bt, raw


# ══════════════════════════════════════════════════════════
# Step5: 精度分析
# ══════════════════════════════════════════════════════════

def analyze(bt):
    """
    ★修正①: analyze() 自体は変更なし。
    呼び出し元が bt_raw を渡すか bt_corrected を渡すかで意味が変わる。
    - correction_table 生成用 → bt_raw を渡す（main() で使い分け）
    - Excelレポート用         → bt_corrected を渡す
    """
    bt = _fix_bt_dtypes(bt.copy())
    bt_valid = bt[bt["結果データあり"] == True].copy()
    bt_buy   = bt_valid[bt_valid["見送り推奨"] == 0].copy()

    def stats(df):
        n = len(df)
        if n == 0:
            return {"レース数":0,"的中数":0,"的中率":0,"総投資":0,"総回収":0,"ROI":0}
        hits = df["的中"].sum(); inv = df["投資額"].sum(); ret = df["回収額"].sum()
        return {
            "レース数": n, "的中数": int(hits),
            "的中率": round(hits/n*100, 1),
            "総投資": int(inv), "総回収": int(ret),
            "ROI": round((ret-inv)/inv*100, 1) if inv > 0 else 0,
        }

    s = {}
    s["全体"]         = stats(bt_buy)
    s["会場別"]       = {v: stats(g) for v, g in bt_buy.groupby("会場")}
    s["戦略別"]       = {v: stats(g) for v, g in bt_buy.groupby("戦略")}
    s["3択verdict別"] = {str(v): stats(g) for v, g in bt_buy.groupby("3択verdict")}
    s["レースランク別"] = {r: stats(bt_buy[bt_buy["レースランク"]==r]) for r in ["S","A","B","C","D"]}
    s["スコア帯別"]   = {}
    for lo, hi in [(0,30),(30,50),(50,70),(70,100),(100,999)]:
        g = bt_buy[(bt_buy["レーススコア"]>=lo)&(bt_buy["レーススコア"]<hi)]
        s["スコア帯別"][f"{lo}〜{hi}"] = stats(g)
    s["脅威スコア帯別"] = {}
    for lo, hi in [(0,12),(12,16),(16,20),(20,999)]:
        g = bt_buy[(bt_buy["脅威合計"]>=lo)&(bt_buy["脅威合計"]<hi)]
        s["脅威スコア帯別"][f"{lo}〜{hi}"] = stats(g)
    s["1号艇脆弱性帯別"] = {}
    for lo, hi in [(0,25),(25,30),(30,35),(35,100)]:
        g = bt_buy[(bt_buy["1号艇脆弱性"]>=lo)&(bt_buy["1号艇脆弱性"]<hi)]
        s["1号艇脆弱性帯別"][f"{lo}〜{hi}"] = stats(g)

    if "シナリオ種別" in bt_buy.columns and bt_buy["シナリオ種別"].str.strip().ne("").any():
        s["シナリオ種別別"] = {
            str(v): stats(g)
            for v, g in bt_buy[bt_buy["シナリオ種別"].str.strip().ne("")].groupby("シナリオ種別")
        }
    if "展開パターン" in bt_buy.columns and bt_buy["展開パターン"].str.strip().ne("").any():
        s["展開パターン別"] = {
            str(v): stats(g)
            for v, g in bt_buy[bt_buy["展開パターン"].str.strip().ne("")].groupby("展開パターン")
        }
    bt_skip = bt_valid[bt_valid["見送り推奨"]==1]
    if len(bt_skip):
        n_skip = len(bt_skip); n_rough = (bt_skip["払戻金"] >= 3000).sum()
        s["見送り精度"] = {
            "見送りレース数": n_skip,
            "そのうち荒れ(3000円超)": int(n_rough),
            "見送り正解率(荒れ回避)": round(n_rough/n_skip*100,1),
        }
    return s


# ══════════════════════════════════════════════════════════
# Step6: 補正係数テーブル
# ══════════════════════════════════════════════════════════

def generate_correction_table(summary_raw: dict, summary_corrected: dict | None = None):
    """
    ★修正①②: correction_table は summary_raw（補正前統計）から生成する。

    Parameters
    ----------
    summary_raw       : bt_raw から生成した summary（補正前・全レース対象）
    summary_corrected : bt_corrected から生成した summary（補正後・Excelレポートと同じ）
                        省略可。省略時は summary_raw を両方に使う。

    EMAの起点:
      - 毎回 summary_raw（補正前の素の統計）から新値を計算するため、
        ループを何度回しても誤差は累積しない（★修正②）。
      - 前回 correction_table.json があれば EMA で滑らかに更新する。
    """
    base = summary_raw.get("全体", {}); base_hit = base.get("的中率", 0)

    def weight(hr):
        if base_hit == 0: return 1.0
        return round(max(0.5, min(2.0, hr / base_hit)), 3)

    score_threshold = 0
    for band, s in sorted(summary_raw.get("スコア帯別", {}).items()):
        lo = int(band.split("〜")[0])
        if s.get("ROI", -999) > -10:
            score_threshold = lo; break

    # ── EMA更新: 前回の correction_table.json を読んで平滑化 ─────────
    # ★修正②: EMAの起点は常に「補正前バックテスト」の新値。
    #           prev（前回テーブル）が存在する場合のみ平滑化する。
    #           これによりループ回数に依存した係数のドリフトを防ぐ。
    EMA_ALPHA = 0.3
    prev = _load_json(CORRECTION_JSON)

    def ema_weight(new_w: float, key: str, axis: str) -> float:
        old_w = prev.get(axis, {}).get(key)
        if old_w is None:
            return new_w
        return round(old_w * (1 - EMA_ALPHA) + new_w * EMA_ALPHA, 3)

    MIN_VENUE_SAMPLES = 30

    def venue_weight_safe(venue: str, s: dict) -> float:
        if s.get("レース数", 0) < MIN_VENUE_SAMPLES:
            return 1.0
        raw = weight(s.get("的中率", 0))
        return ema_weight(raw, venue, "venue_weight")

    venue_stats = {}
    global_races = base.get("レース数", 0)

    # ★修正①: venue_stats の数値も補正前 summary_raw ベースで算出
    for v, s in summary_raw.get("会場別", {}).items():
        n     = s.get("レース数", 0)
        hr    = s.get("的中率", 0)
        roi   = s.get("ROI", 0)
        inv   = s.get("総投資", 0)
        ret   = s.get("総回収", 0)
        recovery = round(ret / inv, 4) if inv > 0 else 0.0
        p = hr / 100.0
        reliability = round(1.0 - (1.96 * ((p * (1 - p) / n) ** 0.5) / max(p, 0.001)), 3) \
                      if n >= 5 else 0.0
        needs_boost   = (hr < base_hit - 3.0) and (n >= MIN_VENUE_SAMPLES)
        insufficient  = n < MIN_VENUE_SAMPLES

        venue_stats[v] = {
            "n_races":           n,
            "hit_rate":          hr,
            "recovery_rate":     recovery,
            "roi":               roi,
            "reliability":       max(0.0, reliability),
            "weight_multiplier": venue_weight_safe(v, s),
            "needs_boost":       needs_boost,
            "insufficient_data": insufficient,
            "judgment": ("⚠ データ不足" if insufficient
                         else "❌ 要強化" if needs_boost
                         else "✅ 維持"),
        }

    new_table = {
        "generated_at":    datetime.now().strftime("%Y-%m-%d %H:%M"),
        # ★修正①: base_hit_rate / base_roi / base_recovery は補正前の値
        "base_hit_rate":   base.get("的中率", 0),
        "base_roi":        base.get("ROI", 0),
        "base_recovery":   round(base.get("総回収", 0) / base.get("総投資", 1), 4)
                           if base.get("総投資", 0) > 0 else 0.0,
        "total_races":     base.get("レース数", 0),
        "score_threshold": score_threshold,
        # ★修正①: rank/strategy/threat/vuln の各重みも補正前統計から算出
        "rank_weight":     {r: ema_weight(weight(s.get("的中率", 0)), r, "rank_weight")
                            for r, s in summary_raw.get("レースランク別", {}).items()},
        "strategy_weight": {k: ema_weight(weight(s.get("的中率", 0)), k, "strategy_weight")
                            for k, s in summary_raw.get("戦略別", {}).items()},
        "threat_penalty":  {k: ema_weight(weight(s.get("的中率", 0)), k, "threat_penalty")
                            for k, s in summary_raw.get("脅威スコア帯別", {}).items()},
        "vuln_penalty":    {k: ema_weight(weight(s.get("的中率", 0)), k, "vuln_penalty")
                            for k, s in summary_raw.get("1号艇脆弱性帯別", {}).items()},
        "venue_weight":    {v: d["weight_multiplier"] for v, d in venue_stats.items()},
        "venue_stats":     venue_stats,
        "skip_accuracy":   summary_raw.get("見送り精度", {}).get("見送り正解率(荒れ回避)", 0),
        # ★修正①: 補正後の実績も参考値として記録（Excelレポートと対応）
        "corrected_hit_rate": (summary_corrected or {}).get("全体", {}).get("的中率", None),
        "corrected_roi":      (summary_corrected or {}).get("全体", {}).get("ROI", None),
        "comment": (
            "backtest_engine.py v4 自動生成。"
            "★補正前(bt_raw)統計から係数を計算。EMA平滑化(α=0.3)・最小nフィルタ(n<30→補正なし)適用済み。"
            "apply_correction.py が読み込む。"
        ),
    }
    return new_table


# ══════════════════════════════════════════════════════════
# Step7: Excelレポート（既存のまま・引数変更なし）
# ══════════════════════════════════════════════════════════

def write_report(bt, summary, correction, output_path):
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.cell.cell import WriteOnlyCell
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        print("  [!] pip install openpyxl が必要です"); return

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
    HIT_FILL  = PatternFill("solid", fgColor="C6EFCE")
    MISS_FILL = PatternFill("solid", fgColor="FCE4D6")
    SKIP_FILL = PatternFill("solid", fgColor="FFFFCC")
    ND_FILL   = PatternFill("solid", fgColor="EEEEEE")
    ROI_POS   = PatternFill("solid", fgColor="C6EFCE")
    ROI_NEG   = PatternFill("solid", fgColor="FCE4D6")
    SEC_FILL  = PatternFill("solid", fgColor="D9E1F2")
    RED_FILL  = PatternFill("solid", fgColor="FCE4D6")
    THIN      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_ALIGN = Alignment(horizontal="center", wrap_text=True)

    def _hdr_row(ws, values):
        row = []
        for v in values:
            c = WriteOnlyCell(ws, value=v)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = HDR_ALIGN; c.border = BORDER
            row.append(c)
        ws.append(row)

    def _data_row(ws, values, fill=None):
        row = []
        for v in values:
            c = WriteOnlyCell(ws, value=v)
            if fill: c.fill = fill
            c.border = BORDER
            row.append(c)
        ws.append(row)

    wb = Workbook(write_only=True)
    ws1 = wb.create_sheet("全レース突合")
    ws1.freeze_panes = "A2"

    sh1_cols = ["日付","会場","レース番号","戦略","レースランク","レーススコア",
                "見送り推奨","3択verdict","脅威合計","1号艇脆弱性","最大脅威艇",
                "生成点数","実結果","的中","的中買い目","払戻金",
                "投資額","回収額","収支","買い目リスト","結果データあり"]
    sh1_cols = [c for c in sh1_cols if c in bt.columns]
    _hdr_row(ws1, sh1_cols)

    ci = {c: i for i, c in enumerate(sh1_cols)}
    _has_res  = "結果データあり" in ci
    _has_skip = "見送り推奨"     in ci
    _has_hit  = "的中"           in ci

    n_rows = len(bt)
    fill_arr = [None] * n_rows
    bt_sub = bt[sh1_cols]

    if _has_res:
        no_result_mask = ~bt_sub["結果データあり"].to_numpy(dtype=bool)
    else:
        no_result_mask = np.zeros(n_rows, dtype=bool)

    if _has_skip:
        skip_mask = bt_sub["見送り推奨"].to_numpy(dtype=float) == 1
    else:
        skip_mask = np.zeros(n_rows, dtype=bool)

    if _has_hit:
        raw_hit = bt_sub["的中"].to_numpy(dtype=object)
        hit_mask  = np.array([str(v).lower() == "true"  for v in raw_hit])
        miss_mask = np.array([str(v).lower() == "false" for v in raw_hit])
    else:
        hit_mask  = np.zeros(n_rows, dtype=bool)
        miss_mask = np.zeros(n_rows, dtype=bool)

    for i in range(n_rows):
        if no_result_mask[i]:
            fill_arr[i] = ND_FILL
        elif skip_mask[i]:
            fill_arr[i] = SKIP_FILL
        elif hit_mask[i]:
            fill_arr[i] = HIT_FILL
        elif miss_mask[i]:
            fill_arr[i] = MISS_FILL

    for i, row_t in enumerate(bt_sub.itertuples(index=False, name=None)):
        fill = fill_arr[i]
        row_cells = []
        for v in row_t:
            if isinstance(v, (np.integer,)): v = int(v)
            elif isinstance(v, (np.floating,)): v = float(v)
            elif isinstance(v, (np.bool_,)): v = bool(v)
            c = WriteOnlyCell(ws1, value=v)
            if fill: c.fill = fill
            c.border = BORDER
            row_cells.append(c)
        ws1.append(row_cells)

    # シート2: 精度サマリー
    ws2 = wb.create_sheet("精度サマリー")
    _hdr_row(ws2, ["カテゴリ","グループ","レース数","的中数","的中率%","総投資","総回収","ROI%"])
    for cat in ["全体","会場別","戦略別","レースランク別","スコア帯別","脅威スコア帯別","1号艇脆弱性帯別"]:
        d = summary.get(cat, {})
        if isinstance(d, dict) and "レース数" in d:
            s = d; roi = s.get("ROI",0)
            fill = ROI_POS if roi >= 0 else ROI_NEG
            _data_row(ws2, [cat,"全体",s["レース数"],s["的中数"],s["的中率"],
                            s["総投資"],s["総回収"],roi], fill)
        else:
            for grp, s in (d.items() if isinstance(d, dict) else []):
                roi = s.get("ROI",0)
                fill = ROI_POS if roi >= 0 else ROI_NEG
                _data_row(ws2, [cat,str(grp),s.get("レース数",0),s.get("的中数",0),
                                s.get("的中率",0),s.get("総投資",0),
                                s.get("総回収",0),roi], fill)

    # シート3: 会場別サマリー（改修C互換）
    ws3 = wb.create_sheet("会場別サマリー")
    _hdr_row(ws3, ["会場","レース数","的中率%","回収率","ROI%","信頼度","判定","データ不足","要強化"])
    for v, vs in sorted(correction.get("venue_stats", {}).items(),
                        key=lambda x: x[1].get("hit_rate", 0)):
        insuf = vs.get("insufficient_data", False)
        boost = vs.get("needs_boost", False)
        fill  = RED_FILL if boost else (ND_FILL if insuf else None)
        _data_row(ws3, [
            v, vs["n_races"], vs["hit_rate"], vs["recovery_rate"],
            vs["roi"], vs["reliability"], vs["judgment"],
            "⚠" if insuf else "", "❌" if boost else "",
        ], fill)

    # シート4: 見送り精度
    ws_skip = wb.create_sheet("見送り精度")
    _hdr_row(ws_skip, ["指標","値"])
    for k, v in summary.get("見送り精度", {}).items():
        _data_row(ws_skip, [k, v])

    # シート5: 補正係数テーブル
    ws4 = wb.create_sheet("補正係数テーブル")
    tc = WriteOnlyCell(ws4, value="補正係数テーブル")
    tc.font = Font(bold=True, size=12, color="1F4E79")
    ws4.append([tc])
    ws4.append([""])
    _hdr_row(ws4, ["キー", "値"])

    def flatten(d, prefix=""):
        rows = []
        for k, v in d.items():
            key = f"{prefix}{k}"
            if isinstance(v, dict): rows.extend(flatten(v, key + "."))
            else: rows.append([key, str(v)])
        return rows
    for row in flatten(correction):
        _data_row(ws4, row)

    wb.save(str(output_path))
    print(f"  [OK] Excelレポート: {output_path}")


# ══════════════════════════════════════════════════════════
# main
# ══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="競艇バックテストエンジン v4")
    parser.add_argument("--venue",       type=str, default=None)
    parser.add_argument("--force",       action="store_true")
    parser.add_argument("--chikuseki",   type=str, default=None)
    parser.add_argument("--results_dir", type=str, default=None)
    parser.add_argument("--output",      type=str, default=None)
    args = parser.parse_args()

    chikuseki_dir = Path(args.chikuseki) if args.chikuseki else CHIKUSEKI_DIR
    results_dir   = Path(args.results_dir) if args.results_dir else RESULTS_DIR
    output_path   = Path(args.output)      if args.output      else OUTPUT_EXCEL

    print(); sep()
    print("  競艇バックテストエンジン v4  ★補正前統計ベース★")
    sep()

    print()
    print("[Step 1]  data_csv/ スキャン中...")
    if not results_dir.exists():
        print(f"  [NG] フォルダが見つかりません: {results_dir}")
        sys.exit(1)

    available_months = scan_available_months(results_dir)
    if not available_months:
        print(f"  [NG] *_payouts.csv が1件もありません: {results_dir}")
        sys.exit(1)
    print(f"  利用可能な月: {', '.join(available_months)}")

    print()
    print("[Step 2]  差分チェック（新規データ検出）...")
    cache      = {} if args.force else _load_json(CACHE_FILE)
    done_months= set(cache.get("processed_months", []))
    new_months = [m for m in available_months if m not in done_months]

    if args.force:
        print(f"  [--force] 全月を再計算します")
    elif not new_months:
        print("  新しい月のデータはありません（--force で強制再計算できます）")
    else:
        print(f"  新規取込: {', '.join(new_months)}")

    print()
    print("[Step 3]  結果CSV読み込み中...")
    res_df, pay_df = collect_all_results(results_dir, available_months)
    print(f"  results: {len(res_df)}行 / payouts: {len(pay_df)}行")

    print()
    print("[Step 4]  蓄積CSV読み込み中...")
    if not chikuseki_dir.exists():
        print(f"  [NG] 蓄積フォルダが見つかりません: {chikuseki_dir}")
        sys.exit(1)
    chikuseki_df = load_chikuseki(chikuseki_dir, venue_filter=args.venue)

    venue_col     = "会場" if "会場" in chikuseki_df.columns else "_venue_file"
    chiku_venues  = set(chikuseki_df[venue_col].unique())
    result_venues = set(pay_df["会場名"].unique()) if len(pay_df) else set()
    matchable     = chiku_venues & result_venues
    print()
    print(f"  蓄積CSV 会場 : {sorted(chiku_venues)}")
    print(f"  突合可能 会場: {sorted(matchable)}")

    print()
    print("[Step 5]  バックテスト実行中...")
    # ★修正①: run_backtest が (bt_corrected, bt_raw) のタプルを返す
    bt, bt_raw = run_backtest(
        chikuseki_df, res_df, pay_df,
        new_months=new_months if not args.force else available_months,
        force=args.force
    )
    matched = int(bt["結果データあり"].sum())
    total   = len(bt)
    print(f"  [OK] {total}レース処理 / {matched}レース突合 ({matched/total*100:.1f}%)")

    if matched == 0:
        print()
        print("  [!] 突合レースが0件です。")
        print("      蓄積CSVの「会場」列と結果CSVの「会場名」列の値が一致しているか確認してください。")
        print(f"      蓄積: {sorted(chiku_venues)}")
        print(f"      結果: {sorted(result_venues)[:10]}")
        sys.exit(1)

    print()
    print("[Step 6]  精度分析中...")
    # ★修正①: summary_raw（補正前）と summary_corrected（補正後）を両方計算
    summary_raw       = analyze(bt_raw)
    summary_corrected = analyze(bt)

    base_raw  = summary_raw.get("全体", {})
    base_corr = summary_corrected.get("全体", {})
    print(f"""
  ┌──────────────────────────────────────────────────────────┐
  │  バックテスト結果サマリー                                  │
  ├────────────────────────────┬────────────┬────────────────┤
  │  指標                      │  補正前(生)│  補正後(現状)  │
  ├────────────────────────────┼────────────┼────────────────┤
  │  突合レース数               │{base_raw.get('レース数',0):>10} R│{base_corr.get('レース数',0):>14} R│
  │  的中数                    │{base_raw.get('的中数',0):>10} R│{base_corr.get('的中数',0):>14} R│
  │  的中率                    │{base_raw.get('的中率',0):>9.1f} %│{base_corr.get('的中率',0):>13.1f} %│
  │  総投資                    │{base_raw.get('総投資',0):>10,} 円│{base_corr.get('総投資',0):>12,} 円│
  │  総回収                    │{base_raw.get('総回収',0):>10,} 円│{base_corr.get('総回収',0):>12,} 円│
  │  ROI                       │{base_raw.get('ROI',0):>+9.1f} %│{base_corr.get('ROI',0):>+13.1f} %│
  └────────────────────────────┴────────────┴────────────────┘
  ※ 補正前(生) → correction_table.json 生成に使用
  ※ 補正後(現状) → Excelレポートに出力""")

    print()
    print("  ■ 会場別（補正前）")
    for v, s in summary_raw.get("会場別", {}).items():
        n=s.get("レース数",0); hit=s.get("的中率",0); roi=s.get("ROI",0)
        print(f"    {v:<6}  的中率{hit:5.1f}%  ROI:{roi:+6.1f}%  ({n}R)")

    print()
    print("  ■ レースランク別（補正前）")
    for rank in ["S","A","B","C","D"]:
        s=summary_raw["レースランク別"].get(rank,{})
        n=s.get("レース数",0); hit=s.get("的中率",0); roi=s.get("ROI",0)
        bar="█"*int(hit/2)
        print(f"    {rank}: {hit:5.1f}%  ROI:{roi:+6.1f}%  ({n}R)  {bar}")

    print()
    print("  ■ 戦略別（補正前）")
    for strat, s in summary_raw["戦略別"].items():
        n=s.get("レース数",0); hit=s.get("的中率",0); roi=s.get("ROI",0)
        label=strat[:22]+"..." if len(strat)>22 else strat
        print(f"    {label:<26}  {hit:5.1f}%  ROI:{roi:+6.1f}%  ({n}R)")

    print()
    print("[Step 7]  補正係数テーブル更新...")
    # ★修正①②: summary_raw を渡すことで自己参照ループを断ち切る
    correction = generate_correction_table(summary_raw, summary_corrected)

    # アーカイブ保存
    _archive_dir = CORRECTION_JSON.parent / "archive"
    _archive_dir.mkdir(exist_ok=True)
    _ym = datetime.now().strftime("%Y%m")
    _archive_path = _archive_dir / f"correction_table_{_ym}.json"
    if CORRECTION_JSON.exists():
        shutil.copy(str(CORRECTION_JSON), str(_archive_path))
        print(f"  [OK] アーカイブ保存: {_archive_path}")

    _save_json(CORRECTION_JSON, correction)
    print(f"  [OK] correction_table.json: {CORRECTION_JSON}")
    print(f"       補正前的中率: {correction['base_hit_rate']}%  "
          f"補正前ROI: {correction['base_roi']}%  "
          f"補正前回収率: {correction.get('base_recovery', 0):.4f}  "
          f"スコア閾値: {correction['score_threshold']}")
    if correction.get("corrected_hit_rate") is not None:
        print(f"       補正後的中率: {correction['corrected_hit_rate']}%  "
              f"補正後ROI: {correction['corrected_roi']}%（参考値）")

    print()
    print("  ■ 会場別サマリー")
    print(f"  {'会場':<8} {'n':>5} {'的中率':>7} {'回収率':>7} {'信頼度':>6} {'判定'}")
    print("  " + "-" * 54)
    for v, vs in sorted(correction.get("venue_stats", {}).items(),
                        key=lambda x: x[1].get("hit_rate", 0)):
        print(f"  {v:<8} {vs['n_races']:>5} {vs['hit_rate']:>6.1f}% "
              f"{vs['recovery_rate']:>6.3f}  {vs['reliability']:>5.2f}  "
              f"{vs['judgment']}")

    print()
    print("[Step 8]  Excelレポート出力...")
    # レポートは補正後の bt を使う（実運用の実績と対応）
    write_report(bt, summary_corrected, correction, output_path)

    cache["processed_months"] = available_months
    cache["last_run"]         = datetime.now().strftime("%Y-%m-%d %H:%M")
    cache["total_races"]      = total
    cache["matched_races"]    = matched
    _save_json(CACHE_FILE, cache)

    print()
    sep()
    print("  完了！")
    sep()
    print(f"  レポート      : {output_path}")
    print(f"  補正テーブル  : {CORRECTION_JSON}")
    print(f"  生データキャッシュ: {BT_RAW_CACHE_CSV}")
    if new_months:
        print(f"  今回取り込んだ月: {', '.join(new_months)}")
    print()
    print("  【次回】月別CSVを data_csv/ に追加するだけで自動取込されます")
    print("  python backtest_engine.py")
    print()


if __name__ == "__main__":
    main()
